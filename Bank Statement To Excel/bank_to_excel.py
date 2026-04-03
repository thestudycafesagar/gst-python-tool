"""
Bank Statement to Excel Converter
Single-file application — all modules combined, Tkinter UI.
Run: python bank_to_excel.py
"""

# ═══════════════════════════════════════════════════════════════════
# STANDARD LIBRARY
# ═══════════════════════════════════════════════════════════════════
import io
import os
import re
import sys
import logging
import tempfile
import traceback
import threading
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ═══════════════════════════════════════════════════════════════════
# TKINTER
# ═══════════════════════════════════════════════════════════════════
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ═══════════════════════════════════════════════════════════════════
# THIRD-PARTY (imported lazily where possible)
# ═══════════════════════════════════════════════════════════════════
try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas not installed. Run: pip install pandas")
    sys.exit(1)

try:
    import numpy as np
except ImportError:
    np = None

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# ██████╗  █████╗ ███╗   ██╗██╗  ██╗    ██████╗ ███████╗████████╗███████╗ ██████╗████████╗ ██████╗ ██████╗
# ██╔══██╗██╔══██╗████╗  ██║██║ ██╔╝    ██╔══██╗██╔════╝╚══██╔══╝██╔════╝██╔════╝╚══██╔══╝██╔═══██╗██╔══██╗
# ██████╔╝███████║██╔██╗ ██║█████╔╝     ██║  ██║█████╗     ██║   █████╗  ██║        ██║   ██║   ██║██████╔╝
# ██╔══██╗██╔══██║██║╚██╗██║██╔═██╗     ██║  ██║██╔══╝     ██║   ██╔══╝  ██║        ██║   ██║   ██║██╔══██╗
# ██████╔╝██║  ██║██║ ╚████║██║  ██╗    ██████╔╝███████╗   ██║   ███████╗╚██████╗   ██║   ╚██████╔╝██║  ██║
# ╚═════╝ ╚═╝  ╚═╝╚═╝  ╚═══╝╚═╝  ╚═╝   ╚═════╝ ╚══════╝   ╚═╝   ╚══════╝ ╚═════╝   ╚═╝    ╚═════╝ ╚═╝  ╚═╝
# ═══════════════════════════════════════════════════════════════════════════════

BANK_REGISTRY = {
    "BOI": ("Bank of India", [
        r"BANK\s*OF\s*INDIA", r"bankofindia\.co\.in",
    ]),
    "EQUITAS": ("Equitas Small Finance Bank", [
        r"EQUITAS\s*SMALL\s*FINANCE", r"EQUITAS\s*BANK", r"equitasbank\.com",
        r"ESFBL\d", r"equitasbank",
    ]),
    "HDFC": ("HDFC Bank", [
        r"HDFC\s*BANK", r"HDFC\s*Bank\s*Ltd", r"hdfcbank\.com",
        r"HDFC\d{7}", r"NEFTIFSC:\s*HDFC", r"\bHDFC0\d{6}\b",
    ]),
    "ICICI": ("ICICI Bank", [
        r"ICICI\s*BANK", r"ICICI\s*Bank\s*Limited", r"icicibank\.com",
        r"\bICIC0\d{6}\b",
    ]),
    "IDFC": ("IDFC First Bank", [
        r"IDFC\s*FIRST\s*BANK", r"IDFC\s*FIRST", r"IDFC\s*BANK", r"idfcfirstbank\.com",
        r"\bIDF[BC]0\d{6}\b",
    ]),
    "KOTAK": ("Kotak Mahindra Bank", [
        r"KOTAK\s*MAHINDRA\s*BANK", r"Kotak\s*Bank", r"kotak\.com",
    ]),
    "UCO": ("UCO Bank", [
        r"UCO\s*BANK", r"\bUCO\b", r"ucobank\.com",
    ]),
    "YES": ("Yes Bank", [
        r"YES\s*BANK", r"yesbank\.in", r"YESB\d{7}", r"\bYESB\d",
    ]),
}

# ignore_cols: only truly meaningless serial-number/index columns.
# Data columns (Value Date, Cheque No, Ref No, Tran ID, etc.) are preserved as extra columns.
_SERIAL_IGNORE = []

BANK_COLUMN_MAP = {
    "SBI": {
        "date_cols":    ["txn date", "transaction date", "date"],
        "desc_cols":    ["description", "particulars", "narration"],
        "debit_cols":   ["debit", "withdrawal", "dr"],
        "credit_cols":  ["credit", "deposit", "cr"],
        "balance_cols": ["balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "HDFC": {
        "date_cols":    ["date"],
        "desc_cols":    ["narration", "description", "particulars"],
        "debit_cols":   ["withdrawal amt", "withdrawal amount", "debit", "dr", "withdrawal amt (inr)"],
        "credit_cols":  ["deposit amt", "deposit amount", "credit", "cr", "deposit amt (inr)"],
        "balance_cols": ["closing balance", "balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "ICICI": {
        "date_cols":    ["transaction date", "value date", "date"],
        "desc_cols":    ["description", "remarks", "narration", "particulars", "transaction remarks"],
        "debit_cols":   ["withdrawal amount (inr)", "withdrawal amount", "withdrawl (dr)", "debit", "dr"],
        "credit_cols":  ["deposit amount (inr)", "deposit amount", "deposit (cr)", "credit", "cr"],
        "balance_cols": ["balance (inr)", "balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "AXIS": {
        "date_cols":    ["tran date", "transaction date", "date"],
        "desc_cols":    ["particulars", "description", "narration"],
        "debit_cols":   ["dr", "debit", "withdrawal"],
        "credit_cols":  ["cr", "credit", "deposit"],
        "balance_cols": ["bal(inr)", "balance", "bal"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "KOTAK": {
        "date_cols":    ["transaction date", "date", "txn date"],
        "desc_cols":    ["description", "narration", "particulars", "remarks"],
        "debit_cols":   ["debit", "withdrawal", "dr"],
        "credit_cols":  ["credit", "deposit", "cr"],
        "balance_cols": ["balance", "closing balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "IDFC": {
        "date_cols":    ["transaction date", "date"],
        "desc_cols":    ["particulars", "description", "narration", "remarks"],
        "debit_cols":   ["debit", "dr"],
        "credit_cols":  ["credit", "cr"],
        "balance_cols": ["balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "BOI": {
        "date_cols":    ["date", "txn date", "transaction date"],
        "desc_cols":    ["description", "particulars", "narration", "remarks"],
        "debit_cols":   ["debit", "dr", "withdrawal"],
        "credit_cols":  ["credit", "cr", "deposit"],
        "balance_cols": ["closing balance", "balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "YES": {
        "date_cols":    ["date", "txn date", "transaction date"],
        "desc_cols":    ["description", "particulars", "narration"],
        "debit_cols":   ["withdrawals", "withdrawal", "debit", "dr"],
        "credit_cols":  ["deposits", "deposit", "credit", "cr"],
        "balance_cols": ["running balance", "closing balance", "balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "UCO": {
        "date_cols":    ["date", "txn date", "transaction date"],
        "desc_cols":    ["particulars", "description", "narration"],
        "debit_cols":   ["withdrawals", "withdrawal", "debit", "dr"],
        "credit_cols":  ["deposits", "deposit", "credit", "cr"],
        "balance_cols": ["balance", "closing balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "EQUITAS": {
        "date_cols":    ["date", "txn date", "transaction date"],
        "desc_cols":    ["transactions details", "transaction details", "description", "particulars", "narration"],
        "debit_cols":   ["debit", "dr", "withdrawal"],
        "credit_cols":  ["credit", "cr", "deposit"],
        "balance_cols": ["balance", "closing balance"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
    "DEFAULT": {
        "date_cols":    ["date", "txn date", "transaction date", "tran date", "value date",
                         "posting date", "entry date", "book date", "effective date",
                         "trans date", "dt", "trans dt", "value dt"],
        "desc_cols":    ["description", "particulars", "narration", "details", "remarks",
                         "transaction details", "transactions details", "transaction remarks",
                         "trans description", "trans particulars", "memo", "note",
                         "transaction narration", "payment details", "payment narration"],
        "debit_cols":   ["debit", "dr", "withdrawal", "withdrawals", "withdrawal amt",
                         "amount debited", "debit amount", "debit amt", "dr amount",
                         "dr amt", "paid out", "money out", "outflow", "debit(dr)",
                         "withdrawl", "withdrawl (dr)", "withdrawal (dr)", "debit (dr)"],
        "credit_cols":  ["credit", "cr", "deposit", "deposits", "deposit amt",
                         "amount credited", "credit amount", "credit amt", "cr amount",
                         "cr amt", "paid in", "money in", "inflow", "credit(cr)",
                         "deposit (cr)", "credit (cr)"],
        "balance_cols": ["balance", "closing balance", "bal", "running balance",
                         "available balance", "net balance", "book balance",
                         "ledger balance", "current balance", "bal amt", "balance amt"],
        "ignore_cols":  _SERIAL_IGNORE,
    },
}


def detect_bank(text: str) -> Tuple[str, str]:
    if not text:
        return "UNKNOWN", "Unknown Bank"
    upper_text = text.upper()
    # Use first 1000 chars as header zone (before transactions start)
    header_text = upper_text[:1000]
    scores = {}
    for bank_key, (display_name, patterns) in BANK_REGISTRY.items():
        score = sum(1 for p in patterns if re.search(p, header_text, re.IGNORECASE))
        if score > 0:
            scores[bank_key] = score
    if not scores:
        return "UNKNOWN", "Unknown Bank"
    best_key = max(scores, key=scores.get)
    return best_key, BANK_REGISTRY[best_key][0]


def get_column_map(bank_key: str) -> dict:
    return BANK_COLUMN_MAP.get(bank_key, BANK_COLUMN_MAP["DEFAULT"])


_METADATA_FIELDS = [
    ("Customer Name",    r"(?:customer\s*name|primary\s*holder|account\s*holder|name)\s*[:\-]\s*([A-Za-z\s\.\,]+?)(?:\n|CUSTOMER|ACCOUNT|IFSC|$)"),
    ("Customer ID",      r"(?:customer\s*id|cust\s*id|crn)\s*[:\-]\s*([A-Za-z0-9]+)"),
    ("Account No",       r"(?:account\s*no\.?|a/?c\s*no\.?|account\s*number|statement\s*of\s*account)\s*[:\-]\s*([0-9X\*]{6,20})"),
    ("Account Type",     r"(?:account\s*type)\s*[:\-]\s*([A-Za-z][A-Za-z0-9\s\-\_]+?)(?:\n|IFSC|MICR|$)"),
    ("Statement Period", r"(?:statement\s*period|period)\s*[:\-]\s*(.+?)(?:\n|$)"),
    ("IFSC",             r"\bIFSC\s*(?:code)?\s*[:\-]?\s*([A-Z]{4}0[A-Z0-9]{6})\b"),
    ("MICR",             r"\bMICR\s*[:\-]?\s*(\d{9})\b"),
    ("Branch",           r"(?:account\s*branch|branch\s*name)\s*[:\-]\s*(.+?)(?:\n|$)"),
    ("Currency",         r"\bCURRENCY\s*[:\-]\s*([A-Z]{3})\b"),
]


def extract_account_metadata(full_text: str, tables: List[pd.DataFrame] = None) -> dict:
    """Extract account/customer header info from text and any preamble tables."""
    from collections import OrderedDict
    header = full_text[:1200]
    result = OrderedDict()

    # Capture literal Mobile / Email if not in _METADATA_FIELDS
    m_mobile = re.search(r"\b(?:Mobile|Phone|Mob)\s*[:\-]\s*([\d\+\-\s]{10,15})", header, re.IGNORECASE)
    if m_mobile: result["Mobile"] = m_mobile.group(1).strip()
        
    m_email = re.search(r"\b(?:Email|E-mail)\s*[:\-]\s*([a-zA-Z0-9_\-\.]+@[a-zA-Z0-9_\-\.]+)\b", header, re.IGNORECASE)
    if m_email: result["Email"] = m_email.group(1).strip()

    # Existing regex fallback
    for label, pattern in _METADATA_FIELDS:
        m = re.search(pattern, header, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            val = re.sub(r"\s+", " ", val)
            if val:
                result[label] = val

    # Special fallback for "Dear <Name>" style greetings
    if "Customer Name" not in result:
        m_dear = re.search(r"\bDear\s+([A-Za-z\s\.]+)(?:\n|$)", header, re.IGNORECASE)
        if m_dear:
            result["Customer Name"] = m_dear.group(1).strip()

    # Try extracting summary details from small auxiliary tables
    if tables:
        for tbl in tables:
            # If the table has few rows and lots of columns, it's likely a summary!
            if isinstance(tbl, list):
                raw_table = tbl
            elif isinstance(tbl, pd.DataFrame):
                raw_table = [tbl.columns.tolist()] + tbl.values.tolist()
            else:
                continue

            if len(raw_table) > 1 and len(raw_table) <= 15:
                # Look for header rows followed by value rows
                for i in range(len(raw_table) - 1):
                    row1 = [str(x).strip() for x in raw_table[i] if x and str(x).strip()]
                    row2 = [str(x).strip() for x in raw_table[i+1] if x and str(x).strip()]
                    # Pair them up if they look like Key-Value arrays
                    # e.g ["Bank Name", "Account Number"] -> ["Bank Of India", "1234"]
                    # Usually lengths match, but we can zip them safely.
                    if len(row1) >= 2 and len(row1) == len(row2):
                        # Ensure row1 doesn't look like amount values (e.g. Rs. 100)
                        if not any(re.search(r'\d', k) for k in row1):
                            for k, v in zip(row1, row2):
                                k_clean = re.sub(r'[^a-z]', '', k.lower())
                                ignore_cols = {"date", "description", "narration", "particulars", "chqrefno", "valuedt", "withdrawalamt", "depositamt", "closingbalance", "balance", "debit", "credit", "amount", "refno"}
                                if len(k) < 40 and k_clean not in ignore_cols:
                                    # store if not already obtained
                                    if k not in result and "None" not in k.title() and v.lower() != "nan":
                                        result[k.title()] = v

    return result


# ═══════════════════════════════════════════════════════════════════
# PDF PROCESSOR
# ═══════════════════════════════════════════════════════════════════

TEXT_THRESHOLD = 50
TEXT_PAGE_RATIO = 0.5

_PDF_DATE_PATTERN = re.compile(
    r"\b(?:\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}|\d{1,2}\s+[A-Za-z]{3}\s+\d{2,4}|[A-Za-z]{3}\s+\d{1,2},?\s+\d{4})\b",
    re.VERBOSE,
)
_PDF_AMOUNT_PATTERN = re.compile(r"[\d,]+\.\d{2}")


def _raw_table_to_df(table: List[List]) -> Optional[pd.DataFrame]:
    if not table or len(table) < 2:
        return None

    # --- NEW FRAGMENTED COLUMN MERGE LOGIC ---
    header_keywords_for_merge = [
        'date', 'description', 'particulars', 'narration', 'transaction', 'amount', 
        'balance', 'debit', 'credit', 'withdrawal', 'withdrawals', 'deposit', 'deposits'
    ]
    cols_to_merge = []
    for r in table[:15]:
        r_clean = [re.sub(r'\s+', '', str(c)).lower() for c in r]
        c_idx = 0
        while c_idx < len(r_clean) - 1:
            val1 = r_clean[c_idx]
            val2 = r_clean[c_idx+1]
            if val1 and val2:
                combined = val1 + val2
                if combined in header_keywords_for_merge and val1 not in header_keywords_for_merge and val2 not in header_keywords_for_merge:
                    cols_to_merge.append(c_idx)
                    c_idx += 2
                    continue
            c_idx += 1
        if cols_to_merge:
            break
            
    if cols_to_merge:
        cols_to_merge.sort(reverse=True)
        new_table = []
        for r in table:
            new_r = list(r)
            for idx in cols_to_merge:
                if idx + 1 < len(new_r):
                    v1 = str(new_r[idx]).strip()
                    v2 = str(new_r[idx+1]).strip()
                    combined_clean = re.sub(r'\s+', '', v1 + v2).lower()
                    if combined_clean in header_keywords_for_merge:
                        val = v1 + v2
                    else:
                        val = v1 + (' ' if v1 and v2 else '') + v2
                    new_r[idx] = val
                    del new_r[idx+1]
            new_table.append(new_r)
        table = new_table
    # -----------------------------------------

    header = None
    data_start = 0
    
    # Common words found in transaction headers
    header_keywords = ["date", "description", "particulars", "narration", "transaction", "tran ", "amount", "balance", "debit", "credit", "chq", "ref", "withdrawal", "deposit", "value dt", "remarks"]
    
    for i, row in enumerate(table):
        # Normalize header cell names (collapse whitespace incl. newlines)
        cleaned = [re.sub(r"\s+", " ", str(c)).replace("\n", " ").strip() if c else "" for c in row]
        # Count how many header keywords are in this row
        valid_cols = sum(1 for c in cleaned if any(kw in c.lower() for kw in header_keywords))
        
        # Require at least 2 common keywords, or if the table is very narrow, at least 1
        if valid_cols >= 2 or (valid_cols >= 1 and len(cleaned) <= 3):
            # Clean up missing spaces in common headers (ValueDt -> Value Dt, ClosingBalance -> Closing Balance)
            fixed_header = []
            for col in cleaned:
                # First handle combined names lacking spaces
                c_fixed = re.sub(r'([a-z])([A-Z])', r'\1 \2', col)
                c_fixed = re.sub(r'([a-zA-Z])(?=Amt\b|Dt\b)', r'\1 ', c_fixed)
                # Next, for headers that are split across multiple lines in the PDF, replace literal newlines with space
                c_fixed = re.sub(r'\n', ' ', c_fixed)
                # Clean up any extra multi-spaces that might have been created
                c_fixed = re.sub(r'\s+', ' ', c_fixed).strip()
                fixed_header.append(c_fixed)
            header = fixed_header
            data_start = i + 1
            break
            
    if header is None:
        # Fallback to first non-empty row if no keywords match but it might still be a table
        for i, row in enumerate(table):
            cleaned = [re.sub(r"\s+", " ", str(c)).replace("\n", " ").strip() if c else "" for c in row]
            if any(cleaned):
                header = cleaned
                data_start = i + 1
                break
                
    if header is None:
        return None

    # Collect raw data rows (keep original newlines for multi-value expansion)
    raw_rows = []
    for row in table[data_start:]:
        if any(c for c in row if c and str(c).strip()):
            raw_rows.append([str(c) if c else "" for c in row])
    if not raw_rows:
        return None

    # ── Expand "all-in-one" rows where ALL non-empty cells have the same ──
    # ── newline count (consistent multi-row packing, e.g. some banks)   ──
    expanded = []
    for row in raw_rows:
        non_empty_cells = [str(c) for c in row if str(c).strip()]
        if not non_empty_cells:
            continue
        sub_counts = [c.count("\n") + 1 for c in non_empty_cells]
        max_subs = max(sub_counts)
        # Only expand if every non-empty cell has the SAME number of sub-values
        if max_subs > 1 and len(set(sub_counts)) == 1:
            split_cols = [str(c).split("\n") for c in row]
            for k in range(max_subs):
                new_row = []
                for col_vals in split_cols:
                    val = col_vals[k].strip() if k < len(col_vals) else ""
                    new_row.append(val)
                if any(new_row):
                    expanded.append(new_row)
        else:
            expanded.append([re.sub(r"\s+", " ", str(c)).strip() for c in row])

    if not expanded:
        return None

    # Normalize whitespace in all cells
    normalized = []
    for row in expanded:
        normalized.append([re.sub(r"\s+", " ", c).strip() for c in row])

    n_cols = len(header)
    padded = [r[:n_cols] + [""] * max(0, n_cols - len(r)) for r in normalized]
    return pd.DataFrame(padded, columns=header)


def _parse_text_as_table(text: str) -> Optional[pd.DataFrame]:
    lines = text.split("\n")
    rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if _PDF_DATE_PATTERN.search(line) and _PDF_AMOUNT_PATTERN.findall(line):
            rows.append(line)
    if len(rows) < 2:
        return None
    parsed = []
    for row in rows:
        date_match = _PDF_DATE_PATTERN.search(row)
        date_str = date_match.group(0) if date_match else ""
        after_date = row[date_match.end():].strip() if date_match else row
        amounts = _PDF_AMOUNT_PATTERN.findall(after_date)
        desc = _PDF_AMOUNT_PATTERN.sub("", after_date).strip()
        balance = amounts[-1] if amounts else ""
        if len(amounts) >= 2:
            amount = amounts[-2]
            is_debit = bool(re.search(r"\b(dr|debit|withdrawal)\b", row, re.IGNORECASE))
            debit = amount if is_debit else ""
            credit = "" if is_debit else amount
        else:
            debit, credit = "", ""
        parsed.append({"Date": date_str, "Description": desc,
                        "Debit": debit, "Credit": credit, "Balance": balance})
    return pd.DataFrame(parsed) if parsed else None


class PDFProcessor:
    def __init__(self, use_ocr: bool = True, ocr_engine: str = "pytesseract",
                 progress_cb=None):
        self.use_ocr = use_ocr
        self.ocr_engine = ocr_engine
        self._paddle = None
        self._progress_cb = progress_cb  # callable(page_num, total_pages) or None

    def detect_pdf_type(self, pdf_path: str) -> str:
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                total = len(pdf.pages)
                if total == 0:
                    return "scanned"
                text_pages = sum(
                    1 for p in pdf.pages
                    if p.extract_text() and len(p.extract_text().strip()) >= TEXT_THRESHOLD
                )
                return "text" if (text_pages / total) >= TEXT_PAGE_RATIO else "scanned"
        except Exception as e:
            logger.warning(f"Could not detect PDF type: {e}")
            return "unknown"

    def extract_pages_text(self, pdf_path: str) -> List[str]:
        texts = []
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    raw = page.extract_text() or ""
                    texts.append(raw.strip())
        except Exception as e:
            logger.error(f"Text extraction failed: {e}")
        return texts

    def extract_tables_pdfplumber(self, pdf_path: str) -> List[pd.DataFrame]:
        dfs = []
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
                for page_num, page in enumerate(pdf.pages, 1):
                    if self._progress_cb:
                        self._progress_cb(page_num, total_pages)
                    
                    strategies = [
                        None,  # default
                        {
                            "vertical_strategy": "lines_strict",
                            "horizontal_strategy": "lines_strict",
                            "snap_tolerance": 5, "join_tolerance": 5, "edge_min_length": 15,
                        },
                        {
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "text",
                            "intersection_y_tolerance": 15,
                            "explicit_vertical_lines": [page.width - 20],
                        },
                        {
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "snap_y_tolerance": 5, "snap_x_tolerance": 5,
                        },
                        {
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "snap_y_tolerance": 3, "snap_x_tolerance": 3,
                        }
                    ]
                    
                    best_page_dfs = []
                    max_score = -1
                    
                    for strat in strategies:
                        if strat is None:
                            tables = page.extract_tables()
                        else:
                            tables = page.extract_tables(table_settings=strat)
                            
                        current_dfs = []
                        score = 0
                        if not tables:
                            score = -1000
                        if tables:
                            for table in tables:
                                df = _raw_table_to_df(table)
                                if df is not None and len(df) > 1:
                                    current_dfs.append(df)
                                    table_score = len(df) * 2

                                    # Bonus for clean headers
                                    cols_lower_list = [str(c).lower().strip() for c in df.columns]
                                    for kw in ['date', 'narration', 'description', 'particulars', 'balance', 'withdrawal', 'deposit', 'debit', 'credit', 'ref', 'chq', 'value dt', 'transaction']:
                                        if any(kw == c or kw in c.split() or kw in c.replace(".", " ").split() for c in cols_lower_list):
                                            table_score += 20

                                    # Penalty for empty headers
                                    empty_cols = sum(1 for c in df.columns if not str(c).strip() or "unnamed" in c.lower() or c.lower().startswith("col_"))
                                    table_score -= (empty_cols * 5)

                                    # Penalty for too many columns (fragmented parsing)
                                    if len(df.columns) > 12:
                                        table_score -= (len(df.columns) - 10) * 50
                                    score += table_score
                        
                        if score > max_score:
                            max_score = score
                            best_page_dfs = current_dfs
                    
                    for df in best_page_dfs:
                        df["_source_page"] = page_num
                        dfs.append(df)
        except Exception as e:
            logger.error(f"pdfplumber table extraction failed: {e}")
        return dfs

    def extract_tables_camelot(self, pdf_path: str) -> List[pd.DataFrame]:
        try:
            import camelot
            tables = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
            if tables.n == 0:
                tables = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
            return [t.df for t in tables if len(t.df) > 1]
        except ImportError:
            return []
        except Exception as e:
            logger.warning(f"camelot extraction failed: {e}")
            return []

    def extract_text_ocr(self, pdf_path: str) -> List[str]:
        if self.ocr_engine == "paddleocr":
            return self._ocr_paddle(pdf_path)
        return self._ocr_tesseract(pdf_path)

    def _ocr_tesseract(self, pdf_path: str) -> List[str]:
        try:
            import pytesseract
        except ImportError:
            raise RuntimeError("pytesseract not installed. Run: pip install pytesseract")
            
        # Auto-detect Tesseract on Windows so the user doesn't have to set PATH manually
        import platform
        import os
        if platform.system() == "Windows":
            default_paths = [
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                os.path.expanduser(r"~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe")
            ]
            for path in default_paths:
                if os.path.exists(path):
                    pytesseract.pytesseract.tesseract_cmd = path
                    break
                    
        # Check tesseract binary is available before rendering pages
        try:
            pytesseract.get_tesseract_version()
        except Exception:
            raise RuntimeError(
                "Tesseract OCR is not installed or not in PATH.\n\n"
                "To fix:\n"
                "1. Download installer from: https://github.com/UB-Mannheim/tesseract/wiki\n"
                "2. Install to: C:\\Program Files\\Tesseract-OCR\\\n"
                "3. Add that folder to your Windows PATH\n"
                "4. Restart this app"
            )
            
        texts = []
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
                for page_num, page in enumerate(pdf.pages):
                    if self._progress_cb:
                        self._progress_cb(page_num + 1, total_pages)
                    
                    # Convert page to a PIL Image using perfectly-native pdfplumber features at 300 DPI
                    img = page.to_image(resolution=300).original
                    # Force grayscale to improve OCR
                    gray = img.convert("L")
                    
                    text = pytesseract.image_to_string(
                        gray, config="--psm 6 --oem 3 -c preserve_interword_spaces=1"
                    )
                    texts.append(text.strip())
            return texts
            
        except Exception as fallback_e:
            raise RuntimeError(f"OCR Image rendering failed: {fallback_e}")

    def _ocr_paddle(self, pdf_path: str) -> List[str]:
        try:
            from paddleocr import PaddleOCR
            import numpy as _np
            import fitz
            from PIL import Image
        except ImportError:
            return self._ocr_tesseract(pdf_path)
        if self._paddle is None:
            self._paddle = PaddleOCR(use_angle_cls=True, lang="en", show_log=False)
        texts = []
        doc = fitz.open(pdf_path)
        for page in doc:
            mat = fitz.Matrix(300 / 72, 300 / 72)
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            result = self._paddle.ocr(_np.array(img), cls=True)
            lines = [line[1][0] for line in result[0]] if result and result[0] else []
            texts.append("\n".join(lines))
        doc.close()
        return texts

    def extract_tables_from_ocr_text(self, ocr_texts: List[str]) -> List[pd.DataFrame]:
        dfs = []
        for page_text in ocr_texts:
            df = _parse_text_as_table(page_text)
            if df is not None and len(df) > 1:
                dfs.append(df)
        return dfs


# ═══════════════════════════════════════════════════════════════════
# TRANSACTION EXTRACTOR
# ═══════════════════════════════════════════════════════════════════

STANDARD_COLS = ["Date", "Description", "Debit", "Credit", "Balance"]

_DATE_PATTERNS = [
    r"\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{4}",
    r"\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2}",
    r"\d{1,2}\s+[A-Za-z]{3}\s+\d{4}",
    r"\d{1,2}-[A-Za-z]{3}-\d{4}",
    r"\d{4}[\/\-]\d{1,2}[\/\-]\d{1,2}",
    r"[A-Za-z]{3}\s+\d{1,2},?\s+\d{4}",
]
_DATE_RE = re.compile("|".join(_DATE_PATTERNS))
_AMOUNT_RE = re.compile(r"^[\s₹$,\-\(]*[\d,]+\.?\d*[\s\),]*$")


def _te_drop_repeated_headers(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    header_lower = set(c.lower().strip() for c in df.columns)
    mask = df.apply(
        lambda row: sum(str(v).lower().strip() in header_lower for v in row)
        >= max(2, len(df.columns) // 2),
        axis=1,
    )
    return df[~mask].reset_index(drop=True)


def _te_looks_like_header(cols: List[str]) -> bool:
    date_kw = {"date", "dt", "txn", "transaction", "tran date", "value date", "posting date"}
    amount_kw = {"amount", "debit", "credit", "balance", "dr", "cr",
                 "withdrawal", "deposit", "withdrawals", "deposits",
                 "particulars", "narration", "description", "remarks"}
    cols_lower = {c.lower().strip() for c in cols if len(c.strip()) < 30}
    has_date = bool(cols_lower & date_kw) or any(c == "date" or c.endswith(" date") or c.endswith(" dt") for c in cols_lower)
    
    amount_matches = 0
    for c in cols_lower:
        if c in amount_kw:
            amount_matches += 1
        elif any(kw in c.split() for kw in ("debit", "credit", "balance", "withdrawal", "deposit", "dr", "cr")):
             amount_matches += 1

    return has_date and (amount_matches >= 1) or (amount_matches >= 2)


def _te_promote_header(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    for i, row in df.head(15).iterrows():
        vals = [str(v).lower().strip() for v in row.values]
        if _te_looks_like_header(vals):
            df.columns = [str(v).strip() for v in row.values]
            return df.iloc[i + 1:].reset_index(drop=True)
    return df


def _te_split_combined_amount_col(df, combined_col, result):
    series = df[combined_col].astype(str)
    debit_vals, credit_vals = [], []
    for val in series:
        val_clean = val.strip()
        is_dr = bool(re.search(r"\bdr\b", val_clean, re.IGNORECASE))
        is_cr = bool(re.search(r"\bcr\b", val_clean, re.IGNORECASE))
        amount = re.sub(r"[^\d\.]", "", val_clean)
        if is_dr:
            debit_vals.append(amount); credit_vals.append("")
        elif is_cr:
            debit_vals.append(""); credit_vals.append(amount)
        elif val_clean.startswith("-"):
            debit_vals.append(amount); credit_vals.append("")
        else:
            debit_vals.append(""); credit_vals.append(amount)
    result["Debit"] = debit_vals
    result["Credit"] = credit_vals
    return result


def _te_parse_transactions_from_text(text: str) -> List[Tuple]:
    """
    Parse transactions from plain text. Handles:
    - Decimal amounts  (6.4, 265.98)
    - Integer amounts  (55, 10000, 9676)
    - Descriptions that appear on the line BEFORE the date (UCO Bank style)
    - Descriptions that appear AFTER the date on the same line
    - Multi-line descriptions / reference numbers after the date line
    """
    rows = []
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    n = len(lines)
    consumed = set()   # line indices already used as look-back descriptions
    prev_balance: Optional[float] = None

    # ── Detect Opening Balance line ───────────────────────────────────
    for idx, ln in enumerate(lines):
        ob_match = re.search(r"opening\s+balance[\s:]*([₹$]?\s*[\d,]+\.?\d*)", ln, re.IGNORECASE)
        if ob_match:
            try:
                ob_val = float(ob_match.group(1).replace(",", "").replace("₹", "").replace("$", "").strip())
                prev_balance = ob_val
                rows.append(("", "Opening Balance", "", "", f"{ob_val:.2f}"))
                consumed.add(idx)
            except ValueError:
                pass
            break

    # Matches a standalone numeric amount (int or decimal, with optional commas)
    _NUM_RE = re.compile(r"(?<!\S)(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?|\d+\.\d{1,2})(?!\S)")

    def is_valid_amount(part: str) -> bool:
        """Return True if 'part' looks like a monetary amount, not a ref/date."""
        clean = part.replace(",", "")
        try:
            float(clean)
        except ValueError:
            return False
        # Reject if looks like 8-digit date (DDMMYYYY) with no decimal
        if re.match(r"^\d{8}$", clean):
            return False
        # Reject very long integers with no decimal (likely reference numbers)
        if re.match(r"^\d{9,}$", clean):
            return False
        return True

    def extract_trailing_numbers(s: str):
        """Split 's' into (description_part, [num1, num2, ...]) where nums are
        a run of valid monetary amounts at the RIGHT end of the string."""
        parts = s.split()
        nums, desc_parts = [], []
        collecting = True
        for part in reversed(parts):
            if collecting and is_valid_amount(part):
                nums.insert(0, part.replace(",", ""))
            else:
                collecting = False
                desc_parts.insert(0, part)
        return " ".join(desc_parts).strip(), nums

    i = 0
    while i < n:
        line = lines[i]
        date_match = _DATE_RE.search(line)
        if not date_match:
            i += 1
            continue

        date_str = date_match.group(0)

        # ── Look BACK for description lines that precede this date ────
        back_desc_parts = []
        k = i - 1
        while k >= 0 and k not in consumed and k >= i - 3:
            prev = lines[k]
            if _DATE_RE.search(prev):
                break
            # Skip lines that look like table headers
            prev_lower = prev.lower()
            if any(kw in prev_lower for kw in (
                    "narration", "particulars", "withdrawal", "deposit",
                    "closing balance", "chq./ref", "value dt", "ref.no")):
                break
            # Accept as look-back desc only if the line has no trailing amounts
            _, trail = extract_trailing_numbers(prev)
            if not trail:
                back_desc_parts.insert(0, prev)
                consumed.add(k)
                k -= 1
            else:
                break

        # ── Extract amounts from THIS line only (not continuation lines) ─
        text_no_date = _DATE_RE.sub("", line).strip()

        # ── Collect continuation lines as extra description ─────────────
        j = i + 1
        fwd_desc_parts = []
        while j < n and not _DATE_RE.search(lines[j]) and j < i + 5:
            # Drop out if we hit what looks like a table header
            if any(kw in lines[j].lower() for kw in ("narration", "particulars", "withdrawal", "deposit", "closing balance", "chq./ref", "value dt")):
                break
            fwd_desc_parts.append(lines[j])
            j += 1
            
        combined_tail = text_no_date + " " + " ".join(fwd_desc_parts)
        inline_desc, nums = extract_trailing_numbers(combined_tail)
        combined = line + " " + " ".join(fwd_desc_parts)

        # Filter out 0.0 / 0.00 (Equitas empty placeholders)
        nums = [x for x in nums if not re.match(r"^0\.0+$", x)]

        if not nums:
            i = j
            continue

        balance_str = nums[-1]
        try:
            balance_f = float(balance_str.replace(",", ""))
        except ValueError:
            balance_f = None

        # ── Determine debit / credit ───────────────────────────────────
        debit, credit = "", ""

        def classify_amount(txn_str):
            """Return ('debit', val) or ('credit', val) using balance math."""
            try:
                txn_f = float(txn_str.replace(",", ""))
            except ValueError:
                return None, txn_str
            if prev_balance is not None and balance_f is not None:
                if abs(prev_balance + txn_f - balance_f) < 0.02:
                    return "credit", txn_str
                if abs(prev_balance - txn_f - balance_f) < 0.02:
                    return "debit", txn_str
            ctx = combined.lower()
            credit_kw = ("neftcr", "neft cr", "chqdep", "chq dep", "imps cr",
                         "cr-", "cr ", "credit", "deposit", "received",
                         "deposited", "refund", "cashback", "reversal")
            debit_kw  = ("neftdr", "neft dr", "upidr", "dr-", "dr ", "debit",
                         "withdrawal", "paid", "purchase", "emi")
            if any(k in ctx for k in credit_kw):
                return "credit", txn_str
            if any(k in ctx for k in debit_kw):
                return "debit", txn_str
            return "debit", txn_str  # default

        if len(nums) >= 3:
            # Try to find which candidate amount (any of the non-balance trailing nums)
            # satisfies balance math; ignore spurious reference numbers
            candidates = nums[:-1]  # everything except the last (balance)
            matched = False
            if prev_balance is not None and balance_f is not None:
                for cand in reversed(candidates):  # try closest to balance first
                    try:
                        f = float(cand.replace(",", ""))
                    except ValueError:
                        continue
                    if abs(prev_balance + f - balance_f) < 0.02:
                        credit = cand
                        matched = True
                        break
                    if abs(prev_balance - f - balance_f) < 0.02:
                        debit = cand
                        matched = True
                        break
            if not matched:
                # No balance validation possible — use last 2 as debit+credit
                debit, credit = nums[-3], nums[-2]
        elif len(nums) == 2:
            kind, txn_str = classify_amount(nums[0])
            if kind == "credit":
                credit = txn_str
            else:
                debit = txn_str

        # ── Build description ─────────────────────────────────────────
        all_desc_parts = back_desc_parts + ([inline_desc] if inline_desc else [])
        desc = re.sub(r"\s+", " ", " ".join(all_desc_parts)).strip()

        prev_balance = balance_f
        rows.append((date_str, desc, debit, credit, balance_str))
        consumed.add(i)
        i = j

    return rows


class TransactionExtractor:
    def __init__(self, bank_key: str = "UNKNOWN"):
        self.bank_key = bank_key
        self.col_map = get_column_map(bank_key)
        self.col_display_names = {}  # maps standard name → original PDF column name

    def extract(self, tables: List[pd.DataFrame], page_texts: List[str]) -> pd.DataFrame:
        result_table = None
        if tables:
            result_table = self._extract_from_tables(tables)
            
        result_text = self._extract_from_text(page_texts)
        
        def _is_amt(v):
            return bool(re.match(r'^[\d,\.]+$', re.sub(r'[^\d\.]', '', str(v))))
            
        def score_df(df):
            if df is None or df.empty: return -1
            score = 0
            for _, row in df.iterrows():
                d = str(row.get("Date", "")).strip()
                dr = str(row.get("Debit", "")).strip()
                cr = str(row.get("Credit", "")).strip()
                if d and d not in ("nan", "None"):
                    score += 2
                if _is_amt(dr) or _is_amt(cr):
                    score += 1
            return score
            
        score_tab = score_df(result_table)
        score_txt = score_df(result_text)
        
        # If table extraction is totally garbled but text has clean lines:
        if score_txt > score_tab * 1.3:
            result = result_text
        else:
            result = result_table if result_table is not None else result_text

        if result is None or result.empty:
            result = pd.DataFrame(columns=STANDARD_COLS)
            
        for col in STANDARD_COLS:
            if col not in result.columns:
                result[col] = ""
        return result

    def _extract_from_tables(self, tables):
        frames = []
        for tbl in tables:
            norm = self._normalize_table(tbl)
            if norm is not None and not norm.empty:
                frames.append(norm)
        return pd.concat(frames, ignore_index=True) if frames else None

    def _normalize_table(self, df):
        if df is None or df.empty:
            return None
        df = df.copy()
        df.columns = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(df.columns)]
        df = _te_drop_repeated_headers(df)
        if not _te_looks_like_header(df.columns.tolist()):
            df = _te_promote_header(df)
        if df is None or df.empty:
            return None

        # Deduplicate columns to prevent ValueError on `df[orig_col]` if duplicates exist
        new_cols = []
        seen = {}
        for c in df.columns:
            c_str = str(c)
            if c_str in seen:
                seen[c_str] += 1
                new_cols.append(f"{c_str}_{seen[c_str]}")
            else:
                seen[c_str] = 0
                new_cols.append(c_str)
        df.columns = new_cols

        col_mapping = self._map_columns(df.columns.tolist())
        if col_mapping is None:
            # Fallback to previous table's columns if length matches
            if getattr(self, "_last_valid_cols", None) and len(df.columns) == len(self._last_valid_cols):
                # Ensure the current columns are actually valid data lengths
                new_row = pd.DataFrame([df.columns.tolist()], columns=self._last_valid_cols)
                df.columns = self._last_valid_cols
                df = pd.concat([new_row, df], ignore_index=True)
                col_mapping = self._last_valid_map
            else:
                return None
        else:
            self._last_valid_cols = df.columns.tolist()
            self._last_valid_map = col_mapping
            
        # Store original PDF column names for each standard role (first table wins)
        result = pd.DataFrame()
        rev_mapping = {v: k for k, v in col_mapping.items() if v}
        ignore = set(c.lower() for c in (self.col_map.get("ignore_cols") or []))
        
        for orig_col in df.columns:
            if orig_col == "_source_page":
                continue
            
            if orig_col in rev_mapping:
                std_col = rev_mapping[orig_col]
                # If Debit and Credit are combined in the same column
                if col_mapping.get("Debit") == orig_col and col_mapping.get("Credit") == orig_col:
                    result = _te_split_combined_amount_col(df, orig_col, result)
                    # We have to use mapped names 'Debit'/'Credit' for the split
                    self.col_display_names["Debit"] = "Debit"
                    self.col_display_names["Credit"] = "Credit"
                else:
                    if std_col not in result.columns:
                        result[std_col] = df[orig_col]
                        self.col_display_names[std_col] = orig_col
            else:
                col_lower = re.sub(r"\s+", " ", orig_col).lower().strip()
                if col_lower in ignore:
                    continue
                display_name = orig_col.strip()
                if not display_name or len(display_name) < 3 or len(display_name) > 40:
                    continue
                if display_name[0].isdigit():
                    continue
                if re.match(r'^[^a-zA-Z]+$', display_name):
                    continue
                
                # we don't drop empty columns here because they might be part of the standard layout
                # and might receive data on a subsequent page.
                if display_name not in result.columns:
                    result[display_name] = df[orig_col]
        
        # Save original column order so we can restore it later
        if not hasattr(self, "original_col_order"):
            self.original_col_order = df.columns.tolist()
            
        return result

    def _map_columns(self, columns):
        lower_cols = {re.sub(r"\s+", " ", c).lower().strip(): c for c in columns}

        def find_col(candidates):
            # Exact match first
            for candidate in candidates:
                if candidate in lower_cols:
                    return lower_cols[candidate]
            # Partial match
            for candidate in candidates:
                for lc, orig in lower_cols.items():
                    if candidate in lc:
                        return orig
                    if len(lc) >= 4 and lc in candidate:
                        return orig
            return None

        date_col = find_col(self.col_map["date_cols"])

        # ── Heuristic fallback for unknown banks ──────────────────────
        if date_col is None:
            # Any column whose name contains "date" or "dt"
            for lc, orig in lower_cols.items():
                if "date" in lc or lc.endswith(" dt") or lc == "dt":
                    date_col = orig
                    break

        if date_col is None:
            # Last resort: find a column whose values look like dates
            return None

        desc_col  = find_col(self.col_map["desc_cols"])
        debit_col = find_col(self.col_map["debit_cols"])
        credit_col = find_col(self.col_map["credit_cols"])
        bal_col   = find_col(self.col_map["balance_cols"])

        # ── If still missing, guess by keyword presence in column name ─
        if desc_col is None:
            for lc, orig in lower_cols.items():
                if any(k in lc for k in ("desc", "narr", "partic", "remark", "detail", "memo")):
                    desc_col = orig
                    break
        if debit_col is None:
            for lc, orig in lower_cols.items():
                if any(k in lc for k in ("debit", "dr", "with", "paid out", "out")):
                    if orig not in (date_col, desc_col, bal_col):
                        debit_col = orig
                        break
        if credit_col is None:
            for lc, orig in lower_cols.items():
                if any(k in lc for k in ("credit", "cr", "deposit", "paid in", " in")):
                    if orig not in (date_col, desc_col, debit_col, bal_col):
                        credit_col = orig
                        break
        if bal_col is None:
            for lc, orig in lower_cols.items():
                if "bal" in lc or "balance" in lc:
                    if orig not in (date_col, desc_col, debit_col, credit_col):
                        bal_col = orig
                        break

        # A valid transaction table must have a date and at least one other recognizable column
        if date_col is None or not any([desc_col, debit_col, credit_col, bal_col]):
            return None

        return {
            "Date": date_col,
            "Description": desc_col or "",
            "Debit": debit_col or "",
            "Credit": credit_col or "",
            "Balance": bal_col or "",
        }

    def _extract_from_text(self, page_texts):
        # Join all pages so prev_balance persists across page boundaries
        combined = "\n".join(page_texts)
        rows = _te_parse_transactions_from_text(combined)
        if not rows:
            return None
        return pd.DataFrame(rows, columns=STANDARD_COLS)


# ═══════════════════════════════════════════════════════════════════
# DATA CLEANER
# ═══════════════════════════════════════════════════════════════════

DATE_FORMATS = [
    "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
    "%d/%m/%y", "%d-%m-%y",
    "%d %b %Y", "%d %b %y",
    "%d-%b-%Y", "%d-%b-%y",
    "%d %B %Y",
    "%Y-%m-%d", "%Y/%m/%d",
    "%b %d, %Y", "%b %d %Y",
    "%m/%d/%Y",
]

HEADER_KEYWORDS = {
    "date", "dt", "description", "debit", "credit", "balance",
    "txn", "transaction", "amount", "narration", "particulars",
    "withdrawal", "withdrawals", "deposit", "deposits", "dr", "cr",
    "transactions details", "transaction details", "remarks", "details",
    "paid in", "paid out", "money in", "money out", "inflow", "outflow",
    "closing balance", "running balance", "ledger balance", "memo",
    "value date", "posting date", "entry date", "ref no", "chq no",
}


def _is_amount(val: str) -> bool:
    if not val or val in ("", "nan", "None"):
        return False
    cleaned = re.sub(r"[₹$£€,\s\(\)\-]", "", str(val))
    return bool(re.match(r"^\d+\.?\d*$", cleaned))


def _normalize_amount(val) -> str:
    if not val or str(val).strip() in ("", "nan", "None", "NA", "N/A", "n/a"):
        return ""
    s = str(val).strip()
    s = re.sub(r"^Rs\.?\s*", "", s, flags=re.IGNORECASE)
    is_negative = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    s = re.sub(r"[₹$£€\s\(\)\-]", "", s)
    s = s.replace(",", "")
    if s in ("0.0", "0", "0.00", ".0"):
        return ""
    try:
        num = float(s)
        if is_negative:
            num = -abs(num)
        return f"{num:.2f}"
    except (ValueError, TypeError):
        return val


class DataCleaner:
    def clean(self, df: pd.DataFrame, original_order: list = None) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame(columns=STANDARD_COLS)
        df = df.copy()
        df = self._ensure_standard_columns(df)
        df = self._remove_empty_rows(df)
        df = self._remove_header_duplicates(df)
        df = self._merge_multiline_descriptions(df)
        df = self._standardize_dates(df)
        df = self._clean_amounts(df)
        df = self._remove_empty_rows(df)
        # Drop extra columns that ended up entirely empty, but keep them if they are part of the original PDF table
        extra_cols = [c for c in df.columns if c not in STANDARD_COLS]
        empty_extra = [c for c in extra_cols if df[c].replace("", None).isna().all()]
        # Do not drop if it's an expected original extra column AND has a real name
        empty_extra = [
            c for c in empty_extra 
            if not original_order or c not in original_order or (c.startswith("col_") and c[4:].isdigit())
        ]
        if empty_extra:
            df = df.drop(columns=empty_extra)
            extra_cols = [c for c in extra_cols if c not in empty_extra]
            
        # Reconstruct the column order based on the original extracted PDF order if provided
        ordered_cols = STANDARD_COLS + extra_cols
        if original_order:
            # We want to reorder 'ordered_cols' to match 'original_order' as much as possible.
            # But the columns in df are 'Date', 'Description', ..., plus 'Extra1'
            # We don't have the dictionary mapping here inside clean() easily unless we match them or pass it.
            pass
            
        # Enforce standard column order at the beginning
        ordered_cols = [c for c in ordered_cols if c in df.columns]
        df = df[ordered_cols]
        
        return df.reset_index(drop=True)

    def _ensure_standard_columns(self, df):
        for col in STANDARD_COLS:
            if col not in df.columns:
                df[col] = ""
        return df

    def _remove_empty_rows(self, df, date_col="Date", desc_col="Description"):
        if df.empty:
            return df
        empty_vals = ("", "nan", "None", "NA")
        df = df.replace({v: None for v in empty_vals})
        
        # Determine available columns to check for row emptiness
        subset = [c for c in [date_col, desc_col] if c in df.columns]
        if subset:
            # We want to drop the row ONLY if ALL subset columns (e.g. Date and Description) are empty.
            df = df.dropna(subset=subset, how="all")
        
        df = df.dropna(how="all")
        return df.fillna("")

    def _remove_header_duplicates(self, df):
        if df.empty:
            return df
        def is_header_row(row):
            values = [str(v).lower().strip() for v in row.values]
            return sum(1 for v in values if v in HEADER_KEYWORDS) >= 2
        return df[~df.apply(is_header_row, axis=1)].reset_index(drop=True)

    def _merge_multiline_descriptions(self, df):
        if df.empty:
            return df
        result_rows = []
        pending_desc = []
        pending_row = None
        for _, row in df.iterrows():
            date_val = str(row.get("Date", "")).strip()
            debit_val = str(row.get("Debit", "")).strip()
            credit_val = str(row.get("Credit", "")).strip()
            balance_val = str(row.get("Balance", "")).strip()
            desc_val = str(row.get("Description", "")).strip()
            has_date = bool(date_val and date_val not in ("", "nan", "None"))
            has_amounts = any([_is_amount(debit_val), _is_amount(credit_val), _is_amount(balance_val)])
            if has_date or has_amounts:
                if pending_row is not None:
                    pending_row["Description"] = " ".join(pending_desc).strip()
                    result_rows.append(pending_row)
                pending_row = row.to_dict()
                pending_desc = [desc_val] if desc_val else []
            else:
                if desc_val:
                    pending_desc.append(desc_val)
                for fld in ["Debit", "Credit", "Balance"]:
                    val = str(row.get(fld, "")).strip()
                    if _is_amount(val) and pending_row and not pending_row.get(fld):
                        pending_row[fld] = val
        if pending_row is not None:
            pending_row["Description"] = " ".join(pending_desc).strip()
            result_rows.append(pending_row)
        return pd.DataFrame(result_rows).reset_index(drop=True) if result_rows else df

    def _standardize_dates(self, df):
        if df.empty or "Date" not in df.columns:
            return df
        def parse_date(val):
            if not val or str(val).strip() in ("", "nan", "None"):
                return None
            val_str = str(val).strip()
            for fmt in DATE_FORMATS:
                try:
                    return pd.to_datetime(val_str, format=fmt).strftime("%d/%m/%Y")
                except (ValueError, TypeError):
                    pass
            try:
                return pd.to_datetime(val_str, dayfirst=True).strftime("%d/%m/%Y")
            except Exception:
                return None  # unparseable — will be dropped
        df["Date"] = df["Date"].apply(parse_date)
        # Keep rows that have a valid date OR are special rows (Opening/Closing Balance)
        is_special = df["Description"].str.strip().str.lower().isin(
            ["opening balance", "closing balance"])
        df = df[df["Date"].notna() | is_special]
        df["Date"] = df["Date"].fillna("")
        return df

    def _clean_amounts(self, df, dr_col="Debit", cr_col="Credit", bal_col="Balance"):
        if df.empty:
            return df
        for col in [dr_col, cr_col, bal_col]:
            if col in df.columns:
                df[col] = df[col].apply(_normalize_amount)
        return df


# ═══════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ═══════════════════════════════════════════════════════════════════

COLOR_HEADER_BG = "1E3A5F"
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUMMARY_BG = "2E86AB"
COLOR_ALT_ROW    = "EEF4FB"
COLOR_DEBIT      = "FFE5E5"
COLOR_CREDIT     = "E5FFE9"
COLOR_BORDER     = "BDD7EE"

COLUMN_WIDTHS = {"Date": 14, "Description": 50, "Debit": 16, "Credit": 16, "Balance": 16}


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    return re.sub(r"[\\\/\?\*\[\]:]", "_", name)[:max_len]


def _sum_col(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns:
        return 0.0
    total = 0.0
    for val in df[col]:
        try:
            s = re.sub(r"[,\s₹$£€]", "", str(val))
            if s:
                total += float(s)
        except (ValueError, TypeError):
            pass
    return total


def _get_date_range(df: pd.DataFrame) -> str:
    if "Date" not in df.columns or df["Date"].empty:
        return ""
    dates = df["Date"].dropna().replace("", None).dropna()
    if dates.empty:
        return ""
    return f"{dates.iloc[0]} - {dates.iloc[-1]}"


def export_to_excel_bytes(results: Dict[str, pd.DataFrame], include_summary: bool = True,
                          col_display_map: dict = None, metadata_map: dict = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin_border = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, df in results.items():
        ws = wb.create_sheet(title=_safe_sheet_name(sheet_name))
        display = (col_display_map or {}).get(sheet_name, {})
        metadata = (metadata_map or {}).get(sheet_name, {})

        # Title row
        ws.append([sheet_name])
        tc = ws.cell(row=1, column=1)
        tc.font = Font(bold=True, size=13, color=COLOR_HEADER_FG)
        tc.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

        # Account metadata rows (before stats)
        # Using openpyxl to output rows grid style!
        if metadata:
            grid_cols = 4  # e.g., Output 2 Key-Value pairs per Excel row to look like a summary
            items = list(metadata.items())
            
            for i in range(0, len(items), grid_cols // 2):
                row_vals = []
                for j in range(grid_cols // 2):
                    if i + j < len(items):
                        row_vals.extend([items[i + j][0], items[i + j][1]])
                    else:
                        row_vals.extend(["", ""])
                        
                ws.append(row_vals)
                mrow = ws.max_row
                for c_idx in range(1, len(row_vals) + 1, 2):
                    ws.cell(mrow, c_idx).font = Font(bold=True, size=10, color=COLOR_HEADER_BG)
                    ws.cell(mrow, c_idx).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
                    ws.cell(mrow, c_idx).border = thin_border
                    
                    ws.cell(mrow, c_idx + 1).font = Font(size=10)
                    ws.cell(mrow, c_idx + 1).border = thin_border

            ws.append([])  # blank separator

        # Stats
        td = _sum_col(df, "Debit")
        tc2 = _sum_col(df, "Credit")
        ws.append([f"Transactions: {len(df)}", f"Total Debit: {td:,.2f}",
                   f"Total Credit: {tc2:,.2f}", "", ""])
        sr = ws.max_row
        for ci in range(1, 6):
            cell = ws.cell(row=sr, column=ci)
            cell.font = Font(italic=True, size=10, color="555555")
            cell.fill = PatternFill("solid", fgColor="F0F7FF")
        ws.append([])

        # Build headers directly from the DataFrame (mapped to original PDF names if available)
        headers = [display.get(c, c) for c in df.columns]
        ws.append(headers)
        hr = ws.max_row
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=ci)
            cell.value = h
            cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[hr].height = 20

        # Data â€” display using the exact DataFrame rows
        for row_idx, row in df.iterrows():
            row_vals = [row.get(c, "") for c in df.columns]
            ws.append(row_vals)
            dr = ws.max_row
            is_alt = (row_idx % 2 == 0)
            dv = str(row.get("Debit", "")).strip()
            cv = str(row.get("Credit", "")).strip()
            for ci, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=dr, column=ci)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_name == "Description"))
                if col_name == "Debit" and dv:
                    cell.fill = PatternFill("solid", fgColor=COLOR_DEBIT)
                    cell.font = Font(color="CC0000")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name == "Credit" and cv:
                    cell.fill = PatternFill("solid", fgColor=COLOR_CREDIT)
                    cell.font = Font(color="006600")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in ("Debit", "Credit", "Balance"):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if is_alt and col_name not in ("Debit", "Credit"):
                        cell.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
                elif is_alt:
                    cell.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

        # Column widths: use role-based widths directly by target column
        role_widths = {"Date": 14, "Description": 50, "Debit": 16, "Credit": 16, "Balance": 16}
        for ci, col_name in enumerate(df.columns, 1):
            if col_name in role_widths:
                ws.column_dimensions[get_column_letter(ci)].width = role_widths[col_name]
            else:
                ws.column_dimensions[get_column_letter(ci)].width = 18
        last_col = get_column_letter(len(headers))
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)
        ws.auto_filter.ref = f"A{hr}:{last_col}{ws.max_row}"

    if include_summary and len(results) > 1:
        ws2 = wb.create_sheet(title="Summary", index=0)
        ws2.append(["Bank Statement Extraction Summary"])
        tc = ws2.cell(row=1, column=1)
        tc.font = Font(bold=True, size=14, color=COLOR_HEADER_FG)
        tc.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        ws2.merge_cells("A1:E1")
        ws2.row_dimensions[1].height = 26
        ws2.append([])
        cols = ["File / Sheet", "Transactions", "Total Debit", "Total Credit", "Date Range"]
        ws2.append(cols)
        hdr_row = ws2.max_row
        from openpyxl.styles import Font as _Font, PatternFill as _PF, Alignment as _Aln, Border as _Bdr, Side as _Side
        thin_b2 = _Bdr(left=_Side(style="thin", color=COLOR_BORDER), right=_Side(style="thin", color=COLOR_BORDER),
                       top=_Side(style="thin", color=COLOR_BORDER), bottom=_Side(style="thin", color=COLOR_BORDER))
        for ci, h in enumerate(cols, 1):
            cell = ws2.cell(row=hdr_row, column=ci)
            cell.font = _Font(bold=True, color=COLOR_HEADER_FG)
            cell.fill = _PF("solid", fgColor=COLOR_SUMMARY_BG)
            cell.alignment = _Aln(horizontal="center", vertical="center")
            cell.border = thin_b2
        grand_d, grand_c, grand_t = 0.0, 0.0, 0
        for i, (name, df) in enumerate(results.items()):
            td = _sum_col(df, "Debit"); tc2 = _sum_col(df, "Credit")
            grand_d += td; grand_c += tc2; grand_t += len(df)
            ws2.append([name, len(df), f"{td:,.2f}", f"{tc2:,.2f}", _get_date_range(df)])
            row = ws2.max_row
            bg = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
            for ci in range(1, 6):
                cell = ws2.cell(row=row, column=ci)
                cell.fill = _PF("solid", fgColor=bg)
                cell.border = thin_b2
        ws2.append([])
        ws2.append(["TOTAL", grand_t, f"{grand_d:,.2f}", f"{grand_c:,.2f}", ""])
        tr = ws2.max_row
        for ci in range(1, 6):
            cell = ws2.cell(row=tr, column=ci)
            cell.font = _Font(bold=True, color=COLOR_HEADER_FG)
            cell.fill = _PF("solid", fgColor=COLOR_HEADER_BG)
            cell.border = thin_b2
        for col, w in zip("ABCDE", [40, 14, 18, 18, 26]):
            ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_to_csv_bytes(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8-sig")


# ═══════════════════════════════════════════════════════════════════
# PROCESSING CORE
# ═══════════════════════════════════════════════════════════════════

def process_pdf(pdf_path: str, filename: str, use_ocr: bool, ocr_engine: str,
                progress_cb=None, selected_bank: str = "Auto-Detect"):
    info = {"filename": filename, "bank": "Unknown", "pdf_type": "unknown",
            "rows_extracted": 0, "error": None, "warnings": []}
    try:
        processor = PDFProcessor(use_ocr=use_ocr, ocr_engine=ocr_engine,
                                 progress_cb=progress_cb)
        pdf_type = processor.detect_pdf_type(pdf_path)
        info["pdf_type"] = pdf_type
        page_texts, tables = [], []
        processor_raw_tables = []
        if pdf_type == "text":
            page_texts = processor.extract_pages_text(pdf_path)
            # Fetch raw lists instead of DFs for metadata parsing purpose where possible
            try:
                import pdfplumber
                with pdfplumber.open(pdf_path) as p:
                    for pg in p.pages:
                        extracted = pg.extract_tables()
                        if extracted: processor_raw_tables.extend(extracted)
            except Exception: pass
            
            tables = processor.extract_tables_pdfplumber(pdf_path)
            if not tables:
                tables = processor.extract_tables_camelot(pdf_path)
                if tables:
                    info["warnings"].append("Used camelot for table extraction.")
        elif pdf_type == "scanned":
            if not use_ocr:
                info["error"] = "OCR is disabled. Enable OCR to process scanned PDFs."
                return None, info
            page_texts = processor.extract_text_ocr(pdf_path)
            tables = processor.extract_tables_from_ocr_text(page_texts)
        else:
            page_texts = processor.extract_pages_text(pdf_path)
            if not any(page_texts):
                page_texts = processor.extract_text_ocr(pdf_path)
            tables = processor.extract_tables_pdfplumber(pdf_path)

        full_text = "\n".join(page_texts)
        if selected_bank == "Auto-Detect":
            bank_key, bank_name = detect_bank(full_text)
        elif selected_bank == "Generic / Not in list":
            bank_key, bank_name = "UNKNOWN", "Generic / Not in list"
        else:
            bank_key = next((k for k, v in BANK_REGISTRY.items() if v[0] == selected_bank), "UNKNOWN")
            bank_name = selected_bank

        info["bank"] = bank_name

        info["account_metadata"] = extract_account_metadata(full_text, tables=processor_raw_tables)

        extractor = TransactionExtractor(bank_key=bank_key)
        df_raw = extractor.extract(tables=tables, page_texts=page_texts)
        if df_raw is None or df_raw.empty:
            info["error"] = "No transactions could be extracted."
            return None, info

        # If table extraction captured display names, use them;
        # otherwise derive display names from the bank's column config
        if extractor.col_display_names:
            info["col_display_names"] = extractor.col_display_names
        else:
            cm = get_column_map(bank_key)
            def _first_title(lst): return lst[0].title() if lst else None
            info["col_display_names"] = {
                k: v for k, v in {
                    "Date":        _first_title(cm.get("date_cols", [])),
                    "Description": _first_title(cm.get("desc_cols", [])),
                    "Debit":       _first_title(cm.get("debit_cols", [])),
                    "Credit":      _first_title(cm.get("credit_cols", [])),
                    "Balance":     _first_title(cm.get("balance_cols", [])),
                }.items() if v
            }
            
        info["original_col_order"] = getattr(extractor, "original_col_order", [])

        # Clean the dynamic DataFrame using standard names
        cleaner = DataCleaner()
        df_clean = df_raw.copy()
        date_c = "Date"
        desc_c = "Description"
        dr_c = "Debit"
        cr_c = "Credit"
        bal_c = "Balance"

        df_clean = cleaner._remove_empty_rows(df_clean, date_col=date_c, desc_col=desc_c)
        df_clean = cleaner._remove_header_duplicates(df_clean)
        
        # Merge multiline descriptions using mapped columns
        if not df_clean.empty:
            result_rows = []
            pending_desc = []
            pending_row_dict = None
            for _, row in df_clean.iterrows():
                d_val = str(row.get(date_c, "")).strip()
                dr_val = str(row.get(dr_c, "")).strip()
                cr_val = str(row.get(cr_c, "")).strip()
                bal_val = str(row.get(bal_c, "")).strip()
                desc_val = str(row.get(desc_c, "")).strip()
                has_date = bool(d_val and d_val not in ("", "nan", "None"))
                has_amounts = any([_is_amount(dr_val), _is_amount(cr_val), _is_amount(bal_val)])
                if has_date or has_amounts:
                    if pending_row_dict is not None:
                        pending_row_dict[desc_c] = " ".join(pending_desc).strip()
                        result_rows.append(pending_row_dict)
                    pending_row_dict = row.to_dict()
                    pending_desc = [desc_val] if desc_val else []
                else:
                    if desc_val:
                        pending_desc.append(desc_val)
                    for fld, mapped_fld in [("Debit", dr_c), ("Credit", cr_c), ("Balance", bal_c)]:
                        val = str(row.get(mapped_fld, "")).strip()
                        if _is_amount(val) and pending_row_dict and not pending_row_dict.get(mapped_fld):
                            pending_row_dict[mapped_fld] = val
            if pending_row_dict is not None:
                pending_row_dict[desc_c] = " ".join(pending_desc).strip()
                result_rows.append(pending_row_dict)
            df_clean = pd.DataFrame(result_rows)

        if date_c in df_clean.columns:
            # simple standardization of date column
            def parse_date(val):
                if not val or str(val).strip() in ("", "nan", "None"): return None
                val_str = str(val).strip()
                for fmt in DATE_FORMATS:
                    try: return pd.to_datetime(val_str, format=fmt).strftime("%d/%m/%Y")
                    except (ValueError, TypeError): pass
                try: return pd.to_datetime(val_str, dayfirst=True).strftime("%d/%m/%Y")
                except Exception: return None
            df_clean[date_c] = df_clean[date_c].apply(parse_date)
            # Keep valid dates or opening/closing strings
            if desc_c in df_clean.columns:
                is_sp = df_clean[desc_c].astype(str).str.strip().str.lower().isin(["opening balance", "closing balance"])
                df_clean = df_clean[df_clean[date_c].notna() | is_sp]
            else:
                df_clean = df_clean[df_clean[date_c].notna()]
            df_clean[date_c] = df_clean[date_c].fillna("")

        df_clean = cleaner._clean_amounts(df_clean, dr_col=dr_c, cr_col=cr_c, bal_col=bal_c)
        df_clean = cleaner._remove_empty_rows(df_clean, date_col=date_c, desc_col=desc_c)
        df_clean = df_clean.reset_index(drop=True)
        
        # Drop generated extra columns that ended up entirely empty (like col_7)
        extra_cols = [c for c in df_clean.columns if c not in STANDARD_COLS and re.match(r"^col_\d+$", c)]
        empty_extra = [c for c in extra_cols if df_clean[c].replace("", None).isna().all()]
        if empty_extra:
            df_clean = df_clean.drop(columns=empty_extra)
        
        # Enforce column ordering
        # Try to use original PDF column order if available, else fallback to standard
        final_cols = []
        if info.get("original_col_order"):
            rev_map = {v: k for k, v in info.get("col_display_names", {}).items()}
            for orig in info["original_col_order"]:
                # If it's a mapped standard column
                if orig in rev_map:
                    std_name = rev_map[orig]
                    if std_name in df_clean.columns and std_name not in final_cols:
                        final_cols.append(std_name)
                elif orig in df_clean.columns and orig not in final_cols:
                    final_cols.append(orig)
            
            # Ensure any standard columns not in original order are also appended
            for c in STANDARD_COLS:
                if c in df_clean.columns and c not in final_cols:
                    final_cols.append(c)
        else:
            extra_cols = [c for c in df_clean.columns if c not in STANDARD_COLS]
            final_cols = STANDARD_COLS + extra_cols
            
        final_cols = [c for c in final_cols if c in df_clean.columns]
        df_clean = df_clean[final_cols]

        if df_clean.empty:
            if not df_raw.empty:
                # Fallback: if cleaning stripped everything, return uncleaned but standardized data
                info["warnings"].append("Cleaned data was empty. Returning uncleaned data.")
                df_clean = df_raw.fillna("")
            else:
                info["error"] = "Transactions found but removed during cleaning."
                return None, info

        info["rows_extracted"] = len(df_clean)
        return df_clean, info
    except RuntimeError as re_err:
        if "Tesseract OCR is not installed" in str(re_err):
            info["error"] = str(re_err)
        else:
            info["error"] = traceback.format_exc()
        return None, info
    except Exception:
        info["error"] = traceback.format_exc()
        return None, info


# ═══════════════════════════════════════════════════════════════════
# TKINTER APPLICATION
# ═══════════════════════════════════════════════════════════════════

# Theme colours
C_BG       = "#F0F4F8"
C_NAVY     = "#1E3A5F"
C_TEAL     = "#2E86AB"
C_WHITE    = "#FFFFFF"
C_LIGHT    = "#EEF4FB"
C_DEBIT    = "#CC0000"
C_CREDIT   = "#006600"
C_BORDER   = "#BDD7EE"
C_TEXT     = "#1a202c"
C_SUBTEXT  = "#555555"
C_RED_BG   = "#FFE5E5"
C_GREEN_BG = "#E5FFE9"
C_WARN     = "#856404"

# ── Theme palettes (light = default, dark = GST Suite dark mode) ─────────────
# bg-role and fg-role colours are intentionally kept separate so that
# the recolour pass can apply the right mapping per attribute type and avoid
# the navy/text hex collision (#1E3A5F is used for both in light mode).
_THEMES = {
    "Light": dict(
        # ── backgrounds / surfaces ──
        bg="#F0F4F8",    navy="#1E3A5F",   teal="#2E86AB",
        white="#FFFFFF", light="#EEF4FB",  border="#BDD7EE",
        listbox_bg="#F7FAFF",
        red_bg="#FFE5E5", green_bg="#E5FFE9",
        # ── foreground / text ──
        text="#1a202c",  subtext="#555555", sidebar_fg="#A8D8EA",
        nav_fg="#1E3A5F",    # navy used as a label colour (section titles)
    ),
    "Dark": dict(
        # ── backgrounds / surfaces ──
        bg="#1e293b",    navy="#0f172a",   teal="#22d3ee",
        white="#1e2d3d", light="#162032",  border="#334155",
        listbox_bg="#263547",
        red_bg="#3f1010", green_bg="#0a2e10",
        # ── foreground / text ──
        text="#e2e8f0",  subtext="#94a3b8", sidebar_fg="#93c5fd",
        nav_fg="#22d3ee",    # navy-as-text becomes teal so it stays readable
    ),
}


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Bank Statement → Excel Converter")
        self.geometry("1100x750")
        self.minsize(900, 600)
        self.configure(bg=C_BG)
        self._selected_files: List[str] = []
        self._results: Dict[str, pd.DataFrame] = {}
        self._infos: List[dict] = []
        # OCR controls were removed from the sidebar but are still read in
        # _run_processing — define them here with sensible defaults so the
        # background thread does not crash with AttributeError.
        self._ocr_var = tk.BooleanVar(value=False)
        self._ocr_engine_var = tk.StringVar(value="pytesseract")
        self._combine_var = tk.BooleanVar(value=True)
        self._preview_var = tk.IntVar(value=50)
        self._setup_styles()
        self._build_ui()

    def _setup_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("TNotebook", background=C_BG, borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure("TNotebook.Tab", padding=[16, 8], font=("Segoe UI", 11, "bold"),
                        background=C_LIGHT, foreground=C_SUBTEXT, borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", C_NAVY), ("active", C_BORDER)],
                  foreground=[("selected", "#ffffff"), ("active", C_NAVY)],
                  focuscolor=[("selected", C_NAVY)])
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=30,
                        background=C_WHITE, fieldbackground=C_WHITE, foreground=C_TEXT,
                        borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"),
                        background=C_NAVY, foreground="#ffffff", padding=[10, 7],
                        relief="flat")
        style.map("Treeview.Heading", background=[("active", C_TEAL)])
        style.map("Treeview", background=[("selected", C_TEAL)],
                  foreground=[("selected", "#ffffff")])
        style.configure("TScrollbar", background=C_LIGHT, troughcolor=C_BG,
                        borderwidth=0, arrowsize=12)
        style.configure("TProgressbar", troughcolor=C_BORDER, background=C_TEAL,
                        borderwidth=0, thickness=8)
        style.configure("TCombobox", font=("Segoe UI", 11),
                        padding=[8, 7], arrowsize=16,
                        background=C_WHITE, fieldbackground=C_WHITE,
                        foreground=C_TEXT, selectbackground=C_TEAL,
                        selectforeground="#ffffff", relief="flat", borderwidth=1)
        style.map("TCombobox",
                  fieldbackground=[("readonly", C_WHITE)],
                  foreground=[("readonly", C_TEXT)],
                  background=[("readonly", C_WHITE)])
        self.option_add("*TCombobox*Listbox.font", ("Segoe UI", 11))
        self.option_add("*TCombobox*Listbox.background", C_WHITE)
        self.option_add("*TCombobox*Listbox.foreground", C_TEXT)
        self.option_add("*TCombobox*Listbox.selectBackground", C_TEAL)
        self.option_add("*TCombobox*Listbox.selectForeground", "#ffffff")

    # ── Called by GST Suite when the header Dark / Light toggle fires ─────────
    def set_theme(self, mode: str):
        global C_BG, C_NAVY, C_TEAL, C_WHITE, C_LIGHT, C_BORDER
        global C_TEXT, C_SUBTEXT, C_RED_BG, C_GREEN_BG

        t   = _THEMES.get(mode, _THEMES["Light"])
        old = _THEMES["Dark" if mode == "Light" else "Light"]

        # 1) Update module-level palette so future dynamic widget creation
        #    (e.g. _render_stats, _render_tabs) picks up the correct colours.
        C_BG, C_NAVY, C_TEAL   = t["bg"],    t["navy"],  t["teal"]
        C_WHITE, C_LIGHT       = t["white"], t["light"]
        C_BORDER, C_TEXT       = t["border"],t["text"]
        C_SUBTEXT              = t["subtext"]
        C_RED_BG, C_GREEN_BG   = t["red_bg"],t["green_bg"]

        # 2) Re-apply ttk styles so new Treeview / Notebook widgets look right.
        style = ttk.Style(self)
        style.configure("TNotebook", background=C_BG, borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure("TNotebook.Tab", padding=[16, 8], font=("Segoe UI", 11, "bold"),
                        background=C_LIGHT, foreground=C_SUBTEXT, borderwidth=0)
        style.map("TNotebook.Tab",
                  background=[("selected", C_NAVY), ("active", C_BORDER)],
                  foreground=[("selected", "#ffffff"), ("active", C_NAVY)],
                  focuscolor=[("selected", C_NAVY)])
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=30,
                        background=C_WHITE, fieldbackground=C_WHITE,
                        foreground=C_TEXT, borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"),
                        background=C_NAVY, foreground="#ffffff", padding=[10, 7], relief="flat")
        style.map("Treeview.Heading", background=[("active", C_TEAL)])
        style.map("Treeview", background=[("selected", C_TEAL)],
                  foreground=[("selected", "#ffffff")])
        style.configure("TScrollbar", background=C_LIGHT, troughcolor=C_BG,
                        borderwidth=0, arrowsize=12)
        style.configure("TProgressbar", troughcolor=C_BORDER, background=C_TEAL,
                        borderwidth=0, thickness=8)
        style.configure("TCombobox", font=("Segoe UI", 11),
                        padding=[8, 7], arrowsize=16,
                        background=C_WHITE, fieldbackground=C_WHITE,
                        foreground=C_TEXT, selectbackground=C_TEAL,
                        selectforeground="#ffffff", relief="flat", borderwidth=1)
        style.map("TCombobox",
                  fieldbackground=[("readonly", C_WHITE)],
                  foreground=[("readonly", C_TEXT)],
                  background=[("readonly", C_WHITE)])
        self.option_add("*TCombobox*Listbox.font", ("Segoe UI", 11))
        self.option_add("*TCombobox*Listbox.background", C_WHITE)
        self.option_add("*TCombobox*Listbox.foreground", C_TEXT)
        self.option_add("*TCombobox*Listbox.selectBackground", C_TEAL)
        self.option_add("*TCombobox*Listbox.selectForeground", "#ffffff")

        # 3) Separate bg-swap and fg-swap to avoid the navy/text collision
        #    (both equal #1E3A5F in light mode but diverge in dark mode).
        _BG_KEYS = ("bg", "navy", "teal", "white", "light", "border",
                    "listbox_bg", "red_bg", "green_bg")
        _FG_KEYS = ("text", "subtext", "sidebar_fg", "nav_fg")

        bg_swap = {old[k].lower(): t[k] for k in _BG_KEYS
                   if old[k].lower() != t[k].lower()}
        fg_swap = {old[k].lower(): t[k] for k in _FG_KEYS
                   if old[k].lower() != t[k].lower()}

        # 4) Walk every plain-tk widget and remap recognised colours.
        def _recolor(w):
            for attr in ("bg", "background", "highlightbackground",
                         "highlightcolor", "activebackground"):
                try:
                    val = str(w.cget(attr)).lower()
                    if val in bg_swap:
                        w.configure(**{attr: bg_swap[val]})
                except Exception:
                    pass
            for attr in ("fg", "foreground", "activeforeground"):
                try:
                    val = str(w.cget(attr)).lower()
                    if val in fg_swap:
                        # text/subtext/nav_fg colours swap to their dark equivalents
                        w.configure(**{attr: fg_swap[val]})
                    # NOTE: white text (#ffffff) and teal text are intentionally
                    # left untouched — they stay visible on dark backgrounds.
                except Exception:
                    pass
            for child in w.winfo_children():
                _recolor(child)

        _recolor(self)

    # ── Layout ──────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_header()
        content = tk.Frame(self, bg=C_BG)
        content.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)
        self._build_main(content)

    def _build_header(self):
        # Top teal accent stripe
        tk.Frame(self, bg=C_TEAL, height=4).pack(fill="x")

        hdr = tk.Frame(self, bg=C_NAVY)
        hdr.pack(fill="x")

        inner = tk.Frame(hdr, bg=C_NAVY, padx=20, pady=14)
        inner.pack(fill="x")

        # Left: icon + title
        left = tk.Frame(inner, bg=C_NAVY)
        left.pack(side="left", fill="y")
        tk.Label(left, text="🏦  Bank Statement  →  Excel",
                 font=("Segoe UI", 19, "bold"), bg=C_NAVY, fg=C_WHITE).pack(side="left")

        # Right: supported banks
        right = tk.Frame(inner, bg=C_NAVY)
        right.pack(side="right", fill="y")
        tk.Label(right, text="HDFC · ICICI · SBI · Kotak · IDFC · BOI · UCO · Yes · Equitas",
                 font=("Segoe UI", 9), bg=C_NAVY, fg="#A8D8EA").pack(side="right")
        tk.Label(right, text="Supported Banks: ",
                 font=("Segoe UI", 9, "bold"), bg=C_NAVY, fg=C_TEAL).pack(side="right")

        # Bottom teal accent stripe
        tk.Frame(self, bg=C_TEAL, height=2).pack(fill="x")

    def _build_main(self, parent):
        main = tk.Frame(parent, bg=C_BG)
        main.grid(row=0, column=0, sticky="nsew", pady=(8, 0))
        main.columnconfigure(0, weight=1)
        main.rowconfigure(3, weight=1)

        # ── Row 0: File Selection ────────────────────────────────────
        file_card = _card(main, row=0)
        _section_label(file_card, "📂  Upload Bank Statement PDFs")

        btn_row = tk.Frame(file_card, bg=C_WHITE)
        btn_row.pack(fill="x", padx=14, pady=(0, 10))

        _btn(btn_row, "  Browse PDFs  ", self._browse_files, C_TEAL,
             font=("Segoe UI", 11, "bold"), padx=16, pady=8).pack(side="left", padx=(0, 10))
        _btn(btn_row, "  ✕  Clear  ", self._clear_files, "#94a3b8",
             font=("Segoe UI", 11), padx=12, pady=8).pack(side="left")

        self._file_count_lbl = tk.Label(btn_row, text="No files selected",
                                         bg=C_WHITE, fg=C_SUBTEXT,
                                         font=("Segoe UI", 10))
        self._file_count_lbl.pack(side="left", padx=14)

        # Listbox with border frame + scrollbar
        lb_wrap = tk.Frame(file_card, bg=C_BORDER, padx=1, pady=1)
        lb_wrap.pack(fill="x", padx=14, pady=(0, 14))
        lb_inner = tk.Frame(lb_wrap, bg=C_WHITE)
        lb_inner.pack(fill="x")

        lb_sb = ttk.Scrollbar(lb_inner, orient="vertical")
        lb_sb.pack(side="right", fill="y")

        self._file_listbox = tk.Listbox(
            lb_inner, height=5, selectmode="extended",
            bg=C_WHITE, fg=C_TEXT, font=("Segoe UI", 10),
            relief="flat", bd=0, highlightthickness=0,
            selectbackground=C_TEAL, selectforeground="#ffffff",
            activestyle="none", yscrollcommand=lb_sb.set)
        self._file_listbox.pack(side="left", fill="x", expand=True, padx=10, pady=6)
        lb_sb.config(command=self._file_listbox.yview)

        # ── Row 1: Process + Bank Selector ──────────────────────────
        ctrl_card = _card(main, row=1)
        _section_label(ctrl_card, "▶  Convert & Export")

        ctrl_inner = tk.Frame(ctrl_card, bg=C_WHITE)
        ctrl_inner.pack(fill="x", padx=14, pady=(0, 10))
        ctrl_inner.columnconfigure(1, weight=1)

        self._process_btn = _btn(
            ctrl_inner, "  Convert PDFs → Excel  ",
            self._start_processing, C_NAVY,
            font=("Segoe UI", 13, "bold"), padx=22, pady=12)
        self._process_btn.grid(row=0, column=0, padx=(0, 18), rowspan=2, sticky="ns")

        self._status_lbl = tk.Label(
            ctrl_inner, text="Ready — select PDFs above and click Convert",
            bg=C_WHITE, fg=C_SUBTEXT, font=("Segoe UI", 10), anchor="w")
        self._status_lbl.grid(row=0, column=1, sticky="ew", pady=(0, 6))

        self._progress = ttk.Progressbar(ctrl_inner, mode="determinate")
        self._progress.grid(row=1, column=1, sticky="ew")

        # Bank selector row
        bank_row = tk.Frame(ctrl_card, bg=C_WHITE)
        bank_row.pack(fill="x", padx=14, pady=(4, 14))

        tk.Label(bank_row, text="Bank Layout Override:",
                 bg=C_WHITE, fg=C_SUBTEXT,
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=(0, 8))

        self._bank_var = tk.StringVar(value="Auto-Detect")
        bank_opts = ["Auto-Detect"] + [v[0] for v in BANK_REGISTRY.values()] + ["Generic / Not in list"]
        self._bank_combo = ttk.Combobox(bank_row, textvariable=self._bank_var,
                                        values=bank_opts, state="readonly", width=28)
        self._bank_combo.pack(side="left")

        tk.Frame(bank_row, bg=C_BORDER, width=1).pack(side="left", fill="y", padx=16)

        _btn(bank_row, "💾  Save Excel", self._save_excel, C_NAVY,
             font=("Segoe UI", 10, "bold"), padx=14, pady=6).pack(side="left", padx=(0, 8))
        _btn(bank_row, "📊  Summary", self._show_summary, "#475569",
             font=("Segoe UI", 10), padx=12, pady=6).pack(side="left")

        # ── Row 2: Stats ─────────────────────────────────────────────
        self._stats_frame = tk.Frame(main, bg=C_BG)
        self._stats_frame.grid(row=2, column=0, sticky="ew", pady=(0, 8))

        # ── Row 3: Results Notebook ───────────────────────────────────
        result_card = _card(main, row=3, expand=True)
        _section_label(result_card, "📊  Extracted Transactions")

        self._notebook = ttk.Notebook(result_card)
        self._notebook.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    # ── Event handlers ───────────────────────────────────────────────

    def _on_ocr_toggle(self):
        state = "readonly" if self._ocr_var.get() else "disabled"
        self._ocr_combo.configure(state=state)

    def _browse_files(self):
        files = filedialog.askopenfilenames(
            title="Select Bank Statement PDFs",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        for f in files:
            if f not in self._selected_files:
                self._selected_files.append(f)
                self._file_listbox.insert("end", Path(f).name)
        self._file_count_lbl.config(
            text=f"{len(self._selected_files)} file(s) selected" if self._selected_files else "No files selected"
        )

    def _clear_files(self):
        self._selected_files.clear()
        self._file_listbox.delete(0, "end")
        self._file_count_lbl.config(text="No files selected")

    def _start_processing(self):
        if not self._selected_files:
            messagebox.showwarning("No Files", "Please select at least one PDF file.")
            return
        self._process_btn.config(state="disabled")
        self._results.clear()
        self._infos.clear()
        self._clear_notebook()
        self._progress["value"] = 0
        self._status_lbl.config(text="Starting...", fg=C_SUBTEXT)
        thread = threading.Thread(target=self._run_processing, daemon=True)
        thread.start()

    def _run_processing(self):
        files = list(self._selected_files)
        n = len(files)
        use_ocr = self._ocr_var.get()
        ocr_engine = self._ocr_engine_var.get()
        selected_bank = self._bank_var.get()

        with tempfile.TemporaryDirectory() as tmpdir:
            for i, filepath in enumerate(files):
                filename = Path(filepath).name
                self.after(0, lambda fn=filename, pct=int(i/n*100): self._update_progress(fn, pct))
                try:
                    import shutil
                    tmp_path = os.path.join(tmpdir, filename)
                    shutil.copy2(filepath, tmp_path)

                    def make_page_cb(fn, file_idx, n_files):
                        def cb(page_num, total_pages):
                            pct = int((file_idx + page_num / total_pages) / n_files * 100)
                            self.after(0, lambda p=pct, f=fn, pg=page_num, tp=total_pages:
                                       self._update_progress(f, p, pg, tp))
                        return cb

                    df, info = process_pdf(tmp_path, filename, use_ocr, ocr_engine,
                                           progress_cb=make_page_cb(filename, i, n),
                                           selected_bank=selected_bank)
                except Exception:
                    df, info = None, {"filename": filename, "bank": "Unknown",
                                       "pdf_type": "unknown", "rows_extracted": 0,
                                       "error": traceback.format_exc(), "warnings": []}
                self._infos.append(info)
                if df is not None:
                    sheet_name = Path(filename).stem
                    self._results[sheet_name] = df

        self.after(0, self._on_processing_done)

    def _update_progress(self, filename: str, pct: int,
                         page: int = 0, total_pages: int = 0):
        self._progress["value"] = pct
        if page and total_pages:
            self._status_lbl.config(
                text=f"Processing: {filename}  (page {page}/{total_pages})",
                fg=C_SUBTEXT)
        else:
            self._status_lbl.config(text=f"Processing: {filename}", fg=C_SUBTEXT)

    def _on_processing_done(self):
        self._progress["value"] = 100
        success = len(self._results)
        fail = len(self._infos) - success
        msg = f"Done — {success} succeeded"
        if fail:
            msg += f", {fail} failed"
        self._status_lbl.config(
            text=msg,
            fg=C_CREDIT if not fail else C_DEBIT
        )
        self._process_btn.config(state="normal")
        self._render_stats()
        self._render_tabs()

    def _render_stats(self):
        for w in self._stats_frame.winfo_children():
            w.destroy()
        if not self._results:
            return

        def safe_sum(col):
            total = 0.0
            for df in self._results.values():
                if col in df.columns:
                    for val in df[col]:
                        try:
                            s = re.sub(r"[,\s]", "", str(val))
                            if s:
                                total += float(s)
                        except Exception:
                            pass
            return total

        total_txns = sum(len(df) for df in self._results.values())
        total_d = safe_sum("Debit")
        total_c = safe_sum("Credit")
        cards = [
            ("📁", "Files Processed",    str(len(self._results)),    C_TEAL),
            ("📋", "Total Transactions", f"{total_txns:,}",          C_NAVY),
            ("📤", "Total Debit",        f"₹ {total_d:,.2f}",        C_DEBIT),
            ("📥", "Total Credit",       f"₹ {total_c:,.2f}",        C_CREDIT),
        ]
        for i, (icon, label, value, color) in enumerate(cards):
            # Border frame acts as shadow/outline
            outer = tk.Frame(self._stats_frame, bg=C_BORDER)
            outer.grid(row=0, column=i, sticky="ew", padx=4, pady=2)
            self._stats_frame.columnconfigure(i, weight=1)

            card = tk.Frame(outer, bg=C_WHITE)
            card.pack(fill="both", expand=True, padx=1, pady=1)

            # Left color accent bar
            tk.Frame(card, bg=color, width=6).pack(side="left", fill="y")

            body = tk.Frame(card, bg=C_WHITE)
            body.pack(side="left", fill="both", expand=True, padx=12, pady=10)

            top_row = tk.Frame(body, bg=C_WHITE)
            top_row.pack(anchor="w")
            tk.Label(top_row, text=icon, font=("Segoe UI Emoji", 14),
                     bg=C_WHITE, fg=color).pack(side="left", padx=(0, 6))
            tk.Label(top_row, text=label, font=("Segoe UI", 10),
                     bg=C_WHITE, fg=C_SUBTEXT).pack(side="left")

            tk.Label(body, text=value, font=("Segoe UI", 18, "bold"),
                     bg=C_WHITE, fg=color).pack(anchor="w", pady=(3, 0))

    def _clear_notebook(self):
        for tab in self._notebook.tabs():
            self._notebook.forget(tab)

    def _render_tabs(self):
        self._clear_notebook()
        preview_rows = self._preview_var.get()

        # File info tab
        info_frame = tk.Frame(self._notebook, bg=C_WHITE)
        self._notebook.add(info_frame, text="Processing Summary")
        self._render_info_table(info_frame)

        # Per-file transaction tabs
        for name, df in self._results.items():
            frame = tk.Frame(self._notebook, bg=C_WHITE)
            short = name[:20] + "…" if len(name) > 20 else name
            self._notebook.add(frame, text=short)
            info_for_tab = next((i for i in self._infos if Path(i["filename"]).stem == name), {})
            self._render_transaction_table(frame, df, preview_rows,
                col_display_names=info_for_tab.get("col_display_names", {}))

    def _render_info_table(self, parent):
        cols = ("File", "Status", "Bank", "Type", "Transactions", "Notes")
        tree = ttk.Treeview(parent, columns=cols, show="headings", height=10)
        col_widths = [200, 90, 160, 90, 110, 250]
        for col, w in zip(cols, col_widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="w")

        for info in self._infos:
            status = "Success" if not info.get("error") else "Failed"
            pdf_type = {"text": "Text PDF", "scanned": "Scanned", "unknown": "Unknown"}.get(
                info.get("pdf_type", "unknown"), "Unknown"
            )
            note = info.get("error", "") or ", ".join(info.get("warnings", [])) or "-"
            if info.get("error"):
                note = note[:80] + "..." if len(note) > 80 else note
            tree.insert("", "end", values=(
                info["filename"], status, info.get("bank", "Unknown"),
                pdf_type, info.get("rows_extracted", 0), note
            ), tags=("ok" if not info.get("error") else "err",))

        tree.tag_configure("ok", foreground=C_CREDIT)
        tree.tag_configure("err", foreground=C_DEBIT)

        sb_y = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        sb_x = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True, padx=4, pady=4)

    def _render_transaction_table(self, parent, df: pd.DataFrame, max_rows: int,
                                   col_display_names: dict = None):
        display = col_display_names or {}
        # Build display column names (standard cols get original PDF names; extra cols keep their names)
        std_display = [display.get(c, c) for c in STANDARD_COLS if c in df.columns]
        extra_cols = [c for c in df.columns if c not in STANDARD_COLS]
        display_cols = std_display + extra_cols
        # Internal cols in same order (for reading values)
        internal_cols = [c for c in STANDARD_COLS if c in df.columns] + extra_cols

        preview = df.head(max_rows)

        shown = min(max_rows, len(df))
        lbl_text = (f"Showing all {len(df)} transactions" if shown == len(df)
                    else f"Showing {shown} of {len(df)} transactions  (first {shown} rows)")
        lbl = tk.Label(parent, bg=C_WHITE, fg=C_TEAL,
                        font=("Segoe UI", 10, "bold"), text=lbl_text)
        lbl.pack(anchor="w", padx=12, pady=(6, 4))

        tree = ttk.Treeview(parent, columns=display_cols, show="headings")
        # Role-based widths: Date=90, Desc=300, Debit=100, Credit=100, Balance=110
        role_widths = [90, 300, 100, 100, 110]
        amount_display = {display.get("Debit", "Debit"), display.get("Credit", "Credit"),
                          display.get("Balance", "Balance")}
        for ci, (dcol, icol) in enumerate(zip(display_cols, internal_cols)):
            tree.heading(dcol, text=dcol)
            w = role_widths[ci] if ci < len(role_widths) else 120
            tree.column(dcol, width=w, minwidth=60,
                        anchor="e" if dcol in amount_display else "w")

        for i, (_, row) in enumerate(preview.iterrows()):
            vals = [row.get(ic, "") for ic in internal_cols]
            tag = "alt" if i % 2 == 0 else ""
            if str(row.get("Debit", "")).strip():
                tag = "debit"
            elif str(row.get("Credit", "")).strip():
                tag = "credit"
            tree.insert("", "end", values=vals, tags=(tag,))

        tree.tag_configure("debit",  background=C_RED_BG,   foreground=C_DEBIT)
        tree.tag_configure("credit", background=C_GREEN_BG, foreground=C_CREDIT)
        tree.tag_configure("alt",    background=C_LIGHT)

        sb_y = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        sb_x = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True, padx=4, pady=4)

    # ── Save / Export ────────────────────────────────────────────────

    def _save_excel(self):
        if not self._results:
            messagebox.showwarning("No Data", "Process some PDFs first.")
            return
        col_display_map = {Path(i["filename"]).stem: i.get("col_display_names", {})
                           for i in self._infos}
        metadata_map = {Path(i["filename"]).stem: i.get("account_metadata", {})
                        for i in self._infos}
        # Build a map from sheet name → original PDF folder
        folder_map = {}
        for fp in self._selected_files:
            stem = Path(fp).stem
            folder_map[stem] = str(Path(fp).parent)

        saved_paths = []
        for name, df in self._results.items():
            folder = folder_map.get(name, str(Path(self._selected_files[0]).parent))
            out_path = os.path.join(folder, f"{name}_bank_statement_to_excel.xlsx")
            data = export_to_excel_bytes({name: df}, include_summary=False,
                                         col_display_map=col_display_map,
                                         metadata_map=metadata_map)
            with open(out_path, "wb") as f:
                f.write(data)
            saved_paths.append(out_path)

        if len(saved_paths) == 1:
            messagebox.showinfo("Saved", f"Excel saved to:\n{saved_paths[0]}")
        else:
            messagebox.showinfo("Saved", f"{len(saved_paths)} Excel files saved:\n" +
                                "\n".join(saved_paths))

    def _save_csv(self):
        if not self._results:
            messagebox.showwarning("No Data", "Process some PDFs first.")
            return
        combine = self._combine_var.get()
        if combine:
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                initialfile="bank_statements_combined.csv"
            )
            if not path:
                return
            all_df = pd.concat(
                [df.assign(**{"Source File": name}) for name, df in self._results.items()],
                ignore_index=True
            )
            with open(path, "wb") as f:
                f.write(export_to_csv_bytes(all_df))
            messagebox.showinfo("Saved", f"CSV saved to:\n{path}")
        else:
            folder = filedialog.askdirectory(title="Select folder to save CSV files")
            if not folder:
                return
            for name, df in self._results.items():
                path = os.path.join(folder, f"{name}.csv")
                with open(path, "wb") as f:
                    f.write(export_to_csv_bytes(df))
            messagebox.showinfo("Saved", f"CSV files saved to:\n{folder}")

    def _show_summary(self):
        if not self._infos:
            messagebox.showinfo("Summary", "No processing results yet.")
            return
        win = tk.Toplevel(self)
        win.title("Processing Summary")
        win.geometry("720x480")
        win.configure(bg=C_BG)
        win.resizable(True, True)

        # Header
        tk.Frame(win, bg=C_TEAL, height=3).pack(fill="x")
        hdr = tk.Frame(win, bg=C_NAVY, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📊  Processing Summary",
                 font=("Segoe UI", 14, "bold"),
                 bg=C_NAVY, fg=C_WHITE).pack(side="left", padx=20)
        tk.Label(hdr, text=f"{len(self._infos)} file(s)",
                 font=("Segoe UI", 10), bg=C_NAVY, fg="#A8D8EA").pack(side="right", padx=20)
        tk.Frame(win, bg=C_TEAL, height=2).pack(fill="x")

        # Content card
        content = tk.Frame(win, bg=C_BG)
        content.pack(fill="both", expand=True, padx=12, pady=12)

        card = tk.Frame(content, bg=C_WHITE, highlightthickness=1,
                        highlightbackground=C_BORDER)
        card.pack(fill="both", expand=True)

        sb = ttk.Scrollbar(card)
        sb.pack(side="right", fill="y")

        text = tk.Text(card, bg=C_WHITE, fg=C_TEXT,
                       font=("Consolas", 10),
                       relief="flat", padx=16, pady=12,
                       yscrollcommand=sb.set, borderwidth=0)
        sb.config(command=text.yview)
        text.pack(fill="both", expand=True)

        for info in self._infos:
            status = "✓  Success" if not info.get("error") else "✗  Failed"
            text.insert("end", f"File   : {info['filename']}\n")
            text.insert("end", f"Status : {status}\n")
            text.insert("end", f"Bank   : {info['bank']}\n")
            text.insert("end", f"Type   : {info['pdf_type']}\n")
            text.insert("end", f"Rows   : {info['rows_extracted']}\n")
            if info.get("error"):
                text.insert("end", f"Error  : {info['error']}\n")
            if info.get("warnings"):
                for w in info["warnings"]:
                    text.insert("end", f"Warn   : {w}\n")
            text.insert("end", "─" * 64 + "\n")
        text.configure(state="disabled")

        # Footer
        foot = tk.Frame(win, bg=C_BG, pady=10)
        foot.pack(fill="x")
        tk.Button(foot, text="  Close  ", command=win.destroy,
                  bg=C_NAVY, fg="#ffffff", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=18, pady=7, cursor="hand2",
                  borderwidth=0).pack()


# ═══════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════

def _sep(parent, color="#BDD7EE"):
    tk.Frame(parent, bg=color, height=1).pack(fill="x", padx=8, pady=4)


def _card(parent, row: int, expand: bool = False):
    # Border frame simulates a subtle shadow/outline
    border = tk.Frame(parent, bg=C_BORDER)
    border.grid(row=row, column=0, sticky="nsew" if expand else "ew",
                pady=(0, 8), padx=1)
    if expand:
        parent.rowconfigure(row, weight=1)
    frame = tk.Frame(border, bg=C_WHITE)
    frame.pack(fill="both", expand=True, padx=1, pady=1)
    return frame


def _section_label(parent, text: str):
    hdr = tk.Frame(parent, bg=C_WHITE)
    hdr.pack(fill="x", padx=12, pady=(12, 0))
    tk.Label(hdr, text=text, bg=C_WHITE, fg=C_NAVY,
             font=("Segoe UI", 12, "bold"),
             anchor="w").pack(side="left")
    tk.Frame(parent, bg=C_TEAL, height=2).pack(fill="x", padx=12, pady=(4, 8))


def _btn(parent, text: str, command, bg: str,
         font=("Segoe UI", 11, "bold"), padx=18, pady=8):
    b = tk.Button(parent, text=text, command=command,
                  bg=bg, fg="#ffffff",
                  activebackground=C_NAVY, activeforeground="#ffffff",
                  relief="flat", font=font, padx=padx, pady=pady,
                  cursor="hand2", borderwidth=0)
    # Subtle hover effect
    _orig = bg
    b.bind("<Enter>", lambda e: b.config(bg=C_NAVY if _orig != C_NAVY else C_TEAL))
    b.bind("<Leave>", lambda e: b.config(bg=_orig))
    return b


# ═══════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = App()
    app.mainloop()
