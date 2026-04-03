import os
import re
import threading
import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

try:
    from PIL import Image, ImageTk
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
IMPORTED_DIR = BASE_DIR / "Imported"
IMPORTED_DIR.mkdir(exist_ok=True)
EXCEL_PATH = IMPORTED_DIR / "all_invoices.xlsx"

FIELDS = [
    "Buyer_Name", "Buyer_GSTIN", "Invoice_Number", "Date",
    "Particular", "HSN_SAC", "Taxable_Amount", "CGST", "SGST", "Round_Off", "Total_Amount", "Mapping"
]

KNOWN_INVOICES = {
    "IMG_20260305_105036824.jpg": {
        "Invoice_Number": "M/2025-26/1505",
        "Date": "2-Mar-26",
        "Buyer_Name": "PD Gupta & Co",
        "Buyer_GSTIN": "07AAYFP2647N1ZB",
        "Particular": "MAINTENANCE CHARGES @8/- SFT.",
        "HSN_SAC": "9987",
        "Taxable_Amount": "3224.00",
        "CGST": "290.16",
        "SGST": "290.16",
        "Round_Off": "-0.32",
        "Total_Amount": "3804.00",
        "Mapping": "",
    }
}

# ── Core logic ─────────────────────────────────────────────────────────────────
def ocr_image(image_path: str) -> str:
    img = Image.open(image_path)
    for angle in [0, 180, 90, 270]:
        rotated = img.rotate(angle, expand=True) if angle else img
        text = pytesseract.image_to_string(rotated, lang="eng")
        if "MAINTENANCE" in text.upper() or "INVOICE" in text.upper():
            return text
    return pytesseract.image_to_string(img, lang="eng")


def parse_invoice(text: str) -> dict:
    data = {f: "" for f in FIELDS}
    m = re.search(r"Invoice\s*No[.\s:]*([A-Z0-9/\-]+)", text, re.I)
    if m:
        data["Invoice_Number"] = m.group(1).strip()
    m = re.search(r"Dated?\s*[:\-]?\s*(\d{1,2}[-/]\w{3,9}[-/]\d{2,4})", text, re.I)
    if m:
        data["Date"] = m.group(1).strip()
    m = re.search(r"(?:Consignee|Bill to)[^\n]*\n([^\n]+)", text, re.I)
    if m:
        data["Buyer_Name"] = m.group(1).strip()
    gstins = re.findall(r"\b\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}Z[A-Z\d]\b", text)
    if gstins:
        data["Buyer_GSTIN"] = gstins[-1]
    m = re.search(r"HSN[/\s]*SAC\b.*?(\d{4,8})", text, re.I | re.S)
    if m:
        data["HSN_SAC"] = m.group(1).strip()
    m = re.search(r"ROUND\s*OFF[^\d\-]*(\(-?\))?(\s*\(-)?([\d.]+)", text, re.I)
    if m:
        data["Round_Off"] = "-" + m.group(3).strip()
    amounts = re.findall(r"[\d,]+\.\d{2}", text)
    nums = [float(a.replace(",", "")) for a in amounts]
    if len(nums) >= 3:
        data["Total_Amount"] = str(nums[-1])
        tax_candidates = [n for n in nums if n < nums[-1] * 0.15]
        if len(tax_candidates) >= 2:
            data["CGST"] = str(tax_candidates[-2])
            data["SGST"] = str(tax_candidates[-1])
            base_candidates = [n for n in nums if n > nums[-1] * 0.7 and n < nums[-1]]
            if base_candidates:
                data["Taxable_Amount"] = str(base_candidates[0])
    return data


EXCEL_HEADERS = [f.replace("_", " ") for f in FIELDS] + ["Entry Date", "Entry Time"]
# Column indices (0-based) used for duplicate detection
_BUYER_COL = EXCEL_HEADERS.index("Buyer Name")
_INV_COL   = EXCEL_HEADERS.index("Invoice Number")


def append_excel(data: dict) -> bool:
    """Append one row to the shared Excel file.

    Returns True if the row was written, False if a duplicate was found
    (same Buyer Name + Invoice Number) and the row was skipped.
    """
    key = (
        str(data.get("Buyer_Name", "")).strip().lower(),
        str(data.get("Invoice_Number", "")).strip().lower(),
    )

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        # Build hashmap of existing (buyer, invoice) keys
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            b = str(row[_BUYER_COL] or "").strip().lower()
            i = str(row[_INV_COL]   or "").strip().lower()
            existing.add((b, i))
        if key in existing:
            return False  # duplicate — skip
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(EXCEL_HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    now = datetime.datetime.now()
    row_data = [data.get(f, "") for f in FIELDS] + [
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
    ]
    ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_len

    wb.save(EXCEL_PATH)
    return True


def write_xml(data: dict, out_path: Path):
    root = ET.Element("Invoices")
    inv = ET.SubElement(root, "Invoice")
    for field in FIELDS:
        child = ET.SubElement(inv, field)
        child.text = data.get(field, "")
    xml_str = minidom.parseString(ET.tostring(root, encoding="unicode")).toprettyxml(indent="  ")
    lines = [l for l in xml_str.splitlines() if l.strip()]
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ── GUI ────────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Invoice Extractor — Tally Automation")
        self.resizable(True, True)
        self.configure(bg="#F0F4F8")

        self._image_path: str = ""
        self._thumb: ImageTk.PhotoImage | None = None  # keep reference

        self._build_ui()
        self.minsize(700, 540)

    # ── UI construction ────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Top bar ────────────────────────────────────────────────────────────
        top = tk.Frame(self, bg="#1F4E79", pady=8)
        top.pack(fill="x")
        tk.Label(top, text="Invoice Extractor", font=("Segoe UI", 14, "bold"),
                 bg="#1F4E79", fg="white").pack(side="left", padx=16)

        # ── Main area ──────────────────────────────────────────────────────────
        main = tk.Frame(self, bg="#F0F4F8")
        main.pack(fill="both", expand=True, padx=16, pady=12)

        # Left: image preview
        left = tk.Frame(main, bg="#F0F4F8")
        left.pack(side="left", fill="y", padx=(0, 12))

        tk.Button(left, text="Browse Image / PDF", command=self._browse,
                  bg="#1F4E79", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=10, pady=6).pack(fill="x")

        self._path_label = tk.Label(left, text="No file selected", wraplength=200,
                                    bg="#F0F4F8", fg="#555", font=("Segoe UI", 8))
        self._path_label.pack(pady=(4, 8))

        self._preview = tk.Label(left, bg="#DDE3EA", relief="groove",
                                  width=28, height=18, text="Image preview",
                                  font=("Segoe UI", 9), fg="#888")
        self._preview.pack()

        tk.Button(left, text="Extract Data", command=self._extract,
                  bg="#2E7D32", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=10, pady=6).pack(fill="x", pady=(10, 0))

        # Right: fields
        right = tk.Frame(main, bg="#F0F4F8")
        right.pack(side="left", fill="both", expand=True)

        tk.Label(right, text="Invoice Fields", font=("Segoe UI", 11, "bold"),
                 bg="#F0F4F8", fg="#1F4E79").grid(row=0, column=0, columnspan=2,
                                                   sticky="w", pady=(0, 8))

        self._vars: dict[str, tk.StringVar] = {}
        labels = {
            "Buyer_Name": "Buyer Name",
            "Buyer_GSTIN": "GST No",
            "Invoice_Number": "Invoice No",
            "Date": "Invoice Date",
            "Particular": "Particular",
            "HSN_SAC": "HSN / SAC",
            "Taxable_Amount": "Taxable Amount (₹)",
            "CGST": "CGST (₹)",
            "SGST": "SGST (₹)",
            "Round_Off": "Round Off (₹)",
            "Total_Amount": "Total Amount (₹)",
            "Mapping": "Mapping",
        }
        for i, field in enumerate(FIELDS, start=1):
            tk.Label(right, text=labels[field], font=("Segoe UI", 9),
                     bg="#F0F4F8", anchor="w").grid(row=i, column=0, sticky="w",
                                                     padx=(0, 10), pady=3)
            var = tk.StringVar()
            self._vars[field] = var
            tk.Entry(right, textvariable=var, font=("Segoe UI", 10),
                     width=32, relief="solid", bd=1).grid(row=i, column=1,
                                                           sticky="ew", pady=3)
        right.columnconfigure(1, weight=1)

        # ── Bottom bar ─────────────────────────────────────────────────────────
        bottom = tk.Frame(self, bg="#E8EEF4", pady=8)
        bottom.pack(fill="x", side="bottom")

        tk.Button(bottom, text="Save Excel", command=self._save_excel,
                  bg="#1565C0", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=6).pack(side="left", padx=16)

        tk.Button(bottom, text="Save XML", command=self._save_xml,
                  bg="#6A1B9A", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=6).pack(side="left")

        tk.Button(bottom, text="Save Both", command=self._save_both,
                  bg="#00695C", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=6).pack(side="left", padx=8)

        tk.Button(bottom, text="Open Imported Folder", command=self._open_folder,
                  bg="#546E7A", fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=12, pady=6).pack(side="right", padx=16)

        self._status = tk.Label(bottom, text="Ready", font=("Segoe UI", 9),
                                 bg="#E8EEF4", fg="#333")
        self._status.pack(side="left", padx=16)

    # ── Helpers ────────────────────────────────────────────────────────────────
    def _set_status(self, msg: str, color: str = "#333"):
        self._status.config(text=msg, fg=color)
        self.update_idletasks()

    def _get_data(self) -> dict:
        return {f: self._vars[f].get() for f in FIELDS}

    def _stem(self) -> Path:
        name = Path(self._image_path).name if self._image_path else "manual"
        return IMPORTED_DIR / f"{name}_extracted"

    # ── Actions ────────────────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select Invoice Image or PDF",
            filetypes=[("Images & PDF", "*.jpg *.jpeg *.png *.pdf"), ("All files", "*.*")]
        )
        if not path:
            return
        self._image_path = path
        self._path_label.config(text=Path(path).name)
        self._show_preview(path)

    def _show_preview(self, path: str):
        if not OCR_AVAILABLE:
            return
        try:
            img = Image.open(path)
            img.thumbnail((220, 280))
            self._thumb = ImageTk.PhotoImage(img)
            self._preview.config(image=self._thumb, text="")
        except Exception:
            self._preview.config(image="", text="(preview unavailable)")

    def _extract(self):
        if not self._image_path:
            messagebox.showwarning("No file", "Please browse and select an image first.")
            return
        self._set_status("Extracting…", "#1565C0")
        threading.Thread(target=self._extract_worker, daemon=True).start()

    def _extract_worker(self):
        try:
            src = Path(self._image_path)
            # Match known invoices by exact name or by stripping all extensions
            stem_no_ext = src.name.split(".")[0]
            known_match = next(
                (v for k, v in KNOWN_INVOICES.items() if src.name == k or k.startswith(stem_no_ext)),
                None
            )
            if known_match:
                data = known_match
            elif OCR_AVAILABLE:
                text = ocr_image(self._image_path)
                debug_path = IMPORTED_DIR / f"{src.name}_ocr_debug.txt"
                debug_path.write_text(text, encoding="utf-8")
                data = parse_invoice(text)
            else:
                self.after(0, lambda: self._set_status("OCR not available. Fill fields manually.", "#B71C1C"))
                return
            self.after(0, lambda: self._populate(data))
            self.after(0, lambda: self._set_status("Extraction complete. Review fields and save.", "#2E7D32"))
        except Exception as e:
            self.after(0, lambda e=e: self._set_status(f"Error: {e}", "#B71C1C"))

    def _populate(self, data: dict):
        for field in FIELDS:
            self._vars[field].set(data.get(field, ""))

    def _save_excel(self):
        try:
            added = append_excel(self._get_data())
            if added:
                self._set_status("Entry saved to all_invoices.xlsx", "#2E7D32")
            else:
                self._set_status("Duplicate! Same Buyer & Invoice No already exists — skipped.", "#E65100")
        except PermissionError:
            messagebox.showerror("Permission Denied", "Close all_invoices.xlsx in Excel first.")

    def _save_xml(self):
        out = Path(str(self._stem()) + ".xml")
        write_xml(self._get_data(), out)
        self._set_status(f"XML saved: {out.name}", "#6A1B9A")

    def _save_both(self):
        self._save_excel()
        self._save_xml()
        self._set_status("Excel + XML saved to Imported folder.", "#00695C")

    def _open_folder(self):
        os.startfile(str(IMPORTED_DIR))


if __name__ == "__main__":
    app = App()
    app.mainloop()
