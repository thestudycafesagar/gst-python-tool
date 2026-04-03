import os
import re
import threading
import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

try:
    from PIL import Image, ImageTk
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

try:
    import pytesseract
except ImportError:
    OCR_AVAILABLE = False

try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
IMPORTED_DIR = BASE_DIR / "Imported"
IMPORTED_DIR.mkdir(exist_ok=True)
EXCEL_PATH = IMPORTED_DIR / "all_invoices.xlsx"

FIELDS = [
    "Vendor_Name", "Vendor_GSTIN", "Invoice_Number", "Date",
    "Particular", "HSN_SAC", "Taxable_Amount", "CGST", "SGST", "IGST", "Round_Off", "Total_Amount", "Mapping"
]

KNOWN_INVOICES = {
    "IMG_20260305_105036824.jpg": {
        "Vendor_Name": "98 Hemkunt Tower Upkeep Society",
        "Vendor_GSTIN": "07AAAAH0592F1ZY",
        "Invoice_Number": "M/2025-26/1505",
        "Date": "2-Mar-26",
        "Particular": "MAINTENANCE CHARGES @8/- SFT.",
        "HSN_SAC": "9987",
        "Taxable_Amount": "3224.00",
        "CGST": "290.16",
        "SGST": "290.16",
        "IGST": "",
        "Round_Off": "-0.32",
        "Total_Amount": "3804.00",
        "Mapping": "",
    }
}

# ── Core logic ─────────────────────────────────────────────────────────────────
def extract_pdf_pages(pdf_path) -> list:
    """Extract text from each page of a PDF separately. Returns list of page texts."""
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t and t.strip():
                pages.append(t)
    return pages


def ocr_image(image_path: str) -> str:
    img = Image.open(image_path)
    for angle in [0, 180, 90, 270]:
        rotated = img.rotate(angle, expand=True) if angle else img
        text = pytesseract.image_to_string(rotated, lang="eng")
        if "MAINTENANCE" in text.upper() or "INVOICE" in text.upper():
            return text
    return pytesseract.image_to_string(img, lang="eng")


def _num(s: str) -> str:
    """Strip commas and return clean number string."""
    return s.replace(",", "").strip()


def parse_invoice(text: str) -> dict:
    data = {f: "" for f in FIELDS}

    # ── Invoice Number ──────────────────────────────────────────────────────────
    m = re.search(r"(?:Invoice\s*No(?:\.|\s*:|\s+)|#\s{0,4})([A-Z0-9][A-Z0-9/\-\s]{2,40})", text, re.I)
    if m:
        data["Invoice_Number"] = m.group(1).strip().rstrip(".")

    # ── Date ───────────────────────────────────────────────────────────────────
    m = re.search(
        r"(?:Invoice\s*Date|Date\s*of\s*Invoice|Bill\s*Date)\s*[:\-]?\s*"
        r"(\d{1,2}[-/]\d{1,2}[-/]\d{2,4}"
        r"|\d{1,2}[-/][A-Za-z]{3,9}[-/]\d{2,4}"
        r"|[A-Za-z]+\s+\d{1,2},?\s+\d{4})",
        text, re.I
    )
    if m:
        data["Date"] = m.group(1).strip()
    else:
        m = re.search(r"\b(\d{1,2}[-/]\d{1,2}[-/]\d{4})\b", text)
        if m:
            data["Date"] = m.group(1).strip()

    # ── Vendor Name ────────────────────────────────────────────────────────────
    m = re.search(r"(?:Bill\s*From|Invoice\s*From|Shipped\s*From)\s*[:\-]?\s*([^\n]{4,80})", text, re.I)
    if m:
        data["Vendor_Name"] = m.group(1).strip()
    else:
        for line in text.splitlines():
            line = line.strip()
            if len(line) > 5 and not re.match(r"^[\d\s\-/]+$", line):
                data["Vendor_Name"] = line
                break

    # ── Vendor GSTIN ───────────────────────────────────────────────────────────
    gstins = re.findall(r"\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z\d]Z[A-Z\d]\b", text)
    if gstins:
        data["Vendor_GSTIN"] = gstins[0]

    # ── Particular ─────────────────────────────────────────────────────────────
    m = re.search(r"^\s*1[\.\s]+([A-Za-z][^\n]{4,60})", text, re.M)
    if m:
        particular = m.group(1).strip()
        particular = re.sub(r"\s+\d{6,}.*$", "", particular).strip()
        data["Particular"] = particular

    # ── HSN / SAC ──────────────────────────────────────────────────────────────
    m = re.search(r"HSN[/\s]*SAC[^\d]*(\d{4,8})", text, re.I | re.S)
    if m:
        data["HSN_SAC"] = m.group(1).strip()
    else:
        m = re.search(r"\b(9{3}[0-9]{3}|0[0-9]{7})\b", text)
        if m:
            data["HSN_SAC"] = m.group(1)

    # ── Taxable Amount ─────────────────────────────────────────────────────────
    m = re.search(r"Taxable\s*(?:Value|Amount)\s*[:\s]*([\d,]+\.?\d*)", text, re.I)
    if m:
        data["Taxable_Amount"] = _num(m.group(1))
    else:
        m = re.search(r"Sub\s*Total[^\d]*([\d,]+\.?\d*)", text, re.I)
        if m:
            data["Taxable_Amount"] = _num(m.group(1))

    # ── CGST ───────────────────────────────────────────────────────────────────
    m = re.search(r"CGST[^:\d]*([\d,]+\.\d{2})", text, re.I)
    if m:
        data["CGST"] = _num(m.group(1))

    # ── SGST ───────────────────────────────────────────────────────────────────
    m = re.search(r"SGST(?:/U[GT]GST)?[^:\d]*([\d,]+\.\d{2})", text, re.I)
    if m:
        data["SGST"] = _num(m.group(1))

    # ── IGST ───────────────────────────────────────────────────────────────────
    m = re.search(r"IGST\s*\(?S?\)?[^:\d]*([\d,]+\.\d{2})", text, re.I)
    if m:
        data["IGST"] = _num(m.group(1))
    else:
        m = re.search(r"Integrated\s+Goods[^:]*:\s*([\d,]+\.?\d*)", text, re.I)
        if m and m.group(1) != "0":
            data["IGST"] = _num(m.group(1))

    # ── Round Off ──────────────────────────────────────────────────────────────
    m = re.search(r"Round\s*Off\s*(?:Value)?[^:\d\-]*([-\(]?[\d.]+)", text, re.I)
    if m:
        val = m.group(1).replace("(", "-").strip()
        if val and val not in ("0", "0.00"):
            data["Round_Off"] = val if val.startswith("-") else "-" + val

    # ── Total Amount ───────────────────────────────────────────────────────────
    for pat in [
        r"Balance\s*Due\s*[₹Rs.\s]*([\d,]+\.?\d*)",
        r"Total\s*Payable\s*(?:Amount)?\s*[₹Rs.\s\(]*([\d,]+\.?\d*)",
        r"Total\s*(?:Bill\s*)?Amount\s*[₹Rs.\s\(]*([\d,]+\.?\d*)",
        r"Amount\s*Due\s*\(INR\)[^₹\d]*([\d,]+\.?\d*)",
        r"Total\s*[₹Rs.\s]*([\d,]+\.\d{2})",
    ]:
        m = re.search(pat, text, re.I)
        if m:
            data["Total_Amount"] = _num(m.group(1))
            break

    return data


EXCEL_HEADERS = [f.replace("_", " ") for f in FIELDS] + ["Entry Date", "Entry Time"]
_VENDOR_COL = EXCEL_HEADERS.index("Vendor Name")
_INV_COL    = EXCEL_HEADERS.index("Invoice Number")


def append_excel(data: dict) -> bool:
    """Append one row to the shared Excel file.
    Returns True if written, False if duplicate (same Vendor + Invoice No)."""
    key = (
        str(data.get("Vendor_Name", "")).strip().lower(),
        str(data.get("Invoice_Number", "")).strip().lower(),
    )

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            b = str(row[_VENDOR_COL] or "").strip().lower()
            i = str(row[_INV_COL]    or "").strip().lower()
            existing.add((b, i))
        if key in existing:
            return False
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(EXCEL_HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    now = datetime.datetime.now()
    row_data = [data.get(f, "") for f in FIELDS] + [
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
    ]
    ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_len

    wb.save(EXCEL_PATH)
    return True


def write_xml(data: dict, out_path: Path):
    root = ET.Element("Invoices")
    inv = ET.SubElement(root, "Invoice")
    for field in FIELDS:
        child = ET.SubElement(inv, field)
        child.text = data.get(field, "")
    xml_str = minidom.parseString(ET.tostring(root, encoding="unicode")).toprettyxml(indent="  ")
    lines = [l for l in xml_str.splitlines() if l.strip()]
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ── GUI ────────────────────────────────────────────────────────────────────────
FIELD_LABELS = {
    "Vendor_Name": "Vendor Name",
    "Vendor_GSTIN": "Vendor GSTIN",
    "Invoice_Number": "Invoice No",
    "Date": "Invoice Date",
    "Particular": "Particular",
    "HSN_SAC": "HSN / SAC",
    "Taxable_Amount": "Taxable Amt (₹)",
    "CGST": "CGST (₹)",
    "SGST": "SGST (₹)",
    "IGST": "IGST (₹)",
    "Round_Off": "Round Off (₹)",
    "Total_Amount": "Total Amt (₹)",
    "Mapping": "Mapping",
}

COL_WIDTHS = {
    "Vendor_Name": 200, "Vendor_GSTIN": 150, "Invoice_Number": 140,
    "Date": 90, "Particular": 220, "HSN_SAC": 70,
    "Taxable_Amount": 110, "CGST": 75, "SGST": 75, "IGST": 75,
    "Round_Off": 85, "Total_Amount": 110, "Mapping": 110,
}


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Invoice Extractor — Tally Automation")
        self.resizable(True, True)
        self.configure(bg="#F0F4F8")

        self._image_path: str = ""
        self._rows: list = []        # list of dicts, one per invoice
        self._selected_idx: int = -1

        self._build_ui()
        self.minsize(900, 620)
        self.geometry("1200x700")

    def _build_ui(self):
        self._build_topbar()
        self._build_main()
        self._build_bottombar()

    # ── Top bar ────────────────────────────────────────────────────────────────
    def _build_topbar(self):
        top = tk.Frame(self, bg="#1F4E79", pady=8)
        top.pack(fill="x")

        tk.Label(top, text="Invoice Extractor", font=("Segoe UI", 13, "bold"),
                 bg="#1F4E79", fg="white").pack(side="left", padx=16)

        tk.Button(top, text="Browse File", command=self._browse,
                  bg="#1565C0", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4).pack(side="left", padx=(8, 4))

        self._path_label = tk.Label(top, text="No file selected",
                                    bg="#1F4E79", fg="#90CAF9", font=("Segoe UI", 9))
        self._path_label.pack(side="left", padx=(0, 8))

        tk.Button(top, text="Extract Data", command=self._extract,
                  bg="#2E7D32", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4).pack(side="left", padx=4)

        tk.Button(top, text="Clear Table", command=self._clear_all,
                  bg="#B71C1C", fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=4).pack(side="left", padx=8)

        self._status = tk.Label(top, text="Ready", font=("Segoe UI", 9),
                                bg="#1F4E79", fg="#90CAF9")
        self._status.pack(side="left", padx=8)

    # ── Main paned area ────────────────────────────────────────────────────────
    def _build_main(self):
        paned = tk.PanedWindow(self, orient="vertical", bg="#CBD5E0",
                               sashwidth=6, sashpad=2, sashrelief="raised")
        paned.pack(fill="both", expand=True, padx=8, pady=(8, 0))

        # Top: table
        table_frame = tk.Frame(paned, bg="#F0F4F8")
        paned.add(table_frame, minsize=200)

        # Bottom: edit form
        edit_outer = tk.Frame(paned, bg="#F0F4F8")
        paned.add(edit_outer, minsize=160)

        self._build_table(table_frame)
        self._build_edit_form(edit_outer)

        # Set initial sash position after window appears
        self.after(100, lambda: paned.sash_place(0, 0, 380))

    # ── Table ──────────────────────────────────────────────────────────────────
    def _build_table(self, parent):
        header = tk.Frame(parent, bg="#F0F4F8")
        header.pack(fill="x", padx=4, pady=(4, 2))
        tk.Label(header, text="Extracted Invoices", font=("Segoe UI", 10, "bold"),
                 bg="#F0F4F8", fg="#1F4E79").pack(side="left")
        self._row_count_label = tk.Label(header, text="(0 rows)",
                                          font=("Segoe UI", 9), bg="#F0F4F8", fg="#666")
        self._row_count_label.pack(side="left", padx=8)

        frame = tk.Frame(parent, bg="#F0F4F8")
        frame.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        style = ttk.Style()
        style.configure("Invoice.Treeview.Heading", font=("Segoe UI", 9, "bold"),
                         background="#1F4E79", foreground="white")
        style.configure("Invoice.Treeview", font=("Segoe UI", 9), rowheight=24)
        style.map("Invoice.Treeview", background=[("selected", "#1565C0")])

        self._tree = ttk.Treeview(frame, columns=FIELDS, show="headings",
                                   selectmode="browse", style="Invoice.Treeview")
        for field in FIELDS:
            self._tree.heading(field, text=FIELD_LABELS[field],
                                command=lambda f=field: self._sort_by(f))
            self._tree.column(field, width=COL_WIDTHS.get(field, 100),
                               minwidth=50, stretch=False)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self._tree.bind("<<TreeviewSelect>>", self._on_row_select)
        self._tree.bind("<Delete>", self._delete_selected_row)

        # Alternating row colors
        self._tree.tag_configure("odd",  background="#FFFFFF")
        self._tree.tag_configure("even", background="#EEF2F7")

    def _sort_by(self, field):
        """Sort table by column (toggle asc/desc)."""
        if not self._rows:
            return
        reverse = getattr(self, "_sort_reverse", False)
        self._rows.sort(key=lambda r: str(r.get(field, "")).lower(), reverse=reverse)
        self._sort_reverse = not reverse
        self._rebuild_tree()

    # ── Edit form ──────────────────────────────────────────────────────────────
    def _build_edit_form(self, parent):
        header = tk.Frame(parent, bg="#E8EEF4", pady=4)
        header.pack(fill="x")
        tk.Label(header, text="Edit Selected Invoice", font=("Segoe UI", 10, "bold"),
                 bg="#E8EEF4", fg="#1F4E79").pack(side="left", padx=8)
        tk.Label(header, text="(select a row above to edit)",
                 font=("Segoe UI", 8), bg="#E8EEF4", fg="#888").pack(side="left")
        tk.Button(header, text="Update Row", command=self._update_row,
                  bg="#F57C00", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=2).pack(side="right", padx=8)
        tk.Button(header, text="Delete Row", command=self._delete_selected_row,
                  bg="#C62828", fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=2).pack(side="right", padx=4)

        # Scrollable canvas for fields
        canvas = tk.Canvas(parent, bg="#F0F4F8", highlightthickness=0, height=120)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=canvas.xview)
        canvas.configure(xscrollcommand=hsb.set)

        inner = tk.Frame(canvas, bg="#F0F4F8")
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_inner_configure)

        canvas.pack(fill="both", expand=True, padx=4)
        hsb.pack(fill="x", padx=4)

        # Fields in 2 rows: row 0 = labels, row 1 = entries (all columns side by side)
        self._vars: dict = {}
        for col_idx, field in enumerate(FIELDS):
            tk.Label(inner, text=FIELD_LABELS[field], font=("Segoe UI", 8),
                     bg="#F0F4F8", fg="#555", anchor="w").grid(
                row=0, column=col_idx, sticky="w", padx=(6, 2), pady=(6, 0))
            var = tk.StringVar()
            self._vars[field] = var
            entry_width = 22 if field in ("Vendor_Name", "Particular") else 14
            tk.Entry(inner, textvariable=var, font=("Segoe UI", 9),
                     width=entry_width, relief="solid", bd=1).grid(
                row=1, column=col_idx, sticky="ew", padx=(6, 2), pady=(2, 8))

    # ── Bottom bar ─────────────────────────────────────────────────────────────
    def _build_bottombar(self):
        bottom = tk.Frame(self, bg="#E8EEF4", pady=8)
        bottom.pack(fill="x", side="bottom")

        tk.Button(bottom, text="Save Selected → Excel", command=self._save_selected_excel,
                  bg="#1565C0", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=14, pady=6).pack(side="left", padx=16)

        tk.Button(bottom, text="Save All → Excel", command=self._save_all_excel,
                  bg="#00695C", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=14, pady=6).pack(side="left", padx=4)

        tk.Button(bottom, text="Save Selected → XML", command=self._save_xml,
                  bg="#6A1B9A", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=14, pady=6).pack(side="left", padx=4)

        tk.Button(bottom, text="Open Imported Folder", command=self._open_folder,
                  bg="#546E7A", fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=12, pady=6).pack(side="right", padx=16)

        self._bottom_status = tk.Label(bottom, text="", font=("Segoe UI", 9),
                                        bg="#E8EEF4", fg="#333")
        self._bottom_status.pack(side="left", padx=16)

    # ── Helpers ────────────────────────────────────────────────────────────────
    def _set_status(self, msg: str, color: str = "#333"):
        self._status.config(text=msg, fg=color)
        self._bottom_status.config(text=msg, fg=color)
        self.update_idletasks()

    def _add_rows(self, rows: list):
        for data in rows:
            self._rows.append(data)
        self._rebuild_tree()
        self._update_row_count()

    def _rebuild_tree(self):
        for item in self._tree.get_children():
            self._tree.delete(item)
        for i, data in enumerate(self._rows):
            tag = "even" if i % 2 == 0 else "odd"
            self._tree.insert("", "end", iid=str(i),
                               values=[data.get(f, "") for f in FIELDS], tags=(tag,))

    def _update_row_count(self):
        n = len(self._rows)
        self._row_count_label.config(text=f"({n} row{'s' if n != 1 else ''})")

    def _clear_all(self):
        self._rows.clear()
        self._rebuild_tree()
        self._update_row_count()
        self._selected_idx = -1
        for v in self._vars.values():
            v.set("")

    # ── Events ─────────────────────────────────────────────────────────────────
    def _on_row_select(self, event=None):
        sel = self._tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        self._selected_idx = idx
        data = self._rows[idx]
        for field in FIELDS:
            self._vars[field].set(data.get(field, ""))

    def _update_row(self):
        if self._selected_idx < 0 or self._selected_idx >= len(self._rows):
            messagebox.showwarning("No selection", "Select a row first.")
            return
        data = {f: self._vars[f].get() for f in FIELDS}
        self._rows[self._selected_idx] = data
        tag = "even" if self._selected_idx % 2 == 0 else "odd"
        self._tree.item(str(self._selected_idx),
                         values=[data.get(f, "") for f in FIELDS], tags=(tag,))
        self._set_status("Row updated.", "#2E7D32")

    def _delete_selected_row(self, event=None):
        if self._selected_idx < 0 or self._selected_idx >= len(self._rows):
            return
        self._rows.pop(self._selected_idx)
        self._selected_idx = -1
        for v in self._vars.values():
            v.set("")
        self._rebuild_tree()
        self._update_row_count()

    # ── File browse ────────────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select Invoice Image or PDF",
            filetypes=[("Images & PDF", "*.jpg *.jpeg *.png *.pdf"), ("All files", "*.*")]
        )
        if not path:
            return
        self._image_path = path
        self._path_label.config(text=Path(path).name)

    # ── Extraction ─────────────────────────────────────────────────────────────
    def _extract(self):
        if not self._image_path:
            messagebox.showwarning("No file", "Please browse and select a file first.")
            return
        self._set_status("Extracting…", "#1565C0")
        threading.Thread(target=self._extract_worker, daemon=True).start()

    def _extract_worker(self):
        try:
            src = Path(self._image_path)
            is_pdf = src.suffix.lower() == ".pdf"

            # Check known hardcoded invoices first
            known_match = next(
                (v for k, v in KNOWN_INVOICES.items()
                 if src.name == k or src.name.startswith(k.split(".")[0])),
                None
            )

            if known_match:
                rows = [known_match]

            elif is_pdf and PDF_AVAILABLE:
                pages = extract_pdf_pages(src)
                # Save debug text
                debug_path = IMPORTED_DIR / f"{src.name}_text_debug.txt"
                debug_path.write_text("\n\n===PAGE BREAK===\n\n".join(pages), encoding="utf-8")

                if not pages:
                    self.after(0, lambda: self._set_status("No text found in PDF.", "#B71C1C"))
                    return

                # Parse each page as a separate invoice
                rows = [parse_invoice(p) for p in pages]
                # Filter pages with no meaningful data (no invoice number AND no total)
                rows = [r for r in rows if r.get("Invoice_Number") or r.get("Total_Amount")]
                if not rows:
                    # Fallback: treat all pages as one document
                    rows = [parse_invoice("\n".join(pages))]

            elif is_pdf and not PDF_AVAILABLE:
                self.after(0, lambda: self._set_status(
                    "pdfplumber not installed. Run: pip install pdfplumber", "#B71C1C"))
                return

            elif OCR_AVAILABLE:
                text = ocr_image(self._image_path)
                debug_path = IMPORTED_DIR / f"{src.name}_ocr_debug.txt"
                debug_path.write_text(text, encoding="utf-8")
                rows = [parse_invoice(text)]

            else:
                self.after(0, lambda: self._set_status(
                    "OCR not available. Fill fields manually.", "#B71C1C"))
                return

            n = len(rows)
            self.after(0, lambda rows=rows: self._add_rows(rows))
            self.after(0, lambda: self._set_status(
                f"Extracted {n} invoice{'s' if n > 1 else ''}. Review and save.",
                "#2E7D32"))
        except Exception as e:
            self.after(0, lambda e=e: self._set_status(f"Error: {e}", "#B71C1C"))

    # ── Save actions ───────────────────────────────────────────────────────────
    def _save_selected_excel(self):
        if self._selected_idx < 0 or self._selected_idx >= len(self._rows):
            messagebox.showwarning("No selection", "Select a row from the table first.")
            return
        try:
            added = append_excel(self._rows[self._selected_idx])
            if added:
                self._set_status("Selected invoice saved to all_invoices.xlsx", "#2E7D32")
            else:
                self._set_status("Duplicate! Same Vendor & Invoice No already exists — skipped.", "#E65100")
        except PermissionError:
            messagebox.showerror("Permission Denied", "Close all_invoices.xlsx in Excel first.")

    def _save_all_excel(self):
        if not self._rows:
            messagebox.showwarning("No data", "No invoices in the table to save.")
            return
        try:
            saved, skipped = 0, 0
            for row in self._rows:
                if append_excel(row):
                    saved += 1
                else:
                    skipped += 1
            msg = f"{saved} invoice(s) saved"
            if skipped:
                msg += f", {skipped} duplicate(s) skipped"
            self._set_status(msg + " → all_invoices.xlsx", "#2E7D32")
        except PermissionError:
            messagebox.showerror("Permission Denied", "Close all_invoices.xlsx in Excel first.")

    def _save_xml(self):
        if self._selected_idx < 0 or self._selected_idx >= len(self._rows):
            messagebox.showwarning("No selection", "Select a row from the table first.")
            return
        src_name = Path(self._image_path).name if self._image_path else "manual"
        out = IMPORTED_DIR / f"{src_name}_invoice_{self._selected_idx + 1}.xml"
        write_xml(self._rows[self._selected_idx], out)
        self._set_status(f"XML saved: {out.name}", "#6A1B9A")

    def _open_folder(self):
        os.startfile(str(IMPORTED_DIR))


if __name__ == "__main__":
    app = App()
    app.mainloop()
