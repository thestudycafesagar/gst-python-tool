import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import threading
import win32com.client
import pythoncom
import calendar
from datetime import date, datetime, timedelta

try:
    import openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

# ═══════════════════════════════════════════════════════════════
#  TEMPLATE DEFINITIONS 
# ═══════════════════════════════════════════════════════════════

class Template:
    def __init__(self, name, icon, config_fields, recipient_cols,
                 defaults, build_subject, build_body, has_attachment=False):
        self.name            = name
        self.icon            = icon
        self.config_fields   = config_fields    # [(label, key), ...]
        self.recipient_cols  = recipient_cols   # extra cols after Name, Email
        self.defaults        = defaults
        self.build_subject   = build_subject
        self.build_body      = build_body
        self.has_attachment  = has_attachment

    @property
    def all_cols(self):
        return ["Name", "Email"] + self.recipient_cols


# ── Template 1: GST Return ──────────────────────────────────────
def _gst_subject(cfg, row):
    return f'Share "{cfg["return_type"]}" Return Data for the Month of "{cfg["month"]}"'

def _gst_body(cfg, row):
    name = row.get("Name", "Recipient")
    return (
        f"Dear {name} Ji,\n\n"
        f"Please provide {cfg['return_type']} Data for the month of {cfg['month']} to File the Return.\n\n"
        f"Kindly send data upto {cfg['data_deadline']} so that we can also get time to verify the data.\n\n"
        f"Please Note that Last Date to file the Return is {cfg['filing_deadline']} so Request you to "
        f"send the Data Maximum By {cfg['data_deadline']}.\n\n"
        f"In Case of Any Query please feel free to call on +91-{cfg['phone']}.\n\n"
        f"Thanks & Regards\n{cfg['sender']}"
    )

TEMPLATE_GST = Template(
    name="GST Return Data Request",
    icon="📋",
    config_fields=[
        ("Month of Return:",          "month"),
        ("Return Type:",              "return_type"),
        ("Data Submission Deadline:", "data_deadline"),
        ("Last Date of Filing:",      "filing_deadline"),
        ("Sender Name:",              "sender"),
        ("Contact Phone:",            "phone"),
        ("CC Email Address:",         "cc"),
    ],
    recipient_cols=[],
    defaults={
        "month":           "March",
        "return_type":     "GST",
        "data_deadline":   "20 Mar 2026",
        "filing_deadline": "31 Mar 2026",
        "sender":          "Rohit Sharma",
        "phone":           "9876543210",
        "cc":              "cc@yourfirm.com",
    },
    build_subject=_gst_subject,
    build_body=_gst_body,
)


# ── Template 2: Invoice Sender ──────────────────────────────────
def _inv_subject(cfg, row):
    return f"Invoice for {row.get('Service','')} for the Period {row.get('Period','')}"

def _inv_body(cfg, row):
    name   = row.get("Name", "Recipient")
    svc    = row.get("Service", "")
    period = row.get("Period", "")
    amount = row.get("Invoice Amount", "")
    return (
        f"Dear {name},\n\n"
        f"I trust this message finds you well. We appreciate your continued partnership with us "
        f"and we are grateful for the opportunity to provide our services to you.\n\n"
        f"As per our agreement, we are pleased to send you the invoice for the {svc} availed for "
        f"the Period {period}. Please note that Invoice Amount is Rs. {amount}.\n\n"
        f"Please ensure that payment is made by the due date mentioned in the invoice. You can make "
        f"the payment through [Bank/Cheque/UPI etc.], and if you have any questions or require further "
        f"clarification regarding the invoice, please do not hesitate to reach out to our Email id.\n\n"
        f"Your satisfaction is our top priority, and we are committed to delivering exceptional service. "
        f"We look forward to your prompt payment and to continue serving your needs.\n\n"
        f"Thank you for choosing us as your service provider. We value your business and are here to "
        f"assist you with any requirements you may have.\n\n"
        f"Please find attached the invoice for your reference.\n\n"
        f"Should you have any queries or require further assistance, please feel free to contact us "
        f"at +91-{cfg['phone']}.\n\n"
        f"Sincerely,\n{cfg['org_name']}"
    )

TEMPLATE_INVOICE = Template(
    name="Invoice Sender",
    icon="🧾",
    config_fields=[
        ("CC Email Address:", "cc"),
        ("Organisation Name:", "org_name"),
        ("Contact Phone:",     "phone"),
    ],
    recipient_cols=["Invoice Amount", "Period", "Service", "Attachment File"],
    defaults={
        "cc":       "cc@yourfirm.com",
        "org_name": "Your Firm Name",
        "phone":    "9876543210",
    },
    build_subject=_inv_subject,
    build_body=_inv_body,
    has_attachment=True,
)


# ── Template 3: Outstanding Payment ────────────────────────────
def _pay_subject(cfg, row):
    name   = row.get("Name", "")
    period = row.get("Period", "")
    amount = row.get("Outstanding Amount", "")
    return f"{name}, Outstanding Payment Request for Services Utilized till {period} of Rs. {amount}."

def _pay_body(cfg, row):
    name    = row.get("Name", "Recipient")
    period  = row.get("Period", "")
    amount  = row.get("Outstanding Amount", "")
    service = row.get("Service Type", "")
    return (
        f"Dear {name},\n\n"
        f"I hope this email finds you well.\n\n"
        f"I am writing on behalf of {cfg['org_name']} to follow up on an outstanding payment against "
        f"the {service} services rendered to Your Organisation. As of {period}, there remains an unpaid "
        f"balance of Rs. {amount}.\n\n"
        f"We value our relationship with you and have always strived to offer our best services. "
        f"We would appreciate it if the pending amount could be cleared at the earliest to maintain "
        f"the smooth functioning of our business relationship.\n\n"
        f"Kindly verify this against your records, and if there are any discrepancies or clarifications "
        f"required, please let us know within the next {cfg['days_to_respond']}. If you have already "
        f"processed this payment, kindly share the payment details or the remittance advice, so we can "
        f"reconcile it in our system.\n\n"
        f"Please share the payment confirmation with us once completed.\n\n"
        f"We appreciate your prompt attention to this matter and look forward to continuing our collaboration.\n\n"
        f"Please Note that if Payment is not Received by {cfg['pay_deadline']} date then Interest of "
        f"{cfg['interest_rate']} will be levied on the Payment.\n\n"
        f"Therefore, Request you to clear the Payment Maximum By {cfg['pay_deadline']}.\n\n"
        f"In Case of Any Query please feel free to call on +91-{cfg['phone']}.\n\n"
        f"Thank you for your understanding and cooperation.\n\n"
        f"Regards\n{cfg['org_name']}"
    )

TEMPLATE_PAYMENT = Template(
    name="Outstanding Payment Reminder",
    icon="💰",
    config_fields=[
        ("CC Email Address:",    "cc"),
        ("Days to Respond:",     "days_to_respond"),
        ("Payment Deadline:",    "pay_deadline"),
        ("Interest Rate:",       "interest_rate"),
        ("Organisation Name:",   "org_name"),
        ("Contact Phone:",       "phone"),
    ],
    recipient_cols=["Period", "Outstanding Amount", "Service Type"],
    defaults={
        "cc":              "cc@yourfirm.com",
        "days_to_respond": "7 days",
        "pay_deadline":    "31 Mar 2026",
        "interest_rate":   "18% p.a.",
        "org_name":        "Your Firm Name",
        "phone":           "9876543210",
    },
    build_subject=_pay_subject,
    build_body=_pay_body,
)

ALL_TEMPLATES = [TEMPLATE_GST, TEMPLATE_INVOICE, TEMPLATE_PAYMENT]


# ═══════════════════════════════════════════════════════════════
#  CUSTOM EMAIL — Template Model & Store
# ═══════════════════════════════════════════════════════════════

import re as _re
import json as _json

class CustomTemplate:
    def __init__(self, name, subject_tpl, body_tpl, cc="",
                 has_attachment=False, attachment_ext="pdf"):
        self.name           = name
        self.icon           = "✏"
        self.subject_tpl    = subject_tpl
        self.body_tpl       = body_tpl
        self.cc             = cc
        self.has_attachment = has_attachment
        self.attachment_ext = attachment_ext

    @property
    def placeholders(self):
        found = _re.findall(r'\{(\w+)\}', self.subject_tpl + "\n" + self.body_tpl)
        seen, result = set(), []
        for p in found:
            if p not in seen:
                seen.add(p); result.append(p)
        return result

    @property
    def all_cols(self):
        phs = self.placeholders
        ph_lower = [p.lower() for p in phs]
        cols = []
        if 'name'  not in ph_lower: cols.append('Name')
        if 'email' not in ph_lower: cols.append('Email')
        cols.extend(phs)
        if self.has_attachment and 'attachment file' not in [c.lower() for c in cols]:
            cols.append('Attachment File')
        return cols or ['Name', 'Email']

    def build_subject(self, cfg, row):
        combined = {**cfg, **row}
        return _re.sub(r'\{(\w+)\}', lambda m: str(combined.get(m.group(1), m.group(0))), self.subject_tpl)

    def build_body(self, cfg, row):
        combined = {**cfg, **row}
        return _re.sub(r'\{(\w+)\}', lambda m: str(combined.get(m.group(1), m.group(0))), self.body_tpl)

    def to_dict(self):
        return {'name': self.name, 'subject': self.subject_tpl, 'body': self.body_tpl,
                'cc': self.cc, 'has_attachment': self.has_attachment, 'attachment_ext': self.attachment_ext}

    @classmethod
    def from_dict(cls, d):
        return cls(name=d.get('name','Unnamed'), subject_tpl=d.get('subject',''),
                   body_tpl=d.get('body',''), cc=d.get('cc',''),
                   has_attachment=d.get('has_attachment', False),
                   attachment_ext=d.get('attachment_ext','pdf'))


class CustomTemplateStore:
    def __init__(self, path):
        self._path = path
        self._templates = []
        self._load()

    def _load(self):
        if not os.path.exists(self._path):
            return
        try:
            with open(self._path, 'r', encoding='utf-8') as f:
                data = _json.load(f)
            self._templates = [CustomTemplate.from_dict(t) for t in data.get('templates', [])]
        except Exception:
            self._templates = []

    def save_all(self):
        with open(self._path, 'w', encoding='utf-8') as f:
            _json.dump({'templates': [t.to_dict() for t in self._templates]},
                        f, indent=2, ensure_ascii=False)

    @property
    def templates(self):
        return list(self._templates)

    def names(self):
        return [t.name for t in self._templates]

    def add(self, template):
        self._templates.append(template)
        self.save_all()
        return len(self._templates) - 1

    def update(self, idx, template):
        if 0 <= idx < len(self._templates):
            self._templates[idx] = template
            self.save_all()

    def delete(self, idx):
        if 0 <= idx < len(self._templates):
            del self._templates[idx]
            self.save_all()


# ═══════════════════════════════════════════════════════════════
#  SEND ENGINE
# ═══════════════════════════════════════════════════════════════

def send_emails(template, cfg, recipients, attachment_folder, log_cb, done_cb):
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
    except Exception as exc:
        log_cb(f"[ERROR] Could not connect to Outlook: {exc}")
        done_cb(success=False)
        pythoncom.CoUninitialize()
        return

    failed = []
    for idx, row in enumerate(recipients, 1):
        name  = str(row.get("Name", "")).strip()
        email = str(row.get("Email", "")).strip()
        if not email:
            log_cb(f"[SKIP]  Row {idx} — empty email.")
            continue
        try:
            mail            = outlook.CreateItem(0)
            mail.To         = email
            mail.CC         = cfg.get("cc", "")
            mail.Subject    = template.build_subject(cfg, row)
            body_text       = template.build_body(cfg, row)
            # Use HTMLBody so Outlook doesn't override with its default HTML/signature
            html_body       = "<html><body><pre style='font-family:Calibri,sans-serif;font-size:11pt;white-space:pre-wrap'>" \
                              + body_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") \
                              + "</pre></body></html>"
            mail.HTMLBody   = html_body

            if template.has_attachment and attachment_folder:
                fname = str(row.get("Attachment File", "")).strip()
                if fname:
                    import os
                    base = fname[:-4] if fname.lower().endswith(".pdf") else fname
                    path = os.path.join(attachment_folder, base + ".pdf")
                    if os.path.exists(path):
                        mail.Attachments.Add(path)
                    else:
                        log_cb(f"[WARN]  {name} — attachment not found: {base}.pdf")

            mail.Send()
            log_cb(f"[SENT]  {name} <{email}>")
        except Exception as exc:
            log_cb(f"[ERROR] {name} <{email}> — {exc}")
            failed.append(email)

    pythoncom.CoUninitialize()
    done_cb(success=True, failed=failed)


# ═══════════════════════════════════════════════════════════════
#  GUI  —  Design Tokens & Setup
# ═══════════════════════════════════════════════════════════════
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import calendar
import os
import threading

# Set default theme
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

# Fonts
F_TITLE    = ("Segoe UI", 16, "bold")
F_BADGE    = ("Segoe UI", 12, "bold")
F_LABEL    = ("Segoe UI", 13)
F_LABEL_SML= ("Segoe UI", 11)
F_INPUT    = ("Segoe UI", 13)
F_BTN      = ("Segoe UI", 11, "bold")
F_BTN_LG   = ("Segoe UI", 12, "bold")
F_MONO     = ("Consolas", 12)
F_MONO_LG  = ("Consolas", 13)
F_TAB      = ("Segoe UI", 11, "bold")
F_COL_HDR  = ("Segoe UI", 11, "bold")
DEMO_VIDEO_URL = "https://www.youtube.com/watch?v=MkLX85XU5rU"

class CustomDatePicker(ctk.CTkToplevel):
    def __init__(self, parent, target_var):
        super().__init__(parent)
        self.target_var = target_var
        self.title("Select Date")
        self.geometry("320x360")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self.current_date = datetime.now()
        
        self._build_ui()
        self._update_calendar()
        
    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", pady=10, padx=10)
        
        ctk.CTkButton(header, text="<", width=30, command=self._prev_month).pack(side="left")
        self.month_lbl = ctk.CTkLabel(header, text="", font=F_BTN_LG)
        self.month_lbl.pack(side="left", expand=True)
        ctk.CTkButton(header, text=">", width=30, command=self._next_month).pack(side="right")
        
        self.cal_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.cal_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for c, day in enumerate(days):
            ctk.CTkLabel(self.cal_frame, text=day, font=F_LABEL_SML, text_color="gray").grid(row=0, column=c, padx=3, pady=3)
            
        self.day_buttons = []
        for r in range(6):
            row_btns = []
            for c in range(7):
                btn = ctk.CTkButton(self.cal_frame, text="", width=35, height=35, fg_color="transparent", text_color=("black", "white"), font=F_LABEL, hover_color=("gray80", "gray30"), command=lambda r=r, c=c: self._select_day(r, c))
                btn.grid(row=r+1, column=c, padx=2, pady=2)
                row_btns.append(btn)
            self.day_buttons.append(row_btns)
            
    def _update_calendar(self):
        self.month_lbl.configure(text=self.current_date.strftime("%B %Y"))
        
        cal = calendar.monthcalendar(self.current_date.year, self.current_date.month)
        
        for r in range(6):
            for c in range(7):
                if r < len(cal) and cal[r][c] != 0:
                    day = cal[r][c]
                    self.day_buttons[r][c].configure(text=str(day), state="normal")
                else:
                    self.day_buttons[r][c].configure(text="", state="disabled")
                    
    def _prev_month(self):
        first = self.current_date.replace(day=1)
        prev_month = first - timedelta(days=1)
        self.current_date = prev_month.replace(day=1)
        self._update_calendar()
        
    def _next_month(self):
        days_in_month = calendar.monthrange(self.current_date.year, self.current_date.month)[1]
        last = self.current_date.replace(day=days_in_month)
        next_month = last + timedelta(days=1)
        self.current_date = next_month.replace(day=1)
        self._update_calendar()
        
    def _select_day(self, r, c):
        day = int(self.day_buttons[r][c].cget("text"))
        selected = datetime(self.current_date.year, self.current_date.month, day)
        self.target_var.set(selected.strftime("%d %b %Y"))
        self.destroy()

class BulkMailApp(ctk.CTk):
    def __init__(self, hide_switcher=False):
        super().__init__()
        self.title("Bulk Outlook Mailer")
        self.geometry("900x700")
        self.minsize(740, 560)
        self._hide_switcher = hide_switcher

        self._active_template  = ALL_TEMPLATES[0]
        self._cfg_vars         = {}
        self._attachment_folder = ""
        self._sending          = False
        self._add_vars         = {}

        self._build_ui()
        self._switch_template(0)

    def set_theme(self, mode: str):
        ctk.set_appearance_mode(mode)
        style = ttk.Style()
        is_dark = mode == "Dark"
        bg_color = "#1e2d3d" if is_dark else "white"
        fg_color = "white" if is_dark else "black"
        hl_color = "#1e3a6e" if is_dark else "#DBEAFE"
        hdr_bg = "#1e293b" if is_dark else "#F1F5F9"
        style.configure("Custom.Treeview", background=bg_color, foreground=fg_color, fieldbackground=bg_color)
        style.map("Custom.Treeview", background=[("selected", hl_color)], foreground=[("selected", fg_color)])
        style.configure("Custom.Treeview.Heading", background=hdr_bg, foreground=fg_color)

    def _build_ui(self):
        self.bottom_bar = ctk.CTkFrame(self, height=60, corner_radius=0)
        self.bottom_bar.pack(fill="x", side="bottom")

        self._status_lbl = ctk.CTkLabel(self.bottom_bar, text="", font=F_LABEL_SML, text_color="gray")
        self._status_lbl.pack(side="left", padx=20)

        ctk.CTkButton(self.bottom_bar, text="  Send All Emails  →", font=F_BTN_LG, height=45, fg_color="#2563EB", hover_color="#1D4ED8", command=self._start_send).pack(side="right", padx=(10, 20), pady=10)
        ctk.CTkButton(self.bottom_bar, text="Preview Email", height=45, fg_color="transparent", border_width=1, text_color=("black", "white"), command=self._show_preview).pack(side="right", padx=(5, 0), pady=10)
        ctk.CTkButton(self.bottom_bar, text="Check Data", height=45, fg_color="transparent", border_width=1, text_color=("black", "white"), command=self._check_data).pack(side="right", padx=(5, 0), pady=10)
        ctk.CTkButton(self.bottom_bar, text="▶ View Demo", height=45, fg_color="#DC2626", hover_color="#B91C1C", text_color="white", command=self.open_demo_link).pack(side="right", padx=(5, 0), pady=10)

        self._nb = ctk.CTkTabview(self)
        self._nb.pack(fill="both", expand=True, padx=20, pady=(10, 10))

        self._tab_config     = self._nb.add("Configuration")
        self._tab_recipients = self._nb.add("Recipients")
        self._tab_log        = self._nb.add("Log")

        self._build_log_tab()

    def _rebuild_config_tab(self):
        for w in self._tab_config.winfo_children():
            w.destroy()
        self._cfg_vars = {}
        t = self._active_template

        scroll = ctk.CTkScrollableFrame(self._tab_config, fg_color="transparent", corner_radius=0)
        scroll.pack(fill="both", expand=True)
        scroll.columnconfigure(0, weight=1)

        form_card = ctk.CTkFrame(scroll, corner_radius=10)
        form_card.pack(fill="x", anchor="n", padx=30, pady=20)
        form_card.columnconfigure(1, weight=1)

        self._row_i = 0
        lbl_text = "Email Template" if not self._hide_switcher else f"{t.icon}  {t.name} Config"
        ctk.CTkLabel(form_card, text=lbl_text, font=F_TITLE).grid(row=self._row_i, column=0, sticky="nw", pady=20, padx=30)

        has_cb_content = (not self._hide_switcher) or hasattr(self, "_open_email_settings")
        if has_cb_content:
            cb_frame = ctk.CTkFrame(form_card, fg_color="transparent")
            cb_frame.grid(row=self._row_i, column=1, sticky="nw", pady=20)

            if not self._hide_switcher:
                self._tmpl_var = ctk.StringVar(value=f"{t.icon}  {t.name}")
                names = [f"{tmpl.icon}  {tmpl.name}" for tmpl in ALL_TEMPLATES]
                cb = ctk.CTkComboBox(cb_frame, variable=self._tmpl_var, values=names, width=300, command=self._on_template_change)
                cb.pack(side="left")
                self._tmpl_cb = cb

            if hasattr(self, "_open_email_settings"):
                ctk.CTkButton(cb_frame, text="⚙ Email Settings", fg_color="#DBEAFE", text_color="#2563EB", hover_color="#BFDBFE", command=self._open_email_settings).pack(side="right", padx=15)

        self._row_i += 1

        def validate_phone(P):
            if P == "": return True
            if P.isdigit() and len(P) <= 10: return True
            return False
        vcmd_phone = (self.register(validate_phone), "%P")

        for i, (label, key) in enumerate(t.config_fields):
            ctk.CTkLabel(form_card, text=label, font=F_LABEL).grid(row=self._row_i+i, column=0, sticky="w", pady=10, padx=30)
            var = ctk.StringVar(value=t.defaults.get(key, ""))
            self._cfg_vars[key] = var

            if key == "phone":
                inner_f = ctk.CTkFrame(form_card, fg_color="transparent")
                inner_f.grid(row=self._row_i+i, column=1, sticky="ew", pady=10, padx=(0, 30))
                ctk.CTkLabel(inner_f, text="+91", font=F_INPUT).pack(side="left", padx=(0, 10))
                e = ctk.CTkEntry(inner_f, textvariable=var, font=F_INPUT, validate="key", validatecommand=vcmd_phone)
                e.pack(side="left", fill="x", expand=True)
            elif "deadline" in key.lower() or "date" in key.lower():
                inner_f = ctk.CTkFrame(form_card, fg_color="transparent")
                inner_f.grid(row=self._row_i+i, column=1, sticky="ew", pady=10, padx=(0, 30))
                e = ctk.CTkEntry(inner_f, textvariable=var, font=F_INPUT)
                e.pack(side="left", fill="x", expand=True)
                btn = ctk.CTkButton(inner_f, text="📅", width=40, fg_color="transparent", text_color=("black", "white"), border_width=1, command=lambda v=var: CustomDatePicker(self, v))
                btn.pack(side="right", padx=(5, 0))
            else:
                e = ctk.CTkEntry(form_card, textvariable=var, font=F_INPUT)
                e.grid(row=self._row_i+i, column=1, sticky="ew", pady=10, padx=(0, 30))

    def _rebuild_recipients_tab(self):
        for w in self._tab_recipients.winfo_children():
            w.destroy()
        t = self._active_template
        cols = t.all_cols

        toolbar = ctk.CTkFrame(self._tab_recipients, fg_color="transparent")
        toolbar.pack(fill="x", pady=10)

        ctk.CTkButton(toolbar, text="📂  Import from Excel", command=self._import_excel, width=150).pack(side="left", padx=(0, 10))
        ctk.CTkButton(toolbar, text="⬇  Download Sample", fg_color="transparent", border_width=1, text_color=("black", "white"), command=self._download_sample, width=150).pack(side="left", padx=(0, 10))
        ctk.CTkButton(toolbar, text="🗑  Clear All", fg_color="#EF4444", hover_color="#DC2626", text_color="white", command=self._clear_recipients, width=100).pack(side="left", padx=(0, 10))

        if t.has_attachment:
            self._folder_lbl = ctk.CTkLabel(toolbar, text="(not selected)", fg_color=("gray90", "gray20"), corner_radius=5, width=200)
            self._folder_lbl.pack(side="left", padx=(10, 10))
            ctk.CTkButton(toolbar, text="Browse…", width=80, fg_color="transparent", border_width=1, text_color=("black", "white"), command=self._pick_folder).pack(side="left")

        self._tree_frame = ctk.CTkFrame(self._tab_recipients, fg_color="transparent")
        self._tree_frame.pack(fill="both", expand=True, pady=(10, 0))

        style = ttk.Style()
        style.theme_use("clam")
        
        is_dark = ctk.get_appearance_mode() == "Dark"
        bg_color = "#1e2d3d" if is_dark else "white"
        fg_color = "white" if is_dark else "black"
        hl_color = "#1e3a6e" if is_dark else "#DBEAFE"
        hdr_bg = "#1e293b" if is_dark else "#F1F5F9"
        
        style.configure("Custom.Treeview", background=bg_color, foreground=fg_color, fieldbackground=bg_color, rowheight=30, borderwidth=0, font=F_LABEL_SML)
        style.map("Custom.Treeview", background=[("selected", hl_color)], foreground=[("selected", "white")])
        style.configure("Custom.Treeview.Heading", background=hdr_bg, foreground=fg_color, font=F_BTN, padding=5)

        self._tree = ttk.Treeview(self._tree_frame, columns=cols, show="headings", style="Custom.Treeview")
        for col in cols:
            self._tree.heading(col, text=col)
            self._tree.column(col, width=150, minwidth=100)

        vsb = ttk.Scrollbar(self._tree_frame, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(self._tree_frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self._tree_frame.rowconfigure(0, weight=1)
        self._tree_frame.columnconfigure(0, weight=1)

        self._tree.bind("<Control-v>", self._on_paste)
        self._tree.bind("<Control-V>", self._on_paste)
        self._tree.bind("<Double-1>", self._on_double_click)
        self._tree.bind("<Delete>", self._on_delete)
        self._tree.bind("<BackSpace>", self._on_delete)

        bot = ctk.CTkFrame(self._tab_recipients, fg_color="transparent")
        bot.pack(fill="x", pady=(5, 0))
        self._count_lbl = ctk.CTkLabel(bot, text="0 recipients", font=F_LABEL_SML)
        self._count_lbl.pack(side="left")
        ctk.CTkLabel(bot, text="Ctrl+V: Paste | Double-click: Edit | Del: Remove", font=F_LABEL_SML, text_color="gray").pack(side="right")

    def _on_paste(self, event):
        try:
            raw = self.clipboard_get()
        except tk.TclError:
            return "break"
        lines = [ln for ln in raw.splitlines() if ln.strip()]
        if not lines:
            return "break"
        cols = self._active_template.all_cols
        for line in lines:
            parts = line.split("\t")
            parts = (parts + [""] * len(cols))[:len(cols)]
            self._tree.insert("", "end", values=parts)
        self._update_count()
        return "break"

    def _on_delete(self, event):
        selected = self._tree.selection()
        for item in selected:
            self._tree.delete(item)
        self._update_count()

    def _on_double_click(self, event):
        region = self._tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        column = self._tree.identify_column(event.x)
        row_id = self._tree.identify_row(event.y)
        if not column or not row_id: return
        
        col_index = int(column[1:]) - 1
        x, y, width, height = self._tree.bbox(row_id, column)
        val = self._tree.set(row_id, column)

        entry = tk.Entry(self._tree, font=F_LABEL_SML)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, val)
        entry.focus()

        def save_edit(event=None):
            new_val = entry.get()
            self._tree.set(row_id, column, new_val)
            entry.destroy()
            self._update_count()

        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)

    def _build_log_tab(self):
        hdr = ctk.CTkFrame(self._tab_log, fg_color="transparent")
        hdr.pack(fill="x", pady=(10, 5))

        ctk.CTkLabel(hdr, text="Send Log", font=F_BTN).pack(side="left")
        ctk.CTkButton(hdr, text="Clear Log", width=100, fg_color="transparent", border_width=1, text_color=("black", "white"), command=self._clear_log).pack(side="right")

        self._log_box = ctk.CTkTextbox(self._tab_log, font=F_MONO, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    def _on_template_change(self, value=None):
        idx = -1
        names = [f"{tmpl.icon}  {tmpl.name}" for tmpl in ALL_TEMPLATES]
        val = self._tmpl_var.get()
        if val in names:
            idx = names.index(val)
        if idx >= 0:
            self._switch_template(idx)

    def _switch_template(self, idx):
        self._active_template = ALL_TEMPLATES[idx]
        self._attachment_folder = ""
        self._rebuild_config_tab()
        self._rebuild_recipients_tab()
        if hasattr(self, "_status_lbl") and hasattr(self._status_lbl, "configure"):
            self._status_lbl.configure(text="")

    def _get_config(self):
        return {k: v.get() for k, v in self._cfg_vars.items()}

    def _get_recipients(self):
        rows = []
        cols = self._active_template.all_cols
        for item in self._tree.get_children():
            values = self._tree.item(item, 'values')
            row = dict(zip(cols, values))
            if any(row.values()):
                rows.append(row)
        return rows

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="Select Attachment Folder")
        if folder:
            self._attachment_folder = folder
            short = folder if len(folder) < 30 else "…" + folder[-27:]
            self._folder_lbl.configure(text=short)

    def _clear_recipients(self):
        if self._tree.get_children():
            if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all recipients?", parent=self):
                for item in self._tree.get_children():
                    self._tree.delete(item)
                self._update_count()

    def _download_sample(self):
        import shutil
        t = self._active_template
        mapping = {
            "GST Return Data Request": "GST_Return_Recipients.xlsx",
            "Invoice Sender": "Invoice_Sender_Recipients.xlsx",
            "Outstanding Payment Reminder": "Outstanding_Payment_Recipients.xlsx"
        }
        filename = mapping.get(t.name)
        if not filename:
            messagebox.showerror("Error", "No sample file configured for this template.")
            return

        base_dir = os.path.dirname(os.path.abspath(__file__))
        src_path = os.path.join(base_dir, "Input For Email Upload", filename)

        if not os.path.exists(src_path):
            messagebox.showerror("Not Found", f"Sample file not found at:\n{src_path}")
            return

        save_path = filedialog.asksaveasfilename(
            title="Save Sample Excel File",
            initialfile=filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            parent=self
        )
        if save_path:
            try:
                shutil.copy2(src_path, save_path)
                messagebox.showinfo("Success", f"Sample saved successfully to:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file:\n{str(e)}")

    def _import_excel(self):
        if not _OPENPYXL:
            messagebox.showerror("Missing Library", "openpyxl is not installed.\n\nRun: pip install openpyxl")
            return
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            messagebox.showerror("Read Error", f"Could not read Excel file:\n{exc}")
            return
        if not rows:
            messagebox.showwarning("Empty File", "The Excel file has no data.")
            return

        cols = self._active_template.all_cols
        cols_lower = [c.lower() for c in cols]

        first_row = [str(v).strip() if v is not None else "" for v in rows[0]]
        header_map = {}
        for xi, hdr in enumerate(first_row):
            hl = hdr.lower()
            if hl in cols_lower:
                header_map[xi] = cols_lower.index(hl)

        if header_map:
            data_rows = rows[1:]
        else:
            data_rows = rows
            header_map = {xi: xi for xi in range(min(len(first_row), len(cols)))}

        for item in self._tree.get_children():
            self._tree.delete(item)

        for row in data_rows:
            if not any(v for v in row if v is not None and str(v).strip()):
                continue
            item_vals = [""] * len(cols)
            for xi, ci in header_map.items():
                val = row[xi] if xi < len(row) else ""
                item_vals[ci] = "" if val is None else str(val).strip()
            self._tree.insert("", "end", values=item_vals)

        self._update_count()

    def _update_count(self):
        n = len(self._tree.get_children())
        self._count_lbl.configure(text=f"{n} recipient{'s' if n != 1 else ''}")

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def open_demo_link(self):
        import webbrowser
        webbrowser.open_new_tab(DEMO_VIDEO_URL)

    def _show_preview(self):
        cfg  = self._get_config()
        recs = self._get_recipients()
        t    = self._active_template

        sample_row = recs[0] if recs else {c: f"<{c}>" for c in t.all_cols}
        subject    = t.build_subject(cfg, sample_row)
        body       = t.build_body(cfg, sample_row)

        win = ctk.CTkToplevel(self)
        win.title("Email Preview")
        win.geometry("660x560")
        win.grab_set()

        hdr = ctk.CTkFrame(win, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text=f"{t.icon}  {t.name}", font=F_BADGE).pack(side="left", padx=20, pady=10)

        content = ctk.CTkFrame(win, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20, pady=10)

        ctk.CTkLabel(content, text="SUBJECT", font=("Segoe UI", 10, "bold"), text_color="gray").pack(anchor="w")
        ctk.CTkLabel(content, text=subject, font=F_LABEL, wraplength=600, justify="left").pack(anchor="w", pady=(0, 20))

        ctk.CTkLabel(content, text="BODY", font=("Segoe UI", 10, "bold"), text_color="gray").pack(anchor="w")
        body_box = ctk.CTkTextbox(content, font=("Segoe UI", 12))
        body_box.pack(fill="both", expand=True)
        body_box.insert("end", body)
        body_box.configure(state="disabled")

        ctk.CTkButton(win, text="Close", command=win.destroy).pack(pady=10)

    def _log(self, message: str, tag="info"):
        self._log_box.configure(state="normal")
        if tag == "sent" or "[SENT]" in message:
            color = "#4ADE80"
            tag_name = "sent"
        elif tag == "error" or "[ERROR]" in message:
            color = "#F87171"
            tag_name = "error"
        elif tag == "skip" or "[SKIP]" in message:
            color = "#FBBF24"
            tag_name = "skip"
        elif tag == "warn" or "[WARN]" in message:
            color = "#FB923C"
            tag_name = "warn"
        else:
            color = None
            tag_name = "info"
            
        if color:
            self._log_box.tag_config(tag_name, foreground=color)
            self._log_box.insert("end", message + "\n", tag_name)
        else:
            self._log_box.insert("end", message + "\n")
            
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _log_safe(self, msg):
        self.after(0, self._log, msg)

    def _check_data(self):
        recs = self._get_recipients()
        cols = self._active_template.all_cols

        win = ctk.CTkToplevel(self)
        win.title("Check Data")
        win.geometry("720x480")
        win.grab_set()

        hdr = ctk.CTkFrame(win, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Recipient Preview", font=F_BADGE).pack(side="left", padx=20, pady=10)
        ctk.CTkLabel(hdr, text=f"{len(recs)} row(s)", font=F_BADGE).pack(side="right", padx=20)

        frame = ctk.CTkFrame(win)
        frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        tree = ttk.Treeview(frame, columns=cols, show="headings", style="Custom.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(90, 640 // len(cols)), minwidth=80)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        for r in recs:
            tree.insert("", "end", values=[r.get(c, "") for c in cols])

        ctk.CTkButton(win, text="Close", command=win.destroy).pack(pady=10)

    def _start_send(self):
        if self._sending:
            messagebox.showwarning("Busy", "Already sending. Please wait.")
            return
        recs = self._get_recipients()
        if not recs:
            messagebox.showwarning("No Recipients", "No recipient rows found.\n\nTip: Use 'Check Data' button to verify your data is parsed correctly.")
            return

        t = self._active_template
        if t.has_attachment and not self._attachment_folder:
            if not messagebox.askyesno("No folder", "No attachment folder selected.\nContinue anyway (no attachments)?"):
                return
                
        msg = f"Send {len(recs)} email(s) using template:\n'{t.name}'?\n\n"
            
        if not messagebox.askyesno("Confirm Send", msg):
            return

        self._sending = True
        self._nb.set("Log")
        self._log(f"── {t.icon} {t.name} — {len(recs)} recipient(s) ──")
        threading.Thread(
            target=send_emails,
            args=(t, self._get_config(), recs,
                  self._attachment_folder,
                  self._log_safe, self._on_done),
            daemon=True
        ).start()

    def _on_done(self, success=True, failed=None):
        failed = failed or []
        def _show():
            self._sending = False
            if success:
                msg = "All emails are sent."
                if failed:
                    msg += f"\n\nFailed ({len(failed)}):\n" + "\n".join(failed)
                self._log("── Done ──")
                messagebox.showinfo("Done", msg)
            else:
                messagebox.showerror("Outlook Error", "Could not connect to Outlook.\n\n1. Make sure Outlook is installed and open.\n2. Note: The 'New Outlook' (Modern App) does not support COM automation. You must use the classic Outlook Desktop App.")
        self.after(0, _show)

class GSTReturnMailApp(BulkMailApp):
    def __init__(self):
        super().__init__(hide_switcher=True)
        self._switch_template(0)

class InvoiceSenderMailApp(BulkMailApp):
    def __init__(self):
        super().__init__(hide_switcher=True)
        self._switch_template(1)

class PaymentReminderMailApp(BulkMailApp):
    def __init__(self):
        super().__init__(hide_switcher=True)
        self._switch_template(2)


# ═══════════════════════════════════════════════════════════════
#  CUSTOM EMAIL SEND ENGINE  (Outlook COM — any attachment type)
# ═══════════════════════════════════════════════════════════════

def send_custom_emails(template, recipients, attachment_folder, log_cb, done_cb):
    import glob as _glob
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
    except Exception as exc:
        log_cb(f"[ERROR] Could not connect to Outlook: {exc}")
        done_cb(success=False)
        pythoncom.CoUninitialize()
        return

    cfg    = {"cc": template.cc}
    failed = []

    for idx, row in enumerate(recipients, 1):
        email = str(row.get("Email", "")).strip()
        name  = str(row.get("Name",  "")).strip()
        if not email:
            log_cb(f"[SKIP]  Row {idx} — empty email.")
            continue
        try:
            mail          = outlook.CreateItem(0)
            mail.To       = email
            mail.CC       = template.cc
            mail.Subject  = template.build_subject(cfg, row)
            body_text     = template.build_body(cfg, row)
            mail.HTMLBody = (
                "<html><body><pre style='font-family:Calibri,sans-serif;font-size:11pt;white-space:pre-wrap'>"
                + body_text.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                + "</pre></body></html>"
            )

            if template.has_attachment and attachment_folder:
                fname = str(row.get("Attachment File", "")).strip()
                if fname:
                    ext = template.attachment_ext.lower().lstrip(".")
                    candidates = [os.path.join(attachment_folder, fname)]
                    if ext == "any":
                        candidates = _glob.glob(os.path.join(attachment_folder, fname + ".*")) + candidates
                    else:
                        if not fname.lower().endswith("." + ext):
                            candidates.append(os.path.join(attachment_folder, fname + "." + ext))
                    att_path = next((p for p in candidates if os.path.exists(p)), None)
                    if att_path:
                        mail.Attachments.Add(att_path)
                    else:
                        log_cb(f"[WARN]  {name} — attachment not found: {fname}")

            mail.Send()
            log_cb(f"[SENT]  {name} <{email}>")

        except Exception as exc:
            log_cb(f"[ERROR] {name} <{email}> — {exc}")
            failed.append(email)

    pythoncom.CoUninitialize()
    if failed:
        log_cb(f"\n[DONE] Finished with {len(failed)} error(s).")
        done_cb(success=False, failed=failed)
    else:
        log_cb("\n[DONE] All emails sent successfully!")
        done_cb(success=True)


# ═══════════════════════════════════════════════════════════════
#  CUSTOM EMAIL UI  (Outlook version)
# ═══════════════════════════════════════════════════════════════

class CustomMailApp(ctk.CTk):
    _STORE_FILENAME = "custom_email_templates.json"

    def __init__(self):
        super().__init__()
        self.title("Custom Email Builder — Outlook")
        self.geometry("1000x760")
        self.minsize(840, 620)

        store_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), self._STORE_FILENAME)
        self._store           = CustomTemplateStore(store_path)
        self._active_idx      = -1
        self._attachment_folder = ""
        self._sending         = False
        self._recipients_cols = ["Name", "Email"]
        self._detected_cols   = []
        self._debounce_id     = None

        self._build_ui()
        self._reload_dropdown()
        if self._store.templates:
            self._select_stored(0)
        else:
            self._new_template()

    def set_theme(self, mode):
        ctk.set_appearance_mode(mode)

    # ── UI skeleton ──────────────────────────────────────────────

    def _build_ui(self):
        top = ctk.CTkFrame(self, height=52, corner_radius=0, fg_color=("gray88","gray16"))
        top.pack(fill="x", side="top")
        top.pack_propagate(False)

        ctk.CTkLabel(top, text="Template:", font=F_BTN).pack(side="left", padx=(18,5), pady=14)
        self._tmpl_var = ctk.StringVar()
        self._tmpl_cb  = ctk.CTkComboBox(top, variable=self._tmpl_var, values=[],
                                          width=240, command=self._on_dropdown_change, state="readonly")
        self._tmpl_cb.pack(side="left", padx=(0,8), pady=14)

        ctk.CTkButton(top, text="+ New",    width=72,  height=32,
                       command=self._new_template).pack(side="left", padx=3)
        ctk.CTkButton(top, text="💾 Save",  width=82,  height=32,
                       fg_color="#059669", hover_color="#047857",
                       command=self._save_template).pack(side="left", padx=3)
        ctk.CTkButton(top, text="🗑 Delete", width=82, height=32,
                       fg_color="#DC2626", hover_color="#B91C1C",
                       command=self._delete_template).pack(side="left", padx=3)

        bot = ctk.CTkFrame(self, height=62, corner_radius=0)
        bot.pack(fill="x", side="bottom")
        bot.pack_propagate(False)
        self._status_lbl = ctk.CTkLabel(bot, text="", font=F_LABEL_SML, text_color="gray")
        self._status_lbl.pack(side="left", padx=20)
        ctk.CTkButton(bot, text="  Send All Emails  →", font=F_BTN_LG, height=44,
                       fg_color="#2563EB", hover_color="#1D4ED8",
                       command=self._start_send).pack(side="right", padx=(10,20), pady=9)
        ctk.CTkButton(bot, text="Preview Email", height=44, fg_color="transparent",
                       border_width=1, text_color=("black","white"),
                       command=self._show_preview).pack(side="right", padx=(5,0), pady=9)
        ctk.CTkButton(bot, text="Check Data", height=44, fg_color="transparent",
                       border_width=1, text_color=("black","white"),
                       command=self._check_data).pack(side="right", padx=(5,0), pady=9)

        self._nb = ctk.CTkTabview(self)
        self._nb.pack(fill="both", expand=True, padx=18, pady=(10,8))
        self._tab_builder    = self._nb.add("✏  Builder")
        self._tab_recipients = self._nb.add("Recipients")
        self._tab_log        = self._nb.add("Log")

        self._build_builder_tab()
        self._build_recipients_tab(["Name", "Email"])
        self._build_log_tab()

    # ── Builder tab ──────────────────────────────────────────────

    def _build_builder_tab(self):
        scroll = ctk.CTkScrollableFrame(self._tab_builder, fg_color="transparent")
        scroll.pack(fill="both", expand=True)
        scroll.columnconfigure(1, weight=1)
        r = 0

        ctk.CTkLabel(scroll, text="Template Name:", font=F_LABEL).grid(
            row=r, column=0, sticky="w", padx=(20,10), pady=(20,8))
        self._name_var = ctk.StringVar(value="My Custom Template")
        ctk.CTkEntry(scroll, textvariable=self._name_var, font=F_INPUT).grid(
            row=r, column=1, sticky="ew", padx=(0,20), pady=(20,8))
        r += 1

        ctk.CTkFrame(scroll, height=1, fg_color=("gray80","gray30")).grid(
            row=r, column=0, columnspan=2, sticky="ew", padx=20, pady=4)
        r += 1

        ctk.CTkLabel(scroll, text="Subject:", font=F_LABEL).grid(
            row=r, column=0, sticky="w", padx=(20,10), pady=8)
        self._subject_var = ctk.StringVar()
        self._subject_var.trace_add("write", self._on_tpl_change_trace)
        ctk.CTkEntry(scroll, textvariable=self._subject_var, font=F_INPUT,
                      placeholder_text="e.g.  Invoice for {Client} — {Month}").grid(
            row=r, column=1, sticky="ew", padx=(0,20), pady=8)
        r += 1

        ctk.CTkLabel(scroll, text="Body:", font=F_LABEL).grid(
            row=r, column=0, sticky="nw", padx=(20,10), pady=(8,4))
        self._body_text = ctk.CTkTextbox(scroll, font=("Segoe UI",12), height=220)
        self._body_text.grid(row=r, column=1, sticky="ew", padx=(0,20), pady=(8,4))
        self._body_text.bind("<KeyRelease>", self._on_tpl_change_event)
        r += 1

        ctk.CTkLabel(scroll,
            text="Tip: Use {Name}, {Email}, {Amount}, {Month} etc. as placeholders.\n"
                 "Name and Email are always included. Every unique {placeholder} becomes an Excel column.",
            font=("Segoe UI",10), text_color="gray", justify="left").grid(
            row=r, column=1, sticky="w", padx=(0,20), pady=(0,8))
        r += 1

        ctk.CTkFrame(scroll, height=1, fg_color=("gray80","gray30")).grid(
            row=r, column=0, columnspan=2, sticky="ew", padx=20, pady=4)
        r += 1

        ctk.CTkLabel(scroll, text="CC Email:", font=F_LABEL).grid(
            row=r, column=0, sticky="w", padx=(20,10), pady=8)
        self._cc_var = ctk.StringVar()
        ctk.CTkEntry(scroll, textvariable=self._cc_var, font=F_INPUT,
                      placeholder_text="cc@example.com  (optional, comma-separated)").grid(
            row=r, column=1, sticky="ew", padx=(0,20), pady=8)
        r += 1

        att_row = ctk.CTkFrame(scroll, fg_color="transparent")
        att_row.grid(row=r, column=0, columnspan=2, sticky="ew", padx=20, pady=8)
        self._has_att_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(att_row, text="Attach a file per recipient",
                         variable=self._has_att_var, font=F_LABEL,
                         command=self._on_att_toggle).pack(side="left", padx=(0,20))
        ctk.CTkLabel(att_row, text="Extension:", font=F_LABEL_SML).pack(side="left", padx=(0,6))
        self._ext_var = ctk.StringVar(value="pdf")
        self._ext_cb  = ctk.CTkComboBox(att_row, variable=self._ext_var, width=100,
                                         values=["pdf","xlsx","xls","docx","doc","jpg","png","any"],
                                         state="disabled")
        self._ext_cb.pack(side="left")
        ctk.CTkLabel(att_row,
            text="  (column 'Attachment File' will be added to Excel)",
            font=("Segoe UI",10), text_color="gray").pack(side="left", padx=6)
        r += 1

        ctk.CTkFrame(scroll, height=1, fg_color=("gray80","gray30")).grid(
            row=r, column=0, columnspan=2, sticky="ew", padx=20, pady=4)
        r += 1

        col_card = ctk.CTkFrame(scroll, corner_radius=10, border_width=1)
        col_card.grid(row=r, column=0, columnspan=2, sticky="ew", padx=20, pady=12)

        col_hdr = ctk.CTkFrame(col_card, fg_color="transparent")
        col_hdr.pack(fill="x", padx=16, pady=(14,4))
        ctk.CTkLabel(col_hdr, text="📋  Detected Excel Columns", font=F_BTN).pack(side="left")
        ctk.CTkButton(col_hdr, text="Apply & Rebuild Recipients Tab ↓",
                       width=230, height=30, fg_color="#7C3AED", hover_color="#6D28D9",
                       command=self._apply_to_recipients).pack(side="right")

        self._cols_lbl = ctk.CTkLabel(col_card,
            text="(type in Subject / Body above to detect columns)",
            font=F_LABEL_SML, text_color="gray", justify="left")
        self._cols_lbl.pack(anchor="w", padx=16, pady=(0,14))

    def _on_att_toggle(self):
        self._ext_cb.configure(state="normal" if self._has_att_var.get() else "disabled")

    def _on_tpl_change_trace(self, *args):
        if self._debounce_id:
            self.after_cancel(self._debounce_id)
        self._debounce_id = self.after(400, self._refresh_cols_preview)

    def _on_tpl_change_event(self, event=None):
        if self._debounce_id:
            self.after_cancel(self._debounce_id)
        self._debounce_id = self.after(400, self._refresh_cols_preview)

    def _extract_placeholders(self):
        subject = self._subject_var.get()
        body    = self._body_text.get("1.0", "end")
        found   = _re.findall(r'\{(\w+)\}', subject + body)
        seen, unique = set(), []
        for p in found:
            if p not in seen:
                seen.add(p); unique.append(p)
        return unique

    def _build_cols_from_phs(self, phs, has_att):
        ph_lower = [p.lower() for p in phs]
        cols = []
        if 'name'  not in ph_lower: cols.append('Name')
        if 'email' not in ph_lower: cols.append('Email')
        cols.extend(phs)
        if has_att and 'attachment file' not in [c.lower() for c in cols]:
            cols.append('Attachment File')
        return cols or ['Name', 'Email']

    def _refresh_cols_preview(self):
        phs  = self._extract_placeholders()
        cols = self._build_cols_from_phs(phs, self._has_att_var.get())
        self._detected_cols = cols
        self._cols_lbl.configure(
            text=("  →  ".join(cols)) if cols else "(no placeholders detected)")

    def _apply_to_recipients(self):
        phs  = self._extract_placeholders()
        cols = self._build_cols_from_phs(phs, self._has_att_var.get())
        if not phs:
            messagebox.showwarning("No Placeholders",
                "Add {placeholders} in Subject or Body first.\n"
                "Example:  Dear {Name}, your invoice amount is {Amount}.",
                parent=self)
            return
        if hasattr(self, '_tree') and self._tree.get_children():
            if not messagebox.askyesno("Rebuild Recipients",
                    "This will clear all recipient data and rebuild columns.\nContinue?",
                    parent=self):
                return
        self._recipients_cols   = cols
        self._attachment_folder = ""
        self._build_recipients_tab(cols)
        self._nb.set("Recipients")

    # ── Recipients tab ───────────────────────────────────────────

    def _build_recipients_tab(self, cols=None):
        for w in self._tab_recipients.winfo_children():
            w.destroy()
        if cols is None:
            cols = getattr(self, '_recipients_cols', ['Name','Email'])
        self._recipients_cols = list(cols)
        has_att = "Attachment File" in cols

        toolbar = ctk.CTkFrame(self._tab_recipients, fg_color="transparent")
        toolbar.pack(fill="x", pady=10)
        ctk.CTkButton(toolbar, text="📂  Import from Excel",
                       command=self._import_excel, width=165).pack(side="left", padx=(0,8))
        ctk.CTkButton(toolbar, text="⬇  Download Sample Excel",
                       fg_color="transparent", border_width=1,
                       text_color=("black","white"),
                       command=self._download_sample, width=185).pack(side="left", padx=(0,8))
        ctk.CTkButton(toolbar, text="🗑  Clear All",
                       fg_color="#EF4444", hover_color="#DC2626", text_color="white",
                       command=self._clear_recipients, width=100).pack(side="left", padx=(0,8))

        if has_att:
            self._folder_lbl = ctk.CTkLabel(toolbar, text="(no folder selected)",
                                              fg_color=("gray90","gray20"),
                                              corner_radius=5, width=200)
            self._folder_lbl.pack(side="left", padx=(8,6))
            ctk.CTkButton(toolbar, text="Browse Folder…", width=120,
                           fg_color="transparent", border_width=1,
                           text_color=("black","white"),
                           command=self._pick_folder).pack(side="left")

        tree_fr = ctk.CTkFrame(self._tab_recipients, fg_color="transparent")
        tree_fr.pack(fill="both", expand=True, pady=(10,0))

        style = ttk.Style()
        style.theme_use("clam")
        is_dark = ctk.get_appearance_mode() == "Dark"
        bg_c = "#1e2d3d" if is_dark else "white"
        fg_c = "white"   if is_dark else "black"
        hl_c = "#1e3a6e" if is_dark else "#DBEAFE"
        hd_c = "#1e293b" if is_dark else "#F1F5F9"
        style.configure("Custom.Treeview", background=bg_c, foreground=fg_c,
                         fieldbackground=bg_c, rowheight=30, borderwidth=0, font=F_LABEL_SML)
        style.map("Custom.Treeview",
                   background=[("selected", hl_c)], foreground=[("selected","white")])
        style.configure("Custom.Treeview.Heading",
                         background=hd_c, foreground=fg_c, font=F_BTN, padding=5)

        self._tree = ttk.Treeview(tree_fr, columns=cols, show="headings", style="Custom.Treeview")
        for col in cols:
            self._tree.heading(col, text=col)
            self._tree.column(col, width=max(140, 800 // len(cols)), minwidth=100)

        vsb = ttk.Scrollbar(tree_fr, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_fr, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_fr.rowconfigure(0, weight=1)
        tree_fr.columnconfigure(0, weight=1)

        self._tree.bind("<Control-v>",  self._on_paste)
        self._tree.bind("<Control-V>",  self._on_paste)
        self._tree.bind("<Double-1>",   self._on_double_click)
        self._tree.bind("<Delete>",     self._on_delete)
        self._tree.bind("<BackSpace>",  self._on_delete)

        bot = ctk.CTkFrame(self._tab_recipients, fg_color="transparent")
        bot.pack(fill="x", pady=(5,0))
        self._count_lbl = ctk.CTkLabel(bot, text="0 recipients", font=F_LABEL_SML)
        self._count_lbl.pack(side="left")
        ctk.CTkLabel(bot, text="Ctrl+V: Paste from Excel  |  Double-click: Edit  |  Del: Remove",
                      font=F_LABEL_SML, text_color="gray").pack(side="right")

    # ── Log tab ──────────────────────────────────────────────────

    def _build_log_tab(self):
        hdr = ctk.CTkFrame(self._tab_log, fg_color="transparent")
        hdr.pack(fill="x", pady=(10,5))
        ctk.CTkLabel(hdr, text="Send Log", font=F_BTN).pack(side="left")
        ctk.CTkButton(hdr, text="Clear Log", width=100, fg_color="transparent",
                       border_width=1, text_color=("black","white"),
                       command=self._clear_log).pack(side="right")
        self._log_box = ctk.CTkTextbox(self._tab_log, font=F_MONO, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    # ── Template management ──────────────────────────────────────

    def _reload_dropdown(self):
        names = self._store.names() or ["(no saved templates)"]
        self._tmpl_cb.configure(values=names)

    def _on_dropdown_change(self, value=None):
        names = self._store.names()
        val   = self._tmpl_var.get()
        if val in names:
            self._select_stored(names.index(val))

    def _select_stored(self, idx):
        templates = self._store.templates
        if not (0 <= idx < len(templates)):
            return
        self._active_idx = idx
        t = templates[idx]
        self._tmpl_var.set(t.name)
        self._name_var.set(t.name)
        self._subject_var.set(t.subject_tpl)
        self._body_text.delete("1.0", "end")
        self._body_text.insert("1.0", t.body_tpl)
        self._cc_var.set(t.cc)
        self._has_att_var.set(t.has_attachment)
        self._ext_var.set(t.attachment_ext or "pdf")
        self._on_att_toggle()
        cols = self._build_cols_from_phs(t.placeholders, t.has_attachment)
        self._detected_cols   = cols
        self._recipients_cols = cols
        self._cols_lbl.configure(text=("  →  ".join(cols)) if cols else "(none)")
        self._attachment_folder = ""
        self._build_recipients_tab(cols)

    def _new_template(self):
        self._active_idx = -1
        self._tmpl_var.set("")
        self._name_var.set("New Template")
        self._subject_var.set("")
        self._body_text.delete("1.0", "end")
        self._cc_var.set("")
        self._has_att_var.set(False)
        self._ext_var.set("pdf")
        self._on_att_toggle()
        self._detected_cols   = []
        self._recipients_cols = ["Name", "Email"]
        self._cols_lbl.configure(text="(type in Subject / Body above to detect columns)")
        self._build_recipients_tab(["Name", "Email"])

    def _save_template(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showwarning("Name Required", "Please enter a template name.", parent=self)
            return
        t = CustomTemplate(
            name=name,
            subject_tpl=self._subject_var.get().strip(),
            body_tpl=self._body_text.get("1.0","end").rstrip(),
            cc=self._cc_var.get().strip(),
            has_attachment=self._has_att_var.get(),
            attachment_ext=self._ext_var.get() if self._has_att_var.get() else "pdf",
        )
        if 0 <= self._active_idx < len(self._store.templates):
            self._store.update(self._active_idx, t)
        else:
            self._active_idx = self._store.add(t)
        self._reload_dropdown()
        self._tmpl_var.set(name)
        self._status_lbl.configure(text=f"✓ Template '{name}' saved.")

    def _delete_template(self):
        if not (0 <= self._active_idx < len(self._store.templates)):
            messagebox.showinfo("Nothing to Delete", "No saved template selected.", parent=self)
            return
        name = self._store.templates[self._active_idx].name
        if not messagebox.askyesno("Confirm Delete", f"Delete template '{name}'?", parent=self):
            return
        self._store.delete(self._active_idx)
        self._active_idx = -1
        self._reload_dropdown()
        if self._store.templates:
            self._select_stored(0)
        else:
            self._new_template()

    # ── Recipients helpers ───────────────────────────────────────

    def _on_paste(self, event):
        try:
            raw = self.clipboard_get()
        except tk.TclError:
            return "break"
        lines = [ln for ln in raw.splitlines() if ln.strip()]
        if not lines:
            return "break"
        cols = self._recipients_cols
        for line in lines:
            parts = line.split("\t")
            parts = (parts + [""] * len(cols))[:len(cols)]
            self._tree.insert("", "end", values=parts)
        self._update_count()
        return "break"

    def _on_delete(self, event):
        for item in self._tree.selection():
            self._tree.delete(item)
        self._update_count()

    def _on_double_click(self, event):
        region = self._tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        column = self._tree.identify_column(event.x)
        row_id = self._tree.identify_row(event.y)
        if not column or not row_id:
            return
        x, y, width, height = self._tree.bbox(row_id, column)
        val   = self._tree.set(row_id, column)
        entry = tk.Entry(self._tree, font=F_LABEL_SML)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, val)
        entry.focus()
        def _save(e=None):
            self._tree.set(row_id, column, entry.get())
            entry.destroy()
        entry.bind("<Return>",   _save)
        entry.bind("<FocusOut>", _save)

    def _clear_recipients(self):
        if self._tree.get_children():
            if messagebox.askyesno("Confirm Clear", "Clear all recipients?", parent=self):
                for item in self._tree.get_children():
                    self._tree.delete(item)
                self._update_count()

    def _pick_folder(self):
        folder = filedialog.askdirectory(title="Select Attachment Folder")
        if folder:
            self._attachment_folder = folder
            short = folder if len(folder) < 34 else "…" + folder[-31:]
            self._folder_lbl.configure(text=short)

    def _update_count(self):
        n = len(self._tree.get_children())
        self._count_lbl.configure(text=f"{n} recipient{'s' if n != 1 else ''}")

    def _get_recipients(self):
        cols, rows = self._recipients_cols, []
        for item in self._tree.get_children():
            vals = self._tree.item(item, "values")
            row  = dict(zip(cols, vals))
            if any(row.values()):
                rows.append(row)
        return rows

    def _download_sample(self):
        if not _OPENPYXL:
            messagebox.showerror("Missing Library",
                "openpyxl is not installed.\n\nRun:  pip install openpyxl", parent=self)
            return
        cols = self._recipients_cols
        if not cols:
            messagebox.showwarning("No Columns", "Apply the template to Recipients first.", parent=self)
            return
        name   = self._name_var.get().strip() or "Custom_Email"
        safe   = "".join(c for c in name if c.isalnum() or c in " _-").strip().replace(" ","_")
        fname  = f"{safe}_Recipients.xlsx"
        save_p = filedialog.asksaveasfilename(
            title="Save Sample Excel", initialfile=fname,
            defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx"),("All files","*.*")], parent=self)
        if not save_p:
            return
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recipients"
            hdr_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hdr_fill  = PatternFill("solid", fgColor="2563EB")
            ctr_align = Alignment(horizontal="center", vertical="center")
            side      = Side(style="thin", color="D1D5DB")
            border    = Border(left=side, right=side, top=side, bottom=side)
            smp_fill  = PatternFill("solid", fgColor="EFF6FF")
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = ctr_align; cell.border = border
                ws.column_dimensions[cell.column_letter].width = max(18, len(col)+4)
            sample = {"Name":"John Doe","Email":"john@example.com","Attachment File":"john_document"}
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=2, column=ci, value=sample.get(col, f"<{col}>"))
                cell.fill = smp_fill; cell.border = border
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 20
            ws.freeze_panes = "A2"
            wb.save(save_p)
            messagebox.showinfo("Saved", f"Sample Excel saved:\n{save_p}", parent=self)
        except Exception as exc:
            messagebox.showerror("Error", f"Could not save Excel:\n{exc}", parent=self)

    def _import_excel(self):
        if not _OPENPYXL:
            messagebox.showerror("Missing Library",
                "openpyxl is not installed.\n\nRun:  pip install openpyxl", parent=self)
            return
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files","*.xlsx *.xls"),("All files","*.*")])
        if not path:
            return
        try:
            wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            messagebox.showerror("Read Error", f"Could not read Excel:\n{exc}", parent=self)
            return
        if not rows:
            messagebox.showwarning("Empty File", "The Excel file has no data.", parent=self)
            return
        cols       = self._recipients_cols
        cols_lower = [c.lower() for c in cols]
        first_row  = [str(v).strip() if v is not None else "" for v in rows[0]]
        header_map = {}
        for xi, hdr in enumerate(first_row):
            if hdr.lower() in cols_lower:
                header_map[xi] = cols_lower.index(hdr.lower())
        data_rows = rows[1:] if header_map else rows
        if not header_map:
            header_map = {xi: xi for xi in range(min(len(first_row), len(cols)))}
        for item in self._tree.get_children():
            self._tree.delete(item)
        for row in data_rows:
            if not any(v for v in row if v is not None and str(v).strip()):
                continue
            vals = [""] * len(cols)
            for xi, ci in header_map.items():
                val = row[xi] if xi < len(row) else ""
                vals[ci] = "" if val is None else str(val).strip()
            self._tree.insert("", "end", values=vals)
        self._update_count()

    # ── Log helpers ──────────────────────────────────────────────

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def _log(self, message, tag="info"):
        self._log_box.configure(state="normal")
        if "[SENT]"  in message: color, tname = "#4ADE80", "sent"
        elif "[ERROR]" in message: color, tname = "#F87171", "error"
        elif "[SKIP]"  in message: color, tname = "#FBBF24", "skip"
        elif "[WARN]"  in message: color, tname = "#FB923C", "warn"
        else: color, tname = None, "info"
        if color:
            self._log_box.tag_config(tname, foreground=color)
            self._log_box.insert("end", message + "\n", tname)
        else:
            self._log_box.insert("end", message + "\n")
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _log_safe(self, msg):
        self.after(0, self._log, msg)

    # ── Preview & Check ──────────────────────────────────────────

    def _build_template_obj(self):
        return CustomTemplate(
            name=self._name_var.get().strip() or "Custom",
            subject_tpl=self._subject_var.get().strip(),
            body_tpl=self._body_text.get("1.0","end").rstrip(),
            cc=self._cc_var.get().strip(),
            has_attachment=self._has_att_var.get(),
            attachment_ext=self._ext_var.get() if self._has_att_var.get() else "pdf",
        )

    def _show_preview(self):
        t    = self._build_template_obj()
        recs = self._get_recipients()
        cfg  = {"cc": t.cc}
        sample  = recs[0] if recs else {c: f"<{c}>" for c in self._recipients_cols}
        subject = t.build_subject(cfg, sample)
        body    = t.build_body(cfg, sample)

        win = ctk.CTkToplevel(self)
        win.title("Email Preview")
        win.geometry("700x560")
        win.grab_set()

        hdr = ctk.CTkFrame(win, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text=f"✏  {t.name}", font=F_BADGE).pack(side="left", padx=20, pady=10)
        if recs:
            ctk.CTkLabel(hdr, text=f"(preview: {sample.get('Name','—')})",
                          font=F_LABEL_SML, text_color="gray").pack(side="right", padx=20)

        content = ctk.CTkFrame(win, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20, pady=10)
        ctk.CTkLabel(content, text="SUBJECT",
                      font=("Segoe UI",10,"bold"), text_color="gray").pack(anchor="w")
        ctk.CTkLabel(content, text=subject, font=F_LABEL,
                      wraplength=640, justify="left").pack(anchor="w", pady=(0,16))
        ctk.CTkLabel(content, text="BODY",
                      font=("Segoe UI",10,"bold"), text_color="gray").pack(anchor="w")
        body_box = ctk.CTkTextbox(content, font=("Segoe UI",12))
        body_box.pack(fill="both", expand=True)
        body_box.insert("end", body)
        body_box.configure(state="disabled")
        ctk.CTkButton(win, text="Close", command=win.destroy).pack(pady=10)

    def _check_data(self):
        recs = self._get_recipients()
        cols = self._recipients_cols
        win  = ctk.CTkToplevel(self)
        win.title("Check Data")
        win.geometry("740x480")
        win.grab_set()

        hdr = ctk.CTkFrame(win, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Recipient Preview", font=F_BADGE).pack(side="left", padx=20, pady=10)
        ctk.CTkLabel(hdr, text=f"{len(recs)} row(s)", font=F_BADGE).pack(side="right", padx=20)

        frame = ctk.CTkFrame(win)
        frame.pack(fill="both", expand=True, padx=20, pady=20)
        tree = ttk.Treeview(frame, columns=cols, show="headings", style="Custom.Treeview")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(90, 640 // len(cols)), minwidth=80)
        vsb = ttk.Scrollbar(frame, orient="vertical",   command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        for r in recs:
            tree.insert("", "end", values=[r.get(c,"") for c in cols])
        ctk.CTkButton(win, text="Close", command=win.destroy).pack(pady=10)

    # ── Send ─────────────────────────────────────────────────────

    def _start_send(self):
        if self._sending:
            messagebox.showwarning("Busy", "Already sending. Please wait.", parent=self)
            return
        recs = self._get_recipients()
        if not recs:
            messagebox.showwarning("No Recipients",
                "No recipients found.\n\nImport an Excel file or paste data in the Recipients tab.",
                parent=self)
            return
        t = self._build_template_obj()
        if not t.subject_tpl.strip():
            messagebox.showwarning("No Subject",
                "Please enter an email subject in the Builder tab.", parent=self)
            return
        if t.has_attachment and not self._attachment_folder:
            if not messagebox.askyesno("No Folder",
                    "No attachment folder selected.\nContinue without attachments?", parent=self):
                return
        if not messagebox.askyesno("Confirm Send",
                f"Send {len(recs)} email(s)?\nTemplate: '{t.name}'", parent=self):
            return

        self._sending = True
        self._nb.set("Log")
        self._log(f"── ✏ {t.name} — {len(recs)} recipient(s) ──")
        threading.Thread(
            target=send_custom_emails,
            args=(t, recs, self._attachment_folder, self._log_safe, self._on_done),
            daemon=True,
        ).start()

    def _on_done(self, success=True, failed=None):
        failed = failed or []
        def _show():
            self._sending = False
            if success:
                msg = "All emails sent successfully!"
                if failed:
                    msg += f"\n\nFailed ({len(failed)}):\n" + "\n".join(failed)
                self._log("── Done ──")
                messagebox.showinfo("Done", msg, parent=self)
            else:
                messagebox.showerror("Outlook Error",
                    "Could not connect to Outlook.\n\n"
                    "1. Make sure Outlook is installed and open.\n"
                    "2. The 'New Outlook' (Modern App) does not support COM automation.\n"
                    "   Please use the classic Outlook Desktop App.",
                    parent=self)
        self.after(0, _show)


if __name__ == "__main__":
    app = BulkMailApp()
    app.mainloop()
