"""
PDF Utilities – Modern Tkinter Desktop App
Layout: Header | Sidebar | Main Content | Status Bar
Supports: Dark Mode / Light Mode via ThemeManager
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, colorchooser
import tkinter.font as tkfont
import fitz          # PyMuPDF
import os
import re
import threading
import pdfplumber
import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════════
#  THEME MANAGER  –  single source of truth for every color
# ═══════════════════════════════════════════════════════════════════

class ThemeManager:
    """
    Centralised colour registry.
    Usage:  TM["accent"]  or  TM.get("hover", "#000")
    Call    TM.toggle()   then  TM.retheme(root_widget)  to switch modes.
    """

    DARK: dict = {
        # ── Backgrounds ──────────────────────────────────────
        "bg":           "#1e1e2e",   # page / window background
        "sidebar":      "#181825",   # sidebar
        "surface":      "#252538",   # cards & panels
        "surface2":     "#2a2a3e",   # inner panels / list bg
        "header_bg":    "#13131f",   # top header bar
        # ── Text ─────────────────────────────────────────────
        "text":         "#cdd6f4",   # primary
        "subtext":      "#a6adc8",   # secondary
        "text_dim":     "#6c7086",   # disabled / placeholder
        # ── Accent ───────────────────────────────────────────
        "accent":       "#7c6af7",   # brand / primary CTA
        "accent2":      "#6b5ce7",   # hover / pressed
        "accent_fg":    "#ffffff",   # text ON accent bg
        # ── Semantic ─────────────────────────────────────────
        "success":      "#a6e3a1",
        "warning":      "#f9e2af",
        "error":        "#f38ba8",
        "danger":       "#c0392b",
        "dangerh":      "#e74c3c",
        # ── Chrome ───────────────────────────────────────────
        "border":       "#313244",
        "separator":    "#45475a",
        "hover":        "#2d2d44",
        "active":       "#363654",
        # ── Inputs ───────────────────────────────────────────
        "input_bg":     "#1a1a2e",
        "input_border": "#494960",
        # ── Buttons ──────────────────────────────────────────
        "sec_btn":      "#4a4a6a",
        "sec_btnH":     "#5a5a7a",
    }

    LIGHT: dict = {
        # ── Backgrounds ──────────────────────────────────────
        "bg":           "#f0f2f9",   # soft blue-grey page — gives cards depth
        "sidebar":      "#e4e8f5",   # clearly distinct from bg
        "surface":      "#ffffff",   # pure white cards
        "surface2":     "#f4f6fb",   # inner panel / list bg
        "header_bg":    "#eaecfa",   # clean distinct header
        # ── Text ─────────────────────────────────────────────
        "text":         "#1a1a2e",   # rich near-black — high contrast
        "subtext":      "#56587a",   # clearly readable secondary
        "text_dim":     "#9090a8",   # placeholder / disabled
        # ── Accent ───────────────────────────────────────────
        "accent":       "#6c5ce7",   # deeper, richer purple
        "accent2":      "#5a4cd4",   # pressed / hover
        "accent_fg":    "#ffffff",
        # ── Semantic ─────────────────────────────────────────
        "success":      "#2da44e",
        "warning":      "#d4890a",
        "error":        "#cf222e",
        "danger":       "#b91c1c",
        "dangerh":      "#dc2626",
        # ── Chrome ───────────────────────────────────────────
        "border":       "#c8cce0",   # visible card borders
        "separator":    "#bec2d8",   # clear section dividers
        "hover":        "#eaecfa",   # gentle hover tint
        "active":       "#d8dcf5",   # pressed state
        # ── Inputs ───────────────────────────────────────────
        "input_bg":     "#f8f9fe",   # clean, distinct from surface
        "input_border": "#b8bcd4",   # clearly visible input border
        # ── Buttons ──────────────────────────────────────────
        "sec_btn":      "#6868aa",   # rich medium purple — white text AA-readable
        "sec_btnH":     "#575792",   # darker hover
    }

    def __init__(self, mode: str = "dark"):
        self.mode = mode
        self._c   = dict(self.DARK if mode == "dark" else self.LIGHT)

    def __getitem__(self, key: str) -> str:
        return self._c[key]

    def get(self, key: str, default: str = "#ff00ff") -> str:
        return self._c.get(key, default)

    def is_dark(self) -> bool:
        return self.mode == "dark"

    def toggle(self):
        """Switch mode and rebuild colour dict."""
        self.mode = "light" if self.mode == "dark" else "dark"
        self._c   = dict(self.DARK if self.mode == "dark" else self.LIGHT)

    # Structural bg keys whose mapping must win over accent/text collisions
    _BG_PRIORITY = (
        "bg", "sidebar", "surface", "surface2", "header_bg",
        "input_bg", "hover", "active", "border", "separator", "input_border",
        "sec_btn", "sec_btnH",
    )

    def retheme(self, root: tk.Widget):
        """Walk the widget tree and remap every colour to the new palette."""
        src = self.LIGHT if self.is_dark() else self.DARK
        dst = self.DARK  if self.is_dark() else self.LIGHT
        # First pass: all keys
        cmap = {src[k]: dst[k] for k in src}
        # Second pass: structural background colours win over any collision
        # (e.g. LIGHT surface="#ffffff" and accent_fg="#ffffff" both map to
        #  the same key — surface must beat accent_fg)
        for k in self._BG_PRIORITY:
            if k in src:
                cmap[src[k]] = dst[k]
        _walk_retheme(root, cmap)


# ── global singleton ─────────────────────────────────────────────
TM = ThemeManager("dark")


# ── colour-walk helpers ──────────────────────────────────────────
_WALK_OPTS = (
    "bg", "fg",
    "selectbackground", "selectforeground",
    "selectcolor",
    "activebackground", "activeforeground",
    "insertbackground", "troughcolor",
    "highlightbackground", "highlightcolor",
)

def _norm_color(widget: tk.Widget, raw: str) -> str:
    """Normalise any Tk colour string → lowercase #rrggbb."""
    try:
        r, g, b = widget.winfo_rgb(raw)
        return "#{:02x}{:02x}{:02x}".format(r >> 8, g >> 8, b >> 8)
    except Exception:
        return str(raw).lower()

def _walk_retheme(widget: tk.Widget, cmap: dict):
    """Recursively remap widget colours using cmap."""
    for opt in _WALK_OPTS:
        try:
            raw = widget.cget(opt)
            if not raw:
                continue
            n = _norm_color(widget, raw)
            if n in cmap:
                widget.configure(**{opt: cmap[n]})
        except Exception:
            pass
    if isinstance(widget, tk.Canvas):
        for item in widget.find_all():
            for opt in ("fill", "outline"):
                try:
                    raw = widget.itemcget(item, opt)
                    n   = _norm_color(widget, raw)
                    if n in cmap:
                        widget.itemconfigure(item, **{opt: cmap[n]})
                except Exception:
                    pass
    for child in widget.winfo_children():
        _walk_retheme(child, cmap)


# ═══════════════════════════════════════════════════════════════════
#  TYPOGRAPHY
# ═══════════════════════════════════════════════════════════════════

F_BODY    = ("Segoe UI", 10)
F_BOLD    = ("Segoe UI", 10, "bold")
F_H1      = ("Segoe UI", 15, "bold")
F_H2      = ("Segoe UI", 12, "bold")
F_H3      = ("Segoe UI", 10, "bold")
F_SMALL   = ("Segoe UI", 9)
F_CAPTION = ("Segoe UI", 8)
F_MONO    = ("Consolas", 9)


# ═══════════════════════════════════════════════════════════════════
#  MODERN WIDGETS
# ═══════════════════════════════════════════════════════════════════

class ModernButton(tk.Canvas):
    """
    Canvas-based button with rounded corners and smooth hover effect.
    Supports: bg, fg, hover_bg, font, padx, pady, radius, width_hint
    """

    def __init__(self, parent, text: str, command=None,
                 radius: int = 8,
                 bg: str = None, fg: str = None, hover_bg: str = None,
                 font=F_BOLD, padx: int = 18, pady: int = 8,
                 min_width: int = 0, **kw):
        self._bg      = bg      or TM["accent"]
        self._fg      = fg      or TM["accent_fg"]
        self._hover   = hover_bg or TM["accent2"]
        self._font    = font
        self._text    = text
        self._cmd     = command
        self._r       = radius
        self._padx    = padx
        self._pady    = pady
        self._pressed = False

        # measure text to auto-size
        f  = tkfont.Font(family=font[0], size=font[1],
                         weight=font[2] if len(font) > 2 else "normal")
        tw = f.measure(text)
        th = f.metrics("linespace")
        w  = max(tw + 2 * padx, min_width)
        h  = th + 2 * pady

        super().__init__(parent, width=w, height=h,
                         bg=parent["bg"], highlightthickness=0,
                         cursor="hand2", **kw)
        self._btn_w = w
        self._btn_h = h
        self._draw(self._bg)

        self.bind("<Enter>",          self._on_enter)
        self.bind("<Leave>",          self._on_leave)
        self.bind("<Button-1>",       self._on_press)
        self.bind("<ButtonRelease-1>",self._on_release)

    def _draw(self, fill: str):
        self.delete("all")
        r, w, h = self._r, self._btn_w, self._btn_h
        # Rounded rectangle via smooth polygon
        self.create_polygon(
            r,   0,   w-r, 0,
            w,   0,   w,   r,
            w,   h-r, w,   h,
            w-r, h,   r,   h,
            0,   h,   0,   h-r,
            0,   r,   0,   0,
            r,   0,
            smooth=True, fill=fill, outline=""
        )
        self.create_text(self._btn_w // 2, self._btn_h // 2, text=self._text,
                         fill=self._fg, font=self._font, anchor="center")

    def _on_enter(self, _):  self._draw(self._hover)
    def _on_leave(self, _):  self._draw(self._pressed and self._hover or self._bg)
    def _on_press(self, _):  self._pressed = True;  self._draw(self._hover)
    def _on_release(self, _):
        self._pressed = False
        self._draw(self._hover)
        if self._cmd: self._cmd()

    def configure_theme(self, bg, fg, hover):
        self._bg, self._fg, self._hover = bg, fg, hover
        self._draw(self._bg)
            

# ── Separator line ────────────────────────────────────────────────
def HSep(parent, pady=(8, 8)):
    tk.Frame(parent, bg=TM["separator"], height=1).pack(
        fill="x", padx=16, pady=pady)


# ── Section label inside a card ──────────────────────────────────
def SectionLabel(parent, text: str):
    tk.Label(parent, text=text, font=F_H3, bg=TM["surface"],
             fg=TM["accent"]).pack(anchor="w", padx=14, pady=(12, 4))


# ── Styled radio / check helpers ─────────────────────────────────
def _radio_opts(parent_bg: str) -> dict:
    """
    indicatoron=False makes the whole button a toggle chip.
    Selected  → accent bg, white text.
    Unselected → surface bg, subtext fg.
    selectcolor is the background when selected.
    """
    return dict(
        bg=parent_bg,
        fg=TM["subtext"],
        selectcolor=TM["accent"],       # chip bg when selected
        activebackground=TM["hover"],
        activeforeground=TM["text"],
        font=F_BOLD,    
        relief="flat",
        bd=0,
        padx=12,
        pady=6,
        highlightthickness=0,
        indicatoron=False,              # no circle, entire chip is the toggle
        overrelief="flat",
    )

def _check_opts(parent_bg: str) -> dict:
    """Common kwargs for tk.Checkbutton — high-visibility style."""
    return dict(
        bg=parent_bg,
        fg=TM["text"],
        selectcolor=TM["accent"],       # accent fill when checked
        activebackground=TM["hover"],
        activeforeground=TM["text"],
        font=F_BOLD,
        relief="flat",
        bd=0,
        highlightthickness=0,
    )


# ── Card container ───────────────────────────────────────────────
class Card(tk.Frame):
    """Styled surface panel with subtle border."""

    def __init__(self, parent, **kw):
        super().__init__(parent,
                         bg=TM["surface"],
                         highlightbackground=TM["border"],
                         highlightthickness=1,
                         **kw)


# ── Styled entry wrapper ─────────────────────────────────────────
class StyledEntry(tk.Frame):
    """Entry field with focus-ring border and consistent styling."""

    def __init__(self, parent, textvariable=None, width=None, **kw):
        super().__init__(parent,
                         bg=TM["input_border"],
                         highlightthickness=0, bd=0)
        inner = tk.Frame(self, bg=TM["input_bg"], bd=0)
        inner.pack(fill="both", expand=True, padx=1, pady=1)
        cfg = dict(textvariable=textvariable,
                   font=F_BODY,
                   bg=TM["input_bg"],
                   fg=TM["text"],
                   insertbackground=TM["text"],
                   relief="flat", bd=6)
        if width:
            cfg["width"] = width
        cfg.update(kw)

        self.entry = tk.Entry(inner, **cfg)
        self.entry.pack(fill="both", expand=True)

        self.entry.bind("<FocusIn>",  self._focus_in)
        self.entry.bind("<FocusOut>", self._focus_out)

    def _focus_in(self, _):
        self.configure(bg=TM["accent"])
    def _focus_out(self, _):
        self.configure(bg=TM["input_border"])

    def get(self):       return self.entry.get()
    def set(self, v):    self.entry.delete(0, "end"); self.entry.insert(0, v)
    def configure(self, **kw):
        super().configure(**kw)


# ── Top navigation tab item ──────────────────────────────────────
class TabItem(tk.Frame):
    """
    Top navigation button — full-height, bottom accent bar when active.
    """

    def __init__(self, parent, icon: str, label: str,
                 on_click, idx: int, **kw):
        super().__init__(parent, bg=TM["header_bg"],
                         cursor="hand2", **kw)
        self._idx      = idx
        self._on_click = on_click
        self._active   = False

        # container to center content vertically
        cnt = tk.Frame(self, bg=TM["header_bg"])
        cnt.pack(expand=True, fill="both", padx=16, pady=(12, 10))

        # icon
        self._icon_lbl = tk.Label(cnt, text=icon, font=("Segoe UI Emoji", 13),
                                   bg=TM["header_bg"], fg=TM["subtext"],
                                   anchor="center")
        self._icon_lbl.pack(side="left", padx=(0, 6))

        # label
        self._text_lbl = tk.Label(cnt, text=label, font=F_BOLD,
                                   bg=TM["header_bg"], fg=TM["subtext"],
                                   anchor="w")
        self._text_lbl.pack(side="left")

        # bottom accent bar
        self._bar = tk.Frame(self, bg=TM["header_bg"], height=3)
        self._bar.pack(side="bottom", fill="x")

        self._all_widgets = (self, cnt, self._icon_lbl, self._text_lbl)
        for w in self._all_widgets:
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>", self._enter)
            w.bind("<Leave>", self._leave)

    def _click(self, _): self._on_click(self._idx)

    def _bg(self): return TM["active"] if self._active else TM["header_bg"]

    def _enter(self, _):
        if not self._active:
            c = TM["hover"]
            for w in self._all_widgets:
                w.configure(bg=c)
            self._bar.configure(bg=c)

    def _leave(self, _):
        c = self._bg()
        for w in self._all_widgets:
            w.configure(bg=c)
        self._bar.configure(bg=TM["accent"] if self._active else c)

    def set_active(self, active: bool):
        self._active = active
        c = self._bg()
        for w in self._all_widgets:
            w.configure(bg=c)
        self._bar.configure(bg=TM["accent"] if active else c)
        fg = TM["text"] if active else TM["subtext"]
        self._icon_lbl.configure(fg=fg)
        self._text_lbl.configure(fg=fg)



# ═══════════════════════════════════════════════════════════════════
#  FILE LIST WIDGET  (shared by Merge & Compress tabs)
# ═══════════════════════════════════════════════════════════════════

def _ask_pdfs():
    return filedialog.askopenfilenames(title="Select PDF files",
                                       filetypes=[("PDF Files", "*.pdf")])
def _ask_pdf():
    return filedialog.askopenfilename(title="Select a PDF file",
                                      filetypes=[("PDF Files", "*.pdf")])
def _ask_save(title="Save PDF as", initial="output.pdf"):
    d = os.path.dirname(initial) if os.path.isabs(initial) else None
    f = os.path.basename(initial) if initial else "output.pdf"
    return filedialog.asksaveasfilename(
        title=title, defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")],
        initialfile=f, **({"initialdir": d} if d else {}))
def _ask_dir():
    return filedialog.askdirectory(title="Select output folder")

def _human_size(path):
    sz = os.path.getsize(path)
    for u in ("B","KB","MB","GB"):
        if sz < 1024: return f"{sz:.1f} {u}"
        sz /= 1024
    return f"{sz:.1f} TB"

def _page_count(path):
    try:
        with fitz.open(path) as d: return d.page_count
    except: return "?"


class FileListWidget(tk.Frame):
    """Reusable file list with Add / Remove / Move Up / Move Down buttons."""

    def __init__(self, parent, allow_multi=True, show_info=True, **kw):
        super().__init__(parent, bg=TM["surface2"], **kw)
        self.allow_multi = allow_multi
        self.show_info   = show_info
        self.files: list = []
        self._build()

    def _build(self):
        # listbox area
        lb_f = tk.Frame(self,
                        bg=TM["input_border"],
                        highlightthickness=0)
        lb_f.pack(fill="both", expand=True)

        inner = tk.Frame(lb_f, bg=TM["input_bg"])
        inner.pack(fill="both", expand=True, padx=1, pady=1)

        self.lb = tk.Listbox(inner,
                              bg=TM["input_bg"], fg=TM["text"],
                              selectbackground=TM["accent"],
                              selectforeground=TM["accent_fg"],
                              font=F_BODY, bd=0,
                              highlightthickness=0,
                              activestyle="none")
        sb = tk.Scrollbar(inner, orient="vertical", command=self.lb.yview,
                          bg=TM["surface2"], troughcolor=TM["surface2"],
                          relief="flat", bd=0)
        self.lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.lb.pack(fill="both", expand=True, padx=4, pady=4)

        # toolbar
        bar = tk.Frame(self, bg=TM["surface2"])
        bar.pack(fill="x", pady=(6, 0))

        def _btn(parent, text, cmd, bg, hover):
            b = ModernButton(parent, text, cmd, bg=bg, hover_bg=hover,
                             font=F_SMALL, padx=10, pady=5, radius=6)
            b.pack(side="left", padx=(0, 4))
            return b

        _btn(bar, "+ Add",    self._add,    TM["accent"],   TM["accent2"])
        _btn(bar, "✕ Remove", self._remove, TM["danger"],   TM["dangerh"])
        _btn(bar, "▲ Page Up",   self._up,     TM["sec_btn"],  TM["sec_btnH"])
        _btn(bar, "▼ Page Down", self._down,   TM["sec_btn"],  TM["sec_btnH"])

        clr = ModernButton(bar, "Clear", self._clear,
                           bg=TM["sec_btn"], hover_bg=TM["sec_btnH"],
                           font=F_SMALL, padx=10, pady=5, radius=6)
        clr.pack(side="right")

    def _label(self, p):
        name = os.path.basename(p)
        if self.show_info:
            return f"  {name}   [{_page_count(p)} pages · {_human_size(p)}]"
        return f"  {name}"

    def _add(self):
        paths = _ask_pdfs() if self.allow_multi else (lambda p: (p,) if p else ())(_ask_pdf())
        for p in paths:
            if p and p not in self.files:
                self.files.append(p)
                self.lb.insert("end", self._label(p))

    def _remove(self):
        for i in reversed(self.lb.curselection()):
            self.lb.delete(i); self.files.pop(i)

    def _up(self):
        sel = self.lb.curselection()
        if not sel or sel[0] == 0: return
        i = sel[0]
        self.files[i], self.files[i-1] = self.files[i-1], self.files[i]
        self._refresh(); self.lb.select_set(i-1)

    def _down(self):
        sel = self.lb.curselection()
        if not sel or sel[0] >= len(self.files)-1: return
        i = sel[0]
        self.files[i], self.files[i+1] = self.files[i+1], self.files[i]
        self._refresh(); self.lb.select_set(i+1)

    def _clear(self):
        self.files.clear(); self.lb.delete(0, "end")

    def _refresh(self):
        self.lb.delete(0, "end")
        for p in self.files: self.lb.insert("end", self._label(p))


# ═══════════════════════════════════════════════════════════════════
#  STATUS BAR
# ═══════════════════════════════════════════════════════════════════

class StatusBar(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=TM["sidebar"], height=30)
        self.pack(fill="x", side="bottom")
        self.pack_propagate(False)

        self._lbl = tk.Label(self, text="Ready", font=F_CAPTION,
                              bg=TM["sidebar"], fg=TM["subtext"], anchor="w")
        self._lbl.pack(side="left", padx=16, fill="x", expand=True)

        from tkinter import ttk
        s = ttk.Style()
        s.configure("SB.Horizontal.TProgressbar",
                    troughcolor=TM["sidebar"], background=TM["accent"],
                    thickness=3)
        self._pb = ttk.Progressbar(self, mode="indeterminate",
                                    style="SB.Horizontal.TProgressbar",
                                    length=160)

    def set(self, msg, color=None):
        self._lbl.configure(text=msg, fg=color or TM["subtext"])
        self.update_idletasks()

    def busy(self, msg="Working…"):
        self.set(msg, TM["warning"])
        self._pb.pack(side="right", padx=12, pady=4)
        self._pb.start(10)

    def done(self, msg="Done", color=None):
        self._pb.stop(); self._pb.pack_forget()
        self.set(msg, color or TM["success"])


# ═══════════════════════════════════════════════════════════════════
#  TAB PAGES
# ═══════════════════════════════════════════════════════════════════

# ── shared helpers ───────────────────────────────────────────────

def _page_title(parent, title: str, subtitle: str):
    tk.Label(parent, text=title, font=F_H1,
             bg=TM["bg"], fg=TM["text"]).pack(anchor="w", padx=20, pady=(18, 2))
    tk.Label(parent, text=subtitle, font=F_SMALL,
             bg=TM["bg"], fg=TM["subtext"]).pack(anchor="w", padx=20, pady=(0, 12))

def _src_row(parent, var, info_var=None, on_browse=None):
    """Standard source-file input row."""
    row = Card(parent)
    row.pack(fill="x", padx=20, pady=(0, 10))

    tk.Label(row, text="Source PDF:", font=F_BOLD,
             bg=TM["surface"], fg=TM["text"],
             padx=14, pady=10).pack(side="left")

    e = tk.Entry(row, textvariable=var, font=F_BODY,
                 bg=TM["input_bg"], fg=TM["text"],
                 insertbackground=TM["text"], relief="flat", bd=4)
    e.pack(side="left", fill="x", expand=True, pady=10)

    b = ModernButton(row, "Browse…", on_browse,
                     bg=TM["sec_btn"], hover_bg=TM["sec_btnH"],
                     font=F_SMALL, padx=12, pady=5, radius=6)
    b.pack(side="right", padx=10, pady=10)
    return row

def _resolve_pages(spec: str, n: int) -> list:
    if spec.strip().lower() == "all":
        return list(range(n))
    pages = set()
    for p in spec.split(","):
        p = p.strip()
        if "-" in p:
            a, b = p.split("-", 1)
            pages.update(range(int(a)-1, int(b)))
        elif p:
            pages.add(int(p)-1)
    return sorted(x for x in pages if 0 <= x < n)


# ─────────────────────────────────────────────────────────────────
#  MERGE TAB
# ─────────────────────────────────────────────────────────────────

class MergeTab(tk.Frame):

    def __init__(self, parent, status: StatusBar):
        super().__init__(parent, bg=TM["bg"])
        self.status = status
        self._build()

    def _build(self):
        _page_title(self, "⊕  Merge PDF Files",
                    "Add files, reorder with ▲▼, configure options then merge.")

        body = tk.Frame(self, bg=TM["bg"])
        body.pack(fill="both", expand=True, padx=20, pady=(0, 12))

        # ── left: file list ──────────────────────────────────
        left = tk.Frame(body, bg=TM["bg"])
        left.pack(side="left", fill="both", expand=True, padx=(0, 10))

        tk.Label(left, text="Files  (reorder with Page Up/Down)", font=F_H3,
                 bg=TM["bg"], fg=TM["text"]).pack(anchor="w", pady=(0, 6))

        self.fl = FileListWidget(left)
        self.fl.pack(fill="both", expand=True)

        # ── right: settings card ─────────────────────────────
        right = Card(body)
        right.pack(side="right", fill="y", ipadx=4, ipady=4)

        def sec(t): SectionLabel(right, t)


        # ── action button ─────────────────────────────────────
        btn_row = tk.Frame(self, bg=TM["bg"])
        btn_row.pack(pady=8)
        ModernButton(btn_row, "  ⊕  Merge PDFs  ", self._merge,
                     min_width=180, pady=10).pack()

    def _merge(self):
        if not self.fl.files:
            messagebox.showwarning("No Files", "Add at least one PDF file."); return
        import time
        out = os.path.join(os.path.dirname(self.fl.files[0]), f"merged_output_{int(time.time())}.pdf")
        
        self.status.busy("Merging PDFs…")

        def run():
            try:
                merged = fitz.open()
                for path in self.fl.files:
                    src = fitz.open(path)
                    merged.insert_pdf(src)
                    src.close()
                merged.save(out, garbage=4, deflate=True)
                merged.close()
                self.status.done(f"Merged → {os.path.basename(out)}")
                messagebox.showinfo("Done", f"Saved to:\n{out}")
            except Exception as e:
                self.status.done(str(e), TM["error"])
                messagebox.showerror("Error", str(e))

        threading.Thread(target=run, daemon=True).start()


# ─────────────────────────────────────────────────────────────────
#  SPLIT TAB
# ─────────────────────────────────────────────────────────────────

class SplitTab(tk.Frame):

    def __init__(self, parent, status: StatusBar):
        super().__init__(parent, bg=TM["bg"])
        self.status   = status
        self.src_path = tk.StringVar()
        self._build()

    def _build(self):
        _page_title(self, "✂  Split PDF",
                    "Split by custom ranges, every N pages, or individual pages.")

        _src_row(self, self.src_path, on_browse=self._browse)

        self._info = tk.Label(self, text="", font=F_SMALL,
                               bg=TM["bg"], fg=TM["subtext"])
        self._info.pack(anchor="w", padx=20, pady=(0, 8))

        # mode card
        mode_card = Card(self)
        mode_card.pack(fill="x", padx=20, pady=(0, 10))

        self.mode = tk.StringVar(value="range")
        modes = [("Custom page ranges  (e.g. 1-3, 5, 7-9)", "range"),
                 ("Every N pages", "every"),
                 ("Individual pages (one PDF per page)", "individual")]
        mode_row = tk.Frame(mode_card, bg=TM["surface"])
        mode_row.pack(anchor="w", padx=16, pady=(4, 12))
        for text, val in modes:
            tk.Radiobutton(mode_row, text=text, variable=self.mode,
                           value=val, command=self._mode_changed,
                           **_radio_opts(TM["surface"])
                           ).pack(side="left", padx=(0, 6))

        # options area
        self._opt = tk.Frame(self, bg=TM["bg"])
        self._opt.pack(fill="x", padx=20, pady=(0, 8))

        # range frame
        self._rf = tk.Frame(self._opt, bg=TM["bg"])
        tk.Label(self._rf, text="Ranges:", font=F_BOLD,
                 bg=TM["bg"], fg=TM["text"]).pack(side="left", padx=(0, 8))
        self._range_var = tk.StringVar(value="1-3, 4-6")
        tk.Entry(self._rf, textvariable=self._range_var, font=F_BODY,
                 bg=TM["input_bg"], fg=TM["text"],
                 insertbackground=TM["text"],
                 relief="flat", bd=5, width=28).pack(side="left")

        # every-n frame
        self._ef = tk.Frame(self._opt, bg=TM["bg"])
        tk.Label(self._ef, text="Every N pages:", font=F_BOLD,
                 bg=TM["bg"], fg=TM["text"]).pack(side="left", padx=(0, 8))
        self._every = tk.IntVar(value=1)
        tk.Spinbox(self._ef, from_=1, to=9999, textvariable=self._every,
                   font=F_BODY, bg=TM["input_bg"], fg=TM["text"],
                   buttonbackground=TM["input_border"],
                   relief="flat", bd=4, width=8).pack(side="left")

        # output dir removed to automatically save in the same dir

        btn_row = tk.Frame(self, bg=TM["bg"])
        btn_row.pack(pady=8)
        ModernButton(btn_row, "  ✂  Split PDF  ", self._split,
                     min_width=160, pady=10).pack()

        self._mode_changed()

    def _browse(self):
        p = _ask_pdf()
        if p:
            self.src_path.set(p)
            with fitz.open(p) as d:
                self._info.configure(
                    text=f"  {d.page_count} pages · {_human_size(p)}")

    def _mode_changed(self):
        self._rf.pack_forget(); self._ef.pack_forget()
        m = self.mode.get()
        if m == "range":  self._rf.pack(anchor="w")
        elif m == "every": self._ef.pack(anchor="w")

    def _split(self):
        src = self.src_path.get().strip()
        if not src or not os.path.isfile(src):
            messagebox.showwarning("No File", "Select a source PDF."); return
        out_dir = os.path.dirname(src)
        
        os.makedirs(out_dir, exist_ok=True)
        mode = self.mode.get()
        self.status.busy("Splitting…")

        def run():
            try:
                base = os.path.splitext(os.path.basename(src))[0]
                doc  = fitz.open(src); n = doc.page_count; saved = 0

                def _save(pg_from, pg_to, label):
                    nonlocal saved
                    o = fitz.open()
                    o.insert_pdf(doc, from_page=pg_from, to_page=pg_to)
                    o.save(os.path.join(out_dir, label), garbage=4, deflate=True)
                    o.close(); saved += 1

                if mode == "individual":
                    for i in range(n):
                        _save(i, i, f"{base}_page{i+1:04d}.pdf")
                elif mode == "every":
                    step = max(1, self._every.get())
                    for s in range(0, n, step):
                        e = min(s+step-1, n-1)
                        _save(s, e, f"{base}_pages{s+1}-{e+1}.pdf")
                else:
                    for idx, part in enumerate(self._range_var.get().split(","), 1):
                        part = part.strip()
                        if "-" in part:
                            a, b = [int(x)-1 for x in part.split("-",1)]
                        else:
                            a = b = int(part)-1
                        a, b = max(0,a), min(b, n-1)
                        _save(a, b, f"{base}_part{idx}_{a+1}-{b+1}.pdf")

                doc.close()
                self.status.done(f"Split → {saved} files in {out_dir}")
                messagebox.showinfo("Done", f"{saved} file(s) → {out_dir}")
            except Exception as e:
                self.status.done(str(e), TM["error"])
                messagebox.showerror("Error", str(e))

        threading.Thread(target=run, daemon=True).start()


# ─────────────────────────────────────────────────────────────────
#  EXTRACTOR TAB
# ─────────────────────────────────────────────────────────────────

class ExtractorTab(tk.Frame):

    def __init__(self, parent, status: StatusBar):
        super().__init__(parent, bg=TM["bg"])
        self.status   = status
        self.src_path = tk.StringVar()
        self._build()

    def _build(self):
        _page_title(self, "⊙  PDF Extractor",
                    "Extract text, images, or specific pages from a PDF.")

        _src_row(self, self.src_path, on_browse=self._browse)
        self._info = tk.Label(self, text="", font=F_SMALL,
                               bg=TM["bg"], fg=TM["subtext"])
        self._info.pack(anchor="w", padx=20, pady=(0, 8))

        # ── inline sub-tab bar ───────────────────────────────
        sub_bar = tk.Frame(self, bg=TM["sidebar"])
        sub_bar.pack(fill="x", padx=20)
        tk.Frame(self, bg=TM["separator"], height=1).pack(fill="x", padx=20)

        self._sub_content = tk.Frame(self, bg=TM["bg"])
        self._sub_content.pack(fill="both", expand=True, padx=20, pady=8)

        self._sub_btns  = []
        self._sub_inds  = []
        self._sub_pages = []
        self._active_sub = -1

        def _show_sub(idx):
            for f in self._sub_pages: f.pack_forget()
            for i,(b,ind) in enumerate(zip(self._sub_btns, self._sub_inds)):
                on = (i == idx)
                b.configure(fg=TM["text"] if on else TM["subtext"])
                if on: ind.pack(fill="x", side="bottom")
                else:  ind.pack_forget()
            self._sub_pages[idx].pack(fill="both", expand=True)
            self._active_sub = idx

        def _make_sub(label):
            wrap = tk.Frame(sub_bar, bg=TM["sidebar"])
            wrap.pack(side="left", fill="y")
            b = tk.Button(wrap, text=f"  {label}  ", font=F_H3,
                          bg=TM["sidebar"], fg=TM["subtext"],
                          relief="flat", bd=0, padx=4, cursor="hand2",
                          activebackground=TM["sidebar"],
                          activeforeground=TM["text"],
                          command=lambda i=len(self._sub_btns): _show_sub(i))
            b.pack(fill="both", expand=True)
            ind = tk.Frame(wrap, bg=TM["accent"], height=2)
            b.bind("<Enter>", lambda e,_b=b,_i=len(self._sub_btns):
                   (_b.configure(fg=TM["text"]) if _i!=self._active_sub else None))
            b.bind("<Leave>", lambda e,_b=b,_i=len(self._sub_btns):
                   (_b.configure(fg=TM["subtext"]) if _i!=self._active_sub else None))
            self._sub_btns.append(b); self._sub_inds.append(ind)
            frm = tk.Frame(self._sub_content, bg=TM["bg"])
            self._sub_pages.append(frm)
            return frm

        # ── TEXT sub-tab ──────────────────────────────────────
        tt = _make_sub("Extract Text")

        r1 = tk.Frame(tt, bg=TM["bg"]); r1.pack(fill="x", pady=4)
        tk.Label(r1, text="Pages:", font=F_BOLD,
                 bg=TM["bg"], fg=TM["text"]).pack(side="left", padx=(0,8))
        self._txt_pg = tk.StringVar(value="all")
        tk.Entry(r1, textvariable=self._txt_pg, font=F_BODY,
                 bg=TM["input_bg"], fg=TM["text"],
                 insertbackground=TM["text"], relief="flat", bd=5,
                 width=20).pack(side="left")
        tk.Label(r1, text="  all · 1,3 · 2-5", font=F_SMALL,
                 bg=TM["bg"], fg=TM["subtext"]).pack(side="left")

        self._txt_fmt = tk.StringVar(value="txt")
        fmt_row = tk.Frame(tt, bg=TM["bg"]); fmt_row.pack(anchor="w", pady=(4,6))
        for val, lbl in (("txt","Plain text"), ("blocks","Text blocks")):
            tk.Radiobutton(fmt_row, text=lbl, variable=self._txt_fmt, value=val,
                           **_radio_opts(TM["bg"])
                           ).pack(side="left", padx=(0,6))

        self._preview = scrolledtext.ScrolledText(
            tt, font=F_MONO, bg=TM["surface2"], fg=TM["text"],
            insertbackground=TM["text"], relief="flat", bd=0, height=7)
        self._preview.pack(fill="both", expand=True, pady=8)

        tr = tk.Frame(tt, bg=TM["bg"]); tr.pack(fill="x")
        ModernButton(tr, "Preview", self._preview_text,
                     bg=TM["sec_btn"], hover_bg=TM["sec_btnH"],
                     font=F_SMALL, padx=14, pady=6, radius=6).pack(side="left", padx=(0,6))
        ModernButton(tr, "Save .txt", self._save_text,
                     font=F_SMALL, padx=14, pady=6, radius=6).pack(side="left")

        # ── IMAGES sub-tab ────────────────────────────────────
        it = _make_sub("Extract Images")

        r2 = tk.Frame(it, bg=TM["bg"]); r2.pack(fill="x", pady=4)
        tk.Label(r2, text="Pages:", font=F_BOLD,
                 bg=TM["bg"], fg=TM["text"]).pack(side="left", padx=(0,8))
        self._img_pg = tk.StringVar(value="all")
        tk.Entry(r2, textvariable=self._img_pg, font=F_BODY,
                 bg=TM["input_bg"], fg=TM["text"],
                 insertbackground=TM["text"], relief="flat", bd=5,
                 width=20).pack(side="left")

        fr = tk.Frame(it, bg=TM["bg"]); fr.pack(anchor="w", pady=(6,8))
        tk.Label(fr, text="Format:", font=F_BOLD,
                 bg=TM["bg"], fg=TM["text"]).pack(side="left", padx=(0,8))
        self._img_fmt = tk.StringVar(value="png")
        for f in ("png","jpeg"):
            tk.Radiobutton(fr, text=f.upper(), variable=self._img_fmt, value=f,
                           **_radio_opts(TM["bg"])
                           ).pack(side="left", padx=(0,6))

        ModernButton(it, "  Extract Images  ", self._extract_images,
                     font=F_SMALL, padx=14, pady=6, radius=6).pack(anchor="w")

        # ── PAGES sub-tab ─────────────────────────────────────
        pt = _make_sub("Extract Pages")

        tk.Label(pt, text="Page ranges:", font=F_BOLD,
                 bg=TM["bg"], fg=TM["text"]).pack(anchor="w", pady=(6,2))
        self._pg_ranges = tk.StringVar(value="1-3, 5, 7-9")
        tk.Entry(pt, textvariable=self._pg_ranges, font=F_BODY,
                 bg=TM["input_bg"], fg=TM["text"],
                 insertbackground=TM["text"], relief="flat", bd=5).pack(
            fill="x", pady=(0,8))

        ModernButton(pt, "  Extract Pages  ", self._extract_pages,
                     font=F_SMALL, padx=14, pady=6, radius=6).pack(anchor="w", pady=4)

        _show_sub(0)

    def _browse(self):
        p = _ask_pdf()
        if p:
            self.src_path.set(p)
            with fitz.open(p) as d:
                self._info.configure(
                    text=f"  {d.page_count} pages · {_human_size(p)}")

    def _preview_text(self):
        src = self.src_path.get().strip()
        if not src: messagebox.showwarning("No File","Select a PDF."); return
        try:
            doc = fitz.open(src)
            pages = _resolve_pages(self._txt_pg.get(), doc.page_count)
            text = ""
            for i in pages:
                text += f"\n── Page {i+1} ──\n"
                if self._txt_fmt.get() == "blocks":
                    text += "\n".join(b[4] for b in doc[i].get_text("blocks") if b[6]==0)
                else:
                    text += doc[i].get_text()
            doc.close()
            self._preview.delete("1.0","end")
            self._preview.insert("1.0", text.strip())
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _save_text(self):
        src = self.src_path.get().strip()
        if not src: messagebox.showwarning("No File","Select a PDF."); return
        import time
        out_dir = os.path.dirname(src)
        out = os.path.join(out_dir, f"extracted_text_{int(time.time())}.txt")
        self.status.busy("Extracting text…")
        def run():
            try:
                doc = fitz.open(src)
                pages = _resolve_pages(self._txt_pg.get(), doc.page_count)
                with open(out,"w",encoding="utf-8") as f:
                    for i in pages:
                        f.write(f"\n{'─'*60}\nPage {i+1}\n{'─'*60}\n")
                        f.write(doc[i].get_text())
                doc.close()
                self.status.done(f"Saved → {os.path.basename(out)}")
                messagebox.showinfo("Done", f"Saved to:\n{out}")
            except Exception as e:
                self.status.done(str(e), TM["error"]); messagebox.showerror("Error",str(e))
        threading.Thread(target=run, daemon=True).start()

    def _extract_images(self):
        src = self.src_path.get().strip()
        if not src: messagebox.showwarning("No File","Select a PDF."); return
        import time
        out_dir = os.path.join(os.path.dirname(src), f"extracted_images_{int(time.time())}")
        os.makedirs(out_dir, exist_ok=True)
        fmt = self._img_fmt.get()
        self.status.busy("Extracting images…")

        def _pil_save(raw_bytes, out_path, target_fmt):
            """Convert raw image bytes to target format using PIL."""
            from PIL import Image
            import io
            img = Image.open(io.BytesIO(raw_bytes))
            if target_fmt == "png":
                if img.mode not in ("RGB", "RGBA", "L", "LA", "P"):
                    img = img.convert("RGB")
                img.save(out_path, "PNG")
            else:  # jpeg
                if img.mode in ("RGBA", "LA"):
                    img = img.convert("RGB")
                elif img.mode == "CMYK":
                    img = img.convert("RGB")
                elif img.mode == "P":
                    img = img.convert("RGB")
                elif img.mode not in ("RGB", "L"):
                    img = img.convert("RGB")
                img.save(out_path, "JPEG", quality=92)

        def _fitz_save(doc, xref, out_path, target_fmt):
            """Convert pixmap using PyMuPDF as fallback."""
            pix = fitz.Pixmap(doc, xref)
            cs = pix.colorspace
            cs_name = getattr(cs, "name", "") if cs else ""
            safe = {"DeviceGray", "DeviceRGB"}
            if cs_name not in safe:
                pix = fitz.Pixmap(fitz.csRGB, pix)
            if pix.alpha:
                pix = fitz.Pixmap(pix, 0)
            pix.save(out_path)

        def run():
            try:
                from PIL import Image
                import io
                has_pil = True
            except ImportError:
                has_pil = False
            try:
                doc = fitz.open(src)
                pages = _resolve_pages(self._img_pg.get(), doc.page_count)
                count = 0
                save_ext = "jpg" if fmt == "jpeg" else fmt
                for i in pages:
                    for j, img in enumerate(doc.get_page_images(i, full=True)):
                        xref = img[0]
                        out_path = os.path.join(out_dir,
                                    f"page{i+1:04d}_img{j+1:03d}.{save_ext}")
                        saved = False
                        # PIL approach: most robust (handles ICC, CMYK, etc.)
                        if has_pil:
                            try:
                                img_data = doc.extract_image(xref)
                                _pil_save(img_data["image"], out_path, fmt)
                                saved = True
                                count += 1
                            except Exception:
                                pass
                        # PyMuPDF fallback
                        if not saved:
                            try:
                                _fitz_save(doc, xref, out_path, fmt)
                                count += 1
                            except Exception:
                                pass
                doc.close()
                self.status.done(f"Extracted {count} image(s)")
                messagebox.showinfo("Done", f"{count} image(s) → {out_dir}")
            except Exception as e:
                self.status.done(str(e), TM["error"]); messagebox.showerror("Error",str(e))
        threading.Thread(target=run, daemon=True).start()

    def _extract_pages(self):
        src = self.src_path.get().strip()
        if not src: messagebox.showwarning("No File","Select a PDF."); return
        import time
        out = os.path.join(os.path.dirname(src), f"extracted_pages_{int(time.time())}.pdf")
        self.status.busy("Extracting pages…")
        def run():
            try:
                doc   = fitz.open(src)
                pages = _resolve_pages(self._pg_ranges.get(), doc.page_count)
                o = fitz.open()
                for i in pages: o.insert_pdf(doc, from_page=i, to_page=i)
                o.save(out, garbage=4, deflate=True); o.close(); doc.close()
                self.status.done(f"Extracted {len(pages)} page(s)")
                messagebox.showinfo("Done", f"{len(pages)} page(s) → {out}")
            except Exception as e:
                self.status.done(str(e), TM["error"]); messagebox.showerror("Error",str(e))
        threading.Thread(target=run, daemon=True).start()


# ─────────────────────────────────────────────────────────────────
#  COMPRESS TAB
# ─────────────────────────────────────────────────────────────────

class CompressorTab(tk.Frame):

    def __init__(self, parent, status: StatusBar):
        super().__init__(parent, bg=TM["bg"])
        self.status = status
        self._build()

    def _build(self):
        _page_title(self, "⊜  PDF Compressor",
                    "Reduce file size by compressing images and cleaning unused objects.")

        # Pack bottom elements FIRST so they're always visible
        btn_row = tk.Frame(self, bg=TM["bg"])
        btn_row.pack(side="bottom", pady=8)
        ModernButton(btn_row, "  ⊜  Compress PDFs  ", self._compress,
                     min_width=180, pady=10).pack()

        self._results = scrolledtext.ScrolledText(
            self, font=F_MONO, bg=TM["surface2"], fg=TM["text"],
            insertbackground=TM["text"], relief="flat", bd=0, height=4)
        self._results.pack(side="bottom", fill="x", padx=20, pady=(0, 10))

        body = tk.Frame(self, bg=TM["bg"])
        body.pack(fill="both", expand=True, padx=20)

        # file list – listbox fixed to 5 rows so toolbar + settings always visible
        tk.Label(body, text="Files to compress", font=F_H3,
                 bg=TM["bg"], fg=TM["text"]).pack(anchor="w", pady=(0,4))
        self.fl = FileListWidget(body)
        self.fl.lb.configure(height=5)
        self.fl.pack(fill="x")

        # settings card
        cfg = Card(body)
        cfg.pack(fill="x", pady=(10, 0))

        # Compression level
        SectionLabel(cfg, "Compression Level")
        lrow = tk.Frame(cfg, bg=TM["surface"]); lrow.pack(fill="x", padx=14, pady=(0,10))
        self._level = tk.StringVar(value="balanced")
        for lbl, val in (("🏆  Best Quality", "quality"),
                         ("⚖  Balanced",      "balanced"),
                         ("📦  Smallest Size", "smallest")):
            tk.Radiobutton(lrow, text=lbl, variable=self._level, value=val,
                           **_radio_opts(TM["surface"])
                           ).pack(side="left", padx=(0,6))

        self._out_dir = tk.StringVar(value="")

    def _compress(self):
        if not self.fl.files:
            messagebox.showwarning("No Files","Add files first."); return
        level = self._level.get()
        if level == "balanced":
            jpeg_q, max_px = 65, 0
        elif level == "smallest":
            jpeg_q, max_px = 40, 1200
        else:
            jpeg_q, max_px = 0, 0
        out_dir = self._out_dir.get().strip()
        self._results.delete("1.0","end")
        self.status.busy(f"Compressing {len(self.fl.files)} file(s)…")

        def run():
            import pikepdf
            from PIL import Image
            import io
            out = []
            for path in self.fl.files:
                try:
                    fname = os.path.basename(path)
                    if out_dir:
                        os.makedirs(out_dir, exist_ok=True)
                        dst = os.path.join(out_dir, fname)
                    else:
                        base, ext = os.path.splitext(path)
                        dst = base + "_compressed" + ext

                    pdf = pikepdf.open(path)

                    if level != "quality":
                        seen = set()
                        for page in pdf.pages:
                            try:
                                xobjs = page.resources.get("/XObject", {})
                            except Exception:
                                continue
                            for key in xobjs:
                                try:
                                    xobj = xobjs[key]
                                    if xobj.get("/Subtype") != "/Image":
                                        continue
                                    objgen = xobj.objgen
                                    if objgen in seen:
                                        continue
                                    seen.add(objgen)
                                    w = int(xobj["/Width"])
                                    h = int(xobj["/Height"])
                                    if w < 100 or h < 100:
                                        continue
                                    # Skip unsupported image filters (hang risk)
                                    raw_filter = xobj.get("/Filter")
                                    if raw_filter is not None:
                                        f = str(raw_filter)
                                        if any(x in f for x in ("JBIG2", "CCITTFax", "JBIG2Decode", "CCITTFaxDecode")):
                                            continue
                                    # Decode with pikepdf PdfImage → PIL
                                    pil_img = pikepdf.PdfImage(xobj).as_pil_image()
                                    orig_size = len(xobj.read_raw_bytes())
                                    # Normalise colorspace
                                    if pil_img.mode in ("RGBA", "LA", "P"):
                                        pil_img = pil_img.convert("RGB")
                                    elif pil_img.mode == "CMYK":
                                        pil_img = pil_img.convert("RGB")
                                    elif pil_img.mode not in ("RGB", "L"):
                                        pil_img = pil_img.convert("RGB")

                                    # Downscale for Smallest mode
                                    if max_px > 0 and max(w, h) > max_px:
                                        scale = max_px / max(w, h)
                                        pil_img = pil_img.resize(
                                            (max(1, int(w*scale)),
                                             max(1, int(h*scale))),
                                            Image.LANCZOS)

                                    buf = io.BytesIO()
                                    pil_img.save(buf, "JPEG", quality=jpeg_q, optimize=True)
                                    new_bytes = buf.getvalue()

                                    # Only replace if genuinely smaller
                                    if len(new_bytes) < orig_size:
                                        xobj.write(new_bytes, filter=pikepdf.Name("/DCTDecode"))
                                        cs = (pikepdf.Name("/DeviceGray")
                                              if pil_img.mode == "L"
                                              else pikepdf.Name("/DeviceRGB"))
                                        xobj["/ColorSpace"] = cs
                                        if "/DecodeParms" in xobj:
                                            del xobj["/DecodeParms"]
                                except Exception:
                                    pass

                    pdf.save(dst,
                             compress_streams=True,
                             recompress_flate=True,
                             object_stream_mode=pikepdf.ObjectStreamMode.generate)
                    pdf.close()
                    before = os.path.getsize(path)
                    after  = os.path.getsize(dst)
                    ratio  = (1 - after / before) * 100 if before else 0
                    out.append(f"✓ {fname}\n"
                               f"  {_human_size(path)} → {_human_size(dst)}"
                               f"  ({ratio:+.1f}% reduction)\n"
                               f"  Saved to: {dst}\n\n")
                except Exception as e:
                    out.append(f"✗ {os.path.basename(path)}: {e}\n\n")
            self._results.insert("end", "".join(out))
            self.status.done(f"Compression complete · {len(self.fl.files)} file(s)")

        threading.Thread(target=run, daemon=True).start()


# ─────────────────────────────────────────────────────────────────
#  REDACT TAB  –  Challan Extractor Style
# ─────────────────────────────────────────────────────────────────

# Colour palette for the Redact tab (self-contained)
_RC = {
    "bg":      "#0d1117",
    "panel":   "#161b22",
    "header":  "#1f2937",
    "accent1": "#e94560",
    "accent2": "#f59e0b",
    "accent3": "#10b981",
    "accent4": "#6366f1",
    "accent5": "#ec4899",
    "accent6": "#06b6d4",
    "txt":     "#f0f6fc",
    "sub":     "#8b949e",
    "border":  "#30363d",
    "canvas":  "#21262d",
    "redact":  "#000000",
}


class RedactTab(tk.Frame):

    def __init__(self, parent, status: "StatusBar"):
        super().__init__(parent, bg=_RC["bg"])
        self.status = status

        # ── App state ──
        self.pdf_path     = None
        self.pdf_doc      = None
        self.current_page = 0
        self.total_pages  = 0
        self.zoom         = 1.5
        self.redactions   = {}   # {page_index: [fitz.Rect, ...]}

        # ── Drawing state ──
        self.drawing    = False
        self.draw_start = (0, 0)
        self.temp_id    = None
        self.img_x0     = 10
        self.img_y0     = 10
        self._tk_img    = None

        # ── Auto-redact checkbox variables ──
        self.cb_gstn   = tk.BooleanVar()
        self.cb_pan    = tk.BooleanVar()
        self.cb_name   = tk.BooleanVar()
        self.cb_notice = tk.BooleanVar()
        self.cb_date   = tk.BooleanVar()

        self._build_ui()
        self.status.set("Redact Tab ready  |  Open a PDF file to begin")

    # ═══════════════════════════════  UI BUILD  ════════════════════════════

    def _build_ui(self):
        body = tk.Frame(self, bg=_RC["bg"])
        body.pack(fill="both", expand=True)

        sidebar = tk.Frame(body, bg=_RC["panel"], width=248)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        self._build_sidebar(sidebar)

        tk.Frame(body, bg=_RC["border"], width=1).pack(side="left", fill="y")

        viewer = tk.Frame(body, bg=_RC["canvas"])
        viewer.pack(side="left", fill="both", expand=True)
        self._build_viewer(viewer)

    # ── Sidebar ───────────────────────────────────────────────────────────

    def _build_sidebar(self, parent):
        self._scrolled_sidebar(parent)

    def _scrolled_sidebar(self, parent):
        from tkinter import ttk as _ttk
        canvas = tk.Canvas(parent, bg=_RC["panel"], highlightthickness=0)
        scroll = _ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)

        scroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=_RC["panel"])
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_resize(e):
            canvas.itemconfig(win_id, width=e.width)
        canvas.bind("<Configure>", _on_resize)

        def _on_frame_resize(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_frame_resize)

        def _wheel(e):
            canvas.yview_scroll(-1 * (e.delta // 120), "units")
        canvas.bind_all("<MouseWheel>", _wheel)

        self._fill_sidebar(inner)

    def _fill_sidebar(self, p):
        pad = dict(padx=12, pady=4)

        self._sec(p, "📁  FILE OPERATIONS")
        self._btn(p, "📂  Open PDF",         self.open_pdf,  _RC["accent3"], pad=pad)
        self._btn(p, "💾  Save Redacted PDF", self.save_pdf,  _RC["accent1"], pad=pad)

        self._sec(p, "🔖  NAVIGATION")
        nav = tk.Frame(p, bg=_RC["panel"])
        nav.pack(fill="x", **pad)
        self._nbtn(nav, "◀ Prev", self.prev_page, _RC["accent4"]).pack(side="left", expand=True, fill="x", padx=(0, 3))
        self._nbtn(nav, "Next ▶", self.next_page, _RC["accent4"]).pack(side="left", expand=True, fill="x", padx=(3, 0))

        self._page_lbl = tk.Label(p, text="Page  —  /  —",
                                  bg=_RC["panel"], fg=_RC["txt"],
                                  font=("Segoe UI", 10, "bold"))
        self._page_lbl.pack(pady=3)

        zrow = tk.Frame(p, bg=_RC["panel"])
        zrow.pack(fill="x", padx=12, pady=2)
        tk.Label(zrow, text="Zoom:", bg=_RC["panel"], fg=_RC["sub"],
                 font=("Segoe UI", 9)).pack(side="left")
        self._zoom_lbl = tk.Label(zrow, text=f"{int(self.zoom * 100)}%",
                                  bg=_RC["panel"], fg=_RC["accent2"],
                                  font=("Segoe UI", 9, "bold"))
        self._zoom_lbl.pack(side="left", padx=6)
        self._nbtn(zrow, " + ", self.zoom_in,  _RC["accent2"], w=3).pack(side="right", padx=2)
        self._nbtn(zrow, " – ", self.zoom_out, _RC["accent2"], w=3).pack(side="right", padx=2)

        self._sec(p, "✏️  MANUAL REDACTION")
        tk.Label(p, text="Draw black rectangles on the PDF\nto cover sensitive content",
                 bg=_RC["panel"], fg=_RC["sub"], font=("Segoe UI", 9),
                 justify="left").pack(anchor="w", padx=12, pady=2)
        self._btn(p, "🗑️  Clear This Page", self.clear_page, _RC["accent5"], pad=pad)
        self._btn(p, "🗑️  Clear All Pages", self.clear_all,  "#7f1d1d",      pad=pad)

        self._sec(p, "🤖  AUTO-REDACT FIELDS")
        tk.Label(p, text="Tick fields to auto-detect & blackout\nacross ALL pages:",
                 bg=_RC["panel"], fg=_RC["sub"], font=("Segoe UI", 9),
                 justify="left").pack(anchor="w", padx=12, pady=2)

        fields = [
            ("GSTN Number",   self.cb_gstn,   _RC["accent1"]),
            ("PAN Number",    self.cb_pan,    "#a855f7"),
            ("Name",          self.cb_name,   _RC["accent4"]),
            ("Notice Number", self.cb_notice, _RC["accent2"]),
            ("Date",          self.cb_date,   _RC["accent3"]),
        ]
        for label, var, col in fields:
            row = tk.Frame(p, bg=_RC["panel"])
            row.pack(fill="x", padx=12, pady=1)
            tk.Frame(row, bg=col, width=4).pack(side="left", fill="y", padx=(0, 6))
            cb = tk.Checkbutton(row, text=label, variable=var,
                                bg=_RC["panel"], fg=_RC["txt"],
                                selectcolor=_RC["bg"],
                                activebackground=_RC["panel"],
                                activeforeground=_RC["txt"],
                                font=("Segoe UI", 10),
                                cursor="hand2")
            cb.pack(side="left")

        self._btn(p, "⚡  Apply Auto-Redact", self.apply_auto_redact, _RC["accent6"], pad=pad)

        self._sec(p, "ℹ️  HOW TO USE")
        help_txt = (
            "1. Open a PDF file.\n"
            "2. Draw boxes on sensitive\n   text to blackout manually.\n"
            "3. OR tick checkboxes above\n   and click Apply Auto-Redact.\n"
            "4. Navigate pages as needed.\n"
            "5. Save the redacted PDF."
        )
        tk.Label(p, text=help_txt, bg=_RC["panel"], fg=_RC["sub"],
                 font=("Segoe UI", 9), justify="left"
                 ).pack(anchor="w", padx=12, pady=4)


    # ── PDF Viewer ────────────────────────────────────────────────────────

    def _build_viewer(self, parent):
        from tkinter import ttk as _ttk
        tbar = tk.Frame(parent, bg=_RC["header"], height=36)
        tbar.pack(fill="x")
        tbar.pack_propagate(False)
        tk.Label(tbar, text="📄  PDF Viewer  –  Draw rectangles to redact",
                 bg=_RC["header"], fg=_RC["sub"], font=("Segoe UI", 9)
                 ).pack(side="left", padx=12, pady=8)

        cf = tk.Frame(parent, bg=_RC["canvas"])
        cf.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(cf, bg=_RC["canvas"],
                                cursor="crosshair", highlightthickness=0)
        vs = _ttk.Scrollbar(cf, orient="vertical",   command=self.canvas.yview)
        hs = _ttk.Scrollbar(cf, orient="horizontal", command=self.canvas.xview)

        self.canvas.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)

        vs.pack(side="right",  fill="y")
        hs.pack(side="bottom", fill="x")
        self.canvas.pack(fill="both", expand=True)

        self.canvas.bind("<ButtonPress-1>",   self._on_press)
        self.canvas.bind("<B1-Motion>",       self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)

        self._welcome = tk.Label(
            self.canvas,
            text="📂   Open a PDF file to begin\n\nUse the sidebar to navigate and redact",
            bg=_RC["canvas"], fg=_RC["sub"],
            font=("Segoe UI", 14), justify="center"
        )
        self._welcome.place(relx=0.5, rely=0.45, anchor="center")

    # ═══════════════════════════  RENDERING  ══════════════════════════════

    def render_page(self):
        if not self.pdf_doc:
            return
        from PIL import Image, ImageTk as _ImageTk
        page = self.pdf_doc[self.current_page]
        mat  = fitz.Matrix(self.zoom, self.zoom)
        pix  = page.get_pixmap(matrix=mat, alpha=False)

        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        self._tk_img = _ImageTk.PhotoImage(img)

        self.canvas.delete("all")
        if hasattr(self, "_welcome"):
            self._welcome.place_forget()

        self.update_idletasks()
        cw = self.canvas.winfo_width()
        self.img_x0 = max(10, (cw - pix.width) // 2)
        self.img_y0 = 10

        sr_w = max(pix.width  + 2 * self.img_x0, cw)
        sr_h = max(pix.height + 2 * self.img_y0, self.canvas.winfo_height())
        self.canvas.configure(scrollregion=(0, 0, sr_w, sr_h))
        self.canvas.create_image(self.img_x0, self.img_y0,
                                  anchor="nw", image=self._tk_img)

        for rect in self.redactions.get(self.current_page, []):
            self._draw_rect_canvas(rect)

        self._page_lbl.config(text=f"Page  {self.current_page + 1}  /  {self.total_pages}")
        self.status.set(
            f"Page {self.current_page + 1} of {self.total_pages}  |  "
            f"Zoom {int(self.zoom * 100)}%  |  "
            f"Marks: {sum(len(v) for v in self.redactions.values())}"
        )

    def _draw_rect_canvas(self, pdf_rect: fitz.Rect):
        x0 = pdf_rect.x0 * self.zoom + self.img_x0
        y0 = pdf_rect.y0 * self.zoom + self.img_y0
        x1 = pdf_rect.x1 * self.zoom + self.img_x0
        y1 = pdf_rect.y1 * self.zoom + self.img_y0
        self.canvas.create_rectangle(x0, y0, x1, y1,
                                      fill=_RC["redact"], outline=_RC["redact"])

    # ═══════════════════════════  MOUSE  ══════════════════════════════════

    def _on_press(self, event):
        if not self.pdf_doc:
            return
        self.drawing    = True
        cx, cy          = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        self.draw_start = (cx, cy)
        self.temp_id    = None

    def _on_drag(self, event):
        if not self.drawing:
            return
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)
        if self.temp_id:
            self.canvas.delete(self.temp_id)
        self.temp_id = self.canvas.create_rectangle(
            self.draw_start[0], self.draw_start[1], cx, cy,
            fill="", outline=_RC["accent2"], width=2, dash=(4, 2)
        )

    def _on_release(self, event):
        if not self.drawing or not self.pdf_doc:
            return
        self.drawing = False

        if self.temp_id:
            self.canvas.delete(self.temp_id)
            self.temp_id = None

        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)

        x0, y0 = min(self.draw_start[0], cx), min(self.draw_start[1], cy)
        x1, y1 = max(self.draw_start[0], cx), max(self.draw_start[1], cy)

        if abs(x1 - x0) < 4 or abs(y1 - y0) < 4:
            return

        pdf_rect = fitz.Rect(
            (x0 - self.img_x0) / self.zoom,
            (y0 - self.img_y0) / self.zoom,
            (x1 - self.img_x0) / self.zoom,
            (y1 - self.img_y0) / self.zoom,
        )
        self.redactions.setdefault(self.current_page, []).append(pdf_rect)
        self._draw_rect_canvas(pdf_rect)
        self.status.set(
            f"Redaction added on page {self.current_page + 1}  |  "
            f"Total marks: {sum(len(v) for v in self.redactions.values())}"
        )

    # ═══════════════════════════  NAVIGATION  ═════════════════════════════

    def prev_page(self):
        if self.pdf_doc and self.current_page > 0:
            self.current_page -= 1
            self.render_page()

    def next_page(self):
        if self.pdf_doc and self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.render_page()

    def zoom_in(self):
        self.zoom = min(4.0, round(self.zoom + 0.25, 2))
        self._zoom_lbl.config(text=f"{int(self.zoom * 100)}%")
        self.render_page()

    def zoom_out(self):
        self.zoom = max(0.5, round(self.zoom - 0.25, 2))
        self._zoom_lbl.config(text=f"{int(self.zoom * 100)}%")
        self.render_page()

    # ═══════════════════════════  FILE OPS  ═══════════════════════════════

    def open_pdf(self):
        path = filedialog.askopenfilename(
            title="Open PDF File",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if not path:
            return
        try:
            if self.pdf_doc:
                self.pdf_doc.close()
            self.pdf_doc      = fitz.open(path)
            self.pdf_path     = path
            self.current_page = 0
            self.total_pages  = len(self.pdf_doc)
            self.redactions   = {}
            self.update_idletasks()
            self.render_page()
            self.status.set(
                f"Opened: {os.path.basename(path)}  |  {self.total_pages} page(s)"
            )
        except Exception as exc:
            messagebox.showerror("Open Error", f"Could not open PDF:\n{exc}")

    def save_pdf(self):
        if not self.pdf_doc:
            messagebox.showwarning("No PDF", "Please open a PDF file first.")
            return
        if not self.redactions:
            messagebox.showwarning("No Marks",
                "No redaction marks found.\n"
                "Draw rectangles or run Auto-Redact first.")
            return

        import time
        default_name = f"redacted_{int(time.time())}_" + os.path.basename(self.pdf_path)
        save_path = os.path.join(os.path.dirname(self.pdf_path), default_name)

        try:
            doc = fitz.open(self.pdf_path)
            total_applied = 0
            for pg_num, rects in self.redactions.items():
                pg = doc[pg_num]
                for r in rects:
                    pg.add_redact_annot(r, fill=(0, 0, 0))
                pg.apply_redactions()
                total_applied += len(rects)
            doc.save(save_path, garbage=4, deflate=True)
            doc.close()
            self.status.done(f"Saved: {os.path.basename(save_path)}")
            messagebox.showinfo(
                "Saved Successfully ✅",
                f"Redacted PDF saved!\n\n"
                f"File : {os.path.basename(save_path)}\n"
                f"Pages redacted : {len(self.redactions)}\n"
                f"Total marks    : {total_applied}"
            )
        except Exception as exc:
            messagebox.showerror("Save Error", f"Could not save PDF:\n{exc}")

    # ═══════════════════════════  CLEAR  ══════════════════════════════════

    def clear_page(self):
        if not self.pdf_doc:
            return
        if self.current_page in self.redactions:
            del self.redactions[self.current_page]
        self.render_page()
        self.status.set(f"Cleared all marks on page {self.current_page + 1}")

    def clear_all(self):
        if not self.pdf_doc:
            return
        if not messagebox.askyesno("Clear All Marks",
                                    "Remove ALL redaction marks from ALL pages?\n"
                                    "This cannot be undone."):
            return
        self.redactions = {}
        self.render_page()
        self.status.set("All redaction marks cleared")

    # ═══════════════════════  AUTO-REDACT  ════════════════════════════════

    def apply_auto_redact(self):
        if not self.pdf_doc:
            messagebox.showwarning("No PDF", "Please open a PDF file first.")
            return

        checks = {
            "GSTN":      self.cb_gstn.get(),
            "PAN":       self.cb_pan.get(),
            "Name":      self.cb_name.get(),
            "Notice No": self.cb_notice.get(),
            "Date":      self.cb_date.get(),
        }
        if not any(checks.values()):
            messagebox.showwarning("Nothing Selected",
                "Please tick at least one field to auto-redact.")
            return

        self.status.busy("Running auto-redact – please wait…")

        total = 0
        for pg_idx in range(self.total_pages):
            page  = self.pdf_doc[pg_idx]
            found = []

            if checks["GSTN"]:
                found += self._find_regex(page,
                    r'\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9]\b')

            if checks["PAN"]:
                found += self._find_regex(page,
                    r'\b[A-Z]{5}[0-9]{4}[A-Z]\b')

            if checks["Name"]:
                found += self._find_label_line(page,
                    ["Name", "Party Name", "Taxpayer Name",
                     "Trade Name", "Legal Name", "Applicant Name",
                     "Name of Tax Payer"])

            if checks["Notice No"]:
                found += self._find_label_line(page,
                    ["Notice No", "Notice Number", "Reference No",
                     "Reference Number", "ARN", "DIN", "Token No",
                     "Application No", "Challan No", "Document No"])

            if checks["Date"]:
                found += self._find_regex(page,
                    r'\b\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}\b')
                found += self._find_regex(page,
                    r'\b\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2}\b')
                found += self._find_regex(page,
                    r'\b\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|'
                    r'Sep|Oct|Nov|Dec)[a-z]*[\.,\s]+\d{2,4}\b',
                    flags=re.IGNORECASE)

            if found:
                self.redactions.setdefault(pg_idx, []).extend(found)
                total += len(found)

        self.render_page()
        self.status.done(
            f"Auto-Redact complete  |  {total} item(s) marked across "
            f"{len(self.redactions)} page(s)"
        )
        messagebox.showinfo(
            "Auto-Redact Complete ✅",
            f"Found and marked {total} item(s) across "
            f"{len(self.redactions)} page(s).\n\n"
            "Review all pages, then click 'Save Redacted PDF'."
        )

    # ── Pattern helpers ───────────────────────────────────────────────────

    def _find_regex(self, page: fitz.Page, pattern: str, flags: int = 0) -> list:
        full_text = page.get_text()
        rects     = []
        for m in re.finditer(pattern, full_text, flags):
            txt = m.group().strip()
            if txt:
                for r in page.search_for(txt):
                    rects.append(r)
        return rects

    def _find_label_line(self, page: fitz.Page, labels: list) -> list:
        rects     = []
        page_dict = page.get_text("dict")
        for block in page_dict.get("blocks", []):
            if "lines" not in block:
                continue
            for line in block["lines"]:
                line_text = " ".join(
                    span["text"] for span in line.get("spans", [])
                ).strip()
                for lbl in labels:
                    if lbl.lower() in line_text.lower():
                        bbox = fitz.Rect(line["bbox"])
                        bbox = fitz.Rect(bbox.x0 - 1, bbox.y0 - 1,
                                         bbox.x1 + 1, bbox.y1 + 1)
                        rects.append(bbox)
                        break
        return rects

    # ═══════════════════════  UI HELPERS  ═════════════════════════════════

    @staticmethod
    def _btn(parent, text, cmd, bg, pad=None):
        if pad is None:
            pad = dict(padx=10, pady=3)
        b = tk.Button(parent, text=text, command=cmd,
                      bg=bg, fg="white", relief="flat",
                      font=("Segoe UI", 10, "bold"),
                      cursor="hand2", pady=7, padx=8,
                      activeforeground="white")
        b.pack(fill="x", **pad)
        import colorsys
        def _darken(col):
            col = col.lstrip("#")
            r, g, b_ = int(col[0:2], 16)/255, int(col[2:4], 16)/255, int(col[4:6], 16)/255
            h, s, v  = colorsys.rgb_to_hsv(r, g, b_)
            r2, g2, b2 = colorsys.hsv_to_rgb(h, s, max(0, v - 0.12))
            return "#{:02x}{:02x}{:02x}".format(int(r2*255), int(g2*255), int(b2*255))
        hover = _darken(bg)
        b.bind("<Enter>", lambda e: b.config(bg=hover))
        b.bind("<Leave>", lambda e: b.config(bg=bg))
        return b

    @staticmethod
    def _nbtn(parent, text, cmd, bg, w=None):
        kw = dict(width=w) if w else {}
        b = tk.Button(parent, text=text, command=cmd,
                      bg=bg, fg="white", relief="flat",
                      font=("Segoe UI", 9, "bold"),
                      cursor="hand2", pady=5, **kw)
        return b

    @staticmethod
    def _sec(parent, label):
        f = tk.Frame(parent, bg=_RC["panel"])
        f.pack(fill="x", padx=0, pady=(14, 2))
        tk.Label(f, text=label, bg=_RC["panel"], fg=_RC["accent2"],
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=10)
        tk.Frame(parent, bg=_RC["border"], height=1).pack(fill="x", padx=10)


# ── (RedactEditorWindow removed – all editing is now inline in RedactTab) ──


# ─────────────────────────────────────────────────────────────────
#  PDF → EXCEL TAB
# ─────────────────────────────────────────────────────────────────

class PDFToExcelTab(tk.Frame):

    def __init__(self, parent, status: "StatusBar"):
        super().__init__(parent, bg=TM["bg"])
        self.status   = status
        self.src_path = tk.StringVar()
        self.dst_dir  = tk.StringVar()
        self._build()

    def _build(self):
        _page_title(self, "⊞  PDF → Excel",
                    "Convert any PDF to an Excel workbook (.xlsx) with text and tables.")

        # ── Source PDF row ────────────────────────────────────
        _src_row(self, self.src_path, on_browse=self._browse_pdf)

        # ── Convert button ────────────────────────────────────
        btn_row = tk.Frame(self, bg=TM["bg"])
        btn_row.pack(pady=8)
        ModernButton(btn_row, "  ⊞  Convert to Excel  ", self._convert,
                     min_width=200, pady=10).pack()

        # ── Log area ──────────────────────────────────────────
        log_card = Card(self)
        log_card.pack(fill="both", expand=True, padx=20, pady=(0, 12))
        tk.Label(log_card, text="Log", font=F_H3,
                 bg=TM["surface"], fg=TM["accent"],
                 padx=14, pady=6).pack(anchor="w")
        self._log = scrolledtext.ScrolledText(
            log_card, font=F_MONO, bg=TM["surface2"], fg=TM["text"],
            insertbackground=TM["text"], relief="flat", bd=0,
            height=10, state="disabled")
        self._log.pack(fill="both", expand=True, padx=1, pady=(0, 1))

    # ── helpers ──────────────────────────────────────────────

    def _browse_pdf(self):
        p = _ask_pdf()
        if p:
            self.src_path.set(p)
            # default output folder to same dir as PDF
            if not self.dst_dir.get():
                self.dst_dir.set(os.path.dirname(p))

    def _log_line(self, msg: str):
        self._log.configure(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.configure(state="disabled")

    def _clear_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    # ── conversion logic ────────────────────────────────────

    def _convert(self):
        src = self.src_path.get().strip()
        if not src or not os.path.isfile(src):
            messagebox.showwarning("No Source", "Please select a valid PDF file.")
            return

        import time
        dst_dir = os.path.dirname(src)
        base_name = os.path.splitext(os.path.basename(src))[0]
        out_path  = os.path.join(dst_dir, f"{base_name}_converted_{int(time.time())}.xlsx")

        if os.path.exists(out_path):
            if not messagebox.askyesno("Overwrite?",
                    f"'{os.path.basename(out_path)}' already exists. Overwrite?"):
                return

        self.status.busy("Converting PDF to Excel…")
        self._clear_log()

        def _run():
            try:
                data = self._extract(src)
                if not data:
                    self.after(0, lambda: (
                        self.status.done("No data extracted.", TM["error"]),
                        messagebox.showerror("Error", "No data could be extracted from the PDF.")
                    ))
                    return
                ok = self._write_xlsx(data, out_path)
                if ok:
                    self.after(0, lambda: (
                        self.status.done("Conversion complete."),
                        messagebox.showinfo("Done",
                            f"Excel file saved:\n{out_path}")
                    ))
                else:
                    self.after(0, lambda: self.status.done("Failed.", TM["error"]))
            except Exception as exc:
                self.after(0, lambda: (
                    self.status.done(f"Error: {exc}", TM["error"]),
                    messagebox.showerror("Error", str(exc))
                ))

        threading.Thread(target=_run, daemon=True).start()

    def _extract(self, pdf_path: str) -> list:
        all_data = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                self.after(0, lambda n=len(pdf.pages):
                           self._log_line(f"Total pages: {n}"))
                for page_num, page in enumerate(pdf.pages, 1):
                    self.after(0, lambda p=page_num:
                               self._log_line(f"Processing page {p}…"))
                    all_data.append(f"--- PAGE {page_num} ---")
                    text = page.extract_text()
                    if text:
                        all_data.extend(
                            line.strip() for line in text.split("\n") if line.strip()
                        )
                    tables = page.extract_tables()
                    if tables:
                        all_data.append("")
                        all_data.append(f"--- TABLES IN PAGE {page_num} ---")
                        for t_idx, table in enumerate(tables, 1):
                            all_data.append("")
                            all_data.append(f"Table {t_idx}:")
                            for row in (table if isinstance(table, list) else []):
                                if isinstance(row, list):
                                    all_data.append(
                                        " | ".join(str(c) if c else "" for c in row)
                                    )
                                else:
                                    all_data.append(str(row))
        except Exception as exc:
            self.after(0, lambda: self._log_line(f"ERROR: {exc}"))
            return []
        self.after(0, lambda n=len(all_data):
                   self._log_line(f"Extracted {n} items."))
        return all_data

    def _write_xlsx(self, data: list, out_path: str) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PDF Content"

            sec_fill   = PatternFill("solid", fgColor="D9E1F2")
            sec_font   = XLFont(bold=True, size=10, color="000000")
            thin_side  = Side(style="thin")
            bdr        = Border(left=thin_side, right=thin_side,
                                top=thin_side,  bottom=thin_side)

            ws["A1"] = "PDF Content Extracted"
            ws["A1"].font = XLFont(bold=True, size=14)
            ws["A1"].fill = PatternFill("solid", fgColor="E7E6E6")
            ws.merge_cells("A1:D1")

            ws["A2"] = "Source File:"
            ws["B2"] = os.path.basename(out_path).replace("_converted.xlsx", ".pdf")
            ws["A2"].font = XLFont(italic=True)

            row = 4
            for item in data:
                s = str(item) if item is not None else ""
                if s.startswith("---"):
                    ws[f"A{row}"] = s
                    ws[f"A{row}"].fill = sec_fill
                    ws[f"A{row}"].font = sec_font
                    ws.merge_cells(f"A{row}:D{row}")
                    ws[f"A{row}"].alignment = Alignment(vertical="top")
                elif "|" in s and len(s) > 3:
                    for ci, val in enumerate(s.split("|"), 1):
                        c = ws.cell(row=row, column=ci, value=val.strip())
                        c.alignment = Alignment(vertical="top")
                        c.border = bdr
                elif ":" in s and len(s) > 3:
                    parts = s.split(":", 1)
                    for ci, val in enumerate(parts, 1):
                        c = ws.cell(row=row, column=ci, value=val.strip())
                        c.alignment = Alignment(vertical="top")
                        c.border = bdr
                elif s.strip():
                    ws[f"A{row}"] = s
                    ws[f"A{row}"].alignment = Alignment(vertical="top")
                    ws[f"A{row}"].border = bdr
                    ws.merge_cells(f"A{row}:D{row}")
                else:
                    row += 1
                    continue
                row += 1

            for col, w in zip("ABCD", (35, 40, 40, 40)):
                ws.column_dimensions[col].width = w

            wb.save(out_path)
            self.after(0, lambda: self._log_line(f"Saved: {out_path}"))
            return True
        except Exception as exc:
            self.after(0, lambda: self._log_line(f"WRITE ERROR: {exc}"))
            return False


# ═══════════════════════════════════════════════════════════════════
#  HEADER
# ═══════════════════════════════════════════════════════════════════

class Header(tk.Frame):
    """
    Top navigation bar: logo, title, theme toggle.
    """

    def __init__(self, parent, toggle_cb):
        super().__init__(parent, bg=TM["header_bg"], height=56)
        self.pack(fill="x")
        self.pack_propagate(False)

        # ── App name ──────────────────────────────────────────
        tk.Label(self, text="PDF Utilities", font=("Segoe UI", 14, "bold"),
                 bg=TM["header_bg"], fg=TM["text"]).pack(side="left", padx=(20, 0))
        tk.Label(self, text="  ·  Merge · Split · Extract · Compress · Redact",
                 font=F_SMALL, bg=TM["header_bg"],
                 fg=TM["subtext"]).pack(side="left")

    def update_toggle_label(self, dark_mode: bool):
        pass


# ═══════════════════════════════════════════════════════════════════
#  MODULE TAB BAR
# ═══════════════════════════════════════════════════════════════════

class ModuleTabBar(tk.Frame):
    """
    Horizontal tab bar for module navigation.
    """

    _NAV = [
        ("⊕", "Merge"),
        ("✂", "Split"),
        ("⊙", "Extract"),
        ("⊜", "Compress"),
        ("⬛", "Redact"),
        ("⊞", "PDF→Excel"),
    ]

    def __init__(self, parent, on_select):
        super().__init__(parent, bg=TM["header_bg"])
        self._items: list[TabItem] = []
        self._on_select = on_select
        self._build()

    def _build(self):
        # Tools label (optional, can be omitted if too crowded)
        tk.Label(self, text="TOOLS ", font=F_CAPTION,
                 bg=TM["header_bg"], fg=TM["text_dim"]).pack(side="left", padx=(20, 10))

        for idx, (icon, label) in enumerate(self._NAV):
            item = TabItem(self, icon, label, self._click, idx)
            item.pack(side="left", fill="y")
            self._items.append(item)

        # Version hint at the far right
        tk.Label(self, text="v1.0.0 ", font=F_CAPTION,
                 bg=TM["header_bg"], fg=TM["text_dim"]).pack(side="right", padx=20)

    def _click(self, idx: int):
        for i, item in enumerate(self._items):
            item.set_active(i == idx)
        self._on_select(idx)

    def select(self, idx: int):
        self._click(idx)



# ═══════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════════

class PDFUtilitiesApp(tk.Tk):
    """
    Main app window.
    Layout:
        Header  (top, full width)
        ├── Sidebar (left, fixed 210 px)
        └── Content area (right, fills remaining)
        StatusBar (bottom, full width)
    """

    def __init__(self):
        super().__init__()
        self.title("PDF Utilities")
        self.geometry("1080x720")
        self.minsize(860, 580)
        self.configure(bg=TM["bg"])
        self._dark = True
        self._build()

    def _build(self):
        # Status bar first (packs to bottom)
        self.status = StatusBar(self)

        # Header
        self.header = Header(self, None)

        # Tab bar below header
        self.tab_bar = ModuleTabBar(self, self._show_page)
        self.tab_bar.pack(fill="x")

        # thin separator under tab bar
        tk.Frame(self, bg=TM["separator"], height=1).pack(fill="x")

        # content area (fills bottom)
        self._content = tk.Frame(self, bg=TM["bg"])
        self._content.pack(fill="both", expand=True)


        # Build all pages
        self._pages: list[tk.Frame] = [
            MergeTab(self._content,        self.status),
            SplitTab(self._content,        self.status),
            ExtractorTab(self._content,    self.status),
            CompressorTab(self._content,   self.status),
            RedactTab(self._content,       self.status),
            PDFToExcelTab(self._content,   self.status),
        ]

        # Show first page
        self.tab_bar.select(0)


    def _show_page(self, idx: int):
        for pg in self._pages:
            pg.pack_forget()
        self._pages[idx].pack(fill="both", expand=True)

    def set_theme(self, mode: str):
        is_dark = (mode.lower() == "dark")
        if TM.is_dark() != is_dark:
            TM.toggle()
            TM.retheme(self)
            self._dark = is_dark
            self.configure(bg=TM["bg"])


# ═══════════════════════════════════════════════════════════════════
#  INDIVIDUAL TOOL WRAPPERS (for GST Suite integration)
# ═══════════��═══════════════════════════════════════════════════════

class BasePDFTool(tk.Tk):
    def __init__(self, tab_class):
        super().__init__()
        self.configure(bg=TM["bg"])
        self.status = StatusBar(self)
        self.tool   = tab_class(self, self.status)
        self.tool.pack(fill="both", expand=True)

    def set_theme(self, mode: str):
        is_dark = (mode.lower() == "dark")
        if TM.is_dark() != is_dark:
            TM.toggle()
            TM.retheme(self)
            self.configure(bg=TM["bg"])

class MergeApp(BasePDFTool):
    def __init__(self): super().__init__(MergeTab)

class SplitApp(BasePDFTool):
    def __init__(self): super().__init__(SplitTab)

class ExtractApp(BasePDFTool):
    def __init__(self): super().__init__(ExtractorTab)

class CompressApp(BasePDFTool):
    def __init__(self): super().__init__(CompressorTab)

class RedactApp(BasePDFTool):
    def __init__(self): super().__init__(RedactTab)

class PDFToExcelApp(BasePDFTool):
    def __init__(self): super().__init__(PDFToExcelTab)


# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = PDFUtilitiesApp()
    app.mainloop()
