"""
TallyBankPro - Bank Statement to Tally Payment/Receipt Voucher Creator
Reads bank statement Excel (template format) and pushes Payment/Receipt
vouchers to TallyPrime via HTTP API.
Author: Studycafe | Built with CustomTkinter
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
import openpyxl
import os
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta
import urllib.request
from urllib.error import HTTPError, URLError
import json
import threading
import re
import html
import shutil

# ─── Theme (matches TallySalesPro) ──────────────────────────────────────
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

COLORS = {
    "bg_dark": ("#F0F4F8", "#0F172A"),
    "bg_card": ("#FFFFFF", "#1E293B"),
    "bg_card_hover": ("#E2E8F0", "#334155"),
    "bg_input": ("#F1F5F9", "#1E293B"),
    "accent": ("#2563EB", "#3B82F6"),
    "accent_hover": ("#1D4ED8", "#2563EB"),
    "success": ("#059669", "#10B981"),
    "warning": ("#D97706", "#F59E0B"),
    "error": ("#DC2626", "#EF4444"),
    "text_primary": ("#0F172A", "#F1F5F9"),
    "text_secondary": ("#475569", "#CBD5E1"),
    "text_muted": ("#64748B", "#94A3B8"),
    "border": ("#E2E8F0", "#334155"),
    "table_header": ("#1E293B", "#0F172A"),
}

ACCENT = COLORS["accent"]
ACCENT_HOVER = COLORS["accent_hover"]
SUCCESS = COLORS["success"]
DANGER = COLORS["error"]
TEXT_PRIMARY = COLORS["text_primary"]
TEXT_MUTED = COLORS["text_muted"]

PUSH_REQUEST_TIMEOUT_SEC = 300
PUSH_BATCH_SIZE = 40


# ─── XML / Tally Helpers ────────────────────────────────────────────────

def xml_escape(s: str) -> str:
    if not s:
        return ""
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&apos;")


def fmt_amt(num: float) -> str:
    return f"{num:.2f}"


def _parse_statement_datetime(dt):
    if dt in (None, ""):
        return None
    if isinstance(dt, datetime):
        return dt
    if isinstance(dt, date):
        return datetime(dt.year, dt.month, dt.day)
    if isinstance(dt, (int, float)) and not isinstance(dt, bool):
        try:
            if float(dt) > 1000:
                return datetime(1899, 12, 30) + timedelta(days=float(dt))
        except Exception:
            return None

    text = str(dt).strip()
    if not text:
        return None
    text = re.sub(r"[\u2010-\u2015\u2212]", "-", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace(",", "")

    broken_year_match = re.match(
        r"^(\d{1,2})([\-/\s])([A-Za-z]{3,9}|\d{1,2})([\-/\s])(\d{3})$",
        text,
    )
    if broken_year_match:
        day, sep1, month, sep2, year3 = broken_year_match.groups()
        if year3.startswith("20"):
            decade = (datetime.today().year % 100) // 10
            year4 = f"20{decade}{year3[-1]}"
        else:
            year4 = f"20{year3[-2:]}"
        text = f"{day}{sep1}{month}{sep2}{year4}"

    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if re.fullmatch(r"\d{8}", text):
        if 1900 <= int(text[:4]) <= 2100:
            try:
                return datetime.strptime(text, "%Y%m%d")
            except ValueError:
                return None
        try:
            return datetime.strptime(f"{text[4:8]}{text[2:4]}{text[:2]}", "%Y%m%d")
        except ValueError:
            return None

    candidates = [text]
    if " " in text:
        candidates.append(text.split(" ", 1)[0])
    formats = (
        "%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y",
        "%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y", "%d-%B-%y",
        "%d %b %Y", "%d %b %y", "%d %B %Y", "%d %B %y",
        "%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S",
    )
    for candidate in candidates:
        for fmt in formats:
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def tally_date(dt) -> str:
    today = datetime.today().strftime("%Y%m%d")
    parsed = _parse_statement_datetime(dt)
    if parsed:
        return parsed.strftime("%Y%m%d")
    return today


def _derive_import_period(rows: list, use_today_date: bool = False) -> tuple:
    if use_today_date:
        today = datetime.today().strftime("%Y%m%d")
        return today, today, today

    parsed_dates = []
    for r in rows or []:
        source_date = r.get("DATE") or r.get("Date") or r.get("date") or ""
        parsed = _parse_statement_datetime(source_date)
        if parsed is not None:
            parsed_dates.append(parsed)

    if not parsed_dates:
        today = datetime.today().strftime("%Y%m%d")
        return today, today, today

    start = min(parsed_dates).strftime("%Y%m%d")
    end = max(parsed_dates).strftime("%Y%m%d")
    return start, end, end


def _count_voucher_entries(rows: list) -> tuple:
    payment_count = 0
    receipt_count = 0

    for r in rows or []:
        dr = r.get("Debit") or r.get("DEBIT") or r.get("debit") or 0
        cr = r.get("Credit") or r.get("CREDIT") or r.get("credit") or 0
        try:
            dr_val = float(dr) if dr not in (None, "", "None") else 0.0
        except (TypeError, ValueError):
            dr_val = 0.0
        try:
            cr_val = float(cr) if cr not in (None, "", "None") else 0.0
        except (TypeError, ValueError):
            cr_val = 0.0

        if dr_val > 0:
            payment_count += 1
        if cr_val > 0:
            receipt_count += 1

    return payment_count, receipt_count


def _normalize_company_name(value) -> str:
    text = html.unescape(str(value or ""))
    text = text.replace("\x00", "")
    text = re.sub(r"[\x01-\x1F\x7F]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _company_key(value) -> str:
    return _normalize_company_name(value).upper()


def _is_valid_company_name(value) -> bool:
    name = _normalize_company_name(value)
    if not name:
        return False
    return not name.isdigit()


def _build_tally_url(host: str, port: str) -> str:
    host_text = (host or "localhost").strip()
    port_text = (port or "9000").strip()
    if host_text.startswith("http://"):
        host_text = host_text[7:]
    elif host_text.startswith("https://"):
        host_text = host_text[8:]
    host_text = host_text.strip("/") or "localhost"
    if "/" in host_text:
        host_text = host_text.split("/", 1)[0]
    if not port_text.isdigit():
        raise ValueError("Port must be numeric.")
    return f"http://{host_text}:{port_text}"


def _post_tally_xml(tally_url: str, xml_payload: str, timeout: float = 15.0) -> str:
    request = urllib.request.Request(
        tally_url,
        data=xml_payload.encode("utf-8"),
        headers={"Content-Type": "application/xml"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="replace")


def push_to_tally(
    xml_str: str,
    host: str = "localhost",
    port: int = 9000,
    timeout: float = PUSH_REQUEST_TIMEOUT_SEC,
) -> str:
    url = f"http://{host}:{port}"
    req = urllib.request.Request(url, data=xml_str.encode("utf-8"),
                                 headers={"Content-Type": "application/xml"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8")


def _company_static_block(
    company: str,
    from_date: str = "",
    to_date: str = "",
    current_date: str = "",
) -> str:
    parts = []
    selected = str(company or "").strip()
    if selected:
        parts.append(f"<SVCURRENTCOMPANY>{xml_escape(selected)}</SVCURRENTCOMPANY>")
    if from_date:
        parts.append(f"<SVFROMDATE TYPE=\"Date\">{from_date}</SVFROMDATE>")
    if to_date:
        parts.append(f"<SVTODATE TYPE=\"Date\">{to_date}</SVTODATE>")
    if current_date:
        parts.append(f"<SVCURRENTDATE TYPE=\"Date\">{current_date}</SVCURRENTDATE>")

    if not parts:
        return ""
    return f"   <STATICVARIABLES>{''.join(parts)}</STATICVARIABLES>"


def _extract_company_names(response_text: str) -> set:
    names = set()
    try:
        root = ET.fromstring(response_text)
        for node in root.iter():
            tag = str(node.tag or "").upper()
            txt = _normalize_company_name(node.text)
            attr_name = _normalize_company_name(node.attrib.get("NAME") or "")
            if tag in {"COMPANYNAME", "SVCURRENTCOMPANY", "CURRENTCOMPANY"} and _is_valid_company_name(txt):
                names.add(txt)
            if "COMPANY" in tag and _is_valid_company_name(attr_name):
                names.add(attr_name)
            if tag == "COMPANY" and _is_valid_company_name(txt):
                names.add(txt)
    except ET.ParseError:
        pass
    patterns = [
        r'COMPANY[^>]*NAME="([^"]+)"',
        r"<COMPANYNAME>(.*?)</COMPANYNAME>",
        r"<SVCURRENTCOMPANY>(.*?)</SVCURRENTCOMPANY>",
    ]
    for pattern in patterns:
        for match in re.findall(pattern, response_text, flags=re.IGNORECASE | re.DOTALL):
            value = _normalize_company_name(match)
            if _is_valid_company_name(value):
                names.add(value)
    return names


def _fetch_tally_companies(tally_url: str, timeout: float = 15.0) -> dict:
    requests_xml = [
        (
            "report-list-companies",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>",
        ),
        (
            "collection-company",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
            "<TYPE>Collection</TYPE><ID>Company Collection</ID></HEADER><BODY><DESC>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "<TDL><TDLMESSAGE><COLLECTION NAME='Company Collection'>"
            "<TYPE>Company</TYPE><FETCH>Name</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
            "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>",
        ),
    ]
    companies = set()
    errors = []
    for label, xml_payload in requests_xml:
        try:
            response_text = _post_tally_xml(tally_url, xml_payload, timeout=timeout)
            companies.update(_extract_company_names(response_text))
        except Exception as exc:
            errors.append(f"{label}: {exc}")
    sorted_companies = sorted(companies, key=lambda x: _company_key(x))
    if sorted_companies:
        return {"success": True, "companies": sorted_companies}
    err = "; ".join(errors) if errors else "No companies returned by Tally."
    return {"success": False, "error": err, "companies": []}


def _check_tally_connection(tally_url: str, timeout: float = 5.0) -> dict:
    probe_xml = (
        "<ENVELOPE><HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
        "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
        "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>"
    )
    try:
        _post_tally_xml(tally_url, probe_xml, timeout=timeout)
        return {"connected": True}
    except HTTPError as exc:
        return {"connected": False, "error": f"HTTP {exc.code}"}
    except URLError:
        return {"connected": False, "error": "ConnectionError"}
    except Exception as exc:
        return {"connected": False, "error": str(exc)}


def _safe_int(text, default=0) -> int:
    try:
        return int(float(str(text).strip()))
    except (TypeError, ValueError):
        return default


def _parse_tally_response_details(response_text: str) -> dict:
    details = {
        "success": False, "created": 0, "altered": 0, "deleted": 0,
        "lastvchid": 0, "lastmid": 0, "combined": 0, "ignored": 0,
        "errors": 0, "cancelled": 0, "exceptions": 0,
        "line_errors": [], "error": "",
    }
    try:
        root = ET.fromstring(response_text)
    except ET.ParseError as exc:
        details["error"] = f"Could not parse Tally response: {exc}"
        return details

    def _count(tag_name: str) -> int:
        node = root.find(f".//{tag_name}")
        return _safe_int(node.text if node is not None else 0)

    details["created"] = _count("CREATED")
    details["altered"] = _count("ALTERED")
    details["deleted"] = _count("DELETED")
    details["lastvchid"] = _count("LASTVCHID")
    details["lastmid"] = _count("LASTMID")
    details["combined"] = _count("COMBINED")
    details["ignored"] = _count("IGNORED")
    details["errors"] = _count("ERRORS")
    details["cancelled"] = _count("CANCELLED")
    details["exceptions"] = _count("EXCEPTIONS")

    line_errors = []
    seen = set()
    for node in root.findall(".//LINEERROR"):
        text = (node.text or "").strip()
        if text and text not in seen:
            seen.add(text)
            line_errors.append(text)
    details["line_errors"] = line_errors
    details["success"] = (
        details["errors"] == 0
        and details["exceptions"] == 0
        and not details["line_errors"]
    )
    if not details["success"] and details["line_errors"]:
        details["error"] = details["line_errors"][0]
    return details


# ─── Fetch Bank Ledgers from Tally ──────────────────────────────────────

def _fetch_bank_ledgers(tally_url: str, company: str = "", timeout: float = 15.0) -> dict:
    """Fetch ledgers under 'Bank Accounts' and 'Cash-in-Hand' groups from Tally."""
    static = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if company:
        static += f"<SVCURRENTCOMPANY>{xml_escape(company)}</SVCURRENTCOMPANY>"
    static += "</STATICVARIABLES>"

    # Use TDL collection to fetch ledgers under Bank Accounts and Cash-in-Hand
    xml_payload = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>BankLedgerCollection</ID></HEADER><BODY><DESC>"
        f"{static}"
        "<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='BankLedgerCollection'>"
        "<TYPE>Ledger</TYPE>"
        "<NATIVEMETHOD>Name</NATIVEMETHOD>"
        "<NATIVEMETHOD>Parent</NATIVEMETHOD>"
        "<NATIVEMETHOD>ClosingBalance</NATIVEMETHOD>"
        "<FILTER>BankOrCashFilter</FILTER>"
        "</COLLECTION>"
        "<SYSTEM TYPE='Formulae' NAME='BankOrCashFilter'>"
        "$$IsSameOrBelongsTo:$Parent:$$GroupSundryParent:BankAccounts OR "
        "$$IsSameOrBelongsTo:$Parent:$$GroupSundryParent:CashinHand"
        "</SYSTEM>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    # Fallback: simple collection without filter
    xml_fallback = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>AllLedgerCollection</ID></HEADER><BODY><DESC>"
        f"{static}"
        "<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='AllLedgerCollection'>"
        "<TYPE>Ledger</TYPE>"
        "<NATIVEMETHOD>Name</NATIVEMETHOD>"
        "<NATIVEMETHOD>Parent</NATIVEMETHOD>"
        "</COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    bank_groups = {"bank accounts", "bank account", "bank", "bank ods", "bank od accounts"}
    cash_groups = {"cash-in-hand", "cash in hand"}
    all_bank_groups = bank_groups | cash_groups

    def _is_valid_xml_codepoint(codepoint: int) -> bool:
        return (
            codepoint in (0x9, 0xA, 0xD)
            or (0x20 <= codepoint <= 0xD7FF)
            or (0xE000 <= codepoint <= 0xFFFD)
            or (0x10000 <= codepoint <= 0x10FFFF)
        )

    def _sanitize_tally_xml(text: str) -> str:
        if not text:
            return ""

        # Strip raw control characters disallowed in XML 1.0.
        cleaned = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

        # Drop malformed/invalid numeric character references that break XML parser.
        entity_pattern = re.compile(r"&#(x[0-9A-Fa-f]+|\d+);")

        def fix_entity(match):
            token = match.group(1)
            try:
                if token.lower().startswith("x"):
                    cp = int(token[1:], 16)
                else:
                    cp = int(token)
            except ValueError:
                return ""
            return match.group(0) if _is_valid_xml_codepoint(cp) else ""

        return entity_pattern.sub(fix_entity, cleaned)

    def _tag_name(tag) -> str:
        raw = str(tag or "")
        if "}" in raw:
            raw = raw.split("}", 1)[1]
        return raw.upper()

    def _extract_ledgers(resp_text: str, filter_groups=True) -> list:
        ledgers = []
        parse_candidates = [resp_text, _sanitize_tally_xml(resp_text)]
        parsed = False

        for candidate in parse_candidates:
            if parsed or not candidate:
                continue
            try:
                root = ET.fromstring(candidate)
            except ET.ParseError:
                continue

            parsed = True
            for ledger_node in root.iter():
                if _tag_name(ledger_node.tag) != "LEDGER":
                    continue

                name = _normalize_company_name(ledger_node.attrib.get("NAME") or "")
                parent = ""

                for child in ledger_node:
                    tag = _tag_name(child.tag)
                    text = _normalize_company_name(child.text)
                    if tag in {"NAME", "LEDGERNAME"} and text:
                        name = text
                    elif tag in {"PARENT", "PARENTGROUP"} and text:
                        parent = text

                if not name:
                    continue
                if filter_groups and parent.casefold() not in all_bank_groups:
                    continue
                ledgers.append({"name": name, "parent": parent})

        # Regex fallback for malformed XML that still couldn't be parsed.
        if not ledgers:
            try:
                ledger_pattern = re.compile(
                    r'<LEDGER\b[^>]*\bNAME="([^"]*)"[^>]*>(.*?)</LEDGER>',
                    flags=re.IGNORECASE | re.DOTALL,
                )
                parent_pattern = re.compile(
                    r'<PARENT(?:\s+[^>]*)?>(.*?)</PARENT>',
                    flags=re.IGNORECASE | re.DOTALL,
                )
                source = _sanitize_tally_xml(resp_text)
                for match in ledger_pattern.finditer(source):
                    name = _normalize_company_name(match.group(1))
                    block = match.group(2)
                    parent_match = parent_pattern.search(block)
                    parent = _normalize_company_name(parent_match.group(1) if parent_match else "")
                    if not name:
                        continue
                    if filter_groups and parent.casefold() not in all_bank_groups:
                        continue
                    ledgers.append({"name": name, "parent": parent})
            except Exception:
                pass
        return ledgers

    errors = []

    # Try filtered query first
    try:
        resp = _post_tally_xml(tally_url, xml_payload, timeout=timeout)
        # Keep client-side filtering enabled because some Tally setups ignore the TDL filter.
        ledgers = _extract_ledgers(resp, filter_groups=True)
        if ledgers:
            seen = set()
            unique = []
            for l in ledgers:
                key = l["name"].upper()
                if key not in seen:
                    seen.add(key)
                    unique.append(l)
            return {"success": True, "ledgers": sorted(unique, key=lambda x: x["name"].upper())}
    except Exception as exc:
        errors.append(f"filtered: {exc}")

    # Fallback: fetch all ledgers and filter client-side
    try:
        resp = _post_tally_xml(tally_url, xml_fallback, timeout=timeout)
        ledgers = _extract_ledgers(resp, filter_groups=True)
        if ledgers:
            seen = set()
            unique = []
            for l in ledgers:
                key = l["name"].upper()
                if key not in seen:
                    seen.add(key)
                    unique.append(l)
            return {"success": True, "ledgers": sorted(unique, key=lambda x: x["name"].upper())}
    except Exception as exc:
        errors.append(f"fallback: {exc}")

    err = "; ".join(errors) if errors else "No bank ledgers found."
    return {"success": False, "error": err, "ledgers": []}


# ─── Fetch Next Voucher Number ──────────────────────────────────────────

def _fetch_next_voucher_number(tally_url: str, company: str, voucher_type: str, timeout: float = 15.0) -> dict:
    static = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    static += f"<SVVOUCHERTYPENAME>{xml_escape(voucher_type)}</SVVOUCHERTYPENAME>"
    if company:
        static += f"<SVCURRENTCOMPANY>{xml_escape(company)}</SVCURRENTCOMPANY>"
    static += "</STATICVARIABLES>"

    xml_payload = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
        "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Vouchers</REPORTNAME>"
        f"{static}</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>"
    )

    try:
        resp = _post_tally_xml(tally_url, xml_payload, timeout=timeout)
        numbers = []
        for match in re.findall(r"<VOUCHERNUMBER>(.*?)</VOUCHERNUMBER>", resp, flags=re.IGNORECASE | re.DOTALL):
            text = str(match or "").strip()
            if text.endswith(".0") and text[:-2].isdigit():
                text = text[:-2]
            if text.isdigit():
                numbers.append(int(text))
        try:
            root = ET.fromstring(resp)
            for node in root.iter():
                if "VOUCHERNUM" in str(node.tag or "").upper():
                    t = str(node.text or "").strip()
                    if t.endswith(".0") and t[:-2].isdigit():
                        t = t[:-2]
                    if t.isdigit():
                        numbers.append(int(t))
        except ET.ParseError:
            pass

        if numbers:
            last = max(numbers)
            return {"success": True, "last_number": last, "next_number": last + 1}
        return {"success": True, "last_number": 0, "next_number": 1}
    except Exception as exc:
        return {"success": False, "last_number": 0, "next_number": 0, "error": str(exc)}


# ─── Generate Payment/Receipt Voucher XML ───────────────────────────────

def generate_bank_voucher_xml(
    rows: list,
    company: str,
    bank_ledger: str,
    use_today_date: bool = False,
    payment_start_vno: int = None,
    receipt_start_vno: int = None,
) -> str:
    """
    Generate Payment and Receipt vouchers from bank statement rows.
    Debit column → Payment voucher (money out: Bank CR, Contra Ledger DR)
    Credit column → Receipt voucher (money in: Bank DR, Contra Ledger CR)
    """
    period_from, period_to, period_current = _derive_import_period(rows, use_today_date)

    lines = []
    a = lines.append
    company_static = _company_static_block(
        company,
        from_date=period_from,
        to_date=period_to,
        current_date=period_current,
    )
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    bank_esc = xml_escape(bank_ledger)
    payment_counter = 0
    receipt_counter = 0

    for idx, r in enumerate(rows):
        if use_today_date:
            dt = datetime.today().strftime("%Y%m%d")
        else:
            source_date = r.get("DATE") or r.get("Date") or r.get("date") or ""
            parsed_source_date = _parse_statement_datetime(source_date)
            if parsed_source_date is None:
                raise ValueError(
                    f"Invalid DATE in row {idx + 1}: '{source_date}'. "
                    "Use Excel template date style like 01-Apr-2024."
                )
            dt = parsed_source_date.strftime("%Y%m%d")

        contra_ledger = str(r.get("LEDGER") or r.get("Ledger") or "Suspense A/c").strip()
        if not contra_ledger:
            contra_ledger = "Suspense A/c"
        contra_esc = xml_escape(contra_ledger)

        description = str(r.get("DESCRIPTION") or r.get("Description") or "").strip()
        cheque_no = str(r.get("CHEQUE NO.") or r.get("ChequeNo") or r.get("Cheque No") or "").strip()
        if cheque_no in ("None", "none", ""):
            cheque_no = ""

        narration_parts = []
        if description:
            narration_parts.append(description)
        if cheque_no:
            narration_parts.append(f"Chq: {cheque_no}")
        narration = xml_escape(" | ".join(narration_parts))

        # Parse debit and credit amounts
        debit_raw = r.get("Debit") or r.get("DEBIT") or r.get("debit") or 0
        credit_raw = r.get("Credit") or r.get("CREDIT") or r.get("credit") or 0

        try:
            debit_amt = float(debit_raw) if debit_raw not in (None, "", "None") else 0.0
        except (TypeError, ValueError):
            debit_amt = 0.0
        try:
            credit_amt = float(credit_raw) if credit_raw not in (None, "", "None") else 0.0
        except (TypeError, ValueError):
            credit_amt = 0.0

        if debit_amt <= 0 and credit_amt <= 0:
            continue  # Skip empty rows

        if debit_amt > 0:
            # PAYMENT voucher: money going OUT of bank
            vch_type = "Payment"
            amount = debit_amt
            payment_counter += 1
            if payment_start_vno is not None:
                vno = str(payment_start_vno + payment_counter - 1)
            else:
                vno = str(r.get("VoucherNo") or r.get("VchNo") or payment_counter)

            a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
            a(f'    <VOUCHER VCHTYPE="{vch_type}" ACTION="Create" OBJVIEW="Accounting Voucher View">')
            a(f'     <DATE>{dt}</DATE>')
            a(f'     <VOUCHERTYPENAME>{vch_type}</VOUCHERTYPENAME>')
            a(f'     <VOUCHERNUMBER>{xml_escape(vno)}</VOUCHERNUMBER>')
            a(f'     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>')
            a(f'     <PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW>')
            if narration:
                a(f'     <NARRATION>{narration}</NARRATION>')

            # Contra Ledger - DEBIT (money goes TO this ledger)
            a('     <ALLLEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{contra_esc}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(amount)}</AMOUNT>')
            a('     </ALLLEDGERENTRIES.LIST>')

            # Bank Ledger - CREDIT (money goes OUT of bank)
            a('     <ALLLEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{bank_esc}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(amount)}</AMOUNT>')
            a('     </ALLLEDGERENTRIES.LIST>')

            a('    </VOUCHER>')
            a('   </TALLYMESSAGE>')

        if credit_amt > 0:
            # RECEIPT voucher: money coming IN to bank
            vch_type = "Receipt"
            amount = credit_amt
            receipt_counter += 1
            if receipt_start_vno is not None:
                vno = str(receipt_start_vno + receipt_counter - 1)
            else:
                vno = str(r.get("VoucherNo") or r.get("VchNo") or receipt_counter)

            a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
            a(f'    <VOUCHER VCHTYPE="{vch_type}" ACTION="Create" OBJVIEW="Accounting Voucher View">')
            a(f'     <DATE>{dt}</DATE>')
            a(f'     <VOUCHERTYPENAME>{vch_type}</VOUCHERTYPENAME>')
            a(f'     <VOUCHERNUMBER>{xml_escape(vno)}</VOUCHERNUMBER>')
            a(f'     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>')
            a(f'     <PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW>')
            if narration:
                a(f'     <NARRATION>{narration}</NARRATION>')

            # Bank Ledger - DEBIT (money comes IN to bank)
            a('     <ALLLEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{bank_esc}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(amount)}</AMOUNT>')
            a('     </ALLLEDGERENTRIES.LIST>')

            # Contra Ledger - CREDIT (money comes FROM this ledger)
            a('     <ALLLEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{contra_esc}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(amount)}</AMOUNT>')
            a('     </ALLLEDGERENTRIES.LIST>')

            a('    </VOUCHER>')
            a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


# ─── Generate Ledger Master XML (for auto-creating missing ledgers) ─────

def generate_ledger_xml(ledgers: list, company: str) -> str:
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>All Masters</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')
    for led in ledgers:
        name = xml_escape(led["Name"])
        parent = xml_escape(led.get("Parent", "Suspense A/c"))
        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <LEDGER NAME="{name}" ACTION="Create">')
        a(f'     <NAME>{name}</NAME>')
        a(f'     <PARENT>{parent}</PARENT>')
        a('    </LEDGER>')
        a('   </TALLYMESSAGE>')
    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


STATEMENT_HEADERS = ["DATE", "CHEQUE NO.", "DESCRIPTION", "LEDGER", "Debit", "Credit", "Balance"]

_PDF_DATE_PATTERNS = [
    re.compile(
        r"^(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4})\b",
        re.IGNORECASE,
    ),
    re.compile(r"^(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b"),
    re.compile(r"^(\d{4}[/\-]\d{1,2}[/\-]\d{1,2})\b"),
    re.compile(
        r"^(\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*-\d{2,4})\b",
        re.IGNORECASE,
    ),
]

# Handles rows like "1 08/03/2026 narration ..."
_PDF_SERIAL_DATE_PATTERN = re.compile(r"^\d+\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+")

_PDF_DRCR_PATTERN = re.compile(
    r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+(.*?)\s+\b(DR|CR)\b\s+([\d,]+\.\d{1,2})\s+([\d,]+\.\d{1,2})\s*$"
)

# Handles rows like "12/03/2026 narration 1000.00(Dr) 25000.00(Cr)"
_PDF_DRCR_SUFFIX_PATTERN = re.compile(
    r"(\d{1,2}[\-/]\d{1,2}[\-/]\d{2,4})\s+"
    r"(.*?)\s+"
    r"([\d,]+\.\d{2})\((\w+)\)\s+"
    r"([\d,]+\.\d{2})\((\w+)\)\s*$"
)

_PDF_TRAILING_3_AMOUNTS = re.compile(
    r"([\d,]*\d+(?:\.\d{1,2})?)\s+([\d,]*\d+(?:\.\d{1,2})?)\s+([\d,]*\d+(?:\.\d{1,2})?)\s*$"
)
_PDF_TRAILING_2_AMOUNTS = re.compile(
    r"([\d,]*\d+(?:\.\d{1,2})?)\s+([\d,]*\d+(?:\.\d{1,2})?)\s*$"
)

_PDF_SKIP_PATTERNS = [
    re.compile(r"^\s*(?:Sr\s*$|No\.\s*$)", re.IGNORECASE),
    re.compile(r"^\s*Date\s+(?:Particulars|Description|Narration|Transaction)", re.IGNORECASE),
    re.compile(r"^\s*(?:This is a computer|Registered office|Acronyms:|Grievance|Page\s+\d)", re.IGNORECASE),
    re.compile(r"^\s*\*(?:Service Tax|Any discrepancies|This is a computer)", re.IGNORECASE),
    re.compile(r"https?://", re.IGNORECASE),
    re.compile(r"^\s*(?:No\.\s*Transactio|Amount\(INR|Balance\(INR)", re.IGNORECASE),
    re.compile(r"^\s*(?:DETAILED STATEMENT|Transactions List|STATEMENT SUMMARY|Opening Balance|Statement of account)", re.IGNORECASE),
    re.compile(r"^\s*(?:Generated On|Generated By|Requesting Branch|State account branch|HDFC Bank GSTIN)", re.IGNORECASE),
    re.compile(r"^\s*(?:Account Branch|City\b|Phone|OD Limit|Currency|Email|Cust ID|Account No|A/C Open|Account Status|RTGS|Branch Code|Account Type|Nomination|From\s*:|Period\b|Cust Reln)", re.IGNORECASE),
    re.compile(r"^\s*(?:Customer Name|Account Number|Report Generated|ADDRESS\s*:)", re.IGNORECASE),
    re.compile(r"^\s*(?:e-Pass Sheet|\(Report is generated)", re.IGNORECASE),
    re.compile(r"^\s*(?:HDFC BANK LIMITED|Contents of this statement)", re.IGNORECASE),
    re.compile(r"^\s*(?:Dr Count|Cr Count|Debits\b|Credits\b|Closing Bal)", re.IGNORECASE),
    re.compile(r"^\s*(?:M/S\.|JOINT HOLDERS)", re.IGNORECASE),
    re.compile(r"^\s*(?:Dear\s+\w|Customer ID|S/O:|Mob:|Email:)", re.IGNORECASE),
    re.compile(r"^\s*(?:Last \d+ Transactions|Account No:)", re.IGNORECASE),
    re.compile(r"^\s*(?:Date\s+Transactions\s+Details|Date\s+Narration\s+Chq)", re.IGNORECASE),
    re.compile(r"^\s*(?:Statement Summary|Opening Balance|Total Withdrawal|Total Deposit|Closing Balance|Withdrawal Count|Deposit Count)", re.IGNORECASE),
    re.compile(r"^\s*(?:End of Statement|Any discrepancy|system generated)", re.IGNORECASE),
    re.compile(r"^\s*(?:Nominee|Branch Address|Branch Phone|MICR Code|IFSC Code)", re.IGNORECASE),
]

_PDF_FOOTER_NOISE = [
    re.compile(r"(?:Phones?\s+Banking|reachus@|www\.|\.com|\.in)", re.IGNORECASE),
    re.compile(r"(?:Registered\s+office|Corporate\s+Identify|CIN\s*:)", re.IGNORECASE),
    re.compile(r"(?:Acronyms|MICR|IFSC\s+Code\s*:|NEFT\s*:|RTGS\s*:)", re.IGNORECASE),
    re.compile(r"(?:Grievance\s+Officer|https?://)", re.IGNORECASE),
    re.compile(r"(?:Service\s+Tax\s+Registration|discrepancies)", re.IGNORECASE),
    re.compile(r"(?:computer generated|not require signature|system generated)", re.IGNORECASE),
    re.compile(r"(?:Page\s+\d+\s+of\s+\d+)", re.IGNORECASE),
]

# Column matching to normalize table-based statements into template fields.
_PDF_COLUMN_KEYWORDS = {
    "date": ("date", "txn date", "transaction date", "value date"),
    "description": ("description", "narration", "particular", "transaction details", "details", "remarks"),
    "cheque": ("cheque", "chq", "cheque no", "instrument", "reference", "ref", "utr"),
    "drcr": ("dr/cr", "cr/dr", "dr cr", "drcr", "txn type", "tran type", "type"),
    "debit": ("debit", "withdrawal", "withdraw", "dr", "paid out"),
    "credit": ("credit", "credi", "deposit", "cr", "paid in"),
    "amount": ("amount", "txn amount", "transaction amount", "amt"),
    "balance": ("balance", "running balance", "closing balance", "avail bal", "avl bal"),
}


def _normalize_pdf_date_text(text: str) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""

    raw = re.sub(r"[\u2010-\u2015\u2212]", "-", raw)
    raw = re.sub(r"\s+", " ", raw).strip()

    m = re.match(r"^(\d{1,2})([\-/\s])([A-Za-z]{3,9}|\d{1,2})([\-/\s])(\d{3})$", raw)
    if m:
        day, sep1, month, sep2, year3 = m.groups()
        if year3.startswith("20"):
            decade = (datetime.today().year % 100) // 10
            year4 = f"20{decade}{year3[-1]}"
        else:
            year4 = f"20{year3[-2:]}"
        raw = f"{day}{sep1}{month}{sep2}{year4}"

    return raw


def _format_date_for_template(text: str) -> str:
    parsed = _parse_statement_datetime(text)
    if parsed:
        return parsed.strftime("%d-%b-%Y")
    return str(text or "").strip()


def _pdf_parse_amount(text: str) -> float:
    source = str(text or "").strip()
    if not source:
        return 0.0

    # Reject mixed identifier text like UPI refs, emails, and URLs.
    non_amount = re.sub(r"(?i)\b(dr|cr|rs|inr)\b", "", source)
    if re.search(r"[A-Za-z@/:|]", non_amount):
        return 0.0

    cleaned = re.sub(r"(?i)\b(dr|cr)\b", "", source)
    cleaned = re.sub(r"(?i)\b(?:rs|inr)\s*\.?\s*", "", cleaned)
    cleaned = re.sub(r"\((?:Dr|Cr)\)", "", cleaned, flags=re.IGNORECASE)
    cleaned = cleaned.replace(",", "").strip()
    cleaned = re.sub(r"[^0-9.\-()]", "", cleaned)
    if not cleaned:
        return 0.0

    digits_only = re.sub(r"\D", "", cleaned)
    if "." not in cleaned and len(digits_only) > 11:
        return 0.0

    # Normalize malformed decimal strings such as ".126.0" -> "126.0".
    if cleaned.count(".") > 1:
        negative = cleaned.startswith("-")
        body = cleaned[1:] if negative else cleaned
        parts = body.split(".")
        body = "".join(parts[:-1]) + "." + parts[-1]
        cleaned = ("-" if negative else "") + body

    cleaned = re.sub(r"^\.(\d)", r"0.\1", cleaned)

    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    try:
        return abs(float(cleaned))
    except ValueError:
        return 0.0


def _pdf_collect_trailing_amount_tokens(text: str, max_tokens: int = 3) -> tuple:
    """
    Collect amount-like tokens from the end of a statement line.
    Stops at the first non-amount token, so date fragments like 09/04/25
    are not mistaken as trailing amounts.
    Returns (tokens_left_to_right, middle_text_before_amounts).
    """
    parts = re.split(r"\s+", str(text or "").strip())
    if not parts:
        return [], ""

    amounts_rev = []
    idx = len(parts) - 1

    while idx >= 0 and len(amounts_rev) < max_tokens:
        token = parts[idx].strip()
        if _pdf_is_amount_cell(token):
            amounts_rev.append(token)
            idx -= 1
            continue
        break

    if len(amounts_rev) < 2:
        return [], str(text or "").strip()

    amount_tokens = list(reversed(amounts_rev))
    middle = " ".join(parts[: idx + 1]).strip()
    return amount_tokens, middle


def _pdf_is_amount_cell(text: str) -> bool:
    raw = str(text or "").strip()
    if not raw or not re.search(r"\d", raw):
        return False
    if _parse_statement_datetime(raw):
        return False

    compact = re.sub(r"(?i)\b(inr|rs\.?)\b", "", raw)
    compact = re.sub(r"\s+", "", compact)
    match_ok = bool(
        re.fullmatch(
            r"\(?[+\-]?[\d,]+(?:\.\d{1,2})?\)?(?:\(?\s*(?:dr|cr)\s*\)?)?",
            compact,
            flags=re.IGNORECASE,
        )
    )
    if not match_ok:
        return False

    # Avoid treating long IDs (e.g. UPI RRN/account numbers) as amounts.
    digits_only = re.sub(r"\D", "", compact)
    if "." not in compact and len(digits_only) > 11:
        return False
    return True


def _pdf_match_date_prefix(line: str):
    stripped = line.strip()
    for pat in _PDF_DATE_PATTERNS:
        m = pat.match(stripped)
        if m:
            date_token = _normalize_pdf_date_text(m.group(1))
            return date_token, stripped[m.end():].strip()

    m = _PDF_SERIAL_DATE_PATTERN.match(stripped)
    if m:
        date_token = _normalize_pdf_date_text(m.group(1))
        return date_token, stripped[m.end():].strip()
    return None


def _pdf_is_skip_line(line: str) -> bool:
    return any(pat.search(line) for pat in _PDF_SKIP_PATTERNS)


def _pdf_is_footer_noise(line: str) -> bool:
    return any(pat.search(line) for pat in _PDF_FOOTER_NOISE)


def _normalize_pdf_header_token(text: str) -> str:
    token = str(text or "").strip().lower()
    if not token:
        return ""
    token = token.replace("&", " and ")
    token = re.sub(r"[\(\)\[\]\.:]", " ", token)
    token = re.sub(r"[^a-z0-9/\- ]", " ", token)
    token = re.sub(r"\s+", " ", token).strip()
    return token


def _pdf_detect_table_column_map(header_cells: list) -> tuple:
    mapping = {key: None for key in _PDF_COLUMN_KEYWORDS.keys()}
    score = 0

    for idx, cell in enumerate(header_cells):
        token = _normalize_pdf_header_token(cell)
        if not token:
            continue

        for key, keywords in _PDF_COLUMN_KEYWORDS.items():
            if mapping[key] is not None:
                continue
            if any(word in token for word in keywords):
                mapping[key] = idx
                score += 1
                break

    return mapping, score


def _pdf_table_cell(cells: list, idx) -> str:
    if idx is None:
        return ""
    if idx < 0 or idx >= len(cells):
        return ""
    return str(cells[idx] or "").strip()


def _pdf_parse_drcr_side(text: str) -> str:
    source = str(text or "")
    if not source:
        return ""

    has_dr = bool(re.search(r"\bDR\b", source, flags=re.IGNORECASE))
    has_cr = bool(re.search(r"\bCR\b", source, flags=re.IGNORECASE))

    if has_dr and not has_cr:
        return "DR"
    if has_cr and not has_dr:
        return "CR"

    normalized = _normalize_pdf_header_token(source)
    if normalized in {"d", "db", "debit"}:
        return "DR"
    if normalized in {"c", "cr", "credit"}:
        return "CR"
    return ""


def _pdf_append_continuation(row: dict, extra: str):
    extra_text = re.sub(r"\s+", " ", str(extra or "")).strip()
    if not extra_text:
        return
    current = str(row.get("DESCRIPTION") or "").strip()
    row["DESCRIPTION"] = f"{current} {extra_text}".strip() if current else extra_text


def _configure_tesseract_cmd() -> str:
    try:
        import pytesseract
    except ImportError:
        return ""

    candidates = [
        os.environ.get("TESSERACT_CMD", "").strip(),
        shutil.which("tesseract") or "",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "Programs", "Tesseract-OCR", "tesseract.exe"),
    ]

    for candidate in candidates:
        if candidate and os.path.isfile(candidate):
            pytesseract.pytesseract.tesseract_cmd = candidate
            return candidate
    return ""


def _extract_text_lines(filepath: str) -> tuple:
    """Extract lines from text PDF, or OCR image-based PDFs. Returns (lines, used_ocr)."""
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("Install pdfplumber: pip install pdfplumber") from exc

    lines = []
    used_ocr = False

    with pdfplumber.open(filepath) as pdf:
        if not pdf.pages:
            raise ValueError("PDF has no pages.")

        first_page = pdf.pages[0]
        has_chars = bool(first_page.chars)
        first_text = (first_page.extract_text() or "").strip()

        if has_chars and len(first_text) > 50:
            for page in pdf.pages:
                lines.extend((page.extract_text() or "").split("\n"))
        else:
            used_ocr = True

    if used_ocr:
        try:
            import pytesseract
        except ImportError as exc:
            raise RuntimeError(
                "This PDF is image-based and requires OCR.\n"
                "Install: pip install pytesseract\n"
                "Also install Tesseract OCR on your system."
            ) from exc

        tesseract_cmd = _configure_tesseract_cmd()
        if not tesseract_cmd:
            raise RuntimeError(
                "OCR engine not found. Install Tesseract OCR and ensure it is in PATH.\n"
                "If installed already, set TESSERACT_CMD to full path, for example:\n"
                "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
            )

        images = None
        pdf2image_err = ""
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(filepath, dpi=250)
        except Exception as exc:
            pdf2image_err = str(exc)

        if images is None:
            pdfium_err = ""
            try:
                import pypdfium2 as pdfium

                images = []
                pdf_doc = pdfium.PdfDocument(filepath)
                for page in pdf_doc:
                    pil_img = page.render(scale=2.5).to_pil()
                    images.append(pil_img)
            except Exception as exc:
                pdfium_err = str(exc)
                raise RuntimeError(
                    "Could not convert PDF pages for OCR.\n"
                    f"pdf2image error: {pdf2image_err}\n"
                    f"pypdfium2 fallback error: {pdfium_err}\n\n"
                    "Fix options:\n"
                    "1) Install Poppler and add it to PATH (for pdf2image), or\n"
                    "2) Install pypdfium2: pip install pypdfium2"
                )

        try:
            for img in images:
                text = pytesseract.image_to_string(img, config="--psm 6")
                lines.extend(text.strip().split("\n"))
        except pytesseract.TesseractNotFoundError as exc:
            raise RuntimeError(
                "OCR engine not found. Install Tesseract OCR and ensure it is in PATH.\n"
                "Windows install: https://github.com/UB-Mannheim/tesseract/wiki"
            ) from exc

    if not lines or all(not line.strip() for line in lines):
        raise ValueError("PDF contains no extractable text and OCR produced no results.")

    return lines, used_ocr


def _pdf_strategy_table_columnmap(filepath: str) -> list:
    """Strategy: Parse structured tables by mapping column names to required fields."""
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []
    prev_balance = 0.0

    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    cleaned_rows = []
                    for raw_row in table or []:
                        if not raw_row:
                            continue
                        cells = [str(c or "").replace("\n", " ").strip() for c in raw_row]
                        if any(cells):
                            cleaned_rows.append(cells)
                    if not cleaned_rows:
                        continue

                    best_header_idx = -1
                    best_mapping = None
                    best_score = -1

                    for idx, candidate in enumerate(cleaned_rows[:4]):
                        mapping, score = _pdf_detect_table_column_map(candidate)
                        has_date = mapping.get("date") is not None
                        has_amount_column = any(
                            mapping.get(k) is not None
                            for k in ("debit", "credit", "amount")
                        )
                        if has_date and has_amount_column and score > best_score:
                            best_score = score
                            best_header_idx = idx
                            best_mapping = mapping

                    if best_mapping is None:
                        continue

                    for cells in cleaned_rows[best_header_idx + 1:]:
                        date_text = _pdf_table_cell(cells, best_mapping.get("date"))
                        desc_text = _pdf_table_cell(cells, best_mapping.get("description"))
                        cheque_text = _pdf_table_cell(cells, best_mapping.get("cheque"))
                        drcr_text = _pdf_table_cell(cells, best_mapping.get("drcr"))

                        parsed_date = _parse_statement_datetime(date_text)
                        if parsed_date is None:
                            continuation = " ".join(x for x in [desc_text, cheque_text] if x).strip()
                            if continuation and rows:
                                _pdf_append_continuation(rows[-1], continuation)
                            continue

                        debit = _pdf_parse_amount(_pdf_table_cell(cells, best_mapping.get("debit")))
                        credit = _pdf_parse_amount(_pdf_table_cell(cells, best_mapping.get("credit")))
                        amount = _pdf_parse_amount(_pdf_table_cell(cells, best_mapping.get("amount")))
                        balance = _pdf_parse_amount(_pdf_table_cell(cells, best_mapping.get("balance")))

                        if debit <= 0 and credit <= 0 and amount > 0:
                            side = _pdf_parse_drcr_side(drcr_text)
                            if side == "DR":
                                debit = amount
                            elif side == "CR":
                                credit = amount
                            elif balance > 0 and prev_balance > 0:
                                if balance >= prev_balance:
                                    credit = amount
                                else:
                                    debit = amount
                            else:
                                debit = amount

                        if debit <= 0 and credit <= 0:
                            continue

                        description = desc_text or "Bank Transaction"
                        row = {
                            "DATE": parsed_date.strftime("%d-%b-%Y"),
                            "CHEQUE NO.": cheque_text,
                            "DESCRIPTION": description,
                            "LEDGER": "Suspense A/c",
                            "Debit": debit if debit > 0 else "",
                            "Credit": credit if credit > 0 else "",
                            "Balance": balance if balance > 0 else "",
                        }
                        rows.append(row)

                        if balance > 0:
                            prev_balance = balance
    except Exception:
        pass

    return rows


def _pdf_strategy_table(filepath: str) -> list:
    """Strategy: Generic table extraction fallback when column headers are unclear."""
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []
    prev_balance = 0.0

    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    for raw_row in table or []:
                        if not raw_row:
                            continue

                        cells = [str(c or "").replace("\n", " ").strip() for c in raw_row if str(c or "").strip()]
                        if len(cells) < 2:
                            continue

                        date_text = ""
                        for cell in cells:
                            if _parse_statement_datetime(cell):
                                date_text = cell
                                break
                        if not date_text:
                            continue

                        amount_cells = [cell for cell in cells if _pdf_is_amount_cell(cell)]
                        amounts = [_pdf_parse_amount(cell) for cell in amount_cells if _pdf_parse_amount(cell) > 0]

                        if not amounts:
                            continue

                        desc_parts = []
                        for cell in cells:
                            if cell == date_text:
                                continue
                            if _pdf_is_amount_cell(cell):
                                continue
                            desc_parts.append(cell)
                        description = " ".join(desc_parts).strip() or "Bank Transaction"

                        debit = 0.0
                        credit = 0.0
                        balance = 0.0

                        if len(amounts) >= 3:
                            debit = amounts[0]
                            credit = amounts[1]
                            balance = amounts[2]
                        elif len(amounts) == 2:
                            amount, balance = amounts
                            if balance >= prev_balance:
                                credit = amount
                            else:
                                debit = amount
                        else:
                            amount = amounts[0]
                            side = ""
                            for cell in amount_cells:
                                side = _pdf_parse_drcr_side(cell)
                                if side:
                                    break
                            if side == "CR":
                                credit = amount
                            else:
                                debit = amount

                        if debit <= 0 and credit <= 0:
                            continue

                        rows.append(
                            {
                                "DATE": _format_date_for_template(date_text),
                                "CHEQUE NO.": "",
                                "DESCRIPTION": description,
                                "LEDGER": "Suspense A/c",
                                "Debit": debit if debit > 0 else "",
                                "Credit": credit if credit > 0 else "",
                                "Balance": balance if balance > 0 else "",
                            }
                        )

                        if balance > 0:
                            prev_balance = balance
    except Exception:
        pass

    return rows


def _pdf_strategy_drcr(lines: list) -> list:
    """Strategy: DR/CR indicator rows (common in ICICI and Axis formats)."""
    rows = []
    for line in lines:
        stripped = line.strip()
        if not stripped or _pdf_is_skip_line(stripped):
            continue

        match = _PDF_DRCR_PATTERN.search(stripped)
        if not match:
            continue

        date_text, mid, drcr, amount_text, balance_text = match.groups()
        description = re.sub(
            r"^\d{1,2}:\d{2}:\d{2}\s*(AM|PM)?\s*[-\u2013]?\s*",
            "",
            mid.strip(),
            flags=re.IGNORECASE,
        ).lstrip("- ").strip()

        amount = _pdf_parse_amount(amount_text)
        balance = _pdf_parse_amount(balance_text)
        if amount <= 0:
            continue

        rows.append(
            {
                "DATE": _format_date_for_template(date_text),
                "CHEQUE NO.": "",
                "DESCRIPTION": description,
                "LEDGER": "Suspense A/c",
                "Debit": amount if drcr == "DR" else "",
                "Credit": amount if drcr == "CR" else "",
                "Balance": balance if balance > 0 else "",
            }
        )

    return rows


def _pdf_strategy_drcr_suffix(lines: list) -> list:
    """Strategy: amount(Dr)/amount(Cr) style rows (common in Kotak-like layouts)."""
    rows = []
    current_txn = None

    for line in lines:
        stripped = line.strip()
        if not stripped or _pdf_is_skip_line(stripped):
            continue

        m = _PDF_DRCR_SUFFIX_PATTERN.search(stripped)
        if m:
            if current_txn:
                rows.append(current_txn)

            date_text, mid, amount_text, drcr_side, balance_text, _balance_side = m.groups()
            amount = _pdf_parse_amount(amount_text)
            balance = _pdf_parse_amount(balance_text)
            if amount <= 0:
                current_txn = None
                continue

            is_debit = _pdf_parse_drcr_side(drcr_side) == "DR"
            current_txn = {
                "DATE": _format_date_for_template(date_text),
                "CHEQUE NO.": "",
                "DESCRIPTION": mid.strip(),
                "LEDGER": "Suspense A/c",
                "Debit": amount if is_debit else "",
                "Credit": amount if not is_debit else "",
                "Balance": balance if balance > 0 else "",
            }
        elif current_txn is not None and not _pdf_is_footer_noise(stripped):
            _pdf_append_continuation(current_txn, stripped)

    if current_txn:
        rows.append(current_txn)

    return rows


def _pdf_strategy_dateprefix(lines: list) -> list:
    """Strategy: Date-prefix rows with trailing amount columns."""
    raw_transactions = []
    current_txn = None

    for line in lines:
        stripped = line.strip()
        if not stripped or _pdf_is_skip_line(stripped):
            continue

        date_match = _pdf_match_date_prefix(stripped)
        if date_match:
            if current_txn:
                raw_transactions.append(current_txn)
            date_str, rest = date_match
            current_txn = {"date": date_str, "first_line": rest, "continuation": []}
        elif current_txn is not None and not _pdf_is_footer_noise(stripped):
            current_txn["continuation"].append(stripped)

    if current_txn:
        raw_transactions.append(current_txn)

    rows = []
    prev_balance = 0.0

    for txn in raw_transactions:
        first_line = txn["first_line"]
        if not first_line:
            continue

        withdrawal = 0.0
        deposit = 0.0
        balance = 0.0

        amount_tokens, middle = _pdf_collect_trailing_amount_tokens(first_line, max_tokens=3)
        if not amount_tokens:
            continue

        # Strip trailing value-date token that can appear before amount columns.
        middle = re.sub(
            r"\s*\b(?:\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*-\d{2,4})\b\s*$",
            "",
            middle,
            flags=re.IGNORECASE,
        ).strip()

        if len(amount_tokens) >= 3:
            withdrawal = _pdf_parse_amount(amount_tokens[-3])
            deposit = _pdf_parse_amount(amount_tokens[-2])
            balance = _pdf_parse_amount(amount_tokens[-1])

            # If both withdrawal and deposit are positive, treat this as a
            # single-amount row inferred via running balance.
            if withdrawal > 0 and deposit > 0:
                amount1 = deposit
                if balance >= prev_balance:
                    deposit = amount1
                    withdrawal = 0.0
                else:
                    withdrawal = amount1
                    deposit = 0.0
        else:
            amount1 = _pdf_parse_amount(amount_tokens[-2])
            balance = _pdf_parse_amount(amount_tokens[-1])
            if balance >= prev_balance:
                deposit = amount1
            else:
                withdrawal = amount1

        middle = re.sub(r"\s*\|\s*", " ", middle).strip()
        description = middle
        cheque_ref = ""

        if middle:
            parts = middle.rsplit(None, 1)
            if len(parts) == 2 and (re.search(r"\d", parts[1]) or re.search(r"[\-/]", parts[1])):
                description, cheque_ref = parts

        if txn["continuation"]:
            extra = " ".join(c for c in txn["continuation"] if not _pdf_is_footer_noise(c))
            extra = re.sub(r"\s+", " ", extra).strip()
            if extra:
                description = (description + " " + extra).strip() if description else extra

        if withdrawal <= 0 and deposit <= 0:
            continue

        if balance > 0:
            prev_balance = balance

        rows.append(
            {
                "DATE": _format_date_for_template(txn["date"]),
                "CHEQUE NO.": cheque_ref,
                "DESCRIPTION": description,
                "LEDGER": "Suspense A/c",
                "Debit": withdrawal if withdrawal > 0 else "",
                "Credit": deposit if deposit > 0 else "",
                "Balance": balance if balance > 0 else "",
            }
        )

    return rows


def read_pdf_statement(filepath: str) -> tuple:
    """Multi-strategy parser that normalizes statement PDFs into template columns."""
    lines, used_ocr = _extract_text_lines(filepath)

    strategy_results = [
        _pdf_strategy_dateprefix(lines),
        _pdf_strategy_drcr(lines),
        _pdf_strategy_drcr_suffix(lines),
    ]

    best_rows = []
    for rows in strategy_results:
        if len(rows) > len(best_rows):
            best_rows = rows

    # Table strategies are useful for bank PDFs with explicit column headers.
    if len(best_rows) < 3 and not used_ocr:
        table_mapped_rows = _pdf_strategy_table_columnmap(filepath)
        table_fallback_rows = _pdf_strategy_table(filepath)

        if len(table_mapped_rows) > len(best_rows):
            best_rows = table_mapped_rows
        if len(table_fallback_rows) > len(best_rows):
            best_rows = table_fallback_rows

    if best_rows:
        return STATEMENT_HEADERS[:], best_rows

    note = " OCR was used for this image-based PDF." if used_ocr else ""
    raise ValueError(
        "Could not extract transactions."
        f"{note}\n"
        "Supported: Date+Amounts, DR/CR, Amount(Dr/Cr), and column-based tables with OCR fallback.\n"
        "If your bank layout is still unsupported, export Excel from bank portal and load it directly."
    )


PDF_COLUMN_MAPPING_FIELDS = [
    ("date", "Date Column", True),
    ("description", "Description/Narration Column", True),
    ("cheque", "Cheque/Reference Column", False),
    ("debit", "Debit/Withdrawal Column", False),
    ("credit", "Credit/Deposit Column", False),
    ("amount", "Amount Column (single amount statements)", False),
    ("drcr", "DR/CR Type Column", False),
    ("balance", "Balance Column", False),
]

_PDF_INLINE_DATE_PATTERNS = [
    re.compile(r"\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}"),
    re.compile(r"\d{4}[/\-]\d{1,2}[/\-]\d{1,2}"),
    re.compile(r"\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4}", re.IGNORECASE),
    re.compile(r"\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*-\d{2,4}", re.IGNORECASE),
]


def _pdf_make_unique_headers(header_cells: list) -> list:
    headers = []
    used = {}

    for idx, cell in enumerate(header_cells):
        raw = str(cell or "").strip()
        if raw:
            base = re.sub(r"\s+", " ", raw)
        else:
            base = f"Column {idx + 1}"

        count = used.get(base, 0)
        used[base] = count + 1
        headers.append(base if count == 0 else f"{base} ({count + 1})")

    return headers


def _pdf_extract_positional_statement_rows(filepath: str) -> tuple:
    """
    Fallback extractor for fixed-layout statements where table extraction fails,
    typically with columns like:
    DATE | PARTICULARS | CHQ.NO. | WITHDRAWALS | DEPOSITS | BALANCE
    """
    try:
        import pdfplumber
    except ImportError:
        return [], []

    canonical_headers = ["Date", "Particulars", "Chq.No.", "Withdrawals", "Deposits", "Balance"]
    all_rows = []

    def _group_lines(words: list, y_tol: float = 2.5):
        grouped = []
        for word in sorted(words, key=lambda w: (float(w.get("top", 0.0)), float(w.get("x0", 0.0)))):
            top = float(word.get("top", 0.0))
            if not grouped or abs(top - grouped[-1]["top"]) > y_tol:
                grouped.append({"top": top, "words": [word]})
            else:
                grouped[-1]["words"].append(word)
        return grouped

    def _header_positions(line_words: list):
        pos = {}
        for w in sorted(line_words, key=lambda x: float(x.get("x0", 0.0))):
            token = _normalize_pdf_header_token(w.get("text", ""))
            xmid = (float(w.get("x0", 0.0)) + float(w.get("x1", 0.0))) / 2.0

            if "date" in token and "Date" not in pos:
                pos["Date"] = xmid
            elif any(k in token for k in ("particular", "narration", "description")) and "Particulars" not in pos:
                pos["Particulars"] = xmid
            elif any(k in token for k in ("chq", "cheque", "ref")) and "Chq.No." not in pos:
                pos["Chq.No."] = xmid
            elif any(k in token for k in ("withdraw", "debit")) and "Withdrawals" not in pos:
                pos["Withdrawals"] = xmid
            elif any(k in token for k in ("deposit", "credit", "credi")) and "Deposits" not in pos:
                pos["Deposits"] = xmid
            elif "balance" in token and "Balance" not in pos:
                pos["Balance"] = xmid

        return pos

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=2, y_tolerance=2, keep_blank_chars=False) or []
            if not words:
                continue

            lines = _group_lines(words)
            header_idx = -1
            col_pos = {}

            for idx, line in enumerate(lines):
                line_words = line["words"]
                line_text = " ".join(w.get("text", "") for w in sorted(line_words, key=lambda x: float(x.get("x0", 0.0))))
                normalized = _normalize_pdf_header_token(line_text)
                if (
                    "date" in normalized
                    and any(k in normalized for k in ("particular", "narration", "description"))
                    and "balance" in normalized
                    and any(k in normalized for k in ("withdraw", "debit", "deposit", "credit", "credi"))
                ):
                    col_pos = _header_positions(line_words)
                    if "Date" in col_pos and "Particulars" in col_pos and "Balance" in col_pos:
                        header_idx = idx
                        break

            if header_idx < 0:
                continue

            if "Chq.No." not in col_pos:
                left = col_pos.get("Particulars")
                right = col_pos.get("Withdrawals") or col_pos.get("Deposits") or col_pos.get("Balance")
                if left is not None and right is not None and right > left:
                    col_pos["Chq.No."] = (left + right) / 2.0

            ordered_cols = sorted(col_pos.items(), key=lambda kv: kv[1])
            col_names = [name for name, _x in ordered_cols]
            col_x = [x for _name, x in ordered_cols]
            if not col_names:
                continue

            boundaries = []
            for i in range(len(col_x) - 1):
                boundaries.append((col_x[i] + col_x[i + 1]) / 2.0)

            def _bucket(xmid: float) -> str:
                for i, b in enumerate(boundaries):
                    if xmid <= b:
                        return col_names[i]
                return col_names[-1]

            current: dict | None = None

            for line in lines[header_idx + 1 :]:
                line_words = sorted(line["words"], key=lambda x: float(x.get("x0", 0.0)))
                if not line_words:
                    continue

                row = {h: "" for h in canonical_headers}
                for w in line_words:
                    text = str(w.get("text", "") or "").strip()
                    if not text:
                        continue
                    xmid = (float(w.get("x0", 0.0)) + float(w.get("x1", 0.0))) / 2.0
                    key = _bucket(xmid)
                    if key not in row:
                        continue
                    row[key] = (row[key] + " " + text).strip() if row[key] else text

                line_text = " ".join(v for v in row.values() if v).strip()
                if not line_text:
                    continue

                lower = line_text.lower()
                if any(k in lower for k in ("opening balance", "closing balance", "account summary", "statement summary")):
                    continue

                row_date = _pdf_parse_date_value(row.get("Date", ""))
                if row_date is not None:
                    if current:
                        all_rows.append(current)
                    current = row
                elif current:
                    part = str(row.get("Particulars", "") or "").strip()
                    if part:
                        current["Particulars"] = (current.get("Particulars", "") + " " + part).strip()

                    chq = str(row.get("Chq.No.", "") or "").strip()
                    if chq:
                        if not str(current.get("Chq.No.", "") or "").strip():
                            current["Chq.No."] = chq
                        elif chq not in str(current.get("Chq.No.", "")):
                            current["Chq.No."] = (str(current.get("Chq.No.", "")).strip() + " " + chq).strip()

                    # Some PDFs split debit/credit/balance into a separate line without date.
                    for key in ("Withdrawals", "Deposits", "Balance"):
                        cell = str(row.get(key, "") or "").strip()
                        if not cell or str(current.get(key, "") or "").strip():
                            continue
                        if not _pdf_is_amount_cell(cell):
                            continue
                        if _pdf_parse_amount(cell) <= 0:
                            continue
                        current[key] = cell

            if current:
                all_rows.append(current)

    normalized_rows = []
    for row in all_rows:
        normalized_rows.append({h: str(row.get(h, "") or "").strip() for h in canonical_headers})

    # Keep rows where either date or transaction fields are present.
    filtered_rows = []
    for row in normalized_rows:
        if _pdf_parse_date_value(row.get("Date", "")) is not None:
            filtered_rows.append(row)

    return canonical_headers, filtered_rows


def _pdf_extract_raw_table_rows(filepath: str) -> tuple:
    """Extract raw table rows from PDF before mapping into template columns."""
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("Install pdfplumber: pip install pdfplumber") from exc

    all_headers = []
    all_rows = []

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                cleaned_rows = []
                for raw_row in table or []:
                    if not raw_row:
                        continue
                    cells = [str(c or "").replace("\n", " ").strip() for c in raw_row]
                    if any(cells):
                        cleaned_rows.append(cells)

                if not cleaned_rows:
                    continue

                best_header_idx = -1
                best_score = -1
                # Some statements place account profile text before actual table headers.
                # Scan deeper than the first few rows so headers like
                # "Date / Transactions Details / Debit / Credit / Balance"
                # are still detected.
                scan_limit = min(len(cleaned_rows), 40)
                for row_idx, row_cells in enumerate(cleaned_rows[:scan_limit]):
                    _, score = _pdf_detect_table_column_map(row_cells)
                    if score > best_score:
                        best_score = score
                        best_header_idx = row_idx

                # Require at least two matched header groups to avoid
                # misclassifying data/noise rows as headers.
                if best_score >= 2:
                    table_headers = _pdf_make_unique_headers(cleaned_rows[best_header_idx])
                    data_rows = cleaned_rows[best_header_idx + 1 :]
                else:
                    max_cols = max(len(r) for r in cleaned_rows)
                    table_headers = [f"Column {i + 1}" for i in range(max_cols)]
                    data_rows = cleaned_rows

                for h in table_headers:
                    if h not in all_headers:
                        all_headers.append(h)

                for row in data_rows:
                    row_dict = {}
                    for i, header in enumerate(table_headers):
                        row_dict[header] = str(row[i] if i < len(row) else "").strip()
                    if any(str(v or "").strip() for v in row_dict.values()):
                        all_rows.append(row_dict)

    normalized_rows = []
    for row in all_rows:
        normalized_rows.append({h: row.get(h, "") for h in all_headers})

    if normalized_rows:
        return all_headers, normalized_rows

    pos_headers, pos_rows = _pdf_extract_positional_statement_rows(filepath)
    if pos_rows:
        return pos_headers, pos_rows

    return all_headers, normalized_rows


def _pdf_parse_date_value(text: str):
    value = str(text or "").strip()
    parsed = _parse_statement_datetime(value)
    if parsed is None and value:
        tokens = _pdf_extract_date_tokens(value)
        if tokens:
            parsed = _parse_statement_datetime(tokens[0])
    return parsed


def _pdf_guess_column_mapping(headers: list, rows: list = None) -> dict:
    mapping = {field: None for field, _label, _required in PDF_COLUMN_MAPPING_FIELDS}
    if not headers:
        return mapping

    normalized = {h: _normalize_pdf_header_token(h) for h in headers}

    def _header_has_keyword(token: str, word: str) -> bool:
        if not token or not word:
            return False
        # Prevent false positives like "description" matching "cr".
        if len(word) <= 2:
            return bool(re.search(rf"(?<![a-z0-9]){re.escape(word)}(?![a-z0-9])", token))
        return word in token

    def candidates(keyword_group: str):
        words = _PDF_COLUMN_KEYWORDS.get(keyword_group, ())
        return [h for h, token in normalized.items() if any(_header_has_keyword(token, word) for word in words)]

    # Header-only fallback when row data is not available.
    if not rows:
        for key in ("date", "description", "cheque", "debit", "credit", "amount", "drcr", "balance"):
            cand = candidates(key)
            mapping[key] = cand[0] if cand else None
        return mapping

    sample_rows = [r for r in rows[:500] if any(str(v or "").strip() for v in r.values())]
    if not sample_rows:
        for key in ("date", "description", "cheque", "debit", "credit", "amount", "drcr", "balance"):
            cand = candidates(key)
            mapping[key] = cand[0] if cand else None
        return mapping

    def pick_best(keyword_group: str, scorer):
        cand = candidates(keyword_group)
        if not cand:
            return None

        best = None
        best_score = -1
        for header in cand:
            score = scorer(header)
            token = normalized.get(header, "")

            # Prefer transactional columns over summary columns.
            if keyword_group in {"debit", "credit"} and ("opening" in token or "closing" in token):
                score -= 2
            if keyword_group == "balance" and token == "opening balance":
                score -= 2

            if score > best_score:
                best = header
                best_score = score

        if best_score <= 0:
            return None
        return best

    def score_date(header):
        count = 0
        for row in sample_rows:
            if _pdf_parse_date_value(row.get(header, "")) is not None:
                count += 1
        return count

    mapping["date"] = pick_best("date", score_date)
    if mapping["date"] is None:
        date_cand = candidates("date")
        mapping["date"] = date_cand[0] if date_cand else None

    date_rows = []
    if mapping["date"]:
        for row in sample_rows:
            if _pdf_parse_date_value(row.get(mapping["date"], "")) is not None:
                date_rows.append(row)
    if not date_rows:
        date_rows = sample_rows

    def score_text(header):
        count = 0
        for row in date_rows:
            val = str(row.get(header, "") or "").strip()
            if not val:
                continue
            if _pdf_is_amount_cell(val):
                continue
            if _pdf_parse_date_value(val) is not None:
                continue
            count += 1
        return count

    def score_amount(header):
        count = 0
        for row in date_rows:
            val = str(row.get(header, "") or "").strip()
            if not val:
                continue
            if _pdf_parse_amount(val) > 0:
                count += 1
        return count

    def score_side(header):
        count = 0
        for row in date_rows:
            side = _pdf_parse_drcr_side(row.get(header, ""))
            if side in {"DR", "CR"}:
                count += 1
        return count

    mapping["description"] = pick_best("description", score_text)
    mapping["cheque"] = pick_best("cheque", score_text)
    mapping["debit"] = pick_best("debit", score_amount)
    mapping["credit"] = pick_best("credit", score_amount)
    mapping["drcr"] = pick_best("drcr", score_side)
    mapping["balance"] = pick_best("balance", score_amount)

    amount_pick = pick_best("amount", score_amount)
    if amount_pick and amount_pick not in {mapping["debit"], mapping["credit"]}:
        mapping["amount"] = amount_pick
    else:
        mapping["amount"] = None

    return mapping


def _pdf_extract_date_tokens(text: str) -> list:
    source = str(text or "")
    tokens = []
    seen = set()

    for pattern in _PDF_INLINE_DATE_PATTERNS:
        for match in pattern.finditer(source):
            token = _normalize_pdf_date_text(match.group(0))
            if not token:
                continue
            key = token.upper()
            if key in seen:
                continue
            seen.add(key)
            tokens.append(token)

    return tokens


def _pdf_is_stage_quality_low(headers: list, rows: list) -> bool:
    """Detect unusable staged tables where many transactions are merged into single cells."""
    if not headers or not rows:
        return True

    guessed = _pdf_guess_column_mapping(headers, rows)
    date_col = guessed.get("date")
    desc_col = guessed.get("description")
    if not date_col or not desc_col:
        return True

    sample = []
    for row in rows[:80]:
        if any(str(v or "").strip() for v in row.values()):
            sample.append(row)
    if not sample:
        return True

    date_values = [str(r.get(date_col, "") or "").strip() for r in sample if str(r.get(date_col, "") or "").strip()]
    if not date_values:
        return True

    parsed_count = 0
    packed_count = 0
    for value in date_values:
        tokens = _pdf_extract_date_tokens(value)
        if len(tokens) > 1:
            packed_count += 1

        parsed = _parse_statement_datetime(value)
        if parsed is None and tokens:
            parsed = _parse_statement_datetime(tokens[0])
        if parsed is not None:
            parsed_count += 1

    parse_ratio = parsed_count / max(1, len(date_values))
    packed_ratio = packed_count / max(1, len(date_values))

    return parse_ratio < 0.45 or packed_ratio > 0.30


def _pdf_build_mapping_stage_from_template_rows(parsed_rows: list) -> tuple:
    headers = ["Date", "Narration", "Chq./Ref.No.", "WithdrawalAmt.", "DepositAmt.", "ClosingBalance"]
    rows = []
    for row in parsed_rows or []:
        rows.append(
            {
                "Date": row.get("DATE", "") or "",
                "Narration": row.get("DESCRIPTION", "") or "",
                "Chq./Ref.No.": row.get("CHEQUE NO.", "") or "",
                "WithdrawalAmt.": row.get("Debit", "") or "",
                "DepositAmt.": row.get("Credit", "") or "",
                "ClosingBalance": row.get("Balance", "") or "",
            }
        )

    suggested_mapping = {
        "date": "Date",
        "description": "Narration",
        "cheque": "Chq./Ref.No.",
        "debit": "WithdrawalAmt.",
        "credit": "DepositAmt.",
        "amount": None,
        "drcr": None,
        "balance": "ClosingBalance",
    }
    return headers, rows, suggested_mapping


def _pdf_pick_mapped_value(raw_row: dict, mapping: dict, key: str) -> str:
    source_col = mapping.get(key)
    if not source_col:
        return ""
    return str(raw_row.get(source_col, "") or "").strip()


def _pdf_apply_column_mapping(raw_rows: list, mapping: dict) -> list:
    date_col = mapping.get("date")
    desc_col = mapping.get("description")
    if not date_col or not desc_col:
        raise ValueError("Please map both Date and Description columns.")

    mapped_rows = []
    prev_balance = 0.0

    def _has_amount_data(item: dict) -> bool:
        return (
            float(item.get("debit", 0.0)) > 0
            or float(item.get("credit", 0.0)) > 0
            or float(item.get("amount", 0.0)) > 0
        )

    pending = None

    def _start_pending(parsed_date, desc_text, cheque_text, debit, credit, amount, drcr_side, balance):
        return {
            "date": parsed_date,
            "description": desc_text,
            "cheque": cheque_text,
            "debit": debit,
            "credit": credit,
            "amount": amount,
            "drcr": drcr_side,
            "balance": balance,
        }

    def _merge_into_pending(item, desc_text, cheque_text, debit, credit, amount, drcr_side, balance):
        if desc_text:
            current_desc = str(item.get("description") or "").strip()
            item["description"] = f"{current_desc} {desc_text}".strip() if current_desc else desc_text

        if cheque_text:
            current_chq = str(item.get("cheque") or "").strip()
            if not current_chq:
                item["cheque"] = cheque_text
            elif cheque_text != current_chq:
                item["cheque"] = f"{current_chq} {cheque_text}".strip()

        if float(item.get("debit", 0.0)) <= 0 and debit > 0:
            item["debit"] = debit
        if float(item.get("credit", 0.0)) <= 0 and credit > 0:
            item["credit"] = credit
        if float(item.get("amount", 0.0)) <= 0 and amount > 0:
            item["amount"] = amount
        if not item.get("drcr") and drcr_side:
            item["drcr"] = drcr_side
        if float(item.get("balance", 0.0)) <= 0 and balance > 0:
            item["balance"] = balance

    def _flush_pending(item):
        nonlocal prev_balance
        if not item or not item.get("date"):
            return

        debit = float(item.get("debit", 0.0) or 0.0)
        credit = float(item.get("credit", 0.0) or 0.0)
        amount = float(item.get("amount", 0.0) or 0.0)
        drcr_side = str(item.get("drcr") or "")
        balance = float(item.get("balance", 0.0) or 0.0)

        if debit <= 0 and credit <= 0 and amount > 0:
            if drcr_side == "DR":
                debit = amount
            elif drcr_side == "CR":
                credit = amount
            elif balance > 0 and prev_balance > 0:
                if balance >= prev_balance:
                    credit = amount
                else:
                    debit = amount
            else:
                debit = amount

        if debit <= 0 and credit <= 0:
            return

        row = {
            "DATE": item["date"].strftime("%d-%b-%Y"),
            "CHEQUE NO.": str(item.get("cheque") or "").strip(),
            "DESCRIPTION": (str(item.get("description") or "").strip() or "Bank Transaction"),
            "LEDGER": "Suspense A/c",
            "Debit": debit if debit > 0 else "",
            "Credit": credit if credit > 0 else "",
            "Balance": balance if balance > 0 else "",
        }
        mapped_rows.append(row)

        if balance > 0:
            prev_balance = balance

    for raw_row in raw_rows:
        date_text = _normalize_pdf_date_text(_pdf_pick_mapped_value(raw_row, mapping, "date"))
        parsed_date = _pdf_parse_date_value(date_text)
        desc_text = _pdf_pick_mapped_value(raw_row, mapping, "description")
        cheque_text = _pdf_pick_mapped_value(raw_row, mapping, "cheque")

        debit = _pdf_parse_amount(_pdf_pick_mapped_value(raw_row, mapping, "debit"))
        credit = _pdf_parse_amount(_pdf_pick_mapped_value(raw_row, mapping, "credit"))
        amount = _pdf_parse_amount(_pdf_pick_mapped_value(raw_row, mapping, "amount"))
        drcr_side = _pdf_parse_drcr_side(_pdf_pick_mapped_value(raw_row, mapping, "drcr"))
        balance = _pdf_parse_amount(_pdf_pick_mapped_value(raw_row, mapping, "balance"))

        if pending is None:
            if parsed_date is None:
                continue
            pending = _start_pending(parsed_date, desc_text, cheque_text, debit, credit, amount, drcr_side, balance)
            continue

        pending_has_amount = _has_amount_data(pending)

        if parsed_date is not None:
            # Start a new transaction when we already have amount data in pending.
            if pending_has_amount:
                _flush_pending(pending)
                pending = _start_pending(parsed_date, desc_text, cheque_text, debit, credit, amount, drcr_side, balance)
                continue

            # Pending has no amount yet: treat same-date row as continuation,
            # different date as new transaction boundary.
            same_date = pending.get("date") is not None and pending["date"].date() == parsed_date.date()
            if not same_date:
                _flush_pending(pending)
                pending = _start_pending(parsed_date, desc_text, cheque_text, debit, credit, amount, drcr_side, balance)
                continue

            _merge_into_pending(pending, desc_text, cheque_text, debit, credit, amount, drcr_side, balance)
            continue

        # No date in current row: continuation of pending.
        _merge_into_pending(pending, desc_text, cheque_text, debit, credit, amount, drcr_side, balance)

    _flush_pending(pending)

    return mapped_rows


def _pdf_write_stage_excel(headers: list, rows: list, source_pdf_path: str) -> str:
    stage_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "pdf_staging")
    os.makedirs(stage_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(source_pdf_path))[0]
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(stage_dir, f"{base_name}_staged_{stamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF_Staged"

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    wb.save(out_path)
    wb.close()
    return out_path


def prepare_pdf_mapping_payload(filepath: str) -> dict:
    """
    Backend staging flow:
    1) Extract raw table rows from PDF.
    2) Save staged Excel-like sheet in backend.
    3) Return headers + rows + suggested mapping for user confirmation.
    """
    raw_headers, raw_rows = _pdf_extract_raw_table_rows(filepath)
    mapping_suggestion = {}
    stage_mode = "table"
    fallback_template_rows = []

    if raw_rows:
        headers = raw_headers or [f"Column {i + 1}" for i in range(max(len(r.keys()) for r in raw_rows))]
        rows = raw_rows
        mapping_suggestion = _pdf_guess_column_mapping(headers, rows)

        # Keep raw headers for user mapping first; prepare fallback only if needed.
        need_fallback = _pdf_is_stage_quality_low(headers, rows)
        date_col = mapping_suggestion.get("date")
        desc_col = mapping_suggestion.get("description")
        if date_col and desc_col:
            try:
                preflight_rows = _pdf_apply_column_mapping(rows, mapping_suggestion)
                if len(preflight_rows) == 0:
                    need_fallback = True
            except Exception:
                need_fallback = True
        else:
            need_fallback = True

        if need_fallback:
            try:
                _headers, parsed_rows = read_pdf_statement(filepath)
                fallback_template_rows = parsed_rows or []
                if fallback_template_rows:
                    stage_mode = "table-with-fallback"
            except Exception:
                fallback_template_rows = []
    else:
        try:
            _headers, parsed_rows = read_pdf_statement(filepath)
            if not parsed_rows:
                raise ValueError("No line parser rows")
            headers, rows, mapping_suggestion = _pdf_build_mapping_stage_from_template_rows(parsed_rows)
            stage_mode = "line-fallback"
        except Exception:
            pos_headers, pos_rows = _pdf_extract_positional_statement_rows(filepath)
            if not pos_rows:
                raise ValueError("Could not extract any rows from PDF.")
            headers = pos_headers
            rows = pos_rows
            mapping_suggestion = _pdf_guess_column_mapping(headers, rows)
            stage_mode = "positional-fallback"

    staged_excel_path = _pdf_write_stage_excel(headers, rows, filepath)

    return {
        "headers": headers,
        "rows": rows,
        "suggested_mapping": mapping_suggestion,
        "staged_excel_path": staged_excel_path,
        "stage_mode": stage_mode,
        "fallback_template_rows": fallback_template_rows,
    }


# ─── Read Excel ─────────────────────────────────────────────────────────

def read_excel(filepath: str, sheet: str = None) -> tuple:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(c or "").strip() for c in first_row]
    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        vals = list(vals[:len(headers)])
        if len(vals) < len(headers):
            vals.extend([None] * (len(headers) - len(vals)))
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    wb.close()
    return headers, rows


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════════════════

class TallyBankApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("TallyBankPro — Bank Statement to Tally Voucher Creator")
        self.geometry("1120x720")
        self.minsize(960, 640)
        self.configure(fg_color=COLORS["bg_dark"])

        self.loaded_rows = []
        self.loaded_headers = []
        self.fp_var = ctk.StringVar(value="")
        self.company_placeholder = "Auto (Loaded Company)"
        self.company_var = ctk.StringVar(value=self.company_placeholder)
        self.bank_ledger_var = ctk.StringVar(value="")
        self.tally_host_var = ctk.StringVar(value="localhost")
        self.tally_port_var = ctk.StringVar(value="9000")
        self.use_today_date_var = ctk.BooleanVar(value=False)
        self.status_var = ctk.StringVar(value="Ready")
        self.connection_status_var = ctk.StringVar(value="Connection: Not checked")
        self.company_status_var = ctk.StringVar(value="Companies: Not fetched")
        self.bank_status_var = ctk.StringVar(value="Banks: Not fetched")
        self.fetched_companies = []
        self.fetched_banks = []
        self._company_fetch_running = False
        self._bank_fetch_running = False
        self._file_load_running = False
        self._preview_trees = {}
        self._preview_info_labels = {}
        self._preview_summary_labels = {}
        self._browse_buttons = {}
        self._path_vars = {}
        self._loaded_source = "excel"
        self._push_running = False
        self._push_overlay = None
        self._push_message_var = ctk.StringVar(value="")
        self.debug_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tally_bank_debug.log")

        self._build_ui()

    def _resolve_theme_color(self, key: str):
        value = COLORS.get(key)
        if isinstance(value, tuple):
            mode = ctk.get_appearance_mode().lower()
            return value[1] if mode == "dark" else value[0]
        return value

    def _apply_ttk_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("default")
        except tk.TclError:
            pass

        tree_bg = self._resolve_theme_color("bg_card")
        field_bg = self._resolve_theme_color("bg_input")
        heading_bg = self._resolve_theme_color("table_header")
        text_fg = self._resolve_theme_color("text_primary")
        border_fg = self._resolve_theme_color("border")
        selected_bg = self._resolve_theme_color("accent")

        style.configure(
            "Treeview",
            background=tree_bg,
            foreground=text_fg,
            fieldbackground=field_bg,
            bordercolor=border_fg,
            font=("Segoe UI", 10),
            rowheight=26,
        )
        style.configure(
            "Treeview.Heading",
            background=heading_bg,
            foreground="#FFFFFF",
            font=("Segoe UI", 10, "bold"),
        )
        style.map(
            "Treeview",
            background=[("selected", selected_bg)],
            foreground=[("selected", "#FFFFFF")],
        )

    def set_theme(self, mode: str):
        try:
            ctk.set_appearance_mode(mode)
        except Exception:
            pass
        self._apply_ttk_styles()

    def _get_tally_url(self):
        return _build_tally_url(self.tally_host_var.get(), self.tally_port_var.get())

    def _get_selected_company(self):
        selected = _normalize_company_name(self.company_var.get())
        if not selected or _company_key(selected) == _company_key(self.company_placeholder):
            if len(self.fetched_companies) == 1:
                return self.fetched_companies[0]
            return ""
        return selected

    def _append_debug_log(self, mode, target_company, xml_payload, response_text, parsed, note=""):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_lines = [
            "=" * 96,
            f"[{timestamp}] mode={mode} company={target_company or 'Loaded'} note={note}",
            f"summary: created={parsed.get('created', 0)} altered={parsed.get('altered', 0)} "
            f"errors={parsed.get('errors', 0)} exceptions={parsed.get('exceptions', 0)}",
        ]
        line_errors = parsed.get("line_errors") or []
        if line_errors:
            log_lines.append("line_errors:")
            for err in line_errors:
                log_lines.append(f"- {err}")
        log_lines.append("response:")
        log_lines.append(response_text[:12000])
        log_lines.append("xml:")
        log_lines.append(xml_payload[:12000])
        log_lines.append("\n")
        with open(self.debug_log_path, "a", encoding="utf-8") as f:
            f.write("\n".join(log_lines))

    # ── UI ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ── Settings Card ──
        settings_card = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], border_width=1,
                                      border_color=COLORS["border"], corner_radius=12)
        settings_card.grid(row=1, column=0, sticky="ew", padx=16, pady=(10, 8))

        row_1 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_1.pack(fill="x", padx=14, pady=(12, 6))
        ctk.CTkLabel(row_1, text="Host", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        ctk.CTkEntry(row_1, textvariable=self.tally_host_var, width=140, height=32,
                      fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                      text_color=COLORS["text_primary"]).pack(side="left", padx=(6, 12))
        ctk.CTkLabel(row_1, text="Port", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        ctk.CTkEntry(row_1, textvariable=self.tally_port_var, width=90, height=32,
                      fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                      text_color=COLORS["text_primary"]).pack(side="left", padx=(6, 12))
        self.connection_test_btn = ctk.CTkButton(
            row_1, text="Test Connection", width=130, height=32, font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["warning"], hover_color="#B45309", text_color="#FFFFFF", corner_radius=8,
            command=self._check_connection_thread)
        self.connection_test_btn.pack(side="right")

        # Company row
        row_2 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_2.pack(fill="x", padx=14, pady=(0, 6))
        ctk.CTkLabel(row_2, text="Target Company", font=("Segoe UI", 10),
                      text_color=COLORS["text_secondary"]).pack(side="left")
        self.company_combo = ctk.CTkComboBox(
            row_2, values=[self.company_placeholder], variable=self.company_var,
            width=340, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            button_color=COLORS["accent"], button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10), command=self._on_company_changed)
        self.company_combo.set(self.company_placeholder)
        self.company_combo.pack(side="left", padx=(10, 8), fill="x", expand=True)
        self.company_refresh_btn = ctk.CTkButton(
            row_2, text="Refresh", width=96, height=34, font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], corner_radius=8,
            command=lambda: self._fetch_companies_thread())
        self.company_refresh_btn.pack(side="right")

        # Bank Ledger row
        row_3 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_3.pack(fill="x", padx=14, pady=(0, 6))
        ctk.CTkLabel(row_3, text="Bank Ledger", font=("Segoe UI", 10, "bold"),
                      text_color=COLORS["accent"]).pack(side="left")
        self.bank_combo = ctk.CTkComboBox(
            row_3, values=["-- Select Bank --"], variable=self.bank_ledger_var,
            width=340, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["accent"],
            button_color="#059669", button_hover_color="#047857",
            font=("Segoe UI", 10, "bold"))
        self.bank_combo.set("-- Select Bank --")
        self.bank_combo.pack(side="left", padx=(10, 8), fill="x", expand=True)
        self.bank_refresh_btn = ctk.CTkButton(
            row_3, text="Fetch Banks", width=110, height=34, font=("Segoe UI", 10, "bold"),
            fg_color="#059669", hover_color="#047857", text_color="#FFFFFF", corner_radius=8,
            command=self._fetch_banks_thread)
        self.bank_refresh_btn.pack(side="right")

        # Status labels
        status_row = ctk.CTkFrame(settings_card, fg_color="transparent")
        status_row.pack(fill="x", padx=14, pady=(0, 6))
        self.connection_status_label = ctk.CTkLabel(
            status_row, textvariable=self.connection_status_var,
            font=("Segoe UI", 10), text_color=COLORS["text_muted"])
        self.connection_status_label.pack(side="left", padx=(0, 15))
        self.company_status_label = ctk.CTkLabel(
            status_row, textvariable=self.company_status_var,
            font=("Segoe UI", 10), text_color=COLORS["text_muted"])
        self.company_status_label.pack(side="left", padx=(0, 15))
        self.bank_status_label = ctk.CTkLabel(
            status_row, textvariable=self.bank_status_var,
            font=("Segoe UI", 10), text_color=COLORS["text_muted"])
        self.bank_status_label.pack(side="left", padx=(0, 15))

        self.today_date_checkbox = ctk.CTkCheckBox(
            settings_card, text="Use Today Date For Vouchers (ignore source statement date)",
            variable=self.use_today_date_var, font=("Segoe UI", 10, "bold"),
            text_color=COLORS["text_secondary"], fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"], border_color=COLORS["border"])
        self.today_date_checkbox.pack(anchor="w", padx=14, pady=(0, 10))

        # ── Main Content ──
        content_card = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], border_width=1,
                                     border_color=COLORS["border"], corner_radius=12)
        content_card.grid(row=2, column=0, sticky="nsew", padx=16, pady=(0, 10))

        self._apply_ttk_styles()

        self.source_tabs = ctk.CTkTabview(
            content_card,
            fg_color="transparent",
            segmented_button_fg_color=COLORS["bg_input"],
            segmented_button_selected_color=COLORS["accent"],
            segmented_button_selected_hover_color=COLORS["accent_hover"],
            segmented_button_unselected_color=COLORS["bg_input"],
            segmented_button_unselected_hover_color=COLORS["bg_card_hover"],
        )
        self.source_tabs.pack(fill="both", expand=True, padx=10, pady=(10, 5))

        excel_tab = self.source_tabs.add("Excel Statement")
        self._build_source_tab(excel_tab, source="excel")
        self.source_tabs.set("Excel Statement")

        # Fixed bottom action bar so primary buttons remain visible on smaller window heights.
        action_bar = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], border_width=1,
                                   border_color=COLORS["border"], corner_radius=12)
        action_bar.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 10))

        action_left = ctk.CTkFrame(action_bar, fg_color="transparent")
        action_left.pack(side="left", padx=10, pady=10)

        self.save_xml_btn = ctk.CTkButton(
            action_left,
            text="💾  Save XML File",
            fg_color=SUCCESS,
            hover_color="#15803D",
            width=170,
            command=lambda: self._generate("save"),
        )
        self.save_xml_btn.pack(side="left", padx=(0, 10))

        self.push_tally_btn = ctk.CTkButton(
            action_left,
            text="🚀  Push to Tally",
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER,
            width=170,
            command=lambda: self._generate("push"),
        )
        self.push_tally_btn.pack(side="left")

        self.template_btn = ctk.CTkButton(
            action_bar,
            text="📋  Download Template",
            fg_color="#94A3B8",
            hover_color="#64748B",
            text_color="#FFFFFF",
            width=170,
            command=self._save_template,
        )
        self.template_btn.pack(side="right", padx=10, pady=10)

        # Status bar
        status_bar = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=0, height=32)
        status_bar.grid(row=4, column=0, sticky="ew")
        status_bar.grid_propagate(False)
        ctk.CTkLabel(status_bar, textvariable=self.status_var, font=("Segoe UI", 10),
                      text_color=COLORS["text_muted"]).pack(side="left", padx=16)

        self.after(200, lambda: self._fetch_companies_thread(silent=True))

    def _build_source_tab(self, parent, source: str):
        load_frame = ctk.CTkFrame(parent, fg_color="transparent")
        load_frame.pack(fill="x", padx=4, pady=(8, 5))

        path_var = ctk.StringVar()
        self._path_vars[source] = path_var
        self.fp_var = path_var

        placeholder = (
            "Select Bank Statement Excel (.xlsx) — Template: DATE | CHEQUE NO. | DESCRIPTION | LEDGER | Debit | Credit | Balance"
        )
        ctk.CTkEntry(
            load_frame,
            textvariable=path_var,
            placeholder_text=placeholder,
            width=600,
            state="readonly",
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
        ).pack(side="left", padx=(0, 8), fill="x", expand=True)

        browse_btn = ctk.CTkButton(
            load_frame,
            text="📂 Browse Excel",
            command=self._browse_file,
            width=120,
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER,
        )
        browse_btn.pack(side="left", padx=(0, 8))
        self._browse_buttons[source] = browse_btn

        info_frame = ctk.CTkFrame(parent, fg_color="transparent")
        info_frame.pack(fill="x", padx=4, pady=(0, 5))
        info_label = ctk.CTkLabel(info_frame, text="", font=("Segoe UI", 11), text_color=TEXT_MUTED)
        info_label.pack(side="left")
        summary_label = ctk.CTkLabel(
            info_frame,
            text="",
            font=("Segoe UI", 11, "bold"),
            text_color=COLORS["accent"],
        )
        summary_label.pack(side="right")
        self._preview_info_labels[source] = info_label
        self._preview_summary_labels[source] = summary_label

        tree_frame = ctk.CTkFrame(parent, fg_color=COLORS["bg_dark"], corner_radius=8,
                                   border_width=1, border_color=COLORS["border"])
        tree_frame.pack(fill="both", expand=True, padx=4, pady=5)

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        tree = ttk.Treeview(
            tree_frame,
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
        )
        tree_scroll_y.config(command=tree.yview)
        tree_scroll_x.config(command=tree.xview)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")

        self._preview_trees[source] = tree

    # ── Company Dropdown ────────────────────────────────────────────────

    def _set_company_dropdown(self, companies, keep_selection=True):
        current = _normalize_company_name(self.company_var.get()) if keep_selection else ""
        cleaned = []
        seen = set()
        for name in companies or []:
            normalized = _normalize_company_name(name)
            if not _is_valid_company_name(normalized):
                continue
            key = _company_key(normalized)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(normalized)
        cleaned = sorted(cleaned, key=lambda x: _company_key(x))
        values = [self.company_placeholder] + cleaned
        self.company_combo.configure(values=values)
        if current and _company_key(current) in {_company_key(x) for x in cleaned}:
            self.company_combo.set(current)
            self.company_var.set(current)
        else:
            self.company_combo.set(self.company_placeholder)
            self.company_var.set(self.company_placeholder)
        self.fetched_companies = cleaned
        self.company_status_var.set(f"Companies: {len(cleaned)} available")

    def _on_company_changed(self, _=None):
        """When company changes, auto-fetch banks."""
        selected_company = self._get_selected_company()
        if not selected_company and len(self.fetched_companies) > 1:
            return
        self.after(100, self._fetch_banks_thread)

    def _fetch_companies_thread(self, silent=False):
        if self._company_fetch_running:
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return
        self._company_fetch_running = True
        self.company_refresh_btn.configure(state="disabled", text="Fetching...")

        def worker():
            result = _fetch_tally_companies(tally_url, timeout=15)

            def done():
                self._company_fetch_running = False
                self.company_refresh_btn.configure(state="normal", text="Refresh")
                if result.get("success"):
                    companies = result.get("companies", [])
                    self._set_company_dropdown(companies)
                    self.status_var.set(f"Fetched {len(companies)} company(s)")
                    # Auto-fetch banks
                    self.after(200, self._fetch_banks_thread)
                else:
                    self.company_status_var.set("Companies: Fetch failed")
                    self.company_status_label.configure(text_color=COLORS["error"])
                    if not silent:
                        messagebox.showwarning("Fetch Failed", str(result.get("error", "")))
            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _check_connection_thread(self):
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return
        self.connection_test_btn.configure(state="disabled", text="Checking...")
        self.connection_status_var.set("Connection: Checking...")
        self.connection_status_label.configure(text_color=COLORS["warning"])

        def worker():
            result = _check_tally_connection(tally_url, timeout=8)

            def done():
                self.connection_test_btn.configure(state="normal", text="Test Connection")
                if result.get("connected"):
                    self.connection_status_var.set(f"Connection: Connected ({tally_url})")
                    self.connection_status_label.configure(text_color=COLORS["success"])
                    self.status_var.set("Connected to Tally")
                    self._fetch_companies_thread(silent=True)
                else:
                    err = str(result.get("error", "Unknown"))
                    self.connection_status_var.set("Connection: Offline")
                    self.connection_status_label.configure(text_color=COLORS["error"])
                    messagebox.showwarning("Connection Failed", err)
            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    # ── Bank Ledger Fetch ───────────────────────────────────────────────

    def _fetch_banks_thread(self):
        if self._bank_fetch_running:
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError:
            return
        company = self._get_selected_company()
        self._bank_fetch_running = True
        self.bank_refresh_btn.configure(state="disabled", text="Fetching...")
        self.bank_status_var.set("Banks: Fetching...")
        self.bank_status_label.configure(text_color=COLORS["warning"])

        def worker():
            result = _fetch_bank_ledgers(tally_url, company=company, timeout=15)

            def done():
                self._bank_fetch_running = False
                self.bank_refresh_btn.configure(state="normal", text="Fetch Banks")
                if result.get("success"):
                    ledgers = result.get("ledgers", [])
                    names = [l["name"] for l in ledgers]
                    self.fetched_banks = names
                    values = ["-- Select Bank --"] + names
                    self.bank_combo.configure(values=values)
                    if names:
                        self.bank_combo.set(names[0])
                        self.bank_ledger_var.set(names[0])
                    self.bank_status_var.set(f"Banks: {len(names)} found")
                    self.bank_status_label.configure(text_color=COLORS["success"])
                    self.status_var.set(f"Fetched {len(names)} bank/cash ledger(s)")
                else:
                    self.bank_status_var.set("Banks: Fetch failed")
                    self.bank_status_label.configure(text_color=COLORS["error"])
            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    # ── File Loading ────────────────────────────────────────────────────

    def _browse_file(self):
        if self._file_load_running:
            self.status_var.set("Please wait, file is still loading...")
            return
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm *.xls")])
        if f:
            self._loaded_source = "excel"
            self.source_tabs.set("Excel Statement")
            if "excel" in self._path_vars:
                self._path_vars["excel"].set(f)
            self._load_preview(f, source="excel")

    def _browse_pdf_file(self):
        messagebox.showinfo("PDF Disabled", "PDF statement import has been removed. Please use Excel input.")

    def _set_file_loading_state(self, is_loading: bool):
        self._file_load_running = is_loading
        button_state = "disabled" if is_loading else "normal"
        for btn in self._browse_buttons.values():
            if is_loading:
                btn.configure(state=button_state, text="Loading...")
            else:
                btn.configure(state=button_state, text="📂 Browse Excel")
        self.save_xml_btn.configure(state=button_state)
        self.push_tally_btn.configure(state=button_state)

    def _set_push_loading_state(self, is_loading: bool, message: str = ""):
        self._push_running = is_loading
        button_state = "disabled" if is_loading else "normal"
        self.save_xml_btn.configure(state=button_state)
        self.push_tally_btn.configure(state=button_state)

        for btn in self._browse_buttons.values():
            btn.configure(state=button_state)

        if is_loading:
            self._push_message_var.set(message or "Posting to Tally...")
            if self._push_overlay is None or not self._push_overlay.winfo_exists():
                overlay = ctk.CTkToplevel(self)
                overlay.title("Please Wait")
                overlay.geometry("420x170")
                overlay.transient(self)
                overlay.grab_set()
                overlay.resizable(False, False)

                frame = ctk.CTkFrame(overlay, fg_color=COLORS["bg_card"], corner_radius=10)
                frame.pack(fill="both", expand=True, padx=12, pady=12)

                ctk.CTkLabel(
                    frame,
                    text="Pushing Vouchers To Tally",
                    font=("Segoe UI", 15, "bold"),
                    text_color=COLORS["text_primary"],
                ).pack(pady=(12, 8))

                ctk.CTkLabel(
                    frame,
                    textvariable=self._push_message_var,
                    font=("Segoe UI", 11),
                    text_color=COLORS["text_muted"],
                    wraplength=360,
                    justify="center",
                ).pack(pady=(0, 10))

                progress = ctk.CTkProgressBar(frame, mode="indeterminate", width=320)
                progress.pack(pady=(0, 10))
                progress.start()

                overlay.protocol("WM_DELETE_WINDOW", lambda: None)
                self._push_overlay = overlay
            else:
                self._push_overlay.deiconify()
                self._push_overlay.lift()
        else:
            if self._push_overlay is not None and self._push_overlay.winfo_exists():
                self._push_overlay.grab_release()
                self._push_overlay.destroy()
            self._push_overlay = None
            self._push_message_var.set("")

    def _build_preview_result(self, headers, rows):
        preview_rows = []
        for row in rows[:500]:
            preview_rows.append([
                "" if row.get(h, "") is None else str(row.get(h, ""))
                for h in headers
            ])

        total_debit = 0.0
        total_credit = 0.0
        payment_count = 0
        receipt_count = 0
        for row in rows:
            dr = row.get("Debit") or row.get("DEBIT") or row.get("debit") or 0
            cr = row.get("Credit") or row.get("CREDIT") or row.get("credit") or 0
            try:
                dr_val = float(dr) if dr not in (None, "", "None") else 0.0
            except (TypeError, ValueError):
                dr_val = 0.0
            try:
                cr_val = float(cr) if cr not in (None, "", "None") else 0.0
            except (TypeError, ValueError):
                cr_val = 0.0

            if dr_val > 0:
                total_debit += dr_val
                payment_count += 1
            if cr_val > 0:
                total_credit += cr_val
                receipt_count += 1

        return {
            "headers": headers,
            "rows": rows,
            "preview_rows": preview_rows,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "payment_count": payment_count,
            "receipt_count": receipt_count,
        }

    def _apply_preview_result(self, source, source_label, result, info_label, summary_label, tree, filepath=""):
        headers = result["headers"]
        rows = result["rows"]
        self.loaded_headers = headers
        self.loaded_rows = rows
        self._loaded_source = source

        if tree:
            tree.delete(*tree.get_children())
            tree["columns"] = headers
            for h in headers:
                tree.heading(h, text=h)
                tree.column(h, width=120, minwidth=60)

            for values in result["preview_rows"]:
                tree.insert("", "end", values=values)

        if info_label:
            if len(rows) > 500:
                info_label.configure(
                    text=f"✅ Loaded {len(rows)} rows, {len(headers)} columns (showing first 500 in preview)"
                )
            else:
                info_label.configure(text=f"✅ Loaded {len(rows)} rows, {len(headers)} columns")
        if summary_label:
            summary_label.configure(
                text=(
                    f"Payments: {result['payment_count']} (₹{result['total_debit']:,.2f})  |  "
                    f"Receipts: {result['receipt_count']} (₹{result['total_credit']:,.2f})"
                )
            )

        base_name = os.path.basename(filepath) if filepath else ""
        suffix = f": {base_name}" if base_name else ""
        self.status_var.set(f"Loaded {source_label}{suffix} — {len(rows)} rows")

    def _open_pdf_column_mapping_dialog(self, filepath, payload, info_label, summary_label, tree):
        messagebox.showinfo("PDF Disabled", "PDF statement import has been removed. Please use Excel input.")

    def _load_preview(self, filepath, source: str = "excel"):
        if source != "excel":
            raise ValueError("PDF statement import has been removed. Please use Excel input.")

        info_label = self._preview_info_labels.get(source)
        summary_label = self._preview_summary_labels.get(source)
        tree = self._preview_trees.get(source)
        source_label = "Excel"

        self._set_file_loading_state(True)
        if info_label:
            info_label.configure(text=f"Loading {source_label} preview...")
        if summary_label:
            summary_label.configure(text="")
        self.status_var.set(f"Loading {source_label}: {os.path.basename(filepath)}")

        def worker():
            try:
                headers, rows = read_excel(filepath)
                result = {"ok": True, **self._build_preview_result(headers, rows)}
            except Exception as exc:
                result = {"ok": False, "error": str(exc)}

            def done():
                self._set_file_loading_state(False)
                if not result.get("ok"):
                    if info_label:
                        info_label.configure(text="")
                    if summary_label:
                        summary_label.configure(text="")
                    self.status_var.set("Ready")
                    messagebox.showerror("Error", str(result.get("error", "Unknown error")))
                    return

                self._apply_preview_result(
                    source,
                    source_label,
                    result,
                    info_label,
                    summary_label,
                    tree,
                    filepath,
                )

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    # ── Save Template ───────────────────────────────────────────────────

    def _save_template(self):
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile="BankStatement_Template.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not out:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BankStatement"
        headers = ["DATE", "CHEQUE NO.", "DESCRIPTION", "LEDGER", "Debit", "Credit", "Balance"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
        # Sample rows
        sample = [
            [datetime(2026, 4, 1), "", "IMPS-Transfer-ABC", "Suspense A/c", 20000, None, 20000],
            [datetime(2026, 4, 2), "123456", "NEFT CR-XYZ Ltd", "Suspense A/c", None, 5000, 15000],
        ]
        for row_idx, row_data in enumerate(sample, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        # Auto width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
        wb.save(out)
        messagebox.showinfo("Template Saved", f"Template saved to:\n{out}")

    # ── Generate & Push ─────────────────────────────────────────────────

    def _generate(self, action):
        if self._push_running:
            self.status_var.set("Push already in progress. Please wait...")
            return

        company = self._get_selected_company()
        bank_ledger = self.bank_ledger_var.get().strip()
        use_today_date = bool(self.use_today_date_var.get())

        if not self.loaded_rows:
            messagebox.showwarning("No Data", "Load a bank statement Excel file first.")
            return

        if not bank_ledger or bank_ledger == "-- Select Bank --":
            messagebox.showwarning("Select Bank", "Please select a Bank Ledger from the dropdown.\n\n"
                                    "Click 'Fetch Banks' to load bank ledgers from Tally.")
            return

        if action == "push" and not company and len(self.fetched_companies) > 1:
            messagebox.showwarning("Select Company",
                                    "Multiple companies detected. Please select the target company.")
            return

        try:
            if action == "save":
                xml = generate_bank_voucher_xml(
                    self.loaded_rows, company, bank_ledger,
                    use_today_date=use_today_date)
                out = filedialog.asksaveasfilename(
                    defaultextension=".xml", initialfile="Tally_BankVouchers.xml",
                    filetypes=[("XML", "*.xml")])
                if out:
                    with open(out, "w", encoding="utf-8") as f:
                        f.write(xml)
                    self.status_var.set(f"Saved XML: {out}")
                    messagebox.showinfo("Success", f"XML saved successfully!\n{out}")
            else:
                tally_url = self._get_tally_url()
                host, port_text = tally_url.rsplit(":", 1)
                host = host.replace("http://", "", 1)
                target_company = company or "Loaded company in Tally"
                rows_snapshot = list(self.loaded_rows)

                self._set_push_loading_state(True, "Preparing vouchers...")
                self.status_var.set(f"Posting to Tally ({target_company}, Bank: {bank_ledger})...")

                def worker():
                    result = {"ok": False, "error": "Unknown error"}
                    try:
                        self.after(0, lambda: self._push_message_var.set("Checking and creating missing ledgers..."))

                        unique_ledgers = set()
                        for r in rows_snapshot:
                            ledger = str(r.get("LEDGER") or r.get("Ledger") or "").strip()
                            if ledger and ledger.lower() != bank_ledger.lower():
                                unique_ledgers.add(ledger)

                        if unique_ledgers:
                            ledger_defs = [{"Name": n, "Parent": "Suspense A/c"} for n in unique_ledgers]
                            ledger_xml = generate_ledger_xml(ledger_defs, company)
                            try:
                                ledger_resp = push_to_tally(ledger_xml, host, int(port_text))
                                ledger_parsed = _parse_tally_response_details(ledger_resp)
                                self._append_debug_log(
                                    "auto-ledger",
                                    target_company,
                                    ledger_xml,
                                    ledger_resp,
                                    ledger_parsed,
                                    note=f"bank_contra_ledgers={len(ledger_defs)}",
                                )
                            except Exception:
                                pass

                        self.after(0, lambda: self._push_message_var.set("Getting next voucher numbers..."))
                        payment_start = None
                        receipt_start = None
                        try:
                            pmt_result = _fetch_next_voucher_number(tally_url, company, "Payment", timeout=15)
                            if pmt_result.get("success"):
                                payment_start = pmt_result.get("next_number")
                        except Exception:
                            pass
                        try:
                            rct_result = _fetch_next_voucher_number(tally_url, company, "Receipt", timeout=15)
                            if rct_result.get("success"):
                                receipt_start = rct_result.get("next_number")
                        except Exception:
                            pass

                        self.after(0, lambda: self._push_message_var.set("Preparing XML batches..."))
                        batches = [
                            rows_snapshot[i:i + PUSH_BATCH_SIZE]
                            for i in range(0, len(rows_snapshot), PUSH_BATCH_SIZE)
                        ]
                        total_batches = max(1, len(batches))
                        payment_cursor = payment_start
                        receipt_cursor = receipt_start

                        created_total = 0
                        altered_total = 0
                        ignored_total = 0

                        for batch_idx, batch_rows in enumerate(batches, start=1):
                            self.after(
                                0,
                                lambda i=batch_idx, n=total_batches: self._push_message_var.set(
                                    f"Pushing batch {i}/{n} to Tally..."
                                ),
                            )

                            xml = generate_bank_voucher_xml(
                                batch_rows,
                                company,
                                bank_ledger,
                                use_today_date=use_today_date,
                                payment_start_vno=payment_cursor,
                                receipt_start_vno=receipt_cursor,
                            )

                            resp = push_to_tally(
                                xml,
                                host,
                                int(port_text),
                                timeout=PUSH_REQUEST_TIMEOUT_SEC,
                            )
                            parsed = _parse_tally_response_details(resp)
                            self._append_debug_log(
                                "bank-voucher",
                                target_company,
                                xml,
                                resp,
                                parsed,
                                note=(
                                    f"bank={bank_ledger}, pmt_start={payment_cursor}, "
                                    f"rct_start={receipt_cursor}, batch={batch_idx}/{total_batches}, "
                                    f"rows={len(batch_rows)}"
                                ),
                            )

                            if not parsed.get("success"):
                                detail = parsed.get("error") or "Unknown"
                                if parsed.get("line_errors"):
                                    detail = parsed["line_errors"][0]
                                result = {
                                    "ok": False,
                                    "target_company": target_company,
                                    "bank_ledger": bank_ledger,
                                    "parsed": parsed,
                                    "detail": f"{detail} (batch {batch_idx}/{total_batches})",
                                }
                                break

                            created_total += parsed.get("created", 0)
                            altered_total += parsed.get("altered", 0)
                            ignored_total += parsed.get("ignored", 0)

                            if payment_cursor is not None or receipt_cursor is not None:
                                p_count, r_count = _count_voucher_entries(batch_rows)
                                if payment_cursor is not None:
                                    payment_cursor += p_count
                                if receipt_cursor is not None:
                                    receipt_cursor += r_count
                        else:
                            result = {
                                "ok": True,
                                "target_company": target_company,
                                "bank_ledger": bank_ledger,
                                "parsed": {
                                    "created": created_total,
                                    "altered": altered_total,
                                    "ignored": ignored_total,
                                    "errors": 0,
                                    "exceptions": 0,
                                },
                            }
                    except Exception as exc:
                        result = {
                            "ok": False,
                            "target_company": target_company,
                            "bank_ledger": bank_ledger,
                            "error": str(exc),
                        }

                    def done():
                        self._set_push_loading_state(False)
                        if result.get("ok"):
                            parsed = result.get("parsed", {})
                            self.status_var.set(f"Posted to Tally ({result.get('target_company', target_company)})")
                            messagebox.showinfo(
                                "Tally Response",
                                "Bank vouchers posted successfully!\n\n"
                                f"Target Company: {result.get('target_company', target_company)}\n"
                                f"Bank Ledger: {result.get('bank_ledger', bank_ledger)}\n"
                                f"Created: {parsed.get('created', 0)}\n"
                                f"Altered: {parsed.get('altered', 0)}\n"
                                f"Ignored: {parsed.get('ignored', 0)}",
                            )
                            return

                        parsed = result.get("parsed", {})
                        detail = result.get("detail") or result.get("error") or "Unknown"
                        hint = ""
                        if "out of range" in detail.lower():
                            hint = (
                                "\n\nTip: This usually means the target company's Books Beginning date "
                                "is after the statement date. In Tally, open company settings and set "
                                "Books Beginning From on/before the statement start date."
                            )
                        elif "timed out" in detail.lower() or "timeout" in detail.lower():
                            hint = (
                                "\n\nTip: Tally may still be processing a large push. Wait 1-2 minutes, "
                                "verify created vouchers in Tally, then push only the remaining rows if needed."
                            )
                        self.status_var.set("Push failed (see debug log)")
                        messagebox.showerror(
                            "Push Failed",
                            f"Tally returned an error.\n\n"
                            f"Target Company: {result.get('target_company', target_company)}\n"
                            f"Bank: {result.get('bank_ledger', bank_ledger)}\n"
                            f"Errors: {parsed.get('errors', 0)}\n"
                            f"Detail: {detail}{hint}\n\n"
                            f"Debug Log: {self.debug_log_path}",
                        )

                    self.after(0, done)

                threading.Thread(target=worker, daemon=True).start()
                return

        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
        except Exception as e:
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", str(e))


# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = TallyBankApp()
    app.mainloop()