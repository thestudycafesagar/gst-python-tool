"""
TallySalesPro - Professional Sales Voucher & Master Creator for TallyPrime
Converts Excel sales data to Tally XML (Accounting & Inventory voucher modes)
Also creates Ledger Masters and Stock Item Masters.
Author: Studycafe | Built with CustomTkinter
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
import openpyxl
import os
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta
import urllib.request
from urllib.error import HTTPError, URLError
import json
import threading
import re
import html
import webbrowser
import io
import base64
import time
from urllib.parse import unquote

try:
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    _REQUESTS_AVAILABLE = True
except ImportError:
    _REQUESTS_AVAILABLE = False

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, WebDriverException
    from webdriver_manager.chrome import ChromeDriverManager
    _SELENIUM_AVAILABLE = True
except ImportError:
    _SELENIUM_AVAILABLE = False

try:
    from PIL import Image, ImageTk
    _PIL_AVAILABLE = True
except ImportError:
    _PIL_AVAILABLE = False

# ─── Theme ──────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

COLORS = {
    "bg_dark": ("#F0F4F8", "#0F172A"),
    "bg_card": ("#FFFFFF", "#1E293B"),
    "bg_card_hover": ("#E2E8F0", "#334155"),
    "bg_input": ("#F1F5F9", "#1E293B"),
    "accent": ("#2563EB", "#3B82F6"),
    "accent_hover": ("#1D4ED8", "#2563EB"),
    "success": ("#059669", "#10B981"),
    "warning": ("#D97706", "#F59E0B"),
    "error": ("#DC2626", "#EF4444"),
    "text_primary": ("#0F172A", "#F1F5F9"),
    "text_secondary": ("#475569", "#CBD5E1"),
    "text_muted": ("#64748B", "#94A3B8"),
    "border": ("#E2E8F0", "#334155"),
    "table_header": ("#1E293B", "#0F172A"),
}

ACCENT = COLORS["accent"]
ACCENT_HOVER = COLORS["accent_hover"]
SUCCESS = COLORS["success"]
DANGER = COLORS["error"]
SURFACE = COLORS["bg_card"]
SURFACE2 = COLORS["bg_input"]
TEXT_PRIMARY = COLORS["text_primary"]
TEXT_MUTED = COLORS["text_muted"]
CARD_BG = COLORS["bg_dark"]
SUSPENSE_LEDGER = "Suspense A/c"

LEDGER_PARENT_OPTIONS = [
    "Sundry Debtors",
    "Sundry Creditors",
    "Sales Accounts",
    "Purchase Accounts",
    "Duties & Taxes",
    "Direct Incomes",
    "Indirect Incomes",
    "Direct Expenses",
    "Indirect Expenses",
    "Fixed Assets",
]

LEDGER_GST_APPLICABLE_OPTIONS = [
    "Applicable",
    "Not Applicable",
]

LEDGER_STATE_OPTIONS = [
    "Not Applicable",
    "Andaman & Nicobar Islands",
    "Andhra Pradesh",
    "Arunachal Pradesh",
    "Assam",
    "Bihar",
    "Chandigarh",
    "Chhattisgarh",
    "Dadra & Nagar Haveli and Daman & Diu",
    "Delhi",
    "Goa",
    "Gujarat",
    "Haryana",
    "Himachal Pradesh",
    "Jammu & Kashmir",
    "Jharkhand",
    "Karnataka",
    "Kerala",
    "Ladakh",
    "Lakshadweep",
    "Madhya Pradesh",
    "Maharashtra",
    "Manipur",
    "Meghalaya",
    "Mizoram",
    "Nagaland",
    "Odisha",
    "Puducherry",
    "Punjab",
    "Rajasthan",
    "Sikkim",
    "Tamil Nadu",
    "Telangana",
    "Tripura",
    "Uttarakhand",
    "Uttar Pradesh",
    "West Bengal",
]

# ─── Tally XML Helpers ──────────────────────────────────────────────────────

def xml_escape(s: str) -> str:
    if not s:
        return ""
    return s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;").replace("'","&apos;")

def fmt_amt(num: float) -> str:
    return f"{num:.2f}"


def _append_round_off_entry(a, round_off_ledger: str, round_off: float, is_purchase_mode: bool):
    """Append a round-off ledger entry to the XML.

    round_off = rounded_total - actual_total  (signed, e.g. +0.40 or -0.60).
    Sales:    positive round_off → extra collected  → credit (income).
    Purchase: positive round_off → extra paid       → debit  (expense).
    """
    if not round_off_ledger or abs(round_off) < 0.005:
        return
    esc = xml_escape(round_off_ledger)
    # XOR: sales+positive → credit; purchase+positive → debit
    is_credit = (round_off > 0) != is_purchase_mode
    if is_credit:
        a('     <LEDGERENTRIES.LIST>')
        a(f'      <LEDGERNAME>{esc}</LEDGERNAME>')
        a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
        a(f'      <AMOUNT>{fmt_amt(abs(round_off))}</AMOUNT>')
        a('     </LEDGERENTRIES.LIST>')
    else:
        a('     <LEDGERENTRIES.LIST>')
        a(f'      <LEDGERNAME>{esc}</LEDGERNAME>')
        a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
        a(f'      <AMOUNT>-{fmt_amt(abs(round_off))}</AMOUNT>')
        a('     </LEDGERENTRIES.LIST>')

def tally_date(dt) -> str:
    today = datetime.today().strftime("%Y%m%d")

    if dt in (None, ""):
        return today
    if isinstance(dt, datetime):
        return dt.strftime("%Y%m%d")
    if isinstance(dt, date):
        return dt.strftime("%Y%m%d")

    # Excel serial date support.
    if isinstance(dt, (int, float)) and not isinstance(dt, bool):
        try:
            if float(dt) > 1000:
                return (datetime(1899, 12, 30) + timedelta(days=float(dt))).strftime("%Y%m%d")
        except Exception:
            pass

    text = str(dt).strip()
    if not text:
        return today

    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]

    # Already in compact numeric form.
    if re.fullmatch(r"\d{8}", text):
        if 1900 <= int(text[:4]) <= 2100:
            return text
        return f"{text[4:8]}{text[2:4]}{text[:2]}"

    candidates = [text]
    if " " in text:
        candidates.append(text.split(" ", 1)[0])

    formats = (
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d-%m-%y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d-%B-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    )
    for candidate in candidates:
        for fmt in formats:
            try:
                return datetime.strptime(candidate, fmt).strftime("%Y%m%d")
            except ValueError:
                continue

    return today


def _normalize_manual_date_to_tally(date_text: str) -> str:
    text = str(date_text or "").strip()
    if not text:
        raise ValueError("Custom date is empty.")

    compact = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\d{8}", compact):
        for fmt in ("%Y%m%d", "%d%m%Y"):
            try:
                return datetime.strptime(compact, fmt).strftime("%Y%m%d")
            except ValueError:
                continue

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            continue

    raise ValueError("Invalid custom date format. Use DD/MM/YYYY, DD-MM-YYYY, or YYYY-MM-DD.")

def push_to_tally(xml_str: str, host: str = "localhost", port: int = 9000) -> str:
    """Send XML to TallyPrime HTTP API and return response."""
    url = f"http://{host}:{port}"
    req = urllib.request.Request(url, data=xml_str.encode("utf-8"),
                                 headers={"Content-Type":"application/xml"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8")


def _row_get(row: dict, key: str, default=None):
    """Read a row value by key while tolerating header spacing/case differences."""
    if key in row:
        value = row.get(key)
        return default if value is None else value

    target = re.sub(r"\s+", "", str(key or "")).lower()
    for raw_key, raw_value in row.items():
        normalized = re.sub(r"\s+", "", str(raw_key or "")).lower()
        if normalized == target:
            return default if raw_value is None else raw_value
    return default


def _row_text(row: dict, key: str, default: str = "") -> str:
    value = _row_get(row, key, default)
    if value is None:
        return default
    return str(value).strip()


def _row_text_any(row: dict, keys: list, default: str = "") -> str:
    for key in keys or []:
        value = _row_text(row, key, "")
        if value:
            return value
    return default


def _resolve_party_ledger(row: dict, is_purchase_mode: bool = False) -> str:
    party_ledger = _row_text(row, "PartyLedger")
    if party_ledger:
        return party_ledger

    if is_purchase_mode:
        fallback_keys = [
            "PartyName",
            "Party Name",
            "SupplierName",
            "Supplier",
            "VendorName",
            "Vendor",
            "BillToName",
            "Bill To Name",
            "Party",
        ]
    else:
        fallback_keys = [
            "PartyName",
            "Party Name",
            "BuyerName",
            "CustomerName",
            "Customer",
            "BillToName",
            "Bill To Name",
            "Party",
        ]
    return _row_text_any(row, fallback_keys, "")


def _row_float(row: dict, key: str, default: float = 0.0) -> float:
    value = _row_get(row, key, default)
    if value in (None, ""):
        return float(default)
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _row_voucher_number(row: dict, default: str = "") -> str:
    return (
        _row_text(row, "InvoiceNo")
        or _row_text(row, "VoucherNo")
        or _row_text(row, "BillNo")
        or default
    )


def _row_invoice_reference(row: dict, default: str = "") -> str:
    return (
        _row_text(row, "ReferenceNo")
        or _row_text(row, "VoucherNo")
        or _row_text(row, "SupplierInvoiceNo")
        or _row_text(row, "BillNo")
        or default
    )


def _is_effectively_blank_ledger(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return True

    key = text.casefold()
    if key in {
        "na",
        "n/a",
        "none",
        "null",
        "nil",
        "not applicable",
        "* not applicable",
        "-",
        "--",
    }:
        return True

    compact = text.replace(",", "")
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", compact):
        try:
            return float(compact) == 0.0
        except ValueError:
            return False

    return False


def _ledger_or_suspense(value: str, fallback: str = SUSPENSE_LEDGER) -> str:
    text = str(value or "").strip()
    if _is_effectively_blank_ledger(text):
        text = ""
    return text or fallback


def _is_party_parent(parent: str) -> bool:
    key = re.sub(r"\s+", " ", str(parent or "")).strip().casefold()
    return key in {"sundry debtors", "sundry creditors"}


def _is_duties_parent(parent: str) -> bool:
    key = re.sub(r"\s+", " ", str(parent or "")).strip().casefold()
    return key in {"duties & taxes", "duties and taxes", "duty"}


def _current_fy_start() -> str:
    today = date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    return f"{fy_start_year}0401"


def _state_name_from_gstin(gstin: str) -> str:
    gstin_text = str(gstin or "").strip().upper()
    state_map = {
        "01": "Jammu And Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
        "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
        "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
        "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
        "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
        "24": "Gujarat", "25": "Daman And Diu", "26": "Dadra And Nagar Haveli And Daman And Diu",
        "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
        "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman And Nicobar Islands", "36": "Telangana",
        "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory", "99": "Centre Jurisdiction",
    }
    return state_map.get(gstin_text[:2], "")


def _normalize_gst_applicable(value: str, gstin: str = "") -> str:
    raw = str(value or "").strip()
    key = raw.casefold()
    if key in {"applicable", "yes", "y", "true", "1", "registered", "regular", "gst applicable"}:
        return "Applicable"
    if key in {"not applicable", "no", "n", "false", "0", "na", "n/a", "notapplicable"}:
        return "Not Applicable"
    if gstin:
        return "Applicable"
    return raw


def _normalize_gst_registration_type(value: str, gstin: str = "", gst_applicable: str = "") -> str:
    raw = str(value or "").strip()
    if not raw:
        if gstin or str(gst_applicable).strip().casefold() == "applicable":
            return "Regular"
        return ""

    key = raw.casefold()
    mapping = {
        "regular": "Regular",
        "registered": "Regular",
        "composition": "Composition",
        "consumer": "Consumer",
        "unregistered": "Unregistered",
        "sez": "SEZ",
        "sez unit": "SEZ",
        "sez developer": "SEZ",
        "overseas": "Overseas",
    }
    return mapping.get(key, raw)


def _normalize_state_for_ledger(value: str) -> str:
    text = str(value or "").strip()
    if text.casefold() in {"not applicable", "* not applicable", "na", "n/a"}:
        return ""
    return text


def _company_static_block(company: str) -> str:
    selected = str(company or "").strip()
    if not selected:
        return ""
    return f"   <STATICVARIABLES><SVCURRENTCOMPANY>{xml_escape(selected)}</SVCURRENTCOMPANY></STATICVARIABLES>"


def _collect_party_context(
    row: dict,
    party_ledger: str,
    allow_place_of_supply_column: bool = True,
) -> dict:
    party_ledger_raw = _ledger_or_suspense(party_ledger)
    party_name_raw = _row_text_any(
        row,
        [
            "PartyName",
            "BuyerName",
            "SupplierName",
            "BillToName",
            "Party",
        ],
        default=party_ledger_raw,
    )
    mailing_name_raw = _row_text_any(
        row,
        [
            "PartyMailingName",
            "MailingName",
            "BillingName",
            "Supplier",
            "Bill To Name",
        ],
        default=party_name_raw,
    )
    gstin_raw = _row_text_any(
        row,
        [
            "PartyGSTIN",
            "GSTIN",
            "GSTIN/UIN",
            "GSTIN UIN",
            "Party GSTIN",
            "SupplierGSTIN",
            "Supplier GSTIN",
            "GST No",
            "GST Number",
        ],
    ).upper()

    # Party's own state — look only at direct state columns, NOT PlaceOfSupply
    # (PlaceOfSupply is a delivery/billing concept for the sales side).
    state_raw = _row_text_any(
        row,
        [
            "PartyState",
            "State",
            "StateName",
            "State Name",
        ],
    )
    state_raw = _normalize_state_for_ledger(state_raw)
    if not state_raw and gstin_raw:
        state_raw = _state_name_from_gstin(gstin_raw)

    # PlaceOfSupply is valid for sales delivery context. For purchases,
    # do not let Excel POS override party state/GSTIN-derived state.
    if allow_place_of_supply_column:
        place_raw = _row_text_any(
            row,
            [
                "PlaceOfSupply",
                "Place Of Supply",
                "Place of Supply",
                "POS",
                "StateOfSupply",
            ],
            default=state_raw,
        )
    else:
        place_raw = state_raw
    place_raw = _normalize_state_for_ledger(place_raw)
    if not place_raw and gstin_raw:
        place_raw = _state_name_from_gstin(gstin_raw)

    # For unregistered/no-GSTIN parties: if State is absent but PlaceOfSupply is
    # provided, use PlaceOfSupply as State too. Tally needs STATENAME in the voucher
    # party block to avoid "Uncertain" in GSTR-1 when the ledger master has no state.
    if not state_raw and place_raw and not gstin_raw:
        state_raw = place_raw

    country_raw = _row_text_any(
        row,
        [
            "PartyCountry",
            "Country",
            "Country Name",
            "CountryOfResidence",
        ],
        default="India",
    )
    pincode_raw = _row_text_any(
        row,
        [
            "PartyPincode",
            "Pincode",
            "PinCode",
            "PIN",
            "PIN Code",
            "PostalCode",
            "Postal Code",
        ],
    )
    address1_raw = _row_text_any(
        row,
        [
            "PartyAddress1",
            "PartyAddressLine1",
            "Address1",
            "Address Line 1",
            "Address Line1",
            "AddressLine1",
            "BillToAddress",
            "Bill To Address",
            "Address",
        ],
    )
    address2_raw = _row_text_any(
        row,
        [
            "PartyAddress2",
            "PartyAddressLine2",
            "Address2",
            "Address Line 2",
            "Address Line2",
            "AddressLine2",
        ],
    )

    gst_app_raw = _row_text_any(
        row,
        [
            "GSTApplicable",
            "GST Applicable",
            "IsGSTApplicable",
            "GST",
        ],
    )
    reg_type_raw = _row_text_any(
        row,
        [
            "GSTRegistrationType",
            "GST Registration Type",
            "GST Reg Type",
            "RegistrationType",
            "Registration Type",
            "RegType",
            "Reg Type",
        ],
    )
    reg_type = _normalize_gst_registration_type(reg_type_raw, gstin=gstin_raw, gst_applicable=gst_app_raw)
    if reg_type.casefold() == "regular" and not gstin_raw:
        reg_type = ""

    return {
        "party_ledger": party_ledger_raw,
        "party_name": party_name_raw,
        "mailing_name": mailing_name_raw,
        "gstin": gstin_raw,
        "state": state_raw,
        "place_of_supply": place_raw,
        "country": country_raw or "India",
        "pincode": pincode_raw,
        "address1": address1_raw,
        "address2": address2_raw,
        "registration_type": reg_type,
    }


def _append_invoice_party_context_xml(
    add_line,
    party_context: dict,
    include_basic_buyer: bool = False,
    include_state: bool = True,
    include_place_of_supply: bool = True,
    place_of_supply_override=None,
) -> None:
    party_name = xml_escape(party_context.get("party_name", ""))
    mailing_name = xml_escape(party_context.get("mailing_name", "") or party_context.get("party_name", ""))
    party_gstin = xml_escape(party_context.get("gstin", ""))
    party_state = xml_escape(party_context.get("state", "")) if include_state else ""
    if place_of_supply_override is None:
        place_source = party_context.get("place_of_supply", "")
    else:
        place_source = place_of_supply_override
    place_of_supply = xml_escape(place_source) if include_place_of_supply else ""
    country = xml_escape(party_context.get("country", "") or "India")
    pincode = xml_escape(party_context.get("pincode", ""))
    address1 = xml_escape(party_context.get("address1", ""))
    address2 = xml_escape(party_context.get("address2", ""))
    reg_type_raw = str(party_context.get("registration_type", "") or "").strip()
    party_gstin_raw = str(party_context.get("gstin", "") or "").strip()
    country_raw = str(party_context.get("country", "") or "India").strip()

    # Tally marks vouchers as uncertain when GST registration is omitted.
    # Always send an explicit registration type for invoice vouchers.
    if not reg_type_raw:
        if party_gstin_raw:
            reg_type_raw = "Regular"
        elif country_raw and country_raw.casefold() not in {"india"}:
            reg_type_raw = "Overseas"
        else:
            reg_type_raw = "Unregistered/Consumer"
    # Tally requires "Unregistered/Consumer" (not just "Unregistered") for the
    # GSTREGISTRATIONTYPE field — using the short form leaves entries in "Uncertain".
    if reg_type_raw.casefold() == "unregistered":
        reg_type_raw = "Unregistered/Consumer"

    reg_type = xml_escape(reg_type_raw)

    if address1 or address2:
        add_line('     <ADDRESS.LIST TYPE="String">')
        if address1:
            add_line(f'      <ADDRESS>{address1}</ADDRESS>')
        if address2:
            add_line(f'      <ADDRESS>{address2}</ADDRESS>')
        add_line('     </ADDRESS.LIST>')

    if reg_type:
        add_line(f'     <GSTREGISTRATIONTYPE>{reg_type}</GSTREGISTRATIONTYPE>')
    # VATDEALERTYPE=Regular is required for all parties (including unregistered)
    # to prevent Tally from marking the entry as uncertain in GSTR-1.
    add_line('     <VATDEALERTYPE>Regular</VATDEALERTYPE>')

    if party_state:
        add_line(f'     <STATENAME>{party_state}</STATENAME>')
        add_line(f'     <PARTYSTATENAME>{party_state}</PARTYSTATENAME>')

    add_line(f'     <COUNTRYOFRESIDENCE>{country}</COUNTRYOFRESIDENCE>')

    if party_gstin:
        add_line(f'     <PARTYGSTIN>{party_gstin}</PARTYGSTIN>')

    if place_of_supply:
        add_line(f'     <PLACEOFSUPPLY>{place_of_supply}</PLACEOFSUPPLY>')

    if party_name:
        add_line(f'     <PARTYNAME>{party_name}</PARTYNAME>')
        add_line(f'     <BASICBASEPARTYNAME>{party_name}</BASICBASEPARTYNAME>')
        if include_basic_buyer:
            add_line(f'     <BASICBUYERNAME>{party_name}</BASICBUYERNAME>')

    if mailing_name:
        add_line(f'     <PARTYMAILINGNAME>{mailing_name}</PARTYMAILINGNAME>')
    # Consignee fields — Tally reads CONSIGNEESTATENAME for GSTR-1 state resolution
    # on unregistered party sales. Without it the entry stays "Uncertain".
    consignee_name = mailing_name or party_name
    if consignee_name:
        add_line(f'     <CONSIGNEEMAILINGNAME>{consignee_name}</CONSIGNEEMAILINGNAME>')
    consignee_state = party_state or place_of_supply
    if consignee_state:
        add_line(f'     <CONSIGNEESTATENAME>{consignee_state}</CONSIGNEESTATENAME>')
    add_line(f'     <CONSIGNEECOUNTRYNAME>{country}</CONSIGNEECOUNTRYNAME>')

    if pincode:
        add_line(f'     <PARTYPINCODE>{pincode}</PARTYPINCODE>')


def _state_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _is_gstin_like(value: str) -> bool:
    text = str(value or "").strip().upper()
    return bool(re.fullmatch(r"\d{2}[A-Z0-9]{13}", text))


def _pick_company_gst_registration(registrations: list, preferred_state: str = "") -> dict:
    rows = list(registrations or [])
    if not rows:
        return {}

    preferred_key = _state_key(preferred_state)

    def score(entry: dict) -> tuple:
        state_key = _state_key(entry.get("state", ""))
        name = str(entry.get("name", "") or "").strip()
        gstin = str(entry.get("gstin", "") or "").strip().upper()

        return (
            1 if (preferred_key and preferred_key == state_key) else 0,
            1 if (name and not _is_gstin_like(name)) else 0,
            1 if bool(gstin) else 0,
        )

    return max(rows, key=score)


def _resolve_company_gst_state(registrations: list, preferred_state: str = "") -> str:
    selected = _pick_company_gst_registration(registrations, preferred_state=preferred_state)
    if not selected:
        return ""
    return str(selected.get("state", "") or "").strip()


def _append_company_gst_context_xml(
    add_line,
    party_context: dict,
    company_gst_registrations: list = None,
    prefer_party_state: bool = True,
) -> None:
    registrations = list(company_gst_registrations or [])
    if not registrations:
        return

    preferred_state = ""
    if prefer_party_state:
        preferred_state = (
            str(party_context.get("place_of_supply", "") or "").strip()
            or str(party_context.get("state", "") or "").strip()
        )
    selected = _pick_company_gst_registration(registrations, preferred_state=preferred_state)
    if not selected:
        return

    reg_name_raw = str(selected.get("name", "") or "").strip()
    reg_gstin_raw = str(selected.get("gstin", "") or "").strip().upper()
    reg_state_raw = str(selected.get("state", "") or "").strip()

    if not reg_name_raw and reg_state_raw:
        reg_name_raw = f"{reg_state_raw} Registration"
    if not reg_name_raw and reg_gstin_raw:
        reg_name_raw = reg_gstin_raw

    reg_name = xml_escape(reg_name_raw)
    reg_gstin = xml_escape(reg_gstin_raw)
    reg_state = xml_escape(reg_state_raw)

    if reg_name and reg_gstin:
        add_line(
            f'     <GSTREGISTRATION TAXTYPE="GST" TAXREGISTRATION="{reg_gstin}">{reg_name}</GSTREGISTRATION>'
        )
        add_line(f'     <CMPGSTIN>{reg_gstin}</CMPGSTIN>')
        add_line('     <CMPGSTREGISTRATIONTYPE>Regular</CMPGSTREGISTRATIONTYPE>')
    if reg_state:
        add_line(f'     <CMPGSTSTATE>{reg_state}</CMPGSTSTATE>')


def _append_tax_object_allocation_xml(add_line, tax_classification: str) -> None:
    tax_class = xml_escape(str(tax_classification or "").strip())
    if not tax_class:
        return
    # Many Tally companies do not define Tax Classification masters named
    # exactly IGST/CGST/SGST. Emitting those names causes import exception.
    if tax_class.casefold() in {"igst", "cgst", "sgst", "utgst", "cess"}:
        return
    add_line('      <TAXOBJECTALLOCATIONS.LIST>')
    add_line('       <TAXOBJECTALLOCATIONS>')
    add_line('        <TAXTYPE>GST</TAXTYPE>')
    add_line('        <TAXABILITY>Taxable</TAXABILITY>')
    add_line(f'        <TAXCLASSIFICATIONNAME>{tax_class}</TAXCLASSIFICATIONNAME>')
    add_line('       </TAXOBJECTALLOCATIONS>')
    add_line('      </TAXOBJECTALLOCATIONS.LIST>')


def _gst_transaction_type(reg_type: str, gstin: str = "") -> str:
    """Return the GSTTRANSACTIONTYPE value Tally needs to avoid 'Uncertain' in GSTR.
    Rules:
      - Regular registered party with GSTIN  -> 'Tax Invoice' (B2B)
      - Composition/Consumer/Unregistered    -> 'Unregistered'
      - Overseas / non-India                 -> 'Overseas'
      - SEZ                                  -> 'SEZ exports with payment'
      - Empty / unknown                      -> 'Unregistered'
    """
    key = str(reg_type or "").strip().casefold()
    if key in {"regular", "registered"} and gstin:
        return "Tax Invoice"
    if key in {"composition", "consumer", "unregistered", ""}:
        return "Unregistered"
    if key in {"overseas"}:
        return "Overseas"
    if key in {"sez", "sez unit", "sez developer"}:
        return "SEZ exports with payment"
    # fallback
    if gstin:
        return "Tax Invoice"
    return "Unregistered"


def _pick_tax_ledger_name(row: dict, ledger_keys: list, rate_value: float, default_name: str, amount_value: float = 0.0) -> str:
    tax_ledger_raw = _row_text_any(row, ledger_keys, "")
    if _is_effectively_blank_ledger(tax_ledger_raw):
        tax_ledger_raw = ""
    if (rate_value > 0 or amount_value > 0) and not tax_ledger_raw:
        tax_ledger_raw = default_name
    return _ledger_or_suspense(tax_ledger_raw)


def _normalize_company_name(value) -> str:
    text = html.unescape(str(value or ""))
    text = text.replace("\x00", "")
    text = re.sub(r"[\x01-\x1F\x7F]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _company_key(value) -> str:
    return _normalize_company_name(value).upper()


def _is_valid_company_name(value) -> bool:
    name = _normalize_company_name(value)
    if not name:
        return False
    return not name.isdigit()


def _build_tally_url(host: str, port: str) -> str:
    host_text = (host or "localhost").strip()
    port_text = (port or "9000").strip()
    if host_text.startswith("http://"):
        host_text = host_text[7:]
    elif host_text.startswith("https://"):
        host_text = host_text[8:]
    host_text = host_text.strip("/") or "localhost"
    if "/" in host_text:
        host_text = host_text.split("/", 1)[0]
    if not port_text.isdigit():
        raise ValueError("Port must be numeric.")
    return f"http://{host_text}:{port_text}"


def _post_tally_xml(tally_url: str, xml_payload: str, timeout: float = 15.0) -> str:
    request = urllib.request.Request(
        tally_url,
        data=xml_payload.encode("utf-8"),
        headers={"Content-Type": "application/xml"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="replace")


def _fetch_voucher_types_from_tally(
    tally_url: str,
    company: str = "",
    prefix: str = "",
    timeout: float = 10.0,
) -> list:
    """
    Fetch all voucher type names from Tally for the given company.
    If prefix is provided (e.g. "Sales" or "Purchase"), only matching names are returned.
    Returns a list of voucher type name strings, or empty list on failure.
    """
    static = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if company:
        static += f"<SVCURRENTCOMPANY>{xml_escape(company)}</SVCURRENTCOMPANY>"
    static += "</STATICVARIABLES>"

    payload = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>AllVoucherTypes</ID></HEADER>"
        "<BODY><DESC>"
        f"{static}"
        "<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='AllVoucherTypes'>"
        "<TYPE>VoucherType</TYPE>"
        "<FETCH>Name</FETCH>"
        "</COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    names = []
    try:
        resp = _post_tally_xml(tally_url, payload, timeout=timeout)
        try:
            root = ET.fromstring(resp)
            for node in root.iter():
                tag = str(node.tag or "")
                if "}" in tag:
                    tag = tag.split("}", 1)[1]
                tag = tag.upper()
                if tag == "VOUCHERTYPE":
                    name = (node.attrib.get("NAME") or "").strip()
                    if not name:
                        for child in node:
                            ctag = str(child.tag or "")
                            if "}" in ctag:
                                ctag = ctag.split("}", 1)[1]
                            if ctag.upper() == "NAME" and child.text:
                                name = child.text.strip()
                                break
                    if name and name not in names:
                        names.append(name)
        except ET.ParseError:
            for m in re.findall(r'<VOUCHERTYPE[^>]+NAME="([^"]*)"', resp, re.IGNORECASE):
                n = m.strip()
                if n and n not in names:
                    names.append(n)
    except Exception:
        pass

    if prefix:
        prefix_lower = prefix.strip().lower()
        names = [n for n in names if n.lower().startswith(prefix_lower)]

    return names


def _check_tally_connection(tally_url: str, timeout: float = 5.0) -> dict:
    probe_xml = (
        "<ENVELOPE><HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
        "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
        "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>"
    )
    try:
        _post_tally_xml(tally_url, probe_xml, timeout=timeout)
        return {"connected": True}
    except HTTPError as exc:
        return {"connected": False, "error": f"HTTP {exc.code}"}
    except URLError:
        return {"connected": False, "error": "ConnectionError"}
    except Exception as exc:
        return {"connected": False, "error": str(exc)}


def _extract_company_names(response_text: str) -> set:
    names = set()
    try:
        root = ET.fromstring(response_text)
        for node in root.iter():
            tag = str(node.tag or "").upper()
            txt = _normalize_company_name(node.text)
            attr_name = _normalize_company_name(node.attrib.get("NAME") or "")
            if tag in {"COMPANYNAME", "SVCURRENTCOMPANY", "CURRENTCOMPANY"} and _is_valid_company_name(txt):
                names.add(txt)
            if "COMPANY" in tag and _is_valid_company_name(attr_name):
                names.add(attr_name)
            if tag == "COMPANY" and _is_valid_company_name(txt):
                names.add(txt)
    except ET.ParseError:
        pass

    patterns = [
        r'COMPANY[^>]*NAME="([^"]+)"',
        r"<COMPANYNAME>(.*?)</COMPANYNAME>",
        r"<SVCURRENTCOMPANY>(.*?)</SVCURRENTCOMPANY>",
        r"<COMPANY[^>]*>.*?<NAME>(.*?)</NAME>",
    ]
    for pattern in patterns:
        for match in re.findall(pattern, response_text, flags=re.IGNORECASE | re.DOTALL):
            value = _normalize_company_name(match)
            if _is_valid_company_name(value):
                names.add(value)
    return names


def _fetch_tally_companies(tally_url: str, timeout: float = 15.0) -> dict:
    requests_xml = [
        (
            "report-list-companies",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>",
        ),
        (
            "collection-company",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
            "<TYPE>Collection</TYPE><ID>Company Collection</ID></HEADER><BODY><DESC>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "<TDL><TDLMESSAGE><COLLECTION NAME='Company Collection'>"
            "<TYPE>Company</TYPE><FETCH>Name</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
            "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>",
        ),
    ]

    companies = set()
    errors = []
    for label, xml_payload in requests_xml:
        try:
            response_text = _post_tally_xml(tally_url, xml_payload, timeout=timeout)
            companies.update(_extract_company_names(response_text))
        except HTTPError as exc:
            errors.append(f"{label}: HTTP {exc.code}")
        except URLError:
            errors.append(f"{label}: ConnectionError")
        except Exception as exc:
            errors.append(f"{label}: {exc}")

    sorted_companies = sorted(companies, key=lambda x: _company_key(x))
    if sorted_companies:
        return {"success": True, "companies": sorted_companies}

    err = "; ".join(errors) if errors else "No companies returned by Tally."
    return {"success": False, "error": err, "companies": []}


def _extract_ledger_names(response_text: str) -> set:
    names = set()

    try:
        root = ET.fromstring(response_text)
        for node in root.iter():
            tag = str(node.tag or "").upper()
            if tag == "LEDGER":
                attr_name = _normalize_company_name(node.attrib.get("NAME") or "")
                if attr_name:
                    names.add(attr_name)
                for child in list(node):
                    child_tag = str(child.tag or "").upper()
                    child_text = _normalize_company_name(child.text)
                    if child_tag == "NAME" and child_text:
                        names.add(child_text)
    except ET.ParseError:
        pass

    for match in re.findall(r"<LEDGER\b[^>]*\bNAME=\"([^\"]+)\"", response_text, flags=re.IGNORECASE):
        value = _normalize_company_name(match)
        if value:
            names.add(value)

    return names


def _fetch_existing_ledger_names(tally_url: str, company_name: str = "", timeout: float = 15.0) -> dict:
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>Ledger Name Lookup</ID></HEADER>"
        f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='Ledger Name Lookup'><TYPE>Ledger</TYPE>"
        "<FETCH>Name,Parent</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD></COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    try:
        response_text = _post_tally_xml(tally_url, request_xml, timeout=timeout)
    except HTTPError as exc:
        return {"success": False, "error": f"HTTP {exc.code}", "ledgers": set()}
    except URLError:
        return {"success": False, "error": "ConnectionError", "ledgers": set()}
    except Exception as exc:
        return {"success": False, "error": str(exc), "ledgers": set()}

    ledgers = _extract_ledger_names(response_text)
    if ledgers:
        return {"success": True, "ledgers": ledgers}

    return {
        "success": False,
        "error": "No ledgers returned for selected company.",
        "ledgers": set(),
    }


def _extract_company_gst_registrations(response_text: str) -> list:
    registrations = []
    seen = set()

    try:
        root = ET.fromstring(response_text)
        for tax_unit in root.findall(".//TAXUNIT"):
            tax_type = str(tax_unit.attrib.get("TAXTYPE") or tax_unit.findtext("TAXTYPE") or "").strip().upper()
            if tax_type and tax_type != "GST":
                continue

            name_raw = str(tax_unit.attrib.get("NAME") or tax_unit.findtext("NAME") or "").strip()
            gstin_raw = str(
                tax_unit.attrib.get("TAXREGISTRATION")
                or tax_unit.findtext("GSTREGNUMBER")
                or tax_unit.findtext("GSTIN")
                or ""
            ).strip().upper()
            state_raw = str(tax_unit.findtext("STATENAME") or "").strip()

            if not gstin_raw and _is_gstin_like(name_raw):
                gstin_raw = name_raw.upper()
            if not state_raw and gstin_raw:
                state_raw = _state_name_from_gstin(gstin_raw)

            if not gstin_raw:
                continue

            name = _normalize_company_name(name_raw or gstin_raw)
            if not name:
                continue

            key = (name.casefold(), gstin_raw)
            if key in seen:
                continue
            seen.add(key)

            registrations.append(
                {
                    "name": name,
                    "gstin": gstin_raw,
                    "state": state_raw,
                }
            )
    except ET.ParseError:
        pass

    return registrations


def _fetch_company_gst_registrations(tally_url: str, company_name: str = "", timeout: float = 15.0) -> dict:
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>Tax Unit Lookup</ID></HEADER>"
        f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='Tax Unit Lookup'><TYPE>TaxUnit</TYPE>"
        "<FETCH>Name,TaxType,TaxRegistration,GSTRegNumber,StateName,UseFor</FETCH>"
        "<NATIVEMETHOD>Name</NATIVEMETHOD></COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    try:
        response_text = _post_tally_xml(tally_url, request_xml, timeout=timeout)
    except HTTPError as exc:
        return {"success": False, "error": f"HTTP {exc.code}", "registrations": []}
    except URLError:
        return {"success": False, "error": "ConnectionError", "registrations": []}
    except Exception as exc:
        return {"success": False, "error": str(exc), "registrations": []}

    registrations = _extract_company_gst_registrations(response_text)
    if registrations:
        return {"success": True, "registrations": registrations}

    return {
        "success": False,
        "error": "No GST registrations returned for selected company.",
        "registrations": [],
    }


def _safe_int(text, default=0) -> int:
    try:
        return int(float(str(text).strip()))
    except (TypeError, ValueError):
        return default


def _parse_tally_response_details(response_text: str) -> dict:
    details = {
        "success": False,
        "created": 0,
        "altered": 0,
        "deleted": 0,
        "lastvchid": 0,
        "lastmid": 0,
        "combined": 0,
        "ignored": 0,
        "errors": 0,
        "cancelled": 0,
        "exceptions": 0,
        "line_errors": [],
        "error": "",
    }

    try:
        root = ET.fromstring(response_text)
    except ET.ParseError as exc:
        details["error"] = f"Could not parse Tally response: {exc}"
        return details

    def _count(tag_name: str) -> int:
        node = root.find(f".//{tag_name}")
        return _safe_int(node.text if node is not None else 0)

    details["created"] = _count("CREATED")
    details["altered"] = _count("ALTERED")
    details["deleted"] = _count("DELETED")
    details["lastvchid"] = _count("LASTVCHID")
    details["lastmid"] = _count("LASTMID")
    details["combined"] = _count("COMBINED")
    details["ignored"] = _count("IGNORED")
    details["errors"] = _count("ERRORS")
    details["cancelled"] = _count("CANCELLED")
    details["exceptions"] = _count("EXCEPTIONS")

    line_errors = []
    seen = set()
    for node in root.findall(".//LINEERROR"):
        text = (node.text or "").strip()
        if text and text not in seen:
            seen.add(text)
            line_errors.append(text)
    details["line_errors"] = line_errors

    details["success"] = (
        details["errors"] == 0
        and details["exceptions"] == 0
        and not details["line_errors"]
    )
    if not details["success"] and details["line_errors"]:
        details["error"] = details["line_errors"][0]
    return details


def _ledger_name_key(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _filter_out_existing_ledgers(ledger_defs: list, existing_ledger_names) -> list:
    existing_keys = {
        _ledger_name_key(name)
        for name in (existing_ledger_names or set())
        if str(name or "").strip()
    }
    filtered = []
    seen = set()

    for entry in ledger_defs or []:
        name = str((entry or {}).get("Name", "") or "").strip()
        key = _ledger_name_key(name)
        if not key or key in seen or key in existing_keys:
            continue
        seen.add(key)
        filtered.append(entry)

    return filtered


def _extract_stock_item_balances(response_text: str) -> dict:
    balances = {}

    try:
        root = ET.fromstring(response_text)
    except ET.ParseError:
        return balances

    for stock_item in root.findall(".//STOCKITEM"):
        name = str(
            stock_item.attrib.get("NAME")
            or stock_item.findtext("NAME")
            or ""
        ).strip()
        if not name:
            continue
        base_unit = str(stock_item.findtext("BASEUNITS") or "").strip()
        closing_balance = str(stock_item.findtext("CLOSINGBALANCE") or "").strip()
        qty_match = re.search(r"-?\d+(?:\.\d+)?", closing_balance)
        balances[_ledger_name_key(name)] = {
            "name": name,
            "base_unit": base_unit,
            "closing_balance": closing_balance,
            "quantity": float(qty_match.group(0)) if qty_match else 0.0,
        }

    return balances


def _fetch_stock_item_balances(tally_url: str, company_name: str = "", timeout: float = 15.0) -> dict:
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>Stock Item Balance Lookup</ID></HEADER>"
        f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='Stock Item Balance Lookup'><TYPE>StockItem</TYPE>"
        "<FETCH>Name,BaseUnits,ClosingBalance</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD></COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    try:
        response_text = _post_tally_xml(tally_url, request_xml, timeout=timeout)
    except HTTPError as exc:
        return {"success": False, "error": f"HTTP {exc.code}", "items": {}}
    except URLError:
        return {"success": False, "error": "ConnectionError", "items": {}}
    except Exception as exc:
        return {"success": False, "error": str(exc), "items": {}}

    items = _extract_stock_item_balances(response_text)
    if items:
        return {"success": True, "items": items}

    return {"success": False, "error": "No stock items returned for selected company.", "items": {}}


def _find_conflicting_row_voucher_numbers(rows: list, existing_numbers) -> list:
    existing_keys = {
        _normalize_voucher_number_text(value)
        for value in (existing_numbers or set())
        if _normalize_voucher_number_text(value)
    }
    conflicts = []
    seen = set()

    for idx, row in enumerate(rows or []):
        voucher_no = _normalize_voucher_number_text(_row_voucher_number(row))
        if not voucher_no or voucher_no not in existing_keys or voucher_no in seen:
            continue
        seen.add(voucher_no)
        conflicts.append(f"row {idx + 1} voucher '{voucher_no}' already exists")

    return conflicts


def _diagnose_item_stock_issues(rows: list, stock_items: dict) -> list:
    issues = []

    for idx, row in enumerate(rows or []):
        item_name_raw = (
            _row_text(row, "ItemName")
            or _row_text(row, "Item")
            or _row_text(row, "StockItem")
            or _row_text(row, "ProductName")
            or _row_text(row, "SalesLedger")
        )
        if not item_name_raw:
            continue

        requested_qty = _row_float(row, "Quantity", 0.0)
        if requested_qty <= 0:
            requested_qty = _row_float(row, "Qty", 0.0)
        if requested_qty <= 0:
            requested_qty = _row_float(row, "Unit", 0.0)

        item = (stock_items or {}).get(_ledger_name_key(item_name_raw))
        if item is None:
            issues.append(f"row {idx + 1} stock item '{item_name_raw}' does not exist in Tally")
            continue

        if requested_qty > float(item.get("quantity", 0.0) or 0.0) + 1e-9:
            requested_unit = (
                _row_text(row, "Per", "")
                or _row_text(row, "UOM", "")
                or _row_text(row, "Unit", "")
                or item.get("base_unit", "")
            )
            requested_unit = _normalize_stock_unit_name(requested_unit) or str(item.get("base_unit", "") or "").strip()
            available_text = str(item.get("closing_balance", "") or "").strip()
            if not available_text:
                available_text = f"{fmt_amt(float(item.get('quantity', 0.0) or 0.0))} {item.get('base_unit', '')}".strip()
            issues.append(
                f"row {idx + 1} item '{item.get('name', item_name_raw)}' needs {fmt_amt(requested_qty)} {requested_unit} "
                f"but Tally stock is {available_text}"
            )

    return issues


def _infer_tax_type_from_ledger_name(ledger_name: str) -> str:
    key = _ledger_name_key(ledger_name)
    if not key:
        return ""
    if "igst" in key or "integrated tax" in key:
        return "Integrated Tax"
    if "cgst" in key or "central tax" in key:
        return "Central Tax"
    if "sgst" in key or "utgst" in key or "state tax" in key or "ut tax" in key or "state/ut tax" in key:
        return "State Tax"
    if "cess" in key:
        return "Cess"
    return ""


def _is_protected_gst_tax_ledger(ledger_name: str, parent_name: str, tax_type: str = "") -> bool:
    if not _is_duties_parent(parent_name):
        return False

    key = _ledger_name_key(ledger_name)
    tax_key = _ledger_name_key(tax_type)
    if not key and not tax_key:
        return False

    if tax_key in {"integrated tax", "central tax", "state tax", "state/ut tax", "ut tax", "cess"}:
        return True

    if key in {"igst", "cgst", "sgst", "utgst", "cess", "integrated tax", "central tax", "state tax"}:
        return True

    marker_tokens = (
        "igst",
        "cgst",
        "sgst",
        "utgst",
        "integrated tax",
        "central tax",
        "state tax",
        "ut tax",
        "state/ut tax",
        "cess",
        "gst",
    )
    return any(token in key for token in marker_tokens)


def _collect_party_ledger_definition(row: dict, is_purchase_mode: bool) -> dict:
    party_ledger_raw = _resolve_party_ledger(row, is_purchase_mode)
    if not party_ledger_raw:
        return {}

    party_context = _collect_party_context(
        row,
        party_ledger_raw,
        allow_place_of_supply_column=not is_purchase_mode,
    )
    party_name_raw = str(party_context.get("party_ledger", "") or party_ledger_raw).strip()
    if not party_name_raw:
        return {}

    gstin_raw = str(party_context.get("gstin", "") or "").strip().upper()
    gst_app_raw = _row_text_any(
        row,
        [
            "GSTApplicable",
            "GST Applicable",
            "IsGSTApplicable",
            "GST",
        ],
    )
    gst_applicable = _normalize_gst_applicable(gst_app_raw, gstin=gstin_raw)
    if not gst_applicable:
        gst_applicable = "Applicable" if gstin_raw else "Not Applicable"
    if not gstin_raw and str(gst_applicable).strip().casefold() == "applicable":
        gst_applicable = "Not Applicable"

    reg_type_raw = _row_text_any(
        row,
        [
            "GSTRegistrationType",
            "GST Registration Type",
            "GST Reg Type",
            "RegistrationType",
            "Registration Type",
            "RegType",
            "Reg Type",
        ],
    )
    gst_reg_type = _normalize_gst_registration_type(
        reg_type_raw,
        gstin=gstin_raw,
        gst_applicable=gst_applicable,
    )
    if not gst_reg_type:
        gst_reg_type = str(party_context.get("registration_type", "") or "").strip()
    if gst_reg_type.casefold() == "regular" and not gstin_raw:
        gst_reg_type = ""

    state_raw = _normalize_state_for_ledger(
        str(party_context.get("state", "") or party_context.get("place_of_supply", "") or "")
    )
    if not state_raw and gstin_raw:
        state_raw = _state_name_from_gstin(gstin_raw)

    return {
        "Name": party_name_raw,
        "Parent": "Sundry Creditors" if is_purchase_mode else "Sundry Debtors",
        "GSTApplicable": gst_applicable,
        "GSTIN": gstin_raw,
        "PAN": _pan_from_gstin(gstin_raw),
        "StateOfSupply": state_raw,
        "Address1": str(party_context.get("address1", "") or "").strip(),
        "Address2": str(party_context.get("address2", "") or "").strip(),
        "MailingName": str(
            party_context.get("mailing_name", "")
            or party_context.get("party_name", "")
            or party_name_raw
        ).strip(),
        "Country": str(party_context.get("country", "") or "India").strip() or "India",
        "Pincode": str(party_context.get("pincode", "") or "").strip(),
        "Billwise": "Yes",
        "GSTRegistrationType": gst_reg_type,
        "TypeOfTaxation": "",
        "GSTRate": "",
    }


def _merge_ledger_definitions(existing: dict, candidate: dict) -> None:
    if not candidate:
        return

    parent_priority = {
        "": 0,
        "sales accounts": 1,
        "purchase accounts": 1,
        "sundry debtors": 2,
        "sundry creditors": 2,
        "duties & taxes": 3,
        "duties and taxes": 3,
    }

    existing_parent = str(existing.get("Parent", "") or "").strip()
    candidate_parent = str(candidate.get("Parent", "") or "").strip()
    existing_rank = parent_priority.get(existing_parent.casefold(), 1 if existing_parent else 0)
    candidate_rank = parent_priority.get(candidate_parent.casefold(), 1 if candidate_parent else 0)
    if candidate_rank > existing_rank:
        existing["Parent"] = candidate_parent
    elif not existing_parent and candidate_parent:
        existing["Parent"] = candidate_parent

    for field, value in candidate.items():
        if field in {"Name", "Parent"}:
            continue
        if existing.get(field) in (None, "") and value not in (None, ""):
            existing[field] = value


def _extract_missing_ledgers_from_line_errors(line_errors: list) -> list:
    missing = []
    seen = set()
    patterns = [
        re.compile(r"Ledger\s+'([^']+)'\s+does\s+not\s+exist", re.IGNORECASE),
        re.compile(r'Ledger\s+"([^"]+)"\s+does\s+not\s+exist', re.IGNORECASE),
    ]

    for err in line_errors or []:
        text = str(err or "")
        for pattern in patterns:
            for raw_name in pattern.findall(text):
                name = str(raw_name or "").strip()
                key = _ledger_name_key(name)
                if not name or key in seen:
                    continue
                seen.add(key)
                missing.append(name)
    return missing


def _collect_auto_voucher_ledgers(rows: list, mode: str) -> list:
    """Collect voucher ledgers so missing ledgers can be pre-created."""
    entries = {}
    is_purchase_mode = mode in {"purchase_accounting", "purchase_item"}

    def add_entry(
        name: str,
        parent: str,
        tax_type: str = "",
        gst_rate: str = "",
        extra_fields: dict = None,
    ):
        ledger_name = str(name or "").strip()
        if _is_effectively_blank_ledger(ledger_name):
            return

        key = _ledger_name_key(ledger_name)
        if not key:
            return

        candidate = {
            "Name": ledger_name,
            "Parent": parent,
            "GSTApplicable": "",
            "GSTIN": "",
            "StateOfSupply": "",
            "TypeOfTaxation": tax_type or "",
            "GSTRate": str(gst_rate).strip() if gst_rate not in (None, "") else "",
        }

        if extra_fields:
            for field, value in extra_fields.items():
                if field in {"Name", "Parent"}:
                    continue
                if value in (None, ""):
                    continue
                candidate[field] = value

        existing = entries.get(key)
        if existing is None:
            entries[key] = candidate
            return

        _merge_ledger_definitions(existing, candidate)

    for r in rows or []:
        party_entry = _collect_party_ledger_definition(r, is_purchase_mode)
        if party_entry:
            add_entry(
                party_entry.get("Name", ""),
                party_entry.get("Parent", ""),
                extra_fields=party_entry,
            )

        if is_purchase_mode:
            purchase_ledger = (
                _row_text(r, "PurchaseLedger")
                or _row_text(r, "PurchaseAccount")
                or _row_text(r, "Purchase Ledger")
                or _row_text(r, "ExpenseLedger")
                or _row_text(r, "SalesLedger")
            )
            add_entry(purchase_ledger, "Purchase Accounts")
        else:
            sales_ledger = (
                _row_text(r, "SalesLedger")
                or _row_text(r, "SalesAccount")
                or _row_text(r, "Sales Ledger")
                or _row_text(r, "IncomeLedger")
            )
            add_entry(sales_ledger, "Sales Accounts")

        cgst_rate_val = _row_float(r, "CGSTRate", 0.0)
        sgst_rate_val = _row_float(r, "SGSTRate", 0.0)
        igst_rate_val = _row_float(r, "IGSTRate", 0.0)

        cgst_ledger_name = _row_text_any(
            r,
            ["CGSTLedger", "CGST Ledger", "CentralTaxLedger", "Central Tax Ledger", "Central Tax"],
            "",
        )
        sgst_ledger_name = _row_text_any(
            r,
            ["SGSTLedger", "SGST Ledger", "StateTaxLedger", "State Tax Ledger", "State Tax", "UTGSTLedger", "UTGST Ledger"],
            "",
        )
        igst_ledger_name = _row_text_any(
            r,
            ["IGSTLedger", "IGST Ledger", "IntegratedTaxLedger", "Integrated Tax Ledger", "Integrated Tax"],
            "",
        )

        if _is_effectively_blank_ledger(cgst_ledger_name):
            cgst_ledger_name = ""
        if _is_effectively_blank_ledger(sgst_ledger_name):
            sgst_ledger_name = ""
        if _is_effectively_blank_ledger(igst_ledger_name):
            igst_ledger_name = ""

        if cgst_rate_val > 0 and not cgst_ledger_name:
            cgst_ledger_name = "CGST"
        if sgst_rate_val > 0 and not sgst_ledger_name:
            sgst_ledger_name = "SGST"
        if igst_rate_val > 0 and not igst_ledger_name:
            igst_ledger_name = "IGST"

        add_entry(cgst_ledger_name, "Duties & Taxes", "Central Tax", _row_text(r, "CGSTRate"))
        add_entry(sgst_ledger_name, "Duties & Taxes", "State Tax", _row_text(r, "SGSTRate"))
        add_entry(igst_ledger_name, "Duties & Taxes", "Integrated Tax", _row_text(r, "IGSTRate"))

    return list(entries.values())


def _build_missing_ledger_defs(line_errors: list, rows: list, mode: str) -> list:
    missing_names = _extract_missing_ledgers_from_line_errors(line_errors)
    if not missing_names:
        return []

    is_purchase_mode = mode in {"purchase_accounting", "purchase_item"}

    party_keys = set()
    party_defs = {}
    purchase_keys = set()
    sales_keys = set()
    tax_type_map = {}
    tax_rate_map = {}

    for r in rows or []:
        party_entry = _collect_party_ledger_definition(r, is_purchase_mode)
        if party_entry:
            party_key = _ledger_name_key(party_entry.get("Name", ""))
            if party_key:
                party_keys.add(party_key)
                existing_party = party_defs.get(party_key)
                if existing_party is None:
                    party_defs[party_key] = party_entry
                else:
                    _merge_ledger_definitions(existing_party, party_entry)

        purchase_ledger = (
            _row_text(r, "PurchaseLedger")
            or _row_text(r, "PurchaseAccount")
            or _row_text(r, "Purchase Ledger")
            or _row_text(r, "ExpenseLedger")
            or _row_text(r, "SalesLedger")
        )
        if purchase_ledger:
            purchase_keys.add(_ledger_name_key(purchase_ledger))

        sales_ledger = (
            _row_text(r, "SalesLedger")
            or _row_text(r, "SalesAccount")
            or _row_text(r, "Sales Ledger")
            or _row_text(r, "IncomeLedger")
        )
        if sales_ledger:
            sales_keys.add(_ledger_name_key(sales_ledger))

        tax_configs = [
            (
                "Central Tax",
                "CGST",
                ["CGSTLedger", "CGST Ledger", "CentralTaxLedger", "Central Tax Ledger", "Central Tax"],
                ["CGSTRate", "CGST Rate"],
            ),
            (
                "State Tax",
                "SGST",
                ["SGSTLedger", "SGST Ledger", "StateTaxLedger", "State Tax Ledger", "State Tax", "UTGSTLedger", "UTGST Ledger"],
                ["SGSTRate", "SGST Rate"],
            ),
            (
                "Integrated Tax",
                "IGST",
                ["IGSTLedger", "IGST Ledger", "IntegratedTaxLedger", "Integrated Tax Ledger", "Integrated Tax"],
                ["IGSTRate", "IGST Rate"],
            ),
        ]
        for tax_type, default_ledger, ledger_cols, rate_cols in tax_configs:
            tax_name = _row_text_any(r, ledger_cols, "")
            if _is_effectively_blank_ledger(tax_name):
                tax_name = ""
            rate_val = ""
            rate_num = 0.0
            for rate_col in rate_cols:
                candidate_rate_val = _row_text(r, rate_col, "")
                if candidate_rate_val:
                    rate_val = candidate_rate_val
                candidate_rate_num = _row_float(r, rate_col, 0.0)
                if candidate_rate_num > 0:
                    rate_num = candidate_rate_num
                    break

            if not tax_name and rate_num > 0:
                tax_name = default_ledger
            if not tax_name:
                continue

            key = _ledger_name_key(tax_name)
            tax_type_map.setdefault(key, tax_type)
            if rate_val and key not in tax_rate_map:
                tax_rate_map[key] = rate_val

    entries = []
    seen = set()
    for ledger_name in missing_names:
        if _is_effectively_blank_ledger(ledger_name):
            continue
        key = _ledger_name_key(ledger_name)
        if key in seen:
            continue
        seen.add(key)

        tax_type = ""
        gst_rate = ""
        if key in party_keys:
            party_entry = dict(party_defs.get(key) or {})
            if party_entry:
                party_entry["Name"] = party_entry.get("Name") or ledger_name
                entries.append(party_entry)
                continue
            parent = "Sundry Creditors" if is_purchase_mode else "Sundry Debtors"
        elif key in purchase_keys:
            parent = "Purchase Accounts"
        elif key in sales_keys:
            parent = "Sales Accounts"
        elif key in tax_type_map or "gst" in key or "tax" in key:
            parent = "Duties & Taxes"
            tax_type = tax_type_map.get(key, "")
            gst_rate = tax_rate_map.get(key, "")
            if not tax_type:
                tax_type = _infer_tax_type_from_ledger_name(ledger_name)
        else:
            parent = "Purchase Accounts" if is_purchase_mode else "Sales Accounts"

        entries.append(
            {
                "Name": ledger_name,
                "Parent": parent,
                "GSTApplicable": "",
                "GSTIN": "",
                "StateOfSupply": "",
                "Address1": "",
                "Address2": "",
                "MailingName": "",
                "Country": "India",
                "Pincode": "",
                "Billwise": "No",
                "GSTRegistrationType": "",
                "TypeOfTaxation": tax_type,
                "GSTRate": str(gst_rate).strip() if gst_rate not in (None, "") else "",
            }
        )

    return entries


def _extract_numeric_voucher_no(value):
    text = _normalize_voucher_number_text(value)
    if text.isdigit():
        return int(text)
    return None


def _normalize_voucher_number_text(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _increment_voucher_number_text(value: str):
    text = _normalize_voucher_number_text(value)
    if not text:
        return None
    if text.isdigit():
        return str(int(text) + 1)

    match = re.search(r"(\d+)(?!.*\d)", text)
    if not match:
        return None

    digits = match.group(1)
    start, end = match.span(1)
    return f"{text[:start]}{str(int(digits) + 1).zfill(len(digits))}{text[end:]}"


def _voucher_number_with_offset(start_value, offset: int) -> str:
    text = _normalize_voucher_number_text(start_value)
    if not text:
        return ""
    if offset <= 0:
        return text
    if text.isdigit():
        return str(int(text) + offset)

    current = text
    for idx in range(offset):
        next_value = _increment_voucher_number_text(current)
        if not next_value:
            return f"{text}-{idx + 2}"
        current = next_value
    return current


def _extract_voucher_records(response_text: str) -> list:
    records = []

    try:
        root = ET.fromstring(response_text)
    except ET.ParseError:
        return records

    for voucher in root.findall(".//VOUCHER"):
        voucher_type_name = str(
            voucher.findtext("VOUCHERTYPENAME")
            or voucher.attrib.get("VCHTYPE")
            or ""
        ).strip()
        voucher_number = _normalize_voucher_number_text(voucher.findtext("VOUCHERNUMBER"))
        voucher_date = str(voucher.findtext("DATE") or "").strip()
        master_id = str(voucher.findtext("MASTERID") or "").strip()
        if not voucher_type_name and not voucher_number:
            continue
        records.append(
            {
                "voucher_type": voucher_type_name,
                "voucher_number": voucher_number,
                "date": voucher_date,
                "master_id": master_id,
            }
        )

    return records


def _extract_voucher_numbers(response_text: str) -> list:
    found = []
    seen = set()

    def add_value(raw):
        num = _extract_numeric_voucher_no(raw)
        if num is None or num in seen:
            return
        seen.add(num)
        found.append(num)

    try:
        root = ET.fromstring(response_text)
        for node in root.iter():
            tag = str(node.tag or "").upper()
            if "VOUCHERNUM" in tag:
                add_value(node.text)
            for key, raw_val in node.attrib.items():
                if "VOUCHERNUM" in str(key or "").upper():
                    add_value(raw_val)
    except ET.ParseError:
        pass

    for match in re.findall(r"<VOUCHERNUMBER>(.*?)</VOUCHERNUMBER>", response_text, flags=re.IGNORECASE | re.DOTALL):
        add_value(match)

    return found


def _fetch_next_voucher_number(tally_url: str, company_name: str = "", voucher_type: str = "Sales", timeout: float = 15.0) -> dict:
    selected_company = _normalize_company_name(company_name)
    voucher_type_text = str(voucher_type or "Sales").strip() or "Sales"
    escaped_voucher_type = xml_escape(voucher_type_text)

    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    static_vars += f"<SVVOUCHERTYPENAME>{escaped_voucher_type}</SVVOUCHERTYPENAME>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>Voucher Number Collection</ID></HEADER><BODY><DESC>"
        f"{static_vars}"
        "<TDL><TDLMESSAGE><COLLECTION NAME='Voucher Number Collection'>"
        "<TYPE>Voucher</TYPE><FETCH>VoucherNumber,VoucherTypeName,Date,MasterID</FETCH>"
        "<FILTERS>ExactVoucherType</FILTERS>"
        "</COLLECTION>"
        f"<SYSTEM TYPE='Formulae' NAME='ExactVoucherType'>$VoucherTypeName = &quot;{escaped_voucher_type}&quot;</SYSTEM>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    try:
        response_text = _post_tally_xml(tally_url, request_xml, timeout=timeout)
    except Exception as exc:
        return {
            "success": False,
            "last_number": "",
            "next_number": "",
            "existing_numbers": set(),
            "error": str(exc) or "Could not fetch voucher number from Tally.",
        }

    records = [
        record
        for record in _extract_voucher_records(response_text)
        if _ledger_name_key(record.get("voucher_type", "")) == _ledger_name_key(voucher_type_text)
    ]
    existing_numbers = {
        _normalize_voucher_number_text(record.get("voucher_number", ""))
        for record in records
        if _normalize_voucher_number_text(record.get("voucher_number", ""))
    }

    if records:
        latest_record = max(
            records,
            key=lambda record: (
                _safe_int(record.get("date", 0), 0),
                _safe_int(record.get("master_id", 0), 0),
            ),
        )
        last_number = _normalize_voucher_number_text(latest_record.get("voucher_number", ""))
        next_number = _increment_voucher_number_text(last_number)
        if not next_number:
            numeric_numbers = [
                _extract_numeric_voucher_no(record.get("voucher_number", ""))
                for record in records
            ]
            numeric_numbers = [num for num in numeric_numbers if num is not None]
            if numeric_numbers:
                next_number = str(max(numeric_numbers) + 1)
        if not next_number:
            next_number = "1"
        return {
            "success": True,
            "last_number": last_number,
            "next_number": next_number,
            "existing_numbers": existing_numbers,
            "error": "",
        }

    if response_text.strip():
        return {
            "success": True,
            "last_number": "",
            "next_number": "1",
            "existing_numbers": existing_numbers,
            "error": "",
        }

    return {
        "success": False,
        "last_number": "",
        "next_number": "",
        "existing_numbers": set(),
        "error": "Could not fetch voucher number from Tally.",
    }


# ═══════════════════════════════════════════════════════════════════════════
#  GENERATE SALES XML  –  ACCOUNTING MODE  (mirrors original VBA logic)
# ═══════════════════════════════════════════════════════════════════════════

def generate_accounting_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
    start_voucher_number=None,
    company_gst_registrations: list = None,
    voucher_type: str = "Sales",
    round_off_ledger: str = "",
) -> str:
    """rows = list of dicts with keys matching Excel columns."""
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    resolved_mode = str(date_mode or ("current" if use_today_date else "excel")).strip().lower()
    if resolved_mode not in {"current", "excel", "custom"}:
        resolved_mode = "current" if use_today_date else "excel"
    resolved_custom_date = _normalize_manual_date_to_tally(custom_tally_date) if resolved_mode == "custom" else ""

    rows = _consolidate_accounting_rows(rows, resolved_mode, resolved_custom_date)

    for idx, r in enumerate(rows):
        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt       = tally_date(source_date)
        excel_vno = _row_voucher_number(r, "")
        if excel_vno:
            vno_raw = excel_vno
        elif start_voucher_number is not None:
            vno_raw = _voucher_number_with_offset(start_voucher_number, idx) or str(start_voucher_number)
        else:
            vno_raw = ""
        vno      = xml_escape(vno_raw)
        invoice_ref_raw = _row_invoice_reference(r, vno_raw)
        invoice_ref = xml_escape(invoice_ref_raw)
        party_raw = _ledger_or_suspense(_resolve_party_ledger(r, is_purchase_mode=True))
        party_context = _collect_party_context(r, party_raw)
        party = xml_escape(party_raw)
        narr     = xml_escape(_row_text(r, "Narration"))

        # Build one leg per row (first row + any consolidated duplicates).
        # Each leg contributes its own sales ledger entry; GST is merged by ledger name.
        extra_legs = r.get("acct_legs") or []
        all_legs = [r] + extra_legs

        sales_legs = []   # list of {sales_raw, taxable}
        gst_cgst   = {}   # {led_name: total_amount}
        gst_sgst   = {}
        gst_igst   = {}

        for leg_r in all_legs:
            leg_taxable        = _row_float(leg_r, "TaxableValue", 0.0)
            leg_cgst_r         = _row_float(leg_r, "CGSTRate", 0.0)
            leg_cgst_explicit  = _row_float(leg_r, "CGST Amount", 0.0)
            leg_sgst_r         = _row_float(leg_r, "SGSTRate", 0.0)
            leg_sgst_explicit  = _row_float(leg_r, "SGST Amount", 0.0)
            leg_igst_r         = _row_float(leg_r, "IGSTRate", 0.0)
            leg_igst_explicit  = _row_float(leg_r, "IGST Amount", 0.0)

            leg_cgst = round(leg_taxable * leg_cgst_r / 100, 2) if leg_cgst_r > 0 else leg_cgst_explicit
            leg_sgst = round(leg_taxable * leg_sgst_r / 100, 2) if leg_sgst_r > 0 else leg_sgst_explicit
            leg_igst = round(leg_taxable * leg_igst_r / 100, 2) if leg_igst_r > 0 else leg_igst_explicit

            leg_sales_raw = _ledger_or_suspense(_row_text(leg_r, "SalesLedger"))
            leg_cgst_name = _pick_tax_ledger_name(leg_r, ["CGSTLedger", "CGST Ledger", "CentralTaxLedger", "Central Tax Ledger", "Central Tax"], leg_cgst_r, "CGST", leg_cgst_explicit)
            leg_sgst_name = _pick_tax_ledger_name(leg_r, ["SGSTLedger", "SGST Ledger", "StateTaxLedger", "State Tax Ledger", "State Tax", "UTGSTLedger", "UTGST Ledger"], leg_sgst_r, "SGST", leg_sgst_explicit)
            leg_igst_name = _pick_tax_ledger_name(leg_r, ["IGSTLedger", "IGST Ledger", "IntegratedTaxLedger", "Integrated Tax Ledger", "Integrated Tax"], leg_igst_r, "IGST", leg_igst_explicit)

            sales_legs.append({"sales_raw": leg_sales_raw, "taxable": leg_taxable})
            if leg_cgst:
                gst_cgst[leg_cgst_name] = gst_cgst.get(leg_cgst_name, 0.0) + leg_cgst
            if leg_sgst:
                gst_sgst[leg_sgst_name] = gst_sgst.get(leg_sgst_name, 0.0) + leg_sgst
            if leg_igst:
                gst_igst[leg_igst_name] = gst_igst.get(leg_igst_name, 0.0) + leg_igst

        total_taxable = sum(l["taxable"] for l in sales_legs)
        total_gst     = sum(gst_cgst.values()) + sum(gst_sgst.values()) + sum(gst_igst.values())
        grand_total   = total_taxable + total_gst
        _ro_amt   = round(round(grand_total, 0) - grand_total, 2) if round_off_ledger else 0.0
        _ro_total = round(grand_total + _ro_amt, 2)

        _row_vtype = str(_row_get(r, "VoucherType", "") or "").strip()
        effective_vch_type = _row_vtype or str(voucher_type or "Sales").strip() or "Sales"
        _vch_type_esc = xml_escape(effective_vch_type)
        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <VOUCHER VCHTYPE="{_vch_type_esc}" ACTION="Create" OBJVIEW="Invoice Voucher View">')
        a(f'     <DATE>{dt}</DATE>')
        a(f'     <VOUCHERTYPENAME>{_vch_type_esc}</VOUCHERTYPENAME>')
        a(f'     <VOUCHERNUMBER>{vno}</VOUCHERNUMBER>')
        a(f'     <PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>')
        _append_invoice_party_context_xml(
            a, party_context, include_basic_buyer=True,
        )
        _append_company_gst_context_xml(a, party_context, company_gst_registrations)
        a(f'     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>')
        a('     <NUMBERINGSTYLE>Manual</NUMBERINGSTYLE>')
        a('     <ISINVOICE>Yes</ISINVOICE>')
        a('     <PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW>')
        a('     <VCHENTRYMODE>Accounting Invoice</VCHENTRYMODE>')
        # ISGSTOVERRIDDEN=No lets Tally derive transaction type from GSTREGISTRATIONTYPE;
        # setting Yes was causing the "Accept?" prompt in GSTR-1 uncertain screen.
        a('     <ISGSTOVERRIDDEN>No</ISGSTOVERRIDDEN>')
        # VCHSTATUSISREACCEPHSNSIXONEDONE=Yes is the acceptance flag Tally sets internally
        # when a user "accepts" party details — without it entries stay in Uncertain.
        a('     <VCHSTATUSISREACCEPHSNSIXONEDONE>Yes</VCHSTATUSISREACCEPHSNSIXONEDONE>')
        a('     <VCHGSTSTATUSISUNCERTAIN>No</VCHGSTSTATUSISUNCERTAIN>')
        a('     <VCHGSTSTATUSISINCLUDED>Yes</VCHGSTSTATUSISINCLUDED>')
        a('     <VCHGSTSTATUSISAPPLICABLE>Yes</VCHGSTSTATUSISAPPLICABLE>')
        _gst_txn_type = _gst_transaction_type(
            party_context.get("registration_type", ""),
            party_context.get("gstin", ""),
        )
        a(f'     <GSTTRANSACTIONTYPE>{xml_escape(_gst_txn_type)}</GSTTRANSACTIONTYPE>')
        if invoice_ref:
            a(f'     <REFERENCE>{invoice_ref}</REFERENCE>')
        if narr:
            a(f'     <NARRATION>{narr}</NARRATION>')

        _party_amt = f"-{fmt_amt(_ro_total)}"

        # Party – Debit
        a('     <LEDGERENTRIES.LIST>')
        a(f'      <LEDGERNAME>{party}</LEDGERNAME>')
        a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
        a(f'      <AMOUNT>{_party_amt}</AMOUNT>')
        a('      <BILLALLOCATIONS.LIST>')
        a(f'       <NAME>{invoice_ref or vno}</NAME>')
        a('       <BILLTYPE>New Ref</BILLTYPE>')
        a(f'       <AMOUNT>{_party_amt}</AMOUNT>')
        a('      </BILLALLOCATIONS.LIST>')
        a('     </LEDGERENTRIES.LIST>')

        # Sales – one Credit entry per leg (supports multiple income ledgers per invoice)
        for leg in sales_legs:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(leg["sales_raw"])}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(leg["taxable"])}</AMOUNT>')
            a('     </LEDGERENTRIES.LIST>')

        # CGST – aggregated by ledger name
        for led_name, amt in gst_cgst.items():
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(led_name)}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "CGST")
            a('     </LEDGERENTRIES.LIST>')
        # SGST
        for led_name, amt in gst_sgst.items():
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(led_name)}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "SGST")
            a('     </LEDGERENTRIES.LIST>')
        # IGST
        for led_name, amt in gst_igst.items():
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(led_name)}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "IGST")
            a('     </LEDGERENTRIES.LIST>')

        _append_round_off_entry(a, round_off_ledger, _ro_amt, is_purchase_mode=False)
        a('    </VOUCHER>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
#  GENERATE SALES XML  –  ITEM / INVENTORY MODE
# ═══════════════════════════════════════════════════════════════════════════

def generate_item_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
    start_voucher_number=None,
    fallback_sales_ledger: str = SUSPENSE_LEDGER,
    company_gst_registrations: list = None,
    legacy_invoice_context: bool = False,
    voucher_type: str = "Sales",
    round_off_ledger: str = "",
) -> str:
    """
    Item-mode sales voucher. Each row needs additional columns:
      ItemName, Quantity, Rate, Per (unit), GodownName (optional)
    Uses ALLINVENTORYENTRIES.LIST for stock items +
         LEDGERENTRIES.LIST for accounting legs (party, tax ledgers).
    """
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    resolved_mode = str(date_mode or ("current" if use_today_date else "excel")).strip().lower()
    if resolved_mode not in {"current", "excel", "custom"}:
        resolved_mode = "current" if use_today_date else "excel"
    resolved_custom_date = _normalize_manual_date_to_tally(custom_tally_date) if resolved_mode == "custom" else ""

    def _name_key(value: str) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip().lower()

    rows = _consolidate_item_rows(rows, resolved_mode, resolved_custom_date)

    for idx, r in enumerate(rows):
        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt       = tally_date(source_date)
        excel_vno = _row_voucher_number(r, "")
        if excel_vno:
            vno_raw = excel_vno
        elif start_voucher_number is not None:
            vno_raw = _voucher_number_with_offset(start_voucher_number, idx) or str(start_voucher_number)
        else:
            vno_raw = ""
        vno      = xml_escape(vno_raw)
        invoice_ref_raw = _row_invoice_reference(r, vno_raw)
        invoice_ref = xml_escape(invoice_ref_raw)
        party_raw = _ledger_or_suspense(_row_text(r, "PartyLedger"))
        party_context = _collect_party_context(r, party_raw)
        party = xml_escape(party_raw)
        taxable  = _row_float(r, "TaxableValue", 0.0)

        # ── Items: multi-item list (from popup/consolidation) or single-item from row fields ──
        _raw_items = r.get("items") or []
        if _raw_items:
            inv_items = []
            for _it in _raw_items:
                _it_name_raw = str(_it.get("ItemName") or "")
                if not _it_name_raw:
                    raise ValueError(f"Sales item row {idx + 1}: item name missing in items list.")
                _it_qty = float(_it.get("Quantity") or 0)
                if _it_qty <= 0:
                    raise ValueError(f"Sales item row {idx + 1}: quantity must be > 0.")
                _it_rate = float(_it.get("Rate") or 0)
                _it_per  = xml_escape(_normalize_stock_unit_name(str(_it.get("Per") or "Nos")) or "Nos")
                _it_godown = xml_escape(str(_it.get("GodownName") or "Main Location"))
                
                _s_led_raw = str(_it.get("SalesLedger") or fallback_sales_ledger)
                _s_led_raw = _ledger_or_suspense(_s_led_raw, fallback_sales_ledger)
                _it_sled = xml_escape(_s_led_raw)
                
                _it_amt  = round(_it_qty * _it_rate, 2) if _it_rate > 0 else 0.0
                inv_items.append({
                    "item_name": xml_escape(_it_name_raw),
                    "qty":       _it_qty,
                    "rate":      _it_rate if _it_rate > 0 else (_it_amt / _it_qty if _it_qty > 0 else 0.0),
                    "per_unit":  _it_per,
                    "godown":    _it_godown,
                    "sled":      _it_sled,
                    "amt":       _it_amt,
                })
            item_total = sum(it["amt"] for it in inv_items)
        else:
            item_name_raw = (
                _row_text(r, "ItemName")
                or _row_text(r, "Item")
                or _row_text(r, "StockItem")
                or _row_text(r, "ProductName")
                or _row_text(r, "SalesLedger")
            )
            if not item_name_raw:
                raise ValueError(f"Item row {idx + 1}: item name is missing (ItemName/Item/SalesLedger).")
            item_name = xml_escape(item_name_raw)
            item_name_key = _name_key(item_name_raw)

            qty = _row_float(r, "Quantity", 0.0)
            if qty <= 0:
                qty = _row_float(r, "Qty", 0.0)
            if qty <= 0:
                qty = _row_float(r, "Unit", 0.0)
            if qty <= 0:
                raise ValueError(f"Item row {idx + 1}: quantity is missing/zero (Quantity/Qty/Unit).")

            rate = _row_float(r, "Rate", 0.0)
            if rate <= 0 and taxable > 0 and qty > 0:
                rate = taxable / qty
            per_unit_raw = (
                _row_text(r, "Per", "")
                or _row_text(r, "UOM", "")
                or _row_text(r, "Unit", "")
                or "Nos"
            )
            per_unit = xml_escape(_normalize_stock_unit_name(per_unit_raw) or "Nos")
            godown   = xml_escape(_row_text(r, "GodownName", "Main Location") or "Main Location")

            explicit_sales_ledger = (
                _row_text(r, "SalesAccount")
                or _row_text(r, "Sales Ledger")
                or _row_text(r, "IncomeLedger")
            )
            default_sales_ledger = _ledger_or_suspense(fallback_sales_ledger)
            if _name_key(default_sales_ledger) == item_name_key:
                for candidate in ("Sales Account", "Sales", "Sales A/c", "Sales Ledger"):
                    if _name_key(candidate) != item_name_key:
                        default_sales_ledger = candidate
                        break

            sales_ledger_raw = explicit_sales_ledger or _row_text(r, "SalesLedger") or default_sales_ledger
            if _name_key(sales_ledger_raw) == item_name_key:
                sales_ledger_raw = default_sales_ledger
            sales_ledger_raw = _ledger_or_suspense(sales_ledger_raw, default_sales_ledger)
            if _name_key(sales_ledger_raw) == item_name_key:
                raise ValueError(
                    f"Item row {idx + 1}: sales ledger cannot be same as item '{item_name_raw}'. "
                    "Provide SalesAccount/IncomeLedger in Excel or use a valid fallback sales ledger."
                )
            sales = xml_escape(sales_ledger_raw)
            item_amt  = round(qty * rate, 2) if qty and rate else taxable
            inv_items = [{
                "item_name": item_name,
                "qty":       qty,
                "rate":      rate,
                "per_unit":  per_unit,
                "godown":    godown,
                "sled":      sales,
                "amt":       item_amt,
            }]
            item_total = item_amt

        cgst_r   = _row_float(r, "CGSTRate", 0.0)
        cgst_amt_explicit = _row_float(r, "CGST Amount", 0.0)
        cgst_led = xml_escape(_pick_tax_ledger_name(
            r,
            ["CGSTLedger", "CGST Ledger", "CentralTaxLedger", "Central Tax Ledger", "Central Tax"],
            cgst_r,
            "CGST",
            cgst_amt_explicit,
        ))
        sgst_r   = _row_float(r, "SGSTRate", 0.0)
        sgst_amt_explicit = _row_float(r, "SGST Amount", 0.0)
        sgst_led = xml_escape(_pick_tax_ledger_name(
            r,
            ["SGSTLedger", "SGST Ledger", "StateTaxLedger", "State Tax Ledger", "State Tax", "UTGSTLedger", "UTGST Ledger"],
            sgst_r,
            "SGST",
            sgst_amt_explicit,
        ))
        igst_r   = _row_float(r, "IGSTRate", 0.0)
        igst_amt_explicit = _row_float(r, "IGST Amount", 0.0)
        igst_led = xml_escape(_pick_tax_ledger_name(
            r,
            ["IGSTLedger", "IGST Ledger", "IntegratedTaxLedger", "Integrated Tax Ledger", "Integrated Tax"],
            igst_r,
            "IGST",
            igst_amt_explicit,
        ))
        narr     = xml_escape(_row_text(r, "Narration"))
        hsn_code = xml_escape(_row_text(r, "HSNCode"))

        taxable_for_tax = taxable if taxable > 0 else item_total
        cgst_amt = round(taxable_for_tax * cgst_r / 100, 2) if cgst_r > 0 else cgst_amt_explicit
        sgst_amt = round(taxable_for_tax * sgst_r / 100, 2) if sgst_r > 0 else sgst_amt_explicit
        igst_amt = round(taxable_for_tax * igst_r / 100, 2) if igst_r > 0 else igst_amt_explicit
        total    = item_total + cgst_amt + sgst_amt + igst_amt
        _ro_amt_i   = round(round(total, 0) - total, 2) if round_off_ledger else 0.0
        _ro_total_i = round(total + _ro_amt_i, 2)

        _row_vtype_i = str(_row_get(r, "VoucherType", "") or "").strip()
        _vch_type_esc = xml_escape(_row_vtype_i or str(voucher_type or "Sales").strip() or "Sales")
        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <VOUCHER VCHTYPE="{_vch_type_esc}" ACTION="Create" OBJVIEW="Invoice Voucher View">')
        a(f'     <DATE>{dt}</DATE>')
        a(f'     <VOUCHERTYPENAME>{_vch_type_esc}</VOUCHERTYPENAME>')
        a(f'     <VOUCHERNUMBER>{vno}</VOUCHERNUMBER>')
        a(f'     <PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>')
        if legacy_invoice_context:
            party_gstin = xml_escape(party_context.get("gstin", ""))
            place_of_supply = xml_escape(party_context.get("place_of_supply", ""))
            if party_gstin:
                a(f'     <PARTYGSTIN>{party_gstin}</PARTYGSTIN>')
            if place_of_supply:
                a(f'     <PLACEOFSUPPLY>{place_of_supply}</PLACEOFSUPPLY>')
        else:
            _append_invoice_party_context_xml(a, party_context, include_basic_buyer=True)
            _append_company_gst_context_xml(a, party_context, company_gst_registrations)
        a(f'     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>')
        a('     <NUMBERINGSTYLE>Manual</NUMBERINGSTYLE>')
        a('     <ISINVOICE>Yes</ISINVOICE>')
        a('     <PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW>')
        a('     <VCHENTRYMODE>Item Invoice</VCHENTRYMODE>')
        if not legacy_invoice_context:
            a('     <ISGSTOVERRIDDEN>No</ISGSTOVERRIDDEN>')
            a('     <VCHSTATUSISREACCEPHSNSIXONEDONE>Yes</VCHSTATUSISREACCEPHSNSIXONEDONE>')
            a('     <VCHGSTSTATUSISUNCERTAIN>No</VCHGSTSTATUSISUNCERTAIN>')
            a('     <VCHGSTSTATUSISINCLUDED>Yes</VCHGSTSTATUSISINCLUDED>')
            a('     <VCHGSTSTATUSISAPPLICABLE>Yes</VCHGSTSTATUSISAPPLICABLE>')
            _gst_txn_type = _gst_transaction_type(
                party_context.get("registration_type", ""),
                party_context.get("gstin", ""),
            )
            a(f'     <GSTTRANSACTIONTYPE>{xml_escape(_gst_txn_type)}</GSTTRANSACTIONTYPE>')
        if invoice_ref:
            a(f'     <REFERENCE>{invoice_ref}</REFERENCE>')
        if narr:
            a(f'     <NARRATION>{narr}</NARRATION>')

        # ── Inventory entry ──
        for _it in inv_items:
            a('     <ALLINVENTORYENTRIES.LIST>')
            a(f'      <STOCKITEMNAME>{_it["item_name"]}</STOCKITEMNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <RATE>{fmt_amt(_it["rate"])}/{_it["per_unit"]}</RATE>')
            a(f'      <AMOUNT>{fmt_amt(_it["amt"])}</AMOUNT>')
            a(f'      <ACTUALQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</ACTUALQTY>')
            a(f'      <BILLEDQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</BILLEDQTY>')
            a('      <BATCHALLOCATIONS.LIST>')
            a(f'       <GODOWNNAME>{_it["godown"]}</GODOWNNAME>')
            a(f'       <AMOUNT>{fmt_amt(_it["amt"])}</AMOUNT>')
            a(f'       <ACTUALQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</ACTUALQTY>')
            a(f'       <BILLEDQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</BILLEDQTY>')
            a('      </BATCHALLOCATIONS.LIST>')
            a('      <ACCOUNTINGALLOCATIONS.LIST>')
            a(f'       <LEDGERNAME>{_it["sled"]}</LEDGERNAME>')
            a('       <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'       <AMOUNT>{fmt_amt(_it["amt"])}</AMOUNT>')
            a('      </ACCOUNTINGALLOCATIONS.LIST>')
            a('     </ALLINVENTORYENTRIES.LIST>')

        # ── Ledger entries (Party DR) with bill allocation ──
        a('     <LEDGERENTRIES.LIST>')
        a(f'      <LEDGERNAME>{party}</LEDGERNAME>')
        a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
        a(f'      <AMOUNT>-{fmt_amt(_ro_total_i)}</AMOUNT>')
        a('      <BILLALLOCATIONS.LIST>')
        a(f'       <NAME>{invoice_ref or vno}</NAME>')
        a('       <BILLTYPE>New Ref</BILLTYPE>')
        a(f'       <AMOUNT>-{fmt_amt(_ro_total_i)}</AMOUNT>')
        a('      </BILLALLOCATIONS.LIST>')
        a('     </LEDGERENTRIES.LIST>')

        # ── Tax ledgers ──
        if cgst_amt:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{cgst_led}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(cgst_amt)}</AMOUNT>')
            a('     </LEDGERENTRIES.LIST>')
        if sgst_amt:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{sgst_led}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(sgst_amt)}</AMOUNT>')
            a('     </LEDGERENTRIES.LIST>')
        if igst_amt:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{igst_led}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(igst_amt)}</AMOUNT>')
            a('     </LEDGERENTRIES.LIST>')

        _append_round_off_entry(a, round_off_ledger, _ro_amt_i, is_purchase_mode=False)
        a('    </VOUCHER>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


def generate_purchase_accounting_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
    start_voucher_number=None,
    company_gst_registrations: list = None,
    voucher_type: str = "Purchase",
    round_off_ledger: str = "",
) -> str:
    """Purchase accounting invoice XML (mirror of sales with debit/credit reversed)."""
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    resolved_mode = str(date_mode or ("current" if use_today_date else "excel")).strip().lower()
    if resolved_mode not in {"current", "excel", "custom"}:
        resolved_mode = "current" if use_today_date else "excel"
    resolved_custom_date = _normalize_manual_date_to_tally(custom_tally_date) if resolved_mode == "custom" else ""
    purchase_company_state = _resolve_company_gst_state(company_gst_registrations or [], preferred_state="")

    rows = _consolidate_accounting_rows(rows, resolved_mode, resolved_custom_date)

    for idx, r in enumerate(rows):
        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)
        excel_vno = _row_voucher_number(r, "")
        if excel_vno:
            vno_raw = excel_vno
        elif start_voucher_number is not None:
            vno_raw = _voucher_number_with_offset(start_voucher_number, idx) or str(start_voucher_number)
        else:
            vno_raw = ""
        vno = xml_escape(vno_raw)
        supplier_invoice_raw = _row_invoice_reference(r, vno_raw)
        supplier_invoice = xml_escape(supplier_invoice_raw)

        party_raw = _ledger_or_suspense(_row_text(r, "PartyLedger"))
        party_context = _collect_party_context(
            r,
            party_raw,
            allow_place_of_supply_column=False,
        )
        party = xml_escape(party_raw)
        narr = xml_escape(_row_text(r, "Narration"))

        # Build one leg per row (first row + any consolidated duplicates).
        # Each leg contributes its own purchase ledger entry; GST/TDS merged by ledger name.
        extra_legs = r.get("acct_legs") or []
        all_legs = [r] + extra_legs

        purchase_legs = []   # list of {purchase_raw, taxable}
        gst_cgst      = {}   # {led_name: total_amount}
        gst_sgst      = {}
        gst_igst      = {}
        tds_by_led    = {}   # {led_name: total_amount}

        for leg_r in all_legs:
            leg_taxable       = _row_float(leg_r, "TaxableValue", 0.0)
            leg_cgst_r        = _row_float(leg_r, "CGSTRate", 0.0)
            leg_cgst_explicit = _row_float(leg_r, "CGST Amount", 0.0)
            leg_sgst_r        = _row_float(leg_r, "SGSTRate", 0.0)
            leg_sgst_explicit = _row_float(leg_r, "SGST Amount", 0.0)
            leg_igst_r        = _row_float(leg_r, "IGSTRate", 0.0)
            leg_igst_explicit = _row_float(leg_r, "IGST Amount", 0.0)

            leg_cgst = round(leg_taxable * leg_cgst_r / 100, 2) if leg_cgst_r > 0 else leg_cgst_explicit
            leg_sgst = round(leg_taxable * leg_sgst_r / 100, 2) if leg_sgst_r > 0 else leg_sgst_explicit
            leg_igst = round(leg_taxable * leg_igst_r / 100, 2) if leg_igst_r > 0 else leg_igst_explicit

            leg_purchase_raw = (
                _row_text(leg_r, "PurchaseLedger")
                or _row_text(leg_r, "PurchaseAccount")
                or _row_text(leg_r, "Purchase Ledger")
                or _row_text(leg_r, "ExpenseLedger")
                or _row_text(leg_r, "SalesLedger")
            )
            leg_purchase_raw = _ledger_or_suspense(leg_purchase_raw)

            leg_cgst_name = _pick_tax_ledger_name(leg_r, ["CGSTLedger", "CGST Ledger", "CentralTaxLedger", "Central Tax Ledger", "Central Tax"], leg_cgst_r, "CGST", leg_cgst_explicit)
            leg_sgst_name = _pick_tax_ledger_name(leg_r, ["SGSTLedger", "SGST Ledger", "StateTaxLedger", "State Tax Ledger", "State Tax", "UTGSTLedger", "UTGST Ledger"], leg_sgst_r, "SGST", leg_sgst_explicit)
            leg_igst_name = _pick_tax_ledger_name(leg_r, ["IGSTLedger", "IGST Ledger", "IntegratedTaxLedger", "Integrated Tax Ledger", "Integrated Tax"], leg_igst_r, "IGST", leg_igst_explicit)

            # TDS per leg
            leg_tds_ledger_raw = _row_text(leg_r, "TDSLedger") or _row_text(leg_r, "TDS Ledger") or _row_text(leg_r, "Tds Ledger")
            leg_tds_rate       = _row_float(leg_r, "TDSRate", 0.0) or _row_float(leg_r, "TDS Rate", 0.0)
            leg_tds_raw        = _row_float(leg_r, "TDSAmount", 0.0) or _row_float(leg_r, "TDS Amount", 0.0)
            if leg_tds_ledger_raw and leg_tds_raw <= 0 and leg_tds_rate > 0:
                leg_tds = round(leg_taxable * leg_tds_rate / 100, 2)
            else:
                leg_tds = abs(leg_tds_raw)

            purchase_legs.append({"purchase_raw": leg_purchase_raw, "taxable": leg_taxable})
            if leg_cgst:
                gst_cgst[leg_cgst_name] = gst_cgst.get(leg_cgst_name, 0.0) + leg_cgst
            if leg_sgst:
                gst_sgst[leg_sgst_name] = gst_sgst.get(leg_sgst_name, 0.0) + leg_sgst
            if leg_igst:
                gst_igst[leg_igst_name] = gst_igst.get(leg_igst_name, 0.0) + leg_igst
            if leg_tds_ledger_raw and leg_tds > 0:
                tds_by_led[leg_tds_ledger_raw] = tds_by_led.get(leg_tds_ledger_raw, 0.0) + leg_tds

        total_taxable = sum(l["taxable"] for l in purchase_legs)
        total_gst     = sum(gst_cgst.values()) + sum(gst_sgst.values()) + sum(gst_igst.values())
        total_tds     = sum(tds_by_led.values())
        grand_total   = total_taxable + total_gst
        _ro_amt_pa    = round(round(grand_total, 0) - grand_total, 2) if round_off_ledger else 0.0
        _ro_grand_pa  = round(grand_total + _ro_amt_pa, 2)
        # Vendor is credited the NET amount (invoice total minus TDS deducted at source).
        # TDS Payable is a separate Credit entry.  Using the net party amount ensures
        # Tally's Invoice Voucher View does not back-compute GST from an inflated figure.
        party_total   = _ro_grand_pa - total_tds if total_tds > 0 else _ro_grand_pa

        _vch_type_esc = xml_escape(str(voucher_type or "Purchase").strip() or "Purchase")

        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <VOUCHER VCHTYPE="{_vch_type_esc}" ACTION="Create" OBJVIEW="Invoice Voucher View">')
        a(f'     <DATE>{dt}</DATE>')
        a(f'     <VOUCHERTYPENAME>{_vch_type_esc}</VOUCHERTYPENAME>')
        a(f'     <VOUCHERNUMBER>{vno}</VOUCHERNUMBER>')
        a(f'     <PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>')
        _append_invoice_party_context_xml(
            a,
            party_context,
            include_basic_buyer=False,
            include_place_of_supply=bool(purchase_company_state),
            place_of_supply_override=purchase_company_state,
        )
        _append_company_gst_context_xml(
            a,
            party_context,
            company_gst_registrations,
            prefer_party_state=False,
        )
        a(f'     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>')
        a('     <ISINVOICE>Yes</ISINVOICE>')
        a('     <PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW>')
        a('     <VCHENTRYMODE>Accounting Invoice</VCHENTRYMODE>')
        a('     <ISGSTOVERRIDDEN>Yes</ISGSTOVERRIDDEN>')
        # GSTTRANSACTIONTYPE required to avoid 'Uncertain' in GSTR-3B
        _gst_txn_type = _gst_transaction_type(
            party_context.get("registration_type", ""),
            party_context.get("gstin", ""),
        )
        a(f'     <GSTTRANSACTIONTYPE>{xml_escape(_gst_txn_type)}</GSTTRANSACTIONTYPE>')
        if supplier_invoice:
            a(f'     <REFERENCE>{supplier_invoice}</REFERENCE>')
        if narr:
            a(f'     <NARRATION>{narr}</NARRATION>')

        # Party - Credit (net of TDS if any)
        a('     <LEDGERENTRIES.LIST>')
        a(f'      <LEDGERNAME>{party}</LEDGERNAME>')
        a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
        a(f'      <AMOUNT>{fmt_amt(party_total)}</AMOUNT>')
        a('      <BILLALLOCATIONS.LIST>')
        a(f'       <NAME>{supplier_invoice or vno}</NAME>')
        a('       <BILLTYPE>New Ref</BILLTYPE>')
        a(f'       <AMOUNT>{fmt_amt(party_total)}</AMOUNT>')
        a('      </BILLALLOCATIONS.LIST>')
        a('     </LEDGERENTRIES.LIST>')

        # Purchase – one Debit entry per leg (supports multiple expense ledgers per invoice)
        for leg in purchase_legs:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(leg["purchase_raw"])}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(leg["taxable"])}</AMOUNT>')
            a('     </LEDGERENTRIES.LIST>')

        # CGST – aggregated by ledger name
        for led_name, amt in gst_cgst.items():
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(led_name)}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "CGST")
            a('     </LEDGERENTRIES.LIST>')
        # SGST
        for led_name, amt in gst_sgst.items():
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(led_name)}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "SGST")
            a('     </LEDGERENTRIES.LIST>')
        # IGST
        for led_name, amt in gst_igst.items():
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(led_name)}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "IGST")
            a('     </LEDGERENTRIES.LIST>')

        # TDS Payable – ISDEEMEDPOSITIVE=Yes with a POSITIVE amount.
        # Tally's purchase voucher convention: TDS is "deemed positive" (placed in the
        # debit bucket) but the AMOUNT is positive so the ledger receives a credit
        # (liability increases).  Party is credited the net (grand_total − TDS) so the
        # voucher balances: debit_bucket = -(taxable+GST) + TDS, credit = +(grand_total-TDS).
        for led_name, amt in tds_by_led.items():
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{xml_escape(led_name)}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(amt)}</AMOUNT>')
            a('     </LEDGERENTRIES.LIST>')

        _append_round_off_entry(a, round_off_ledger, _ro_amt_pa, is_purchase_mode=True)
        a('    </VOUCHER>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


def _consolidate_accounting_rows(rows: list, resolved_mode: str, resolved_custom_date: str) -> list:
    """
    Group accounting-mode rows with the same (date, invoice/ref number, GSTIN)
    into a single voucher.  Each duplicate row is appended to 'acct_legs' on the
    first row so that the XML emitter can produce multiple sales/purchase ledger
    entries within one voucher.
    """
    consolidated_rows = []
    group_map = {}
    for r in rows:
        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)

        vno_raw = _row_voucher_number(r)
        ref_raw = _row_invoice_reference(r, vno_raw)
        inv_key = str(ref_raw or vno_raw).strip().lower()

        gstin_raw = _row_text(r, "PartyGSTIN") or _row_text(r, "GSTIN")
        gstin_key = str(gstin_raw).strip().lower()

        party_raw = _ledger_or_suspense(_row_text(r, "PartyLedger"))
        key = (dt, inv_key, gstin_key)

        if key in group_map:
            existing_r = group_map[key]
            existing_party = _ledger_or_suspense(_resolve_party_ledger(existing_r, is_purchase_mode=True))
            if party_raw.strip().lower() != existing_party.strip().lower():
                raise ValueError(
                    f"Conflicting Party Ledgers for Invoice '{ref_raw or vno_raw}': "
                    f"'{existing_party}' vs '{party_raw}'"
                )
            if "acct_legs" not in existing_r:
                existing_r["acct_legs"] = []
            existing_r["acct_legs"].append(dict(r))
        else:
            new_r = dict(r)
            group_map[key] = new_r
            consolidated_rows.append(new_r)
    return consolidated_rows


def _consolidate_item_rows(rows: list, resolved_mode: str, resolved_custom_date: str) -> list:
    consolidated_rows = []
    group_map = {}
    for r in rows:
        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)
        
        vno_raw = _row_voucher_number(r)
        supplier_invoice_raw = _row_invoice_reference(r, vno_raw)
        inv_key = str(supplier_invoice_raw or vno_raw).strip().lower()
        
        gstin_raw = _row_text(r, "PartyGSTIN") or _row_text(r, "GSTIN")
        gstin_key = str(gstin_raw).strip().lower()
        
        party_raw = _ledger_or_suspense(_resolve_party_ledger(r, is_purchase_mode=True))
        
        key = (dt, inv_key, gstin_key)
        
        if key in group_map:
            existing_r = group_map[key]
            existing_party = _ledger_or_suspense(_resolve_party_ledger(existing_r, is_purchase_mode=True))
            if party_raw.strip().lower() != existing_party.strip().lower():
                raise ValueError(f"Conflicting Party Ledgers for Invoice '{supplier_invoice_raw or vno_raw}': '{existing_party}' vs '{party_raw}'")
            
            # Merge amounts
            existing_r["TaxableValue"] = _row_float(existing_r, "TaxableValue", 0.0) + _row_float(r, "TaxableValue", 0.0)
            existing_r["IGST Amount"] = _row_float(existing_r, "IGST Amount", 0.0) + _row_float(r, "IGST Amount", 0.0)
            existing_r["CGST Amount"] = _row_float(existing_r, "CGST Amount", 0.0) + _row_float(r, "CGST Amount", 0.0)
            existing_r["SGST Amount"] = _row_float(existing_r, "SGST Amount", 0.0) + _row_float(r, "SGST Amount", 0.0)
            existing_r["Cess Amount"] = _row_float(existing_r, "Cess Amount", 0.0) + _row_float(r, "Cess Amount", 0.0)
            existing_r["Invoice Value"] = _row_float(existing_r, "Invoice Value", 0.0) + _row_float(r, "Invoice Value", 0.0)
            
            if "items" not in existing_r:
                existing_r["items"] = []
                _orig_it = {
                    "ItemName": _row_text(existing_r, "ItemName") or _row_text(existing_r, "Item") or _row_text(existing_r, "StockItem") or _row_text(existing_r, "ProductName"),
                    "Quantity": _row_float(existing_r, "Quantity", 0.0) or _row_float(existing_r, "Qty", 0.0) or _row_float(existing_r, "Unit", 0.0),
                    "Rate": _row_float(existing_r, "Rate", 0.0),
                    "Per": _row_text(existing_r, "Per") or _row_text(existing_r, "UOM") or _row_text(existing_r, "Unit"),
                    "GodownName": _row_text(existing_r, "GodownName"),
                    "PurchaseLedger": _row_text(existing_r, "PurchaseLedger"),
                    "SalesLedger": _row_text_any(existing_r, ["SalesAccount", "Sales Ledger", "IncomeLedger", "SalesLedger"]),
                }
                if _orig_it["ItemName"]:
                    existing_r["items"].append(_orig_it)
            
            _new_raw_items = r.get("items") or []
            if _new_raw_items:
                existing_r["items"].extend(_new_raw_items)
            else:
                _new_it = {
                    "ItemName": _row_text(r, "ItemName") or _row_text(r, "Item") or _row_text(r, "StockItem") or _row_text(r, "ProductName"),
                    "Quantity": _row_float(r, "Quantity", 0.0) or _row_float(r, "Qty", 0.0) or _row_float(r, "Unit", 0.0),
                    "Rate": _row_float(r, "Rate", 0.0),
                    "Per": _row_text(r, "Per") or _row_text(r, "UOM") or _row_text(r, "Unit"),
                    "GodownName": _row_text(r, "GodownName"),
                    "PurchaseLedger": _row_text(r, "PurchaseLedger"),
                    "SalesLedger": _row_text_any(r, ["SalesAccount", "Sales Ledger", "IncomeLedger", "SalesLedger"]),
                }
                if _new_it["ItemName"]:
                    existing_r["items"].append(_new_it)
                    
        else:
            new_r = dict(r)
            group_map[key] = new_r
            consolidated_rows.append(new_r)
            
    return consolidated_rows


def generate_purchase_item_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
    start_voucher_number=None,
    fallback_purchase_ledger: str = SUSPENSE_LEDGER,
    company_gst_registrations: list = None,
    company_gstin: str = "",
    voucher_type: str = "Purchase",
    round_off_ledger: str = "",
) -> str:
    """Purchase item invoice XML (inventory + accounting allocations)."""
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    resolved_mode = str(date_mode or ("current" if use_today_date else "excel")).strip().lower()
    if resolved_mode not in {"current", "excel", "custom"}:
        resolved_mode = "current" if use_today_date else "excel"
    resolved_custom_date = _normalize_manual_date_to_tally(custom_tally_date) if resolved_mode == "custom" else ""
    purchase_company_state = _resolve_company_gst_state(company_gst_registrations or [], preferred_state="")

    def _name_key(value: str) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip().lower()

    rows = _consolidate_item_rows(rows, resolved_mode, resolved_custom_date)

    for idx, r in enumerate(rows):
        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)
        excel_vno = _row_voucher_number(r, "")
        if excel_vno:
            vno_raw = excel_vno
        elif start_voucher_number is not None:
            vno_raw = _voucher_number_with_offset(start_voucher_number, idx) or str(start_voucher_number)
        else:
            vno_raw = ""
        vno = xml_escape(vno_raw)
        supplier_invoice_raw = _row_invoice_reference(r, vno_raw)
        supplier_invoice = xml_escape(supplier_invoice_raw)

        party_raw = _ledger_or_suspense(_resolve_party_ledger(r, is_purchase_mode=True))
        party_context = _collect_party_context(
            r,
            party_raw,
            allow_place_of_supply_column=False,
        )
        party = xml_escape(party_raw)
        taxable = _row_float(r, "TaxableValue", 0.0)

        # ── Items: multi-item list (from popup) or single-item from row fields ──
        _raw_items = r.get("items") or []
        if _raw_items:
            inv_items = []
            for _it in _raw_items:
                _it_name_raw = str(_it.get("ItemName") or "")
                if not _it_name_raw:
                    raise ValueError(f"Purchase item row {idx + 1}: item name missing in items list.")
                _it_qty = float(_it.get("Quantity") or 0)
                if _it_qty <= 0:
                    raise ValueError(f"Purchase item row {idx + 1}: quantity must be > 0.")
                _it_rate = float(_it.get("Rate") or 0)
                _it_per  = xml_escape(_normalize_stock_unit_name(str(_it.get("Per") or "Nos")) or "Nos")
                _it_godown = xml_escape(str(_it.get("GodownName") or "Main Location"))
                _it_pled = xml_escape(_ledger_or_suspense(str(_it.get("PurchaseLedger") or fallback_purchase_ledger)))
                _it_amt  = round(_it_qty * _it_rate, 2) if _it_rate > 0 else 0.0
                inv_items.append({
                    "item_name": xml_escape(_it_name_raw),
                    "qty":       _it_qty,
                    "rate":      _it_rate if _it_rate > 0 else (_it_amt / _it_qty if _it_qty > 0 else 0.0),
                    "per_unit":  _it_per,
                    "godown":    _it_godown,
                    "pled":      _it_pled,
                    "amt":       _it_amt,
                })
            item_total = sum(it["amt"] for it in inv_items)
        else:
            # Single-item: extract from row fields (original behavior)
            item_name_raw = (
                _row_text(r, "ItemName")
                or _row_text(r, "Item")
                or _row_text(r, "StockItem")
                or _row_text(r, "ProductName")
                or _row_text(r, "PurchaseLedger")
                or _row_text(r, "SalesLedger")
            )
            if not item_name_raw:
                raise ValueError(f"Purchase item row {idx + 1}: item name is missing.")
            item_name_key = _name_key(item_name_raw)

            qty = _row_float(r, "Quantity", 0.0)
            if qty <= 0:
                qty = _row_float(r, "Qty", 0.0)
            if qty <= 0:
                qty = _row_float(r, "Unit", 0.0)
            if qty <= 0:
                raise ValueError(f"Purchase item row {idx + 1}: quantity is missing/zero.")

            rate = _row_float(r, "Rate", 0.0)
            if rate <= 0 and taxable > 0 and qty > 0:
                rate = taxable / qty
            per_unit_raw = (
                _row_text(r, "Per", "")
                or _row_text(r, "UOM", "")
                or _row_text(r, "Unit", "")
                or "Nos"
            )
            per_unit = xml_escape(_normalize_stock_unit_name(per_unit_raw) or "Nos")
            godown = xml_escape(_row_text(r, "GodownName", "Main Location") or "Main Location")

            explicit_purchase_ledger = (
                _row_text(r, "PurchaseAccount")
                or _row_text(r, "Purchase Ledger")
                or _row_text(r, "ExpenseLedger")
                or _row_text(r, "PurchaseLedger")
            )
            default_purchase_ledger = _ledger_or_suspense(fallback_purchase_ledger)
            if _name_key(default_purchase_ledger) == item_name_key:
                for candidate in ("Purchase Account", "Purchase", "Purchase A/c", "Purchase Ledger"):
                    if _name_key(candidate) != item_name_key:
                        default_purchase_ledger = candidate
                        break

            purchase_ledger_raw = (
                explicit_purchase_ledger
                or _row_text(r, "PurchaseLedger")
                or _row_text(r, "SalesLedger")
                or default_purchase_ledger
            )
            if _name_key(purchase_ledger_raw) == item_name_key:
                purchase_ledger_raw = default_purchase_ledger
            purchase_ledger_raw = _ledger_or_suspense(purchase_ledger_raw, default_purchase_ledger)
            if _name_key(purchase_ledger_raw) == item_name_key:
                raise ValueError(
                    f"Purchase item row {idx + 1}: purchase ledger cannot match item '{item_name_raw}'."
                )
            _s_amt = round(qty * rate, 2) if qty and rate else taxable
            inv_items = [{
                "item_name": xml_escape(item_name_raw),
                "qty":       qty,
                "rate":      rate,
                "per_unit":  per_unit,
                "godown":    godown,
                "pled":      xml_escape(purchase_ledger_raw),
                "amt":       _s_amt,
            }]
            item_total = _s_amt

        cgst_r = _row_float(r, "CGSTRate", 0.0)
        cgst_amt_explicit = _row_float(r, "CGST Amount", 0.0)
        cgst_led = xml_escape(_pick_tax_ledger_name(
            r,
            ["CGSTLedger", "CGST Ledger", "CentralTaxLedger", "Central Tax Ledger", "Central Tax"],
            cgst_r,
            "CGST",
            cgst_amt_explicit,
        ))
        sgst_r = _row_float(r, "SGSTRate", 0.0)
        sgst_amt_explicit = _row_float(r, "SGST Amount", 0.0)
        sgst_led = xml_escape(_pick_tax_ledger_name(
            r,
            ["SGSTLedger", "SGST Ledger", "StateTaxLedger", "State Tax Ledger", "State Tax", "UTGSTLedger", "UTGST Ledger"],
            sgst_r,
            "SGST",
            sgst_amt_explicit,
        ))
        igst_r = _row_float(r, "IGSTRate", 0.0)
        igst_amt_explicit = _row_float(r, "IGST Amount", 0.0)
        igst_led = xml_escape(_pick_tax_ledger_name(
            r,
            ["IGSTLedger", "IGST Ledger", "IntegratedTaxLedger", "Integrated Tax Ledger", "Integrated Tax"],
            igst_r,
            "IGST",
            igst_amt_explicit,
        ))
        narr = xml_escape(_row_text(r, "Narration"))

        # Taxes use GSTR-2B taxable (authoritative); item_total balances debit/credit
        taxable_for_tax = taxable if taxable > 0 else item_total
        cgst_amt = round(taxable_for_tax * cgst_r / 100, 2) if cgst_r > 0 else cgst_amt_explicit
        sgst_amt = round(taxable_for_tax * sgst_r / 100, 2) if sgst_r > 0 else sgst_amt_explicit
        igst_amt = round(taxable_for_tax * igst_r / 100, 2) if igst_r > 0 else igst_amt_explicit
        total = item_total + cgst_amt + sgst_amt + igst_amt
        _ro_amt_pi   = round(round(total, 0) - total, 2) if round_off_ledger else 0.0
        _ro_total_pi = round(total + _ro_amt_pi, 2)

        # TDS fields
        tds_ledger_raw = _row_text(r, "TDSLedger") or _row_text(r, "TDS Ledger") or _row_text(r, "Tds Ledger")
        tds_rate       = _row_float(r, "TDSRate", 0.0) or _row_float(r, "TDS Rate", 0.0)
        tds_amount_raw = _row_float(r, "TDSAmount", 0.0) or _row_float(r, "TDS Amount", 0.0)
        if tds_ledger_raw and tds_amount_raw <= 0 and tds_rate > 0:
            tds_amount = round(taxable_for_tax * tds_rate / 100, 2)
        else:
            tds_amount = abs(tds_amount_raw)
        tds_led = xml_escape(tds_ledger_raw)
        # Vendor is credited the net (invoice total minus TDS deducted at source).
        party_total = _ro_total_pi - tds_amount if (tds_led and tds_amount > 0) else _ro_total_pi

        _vch_type_esc = xml_escape(str(voucher_type or "Purchase").strip() or "Purchase")

        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <VOUCHER VCHTYPE="{_vch_type_esc}" ACTION="Create" OBJVIEW="Invoice Voucher View">')
        a(f'     <DATE>{dt}</DATE>')
        a(f'     <VOUCHERTYPENAME>{_vch_type_esc}</VOUCHERTYPENAME>')
        a(f'     <VOUCHERNUMBER>{vno}</VOUCHERNUMBER>')
        a(f'     <PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>')
        _append_invoice_party_context_xml(
            a,
            party_context,
            include_basic_buyer=False,
            include_place_of_supply=bool(purchase_company_state),
            place_of_supply_override=purchase_company_state,
        )
        _append_company_gst_context_xml(
            a,
            party_context,
            company_gst_registrations,
            prefer_party_state=False,
        )
        # Fallback: when Tally returned no Tax Units, emit CMPGSTIN directly from
        # company_gstin so the voucher is never left without a registration block
        # (missing block causes "Uncertain" in GSTR-3B and requires Enter twice).
        if not (company_gst_registrations or []) and company_gstin:
            _co_gstin = xml_escape(company_gstin.strip().upper())
            _co_state_name = _state_name_from_gstin(company_gstin)
            _co_reg_name = xml_escape(
                f"{_co_state_name} Registration" if _co_state_name else company_gstin.strip().upper()
            )
            a(f'     <GSTREGISTRATION TAXTYPE="GST" TAXREGISTRATION="{_co_gstin}">{_co_reg_name}</GSTREGISTRATION>')
            a(f'     <CMPGSTIN>{_co_gstin}</CMPGSTIN>')
            a('     <CMPGSTREGISTRATIONTYPE>Regular</CMPGSTREGISTRATIONTYPE>')
            if _co_state_name:
                a(f'     <CMPGSTSTATE>{xml_escape(_co_state_name)}</CMPGSTSTATE>')
        a(f'     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>')
        a('     <ISINVOICE>Yes</ISINVOICE>')
        a('     <PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW>')
        a('     <VCHENTRYMODE>Item Invoice</VCHENTRYMODE>')
        a('     <ISGSTOVERRIDDEN>Yes</ISGSTOVERRIDDEN>')
        # GSTTRANSACTIONTYPE required to avoid 'Uncertain' in GSTR-3B
        _gst_txn_type = _gst_transaction_type(
            party_context.get("registration_type", ""),
            party_context.get("gstin", ""),
        )
        a(f'     <GSTTRANSACTIONTYPE>{xml_escape(_gst_txn_type)}</GSTTRANSACTIONTYPE>')
        if supplier_invoice:
            a(f'     <REFERENCE>{supplier_invoice}</REFERENCE>')
        if narr:
            a(f'     <NARRATION>{narr}</NARRATION>')

        for _it in inv_items:
            a('     <ALLINVENTORYENTRIES.LIST>')
            a(f'      <STOCKITEMNAME>{_it["item_name"]}</STOCKITEMNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <RATE>{fmt_amt(_it["rate"])}/{_it["per_unit"]}</RATE>')
            a(f'      <AMOUNT>-{fmt_amt(_it["amt"])}</AMOUNT>')
            a(f'      <ACTUALQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</ACTUALQTY>')
            a(f'      <BILLEDQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</BILLEDQTY>')
            a('      <BATCHALLOCATIONS.LIST>')
            a(f'       <GODOWNNAME>{_it["godown"]}</GODOWNNAME>')
            a(f'       <AMOUNT>-{fmt_amt(_it["amt"])}</AMOUNT>')
            a(f'       <ACTUALQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</ACTUALQTY>')
            a(f'       <BILLEDQTY>{fmt_amt(_it["qty"])} {_it["per_unit"]}</BILLEDQTY>')
            a('      </BATCHALLOCATIONS.LIST>')
            a('      <ACCOUNTINGALLOCATIONS.LIST>')
            a(f'       <LEDGERNAME>{_it["pled"]}</LEDGERNAME>')
            a('       <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'       <AMOUNT>-{fmt_amt(_it["amt"])}</AMOUNT>')
            a('      </ACCOUNTINGALLOCATIONS.LIST>')
            a('     </ALLINVENTORYENTRIES.LIST>')

        # Party - Credit (net of TDS so vendor balance shows actual payable)
        a('     <LEDGERENTRIES.LIST>')
        a(f'      <LEDGERNAME>{party}</LEDGERNAME>')
        a('      <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>')
        a(f'      <AMOUNT>{fmt_amt(party_total)}</AMOUNT>')
        a('      <BILLALLOCATIONS.LIST>')
        a(f'       <NAME>{supplier_invoice or vno}</NAME>')
        a('       <BILLTYPE>New Ref</BILLTYPE>')
        a(f'       <AMOUNT>{fmt_amt(party_total)}</AMOUNT>')
        a('      </BILLALLOCATIONS.LIST>')
        a('     </LEDGERENTRIES.LIST>')

        # Taxes - Debit (ITC claim; ISDEEMEDPOSITIVE=Yes, negative amount = debit)
        if cgst_amt:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{cgst_led}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(cgst_amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "CGST")
            a('     </LEDGERENTRIES.LIST>')
        if sgst_amt:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{sgst_led}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(sgst_amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "SGST")
            a('     </LEDGERENTRIES.LIST>')
        if igst_amt:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{igst_led}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>-{fmt_amt(igst_amt)}</AMOUNT>')
            _append_tax_object_allocation_xml(a, "IGST")
            a('     </LEDGERENTRIES.LIST>')

        # TDS Payable – ISDEEMEDPOSITIVE=Yes, AMOUNT positive (Tally purchase convention).
        if tds_led and tds_amount > 0:
            a('     <LEDGERENTRIES.LIST>')
            a(f'      <LEDGERNAME>{tds_led}</LEDGERNAME>')
            a('      <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>')
            a(f'      <AMOUNT>{fmt_amt(tds_amount)}</AMOUNT>')
            a('     </LEDGERENTRIES.LIST>')

        _append_round_off_entry(a, round_off_ledger, _ro_amt_pi, is_purchase_mode=True)
        a('    </VOUCHER>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
#  GENERATE LEDGER MASTER XML
# ═══════════════════════════════════════════════════════════════════════════

def generate_ledger_xml(ledgers: list, company: str, alter_existing: bool = False) -> str:
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>All Masters</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    for led in ledgers:
        name_raw = str(led.get("Name", "") or "").strip()
        if not name_raw:
            continue
        parent_raw = str(led.get("Parent", "Sundry Debtors") or "Sundry Debtors").strip()
        gstin_raw = str(led.get("GSTIN", "") or "").strip().upper()
        state_raw = _normalize_state_for_ledger(
            str(led.get("StateOfSupply", "") or "").strip()
            or str(led.get("PlaceOfSupply", "") or "").strip()
            or str(led.get("State", "") or "").strip()
        )
        address1_raw = (
            str(led.get("Address1", "") or "").strip()
            or str(led.get("AddressLine1", "") or "").strip()
            or str(led.get("Address Line 1", "") or "").strip()
        )
        address2_raw = (
            str(led.get("Address2", "") or "").strip()
            or str(led.get("AddressLine2", "") or "").strip()
            or str(led.get("Address Line 2", "") or "").strip()
        )
        pincode_raw = (
            str(led.get("Pincode", "") or "").strip()
            or str(led.get("PinCode", "") or "").strip()
            or str(led.get("PIN", "") or "").strip()
            or str(led.get("PostalCode", "") or "").strip()
        )
        pan_raw = (
            str(led.get("PAN", "") or "").strip().upper()
            or _pan_from_gstin(gstin_raw)
        )
        country_raw = (
            str(led.get("Country", "") or "").strip()
            or str(led.get("CountryOfResidence", "") or "").strip()
            or "India"
        )
        mailing_name_raw = (
            str(led.get("MailingName", "") or "").strip()
            or str(led.get("PartyMailingName", "") or "").strip()
            or name_raw
        )
        if not state_raw and gstin_raw:
            state_raw = _state_name_from_gstin(gstin_raw)

        is_party = _is_party_parent(parent_raw)
        is_duties_ledger = _is_duties_parent(parent_raw)
        billwise_raw = (
            str(led.get("Billwise", "") or "").strip()
            or str(led.get("IsBillwise", "") or "").strip()
            or str(led.get("ISBILLWISEON", "") or "").strip()
        )
        if billwise_raw:
            billwise_on = billwise_raw.casefold() in {"yes", "y", "true", "1", "on"}
        else:
            billwise_on = bool(is_party)

        gst_app_raw = str(led.get("GSTApplicable", "") or "").strip()
        gst_app_value = _normalize_gst_applicable(gst_app_raw, gstin=gstin_raw)
        reg_type_raw = (
            str(led.get("GSTRegistrationType", "") or "").strip()
            or str(led.get("RegistrationType", "") or "").strip()
            or str(led.get("RegType", "") or "").strip()
        )
        reg_type = _normalize_gst_registration_type(
            reg_type_raw,
            gstin=gstin_raw,
            gst_applicable=gst_app_value,
        )

        name = xml_escape(name_raw)
        parent = xml_escape(parent_raw)
        gst_app = xml_escape(gst_app_value)
        gstin = xml_escape(gstin_raw)
        state = xml_escape(state_raw)
        address1 = xml_escape(address1_raw)
        address2 = xml_escape(address2_raw)
        pincode = xml_escape(pincode_raw)
        country = xml_escape(country_raw)
        mailing_name = xml_escape(mailing_name_raw)
        tax_type_raw = str(led.get("TypeOfTaxation", "") or "").strip()
        if tax_type_raw.casefold() in {"not applicable", "na", "n/a"}:
            tax_type_raw = ""
        if is_duties_ledger and not tax_type_raw:
            tax_type_raw = _infer_tax_type_from_ledger_name(name_raw)
        is_gst_duty_ledger = _is_protected_gst_tax_ledger(name_raw, parent_raw, tax_type_raw)
        tax_type = xml_escape(tax_type_raw)
        gst_rate = str(led.get("GSTRate", "") or "").strip()
        applicable_from = _current_fy_start()
        ledger_action = "Create Alter" if alter_existing else "Create"

        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <LEDGER NAME="{name}" RESERVEDNAME="" ACTION="{ledger_action}">')
        a(f'     <NAME>{name}</NAME>')
        a(f'     <PARENT>{parent}</PARENT>')
        a(f'     <ISBILLWISEON>{"Yes" if billwise_on else "No"}</ISBILLWISEON>')
        a('     <ISCOSTCENTRESON>No</ISCOSTCENTRESON>')
        a('     <ISINTERESTON>No</ISINTERESTON>')
        a('     <ALLOWINMOBILE>No</ALLOWINMOBILE>')
        a('     <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>')
        a('     <ASORIGINAL>Yes</ASORIGINAL>')
        a('     <AFFECTSSTOCK>No</AFFECTSSTOCK>')
        a('     <CURRENCYNAME>INR</CURRENCYNAME>')
        a(f'     <COUNTRYOFRESIDENCE>{country}</COUNTRYOFRESIDENCE>')

        if is_party and gst_app:
            a(f'     <GSTAPPLICABLE>{gst_app}</GSTAPPLICABLE>')
        if is_party and reg_type:
            a(f'     <GSTREGISTRATIONTYPE>{xml_escape(reg_type)}</GSTREGISTRATIONTYPE>')
        if is_party and gstin:
            a(f'     <PARTYGSTIN>{gstin}</PARTYGSTIN>')
        if state:
            a(f'     <PRIORSTATENAME>{state}</PRIORSTATENAME>')
            if is_party:
                a(f'     <LEDSTATENAME>{state}</LEDSTATENAME>')

        a('     <LANGUAGENAME.LIST>')
        a('      <NAME.LIST TYPE="String">')
        a(f'       <NAME>{name}</NAME>')
        a('      </NAME.LIST>')
        a('      <LANGUAGEID>1033</LANGUAGEID>')
        a('     </LANGUAGENAME.LIST>')

        if is_party and (gstin or reg_type):
            a('     <LEDGSTREGDETAILS.LIST>')
            a(f'      <APPLICABLEFROM>{applicable_from}</APPLICABLEFROM>')
            if reg_type:
                a(f'      <GSTREGISTRATIONTYPE>{xml_escape(reg_type)}</GSTREGISTRATIONTYPE>')
            if state:
                a(f'      <PLACEOFSUPPLY>{state}</PLACEOFSUPPLY>')
            if gstin:
                a(f'      <GSTIN>{gstin}</GSTIN>')
            a('      <ISOTHTERRITORYASSESSEE>No</ISOTHTERRITORYASSESSEE>')
            a('      <CONSIDERPURCHASEFOREXPORT>No</CONSIDERPURCHASEFOREXPORT>')
            a('      <ISTRANSPORTER>No</ISTRANSPORTER>')
            a('      <ISCOMMONPARTY>No</ISCOMMONPARTY>')
            a('     </LEDGSTREGDETAILS.LIST>')

        pan = xml_escape(pan_raw)
        if is_party and (state or gstin or address1 or address2 or country or pincode or pan):
            a('     <LEDMAILINGDETAILS.LIST>')
            if address1 or address2:
                a('      <ADDRESS.LIST TYPE="String">')
                if address1:
                    a(f'       <ADDRESS>{address1}</ADDRESS>')
                if address2:
                    a(f'       <ADDRESS>{address2}</ADDRESS>')
                a('      </ADDRESS.LIST>')
            a(f'      <APPLICABLEFROM>{applicable_from}</APPLICABLEFROM>')
            if pincode:
                a(f'      <PINCODE>{pincode}</PINCODE>')
            a(f'      <MAILINGNAME>{mailing_name}</MAILINGNAME>')
            if pan:
                a(f'      <INCOMETAXNUMBER>{pan}</INCOMETAXNUMBER>')
            if state:
                a(f'      <STATE>{state}</STATE>')
            a(f'      <COUNTRY>{country}</COUNTRY>')
            a('     </LEDMAILINGDETAILS.LIST>')

        if is_duties_ledger:
            # In Tally ledger XML, Type of Duty/Tax is controlled by TAXTYPE.
            # For GST duty ledgers (IGST/CGST/SGST/etc.), this must be GST.
            if is_gst_duty_ledger:
                a('     <TAXTYPE>GST</TAXTYPE>')
            elif tax_type:
                a(f'     <TAXTYPE>{tax_type}</TAXTYPE>')
            if gst_rate:
                a(f'     <GSTRATE>{gst_rate}</GSTRATE>')
        a('    </LEDGER>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
#  GENERATE STOCK ITEM MASTER XML
# ═══════════════════════════════════════════════════════════════════════════

def generate_stockitem_xml(items: list, company: str) -> str:
    """
    items = list of dict: Name, Parent (stock group), Unit, HSNCode,
    GSTRate, GSTApplicable, Description, OpeningQty, OpeningRate, OpeningValue
    """
    lines = []
    a = lines.append
    company_static = _company_static_block(company)

    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>All Masters</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    for item in items:
        name   = xml_escape(item["Name"])
        parent_raw = _normalize_stock_group_name(item.get("Parent", "Primary"))
        parent = xml_escape(parent_raw)
        unit_raw = _normalize_stock_unit_name(item.get("Unit", "Nos"))
        unit = xml_escape(unit_raw)
        hsn    = xml_escape(item.get("HSNCode",""))
        gst_r  = item.get("GSTRate","")
        gst_a  = item.get("GSTApplicable","Applicable")
        desc   = xml_escape(item.get("Description",""))

        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <STOCKITEM NAME="{name}" ACTION="Create">')
        a(f'     <NAME>{name}</NAME>')

        if parent_raw and parent_raw.casefold() != "primary":
            a(f'     <PARENT>{parent}</PARENT>')

        # ✅ FIX STARTS HERE
        a(f'     <BASEUNITS>{unit}</BASEUNITS>')
        a('     <ISADDITIONALUNITS>NO</ISADDITIONALUNITS>')
        # ❌ REMOVED: <ADDITIONALUNITS>
        # ✅ FIX ENDS HERE

        if hsn:
            a(f'     <GSTDETAILS.LIST>')
            a(f'      <HSNCODE>{hsn}</HSNCODE>')
            a(f'      <TAXABILITY>Taxable</TAXABILITY>')
            if gst_r:
                a(f'      <STATEWISEDETAILS.LIST>')
                a(f'       <RATEDETAILS.LIST>')
                a(f'        <GSTRATE>{gst_r}</GSTRATE>')
                a(f'       </RATEDETAILS.LIST>')
                a(f'      </STATEWISEDETAILS.LIST>')
            a(f'     </GSTDETAILS.LIST>')

        if gst_a:
            a(f'     <GSTAPPLICABLE>{xml_escape(gst_a)}</GSTAPPLICABLE>')

        if desc:
            a(f'     <DESCRIPTION>{desc}</DESCRIPTION>')

        a('    </STOCKITEM>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')

    return "\n".join(lines)

def generate_stockgroup_xml(
    group_names: list,
    company: str,
    default_parent: str = "",
    force_primary_parent: bool = False,
) -> str:
    """Create/alter stock groups so stock item parent groups exist before item import."""
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>All Masters</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    seen = set()
    for raw_name in group_names or []:
        group_name = str(raw_name or "").strip()
        if not group_name:
            continue
        key = group_name.casefold()
        if key in seen:
            continue
        seen.add(key)
        if key == "primary":
            continue

        name = xml_escape(group_name)
        parent_name = str(default_parent or "").strip()

        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <STOCKGROUP NAME="{name}" ACTION="Create Alter">')
        a(f'     <NAME>{name}</NAME>')
        # For many Tally setups, omitting PARENT creates group under Primary.
        # Keep explicit parent only when it's not the root Primary group.
        if parent_name and (parent_name.casefold() != "primary" or force_primary_parent):
            a(f'     <PARENT>{xml_escape(parent_name)}</PARENT>')
        a('    </STOCKGROUP>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


def _normalize_stock_group_name(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "Primary"

    # These are ledger groups, not inventory stock groups.
    ledger_like_groups = {
        "indirect income",
        "direct income",
        "indirect expenses",
        "direct expenses",
        "sales accounts",
        "purchase accounts",
        "sundry debtors",
        "sundry creditors",
        "duties & taxes",
        "duties and taxes",
        "bank accounts",
        "cash-in-hand",
        "cash in hand",
    }
    if text.casefold() in ledger_like_groups:
        return "Primary"
    return text


def _normalize_stock_unit_name(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "Nos"

    # Keep user-entered unit unless it is a common alias.
    aliases = {
        "no": "Nos",
        "no.": "Nos",
        "nos": "Nos",
        "nos.": "Nos",
        "number": "Nos",
        "numbers": "Nos",
        "piece": "pcs",
        "pieces": "pcs",
    }
    return aliases.get(text.casefold(), text)


def _collect_required_stock_groups(items: list) -> list:
    required = []
    seen = set()
    for item in items or []:
        group_name = _normalize_stock_group_name(item.get("Parent", ""))
        if not group_name or group_name.casefold() == "primary":
            continue
        key = group_name.casefold()
        if key in seen:
            continue
        seen.add(key)
        required.append(group_name)
    return required


def _collect_required_stock_units(items: list) -> list:
    required = []
    seen = set()
    for item in items or []:
        unit_name = _normalize_stock_unit_name(item.get("Unit", "Nos"))
        if not unit_name:
            continue
        key = unit_name.casefold()
        if key in seen:
            continue
        seen.add(key)
        # "Not Applicable" behaves like built-in selection in many setups.
        if key in {"not applicable", "n/a", "na"}:
            continue
        required.append(unit_name)
    return required


def generate_unit_xml(unit_names: list, company: str) -> str:
    """Create/alter simple unit masters required for stock items."""
    lines = []
    a = lines.append
    company_static = _company_static_block(company)
    a('<?xml version="1.0" encoding="UTF-8"?>')
    a('<ENVELOPE>')
    a(' <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>')
    a(' <BODY><IMPORTDATA>')
    a('  <REQUESTDESC><REPORTNAME>All Masters</REPORTNAME>')
    if company_static:
        a(company_static)
    a('  </REQUESTDESC>')
    a('  <REQUESTDATA>')

    seen = set()
    for raw in unit_names or []:
        unit_name = _normalize_stock_unit_name(raw)
        if not unit_name:
            continue
        key = unit_name.casefold()
        if key in seen:
            continue
        seen.add(key)
        if key in {"not applicable", "n/a", "na"}:
            continue

        unit_esc = xml_escape(unit_name)
        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <UNIT NAME="{unit_esc}" ACTION="Create Alter">')
        a(f'     <NAME>{unit_esc}</NAME>')
        a('     <ISSIMPLEUNIT>Yes</ISSIMPLEUNIT>')
        a('     <DECIMALPLACES>2</DECIMALPLACES>')
        a(f'     <FORMALNAME>{unit_esc}</FORMALNAME>')
        a('    </UNIT>')
        a('   </TALLYMESSAGE>')

    a('  </REQUESTDATA>')
    a(' </IMPORTDATA></BODY>')
    a('</ENVELOPE>')
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
#  JOURNAL ENTRY HELPERS & XML GENERATION
# ═══════════════════════════════════════════════════════════════════════════

def _ledger_or_default(value: str, fallback: str = SUSPENSE_LEDGER) -> str:
    text = str(value or "").strip()
    return text or fallback


def _state_from_gstin(g):
    M = {
        '01': 'Jammu and Kashmir', '02': 'Himachal Pradesh', '03': 'Punjab', '04': 'Chandigarh',
        '05': 'Uttarakhand', '06': 'Haryana', '07': 'Delhi', '08': 'Rajasthan', '09': 'Uttar Pradesh',
        '10': 'Bihar', '11': 'Sikkim', '12': 'Arunachal Pradesh', '13': 'Nagaland', '14': 'Manipur',
        '15': 'Mizoram', '16': 'Tripura', '17': 'Meghalaya', '18': 'Assam', '19': 'West Bengal',
        '20': 'Jharkhand', '21': 'Odisha', '22': 'Chattisgarh', '23': 'Madhya Pradesh', '24': 'Gujarat',
        '26': 'Dadra and Nagar Haveli and Daman and Diu', '27': 'Maharashtra',
        '28': 'Andhra Pradesh', '29': 'Karnataka', '30': 'Goa', '31': 'Lakshadweep', '32': 'Kerala',
        '33': 'Tamil Nadu', '34': 'Puducherry', '35': 'Andaman and Nicobar Islands',
        '36': 'Telangana', '37': 'Andhra Pradesh (New)', '38': 'Ladakh', '97': 'Other Territory',
    }
    return M.get(str(g or '')[:2], '')


def _normalize_ledger_name(value) -> str:
    text = html.unescape(str(value or ""))
    text = text.replace("\x00", "")
    text = re.sub(r"[\x01-\x1F\x7F]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _ledger_key(value) -> str:
    return _normalize_ledger_name(value).upper()


def _fetch_tally_ledgers(tally_url: str, timeout: float = 15.0, company_name: str = "") -> dict:
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml_variants = [
        (
            "collection-ledger",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
            "<TYPE>Collection</TYPE><ID>Ledger Collection</ID></HEADER>"
            f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE><COLLECTION NAME='Ledger Collection'>"
            "<TYPE>Ledger</TYPE><FETCH>Name,Parent</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
            "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>",
        ),
        (
            "report-list-ledgers",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Ledgers</REPORTNAME>"
            f"{static_vars}</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>",
        ),
    ]

    def _extract_from_response(response_text: str):
        all_entries = []
        try:
            root = ET.fromstring(response_text)
            for ledger_node in root.findall(".//LEDGER"):
                name = _normalize_ledger_name(
                    ledger_node.attrib.get("NAME")
                    or ledger_node.findtext("NAME")
                )
                parent = _normalize_ledger_name(
                    ledger_node.attrib.get("PARENT")
                    or ledger_node.findtext("PARENT")
                )
                if name:
                    all_entries.append((name, parent))
        except ET.ParseError:
            pass

        for match in re.findall(r'LEDGER[^>]*NAME="([^"]+)"', response_text, flags=re.IGNORECASE):
            name = _normalize_ledger_name(match)
            if name:
                all_entries.append((name, ""))
        return all_entries

    all_ledgers_map = {}
    errors = []

    for label, payload in request_xml_variants:
        try:
            response_text = _post_tally_xml(tally_url, payload, timeout=timeout)
            entries = _extract_from_response(response_text)
            for name, parent in entries:
                key = _ledger_key(name)
                if key not in all_ledgers_map:
                    all_ledgers_map[key] = {"name": name, "parent": parent}
                elif not all_ledgers_map[key].get("parent") and parent:
                    all_ledgers_map[key]["parent"] = parent
        except HTTPError as exc:
            errors.append(f"{label}: HTTP {exc.code}")
        except URLError:
            errors.append(f"{label}: ConnectionError")
        except Exception as exc:
            errors.append(f"{label}: {exc}")

    ledgers = sorted((v["name"] for v in all_ledgers_map.values() if v.get("name")), key=lambda x: _ledger_key(x))

    party_group_keys = {"SUNDRY CREDITORS", "SUNDRY DEBTORS"}
    party_ledgers = sorted(
        (
            v["name"]
            for v in all_ledgers_map.values()
            if v.get("name") and re.sub(r"\s+", " ", str(v.get("parent") or "")).strip().upper() in party_group_keys
        ),
        key=lambda x: _ledger_key(x),
    )

    if party_ledgers or ledgers:
        return {
            "success": True,
            "ledgers": ledgers,
            "party_ledgers": party_ledgers,
        }

    err = "; ".join(errors) if errors else "No ledgers returned by Tally."
    return {
        "success": False,
        "error": err,
        "ledgers": [],
        "party_ledgers": [],
    }


def _inject_party_states_from_tally(rows: list, state_map: dict) -> list:
    """Return a copy of rows with State/PlaceOfSupply filled from Tally ledger master
    for any row where GSTIN is absent and State/PlaceOfSupply is also absent.
    Uses _row_text_any for case-insensitive column lookup matching _collect_party_context."""
    if not state_map:
        return rows
    result = []
    for r in rows:
        # Use _row_text_any (case-insensitive) to match exactly what _collect_party_context reads
        gstin = _row_text_any(r, [
            "PartyGSTIN", "GSTIN", "GSTIN/UIN", "GSTIN UIN",
            "Party GSTIN", "SupplierGSTIN", "Supplier GSTIN",
            "GST No", "GST Number",
        ]).upper()
        state_in_row = _normalize_state_for_ledger(_row_text_any(r, [
            "PartyState", "State", "StateName", "State Name",
        ]))
        pos_in_row = _normalize_state_for_ledger(_row_text_any(r, [
            "PlaceOfSupply", "Place Of Supply", "Place of Supply",
            "POS", "StateOfSupply",
        ]))
        if gstin or (state_in_row and pos_in_row):
            # Already has GSTIN (state derived from it) or both state and POS — leave as-is
            result.append(r)
            continue
        # Use same key priority as _resolve_party_ledger (sales mode)
        party_raw = _row_text_any(r, [
            "PartyLedger", "PartyName", "Party Name",
            "BuyerName", "CustomerName", "Customer",
            "BillToName", "Bill To Name", "Party",
        ])
        fetched_state = state_map.get(party_raw.upper(), "") if party_raw else ""
        if not fetched_state:
            result.append(r)
            continue
        new_r = dict(r)
        if not state_in_row:
            new_r["State"] = fetched_state
        if not pos_in_row:
            new_r["PlaceOfSupply"] = fetched_state
        result.append(new_r)
    return result


def _fetch_party_ledger_states(tally_url: str, company_name: str = "", timeout: float = 15.0) -> dict:
    """Fetch StateName for every ledger in Tally.
    Returns {ledger_name_upper: state_name} for all ledgers that have a StateName set.
    Used to fill in PlaceOfSupply for unregistered parties that have no GSTIN/State in the Excel.
    """
    selected_company = _normalize_company_name(company_name)
    sc = (f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
          if selected_company else "")
    rq = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>SPELedStateCol</ID></HEADER>"
        "<BODY><DESC>"
        f"<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>{sc}</STATICVARIABLES>"
        "<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='SPELedStateCol'><TYPE>Ledger</TYPE>"
        "<FETCH>Name,StateName,GSTRegistrationType</FETCH></COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )
    try:
        resp = _post_tally_xml(tally_url, rq, timeout=timeout)
    except Exception:
        return {}

    state_map = {}
    try:
        root = ET.fromstring(resp)
        for led in root.findall(".//LEDGER"):
            name = (led.attrib.get("NAME") or led.findtext("NAME") or "").strip()
            state = (led.findtext("STATENAME") or "").strip()
            if name and state:
                state_map[name.upper()] = _normalize_state_for_ledger(state)
    except ET.ParseError:
        pass

    # Fallback: regex parse for NAME/STATENAME pairs
    if not state_map:
        for blk in re.findall(r'<LEDGER\b[^>]*>(.*?)</LEDGER>', resp, re.DOTALL | re.IGNORECASE):
            def _tag(t, b=blk):
                m = re.search(fr'<{t}[^>]*>(.*?)</{t}>', b, re.DOTALL | re.IGNORECASE)
                return m.group(1).strip() if m else ""
            name = _tag("NAME")
            state = _tag("STATENAME")
            if name and state:
                state_map[name.upper()] = _normalize_state_for_ledger(state)

    return state_map


def _fetch_cmp_gst_regs_for_journal(url, company='', timeout=15):
    sc = f'<SVCURRENTCOMPANY>{xml_escape(company)}</SVCURRENTCOMPANY>' if company else ''
    col = ("<COLLECTION NAME='TUL'><TYPE>TaxUnit</TYPE>"
           "<FETCH>Name,TaxType,TaxRegistration,GSTRegNumber,StateName</FETCH></COLLECTION>")
    rq = (
        f'<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>'
        f'<TYPE>Collection</TYPE><ID>TUL</ID></HEADER>'
        f'<BODY><DESC><STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>{sc}'
        f'</STATICVARIABLES><TDL><TDLMESSAGE>{col}</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>'
    )
    try:
        r = _post_tally_xml(url, rq, timeout=timeout)
        regs = []
        for m in re.finditer(r'<TAXUNIT[^>]*>(.*?)</TAXUNIT>', r, re.DOTALL):
            blk = m.group(1)
            def _t(tag, b=blk):
                x = re.search(fr'<{tag}[^>]*>(.*?)</{tag}>', b, re.DOTALL)
                return x.group(1).strip() if x else ''
            gstin = (_t('GSTREGNUMBER') or _t('TAXREGISTRATION') or '').upper()
            state = _t('STATENAME') or ''
            name = _t('NAME') or ''
            if not gstin:
                continue
            if not state:
                state = _state_from_gstin(gstin)
            regs.append({'gstin': gstin, 'state': state, 'name': name})
        return {'success': bool(regs), 'registrations': regs}
    except Exception as e:
        return {'success': False, 'registrations': [], 'error': str(e)}


def _clean_tax_ledger(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.casefold() in {"0", "0.0", "none", "na", "n/a", "-"}:
        return ""
    return text


def _row_reference_number(row: dict, default: str = "") -> str:
    return (
        _row_text(row, "Reference")
        or _row_text(row, "RefNo")
        or _row_text(row, "Ref No")
        or _row_text(row, "BillRef")
        or _row_text(row, "BillNo")
        or _row_text(row, "Bill No")
        or _row_text(row, "InvoiceNo")
        or _row_text(row, "Invoice No")
        or _row_text(row, "SupplierInvoiceNo")
        or _row_text(row, "Supplier Invoice No")
        or default
    )


def _append_common_ledger_flags(add_line, is_party: bool, is_debit: bool = None) -> None:
    if is_debit is None:
        is_debit = not is_party
    add_line("      <GSTCLASS>Not Applicable</GSTCLASS>")
    add_line(f"      <ISDEEMEDPOSITIVE>{'Yes' if is_debit else 'No'}</ISDEEMEDPOSITIVE>")
    add_line("      <LEDGERFROMITEM>No</LEDGERFROMITEM>")
    add_line("      <REMOVEZEROENTRIES>No</REMOVEZEROENTRIES>")
    add_line(f"      <ISPARTYLEDGER>{'Yes' if is_party else 'No'}</ISPARTYLEDGER>")
    add_line("      <GSTOVERRIDDEN>No</GSTOVERRIDDEN>")
    add_line("      <ISGSTASSESSABLEVALUEOVERRIDDEN>No</ISGSTASSESSABLEVALUEOVERRIDDEN>")


def generate_journal_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
    include_voucher_number: bool = True,
    include_bill_allocations: bool = True,
    journal_type: str = "purchase",
    company_gst_registrations: list = None,
    round_off_ledger: str = "",
) -> tuple:
    """
    Purchase: Expense Dr / Tax Dr / Party Cr
    Sale:     Party Dr  / Sales Cr / Tax Cr
    GST State/Country fetched by Tally from party ledger master via GSTREGISTRATIONTYPE=Regular.
    """
    is_sale = str(journal_type or "purchase").strip().lower() == "sale"
    lines = []
    a = lines.append
    company_static = _company_static_block(company)

    a('<?xml version="1.0" encoding="UTF-8"?>')
    a("<ENVELOPE>")
    a(" <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>")
    a(" <BODY><IMPORTDATA>")
    a("  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>")
    if company_static:
        a(company_static)
    a("  </REQUESTDESC>")
    a("  <REQUESTDATA>")

    resolved_mode = str(date_mode or ("current" if use_today_date else "excel")).strip().lower()
    if resolved_mode not in {"current", "excel", "custom"}:
        resolved_mode = "current" if use_today_date else "excel"
    resolved_custom_date = _normalize_manual_date_to_tally(custom_tally_date) if resolved_mode == "custom" else ""

    _cmp_regs = list(company_gst_registrations or [])
    _cmp_gstin = _cmp_state = _cmp_name = ""
    if _cmp_regs:
        _r = _cmp_regs[0]
        _cmp_gstin = xml_escape(str(_r.get("gstin", "")).strip())
        _cmp_state = xml_escape(str(_r.get("state", "")).strip())
        _cmp_name = xml_escape(str(_r.get("name", "")).strip())

    voucher_count = 0

    for idx, r in enumerate(rows):
        taxable = _row_float(r, "TaxableValue", 0.0)
        if taxable <= 0:
            continue

        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)
        excel_vno = _row_voucher_number(r, "")
        if excel_vno:
            vno_raw = excel_vno
        else:
            vno_raw = str(idx + 1)
            
        invoice_ref_raw = _row_invoice_reference(r, vno_raw) # Reference No

        party_raw = _ledger_or_default(_row_text(r, "PartyLedger"))
        particular_raw = (
            _row_text(r, "Particular")
            or _row_text(r, "Particulars")
            or _row_text(r, "SalesLedger")
            or _row_text(r, "Sales Ledger")
            or _row_text(r, "ExpenseLedger")
            or _row_text(r, "PurchaseLedger")
            or "Journal Adjustment"
        )
        particular_raw = _ledger_or_default(particular_raw, "Journal Adjustment")

        cgst_ledger_raw = _clean_tax_ledger(_row_text(r, "CGSTLedger"))
        sgst_ledger_raw = _clean_tax_ledger(_row_text(r, "SGSTLedger"))
        igst_ledger_raw = _clean_tax_ledger(_row_text(r, "IGSTLedger"))

        cgst_rate = _row_float(r, "CGSTRate", 0.0)
        sgst_rate = _row_float(r, "SGSTRate", 0.0)
        igst_rate = _row_float(r, "IGSTRate", 0.0)

        cgst_amt_explicit = _row_float(r, "CGST Amount", 0.0)
        sgst_amt_explicit = _row_float(r, "SGST Amount", 0.0)
        igst_amt_explicit = _row_float(r, "IGST Amount", 0.0)

        cgst_amt = round(taxable * cgst_rate / 100, 2) if cgst_rate > 0 else cgst_amt_explicit
        sgst_amt = round(taxable * sgst_rate / 100, 2) if sgst_rate > 0 else sgst_amt_explicit
        igst_amt = round(taxable * igst_rate / 100, 2) if igst_rate > 0 else igst_amt_explicit

        if not cgst_ledger_raw and cgst_amt > 0:
              cgst_ledger_raw = "CGST"
        if not sgst_ledger_raw and sgst_amt > 0:
              sgst_ledger_raw = "SGST"
        if not igst_ledger_raw and igst_amt > 0:
              igst_ledger_raw = "IGST"

        cgst_amt = cgst_amt if cgst_ledger_raw else 0.0
        sgst_amt = sgst_amt if sgst_ledger_raw else 0.0
        igst_amt = igst_amt if igst_ledger_raw else 0.0

        total = taxable + cgst_amt + sgst_amt + igst_amt
        has_gst = (cgst_amt + sgst_amt + igst_amt) > 0
        _ro_amt_jnl  = round(round(total, 0) - total, 2) if round_off_ledger else 0.0
        _ro_total_jnl = round(total + _ro_amt_jnl, 2)

        # TDS fields (only applied for Purchase journal)
        jnl_tds_ledger_raw = (
            _row_text(r, "TDSLedger")
            or _row_text(r, "TDS Ledger")
            or _row_text(r, "Tds Ledger")
        ) if not is_sale else ""
        jnl_tds_rate = (_row_float(r, "TDSRate", 0.0) or _row_float(r, "TDS Rate", 0.0)) if not is_sale else 0.0
        jnl_tds_amount_raw = (_row_float(r, "TDSAmount", 0.0) or _row_float(r, "TDS Amount", 0.0)) if not is_sale else 0.0
        if jnl_tds_ledger_raw and jnl_tds_amount_raw <= 0 and jnl_tds_rate > 0:
            jnl_tds_amount = round(taxable * jnl_tds_rate / 100, 2)
        else:
            jnl_tds_amount = abs(jnl_tds_amount_raw)
        jnl_tds_led = xml_escape(jnl_tds_ledger_raw)
        jnl_party_total = _ro_total_jnl - jnl_tds_amount if (jnl_tds_led and jnl_tds_amount > 0) else _ro_total_jnl

        bill_reference_raw = _row_reference_number(r, "")
        voucher_reference_raw = bill_reference_raw or (vno_raw if include_voucher_number else "")
        vno = xml_escape(vno_raw)
        reference = xml_escape(voucher_reference_raw)
        bill_reference = xml_escape(bill_reference_raw)
        party = xml_escape(party_raw)
        particular = xml_escape(particular_raw)
        narration = xml_escape(_row_text(r, "Narration"))

        party_gstin_raw = str(_row_text(r, "GSTIN/UIN") or _row_text(r, "PartyGSTIN") or _row_text(r, "GSTIN") or "").strip().upper()
        pos_raw = str(_row_text(r, "PlaceOfSupply") or "").strip()
        party_state_raw = pos_raw or (_state_from_gstin(party_gstin_raw) if party_gstin_raw else "")
        place_of_supply = xml_escape(party_state_raw if is_sale else (pos_raw or _cmp_state))
        party_gstin = xml_escape(party_gstin_raw)
        party_state = xml_escape(party_state_raw)

        voucher_count += 1
        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a('    <VOUCHER REMOTEID="" VCHTYPE="Journal" ACTION="Create" OBJVIEW="Accounting Voucher View">')
        a(f"     <DATE>{dt}</DATE>")
        a("     <VOUCHERTYPENAME>Journal</VOUCHERTYPENAME>")
        a("     <PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW>")
        a("     <VCHENTRYMODE>Accounting Voucher View</VCHENTRYMODE>")
        a("     <ISINVOICE>No</ISINVOICE>")
        a(f"     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>")
        a("     <ISELIGIBLEFORITC>No</ISELIGIBLEFORITC>")

        if has_gst:
            a("     <ISGSTOVERRIDDEN>No</ISGSTOVERRIDDEN>")
            a("     <GSTTRANSACTIONTYPE>Tax Invoice</GSTTRANSACTIONTYPE>")
            a("     <GSTREGISTRATIONTYPE>Regular</GSTREGISTRATIONTYPE>")
            if party_gstin:
                a(f'     <PARTYGSTIN>{party_gstin}</PARTYGSTIN>')
            if _cmp_gstin and _cmp_name:
                a(f'     <GSTREGISTRATION TAXTYPE="GST" TAXREGISTRATION="{_cmp_gstin}">{_cmp_name}</GSTREGISTRATION>')
                a(f'     <CMPGSTIN>{_cmp_gstin}</CMPGSTIN>')
                a('     <CMPGSTREGISTRATIONTYPE>Regular</CMPGSTREGISTRATIONTYPE>')
            if _cmp_state:
                a(f'     <CMPGSTSTATE>{_cmp_state}</CMPGSTSTATE>')
            if place_of_supply:
                a(f'     <PLACEOFSUPPLY>{place_of_supply}</PLACEOFSUPPLY>')
        if reference:
            a(f"     <REFERENCE>{reference}</REFERENCE>")
        if include_voucher_number and vno:
            a(f"     <VOUCHERNUMBER>{vno}</VOUCHERNUMBER>")
        if narration:
            a(f"     <NARRATION>{narration}</NARRATION>")

        if is_sale:
            # ---- Sale Journal: Party Dr / Income Cr / Output Tax Cr ----
            a("     <LEDGERENTRIES.LIST>")
            a(f"      <LEDGERNAME>{party}</LEDGERNAME>")
            _append_common_ledger_flags(a, is_party=True, is_debit=True)
            if has_gst:
                a("      <GSTREGISTRATIONTYPE>Regular</GSTREGISTRATIONTYPE>")
                if party_gstin:
                    a(f"      <GSTIN>{party_gstin}</GSTIN>")
                if party_state:
                    a(f"      <STATENAME>{party_state}</STATENAME>")
                a("      <COUNTRYOFRESIDENCE>India</COUNTRYOFRESIDENCE>")
            a(f"      <AMOUNT>-{fmt_amt(_ro_total_jnl)}</AMOUNT>")
            if include_bill_allocations and bill_reference:
                a("      <BILLALLOCATIONS.LIST>")
                a(f"       <NAME>{bill_reference}</NAME>")
                a("       <BILLTYPE>New Ref</BILLTYPE>")
                a(f"       <AMOUNT>-{fmt_amt(_ro_total_jnl)}</AMOUNT>")
                a("      </BILLALLOCATIONS.LIST>")
            a("     </LEDGERENTRIES.LIST>")

            a("     <LEDGERENTRIES.LIST>")
            a(f"      <LEDGERNAME>{particular}</LEDGERNAME>")
            _append_common_ledger_flags(a, is_party=False, is_debit=False)
            a(f"      <AMOUNT>{fmt_amt(taxable)}</AMOUNT>")
            a("     </LEDGERENTRIES.LIST>")

            for _ln, _la in [(cgst_ledger_raw, cgst_amt), (sgst_ledger_raw, sgst_amt), (igst_ledger_raw, igst_amt)]:
                if _la > 0 and _ln:
                    a("     <LEDGERENTRIES.LIST>")
                    a(f"      <LEDGERNAME>{xml_escape(_ln)}</LEDGERNAME>")
                    _append_common_ledger_flags(a, is_party=False, is_debit=False)
                    a(f"      <AMOUNT>{fmt_amt(_la)}</AMOUNT>")
                    a("     </LEDGERENTRIES.LIST>")

            _append_round_off_entry(a, round_off_ledger, _ro_amt_jnl, is_purchase_mode=False)

        else:
            # ---- Purchase Journal: Expense Dr / Input Tax Dr / Party Cr ----
            a("     <LEDGERENTRIES.LIST>")
            a(f"      <LEDGERNAME>{particular}</LEDGERNAME>")
            _append_common_ledger_flags(a, is_party=False)
            a(f"      <AMOUNT>-{fmt_amt(taxable)}</AMOUNT>")
            a("     </LEDGERENTRIES.LIST>")

            if cgst_amt > 0 and cgst_ledger_raw:
                a("     <LEDGERENTRIES.LIST>")
                a(f"      <LEDGERNAME>{xml_escape(cgst_ledger_raw)}</LEDGERNAME>")
                _append_common_ledger_flags(a, is_party=False)
                a(f"      <AMOUNT>-{fmt_amt(cgst_amt)}</AMOUNT>")
                a("     </LEDGERENTRIES.LIST>")

            if sgst_amt > 0 and sgst_ledger_raw:
                a("     <LEDGERENTRIES.LIST>")
                a(f"      <LEDGERNAME>{xml_escape(sgst_ledger_raw)}</LEDGERNAME>")
                _append_common_ledger_flags(a, is_party=False)
                a(f"      <AMOUNT>-{fmt_amt(sgst_amt)}</AMOUNT>")
                a("     </LEDGERENTRIES.LIST>")

            if igst_amt > 0 and igst_ledger_raw:
                a("     <LEDGERENTRIES.LIST>")
                a(f"      <LEDGERNAME>{xml_escape(igst_ledger_raw)}</LEDGERNAME>")
                _append_common_ledger_flags(a, is_party=False)
                a(f"      <AMOUNT>-{fmt_amt(igst_amt)}</AMOUNT>")
                a("     </LEDGERENTRIES.LIST>")

            # TDS Payable - Credit
            if jnl_tds_led and jnl_tds_amount > 0:
                a("     <LEDGERENTRIES.LIST>")
                a(f"      <LEDGERNAME>{jnl_tds_led}</LEDGERNAME>")
                _append_common_ledger_flags(a, is_party=False, is_debit=False)
                a(f"      <AMOUNT>{fmt_amt(jnl_tds_amount)}</AMOUNT>")
                a("     </LEDGERENTRIES.LIST>")

            _append_round_off_entry(a, round_off_ledger, _ro_amt_jnl, is_purchase_mode=True)

            a("     <LEDGERENTRIES.LIST>")
            a(f"      <LEDGERNAME>{party}</LEDGERNAME>")
            _append_common_ledger_flags(a, is_party=True)
            if has_gst:
                a("      <GSTREGISTRATIONTYPE>Regular</GSTREGISTRATIONTYPE>")
                if party_gstin:
                    a(f"      <GSTIN>{party_gstin}</GSTIN>")
                if party_state:
                    a(f"      <STATENAME>{party_state}</STATENAME>")
                a("      <COUNTRYOFRESIDENCE>India</COUNTRYOFRESIDENCE>")
            a(f"      <AMOUNT>{fmt_amt(jnl_party_total)}</AMOUNT>")
            if include_bill_allocations and bill_reference:
                a("      <BILLALLOCATIONS.LIST>")
                a(f"       <NAME>{bill_reference}</NAME>")
                a("       <BILLTYPE>New Ref</BILLTYPE>")
                a(f"       <AMOUNT>{fmt_amt(jnl_party_total)}</AMOUNT>")
                a("      </BILLALLOCATIONS.LIST>")
            a("     </LEDGERENTRIES.LIST>")

        a("    </VOUCHER>")
        a("   </TALLYMESSAGE>")

    a("  </REQUESTDATA>")
    a(" </IMPORTDATA></BODY>")
    a("</ENVELOPE>")
    return "\n".join(lines), voucher_count


# ═══════════════════════════════════════════════════════════════════════════
#  CREDIT / DEBIT NOTE XML GENERATION
# ═══════════════════════════════════════════════════════════════════════════

def _normalize_note_type(value: str) -> str:
    text = re.sub(r"\([^)]*\)", "", str(value or "")).strip().casefold()
    if text in {"debit note", "debit", "debitnote"}:
        return "Debit Note"
    return "Credit Note"


def _clean_note_tax_ledger(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.casefold() in {"0", "0.0", "none", "na", "n/a", "-"}:
        return ""
    return text


def _collect_auto_note_ledgers(rows: list, is_debit_note: bool = False) -> list:
    """Collect ledger defs from Credit/Debit Note rows for pre-creation."""
    entries = {}

    def _add(name, parent, tax_type="", gst_rate="", extra=None):
        ledger_name = str(name or "").strip()
        if _is_effectively_blank_ledger(ledger_name):
            return
        key = _ledger_name_key(ledger_name)
        if not key:
            return
        candidate = {
            "Name": ledger_name,
            "Parent": parent,
            "GSTApplicable": "",
            "GSTIN": "",
            "StateOfSupply": "",
            "TypeOfTaxation": tax_type or "",
            "GSTRate": str(gst_rate).strip() if gst_rate not in (None, "") else "",
        }
        if extra:
            for f, v in extra.items():
                if f in {"Name", "Parent"} or v in (None, ""):
                    continue
                candidate[f] = v
        existing = entries.get(key)
        if existing is None:
            entries[key] = candidate
        else:
            _merge_ledger_definitions(existing, candidate)

    for r in rows or []:
        party_name = str(
            _row_text(r, "PartyLedger") or _row_text(r, "Party Ledger") or ""
        ).strip()
        gstin_raw = str(
            _row_text(r, "GSTIN/UIN") or _row_text(r, "GSTIN") or _row_text(r, "PartyGSTIN") or ""
        ).strip().upper()

        if party_name and not _is_effectively_blank_ledger(party_name):
            party_parent = "Sundry Creditors" if is_debit_note else "Sundry Debtors"
            party_def = {
                "Name": party_name,
                "Parent": party_parent,
                "GSTIN": gstin_raw,
                "PAN": _pan_from_gstin(gstin_raw),
                "GSTApplicable": "Applicable" if gstin_raw else "Not Applicable",
                "GSTRegistrationType": "Regular" if gstin_raw else "Unregistered/Consumer",
                "StateOfSupply": _state_name_from_gstin(gstin_raw) if gstin_raw else "",
                "MailingName": party_name,
                "Country": "India",
                "Billwise": "Yes",
                "Pincode": "",
                "Address1": "",
                "Address2": "",
                "TypeOfTaxation": "",
                "GSTRate": "",
            }
            _add(party_name, party_parent, extra=party_def)

        particular = (
            _row_text(r, "Particular") or _row_text(r, "Particulars")
            or _row_text(r, "SalesLedger") or _row_text(r, "Sales Ledger")
            or _row_text(r, "Purchase Ledger") or _row_text(r, "PurchaseLedger")
        )
        if particular and not _is_effectively_blank_ledger(particular):
            _add(particular, "Purchase Accounts" if is_debit_note else "Sales Accounts")

        cgst_l = _clean_note_tax_ledger(_row_text(r, "CGSTLedger"))
        sgst_l = _clean_note_tax_ledger(_row_text(r, "SGSTLedger"))
        igst_l = _clean_note_tax_ledger(_row_text(r, "IGSTLedger"))
        cgst_r = _row_float(r, "CGSTRate", 0.0)
        sgst_r = _row_float(r, "SGSTRate", 0.0)
        igst_r = _row_float(r, "IGSTRate", 0.0)
        if not cgst_l and cgst_r > 0:
            cgst_l = "CGST"
        if not sgst_l and sgst_r > 0:
            sgst_l = "SGST"
        if not igst_l and igst_r > 0:
            igst_l = "IGST"
        _add(cgst_l, "Duties & Taxes", "Central Tax", _row_text(r, "CGSTRate"))
        _add(sgst_l, "Duties & Taxes", "State Tax", _row_text(r, "SGSTRate"))
        _add(igst_l, "Duties & Taxes", "Integrated Tax", _row_text(r, "IGSTRate"))

    return list(entries.values())


def _consolidate_note_rows(rows: list, resolved_mode: str, resolved_custom_date: str) -> list:
    """
    Merge Credit/Debit Note rows with same (Date, GSTIN, VoucherNo, PartyLedger)
    into one voucher entry. Accumulated tax amounts stored as _note_*_amt.
    Items list stored for item-mode multi-line vouchers.
    """
    consolidated = []
    group_map = {}

    for r in rows:
        taxable = _row_float(r, "TaxableValue", 0.0)
        if taxable <= 0:
            continue

        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)

        vno_raw = _row_voucher_number(r, "")
        gstin_raw = (
            _row_text(r, "GSTIN/UIN") or _row_text(r, "GSTIN") or _row_text(r, "PartyGSTIN") or ""
        ).upper().strip()
        party_raw = str(_row_text(r, "PartyLedger") or _row_text(r, "Party Ledger") or "").strip().lower()

        key = (dt, gstin_raw.lower(), str(vno_raw).strip().lower(), party_raw)

        cgst_rate = _row_float(r, "CGSTRate", 0.0)
        sgst_rate = _row_float(r, "SGSTRate", 0.0)
        igst_rate = _row_float(r, "IGSTRate", 0.0)
        cgst_x = _row_float(r, "CGST Amount", 0.0)
        sgst_x = _row_float(r, "SGST Amount", 0.0)
        igst_x = _row_float(r, "IGST Amount", 0.0)
        row_cgst = round(taxable * cgst_rate / 100, 2) if cgst_rate > 0 else cgst_x
        row_sgst = round(taxable * sgst_rate / 100, 2) if sgst_rate > 0 else sgst_x
        row_igst = round(taxable * igst_rate / 100, 2) if igst_rate > 0 else igst_x

        item_name = _row_text_any(r, ["Item Name", "ItemName", "Item", "StockItem", "ProductName"], "")
        particular = _row_text(r, "Particular") or _row_text(r, "Particulars") or ""

        if key in group_map:
            base = group_map[key]
            base["TaxableValue"] = _row_float(base, "TaxableValue", 0.0) + taxable
            base["_note_cgst_amt"] = base.get("_note_cgst_amt", 0.0) + row_cgst
            base["_note_sgst_amt"] = base.get("_note_sgst_amt", 0.0) + row_sgst
            base["_note_igst_amt"] = base.get("_note_igst_amt", 0.0) + row_igst
            # Zero out per-row rates so generate uses accumulated amounts
            base["CGSTRate"] = 0
            base["SGSTRate"] = 0
            base["IGSTRate"] = 0

            # Initialise items list on first merge
            if "items" not in base:
                base["items"] = []
                orig_tv = base.get("_note_first_taxable", taxable)
                orig_item = {
                    "ItemName": _row_text_any(base, ["Item Name","ItemName","Item","StockItem","ProductName"], ""),
                    "Quantity": _row_float(base, "Quantity", 0.0) or _row_float(base, "Qty", 0.0) or 1.0,
                    "Rate": _row_float(base, "Rate", 0.0),
                    "Per": _row_text_any(base, ["Unit","UOM","Per"], "") or "Nos",
                    "GodownName": _row_text(base, "GodownName") or "Main Location",
                    "Particular": _row_text(base, "Particular") or _row_text(base, "Particulars") or "",
                    "TaxableValue": orig_tv,
                }
                if orig_item["ItemName"]:
                    base["items"].append(orig_item)
                # Accounting entries for accounting mode
                base["accounting_entries"] = [{
                    "Particular": orig_item["Particular"],
                    "TaxableValue": orig_tv,
                }]

            new_item = {
                "ItemName": item_name,
                "Quantity": _row_float(r, "Quantity", 0.0) or _row_float(r, "Qty", 0.0) or 1.0,
                "Rate": _row_float(r, "Rate", 0.0),
                "Per": _row_text_any(r, ["Unit","UOM","Per"], "") or "Nos",
                "GodownName": _row_text(r, "GodownName") or "Main Location",
                "Particular": particular,
                "TaxableValue": taxable,
            }
            if item_name:
                base["items"].append(new_item)
            if "accounting_entries" in base:
                base["accounting_entries"].append({"Particular": particular, "TaxableValue": taxable})
        else:
            new_r = dict(r)
            new_r["_note_first_taxable"] = taxable
            new_r["_note_cgst_amt"] = row_cgst
            new_r["_note_sgst_amt"] = row_sgst
            new_r["_note_igst_amt"] = row_igst
            group_map[key] = new_r
            consolidated.append(new_r)

    return consolidated


def generate_note_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
    voucher_type: str = "Credit Note",
    company_gst_registrations: list = None,
    entry_mode: str = "accounting",
    round_off_ledger: str = "",
) -> tuple:
    """
    Credit/Debit Note accounting or item:
    - Credit Note: party credited (ISDEEMEDPOSITIVE=No), particular/tax debited (Yes)
    - Debit Note:  party debited  (ISDEEMEDPOSITIVE=Yes), particular/tax credited (No)
    """
    normalized_type = _normalize_note_type(voucher_type)
    is_debit_note = (normalized_type == "Debit Note")
    default_particular_ledger = f"{normalized_type} Account"
    resolved_entry_mode = str(entry_mode or "accounting").strip().lower()
    if resolved_entry_mode not in {"accounting", "item"}:
        resolved_entry_mode = "accounting"
    is_item_mode = (resolved_entry_mode == "item")

    lines_out = []
    a = lines_out.append
    company_static = _company_static_block(company)

    a('<?xml version="1.0" encoding="UTF-8"?>')
    a("<ENVELOPE>")
    a(" <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>")
    a(" <BODY><IMPORTDATA>")
    a("  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>")
    if company_static:
        a(company_static)
    a("  </REQUESTDESC>")
    a("  <REQUESTDATA>")

    voucher_count = 0
    resolved_mode = str(date_mode or ("current" if use_today_date else "excel")).strip().lower()
    if resolved_mode not in {"current", "excel", "custom"}:
        resolved_mode = "current" if use_today_date else "excel"
    resolved_custom_date = _normalize_manual_date_to_tally(custom_tally_date) if resolved_mode == "custom" else ""

    # Merge rows with same Date+GSTIN+VoucherNo+PartyLedger into one voucher
    rows = _consolidate_note_rows(rows, resolved_mode, resolved_custom_date)

    for idx, r in enumerate(rows):
        taxable = _row_float(r, "TaxableValue", 0.0)
        if taxable <= 0:
            continue

        if resolved_mode == "current":
            from datetime import datetime as _dt
            source_date = _dt.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)

        vno_raw = _row_voucher_number(r, str(idx + 1))
        party_raw = _ledger_or_suspense(_row_text(r, "PartyLedger") or _row_text(r, "Party Ledger"))
        particular_raw = (
            _row_text(r, "Particular") or _row_text(r, "Particulars")
            or _row_text(r, "SalesLedger") or _row_text(r, "Sales Ledger")
            or _row_text(r, "Purchase Ledger") or default_particular_ledger
        )
        particular_raw = _ledger_or_suspense(particular_raw) or default_particular_ledger

        cgst_ledger_raw = _clean_note_tax_ledger(_row_text(r, "CGSTLedger"))
        sgst_ledger_raw = _clean_note_tax_ledger(_row_text(r, "SGSTLedger"))
        igst_ledger_raw = _clean_note_tax_ledger(_row_text(r, "IGSTLedger"))

        cgst_rate = _row_float(r, "CGSTRate", 0.0)
        sgst_rate = _row_float(r, "SGSTRate", 0.0)
        igst_rate = _row_float(r, "IGSTRate", 0.0)

        # Use pre-accumulated amounts from consolidation when available
        if "_note_cgst_amt" in r:
            cgst_amt = r["_note_cgst_amt"]
            sgst_amt = r["_note_sgst_amt"]
            igst_amt = r["_note_igst_amt"]
        else:
            cgst_amt_explicit = _row_float(r, "CGST Amount", 0.0)
            sgst_amt_explicit = _row_float(r, "SGST Amount", 0.0)
            igst_amt_explicit = _row_float(r, "IGST Amount", 0.0)
            cgst_amt = round(taxable * cgst_rate / 100, 2) if cgst_rate > 0 else cgst_amt_explicit
            sgst_amt = round(taxable * sgst_rate / 100, 2) if sgst_rate > 0 else sgst_amt_explicit
            igst_amt = round(taxable * igst_rate / 100, 2) if igst_rate > 0 else igst_amt_explicit

        if not cgst_ledger_raw and cgst_amt > 0:
            cgst_ledger_raw = "CGST"
        if not sgst_ledger_raw and sgst_amt > 0:
            sgst_ledger_raw = "SGST"
        if not igst_ledger_raw and igst_amt > 0:
            igst_ledger_raw = "IGST"
             
        cgst_amt = cgst_amt if cgst_ledger_raw else 0.0
        sgst_amt = sgst_amt if sgst_ledger_raw else 0.0
        igst_amt = igst_amt if igst_ledger_raw else 0.0
        
        total = taxable + cgst_amt + sgst_amt + igst_amt
        _ro_amt_n  = round(round(total, 0) - total, 2) if round_off_ledger else 0.0
        _ro_total_n = round(total + _ro_amt_n, 2)

        vno = xml_escape(vno_raw)
        party = xml_escape(party_raw)
        particular = xml_escape(particular_raw)
        narration = xml_escape(_row_text(r, "Narration"))
        gstin_raw = (_row_text(r, "GSTIN/UIN") or _row_text(r, "GSTIN") or _row_text(r, "PartyGSTIN")).upper()
        gstin = xml_escape(gstin_raw)

        if is_item_mode:
            item_name_raw = _row_text_any(
                r,
                ["Item Name", "ItemName", "Item", "StockItem", "ProductName"],
                "",
            )
            if not item_name_raw:
                raise ValueError(f"{normalized_type} row {idx + 1}: item name is missing.")
            qty = _row_float(r, "Quantity", 0.0) or _row_float(r, "Qty", 0.0) or _row_float(r, "Unit", 0.0)
            if qty <= 0:
                qty = 1.0
            rate = _row_float(r, "Rate", 0.0)
            if rate <= 0 and taxable > 0 and qty > 0:
                rate = taxable / qty
            per_unit_raw = _row_text_any(r, ["Unit", "UOM", "Per"], "") or "Nos"
            per_unit = xml_escape(_normalize_stock_unit_name(per_unit_raw) or "Nos")
            godown = xml_escape(_row_text(r, "GodownName", "Main Location") or "Main Location")
            item_name = xml_escape(item_name_raw)
            item_amt = round(qty * rate, 2) if qty and rate else taxable

        place_raw = _row_text(r, "PlaceOfSupply") or _row_text(r, "Place Of Supply") or _state_name_from_gstin(gstin_raw)
        state_xml = xml_escape(place_raw)

        # Sign convention
        party_is_deemed_positive = "Yes" if is_debit_note else "No"
        party_amount = -_ro_total_n if is_debit_note else _ro_total_n
        counter_is_deemed_positive = "No" if is_debit_note else "Yes"
        taxable_amount = taxable if is_debit_note else -taxable
        cgst_amount = cgst_amt if is_debit_note else -cgst_amt
        sgst_amount = sgst_amt if is_debit_note else -sgst_amt
        igst_amount = igst_amt if is_debit_note else -igst_amt

        voucher_count += 1
        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <VOUCHER VCHTYPE="{normalized_type}" ACTION="Create" OBJVIEW="Invoice Voucher View">')
        a(f"     <DATE>{dt}</DATE>")
        a(f"     <VOUCHERTYPENAME>{normalized_type}</VOUCHERTYPENAME>")
        a(f"     <VOUCHERNUMBER>{vno}</VOUCHERNUMBER>")
        a(f"     <PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>")
        a(f"     <PARTYNAME>{party}</PARTYNAME>")

        if not is_debit_note:
            a(f"     <BASICBUYERNAME>{party}</BASICBUYERNAME>")
            if state_xml:
                a(f"     <STATENAME>{state_xml}</STATENAME>")
                a(f"     <PLACEOFSUPPLY>{state_xml}</PLACEOFSUPPLY>")
        else:
            _cmp_regs = list(company_gst_registrations or [])
            _cmp_state = _cmp_gstin = _cmp_name = ""
            if _cmp_regs:
                _cr = _cmp_regs[0]
                _cmp_gstin = xml_escape(str(_cr.get("gstin", "") or "").strip())
                _cmp_state = xml_escape(str(_cr.get("state", "") or "").strip())
                _cmp_name  = xml_escape(str(_cr.get("name", "") or "").strip())
            if state_xml:
                a(f"     <STATENAME>{state_xml}</STATENAME>")
            if _cmp_state:
                a(f"     <PLACEOFSUPPLY>{_cmp_state}</PLACEOFSUPPLY>")
            _dn_reg = "Regular" if gstin_raw else "Unregistered"
            a(f"     <GSTREGISTRATIONTYPE>{_dn_reg}</GSTREGISTRATIONTYPE>")
            if gstin_raw:
                a("     <VATDEALERTYPE>Regular</VATDEALERTYPE>")
            if _cmp_gstin and _cmp_name:
                a(f'     <GSTREGISTRATION TAXTYPE="GST" TAXREGISTRATION="{_cmp_gstin}">{_cmp_name}</GSTREGISTRATION>')
                a(f'     <CMPGSTIN>{_cmp_gstin}</CMPGSTIN>')
                a('     <CMPGSTREGISTRATIONTYPE>Regular</CMPGSTREGISTRATIONTYPE>')
            if _cmp_state:
                a(f'     <CMPGSTSTATE>{_cmp_state}</CMPGSTSTATE>')

        a("     <COUNTRYOFRESIDENCE>India</COUNTRYOFRESIDENCE>")
        a(f"     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>")
        a("     <ISINVOICE>Yes</ISINVOICE>")
        a("     <PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW>")
        a(f"     <VCHENTRYMODE>{'Item Invoice' if is_item_mode else 'Accounting Invoice'}</VCHENTRYMODE>")
        a("     <NUMBERINGSTYLE>Manual</NUMBERINGSTYLE>")
        a("     <ISGSTOVERRIDDEN>No</ISGSTOVERRIDDEN>")
        a("     <VCHSTATUSISREACCEPHSNSIXONEDONE>Yes</VCHSTATUSISREACCEPHSNSIXONEDONE>")
        a("     <VCHGSTSTATUSISUNCERTAIN>No</VCHGSTSTATUSISUNCERTAIN>")
        a("     <VCHGSTSTATUSISINCLUDED>Yes</VCHGSTSTATUSISINCLUDED>")
        a("     <VCHGSTSTATUSISAPPLICABLE>Yes</VCHGSTSTATUSISAPPLICABLE>")
        if is_debit_note:
            _dn_gst_txn = "Tax Invoice" if gstin_raw else "Unregistered"
            a(f"     <GSTTRANSACTIONTYPE>{_dn_gst_txn}</GSTTRANSACTIONTYPE>")
        if gstin:
            a(f"     <PARTYGSTIN>{gstin}</PARTYGSTIN>")
        if narration:
            a(f"     <NARRATION>{narration}</NARRATION>")

        # Party ledger entry
        a("     <LEDGERENTRIES.LIST>")
        a(f"      <LEDGERNAME>{party}</LEDGERNAME>")
        a(f"      <ISDEEMEDPOSITIVE>{party_is_deemed_positive}</ISDEEMEDPOSITIVE>")
        a(f"      <AMOUNT>{fmt_amt(party_amount)}</AMOUNT>")
        a("     </LEDGERENTRIES.LIST>")

        if is_item_mode:
            inv_is_deemed = counter_is_deemed_positive
            merged_items = r.get("items")
            if merged_items:
                # Multi-item consolidated voucher
                for it in merged_items:
                    it_name = xml_escape(str(it.get("ItemName") or "").strip())
                    if not it_name:
                        continue
                    it_qty = float(it.get("Quantity") or 1.0)
                    it_rate = float(it.get("Rate") or 0.0)
                    it_per = xml_escape(_normalize_stock_unit_name(str(it.get("Per") or "Nos")) or "Nos")
                    it_godown = xml_escape(str(it.get("GodownName") or "Main Location"))
                    it_particular = xml_escape(str(it.get("Particular") or particular_raw) or default_particular_ledger)
                    it_tv = float(it.get("TaxableValue") or 0.0)
                    it_amt_base = round(it_qty * it_rate, 2) if it_qty and it_rate else it_tv
                    it_inv_amount = -it_amt_base if inv_is_deemed == "Yes" else it_amt_base
                    a("     <ALLINVENTORYENTRIES.LIST>")
                    a(f"      <STOCKITEMNAME>{it_name}</STOCKITEMNAME>")
                    a(f"      <ISDEEMEDPOSITIVE>{inv_is_deemed}</ISDEEMEDPOSITIVE>")
                    a(f"      <RATE>{fmt_amt(it_rate)}/{it_per}</RATE>")
                    a(f"      <AMOUNT>{fmt_amt(it_inv_amount)}</AMOUNT>")
                    a(f"      <ACTUALQTY>{fmt_amt(it_qty)} {it_per}</ACTUALQTY>")
                    a(f"      <BILLEDQTY>{fmt_amt(it_qty)} {it_per}</BILLEDQTY>")
                    a("      <BATCHALLOCATIONS.LIST>")
                    a(f"       <GODOWNNAME>{it_godown}</GODOWNNAME>")
                    a(f"       <AMOUNT>{fmt_amt(it_inv_amount)}</AMOUNT>")
                    a(f"       <ACTUALQTY>{fmt_amt(it_qty)} {it_per}</ACTUALQTY>")
                    a(f"       <BILLEDQTY>{fmt_amt(it_qty)} {it_per}</BILLEDQTY>")
                    a("      </BATCHALLOCATIONS.LIST>")
                    a("      <ACCOUNTINGALLOCATIONS.LIST>")
                    a(f"       <LEDGERNAME>{it_particular}</LEDGERNAME>")
                    a(f"       <ISDEEMEDPOSITIVE>{inv_is_deemed}</ISDEEMEDPOSITIVE>")
                    a(f"       <AMOUNT>{fmt_amt(it_inv_amount)}</AMOUNT>")
                    a("      </ACCOUNTINGALLOCATIONS.LIST>")
                    a("     </ALLINVENTORYENTRIES.LIST>")
            else:
                # Single item (standard path)
                inv_amount = -item_amt if inv_is_deemed == "Yes" else item_amt
                a("     <ALLINVENTORYENTRIES.LIST>")
                a(f"      <STOCKITEMNAME>{item_name}</STOCKITEMNAME>")
                a(f"      <ISDEEMEDPOSITIVE>{inv_is_deemed}</ISDEEMEDPOSITIVE>")
                a(f"      <RATE>{fmt_amt(rate)}/{per_unit}</RATE>")
                a(f"      <AMOUNT>{fmt_amt(inv_amount)}</AMOUNT>")
                a(f"      <ACTUALQTY>{fmt_amt(qty)} {per_unit}</ACTUALQTY>")
                a(f"      <BILLEDQTY>{fmt_amt(qty)} {per_unit}</BILLEDQTY>")
                a("      <BATCHALLOCATIONS.LIST>")
                a(f"       <GODOWNNAME>{godown}</GODOWNNAME>")
                a(f"       <AMOUNT>{fmt_amt(inv_amount)}</AMOUNT>")
                a(f"       <ACTUALQTY>{fmt_amt(qty)} {per_unit}</ACTUALQTY>")
                a(f"       <BILLEDQTY>{fmt_amt(qty)} {per_unit}</BILLEDQTY>")
                a("      </BATCHALLOCATIONS.LIST>")
                a("      <ACCOUNTINGALLOCATIONS.LIST>")
                a(f"       <LEDGERNAME>{particular}</LEDGERNAME>")
                a(f"       <ISDEEMEDPOSITIVE>{inv_is_deemed}</ISDEEMEDPOSITIVE>")
                a(f"       <AMOUNT>{fmt_amt(inv_amount)}</AMOUNT>")
                a("      </ACCOUNTINGALLOCATIONS.LIST>")
                a("     </ALLINVENTORYENTRIES.LIST>")
        else:
            # Particular / income ledger — handle multi-particular merged rows
            acct_entries = r.get("accounting_entries")
            if acct_entries and len(acct_entries) > 1:
                for ae in acct_entries:
                    ae_particular = xml_escape(str(ae.get("Particular") or particular_raw) or default_particular_ledger)
                    ae_tv = float(ae.get("TaxableValue") or 0.0)
                    ae_amount = ae_tv if is_debit_note else -ae_tv
                    a("     <LEDGERENTRIES.LIST>")
                    a(f"      <LEDGERNAME>{ae_particular}</LEDGERNAME>")
                    a(f"      <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
                    a(f"      <AMOUNT>{fmt_amt(ae_amount)}</AMOUNT>")
                    a("     </LEDGERENTRIES.LIST>")
            else:
                a("     <LEDGERENTRIES.LIST>")
                a(f"      <LEDGERNAME>{particular}</LEDGERNAME>")
                a(f"      <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
                a(f"      <AMOUNT>{fmt_amt(taxable_amount)}</AMOUNT>")
                a("     </LEDGERENTRIES.LIST>")

        if cgst_amt > 0 and cgst_ledger_raw:
            a("     <LEDGERENTRIES.LIST>")
            a(f"      <LEDGERNAME>{xml_escape(cgst_ledger_raw)}</LEDGERNAME>")
            a(f"      <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
            a(f"      <AMOUNT>{fmt_amt(cgst_amount)}</AMOUNT>")
            a("     </LEDGERENTRIES.LIST>")

        if sgst_amt > 0 and sgst_ledger_raw:
            a("     <LEDGERENTRIES.LIST>")
            a(f"      <LEDGERNAME>{xml_escape(sgst_ledger_raw)}</LEDGERNAME>")
            a(f"      <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
            a(f"      <AMOUNT>{fmt_amt(sgst_amount)}</AMOUNT>")
            a("     </LEDGERENTRIES.LIST>")

        if igst_amt > 0 and igst_ledger_raw:
            a("     <LEDGERENTRIES.LIST>")
            a(f"      <LEDGERNAME>{xml_escape(igst_ledger_raw)}</LEDGERNAME>")
            a(f"      <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
            a(f"      <AMOUNT>{fmt_amt(igst_amount)}</AMOUNT>")
            a("     </LEDGERENTRIES.LIST>")

        _append_round_off_entry(a, round_off_ledger, _ro_amt_n, is_purchase_mode=is_debit_note)

        a("    </VOUCHER>")
        a("   </TALLYMESSAGE>")

    a("  </REQUESTDATA>")
    a(" </IMPORTDATA></BODY>")
    a("</ENVELOPE>")
    return "\n".join(lines_out), voucher_count


# ═══════════════════════════════════════════════════════════════════════════
#  READ EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def read_excel(filepath: str, sheet: str = None) -> tuple:
    """Returns (headers: list, rows: list[dict])"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    header_rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    first_row = next(header_rows, None) or []
    headers = [str(c or "").strip() for c in first_row]
    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        vals = list(vals[:len(headers)])
        if len(vals) < len(headers):
            vals.extend([None] * (len(headers) - len(vals)))
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    wb.close()
    return headers, rows


# ═══════════════════════════════════════════════════════════════════════════
#  GST PORTAL SEARCHER  (reuses logic from user.py)
# ═══════════════════════════════════════════════════════════════════════════

_GST_SEARCH_URL = "https://services.gst.gov.in/services/searchtp"
_GST_MOBILE_UA = (
    "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Mobile Safari/537.36"
)
_GST_DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
)


def _pan_from_gstin(gstin: str) -> str:
    """Extract 10-character PAN from a 15-character GSTIN (chars 2–11 inclusive)."""
    g = (gstin or "").strip().upper()
    return g[2:12] if len(g) == 15 else ""


def _parse_portal_data_to_ledger_def(portal_data: dict, ledger_name: str, parent: str) -> dict:
    """Convert a GST portal search-result dict into a ledger definition dict."""
    details = portal_data.get("details") or {}
    key_details = details.get("key_details") or {}

    trade_name = str(key_details.get("Trade Name") or "").strip()
    legal_name = str(key_details.get("Legal Name of Business") or "").strip()
    mailing_name = trade_name or legal_name or ledger_name

    ppob = str(key_details.get("Principal Place of Business") or "").strip()
    address1 = ""
    address2 = ""
    state_raw = ""
    pincode_raw = ""

    if ppob:
        parts = [p.strip() for p in ppob.split(",") if p.strip()]
        # last part is pincode if numeric
        if parts and re.fullmatch(r"\d{6}", parts[-1]):
            pincode_raw = parts.pop()
        # second-to-last before pincode is state (e.g. "Delhi")
        if parts:
            state_raw = parts[-1]
        # everything else is address
        if len(parts) > 1:
            mid = parts[:-1]
            address1 = ", ".join(mid[:3])
            address2 = ", ".join(mid[3:])
        elif parts:
            address1 = parts[0]

    taxpayer_type = str(key_details.get("Taxpayer Type") or "").strip()
    gstin_val = str(portal_data.get("gstin") or "").strip().upper()

    reg_type_map = {
        "regular": "Regular",
        "composition": "Composition",
        "unregistered": "Unregistered/Consumer",
        "consumer": "Unregistered/Consumer",
        "input service distributor": "Input Service Distributor",
        "isd": "Input Service Distributor",
        "sez unit": "SEZ",
        "sez developer": "SEZ",
        "embassy / consulate": "Overseas",
        "overseas": "Overseas",
    }
    reg_type = reg_type_map.get(taxpayer_type.lower(), "Regular" if gstin_val else "Unregistered/Consumer")

    normalized_state = _normalize_state_for_ledger(state_raw)
    if not normalized_state and gstin_val:
        normalized_state = _state_name_from_gstin(gstin_val)

    gst_applicable = "Applicable" if gstin_val else "Not Applicable"

    return {
        "Name": ledger_name,
        "Parent": parent,
        "MailingName": mailing_name,
        "GSTIN": gstin_val,
        "PAN": _pan_from_gstin(gstin_val),
        "GSTApplicable": gst_applicable,
        "GSTRegistrationType": reg_type,
        "StateOfSupply": normalized_state,
        "Address1": address1,
        "Address2": address2,
        "Pincode": pincode_raw,
        "Country": "India",
        "Billwise": "Yes",
        "TypeOfTaxation": "",
        "GSTRate": "",
    }


class _GSTPortalSearcher:
    """Minimal GST portal searcher (captcha + form submit). Mirrors user.py logic."""

    def __init__(self):
        self.driver = None
        self.session = self._build_session() if _REQUESTS_AVAILABLE else None
        self.mfp = ""
        self.device_id = ""
        self.captcha_png_bytes = b""

    def _build_session(self):
        session = requests.Session()
        retry = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504],
                      allowed_methods=frozenset(["GET", "POST"]))
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        session.headers.update({"User-Agent": _GST_DEFAULT_UA,
                                 "Accept": "application/json, text/plain, */*"})
        return session

    def _ensure_driver(self):
        if self.driver is not None:
            return self.driver
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument(f"--user-agent={_GST_MOBILE_UA}")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--window-size=1200,900")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--lang=en-US")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        svc = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=svc, options=options)
        self.driver.set_page_load_timeout(60)
        try:
            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"})
        except Exception:
            pass
        return self.driver

    def _trigger_events(self, driver, element):
        try:
            driver.execute_script(
                "const e=arguments[0];e.dispatchEvent(new Event('input',{bubbles:true}));"
                "e.dispatchEvent(new Event('change',{bubbles:true}));"
                "e.dispatchEvent(new Event('blur',{bubbles:true}));", element)
        except Exception:
            pass

    def _find_ready_captcha(self, driver):
        try:
            el = driver.find_element(By.ID, "imgCaptcha")
            src = (el.get_attribute("src") or "").lower()
            if el.is_displayed() and "captcha" in src:
                return el
        except Exception:
            pass
        return False

    def _capture_captcha_bytes(self, driver, el) -> bytes:
        try:
            WebDriverWait(driver, 15).until(
                lambda d: d.execute_script(
                    "const i=arguments[0];return i&&i.complete&&i.naturalWidth>0;", el))
        except Exception:
            pass
        try:
            png = el.screenshot_as_png
            if png and len(png) > 200:
                return png
        except Exception:
            pass
        return b""

    def _sync_cookies(self):
        if not self.driver or not self.session:
            return
        self.session.cookies.clear()
        for c in self.driver.get_cookies():
            if c.get("name"):
                self.session.cookies.set(
                    name=c["name"], value=c.get("value", ""),
                    path=c.get("path", "/"), domain=c.get("domain"))

    def _find_captcha_input(self, driver):
        locators = [(By.ID, "captcha"), (By.ID, "captchaCode"), (By.ID, "captchaText"),
                    (By.NAME, "captcha"), (By.NAME, "captchaCode"),
                    (By.CSS_SELECTOR, "input[ng-model*='captcha' i]"),
                    (By.CSS_SELECTOR, "input[id*='captcha' i]")]
        for by, val in locators:
            try:
                el = driver.find_element(by, val)
                if el.is_displayed():
                    return el
            except Exception:
                pass
        return None

    def load_captcha(self, gstin: str) -> bytes:
        """Navigate to GST search page and return PNG bytes of captcha."""
        if not _SELENIUM_AVAILABLE:
            raise RuntimeError("Selenium not installed. Cannot load captcha.")
        driver = self._ensure_driver()
        driver.get(_GST_SEARCH_URL)
        wait = WebDriverWait(driver, 35)
        try:
            inp = wait.until(EC.element_to_be_clickable((By.ID, "for_gstin")))
            inp.clear()
            inp.send_keys(gstin)
            self._trigger_events(driver, inp)
            el = wait.until(self._find_ready_captcha)
        except TimeoutException:
            raise RuntimeError("Timed out waiting for captcha on GST portal.")
        self.captcha_png_bytes = self._capture_captcha_bytes(driver, el)
        self._sync_cookies()
        return self.captcha_png_bytes

    def fetch(self, gstin: str, captcha_text: str) -> dict:
        """Submit form and return portal data dict."""
        if not _SELENIUM_AVAILABLE:
            raise RuntimeError("Selenium not installed.")
        driver = self._ensure_driver()
        wait = WebDriverWait(driver, 25)
        if "searchtp" not in (driver.current_url or "").lower():
            driver.get(_GST_SEARCH_URL)

        inp = wait.until(EC.element_to_be_clickable((By.ID, "for_gstin")))
        inp.clear()
        inp.send_keys(gstin)
        self._trigger_events(driver, inp)

        cap_inp = None
        for _ in range(40):
            cap_inp = self._find_captcha_input(driver)
            if cap_inp:
                break
            time.sleep(0.3)
        if not cap_inp:
            raise RuntimeError("Captcha input not found on page.")
        cap_inp.clear()
        cap_inp.send_keys(captcha_text)

        btn = wait.until(EC.element_to_be_clickable((By.ID, "lotsearch")))
        btn.click()

        # Wait for result
        end = time.time() + 35
        while time.time() < end:
            for sel in ["#lottable", "#searchResult", ".panel-body", "table"]:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el.text.strip():
                        break
                except Exception:
                    pass
            else:
                time.sleep(0.5)
                continue
            break

        details = driver.execute_script("""
            const getText=e=>e?(e.textContent||'').replace(/\\s+/g,' ').trim():'';
            const result={};const container=document.querySelector('#lottable');
            if(!container)return result;
            const pairs={};
            const cols=container.querySelectorAll('.tbl-format .inner .col-sm-4,.tbl-format .inner .col-sm-12');
            cols.forEach(col=>{
                const label=getText(col.querySelector('strong'));if(!label)return;
                const list=col.querySelector('ul.jurisdictList');
                if(list){const items=Array.from(list.querySelectorAll('li')).map(getText).filter(Boolean);
                    if(items.length){pairs[label]=items;return;}}
                const word=col.querySelector('.wordCls');
                if(word){pairs[label]=getText(word);return;}
                const ps=Array.from(col.querySelectorAll('p')).map(getText).filter(Boolean);
                if(ps.length>1)pairs[label]=ps[ps.length-1];
                else if(ps.length===1)pairs[label]=ps[0];
            });
            result.key_details=pairs;
            return result;
        """) or {}

        return {"gstin": gstin, "fetchedAt": datetime.now().isoformat(), "details": details}

    def quit(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


# ───────────────────────────────────────────────────────────────────────────
#  GST Fetch Dialog  – shown when user clicks "Fetch from Portal"
# ───────────────────────────────────────────────────────────────────────────

class _GSTFetchDialog(ctk.CTkToplevel):
    """Popup for fetching party details from the GST portal."""

    def __init__(self, parent, ledger_name: str, gstin_hint: str, ledger_parent: str):
        super().__init__(parent)
        self.title(f"Fetch from GST Portal — {ledger_name}")
        self.geometry("560x620")
        self.resizable(False, False)
        self.grab_set()
        self.lift()
        self.focus_force()

        self._ledger_name = ledger_name
        self._ledger_parent = ledger_parent
        self._searcher = _GSTPortalSearcher()
        self._captcha_photo = None
        self._portal_data = None
        self._result_def = None   # set when user confirms

        self._gstin_var = tk.StringVar(value=gstin_hint)
        self._captcha_var = tk.StringVar()
        self._status_var = tk.StringVar(value="Enter GSTIN, click Load Captcha.")
        self._name_var = tk.StringVar()
        self._state_var = tk.StringVar()
        self._addr_var = tk.StringVar()
        self._pin_var = tk.StringVar()
        self._reg_type_var = tk.StringVar()

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

    def _build_ui(self):
        pad = {"padx": 14, "pady": 6}
        ctk.CTkLabel(self, text="GSTIN / UIN", anchor="w").pack(fill="x", **pad)
        gstin_row = ctk.CTkFrame(self, fg_color="transparent")
        gstin_row.pack(fill="x", padx=14, pady=(0, 6))
        ctk.CTkEntry(gstin_row, textvariable=self._gstin_var, width=300).pack(side="left", padx=(0, 8))
        ctk.CTkButton(gstin_row, text="Load Captcha", width=130,
                      command=self._on_load_captcha).pack(side="left")

        ctk.CTkLabel(self, text="Captcha Image", anchor="w").pack(fill="x", **pad)
        self._captcha_lbl = ctk.CTkLabel(self, text="— captcha not loaded —",
                                          width=360, height=80, fg_color=("gray90", "gray20"),
                                          corner_radius=6)
        self._captcha_lbl.pack(padx=14, pady=(0, 6))

        ctk.CTkLabel(self, text="Enter Captcha Text", anchor="w").pack(fill="x", **pad)
        cap_row = ctk.CTkFrame(self, fg_color="transparent")
        cap_row.pack(fill="x", padx=14, pady=(0, 6))
        ctk.CTkEntry(cap_row, textvariable=self._captcha_var, width=200).pack(side="left", padx=(0, 8))
        ctk.CTkButton(cap_row, text="Fetch Details", width=130,
                      command=self._on_fetch).pack(side="left")

        ctk.CTkLabel(self, textvariable=self._status_var, anchor="w",
                      text_color=("gray40", "gray70"), font=("Segoe UI", 11)).pack(
            fill="x", padx=14, pady=(0, 8))

        # Fetched data preview (editable)
        sep = ctk.CTkFrame(self, height=1, fg_color=("gray80", "gray30"))
        sep.pack(fill="x", padx=14, pady=(0, 8))
        ctk.CTkLabel(self, text="Fetched Details (editable before creating):",
                      anchor="w", font=("Segoe UI", 11, "bold")).pack(fill="x", padx=14, pady=(0, 4))

        for label, var in [
            ("Mailing Name", self._name_var),
            ("GST Registration Type", self._reg_type_var),
            ("State", self._state_var),
            ("Address", self._addr_var),
            ("Pincode", self._pin_var),
        ]:
            row = ctk.CTkFrame(self, fg_color="transparent")
            row.pack(fill="x", padx=14, pady=2)
            ctk.CTkLabel(row, text=label, width=190, anchor="w").pack(side="left")
            ctk.CTkEntry(row, textvariable=var, width=300).pack(side="left")

        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=14, pady=(12, 8))
        self._confirm_btn = ctk.CTkButton(btn_row, text="Create Ledger with Portal Data",
                                           fg_color=("#059669", "#10B981"),
                                           command=self._on_confirm, state="disabled")
        self._confirm_btn.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkButton(btn_row, text="Cancel", fg_color=("gray60", "gray30"),
                       command=self._on_cancel, width=90).pack(side="left")

    def _on_load_captcha(self):
        gstin = self._gstin_var.get().strip().upper()
        if not gstin:
            self._status_var.set("Enter GSTIN first.")
            return
        if not _SELENIUM_AVAILABLE:
            self._status_var.set("Selenium not installed — cannot load captcha.")
            return
        self._status_var.set("Loading captcha (opening browser)...")
        self._captcha_lbl.configure(text="Loading…")
        threading.Thread(target=self._bg_load_captcha, args=(gstin,), daemon=True).start()

    def _bg_load_captcha(self, gstin):
        try:
            png = self._searcher.load_captcha(gstin)
            self.after(0, self._show_captcha, png)
            self.after(0, self._status_var.set, "Captcha loaded. Enter text and click Fetch Details.")
        except Exception as exc:
            self.after(0, self._status_var.set, f"Error: {exc}")
            self.after(0, lambda: self._captcha_lbl.configure(text="Failed to load captcha"))

    def _show_captcha(self, png_bytes):
        if not _PIL_AVAILABLE or not png_bytes:
            self._captcha_lbl.configure(text="Captcha loaded (PIL not available to render image)")
            return
        try:
            img = Image.open(io.BytesIO(png_bytes))
            img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS if hasattr(Image, 'LANCZOS') else Image.Resampling.LANCZOS)
            self._captcha_photo = ImageTk.PhotoImage(img)
            self._captcha_lbl.configure(image=self._captcha_photo, text="")
        except Exception:
            self._captcha_lbl.configure(text="Captcha loaded (render error)")

    def _on_fetch(self):
        gstin = self._gstin_var.get().strip().upper()
        captcha_text = self._captcha_var.get().strip()
        if not gstin:
            self._status_var.set("GSTIN is required.")
            return
        if not captcha_text:
            self._status_var.set("Enter captcha text first.")
            return
        self._status_var.set("Fetching from GST portal...")
        threading.Thread(target=self._bg_fetch, args=(gstin, captcha_text), daemon=True).start()

    def _bg_fetch(self, gstin, captcha_text):
        try:
            data = self._searcher.fetch(gstin, captcha_text)
            self._portal_data = data
            led_def = _parse_portal_data_to_ledger_def(data, self._ledger_name, self._ledger_parent)
            self.after(0, self._populate_fields, led_def)
            self.after(0, self._status_var.set, "Details fetched. Review and click Create Ledger.")
        except Exception as exc:
            self.after(0, self._status_var.set, f"Fetch failed: {exc}")

    def _populate_fields(self, led_def: dict):
        self._name_var.set(led_def.get("MailingName", "") or led_def.get("Name", ""))
        self._reg_type_var.set(led_def.get("GSTRegistrationType", ""))
        self._state_var.set(led_def.get("StateOfSupply", ""))
        self._addr_var.set(
            ", ".join(filter(None, [led_def.get("Address1", ""), led_def.get("Address2", "")]))
        )
        self._pin_var.set(led_def.get("Pincode", ""))
        self._confirm_btn.configure(state="normal")

    def _on_confirm(self):
        # Build enriched ledger def from (possibly edited) fields
        addr_full = self._addr_var.get().strip()
        addr_parts = [p.strip() for p in addr_full.split(",") if p.strip()]
        led_def = {
            "Name": self._ledger_name,
            "Parent": self._ledger_parent,
            "MailingName": self._name_var.get().strip() or self._ledger_name,
            "GSTIN": self._gstin_var.get().strip().upper(),
            "GSTApplicable": "Applicable" if self._gstin_var.get().strip() else "Not Applicable",
            "GSTRegistrationType": self._reg_type_var.get().strip() or "Regular",
            "StateOfSupply": _normalize_state_for_ledger(self._state_var.get().strip()),
            "Address1": addr_parts[0] if addr_parts else "",
            "Address2": ", ".join(addr_parts[1:]) if len(addr_parts) > 1 else "",
            "Pincode": self._pin_var.get().strip(),
            "Country": "India",
            "Billwise": "Yes",
            "TypeOfTaxation": "",
            "GSTRate": "",
        }
        self._result_def = led_def
        self._searcher.quit()
        self.destroy()

    def _on_cancel(self):
        self._result_def = None
        self._searcher.quit()
        self.destroy()

    def get_result(self):
        return self._result_def


# ───────────────────────────────────────────────────────────────────────────
#  Missing Ledger Dialog  – shown when a party ledger needs to be created
# ───────────────────────────────────────────────────────────────────────────

class _MissingLedgerDialog(ctk.CTkToplevel):
    """
    Popup: "Ledger X not found — Fetch from Portal or Create Manually?"
    Called from main thread. Sets choice_data["choice"] and fires event.
    """

    def __init__(self, parent, led_def: dict, choice_data: dict, event: threading.Event):
        super().__init__(parent)
        self._led_def = led_def
        self._choice_data = choice_data
        self._event = event

        ledger_name = str(led_def.get("Name", "") or "Unknown")
        gstin_hint = str(led_def.get("GSTIN", "") or "")
        parent_account = str(led_def.get("Parent", "") or "Sundry Debtors")

        self.title("Missing Ledger")
        self.geometry("500x280")
        self.resizable(False, False)
        self.grab_set()
        self.lift()
        self.focus_force()
        self.protocol("WM_DELETE_WINDOW", self._choose_manual)

        # Header
        header = ctk.CTkFrame(self, fg_color=("#FEF3C7", "#78350F"), corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(header, text="⚠  Ledger Not Found in Tally",
                      font=("Segoe UI", 13, "bold"),
                      text_color=("#92400E", "#FCD34D")).pack(padx=16, pady=10)

        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=20, pady=12)

        ctk.CTkLabel(body,
                      text=f"The ledger  \"{ledger_name}\"  was not found in Tally.",
                      font=("Segoe UI", 12), wraplength=440, justify="left").pack(anchor="w")
        ctk.CTkLabel(body,
                      text="How would you like to create it?",
                      font=("Segoe UI", 11), text_color=("gray50", "gray60")).pack(
            anchor="w", pady=(4, 0))

        if gstin_hint:
            ctk.CTkLabel(body,
                          text=f"GSTIN: {gstin_hint}",
                          font=("Segoe UI", 10, "italic"),
                          text_color=("gray50", "gray60")).pack(anchor="w")

        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 16))

        portal_enabled = _SELENIUM_AVAILABLE and _PIL_AVAILABLE
        portal_tooltip = "" if portal_enabled else " (Selenium/PIL not installed)"

        ctk.CTkButton(
            btn_frame,
            text=f"🌐  Fetch from GST Portal{portal_tooltip}",
            fg_color=("#2563EB", "#3B82F6"),
            hover_color=("#1D4ED8", "#2563EB"),
            height=44,
            font=("Segoe UI", 12),
            command=lambda: self._choose_fetch(ledger_name, gstin_hint, parent_account),
            state="normal" if portal_enabled else "disabled",
        ).pack(fill="x", pady=(0, 8))

        ctk.CTkButton(
            btn_frame,
            text="✏  Create Manually (Basic Info Only)",
            fg_color=("gray70", "gray30"),
            hover_color=("gray60", "gray40"),
            height=44,
            font=("Segoe UI", 12),
            command=self._choose_manual,
        ).pack(fill="x")

    def _choose_fetch(self, ledger_name, gstin_hint, parent_account):
        self.withdraw()
        dlg = _GSTFetchDialog(self.master, ledger_name, gstin_hint, parent_account)
        self.wait_window(dlg)
        result_def = dlg.get_result()
        if result_def:
            self._choice_data["choice"] = "fetch"
            self._choice_data["def"] = result_def
        else:
            # user cancelled the fetch dialog → fall back to manual
            self._choice_data["choice"] = "manual"
            self._choice_data["def"] = self._led_def
        self._event.set()
        self.destroy()

    def _choose_manual(self):
        self._choice_data["choice"] = "manual"
        self._choice_data["def"] = self._led_def
        self._event.set()
        self.destroy()


def _ask_create_party_ledger(widget, led_def: dict) -> dict:
    """
    Call from a worker thread. Shows _MissingLedgerDialog on the main thread,
    waits for user choice (up to 5 min), and returns the (possibly enriched)
    ledger definition dict.
    """
    event = threading.Event()
    choice_data = {"choice": None, "def": led_def}

    def _show():
        dlg = _MissingLedgerDialog(widget, led_def, choice_data, event)

    widget.after(0, _show)
    event.wait(timeout=300)

    if choice_data["choice"] is None:
        # Timeout – fall back to manual
        choice_data["def"] = led_def
    return choice_data["def"]


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════════════════

class TallySalesApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("TallySalesPro — Sales Voucher & Master Creator")
        self.geometry("1120x720")
        self.minsize(960, 640)
        self.configure(fg_color=COLORS["bg_dark"])

        self.loaded_rows = []
        self.loaded_headers = []
        self.file_path_var = ctk.StringVar()
        self.company_placeholder = "Auto (Loaded Company)"
        self.company_var = ctk.StringVar(value=self.company_placeholder)
        self.tally_host_var = ctk.StringVar(value="localhost")
        self.tally_port_var = ctk.StringVar(value="9000")
        self.voucher_date_mode_var = ctk.StringVar(value="excel")
        self.voucher_custom_date_var = ctk.StringVar(value="")
        self.voucher_date_checks = {
            "current": ctk.BooleanVar(value=False),
            "excel": ctk.BooleanVar(value=True),
            "custom": ctk.BooleanVar(value=False),
        }
        self.status_var = ctk.StringVar(value="Ready")
        self.connection_status_var = ctk.StringVar(value="Connection: Not checked")
        self.company_status_var = ctk.StringVar(value="Companies: Not fetched")
        self.fetched_companies = []
        self._company_fetch_running = False
        self.debug_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tally_push_debug.log")
        self._voucher_info_labels = {}
        self._loaded_rows_by_mode = {}
        self._loaded_headers_by_mode = {}
        self._voucher_load_running = {}
        self._voucher_browse_buttons = {}
        self._voucher_save_buttons = {}
        self._voucher_push_buttons = {}
        self._voucher_template_buttons = {}
        self._voucher_xml_fp_vars = {}
        self._voucher_xml_browse_buttons = {}
        self._voucher_import_buttons = {}  # kept for compat, not used in UI
        self._active_preview_mode = {}     # tracks "excel" or "xml" per mode
        self._voucher_type_vars = {}        # StringVar per mode for selected voucher type
        self._voucher_type_cbs = {}         # CTkComboBox per mode
        self._voucher_type_fetch_buttons = {}  # Fetch button per mode
        self._note_type_var = ctk.StringVar(value="Credit Note (Accounting)")
        self._note_entry_mode = "accounting"
        self.company_gst_registrations = []
        self._note_loaded_rows = []
        self._note_loaded_headers = []
        self.fetched_party_ledgers = []
        self._ledger_fetch_running = False
        # Journal Entry instance variables
        self._jnl_loaded_rows = []
        self._jnl_loaded_headers = []
        self._jnl_manual_rows = []
        self._jnl_manual_form_vars = {}
        self._jnl_manual_party_ledger_combo = None
        self._jnl_manual_party_search_var = ctk.StringVar(value="")
        self._jnl_manual_party_match_label = None
        self._jnl_manual_fetch_ledger_btn = None
        self._jnl_manual_tree = None
        self._jnl_excel_tree = None
        self._jnl_manual_editing_index = None
        self._jnl_manual_update_btn = None
        self._jnl_source_tabs = None
        self._jnl_manual_info_label = None
        self._jnl_excel_info_label = None
        self._journal_type_var = ctk.StringVar(value="purchase")
        self._jnl_manual_party_search_clear_btn = None
        # Note panel manual entry vars
        self._note_manual_rows = []
        self._note_manual_form_vars = {}
        self._note_manual_party_ledger_combo = None
        self._note_manual_party_search_var = ctk.StringVar(value="")
        self._note_manual_party_match_label = None
        self._note_manual_fetch_ledger_btn = None
        self._note_manual_tree = None
        self._note_excel_tree = None
        self._note_manual_editing_index = None
        self._note_manual_update_btn = None
        self._note_source_tabs = None
        self._note_manual_info_label = None
        self._note_excel_info_label = None
        self._note_manual_party_search_clear_btn = None
        self._push_running = False
        self._push_overlay = None
        self._push_message_var = ctk.StringVar(value="")
        self._roundoff_enabled_vars = {}   # BooleanVar per mode
        self._roundoff_ledger_vars = {}    # StringVar per mode
        self._roundoff_all_ledgers = {}    # cached ledger list per mode
        self._roundoff_fetch_btns = {}     # Fetch button per mode
        self._note_roundoff_enabled_var = ctk.BooleanVar(value=False)
        self._note_roundoff_ledger_var = ctk.StringVar(value="")
        self._note_roundoff_all_ledgers = []
        self._jnl_roundoff_enabled_var = ctk.BooleanVar(value=False)
        self._jnl_roundoff_ledger_var = ctk.StringVar(value="")
        self._jnl_roundoff_all_ledgers = []
        self.workflow_demo_url = ""
        self.demo_btn = None
        self.voucher_date_current_cb = None
        self.voucher_date_excel_cb = None
        self.voucher_date_custom_cb = None
        self.voucher_custom_date_entry = None
        self.template_names_by_mode = {
            "accounting": "Template_Sales Accounting Voucher.xlsx",
            "item": "Template_Sales Item Invoice.xlsx",
            "purchase_accounting": "Template_Purchase Accounting Voucher.xlsx",
            "purchase_item": "Template_Purchase Item Voucher.xlsx",
        }

        self._build_ui()

    def _resolve_theme_color(self, key: str):
        value = COLORS.get(key)
        if isinstance(value, tuple):
            mode = ctk.get_appearance_mode().lower()
            return value[1] if mode == "dark" else value[0]
        return value

    def _apply_ttk_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("default")
        except tk.TclError:
            pass

        tree_bg = self._resolve_theme_color("bg_card")
        field_bg = self._resolve_theme_color("bg_input")
        heading_bg = self._resolve_theme_color("table_header")
        text_fg = self._resolve_theme_color("text_primary")
        border_fg = self._resolve_theme_color("border")
        selected_bg = self._resolve_theme_color("accent")

        style.configure(
            "Treeview",
            background=tree_bg,
            foreground=text_fg,
            fieldbackground=field_bg,
            bordercolor=border_fg,
            font=("Segoe UI", 10),
            rowheight=26,
        )
        style.configure(
            "Treeview.Heading",
            background=heading_bg,
            foreground="#FFFFFF",
            font=("Segoe UI", 10, "bold"),
        )
        style.map(
            "Treeview",
            background=[("selected", selected_bg)],
            foreground=[("selected", "#FFFFFF")],
        )

    def set_theme(self, mode: str):
        try:
            ctk.set_appearance_mode(mode)
        except Exception:
            pass
        self._apply_ttk_styles()

    # ── UI BUILDING ─────────────────────────────────────────────────────

    def _build_ui(self):
        self._apply_ttk_styles()

        settings_card = ctk.CTkFrame(
            self,
            fg_color=COLORS["bg_card"],
            border_width=1,
            border_color=COLORS["border"],
            corner_radius=12,
        )
        settings_card.pack(fill="x", padx=16, pady=(10, 8))

        row_1 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_1.pack(fill="x", padx=14, pady=(12, 6))
        ctk.CTkLabel(row_1, text="Host", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        ctk.CTkEntry(
            row_1,
            textvariable=self.tally_host_var,
            width=140,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
        ).pack(side="left", padx=(6, 12))
        ctk.CTkLabel(row_1, text="Port", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        ctk.CTkEntry(
            row_1,
            textvariable=self.tally_port_var,
            width=90,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
        ).pack(side="left", padx=(6, 12))

        self.connection_test_btn = ctk.CTkButton(
            row_1,
            text="Test Connection",
            width=130,
            height=32,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["warning"],
            hover_color="#B45309",
            text_color="#FFFFFF",
            corner_radius=8,
            command=self._check_tally_connection_thread,
        )
        self.connection_test_btn.pack(side="right")

        self.demo_btn = ctk.CTkButton(
            row_1,
            text="▶ View Demo",
            width=132,
            height=32,
            font=("Segoe UI", 10, "bold"),
            fg_color="#DC2626",
            hover_color="#B91C1C",
            text_color="#FFFFFF",
            corner_radius=8,
            command=self._view_workflow_demo,
        )
        self.demo_btn.pack(side="right", padx=(0, 8))

        row_2 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_2.pack(fill="x", padx=14, pady=(0, 6))
        ctk.CTkLabel(row_2, text="Target Company", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        self.company_combo = ctk.CTkComboBox(
            row_2,
            values=[self.company_placeholder],
            variable=self.company_var,
            width=380,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.company_combo.set(self.company_placeholder)
        self.company_combo.pack(side="left", padx=(10, 8), fill="x", expand=True)
        self.company_refresh_btn = ctk.CTkButton(
            row_2,
            text="Refresh",
            width=96,
            height=34,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._fetch_tally_companies_thread,
        )
        self.company_refresh_btn.pack(side="right")

        status_row = ctk.CTkFrame(settings_card, fg_color="transparent")
        status_row.pack(fill="x", padx=14, pady=(0, 10))
        self.connection_status_label = ctk.CTkLabel(
            status_row,
            textvariable=self.connection_status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.connection_status_label.pack(side="left", padx=(0, 15))
        self.company_status_label = ctk.CTkLabel(
            status_row,
            textvariable=self.company_status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.company_status_label.pack(side="left", padx=(0, 15))

        date_mode_row = ctk.CTkFrame(settings_card, fg_color="transparent")
        date_mode_row.pack(fill="x", padx=14, pady=(0, 10))
        ctk.CTkLabel(
            date_mode_row,
            text="Voucher Date",
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
        ).pack(side="left")

        checks_wrap = ctk.CTkFrame(date_mode_row, fg_color="transparent")
        checks_wrap.pack(side="left", padx=(8, 0))

        self.voucher_date_current_cb = ctk.CTkCheckBox(
            checks_wrap,
            text="Current Date",
            variable=self.voucher_date_checks["current"],
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            border_color=COLORS["border"],
            command=lambda: self._set_voucher_date_mode("current"),
        )
        self.voucher_date_current_cb.pack(side="left", padx=(0, 8))

        self.voucher_date_excel_cb = ctk.CTkCheckBox(
            checks_wrap,
            text="Excel Date",
            variable=self.voucher_date_checks["excel"],
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            border_color=COLORS["border"],
            command=lambda: self._set_voucher_date_mode("excel"),
        )
        self.voucher_date_excel_cb.pack(side="left", padx=(0, 8))

        self.voucher_date_custom_cb = ctk.CTkCheckBox(
            checks_wrap,
            text="Custom Date",
            variable=self.voucher_date_checks["custom"],
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            border_color=COLORS["border"],
            command=lambda: self._set_voucher_date_mode("custom"),
        )
        self.voucher_date_custom_cb.pack(side="left")

        self.voucher_custom_date_entry = ctk.CTkEntry(
            checks_wrap,
            textvariable=self.voucher_custom_date_var,
            width=170,
            height=30,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="DD/MM/YYYY",
            font=("Segoe UI", 10),
        )
        self.voucher_custom_date_entry.pack(side="left", padx=(8, 0))
        self._set_voucher_date_mode("excel")

        # --- Shifted Select Mode into date_mode_row ---
        ctk.CTkLabel(date_mode_row, text="Select Mode", font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["text_secondary"]).pack(side="left", padx=(40, 10))

        _panel_options = [
            "📋  Sales Accounting Invoice",
            "📦  Sales Item Invoice",
            "🧾  Purchase Accounting Invoice",
            "🛒  Purchase Item Invoice",
            "📝  Credit Note (Accounting)",
            "📒  Debit Note (Accounting)",
            "📝  Credit Note (Item)",
            "📒  Debit Note (Item)",
            "📓  Journal Entry",
            "🏦  Create Ledgers",
            "📁  Create Stock Items",
        ]
        self._panel_option_to_key = {
            "📋  Sales Accounting Invoice": "accounting",
            "📦  Sales Item Invoice": "item",
            "🧾  Purchase Accounting Invoice": "purchase_accounting",
            "🛒  Purchase Item Invoice": "purchase_item",
            "📝  Credit Note (Accounting)": "credit_note_accounting",
            "📒  Debit Note (Accounting)": "debit_note_accounting",
            "📝  Credit Note (Item)": "credit_note_item",
            "📒  Debit Note (Item)": "debit_note_item",
            "📓  Journal Entry": "journal",
            "🏦  Create Ledgers": "ledger",
            "📁  Create Stock Items": "stock",
        }
        self._panel_var = ctk.StringVar(value=_panel_options[0])
        self._panel_option_menu = ctk.CTkOptionMenu(
            date_mode_row,
            variable=self._panel_var,
            values=_panel_options,
            width=300,
            height=34,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["accent"],
            button_color=COLORS["accent_hover"],
            button_hover_color=COLORS["accent_hover"],
            text_color="#FFFFFF",
            dropdown_fg_color=COLORS["bg_card"],
            dropdown_text_color=COLORS["text_primary"],
            dropdown_hover_color=COLORS["bg_input"],
            corner_radius=8,
            command=self._switch_panel,
        )
        self._panel_option_menu.pack(side="left", padx=(2, 0))



        # ─── Content container (all panels stacked in same grid cell) ────────────
        _content_outer = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=10,
                                       border_width=1, border_color=COLORS["border"])
        _content_outer.pack(fill="both", expand=True, padx=16, pady=(0, 10))
        _content_outer.grid_rowconfigure(0, weight=1)
        _content_outer.grid_columnconfigure(0, weight=1)

        self._panels = {}
        for _key in ("accounting", "item", "purchase_accounting", "purchase_item", "ledger", "stock"):
            _pf = ctk.CTkFrame(_content_outer, fg_color="transparent")
            _pf.grid(row=0, column=0, sticky="nsew")
            self._panels[_key] = _pf

        # Shared Credit/Debit Note panel
        _note_pf = ctk.CTkFrame(_content_outer, fg_color="transparent")
        _note_pf.grid(row=0, column=0, sticky="nsew")
        self._panels["credit_note"] = _note_pf
        self._panels["debit_note"] = _note_pf   # same frame — note_type_var controls type
        self._panels["credit_note_accounting"] = _note_pf
        self._panels["debit_note_accounting"] = _note_pf
        self._panels["credit_note_item"] = _note_pf
        self._panels["debit_note_item"] = _note_pf
        self._panels["note"] = _note_pf

        # Journal Entry panel
        journal_pf = ctk.CTkFrame(_content_outer, fg_color="transparent")
        journal_pf.grid(row=0, column=0, sticky="nsew")
        self._panels["journal"] = journal_pf

        self.tab_acct = self._panels["accounting"]
        self.tab_item = self._panels["item"]
        self.tab_purchase_acct = self._panels["purchase_accounting"]
        self.tab_purchase_item = self._panels["purchase_item"]
        self.tab_ledger = self._panels["ledger"]
        self.tab_stock = self._panels["stock"]

        self._build_voucher_tab(self.tab_acct, mode="accounting")
        self._build_voucher_tab(self.tab_item, mode="item")
        self._build_voucher_tab(self.tab_purchase_acct, mode="purchase_accounting")
        self._build_voucher_tab(self.tab_purchase_item, mode="purchase_item")
        self._build_note_panel(self._panels["note"])
        self._build_journal_panel(self._panels["journal"])
        self._build_ledger_tab()
        self._build_stock_tab()

        # Show first panel
        self._switch_panel(_panel_options[0])

        status_bar = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=0, height=32)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)
        ctk.CTkLabel(
            status_bar,
            textvariable=self.status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        ).pack(side="left", padx=16)

        self.after(200, lambda: self._fetch_tally_companies_thread(silent=True))

    def _get_tally_url(self):
        return _build_tally_url(self.tally_host_var.get(), self.tally_port_var.get())

    def _set_voucher_date_mode(self, selected_mode: str):
        mode = str(selected_mode or "excel").strip().lower()
        if mode not in {"current", "excel", "custom"}:
            mode = "excel"

        self.voucher_date_mode_var.set(mode)
        for key, var in self.voucher_date_checks.items():
            var.set(key == mode)

        if self.voucher_custom_date_entry is not None:
            self.voucher_custom_date_entry.configure(
                state="normal" if (mode == "custom" and not self._push_running) else "disabled"
            )

    def _get_voucher_date_selection(self):
        mode = str(self.voucher_date_mode_var.get() or "excel").strip().lower()
        if mode not in {"current", "excel", "custom"}:
            mode = "excel"
            self._set_voucher_date_mode(mode)

        custom_tally_date = ""
        if mode == "custom":
            custom_raw = (self.voucher_custom_date_var.get() or "").strip()
            if not custom_raw:
                raise ValueError("Enter custom date or select Current Date / Excel Date.")
            custom_tally_date = _normalize_manual_date_to_tally(custom_raw)

        return mode, custom_tally_date

    def _view_workflow_demo(self):
        demo_url = (self.workflow_demo_url or "").strip()
        if demo_url:
            try:
                opened = webbrowser.open(demo_url)
            except webbrowser.Error as exc:
                messagebox.showwarning("View Demo", f"Could not open demo link.\n\n{exc}")
                return
            if not opened:
                messagebox.showwarning("View Demo", "Could not open demo link in your default browser.")
            return

        messagebox.showinfo(
            "View Demo",
            "Demo link is not set yet.\n\nSet self.workflow_demo_url in code to your YouTube link later.",
        )

    def _append_debug_log(self, mode, target_company, xml_payload, response_text, parsed, note=""):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        lines = [
            "=" * 96,
            f"[{timestamp}] mode={mode} company={target_company or 'Loaded company in Tally'} note={note}",
            (
                "summary: "
                f"created={parsed.get('created', 0)} "
                f"altered={parsed.get('altered', 0)} "
                f"deleted={parsed.get('deleted', 0)} "
                f"ignored={parsed.get('ignored', 0)} "
                f"errors={parsed.get('errors', 0)} "
                f"exceptions={parsed.get('exceptions', 0)}"
            ),
        ]
        line_errors = parsed.get("line_errors") or []
        if line_errors:
            lines.append("line_errors:")
            for err in line_errors:
                lines.append(f"- {err}")
        if parsed.get("error"):
            lines.append(f"parsed_error: {parsed.get('error')}")

        lines.append("response:")
        lines.append(response_text[:12000])
        lines.append("xml:")
        lines.append(xml_payload[:12000])
        lines.append("\n")

        with open(self.debug_log_path, "a", encoding="utf-8") as log_file:
            log_file.write("\n".join(lines))

    def _get_selected_company(self):
        selected = _normalize_company_name(self.company_var.get())
        if not selected or _company_key(selected) == _company_key(self.company_placeholder):
            if len(getattr(self, "fetched_companies", [])) == 1:
                return self.fetched_companies[0]
            return ""
        return selected

    def _set_company_dropdown(self, companies, keep_selection=True):
        current = _normalize_company_name(self.company_var.get()) if keep_selection else ""
        cleaned = []
        seen = set()
        for name in companies or []:
            normalized = _normalize_company_name(name)
            if not _is_valid_company_name(normalized):
                continue
            key = _company_key(normalized)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(normalized)

        cleaned = sorted(cleaned, key=lambda x: _company_key(x))
        values = [self.company_placeholder] + cleaned
        self.company_combo.configure(values=values)

        if current and _company_key(current) in {_company_key(x) for x in cleaned}:
            self.company_combo.set(current)
            self.company_var.set(current)
        else:
            self.company_combo.set(self.company_placeholder)
            self.company_var.set(self.company_placeholder)

        self.fetched_companies = cleaned
        self.company_status_var.set(f"Companies: {len(cleaned)} available")
        self.company_status_label.configure(text_color=COLORS["text_muted"])

    def _fetch_tally_companies_thread(self, silent=False):
        if self._company_fetch_running:
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return

        self._company_fetch_running = True
        self.company_refresh_btn.configure(state="disabled", text="Fetching...")
        if not silent:
            self.company_status_var.set("Companies: Fetching...")
            self.company_status_label.configure(text_color=COLORS["warning"])

        def worker():
            result = _fetch_tally_companies(tally_url, timeout=15)

            def done():
                self._company_fetch_running = False
                self.company_refresh_btn.configure(state="normal", text="Refresh")
                if result.get("success"):
                    companies = result.get("companies", [])
                    self._set_company_dropdown(companies, keep_selection=True)
                    self.status_var.set(f"Fetched {len(companies)} company(s) from Tally")
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.company_status_var.set("Companies: Fetch failed")
                    self.company_status_label.configure(text_color=COLORS["error"])
                    self.status_var.set("Company fetch failed")
                    if not silent:
                        messagebox.showwarning("Company Fetch Failed", f"Could not fetch companies from Tally.\n\n{err}")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _check_tally_connection_thread(self):
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return

        self.connection_test_btn.configure(state="disabled", text="Checking...")
        self.connection_status_var.set("Connection: Checking...")
        self.connection_status_label.configure(text_color=COLORS["warning"])

        def worker():
            result = _check_tally_connection(tally_url, timeout=8)

            def done():
                self.connection_test_btn.configure(state="normal", text="Test Connection")
                if result.get("connected"):
                    self.connection_status_var.set(f"Connection: Connected ({tally_url})")
                    self.connection_status_label.configure(text_color=COLORS["success"])
                    self.status_var.set("Connected to Tally")
                    self._fetch_tally_companies_thread(silent=True)
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.connection_status_var.set("Connection: Offline")
                    self.connection_status_label.configure(text_color=COLORS["error"])
                    self.status_var.set(f"Connection failed: {err}")
                    messagebox.showwarning("Connection Failed", f"Could not connect to Tally.\n\n{err}")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    # ── VOUCHER TAB (shared for accounting & item) ──────────────────────

    def _build_voucher_tab(self, parent, mode="accounting"):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(6, weight=1)  # row 6 = preview container

        # Row 0: Template download
        template_row = ctk.CTkFrame(parent, fg_color="transparent")
        template_row.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))
        template_btn = ctk.CTkButton(
            template_row, text="📥  Download Template",
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], width=170,
            command=lambda: self._download_template_for_mode(mode),
        )
        template_btn.pack(side="right")
        self._voucher_template_buttons[mode] = template_btn

        # Row 1: Browse Excel row
        load_frame = ctk.CTkFrame(parent, fg_color="transparent")
        load_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 0))
        fp_var = ctk.StringVar()
        ctk.CTkEntry(load_frame, textvariable=fp_var,
                     placeholder_text="Select Excel file (.xlsx / .xlsm)...",
                     width=500, state="readonly").pack(side="left", padx=(0, 8))

        self._active_preview_mode[mode] = "excel"  # default

        def browse():
            if self._voucher_load_running.get(mode):
                self.status_var.set("Please wait, file is still loading...")
                return
            f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm *.xls")])
            if f:
                fp_var.set(f)
                self._load_preview(f, tree, mode)
                self._show_preview_mode(excel_tree_frame, xml_preview_frame, "excel",
                                        excel_toggle_btn, xml_toggle_btn, mode_key=mode)

        browse_btn = ctk.CTkButton(load_frame, text="Browse Excel", command=browse,
                                    width=90, fg_color=ACCENT, hover_color=ACCENT_HOVER)
        browse_btn.pack(side="left", padx=(0, 8))
        self._voucher_browse_buttons[mode] = browse_btn
        self._voucher_load_running[mode] = False

        if mode in {"item", "purchase_item"}:
            ctk.CTkLabel(load_frame,
                         text="⚠ Requires: ItemName, Quantity, Rate, and Per or Unit/UOM columns",
                         font=("Segoe UI", 11), text_color=COLORS["warning"]).pack(side="left")

        # Row 2: Browse XML row
        xml_row = ctk.CTkFrame(parent, fg_color="transparent")
        xml_row.grid(row=2, column=0, sticky="ew", padx=10, pady=(6, 0))
        xml_fp_var = ctk.StringVar()
        self._voucher_xml_fp_vars[mode] = xml_fp_var
        ctk.CTkEntry(xml_row, textvariable=xml_fp_var,
                     placeholder_text="Or select existing XML file to import/preview...",
                     width=500, state="readonly").pack(side="left", padx=(0, 8))

        def browse_xml():
            f = filedialog.askopenfilename(filetypes=[("XML Files", "*.xml")])
            if f:
                xml_fp_var.set(f)
                self._load_xml_preview(f, xml_text, mode, info_lbl)
                self._show_preview_mode(excel_tree_frame, xml_preview_frame, "xml",
                                        excel_toggle_btn, xml_toggle_btn, mode_key=mode)

        xml_browse_btn = ctk.CTkButton(
            xml_row, text="Browse XML", command=browse_xml, width=90,
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"])
        xml_browse_btn.pack(side="left", padx=(0, 8))
        self._voucher_xml_browse_buttons[mode] = xml_browse_btn
        ctk.CTkLabel(xml_row, text="Browse an existing XML to preview & push it directly to Tally",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left")

        # Row 3: Voucher Type selector (purchase modes only)
        # Sales modes (accounting / item) use per-row VoucherType column from the Excel file.
        _default_vtype = "Sales" if mode in {"accounting", "item"} else "Purchase"
        vtype_var = ctk.StringVar(value=_default_vtype)
        self._voucher_type_vars[mode] = vtype_var

        if mode not in {"accounting", "item"}:
            vtype_row = ctk.CTkFrame(parent, fg_color="transparent")
            vtype_row.grid(row=3, column=0, sticky="ew", padx=10, pady=(6, 0))
            ctk.CTkLabel(vtype_row, text="Voucher Type:", font=("Segoe UI", 12)).pack(side="left", padx=(0, 8))
            vtype_cb = ctk.CTkComboBox(vtype_row, variable=vtype_var, values=[_default_vtype],
                                        width=230, state="readonly")
            vtype_cb.pack(side="left", padx=(0, 8))
            self._voucher_type_cbs[mode] = vtype_cb
            _vtype_prefix = "Purchase"

            def _do_fetch_vtypes(_m=mode, _cb=vtype_cb, _var=vtype_var, _pfx=_vtype_prefix):
                _btn = self._voucher_type_fetch_buttons.get(_m)
                if _btn:
                    _btn.configure(state="disabled", text="Fetching...")
                _url = self._get_tally_url()
                _co = self._get_selected_company()

                def _worker():
                    types = _fetch_voucher_types_from_tally(_url, company=_co, prefix=_pfx)

                    def _update():
                        if types:
                            _cb.configure(values=types)
                            if _var.get() not in types:
                                _var.set(types[0])
                        else:
                            self.status_var.set("No voucher types found — is Tally running?")
                        if _btn:
                            _btn.configure(state="normal", text="Fetch")

                    self.after(0, _update)

                threading.Thread(target=_worker, daemon=True).start()

            fetch_vtype_btn = ctk.CTkButton(
                vtype_row, text="Fetch", width=70,
                fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                text_color=COLORS["text_secondary"],
                command=_do_fetch_vtypes,
            )
            fetch_vtype_btn.pack(side="left")
            self._voucher_type_fetch_buttons[mode] = fetch_vtype_btn
            ctk.CTkLabel(vtype_row, text="Fetch voucher types from Tally",
                         font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left", padx=8)

        # Row 4 (or 3 for sales modes): Auto Round Off
        _roundoff_grid_row = 3 if mode in {"accounting", "item"} else 4
        roundoff_row = ctk.CTkFrame(parent, fg_color="transparent")
        roundoff_row.grid(row=_roundoff_grid_row, column=0, sticky="ew", padx=10, pady=(4, 0))

        ro_enabled_var = ctk.BooleanVar(value=False)
        self._roundoff_enabled_vars[mode] = ro_enabled_var
        ro_ledger_var = ctk.StringVar(value="")
        self._roundoff_ledger_vars[mode] = ro_ledger_var

        ro_cb = ctk.CTkCheckBox(roundoff_row, text="Enable Auto Round Off",
                                variable=ro_enabled_var, font=("Segoe UI", 12))
        ro_cb.pack(side="left", padx=(0, 10))

        ro_entry = ctk.CTkEntry(roundoff_row, textvariable=ro_ledger_var,
                                placeholder_text="Search Round Off Ledger...",
                                width=220, state="disabled",
                                fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                                text_color=COLORS["text_primary"])
        ro_entry.pack(side="left", padx=(0, 6))

        ro_fetch_btn = ctk.CTkButton(roundoff_row, text="Fetch Ledgers", width=110,
                                     fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                                     text_color=COLORS["text_secondary"], state="disabled")
        ro_fetch_btn.pack(side="left", padx=(0, 6))
        self._roundoff_fetch_btns[mode] = ro_fetch_btn

        ctk.CTkLabel(roundoff_row, text="Fractional amounts go to this ledger",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left", padx=4)

        def _toggle_roundoff(_ev=ro_enabled_var, _e=ro_entry, _b=ro_fetch_btn, _v=ro_ledger_var):
            if _ev.get():
                _e.configure(state="normal")
                _b.configure(state="normal")
            else:
                _e.configure(state="disabled")
                _b.configure(state="disabled")
                _v.set("")

        ro_cb.configure(command=_toggle_roundoff)

        def _do_fetch_roundoff(_m=mode, _e=ro_entry, _b=ro_fetch_btn, _v=ro_ledger_var):
            try:
                _url = self._get_tally_url()
            except (ValueError, AttributeError):
                messagebox.showerror("Settings", "Invalid Tally URL. Check host/port in Settings.")
                return
            _co = self._get_selected_company()
            _b.configure(state="disabled", text="Fetching...")

            def _worker():
                result = _fetch_tally_ledgers(_url, timeout=15, company_name=_co)

                def _done():
                    _b.configure(state="normal", text="Fetch Ledgers")
                    if result.get("success"):
                        ledgers = result.get("ledgers") or []
                        self._roundoff_all_ledgers[_m] = ledgers
                        self._show_ledger_picker_popup(_e, _v, ledgers)
                    else:
                        messagebox.showwarning("Fetch Failed",
                                               "Could not fetch ledgers from Tally.\nIs Tally running?")

                self.after(0, _done)

            threading.Thread(target=_worker, daemon=True).start()

        ro_fetch_btn.configure(command=_do_fetch_roundoff)

        def _entry_click(event, _m=mode, _e=ro_entry, _v=ro_ledger_var):
            if ro_enabled_var.get():
                ledgers = self._roundoff_all_ledgers.get(_m, [])
                if ledgers:
                    self._show_ledger_picker_popup(_e, _v, ledgers)

        ro_entry.bind("<Button-1>", _entry_click)

        # Row 5: Preview toggle + info label
        toggle_row = ctk.CTkFrame(parent, fg_color="transparent")
        toggle_row.grid(row=5, column=0, sticky="ew", padx=10, pady=(8, 0))

        excel_toggle_btn = ctk.CTkButton(
            toggle_row, text="📊 Excel Data", width=160, height=30,
            font=("Segoe UI", 10, "bold"), fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"], text_color="#FFFFFF", corner_radius=6)
        excel_toggle_btn.pack(side="left", padx=(0, 4))

        xml_toggle_btn = ctk.CTkButton(
            toggle_row, text="📄 XML Preview", width=150, height=30,
            font=("Segoe UI", 10, "bold"), fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"], text_color=COLORS["text_muted"], corner_radius=6)
        xml_toggle_btn.pack(side="left", padx=(0, 8))

        info_lbl = ctk.CTkLabel(toggle_row, text="", font=("Segoe UI", 11),
                                 text_color=TEXT_MUTED)
        info_lbl.pack(side="left", padx=4)
        self._voucher_info_labels[mode] = info_lbl

        # Row 6: Preview container — Excel treeview + XML text widget stacked
        preview_container = ctk.CTkFrame(parent, fg_color="transparent")
        preview_container.grid(row=6, column=0, sticky="nsew", padx=10, pady=(4, 4))
        preview_container.grid_rowconfigure(0, weight=1)
        preview_container.grid_columnconfigure(0, weight=1)

        # Excel treeview frame
        excel_tree_frame = ctk.CTkFrame(
            preview_container, fg_color=COLORS["bg_dark"], corner_radius=8,
            border_width=1, border_color=COLORS["border"])
        excel_tree_frame.grid(row=0, column=0, sticky="nsew")
        excel_tree_frame.grid_rowconfigure(0, weight=1)
        excel_tree_frame.grid_columnconfigure(0, weight=1)

        tree_scroll_y = ttk.Scrollbar(excel_tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(excel_tree_frame, orient="horizontal")
        tree = ttk.Treeview(excel_tree_frame, show="headings",
                            yscrollcommand=tree_scroll_y.set,
                            xscrollcommand=tree_scroll_x.set)
        tree_scroll_y.config(command=tree.yview)
        tree_scroll_x.config(command=tree.xview)
        tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")

        # XML preview frame
        xml_preview_frame = ctk.CTkFrame(
            preview_container, fg_color=COLORS["bg_dark"], corner_radius=8,
            border_width=1, border_color=COLORS["border"])
        xml_preview_frame.grid(row=0, column=0, sticky="nsew")
        xml_preview_frame.grid_rowconfigure(0, weight=1)
        xml_preview_frame.grid_columnconfigure(0, weight=1)

        _xml_bg = self._resolve_theme_color("bg_card")
        _xml_fg = self._resolve_theme_color("text_primary")
        xml_text = tk.Text(
            xml_preview_frame, wrap="none", font=("Consolas", 10),
            bg=_xml_bg, fg=_xml_fg, insertbackground=_xml_fg,
            relief="flat", borderwidth=0)
        _xsy = ttk.Scrollbar(xml_preview_frame, orient="vertical", command=xml_text.yview)
        _xsx = ttk.Scrollbar(xml_preview_frame, orient="horizontal", command=xml_text.xview)
        xml_text.configure(yscrollcommand=_xsy.set, xscrollcommand=_xsx.set)
        xml_text.grid(row=0, column=0, sticky="nsew")
        _xsy.grid(row=0, column=1, sticky="ns")
        _xsx.grid(row=1, column=0, sticky="ew")
        xml_text.insert("end", "Browse an XML file above to preview its contents here.")
        xml_text.configure(state="disabled")

        # Wire toggle buttons
        excel_toggle_btn.configure(command=lambda: self._show_preview_mode(
            excel_tree_frame, xml_preview_frame, "excel", excel_toggle_btn, xml_toggle_btn, mode_key=mode))
        xml_toggle_btn.configure(command=lambda: self._show_preview_mode(
            excel_tree_frame, xml_preview_frame, "xml", excel_toggle_btn, xml_toggle_btn, mode_key=mode))

        # Show Excel frame by default
        excel_tree_frame.tkraise()

        # Row 7: Action buttons (smart Push: uses active preview mode)
        btn_frame = ctk.CTkFrame(parent, fg_color="transparent")
        btn_frame.grid(row=7, column=0, sticky="ew", padx=10, pady=(4, 10))

        save_btn = ctk.CTkButton(
            btn_frame, text="💾  Save XML File",
            fg_color=SUCCESS, hover_color="#15803D", width=155,
            command=lambda: self._generate(mode, "save", fp_var.get()))
        save_btn.pack(side="left", padx=(0, 8))
        self._voucher_save_buttons[mode] = save_btn

        push_btn = ctk.CTkButton(
            btn_frame, text="🚀  Push to Tally",
            fg_color=ACCENT, hover_color=ACCENT_HOVER, width=165,
            command=lambda: self._smart_push(mode, fp_var, xml_fp_var))
        push_btn.pack(side="left", padx=(0, 8))
        self._voucher_push_buttons[mode] = push_btn

        hint_lbl = ctk.CTkLabel(
            btn_frame, text="Push uses active preview (Excel or XML)",
            font=("Segoe UI", 10), text_color=COLORS["text_muted"])
        hint_lbl.pack(side="left", padx=8)


    def _show_ledger_picker_popup(self, anchor_entry, var, all_ledgers):
        """Searchable popup for picking a ledger name."""
        popup = tk.Toplevel(self)
        popup.title("Select Round Off Ledger")
        popup.geometry("360x340")
        popup.resizable(False, False)
        popup.transient(self)
        popup.grab_set()
        bg = self._resolve_theme_color("bg_card")
        fg = self._resolve_theme_color("text_primary")
        inp_bg = self._resolve_theme_color("bg_input")
        popup.configure(bg=bg)

        search_frame = tk.Frame(popup, bg=bg)
        search_frame.pack(fill="x", padx=8, pady=8)
        tk.Label(search_frame, text="Search:", bg=bg, fg=fg,
                 font=("Segoe UI", 10)).pack(side="left")
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var,
                                font=("Segoe UI", 10), bg=inp_bg, fg=fg,
                                insertbackground=fg, relief="flat", bd=2)
        search_entry.pack(side="left", padx=(6, 0), fill="x", expand=True)

        list_frame = tk.Frame(popup, bg=bg)
        list_frame.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                             font=("Segoe UI", 10), selectmode="single",
                             bg=inp_bg, fg=fg, relief="flat", bd=0,
                             selectbackground=self._resolve_theme_color("accent"),
                             selectforeground="#FFFFFF")
        listbox.pack(fill="both", expand=True)
        scrollbar.config(command=listbox.yview)

        def _filter(*_):
            q = search_var.get().strip().lower()
            listbox.delete(0, "end")
            for name in all_ledgers:
                if not q or q in name.lower():
                    listbox.insert("end", name)

        search_var.trace_add("write", _filter)
        _filter()

        def _select(*_):
            sel = listbox.curselection()
            if sel:
                var.set(listbox.get(sel[0]))
                popup.destroy()

        btn_row = tk.Frame(popup, bg=bg)
        btn_row.pack(fill="x", padx=8, pady=(0, 8))
        tk.Button(btn_row, text="Select", command=_select, relief="flat",
                  font=("Segoe UI", 10, "bold"),
                  bg=self._resolve_theme_color("accent"), fg="#FFFFFF",
                  activebackground=self._resolve_theme_color("accent_hover"),
                  padx=12, pady=4).pack(side="left", padx=(0, 6))
        tk.Button(btn_row, text="Cancel", command=popup.destroy, relief="flat",
                  font=("Segoe UI", 10), bg=inp_bg, fg=fg,
                  padx=12, pady=4).pack(side="left")

        listbox.bind("<Double-Button-1>", _select)
        listbox.bind("<Return>", _select)
        search_entry.focus_set()

    def _set_voucher_loading_state(self, mode: str, is_loading: bool):
        self._voucher_load_running[mode] = is_loading
        state = "disabled" if is_loading else "normal"
        browse_text = "Loading..." if is_loading else "Browse Excel"
        browse_btn = self._voucher_browse_buttons.get(mode)
        if browse_btn:
            browse_btn.configure(state=state, text=browse_text)
        save_btn = self._voucher_save_buttons.get(mode)
        if save_btn:
            save_btn.configure(state=state)
        template_btn = self._voucher_template_buttons.get(mode)
        if template_btn:
            template_btn.configure(state=state)
        push_btn = self._voucher_push_buttons.get(mode)
        if push_btn:
            push_btn.configure(state=state)
        xml_browse_btn = self._voucher_xml_browse_buttons.get(mode)
        if xml_browse_btn:
            xml_browse_btn.configure(state=state)
        import_btn = self._voucher_import_buttons.get(mode)
        if import_btn:
            import_btn.configure(state=state)

    def _set_push_loading_state(self, is_loading: bool, message: str = ""):
        self._push_running = is_loading
        state = "disabled" if is_loading else "normal"

        for btn in self._voucher_browse_buttons.values():
            btn.configure(state=state)
        for btn in self._voucher_save_buttons.values():
            btn.configure(state=state)
        for btn in self._voucher_template_buttons.values():
            btn.configure(state=state)
        for btn in self._voucher_push_buttons.values():
            btn.configure(state=state)
        for btn in self._voucher_xml_browse_buttons.values():
            btn.configure(state=state)
        for btn in self._voucher_type_fetch_buttons.values():
            btn.configure(state=state)
        for btn in self._roundoff_fetch_btns.values():
            btn.configure(state=state)
        if self.demo_btn is not None:
            self.demo_btn.configure(state=state)

        if self.voucher_date_current_cb is not None:
            self.voucher_date_current_cb.configure(state=state)
        if self.voucher_date_excel_cb is not None:
            self.voucher_date_excel_cb.configure(state=state)
        if self.voucher_date_custom_cb is not None:
            self.voucher_date_custom_cb.configure(state=state)

        if is_loading:
            self._push_message_var.set(message or "Posting to Tally...")
            if self._push_overlay is None or not self._push_overlay.winfo_exists():
                overlay = ctk.CTkToplevel(self)
                overlay.title("Please Wait")
                overlay.geometry("420x170")
                overlay.transient(self)
                overlay.grab_set()
                overlay.resizable(False, False)

                frame = ctk.CTkFrame(overlay, fg_color=COLORS["bg_card"], corner_radius=10)
                frame.pack(fill="both", expand=True, padx=12, pady=12)

                ctk.CTkLabel(
                    frame,
                    text="Pushing Entries To Tally",
                    font=("Segoe UI", 15, "bold"),
                    text_color=COLORS["text_primary"],
                ).pack(pady=(12, 8))

                ctk.CTkLabel(
                    frame,
                    textvariable=self._push_message_var,
                    font=("Segoe UI", 11),
                    text_color=COLORS["text_muted"],
                    wraplength=360,
                    justify="center",
                ).pack(pady=(0, 10))

                progress = ctk.CTkProgressBar(frame, mode="indeterminate", width=320)
                progress.pack(pady=(0, 10))
                progress.start()

                overlay.protocol("WM_DELETE_WINDOW", lambda: None)
                self._push_overlay = overlay
            else:
                self._push_overlay.deiconify()
                self._push_overlay.lift()
        else:
            if self._push_overlay is not None and self._push_overlay.winfo_exists():
                try:
                    self._push_overlay.grab_release()
                except Exception:
                    pass
                self._push_overlay.destroy()
            self._push_overlay = None
            self._push_message_var.set("")

        self._set_voucher_date_mode(self.voucher_date_mode_var.get())

        self.update_idletasks()

    def _get_template_definition(self, mode: str) -> dict:
        templates = {
            "accounting": {
                "sheet_name": "Sheet1",
                "headers": [
                    "VoucherType",
                    "Date",
                    "InvoiceNo",
                    "VoucherNo",
                    "PartyLedger",
                    "PlaceOfSupply",
                    "GSTIN/UIN",
                    "SalesLedger",
                    "TaxableValue",
                    "CGSTLedger",
                    "CGST Amount",
                    "SGSTLedger",
                    "SGST Amount",
                    "IGSTLedger",
                    "IGST Amount",
                    "Narration",
                ],
                "sample_rows": [
                    ["Sales", "20-04-2026", "SPL/25-26/024", "155", "Interactive Media Pvt Ltd", "Assam",
                     "18AABCI8307G1ZM", "Lecture Income", 10000, "CGST", 0,
                     "SGST", 0, "IGST Outward", 1800, "Testing"],
                ],
            },
            "item": {
                "sheet_name": "Sheet1",
                "headers": [
                    "VoucherType",
                    "Date",
                    "InvoiceNo",
                    "VoucherNo",
                    "GSTIN",
                    "PartyLedger",
                    "Sales Ledger",
                    "Item Name",
                    "Unit",
                    "Quantity",
                    "Rate",
                    "TaxableValue",
                    "CGSTLedger",
                    "CGST Amount",
                    "SGSTLedger",
                    "SGST Amount",
                    "IGSTLedger",
                    "IGST Amount",
                    "Narration",
                ],
                "sample_rows": [
                    ["Sales", "20-04-2026", "SINV-001", "1", "27AAAPQ1234B1Z3", "Party Ledger Name",
                     "Sales Account", "Item Name", "Nos", 10, 1000, 10000,
                     "CGST", 900, "SGST", 900, "IGST", 0, "Testing"],
                ],
            },
            "purchase_accounting": {
                "sheet_name": "Sheet1",
                "headers": [
                    "Date",
                    "InvoiceNo",
                    "VoucherNo",
                    "SupplierInvoiceNo",
                    "PartyLedger",
                    "PartyName",
                    "PartyMailingName",
                    "PartyAddress1",
                    "PartyAddress2",
                    "PartyPincode",
                    "PartyState",
                    "PlaceOfSupply",
                    "PartyCountry",
                    "GSTApplicable",
                    "GSTRegistrationType",
                    "GSTIN/UIN",
                    "PurchaseLedger",
                    "TaxableValue",
                    "CGSTLedger",
                    "CGST Amount",
                    "SGSTLedger",
                    "SGST Amount",
                    "IGSTLedger",
                    "IGST Amount",
                    "Narration",
                    "TDSLedger",
                    "TDSRate",
                    "TDSAmount",
                ],
                "sample_rows": [
                    ["01/04/2024", "SINV-001", "1", "SINV-001", "PQR Suppliers", "PQR Suppliers", "PQR Suppliers",
                     "789 MIDC Road", "", "411001", "Maharashtra", "Maharashtra", "India",
                     "Applicable", "Regular", "27AAAPQ1234B1Z3", "Purchase Account", 10000,
                     "CGST", 900, "SGST", 900, "IGST", 0, "Being goods purchased from PQR Suppliers",
                     "TDS Payable on Professional", 10, ""],
                    ["02/04/2024", "SINV-002", "2", "SINV-002", "LMN Industries", "LMN Industries", "LMN Industries",
                     "321 Ring Road", "", "302001", "Rajasthan", "Maharashtra", "India",
                     "Applicable", "Regular", "08AAELM9876C1Z1", "Purchase Account", 20000,
                     "CGST", 0, "SGST", 0, "IGST", 3600, "Being goods purchased from LMN Industries",
                     "", "", ""],
                ],
            },
            "purchase_item": {
                "sheet_name": "Sheet1",
                "headers": [
                    "Date",
                    "InvoiceNo",
                    "VoucherNo",
                    "SupplierInvoiceNo",
                    "PartyLedger",
                    "PartyName",
                    "PartyMailingName",
                    "PartyAddress1",
                    "PartyAddress2",
                    "PartyPincode",
                    "PartyState",
                    "PlaceOfSupply",
                    "PartyCountry",
                    "GSTApplicable",
                    "GSTRegistrationType",
                    "GSTIN/UIN",
                    "Purchase Ledger",
                    "Item Name",
                    "Unit",
                    "Quantity",
                    "Rate",
                    "TaxableValue",
                    "CGSTLedger",
                    "CGST Amount",
                    "SGSTLedger",
                    "SGST Amount",
                    "IGSTLedger",
                    "IGST Amount",
                    "Narration",
                ],
                "sample_rows": [],
            },
        }
        return templates.get(mode, {})

    def _download_template_for_mode(self, mode: str):
        template_name = self.template_names_by_mode.get(mode, "Template.xlsx")
        template_definition = self._get_template_definition(mode)
        if not template_definition:
            messagebox.showwarning("Template Not Found", "No template is configured for this mode.")
            return

        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=template_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not out_path:
            return

        try:
            headers = list(template_definition.get("headers", []))
            sample_rows = list(template_definition.get("sample_rows", []))
            sheet_name = str(template_definition.get("sheet_name") or "Template")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            for col_index, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_index, value=header)

            for row_index, row_values in enumerate(sample_rows, 2):
                for col_index, value in enumerate(row_values, 1):
                    ws.cell(row=row_index, column=col_index, value=value)

            wb.save(out_path)
            wb.close()
            self.status_var.set(f"Template saved: {os.path.basename(out_path)}")
            messagebox.showinfo("Template Saved", f"Template saved successfully!\n{out_path}")
        except Exception as exc:
            messagebox.showerror("Template Error", str(exc))

    def _load_preview(self, filepath, tree, mode):
        self._set_voucher_loading_state(mode, True)
        info_label = self._voucher_info_labels.get(mode)
        if info_label:
            info_label.configure(text="Loading Excel preview...")
        self.status_var.set(f"Loading: {os.path.basename(filepath)}")

        def worker():
            try:
                headers, rows = read_excel(filepath)
                preview_rows = []
                for r in rows[:300]:
                    values = []
                    for h in headers:
                        cell_value = _row_get(r, h, "")
                        values.append("" if cell_value is None else str(cell_value))
                    preview_rows.append(values)
                result = {
                    "ok": True,
                    "headers": headers,
                    "rows": rows,
                    "preview_rows": preview_rows,
                }
            except Exception as exc:
                result = {"ok": False, "error": str(exc)}

            def done():
                self._set_voucher_loading_state(mode, False)
                if not result.get("ok"):
                    if info_label:
                        info_label.configure(text="")
                    self.status_var.set("Ready")
                    messagebox.showerror("Error", str(result.get("error", "Unknown error")))
                    return

                headers = result["headers"]
                rows = result["rows"]
                self.loaded_headers = headers
                self.loaded_rows = rows
                self._loaded_headers_by_mode[mode] = headers
                self._loaded_rows_by_mode[mode] = rows

                tree.delete(*tree.get_children())
                tree["columns"] = headers
                for h in headers:
                    tree.heading(h, text=h)
                    tree.column(h, width=100, minwidth=60)

                for values in result["preview_rows"]:
                    tree.insert("", "end", values=values)

                info = f"✅ Loaded {len(rows)} rows, {len(headers)} columns"
                if info_label:
                    info_label.configure(text=info)
                self.status_var.set(f"Loaded: {os.path.basename(filepath)} — {len(rows)} rows")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _check_voucher_types_and_remap(self, rows, unique_vtypes, default_vtype):
        """
        1. Fetch voucher types from Tally (synchronously via background thread + event).
        2. If every VoucherType in Excel is found in Tally → proceed silently, no popup.
        3. If any are NOT found → show popup only for the mismatched ones.
        Returns True to proceed (rows mutated in-place for remapped types), False to cancel.
        """
        # ── Step 1: fetch Tally voucher types in background, wait up to 6 s ──
        url = self._get_tally_url()
        co  = self._get_selected_company()
        fetch_result = {"types": []}
        fetch_done   = threading.Event()

        def _fetch_worker():
            try:
                fetch_result["types"] = _fetch_voucher_types_from_tally(
                    url, company=co, prefix=default_vtype, timeout=6.0) or []
            except Exception:
                fetch_result["types"] = []
            finally:
                fetch_done.set()

        threading.Thread(target=_fetch_worker, daemon=True).start()
        fetch_done.wait(timeout=7)

        tally_types = fetch_result["types"]

        # If Tally is unreachable (empty list) → skip validation, proceed silently
        if not tally_types:
            return True

        tally_set   = {t.strip().lower() for t in tally_types}

        # ── Step 2: find mismatches ──
        mismatched = [
            vt for vt in unique_vtypes
            if vt.strip().lower() not in tally_set
        ]

        # All types matched → proceed without popup
        if not mismatched:
            return True

        # ── Step 3: show popup ONLY for mismatched types ──
        result  = {"proceed": False, "remap": {}}
        bg      = self._resolve_theme_color("bg_dark")
        card    = self._resolve_theme_color("bg_card")
        fg      = self._resolve_theme_color("text_primary")
        muted   = self._resolve_theme_color("text_muted")
        error   = self._resolve_theme_color("error")
        accent  = self._resolve_theme_color("accent")
        inp_bg  = self._resolve_theme_color("bg_input")
        border  = self._resolve_theme_color("border")
        gold    = self._resolve_theme_color("tally_gold")
        hdr_bg  = self._resolve_theme_color("table_header")

        choices = tally_types or unique_vtypes

        popup = ctk.CTkToplevel(self)
        popup.title("Voucher Type Not Found in Tally")
        popup.geometry("600x380")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)
        popup.configure(fg_color=bg)

        ctk.CTkLabel(popup, text="Voucher Type Not Found",
                     font=("Segoe UI", 13, "bold"),
                     text_color=error).pack(anchor="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(popup,
                     text=f"The following voucher type(s) from your Excel were not found in Tally. "
                          f"Please map each to the correct Tally voucher type.",
                     font=("Segoe UI", 10), text_color=muted,
                     wraplength=560).pack(anchor="w", padx=16, pady=(0, 8))

        remap_vars = {}   # {excel_vtype: ctk.StringVar}

        table_frame = ctk.CTkFrame(popup, fg_color=card, corner_radius=8)
        table_frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        hdr = ctk.CTkFrame(table_frame, fg_color=hdr_bg, corner_radius=0, height=30)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="Not Found in Tally", font=("Segoe UI", 10, "bold"),
                     text_color=gold, width=200, anchor="w").pack(side="left", padx=12)
        ctk.CTkLabel(hdr, text="Map To (Tally Voucher Type)", font=("Segoe UI", 10, "bold"),
                     text_color=gold).pack(side="left", padx=8)

        scroll = ctk.CTkScrollableFrame(table_frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=4, pady=4)

        for vt in mismatched:
            var = ctk.StringVar(value=choices[0] if choices else vt)
            remap_vars[vt] = var
            row_f = ctk.CTkFrame(scroll, fg_color="transparent")
            row_f.pack(fill="x", pady=3)
            ctk.CTkLabel(row_f, text=vt, font=("Segoe UI", 11),
                         text_color=error, width=200, anchor="w").pack(side="left", padx=12)
            ctk.CTkComboBox(row_f, variable=var,
                            values=choices if choices else [vt],
                            width=280, font=("Segoe UI", 10),
                            fg_color=inp_bg, border_color=border,
                            button_color=accent).pack(side="left", padx=4)

        btn_row = ctk.CTkFrame(popup, fg_color="transparent")
        btn_row.pack(fill="x", padx=16, pady=(0, 14))

        def _re_fetch():
            re_fetch_btn.configure(state="disabled", text="Fetching...")
            def _w():
                new_types = _fetch_voucher_types_from_tally(
                    url, company=co, prefix=default_vtype, timeout=6.0) or []
                def _done():
                    nonlocal choices
                    choices = new_types or choices
                    for vt, var in remap_vars.items():
                        cb_widget = var._widget if hasattr(var, "_widget") else None
                    # Rebuild comboboxes with fresh list
                    for w in scroll.winfo_children():
                        w.destroy()
                    for vt in mismatched:
                        var = remap_vars[vt]
                        row_f = ctk.CTkFrame(scroll, fg_color="transparent")
                        row_f.pack(fill="x", pady=3)
                        ctk.CTkLabel(row_f, text=vt, font=("Segoe UI", 11),
                                     text_color=error, width=200, anchor="w").pack(side="left", padx=12)
                        ctk.CTkComboBox(row_f, variable=var,
                                        values=choices,
                                        width=280, font=("Segoe UI", 10),
                                        fg_color=inp_bg, border_color=border,
                                        button_color=accent).pack(side="left", padx=4)
                    re_fetch_btn.configure(state="normal", text="Re-Fetch Tally Types")
                self.after(0, _done)
            threading.Thread(target=_w, daemon=True).start()

        re_fetch_btn = ctk.CTkButton(btn_row, text="Re-Fetch Tally Types", width=160,
                                     fg_color=inp_bg,
                                     hover_color=self._resolve_theme_color("bg_card_hover"),
                                     text_color=self._resolve_theme_color("text_secondary"),
                                     command=_re_fetch)
        re_fetch_btn.pack(side="left", padx=(0, 8))

        def _apply():
            result["remap"]   = {vt: var.get() for vt, var in remap_vars.items()}
            result["proceed"] = True
            popup.destroy()

        def _cancel():
            result["proceed"] = False
            popup.destroy()

        ctk.CTkButton(btn_row, text="Apply & Push", width=130,
                      fg_color=accent,
                      hover_color=self._resolve_theme_color("accent_hover"),
                      text_color="#FFFFFF",
                      command=_apply).pack(side="right", padx=(8, 0))
        ctk.CTkButton(btn_row, text="Cancel", width=100,
                      fg_color=inp_bg,
                      hover_color=self._resolve_theme_color("bg_card_hover"),
                      text_color=self._resolve_theme_color("text_secondary"),
                      command=_cancel).pack(side="right")

        popup.protocol("WM_DELETE_WINDOW", _cancel)
        popup.wait_window()

        if not result["proceed"]:
            return False

        # Apply remapping only for the mismatched rows
        remap = result["remap"]
        for r in rows:
            cur = str(_row_get(r, "VoucherType", "") or "").strip()
            if cur in remap:
                r["VoucherType"] = remap[cur]

        return True

    def _generate(self, mode, action, filepath):
        _ = filepath
        if self._voucher_load_running.get(mode):
            messagebox.showinfo("Please Wait", "Excel file is still loading. Please try again in a moment.")
            return

        if action == "push" and self._push_running:
            self.status_var.set("Push already in progress. Please wait...")
            return

        mode_rows = self._loaded_rows_by_mode.get(mode)
        rows_to_use = mode_rows if mode_rows is not None else self.loaded_rows
        company = self._get_selected_company()
        date_mode, custom_tally_date = self._get_voucher_date_selection()
        if not rows_to_use:
            messagebox.showwarning("No Data", "Load an Excel file first.")
            return

        if action == "push" and not company and len(getattr(self, "fetched_companies", [])) > 1:
            self.status_var.set("Select target company before push")
            messagebox.showwarning(
                "Select Target Company",
                "Multiple companies were detected in Tally.\n"
                "Please select the target company from the dropdown before pushing.",
            )
            return

        _default_vtype = "Sales" if mode in {"accounting", "item"} else "Purchase"
        voucher_type = (self._voucher_type_vars.get(mode) or ctk.StringVar(value=_default_vtype)).get() or _default_vtype

        # For sales modes pushing to Tally: validate per-row VoucherType against live Tally types.
        # Only show mismatch popup if a type is not found in Tally — never popup when all match.
        if mode in {"accounting", "item"} and action == "push":
            unique_vtypes = sorted({
                str(_row_get(r, "VoucherType", "") or "").strip()
                for r in rows_to_use
                if str(_row_get(r, "VoucherType", "") or "").strip()
            })
            if unique_vtypes:
                proceed = self._check_voucher_types_and_remap(
                    rows_to_use, unique_vtypes, _default_vtype
                )
                if not proceed:
                    return

        _ro_en = self._roundoff_enabled_vars.get(mode)
        _ro_lv = self._roundoff_ledger_vars.get(mode)
        round_off_ledger = (_ro_lv.get() or "").strip() if (_ro_en and _ro_en.get() and _ro_lv) else ""
        mode_to_file_name = {
            "accounting": "Tally_Sales_Accounting.xml",
            "item": "Tally_Sales_Item.xml",
            "purchase_accounting": "Tally_Purchase_Accounting.xml",
            "purchase_item": "Tally_Purchase_Item.xml",
        }

        def build_voucher_xml(
            selected_mode: str,
            selected_date_mode: str,
            selected_custom_date: str,
            voucher_start,
            rows_data,
            company_gst_registrations=None,
            legacy_item_invoice_context: bool = False,
        ):
            if selected_mode == "accounting":
                return generate_accounting_xml(
                    rows_data,
                    company,
                    date_mode=selected_date_mode,
                    custom_tally_date=selected_custom_date,
                    start_voucher_number=voucher_start,
                    company_gst_registrations=company_gst_registrations,
                    voucher_type=voucher_type,
                    round_off_ledger=round_off_ledger,
                )
            if selected_mode == "item":
                return generate_item_xml(
                    rows_data,
                    company,
                    date_mode=selected_date_mode,
                    custom_tally_date=selected_custom_date,
                    start_voucher_number=voucher_start,
                    company_gst_registrations=company_gst_registrations,
                    legacy_invoice_context=legacy_item_invoice_context,
                    voucher_type=voucher_type,
                    round_off_ledger=round_off_ledger,
                )
            if selected_mode == "purchase_accounting":
                return generate_purchase_accounting_xml(
                    rows_data,
                    company,
                    date_mode=selected_date_mode,
                    custom_tally_date=selected_custom_date,
                    start_voucher_number=voucher_start,
                    company_gst_registrations=company_gst_registrations,
                    voucher_type=voucher_type,
                    round_off_ledger=round_off_ledger,
                )
            if selected_mode == "purchase_item":
                return generate_purchase_item_xml(
                    rows_data,
                    company,
                    date_mode=selected_date_mode,
                    custom_tally_date=selected_custom_date,
                    start_voucher_number=voucher_start,
                    company_gst_registrations=company_gst_registrations,
                    voucher_type=voucher_type,
                    round_off_ledger=round_off_ledger,
                )
            raise ValueError(f"Unsupported mode: {selected_mode}")

        try:
            if action == "save":
                rows_for_save = list(rows_to_use)
                if mode in {"accounting", "item"}:
                    try:
                        _save_url = self._get_tally_url()
                        _save_state_map = _fetch_party_ledger_states(
                            _save_url, company_name=company, timeout=8)
                        if _save_state_map:
                            rows_for_save = _inject_party_states_from_tally(
                                rows_for_save, _save_state_map)
                    except Exception:
                        pass  # Tally not running — save as-is
                xml = build_voucher_xml(mode, date_mode, custom_tally_date, None, rows_for_save)
                out = filedialog.asksaveasfilename(defaultextension=".xml",
                        filetypes=[("XML","*.xml")],
                        initialfile=mode_to_file_name.get(mode, "Tally_Voucher.xml"))
                if out:
                    with open(out, "w", encoding="utf-8") as f:
                        f.write(xml)
                    self.status_var.set(f"Saved XML: {out}")
                    messagebox.showinfo("Success", f"XML saved successfully!\n{out}")
            else:
                tally_url = self._get_tally_url()
                host, port_text = tally_url.rsplit(":", 1)
                host = host.replace("http://", "", 1)
                port_value = int(port_text)
                target_company = company or "Loaded company in Tally"
                date_mode_label = {
                    "current": "current date",
                    "excel": "excel date",
                    "custom": "custom date",
                }.get(date_mode, "excel date")
                rows_snapshot = list(rows_to_use)

                self._set_push_loading_state(True, f"Preparing vouchers for {target_company}...")
                self.status_var.set(f"Posting to Tally ({target_company}, {date_mode_label})...")

                def worker():
                    nonlocal rows_snapshot
                    result = {
                        "ok": False,
                        "target_company": target_company,
                        "parsed": {},
                        "message": "",
                        "detail": "",
                    }
                    try:
                        effective_date_mode = date_mode
                        effective_custom_tally_date = custom_tally_date
                        company_gst_registrations = []
                        existing_ledger_names = set()

                        self.after(0, lambda: self._push_message_var.set("Fetching company GST registrations..."))
                        gst_reg_result = _fetch_company_gst_registrations(
                            tally_url,
                            company_name=company,
                            timeout=15,
                        )
                        if gst_reg_result.get("success"):
                            company_gst_registrations = gst_reg_result.get("registrations", [])

                        self.after(0, lambda: self._push_message_var.set("Checking existing ledgers in Tally..."))
                        existing_ledger_result = _fetch_existing_ledger_names(
                            tally_url,
                            company_name=company,
                            timeout=15,
                        )
                        if existing_ledger_result.get("success"):
                            existing_ledger_names = set(existing_ledger_result.get("ledgers") or set())

                        auto_ledger_defs = _collect_auto_voucher_ledgers(rows_snapshot, mode)
                        auto_ledger_defs_to_create = _filter_out_existing_ledgers(
                            auto_ledger_defs,
                            existing_ledger_names,
                        )
                        if auto_ledger_defs_to_create:
                            # For party ledgers: ask the user how to create each one.
                            # For tax/sales/purchase ledgers: auto-create silently.
                            party_defs_to_ask = []
                            non_party_defs = []
                            for _ld in auto_ledger_defs_to_create:
                                if _is_party_parent(str(_ld.get("Parent", "") or "")):
                                    party_defs_to_ask.append(_ld)
                                else:
                                    non_party_defs.append(_ld)

                            # Ask user for each missing party ledger
                            asked_defs = []
                            for _ld in party_defs_to_ask:
                                self.after(0, lambda: self._push_message_var.set(
                                    f"Waiting for user input: ledger '{_ld.get('Name', '')}' not found..."))
                                enriched = _ask_create_party_ledger(self, _ld)
                                asked_defs.append(enriched)

                            auto_ledger_defs_to_create = non_party_defs + asked_defs

                            self.after(0, lambda: self._push_message_var.set("Creating required ledgers in Tally..."))
                            auto_ledger_xml = generate_ledger_xml(auto_ledger_defs_to_create, company)
                            auto_ledger_resp = push_to_tally(auto_ledger_xml, host, port_value)
                            auto_ledger_parsed = _parse_tally_response_details(auto_ledger_resp)
                            self._append_debug_log(
                                "auto-ledger",
                                target_company,
                                auto_ledger_xml,
                                auto_ledger_resp,
                                auto_ledger_parsed,
                                note=(
                                    f"mode={mode}, ledgers_total={len(auto_ledger_defs)}, "
                                    f"new_ledgers={len(auto_ledger_defs_to_create)}"
                                ),
                            )
                            existing_ledger_names.update(
                                str(entry.get("Name", "") or "").strip()
                                for entry in auto_ledger_defs_to_create
                                if str(entry.get("Name", "") or "").strip()
                            )

                        self.after(0, lambda: self._push_message_var.set("Fetching next voucher number from Tally..."))
                        next_voucher = None
                        voucher_note = "excel_voucher_no"
                        voucher_result = _fetch_next_voucher_number(
                            tally_url,
                            company_name=company,
                            voucher_type=voucher_type,
                            timeout=15,
                        )
                        if voucher_result.get("success"):
                            next_voucher = voucher_result.get("next_number")
                            voucher_note = f"auto_vno_start={next_voucher}"
                        else:
                            voucher_note = f"auto_vno_fetch_failed={voucher_result.get('error', 'Unknown')}"

                        vno_label = f"vno start {next_voucher}" if next_voucher is not None else "vno from excel"
                        self.after(
                            0,
                            lambda: self.status_var.set(
                                f"Posting to Tally ({target_company}, {date_mode_label}, {vno_label})..."
                            ),
                        )

                        # For sales modes: fetch party ledger states from Tally so that
                        # unregistered parties (no GSTIN, no State in Excel) get a valid
                        # PlaceOfSupply — preventing "Uncertain" in GSTR-1.
                        if mode in {"accounting", "item"}:
                            self.after(0, lambda: self._push_message_var.set(
                                "Fetching party states from Tally for unregistered parties..."))
                            _party_state_map = _fetch_party_ledger_states(
                                tally_url, company_name=company, timeout=15)
                            if _party_state_map:
                                rows_snapshot = _inject_party_states_from_tally(
                                    rows_snapshot, _party_state_map)

                        self.after(0, lambda: self._push_message_var.set("Posting voucher data to Tally..."))

                        xml = build_voucher_xml(
                            mode,
                            effective_date_mode,
                            effective_custom_tally_date,
                            next_voucher,
                            rows_snapshot,
                            company_gst_registrations,
                        )
                        resp = push_to_tally(xml, host, port_value)
                        parsed = _parse_tally_response_details(resp)
                        self._append_debug_log(
                            mode,
                            target_company,
                            xml,
                            resp,
                            parsed,
                            note=f"voucher_type={voucher_type}, date_mode={effective_date_mode}, {voucher_note}",
                        )

                        if parsed.get("success"):
                            result = {
                                "ok": True,
                                "target_company": target_company,
                                "parsed": parsed,
                                "message": (
                                    "Posted to Tally successfully.\n\n"
                                    f"Target Company: {target_company}\n"
                                    f"Created: {parsed.get('created', 0)}\n"
                                    f"Altered: {parsed.get('altered', 0)}\n"
                                    f"Ignored: {parsed.get('ignored', 0)}"
                                ),
                            }
                        else:
                            if not result.get("ok"):
                                missing_ledger_defs = _build_missing_ledger_defs(
                                    parsed.get("line_errors") or [],
                                    rows_snapshot,
                                    mode,
                                )
                                missing_ledger_defs = _filter_out_existing_ledgers(
                                    missing_ledger_defs,
                                    existing_ledger_names,
                                )
                                if missing_ledger_defs:
                                    # Ask user for each missing party ledger before creating
                                    final_missing_defs = []
                                    for _mld in missing_ledger_defs:
                                        if _is_party_parent(str(_mld.get("Parent", "") or "")):
                                            self.after(0, lambda: self._push_message_var.set(
                                                f"Waiting for user input: ledger '{_mld.get('Name', '')}' not found..."))
                                            _mld = _ask_create_party_ledger(self, _mld)
                                        final_missing_defs.append(_mld)
                                    missing_ledger_defs = final_missing_defs

                                    self.after(0, lambda: self._push_message_var.set("Creating missing ledgers and retrying..."))
                                    missing_ledger_xml = generate_ledger_xml(missing_ledger_defs, company)
                                    missing_ledger_resp = push_to_tally(missing_ledger_xml, host, port_value)
                                    missing_ledger_parsed = _parse_tally_response_details(missing_ledger_resp)
                                    self._append_debug_log(
                                        "missing-ledger",
                                        target_company,
                                        missing_ledger_xml,
                                        missing_ledger_resp,
                                        missing_ledger_parsed,
                                        note=f"mode={mode}, ledgers={len(missing_ledger_defs)}",
                                    )

                                    blockers = []
                                    for err in missing_ledger_parsed.get("line_errors", []):
                                        err_low = err.lower()
                                        if "already exists" in err_low or "already present" in err_low:
                                            continue
                                        blockers.append(err)

                                    if not blockers:
                                        existing_ledger_names.update(
                                            str(entry.get("Name", "") or "").strip()
                                            for entry in missing_ledger_defs
                                            if str(entry.get("Name", "") or "").strip()
                                        )
                                        self.after(0, lambda: self._push_message_var.set("Retrying voucher post after ledger creation..."))
                                        post_ledger_retry_xml = build_voucher_xml(
                                            mode,
                                            effective_date_mode,
                                            effective_custom_tally_date,
                                            next_voucher,
                                            rows_snapshot,
                                            company_gst_registrations,
                                        )
                                        post_ledger_retry_resp = push_to_tally(post_ledger_retry_xml, host, port_value)
                                        post_ledger_retry_parsed = _parse_tally_response_details(post_ledger_retry_resp)
                                        self._append_debug_log(
                                            mode,
                                            target_company,
                                            post_ledger_retry_xml,
                                            post_ledger_retry_resp,
                                            post_ledger_retry_parsed,
                                            note=(
                                                f"voucher_type={voucher_type}, auto_retry_missing_ledgers, {voucher_note}"
                                            ),
                                        )
                                        if post_ledger_retry_parsed.get("success"):
                                            result = {
                                                "ok": True,
                                                "target_company": target_company,
                                                "parsed": post_ledger_retry_parsed,
                                                "message": (
                                                    "Auto-created missing ledgers and posted to Tally successfully.\n\n"
                                                    f"Target Company: {target_company}\n"
                                                    f"Created: {post_ledger_retry_parsed.get('created', 0)}\n"
                                                    f"Altered: {post_ledger_retry_parsed.get('altered', 0)}\n"
                                                    f"Ignored: {post_ledger_retry_parsed.get('ignored', 0)}"
                                                ),
                                            }
                                        else:
                                            parsed = post_ledger_retry_parsed

                            if (
                                not result.get("ok")
                                and mode == "item"
                                and voucher_type == "Sales"
                                and parsed.get("exceptions", 0) > 0
                                and not (parsed.get("line_errors") or [])
                            ):
                                self.after(
                                    0,
                                    lambda: self._push_message_var.set(
                                        "Retrying sales item invoice with compatibility header..."
                                    ),
                                )
                                compatibility_xml = build_voucher_xml(
                                    mode,
                                    effective_date_mode,
                                    effective_custom_tally_date,
                                    next_voucher,
                                    rows_snapshot,
                                    company_gst_registrations,
                                    legacy_item_invoice_context=True,
                                )
                                compatibility_resp = push_to_tally(compatibility_xml, host, port_value)
                                compatibility_parsed = _parse_tally_response_details(compatibility_resp)
                                self._append_debug_log(
                                    mode,
                                    target_company,
                                    compatibility_xml,
                                    compatibility_resp,
                                    compatibility_parsed,
                                    note=(
                                        f"voucher_type={voucher_type}, compatibility_retry_legacy_item_header, "
                                        f"{voucher_note}, date_mode={effective_date_mode}"
                                    ),
                                )
                                if compatibility_parsed.get("success"):
                                    result = {
                                        "ok": True,
                                        "target_company": target_company,
                                        "parsed": compatibility_parsed,
                                        "message": (
                                            "Posted to Tally successfully after retrying the sales item invoice "
                                            "with the compatibility header.\n\n"
                                            f"Target Company: {target_company}\n"
                                            f"Created: {compatibility_parsed.get('created', 0)}\n"
                                            f"Altered: {compatibility_parsed.get('altered', 0)}\n"
                                            f"Ignored: {compatibility_parsed.get('ignored', 0)}"
                                        ),
                                    }
                                else:
                                    parsed = compatibility_parsed

                            if not result.get("ok"):
                                detail = parsed.get("error") or "Unknown Tally exception"
                                line_errors = parsed.get("line_errors") or []
                                if line_errors:
                                    detail = line_errors[0]
                                result = {
                                    "ok": False,
                                    "target_company": target_company,
                                    "parsed": parsed,
                                    "detail": detail,
                                }

                    except Exception as exc:
                        result = {
                            "ok": False,
                            "target_company": target_company,
                            "parsed": {
                                "created": 0,
                                "altered": 0,
                                "deleted": 0,
                                "ignored": 0,
                                "errors": 1,
                                "exceptions": 1,
                                "line_errors": [],
                                "error": str(exc),
                            },
                            "detail": str(exc),
                        }
                        try:
                            self._append_debug_log(mode, company or "", "", str(exc), result["parsed"], note="python_exception")
                        except Exception:
                            pass

                    def done():
                        self._set_push_loading_state(False)
                        if result.get("ok"):
                            parsed_local = result.get("parsed", {})
                            self.status_var.set(f"Posted to Tally ({result.get('target_company', target_company)})")
                            messagebox.showinfo("Tally Response", result.get("message", "Posted to Tally."))
                            return

                        parsed_local = result.get("parsed", {})
                        detail = result.get("detail") or result.get("error") or "Unknown Tally exception"
                        self.status_var.set("Push failed (see debug log)")
                        messagebox.showerror(
                            "Push Failed",
                            "Tally returned an exception/error while importing voucher.\n\n"
                            f"Target Company: {result.get('target_company', target_company)}\n"
                            f"Errors: {parsed_local.get('errors', 0)}\n"
                            f"Exceptions: {parsed_local.get('exceptions', 0)}\n"
                            f"Detail: {detail}\n\n"
                            f"Debug Log Saved: {self.debug_log_path}",
                        )

                    self.after(0, done)

                threading.Thread(target=worker, daemon=True).start()
                return
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
        except Exception as e:
            self.status_var.set(f"Error: {e}")
            try:
                parsed = {
                    "created": 0,
                    "altered": 0,
                    "deleted": 0,
                    "ignored": 0,
                    "errors": 1,
                    "exceptions": 1,
                    "line_errors": [],
                    "error": str(e),
                }
                self._append_debug_log(mode, company or "", "", str(e), parsed, note="python_exception")
            except Exception:
                pass
            messagebox.showerror("Error", str(e))

    # ── LEDGER CREATION TAB ─────────────────────────────────────────────

    def _build_ledger_tab(self):
        parent = self.tab_ledger

        info = ctk.CTkLabel(parent, text="Create Party / Sales / Tax Ledgers in TallyPrime",
                             font=("Segoe UI", 13, "bold"), text_color=TEXT_PRIMARY)
        info.pack(pady=(10,5))

        main = ctk.CTkFrame(parent, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=10, pady=5)
        main.grid_columnconfigure(0, weight=1, minsize=390)
        main.grid_columnconfigure(1, weight=2)
        main.grid_rowconfigure(0, weight=1)

        # ── Shared state ──────────────────────────────────────────────────
        self._ledger_list = []
        self._ledger_gst_searcher = None   # lazy _GSTPortalSearcher
        self._ledger_captcha_photo = None  # prevent GC
        ledger_edit_index = None

        # ── Helper ────────────────────────────────────────────────────────
        def _set_field(widget, value=""):
            text = str(value or "").strip()
            if isinstance(widget, ctk.CTkComboBox):
                options = list(widget.cget("values") or [])
                if text in options:
                    widget.set(text)
                elif not text and options:
                    widget.set(options[0])
                else:
                    widget.set(text)
                return
            widget.delete(0, "end")
            if text:
                widget.insert(0, text)

        # ── RIGHT: Queue table + action buttons (built first so callbacks can ref it) ──
        right = ctk.CTkFrame(main, fg_color="transparent")
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=1)

        led_tree = ttk.Treeview(right, columns=("Name", "Parent", "GST", "GSTIN", "Rate"),
                                 show="headings", height=8)
        for col_name, col_w in (("Name", 160), ("Parent", 130), ("GST", 90), ("GSTIN", 150), ("Rate", 60)):
            led_tree.heading(col_name, text=col_name)
            led_tree.column(col_name, width=col_w)
        led_tree.grid(row=0, column=0, sticky="nsew", pady=(0, 8))

        self._led_count_label = ctk.CTkLabel(right, text="0 ledger(s) queued",
                                              font=("Segoe UI", 11), text_color=TEXT_MUTED)
        self._led_count_label.grid(row=1, column=0, sticky="w")

        def _enqueue(entry: dict):
            """Add a ledger entry dict to list + tree."""
            self._ledger_list.append(entry)
            led_tree.insert("", "end", values=(
                entry.get("Name", ""), entry.get("Parent", ""),
                entry.get("GSTApplicable", ""), entry.get("GSTIN", ""),
                entry.get("GSTRate", ""),
            ))
            self._led_count_label.configure(text=f"{len(self._ledger_list)} ledger(s) queued")

        def _update_queue_row(idx: int, entry: dict):
            """Update existing queue row at index idx."""
            self._ledger_list[idx] = entry
            item_ids = led_tree.get_children()
            if idx < len(item_ids):
                led_tree.item(item_ids[idx], values=(
                    entry.get("Name", ""), entry.get("Parent", ""),
                    entry.get("GSTApplicable", ""), entry.get("GSTIN", ""),
                    entry.get("GSTRate", ""),
                ))

        # ── LEFT: CTkTabview — Manual | Fetch from GST ───────────────────
        tab_view = ctk.CTkTabview(main, width=400, fg_color=COLORS["bg_card"],
                                   segmented_button_selected_color=COLORS["accent"],
                                   segmented_button_selected_hover_color=COLORS["accent_hover"])
        tab_view.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)

        tab_manual = tab_view.add("✏  Manual")
        tab_fetch  = tab_view.add("🌐  Fetch from GST")

        # ════════════════════════════════════════════════════════════════
        #  MANUAL TAB
        # ════════════════════════════════════════════════════════════════
        form = ctk.CTkScrollableFrame(
            tab_manual,
            fg_color=COLORS["bg_input"],
            corner_radius=8,
            border_width=0,
        )
        form.pack(fill="both", expand=True)
        form.grid_columnconfigure(0, weight=1)

        m_fields = {}

        def _mlabel(txt): ctk.CTkLabel(form, text=txt, font=("Segoe UI", 11),
                                        text_color=COLORS["text_secondary"]).pack(anchor="w", padx=12, pady=(6,0))
        def _mentry(key, ph=""):
            m_fields[key] = ctk.CTkEntry(form, placeholder_text=ph,
                                          fg_color=COLORS["bg_card"], border_color=COLORS["border"],
                                          text_color=COLORS["text_primary"])
            m_fields[key].pack(fill="x", padx=12, pady=(0,2))

        def _mcombo(key, values, default=None):
            m_fields[key] = ctk.CTkComboBox(form, values=values, state="readonly",
                                              fg_color=COLORS["bg_card"], border_color=COLORS["border"],
                                              button_color=COLORS["accent"],
                                              button_hover_color=COLORS["accent_hover"],
                                              text_color=COLORS["text_primary"])
            m_fields[key].set(default or values[0])
            m_fields[key].pack(fill="x", padx=12, pady=(0,2))

        _mlabel("Ledger Name *");      _mentry("led_name",     "e.g. ABC Traders")
        _mlabel("Parent Group *");     _mcombo("led_parent",   LEDGER_PARENT_OPTIONS, LEDGER_PARENT_OPTIONS[0])
        _mlabel("GST Applicable");     _mcombo("led_gst_app",  LEDGER_GST_APPLICABLE_OPTIONS, "Not Applicable")
        _mlabel("GSTIN");              _mentry("led_gstin",    "e.g. 07AAACR1718Q1ZZ")
        _mlabel("PAN / IT No.");       _mentry("led_pan",      "Auto-filled from GSTIN")
        _mlabel("State");              _mcombo("led_state",    LEDGER_STATE_OPTIONS, "Not Applicable")
        _mlabel("Address Line 1");     _mentry("led_addr1",    "e.g. Street / Building")
        _mlabel("Address Line 2");     _mentry("led_addr2",    "e.g. Area / Locality")
        _mlabel("GST Rate %");         _mentry("led_gst_rate", "e.g. 9")

        def _manual_gstin_changed(*_):
            gstin_val = m_fields["led_gstin"].get().strip().upper()
            pan_val = _pan_from_gstin(gstin_val)
            if pan_val:
                _set_field(m_fields["led_pan"], pan_val)

        m_fields["led_gstin"].bind("<FocusOut>", _manual_gstin_changed)
        m_fields["led_gstin"].bind("<KeyRelease>", _manual_gstin_changed)

        def _clear_manual_form():
            _set_field(m_fields["led_name"], "")
            _set_field(m_fields["led_parent"], LEDGER_PARENT_OPTIONS[0])
            _set_field(m_fields["led_gst_app"], "Not Applicable")
            _set_field(m_fields["led_gstin"], "")
            _set_field(m_fields["led_pan"], "")
            _set_field(m_fields["led_state"], "Not Applicable")
            _set_field(m_fields["led_addr1"], "")
            _set_field(m_fields["led_addr2"], "")
            _set_field(m_fields["led_gst_rate"], "")

        def add_manual_ledger():
            nonlocal ledger_edit_index
            name      = m_fields["led_name"].get().strip()
            parent_grp= m_fields["led_parent"].get().strip()
            if not name or not parent_grp:
                messagebox.showwarning("Required", "Ledger Name and Parent Group are required.")
                return
            gst_app  = m_fields["led_gst_app"].get().strip()
            gstin    = m_fields["led_gstin"].get().strip().upper()
            pan_val  = m_fields["led_pan"].get().strip().upper() or _pan_from_gstin(gstin)
            state    = _normalize_state_for_ledger(m_fields["led_state"].get().strip())
            entry = {
                "Name": name, "Parent": parent_grp,
                "GSTApplicable": gst_app,
                "GSTIN": gstin,
                "PAN": pan_val,
                "StateOfSupply": state,
                "Address1": m_fields["led_addr1"].get().strip(),
                "Address2": m_fields["led_addr2"].get().strip(),
                "MailingName": name,
                "Country": "India", "Pincode": "",
                "Billwise": "Yes" if _is_party_parent(parent_grp) else "No",
                "GSTRate": m_fields["led_gst_rate"].get().strip(),
            }
            if ledger_edit_index is None:
                _enqueue(entry)
            else:
                _update_queue_row(ledger_edit_index, entry)
                ledger_edit_index = None
                add_ledger_btn.configure(text="➕  Add to Queue")
                gst_add_btn.configure(text="✅  Add to Queue", fg_color=ACCENT, hover_color=ACCENT_HOVER)
            _clear_manual_form()

        def edit_selected_ledger():
            nonlocal ledger_edit_index
            selected = led_tree.selection()
            if not selected:
                messagebox.showwarning("Select Row", "Select a queued ledger row to edit.")
                return
            idx = led_tree.index(selected[0])
            if idx >= len(self._ledger_list):
                return
            entry = self._ledger_list[idx]
            _set_field(m_fields["led_name"],     entry.get("Name", ""))
            _set_field(m_fields["led_parent"],   entry.get("Parent", ""))
            _set_field(m_fields["led_gst_app"],  entry.get("GSTApplicable", ""))
            _set_field(m_fields["led_gstin"],    entry.get("GSTIN", ""))
            gstin_for_pan = entry.get("GSTIN", "")
            _set_field(m_fields["led_pan"],      entry.get("PAN", "") or _pan_from_gstin(gstin_for_pan))
            _set_field(m_fields["led_state"],    entry.get("StateOfSupply", "Not Applicable") or "Not Applicable")
            _set_field(m_fields["led_addr1"],    entry.get("Address1", ""))
            _set_field(m_fields["led_addr2"],    entry.get("Address2", ""))
            _set_field(m_fields["led_gst_rate"], entry.get("GSTRate", ""))
            ledger_edit_index = idx
            add_ledger_btn.configure(text="💾  Update Selected")
            # Reset Fetch tab button so it doesn't stay in Update mode
            gst_add_btn.configure(text="✅  Add to Queue", fg_color=ACCENT, hover_color=ACCENT_HOVER)
            tab_view.set("✏  Manual")

        add_ledger_btn = ctk.CTkButton(form, text="➕  Add to Queue",
                                        fg_color=ACCENT, hover_color=ACCENT_HOVER,
                                        command=add_manual_ledger)
        add_ledger_btn.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkButton(form, text="✏️  Edit Selected", fg_color="#94A3B8", hover_color="#64748B",
                       text_color="#FFFFFF", command=edit_selected_ledger).pack(fill="x", padx=12, pady=(0, 12))

        # ════════════════════════════════════════════════════════════════
        #  FETCH FROM GST TAB
        # ════════════════════════════════════════════════════════════════
        gst_frame = ctk.CTkScrollableFrame(tab_fetch, fg_color=COLORS["bg_input"],
                                            corner_radius=8, border_width=0)
        gst_frame.pack(fill="both", expand=True)
        gst_frame.grid_columnconfigure(0, weight=1)

        g = {}  # widgets for fetch tab

        def _glabel(txt, bold=False):
            ctk.CTkLabel(gst_frame, text=txt,
                          font=("Segoe UI", 11, "bold" if bold else "normal"),
                          text_color=COLORS["text_secondary"]).pack(anchor="w", padx=12, pady=(6,0))

        def _gentry(key, ph=""):
            g[key] = ctk.CTkEntry(gst_frame, placeholder_text=ph,
                                   fg_color=COLORS["bg_card"], border_color=COLORS["border"],
                                   text_color=COLORS["text_primary"])
            g[key].pack(fill="x", padx=12, pady=(0,2))

        def _gcombo(key, values, default=None):
            g[key] = ctk.CTkComboBox(gst_frame, values=values, state="readonly",
                                      fg_color=COLORS["bg_card"], border_color=COLORS["border"],
                                      button_color=COLORS["accent"],
                                      button_hover_color=COLORS["accent_hover"],
                                      text_color=COLORS["text_primary"])
            g[key].set(default or values[0])
            g[key].pack(fill="x", padx=12, pady=(0,2))

        # ── Step 1: GSTIN + captcha ──
        _glabel("Step 1 — Enter GSTIN", bold=True)
        gstin_row = ctk.CTkFrame(gst_frame, fg_color="transparent")
        gstin_row.pack(fill="x", padx=12, pady=(4, 2))
        g["gstin"] = ctk.CTkEntry(gstin_row, placeholder_text="e.g. 07AAACB2894G1ZP",
                                   fg_color=COLORS["bg_card"], border_color=COLORS["border"],
                                   text_color=COLORS["text_primary"])
        g["gstin"].pack(side="left", fill="x", expand=True, padx=(0,8))
        g["load_cap_btn"] = ctk.CTkButton(gstin_row, text="Load Captcha", width=120,
                                           fg_color=COLORS["accent"],
                                           hover_color=COLORS["accent_hover"],
                                           command=lambda: _gst_load_captcha())
        g["load_cap_btn"].pack(side="left")

        def _gst_gstin_changed(*_):
            gstin_val = g["gstin"].get().strip().upper()
            pan_val = _pan_from_gstin(gstin_val)
            if pan_val and "g_pan" in g:
                _set_field(g["g_pan"], pan_val)

        g["gstin"].bind("<FocusOut>", _gst_gstin_changed)
        g["gstin"].bind("<KeyRelease>", _gst_gstin_changed)

        _glabel("Step 2 — Enter Captcha", bold=True)
        g["captcha_img"] = ctk.CTkLabel(gst_frame, text="— captcha not loaded —",
                                         width=300, height=70,
                                         fg_color=("gray90","gray25"), corner_radius=6)
        g["captcha_img"].pack(padx=12, pady=(4,4))

        cap_row = ctk.CTkFrame(gst_frame, fg_color="transparent")
        cap_row.pack(fill="x", padx=12, pady=(0,4))
        g["captcha_text"] = ctk.CTkEntry(cap_row, placeholder_text="Type captcha here",
                                          fg_color=COLORS["bg_card"], border_color=COLORS["border"],
                                          text_color=COLORS["text_primary"])
        g["captcha_text"].pack(side="left", fill="x", expand=True, padx=(0,8))
        g["fetch_btn"] = ctk.CTkButton(cap_row, text="Fetch Details", width=120,
                                        fg_color=COLORS["success"],
                                        hover_color="#047857",
                                        command=lambda: _gst_fetch())
        g["fetch_btn"].pack(side="left")

        g["status"] = ctk.CTkLabel(gst_frame, text="Enter GSTIN and click Load Captcha.",
                                    anchor="w", font=("Segoe UI", 10),
                                    text_color=("gray45","gray65"), wraplength=340)
        g["status"].pack(fill="x", padx=12, pady=(0,6))

        # Separator
        ctk.CTkFrame(gst_frame, height=1, fg_color=COLORS["border"]).pack(fill="x", padx=12, pady=(2,6))

        # ── Step 3: Fetched / editable fields ──
        _glabel("Step 3 — Review & Edit Details", bold=True)

        _glabel("Ledger Name *");         _gentry("g_name",     "Will be set from Trade Name")
        _glabel("Mailing Name");          _gentry("g_mailing",  "Legal / Trade Name from portal")
        _glabel("PAN / IT No.");          _gentry("g_pan",      "Auto-filled from GSTIN")
        _glabel("Parent Group *");        _gcombo("g_parent",   LEDGER_PARENT_OPTIONS, "Sundry Debtors")
        _glabel("GST Registration Type"); _gcombo("g_reg_type",
            ["Regular", "Unregistered/Consumer", "Composition", "Input Service Distributor",
             "SEZ", "Overseas", "Not Applicable"], "Regular")
        _glabel("GST Applicable");        _gcombo("g_gst_app",  LEDGER_GST_APPLICABLE_OPTIONS, "Applicable")
        _glabel("State");                 _gcombo("g_state",    LEDGER_STATE_OPTIONS, "Not Applicable")
        _glabel("Address Line 1");        _gentry("g_addr1",    "Street / Building")
        _glabel("Address Line 2");        _gentry("g_addr2",    "Area / Locality / City")
        _glabel("Pincode");               _gentry("g_pincode",  "e.g. 110020")

        def _gst_clear_details():
            _set_field(g["g_name"], "")
            _set_field(g["g_mailing"], "")
            _set_field(g["g_pan"], "")
            _set_field(g["g_parent"], "Sundry Debtors")
            _set_field(g["g_reg_type"], "Regular")
            _set_field(g["g_gst_app"], "Applicable")
            _set_field(g["g_state"], "Not Applicable")
            _set_field(g["g_addr1"], "")
            _set_field(g["g_addr2"], "")
            _set_field(g["g_pincode"], "")

        def _gst_populate(led_def: dict):
            _set_field(g["g_name"],     led_def.get("Name", ""))
            _set_field(g["g_mailing"],  led_def.get("MailingName", "") or led_def.get("Name", ""))
            pan_v = led_def.get("PAN", "") or _pan_from_gstin(led_def.get("GSTIN", ""))
            _set_field(g["g_pan"],      pan_v)
            # Keep current parent selection — user should choose
            reg = led_def.get("GSTRegistrationType", "Regular") or "Regular"
            _set_field(g["g_reg_type"], reg)
            gst_app = led_def.get("GSTApplicable", "")
            _set_field(g["g_gst_app"],  gst_app if gst_app else "Applicable")
            _set_field(g["g_state"],    led_def.get("StateOfSupply", "") or "Not Applicable")
            _set_field(g["g_addr1"],    led_def.get("Address1", ""))
            _set_field(g["g_addr2"],    led_def.get("Address2", ""))
            _set_field(g["g_pincode"],  led_def.get("Pincode", ""))

        def _gst_load_captcha():
            if not _SELENIUM_AVAILABLE:
                g["status"].configure(text="Selenium not installed — cannot load captcha.")
                return
            gstin = g["gstin"].get().strip().upper()
            if not gstin:
                g["status"].configure(text="Enter GSTIN first.")
                return
            if self._ledger_gst_searcher is None:
                self._ledger_gst_searcher = _GSTPortalSearcher()
            g["load_cap_btn"].configure(state="disabled", text="Loading…")
            g["status"].configure(text="Opening browser and loading captcha…")

            def _bg():
                try:
                    png = self._ledger_gst_searcher.load_captcha(gstin)
                    self.after(0, _show_cap, png)
                    self.after(0, lambda: g["status"].configure(
                        text="Captcha loaded. Enter the text and click Fetch Details."))
                except Exception as exc:
                    err = str(exc)
                    self.after(0, lambda e=err: g["status"].configure(text=f"Error loading captcha: {e}"))
                finally:
                    self.after(0, lambda: g["load_cap_btn"].configure(state="normal", text="Load Captcha"))

            threading.Thread(target=_bg, daemon=True).start()

        def _show_cap(png_bytes: bytes):
            if not _PIL_AVAILABLE or not png_bytes:
                g["captcha_img"].configure(text="Captcha loaded (install Pillow to render image)")
                return
            try:
                img = Image.open(io.BytesIO(png_bytes))
                img = img.resize((img.width * 2, img.height * 2),
                                  Image.LANCZOS if hasattr(Image, "LANCZOS") else Image.Resampling.LANCZOS)
                self._ledger_captcha_photo = ImageTk.PhotoImage(img)
                g["captcha_img"].configure(image=self._ledger_captcha_photo, text="")
            except Exception:
                g["captcha_img"].configure(text="Captcha image render error")

        def _gst_fetch():
            if not _SELENIUM_AVAILABLE:
                g["status"].configure(text="Selenium not installed — cannot fetch.")
                return
            gstin       = g["gstin"].get().strip().upper()
            captcha_txt = g["captcha_text"].get().strip()
            if not gstin:
                g["status"].configure(text="GSTIN is required.")
                return
            if not captcha_txt:
                g["status"].configure(text="Enter the captcha text first.")
                return
            if self._ledger_gst_searcher is None:
                self._ledger_gst_searcher = _GSTPortalSearcher()
            g["fetch_btn"].configure(state="disabled", text="Fetching…")
            g["status"].configure(text="Fetching details from GST portal…")

            def _bg():
                try:
                    data = self._ledger_gst_searcher.fetch(gstin, captcha_txt)
                    # use Trade/Legal name as default ledger name
                    details   = data.get("details") or {}
                    kd        = details.get("key_details") or {}
                    trade_nm  = str(kd.get("Trade Name") or "").strip()
                    legal_nm  = str(kd.get("Legal Name of Business") or "").strip()
                    led_name  = trade_nm or legal_nm or gstin
                    led_def   = _parse_portal_data_to_ledger_def(data, led_name, "Sundry Debtors")
                    self.after(0, _gst_populate, led_def)
                    ok_msg = f"✅ Fetched: {led_name}. Review details and click Add to Queue."
                    self.after(0, lambda m=ok_msg: g["status"].configure(text=m))
                except Exception as exc:
                    err = str(exc)
                    self.after(0, lambda e=err: g["status"].configure(
                        text=f"Fetch failed: {e}. Try reloading captcha."))
                finally:
                    self.after(0, lambda: g["fetch_btn"].configure(state="normal", text="Fetch Details"))

            threading.Thread(target=_bg, daemon=True).start()

        def _gst_build_entry() -> dict:
            """Build ledger entry dict from the Fetch tab fields."""
            name       = g["g_name"].get().strip()
            parent_grp = g["g_parent"].get().strip()
            gstin_val  = g["gstin"].get().strip().upper()
            pan_val    = g["g_pan"].get().strip().upper() or _pan_from_gstin(gstin_val)
            state_val  = _normalize_state_for_ledger(g["g_state"].get().strip())
            return {
                "Name":                name,
                "Parent":              parent_grp,
                "MailingName":         g["g_mailing"].get().strip() or name,
                "GSTIN":               gstin_val,
                "PAN":                 pan_val,
                "GSTApplicable":       g["g_gst_app"].get().strip(),
                "GSTRegistrationType": g["g_reg_type"].get().strip(),
                "StateOfSupply":       state_val,
                "Address1":            g["g_addr1"].get().strip(),
                "Address2":            g["g_addr2"].get().strip(),
                "Pincode":             g["g_pincode"].get().strip(),
                "Country":             "India",
                "Billwise":            "Yes" if _is_party_parent(parent_grp) else "No",
                "TypeOfTaxation":      "",
                "GSTRate":             "",
            }

        def _gst_reset_form_state():
            """Clear all fetch-tab fields and reset to Add mode."""
            nonlocal ledger_edit_index
            _gst_clear_details()
            _set_field(g["gstin"], "")
            _set_field(g["captcha_text"], "")
            g["captcha_img"].configure(image="", text="— captcha not loaded —")
            self._ledger_captcha_photo = None
            ledger_edit_index = None
            gst_add_btn.configure(text="✅  Add to Queue", fg_color=ACCENT, hover_color=ACCENT_HOVER)
            # Quit searcher; a fresh one is created on the next Load Captcha click
            if self._ledger_gst_searcher:
                self._ledger_gst_searcher.quit()
                self._ledger_gst_searcher = None

        def _gst_add_to_queue():
            nonlocal ledger_edit_index
            name       = g["g_name"].get().strip()
            parent_grp = g["g_parent"].get().strip()
            if not name or not parent_grp:
                messagebox.showwarning("Required", "Ledger Name and Parent Group are required.")
                return
            entry = _gst_build_entry()
            if ledger_edit_index is None:
                _enqueue(entry)
                g["status"].configure(text="Added to queue. You can fetch another GSTIN.")
            else:
                _update_queue_row(ledger_edit_index, entry)
                g["status"].configure(text=f"Updated row {ledger_edit_index + 1}. You can fetch another GSTIN.")
            _gst_reset_form_state()

        def _gst_edit_selected():
            """Load a selected queue row into the Fetch tab for editing."""
            nonlocal ledger_edit_index
            selected = led_tree.selection()
            if not selected:
                messagebox.showwarning("Select Row", "Select a queued ledger row to edit.")
                return
            idx = led_tree.index(selected[0])
            if idx >= len(self._ledger_list):
                return
            e = self._ledger_list[idx]
            # Populate GSTIN (Step 1 field)
            gstin_e = e.get("GSTIN", "")
            _set_field(g["gstin"], gstin_e)
            # Populate Step 3 editable fields
            _set_field(g["g_name"],     e.get("Name", ""))
            _set_field(g["g_mailing"],  e.get("MailingName", "") or e.get("Name", ""))
            _set_field(g["g_pan"],      e.get("PAN", "") or _pan_from_gstin(gstin_e))
            _set_field(g["g_parent"],   e.get("Parent", "Sundry Debtors"))
            _set_field(g["g_reg_type"], e.get("GSTRegistrationType", "") or "Regular")
            _set_field(g["g_gst_app"],  e.get("GSTApplicable", "") or "Applicable")
            _set_field(g["g_state"],    e.get("StateOfSupply", "") or "Not Applicable")
            _set_field(g["g_addr1"],    e.get("Address1", ""))
            _set_field(g["g_addr2"],    e.get("Address2", ""))
            _set_field(g["g_pincode"],  e.get("Pincode", ""))
            # Clear captcha since it's a manual edit (no new fetch needed)
            g["captcha_img"].configure(image="", text="— not needed for edit —")
            g["status"].configure(text=f"Editing row {idx + 1}: \"{e.get('Name', '')}\". Edit fields and click Update.")
            ledger_edit_index = idx
            gst_add_btn.configure(text="💾  Update Selected",
                                   fg_color=COLORS["warning"], hover_color="#B45309")
            # Switch to fetch tab so user sees the fields
            tab_view.set("🌐  Fetch from GST")

        # Buttons at bottom of fetch frame
        gst_add_btn = ctk.CTkButton(gst_frame, text="✅  Add to Queue",
                                     fg_color=ACCENT, hover_color=ACCENT_HOVER,
                                     command=_gst_add_to_queue)
        gst_add_btn.pack(fill="x", padx=12, pady=(12, 4))

        btn2_row = ctk.CTkFrame(gst_frame, fg_color="transparent")
        btn2_row.pack(fill="x", padx=12, pady=(0, 12))
        btn2_row.grid_columnconfigure(0, weight=1)
        btn2_row.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(btn2_row, text="✏️  Edit Selected",
                       fg_color=("#94A3B8","#475569"), hover_color=("#64748B","#334155"),
                       text_color="#FFFFFF", command=_gst_edit_selected
                       ).grid(row=0, column=0, sticky="ew", padx=(0, 4))
        ctk.CTkButton(btn2_row, text="🔄  Clear Fields",
                       fg_color=("gray65","gray35"), hover_color=("gray55","gray45"),
                       command=_gst_reset_form_state
                       ).grid(row=0, column=1, sticky="ew", padx=(4, 0))

        # ── RIGHT: Queue action buttons ───────────────────────────────────
        btn_row = ctk.CTkFrame(right, fg_color="transparent")
        btn_row.grid(row=2, column=0, sticky="ew", pady=(5, 0))
        btn_row.grid_columnconfigure(0, weight=1)
        btn_row.grid_columnconfigure(1, weight=1)

        def clear_ledgers():
            nonlocal ledger_edit_index
            self._ledger_list.clear()
            led_tree.delete(*led_tree.get_children())
            self._led_count_label.configure(text="0 ledger(s) queued")
            ledger_edit_index = None
            add_ledger_btn.configure(text="➕  Add to Queue")
            gst_add_btn.configure(text="✅  Add to Queue", fg_color=ACCENT, hover_color=ACCENT_HOVER)
            _clear_manual_form()

        def export_ledgers(action):
            company = self._get_selected_company()
            if not self._ledger_list:
                messagebox.showwarning("Empty","Add at least one ledger.")
                return
            if action == "save":
                xml = generate_ledger_xml(self._ledger_list, company)
                out = filedialog.asksaveasfilename(defaultextension=".xml",
                        initialfile="Tally_Ledgers.xml", filetypes=[("XML","*.xml")])
                if out:
                    with open(out,"w",encoding="utf-8") as f: f.write(xml)
                    messagebox.showinfo("Saved", f"Ledger XML saved!\n{out}")
            else:
                try:
                    tally_url = self._get_tally_url()
                    host, port_text = tally_url.rsplit(":", 1)
                    host = host.replace("http://", "", 1)
                    target_company = company or "Loaded company in Tally"
                    ledgers_to_push = list(self._ledger_list)
                    existing_ledger_result = _fetch_existing_ledger_names(
                        tally_url,
                        company_name=company,
                        timeout=15,
                    )
                    if existing_ledger_result.get("success"):
                        ledgers_to_push = _filter_out_existing_ledgers(
                            ledgers_to_push,
                            existing_ledger_result.get("ledgers") or set(),
                        )
                    if not ledgers_to_push:
                        self.status_var.set(f"No new ledgers to create in Tally ({target_company})")
                        messagebox.showinfo(
                            "Nothing To Create",
                            "All queued ledgers already exist in Tally, so nothing was changed.",
                        )
                        return

                    xml = generate_ledger_xml(ledgers_to_push, company)
                    self.status_var.set(f"Posting ledgers to Tally ({target_company})...")
                    resp = push_to_tally(xml, host, int(port_text))
                    self.status_var.set(f"Ledgers posted to Tally ({target_company})")
                    messagebox.showinfo("Tally Response", f"Target Company: {target_company}\n\n{resp[:1000]}")
                except ValueError as e:
                    messagebox.showerror("Invalid Settings", str(e))
                except Exception as e:
                    messagebox.showerror("Error", str(e))

        ctk.CTkButton(btn_row, text="💾 Save XML", fg_color=SUCCESS, hover_color="#15803D",
                       command=lambda: export_ledgers("save")).grid(row=0, column=0, sticky="ew", padx=(0,6), pady=(0,6))
        ctk.CTkButton(btn_row, text="🚀 Push to Tally", fg_color=ACCENT, hover_color=ACCENT_HOVER,
                       command=lambda: export_ledgers("push")).grid(row=0, column=1, sticky="ew", padx=(6,0), pady=(0,6))
        ctk.CTkButton(btn_row, text="📂 Import XML", fg_color=COLORS["warning"], hover_color="#B45309",
                       text_color="#FFFFFF",
                       command=lambda: self._import_xml_direct("ledger")).grid(row=0, column=2, sticky="ew", padx=(6,0), pady=(0,6))
        ctk.CTkButton(btn_row, text="🗑 Clear", fg_color=DANGER, hover_color="#B91C1C",
                       command=clear_ledgers).grid(row=1, column=0, sticky="ew", padx=(0,6))

        # Load from Excel
        def load_ledgers_excel():
            f = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xlsm")])
            if not f: return
            try:
                _, rows = read_excel(f)
                for r in rows:
                    entry = {
                        "Name": _row_text(r, "Name") or _row_text(r, "LedgerName") or _row_text(r, "Ledger Name"),
                        "Parent": _row_text(r, "Parent") or _row_text(r, "ParentGroup") or _row_text(r, "Parent Group") or "Sundry Debtors",
                        "GSTApplicable": _row_text(r, "GSTApplicable") or _row_text(r, "GST Applicable") or _row_text(r, "GST"),
                        "GSTIN": (_row_text(r, "GSTIN") or _row_text(r, "PartyGSTIN") or _row_text(r, "Party GSTIN")).upper(),
                        "StateOfSupply": _row_text(r, "State") or _row_text(r, "StateOfSupply") or _row_text(r, "PlaceOfSupply") or _row_text(r, "Place Of Supply"),
                        "MailingName": _row_text(r, "MailingName") or _row_text(r, "PartyMailingName") or _row_text(r, "Mailing Name"),
                        "Country": _row_text(r, "Country") or _row_text(r, "CountryOfResidence") or _row_text(r, "Country Of Residence") or "India",
                        "Pincode": _row_text(r, "Pincode") or _row_text(r, "PinCode") or _row_text(r, "PIN") or _row_text(r, "PostalCode"),
                        "Billwise": _row_text(r, "Billwise") or _row_text(r, "IsBillwise") or _row_text(r, "ISBILLWISEON"),
                        "Address1": _row_text(r, "Address1") or _row_text(r, "Address Line 1") or _row_text(r, "AddressLine1") or _row_text(r, "Address"),
                        "Address2": _row_text(r, "Address2") or _row_text(r, "Address Line 2") or _row_text(r, "AddressLine2"),
                        "TypeOfTaxation": _row_text(r, "TaxType") or _row_text(r, "TypeOfTaxation") or _row_text(r, "Tax Type"),
                        "GSTRate": _row_text(r, "GSTRate") or _row_text(r, "GST Rate") or _row_text(r, "Rate"),
                        "GSTRegistrationType": _row_text(r, "GSTRegistrationType") or _row_text(r, "GST Registration Type") or _row_text(r, "RegistrationType") or _row_text(r, "RegType"),
                    }
                    if entry["Name"]:
                        self._ledger_list.append(entry)
                        led_tree.insert("","end", values=(entry["Name"], entry["Parent"],
                                        entry["GSTApplicable"], entry["GSTIN"], entry["GSTRate"]))
                self._led_count_label.configure(text=f"{len(self._ledger_list)} ledger(s) queued")
            except Exception as e:
                messagebox.showerror("Error", str(e))

        ctk.CTkButton(
            btn_row,
            text="📂 Load Excel",
            fg_color="#94A3B8",
            hover_color="#64748B",
            text_color="#FFFFFF",
            command=load_ledgers_excel,
        ).grid(row=1, column=1, sticky="ew", padx=(6,0))

    # ── STOCK ITEM CREATION TAB ─────────────────────────────────────────

    def _build_stock_tab(self):
        parent = self.tab_stock

        info = ctk.CTkLabel(parent, text="Create Stock Items (Inventory Masters) in TallyPrime",
                             font=("Segoe UI", 13, "bold"), text_color=TEXT_PRIMARY)
        info.pack(pady=(10,5))

        main = ctk.CTkFrame(parent, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=10, pady=5)
        main.grid_columnconfigure(0, weight=1, minsize=300)
        main.grid_columnconfigure(1, weight=2)
        main.grid_rowconfigure(0, weight=1)

        form = ctk.CTkScrollableFrame(
            main,
            fg_color=COLORS["bg_input"],
            corner_radius=10,
            width=360,
            border_width=1,
            border_color=COLORS["border"],
        )
        form.grid(row=0, column=0, sticky="nsew", padx=(0,10))
        form.grid_columnconfigure(0, weight=1)

        fields = {}
        configs = [
            ("Item Name *", "item_name", "e.g. Laptop Dell Inspiron"),
            ("Stock Group", "item_parent", "Primary"),
            ("Unit", "item_unit", "Nos / Pcs / Kg / Ltr"),
            ("HSN/SAC Code", "item_hsn", "e.g. 84713010"),
            ("GST Rate %", "item_gst_rate", "e.g. 18"),
            ("Description", "item_desc", "Optional description"),
        ]
        for label, key, placeholder in configs:
            ctk.CTkLabel(form, text=label, font=("Segoe UI", 11), text_color=COLORS["text_secondary"]).pack(anchor="w", padx=12, pady=(6,0))
            e = ctk.CTkEntry(
                form,
                placeholder_text=placeholder,
                fg_color=COLORS["bg_card"],
                border_color=COLORS["border"],
                text_color=COLORS["text_primary"],
            )
            e.pack(fill="x", padx=12, pady=(0,2))
            fields[key] = e

        self._stock_list = []
        stock_edit_index = None

        def add_item():
            nonlocal stock_edit_index
            name = fields["item_name"].get().strip()
            if not name:
                messagebox.showwarning("Required","Item Name is required.")
                return
            parent_group = _normalize_stock_group_name(fields["item_parent"].get())
            unit_name = _normalize_stock_unit_name(fields["item_unit"].get())
            entry = {
                "Name": name,
                "Parent": parent_group,
                "Unit": unit_name,
                "HSNCode": fields["item_hsn"].get().strip(),
                "GSTRate": fields["item_gst_rate"].get().strip(),
                "GSTApplicable": "Applicable" if fields["item_gst_rate"].get().strip() else "",
                "Description": fields["item_desc"].get().strip(),
            }
            row_values = (name, entry["Parent"], entry["Unit"], entry["HSNCode"], entry["GSTRate"])

            if stock_edit_index is None:
                self._stock_list.append(entry)
                stk_tree.insert("", "end", values=row_values)
            else:
                self._stock_list[stock_edit_index] = entry
                item_ids = stk_tree.get_children()
                if stock_edit_index < len(item_ids):
                    stk_tree.item(item_ids[stock_edit_index], values=row_values)
                stock_edit_index = None
                add_item_btn.configure(text="➕  Add to Queue")

            for v in fields.values():
                v.delete(0, "end")
            self._stk_count_label.configure(text=f"{len(self._stock_list)} item(s) queued")

        def edit_selected_item():
            nonlocal stock_edit_index
            selected = stk_tree.selection()
            if not selected:
                messagebox.showwarning("Select Row", "Select a queued stock item row to edit.")
                return

            item_id = selected[0]
            idx = stk_tree.index(item_id)
            if idx >= len(self._stock_list):
                messagebox.showwarning("Selection Error", "Selected row is out of sync with queue.")
                return

            entry = self._stock_list[idx]
            fields["item_name"].delete(0, "end")
            fields["item_name"].insert(0, entry.get("Name", ""))
            fields["item_parent"].delete(0, "end")
            fields["item_parent"].insert(0, entry.get("Parent", ""))
            fields["item_unit"].delete(0, "end")
            fields["item_unit"].insert(0, entry.get("Unit", ""))
            fields["item_hsn"].delete(0, "end")
            fields["item_hsn"].insert(0, entry.get("HSNCode", ""))
            fields["item_gst_rate"].delete(0, "end")
            fields["item_gst_rate"].insert(0, entry.get("GSTRate", ""))
            fields["item_desc"].delete(0, "end")
            fields["item_desc"].insert(0, entry.get("Description", ""))

            stock_edit_index = idx
            add_item_btn.configure(text="💾  Update Selected")

        add_item_btn = ctk.CTkButton(
            form,
            text="➕  Add to Queue",
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER,
            command=add_item,
        )
        add_item_btn.pack(fill="x", padx=12, pady=(10,6))
        ctk.CTkButton(
            form,
            text="✏️  Edit Selected",
            fg_color="#94A3B8",
            hover_color="#64748B",
            text_color="#FFFFFF",
            command=edit_selected_item,
        ).pack(fill="x", padx=12, pady=(0,12))

        right = ctk.CTkFrame(main, fg_color="transparent")
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=1)

        stk_tree = ttk.Treeview(right, columns=("Name","Group","Unit","HSN","GST%"),
                                 show="headings", height=8)
        for c in ("Name","Group","Unit","HSN","GST%"):
            stk_tree.heading(c, text=c)
            stk_tree.column(c, width=120)
        stk_tree.grid(row=0, column=0, sticky="nsew", pady=(0,8))

        self._stk_count_label = ctk.CTkLabel(right, text="0 item(s) queued",
                                              font=("Segoe UI", 11), text_color=TEXT_MUTED)
        self._stk_count_label.grid(row=1, column=0, sticky="w")

        btn_row = ctk.CTkFrame(right, fg_color="transparent")
        btn_row.grid(row=2, column=0, sticky="ew", pady=(5,0))
        btn_row.grid_columnconfigure(0, weight=1)
        btn_row.grid_columnconfigure(1, weight=1)

        def clear_stock():
            nonlocal stock_edit_index
            self._stock_list.clear()
            stk_tree.delete(*stk_tree.get_children())
            self._stk_count_label.configure(text="0 item(s) queued")
            stock_edit_index = None
            add_item_btn.configure(text="➕  Add to Queue")

        def export_stock(action):
            company = self._get_selected_company()
            if not self._stock_list:
                messagebox.showwarning("Empty","Add at least one stock item.")
                return
            normalized_items = []
            for item in self._stock_list:
                normalized = dict(item)
                normalized["Parent"] = _normalize_stock_group_name(normalized.get("Parent", ""))
                normalized["Unit"] = _normalize_stock_unit_name(normalized.get("Unit", "Nos"))
                normalized_items.append(normalized)

            xml = generate_stockitem_xml(normalized_items, company)
            if action == "save":
                out = filedialog.asksaveasfilename(defaultextension=".xml",
                        initialfile="Tally_StockItems.xml", filetypes=[("XML","*.xml")])
                if out:
                    with open(out,"w",encoding="utf-8") as f: f.write(xml)
                    messagebox.showinfo("Saved", f"Stock Item XML saved!\n{out}")
            else:
                try:
                    tally_url = self._get_tally_url()
                    host, port_text = tally_url.rsplit(":", 1)
                    host = host.replace("http://", "", 1)
                    target_company = company or "Loaded company in Tally"

                    # Ensure required units exist before creating stock items.
                    required_units = _collect_required_stock_units(normalized_items)
                    if required_units:
                        unit_xml = generate_unit_xml(required_units, company)
                        unit_resp = push_to_tally(unit_xml, host, int(port_text))
                        unit_parsed = _parse_tally_response_details(unit_resp)
                        self._append_debug_log(
                            "stock-unit",
                            target_company,
                            unit_xml,
                            unit_resp,
                            unit_parsed,
                            note=f"units={','.join(required_units)}",
                        )
                        unit_blockers = []
                        for err in unit_parsed.get("line_errors", []):
                            err_text = err.lower()
                            if "already exists" in err_text or "already present" in err_text:
                                continue
                            unit_blockers.append(err)
                        if unit_blockers:
                            messagebox.showerror(
                                "Unit Error",
                                "Could not create required unit masters in Tally.\n\n"
                                f"Target Company: {target_company}\n"
                                f"Detail: {unit_blockers[0]}",
                            )
                            return

                    # Ensure parent stock groups exist before creating stock items.
                    required_groups = _collect_required_stock_groups(normalized_items)
                    if required_groups:
                        group_blockers = ["Unknown stock group create failure"]
                        for parent_mode, force_parent in (("", False), ("Primary", True)):
                            group_xml = generate_stockgroup_xml(
                                required_groups,
                                company,
                                default_parent=parent_mode,
                                force_primary_parent=force_parent,
                            )
                            group_resp = push_to_tally(group_xml, host, int(port_text))
                            group_parsed = _parse_tally_response_details(group_resp)
                            self._append_debug_log(
                                "stock-group",
                                target_company,
                                group_xml,
                                group_resp,
                                group_parsed,
                                note=f"parent_mode={parent_mode or 'root'}, groups={','.join(required_groups)}",
                            )
                            blockers = []
                            for err in group_parsed.get("line_errors", []):
                                err_text = err.lower()
                                if "already exists" in err_text or "already present" in err_text:
                                    continue
                                blockers.append(err)
                            if not blockers:
                                group_blockers = []
                                break
                            group_blockers = blockers

                        if group_blockers:
                            messagebox.showerror(
                                "Stock Group Error",
                                "Could not create required stock groups in Tally.\n\n"
                                f"Target Company: {target_company}\n"
                                f"Detail: {group_blockers[0]}",
                            )
                            return

                    self.status_var.set(f"Posting stock items to Tally ({target_company})...")
                    resp = push_to_tally(xml, host, int(port_text))
                    parsed = _parse_tally_response_details(resp)
                    self._append_debug_log("stock-master", target_company, xml, resp, parsed, note="stock_item_push")

                    if parsed.get("success"):
                        self.status_var.set(f"Stock items posted to Tally ({target_company})")
                        messagebox.showinfo(
                            "Tally Response",
                            "Stock items posted successfully.\n\n"
                            f"Target Company: {target_company}\n"
                            f"Created: {parsed.get('created', 0)}\n"
                            f"Altered: {parsed.get('altered', 0)}\n"
                            f"Ignored: {parsed.get('ignored', 0)}",
                        )
                    else:
                        detail = parsed.get("error") or "Unknown Tally error"
                        if parsed.get("line_errors"):
                            detail = parsed["line_errors"][0]
                        self.status_var.set("Stock item push failed")
                        messagebox.showerror(
                            "Push Failed",
                            "Tally returned an exception/error while importing stock items.\n\n"
                            f"Target Company: {target_company}\n"
                            f"Errors: {parsed.get('errors', 0)}\n"
                            f"Exceptions: {parsed.get('exceptions', 0)}\n"
                            f"Detail: {detail}",
                        )
                except ValueError as e:
                    messagebox.showerror("Invalid Settings", str(e))
                except Exception as e:
                    messagebox.showerror("Error", str(e))

        ctk.CTkButton(btn_row, text="💾 Save XML", fg_color=SUCCESS, hover_color="#15803D",
                       command=lambda: export_stock("save")).grid(row=0, column=0, sticky="ew", padx=(0,6), pady=(0,6))
        ctk.CTkButton(btn_row, text="🚀 Push to Tally", fg_color=ACCENT, hover_color=ACCENT_HOVER,
                       command=lambda: export_stock("push")).grid(row=0, column=1, sticky="ew", padx=(6,0), pady=(0,6))
        ctk.CTkButton(btn_row, text="📂 Import XML", fg_color=COLORS["warning"], hover_color="#B45309",
                       text_color="#FFFFFF",
                       command=lambda: self._import_xml_direct("stock")).grid(row=0, column=2, sticky="ew", padx=(6,0), pady=(0,6))
        ctk.CTkButton(btn_row, text="🗑 Clear", fg_color=DANGER, hover_color="#B91C1C",
                       command=clear_stock).grid(row=1, column=0, sticky="ew", padx=(0,6))

        def load_stock_excel():
            f = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xlsm")])
            if not f: return
            try:
                _, rows = read_excel(f)
                for r in rows:
                    parent_group = _normalize_stock_group_name(
                        r.get("Parent", "") or r.get("StockGroup", "") or "Primary"
                    )
                    unit_name = _normalize_stock_unit_name(r.get("Unit", "") or "Nos")
                    entry = {
                        "Name": str(r.get("Name","") or r.get("ItemName","") or ""),
                        "Parent": parent_group,
                        "Unit": unit_name,
                        "HSNCode": str(r.get("HSNCode","") or r.get("HSN","") or ""),
                        "GSTRate": str(r.get("GSTRate","") or ""),
                        "GSTApplicable": "Applicable",
                        "Description": str(r.get("Description","") or ""),
                    }
                    if entry["Name"]:
                        self._stock_list.append(entry)
                        stk_tree.insert("","end", values=(entry["Name"], entry["Parent"],
                                        entry["Unit"], entry["HSNCode"], entry["GSTRate"]))
                self._stk_count_label.configure(text=f"{len(self._stock_list)} item(s) queued")
            except Exception as e:
                messagebox.showerror("Error", str(e))

        ctk.CTkButton(
            btn_row,
            text="📂 Load Excel",
            fg_color="#94A3B8",
            hover_color="#64748B",
            text_color="#FFFFFF",
            command=load_stock_excel,
        ).grid(row=1, column=1, sticky="ew", padx=(6,0))
    # ─── PANEL SWITCHING ──────────────────────────────────────────────────────

    def _switch_panel(self, selected_label: str):
        """Raise the panel matching the dropdown selection."""
        key = self._panel_option_to_key.get(selected_label, "accounting")
        # Update note type var when a note option is selected
        if key in {"credit_note", "credit_note_accounting", "credit_note_item"}:
            is_item = (key == "credit_note_item")
            self._note_entry_mode = "item" if is_item else "accounting"
            self._note_type_var.set("Credit Note (Item)" if is_item else "Credit Note (Accounting)")
            if hasattr(self, "_note_type_display_label"):
                suffix = "Item" if is_item else "Accounting"
                self._note_type_display_label.configure(text=f"Mode: Credit Note ({suffix}) 📝",
                                                         text_color=COLORS["success"])
        elif key in {"debit_note", "debit_note_accounting", "debit_note_item"}:
            is_item = (key == "debit_note_item")
            self._note_entry_mode = "item" if is_item else "accounting"
            self._note_type_var.set("Debit Note (Item)" if is_item else "Debit Note (Accounting)")
            if hasattr(self, "_note_type_display_label"):
                suffix = "Item" if is_item else "Accounting"
                self._note_type_display_label.configure(text=f"Mode: Debit Note ({suffix}) 📒",
                                                         text_color=COLORS["warning"])
        elif key == "journal":
            pass  # no special state to set
        panel = self._panels.get(key)
        if panel:
            panel.tkraise()

    # ─── PREVIEW MODE TOGGLE ──────────────────────────────────────────────────

    def _show_preview_mode(self, excel_frame, xml_frame, mode: str,
                           excel_btn, xml_btn, mode_key: str = ""):
        """Raise the correct preview frame, update toggle styles, track active view."""
        if mode_key:
            self._active_preview_mode[mode_key] = mode
        if mode == "xml":
            xml_frame.tkraise()
            xml_btn.configure(fg_color=COLORS["accent"], text_color="#FFFFFF")
            excel_btn.configure(fg_color=COLORS["bg_input"], text_color=COLORS["text_muted"])
        else:
            excel_frame.tkraise()
            excel_btn.configure(fg_color=COLORS["accent"], text_color="#FFFFFF")
            xml_btn.configure(fg_color=COLORS["bg_input"], text_color=COLORS["text_muted"])

    # ─── XML PREVIEW ──────────────────────────────────────────────────────────

    def _load_xml_preview(self, filepath: str, xml_text_widget, mode: str, info_label):
        """Read an XML file and show its pretty-printed content in the text widget."""
        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                raw = f.read()

            import xml.dom.minidom as _minidom
            try:
                dom = _minidom.parseString(raw.encode("utf-8", errors="replace"))
                pretty = dom.toprettyxml(indent="  ")
                # Strip the <?xml ...?> declaration for cleaner display
                lines = pretty.splitlines()
                if lines and lines[0].startswith("<?xml"):
                    pretty = "\n".join(lines[1:])
            except Exception:
                pretty = raw  # show raw if XML is malformed

            xml_text_widget.configure(state="normal")
            xml_text_widget.delete("1.0", "end")
            xml_text_widget.insert("end", pretty)
            xml_text_widget.configure(state="disabled")

            # Count vouchers for info label
            import xml.etree.ElementTree as _ET
            count_str = ""
            try:
                root = _ET.fromstring(raw)
                vouchers = root.findall(".//VOUCHER")
                masters = root.findall(".//LEDGER") + root.findall(".//STOCKITEM")
                if vouchers:
                    count_str = f"📄 {len(vouchers)} voucher(s)"
                elif masters:
                    count_str = f"📄 {len(masters)} master(s)"
                else:
                    count_str = "📄 XML loaded"
            except Exception:
                count_str = "📄 XML loaded"

            info_text = f"{count_str} — {len(raw):,} bytes"
            if info_label:
                info_label.configure(text=info_text)
            self.status_var.set(f"XML loaded: {os.path.basename(filepath)}")
        except Exception as e:
            messagebox.showerror("XML Load Error", str(e))

    # ─── IMPORT XML (Voucher modes) ───────────────────────────────────────────

    def _import_xml(self, mode: str, xml_fp_var):
        """Push an existing XML file to Tally (used by voucher tabs)."""
        filepath = xml_fp_var.get() if hasattr(xml_fp_var, "get") else str(xml_fp_var)
        if not filepath or not os.path.isfile(filepath):
            messagebox.showwarning("No XML File",
                                   "Please browse and select an XML file first using the Browse XML button.")
            return
        if self._push_running:
            self.status_var.set("Push already in progress. Please wait...")
            return

        host = self.tally_host_var.get() or "localhost"
        port_text = self.tally_port_var.get() or "9000"
        try:
            port = int(port_text)
        except ValueError:
            messagebox.showerror("Invalid Port", f"Invalid port number: {port_text}")
            return

        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                xml_content = f.read()
        except Exception as e:
            messagebox.showerror("File Read Error", str(e))
            return

        self._set_push_loading_state(True, f"Importing XML to Tally...")

        def worker():
            try:
                resp = push_to_tally(xml_content, host, port)
                parsed = _parse_tally_response_details(resp)
                def done():
                    self._set_push_loading_state(False)
                    if parsed.get("success"):
                        self.status_var.set("XML imported to Tally successfully")
                        messagebox.showinfo(
                            "Import Successful",
                            f"XML file imported to Tally successfully.\n\n"
                            f"File: {os.path.basename(filepath)}\n"
                            f"Created: {parsed.get('created', 0)}\n"
                            f"Altered: {parsed.get('altered', 0)}\n"
                            f"Ignored: {parsed.get('ignored', 0)}",
                        )
                    else:
                        err = parsed.get("error") or "Unknown Tally error"
                        self.status_var.set("Import failed")
                        messagebox.showerror("Import Failed",
                                             f"Tally rejected the import:\n\n{err}\n\n"
                                             f"Raw response (first 500 chars):\n{resp[:500]}")
                self.after(0, done)
            except Exception as e:
                def done_err():
                    self._set_push_loading_state(False)
                    messagebox.showerror("Error", str(e))
                self.after(0, done_err)

        threading.Thread(target=worker, daemon=True).start()

    # ─── IMPORT XML DIRECT (Ledger/Stock tabs — browse a file) ───────────────

    def _import_xml_direct(self, panel_hint: str = ""):
        """Browse for an XML file and push it directly to Tally (ledger/stock tabs)."""
        if self._push_running:
            self.status_var.set("Push already in progress. Please wait...")
            return

        f = filedialog.askopenfilename(
            title="Select XML file to import into Tally",
            filetypes=[("XML Files", "*.xml"), ("All Files", "*.*")],
        )
        if not f:
            return

        host = self.tally_host_var.get() or "localhost"
        port_text = self.tally_port_var.get() or "9000"
        try:
            port = int(port_text)
        except ValueError:
            messagebox.showerror("Invalid Port", f"Invalid port number: {port_text}")
            return

        try:
            with open(f, "r", encoding="utf-8", errors="replace") as fh:
                xml_content = fh.read()
        except Exception as e:
            messagebox.showerror("File Read Error", str(e))
            return

        self._set_push_loading_state(True, f"Importing {os.path.basename(f)} to Tally...")

        def worker():
            try:
                resp = push_to_tally(xml_content, host, port)
                parsed = _parse_tally_response_details(resp)
                def done():
                    self._set_push_loading_state(False)
                    if parsed.get("success"):
                        self.status_var.set("XML imported to Tally successfully")
                        messagebox.showinfo(
                            "Import Successful",
                            f"XML file imported to Tally successfully.\n\n"
                            f"File: {os.path.basename(f)}\n"
                            f"Created: {parsed.get('created', 0)}\n"
                            f"Altered: {parsed.get('altered', 0)}\n"
                            f"Ignored: {parsed.get('ignored', 0)}",
                        )
                    else:
                        err = parsed.get("error") or "Unknown Tally error"
                        self.status_var.set("Import failed")
                        messagebox.showerror("Import Failed",
                                             f"Tally rejected the import:\n\n{err}\n\n"
                                             f"Raw response (first 500 chars):\n{resp[:500]}")
                self.after(0, done)
            except Exception as e:
                def done_err():
                    self._set_push_loading_state(False)
                    messagebox.showerror("Error", str(e))
                self.after(0, done_err)

        threading.Thread(target=worker, daemon=True).start()
    # ─── SMART PUSH (context-aware: Excel or XML based on active toggle) ─────

    def _smart_push(self, mode: str, fp_var, xml_fp_var):
        """Push to Tally using whichever preview is currently active."""
        active = self._active_preview_mode.get(mode, "excel")
        if active == "xml":
            self._import_xml(mode, xml_fp_var)
        else:
            self._generate(mode, "push", fp_var.get() if hasattr(fp_var, "get") else str(fp_var))

    # ─── SMART PUSH FOR NOTE PANEL ────────────────────────────────────────────

    def _smart_push_note(self, note_fp_var, note_xml_fp_var):
        """Smart push for the Credit/Debit Note panel."""
        active = self._active_preview_mode.get("note", "excel")
        if active == "xml":
            self._import_xml("note", note_xml_fp_var)
        else:
            self._generate_note("push", note_fp_var.get() if hasattr(note_fp_var, "get") else "")

    # ─── NOTE GENERATION ─────────────────────────────────────────────────────

    def _generate_note(self, action: str, filepath: str):
        if self._push_running:
            self.status_var.set("Push already in progress...")
            return

        note_type = self._note_type_var.get() or "Credit Note"
        note_entry_mode = str(getattr(self, "_note_entry_mode", "accounting") or "accounting").strip().lower()

        # Determine rows based on active tab
        if hasattr(self, '_note_source_tabs') and self._note_source_tabs is not None:
            active_tab = self._note_source_tabs.get()
            if active_tab == "Manual Entry":
                rows = getattr(self, '_note_manual_rows', [])
                source_label = "Manual Entry"
            else:
                rows = getattr(self, '_note_loaded_rows', [])
                source_label = "Excel Upload"
        else:
            rows = getattr(self, '_note_loaded_rows', [])
            source_label = "Excel Upload"

        if not rows:
            messagebox.showwarning("No Data", f"No rows available in {source_label}.")
            return

        company = self._get_selected_company()
        if action == "push" and not company and len(getattr(self, "fetched_companies", [])) > 1:
            messagebox.showwarning("Select Company", "Please select a target company before pushing.")
            return

        try:
            date_mode, custom_tally_date = self._get_voucher_date_selection()

            _cmp_gst_regs = list(self.company_gst_registrations or [])
            if _normalize_note_type(note_type) == "Debit Note" and not _cmp_gst_regs:
                try:
                    _tally_url = self._get_tally_url()
                    _gst_fetch = _fetch_company_gst_registrations(_tally_url, company_name=company, timeout=10)
                    if _gst_fetch.get("success"):
                        _cmp_gst_regs = _gst_fetch.get("registrations", [])
                        self.company_gst_registrations = _cmp_gst_regs
                except Exception:
                    pass

            _note_ro_ledger = (
                self._note_roundoff_ledger_var.get().strip()
                if self._note_roundoff_enabled_var.get()
                else ""
            )
            xml_payload, voucher_count = generate_note_xml(
                rows,
                company=company,
                date_mode=date_mode,
                custom_tally_date=custom_tally_date,
                voucher_type=note_type,
                company_gst_registrations=_cmp_gst_regs,
                entry_mode=note_entry_mode,
                round_off_ledger=_note_ro_ledger,
            )
            if voucher_count <= 0:
                messagebox.showwarning("No Vouchers", "No valid rows (TaxableValue must be > 0).")
                return

            if action == "save":
                stem = note_type.replace(" ", "")
                out = filedialog.asksaveasfilename(
                    defaultextension=".xml",
                    initialfile=f"{stem}.xml",
                    filetypes=[("XML", "*.xml")],
                )
                if not out:
                    return
                with open(out, "w", encoding="utf-8") as f:
                    f.write(xml_payload)
                self.status_var.set(f"{note_type} XML saved: {os.path.basename(out)}")
                messagebox.showinfo("Saved", f"{note_type} XML saved.\n{out}")
                return

            host = (self.tally_host_var.get() or "localhost").strip()
            port = int((self.tally_port_var.get() or "9000").strip())
            tally_url = self._get_tally_url()
            is_debit = (_normalize_note_type(note_type) == "Debit Note")
            note_rows_snap = list(rows)

            self._set_push_loading_state(True, f"Pushing {voucher_count} {note_type} voucher(s)...")
            self.status_var.set("Pushing to Tally...")

            def worker():
                try:
                    # Pre-creation: check & create missing party ledgers
                    self.after(0, lambda: self._push_message_var.set("Checking existing ledgers in Tally..."))
                    _el_result = _fetch_existing_ledger_names(tally_url, company_name=company, timeout=15)
                    existing_note_ledgers = set(_el_result.get("ledgers") or set()) if _el_result.get("success") else set()

                    auto_note_defs = _collect_auto_note_ledgers(note_rows_snap, is_debit_note=is_debit)
                    auto_note_to_create = _filter_out_existing_ledgers(auto_note_defs, existing_note_ledgers)
                    if auto_note_to_create:
                        party_defs_n = [d for d in auto_note_to_create if _is_party_parent(str(d.get("Parent", "")))]
                        non_party_n  = [d for d in auto_note_to_create if not _is_party_parent(str(d.get("Parent", "")))]
                        asked_n = []
                        for _ld in party_defs_n:
                            self.after(0, lambda: self._push_message_var.set("Waiting for user input on missing ledger..."))
                            asked_n.append(_ask_create_party_ledger(self, _ld))
                        auto_note_to_create = non_party_n + asked_n
                        self.after(0, lambda: self._push_message_var.set("Creating required ledgers in Tally..."))
                        push_to_tally(generate_ledger_xml(auto_note_to_create, company), host=host, port=port)
                        existing_note_ledgers.update(
                            str(e.get("Name", "")).strip() for e in auto_note_to_create
                            if str(e.get("Name", "")).strip()
                        )

                    # Push note XML
                    self.after(0, lambda: self._push_message_var.set(f"Pushing {note_type} vouchers..."))
                    resp = push_to_tally(xml_payload, host=host, port=port)
                    parsed = _parse_tally_response_details(resp)

                    # On failure: auto-create missing ledgers from line errors and retry
                    if not parsed.get("success"):
                        line_errs = parsed.get("line_errors", [])
                        if line_errs:
                            miss_mode = "purchase_accounting" if is_debit else "accounting"
                            miss_defs = _build_missing_ledger_defs(line_errs, note_rows_snap, miss_mode)
                            miss_defs = _filter_out_existing_ledgers(miss_defs, existing_note_ledgers)
                            if miss_defs:
                                final_miss = []
                                for _mld in miss_defs:
                                    if _is_party_parent(str(_mld.get("Parent", ""))):
                                        self.after(0, lambda: self._push_message_var.set("Waiting for user input on missing ledger..."))
                                        _mld = _ask_create_party_ledger(self, _mld)
                                    final_miss.append(_mld)
                                self.after(0, lambda: self._push_message_var.set("Creating missing ledgers and retrying..."))
                                push_to_tally(generate_ledger_xml(final_miss, company), host=host, port=port)
                                existing_note_ledgers.update(
                                    str(e.get("Name", "")).strip() for e in final_miss
                                    if str(e.get("Name", "")).strip()
                                )
                                retry_resp = push_to_tally(xml_payload, host=host, port=port)
                                retry_parsed = _parse_tally_response_details(retry_resp)
                                if retry_parsed.get("success"):
                                    parsed = retry_parsed

                    result = {"ok": True, "parsed": parsed}
                except Exception as exc:
                    result = {"ok": False, "error": str(exc)}

                def done():
                    self._set_push_loading_state(False)
                    if not result.get("ok"):
                        messagebox.showerror("Push Failed", str(result.get("error", "Unknown error")))
                        return
                    p = result["parsed"]
                    created, altered = p.get("created", 0), p.get("altered", 0)
                    errors = p.get("errors", 0)
                    summary = f"Created: {created}\nAltered: {altered}\nErrors: {errors}"
                    if p.get("success"):
                        self.status_var.set(f"{note_type} pushed: {created} created, {altered} altered")
                        messagebox.showinfo("Push Successful", summary)
                    else:
                        line_errors = p.get("line_errors", [])
                        if line_errors:
                            summary += "\n\nLine Errors:\n- " + "\n- ".join(line_errors[:8])
                        self.status_var.set(f"{note_type} push completed with errors.")
                        messagebox.showwarning("Push With Errors", summary)
                self.after(0, done)

            threading.Thread(target=worker, daemon=True).start()

        except ValueError as exc:
            messagebox.showerror("Validation Error", str(exc))
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    # ─── NOTE PANEL BUILDER ───────────────────────────────────────────────────

    NOTE_TEMPLATE_HEADERS = [
        "Date", "GSTIN", "VoucherNo", "PartyLedger", "Particular", "Item Name", "Unit",
        "Quantity", "Rate", "TaxableValue", "CGSTLedger", "CGSTRate", "SGSTLedger",
        "SGSTRate", "IGSTLedger", "IGSTRate", "Narration"
    ]

    def _build_note_panel(self, parent):
        """Build the shared Credit / Debit Note panel (tabbed: Excel Upload, Manual Entry, Create Party)."""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)

        # Row 0: Note type indicator + template
        top_row = ctk.CTkFrame(parent, fg_color="transparent")
        top_row.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))

        type_badge = ctk.CTkFrame(top_row, fg_color=COLORS["bg_input"], corner_radius=8)
        type_badge.pack(side="left")
        self._note_type_display_label = ctk.CTkLabel(
            type_badge, text="Mode: Credit Note (Accounting) 📝",
            font=("Segoe UI", 11, "bold"), text_color=COLORS["success"], padx=10, pady=4)
        self._note_type_display_label.pack()

        ctk.CTkLabel(top_row,
                     text="Select Credit/Debit Note (Accounting or Item) from the dropdown above",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left", padx=12)

        note_template_btn = ctk.CTkButton(
            top_row, text="📥  Download Template",
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], width=170,
            command=self._download_note_template)
        note_template_btn.pack(side="right")

        # Row 1: Tabview
        self._note_source_tabs = ctk.CTkTabview(
            parent,
            fg_color="transparent",
            segmented_button_fg_color=COLORS["bg_input"],
            segmented_button_selected_color=COLORS["accent"],
            segmented_button_selected_hover_color=COLORS["accent_hover"],
            segmented_button_unselected_color=COLORS["bg_input"],
            segmented_button_unselected_hover_color=COLORS["bg_card_hover"],
        )
        self._note_source_tabs.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 4))

        excel_tab = self._note_source_tabs.add("Excel Upload")
        manual_tab = self._note_source_tabs.add("Manual Entry")
        party_tab = self._note_source_tabs.add("Create Party Ledger")
        self._note_source_tabs.set("Excel Upload")

        self._build_note_excel_tab(excel_tab)
        self._build_note_manual_tab(manual_tab)
        self._build_note_create_party_tab(party_tab)

        # Row 2: Round off row
        note_ro_row = ctk.CTkFrame(parent, fg_color="transparent")
        note_ro_row.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 2))

        note_ro_cb = ctk.CTkCheckBox(note_ro_row, text="Enable Auto Round Off",
                                     variable=self._note_roundoff_enabled_var,
                                     font=("Segoe UI", 12), text_color=COLORS["text_primary"])
        note_ro_cb.pack(side="left", padx=(0, 10))

        note_ro_entry = ctk.CTkEntry(note_ro_row, textvariable=self._note_roundoff_ledger_var,
                                     placeholder_text="Search Round Off Ledger...",
                                     width=220, state="disabled",
                                     fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                                     text_color=COLORS["text_primary"])
        note_ro_entry.pack(side="left", padx=(0, 6))

        note_ro_fetch_btn = ctk.CTkButton(note_ro_row, text="Fetch Ledgers", width=110,
                                          fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                                          text_color=COLORS["text_secondary"], state="disabled")
        note_ro_fetch_btn.pack(side="left", padx=(0, 6))

        ctk.CTkLabel(note_ro_row, text="Fractional amounts go to this ledger",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left", padx=4)

        def _note_toggle_ro(_e=note_ro_entry, _b=note_ro_fetch_btn, _v=self._note_roundoff_ledger_var):
            if self._note_roundoff_enabled_var.get():
                _e.configure(state="normal"); _b.configure(state="normal")
            else:
                _e.configure(state="disabled"); _b.configure(state="disabled"); _v.set("")
        note_ro_cb.configure(command=_note_toggle_ro)

        def _note_do_fetch_ro(_e=note_ro_entry, _b=note_ro_fetch_btn, _v=self._note_roundoff_ledger_var):
            try:
                _url = self._get_tally_url()
            except (ValueError, AttributeError):
                messagebox.showerror("Settings", "Invalid Tally URL. Check host/port in Settings.")
                return
            _co = self._get_selected_company()
            _b.configure(state="disabled", text="Fetching...")
            def _worker():
                result = _fetch_tally_ledgers(_url, timeout=15, company_name=_co)
                def _done():
                    _b.configure(state="normal", text="Fetch Ledgers")
                    if result.get("success"):
                        self._note_roundoff_all_ledgers = result.get("ledgers") or []
                        self._show_ledger_picker_popup(_e, _v, self._note_roundoff_all_ledgers)
                    else:
                        messagebox.showwarning("Fetch Failed", "Could not fetch ledgers from Tally.\nIs Tally running?")
                self.after(0, _done)
            threading.Thread(target=_worker, daemon=True).start()
        note_ro_fetch_btn.configure(command=_note_do_fetch_ro)

        def _note_ro_entry_click(event, _e=note_ro_entry, _v=self._note_roundoff_ledger_var):
            if self._note_roundoff_enabled_var.get() and self._note_roundoff_all_ledgers:
                self._show_ledger_picker_popup(_e, _v, self._note_roundoff_all_ledgers)
        note_ro_entry.bind("<Button-1>", _note_ro_entry_click)

        # Row 3: Action buttons
        btn_row = ctk.CTkFrame(parent, fg_color="transparent")
        btn_row.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))

        ctk.CTkButton(
            btn_row, text="💾  Save XML File", fg_color=SUCCESS, hover_color="#15803D", width=155,
            command=lambda: self._generate_note("save", "")
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_row, text="🚀  Push to Tally", fg_color=ACCENT, hover_color=ACCENT_HOVER, width=165,
            command=lambda: self._generate_note("push", "")
        ).pack(side="left", padx=(0, 8))

    def _build_note_excel_tab(self, parent):
        """Excel Upload sub-tab for the Note panel."""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(2, weight=1)

        load_frame = ctk.CTkFrame(parent, fg_color="transparent")
        load_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))
        note_fp_var = ctk.StringVar()
        ctk.CTkEntry(load_frame, textvariable=note_fp_var,
                     placeholder_text="Select Credit/Debit Note Excel (.xlsx/.xlsm)...",
                     state="readonly", fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                     ).pack(side="left", padx=(0, 8), fill="x", expand=True)

        def browse_note_excel():
            f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm *.xls")])
            if not f:
                return
            note_fp_var.set(f)
            try:
                headers, rows = read_excel(f)
                self._note_loaded_headers = headers
                self._note_loaded_rows = rows
                if self._note_excel_tree is not None:
                    self._note_excel_tree.delete(*self._note_excel_tree.get_children())
                    self._note_excel_tree["columns"] = headers
                    for col in headers:
                        self._note_excel_tree.heading(col, text=col)
                        self._note_excel_tree.column(col, width=max(80, len(col) * 9), minwidth=50)
                    for row in rows[:300]:
                        self._note_excel_tree.insert("", "end", values=[row.get(h, "") for h in headers])
                if self._note_excel_info_label is not None:
                    self._note_excel_info_label.configure(text=f"📊 {len(rows)} row(s) — {len(headers)} column(s)")
                self.status_var.set(f"Loaded: {os.path.basename(f)}")
            except Exception as exc:
                messagebox.showerror("Load Error", str(exc))

        ctk.CTkButton(load_frame, text="Browse Excel", command=browse_note_excel,
                      width=120, fg_color=ACCENT, hover_color=ACCENT_HOVER).pack(side="left")

        info_frame = ctk.CTkFrame(parent, fg_color="transparent")
        info_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 4))
        self._note_excel_info_label = ctk.CTkLabel(info_frame, text="", font=("Segoe UI", 11),
                                                    text_color=TEXT_MUTED)
        self._note_excel_info_label.pack(side="left")

        tree_frame = ctk.CTkFrame(parent, fg_color=COLORS["bg_dark"],
                                   corner_radius=8, border_width=1, border_color=COLORS["border"])
        tree_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 8))
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        _sy = ttk.Scrollbar(tree_frame, orient="vertical")
        _sx = ttk.Scrollbar(tree_frame, orient="horizontal")
        self._note_excel_tree = ttk.Treeview(tree_frame, show="headings",
                                              yscrollcommand=_sy.set, xscrollcommand=_sx.set)
        _sy.config(command=self._note_excel_tree.yview)
        _sx.config(command=self._note_excel_tree.xview)
        self._note_excel_tree.grid(row=0, column=0, sticky="nsew")
        _sy.grid(row=0, column=1, sticky="ns")
        _sx.grid(row=1, column=0, sticky="ew")

    def _build_note_manual_tab(self, parent):
        """Manual Entry sub-tab for the Note panel."""
        _NOTE_HDRS = self.NOTE_TEMPLATE_HEADERS
        wrapper = ctk.CTkFrame(parent, fg_color="transparent")
        wrapper.pack(fill="both", expand=True, padx=10, pady=8)
        wrapper.grid_columnconfigure(0, weight=1, uniform="note_manual_split")
        wrapper.grid_columnconfigure(1, weight=1, uniform="note_manual_split")
        wrapper.grid_rowconfigure(0, weight=1)

        left_panel = ctk.CTkFrame(wrapper, fg_color=COLORS["bg_card"],
                                   border_width=1, border_color=COLORS["border"], corner_radius=10)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        left_panel.grid_columnconfigure(0, weight=1)
        left_panel.grid_rowconfigure(0, weight=1)

        form_scroll = ctk.CTkScrollableFrame(left_panel, fg_color="transparent", corner_radius=8)
        form_scroll.grid(row=0, column=0, sticky="nsew", padx=8, pady=(8, 6))
        form_card = ctk.CTkFrame(form_scroll, fg_color="transparent")
        form_card.pack(fill="x", padx=2, pady=2)

        fields = [
            ("Date", "Date", "DD-MM-YY"),
            ("VoucherNo", "Voucher No", "Optional"),
            ("GSTIN", "GSTIN", "Optional"),
            ("PartyLedger", "Party Ledger", "Required"),
            ("Particular", "Particular", "Required"),
            ("Item Name", "Item Name", "Required for item mode"),
            ("Quantity", "Quantity", "1"),
            ("Unit", "Unit", "Nos"),
            ("Rate", "Rate", "0"),
            ("TaxableValue", "Taxable Value", "Required Amount"),
            ("CGSTLedger", "CGST Ledger", "Optional"),
            ("CGSTRate", "CGST Rate", "0"),
            ("SGSTLedger", "SGST Ledger", "Optional"),
            ("SGSTRate", "SGST Rate", "0"),
            ("IGSTLedger", "IGST Ledger", "Optional"),
            ("IGSTRate", "IGST Rate", "0"),
            ("Narration", "Narration", "Optional"),
        ]

        cols = 2
        for i, (key, label, placeholder) in enumerate(fields):
            row_idx = i // cols
            col_idx = i % cols
            field_wrap = ctk.CTkFrame(form_card, fg_color="transparent")
            field_wrap.grid(row=row_idx, column=col_idx, sticky="ew", padx=8, pady=6)
            form_card.grid_columnconfigure(col_idx, weight=1)

            if key == "PartyLedger":
                top_line = ctk.CTkFrame(field_wrap, fg_color="transparent")
                top_line.pack(fill="x")
                ctk.CTkLabel(top_line, text=label, font=("Segoe UI", 10),
                             text_color=COLORS["text_secondary"]).pack(side="left")
                self._note_manual_fetch_ledger_btn = ctk.CTkButton(
                    top_line, text="Fetch", width=70, height=26,
                    font=("Segoe UI", 10, "bold"),
                    fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                    text_color=COLORS["text_secondary"],
                    command=self._note_fetch_party_ledgers_thread)
                self._note_manual_fetch_ledger_btn.pack(side="right")

                search_row = ctk.CTkFrame(field_wrap, fg_color="transparent")
                search_row.pack(fill="x", pady=(4, 2))
                search_entry = ctk.CTkEntry(
                    search_row, textvariable=self._note_manual_party_search_var,
                    placeholder_text="Search party ledger...", height=34,
                    fg_color=COLORS["bg_input"], border_color=COLORS["border"], font=("Segoe UI", 10))
                search_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
                search_entry.bind("<KeyRelease>", self._on_note_party_search_change)

                self._note_manual_party_search_clear_btn = ctk.CTkButton(
                    search_row, text="Clear", width=58, height=30,
                    fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                    text_color=COLORS["text_secondary"], font=("Segoe UI", 9, "bold"),
                    command=lambda: (self._note_manual_party_search_var.set(""), self._on_note_party_search_change()))
                self._note_manual_party_search_clear_btn.pack(side="right")

                var = ctk.StringVar(value="")
                combo = ctk.CTkComboBox(field_wrap, values=[""], variable=var, height=36,
                                        fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                                        button_color=COLORS["accent"], button_hover_color=COLORS["accent_hover"],
                                        font=("Segoe UI", 10), state="readonly")
                combo.pack(fill="x", pady=(0, 2))
                self._note_manual_party_match_label = ctk.CTkLabel(
                    field_wrap, text="Type in search box after fetching ledgers",
                    font=("Segoe UI", 9), text_color=COLORS["text_muted"])
                self._note_manual_party_match_label.pack(anchor="w")
                self._note_manual_party_search_var.trace_add("write", lambda *_: self._on_note_party_search_change())
                self._note_manual_party_ledger_combo = combo
                self._note_manual_form_vars[key] = var
                continue

            ctk.CTkLabel(field_wrap, text=label, font=("Segoe UI", 10),
                         text_color=COLORS["text_secondary"]).pack(anchor="w")
            var = ctk.StringVar(value="")
            ctk.CTkEntry(field_wrap, textvariable=var, placeholder_text=placeholder,
                         height=36, fg_color=COLORS["bg_input"], border_color=COLORS["border"]).pack(fill="x")
            self._note_manual_form_vars[key] = var

        if "Date" in self._note_manual_form_vars:
            self._note_manual_form_vars["Date"].set(datetime.today().strftime("%d-%m-%y"))

        btn_row = ctk.CTkFrame(left_panel, fg_color="transparent")
        btn_row.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 8))
        for ci in range(3):
            btn_row.grid_columnconfigure(ci, weight=1)

        add_btn = ctk.CTkButton(btn_row, text="Add Entry", fg_color=ACCENT, hover_color=ACCENT_HOVER,
                                height=34, command=self._note_add_manual_entry)
        add_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6), pady=(0, 6))

        edit_btn = ctk.CTkButton(btn_row, text="Edit Selected", fg_color="#0EA5E9", hover_color="#0284C7",
                                  text_color="#FFFFFF", height=34, command=self._note_edit_selected_manual)
        edit_btn.grid(row=0, column=1, sticky="ew", padx=3, pady=(0, 6))

        self._note_manual_update_btn = ctk.CTkButton(
            btn_row, text="Update Entry", fg_color="#10B981", hover_color="#059669",
            text_color="#FFFFFF", height=34, state="disabled", command=self._note_update_manual_entry)
        self._note_manual_update_btn.grid(row=0, column=2, sticky="ew", padx=(6, 0), pady=(0, 6))

        ctk.CTkButton(btn_row, text="Clear Form", fg_color=COLORS["bg_input"],
                      hover_color=COLORS["bg_card_hover"], text_color=COLORS["text_secondary"],
                      height=34, command=self._note_clear_manual_form).grid(row=1, column=0, sticky="ew", padx=(0, 6))

        ctk.CTkButton(btn_row, text="Remove Selected", fg_color=COLORS["warning"], hover_color="#B45309",
                      text_color="#FFFFFF", height=34, command=self._note_remove_selected_manual).grid(
            row=1, column=1, sticky="ew", padx=3)

        ctk.CTkButton(btn_row, text="Clear All", fg_color=COLORS["error"], hover_color="#B91C1C",
                      text_color="#FFFFFF", height=34, command=self._note_clear_all_manual).grid(
            row=1, column=2, sticky="ew", padx=(6, 0))

        # Right panel: review treeview
        right_panel = ctk.CTkFrame(wrapper, fg_color=COLORS["bg_card"],
                                    border_width=1, border_color=COLORS["border"], corner_radius=10)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(1, weight=1)

        right_header = ctk.CTkFrame(right_panel, fg_color="transparent")
        right_header.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 6))
        ctk.CTkLabel(right_header, text="Review (Excel Format)", font=("Segoe UI", 12, "bold"),
                     text_color=COLORS["text_primary"]).pack(side="left")
        self._note_manual_info_label = ctk.CTkLabel(right_header, text="Manual entries: 0",
                                                     font=("Segoe UI", 11), text_color=TEXT_MUTED)
        self._note_manual_info_label.pack(side="right")

        tree_frame = ctk.CTkFrame(right_panel, fg_color=COLORS["bg_dark"],
                                   corner_radius=8, border_width=1, border_color=COLORS["border"])
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))

        _sy = ttk.Scrollbar(tree_frame, orient="vertical")
        _sx = ttk.Scrollbar(tree_frame, orient="horizontal")
        self._note_manual_tree = ttk.Treeview(tree_frame, show="headings", selectmode="extended",
                                               yscrollcommand=_sy.set, xscrollcommand=_sx.set)
        _sy.config(command=self._note_manual_tree.yview)
        _sx.config(command=self._note_manual_tree.xview)
        _sy.pack(side="right", fill="y")
        _sx.pack(side="bottom", fill="x")
        self._note_manual_tree.pack(fill="both", expand=True)
        self._note_manual_tree.bind("<Double-1>", lambda e: self._note_edit_selected_manual())

        ctk.CTkLabel(right_panel, text="Double-click a row to edit it.",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).grid(
            row=2, column=0, sticky="w", padx=10, pady=(0, 8))

        self._note_populate_tree(self._note_manual_tree, _NOTE_HDRS, [])

    def _build_note_create_party_tab(self, parent):
        """Create Party Ledger sub-tab for the Note panel."""
        wrapper = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        wrapper.pack(fill="both", expand=True, padx=10, pady=8)

        card = ctk.CTkFrame(wrapper, fg_color=COLORS["bg_card"],
                             border_width=1, border_color=COLORS["border"], corner_radius=10)
        card.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(card, text="Create Party Ledger In Tally",
                     font=("Segoe UI", 13, "bold"), text_color=COLORS["text_primary"]).pack(
            anchor="w", padx=12, pady=(12, 4))
        ctk.CTkLabel(card, text="Uses current Host, Port, and selected Target Company.",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(
            anchor="w", padx=12, pady=(0, 8))

        ctk.CTkLabel(card, text="Ledger Name", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=12)
        self._note_create_party_name_entry = ctk.CTkEntry(
            card, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="Required party ledger name",
            font=("Segoe UI", 10))
        self._note_create_party_name_entry.pack(fill="x", padx=12, pady=(4, 8))

        row_parent = ctk.CTkFrame(card, fg_color="transparent")
        row_parent.pack(fill="x", padx=12, pady=(0, 8))
        ctk.CTkLabel(row_parent, text="Parent Group", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self._note_create_party_parent_cb = ctk.CTkComboBox(
            row_parent, values=["Sundry Debtors", "Sundry Creditors"], width=200, height=32,
            fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            button_color=COLORS["accent"], button_hover_color=COLORS["accent_hover"], font=("Segoe UI", 10))
        self._note_create_party_parent_cb.set("Sundry Debtors")
        self._note_create_party_parent_cb.pack(side="right")

        ctk.CTkLabel(card, text="Mailing Name", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=12)
        self._note_create_party_mailing_entry = ctk.CTkEntry(
            card, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="Optional mailing name", font=("Segoe UI", 10))
        self._note_create_party_mailing_entry.pack(fill="x", padx=12, pady=(4, 8))

        row_gst = ctk.CTkFrame(card, fg_color="transparent")
        row_gst.pack(fill="x", padx=12, pady=(0, 8))
        self._note_create_party_gstin_entry = ctk.CTkEntry(
            row_gst, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="GSTIN", font=("Segoe UI", 10))
        self._note_create_party_gstin_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self._note_create_party_pincode_entry = ctk.CTkEntry(
            row_gst, width=130, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="Pincode", font=("Segoe UI", 10))
        self._note_create_party_pincode_entry.pack(side="left", padx=(4, 0))

        row_geo = ctk.CTkFrame(card, fg_color="transparent")
        row_geo.pack(fill="x", padx=12, pady=(0, 8))
        self._note_create_party_state_entry = ctk.CTkEntry(
            row_geo, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="State", font=("Segoe UI", 10))
        self._note_create_party_state_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self._note_create_party_country_entry = ctk.CTkEntry(
            row_geo, width=130, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], font=("Segoe UI", 10))
        self._note_create_party_country_entry.insert(0, "India")
        self._note_create_party_country_entry.pack(side="left", padx=(4, 0))

        self._note_create_party_address1_entry = ctk.CTkEntry(
            card, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="Address line 1", font=("Segoe UI", 10))
        self._note_create_party_address1_entry.pack(fill="x", padx=12, pady=(0, 6))
        self._note_create_party_address2_entry = ctk.CTkEntry(
            card, height=34, fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="Address line 2", font=("Segoe UI", 10))
        self._note_create_party_address2_entry.pack(fill="x", padx=12, pady=(0, 8))

        row_billwise = ctk.CTkFrame(card, fg_color="transparent")
        row_billwise.pack(fill="x", padx=12, pady=(0, 10))
        ctk.CTkLabel(row_billwise, text="Billwise", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self._note_create_party_billwise_cb = ctk.CTkComboBox(
            row_billwise, values=["Yes", "No"], width=120, height=32,
            fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            button_color=COLORS["accent"], button_hover_color=COLORS["accent_hover"], font=("Segoe UI", 10))
        self._note_create_party_billwise_cb.set("Yes")
        self._note_create_party_billwise_cb.pack(side="right")

        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=12, pady=(0, 12))

        self._note_create_party_fetch_btn = ctk.CTkButton(
            btn_row, text="Fetch Party Ledgers", width=160, height=34,
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], command=self._note_fetch_party_ledgers_thread)
        self._note_create_party_fetch_btn.pack(side="left", padx=(0, 8))

        self._note_create_party_clear_btn = ctk.CTkButton(
            btn_row, text="Clear", width=90, height=34,
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], command=self._note_clear_create_party_form)
        self._note_create_party_clear_btn.pack(side="left", padx=(0, 8))

        self._note_create_party_create_btn = ctk.CTkButton(
            btn_row, text="Create Party Ledger", height=34,
            fg_color=COLORS["success"], hover_color="#047857",
            text_color="#FFFFFF", command=self._note_create_party_ledger_thread)
        self._note_create_party_create_btn.pack(side="left", fill="x", expand=True)

        self._note_create_party_status_var = ctk.StringVar(value="Ready to create party ledger")
        ctk.CTkLabel(card, textvariable=self._note_create_party_status_var,
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(
            anchor="w", padx=12, pady=(0, 12))

    # ─── NOTE MANUAL ENTRY METHODS ────────────────────────────────────────────

    def _note_populate_tree(self, tree, headers, rows, limit=500):
        tree.delete(*tree.get_children())
        tree["columns"] = headers
        for h in headers:
            tree.heading(h, text=h)
            tree.column(h, width=max(120, min(260, len(h) * 12)), minwidth=80)
        for idx, row in enumerate(rows[:limit]):
            values = [str(_row_get(row, h, "") or "") for h in headers]
            tree.insert("", "end", iid=str(idx), values=values)

    def _on_note_party_search_change(self, _event=None):
        combo = self._note_manual_party_ledger_combo
        if combo is None:
            return
        if not self.fetched_party_ledgers:
            combo.configure(values=[""])
            combo.set("")
            if self._note_manual_party_match_label is not None:
                self._note_manual_party_match_label.configure(text="No fetched party ledgers yet")
            return
        typed_text = (self._note_manual_party_search_var.get() or "").strip()
        typed = typed_text.casefold()
        current_value = (self._note_manual_form_vars.get("PartyLedger", ctk.StringVar()).get() or "").strip()
        if not typed:
            filtered = self.fetched_party_ledgers[:200]
        else:
            starts = [n for n in self.fetched_party_ledgers if n.casefold().startswith(typed)]
            contains = [n for n in self.fetched_party_ledgers if typed in n.casefold() and n not in starts]
            filtered = (starts + contains)[:200]
        if typed and not filtered:
            combo.configure(values=[""])
            combo.set("")
            if "PartyLedger" in self._note_manual_form_vars:
                self._note_manual_form_vars["PartyLedger"].set("")
            if self._note_manual_party_match_label is not None:
                self._note_manual_party_match_label.configure(text=f"Search '{typed_text}': no match")
            return
        display_values = filtered if filtered else self.fetched_party_ledgers[:200]
        combo.configure(values=display_values)
        if typed and display_values and current_value not in display_values:
            combo.set(display_values[0])
            if "PartyLedger" in self._note_manual_form_vars:
                self._note_manual_form_vars["PartyLedger"].set(display_values[0])
        elif current_value:
            combo.set(current_value)
        if self._note_manual_party_match_label is not None:
            shown = len(display_values)
            total = len(self.fetched_party_ledgers)
            if typed:
                self._note_manual_party_match_label.configure(
                    text=f"Search '{typed_text}': showing {shown} of {total}")
            else:
                self._note_manual_party_match_label.configure(text=f"Showing {shown} of {total} party ledgers")

    def _note_fetch_party_ledgers_thread(self, silent=False):
        if self._ledger_fetch_running:
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            if not silent:
                messagebox.showerror("Invalid Settings", str(exc))
            return
        selected_company = self._get_selected_company()
        self._ledger_fetch_running = True
        if self._note_manual_fetch_ledger_btn is not None:
            self._note_manual_fetch_ledger_btn.configure(state="disabled", text="...")
        if hasattr(self, '_note_create_party_fetch_btn') and self._note_create_party_fetch_btn is not None:
            self._note_create_party_fetch_btn.configure(state="disabled", text="Fetching...")

        def worker():
            result = _fetch_tally_ledgers(tally_url, timeout=15, company_name=selected_company)

            def done():
                self._ledger_fetch_running = False
                if self._note_manual_fetch_ledger_btn is not None:
                    self._note_manual_fetch_ledger_btn.configure(state="normal", text="Fetch")
                if hasattr(self, '_note_create_party_fetch_btn') and self._note_create_party_fetch_btn is not None:
                    self._note_create_party_fetch_btn.configure(state="normal", text="Fetch Party Ledgers")
                if result.get("success"):
                    party_ledgers = result.get("party_ledgers") or result.get("ledgers") or []
                    cleaned = []
                    seen = set()
                    for name in party_ledgers:
                        norm = _normalize_ledger_name(name)
                        if norm and _ledger_key(norm) not in seen:
                            seen.add(_ledger_key(norm))
                            cleaned.append(norm)
                    self.fetched_party_ledgers = sorted(cleaned, key=lambda x: _ledger_key(x))
                    if self._note_manual_party_ledger_combo is not None:
                        self._note_manual_party_ledger_combo.configure(
                            values=self.fetched_party_ledgers[:200] if self.fetched_party_ledgers else [""])
                    self._on_note_party_search_change()
                    self.status_var.set(f"Fetched {len(self.fetched_party_ledgers)} party ledger(s) from Tally")
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.status_var.set("Party ledger fetch failed")
                    if not silent:
                        messagebox.showwarning("Party Ledger Fetch Failed",
                                               f"Could not fetch ledgers from Tally.\n\n{err}")
            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _note_manual_row_from_form(self):
        row = {}
        for header in self.NOTE_TEMPLATE_HEADERS:
            row[header] = self._note_manual_form_vars.get(header, ctk.StringVar()).get().strip()
        return row

    def _note_validate_manual_row(self, row):
        party = str(row.get("PartyLedger", "")).strip()
        particular = str(row.get("Particular", "")).strip()
        if not party:
            messagebox.showwarning("Missing Field", "Party Ledger is required.")
            return None
        if not particular:
            messagebox.showwarning("Missing Field", "Particular is required.")
            return None
        try:
            taxable = float(str(row.get("TaxableValue") or "0").strip())
        except ValueError:
            messagebox.showwarning("Invalid Value", "Taxable Value must be numeric.")
            return None
        if taxable <= 0:
            messagebox.showwarning("Invalid Value", "Taxable Value must be greater than zero.")
            return None

        is_item_mode = str(getattr(self, "_note_entry_mode", "accounting") or "accounting").strip().lower() == "item"
        if is_item_mode:
            item_name = _row_text_any(row, ["Item Name", "ItemName", "Item", "StockItem", "ProductName"], "")
            if not item_name:
                messagebox.showwarning("Missing Field", "Item Name is required for item notes.")
                return None
            qty = _row_float(row, "Quantity", 0.0) or _row_float(row, "Qty", 0.0)
            if qty <= 0:
                qty = 1.0
            rate = _row_float(row, "Rate", 0.0)
            if rate <= 0 and taxable > 0 and qty > 0:
                rate = taxable / qty
            if not _row_text_any(row, ["Unit", "UOM", "Per"], ""):
                row["Unit"] = "Nos"
            row["Quantity"] = str(qty)
            row["Rate"] = str(rate)

        if not str(row.get("Date") or "").strip():
            row["Date"] = datetime.today().strftime("%d-%m-%y")
        for key in ["CGSTRate", "SGSTRate", "IGSTRate"]:
            if not str(row.get(key) or "").strip():
                row[key] = "0"
        return row

    def _note_set_manual_edit_mode(self, index=None):
        self._note_manual_editing_index = index
        if self._note_manual_update_btn is not None:
            self._note_manual_update_btn.configure(state="normal" if index is not None else "disabled")

    def _note_selected_manual_index(self):
        if self._note_manual_tree is None:
            return None
        selected = list(self._note_manual_tree.selection())
        if not selected:
            return None
        try:
            return int(selected[0])
        except ValueError:
            return None

    def _note_refresh_manual_tree(self, focus_index=None):
        if self._note_manual_tree is None:
            return
        self._note_populate_tree(self._note_manual_tree, self.NOTE_TEMPLATE_HEADERS,
                                  self._note_manual_rows, limit=500)
        if focus_index is not None:
            iid = str(focus_index)
            if iid in self._note_manual_tree.get_children():
                self._note_manual_tree.selection_set(iid)
                self._note_manual_tree.focus(iid)
                self._note_manual_tree.see(iid)
        if self._note_manual_info_label is not None:
            self._note_manual_info_label.configure(text=f"Manual entries: {len(self._note_manual_rows)}")

    def _note_add_manual_entry(self):
        row = self._note_manual_row_from_form()
        validated = self._note_validate_manual_row(row)
        if validated is None:
            return
        if not str(validated.get("VoucherNo") or "").strip():
            validated["VoucherNo"] = str(len(self._note_manual_rows) + 1)
        self._note_manual_rows.append(validated)
        self._note_refresh_manual_tree(focus_index=len(self._note_manual_rows) - 1)
        self._note_set_manual_edit_mode(None)
        self.status_var.set(f"Note manual entry added. Total: {len(self._note_manual_rows)}")

    def _note_edit_selected_manual(self):
        idx = self._note_selected_manual_index()
        if idx is None:
            messagebox.showinfo("Edit Entry", "Select one row in the table to edit.")
            return
        if idx < 0 or idx >= len(self._note_manual_rows):
            return
        row = self._note_manual_rows[idx]
        for header in self.NOTE_TEMPLATE_HEADERS:
            value = _row_get(row, header, "")
            if header in self._note_manual_form_vars:
                self._note_manual_form_vars[header].set("" if value is None else str(value))
        self._note_set_manual_edit_mode(idx)
        self.status_var.set(f"Editing note entry #{idx + 1}. Update Entry to save changes.")

    def _note_update_manual_entry(self):
        idx = self._note_manual_editing_index
        if idx is None:
            messagebox.showinfo("Update Entry", "Select and edit a row first.")
            return
        if idx < 0 or idx >= len(self._note_manual_rows):
            self._note_set_manual_edit_mode(None)
            messagebox.showwarning("Update Entry", "Selected row is no longer available.")
            return
        row = self._note_manual_row_from_form()
        validated = self._note_validate_manual_row(row)
        if validated is None:
            return
        if not str(validated.get("VoucherNo") or "").strip():
            validated["VoucherNo"] = str(idx + 1)
        self._note_manual_rows[idx] = validated
        self._note_refresh_manual_tree(focus_index=idx)
        self._note_set_manual_edit_mode(None)
        self.status_var.set(f"Note manual entry #{idx + 1} updated.")

    def _note_clear_manual_form(self):
        keep_date = datetime.today().strftime("%d-%m-%y")
        for key, var in self._note_manual_form_vars.items():
            if key == "Date":
                var.set(keep_date)
            elif key in {"CGSTRate", "SGSTRate", "IGSTRate"}:
                var.set("0")
            else:
                var.set("")
        self._note_set_manual_edit_mode(None)
        self._note_manual_party_search_var.set("")
        self._on_note_party_search_change()

    def _note_remove_selected_manual(self):
        if self._note_manual_tree is None:
            return
        selected = list(self._note_manual_tree.selection())
        if not selected:
            messagebox.showinfo("Remove Entry", "Select at least one row to remove.")
            return
        indexes = []
        for iid in selected:
            try:
                indexes.append(int(iid))
            except ValueError:
                continue
        for idx in sorted(indexes, reverse=True):
            if 0 <= idx < len(self._note_manual_rows):
                self._note_manual_rows.pop(idx)
        self._note_refresh_manual_tree()
        self._note_set_manual_edit_mode(None)
        self.status_var.set(f"Removed. Remaining: {len(self._note_manual_rows)}")

    def _note_clear_all_manual(self):
        if not self._note_manual_rows:
            return
        if not messagebox.askyesno("Clear All", "Remove all note manual entries?"):
            return
        self._note_manual_rows = []
        self._note_refresh_manual_tree()
        self._note_set_manual_edit_mode(None)
        self.status_var.set("All note manual entries cleared.")

    def _note_clear_create_party_form(self):
        for attr in ['_note_create_party_name_entry', '_note_create_party_mailing_entry',
                     '_note_create_party_gstin_entry', '_note_create_party_state_entry',
                     '_note_create_party_pincode_entry', '_note_create_party_address1_entry',
                     '_note_create_party_address2_entry']:
            w = getattr(self, attr, None)
            if w is not None:
                w.delete(0, "end")
        w = getattr(self, '_note_create_party_country_entry', None)
        if w is not None:
            w.delete(0, "end")
            w.insert(0, "India")
        if hasattr(self, '_note_create_party_parent_cb') and self._note_create_party_parent_cb:
            self._note_create_party_parent_cb.set("Sundry Debtors")
        if hasattr(self, '_note_create_party_billwise_cb') and self._note_create_party_billwise_cb:
            self._note_create_party_billwise_cb.set("Yes")
        if hasattr(self, '_note_create_party_status_var'):
            self._note_create_party_status_var.set("Ready to create party ledger")

    def _note_create_party_ledger_thread(self):
        name_entry = getattr(self, '_note_create_party_name_entry', None)
        ledger_name = _normalize_ledger_name(name_entry.get() if name_entry else "")
        if not ledger_name:
            messagebox.showwarning("Missing Field", "Ledger Name is required.")
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return
        selected_company = self._get_selected_company()
        parent_cb = getattr(self, '_note_create_party_parent_cb', None)
        parent_name = parent_cb.get().strip() if parent_cb else "Sundry Debtors"
        mailing_entry = getattr(self, '_note_create_party_mailing_entry', None)
        gstin_entry = getattr(self, '_note_create_party_gstin_entry', None)
        state_entry = getattr(self, '_note_create_party_state_entry', None)
        country_entry = getattr(self, '_note_create_party_country_entry', None)
        pincode_entry = getattr(self, '_note_create_party_pincode_entry', None)
        addr1_entry = getattr(self, '_note_create_party_address1_entry', None)
        addr2_entry = getattr(self, '_note_create_party_address2_entry', None)
        billwise_cb = getattr(self, '_note_create_party_billwise_cb', None)

        mailing_name = mailing_entry.get().strip() if mailing_entry else ""
        gstin = gstin_entry.get().strip() if gstin_entry else ""
        state = state_entry.get().strip() if state_entry else ""
        country = country_entry.get().strip() if country_entry else "India"
        pincode = pincode_entry.get().strip() if pincode_entry else ""
        address1 = addr1_entry.get().strip() if addr1_entry else ""
        address2 = addr2_entry.get().strip() if addr2_entry else ""
        billwise_raw = billwise_cb.get().strip().upper() if billwise_cb else "YES"
        billwise_on = billwise_raw in {"YES", "Y", "TRUE", "1"}

        status_var = getattr(self, '_note_create_party_status_var', None)
        if status_var:
            status_var.set("Creating ledger in Tally...")
        self.status_var.set("Creating party ledger...")

        def worker():
            try:
                result = _create_tally_ledger(
                    tally_url=tally_url,
                    ledger_name=ledger_name,
                    parent_name=parent_name,
                    company_name=selected_company,
                    gstin=gstin,
                    state=state,
                    country=country,
                    pincode=pincode,
                    mailing_name=mailing_name,
                    address1=address1,
                    address2=address2,
                    billwise_on=billwise_on,
                    timeout=30,
                )
            except Exception as exc:
                result = {"success": False, "error": str(exc)}

            def done():
                if result.get("success"):
                    created = int(result.get("created", 0) or 0)
                    altered = int(result.get("altered", 0) or 0)
                    if status_var:
                        status_var.set(f"Created/updated. Created={created}, Altered={altered}")
                    self.status_var.set("Party ledger created successfully")
                    if "PartyLedger" in self._note_manual_form_vars:
                        self._note_manual_form_vars["PartyLedger"].set(ledger_name)
                    self._note_fetch_party_ledgers_thread(silent=True)
                    messagebox.showinfo("Create Party Ledger",
                                        f"Ledger created/updated successfully.\n\nName: {ledger_name}\nParent: {parent_name}")
                else:
                    err = str(result.get("error") or "Ledger creation failed in Tally.")
                    if status_var:
                        status_var.set(f"Create failed: {err}")
                    self.status_var.set("Party ledger create failed")
                    messagebox.showerror("Create Party Ledger", err)
            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    # ─── NOTE TEMPLATE DOWNLOAD ───────────────────────────────────────────────

    def _download_note_template(self):
        note_type = self._note_type_var.get() or "Credit Note"
        headers = self.NOTE_TEMPLATE_HEADERS
        sample = ["22-04-2026", "18AABCI8307G1ZM", "1", "ABC Corp Pvt Ltd", "Sales Account", 20000, "", 0, "", 0, "IGST", 18, "Test Entry for ABC Corp"]
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"Template_{note_type.replace(' ', '_')}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not out:
            return
        try:
            import openpyxl as _xl
            wb = _xl.Workbook()
            ws = wb.active
            ws.title = note_type
            ws.append(headers)
            ws.append(sample)
            wb.save(out)
            messagebox.showinfo("Template Saved", f"{note_type} template saved:\n{out}")
        except Exception as exc:
            messagebox.showerror("Template Error", str(exc))

    # ─── JOURNAL PANEL BUILDER ────────────────────────────────────────────────

    JNL_TEMPLATE_HEADERS = [
        "Date", "GSTIN", "VoucherNo", "PartyLedger", "Particular", "TaxableValue",
        "CGSTLedger", "CGSTRate", "SGSTLedger", "SGSTRate", "IGSTLedger", "IGSTRate",
        "Narration"
    ]

    def _build_journal_panel(self, parent):
        """Build the Journal Entry panel."""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=0)
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_rowconfigure(2, weight=0)
        parent.grid_rowconfigure(3, weight=0)

        # Row 0: Journal Type + header
        top_row = ctk.CTkFrame(parent, fg_color=COLORS["bg_card"],
                                border_width=1, border_color=COLORS["border"], corner_radius=8)
        top_row.grid(row=0, column=0, sticky="ew", padx=10, pady=(6, 0))
        inner = ctk.CTkFrame(top_row, fg_color="transparent")
        inner.pack(fill="x", padx=14, pady=8)

        ctk.CTkLabel(inner, text="Journal Type:", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        for _lbl, _val in (("Purchase  (Dr Expense / Cr Party)", "purchase"),
                            ("Sale  (Dr Party / Cr Income)", "sale")):
            ctk.CTkRadioButton(
                inner, text=_lbl, variable=self._journal_type_var, value=_val,
                font=("Segoe UI", 10), text_color=COLORS["text_secondary"],
                fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
                border_color=COLORS["border"]).pack(side="left", padx=(12, 0))

        ctk.CTkButton(
            inner, text="📥  Download Template",
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], width=170,
            command=self._download_journal_template).pack(side="right")

        # Row 1: Tabview
        self._jnl_source_tabs = ctk.CTkTabview(
            parent,
            fg_color="transparent",
            segmented_button_fg_color=COLORS["bg_input"],
            segmented_button_selected_color=COLORS["accent"],
            segmented_button_selected_hover_color=COLORS["accent_hover"],
            segmented_button_unselected_color=COLORS["bg_input"],
            segmented_button_unselected_hover_color=COLORS["bg_card_hover"],
        )
        self._jnl_source_tabs.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 2))

        excel_tab = self._jnl_source_tabs.add("Excel Upload")
        manual_tab = self._jnl_source_tabs.add("Manual Entry")
        self._jnl_source_tabs.set("Excel Upload")

        self._build_journal_excel_tab(excel_tab)
        self._build_journal_manual_tab(manual_tab)

        # Row 2: Round off row
        jnl_ro_row = ctk.CTkFrame(parent, fg_color="transparent")
        jnl_ro_row.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 2))

        jnl_ro_cb = ctk.CTkCheckBox(jnl_ro_row, text="Enable Auto Round Off",
                                     variable=self._jnl_roundoff_enabled_var,
                                     font=("Segoe UI", 12), text_color=COLORS["text_primary"])
        jnl_ro_cb.pack(side="left", padx=(0, 10))

        jnl_ro_entry = ctk.CTkEntry(jnl_ro_row, textvariable=self._jnl_roundoff_ledger_var,
                                     placeholder_text="Search Round Off Ledger...",
                                     width=220, state="disabled",
                                     fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                                     text_color=COLORS["text_primary"])
        jnl_ro_entry.pack(side="left", padx=(0, 6))

        jnl_ro_fetch_btn = ctk.CTkButton(jnl_ro_row, text="Fetch Ledgers", width=110,
                                          fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                                          text_color=COLORS["text_secondary"], state="disabled")
        jnl_ro_fetch_btn.pack(side="left", padx=(0, 6))

        ctk.CTkLabel(jnl_ro_row, text="Fractional amounts go to this ledger",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left", padx=4)

        def _jnl_toggle_ro(_e=jnl_ro_entry, _b=jnl_ro_fetch_btn, _v=self._jnl_roundoff_ledger_var):
            if self._jnl_roundoff_enabled_var.get():
                _e.configure(state="normal"); _b.configure(state="normal")
            else:
                _e.configure(state="disabled"); _b.configure(state="disabled"); _v.set("")
        jnl_ro_cb.configure(command=_jnl_toggle_ro)

        def _jnl_do_fetch_ro(_e=jnl_ro_entry, _b=jnl_ro_fetch_btn, _v=self._jnl_roundoff_ledger_var):
            try:
                _url = self._get_tally_url()
            except (ValueError, AttributeError):
                messagebox.showerror("Settings", "Invalid Tally URL. Check host/port in Settings.")
                return
            _co = self._get_selected_company()
            _b.configure(state="disabled", text="Fetching...")
            def _worker():
                result = _fetch_tally_ledgers(_url, timeout=15, company_name=_co)
                def _done():
                    _b.configure(state="normal", text="Fetch Ledgers")
                    if result.get("success"):
                        self._jnl_roundoff_all_ledgers = result.get("ledgers") or []
                        self._show_ledger_picker_popup(_e, _v, self._jnl_roundoff_all_ledgers)
                    else:
                        messagebox.showwarning("Fetch Failed", "Could not fetch ledgers from Tally.\nIs Tally running?")
                self.after(0, _done)
            threading.Thread(target=_worker, daemon=True).start()
        jnl_ro_fetch_btn.configure(command=_jnl_do_fetch_ro)

        def _jnl_ro_entry_click(event, _e=jnl_ro_entry, _v=self._jnl_roundoff_ledger_var):
            if self._jnl_roundoff_enabled_var.get() and self._jnl_roundoff_all_ledgers:
                self._show_ledger_picker_popup(_e, _v, self._jnl_roundoff_all_ledgers)
        jnl_ro_entry.bind("<Button-1>", _jnl_ro_entry_click)

        # Row 3: Action buttons
        btn_row = ctk.CTkFrame(parent, fg_color="transparent")
        btn_row.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))

        ctk.CTkButton(
            btn_row, text="💾  Save XML File", fg_color=SUCCESS, hover_color="#15803D", width=155,
            command=lambda: self._generate_journal("save")
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_row, text="🚀  Push to Tally", fg_color=ACCENT, hover_color=ACCENT_HOVER, width=165,
            command=lambda: self._generate_journal("push")
        ).pack(side="left", padx=(0, 8))

    def _build_journal_excel_tab(self, parent):
        """Excel Upload sub-tab for Journal Entry panel."""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(2, weight=1)

        load_frame = ctk.CTkFrame(parent, fg_color="transparent")
        load_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))
        jnl_fp_var = ctk.StringVar()
        ctk.CTkEntry(load_frame, textvariable=jnl_fp_var,
                     placeholder_text="Select Journal Entry Excel (.xlsx/.xlsm/.xls)...",
                     state="readonly", fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                     ).pack(side="left", padx=(0, 8), fill="x", expand=True)

        def browse_jnl_excel():
            f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm *.xls")])
            if not f:
                return
            jnl_fp_var.set(f)
            try:
                headers, rows = read_excel(f)
                self._jnl_loaded_headers = headers
                self._jnl_loaded_rows = rows
                if self._jnl_excel_tree is not None:
                    self._jnl_excel_tree.delete(*self._jnl_excel_tree.get_children())
                    self._jnl_excel_tree["columns"] = headers
                    for col in headers:
                        self._jnl_excel_tree.heading(col, text=col)
                        self._jnl_excel_tree.column(col, width=max(80, len(col) * 9), minwidth=50)
                    for row in rows[:300]:
                        self._jnl_excel_tree.insert("", "end", values=[row.get(h, "") for h in headers])
                if self._jnl_excel_info_label is not None:
                    self._jnl_excel_info_label.configure(
                        text=f"📊 {len(rows)} row(s) — {len(headers)} column(s)")
                self.status_var.set(f"Loaded: {os.path.basename(f)}")
            except Exception as exc:
                messagebox.showerror("Load Error", str(exc))

        ctk.CTkButton(load_frame, text="Browse Excel", command=browse_jnl_excel,
                      width=120, fg_color=ACCENT, hover_color=ACCENT_HOVER).pack(side="left")

        info_frame = ctk.CTkFrame(parent, fg_color="transparent")
        info_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 4))
        self._jnl_excel_info_label = ctk.CTkLabel(info_frame, text="", font=("Segoe UI", 11),
                                                    text_color=TEXT_MUTED)
        self._jnl_excel_info_label.pack(side="left")

        tree_frame = ctk.CTkFrame(parent, fg_color=COLORS["bg_dark"],
                                   corner_radius=8, border_width=1, border_color=COLORS["border"])
        tree_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 8))
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        _sy = ttk.Scrollbar(tree_frame, orient="vertical")
        _sx = ttk.Scrollbar(tree_frame, orient="horizontal")
        self._jnl_excel_tree = ttk.Treeview(tree_frame, show="headings",
                                              yscrollcommand=_sy.set, xscrollcommand=_sx.set)
        _sy.config(command=self._jnl_excel_tree.yview)
        _sx.config(command=self._jnl_excel_tree.xview)
        self._jnl_excel_tree.grid(row=0, column=0, sticky="nsew")
        _sy.grid(row=0, column=1, sticky="ns")
        _sx.grid(row=1, column=0, sticky="ew")

    def _build_journal_manual_tab(self, parent):
        """Manual Entry sub-tab for Journal Entry panel."""
        _JNL_HDRS = self.JNL_TEMPLATE_HEADERS
        wrapper = ctk.CTkFrame(parent, fg_color="transparent")
        wrapper.pack(fill="both", expand=True, padx=10, pady=8)
        wrapper.grid_columnconfigure(0, weight=1, uniform="jnl_manual_split")
        wrapper.grid_columnconfigure(1, weight=1, uniform="jnl_manual_split")
        wrapper.grid_rowconfigure(0, weight=1)

        left_panel = ctk.CTkFrame(wrapper, fg_color=COLORS["bg_card"],
                                   border_width=1, border_color=COLORS["border"], corner_radius=10)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        left_panel.grid_columnconfigure(0, weight=1)
        left_panel.grid_rowconfigure(0, weight=1)

        form_scroll = ctk.CTkScrollableFrame(left_panel, fg_color="transparent", corner_radius=8)
        form_scroll.grid(row=0, column=0, sticky="nsew", padx=8, pady=(8, 6))
        form_card = ctk.CTkFrame(form_scroll, fg_color="transparent")
        form_card.pack(fill="x", padx=2, pady=2)

        fields = [
            ("Date", "Date", "DD-MM-YY"),
            ("VoucherNo", "Voucher No", "Optional"),
            ("PartyLedger", "Party Ledger", "Required"),
            ("Particular", "Particular", "Required"),
            ("TaxableValue", "Taxable Value", "Required Amount"),
            ("CGSTLedger", "CGST Ledger", "Optional"),
            ("CGSTRate", "CGST Rate", "0"),
            ("SGSTLedger", "SGST Ledger", "Optional"),
            ("SGSTRate", "SGST Rate", "0"),
            ("IGSTLedger", "IGST Ledger", "Optional"),
            ("IGSTRate", "IGST Rate", "0"),
            ("Narration", "Narration", "Optional"),
        ]

        cols = 2
        for i, (key, label, placeholder) in enumerate(fields):
            row_idx = i // cols
            col_idx = i % cols
            field_wrap = ctk.CTkFrame(form_card, fg_color="transparent")
            field_wrap.grid(row=row_idx, column=col_idx, sticky="ew", padx=8, pady=6)
            form_card.grid_columnconfigure(col_idx, weight=1)

            if key == "PartyLedger":
                top_line = ctk.CTkFrame(field_wrap, fg_color="transparent")
                top_line.pack(fill="x")
                ctk.CTkLabel(top_line, text=label, font=("Segoe UI", 10),
                             text_color=COLORS["text_secondary"]).pack(side="left")
                self._jnl_manual_fetch_ledger_btn = ctk.CTkButton(
                    top_line, text="Fetch", width=70, height=26,
                    font=("Segoe UI", 10, "bold"),
                    fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                    text_color=COLORS["text_secondary"],
                    command=self._jnl_fetch_party_ledgers_thread)
                self._jnl_manual_fetch_ledger_btn.pack(side="right")

                search_row = ctk.CTkFrame(field_wrap, fg_color="transparent")
                search_row.pack(fill="x", pady=(4, 2))
                search_entry = ctk.CTkEntry(
                    search_row, textvariable=self._jnl_manual_party_search_var,
                    placeholder_text="Search party ledger...", height=34,
                    fg_color=COLORS["bg_input"], border_color=COLORS["border"], font=("Segoe UI", 10))
                search_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
                search_entry.bind("<KeyRelease>", self._on_jnl_party_search_change)

                self._jnl_manual_party_search_clear_btn = ctk.CTkButton(
                    search_row, text="Clear", width=58, height=30,
                    fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                    text_color=COLORS["text_secondary"], font=("Segoe UI", 9, "bold"),
                    command=lambda: (self._jnl_manual_party_search_var.set(""), self._on_jnl_party_search_change()))
                self._jnl_manual_party_search_clear_btn.pack(side="right")

                var = ctk.StringVar(value="")
                combo = ctk.CTkComboBox(field_wrap, values=[""], variable=var, height=36,
                                        fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                                        button_color=COLORS["accent"], button_hover_color=COLORS["accent_hover"],
                                        font=("Segoe UI", 10), state="readonly")
                combo.pack(fill="x", pady=(0, 2))
                self._jnl_manual_party_match_label = ctk.CTkLabel(
                    field_wrap, text="Type in search box after fetching ledgers",
                    font=("Segoe UI", 9), text_color=COLORS["text_muted"])
                self._jnl_manual_party_match_label.pack(anchor="w")
                self._jnl_manual_party_search_var.trace_add("write", lambda *_: self._on_jnl_party_search_change())
                self._jnl_manual_party_ledger_combo = combo
                self._jnl_manual_form_vars[key] = var
                continue

            ctk.CTkLabel(field_wrap, text=label, font=("Segoe UI", 10),
                         text_color=COLORS["text_secondary"]).pack(anchor="w")
            var = ctk.StringVar(value="")
            ctk.CTkEntry(field_wrap, textvariable=var, placeholder_text=placeholder,
                         height=36, fg_color=COLORS["bg_input"], border_color=COLORS["border"]).pack(fill="x")
            self._jnl_manual_form_vars[key] = var

        if "Date" in self._jnl_manual_form_vars:
            self._jnl_manual_form_vars["Date"].set(datetime.today().strftime("%d-%m-%y"))

        btn_row = ctk.CTkFrame(left_panel, fg_color="transparent")
        btn_row.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 8))
        for ci in range(3):
            btn_row.grid_columnconfigure(ci, weight=1)

        add_btn = ctk.CTkButton(btn_row, text="Add Entry", fg_color=ACCENT, hover_color=ACCENT_HOVER,
                                height=28, font=("Segoe UI", 10, "bold"), command=self._jnl_add_manual_entry)
        add_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6), pady=(0, 4))

        edit_btn = ctk.CTkButton(btn_row, text="Edit Selected", fg_color="#0EA5E9", hover_color="#0284C7",
                                  text_color="#FFFFFF", height=28, font=("Segoe UI", 10, "bold"), command=self._jnl_edit_selected_manual)
        edit_btn.grid(row=0, column=1, sticky="ew", padx=3, pady=(0, 4))

        self._jnl_manual_update_btn = ctk.CTkButton(
            btn_row, text="Update Entry", fg_color="#10B981", hover_color="#059669",
            text_color="#FFFFFF", height=28, font=("Segoe UI", 10, "bold"), state="disabled", command=self._jnl_update_manual_entry)
        self._jnl_manual_update_btn.grid(row=0, column=2, sticky="ew", padx=(6, 0), pady=(0, 4))

        ctk.CTkButton(btn_row, text="Clear Form", fg_color=COLORS["bg_input"],
                      hover_color=COLORS["bg_card_hover"], text_color=COLORS["text_secondary"],
                      height=28, font=("Segoe UI", 10, "bold"), command=self._jnl_clear_manual_form).grid(row=1, column=0, sticky="ew", padx=(0, 6))

        ctk.CTkButton(btn_row, text="Remove Selected", fg_color=COLORS["warning"], hover_color="#B45309",
                      text_color="#FFFFFF", height=28, font=("Segoe UI", 10, "bold"), command=self._jnl_remove_selected_manual).grid(
            row=1, column=1, sticky="ew", padx=3)

        ctk.CTkButton(btn_row, text="Clear All", fg_color=COLORS["error"], hover_color="#B91C1C",
                      text_color="#FFFFFF", height=28, font=("Segoe UI", 10, "bold"), command=self._jnl_clear_all_manual).grid(
            row=1, column=2, sticky="ew", padx=(6, 0))

        # Right panel: review treeview
        right_panel = ctk.CTkFrame(wrapper, fg_color=COLORS["bg_card"],
                                    border_width=1, border_color=COLORS["border"], corner_radius=10)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(1, weight=1)

        right_header = ctk.CTkFrame(right_panel, fg_color="transparent")
        right_header.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 6))
        ctk.CTkLabel(right_header, text="Review (Excel Format)", font=("Segoe UI", 12, "bold"),
                     text_color=COLORS["text_primary"]).pack(side="left")
        self._jnl_manual_info_label = ctk.CTkLabel(right_header, text="Manual entries: 0",
                                                     font=("Segoe UI", 11), text_color=TEXT_MUTED)
        self._jnl_manual_info_label.pack(side="right")

        tree_frame = ctk.CTkFrame(right_panel, fg_color=COLORS["bg_dark"],
                                   corner_radius=8, border_width=1, border_color=COLORS["border"])
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))

        _sy = ttk.Scrollbar(tree_frame, orient="vertical")
        _sx = ttk.Scrollbar(tree_frame, orient="horizontal")
        self._jnl_manual_tree = ttk.Treeview(tree_frame, show="headings", selectmode="extended",
                                               yscrollcommand=_sy.set, xscrollcommand=_sx.set)
        _sy.config(command=self._jnl_manual_tree.yview)
        _sx.config(command=self._jnl_manual_tree.xview)
        _sy.pack(side="right", fill="y")
        _sx.pack(side="bottom", fill="x")
        self._jnl_manual_tree.pack(fill="both", expand=True)
        self._jnl_manual_tree.bind("<Double-1>", lambda e: self._jnl_edit_selected_manual())

        ctk.CTkLabel(right_panel, text="Double-click a row to edit it.",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).grid(
            row=2, column=0, sticky="w", padx=10, pady=(0, 8))

        self._jnl_populate_tree(self._jnl_manual_tree, _JNL_HDRS, [])

    # ─── JOURNAL MANUAL ENTRY METHODS ─────────────────────────────────────────

    def _jnl_populate_tree(self, tree, headers, rows, limit=500):
        tree.delete(*tree.get_children())
        tree["columns"] = headers
        for h in headers:
            tree.heading(h, text=h)
            tree.column(h, width=max(120, min(260, len(h) * 12)), minwidth=80)
        for idx, row in enumerate(rows[:limit]):
            values = [str(_row_get(row, h, "") or "") for h in headers]
            tree.insert("", "end", iid=str(idx), values=values)

    def _on_jnl_party_search_change(self, _event=None):
        combo = self._jnl_manual_party_ledger_combo
        if combo is None:
            return
        if not self.fetched_party_ledgers:
            combo.configure(values=[""])
            combo.set("")
            if self._jnl_manual_party_match_label is not None:
                self._jnl_manual_party_match_label.configure(text="No fetched party ledgers yet")
            return
        typed_text = (self._jnl_manual_party_search_var.get() or "").strip()
        typed = typed_text.casefold()
        current_value = (self._jnl_manual_form_vars.get("PartyLedger", ctk.StringVar()).get() or "").strip()
        if not typed:
            filtered = self.fetched_party_ledgers[:200]
        else:
            starts = [n for n in self.fetched_party_ledgers if n.casefold().startswith(typed)]
            contains = [n for n in self.fetched_party_ledgers if typed in n.casefold() and n not in starts]
            filtered = (starts + contains)[:200]
        if typed and not filtered:
            combo.configure(values=[""])
            combo.set("")
            if "PartyLedger" in self._jnl_manual_form_vars:
                self._jnl_manual_form_vars["PartyLedger"].set("")
            if self._jnl_manual_party_match_label is not None:
                self._jnl_manual_party_match_label.configure(text=f"Search '{typed_text}': no match")
            return
        display_values = filtered if filtered else self.fetched_party_ledgers[:200]
        combo.configure(values=display_values)
        if typed and display_values and current_value not in display_values:
            combo.set(display_values[0])
            if "PartyLedger" in self._jnl_manual_form_vars:
                self._jnl_manual_form_vars["PartyLedger"].set(display_values[0])
        elif current_value:
            combo.set(current_value)
        if self._jnl_manual_party_match_label is not None:
            shown = len(display_values)
            total = len(self.fetched_party_ledgers)
            if typed:
                self._jnl_manual_party_match_label.configure(
                    text=f"Search '{typed_text}': showing {shown} of {total}")
            else:
                self._jnl_manual_party_match_label.configure(text=f"Showing {shown} of {total} party ledgers")

    def _jnl_fetch_party_ledgers_thread(self, silent=False):
        if self._ledger_fetch_running:
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            if not silent:
                messagebox.showerror("Invalid Settings", str(exc))
            return
        selected_company = self._get_selected_company()
        self._ledger_fetch_running = True
        if self._jnl_manual_fetch_ledger_btn is not None:
            self._jnl_manual_fetch_ledger_btn.configure(state="disabled", text="...")

        def worker():
            result = _fetch_tally_ledgers(tally_url, timeout=15, company_name=selected_company)

            def done():
                self._ledger_fetch_running = False
                if self._jnl_manual_fetch_ledger_btn is not None:
                    self._jnl_manual_fetch_ledger_btn.configure(state="normal", text="Fetch")
                if result.get("success"):
                    party_ledgers = result.get("party_ledgers") or result.get("ledgers") or []
                    cleaned = []
                    seen = set()
                    for name in party_ledgers:
                        norm = _normalize_ledger_name(name)
                        if norm and _ledger_key(norm) not in seen:
                            seen.add(_ledger_key(norm))
                            cleaned.append(norm)
                    self.fetched_party_ledgers = sorted(cleaned, key=lambda x: _ledger_key(x))
                    if self._jnl_manual_party_ledger_combo is not None:
                        self._jnl_manual_party_ledger_combo.configure(
                            values=self.fetched_party_ledgers[:200] if self.fetched_party_ledgers else [""])
                    self._on_jnl_party_search_change()
                    self.status_var.set(f"Fetched {len(self.fetched_party_ledgers)} party ledger(s) from Tally")
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.status_var.set("Party ledger fetch failed")
                    if not silent:
                        messagebox.showwarning("Party Ledger Fetch Failed",
                                               f"Could not fetch ledgers from Tally.\n\n{err}")
            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _jnl_manual_row_from_form(self):
        row = {}
        for header in self.JNL_TEMPLATE_HEADERS:
            row[header] = self._jnl_manual_form_vars.get(header, ctk.StringVar()).get().strip()
        return row

    def _jnl_validate_manual_row(self, row):
        party = str(row.get("PartyLedger", "")).strip()
        particular = str(row.get("Particular", "")).strip()
        if not party:
            messagebox.showwarning("Missing Field", "Party Ledger is required.")
            return None
        if not particular:
            messagebox.showwarning("Missing Field", "Particular is required.")
            return None
        try:
            taxable = float(str(row.get("TaxableValue") or "0").strip())
        except ValueError:
            messagebox.showwarning("Invalid Value", "Taxable Value must be numeric.")
            return None
        if taxable <= 0:
            messagebox.showwarning("Invalid Value", "Taxable Value must be greater than zero.")
            return None
        if not str(row.get("Date") or "").strip():
            row["Date"] = datetime.today().strftime("%d-%m-%y")
        for key in ["CGSTRate", "SGSTRate", "IGSTRate"]:
            if not str(row.get(key) or "").strip():
                row[key] = "0"
        return row

    def _jnl_set_manual_edit_mode(self, index=None):
        self._jnl_manual_editing_index = index
        if self._jnl_manual_update_btn is not None:
            self._jnl_manual_update_btn.configure(state="normal" if index is not None else "disabled")

    def _jnl_selected_manual_index(self):
        if self._jnl_manual_tree is None:
            return None
        selected = list(self._jnl_manual_tree.selection())
        if not selected:
            return None
        try:
            return int(selected[0])
        except ValueError:
            return None

    def _jnl_refresh_manual_tree(self, focus_index=None):
        if self._jnl_manual_tree is None:
            return
        self._jnl_populate_tree(self._jnl_manual_tree, self.JNL_TEMPLATE_HEADERS,
                                  self._jnl_manual_rows, limit=500)
        if focus_index is not None:
            iid = str(focus_index)
            if iid in self._jnl_manual_tree.get_children():
                self._jnl_manual_tree.selection_set(iid)
                self._jnl_manual_tree.focus(iid)
                self._jnl_manual_tree.see(iid)
        if self._jnl_manual_info_label is not None:
            self._jnl_manual_info_label.configure(text=f"Manual entries: {len(self._jnl_manual_rows)}")

    def _jnl_add_manual_entry(self):
        row = self._jnl_manual_row_from_form()
        validated = self._jnl_validate_manual_row(row)
        if validated is None:
            return
        if not str(validated.get("VoucherNo") or "").strip():
            validated["VoucherNo"] = str(len(self._jnl_manual_rows) + 1)
        self._jnl_manual_rows.append(validated)
        self._jnl_refresh_manual_tree(focus_index=len(self._jnl_manual_rows) - 1)
        self._jnl_set_manual_edit_mode(None)
        self.status_var.set(f"Journal manual entry added. Total: {len(self._jnl_manual_rows)}")

    def _jnl_edit_selected_manual(self):
        idx = self._jnl_selected_manual_index()
        if idx is None:
            messagebox.showinfo("Edit Entry", "Select one row in the table to edit.")
            return
        if idx < 0 or idx >= len(self._jnl_manual_rows):
            return
        row = self._jnl_manual_rows[idx]
        for header in self.JNL_TEMPLATE_HEADERS:
            value = _row_get(row, header, "")
            if header in self._jnl_manual_form_vars:
                self._jnl_manual_form_vars[header].set("" if value is None else str(value))
        self._jnl_set_manual_edit_mode(idx)
        self.status_var.set(f"Editing journal entry #{idx + 1}. Update Entry to save changes.")

    def _jnl_update_manual_entry(self):
        idx = self._jnl_manual_editing_index
        if idx is None:
            messagebox.showinfo("Update Entry", "Select and edit a row first.")
            return
        if idx < 0 or idx >= len(self._jnl_manual_rows):
            self._jnl_set_manual_edit_mode(None)
            messagebox.showwarning("Update Entry", "Selected row is no longer available.")
            return
        row = self._jnl_manual_row_from_form()
        validated = self._jnl_validate_manual_row(row)
        if validated is None:
            return
        if not str(validated.get("VoucherNo") or "").strip():
            validated["VoucherNo"] = str(idx + 1)
        self._jnl_manual_rows[idx] = validated
        self._jnl_refresh_manual_tree(focus_index=idx)
        self._jnl_set_manual_edit_mode(None)
        self.status_var.set(f"Journal manual entry #{idx + 1} updated.")

    def _jnl_clear_manual_form(self):
        keep_date = datetime.today().strftime("%d-%m-%y")
        for key, var in self._jnl_manual_form_vars.items():
            if key == "Date":
                var.set(keep_date)
            elif key in {"CGSTRate", "SGSTRate", "IGSTRate"}:
                var.set("0")
            else:
                var.set("")
        self._jnl_set_manual_edit_mode(None)
        self._jnl_manual_party_search_var.set("")
        self._on_jnl_party_search_change()

    def _jnl_remove_selected_manual(self):
        if self._jnl_manual_tree is None:
            return
        selected = list(self._jnl_manual_tree.selection())
        if not selected:
            messagebox.showinfo("Remove Entry", "Select at least one row to remove.")
            return
        indexes = []
        for iid in selected:
            try:
                indexes.append(int(iid))
            except ValueError:
                continue
        for idx in sorted(indexes, reverse=True):
            if 0 <= idx < len(self._jnl_manual_rows):
                self._jnl_manual_rows.pop(idx)
        self._jnl_refresh_manual_tree()
        self._jnl_set_manual_edit_mode(None)
        self.status_var.set(f"Removed. Remaining: {len(self._jnl_manual_rows)}")

    def _jnl_clear_all_manual(self):
        if not self._jnl_manual_rows:
            return
        if not messagebox.askyesno("Clear All", "Remove all journal manual entries?"):
            return
        self._jnl_manual_rows = []
        self._jnl_refresh_manual_tree()
        self._jnl_set_manual_edit_mode(None)
        self.status_var.set("All journal manual entries cleared.")

    def _download_journal_template(self):
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Template_Journal Voucher.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not out:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(self.JNL_TEMPLATE_HEADERS)
            ws.append(["22-04-2026", "18AABCI8307G1ZM", "1", "ABC Corp Pvt Ltd", "Office Expense 18%", 10000, "", 0, "", 0, "IGST", 18, "Test Entry for ABC Corp"])
            wb.save(out)
            messagebox.showinfo("Template Saved", f"Journal template saved to:\n{out}")
        except Exception as exc:
            messagebox.showerror("Template Error", str(exc))

    # ─── JOURNAL GENERATION ───────────────────────────────────────────────────

    def _generate_journal(self, action: str):
        if self._push_running:
            self.status_var.set("Push already in progress...")
            return

        if hasattr(self, '_jnl_source_tabs') and self._jnl_source_tabs is not None:
            active_tab = self._jnl_source_tabs.get()
            if active_tab == "Manual Entry":
                rows = getattr(self, '_jnl_manual_rows', [])
                source_label = "Manual Entry"
            else:
                rows = getattr(self, '_jnl_loaded_rows', [])
                source_label = "Excel Upload"
        else:
            rows = getattr(self, '_jnl_loaded_rows', [])
            source_label = "Excel Upload"

        if not rows:
            messagebox.showwarning("No Data", f"No rows available in {source_label}.")
            return

        company = self._get_selected_company()
        if action == "push" and not company and len(getattr(self, "fetched_companies", [])) > 1:
            messagebox.showwarning("Select Company", "Please select a target company before pushing.")
            return

        try:
            date_mode, custom_tally_date = self._get_voucher_date_selection()
            journal_type = self._journal_type_var.get() if self._journal_type_var else "purchase"
            _cmp_regs = list(self.company_gst_registrations or [])
            if not _cmp_regs:
                try:
                    _r = _fetch_cmp_gst_regs_for_journal(self._get_tally_url(), company=company, timeout=10)
                    if _r.get("success"):
                        _cmp_regs = _r["registrations"]
                        self.company_gst_registrations = _cmp_regs
                except Exception:
                    pass

            _jnl_ro_ledger = (
                self._jnl_roundoff_ledger_var.get().strip()
                if self._jnl_roundoff_enabled_var.get()
                else ""
            )
            xml_payload, voucher_count = generate_journal_xml(
                rows,
                company=company,
                date_mode=date_mode,
                custom_tally_date=custom_tally_date,
                journal_type=journal_type,
                company_gst_registrations=_cmp_regs,
                round_off_ledger=_jnl_ro_ledger,
            )
            if voucher_count <= 0:
                messagebox.showwarning("No Vouchers", "No valid rows found (TaxableValue must be greater than zero).")
                return

            if action == "save":
                out = filedialog.asksaveasfilename(
                    defaultextension=".xml",
                    initialfile="Journal.xml",
                    filetypes=[("XML", "*.xml")],
                )
                if not out:
                    return
                with open(out, "w", encoding="utf-8") as f:
                    f.write(xml_payload)
                self.status_var.set(f"Journal XML saved: {os.path.basename(out)} ({voucher_count} voucher(s))")
                messagebox.showinfo("Saved", f"Journal XML saved successfully.\n{out}")
                return

            host = (self.tally_host_var.get() or "localhost").strip()
            port_text = (self.tally_port_var.get() or "9000").strip()
            if not port_text.isdigit():
                raise ValueError("Port must be numeric.")
            port = int(port_text)
            tally_url_jnl = self._get_tally_url()
            jnl_is_purchase = (journal_type == "purchase")
            jnl_rows_snap = list(rows)

            self._set_push_loading_state(True, f"Pushing {voucher_count} Journal voucher(s) from {source_label}...")
            self.status_var.set("Pushing to Tally...")

            def worker():
                try:
                    # Pre-creation: check & create missing party ledgers
                    self.after(0, lambda: self._push_message_var.set("Checking existing ledgers in Tally..."))
                    _jel_result = _fetch_existing_ledger_names(tally_url_jnl, company_name=company, timeout=15)
                    existing_jnl_ledgers = set(_jel_result.get("ledgers") or set()) if _jel_result.get("success") else set()

                    auto_jnl_defs = _collect_auto_note_ledgers(jnl_rows_snap, is_debit_note=jnl_is_purchase)
                    auto_jnl_to_create = _filter_out_existing_ledgers(auto_jnl_defs, existing_jnl_ledgers)
                    if auto_jnl_to_create:
                        party_defs_j = [d for d in auto_jnl_to_create if _is_party_parent(str(d.get("Parent", "")))]
                        non_party_j  = [d for d in auto_jnl_to_create if not _is_party_parent(str(d.get("Parent", "")))]
                        asked_j = []
                        for _ld in party_defs_j:
                            self.after(0, lambda: self._push_message_var.set("Waiting for user input on missing ledger..."))
                            asked_j.append(_ask_create_party_ledger(self, _ld))
                        auto_jnl_to_create = non_party_j + asked_j
                        self.after(0, lambda: self._push_message_var.set("Creating required ledgers in Tally..."))
                        push_to_tally(generate_ledger_xml(auto_jnl_to_create, company), host=host, port=port)
                        existing_jnl_ledgers.update(
                            str(e.get("Name", "")).strip() for e in auto_jnl_to_create
                            if str(e.get("Name", "")).strip()
                        )

                    # Push journal XML
                    self.after(0, lambda: self._push_message_var.set("Pushing journal vouchers..."))
                    resp = push_to_tally(xml_payload, host=host, port=port)
                    parsed = _parse_tally_response_details(resp)

                    # On failure: auto-create missing ledgers from line errors and retry
                    if not parsed.get("success"):
                        line_errs = parsed.get("line_errors", [])
                        if line_errs:
                            miss_mode = "purchase_accounting" if jnl_is_purchase else "accounting"
                            miss_defs = _build_missing_ledger_defs(line_errs, jnl_rows_snap, miss_mode)
                            miss_defs = _filter_out_existing_ledgers(miss_defs, existing_jnl_ledgers)
                            if miss_defs:
                                final_miss = []
                                for _mld in miss_defs:
                                    if _is_party_parent(str(_mld.get("Parent", ""))):
                                        self.after(0, lambda: self._push_message_var.set("Waiting for user input on missing ledger..."))
                                        _mld = _ask_create_party_ledger(self, _mld)
                                    final_miss.append(_mld)
                                self.after(0, lambda: self._push_message_var.set("Creating missing ledgers and retrying..."))
                                push_to_tally(generate_ledger_xml(final_miss, company), host=host, port=port)
                                existing_jnl_ledgers.update(
                                    str(e.get("Name", "")).strip() for e in final_miss
                                    if str(e.get("Name", "")).strip()
                                )
                                retry_resp = push_to_tally(xml_payload, host=host, port=port)
                                retry_parsed = _parse_tally_response_details(retry_resp)
                                if retry_parsed.get("success"):
                                    parsed = retry_parsed

                    result = {"ok": True, "parsed": parsed}
                except Exception as exc:
                    result = {"ok": False, "error": str(exc)}

                def done():
                    self._set_push_loading_state(False)
                    if not result.get("ok"):
                        err = str(result.get("error", "Unknown error"))
                        self.status_var.set(f"Push failed: {err}")
                        messagebox.showerror("Push Failed", err)
                        return
                    parsed = result["parsed"]
                    created = parsed.get("created", 0)
                    altered = parsed.get("altered", 0)
                    errors = parsed.get("errors", 0)
                    line_errors = parsed.get("line_errors", [])
                    summary = (f"Created: {created}\nAltered: {altered}\nErrors: {errors}")
                    if parsed.get("success"):
                        self.status_var.set(f"Journal push successful: Created {created}, Altered {altered}")
                        messagebox.showinfo("Push Successful", summary)
                    else:
                        if line_errors:
                            summary += "\n\nLine Errors:\n- " + "\n- ".join(line_errors[:8])
                        self.status_var.set("Journal push completed with errors.")
                        messagebox.showwarning("Push Completed With Errors", summary)
                self.after(0, done)

            threading.Thread(target=worker, daemon=True).start()

        except ValueError as exc:
            messagebox.showerror("Validation Error", str(exc))
        except Exception as exc:
            messagebox.showerror("Error", str(exc))


# ═══════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = TallySalesApp()
    app.mainloop()
