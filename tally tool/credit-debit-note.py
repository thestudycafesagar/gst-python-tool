"""
TallyNoteEntryPro - Credit and Debit Note Voucher Creator for TallyPrime
Supports Excel upload and manual entry, with XML export and direct push to Tally.
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
import re
import html
import threading
import webbrowser
import urllib.request
import xml.etree.ElementTree as ET
from urllib.error import HTTPError, URLError
from datetime import datetime, date, timedelta


# Theme colors are host-driven (GST Suite). Keep all colors dual-mode for auto sync.
COLORS = {
    "bg_dark": ("#F0F4F8", "#0F172A"),
    "bg_card": ("#FFFFFF", "#1E293B"),
    "bg_card_hover": ("#E2E8F0", "#334155"),
    "bg_input": ("#F1F5F9", "#1E293B"),
    "accent": ("#2563EB", "#3B82F6"),
    "accent_hover": ("#1D4ED8", "#2563EB"),
    "success": ("#059669", "#10B981"),
    "warning": ("#D97706", "#F59E0B"),
    "error": ("#DC2626", "#EF4444"),
    "text_primary": ("#0F172A", "#F1F5F9"),
    "text_secondary": ("#475569", "#CBD5E1"),
    "text_muted": ("#64748B", "#94A3B8"),
    "border": ("#E2E8F0", "#334155"),
    "table_header": ("#1E293B", "#0B1220"),
    "table_bg": ("#FFFFFF", "#111827"),
    "table_selected": ("#2563EB", "#3B82F6"),
    "table_selected_fg": ("#FFFFFF", "#F8FAFC"),
}


def _theme_color(name_or_value):
    value = COLORS.get(name_or_value, name_or_value)
    if isinstance(value, tuple):
        mode = ctk.get_appearance_mode().lower()
        return value[1] if mode == "dark" else value[0]
    return value

ACCENT = COLORS["accent"]
ACCENT_HOVER = COLORS["accent_hover"]
SUCCESS = COLORS["success"]
TEXT_MUTED = COLORS["text_muted"]
SUSPENSE_LEDGER = "Suspense A/c"
PUSH_REQUEST_TIMEOUT_SEC = 300

TEMPLATE_HEADERS = [
    "Date",
    "VoucherNo",
    "GSTIN",
    "PartyLedger",
    "Particular",
    "TaxableValue",
    "CGSTLedger",
    "CGSTRate",
    "SGSTLedger",
    "SGSTRate",
    "IGSTLedger",
    "IGSTRate",
    "Narration",
]

TEMPLATE_SAMPLE_ROW = [
    "16-12-25",
    "1",
    "",
    "Interactive Media Pvt Ltd",
    "Lecture Income",
    100000,
    0,
    0,
    0,
    0,
    "IGST Outward",
    18,
    "This is Testing Voucher56",
]


# Helpers
def xml_escape(s: str) -> str:
    if not s:
        return ""
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def fmt_amt(num: float) -> str:
    return f"{num:.2f}"


def tally_date(dt) -> str:
    today = datetime.today().strftime("%Y%m%d")

    if dt in (None, ""):
        return today
    if isinstance(dt, datetime):
        return dt.strftime("%Y%m%d")
    if isinstance(dt, date):
        return dt.strftime("%Y%m%d")

    if isinstance(dt, (int, float)) and not isinstance(dt, bool):
        try:
            if float(dt) > 1000:
                return (datetime(1899, 12, 30) + timedelta(days=float(dt))).strftime("%Y%m%d")
        except Exception:
            pass

    text = str(dt).strip()
    if not text:
        return today

    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]

    if re.fullmatch(r"\d{8}", text):
        if 1900 <= int(text[:4]) <= 2100:
            return text
        return f"{text[4:8]}{text[2:4]}{text[:2]}"

    candidates = [text]
    if " " in text:
        candidates.append(text.split(" ", 1)[0])

    formats = (
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d-%m-%y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d-%B-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    )
    for candidate in candidates:
        for fmt in formats:
            try:
                return datetime.strptime(candidate, fmt).strftime("%Y%m%d")
            except ValueError:
                continue

    return today


def _normalize_manual_date_to_tally(date_text: str) -> str:
    text = str(date_text or "").strip()
    if not text:
        raise ValueError("Custom date is empty.")

    compact = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\d{8}", compact):
        for fmt in ("%Y%m%d", "%d%m%Y"):
            try:
                return datetime.strptime(compact, fmt).strftime("%Y%m%d")
            except ValueError:
                continue

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            continue

    raise ValueError("Invalid custom date format. Use DD/MM/YYYY, DD-MM-YYYY, or YYYY-MM-DD.")


def _row_get(row: dict, key: str, default=None):
    if key in row:
        value = row.get(key)
        return default if value is None else value

    target = re.sub(r"\s+", "", str(key or "")).lower()
    for raw_key, raw_value in row.items():
        normalized = re.sub(r"\s+", "", str(raw_key or "")).lower()
        if normalized == target:
            return default if raw_value is None else raw_value
    return default


def _row_text(row: dict, key: str, default: str = "") -> str:
    value = _row_get(row, key, default)
    if value is None:
        return default
    return str(value).strip()


def _row_float(row: dict, key: str, default: float = 0.0) -> float:
    value = _row_get(row, key, default)
    if value in (None, ""):
        return float(default)
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _row_voucher_number(row: dict, default: str = "") -> str:
    return (
        _row_text(row, "VoucherNo")
        or _row_text(row, "InvoiceNo")
        or _row_text(row, "Invoice No")
        or _row_text(row, "BillNo")
        or default
    )


def _ledger_or_default(value: str, fallback: str = SUSPENSE_LEDGER) -> str:
    text = str(value or "").strip()
    return text or fallback


def _normalize_company_name(value) -> str:
    text = html.unescape(str(value or ""))
    text = text.replace("\x00", "")
    text = re.sub(r"[\x01-\x1F\x7F]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _company_key(value) -> str:
    return _normalize_company_name(value).upper()


def _is_valid_company_name(value) -> bool:
    name = _normalize_company_name(value)
    if not name:
        return False
    return not name.isdigit()


def _build_tally_url(host: str, port: str) -> str:
    host_text = (host or "localhost").strip()
    port_text = (port or "9000").strip()
    if host_text.startswith("http://"):
        host_text = host_text[7:]
    elif host_text.startswith("https://"):
        host_text = host_text[8:]
    host_text = host_text.strip("/") or "localhost"
    if "/" in host_text:
        host_text = host_text.split("/", 1)[0]
    if not port_text.isdigit():
        raise ValueError("Port must be numeric.")
    return f"http://{host_text}:{port_text}"


def _post_tally_xml(tally_url: str, xml_payload: str, timeout: float = 15.0) -> str:
    request = urllib.request.Request(
        tally_url,
        data=xml_payload.encode("utf-8"),
        headers={"Content-Type": "application/xml"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="replace")


def push_to_tally(xml_str: str, host: str = "localhost", port: int = 9000, timeout: float = PUSH_REQUEST_TIMEOUT_SEC) -> str:
    url = f"http://{host}:{port}"
    req = urllib.request.Request(
        url,
        data=xml_str.encode("utf-8"),
        headers={"Content-Type": "application/xml"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def _check_tally_connection(tally_url: str, timeout: float = 5.0) -> dict:
    probe_xml = (
        "<ENVELOPE><HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
        "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
        "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>"
    )
    try:
        _post_tally_xml(tally_url, probe_xml, timeout=timeout)
        return {"connected": True}
    except HTTPError as exc:
        return {"connected": False, "error": f"HTTP {exc.code}"}
    except URLError:
        return {"connected": False, "error": "ConnectionError"}
    except Exception as exc:
        return {"connected": False, "error": str(exc)}


def _extract_company_names(response_text: str) -> set:
    names = set()
    try:
        root = ET.fromstring(response_text)
        for node in root.iter():
            tag = str(node.tag or "").upper()
            txt = _normalize_company_name(node.text)
            attr_name = _normalize_company_name(node.attrib.get("NAME") or "")
            if tag in {"COMPANYNAME", "SVCURRENTCOMPANY", "CURRENTCOMPANY"} and _is_valid_company_name(txt):
                names.add(txt)
            if "COMPANY" in tag and _is_valid_company_name(attr_name):
                names.add(attr_name)
            if tag == "COMPANY" and _is_valid_company_name(txt):
                names.add(txt)
    except ET.ParseError:
        pass

    patterns = [
        r'COMPANY[^>]*NAME="([^"]+)"',
        r"<COMPANYNAME>(.*?)</COMPANYNAME>",
        r"<SVCURRENTCOMPANY>(.*?)</SVCURRENTCOMPANY>",
        r"<COMPANY[^>]*>.*?<NAME>(.*?)</NAME>",
    ]
    for pattern in patterns:
        for match in re.findall(pattern, response_text, flags=re.IGNORECASE | re.DOTALL):
            value = _normalize_company_name(match)
            if _is_valid_company_name(value):
                names.add(value)

    return names


def _fetch_tally_companies(tally_url: str, timeout: float = 15.0) -> dict:
    requests_xml = [
        (
            "report-list-companies",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>",
        ),
        (
            "collection-company",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
            "<TYPE>Collection</TYPE><ID>Company Collection</ID></HEADER><BODY><DESC>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "<TDL><TDLMESSAGE><COLLECTION NAME='Company Collection'>"
            "<TYPE>Company</TYPE><FETCH>Name</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
            "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>",
        ),
    ]

    companies = set()
    errors = []
    for label, xml_payload in requests_xml:
        try:
            response_text = _post_tally_xml(tally_url, xml_payload, timeout=timeout)
            companies.update(_extract_company_names(response_text))
        except HTTPError as exc:
            errors.append(f"{label}: HTTP {exc.code}")
        except URLError:
            errors.append(f"{label}: ConnectionError")
        except Exception as exc:
            errors.append(f"{label}: {exc}")

    sorted_companies = sorted(companies, key=lambda x: _company_key(x))
    if sorted_companies:
        return {"success": True, "companies": sorted_companies}

    err = "; ".join(errors) if errors else "No companies returned by Tally."
    return {"success": False, "error": err, "companies": []}


def _normalize_ledger_name(value) -> str:
    text = html.unescape(str(value or ""))
    text = text.replace("\x00", "")
    text = re.sub(r"[\x01-\x1F\x7F]", " ", text)
    text = re.sub(r"[\s]+", " ", text).strip()
    while len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"', "`"}:
        text = text[1:-1].strip()
    return text


def _ledger_key(value) -> str:
    return _normalize_ledger_name(value).upper()


def _parent_group_key(value) -> str:
    return _normalize_ledger_name(value).replace("&", "and").upper()


def _state_name_from_gstin(gstin: str) -> str:
    gstin_text = _normalize_ledger_name(gstin).upper()
    state_map = {
        "01": "Jammu And Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
        "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
        "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
        "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
        "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
        "24": "Gujarat", "25": "Daman And Diu", "26": "Dadra And Nagar Haveli And Daman And Diu",
        "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
        "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman And Nicobar Islands", "36": "Telangana",
        "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory", "99": "Centre Jurisdiction",
    }
    return state_map.get(gstin_text[:2], "")


def _is_gstin_like(value: str) -> bool:
    text = str(value or "").strip().upper()
    return bool(re.fullmatch(r"\d{2}[A-Z0-9]{13}", text))


def _extract_company_gst_registrations(response_text: str) -> list:
    """Parse TaxUnit collection response and return list of company GST registrations."""
    registrations = []
    seen = set()
    try:
        root = ET.fromstring(response_text)
        for tax_unit in root.findall(".//TAXUNIT"):
            tax_type = str(tax_unit.attrib.get("TAXTYPE") or tax_unit.findtext("TAXTYPE") or "").strip().upper()
            if tax_type and tax_type != "GST":
                continue
            name_raw = str(tax_unit.attrib.get("NAME") or tax_unit.findtext("NAME") or "").strip()
            gstin_raw = str(
                tax_unit.attrib.get("TAXREGISTRATION")
                or tax_unit.findtext("GSTREGNUMBER")
                or tax_unit.findtext("GSTIN")
                or ""
            ).strip().upper()
            state_raw = str(tax_unit.findtext("STATENAME") or "").strip()
            if not gstin_raw and _is_gstin_like(name_raw):
                gstin_raw = name_raw.upper()
            if not state_raw and gstin_raw:
                state_raw = _state_name_from_gstin(gstin_raw)
            if not gstin_raw:
                continue
            name = _normalize_company_name(name_raw or gstin_raw)
            if not name:
                continue
            key = (name.casefold(), gstin_raw)
            if key in seen:
                continue
            seen.add(key)
            registrations.append({"name": name, "gstin": gstin_raw, "state": state_raw})
    except ET.ParseError:
        pass
    return registrations


def _fetch_company_gst_registrations(tally_url: str, company_name: str = "", timeout: float = 15.0) -> dict:
    """Fetch the company's GST registration details from Tally (TaxUnit collection)."""
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"
    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>Tax Unit Lookup</ID></HEADER>"
        f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='Tax Unit Lookup'><TYPE>TaxUnit</TYPE>"
        "<FETCH>Name,TaxType,TaxRegistration,GSTRegNumber,StateName,UseFor</FETCH>"
        "<NATIVEMETHOD>Name</NATIVEMETHOD></COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )
    try:
        response_text = _post_tally_xml(tally_url, request_xml, timeout=timeout)
    except HTTPError as exc:
        return {"success": False, "error": f"HTTP {exc.code}", "registrations": []}
    except URLError:
        return {"success": False, "error": "ConnectionError", "registrations": []}
    except Exception as exc:
        return {"success": False, "error": str(exc), "registrations": []}
    registrations = _extract_company_gst_registrations(response_text)
    if registrations:
        return {"success": True, "registrations": registrations}
    return {"success": False, "error": "No GST registrations returned.", "registrations": []}


def _fetch_tally_ledgers(tally_url: str, timeout: float = 15.0, company_name: str = "") -> dict:
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{xml_escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml_variants = [
        (
            "collection-ledger",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
            "<TYPE>Collection</TYPE><ID>Ledger Collection</ID></HEADER>"
            f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE><COLLECTION NAME='Ledger Collection'>"
            "<TYPE>Ledger</TYPE><FETCH>Name,Parent</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
            "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>",
        ),
        (
            "report-list-ledgers",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Ledgers</REPORTNAME>"
            f"{static_vars}</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>",
        ),
        (
            "report-list-accounts",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Accounts</REPORTNAME>"
            f"{static_vars}</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>",
        ),
    ]

    def _extract_from_response(response_text: str):
        all_entries = []
        try:
            root = ET.fromstring(response_text)
            for node in root.iter():
                tag = str(node.tag or "")
                tag_upper = tag.upper().split("}")[-1]
                if tag_upper != "LEDGER":
                    continue

                name = _normalize_ledger_name(
                    node.attrib.get("NAME")
                    or (node.findtext("NAME") if hasattr(node, "findtext") else "")
                    or node.text
                    or ""
                )
                parent = _normalize_ledger_name(
                    (node.findtext("PARENT") if hasattr(node, "findtext") else "")
                    or node.attrib.get("PARENT")
                    or ""
                )
                if name:
                    all_entries.append((name, parent))
        except ET.ParseError:
            pass

        for match in re.findall(r'LEDGER[^>]*NAME="([^"]+)"', response_text, flags=re.IGNORECASE):
            name = _normalize_ledger_name(match)
            if name:
                all_entries.append((name, ""))
        for match in re.findall(r"<LEDGER[^>]*>.*?<NAME>(.*?)</NAME>", response_text, flags=re.IGNORECASE | re.DOTALL):
            name = _normalize_ledger_name(match)
            if name:
                all_entries.append((name, ""))

        return all_entries

    all_ledgers_map = {}
    errors = []

    for label, payload in request_xml_variants:
        try:
            response_text = _post_tally_xml(tally_url, payload, timeout=timeout)
            entries = _extract_from_response(response_text)
            for name, parent in entries:
                key = _ledger_key(name)
                if not key:
                    continue
                existing = all_ledgers_map.get(key)
                if existing is None:
                    all_ledgers_map[key] = {"name": name, "parent": parent}
                elif not existing.get("parent") and parent:
                    existing["parent"] = parent
        except HTTPError as exc:
            errors.append(f"{label}: HTTP {exc.code}")
        except URLError:
            errors.append(f"{label}: ConnectionError")
        except Exception as exc:
            errors.append(f"{label}: {exc}")

    ledgers = sorted((v["name"] for v in all_ledgers_map.values() if v.get("name")), key=lambda x: _ledger_key(x))

    party_group_keys = {
        "SUNDRY CREDITORS",
        "SUNDRY DEBTORS",
    }
    party_ledgers = sorted(
        (
            v["name"]
            for v in all_ledgers_map.values()
            if v.get("name") and _parent_group_key(v.get("parent")) in party_group_keys
        ),
        key=lambda x: _ledger_key(x),
    )

    if party_ledgers or ledgers:
        return {
            "success": True,
            "ledgers": ledgers,
            "party_ledgers": party_ledgers,
        }

    err = "; ".join(errors) if errors else "No ledgers returned by Tally."
    return {
        "success": False,
        "error": err,
        "ledgers": [],
        "party_ledgers": [],
    }


def _create_tally_ledger(
    tally_url: str,
    ledger_name: str,
    parent_name: str,
    company_name: str = "",
    gstin: str = "",
    state: str = "",
    country: str = "India",
    pincode: str = "",
    mailing_name: str = "",
    address1: str = "",
    address2: str = "",
    billwise_on: bool = True,
    timeout: float = 30.0,
    gst_applicable: str = "",
) -> dict:
    name = _normalize_ledger_name(ledger_name)
    parent = _normalize_ledger_name(parent_name) or "Sundry Debtors"
    if not name:
        return {"success": False, "error": "Ledger name cannot be empty."}

    gstin_clean = _normalize_ledger_name(gstin).upper()
    state_clean = _normalize_ledger_name(state)
    if not state_clean and gstin_clean:
        state_clean = _state_name_from_gstin(gstin_clean)
    country_clean = _normalize_ledger_name(country) or "India"
    pincode_clean = _normalize_ledger_name(pincode)
    mailing_clean = _normalize_ledger_name(mailing_name) or name
    addr1_clean = _normalize_ledger_name(address1)
    addr2_clean = _normalize_ledger_name(address2)

    parent_key = _parent_group_key(parent)
    is_party_ledger = parent_key in {"SUNDRY DEBTORS", "SUNDRY CREDITORS"}

    gst_app_raw = _normalize_ledger_name(gst_applicable)
    gst_app_key = gst_app_raw.casefold()
    if gst_app_key in {"applicable", "yes", "y", "true", "1", "registered", "regular", "gst applicable"}:
        gst_app_text = "Applicable"
    elif gst_app_key in {"not applicable", "no", "n", "false", "0", "na", "n/a", "notapplicable"}:
        gst_app_text = "Not Applicable"
    else:
        gst_app_text = "Applicable" if gstin_clean else "Not Applicable"

    reg_type = "Regular" if (gstin_clean or gst_app_text == "Applicable") else ""

    today = datetime.today().date()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    applicable_from = f"{fy_start_year}0401"

    envelope = ET.Element("ENVELOPE")
    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
    body = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")
    req_desc = ET.SubElement(import_data, "REQUESTDESC")
    ET.SubElement(req_desc, "REPORTNAME").text = "All Masters"

    selected_company = _normalize_company_name(company_name)
    if selected_company:
        static_vars = ET.SubElement(req_desc, "STATICVARIABLES")
        ET.SubElement(static_vars, "SVCURRENTCOMPANY").text = selected_company

    req_data = ET.SubElement(import_data, "REQUESTDATA")
    tally_msg = ET.SubElement(req_data, "TALLYMESSAGE")
    tally_msg.set("xmlns:UDF", "TallyUDF")

    ledger = ET.SubElement(tally_msg, "LEDGER")
    ledger.set("NAME", name)
    ledger.set("RESERVEDNAME", "")
    ledger.set("ACTION", "Create")
    ET.SubElement(ledger, "NAME").text = name
    ET.SubElement(ledger, "PARENT").text = parent
    ET.SubElement(ledger, "ISBILLWISEON").text = "Yes" if (billwise_on and is_party_ledger) else "No"
    ET.SubElement(ledger, "ISCOSTCENTRESON").text = "No"
    ET.SubElement(ledger, "ISINTERESTON").text = "No"
    ET.SubElement(ledger, "ALLOWINMOBILE").text = "No"
    ET.SubElement(ledger, "ISUPDATINGTARGETID").text = "No"
    ET.SubElement(ledger, "ASORIGINAL").text = "Yes"
    ET.SubElement(ledger, "AFFECTSSTOCK").text = "No"
    ET.SubElement(ledger, "CURRENCYNAME").text = "INR"
    ET.SubElement(ledger, "COUNTRYOFRESIDENCE").text = country_clean

    if is_party_ledger:
        ET.SubElement(ledger, "GSTAPPLICABLE").text = gst_app_text
        if reg_type:
            ET.SubElement(ledger, "GSTREGISTRATIONTYPE").text = reg_type
        if gstin_clean:
            ET.SubElement(ledger, "PARTYGSTIN").text = gstin_clean

    if state_clean:
        ET.SubElement(ledger, "PRIORSTATENAME").text = state_clean
        if is_party_ledger:
            ET.SubElement(ledger, "LEDSTATENAME").text = state_clean

    language_list = ET.SubElement(ledger, "LANGUAGENAME.LIST")
    name_list = ET.SubElement(language_list, "NAME.LIST")
    name_list.set("TYPE", "String")
    ET.SubElement(name_list, "NAME").text = name
    ET.SubElement(language_list, "LANGUAGEID").text = "1033"

    if is_party_ledger and (gstin_clean or reg_type):
        gst_list = ET.SubElement(ledger, "LEDGSTREGDETAILS.LIST")
        ET.SubElement(gst_list, "APPLICABLEFROM").text = applicable_from
        if reg_type:
            ET.SubElement(gst_list, "GSTREGISTRATIONTYPE").text = reg_type
        if state_clean:
            ET.SubElement(gst_list, "PLACEOFSUPPLY").text = state_clean
        if gstin_clean:
            ET.SubElement(gst_list, "GSTIN").text = gstin_clean
        ET.SubElement(gst_list, "ISOTHTERRITORYASSESSEE").text = "No"
        ET.SubElement(gst_list, "CONSIDERPURCHASEFOREXPORT").text = "No"
        ET.SubElement(gst_list, "ISTRANSPORTER").text = "No"
        ET.SubElement(gst_list, "ISCOMMONPARTY").text = "No"

    if is_party_ledger and (addr1_clean or addr2_clean or state_clean or country_clean or pincode_clean):
        mailing_list = ET.SubElement(ledger, "LEDMAILINGDETAILS.LIST")
        if addr1_clean or addr2_clean:
            addr_list = ET.SubElement(mailing_list, "ADDRESS.LIST")
            addr_list.set("TYPE", "String")
            if addr1_clean:
                ET.SubElement(addr_list, "ADDRESS").text = addr1_clean
            if addr2_clean:
                ET.SubElement(addr_list, "ADDRESS").text = addr2_clean
        ET.SubElement(mailing_list, "APPLICABLEFROM").text = applicable_from
        ET.SubElement(mailing_list, "MAILINGNAME").text = mailing_clean
        if pincode_clean:
            ET.SubElement(mailing_list, "PINCODE").text = pincode_clean
        if state_clean:
            ET.SubElement(mailing_list, "STATE").text = state_clean
        ET.SubElement(mailing_list, "COUNTRY").text = country_clean

    xml_payload = ET.tostring(envelope, encoding="unicode")
    response_text = _post_tally_xml(tally_url, xml_payload, timeout=timeout)
    parsed = _parse_tally_response_details(response_text)

    created = int(parsed.get("created", 0) or 0)
    altered = int(parsed.get("altered", 0) or 0)
    if parsed.get("success") and (created > 0 or altered > 0):
        return {
            "success": True,
            "created": created,
            "altered": altered,
            "response": parsed,
        }

    return {
        "success": False,
        "error": parsed.get("error") or "Ledger creation failed in Tally.",
        "response": parsed,
    }


def _safe_int(text, default=0) -> int:
    try:
        return int(float(str(text).strip()))
    except (TypeError, ValueError):
        return default


def _parse_tally_response_details(response_text: str) -> dict:
    details = {
        "success": False,
        "created": 0,
        "altered": 0,
        "deleted": 0,
        "lastvchid": 0,
        "lastmid": 0,
        "combined": 0,
        "ignored": 0,
        "errors": 0,
        "cancelled": 0,
        "exceptions": 0,
        "line_errors": [],
        "error": "",
    }

    try:
        root = ET.fromstring(response_text)
    except ET.ParseError as exc:
        details["error"] = f"Could not parse Tally response: {exc}"
        return details

    def _count(tag_name: str) -> int:
        node = root.find(f".//{tag_name}")
        return _safe_int(node.text if node is not None else 0)

    details["created"] = _count("CREATED")
    details["altered"] = _count("ALTERED")
    details["deleted"] = _count("DELETED")
    details["lastvchid"] = _count("LASTVCHID")
    details["lastmid"] = _count("LASTMID")
    details["combined"] = _count("COMBINED")
    details["ignored"] = _count("IGNORED")
    details["errors"] = _count("ERRORS")
    details["cancelled"] = _count("CANCELLED")
    details["exceptions"] = _count("EXCEPTIONS")

    line_errors = []
    seen = set()
    for node in root.findall(".//LINEERROR"):
        text = (node.text or "").strip()
        if text and text not in seen:
            seen.add(text)
            line_errors.append(text)
    details["line_errors"] = line_errors

    details["success"] = (
        details["errors"] == 0
        and details["exceptions"] == 0
        and not details["line_errors"]
    )
    if not details["success"] and details["line_errors"]:
        details["error"] = details["line_errors"][0]

    return details


def _company_static_block(company: str) -> str:
    selected = str(company or "").strip()
    if not selected:
        return ""
    return f"   <STATICVARIABLES><SVCURRENTCOMPANY>{xml_escape(selected)}</SVCURRENTCOMPANY></STATICVARIABLES>"


def _clean_tax_ledger(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.casefold() in {"0", "0.0", "none", "na", "n/a", "-"}:
        return ""
    return text


def _normalize_note_type(value: str) -> str:
    text = str(value or "").strip().casefold()
    if text in {"debit note", "debit", "debitnote"}:
        return "Debit Note"
    return "Credit Note"


def generate_note_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
    voucher_type: str = "Credit Note",
    company_gst_registrations: list = None,
) -> tuple:
    """
    Note accounting logic:
    - Credit Note: party credited, particular/tax debited
    - Debit Note: party debited, particular/tax credited
    """
    normalized_type = _normalize_note_type(voucher_type)
    is_debit_note = normalized_type == "Debit Note"
    default_particular_ledger = f"{normalized_type} Account"

    lines = []
    a = lines.append
    company_static = _company_static_block(company)

    a('<?xml version="1.0" encoding="UTF-8"?>')
    a("<ENVELOPE>")
    a(" <HEADER><TALLYREQUEST>Import Data</TALLYREQUEST></HEADER>")
    a(" <BODY><IMPORTDATA>")
    a("  <REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME>")
    if company_static:
        a(company_static)
    a("  </REQUESTDESC>")
    a("  <REQUESTDATA>")

    voucher_count = 0
    resolved_mode = str(date_mode or ("current" if use_today_date else "excel")).strip().lower()
    if resolved_mode not in {"current", "excel", "custom"}:
        resolved_mode = "current" if use_today_date else "excel"
    resolved_custom_date = _normalize_manual_date_to_tally(custom_tally_date) if resolved_mode == "custom" else ""

    for idx, r in enumerate(rows):
        taxable = _row_float(r, "TaxableValue", 0.0)
        if taxable <= 0:
            continue

        if resolved_mode == "current":
            source_date = datetime.today()
        elif resolved_mode == "custom":
            source_date = resolved_custom_date
        else:
            source_date = _row_get(r, "Date", "")
        dt = tally_date(source_date)
        vno_raw = _row_voucher_number(r, str(idx + 1))

        party_raw = _ledger_or_default(_row_text(r, "PartyLedger"))
        particular_raw = (
            _row_text(r, "Particular")
            or _row_text(r, "Particulars")
            or _row_text(r, "SalesLedger")
            or _row_text(r, "Sales Ledger")
            or _row_text(r, "Purchase Ledger")
            or default_particular_ledger
        )
        particular_raw = _ledger_or_default(particular_raw, default_particular_ledger)

        cgst_ledger_raw = _clean_tax_ledger(_row_text(r, "CGSTLedger"))
        sgst_ledger_raw = _clean_tax_ledger(_row_text(r, "SGSTLedger"))
        igst_ledger_raw = _clean_tax_ledger(_row_text(r, "IGSTLedger"))

        cgst_rate = _row_float(r, "CGSTRate", 0.0)
        sgst_rate = _row_float(r, "SGSTRate", 0.0)
        igst_rate = _row_float(r, "IGSTRate", 0.0)

        cgst_amt = round(taxable * cgst_rate / 100, 2) if cgst_rate > 0 else 0.0
        sgst_amt = round(taxable * sgst_rate / 100, 2) if sgst_rate > 0 else 0.0
        igst_amt = round(taxable * igst_rate / 100, 2) if igst_rate > 0 else 0.0
        total = taxable + cgst_amt + sgst_amt + igst_amt

        vno = xml_escape(vno_raw)
        party = xml_escape(party_raw)
        particular = xml_escape(particular_raw)
        narration = xml_escape(_row_text(r, "Narration"))
        
        gstin_raw = _row_text(r, "GSTIN") or _row_text(r, "PartyGSTIN")
        gstin = xml_escape(gstin_raw)

        # Extract state from GSTIN
        state_name_raw = _state_name_from_gstin(gstin_raw)
        state_xml = xml_escape(state_name_raw)

        party_is_deemed_positive = "Yes" if is_debit_note else "No"
        party_amount = -total if is_debit_note else total
        counter_is_deemed_positive = "No" if is_debit_note else "Yes"
        taxable_amount = taxable if is_debit_note else -taxable
        cgst_amount = cgst_amt if is_debit_note else -cgst_amt
        sgst_amount = sgst_amt if is_debit_note else -sgst_amt
        igst_amount = igst_amt if is_debit_note else -igst_amt

        voucher_count += 1
        a('   <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        a(f'    <VOUCHER VCHTYPE="{normalized_type}" ACTION="Create" OBJVIEW="Invoice Voucher View">')
        a(f"     <DATE>{dt}</DATE>")
        a(f"     <VOUCHERTYPENAME>{normalized_type}</VOUCHERTYPENAME>")
        a(f"     <VOUCHERNUMBER>{vno}</VOUCHERNUMBER>")
        
        # --- FIX: Handle Place of Supply correctly for Credit vs Debit Note ---
        a(f"     <PARTYLEDGERNAME>{party}</PARTYLEDGERNAME>")
        a(f"     <PARTYNAME>{party}</PARTYNAME>")
        
        if not is_debit_note:
            # Credit Note (Sales Return) - Party is Buyer
            a(f"     <BASICBUYERNAME>{party}</BASICBUYERNAME>")
            if state_xml:
                a(f"     <STATENAME>{state_xml}</STATENAME>")
                a(f"     <PLACEOFSUPPLY>{state_xml}</PLACEOFSUPPLY>")
        else:
            # Debit Note (Purchase Return) - Party is Supplier.
            # Tally requires GSTREGISTRATIONTYPE, PLACEOFSUPPLY (= company state),
            # ISGSTOVERRIDDEN and GSTTRANSACTIONTYPE for inward-supply vouchers, otherwise
            # the voucher shows as "Uncertain" in GSTR-3B reports.
            # Mirror what purchase_accounting / purchase_item do in sale_purchase_entry.py:
            #   PLACEOFSUPPLY = place_of_supply_override = purchase_company_state (company state)
            #   (NOT the supplier's state — that caused master-alteration popup earlier)
            _cmp_regs = list(company_gst_registrations or [])
            _cmp_state = ""
            _cmp_gstin = ""
            _cmp_name = ""
            if _cmp_regs:
                _cmp_reg = _cmp_regs[0]
                _cmp_gstin = xml_escape(str(_cmp_reg.get("gstin", "") or "").strip())
                _cmp_state = xml_escape(str(_cmp_reg.get("state", "") or "").strip())
                _cmp_name = xml_escape(str(_cmp_reg.get("name", "") or "").strip())

            if state_xml:
                # STATENAME = the supplier's (party's) state.
                a(f"     <STATENAME>{state_xml}</STATENAME>")

            # PLACEOFSUPPLY for inward supply = company's state (buyer's state).
            # This is what resolves the GSTR-3B "Uncertain" status and suppresses the
            # "GST Registration Details" acceptance popup.  Do NOT use the supplier's
            # state here — that caused the "Mismatch / master alteration" issue.
            if _cmp_state:
                a(f"     <PLACEOFSUPPLY>{_cmp_state}</PLACEOFSUPPLY>")

            # Party GST registration type — mandatory for Tally to resolve company GST context.
            _dn_reg_type = "Regular" if gstin_raw else "Unregistered"
            a(f"     <GSTREGISTRATIONTYPE>{_dn_reg_type}</GSTREGISTRATIONTYPE>")
            if gstin_raw:
                a("     <VATDEALERTYPE>Regular</VATDEALERTYPE>")

            # Embed company GST context (CMPGSTIN / CMPGSTSTATE) to fully suppress
            # Tally's "GST Registration Details" acceptance popup.
            if _cmp_gstin and _cmp_name:
                a(f'     <GSTREGISTRATION TAXTYPE="GST" TAXREGISTRATION="{_cmp_gstin}">{_cmp_name}</GSTREGISTRATION>')
                a(f'     <CMPGSTIN>{_cmp_gstin}</CMPGSTIN>')
                a('     <CMPGSTREGISTRATIONTYPE>Regular</CMPGSTREGISTRATIONTYPE>')
            if _cmp_state:
                a(f'     <CMPGSTSTATE>{_cmp_state}</CMPGSTSTATE>')

        a("      <COUNTRYOFRESIDENCE>India</COUNTRYOFRESIDENCE>")
        # ----------------------------------------------------------------

        a(f"     <EFFECTIVEDATE>{dt}</EFFECTIVEDATE>")
        a("      <ISINVOICE>Yes</ISINVOICE>")
        a("      <PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW>")
        a("      <VCHENTRYMODE>Accounting Invoice</VCHENTRYMODE>")
        if is_debit_note:
            # Required for Tally to accept the debit note without "company details not
            # specified" error and without showing an Accept Confirmation popup.
            a("      <ISGSTOVERRIDDEN>No</ISGSTOVERRIDDEN>")
            _dn_gst_txn = "Tax Invoice" if gstin_raw else "Unregistered"
            a(f"     <GSTTRANSACTIONTYPE>{_dn_gst_txn}</GSTTRANSACTIONTYPE>")
        if gstin:
            a(f"     <PARTYGSTIN>{gstin}</PARTYGSTIN>")
        if narration:
            a(f"     <NARRATION>{narration}</NARRATION>")

        # Party ledger line
        a("      <LEDGERENTRIES.LIST>")
        a(f"       <LEDGERNAME>{party}</LEDGERNAME>")
        a(f"       <ISDEEMEDPOSITIVE>{party_is_deemed_positive}</ISDEEMEDPOSITIVE>")
        a(f"       <AMOUNT>{fmt_amt(party_amount)}</AMOUNT>")
        a("      </LEDGERENTRIES.LIST>")

        # Particular ledger line
        a("      <LEDGERENTRIES.LIST>")
        a(f"       <LEDGERNAME>{particular}</LEDGERNAME>")
        a(f"       <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
        a(f"       <AMOUNT>{fmt_amt(taxable_amount)}</AMOUNT>")
        a("      </LEDGERENTRIES.LIST>")

        if cgst_amt > 0 and cgst_ledger_raw:
            a("      <LEDGERENTRIES.LIST>")
            a(f"       <LEDGERNAME>{xml_escape(cgst_ledger_raw)}</LEDGERNAME>")
            a(f"       <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
            a(f"       <AMOUNT>{fmt_amt(cgst_amount)}</AMOUNT>")
            a("      </LEDGERENTRIES.LIST>")

        if sgst_amt > 0 and sgst_ledger_raw:
            a("      <LEDGERENTRIES.LIST>")
            a(f"       <LEDGERNAME>{xml_escape(sgst_ledger_raw)}</LEDGERNAME>")
            a(f"       <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
            a(f"       <AMOUNT>{fmt_amt(sgst_amount)}</AMOUNT>")
            a("      </LEDGERENTRIES.LIST>")

        if igst_amt > 0 and igst_ledger_raw:
            a("      <LEDGERENTRIES.LIST>")
            a(f"       <LEDGERNAME>{xml_escape(igst_ledger_raw)}</LEDGERNAME>")
            a(f"       <ISDEEMEDPOSITIVE>{counter_is_deemed_positive}</ISDEEMEDPOSITIVE>")
            a(f"       <AMOUNT>{fmt_amt(igst_amount)}</AMOUNT>")
            a("      </LEDGERENTRIES.LIST>")

        a("    </VOUCHER>")
        a("   </TALLYMESSAGE>")

    a("  </REQUESTDATA>")
    a(" </IMPORTDATA></BODY>")
    a("</ENVELOPE>")
    return "\n".join(lines), voucher_count


def generate_credit_note_xml(
    rows: list,
    company: str,
    use_today_date: bool = False,
    date_mode: str = "",
    custom_tally_date: str = "",
) -> tuple:
    return generate_note_xml(
        rows,
        company=company,
        use_today_date=use_today_date,
        date_mode=date_mode,
        custom_tally_date=custom_tally_date,
        voucher_type="Credit Note",
    )


def read_excel(filepath: str, sheet: str = None) -> tuple:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(c or "").strip() for c in first_row]
    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        vals = list(vals[: len(headers)])
        if len(vals) < len(headers):
            vals.extend([None] * (len(headers) - len(vals)))
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    wb.close()
    return headers, rows


class TallyNoteEntryApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("TallyNoteEntryPro - Credit and Debit Note Entry Tool")
        self.geometry("1120x720")
        self.minsize(980, 640)
        self.configure(fg_color=COLORS["bg_dark"])

        self.loaded_rows = []
        self.loaded_headers = []
        self.manual_rows = []

        self.file_path_var = ctk.StringVar(value="")
        self.company_placeholder = "Auto (Loaded Company)"
        self.company_var = ctk.StringVar(value=self.company_placeholder)
        self.tally_host_var = ctk.StringVar(value="localhost")
        self.tally_port_var = ctk.StringVar(value="9000")
        self.voucher_date_mode_var = ctk.StringVar(value="excel")
        self.voucher_custom_date_var = ctk.StringVar(value="")
        self.voucher_date_checks = {
            "current": ctk.BooleanVar(value=False),
            "excel": ctk.BooleanVar(value=True),
            "custom": ctk.BooleanVar(value=False),
        }
        self.note_type_options = ["Credit Note", "Debit Note"]
        self.note_type_var = ctk.StringVar(value="Credit Note")
        self.workflow_demo_url = ""

        self.status_var = ctk.StringVar(value="Ready")
        self.connection_status_var = ctk.StringVar(value="Connection: Not checked")
        self.company_status_var = ctk.StringVar(value="Companies: Not fetched")
        self.party_ledger_status_var = ctk.StringVar(value="Party Ledgers: Not fetched")
        self.create_party_status_var = ctk.StringVar(value="Ready to create party ledger")

        self.fetched_companies = []
        self.company_gst_registrations = []
        self.fetched_party_ledgers = []
        self._company_fetch_running = False
        self._ledger_fetch_running = False
        self._excel_load_running = False
        self._push_running = False
        self._create_party_running = False

        self._push_overlay = None
        self._push_message_var = ctk.StringVar(value="")
        self.debug_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tally_note_entry_debug.log")

        self.excel_tree = None
        self.manual_tree = None
        self.excel_info_label = None
        self.manual_info_label = None
        self.manual_form_vars = {}
        self.manual_action_buttons = []
        self.manual_add_btn = None
        self.manual_update_btn = None
        self.manual_editing_index = None
        self.manual_party_ledger_combo = None
        self.manual_fetch_ledger_btn = None
        self.manual_party_search_var = ctk.StringVar(value="")
        self.manual_party_search_entry = None
        self.manual_party_search_clear_btn = None
        self.manual_party_match_label = None

        self.browse_btn = None
        self.connection_test_btn = None
        self.company_refresh_btn = None
        self.company_combo = None
        self.note_type_combo = None
        self.save_xml_btn = None
        self.push_tally_btn = None
        self.template_btn = None
        self.demo_btn = None
        self.party_ledger_status_label = None
        self.voucher_date_current_cb = None
        self.voucher_date_excel_cb = None
        self.voucher_date_custom_cb = None
        self.voucher_custom_date_entry = None

        self.create_party_name_entry = None
        self.create_party_parent_cb = None
        self.create_party_mailing_entry = None
        self.create_party_gstin_entry = None
        self.create_party_state_entry = None
        self.create_party_country_entry = None
        self.create_party_pincode_entry = None
        self.create_party_address1_entry = None
        self.create_party_address2_entry = None
        self.create_party_billwise_cb = None
        self.create_party_create_btn = None
        self.create_party_fetch_btn = None
        self.create_party_clear_btn = None

        self._build_ui()
        self.set_theme(ctk.get_appearance_mode())

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        settings_card = ctk.CTkFrame(
            self,
            fg_color=COLORS["bg_card"],
            border_width=1,
            border_color=COLORS["border"],
            corner_radius=12,
        )
        settings_card.grid(row=0, column=0, sticky="ew", padx=16, pady=(10, 8))

        row_1 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_1.pack(fill="x", padx=14, pady=(10, 4))
        ctk.CTkLabel(row_1, text="Host", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        ctk.CTkEntry(
            row_1,
            textvariable=self.tally_host_var,
            width=140,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
        ).pack(side="left", padx=(6, 12))

        ctk.CTkLabel(row_1, text="Port", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        ctk.CTkEntry(
            row_1,
            textvariable=self.tally_port_var,
            width=90,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
        ).pack(side="left", padx=(6, 12))

        self.connection_test_btn = ctk.CTkButton(
            row_1,
            text="Test Connection",
            width=130,
            height=32,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["warning"],
            hover_color="#B45309",
            text_color="#FFFFFF",
            corner_radius=8,
            command=self._check_tally_connection_thread,
        )
        self.connection_test_btn.pack(side="right")

        self.demo_btn = ctk.CTkButton(
            row_1,
            text="▶ View Demo",
            width=132,
            height=32,
            font=("Segoe UI", 10, "bold"),
            fg_color="#DC2626",
            hover_color="#B91C1C",
            text_color="#FFFFFF",
            corner_radius=8,
            command=self._view_workflow_demo,
        )
        self.demo_btn.pack(side="right", padx=(0, 8))

        row_date = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_date.pack(fill="x", padx=14, pady=(0, 4))
        ctk.CTkLabel(row_date, text="Voucher Date", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")

        checks_wrap = ctk.CTkFrame(row_date, fg_color="transparent")
        checks_wrap.pack(side="left", padx=(8, 12))

        self.voucher_date_current_cb = ctk.CTkCheckBox(
            checks_wrap,
            text="Current Date",
            variable=self.voucher_date_checks["current"],
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            border_color=COLORS["border"],
            command=lambda: self._set_voucher_date_mode("current"),
        )
        self.voucher_date_current_cb.pack(side="left", padx=(0, 8))

        self.voucher_date_excel_cb = ctk.CTkCheckBox(
            checks_wrap,
            text="Excel Date",
            variable=self.voucher_date_checks["excel"],
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            border_color=COLORS["border"],
            command=lambda: self._set_voucher_date_mode("excel"),
        )
        self.voucher_date_excel_cb.pack(side="left", padx=(0, 8))

        self.voucher_date_custom_cb = ctk.CTkCheckBox(
            checks_wrap,
            text="Custom Date",
            variable=self.voucher_date_checks["custom"],
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            border_color=COLORS["border"],
            command=lambda: self._set_voucher_date_mode("custom"),
        )
        self.voucher_date_custom_cb.pack(side="left")

        self.voucher_custom_date_entry = ctk.CTkEntry(
            row_date,
            textvariable=self.voucher_custom_date_var,
            width=180,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="DD/MM/YYYY",
            font=("Segoe UI", 10),
        )
        self.voucher_custom_date_entry.pack(side="right")
        self._set_voucher_date_mode("excel")

        row_2 = ctk.CTkFrame(settings_card, fg_color="transparent")
        row_2.pack(fill="x", padx=14, pady=(0, 4))

        ctk.CTkLabel(row_2, text="Voucher Type", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        self.note_type_combo = ctk.CTkComboBox(
            row_2,
            values=self.note_type_options,
            variable=self.note_type_var,
            width=140,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
            state="readonly",
        )
        self.note_type_combo.pack(side="left", padx=(8, 16))

        ctk.CTkLabel(row_2, text="Target Company", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")

        self.company_combo = ctk.CTkComboBox(
            row_2,
            values=[self.company_placeholder],
            variable=self.company_var,
            width=340,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.company_combo.set(self.company_placeholder)
        self.company_combo.pack(side="left", padx=(10, 8), fill="x", expand=True)

        self.company_refresh_btn = ctk.CTkButton(
            row_2,
            text="Refresh",
            width=96,
            height=34,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=lambda: self._fetch_tally_companies_thread(),
        )
        self.company_refresh_btn.pack(side="right")

        status_row = ctk.CTkFrame(settings_card, fg_color="transparent")
        status_row.pack(fill="x", padx=14, pady=(0, 8))
        status_row.grid_columnconfigure(0, weight=1)
        status_row.grid_columnconfigure(1, weight=1)
        status_row.grid_columnconfigure(2, weight=1)

        self.connection_status_label = ctk.CTkLabel(
            status_row,
            textvariable=self.connection_status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.connection_status_label.grid(row=0, column=0, sticky="w", padx=(0, 8))

        self.company_status_label = ctk.CTkLabel(
            status_row,
            textvariable=self.company_status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.company_status_label.grid(row=0, column=1, sticky="w", padx=(0, 8))

        self.party_ledger_status_label = ctk.CTkLabel(
            status_row,
            textvariable=self.party_ledger_status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.party_ledger_status_label.grid(row=0, column=2, sticky="w")

        content_card = ctk.CTkFrame(
            self,
            fg_color=COLORS["bg_card"],
            border_width=1,
            border_color=COLORS["border"],
            corner_radius=12,
        )
        content_card.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 10))

        self.source_tabs = ctk.CTkTabview(
            content_card,
            fg_color="transparent",
            segmented_button_fg_color=COLORS["bg_input"],
            segmented_button_selected_color=COLORS["accent"],
            segmented_button_selected_hover_color=COLORS["accent_hover"],
            segmented_button_unselected_color=COLORS["bg_input"],
            segmented_button_unselected_hover_color=COLORS["bg_card_hover"],
        )
        self.source_tabs.pack(fill="both", expand=True, padx=10, pady=(10, 5))

        excel_tab = self.source_tabs.add("Excel Upload")
        manual_tab = self.source_tabs.add("Manual Entry")
        create_party_tab = self.source_tabs.add("Create Party Ledger")

        self._build_excel_tab(excel_tab)
        self._build_manual_tab(manual_tab)
        self._build_create_party_tab(create_party_tab)
        self.source_tabs.set("Excel Upload")

        action_bar = ctk.CTkFrame(
            self,
            fg_color=COLORS["bg_card"],
            border_width=1,
            border_color=COLORS["border"],
            corner_radius=12,
        )
        action_bar.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 10))

        action_left = ctk.CTkFrame(action_bar, fg_color="transparent")
        action_left.pack(side="left", padx=10, pady=10)

        self.save_xml_btn = ctk.CTkButton(
            action_left,
            text="Save XML File",
            fg_color=SUCCESS,
            hover_color="#15803D",
            width=170,
            command=lambda: self._generate("save"),
        )
        self.save_xml_btn.pack(side="left", padx=(0, 10))

        self.push_tally_btn = ctk.CTkButton(
            action_left,
            text="Push to Tally",
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER,
            width=170,
            command=lambda: self._generate("push"),
        )
        self.push_tally_btn.pack(side="left")

        status_bar = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=0, height=32)
        status_bar.grid(row=3, column=0, sticky="ew")
        status_bar.grid_propagate(False)
        ctk.CTkLabel(
            status_bar,
            textvariable=self.status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        ).pack(side="left", padx=16)

        self.after(200, lambda: self._fetch_tally_companies_thread(silent=True))

    def _build_excel_tab(self, parent):
        template_row = ctk.CTkFrame(parent, fg_color="transparent")
        template_row.pack(fill="x", padx=10, pady=(10, 4))

        self.template_btn = ctk.CTkButton(
            template_row,
            text="Download Template",
            fg_color="#94A3B8",
            hover_color="#64748B",
            text_color="#FFFFFF",
            width=170,
            command=self._download_template,
        )
        self.template_btn.pack(side="right")

        load_frame = ctk.CTkFrame(parent, fg_color="transparent")
        load_frame.pack(fill="x", padx=10, pady=(0, 5))

        ctk.CTkEntry(
            load_frame,
            textvariable=self.file_path_var,
            placeholder_text="Select Credit/Debit Note Excel (.xlsx/.xlsm/.xls)",
            width=600,
            state="readonly",
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
        ).pack(side="left", padx=(0, 8), fill="x", expand=True)

        self.browse_btn = ctk.CTkButton(
            load_frame,
            text="Browse Excel",
            command=self._browse_file,
            width=120,
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER,
        )
        self.browse_btn.pack(side="left", padx=(0, 8))

        info_frame = ctk.CTkFrame(parent, fg_color="transparent")
        info_frame.pack(fill="x", padx=10, pady=(0, 5))
        self.excel_info_label = ctk.CTkLabel(info_frame, text="", font=("Segoe UI", 11), text_color=TEXT_MUTED)
        self.excel_info_label.pack(side="left")

        tree_frame = ctk.CTkFrame(
            parent,
            fg_color=COLORS["bg_dark"],
            corner_radius=8,
            border_width=1,
            border_color=COLORS["border"],
        )
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        self.excel_tree = ttk.Treeview(
            tree_frame,
            show="headings",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
        )
        tree_scroll_y.config(command=self.excel_tree.yview)
        tree_scroll_x.config(command=self.excel_tree.xview)
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        self.excel_tree.pack(fill="both", expand=True)

        self._setup_tree_style()

    def _build_manual_tab(self, parent):
        wrapper = ctk.CTkFrame(parent, fg_color="transparent")
        wrapper.pack(fill="both", expand=True, padx=10, pady=8)
        wrapper.grid_columnconfigure(0, weight=1, uniform="manual_split")
        wrapper.grid_columnconfigure(1, weight=1, uniform="manual_split")
        wrapper.grid_rowconfigure(0, weight=1)

        left_panel = ctk.CTkFrame(
            wrapper,
            fg_color=COLORS["bg_card"],
            border_width=1,
            border_color=COLORS["border"],
            corner_radius=10,
        )
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        left_panel.grid_columnconfigure(0, weight=1)
        left_panel.grid_rowconfigure(0, weight=1)

        form_scroll = ctk.CTkScrollableFrame(
            left_panel,
            fg_color="transparent",
            corner_radius=8,
        )
        form_scroll.grid(row=0, column=0, sticky="nsew", padx=8, pady=(8, 6))

        form_card = ctk.CTkFrame(form_scroll, fg_color="transparent")
        form_card.pack(fill="x", padx=2, pady=2)

        fields = [
            ("Date", "Date", "DD-MM-YY"),
            ("VoucherNo", "Voucher No", "Optional"),
            ("GSTIN", "GSTIN", "Optional"),
            ("PartyLedger", "Party Ledger", "Required"),
            ("Particular", "Particular", "Required"),
            ("TaxableValue", "Taxable Value", "Required Amount"),
            ("CGSTLedger", "CGST Ledger", "Optional"),
            ("CGSTRate", "CGST Rate", "0"),
            ("SGSTLedger", "SGST Ledger", "Optional"),
            ("SGSTRate", "SGST Rate", "0"),
            ("IGSTLedger", "IGST Ledger", "Optional"),
            ("IGSTRate", "IGST Rate", "0"),
            ("Narration", "Narration", "Optional"),
        ]

        cols = 2
        for i, (key, label, placeholder) in enumerate(fields):
            row = i // cols
            col = i % cols
            field_wrap = ctk.CTkFrame(form_card, fg_color="transparent")
            field_wrap.grid(row=row, column=col, sticky="ew", padx=8, pady=6)
            form_card.grid_columnconfigure(col, weight=1)

            if key == "PartyLedger":
                top_line = ctk.CTkFrame(field_wrap, fg_color="transparent")
                top_line.pack(fill="x")

                ctk.CTkLabel(
                    top_line,
                    text=label,
                    font=("Segoe UI", 10),
                    text_color=COLORS["text_secondary"],
                ).pack(side="left")

                self.manual_fetch_ledger_btn = ctk.CTkButton(
                    top_line,
                    text="Fetch",
                    width=70,
                    height=26,
                    font=("Segoe UI", 10, "bold"),
                    fg_color=COLORS["bg_input"],
                    hover_color=COLORS["bg_card_hover"],
                    text_color=COLORS["text_secondary"],
                    command=self._fetch_party_ledgers_thread,
                )
                self.manual_fetch_ledger_btn.pack(side="right")

                search_row = ctk.CTkFrame(field_wrap, fg_color="transparent")
                search_row.pack(fill="x", pady=(4, 2))

                self.manual_party_search_entry = ctk.CTkEntry(
                    search_row,
                    textvariable=self.manual_party_search_var,
                    placeholder_text="Search party ledger...",
                    height=34,
                    fg_color=COLORS["bg_input"],
                    border_color=COLORS["border"],
                    font=("Segoe UI", 10),
                )
                self.manual_party_search_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
                self.manual_party_search_entry.bind("<KeyRelease>", self._on_party_ledger_search_change)

                self.manual_party_search_clear_btn = ctk.CTkButton(
                    search_row,
                    text="Clear",
                    width=58,
                    height=30,
                    fg_color=COLORS["bg_input"],
                    hover_color=COLORS["bg_card_hover"],
                    text_color=COLORS["text_secondary"],
                    font=("Segoe UI", 9, "bold"),
                    command=self._clear_party_ledger_search,
                )
                self.manual_party_search_clear_btn.pack(side="right")

                var = ctk.StringVar(value="")
                combo = ctk.CTkComboBox(
                    field_wrap,
                    values=[""],
                    variable=var,
                    height=36,
                    fg_color=COLORS["bg_input"],
                    border_color=COLORS["border"],
                    button_color=COLORS["accent"],
                    button_hover_color=COLORS["accent_hover"],
                    font=("Segoe UI", 10),
                    state="readonly",
                )
                combo.pack(fill="x", pady=(0, 2))

                self.manual_party_match_label = ctk.CTkLabel(
                    field_wrap,
                    text="Type in search box after fetching ledgers",
                    font=("Segoe UI", 9),
                    text_color=COLORS["text_muted"],
                )
                self.manual_party_match_label.pack(anchor="w")

                self.manual_party_search_var.trace_add("write", lambda *_: self._on_party_ledger_search_change())
                self.manual_party_ledger_combo = combo
                self.manual_form_vars[key] = var
                continue

            ctk.CTkLabel(
                field_wrap,
                text=label,
                font=("Segoe UI", 10),
                text_color=COLORS["text_secondary"],
            ).pack(anchor="w")

            var = ctk.StringVar(value="")
            entry = ctk.CTkEntry(
                field_wrap,
                textvariable=var,
                placeholder_text=placeholder,
                height=36,
                fg_color=COLORS["bg_input"],
                border_color=COLORS["border"],
            )
            entry.pack(fill="x")
            self.manual_form_vars[key] = var

        self.manual_form_vars["Date"].set(datetime.today().strftime("%d-%m-%y"))

        btn_row = ctk.CTkFrame(left_panel, fg_color="transparent")
        btn_row.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 8))
        for col_idx in range(3):
            btn_row.grid_columnconfigure(col_idx, weight=1)

        add_btn = ctk.CTkButton(
            btn_row,
            text="Add Entry",
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER,
            height=34,
            command=self._add_manual_entry,
        )
        add_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6), pady=(0, 6))
        self.manual_add_btn = add_btn

        edit_selected_btn = ctk.CTkButton(
            btn_row,
            text="Edit Selected",
            fg_color="#0EA5E9",
            hover_color="#0284C7",
            text_color="#FFFFFF",
            height=34,
            command=self._edit_selected_manual,
        )
        edit_selected_btn.grid(row=0, column=1, sticky="ew", padx=3, pady=(0, 6))

        update_btn = ctk.CTkButton(
            btn_row,
            text="Update Entry",
            fg_color="#10B981",
            hover_color="#059669",
            text_color="#FFFFFF",
            height=34,
            state="disabled",
            command=self._update_manual_entry,
        )
        update_btn.grid(row=0, column=2, sticky="ew", padx=(6, 0), pady=(0, 6))
        self.manual_update_btn = update_btn

        clear_form_btn = ctk.CTkButton(
            btn_row,
            text="Clear Form",
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            height=34,
            command=self._clear_manual_form,
        )
        clear_form_btn.grid(row=1, column=0, sticky="ew", padx=(0, 6))

        remove_selected_btn = ctk.CTkButton(
            btn_row,
            text="Remove Selected",
            fg_color=COLORS["warning"],
            hover_color="#B45309",
            text_color="#FFFFFF",
            height=34,
            command=self._remove_selected_manual,
        )
        remove_selected_btn.grid(row=1, column=1, sticky="ew", padx=3)

        clear_all_btn = ctk.CTkButton(
            btn_row,
            text="Clear All",
            fg_color=COLORS["error"],
            hover_color="#B91C1C",
            text_color="#FFFFFF",
            height=34,
            command=self._clear_all_manual,
        )
        clear_all_btn.grid(row=1, column=2, sticky="ew", padx=(6, 0))

        self.manual_action_buttons = [
            add_btn,
            edit_selected_btn,
            update_btn,
            clear_form_btn,
            remove_selected_btn,
            clear_all_btn,
        ]

        right_panel = ctk.CTkFrame(
            wrapper,
            fg_color=COLORS["bg_card"],
            border_width=1,
            border_color=COLORS["border"],
            corner_radius=10,
        )
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(1, weight=1)

        right_header = ctk.CTkFrame(right_panel, fg_color="transparent")
        right_header.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 6))

        ctk.CTkLabel(
            right_header,
            text="Review (Excel Format)",
            font=("Segoe UI", 12, "bold"),
            text_color=COLORS["text_primary"],
        ).pack(side="left")

        self.manual_info_label = ctk.CTkLabel(
            right_header,
            text="Manual entries: 0",
            font=("Segoe UI", 11),
            text_color=TEXT_MUTED,
        )
        self.manual_info_label.pack(side="right")

        tree_frame = ctk.CTkFrame(
            right_panel,
            fg_color=COLORS["bg_dark"],
            corner_radius=8,
            border_width=1,
            border_color=COLORS["border"],
        )
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        self.manual_tree = ttk.Treeview(
            tree_frame,
            show="headings",
            selectmode="extended",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
        )
        tree_scroll_y.config(command=self.manual_tree.yview)
        tree_scroll_x.config(command=self.manual_tree.xview)
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        self.manual_tree.pack(fill="both", expand=True)
        self.manual_tree.bind("<Double-1>", self._on_manual_tree_double_click)

        ctk.CTkLabel(
            right_panel,
            text="Select a row and click Edit Selected, or double-click to edit it.",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        ).grid(row=2, column=0, sticky="w", padx=10, pady=(0, 8))

        self._setup_tree_style()
        self._populate_tree(self.manual_tree, TEMPLATE_HEADERS, [])

    def _set_party_ledger_values(self, ledgers, keep_current=True):
        cleaned = []
        seen = set()
        for name in ledgers or []:
            normalized = _normalize_ledger_name(name)
            if not normalized:
                continue
            key = _ledger_key(normalized)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(normalized)
        cleaned.sort(key=lambda x: _ledger_key(x))

        self.fetched_party_ledgers = cleaned

        if self.manual_party_ledger_combo is not None:
            current = ""
            if keep_current and self.manual_form_vars.get("PartyLedger") is not None:
                current = (self.manual_form_vars.get("PartyLedger").get() or "").strip()
            values = cleaned[:200] if cleaned else [""]
            self.manual_party_ledger_combo.configure(values=values)
            if current:
                self.manual_form_vars["PartyLedger"].set(current)

        self._on_party_ledger_search_change()

        self.party_ledger_status_var.set(f"Party Ledgers: {len(cleaned)} available")
        if self.party_ledger_status_label is not None:
            self.party_ledger_status_label.configure(text_color=COLORS["text_muted"])

    def _on_party_ledger_search_change(self, _event=None):
        if self.manual_party_ledger_combo is None:
            return
        if not self.fetched_party_ledgers:
            self.manual_party_ledger_combo.configure(values=[""])
            self.manual_party_ledger_combo.set("")
            if self.manual_party_match_label is not None:
                self.manual_party_match_label.configure(text="No fetched party ledgers yet")
            return

        typed_text = (self.manual_party_search_var.get() or "").strip()
        typed = typed_text.casefold()
        current_value = ""
        if self.manual_form_vars.get("PartyLedger") is not None:
            current_value = (self.manual_form_vars["PartyLedger"].get() or "").strip()

        if not typed:
            filtered = self.fetched_party_ledgers[:200]
        else:
            starts = [name for name in self.fetched_party_ledgers if name.casefold().startswith(typed)]
            contains = [
                name
                for name in self.fetched_party_ledgers
                if typed in name.casefold() and name not in starts
            ]
            filtered = (starts + contains)[:200]

        if typed and not filtered:
            self.manual_party_ledger_combo.configure(values=[""])
            self.manual_party_ledger_combo.set("")
            if self.manual_form_vars.get("PartyLedger") is not None:
                self.manual_form_vars["PartyLedger"].set("")
            if self.manual_party_match_label is not None:
                self.manual_party_match_label.configure(text=f"Search '{typed_text}': no matching ledger")
            return

        display_values = filtered if filtered else self.fetched_party_ledgers[:200]
        self.manual_party_ledger_combo.configure(values=display_values)

        if typed and display_values and current_value not in display_values:
            first_match = display_values[0]
            self.manual_party_ledger_combo.set(first_match)
            if self.manual_form_vars.get("PartyLedger") is not None:
                self.manual_form_vars["PartyLedger"].set(first_match)
        elif current_value:
            self.manual_party_ledger_combo.set(current_value)

        if self.manual_party_match_label is not None:
            shown = len(display_values)
            total = len(self.fetched_party_ledgers)
            if typed:
                self.manual_party_match_label.configure(text=f"Search '{typed_text}': showing {shown} of {total}")
            else:
                self.manual_party_match_label.configure(text=f"Showing {shown} of {total} party ledgers")

    def _clear_party_ledger_search(self):
        if self.manual_party_search_var is not None:
            self.manual_party_search_var.set("")
        self._on_party_ledger_search_change()

    def _fetch_party_ledgers_thread(self, silent=False):
        if self._ledger_fetch_running:
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            if not silent:
                messagebox.showerror("Invalid Settings", str(exc))
            return

        selected_company = self._get_selected_company()
        self._ledger_fetch_running = True

        if self.manual_fetch_ledger_btn is not None:
            self.manual_fetch_ledger_btn.configure(state="disabled", text="...")
        if self.create_party_fetch_btn is not None:
            self.create_party_fetch_btn.configure(state="disabled", text="Fetching...")

        self.party_ledger_status_var.set("Party Ledgers: Fetching...")
        if self.party_ledger_status_label is not None:
            self.party_ledger_status_label.configure(text_color=COLORS["warning"])

        def worker():
            result = _fetch_tally_ledgers(
                tally_url,
                timeout=15,
                company_name=selected_company,
            )

            def done():
                self._ledger_fetch_running = False
                if self.manual_fetch_ledger_btn is not None:
                    self.manual_fetch_ledger_btn.configure(state="normal", text="Fetch")
                if self.create_party_fetch_btn is not None:
                    self.create_party_fetch_btn.configure(state="normal", text="Fetch Party Ledgers")

                if result.get("success"):
                    party_ledgers = result.get("party_ledgers") or result.get("ledgers") or []
                    self._set_party_ledger_values(party_ledgers, keep_current=True)
                    self.status_var.set(f"Fetched {len(party_ledgers)} party ledger(s) from Tally")
                    if not silent:
                        self.create_party_status_var.set(f"Party ledgers loaded: {len(party_ledgers)}")
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.party_ledger_status_var.set("Party Ledgers: Fetch failed")
                    if self.party_ledger_status_label is not None:
                        self.party_ledger_status_label.configure(text_color=COLORS["error"])
                    self.status_var.set("Party ledger fetch failed")
                    if not silent:
                        messagebox.showwarning("Party Ledger Fetch Failed", f"Could not fetch ledgers from Tally.\n\n{err}")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _fetch_company_gst_regs_thread(self, silent=False):
        """Fetch company GST registration details from Tally and store them.
        Used to embed CMPGSTIN/CMPGSTSTATE in debit note XML (avoids the
        'GST Registration Details' acceptance popup in TallyPrime).
        """
        try:
            tally_url = self._get_tally_url()
        except ValueError:
            return
        selected_company = self._get_selected_company()

        def worker():
            result = _fetch_company_gst_registrations(
                tally_url,
                company_name=selected_company,
                timeout=15,
            )

            def done():
                if result.get("success"):
                    self.company_gst_registrations = result.get("registrations", [])
                else:
                    self.company_gst_registrations = []

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _build_create_party_tab(self, parent):
        wrapper = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        wrapper.pack(fill="both", expand=True, padx=10, pady=8)

        card = ctk.CTkFrame(
            wrapper,
            fg_color=COLORS["bg_card"],
            border_width=1,
            border_color=COLORS["border"],
            corner_radius=10,
        )
        card.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(
            card,
            text="Create Party Ledger In Tally",
            font=("Segoe UI", 13, "bold"),
            text_color=COLORS["text_primary"],
        ).pack(anchor="w", padx=12, pady=(12, 4))

        ctk.CTkLabel(
            card,
            text="Uses current Host, Port, and selected Target Company.",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        ).pack(anchor="w", padx=12, pady=(0, 8))

        ctk.CTkLabel(card, text="Ledger Name", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(anchor="w", padx=12)
        self.create_party_name_entry = ctk.CTkEntry(
            card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Required party ledger name",
            font=("Segoe UI", 10),
        )
        self.create_party_name_entry.pack(fill="x", padx=12, pady=(4, 8))

        row_parent = ctk.CTkFrame(card, fg_color="transparent")
        row_parent.pack(fill="x", padx=12, pady=(0, 8))
        ctk.CTkLabel(row_parent, text="Parent Group", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        self.create_party_parent_cb = ctk.CTkComboBox(
            row_parent,
            values=["Sundry Debtors", "Sundry Creditors"],
            width=200,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.create_party_parent_cb.set("Sundry Debtors")
        self.create_party_parent_cb.pack(side="right")

        ctk.CTkLabel(card, text="Mailing Name", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(anchor="w", padx=12)
        self.create_party_mailing_entry = ctk.CTkEntry(
            card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Optional mailing name",
            font=("Segoe UI", 10),
        )
        self.create_party_mailing_entry.pack(fill="x", padx=12, pady=(4, 8))

        row_gst = ctk.CTkFrame(card, fg_color="transparent")
        row_gst.pack(fill="x", padx=12, pady=(0, 8))
        self.create_party_gstin_entry = ctk.CTkEntry(
            row_gst,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="GSTIN",
            font=("Segoe UI", 10),
        )
        self.create_party_gstin_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.create_party_pincode_entry = ctk.CTkEntry(
            row_gst,
            width=130,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Pincode",
            font=("Segoe UI", 10),
        )
        self.create_party_pincode_entry.pack(side="left", padx=(4, 0))

        row_geo = ctk.CTkFrame(card, fg_color="transparent")
        row_geo.pack(fill="x", padx=12, pady=(0, 8))
        self.create_party_state_entry = ctk.CTkEntry(
            row_geo,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="State",
            font=("Segoe UI", 10),
        )
        self.create_party_state_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.create_party_country_entry = ctk.CTkEntry(
            row_geo,
            width=130,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 10),
        )
        self.create_party_country_entry.insert(0, "India")
        self.create_party_country_entry.pack(side="left", padx=(4, 0))

        self.create_party_address1_entry = ctk.CTkEntry(
            card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Address line 1",
            font=("Segoe UI", 10),
        )
        self.create_party_address1_entry.pack(fill="x", padx=12, pady=(0, 6))

        self.create_party_address2_entry = ctk.CTkEntry(
            card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Address line 2",
            font=("Segoe UI", 10),
        )
        self.create_party_address2_entry.pack(fill="x", padx=12, pady=(0, 8))

        row_billwise = ctk.CTkFrame(card, fg_color="transparent")
        row_billwise.pack(fill="x", padx=12, pady=(0, 10))
        ctk.CTkLabel(row_billwise, text="Billwise", font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        self.create_party_billwise_cb = ctk.CTkComboBox(
            row_billwise,
            values=["Yes", "No"],
            width=120,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.create_party_billwise_cb.set("Yes")
        self.create_party_billwise_cb.pack(side="right")

        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=12, pady=(0, 12))

        self.create_party_fetch_btn = ctk.CTkButton(
            btn_row,
            text="Fetch Party Ledgers",
            width=160,
            height=34,
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            command=self._fetch_party_ledgers_thread,
        )
        self.create_party_fetch_btn.pack(side="left", padx=(0, 8))

        self.create_party_clear_btn = ctk.CTkButton(
            btn_row,
            text="Clear",
            width=90,
            height=34,
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            command=self._clear_create_party_form,
        )
        self.create_party_clear_btn.pack(side="left", padx=(0, 8))

        self.create_party_create_btn = ctk.CTkButton(
            btn_row,
            text="Create Party Ledger",
            height=34,
            fg_color=COLORS["success"],
            hover_color="#047857",
            text_color="#FFFFFF",
            command=self._create_party_ledger_thread,
        )
        self.create_party_create_btn.pack(side="left", fill="x", expand=True)

        ctk.CTkLabel(
            card,
            textvariable=self.create_party_status_var,
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        ).pack(anchor="w", padx=12, pady=(0, 12))

    def _set_create_party_running_state(self, running: bool):
        self._create_party_running = running
        state = "disabled" if running else "normal"
        for btn in [self.create_party_create_btn, self.create_party_fetch_btn, self.create_party_clear_btn]:
            if btn is not None:
                btn.configure(state=state)

    def _clear_create_party_form(self):
        if self.create_party_name_entry is not None:
            self.create_party_name_entry.delete(0, "end")
        if self.create_party_mailing_entry is not None:
            self.create_party_mailing_entry.delete(0, "end")
        if self.create_party_gstin_entry is not None:
            self.create_party_gstin_entry.delete(0, "end")
        if self.create_party_state_entry is not None:
            self.create_party_state_entry.delete(0, "end")
        if self.create_party_country_entry is not None:
            self.create_party_country_entry.delete(0, "end")
            self.create_party_country_entry.insert(0, "India")
        if self.create_party_pincode_entry is not None:
            self.create_party_pincode_entry.delete(0, "end")
        if self.create_party_address1_entry is not None:
            self.create_party_address1_entry.delete(0, "end")
        if self.create_party_address2_entry is not None:
            self.create_party_address2_entry.delete(0, "end")
        if self.create_party_parent_cb is not None:
            self.create_party_parent_cb.set("Sundry Debtors")
        if self.create_party_billwise_cb is not None:
            self.create_party_billwise_cb.set("Yes")
        self.create_party_status_var.set("Ready to create party ledger")

    def _create_party_ledger_thread(self):
        if self._create_party_running:
            return

        ledger_name = _normalize_ledger_name(self.create_party_name_entry.get() if self.create_party_name_entry else "")
        if not ledger_name:
            messagebox.showwarning("Missing Field", "Ledger Name is required.")
            return

        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return

        selected_company = self._get_selected_company()
        if not selected_company and len(self.fetched_companies) > 1:
            messagebox.showwarning("Select Company", "Please select target company before creating a party ledger.")
            return

        parent_name = self.create_party_parent_cb.get().strip() if self.create_party_parent_cb else "Sundry Debtors"
        mailing_name = self.create_party_mailing_entry.get().strip() if self.create_party_mailing_entry else ""
        gstin = self.create_party_gstin_entry.get().strip() if self.create_party_gstin_entry else ""
        state = self.create_party_state_entry.get().strip() if self.create_party_state_entry else ""
        country = self.create_party_country_entry.get().strip() if self.create_party_country_entry else "India"
        pincode = self.create_party_pincode_entry.get().strip() if self.create_party_pincode_entry else ""
        address1 = self.create_party_address1_entry.get().strip() if self.create_party_address1_entry else ""
        address2 = self.create_party_address2_entry.get().strip() if self.create_party_address2_entry else ""
        billwise_raw = self.create_party_billwise_cb.get().strip().upper() if self.create_party_billwise_cb else "YES"
        billwise_on = billwise_raw in {"YES", "Y", "TRUE", "1"}

        self._set_create_party_running_state(True)
        self.create_party_status_var.set("Creating ledger in Tally...")
        self.status_var.set("Creating party ledger...")

        def worker():
            try:
                result = _create_tally_ledger(
                    tally_url=tally_url,
                    ledger_name=ledger_name,
                    parent_name=parent_name,
                    company_name=selected_company,
                    gstin=gstin,
                    state=state,
                    country=country,
                    pincode=pincode,
                    mailing_name=mailing_name,
                    address1=address1,
                    address2=address2,
                    billwise_on=billwise_on,
                    timeout=30,
                )
            except Exception as exc:
                result = {"success": False, "error": str(exc)}

            def done():
                self._set_create_party_running_state(False)
                if result.get("success"):
                    created = int(result.get("created", 0) or 0)
                    altered = int(result.get("altered", 0) or 0)
                    self.create_party_status_var.set(f"Ledger created/updated. Created={created}, Altered={altered}")
                    self.status_var.set("Party ledger created successfully")
                    if self.manual_form_vars.get("PartyLedger") is not None:
                        self.manual_form_vars["PartyLedger"].set(ledger_name)
                    self._fetch_party_ledgers_thread(silent=True)
                    messagebox.showinfo(
                        "Create Party Ledger",
                        f"Ledger created/updated successfully.\n\nName: {ledger_name}\nParent: {parent_name}",
                    )
                else:
                    err = str(result.get("error") or "Ledger creation failed in Tally.")
                    self.create_party_status_var.set(f"Create failed: {err}")
                    self.status_var.set("Party ledger create failed")
                    messagebox.showerror("Create Party Ledger", err)

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _setup_tree_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            background=_theme_color("table_bg"),
            foreground=_theme_color("text_primary"),
            fieldbackground=_theme_color("table_bg"),
            bordercolor=_theme_color("border"),
            font=("Segoe UI", 10),
            rowheight=26,
        )
        style.map(
            "Treeview",
            background=[("selected", _theme_color("table_selected"))],
            foreground=[("selected", _theme_color("table_selected_fg"))],
        )
        style.configure(
            "Treeview.Heading",
            background=_theme_color("table_header"),
            foreground="#FFFFFF",
            font=("Segoe UI", 10, "bold"),
        )

    def set_theme(self, mode: str):
        normalized = "Dark" if str(mode).lower().startswith("dark") else "Light"
        try:
            ctk.set_appearance_mode(normalized)
        except Exception:
            pass
        self.configure(fg_color=COLORS["bg_dark"])
        self._setup_tree_style()

    def _get_tally_url(self):
        return _build_tally_url(self.tally_host_var.get(), self.tally_port_var.get())

    def _set_voucher_date_mode(self, selected_mode: str):
        mode = str(selected_mode or "excel").strip().lower()
        if mode not in {"current", "excel", "custom"}:
            mode = "excel"

        self.voucher_date_mode_var.set(mode)
        for key, var in self.voucher_date_checks.items():
            var.set(key == mode)

        if self.voucher_custom_date_entry is not None:
            self.voucher_custom_date_entry.configure(
                state="normal" if (mode == "custom" and not self._push_running) else "disabled"
            )

    def _get_voucher_date_selection(self):
        mode = str(self.voucher_date_mode_var.get() or "excel").strip().lower()
        if mode not in {"current", "excel", "custom"}:
            mode = "excel"
            self._set_voucher_date_mode(mode)

        custom_tally_date = ""
        if mode == "custom":
            custom_raw = (self.voucher_custom_date_var.get() or "").strip()
            if not custom_raw:
                raise ValueError("Enter custom date or select Current Date / Excel Date.")
            custom_tally_date = _normalize_manual_date_to_tally(custom_raw)

        return mode, custom_tally_date

    def _view_workflow_demo(self):
        demo_url = (self.workflow_demo_url or "").strip()
        if demo_url:
            try:
                webbrowser.open(demo_url)
                return
            except Exception as exc:
                messagebox.showwarning("View Demo", f"Could not open demo link.\n\n{exc}")
                return

        messagebox.showinfo(
            "View Demo",
            "Demo link is not set yet.\n\nSet self.workflow_demo_url in code to your YouTube link later.",
        )

    def _append_debug_log(self, xml_payload: str, response_text: str, parsed: dict, note: str = ""):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        lines = [
            "=" * 96,
            f"[{timestamp}] note={note}",
            (
                "summary: "
                f"created={parsed.get('created', 0)} "
                f"altered={parsed.get('altered', 0)} "
                f"deleted={parsed.get('deleted', 0)} "
                f"ignored={parsed.get('ignored', 0)} "
                f"errors={parsed.get('errors', 0)} "
                f"exceptions={parsed.get('exceptions', 0)}"
            ),
        ]
        for err in parsed.get("line_errors", []):
            lines.append(f"line_error: {err}")
        lines.append("response:")
        lines.append(response_text[:12000])
        lines.append("xml:")
        lines.append(xml_payload[:12000])
        lines.append("\n")

        with open(self.debug_log_path, "a", encoding="utf-8") as f:
            f.write("\n".join(lines))

    def _populate_tree(self, tree: ttk.Treeview, headers: list, rows: list, limit: int = 300):
        tree.delete(*tree.get_children())
        tree["columns"] = headers
        for h in headers:
            tree.heading(h, text=h)
            tree.column(h, width=max(120, min(260, len(h) * 12)), minwidth=80)

        for idx, row in enumerate(rows[:limit]):
            values = []
            for h in headers:
                value = _row_get(row, h, "")
                values.append("" if value is None else str(value))
            tree.insert("", "end", iid=str(idx), values=values)

    def _set_company_dropdown(self, companies, keep_selection=True):
        current = _normalize_company_name(self.company_var.get()) if keep_selection else ""
        cleaned = []
        seen = set()
        for name in companies or []:
            normalized = _normalize_company_name(name)
            if not _is_valid_company_name(normalized):
                continue
            key = _company_key(normalized)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(normalized)

        cleaned = sorted(cleaned, key=lambda x: _company_key(x))
        values = [self.company_placeholder] + cleaned
        self.company_combo.configure(values=values)

        if current and _company_key(current) in {_company_key(x) for x in cleaned}:
            self.company_combo.set(current)
            self.company_var.set(current)
        else:
            self.company_combo.set(self.company_placeholder)
            self.company_var.set(self.company_placeholder)

        self.fetched_companies = cleaned
        self.company_status_var.set(f"Companies: {len(cleaned)} available")
        self.company_status_label.configure(text_color=COLORS["text_muted"])

    def _get_selected_company(self):
        selected = _normalize_company_name(self.company_var.get())
        if not selected or _company_key(selected) == _company_key(self.company_placeholder):
            if len(self.fetched_companies) == 1:
                return self.fetched_companies[0]
            return ""
        return selected

    def _fetch_tally_companies_thread(self, silent=False):
        if self._company_fetch_running:
            return
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return

        self._company_fetch_running = True
        self.company_refresh_btn.configure(state="disabled", text="Fetching...")
        if not silent:
            self.company_status_var.set("Companies: Fetching...")
            self.company_status_label.configure(text_color=COLORS["warning"])

        def worker():
            result = _fetch_tally_companies(tally_url, timeout=15)

            def done():
                self._company_fetch_running = False
                self.company_refresh_btn.configure(state="normal", text="Refresh")
                if result.get("success"):
                    companies = result.get("companies", [])
                    self._set_company_dropdown(companies, keep_selection=True)
                    self.status_var.set(f"Fetched {len(companies)} company(s) from Tally")
                    self._fetch_party_ledgers_thread(silent=True)
                    # Also fetch company GST registrations (needed for debit note XML).
                    self._fetch_company_gst_regs_thread(silent=True)
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.company_status_var.set("Companies: Fetch failed")
                    self.company_status_label.configure(text_color=COLORS["error"])
                    self.status_var.set("Company fetch failed")
                    if not silent:
                        messagebox.showwarning("Company Fetch Failed", f"Could not fetch companies from Tally.\n\n{err}")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _check_tally_connection_thread(self):
        try:
            tally_url = self._get_tally_url()
        except ValueError as exc:
            messagebox.showerror("Invalid Settings", str(exc))
            return

        self.connection_test_btn.configure(state="disabled", text="Checking...")
        self.connection_status_var.set("Connection: Checking...")
        self.connection_status_label.configure(text_color=COLORS["warning"])

        def worker():
            result = _check_tally_connection(tally_url, timeout=8)

            def done():
                self.connection_test_btn.configure(state="normal", text="Test Connection")
                if result.get("connected"):
                    self.connection_status_var.set(f"Connection: Connected ({tally_url})")
                    self.connection_status_label.configure(text_color=COLORS["success"])
                    self.status_var.set("Connected to Tally")
                    self._fetch_tally_companies_thread(silent=True)
                    self._fetch_party_ledgers_thread(silent=True)
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.connection_status_var.set("Connection: Offline")
                    self.connection_status_label.configure(text_color=COLORS["error"])
                    self.status_var.set(f"Connection failed: {err}")
                    messagebox.showwarning("Connection Failed", f"Could not connect to Tally.\n\n{err}")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _browse_file(self):
        if self._excel_load_running:
            self.status_var.set("Please wait, file is still loading...")
            return

        file_path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm *.xls")])
        if file_path:
            self.file_path_var.set(file_path)
            self._load_preview(file_path)

    def _set_excel_loading_state(self, is_loading: bool):
        self._excel_load_running = is_loading
        state = "disabled" if is_loading else "normal"
        browse_text = "Loading..." if is_loading else "Browse Excel"
        if self.browse_btn is not None:
            self.browse_btn.configure(state=state, text=browse_text)
        if self.save_xml_btn is not None:
            self.save_xml_btn.configure(state=state)
        if self.push_tally_btn is not None:
            self.push_tally_btn.configure(state=state)

    def _set_push_loading_state(self, is_loading: bool, message: str = ""):
        self._push_running = is_loading
        state = "disabled" if is_loading else "normal"

        for btn in [
            self.browse_btn,
            self.connection_test_btn,
            self.company_refresh_btn,
            self.save_xml_btn,
            self.push_tally_btn,
            self.template_btn,
            self.demo_btn,
            self.manual_fetch_ledger_btn,
            self.manual_party_search_clear_btn,
            self.create_party_fetch_btn,
            self.create_party_create_btn,
            self.create_party_clear_btn,
        ]:
            if btn is not None:
                btn.configure(state=state)

        if self.manual_party_search_entry is not None:
            self.manual_party_search_entry.configure(state=state)
        if self.manual_party_ledger_combo is not None:
            self.manual_party_ledger_combo.configure(state="disabled" if is_loading else "readonly")
        if self.note_type_combo is not None:
            self.note_type_combo.configure(state="disabled" if is_loading else "readonly")
        if self.voucher_date_current_cb is not None:
            self.voucher_date_current_cb.configure(state=state)
        if self.voucher_date_excel_cb is not None:
            self.voucher_date_excel_cb.configure(state=state)
        if self.voucher_date_custom_cb is not None:
            self.voucher_date_custom_cb.configure(state=state)
        self._set_voucher_date_mode(self.voucher_date_mode_var.get())

        for btn in self.manual_action_buttons:
            btn.configure(state=state)

        if is_loading:
            self._push_message_var.set(message or "Pushing note vouchers to Tally...")
            if self._push_overlay is None or not self._push_overlay.winfo_exists():
                overlay = ctk.CTkToplevel(self)
                overlay.title("Please Wait")
                overlay.geometry("420x170")
                overlay.transient(self)
                overlay.grab_set()
                overlay.resizable(False, False)

                frame = ctk.CTkFrame(overlay, fg_color=COLORS["bg_card"], corner_radius=10)
                frame.pack(fill="both", expand=True, padx=12, pady=12)

                ctk.CTkLabel(
                    frame,
                    text="Pushing Entries To Tally",
                    font=("Segoe UI", 15, "bold"),
                    text_color=COLORS["text_primary"],
                ).pack(pady=(12, 8))

                ctk.CTkLabel(
                    frame,
                    textvariable=self._push_message_var,
                    font=("Segoe UI", 11),
                    text_color=COLORS["text_muted"],
                    wraplength=360,
                    justify="center",
                ).pack(pady=(0, 10))

                progress = ctk.CTkProgressBar(frame, mode="indeterminate", width=320)
                progress.pack(pady=(0, 10))
                progress.start()

                overlay.protocol("WM_DELETE_WINDOW", lambda: None)
                self._push_overlay = overlay
            else:
                self._push_overlay.deiconify()
                self._push_overlay.lift()
        else:
            if self._push_overlay is not None and self._push_overlay.winfo_exists():
                try:
                    self._push_overlay.grab_release()
                except Exception:
                    pass
                self._push_overlay.destroy()
            self._push_overlay = None
            self._push_message_var.set("")

        self.update_idletasks()

    def _load_preview(self, filepath):
        self._set_excel_loading_state(True)
        if self.excel_info_label is not None:
            self.excel_info_label.configure(text="Loading Excel preview...")
        self.status_var.set(f"Loading: {os.path.basename(filepath)}")

        def worker():
            try:
                headers, rows = read_excel(filepath)
                result = {"ok": True, "headers": headers, "rows": rows}
            except Exception as exc:
                result = {"ok": False, "error": str(exc)}

            def done():
                self._set_excel_loading_state(False)
                if not result.get("ok"):
                    if self.excel_info_label is not None:
                        self.excel_info_label.configure(text="")
                    self.status_var.set("Ready")
                    messagebox.showerror("Error", str(result.get("error", "Unknown error")))
                    return

                self.loaded_headers = result["headers"]
                self.loaded_rows = result["rows"]
                self._populate_tree(self.excel_tree, self.loaded_headers, self.loaded_rows, limit=300)

                if self.excel_info_label is not None:
                    self.excel_info_label.configure(
                        text=f"Rows loaded: {len(self.loaded_rows)} | Headers: {len(self.loaded_headers)}"
                    )
                self.status_var.set(f"Loaded Excel: {os.path.basename(filepath)}")

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

    def _manual_row_from_form(self):
        row = {}
        for header in TEMPLATE_HEADERS:
            row[header] = self.manual_form_vars.get(header, ctk.StringVar(value="")).get().strip()
        return row

    def _validate_manual_row(self, row):
        party = str(row.get("PartyLedger", "")).strip()
        particular = str(row.get("Particular", "")).strip()
        if not party:
            messagebox.showwarning("Missing Field", "Party Ledger is required.")
            return None
        if not particular:
            messagebox.showwarning("Missing Field", "Particular is required.")
            return None

        try:
            taxable = float(str(row.get("TaxableValue") or "0").strip())
        except ValueError:
            messagebox.showwarning("Invalid Value", "Taxable Value must be numeric.")
            return None
        if taxable <= 0:
            messagebox.showwarning("Invalid Value", "Taxable Value must be greater than zero.")
            return None

        if not str(row.get("Date") or "").strip():
            row["Date"] = datetime.today().strftime("%d-%m-%y")

        for key in ["CGSTRate", "SGSTRate", "IGSTRate"]:
            if not str(row.get(key) or "").strip():
                row[key] = "0"

        return row

    def _set_manual_edit_mode(self, index=None):
        self.manual_editing_index = index
        if self.manual_update_btn is not None:
            self.manual_update_btn.configure(state="normal" if index is not None else "disabled")

    def _selected_manual_index(self):
        if self.manual_tree is None:
            return None
        selected = list(self.manual_tree.selection())
        if not selected:
            return None
        try:
            return int(selected[0])
        except ValueError:
            return None

    def _load_manual_row_to_form(self, index: int):
        if index < 0 or index >= len(self.manual_rows):
            return False
        row = self.manual_rows[index]
        for header in TEMPLATE_HEADERS:
            value = _row_get(row, header, "")
            self.manual_form_vars[header].set("" if value is None else str(value))
        return True

    def _on_manual_tree_double_click(self, _event=None):
        self._edit_selected_manual()

    def _edit_selected_manual(self):
        idx = self._selected_manual_index()
        if idx is None:
            messagebox.showinfo("Edit Entry", "Select one row in the table to edit.")
            return
        if not self._load_manual_row_to_form(idx):
            messagebox.showwarning("Edit Entry", "Could not load selected row.")
            return

        self._set_manual_edit_mode(idx)
        self.status_var.set(f"Editing entry #{idx + 1}. Update Entry to save changes.")

    def _update_manual_entry(self):
        idx = self.manual_editing_index
        if idx is None:
            messagebox.showinfo("Update Entry", "Select and edit a row first.")
            return
        if idx < 0 or idx >= len(self.manual_rows):
            self._set_manual_edit_mode(None)
            messagebox.showwarning("Update Entry", "Selected row is no longer available.")
            return

        row = self._manual_row_from_form()
        validated = self._validate_manual_row(row)
        if validated is None:
            return

        if not str(validated.get("VoucherNo") or "").strip():
            validated["VoucherNo"] = str(idx + 1)

        self.manual_rows[idx] = validated
        self._refresh_manual_tree(focus_index=idx)
        self._set_manual_edit_mode(None)
        self.status_var.set(f"Manual entry #{idx + 1} updated.")

    def _clear_manual_form(self):
        keep_date = datetime.today().strftime("%d-%m-%y")
        for key, var in self.manual_form_vars.items():
            if key == "Date":
                var.set(keep_date)
            elif key in {"CGSTRate", "SGSTRate", "IGSTRate"}:
                var.set("0")
            else:
                var.set("")
        self._set_manual_edit_mode(None)
        self._clear_party_ledger_search()

    def _add_manual_entry(self):
        row = self._manual_row_from_form()
        validated = self._validate_manual_row(row)
        if validated is None:
            return

        if not str(validated.get("VoucherNo") or "").strip():
            validated["VoucherNo"] = str(len(self.manual_rows) + 1)

        self.manual_rows.append(validated)
        self._refresh_manual_tree(focus_index=len(self.manual_rows) - 1)
        self._set_manual_edit_mode(None)
        self.status_var.set(f"Manual entry added. Total entries: {len(self.manual_rows)}")

    def _refresh_manual_tree(self, focus_index=None):
        self._populate_tree(self.manual_tree, TEMPLATE_HEADERS, self.manual_rows, limit=500)
        if self.manual_tree is not None and focus_index is not None:
            iid = str(focus_index)
            if iid in self.manual_tree.get_children():
                self.manual_tree.selection_set(iid)
                self.manual_tree.focus(iid)
                self.manual_tree.see(iid)
        if self.manual_info_label is not None:
            self.manual_info_label.configure(text=f"Manual entries: {len(self.manual_rows)}")

    def _remove_selected_manual(self):
        if self.manual_tree is None:
            return
        selected = list(self.manual_tree.selection())
        if not selected:
            messagebox.showinfo("Remove Entry", "Select at least one row to remove.")
            return

        indexes = []
        for iid in selected:
            try:
                indexes.append(int(iid))
            except ValueError:
                continue

        if not indexes:
            return

        for idx in sorted(indexes, reverse=True):
            if 0 <= idx < len(self.manual_rows):
                self.manual_rows.pop(idx)

        self._refresh_manual_tree()
        self._set_manual_edit_mode(None)
        self.status_var.set(f"Selected entries removed. Remaining: {len(self.manual_rows)}")

    def _clear_all_manual(self):
        if not self.manual_rows:
            return
        if not messagebox.askyesno("Clear All", "Remove all manual entries?"):
            return
        self.manual_rows = []
        self._refresh_manual_tree()
        self._set_manual_edit_mode(None)
        self.status_var.set("All manual entries cleared.")

    def _download_template(self):
        note_type = self._get_note_type()
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"Template_{note_type} Voucher.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not out:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
                ws.cell(row=1, column=col_idx, value=header)

            for col_idx, value in enumerate(TEMPLATE_SAMPLE_ROW, 1):
                ws.cell(row=2, column=col_idx, value=value)

            for col in range(1, len(TEMPLATE_HEADERS) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

            wb.save(out)
            wb.close()
            self.status_var.set(f"{note_type} template saved: {os.path.basename(out)}")
            messagebox.showinfo("Template Saved", f"{note_type} template saved to:\n{out}")
        except Exception as exc:
            messagebox.showerror("Template Error", str(exc))

    def _get_note_type(self):
        return _normalize_note_type(self.note_type_var.get() if self.note_type_var is not None else "Credit Note")

    def _active_rows(self):
        active_tab = self.source_tabs.get()
        if active_tab == "Manual Entry":
            return self.manual_rows, "Manual Entry"
        return self.loaded_rows, "Excel Upload"

    def _generate(self, action):
        if self._push_running:
            self.status_var.set("Push already in progress...")
            return

        note_type = self._get_note_type()
        note_file_stem = note_type.replace(" ", "")

        rows, source_label = self._active_rows()
        if not rows:
            messagebox.showwarning("No Data", f"No rows available in {source_label}.")
            return

        company = self._get_selected_company()
        if action == "push" and not company and len(self.fetched_companies) > 1:
            messagebox.showwarning("Select Company", "Please select target company before pushing.")
            return

        try:
            date_mode, custom_tally_date = self._get_voucher_date_selection()

            # For Debit Notes, company GST registration context (CMPGSTIN, CMPGSTSTATE)
            # must be embedded in the XML so Tally does not show the "GST Registration
            # Details of the Company are invalid or not specified" uncertainty.
            # The async fetch may not have completed yet, so we do a synchronous fetch
            # here if registrations are not already cached.
            _cmp_gst_regs = list(self.company_gst_registrations or [])
            if _normalize_note_type(note_type) == "Debit Note" and not _cmp_gst_regs:
                try:
                    _tally_url = self._get_tally_url()
                    _gst_fetch = _fetch_company_gst_registrations(
                        _tally_url, company_name=company, timeout=10
                    )
                    if _gst_fetch.get("success"):
                        _cmp_gst_regs = _gst_fetch.get("registrations", [])
                        self.company_gst_registrations = _cmp_gst_regs
                except Exception:
                    pass  # Proceed without; debit note will still import, may be uncertain

            xml_payload, voucher_count = generate_note_xml(
                rows,
                company=company,
                date_mode=date_mode,
                custom_tally_date=custom_tally_date,
                voucher_type=note_type,
                company_gst_registrations=_cmp_gst_regs,
            )
            if voucher_count <= 0:
                messagebox.showwarning("No Vouchers", "No valid rows found (TaxableValue must be greater than zero).")
                return

            if action == "save":
                out = filedialog.asksaveasfilename(
                    defaultextension=".xml",
                    initialfile=f"{note_file_stem}.xml",
                    filetypes=[("XML", "*.xml")],
                )
                if not out:
                    return
                with open(out, "w", encoding="utf-8") as f:
                    f.write(xml_payload)
                self.status_var.set(f"{note_type} XML saved: {os.path.basename(out)} ({voucher_count} voucher(s))")
                messagebox.showinfo("Saved", f"{note_type} XML saved successfully.\n{out}")
                return

            host = (self.tally_host_var.get() or "localhost").strip()
            port_text = (self.tally_port_var.get() or "9000").strip()
            if not port_text.isdigit():
                raise ValueError("Port must be numeric.")
            port = int(port_text)

            self._set_push_loading_state(True, f"Pushing {voucher_count} {note_type} voucher(s) from {source_label}...")
            self.status_var.set("Pushing to Tally...")

            def worker():
                try:
                    response_text = push_to_tally(xml_payload, host=host, port=port, timeout=PUSH_REQUEST_TIMEOUT_SEC)
                    parsed = _parse_tally_response_details(response_text)
                    self._append_debug_log(xml_payload, response_text, parsed, note=f"source={source_label}; type={note_type}")
                    result = {"ok": True, "parsed": parsed}
                except Exception as exc:
                    result = {"ok": False, "error": str(exc)}

                def done():
                    self._set_push_loading_state(False)
                    if not result.get("ok"):
                        err = str(result.get("error", "Unknown error"))
                        self.status_var.set(f"Push failed: {err}")
                        messagebox.showerror("Push Failed", err)
                        return

                    parsed = result["parsed"]
                    created = parsed.get("created", 0)
                    altered = parsed.get("altered", 0)
                    errors = parsed.get("errors", 0)
                    exceptions = parsed.get("exceptions", 0)
                    line_errors = parsed.get("line_errors", [])

                    summary = (
                        f"Created: {created}\n"
                        f"Altered: {altered}\n"
                        f"Errors: {errors}\n"
                        f"Exceptions: {exceptions}"
                    )

                    if parsed.get("success"):
                        self.status_var.set(f"{note_type} push successful: Created {created}, Altered {altered}")
                        messagebox.showinfo("Push Successful", summary)
                    else:
                        if line_errors:
                            summary += "\n\nLine Errors:\n- " + "\n- ".join(line_errors[:8])
                        self.status_var.set(f"{note_type} push completed with errors. See popup/debug log.")
                        messagebox.showwarning("Push Completed With Errors", summary)

                self.after(0, done)

            threading.Thread(target=worker, daemon=True).start()

        except ValueError as exc:
            messagebox.showerror("Validation Error", str(exc))
        except Exception as exc:
            messagebox.showerror("Error", str(exc))


if __name__ == "__main__":
    app = TallyNoteEntryApp()
    app.mainloop()