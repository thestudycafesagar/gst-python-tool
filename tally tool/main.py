#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════╗
║     GSTR-2B + Tally Sheet → Tally Converter v4.1         ║
║     Universal - Works with ALL GSTR-2B Formats           ║
║     + Tally Sheet → XML  + Party Ledger Mapping          ║
║     By: Studycafe Digital Solutions                      ║
╚══════════════════════════════════════════════════════════╝

Mode 1: GSTR-2B Excel → Tally-Ready Excel + XML (with optional mapping)
Mode 2: Tally Sheet → Tally XML (direct conversion)
"""

import os
import sys
import threading
import datetime
import time
import random
import glob
import re
import html
import webbrowser
import xml.etree.ElementTree as ET
from xml.dom import minidom
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


def _silent_pip_install(package_spec: str):
    os.system(
        f"{sys.executable} -m pip install {package_spec} "
        "--break-system-packages --disable-pip-version-check --no-warn-script-location -q"
    )

try:
    import customtkinter as ctk
except ImportError:
    _silent_pip_install("customtkinter")
    import customtkinter as ctk

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation as OpenpyxlDataValidation
except ImportError:
    _silent_pip_install("openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation as OpenpyxlDataValidation

try:
    import pandas as pd
except ImportError:
    _silent_pip_install("pandas")
    import pandas as pd

try:
    from PIL import Image
except ImportError:
    _silent_pip_install("pillow")
    from PIL import Image

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    _silent_pip_install("selenium webdriver-manager")
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

try:
    import requests
except ImportError:
    _silent_pip_install("requests")
    import requests

try:
    import sys as _sys_spe, os as _os_spe
    _spe_dir = _os_spe.path.dirname(_os_spe.path.abspath(__file__))
    if _spe_dir not in _sys_spe.path:
        _sys_spe.path.insert(0, _spe_dir)
    from sale_purchase_entry import generate_purchase_item_xml as _spe_gen_purchase_item_xml
    from sale_purchase_entry import _fetch_company_gst_registrations as _spe_fetch_gst_regs
    _SPE_AVAILABLE = True
except Exception:
    _spe_gen_purchase_item_xml = None
    _spe_fetch_gst_regs = None
    _SPE_AVAILABLE = False


# ─── THEME & PALETTE ─────────────────────────────────────
COLORS = {
    "bg_dark":       ("#F0F4F8", "#0F172A"),
    "bg_card":       ("#FFFFFF", "#1E293B"),
    "bg_card_hover": ("#E2E8F0", "#334155"),
    "bg_input":      ("#F1F5F9", "#1E293B"),
    "accent":        ("#2563EB", "#3B82F6"),
    "accent_hover":  ("#1D4ED8", "#2563EB"),
    "accent_glow":   ("#3B82F6", "#60A5FA"),
    "success":       ("#059669", "#10B981"),
    "success_bg":    ("#D1FAE5", "#064E3B"),
    "warning":       ("#D97706", "#F59E0B"),
    "warning_bg":    ("#FEF3C7", "#78350F"),
    "error":         ("#DC2626", "#EF4444"),
    "error_bg":      ("#FEE2E2", "#7F1D1D"),
    "text_primary":  ("#0F172A", "#F1F5F9"),
    "text_secondary":("#475569", "#CBD5E1"),
    "text_muted":    ("#64748B", "#94A3B8"),
    "border":        ("#E2E8F0", "#334155"),
    "border_active": ("#2563EB", "#3B82F6"),
    "tally_gold":    ("#F59E0B", "#FBBF24"),
    "tally_dark":    ("#F8FAFC", "#0B1220"),
    "gradient_1":    ("#2563EB", "#1D4ED8"),
    "gradient_2":    ("#4F46E5", "#3730A3"),
    "gradient_3":    ("#7C3AED", "#6D28D9"),
    "table_header":  ("#1E293B", "#0F172A"),
    "table_row_odd": ("#FFFFFF", "#1E293B"),
    "table_row_even":("#F8FAFC", "#0F172A"),
    "table_border":  ("#E2E8F0", "#334155"),
    "xml_accent":    ("#7C3AED", "#8B5CF6"),
    "xml_accent_h":  ("#6D28D9", "#7C3AED"),
}


def _theme_color(name_or_value):
    value = COLORS.get(name_or_value, name_or_value)
    if isinstance(value, tuple):
        mode = ctk.get_appearance_mode().lower()
        return value[1] if mode == "dark" else value[0]
    return value


# ═══════════════════════════════════════════════════════════
#  DYNAMIC COLUMN MAPPER — works with ANY GSTR-2B layout
# ═══════════════════════════════════════════════════════════

class B2BColumnMapper:
    HEADER_PATTERNS = {
        "gstin":         ["gstin of supplier"],
        "trade_name":    ["trade/legal name", "trade name", "legal name"],
        "invoice_no":    ["invoice number", "note number"],
        "invoice_type":  ["invoice type", "note type"],
        "invoice_date":  ["invoice date", "note date"],
        "invoice_value": ["invoice value", "note value"],
        "place_of_supply": ["place of supply"],
        "reverse_charge": ["reverse charge", "supply attract reverse charge"],
        "rate":          ["rate(%)","rate (%)"],
        "taxable_value": ["taxable value"],
        "igst":          ["integrated tax"],
        "cgst":          ["central tax"],
        "sgst":          ["state/ut tax", "state tax", "ut tax"],
        "cess":          ["cess"],
        "filing_period": ["gstr-1/iff/gstr-5 period", "gstr-1/iff period", "filing period", "gstr-1/1a/iff period", "gstr-1/iff/gstr-1a period", "gstr-1/iff/1a period"],
        "filing_date":   ["gstr-1/iff/gstr-5 filing date", "filing date", "gstr-1/1a/iff filing date", "gstr-1/iff/gstr-1a filing date", "gstr-1/iff/1a filing date"],
        "itc_avail":     ["itc availability", "itc avail"],
        "reason":        ["reason"],
        "applicable_pct":["applicable %", "applicable percent"],
    }

    def __init__(self):
        self.column_map = {}
        self.data_start_row = 7
        self.header_row_1 = 5
        self.header_row_2 = 6

    def detect_columns(self, ws) -> dict:
        self.column_map = {}
        self._find_header_rows(ws)
        headers_r1 = {}
        headers_r2 = {}
        for cell in ws[self.header_row_1]:
            if cell.value:
                headers_r1[cell.column - 1] = str(cell.value).strip().lower()
        for cell in ws[self.header_row_2]:
            if cell.value:
                headers_r2[cell.column - 1] = str(cell.value).strip().lower()
        all_headers = {}
        for col_idx in range(ws.max_column):
            parts = []
            if col_idx in headers_r1:
                parts.append(headers_r1[col_idx])
            if col_idx in headers_r2:
                parts.append(headers_r2[col_idx])
            if parts:
                all_headers[col_idx] = " | ".join(parts)
        for field, patterns in self.HEADER_PATTERNS.items():
            for col_idx, header_text in all_headers.items():
                for pattern in patterns:
                    if pattern in header_text:
                        if field not in self.column_map:
                            self.column_map[field] = col_idx
                        break

        # ─── B2BA amendment sheet: positionally assign original vs revised columns ───
        # B2BA always has TWO "Invoice Number" columns: col[0]=Original, col[1]=Revised.
        # Detection uses three independent signals (any one is sufficient):
        #   1. "original invoice" in a combined header  (standard GSTR-2B merged header)
        #   2. "revised invoice"  in a combined header  (alternate wording)
        #   3. Two or more invoice-number columns exist (structural: B2B has exactly 1)
        _all_inv_no_cols = sorted(
            col for col, h in all_headers.items()
            if "invoice number" in h or "note number" in h
        )
        is_b2ba = (
            any("original invoice" in h or "revised invoice" in h for h in all_headers.values())
            or len(_all_inv_no_cols) >= 2
        )
        if is_b2ba:
            inv_no_cols   = _all_inv_no_cols
            inv_date_cols = sorted(
                col for col, h in all_headers.items()
                if "invoice date" in h or "note date" in h
            )
            if len(inv_no_cols) >= 2:
                self.column_map["orig_invoice_no"] = inv_no_cols[0]   # original (for matching)
                self.column_map["invoice_no"]      = inv_no_cols[1]   # revised  (for Tally entry)
            elif inv_no_cols:
                self.column_map["orig_invoice_no"] = inv_no_cols[0]
            if len(inv_date_cols) >= 2:
                self.column_map["orig_invoice_date"] = inv_date_cols[0]
                self.column_map["invoice_date"]      = inv_date_cols[1]  # revised date
            elif inv_date_cols:
                self.column_map["orig_invoice_date"] = inv_date_cols[0]

        gstin_col = self.column_map.get("gstin", 0)
        self.data_start_row = self.header_row_2 + 1
        for row_idx in range(self.header_row_2 + 1, min(self.header_row_2 + 10, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=gstin_col + 1).value
            if val and isinstance(val, str) and len(val.strip()) >= 15:
                self.data_start_row = row_idx
                break
        return self.column_map

    def _find_header_rows(self, ws):
        for row_idx in range(1, min(15, ws.max_row + 1)):
            for cell in ws[row_idx]:
                if cell.value and "gstin of supplier" in str(cell.value).strip().lower():
                    self.header_row_1 = row_idx
                    self.header_row_2 = row_idx + 1
                    return
        self.header_row_1 = 5
        self.header_row_2 = 6

    def get(self, field, default=None):
        return self.column_map.get(field, default)

    def has(self, field) -> bool:
        return field in self.column_map

    def summary(self) -> str:
        lines = [f"Header rows: {self.header_row_1}-{self.header_row_2}, Data starts: row {self.data_start_row}"]
        for field, col_idx in sorted(self.column_map.items(), key=lambda x: x[1]):
            lines.append(f"  Col {col_idx}: {field}")
        return "\n".join(lines)


# ═══════════════════════════════════════════════════════════
#  CORE ENGINE
# ═══════════════════════════════════════════════════════════

class GSTR2BEngine:
    TALLY_HEADERS = [
        "Date", "VoucherNo", "PartyLedger", "PartyName", "PartyGSTIN", "PartyState", "PlaceOfSupply",
        "PurchaseLedger", "TaxableValue", "CGSTLedger", "CGSTRate", "SGSTLedger",
        "SGSTRate", "IGSTLedger", "IGSTRate", "Narration",
        "TDSLedger", "TDSRate", "TDSAmount", "SupplierInvoiceNo", "SupplierInvoiceDate",
        "PartyMailingName", "PartyAddress1", "PartyAddress2", "PartyPincode"
    ]

    def __init__(self):
        self.records = []
        self.company_gstin = ""
        self.company_name = ""
        self.trade_name = ""
        self.financial_year = ""
        self.tax_period = ""
        self.errors = []
        self.warnings = []
        self.stats = self._empty_stats()
        self.party_ledger_map = {}  # party_name → purchase_ledger
        self.party_tds_ledger_map = {}  # party_name → tds_ledger
        self.party_tds_rate_map = {}  # party_name → tds_rate
        self.allowed_tax_rates = (0.0, 5.0, 12.0, 18.0, 28.0, 40.0)
        self.tax_rate_tolerance = 0.30
        # RCM ledger mapping set by UI before generation
        self.rcm_ledger_map = {}  # keys: expense, cgst_inward, sgst_inward, igst_inward, cgst_outward, sgst_outward, igst_outward
        self.itc_map = {}  # invoice_no.upper() → {itc_claimed, has_stock}
        # Company GST registration details fetched from Tally
        # registration_name must match the exact name in Tally's company master
        # (e.g. "Delhi Registration") to prevent vouchers going to Uncertain in GSTR-3B.
        self.company_registration_name = ""   # e.g. "Delhi Registration"
        self.company_registration_state = ""  # e.g. "Delhi"

    def _empty_stats(self):
        return {
            "total_records": 0, "igst_count": 0, "cgst_sgst_count": 0,
            "total_taxable": 0.0, "total_igst": 0.0,
            "total_cgst": 0.0, "total_sgst": 0.0, "total_cess": 0.0,
        }

    # ─── MAPPING SHEET ───

    def load_mapping(self, filepath: str) -> int:
        """Load party→ledger mapping from Excel. Returns count of mappings loaded."""
        self.party_ledger_map = {}
        self.party_tds_ledger_map = {}
        self.party_tds_rate_map = {}
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            # Find header row — look for PartyLedger/PurchaseLedger columns
            party_col = None
            ledger_col = None
            tds_ledger_col = None
            tds_rate_col = None
            header_row = 1

            for row_idx in range(1, min(5, ws.max_row + 1)):
                for cell in ws[row_idx]:
                    val = str(cell.value or "").strip().lower()
                    if val in ("partyledger", "party ledger", "party", "party name", "trade_name", "trade name", "supplier"):
                        party_col = cell.column - 1
                        header_row = row_idx
                    elif val in ("purchaseledger", "purchase ledger", "purchase", "ledger", "ledger name", "expense ledger"):
                        ledger_col = cell.column - 1
                        header_row = row_idx
                    elif val in ("tdsledger", "tds ledger", "tds_ledger"):
                        tds_ledger_col = cell.column - 1
                        header_row = row_idx
                    elif val in ("tdsrate", "tds rate", "tds_rate"):
                        tds_rate_col = cell.column - 1
                        header_row = row_idx

            if party_col is None or ledger_col is None:
                # Try columns A=party, B=ledger as default
                party_col = 0
                ledger_col = 1
                header_row = 0  # no header, start from row 1

            for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
                vals = [c.value for c in row]
                party = str(vals[party_col] or "").strip() if party_col < len(vals) else ""
                ledger = str(vals[ledger_col] or "").strip() if ledger_col < len(vals) else ""
                if party and ledger:
                    key = party.upper()
                    self.party_ledger_map[key] = ledger

                    if tds_ledger_col is not None and tds_ledger_col < len(vals):
                        tds_ledger = str(vals[tds_ledger_col] or "").strip()
                        if tds_ledger:
                            self.party_tds_ledger_map[key] = tds_ledger

                    if tds_rate_col is not None and tds_rate_col < len(vals):
                        rate_val = vals[tds_rate_col]
                        if rate_val not in (None, ""):
                            try:
                                self.party_tds_rate_map[key] = float(rate_val)
                            except (ValueError, TypeError):
                                pass

            wb.close()
            return len(self.party_ledger_map)
        except Exception as e:
            self.errors.append(f"Mapping load failed: {str(e)}")
            return 0

    def get_purchase_ledger(self, party_name: str, default: str = "Purchase Account") -> str:
        """Look up purchase ledger for a party. Falls back to default."""
        if not self.party_ledger_map:
            return default
        return self.party_ledger_map.get(party_name.upper().strip(), default)

    def get_tds_ledger(self, party_name: str, default: str = "") -> str:
        """Look up TDS ledger for a party."""
        key = party_name.upper().strip()
        return self.party_tds_ledger_map.get(key, default)

    def get_tds_rate(self, party_name: str, default= ""):
        """Look up TDS rate for a party."""
        key = party_name.upper().strip()
        return self.party_tds_rate_map.get(key, default)

    def _today_str(self) -> str:
        return datetime.date.today().strftime("%d/%m/%Y")

    def _normalize_date_str(self, value) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, (int, float)):
            try:
                dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(value))
                return dt.strftime("%d/%m/%Y")
            except Exception:
                pass
        text = str(value).strip()
        if not text:
            return ""
        if text.isdigit() and len(text) == 8:
            try:
                # Supports compact formats like 20260330 and 30032026.
                if text[:4].isdigit() and 1900 <= int(text[:4]) <= 2100:
                    return datetime.datetime.strptime(text, "%Y%m%d").strftime("%d/%m/%Y")
                return datetime.datetime.strptime(text, "%d%m%Y").strftime("%d/%m/%Y")
            except ValueError:
                pass
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d", "%d-%b-%Y", "%d-%B-%Y", "%d/%b/%Y", "%d/%B/%Y"):
            try:
                return datetime.datetime.strptime(text, fmt).strftime("%d/%m/%Y")
            except ValueError:
                pass
        try:
            dt = datetime.datetime.fromisoformat(text)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            return text

    def _normalize_state_name(self, value) -> str:
        if value in (None, ""):
            return ""
        raw = str(value).strip()
        if not raw:
            return ""
        raw = raw.replace("_", " ").replace("-", " ").strip()
        if raw[:2].isdigit():
            raw = raw[2:].strip(" -")
        parts = [p for p in raw.split() if p]
        if not parts:
            return ""
        return " ".join(word.capitalize() if word.lower() not in {"and", "of", "ut"} else word.upper() for word in parts)

    def _state_from_gstin(self, gstin: str) -> str:
        gstin = str(gstin or "").strip().upper()
        state_map = {
            "01": "Jammu And Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
            "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
            "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
            "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
            "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
            "24": "Gujarat", "25": "Daman And Diu", "26": "Dadra And Nagar Haveli And Daman And Diu",
            "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
            "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman And Nicobar Islands", "36": "Telangana",
            "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory", "99": "Centre Jurisdiction"
        }
        return state_map.get(gstin[:2], "")

    def _tally_date(self, value, fallback_today=True) -> str:
        date_text = self._normalize_date_str(value)
        if not date_text:
            if fallback_today:
                return datetime.date.today().strftime("%Y%m%d")
            return ""
        parts = date_text.split("/")
        if len(parts) == 3:
            yy, mm, dd = parts[2], parts[1], parts[0]
            candidate = f"{yy}{mm.zfill(2)}{dd.zfill(2)}"
            if candidate.isdigit() and len(candidate) == 8:
                return candidate
        compact = date_text.replace("/", "").replace("-", "")
        if compact.isdigit() and len(compact) == 8:
            return compact
        if fallback_today:
            return datetime.date.today().strftime("%Y%m%d")
        return ""

    def _add_common_ledger_flags(self, node, is_party="No"):
        ET.SubElement(node, "GSTCLASS").text = "Not Applicable"
        ET.SubElement(node, "ISDEEMEDPOSITIVE").text = "No" if is_party == "Yes" else "Yes"
        ET.SubElement(node, "LEDGERFROMITEM").text = "No"
        ET.SubElement(node, "REMOVEZEROENTRIES").text = "No"
        ET.SubElement(node, "ISPARTYLEDGER").text = is_party
        ET.SubElement(node, "GSTOVERRIDDEN").text = "No"
        ET.SubElement(node, "ISGSTASSESSABLEVALUEOVERRIDDEN").text = "No"

    # ─── GSTR-2B PARSING ───

    def parse_gstr2b(self, filepath, progress_callback=None) -> bool:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "Read me" in wb.sheetnames:
                self._parse_readme(wb["Read me"])

            all_sheet_names_upper = {s.upper(): s for s in wb.sheetnames}
            b2b_sheets   = [all_sheet_names_upper[k] for k in ("B2B",)      if k in all_sheet_names_upper]
            cdnr_sheets  = [all_sheet_names_upper[k] for k in ("B2B-CDNR",) if k in all_sheet_names_upper]
            b2ba_sheets  = [all_sheet_names_upper[k] for k in ("B2BA",)     if k in all_sheet_names_upper]

            if not b2b_sheets and not cdnr_sheets and not b2ba_sheets:
                self.errors.append("No B2B, B2B-CDNR, or B2BA sheets found in the uploaded file!")
                return False

            self.records = []
            self.errors = []
            self.warnings = []

            # ── Step 1: Parse B2BA amendments first ──────────────────────────────
            # Collect the set of original invoice numbers that have been amended so
            # that the corresponding B2B rows can be skipped later.
            b2ba_records = []
            b2ba_orig_nos = set()   # normalised original invoice numbers

            for sheet_name in b2ba_sheets:
                ws = wb[sheet_name]
                mapper = B2BColumnMapper()
                mapper.detect_columns(ws)
                if not mapper.has("gstin"):
                    self.warnings.append(f"Could not detect columns in {sheet_name} sheet — skipped.")
                    continue
                for row_idx in range(mapper.data_start_row, ws.max_row + 1):
                    row = [cell.value for cell in ws[row_idx]]
                    gstin_col = mapper.get("gstin", 0)
                    if gstin_col >= len(row) or not row[gstin_col]:
                        continue
                    try:
                        record = self._parse_b2b_row(row, row_idx, mapper)
                        if record:
                            record["is_amendment"] = True
                            record["sheet_type"] = "B2BA"
                            b2ba_records.append(record)
                            orig_no = record.get("orig_invoice_no", "").strip().upper()
                            inv_no  = record.get("invoice_no",      "").strip().upper()
                            # Primary: use the dedicated "original invoice" column when detected.
                            # Fallback: use invoice_no for GSTR-2B formats that have no separate
                            # original column (both B2BA and B2B then carry the same invoice number).
                            if orig_no:
                                b2ba_orig_nos.add(orig_no)
                            elif inv_no:
                                b2ba_orig_nos.add(inv_no)
                    except Exception as e:
                        self.errors.append(f"{sheet_name} Row {row_idx}: {str(e)}")

            if b2ba_sheets and not b2ba_records:
                self.warnings.append("B2BA sheet found but no valid amendment records could be parsed.")
            elif b2ba_records:
                self.warnings.append(
                    f"B2BA: {len(b2ba_records)} amendment record(s) found. "
                    f"{len(b2ba_orig_nos)} original invoice(s) will be replaced."
                )

            # ── Step 2: Parse B2B / B2B-CDNR, skipping amended invoices ─────────
            sheets_to_process = b2b_sheets + cdnr_sheets
            if sheets_to_process:
                total_rows = sum(wb[s].max_row for s in sheets_to_process)
                rows_processed = 0

                for sheet_name in sheets_to_process:
                    _is_cdnr = sheet_name in cdnr_sheets
                    ws = wb[sheet_name]
                    sheet_mapper = B2BColumnMapper()
                    col_map = sheet_mapper.detect_columns(ws)
                    # Store the B2B mapper for UI diagnostics (only the first B2B sheet)
                    if not _is_cdnr and not hasattr(self, "mapper"):
                        self.mapper = sheet_mapper

                    if not sheet_mapper.has("gstin"):
                        self.warnings.append(f"Could not detect 'GSTIN of supplier' column in {sheet_name} sheet!")
                        rows_processed += ws.max_row
                        continue

                    missing_optional = [f for f in ["rate"] if f not in col_map]
                    if missing_optional:
                        self.warnings.append(
                            f"{sheet_name}: Optional columns not found: {', '.join(missing_optional)} — will auto-calculate"
                        )

                    skipped_amendments = 0
                    data_start = sheet_mapper.data_start_row
                    _last_gstin      = ""
                    _last_trade_name = ""
                    _trade_col = sheet_mapper.get("trade_name", 1)
                    for row_idx in range(data_start, ws.max_row + 1):
                        row = [cell.value for cell in ws[row_idx]]
                        gstin_col = sheet_mapper.get("gstin", 0)
                        # GSTR-2B Excel uses merged/blank cells for GSTIN when a supplier
                        # has multiple invoices — forward-fill from the last seen value.
                        cur_gstin = str(row[gstin_col] or "").strip() if gstin_col < len(row) else ""
                        if len(cur_gstin) >= 15:
                            _last_gstin = cur_gstin
                            _last_trade_name = str(row[_trade_col] or "").strip() if _trade_col < len(row) else ""
                        elif _last_gstin:
                            row = list(row)
                            while len(row) <= max(gstin_col, _trade_col):
                                row.append(None)
                            row[gstin_col] = _last_gstin
                            if _trade_col < len(row) and not str(row[_trade_col] or "").strip():
                                row[_trade_col] = _last_trade_name
                        else:
                            rows_processed += 1
                            continue
                        try:
                            record = self._parse_b2b_row(row, row_idx, sheet_mapper)
                            if record:
                                record["sheet_type"] = "CDNR" if _is_cdnr else "B2B"
                                inv_key = record.get("invoice_no", "").strip().upper()
                                if inv_key in b2ba_orig_nos:
                                    # Superseded by a B2BA amendment — skip
                                    skipped_amendments += 1
                                else:
                                    self.records.append(record)
                        except Exception as e:
                            self.errors.append(f"{sheet_name} Row {row_idx}: {str(e)}")

                        rows_processed += 1
                        if progress_callback and rows_processed % 20 == 0:
                            pct = rows_processed / max(1, total_rows)
                            progress_callback(min(pct, 1.0), f"Parsing {sheet_name} row {row_idx}/{ws.max_row}...")

                    if skipped_amendments > 0:
                        self.warnings.append(
                            f"{sheet_name}: {skipped_amendments} invoice(s) skipped — superseded by B2BA amendment(s)."
                        )

            # ── Step 3: Append B2BA amended records ──────────────────────────────
            self.records.extend(b2ba_records)

            self._compute_stats()
            wb.close()
            return len(self.records) > 0
        except Exception as e:
            self.errors.append(f"Failed to open file: {str(e)}")
            import traceback
            self.errors.append(traceback.format_exc())
            return False

    # ─── TALLY SHEET PARSING (for → XML conversion) ───

    def parse_tally_sheet(self, filepath, progress_callback=None) -> bool:
        """Parse an existing tally-ready Excel sheet into records for XML generation."""
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
            ws = wb.active
            self.records = []
            self.errors = []
            self.warnings = []

            headers = [str(cell.value or "").strip() for cell in ws[1]]
            header_map = {h: idx for idx, h in enumerate(headers)}

            expected_any = [("Date", "VoucherDate"), ("PartyLedger",), ("PurchaseLedger",), ("TaxableValue",)]
            for options in expected_any:
                if not any(opt in header_map for opt in options):
                    self.errors.append(f"Invalid tally sheet! Missing one of: {', '.join(options)}")
                    self.errors.append(f"Found headers: {', '.join(headers[:12])}")
                    return False

            def get_val(row, *names, default=None):
                for name in names:
                    idx = header_map.get(name)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        if val not in (None, ""):
                            return val
                return default

            total_rows = ws.max_row
            for row_idx in range(2, total_rows + 1):
                row = [cell.value for cell in ws[row_idx]]
                if not get_val(row, "PartyLedger", "PartyName") and not get_val(row, "TaxableValue", default=0):
                    continue

                voucher_date = self._normalize_date_str(get_val(row, "Date", "VoucherDate", default=""))
                inv_date = self._normalize_date_str(get_val(row, "SupplierInvoiceDate", default=voucher_date))
                taxable = float(get_val(row, "TaxableValue", default=0) or 0)

                cgst_rate = float(get_val(row, "CGSTRate", default=0) or 0)
                sgst_rate = float(get_val(row, "SGSTRate", default=0) or 0)
                igst_ledger = get_val(row, "IGSTLedger", default="")
                igst_rate = float(get_val(row, "IGSTRate", default=0) or 0)

                is_igst = bool(str(igst_ledger or "").strip() and float(igst_rate or 0) > 0)

                if is_igst:
                    igst_amt = round(taxable * igst_rate / 100, 2)
                    cgst_amt = 0.0
                    sgst_amt = 0.0
                    rate = igst_rate
                else:
                    igst_amt = 0.0
                    cgst_amt = round(taxable * cgst_rate / 100, 2)
                    sgst_amt = round(taxable * sgst_rate / 100, 2)
                    rate = cgst_rate + sgst_rate

                party_name = str(get_val(row, "PartyLedger", "PartyName", default="") or "").strip()
                party_gstin = str(get_val(row, "PartyGSTIN", default="") or "").strip().upper()
                party_state = self._normalize_state_name(get_val(row, "PartyState", default="")) or self._state_from_gstin(party_gstin)
                place_of_supply = self._normalize_state_name(get_val(row, "PlaceOfSupply", default=""))
                party_mailing_name = str(get_val(row, "PartyMailingName", "PartyName", "PartyLedger", default=party_name) or "").strip()
                party_address1 = str(get_val(row, "PartyAddress1", default="") or "").strip()
                party_address2 = str(get_val(row, "PartyAddress2", default="") or "").strip()
                party_pincode = str(get_val(row, "PartyPincode", default="") or "").strip()

                tds_rate_val = get_val(row, "TDSRate", default="")
                if tds_rate_val in (None, ""):
                    parsed_tds_rate = ""
                else:
                    try:
                        parsed_tds_rate = float(tds_rate_val)
                    except (ValueError, TypeError):
                        parsed_tds_rate = ""

                tds_amount_val = get_val(row, "TDSAmount", default="")
                if tds_amount_val in (None, ""):
                    parsed_tds_amount = ""
                else:
                    try:
                        # Treat negative input as a deduction amount, but store as positive value.
                        parsed_tds_amount = abs(float(tds_amount_val))
                    except (ValueError, TypeError):
                        parsed_tds_amount = ""

                itc_avail_val = str(
                    get_val(row, "ITCAvailability", "ITC Availability", "ITCAvail", default="Yes") or "Yes"
                ).strip()
                if not itc_avail_val:
                    itc_avail_val = "Yes"
                    
                invoice_type = "Credit Note" if taxable < 0 else "Regular"

                record = {
                    "voucher_date": voucher_date,
                    "voucher_no": str(get_val(row, "VoucherNo", default="") or "").strip(),
                    "gstin": party_gstin,
                    "trade_name": party_name,
                    "party_name": str(get_val(row, "PartyName", "PartyLedger", default=party_name) or "").strip(),
                    "party_mailing_name": party_mailing_name,
                    "party_address1": party_address1,
                    "party_address2": party_address2,
                    "party_pincode": party_pincode,
                    "party_state": party_state,
                    "invoice_no": str(get_val(row, "VoucherNo", default="") or "").strip(),
                    "invoice_type": invoice_type,
                    "invoice_date": inv_date,
                    "invoice_value": taxable + igst_amt + cgst_amt + sgst_amt,
                    "place_of_supply": place_of_supply,
                    "reverse_charge": "No",
                    "rate": rate,
                    "taxable_value": taxable,
                    "igst": igst_amt,
                    "cgst": cgst_amt,
                    "sgst": sgst_amt,
                    "cess": 0.0,
                    "filing_period": "",
                    "itc_avail": itc_avail_val,
                    "row_idx": row_idx,
                    "purchase_ledger": str(get_val(row, "PurchaseLedger", default="Purchase Account") or "Purchase Account").strip(),
                    "narration": str(get_val(row, "Narration", default="") or "").strip(),
                    "tds_ledger": str(get_val(row, "TDSLedger", default="") or "").strip(),
                    "tds_rate": parsed_tds_rate,
                    "tds_amount": parsed_tds_amount,
                    "supplier_invoice_no": str(get_val(row, "SupplierInvoiceNo", "VoucherNo", default="") or "").strip(),
                    "supplier_invoice_date": inv_date,
                }
                self.records.append(record)

                if progress_callback and row_idx % 20 == 0:
                    pct = row_idx / max(1, total_rows)
                    progress_callback(min(pct, 1.0), f"Reading row {row_idx}/{total_rows}...")

            self._compute_stats()
            wb.close()
            return len(self.records) > 0
        except Exception as e:
            self.errors.append(f"Failed to read tally sheet: {str(e)}")
            import traceback
            self.errors.append(traceback.format_exc())
            return False

    def _parse_readme(self, ws):
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=False):
            vals = [c.value for c in row]
            if not vals[0]:
                continue
            label = str(vals[0]).strip()
            value = str(vals[2] or "") if len(vals) > 2 else ""
            if label == "GSTIN":
                self.company_gstin = value
            elif label == "Legal Name":
                self.company_name = value
            elif "Trade Name" in label:
                self.trade_name = value
            elif label == "Financial Year":
                self.financial_year = value
            elif label == "Tax Period":
                self.tax_period = value

    def _safe_get(self, row, field, mapper, default=None):
        col_idx = mapper.get(field)
        if col_idx is None or col_idx >= len(row):
            return default
        return row[col_idx] if row[col_idx] is not None else default

    def _safe_float(self, row, field, mapper, default=0.0):
        val = self._safe_get(row, field, mapper)
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    def _safe_str(self, row, field, mapper, default=""):
        val = self._safe_get(row, field, mapper)
        if val is None:
            return default
        return str(val).strip()

    def _parse_b2b_row(self, row, row_idx, mapper):
        gstin = self._safe_str(row, "gstin", mapper)
        if not gstin or len(gstin) < 15:
            return None
        inv_date_raw = self._safe_get(row, "invoice_date", mapper)
        if isinstance(inv_date_raw, datetime.datetime):
            inv_date = inv_date_raw.strftime("%d/%m/%Y")
        elif inv_date_raw:
            inv_date = str(inv_date_raw)
        else:
            inv_date = ""
            
        taxable = self._safe_float(row, "taxable_value", mapper)
        igst = self._safe_float(row, "igst", mapper)
        cgst = self._safe_float(row, "cgst", mapper)
        sgst = self._safe_float(row, "sgst", mapper)
        cess = self._safe_float(row, "cess", mapper)
        invoice_value = self._safe_float(row, "invoice_value", mapper)
        
        invoice_type = self._safe_str(row, "invoice_type", mapper, "Regular")
        is_credit_note = "credit note" in invoice_type.lower()
        
        if is_credit_note:
            taxable = -abs(taxable) if taxable != 0 else 0.0
            igst = -abs(igst) if igst != 0 else 0.0
            cgst = -abs(cgst) if cgst != 0 else 0.0
            sgst = -abs(sgst) if sgst != 0 else 0.0
            cess = -abs(cess) if cess != 0 else 0.0
            invoice_value = -abs(invoice_value) if invoice_value != 0 else 0.0
            
        if mapper.has("rate"):
            rate = self._safe_float(row, "rate", mapper)
        else:
            total_tax = abs(igst) + abs(cgst) + abs(sgst)
            if abs(taxable) > 0 and total_tax > 0:
                rate = round((total_tax / abs(taxable)) * 100, 0)
            else:
                rate = 0
                
        return {
            "gstin": gstin,
            "trade_name": self._safe_str(row, "trade_name", mapper),
            "invoice_no": self._safe_str(row, "invoice_no", mapper),
            "orig_invoice_no": self._safe_str(row, "orig_invoice_no", mapper) if mapper.has("orig_invoice_no") else "",
            "invoice_type": invoice_type,
            "invoice_date": inv_date,
            "invoice_value": invoice_value,
            "place_of_supply": self._safe_str(row, "place_of_supply", mapper),
            "reverse_charge": self._safe_str(row, "reverse_charge", mapper, "No"),
            "rate": rate,
            "taxable_value": taxable,
            "igst": igst, "cgst": cgst, "sgst": sgst, "cess": cess,
            "filing_period": self._safe_str(row, "filing_period", mapper),
            "itc_avail": self._safe_str(row, "itc_avail", mapper),
            "row_idx": row_idx,
        }

    def _compute_stats(self):
        self.stats = {
            "total_records": len(self.records),
            "igst_count": sum(1 for r in self.records if abs(r["igst"]) > 0),
            "cgst_sgst_count": sum(1 for r in self.records if abs(r["cgst"]) > 0),
            "total_taxable": sum(abs(r["taxable_value"]) for r in self.records),
            "total_igst": sum(abs(r["igst"]) for r in self.records),
            "total_cgst": sum(abs(r["cgst"]) for r in self.records),
            "total_sgst": sum(abs(r["sgst"]) for r in self.records),
            "total_cess": sum(abs(r.get("cess", 0)) for r in self.records),
        }

    def _nearest_allowed_tax_rate(self, rate_value):
        return min(self.allowed_tax_rates, key=lambda r: abs(r - rate_value))

    def validate_tax_configuration(self, records=None):
        """
        Validate GST structure and effective tax percentage.

        Rules:
        - Allowed GST slabs: 0, 5, 12, 18, 28, 40
        - IGST should not coexist with CGST/SGST
        - CGST and SGST should appear together and be equal
        """
        source_records = records if records is not None else self.records
        valid_records = []
        invalid_issues = []

        for rec in source_records:
            taxable = abs(float(rec.get("taxable_value") or 0.0))
            igst_amt = abs(float(rec.get("igst") or 0.0))
            cgst_amt = abs(float(rec.get("cgst") or 0.0))
            sgst_amt = abs(float(rec.get("sgst") or 0.0))
            cess_amt = abs(float(rec.get("cess") or 0.0))

            has_igst = igst_amt > 0.009
            has_cgst = cgst_amt > 0.009
            has_sgst = sgst_amt > 0.009
            reasons = []

            if has_igst and (has_cgst or has_sgst):
                reasons.append("IGST cannot be present together with CGST/SGST.")

            if has_cgst != has_sgst:
                reasons.append("CGST and SGST must both be present (or both zero).")

            if has_cgst and has_sgst and abs(cgst_amt - sgst_amt) > 1.0:
                reasons.append("CGST and SGST amounts are not equal.")

            if taxable <= 0 and (has_igst or has_cgst or has_sgst):
                reasons.append("Tax amount exists but taxable value is zero.")

            if has_igst:
                tax_structure = "IGST"
            elif has_cgst or has_sgst:
                tax_structure = "CGST+SGST"
            else:
                tax_structure = "No GST"

            if taxable > 0:
                if has_igst:
                    computed_rate = (igst_amt / taxable) * 100.0
                else:
                    computed_rate = ((cgst_amt + sgst_amt) / taxable) * 100.0
            else:
                computed_rate = 0.0

            nearest_rate = self._nearest_allowed_tax_rate(computed_rate)
            if not rec.get("is_multi_rate") and abs(computed_rate - nearest_rate) > self.tax_rate_tolerance:
                reasons.append(
                    f"Computed GST rate {computed_rate:.2f}% is not in allowed slabs "
                    "(0, 5, 12, 18, 28, 40)."
                )

            if reasons:
                invalid_issues.append({
                    "row_idx": rec.get("row_idx", ""),
                    "invoice_no": rec.get("invoice_no", ""),
                    "party_name": rec.get("trade_name", ""),
                    "taxable_value": taxable,
                    "igst": igst_amt,
                    "cgst": cgst_amt,
                    "sgst": sgst_amt,
                    "cess": cess_amt,
                    "tax_structure": tax_structure,
                    "computed_rate": round(computed_rate, 4),
                    "sheet_rate": rec.get("rate", ""),
                    "nearest_allowed_rate": nearest_rate,
                    "issue": " | ".join(reasons),
                })
            else:
                valid_records.append(rec)

        return valid_records, invalid_issues

    def _consolidate_invoice_records(self, records: list, default_purchase_ledger: str = "Purchase Account") -> list:
        """Combine rows that belong to the same invoice into a single voucher record."""
        self._consolidation_merge_log = []
        if not records:
            return []

        grouped = []
        key_map = {}

        def _as_float(val, default=0.0):
            try:
                if val in (None, ""):
                    return default
                return float(val)
            except (TypeError, ValueError):
                return default

        def _coalesce_text(current, incoming):
            if str(current or "").strip():
                return current
            return incoming

        def _is_ineligible(value):
            return str(value or "").strip().upper() in {"NO", "N", "INELIGIBLE"}

        def _normalize_invoice_no(value):
            text = str(value or "").strip()
            if not text:
                return ""
            if text.endswith(".0") and text[:-2].isdigit():
                text = text[:-2]
            return text

        def _resolve_purchase_ledger(rec):
            ledger = str(rec.get("purchase_ledger") or "").strip()
            if not ledger:
                party_name = rec.get("trade_name") or rec.get("party_name") or ""
                ledger = self.get_purchase_ledger(party_name, default_purchase_ledger)
            ledger = (ledger or default_purchase_ledger).strip()

            rate_val = rec.get("rate")
            try:
                rate = float(rate_val or 0)
            except (TypeError, ValueError):
                rate = 0.0

            if rate > 0 and re.search(r"\d+(?:\.\d+)?\s*%", ledger):
                rate_clean = int(rate) if abs(rate - int(rate)) < 0.01 else rate
                rate_text = f"{rate_clean:g}%"
                ledger = re.sub(r"\d+(?:\.\d+)?\s*%", rate_text, ledger, count=1)
            return ledger

        for rec in records:
            invoice_no = _normalize_invoice_no(rec.get("invoice_no") or rec.get("supplier_invoice_no") or "")
            if not invoice_no:
                grouped.append(dict(rec))
                continue

            invoice_type_key = str(rec.get("invoice_type") or "Regular").strip().lower()
            sheet_type_key = str(rec.get("sheet_type") or "").strip().upper()
            orig_invoice_key = _normalize_invoice_no(rec.get("orig_invoice_no") or "")
            date_key = self._normalize_date_str(
                rec.get("invoice_date") or rec.get("supplier_invoice_date") or rec.get("voucher_date") or ""
            )

            key = (
                str(rec.get("gstin") or "").strip().upper(),
                str(rec.get("trade_name") or rec.get("party_name") or "").strip().upper(),
                invoice_no.upper(),
                date_key,
                invoice_type_key,
                sheet_type_key,
                orig_invoice_key,
            )

            base = key_map.get(key)
            if base is None:
                base = dict(rec)
                base["invoice_no"] = invoice_no
                if not str(base.get("supplier_invoice_no") or "").strip():
                    base["supplier_invoice_no"] = invoice_no
                if orig_invoice_key and not str(base.get("orig_invoice_no") or "").strip():
                    base["orig_invoice_no"] = orig_invoice_key
                base_ledger = _resolve_purchase_ledger(rec)
                base["purchase_ledger"] = base_ledger
                base["purchase_ledger_splits"] = {base_ledger: _as_float(rec.get("taxable_value"))}
                base["taxable_value"] = _as_float(base.get("taxable_value"))
                base["igst"] = _as_float(base.get("igst"))
                base["cgst"] = _as_float(base.get("cgst"))
                base["sgst"] = _as_float(base.get("sgst"))
                base["cess"] = _as_float(base.get("cess"))
                base["invoice_value"] = _as_float(base.get("invoice_value"))
                key_map[key] = base
                grouped.append(base)
                continue

            # This row is being merged into an existing invoice record
            merge_log = getattr(self, "_consolidation_merge_log", None)
            if merge_log is None:
                self._consolidation_merge_log = []
                merge_log = self._consolidation_merge_log
            party = str(rec.get("trade_name") or rec.get("party_name") or "")[:30]
            merge_log.append(
                f"  Merged row (inv={invoice_no}, party={party}, "
                f"rate={rec.get('rate', '?')}%) into existing invoice record"
            )

            if rec.get("rate") != base.get("rate"):
                base["is_multi_rate"] = True

            base["taxable_value"] = _as_float(base.get("taxable_value")) + _as_float(rec.get("taxable_value"))
            base["igst"] = _as_float(base.get("igst")) + _as_float(rec.get("igst"))
            base["cgst"] = _as_float(base.get("cgst")) + _as_float(rec.get("cgst"))
            base["sgst"] = _as_float(base.get("sgst")) + _as_float(rec.get("sgst"))
            base["cess"] = _as_float(base.get("cess")) + _as_float(rec.get("cess"))

            split_ledger = _resolve_purchase_ledger(rec)
            splits = base.setdefault("purchase_ledger_splits", {})
            splits[split_ledger] = _as_float(splits.get(split_ledger), 0.0) + _as_float(rec.get("taxable_value"))

            base["voucher_date"] = _coalesce_text(base.get("voucher_date"), rec.get("voucher_date") or rec.get("invoice_date"))
            base["voucher_no"] = _coalesce_text(base.get("voucher_no"), rec.get("voucher_no"))
            base["invoice_date"] = _coalesce_text(base.get("invoice_date"), rec.get("invoice_date"))
            base["invoice_type"] = _coalesce_text(base.get("invoice_type"), rec.get("invoice_type"))
            base["party_name"] = _coalesce_text(base.get("party_name"), rec.get("party_name"))
            base["party_mailing_name"] = _coalesce_text(base.get("party_mailing_name"), rec.get("party_mailing_name"))
            base["party_address1"] = _coalesce_text(base.get("party_address1"), rec.get("party_address1"))
            base["party_address2"] = _coalesce_text(base.get("party_address2"), rec.get("party_address2"))
            base["party_pincode"] = _coalesce_text(base.get("party_pincode"), rec.get("party_pincode"))
            base["party_state"] = _coalesce_text(base.get("party_state"), rec.get("party_state"))
            base["place_of_supply"] = _coalesce_text(base.get("place_of_supply"), rec.get("place_of_supply"))
            base["purchase_ledger"] = _coalesce_text(base.get("purchase_ledger"), rec.get("purchase_ledger"))
            base["narration"] = _coalesce_text(base.get("narration"), rec.get("narration"))
            base["tds_ledger"] = _coalesce_text(base.get("tds_ledger"), rec.get("tds_ledger"))
            base["tds_rate"] = _coalesce_text(base.get("tds_rate"), rec.get("tds_rate"))

            supplier_inv_no = _normalize_invoice_no(rec.get("supplier_invoice_no") or "")
            if supplier_inv_no:
                base["supplier_invoice_no"] = _coalesce_text(base.get("supplier_invoice_no"), supplier_inv_no)
            if rec.get("supplier_invoice_date"):
                base["supplier_invoice_date"] = _coalesce_text(base.get("supplier_invoice_date"), rec.get("supplier_invoice_date"))

            if str(rec.get("reverse_charge") or "").strip().upper() in {"YES", "Y"}:
                base["reverse_charge"] = "Yes"

            if _is_ineligible(rec.get("itc_avail")) or _is_ineligible(base.get("itc_avail")):
                base["itc_avail"] = "Ineligible"
            else:
                base["itc_avail"] = _coalesce_text(base.get("itc_avail"), rec.get("itc_avail")) or "Yes"

            if rec.get("has_stock_item"):
                base["has_stock_item"] = True

            base_tds_amount = _as_float(base.get("tds_amount"), None)
            rec_tds_amount = _as_float(rec.get("tds_amount"), None)
            if base_tds_amount is None and rec_tds_amount is None:
                base["tds_amount"] = ""
            else:
                base["tds_amount"] = _as_float(base_tds_amount, 0.0) + _as_float(rec_tds_amount, 0.0)

        for rec in grouped:
            taxable = _as_float(rec.get("taxable_value"))
            igst = _as_float(rec.get("igst"))
            cgst = _as_float(rec.get("cgst"))
            sgst = _as_float(rec.get("sgst"))
            cess = _as_float(rec.get("cess"))
            rec["invoice_value"] = taxable + igst + cgst + sgst + cess
            tax_total = abs(igst) + abs(cgst) + abs(sgst)
            if taxable:
                rec["rate"] = round((tax_total / abs(taxable)) * 100, 2) if tax_total else 0
            else:
                rec["rate"] = 0
        return grouped

    # ─── OUTPUT GENERATORS ───

    def generate_tally_sheet(self, output_path, purchase_ledger="Purchase Account",
                              narration_template="Being purchase from {party} vide Inv {inv} dt {date}",
                              progress_callback=None, records=None) -> bool:
        try:
            source_records = records if records is not None else self.records
            source_records = self._consolidate_invoice_records(source_records, default_purchase_ledger=purchase_ledger)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tally Sheet"
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            data_font = Font(name="Calibri", size=10)
            num_font = Font(name="Calibri", size=10)
            border_style = Border(
                left=Side(style="thin", color="334155"), right=Side(style="thin", color="334155"),
                top=Side(style="thin", color="334155"), bottom=Side(style="thin", color="334155"))
            odd_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            for col_idx, header in enumerate(self.TALLY_HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = header_align; cell.border = border_style
            widths = [14, 18, 35, 35, 18, 18, 18, 24, 16, 14, 12, 14, 12, 14, 12, 55]
            widths += [20, 12, 14, 20, 18, 30, 32, 32, 14]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            total_records = len(source_records)
            for idx, rec in enumerate(source_records):
                row_num = idx + 2
                _trade_name = rec.get("trade_name") or rec.get("party_name") or ""
                _invoice_no = rec.get("invoice_no") or ""
                _invoice_date = rec.get("invoice_date") or ""
                voucher_date = rec.get("voucher_date") or _invoice_date or self._today_str()
                _igst_val = abs(rec.get("igst") or 0)
                _cgst_val = abs(rec.get("cgst") or 0)
                _sgst_val = abs(rec.get("sgst") or 0)
                _rate_val = rec.get("rate") or 0
                _taxable  = rec.get("taxable_value") or 0
                is_igst = _igst_val > 0

                # Determine purchase ledger — mapping takes priority
                rec_ledger = self.get_purchase_ledger(_trade_name, purchase_ledger)
                tds_ledger = rec.get("tds_ledger") or self.get_tds_ledger(_trade_name, "")
                tds_rate = rec.get("tds_rate")
                if tds_rate in (None, ""):
                    tds_rate = self.get_tds_rate(_trade_name, "")
                tds_amount = rec.get("tds_amount", "")
                supplier_invoice_no = rec.get("supplier_invoice_no") or _invoice_no
                supplier_invoice_date = rec.get("supplier_invoice_date") or _invoice_date

                if is_igst:
                    cgst_ledger = 0; cgst_rate = 0; sgst_ledger = 0; sgst_rate = 0
                    igst_ledger = "IGST"; igst_rate = _rate_val
                else:
                    half_rate = _rate_val / 2 if _rate_val > 0 else 0
                    cgst_ledger = "CGST"; cgst_rate = half_rate
                    sgst_ledger = "SGST"; sgst_rate = half_rate
                    igst_ledger = 0; igst_rate = 0
                try:
                    narration = narration_template.format(
                        party=_trade_name, inv=_invoice_no, date=_invoice_date)
                except (KeyError, IndexError):
                    narration = f"Being purchase from {_trade_name} vide Inv {_invoice_no} dt {_invoice_date}"
                if rec.get("is_amendment") and rec.get("orig_invoice_no"):
                    narration += f" [Amends {rec['orig_invoice_no']}]"
                row_data = [voucher_date, rec.get("voucher_no", ""), _trade_name,
                            rec.get("party_name", _trade_name), rec.get("gstin", ""),
                            rec.get("party_state", ""), rec.get("place_of_supply", ""),
                            rec_ledger, _taxable,
                            cgst_ledger, cgst_rate, sgst_ledger, sgst_rate,
                            igst_ledger, igst_rate, narration,
                            tds_ledger, tds_rate, tds_amount, supplier_invoice_no, supplier_invoice_date,
                            rec.get("party_mailing_name", _trade_name), rec.get("party_address1", ""),
                            rec.get("party_address2", ""), rec.get("party_pincode", "")]
                fill = odd_fill if idx % 2 == 0 else even_fill
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.font = num_font if isinstance(value, (int, float)) else data_font
                    cell.border = border_style; cell.fill = fill
                    if isinstance(value, float):
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '#,##0.00'
                    elif isinstance(value, (int,)):
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                if progress_callback and idx % 20 == 0:
                    pct = idx / max(1, total_records)
                    progress_callback(min(pct, 1.0), f"Writing row {idx+1}/{total_records}...")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            wb.save(output_path); wb.close()
            return True
        except Exception as e:
            import traceback
            self.errors.append(f"Excel generation failed: {str(e)}\n{traceback.format_exc()}")
            return False

    def generate_tally_xml(self, output_path, company_name="",
                            purchase_ledger="Purchase Account",
                            narration_template="Being purchase from {party} vide Inv {inv} dt {date}",
                            progress_callback=None, records=None,
                            round_off_ledger: str = "") -> bool:
        try:
            source_records = records if records is not None else self.records
            pre_count = len(source_records)
            source_records = self._consolidate_invoice_records(source_records, default_purchase_ledger=purchase_ledger)
            post_count = len(source_records)
            self._last_consolidation_msg = (
                f"Building XML: {pre_count} rows → {post_count} unique invoice(s) after consolidation"
                + (f" ({pre_count - post_count} multi-rate rows merged)" if pre_count != post_count else "")
            )
            if progress_callback:
                progress_callback(0.05, self._last_consolidation_msg)
            comp_name = company_name or self.trade_name or self.company_name or "My Company"
            envelope = ET.Element("ENVELOPE")
            header = ET.SubElement(envelope, "HEADER")
            ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
            body = ET.SubElement(envelope, "BODY")
            import_data = ET.SubElement(body, "IMPORTDATA")
            req_desc = ET.SubElement(import_data, "REQUESTDESC")
            ET.SubElement(req_desc, "REPORTNAME").text = "Vouchers"
            static_vars = ET.SubElement(req_desc, "STATICVARIABLES")
            ET.SubElement(static_vars, "SVCURRENTCOMPANY").text = comp_name
            req_data = ET.SubElement(import_data, "REQUESTDATA")
            total_records = len(source_records)
            for idx, rec in enumerate(source_records):
                _trade = rec.get("trade_name") or rec.get("party_name") or ""
                _inv   = rec.get("invoice_no") or ""
                _date  = rec.get("invoice_date") or ""
                # For tally sheet records, use stored purchase_ledger & narration
                rec_ledger = rec.get("purchase_ledger") or self.get_purchase_ledger(_trade, purchase_ledger)
                try:
                    rec_narration = rec.get("narration") or narration_template.format(
                        party=_trade, inv=_inv, date=_date)
                except (KeyError, IndexError):
                    rec_narration = f"Being purchase from {_trade} vide Inv {_inv} dt {_date}"
                if rec.get("is_amendment") and rec.get("orig_invoice_no"):
                    rec_narration += f" [Amends {rec['orig_invoice_no']}]"
                self._build_voucher_xml(req_data, rec, rec_ledger, rec_narration, self._today_str(), round_off_ledger=round_off_ledger)
                if progress_callback and idx % 20 == 0:
                    pct = idx / max(1, total_records)
                    progress_callback(min(pct, 1.0), f"Building XML {idx+1}/{total_records}...")
            rough_string = ET.tostring(envelope, encoding="unicode")
            parsed = minidom.parseString(rough_string)
            pretty_xml = parsed.toprettyxml(indent="  ", encoding=None)
            lines = pretty_xml.split("\n")
            if lines[0].startswith("<?xml"):
                lines = lines[1:]
            xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n' + "\n".join(lines)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(xml_content)
            return True
        except Exception as e:
            import traceback
            self.errors.append(f"XML generation failed: {str(e)}\n{traceback.format_exc()}")
            return False

    def _build_voucher_xml(self, parent, rec, purchase_ledger, narration, voucher_date, round_off_ledger: str = ""):
        # RCM invoices get a separate Journal voucher structure
        if str(rec.get("reverse_charge") or "").strip().upper() == "YES":
            self._build_rcm_journal_voucher_xml(parent, rec, purchase_ledger, narration, voucher_date)
            return

        itc_status = str(rec.get("itc_avail") or "Yes").strip().upper()
        is_itc_eligible = itc_status not in {"NO", "N", "INELIGIBLE"}
        if not is_itc_eligible:
            self._build_journal_voucher_xml(parent, rec, purchase_ledger, narration, voucher_date, round_off_ledger=round_off_ledger)
            return

        tally_msg = ET.SubElement(parent, "TALLYMESSAGE")
        tally_msg.set("xmlns:UDF", "TallyUDF")
        voucher = ET.SubElement(tally_msg, "VOUCHER")
        voucher.set("REMOTEID", "")
        
        _inv_type_lower = str(rec.get("invoice_type", "")).lower()
        if "credit note" in _inv_type_lower:
            vch_type = "Debit Note"
        elif "debit note" in _inv_type_lower:
            vch_type = "Credit Note"
        else:
            vch_type = "Purchase"
        is_debit_note = vch_type == "Debit Note"
        is_note = vch_type in ("Debit Note", "Credit Note")
        voucher.set("VCHTYPE", vch_type)
        voucher.set("ACTION", "Create")
        voucher.set("OBJVIEW", "Invoice Voucher View")

        actual_voucher_date = rec.get("voucher_date") or rec.get("invoice_date") or voucher_date
        tally_date = self._tally_date(actual_voucher_date, fallback_today=True)
        ref_date = self._tally_date(
            rec.get("supplier_invoice_date") or rec.get("invoice_date") or actual_voucher_date,
            fallback_today=False,
        )

        party_name = rec.get("party_name") or rec["trade_name"]
        party_ledger = rec["trade_name"]
        party_gstin = str(rec.get("gstin") or "").strip().upper()
        party_state = self._normalize_state_name(rec.get("party_state") or self._state_from_gstin(party_gstin))
        company_gstin = str(self.company_gstin or "").strip().upper()
        company_state = self.company_registration_state or self._state_from_gstin(company_gstin)
        
        # For purchase vouchers (inwards), Place of Supply must be the Company's state,
        # otherwise Tally flags the GST Registration Details as invalid/uncertain.
        place_of_supply = company_state or self._normalize_state_name(rec.get("place_of_supply") or "")
        party_mailing_name = rec.get("party_mailing_name") or party_name
        party_address1 = rec.get("party_address1") or ""
        party_address2 = rec.get("party_address2") or ""
        party_pincode = str(rec.get("party_pincode") or "").strip()
        supplier_invoice_no = rec.get("supplier_invoice_no") or rec["invoice_no"]

        if party_address1 or party_address2:
            address_list = ET.SubElement(voucher, "ADDRESS.LIST")
            address_list.set("TYPE", "String")
            if party_address1:
                ET.SubElement(address_list, "ADDRESS").text = party_address1
            if party_address2:
                ET.SubElement(address_list, "ADDRESS").text = party_address2

        ET.SubElement(voucher, "DATE").text = tally_date
        if ref_date:
            ET.SubElement(voucher, "REFERENCEDATE").text = ref_date
        ET.SubElement(voucher, "GSTREGISTRATIONTYPE").text = "Regular"
        ET.SubElement(voucher, "VATDEALERTYPE").text = "Regular"
        if party_state:
            ET.SubElement(voucher, "STATENAME").text = party_state
        ET.SubElement(voucher, "COUNTRYOFRESIDENCE").text = "India"
        if party_gstin:
            ET.SubElement(voucher, "PARTYGSTIN").text = party_gstin
        if place_of_supply:
            ET.SubElement(voucher, "PLACEOFSUPPLY").text = place_of_supply
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = vch_type
        ET.SubElement(voucher, "PARTYNAME").text = party_name
        if company_gstin:
            # GSTREGISTRATION text must exactly match the registration name in Tally's
            # company master (e.g. "Delhi Registration"). Tally uses this to associate the
            # voucher with the correct company GST registration; without it the voucher goes
            # to "Uncertain Transactions" in GSTR-3B.
            # Use the real name fetched from Tally; fall back to synthetic "{State} Registration".
            reg_name = (
                self.company_registration_name
                or (f"{company_state} Registration" if company_state else company_gstin)
            )
            gst_reg = ET.SubElement(voucher, "GSTREGISTRATION")
            gst_reg.set("TAXTYPE", "GST")
            gst_reg.set("TAXREGISTRATION", company_gstin)
            gst_reg.text = reg_name
            ET.SubElement(voucher, "CMPGSTIN").text = company_gstin
            ET.SubElement(voucher, "CMPGSTREGISTRATIONTYPE").text = "Regular"
            # CMPGSTSTATE: use fetched state, else derive from GSTIN
            _cmp_state = self.company_registration_state or company_state
            if _cmp_state:
                ET.SubElement(voucher, "CMPGSTSTATE").text = _cmp_state
        ET.SubElement(voucher, "PARTYLEDGERNAME").text = party_ledger
        ET.SubElement(voucher, "REFERENCE").text = supplier_invoice_no
        ET.SubElement(voucher, "PARTYMAILINGNAME").text = party_mailing_name
        if party_pincode:
            ET.SubElement(voucher, "PARTYPINCODE").text = party_pincode
        ET.SubElement(voucher, "BASICBASEPARTYNAME").text = party_name
        ET.SubElement(voucher, "PERSISTEDVIEW").text = "Invoice Voucher View"
        ET.SubElement(voucher, "VCHENTRYMODE").text = "Accounting Invoice"
        ET.SubElement(voucher, "ISINVOICE").text = "Yes"
        # ISGSTOVERRIDDEN=No — lets Tally compute GST normally from CMPGSTIN.
        # "Yes" would override Tally's own computation and cause Uncertain in GSTR-3B.
        ET.SubElement(voucher, "ISGSTOVERRIDDEN").text = "No"
        ET.SubElement(voucher, "GSTTRANSACTIONTYPE").text = "Tax Invoice" if party_gstin else "Unregistered"
        ET.SubElement(voucher, "EFFECTIVEDATE").text = tally_date
        ET.SubElement(voucher, "ISELIGIBLEFORITC").text = "Yes"
        ET.SubElement(voucher, "NARRATION").text = narration
        # Only set VOUCHERNUMBER when explicitly provided (not invoice_no fallback).
        # Using invoice_no as voucher number triggers Tally "Prevent Duplicate" blocks
        # on re-import and silently causes Created: 0 with no error.
        vch_no = str(rec.get("voucher_no") or "").strip()
        if vch_no:
            ET.SubElement(voucher, "VOUCHERNUMBER").text = vch_no

        # Round to 2 dp up-front so every formatted ledger amount is stable.
        taxable = round(float(rec.get("taxable_value") or 0), 2)
        igst_amt = round(float(rec.get("igst") or 0), 2)
        cgst_amt = round(float(rec.get("cgst") or 0), 2)
        sgst_amt = round(float(rec.get("sgst") or 0), 2)
        cess_amt = round(float(rec.get("cess") or 0), 2)

        _tname_for_tds = rec.get("trade_name") or rec.get("party_name") or ""
        tds_ledger = rec.get("tds_ledger") or self.get_tds_ledger(_tname_for_tds, "")
        tds_rate = rec.get("tds_rate")
        if tds_rate in (None, ""):
            tds_rate = self.get_tds_rate(_tname_for_tds, 0)
        try:
            tds_rate = float(tds_rate or 0)
        except (ValueError, TypeError):
            tds_rate = 0.0

        tds_amount = rec.get("tds_amount")
        if tds_amount in (None, "") and tds_ledger and tds_rate > 0:
            tds_amount = round(taxable * tds_rate / 100, 2)
        else:
            try:
                raw_tds = round(abs(float(tds_amount or 0)), 2)
                tds_amount = -raw_tds if taxable < 0 else raw_tds
            except (ValueError, TypeError):
                tds_amount = 0.0

        total_amount = round(taxable + igst_amt + cgst_amt + sgst_amt + cess_amt, 2)
        if abs(tds_amount) > abs(total_amount):
            tds_amount = total_amount

        # Pre-compute rounded split amounts so party_amount is their exact complement.
        # This prevents ±0.01 imbalances from float arithmetic that Tally silently rejects.
        splits = rec.get("purchase_ledger_splits") or {}
        rounded_splits = {k: round(float(v), 2) for k, v in splits.items() if round(float(v or 0), 2) != 0}
        purchase_total = round(sum(rounded_splits.values()), 2) if rounded_splits else taxable
        tax_total = round(igst_amt + cgst_amt + sgst_amt + cess_amt, 2)
        # party_amount is the exact sum of all other rounded entries minus TDS
        _ro_base     = round(purchase_total + tax_total, 2)
        _ro_amt_main = round(round(_ro_base, 0) - _ro_base, 2) if round_off_ledger else 0.0
        party_amount = round(_ro_base - round(float(tds_amount or 0), 2) + _ro_amt_main, 2)

        # For Debit Note: party is debited (Dr Supplier), purchase/tax are credited (Cr).
        # For Purchase / Credit Note: party is credited (Cr Supplier), purchase/tax are debited (Dr).
        party_deemed = "Yes" if is_debit_note else "No"
        counter_deemed = "No" if is_debit_note else "Yes"

        pe = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
        ET.SubElement(pe, "LEDGERNAME").text = party_ledger
        self._add_common_ledger_flags(pe, is_party="Yes")
        pe.find("ISDEEMEDPOSITIVE").text = party_deemed
        ET.SubElement(pe, "AMOUNT").text = f"{party_amount:.2f}"
        ba = ET.SubElement(pe, "BILLALLOCATIONS.LIST")
        ET.SubElement(ba, "NAME").text = supplier_invoice_no
        ET.SubElement(ba, "BILLTYPE").text = "New Ref"
        ET.SubElement(ba, "AMOUNT").text = f"{party_amount:.2f}"

        if rounded_splits:
            for ledger_name, ledger_amount in rounded_splits.items():
                pu = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
                ET.SubElement(pu, "LEDGERNAME").text = str(ledger_name)
                self._add_common_ledger_flags(pu, is_party="No")
                pu.find("ISDEEMEDPOSITIVE").text = counter_deemed
                ET.SubElement(pu, "AMOUNT").text = f"{-ledger_amount:.2f}"
        else:
            pu = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(pu, "LEDGERNAME").text = purchase_ledger
            self._add_common_ledger_flags(pu, is_party="No")
            pu.find("ISDEEMEDPOSITIVE").text = counter_deemed
            ET.SubElement(pu, "AMOUNT").text = f"{-taxable:.2f}"

        if abs(igst_amt) > 0:
            ie = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(ie, "LEDGERNAME").text = "IGST"
            self._add_common_ledger_flags(ie, is_party="No")
            ie.find("ISDEEMEDPOSITIVE").text = counter_deemed
            ET.SubElement(ie, "AMOUNT").text = f"{-igst_amt:.2f}"
        else:
            if abs(cgst_amt) > 0:
                ce = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
                ET.SubElement(ce, "LEDGERNAME").text = "CGST"
                self._add_common_ledger_flags(ce, is_party="No")
                ce.find("ISDEEMEDPOSITIVE").text = counter_deemed
                ET.SubElement(ce, "AMOUNT").text = f"{-cgst_amt:.2f}"
            if abs(sgst_amt) > 0:
                se = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
                ET.SubElement(se, "LEDGERNAME").text = "SGST"
                self._add_common_ledger_flags(se, is_party="No")
                se.find("ISDEEMEDPOSITIVE").text = counter_deemed
                ET.SubElement(se, "AMOUNT").text = f"{-sgst_amt:.2f}"
        if abs(cess_amt) > 0:
            cs = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(cs, "LEDGERNAME").text = "Cess"
            self._add_common_ledger_flags(cs, is_party="No")
            cs.find("ISDEEMEDPOSITIVE").text = counter_deemed
            ET.SubElement(cs, "AMOUNT").text = f"{-cess_amt:.2f}"

        if tds_ledger and abs(tds_amount) > 0:
            te = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(te, "LEDGERNAME").text = tds_ledger
            self._add_common_ledger_flags(te, is_party="No")
            te.find("ISDEEMEDPOSITIVE").text = counter_deemed
            ET.SubElement(te, "AMOUNT").text = f"{tds_amount:.2f}"

        if round_off_ledger and abs(_ro_amt_main) >= 0.005:
            ro = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(ro, "LEDGERNAME").text = round_off_ledger
            self._add_common_ledger_flags(ro, is_party="No")
            if _ro_amt_main > 0:
                # Extra paid → debit (expense), same side as purchase entries
                ro.find("ISDEEMEDPOSITIVE").text = counter_deemed
                ET.SubElement(ro, "AMOUNT").text = f"{-_ro_amt_main:.2f}"
            else:
                # Less paid → credit (income), same side as party
                ro.find("ISDEEMEDPOSITIVE").text = party_deemed
                ET.SubElement(ro, "AMOUNT").text = f"{abs(_ro_amt_main):.2f}"

    def _build_journal_voucher_xml(self, parent, rec, purchase_ledger, narration, voucher_date, round_off_ledger: str = ""):
        """
        ITC-ineligible purchase entries are posted as Journal vouchers:
        - Full amount (taxable + GST + cess)
        - No tax breakup ledgers
        - No TDS ledger deduction
        """
        tally_msg = ET.SubElement(parent, "TALLYMESSAGE")
        tally_msg.set("xmlns:UDF", "TallyUDF")
        voucher = ET.SubElement(tally_msg, "VOUCHER")
        voucher.set("REMOTEID", "")
        voucher.set("VCHTYPE", "Journal")
        voucher.set("ACTION", "Create")
        voucher.set("OBJVIEW", "Accounting Voucher View")

        actual_voucher_date = rec.get("voucher_date") or rec.get("invoice_date") or voucher_date
        tally_date = self._tally_date(actual_voucher_date, fallback_today=True)

        supplier_invoice_no = rec.get("supplier_invoice_no") or rec.get("invoice_no") or ""
        party_name = rec.get("party_name") or rec.get("trade_name") or ""

        taxable = round(float(rec.get("taxable_value") or 0), 2)
        igst_amt = round(float(rec.get("igst") or 0), 2)
        cgst_amt = round(float(rec.get("cgst") or 0), 2)
        sgst_amt = round(float(rec.get("sgst") or 0), 2)
        cess_amt = round(float(rec.get("cess") or 0), 2)

        ET.SubElement(voucher, "DATE").text = tally_date
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = "Journal"
        ET.SubElement(voucher, "PERSISTEDVIEW").text = "Accounting Voucher View"
        ET.SubElement(voucher, "VCHENTRYMODE").text = "Accounting Voucher View"
        ET.SubElement(voucher, "ISINVOICE").text = "No"
        ET.SubElement(voucher, "EFFECTIVEDATE").text = tally_date
        ET.SubElement(voucher, "ISELIGIBLEFORITC").text = "No"
        # ISGSTOVERRIDDEN=No — Journal vouchers should not override GST computation.
        ET.SubElement(voucher, "ISGSTOVERRIDDEN").text = "No"
        
        party_gstin = str(rec.get("gstin") or "").strip().upper()
        company_gstin = str(self.company_gstin or "").strip().upper()
        if company_gstin:
            company_state = self.company_registration_state or self._state_from_gstin(company_gstin)
            reg_name = (
                self.company_registration_name
                or (f"{company_state} Registration" if company_state else company_gstin)
            )
            ET.SubElement(voucher, "GSTTRANSACTIONTYPE").text = "Tax Invoice" if party_gstin else "Unregistered"
            ET.SubElement(voucher, "GSTREGISTRATIONTYPE").text = "Regular"
            if party_gstin:
                ET.SubElement(voucher, "PARTYGSTIN").text = party_gstin
            gst_reg = ET.SubElement(voucher, "GSTREGISTRATION")
            gst_reg.set("TAXTYPE", "GST")
            gst_reg.set("TAXREGISTRATION", company_gstin)
            gst_reg.text = reg_name
            ET.SubElement(voucher, "CMPGSTIN").text = company_gstin
            ET.SubElement(voucher, "CMPGSTREGISTRATIONTYPE").text = "Regular"
            if company_state:
                ET.SubElement(voucher, "CMPGSTSTATE").text = company_state
                ET.SubElement(voucher, "PLACEOFSUPPLY").text = company_state
        ET.SubElement(voucher, "NARRATION").text = narration
        if supplier_invoice_no:
            ET.SubElement(voucher, "REFERENCE").text = str(supplier_invoice_no)
        vch_no = str(rec.get("voucher_no") or "").strip()
        if vch_no:
            ET.SubElement(voucher, "VOUCHERNUMBER").text = vch_no

        total_amount = round(taxable + igst_amt + cgst_amt + sgst_amt + cess_amt, 2)
        if total_amount == 0:
            return
        _ro_amt_jnl  = round(round(total_amount, 0) - total_amount, 2) if round_off_ledger else 0.0
        _ro_total_jnl = round(total_amount + _ro_amt_jnl, 2)

        # Dr: Purchase/Expense ledger FIRST (Tally convention: Dr before Cr)
        debit_purchase = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
        ET.SubElement(debit_purchase, "LEDGERNAME").text = purchase_ledger
        self._add_common_ledger_flags(debit_purchase, is_party="No")
        debit_purchase.find("ISDEEMEDPOSITIVE").text = "Yes"
        ET.SubElement(debit_purchase, "AMOUNT").text = f"{-total_amount:.2f}"

        # Round off debit/credit entry
        if round_off_ledger and abs(_ro_amt_jnl) >= 0.005:
            ro = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(ro, "LEDGERNAME").text = round_off_ledger
            self._add_common_ledger_flags(ro, is_party="No")
            if _ro_amt_jnl > 0:
                ro.find("ISDEEMEDPOSITIVE").text = "Yes"
                ET.SubElement(ro, "AMOUNT").text = f"{-_ro_amt_jnl:.2f}"
            else:
                ro.find("ISDEEMEDPOSITIVE").text = "No"
                ET.SubElement(ro, "AMOUNT").text = f"{abs(_ro_amt_jnl):.2f}"

        # Cr: Party ledger (with bill reference) — rounded total
        credit_party = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
        ET.SubElement(credit_party, "LEDGERNAME").text = party_name
        self._add_common_ledger_flags(credit_party, is_party="Yes")
        credit_party.find("ISDEEMEDPOSITIVE").text = "No"
        ET.SubElement(credit_party, "AMOUNT").text = f"{_ro_total_jnl:.2f}"

        if supplier_invoice_no:
            ba = ET.SubElement(credit_party, "BILLALLOCATIONS.LIST")
            ET.SubElement(ba, "NAME").text = str(supplier_invoice_no)
            ET.SubElement(ba, "BILLTYPE").text = "New Ref"
            ET.SubElement(ba, "AMOUNT").text = f"{_ro_total_jnl:.2f}"

    def _build_rcm_journal_voucher_xml(self, parent, rec, purchase_ledger, narration, voucher_date):
        """
        RCM (Reverse Charge Mechanism) Journal voucher.
        Dr: Expense ledger (taxable), Dr: CGST/SGST/IGST Inward RCM
        Cr: Party ledger (taxable with bill ref), Cr: CGST/SGST/IGST Outward RCM
        """
        rcm = self.rcm_ledger_map
        expense_ledger      = rcm.get("expense") or purchase_ledger
        party_ledger_name   = rec.get("trade_name") or ""
        party_name          = rec.get("party_name") or party_ledger_name
        supplier_invoice_no = rec.get("supplier_invoice_no") or rec.get("invoice_no") or ""

        taxable  = round(float(rec.get("taxable_value") or 0), 2)
        igst_amt = round(float(rec.get("igst") or 0), 2)
        cgst_amt = round(float(rec.get("cgst") or 0), 2)
        sgst_amt = round(float(rec.get("sgst") or 0), 2)
        is_igst  = abs(igst_amt) > 0
        has_gst  = (abs(igst_amt) + abs(cgst_amt) + abs(sgst_amt)) > 0

        if taxable == 0 and not has_gst:
            return

        actual_voucher_date = rec.get("voucher_date") or rec.get("invoice_date") or voucher_date
        tally_date = self._tally_date(actual_voucher_date, fallback_today=True)

        party_gstin    = str(rec.get("gstin") or "").strip().upper()
        party_state    = self._normalize_state_name(rec.get("party_state") or self._state_from_gstin(party_gstin))
        place_of_supply = self._normalize_state_name(rec.get("place_of_supply") or "")
        company_gstin  = str(self.company_gstin or "").strip().upper()
        company_state  = self._state_from_gstin(company_gstin)
        company_name   = self.trade_name or self.company_name or "My Company"

        tally_msg = ET.SubElement(parent, "TALLYMESSAGE")
        tally_msg.set("xmlns:UDF", "TallyUDF")
        voucher = ET.SubElement(tally_msg, "VOUCHER")
        voucher.set("REMOTEID", "")
        voucher.set("VCHTYPE", "Journal")
        voucher.set("ACTION", "Create")
        voucher.set("OBJVIEW", "Accounting Voucher View")

        ET.SubElement(voucher, "DATE").text = tally_date
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = "Journal"
        ET.SubElement(voucher, "PERSISTEDVIEW").text = "Accounting Voucher View"
        ET.SubElement(voucher, "VCHENTRYMODE").text = "Accounting Voucher View"
        ET.SubElement(voucher, "ISINVOICE").text = "No"
        ET.SubElement(voucher, "EFFECTIVEDATE").text = tally_date

        # GST fields required for 3B report classification
        if has_gst:
            ET.SubElement(voucher, "ISGSTOVERRIDDEN").text = "No"
            ET.SubElement(voucher, "GSTTRANSACTIONTYPE").text = "Tax Invoice" if party_gstin else "Unregistered"
            ET.SubElement(voucher, "GSTREGISTRATIONTYPE").text = "Regular"
            if party_gstin:
                ET.SubElement(voucher, "PARTYGSTIN").text = party_gstin
            if company_gstin:
                company_state = self.company_registration_state or self._state_from_gstin(company_gstin)
                reg_name = (
                    self.company_registration_name
                    or (f"{company_state} Registration" if company_state else company_gstin)
                )
                gst_reg = ET.SubElement(voucher, "GSTREGISTRATION")
                gst_reg.set("TAXTYPE", "GST")
                gst_reg.set("TAXREGISTRATION", company_gstin)
                gst_reg.text = reg_name
                ET.SubElement(voucher, "CMPGSTIN").text = company_gstin
                ET.SubElement(voucher, "CMPGSTREGISTRATIONTYPE").text = "Regular"
                if company_state:
                    ET.SubElement(voucher, "CMPGSTSTATE").text = company_state
                    ET.SubElement(voucher, "PLACEOFSUPPLY").text = company_state

        ET.SubElement(voucher, "PARTYNAME").text = party_name
        ET.SubElement(voucher, "ISELIGIBLEFORITC").text = "Yes"
        ET.SubElement(voucher, "NARRATION").text = narration
        if supplier_invoice_no:
            ET.SubElement(voucher, "REFERENCE").text = str(supplier_invoice_no)
        vch_no = str(rec.get("voucher_no") or "").strip()
        if vch_no:
            ET.SubElement(voucher, "VOUCHERNUMBER").text = vch_no

        def _dr(ledger_name, amount):
            """Debit entry: ISDEEMEDPOSITIVE=Yes, amount negative."""
            e = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(e, "LEDGERNAME").text = ledger_name
            self._add_common_ledger_flags(e, is_party="No")
            e.find("ISDEEMEDPOSITIVE").text = "Yes"
            ET.SubElement(e, "AMOUNT").text = f"{-abs(amount):.2f}"
            return e

        def _cr_party(ledger_name, amount, inv_no=""):
            """Credit party entry with GST details so Tally can resolve state/country."""
            e = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(e, "LEDGERNAME").text = ledger_name
            self._add_common_ledger_flags(e, is_party="Yes")
            e.find("ISDEEMEDPOSITIVE").text = "No"
            # Party GST details — Tally reads these to populate 3B & balance sheet
            if has_gst:
                ET.SubElement(e, "GSTREGISTRATIONTYPE").text = "Regular"
                if party_gstin:
                    ET.SubElement(e, "GSTIN").text = party_gstin
                if party_state:
                    ET.SubElement(e, "STATENAME").text = party_state
                ET.SubElement(e, "COUNTRYOFRESIDENCE").text = "India"
            ET.SubElement(e, "AMOUNT").text = f"{abs(amount):.2f}"
            if inv_no:
                ba = ET.SubElement(e, "BILLALLOCATIONS.LIST")
                ET.SubElement(ba, "NAME").text = str(inv_no)
                ET.SubElement(ba, "BILLTYPE").text = "New Ref"
                ET.SubElement(ba, "AMOUNT").text = f"{abs(amount):.2f}"
            return e

        def _cr(ledger_name, amount):
            """Credit non-party entry: ISDEEMEDPOSITIVE=No, amount positive."""
            e = ET.SubElement(voucher, "LEDGERENTRIES.LIST")
            ET.SubElement(e, "LEDGERNAME").text = ledger_name
            self._add_common_ledger_flags(e, is_party="No")
            e.find("ISDEEMEDPOSITIVE").text = "No"
            ET.SubElement(e, "AMOUNT").text = f"{abs(amount):.2f}"
            return e

        # Dr Expense / Purchase ledger (taxable amount)
        if taxable:
            _dr(expense_ledger, taxable)

        # Dr GST Inward RCM ledgers
        if is_igst:
            igst_inward = rcm.get("igst_inward") or "IGST Inward RCM"
            if abs(igst_amt) > 0:
                _dr(igst_inward, igst_amt)
        else:
            cgst_inward = rcm.get("cgst_inward") or "CGST Inward RCM"
            sgst_inward = rcm.get("sgst_inward") or "SGST Inward RCM"
            if abs(cgst_amt) > 0:
                _dr(cgst_inward, cgst_amt)
            if abs(sgst_amt) > 0:
                _dr(sgst_inward, sgst_amt)

        # Cr Party ledger (taxable amount with bill ref + GST details)
        if taxable:
            _cr_party(party_ledger_name, taxable, inv_no=supplier_invoice_no)

        # Cr GST Outward RCM ledgers
        if is_igst:
            igst_outward = rcm.get("igst_outward") or "IGST Outward RCM"
            if abs(igst_amt) > 0:
                _cr(igst_outward, igst_amt)
        else:
            cgst_outward = rcm.get("cgst_outward") or "CGST Outward RCM"
            sgst_outward = rcm.get("sgst_outward") or "SGST Outward RCM"
            if abs(cgst_amt) > 0:
                _cr(cgst_outward, cgst_amt)
            if abs(sgst_amt) > 0:
                _cr(sgst_outward, sgst_amt)


# ═══════════════════════════════════════════════════════════
#  UI COMPONENTS
# ═══════════════════════════════════════════════════════════

class StatsCard(ctk.CTkFrame):
    def __init__(self, master, icon, title, value="—", accent=COLORS["accent"], **kw):
        super().__init__(master, fg_color=COLORS["bg_card"], corner_radius=12,
                        border_width=1, border_color=COLORS["border"], **kw)
        self.configure(height=100)
        icon_frame = ctk.CTkFrame(self, fg_color=accent, corner_radius=20, width=40, height=40)
        icon_frame.place(relx=0.06, rely=0.5, anchor="w"); icon_frame.pack_propagate(False)
        ctk.CTkLabel(icon_frame, text=icon, font=("Segoe UI Emoji", 16),
                     text_color="#FFFFFF").place(relx=0.5, rely=0.5, anchor="center")
        text_frame = ctk.CTkFrame(self, fg_color="transparent")
        text_frame.place(relx=0.25, rely=0.5, anchor="w")
        self.title_label = ctk.CTkLabel(text_frame, text=title, font=("Segoe UI", 11),
                                         text_color=COLORS["text_muted"])
        self.title_label.pack(anchor="w")
        self.value_label = ctk.CTkLabel(text_frame, text=value, font=("Segoe UI", 20, "bold"),
                                         text_color=COLORS["text_primary"])
        self.value_label.pack(anchor="w")

    def update_value(self, value):
        self.value_label.configure(text=value)


class LogPanel(ctk.CTkFrame):
    def __init__(self, master, **kw):
        super().__init__(master, fg_color=COLORS["bg_card"], corner_radius=12,
                        border_width=1, border_color=COLORS["border"], **kw)
        header = ctk.CTkFrame(self, fg_color="transparent", height=36)
        header.pack(fill="x", padx=16, pady=(12, 4)); header.pack_propagate(False)
        ctk.CTkLabel(header, text="Activity Log", font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(side="left")
        ctk.CTkButton(header, text="Clear", width=60, height=26, font=("Segoe UI", 11),
                      fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                      text_color=COLORS["text_secondary"], corner_radius=6,
                      command=self.clear_log).pack(side="right")
        self.textbox = ctk.CTkTextbox(self, fg_color=COLORS["bg_dark"], text_color=COLORS["text_secondary"],
                                       font=("Consolas", 11), corner_radius=8, border_width=0,
                                       wrap="word", state="disabled")
        self.textbox.pack(fill="both", expand=True, padx=12, pady=(4, 12))

    def log(self, message, level="info"):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        prefix = {"info":"[i]","success":"[OK]","warning":"[!]","error":"[X]",
                  "process":"[>>]","detect":"[?]","map":"[MAP]"}.get(level," * ")
        self.textbox.configure(state="normal")
        self.textbox.insert("end", f"  [{ts}]  {prefix}  {message}\n")
        self.textbox.see("end"); self.textbox.configure(state="disabled")

    def clear_log(self):
        self.textbox.configure(state="normal")
        self.textbox.delete("1.0", "end"); self.textbox.configure(state="disabled")


class DataPreviewTable(ctk.CTkFrame):
    def __init__(self, master, **kw):
        super().__init__(master, fg_color=COLORS["bg_card"], corner_radius=12,
                         border_width=1, border_color=COLORS["border"], **kw)
        header = ctk.CTkFrame(self, fg_color="transparent", height=36)
        header.pack(fill="x", padx=16, pady=(12, 4))
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="Data Preview",
                     font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(side="left")
        self.count_label = ctk.CTkLabel(header, text="", font=("Segoe UI", 11),
                                        text_color=COLORS["text_muted"])
        self.count_label.pack(side="right")
        self._body = ctk.CTkFrame(self, fg_color="transparent")
        self._body.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self._setup_treeview_style()
        self._show_placeholder("No file loaded")

    def _setup_treeview_style(self):
        hdr_bg  = _theme_color("table_header")
        hdr_fg  = _theme_color("tally_gold")
        row_bg  = _theme_color("table_row_odd")
        sel_bg  = _theme_color("accent")
        txt     = _theme_color("text_primary")
        style = ttk.Style()
        style.theme_use("default")
        style.configure("DP.Treeview",
                        background=row_bg, foreground=txt,
                        fieldbackground=row_bg, rowheight=26,
                        font=("Segoe UI", 10), borderwidth=0)
        style.configure("DP.Treeview.Heading",
                        background=hdr_bg, foreground=hdr_fg,
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("DP.Treeview", background=[("selected", sel_bg)])
        style.map("DP.Treeview.Heading", background=[("active", hdr_bg)])

    def _col(self, name):
        v = COLORS.get(name, name)
        if isinstance(v, tuple):
            m = ctk.get_appearance_mode().lower()
            return v[1] if m == "dark" else v[0]
        return v

    def _clear(self):
        for w in self._body.winfo_children():
            w.destroy()

    def _show_placeholder(self, msg="No data"):
        self._clear()
        ctk.CTkLabel(self._body, text=msg,
                     font=("Segoe UI", 12),
                     text_color=COLORS["text_muted"]).pack(expand=True)

    def _build_tree(self, parent, columns):
        odd_bg  = _theme_color("table_row_odd")
        even_bg = _theme_color("table_row_even")
        card_bg = _theme_color("bg_card")

        container = tk.Frame(parent, bg=card_bg)
        container.pack(fill="both", expand=True)

        vsb = tk.Scrollbar(container, orient="vertical")
        hsb = tk.Scrollbar(container, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        tree = ttk.Treeview(container, columns=columns, show="headings",
                            style="DP.Treeview",
                            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)
        tree.pack(fill="both", expand=True)

        tree.tag_configure("odd",  background=odd_bg)
        tree.tag_configure("even", background=even_bg)

        tree.bind("<MouseWheel>",
                  lambda e: tree.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        return tree

    def _fill_tree(self, tree, columns, rows):
        for col in columns:
            tree.heading(col, text=col)
            w = max(80, min(200, len(str(col)) * 9 + 20))
            tree.column(col, width=w, minwidth=60, stretch=True)
        for i, row in enumerate(rows):
            tag = "odd" if i % 2 == 0 else "even"
            tree.insert("", "end", values=row, tags=(tag,))

    def _render_sheet(self, parent, df):
        if df.empty:
            ctk.CTkLabel(parent, text="(empty sheet)",
                         font=("Segoe UI", 11),
                         text_color=COLORS["text_muted"]).pack(pady=20)
            return
        raw_cols = [str(v) if str(v) != "nan" else "" for v in df.iloc[0]]
        seen = {}
        columns = []
        for c in raw_cols:
            key = c if c else "Col"
            n = seen.get(key, 0)
            seen[key] = n + 1
            columns.append(c if n == 0 else f"{c}.{n}")
        rows = []
        for _, row in df.iloc[1:].iterrows():
            rows.append([("" if str(v) == "nan" else str(v)) for v in row])
        tree = self._build_tree(parent, columns)
        self._fill_tree(tree, columns, rows)

    def load_excel(self, filepath):
        self._clear()
        try:
            sheets = pd.read_excel(filepath, sheet_name=None, header=None, dtype=str)
        except Exception as e:
            self._show_placeholder(f"Cannot read file: {e}")
            return

        _ALLOWED = {
            "B2B", "B2BA", "B2B-CDNR",   # GSTR-2B govt format
            "ITC B2B", "B2B CDNR", "CDNR" # ITC template format
        }
        relevant = {n: df for n, df in sheets.items()
                    if n.strip().upper() in {a.upper() for a in _ALLOWED}}
        if not relevant:
            # Tally sheet or unknown file — show only first sheet
            first_name, first_df = next(iter(sheets.items()))
            relevant = {first_name: first_df}

        total_rows = sum(max(0, len(df) - 1) for df in relevant.values())
        self.count_label.configure(text=f"{total_rows} rows, {len(relevant)} sheet(s)")

        if len(relevant) == 1:
            name, df = next(iter(relevant.items()))
            self._render_sheet(self._body, df)
        else:
            tabs = ctk.CTkTabview(self._body, fg_color=COLORS["bg_dark"],
                                  segmented_button_fg_color=COLORS["bg_dark"],
                                  segmented_button_selected_color=COLORS["accent"],
                                  segmented_button_unselected_color=COLORS["bg_dark"])
            tabs.pack(fill="both", expand=True)
            for name, df in relevant.items():
                tab = tabs.add(str(name)[:20])
                self._render_sheet(tab, df)

    def load_data(self, records):
        self._clear()
        if not records:
            self._show_placeholder("No data loaded")
            return

        self.count_label.configure(text=f"{len(records)} records")
        columns = ["#", "Date", "Invoice No", "Party Name", "Taxable", "IGST", "CGST", "SGST"]
        rows = []
        for i, rec in enumerate(records):
            rows.append([
                str(i + 1),
                rec.get("invoice_date", ""),
                str(rec.get("invoice_no", ""))[:30],
                str(rec.get("trade_name", ""))[:40],
                f"{rec.get('taxable_value', 0):,.2f}",
                f"{rec.get('igst', 0):,.2f}",
                f"{rec.get('cgst', 0):,.2f}",
                f"{rec.get('sgst', 0):,.2f}",
            ])
        tree = self._build_tree(self._body, columns)
        self._fill_tree(tree, columns, rows)


def _get_unique_path(directory, stem, ext):
    path = os.path.join(directory, f"{stem}{ext}")
    if not os.path.exists(path): return path
    n = 1
    while os.path.exists(os.path.join(directory, f"{stem}({n}){ext}")): n += 1
    return os.path.join(directory, f"{stem}({n}){ext}")


def _parse_tally_import_response(response_text):
    try:
        root = ET.fromstring(response_text)
        line_error = root.find(".//LINEERROR")
        if line_error is not None and line_error.text and line_error.text.strip():
            return {"success": False, "error": line_error.text.strip()}

        def _get_count(tag_name):
            node = root.find(f".//{tag_name}")
            return (node.text or "0") if node is not None else "0"

        return {
            "success": True,
            "created": _get_count("CREATED"),
            "altered": _get_count("ALTERED"),
            "deleted": _get_count("DELETED"),
            "cancelled": _get_count("CANCELLED"),
            "errors": _get_count("ERRORS"),
        }
    except ET.ParseError as exc:
        return {
            "success": False,
            "error": f"Could not parse Tally response: {exc}",
        }


def _check_tally_connection(tally_url, timeout=5):
    test_xml = (
        "<ENVELOPE><HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
        "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
        "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>"
    )
    try:
        resp = requests.post(
            tally_url,
            data=test_xml.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=timeout,
        )
        return {"connected": resp.status_code == 200, "status_code": resp.status_code}
    except requests.exceptions.ConnectionError:
        return {"connected": False, "error": "ConnectionError"}
    except requests.exceptions.Timeout:
        return {"connected": False, "error": "Timeout"}
    except Exception as exc:
        return {"connected": False, "error": str(exc)}


def _post_xml_to_tally(tally_url, xml_content, timeout=30):
    try:
        resp = requests.post(
            tally_url,
            data=xml_content.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=timeout,
        )
        if resp.status_code == 200:
            return _parse_tally_import_response(resp.text)
        return {"success": False, "error": f"HTTP Error: {resp.status_code}"}
    except requests.exceptions.ConnectionError:
        return {
            "success": False,
            "error": (
                "Cannot connect to Tally!\n\n"
                "Please ensure:\n"
                "  - TallyPrime / Tally ERP is running\n"
                "  - A company is loaded in Tally\n"
                "  - HTTP Server is enabled on selected port"
            ),
        }
    except requests.exceptions.Timeout:
        return {"success": False, "error": "Request timed out. Tally is taking too long."}
    except Exception as exc:
        return {"success": False, "error": str(exc)}


def _normalize_company_name(value):
    text = _decode_xml_entities(value)
    text = text.replace("\x00", "")
    text = re.sub(r"[\x01-\x1F\x7F]", " ", text)
    text = "".join(ch for ch in text if ch.isprintable())
    return re.sub(r"\s+", " ", text).strip()


def _company_key(value):
    return _normalize_company_name(value).upper()


def _is_valid_company_name(value):
    name = _normalize_company_name(value)
    if not name:
        return False
    # Tally responses can include numeric placeholders like "0" for current company.
    if re.fullmatch(r"\d+", name):
        return False
    return True


def _fetch_tally_companies(tally_url, timeout=15):
    request_xml_variants = [
        (
            "report-list-companies",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Companies</REPORTNAME>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>",
        ),
        (
            "collection-company",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
            "<TYPE>Collection</TYPE><ID>Company Collection</ID></HEADER><BODY><DESC>"
            "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
            "<TDL><TDLMESSAGE><COLLECTION NAME='Company Collection'>"
            "<TYPE>Company</TYPE><FETCH>Name</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
            "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>",
        ),
    ]

    def _extract_from_response(response_text):
        names = set()
        try:
            root = ET.fromstring(response_text)
            for node in root.iter():
                tag = str(node.tag or "").upper()
                text_name = _normalize_company_name(node.text)
                attr_name = _normalize_company_name(node.attrib.get("NAME") or "")

                if tag in {"COMPANYNAME", "SVCURRENTCOMPANY"} and _is_valid_company_name(text_name):
                    names.add(text_name)
                if "COMPANY" in tag and _is_valid_company_name(attr_name):
                    names.add(attr_name)
                if tag in {"COMPANY", "CURRENTCOMPANY"} and _is_valid_company_name(text_name):
                    names.add(text_name)
        except ET.ParseError:
            pass

        for match in re.findall(r'COMPANY[^>]*NAME="([^"]+)"', response_text, flags=re.IGNORECASE):
            name = _normalize_company_name(match)
            if _is_valid_company_name(name):
                names.add(name)
        for match in re.findall(r"<COMPANYNAME>(.*?)</COMPANYNAME>", response_text, flags=re.IGNORECASE | re.DOTALL):
            name = _normalize_company_name(match)
            if _is_valid_company_name(name):
                names.add(name)
        for match in re.findall(r"<SVCURRENTCOMPANY>(.*?)</SVCURRENTCOMPANY>", response_text, flags=re.IGNORECASE | re.DOTALL):
            name = _normalize_company_name(match)
            if _is_valid_company_name(name):
                names.add(name)
        for match in re.findall(r"<COMPANY[^>]*>.*?<NAME>(.*?)</NAME>", response_text, flags=re.IGNORECASE | re.DOTALL):
            name = _normalize_company_name(match)
            if _is_valid_company_name(name):
                names.add(name)

        return names

    def _post_request(xml_payload):
        resp = requests.post(
            tally_url,
            data=xml_payload.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=timeout,
        )
        if resp.status_code != 200:
            return {"success": False, "error": f"HTTP Error: {resp.status_code}", "text": ""}
        try:
            root = ET.fromstring(resp.text)
            line_error = root.find(".//LINEERROR")
            if line_error is not None and line_error.text and line_error.text.strip():
                return {"success": False, "error": line_error.text.strip(), "text": resp.text}
        except ET.ParseError:
            pass
        return {"success": True, "text": resp.text}

    try:
        companies = set()
        errors = []
        for label, payload in request_xml_variants:
            result = _post_request(payload)
            if result.get("success"):
                companies.update(_extract_from_response(result.get("text") or ""))
            else:
                errors.append(f"{label}: {result.get('error', 'Unknown error')}")

        companies = sorted(companies, key=lambda x: _company_key(x))
        if not companies:
            err_text = "; ".join(errors) if errors else "No companies returned by Tally."
            return {"success": False, "error": err_text, "companies": []}
        return {"success": True, "companies": companies}
    except ET.ParseError as exc:
        return {"success": False, "error": f"Could not parse company list from Tally: {exc}", "companies": []}
    except requests.exceptions.ConnectionError:
        return {"success": False, "error": "ConnectionError", "companies": []}
    except requests.exceptions.Timeout:
        return {"success": False, "error": "Timeout", "companies": []}
    except Exception as exc:
        return {"success": False, "error": str(exc), "companies": []}


def _decode_xml_entities(value):
    text = str(value or "")
    # Some Tally responses are double-escaped (&amp;apos;), so decode iteratively.
    for _ in range(3):
        decoded = html.unescape(text)
        if decoded == text:
            break
        text = decoded
    return text


def _normalize_ledger_name(value):
    text = _decode_xml_entities(value)
    text = text.replace("\x00", "")
    text = re.sub(r"[\x01-\x1F\x7F]", " ", text)
    text = "".join(ch for ch in text if ch.isprintable())
    text = re.sub(r"\s+", " ", text).strip()

    # Remove wrapping quotes repeatedly (e.g. "'Phone'", '&apos;Phone&apos;').
    while len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"', "`"}:
        text = text[1:-1].strip()
    return text


def _ledger_key(value):
    return _normalize_ledger_name(value).upper()


_GST_STATE_CODE_MAP = {
    "01": "Jammu And Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
    "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
    "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
    "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
    "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
    "24": "Gujarat", "25": "Daman And Diu", "26": "Dadra And Nagar Haveli And Daman And Diu",
    "27": "Maharashtra", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala",
    "33": "Tamil Nadu", "34": "Puducherry", "35": "Andaman And Nicobar Islands", "36": "Telangana",
    "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory", "99": "Centre Jurisdiction",
}

LEDGER_STATE_OPTIONS = sorted(set(_GST_STATE_CODE_MAP.values()), key=lambda x: x.upper())
LEDGER_COUNTRY_OPTIONS = ["India"]
LEDGER_GST_APPLICABLE_OPTIONS = ["Applicable", "Not Applicable"]


def _state_name_from_gstin(gstin):
    gstin_text = _normalize_ledger_name(gstin).upper()
    code = gstin_text[:2]
    return _GST_STATE_CODE_MAP.get(code, "")


def _pan_from_gstin(gstin: str) -> str:
    """Extract 10-character PAN from a 15-character GSTIN (chars 2–11 inclusive)."""
    g = _normalize_ledger_name(gstin).upper()
    return g[2:12] if len(g) == 15 else ""


def _set_svcurrentcompany(xml_content, company_name):
    target = _normalize_company_name(company_name)
    if not target:
        return xml_content, False

    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError:
        escaped_target = html.escape(target)
        if re.search(r"<SVCURRENTCOMPANY\b[^>]*>.*?</SVCURRENTCOMPANY>", xml_content, flags=re.IGNORECASE | re.DOTALL):
            updated = re.sub(
                r"<SVCURRENTCOMPANY\b[^>]*>.*?</SVCURRENTCOMPANY>",
                f"<SVCURRENTCOMPANY>{escaped_target}</SVCURRENTCOMPANY>",
                xml_content,
                flags=re.IGNORECASE | re.DOTALL,
            )
            return updated, True

        updated = re.sub(
            r"(<STATICVARIABLES\b[^>]*>)",
            r"\1" + f"<SVCURRENTCOMPANY>{escaped_target}</SVCURRENTCOMPANY>",
            xml_content,
            count=1,
            flags=re.IGNORECASE,
        )
        if updated != xml_content:
            return updated, True
        return xml_content, False

    request_desc = root.find(".//REQUESTDESC")
    if request_desc is None:
        return xml_content, False

    static_vars = request_desc.find("STATICVARIABLES")
    if static_vars is None:
        static_vars = ET.SubElement(request_desc, "STATICVARIABLES")

    svc_node = static_vars.find("SVCURRENTCOMPANY")
    if svc_node is None:
        svc_node = ET.SubElement(static_vars, "SVCURRENTCOMPANY")
    svc_node.text = target
    return ET.tostring(root, encoding="unicode"), True


def _extract_missing_ledger_names(error_text):
    text = _decode_xml_entities(error_text)
    patterns = [
        r"ledger\s+'([^']+)'\s+does\s+not\s+exist",
        r'ledger\s+"([^"]+)"\s+does\s+not\s+exist',
        r"ledger\s+([^\n\r]+?)\s+does\s+not\s+exist",
    ]
    found = []
    seen = set()
    for pattern in patterns:
        for match in re.findall(pattern, text, flags=re.IGNORECASE):
            name = _normalize_ledger_name(str(match).strip().strip("! .,:;\"'"))
            if not name:
                continue
            key = _ledger_key(name)
            if key not in seen:
                seen.add(key)
                found.append(name)
    return found


def _extract_ledger_usage_from_xml(xml_content):
    usage = {}
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError:
        return usage

    def add_usage(ledger_name, is_party=False, extra=None):
        name = _normalize_ledger_name(ledger_name)
        if not name:
            return
        rec = usage.setdefault(name, {"is_party": False, "count": 0, "extra": {}})
        rec["count"] += 1
        if is_party:
            rec["is_party"] = True
        if extra:
            for key, value in extra.items():
                clean_value = _normalize_ledger_name(value)
                if clean_value and not rec["extra"].get(key):
                    rec["extra"][key] = clean_value

    for voucher in root.iter("VOUCHER"):
        party_extra = {}
        party_gstin = voucher.findtext("PARTYGSTIN")
        party_state = voucher.findtext("STATENAME")
        party_pin = voucher.findtext("PARTYPINCODE")
        party_country = voucher.findtext("COUNTRYOFRESIDENCE")
        party_mailing = voucher.findtext("PARTYMAILINGNAME")

        if party_gstin:
            party_extra["gstin"] = party_gstin
        if party_state:
            party_extra["state"] = party_state
        if party_pin:
            party_extra["pincode"] = party_pin
        if party_country:
            party_extra["country"] = party_country
        if party_mailing:
            party_extra["mailing_name"] = party_mailing

        address_nodes = voucher.findall("./ADDRESS.LIST/ADDRESS")
        if address_nodes:
            address_lines = [_normalize_ledger_name(node.text) for node in address_nodes if _normalize_ledger_name(node.text)]
            if address_lines:
                party_extra["address1"] = address_lines[0]
            if len(address_lines) > 1:
                party_extra["address2"] = address_lines[1]

        add_usage(voucher.findtext("PARTYLEDGERNAME"), is_party=True, extra=party_extra)
        for led_row in voucher.findall(".//LEDGERENTRIES.LIST"):
            ledger_name = led_row.findtext("LEDGERNAME")
            party_flag = str(led_row.findtext("ISPARTYLEDGER") or "").strip().upper() == "YES"
            add_usage(ledger_name, is_party=party_flag)

    # Fallback for non-standard XML structures.
    if not usage:
        for node in root.iter("LEDGERNAME"):
            add_usage(node.text, is_party=False)
        for node in root.iter("PARTYLEDGERNAME"):
            add_usage(node.text, is_party=True)

    return usage


def _extract_ledger_names_from_xml(xml_content):
    usage = _extract_ledger_usage_from_xml(xml_content)
    if usage:
        return sorted(usage.keys(), key=lambda x: _ledger_key(x))
    return []


def _fetch_tally_ledgers(tally_url, timeout=15, company_name=""):
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{html.escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml_variants = [
        (
            "collection-ledger",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
            "<TYPE>Collection</TYPE><ID>Ledger Collection</ID></HEADER>"
            f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE><COLLECTION NAME='Ledger Collection'>"
            "<TYPE>Ledger</TYPE><FETCH>Name,Parent</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
            "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
        ),
        (
            "report-list-ledgers",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Ledgers</REPORTNAME>"
            f"{static_vars}</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>"
        ),
        (
            "report-list-accounts",
            "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>"
            "<BODY><EXPORTDATA><REQUESTDESC><REPORTNAME>List of Accounts</REPORTNAME>"
            f"{static_vars}</REQUESTDESC></EXPORTDATA></BODY></ENVELOPE>"
        ),
    ]

    def _extract_from_response(response_text):
        names = set()
        try:
            root = ET.fromstring(response_text)
            for node in root.iter():
                tag = str(node.tag or "").upper()
                text_name = _normalize_ledger_name(node.text)
                if tag == "LEDGER":
                    attr_name = _normalize_ledger_name(node.attrib.get("NAME") or "")
                    if attr_name:
                        names.add(attr_name)
                    if text_name:
                        names.add(text_name)
                elif tag in {"LEDGERNAME", "PARTYLEDGERNAME", "DSPLEDGERNAME", "DSPACCNAME"}:
                    if text_name:
                        names.add(text_name)
                elif tag.endswith("LEDGERNAME") and text_name:
                    names.add(text_name)
        except ET.ParseError:
            pass

        # Text fallback for unusual Tally layouts.
        for match in re.findall(r'LEDGER[^>]*NAME="([^"]+)"', response_text, flags=re.IGNORECASE):
            name = _normalize_ledger_name(match)
            if name:
                names.add(name)
        for match in re.findall(r"<LEDGER[^>]*>.*?<NAME>(.*?)</NAME>", response_text, flags=re.IGNORECASE | re.DOTALL):
            name = _normalize_ledger_name(match)
            if name:
                names.add(name)
        return names

    def _post_request(xml_payload):
        resp = requests.post(
            tally_url,
            data=xml_payload.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=timeout,
        )
        if resp.status_code != 200:
            return {"success": False, "error": f"HTTP Error: {resp.status_code}", "text": ""}
        try:
            root = ET.fromstring(resp.text)
            line_error = root.find(".//LINEERROR")
            if line_error is not None and line_error.text and line_error.text.strip():
                return {"success": False, "error": line_error.text.strip(), "text": resp.text}
        except ET.ParseError:
            # Keep processing with regex fallbacks.
            pass
        return {"success": True, "text": resp.text}

    try:
        ledgers = set()
        errors = []
        for label, payload in request_xml_variants:
            result = _post_request(payload)
            if result.get("success"):
                ledgers.update(_extract_from_response(result.get("text") or ""))
            else:
                errors.append(f"{label}: {result.get('error', 'Unknown error')}")

        ledgers = sorted(ledgers, key=lambda x: _ledger_key(x))
        if not ledgers:
            err_text = "; ".join(errors) if errors else "No ledgers returned by Tally."
            return {"success": False, "error": err_text, "ledgers": []}
        return {"success": True, "ledgers": ledgers}
    except ET.ParseError as exc:
        return {"success": False, "error": f"Could not parse ledger list from Tally: {exc}", "ledgers": []}
    except requests.exceptions.ConnectionError:
        return {"success": False, "error": "ConnectionError", "ledgers": []}
    except requests.exceptions.Timeout:
        return {"success": False, "error": "Timeout", "ledgers": []}
    except Exception as exc:
        return {"success": False, "error": str(exc), "ledgers": []}


def _fetch_stock_items_from_tally(tally_url, timeout=15, company_name=""):
    """Return sorted list of stock item names from Tally."""
    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{html.escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>Stock Item Collection</ID></HEADER>"
        f"<BODY><DESC>{static_vars}<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='Stock Item Collection'>"
        "<TYPE>Stock Item</TYPE><FETCH>Name</FETCH><NATIVEMETHOD>Name</NATIVEMETHOD>"
        "</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )

    try:
        resp = requests.post(
            tally_url,
            data=request_xml.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=timeout,
        )
        if resp.status_code != 200:
            return {"success": False, "error": f"HTTP Error: {resp.status_code}", "items": []}

        names = set()
        try:
            root = ET.fromstring(resp.text)
            for node in root.iter():
                tag = str(node.tag or "").upper()
                text_name = _normalize_ledger_name(node.text)
                if tag == "STOCKITEM":
                    attr_name = _normalize_ledger_name(node.attrib.get("NAME") or "")
                    if attr_name:
                        names.add(attr_name)
                    if text_name:
                        names.add(text_name)
                elif tag in {"STOCKITEMNAME", "DSPSTOCKITEMNAME", "NAME"}:
                    if text_name:
                        names.add(text_name)
        except ET.ParseError:
            pass

        for match in re.findall(r'STOCKITEM[^>]*NAME="([^"]+)"', resp.text, flags=re.IGNORECASE):
            name = _normalize_ledger_name(match)
            if name:
                names.add(name)

        names = sorted(names)
        if not names:
            return {"success": False, "error": "No stock items returned by Tally.", "items": []}
        return {"success": True, "items": names}
    except requests.exceptions.ConnectionError:
        return {"success": False, "error": "ConnectionError", "items": []}
    except requests.exceptions.Timeout:
        return {"success": False, "error": "Timeout", "items": []}
    except Exception as exc:
        return {"success": False, "error": str(exc), "items": []}


def _fetch_tally_party_details_by_gstin(tally_url, gstin, timeout=15, company_name=""):
    target_gstin = _normalize_ledger_name(gstin).upper()
    if not target_gstin or len(target_gstin) < 15:
        return {"success": False, "error": "Invalid GSTIN", "details": {}}

    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{html.escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    formula_text = html.escape(f'$PartyGSTIN = "{target_gstin}"')
    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>GSTIN Ledger Lookup</ID></HEADER>"
        "<BODY><DESC>"
        f"{static_vars}"
        "<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='GSTIN Ledger Lookup'>"
        "<TYPE>Ledger</TYPE>"
        "<FETCH>Name,GSTApplicable,PartyGSTIN,GSTIN,GSTREGISTRATIONTYPE,StateName,PriorStateName,Pincode,CountryOfResidence,MailingName,Address,ISBILLWISEON</FETCH>"
        "<FILTERS>ByGSTIN</FILTERS>"
        "</COLLECTION>"
        f"<SYSTEM TYPE='Formulae' NAME='ByGSTIN'>{formula_text}</SYSTEM>"
        "</TDLMESSAGE></TDL>"
        "</DESC></BODY></ENVELOPE>"
    )

    def _extract_ledger_details_from_node(ledger):
        address_lines = []
        for path in [
            "./LEDMAILINGDETAILS.LIST/ADDRESS.LIST/ADDRESS",
            "./ADDRESS.LIST/ADDRESS",
        ]:
            nodes = ledger.findall(path)
            for node in nodes:
                value = _normalize_ledger_name(node.text)
                if value:
                    address_lines.append(value)
            if address_lines:
                break

        name_text = _normalize_ledger_name(ledger.attrib.get("NAME") or ledger.findtext("NAME") or "")
        gstin_text = _normalize_ledger_name(
            ledger.findtext("PARTYGSTIN")
            or ledger.findtext("GSTIN")
            or ledger.findtext("./LEDGSTREGDETAILS.LIST/GSTIN")
            or ""
        ).upper()
        state_text = _normalize_ledger_name(
            ledger.findtext("STATENAME")
            or ledger.findtext("PRIORSTATENAME")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/STATE")
            or ledger.findtext("./LEDGSTREGDETAILS.LIST/PLACEOFSUPPLY")
            or ""
        )
        country_text = _normalize_ledger_name(
            ledger.findtext("COUNTRYOFRESIDENCE")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/COUNTRY")
            or ""
        )
        pincode_text = _normalize_ledger_name(
            ledger.findtext("PINCODE")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/PINCODE")
            or ""
        )
        mailing_name_text = _normalize_ledger_name(
            ledger.findtext("MAILINGNAME")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/MAILINGNAME")
            or name_text
            or ""
        )
        reg_type_text = _normalize_ledger_name(
            ledger.findtext("GSTREGISTRATIONTYPE")
            or ledger.findtext("REGISTRATIONTYPE")
            or ledger.findtext("LEDGERREGISTRATIONTYPE")
            or ledger.findtext("./LEDGSTREGDETAILS.LIST/GSTREGISTRATIONTYPE")
            or ""
        )
        gst_app_text = _normalize_ledger_name(
            ledger.findtext("GSTAPPLICABLE")
            or ledger.findtext("ISGSTAPPLICABLE")
            or ""
        )
        if gst_app_text:
            if gst_app_text.upper() in {"YES", "Y", "TRUE", "1", "APPLICABLE", "GST APPLICABLE"}:
                gst_app_text = "Applicable"
            elif gst_app_text.upper() in {"NO", "N", "FALSE", "0", "NOT APPLICABLE", "NA", "N/A"}:
                gst_app_text = "Not Applicable"
        else:
            gst_app_text = "Applicable" if gstin_text else "Not Applicable"
        billwise_text = _normalize_ledger_name(ledger.findtext("ISBILLWISEON") or "")
        if billwise_text:
            billwise_text = "Yes" if billwise_text.upper() in {"YES", "Y", "TRUE", "1"} else "No"
        else:
            billwise_text = "Yes"

        return {
            "mailing_name": mailing_name_text,
            "address1": address_lines[0] if address_lines else "",
            "address2": address_lines[1] if len(address_lines) > 1 else "",
            "state": state_text,
            "country": country_text,
            "pincode": pincode_text,
            "gstin": gstin_text,
            "gst_applicable": gst_app_text,
            "reg_type": reg_type_text or ("Regular" if gstin_text else "Unknown"),
            "billwise": billwise_text,
            "name": name_text,
        }

    try:
        resp = requests.post(
            tally_url,
            data=request_xml.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=timeout,
        )
        if resp.status_code != 200:
            return {"success": False, "error": f"HTTP Error: {resp.status_code}", "details": {}}

        root = ET.fromstring(resp.text)
        for ledger in root.iter("LEDGER"):
            parsed = _extract_ledger_details_from_node(ledger)
            found_gstin = _normalize_ledger_name(parsed.get("gstin") or "").upper()
            if found_gstin != target_gstin:
                continue

            return {
                "success": True,
                "details": parsed,
            }

        return {"success": False, "error": "GSTIN not found in existing Tally ledgers.", "details": {}}
    except ET.ParseError as exc:
        return {"success": False, "error": f"Could not parse Tally response: {exc}", "details": {}}
    except requests.exceptions.ConnectionError:
        return {"success": False, "error": "ConnectionError", "details": {}}
    except requests.exceptions.Timeout:
        return {"success": False, "error": "Timeout", "details": {}}
    except Exception as exc:
        return {"success": False, "error": str(exc), "details": {}}


def _fetch_tally_ledger_details_by_name(tally_url, ledger_name, timeout=15, company_name=""):
    target_name = _normalize_ledger_name(ledger_name)
    if not target_name:
        return {"success": False, "error": "Ledger name missing", "details": {}}

    selected_company = _normalize_company_name(company_name)
    static_vars = "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
    if selected_company:
        static_vars += f"<SVCURRENTCOMPANY>{html.escape(selected_company)}</SVCURRENTCOMPANY>"
    static_vars += "</STATICVARIABLES>"

    formula_text = html.escape(f'$Name = "{target_name}"')
    request_xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>Name Ledger Lookup</ID></HEADER>"
        "<BODY><DESC>"
        f"{static_vars}"
        "<TDL><TDLMESSAGE>"
        "<COLLECTION NAME='Name Ledger Lookup'>"
        "<TYPE>Ledger</TYPE>"
        "<FETCH>Name,GSTApplicable,PartyGSTIN,GSTIN,GSTREGISTRATIONTYPE,StateName,PriorStateName,Pincode,CountryOfResidence,MailingName,Address,ISBILLWISEON</FETCH>"
        "<FILTERS>ByName</FILTERS>"
        "</COLLECTION>"
        f"<SYSTEM TYPE='Formulae' NAME='ByName'>{formula_text}</SYSTEM>"
        "</TDLMESSAGE></TDL>"
        "</DESC></BODY></ENVELOPE>"
    )

    def _extract_ledger_details_from_node(ledger):
        address_lines = []
        for path in [
            "./LEDMAILINGDETAILS.LIST/ADDRESS.LIST/ADDRESS",
            "./ADDRESS.LIST/ADDRESS",
        ]:
            nodes = ledger.findall(path)
            for node in nodes:
                value = _normalize_ledger_name(node.text)
                if value:
                    address_lines.append(value)
            if address_lines:
                break

        name_text = _normalize_ledger_name(ledger.attrib.get("NAME") or ledger.findtext("NAME") or "")
        gstin_text = _normalize_ledger_name(
            ledger.findtext("PARTYGSTIN")
            or ledger.findtext("GSTIN")
            or ledger.findtext("./LEDGSTREGDETAILS.LIST/GSTIN")
            or ""
        ).upper()
        state_text = _normalize_ledger_name(
            ledger.findtext("STATENAME")
            or ledger.findtext("PRIORSTATENAME")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/STATE")
            or ledger.findtext("./LEDGSTREGDETAILS.LIST/PLACEOFSUPPLY")
            or ""
        )
        country_text = _normalize_ledger_name(
            ledger.findtext("COUNTRYOFRESIDENCE")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/COUNTRY")
            or ""
        )
        pincode_text = _normalize_ledger_name(
            ledger.findtext("PINCODE")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/PINCODE")
            or ""
        )
        mailing_name_text = _normalize_ledger_name(
            ledger.findtext("MAILINGNAME")
            or ledger.findtext("./LEDMAILINGDETAILS.LIST/MAILINGNAME")
            or name_text
            or ""
        )
        reg_type_text = _normalize_ledger_name(
            ledger.findtext("GSTREGISTRATIONTYPE")
            or ledger.findtext("REGISTRATIONTYPE")
            or ledger.findtext("LEDGERREGISTRATIONTYPE")
            or ledger.findtext("./LEDGSTREGDETAILS.LIST/GSTREGISTRATIONTYPE")
            or ""
        )
        gst_app_text = _normalize_ledger_name(
            ledger.findtext("GSTAPPLICABLE")
            or ledger.findtext("ISGSTAPPLICABLE")
            or ""
        )
        if gst_app_text:
            if gst_app_text.upper() in {"YES", "Y", "TRUE", "1", "APPLICABLE", "GST APPLICABLE"}:
                gst_app_text = "Applicable"
            elif gst_app_text.upper() in {"NO", "N", "FALSE", "0", "NOT APPLICABLE", "NA", "N/A"}:
                gst_app_text = "Not Applicable"
        else:
            gst_app_text = "Applicable" if gstin_text else "Not Applicable"
        billwise_text = _normalize_ledger_name(ledger.findtext("ISBILLWISEON") or "")
        if billwise_text:
            billwise_text = "Yes" if billwise_text.upper() in {"YES", "Y", "TRUE", "1"} else "No"
        else:
            billwise_text = "Yes"

        return {
            "mailing_name": mailing_name_text,
            "address1": address_lines[0] if address_lines else "",
            "address2": address_lines[1] if len(address_lines) > 1 else "",
            "state": state_text,
            "country": country_text,
            "pincode": pincode_text,
            "gstin": gstin_text,
            "gst_applicable": gst_app_text,
            "reg_type": reg_type_text or ("Regular" if gstin_text else "Unknown"),
            "billwise": billwise_text,
            "name": name_text,
        }

    try:
        resp = requests.post(
            tally_url,
            data=request_xml.encode("utf-8"),
            headers={"Content-Type": "application/xml"},
            timeout=timeout,
        )
        if resp.status_code != 200:
            return {"success": False, "error": f"HTTP Error: {resp.status_code}", "details": {}}

        root = ET.fromstring(resp.text)
        for ledger in root.iter("LEDGER"):
            parsed = _extract_ledger_details_from_node(ledger)
            parsed_name = _normalize_ledger_name(parsed.get("name") or "")
            if _ledger_key(parsed_name) == _ledger_key(target_name):
                return {"success": True, "details": parsed}

        return {"success": False, "error": "Ledger not found by name.", "details": {}}
    except ET.ParseError as exc:
        return {"success": False, "error": f"Could not parse Tally response: {exc}", "details": {}}
    except requests.exceptions.ConnectionError:
        return {"success": False, "error": "ConnectionError", "details": {}}
    except requests.exceptions.Timeout:
        return {"success": False, "error": "Timeout", "details": {}}
    except Exception as exc:
        return {"success": False, "error": str(exc), "details": {}}


def _guess_parent_for_ledger(ledger_name, usage_meta=None):
    usage_meta = usage_meta or {}
    name = _normalize_ledger_name(ledger_name)
    upper_name = name.upper()
    if usage_meta.get("is_party"):
        return "Sundry Creditors"
    if upper_name in {"IGST", "CGST", "SGST", "CESS"} or "GST" in upper_name or "TDS" in upper_name:
        return "Duties & Taxes"
    if "PURCHASE" in upper_name:
        return "Purchase Accounts"
    return "Purchase Accounts"


def _create_tally_ledger(tally_url, ledger_name, parent_name, timeout=30, is_party=False, extra_info=None, company_name=""):
    extra_info = extra_info or {}
    name = _normalize_ledger_name(ledger_name)
    parent = _normalize_ledger_name(parent_name) or "Purchase Accounts"
    if not name:
        return {"success": False, "error": "Ledger name cannot be empty."}

    today = datetime.date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    applicable_from = f"{fy_start_year}0401"

    country_name = _normalize_ledger_name(extra_info.get("country") or "") or "India"
    state_name = _normalize_ledger_name(extra_info.get("state") or "")
    pincode = _normalize_ledger_name(extra_info.get("pincode") or "")
    gstin = _normalize_ledger_name(extra_info.get("gstin") or "").upper()
    if not state_name and gstin:
        state_name = _state_name_from_gstin(gstin)
    mailing_name = _normalize_ledger_name(extra_info.get("mailing_name") or "") or name
    address1 = _normalize_ledger_name(extra_info.get("address1") or "")
    address2 = _normalize_ledger_name(extra_info.get("address2") or "")
    pan = (
        _normalize_ledger_name(extra_info.get("pan") or "").upper()
        or _pan_from_gstin(gstin)
    )

    parent_key = _ledger_key(parent)
    is_party_ledger = bool(is_party) or parent_key in {"SUNDRY DEBTORS", "SUNDRY CREDITORS"}

    gst_app_raw = _normalize_ledger_name(
        extra_info.get("gst_applicable")
        or extra_info.get("gst_app")
        or extra_info.get("gst")
        or ""
    )
    gst_app_key = gst_app_raw.casefold()
    if gst_app_key in {"applicable", "yes", "y", "true", "1", "registered", "regular", "gst applicable"}:
        gst_applicable = "Applicable"
    elif gst_app_key in {"not applicable", "no", "n", "false", "0", "na", "n/a", "notapplicable"}:
        gst_applicable = "Not Applicable"
    else:
        gst_applicable = "Applicable" if gstin else "Not Applicable"

    reg_type = _normalize_ledger_name(extra_info.get("reg_type") or "")
    reg_key = reg_type.casefold()
    reg_map = {
        "regular": "Regular",
        "registered": "Regular",
        "composition": "Composition",
        "consumer": "Consumer",
        "unregistered": "Unregistered",
        "sez": "SEZ",
        "sez unit": "SEZ",
        "sez developer": "SEZ",
        "overseas": "Overseas",
    }
    if reg_key in reg_map:
        reg_type = reg_map[reg_key]
    elif not reg_type:
        reg_type = "Regular" if (gstin or gst_applicable == "Applicable") else ""

    billwise_raw = _normalize_ledger_name(extra_info.get("billwise") or "")
    if billwise_raw:
        billwise_on = billwise_raw.strip().upper() in {"YES", "Y", "TRUE", "1"}
    else:
        billwise_on = bool(is_party_ledger)

    envelope = ET.Element("ENVELOPE")
    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"
    body = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")
    req_desc = ET.SubElement(import_data, "REQUESTDESC")
    ET.SubElement(req_desc, "REPORTNAME").text = "All Masters"

    target_company = _normalize_company_name(company_name)
    if target_company:
        static_vars = ET.SubElement(req_desc, "STATICVARIABLES")
        ET.SubElement(static_vars, "SVCURRENTCOMPANY").text = target_company

    req_data = ET.SubElement(import_data, "REQUESTDATA")
    tally_msg = ET.SubElement(req_data, "TALLYMESSAGE")
    tally_msg.set("xmlns:UDF", "TallyUDF")

    ledger = ET.SubElement(tally_msg, "LEDGER")
    ledger.set("NAME", name)
    ledger.set("RESERVEDNAME", "")
    ledger.set("ACTION", "Create")
    ET.SubElement(ledger, "NAME").text = name
    ET.SubElement(ledger, "PARENT").text = parent
    ET.SubElement(ledger, "ISBILLWISEON").text = "Yes" if billwise_on else "No"
    ET.SubElement(ledger, "ISCOSTCENTRESON").text = "No"
    ET.SubElement(ledger, "ISINTERESTON").text = "No"
    ET.SubElement(ledger, "ALLOWINMOBILE").text = "No"
    ET.SubElement(ledger, "ISUPDATINGTARGETID").text = "No"
    ET.SubElement(ledger, "ASORIGINAL").text = "Yes"
    ET.SubElement(ledger, "AFFECTSSTOCK").text = "No"
    ET.SubElement(ledger, "CURRENCYNAME").text = "INR"
    ET.SubElement(ledger, "COUNTRYOFRESIDENCE").text = country_name

    if is_party_ledger:
        ET.SubElement(ledger, "GSTAPPLICABLE").text = gst_applicable
        if reg_type:
            ET.SubElement(ledger, "GSTREGISTRATIONTYPE").text = reg_type
        if gstin:
            ET.SubElement(ledger, "PARTYGSTIN").text = gstin

    if state_name:
        ET.SubElement(ledger, "PRIORSTATENAME").text = state_name
        if is_party_ledger:
            ET.SubElement(ledger, "LEDSTATENAME").text = state_name

    language_list = ET.SubElement(ledger, "LANGUAGENAME.LIST")
    name_list = ET.SubElement(language_list, "NAME.LIST")
    name_list.set("TYPE", "String")
    ET.SubElement(name_list, "NAME").text = name
    ET.SubElement(language_list, "LANGUAGEID").text = "1033"

    if is_party_ledger and (gstin or reg_type):
        gst_list = ET.SubElement(ledger, "LEDGSTREGDETAILS.LIST")
        ET.SubElement(gst_list, "APPLICABLEFROM").text = applicable_from
        if reg_type:
            ET.SubElement(gst_list, "GSTREGISTRATIONTYPE").text = reg_type
        if state_name:
            ET.SubElement(gst_list, "PLACEOFSUPPLY").text = state_name
        if gstin:
            ET.SubElement(gst_list, "GSTIN").text = gstin
        ET.SubElement(gst_list, "ISOTHTERRITORYASSESSEE").text = "No"
        ET.SubElement(gst_list, "CONSIDERPURCHASEFOREXPORT").text = "No"
        ET.SubElement(gst_list, "ISTRANSPORTER").text = "No"
        ET.SubElement(gst_list, "ISCOMMONPARTY").text = "No"

    if is_party_ledger and (address1 or address2 or state_name or country_name or pincode or pan):
        mailing_list = ET.SubElement(ledger, "LEDMAILINGDETAILS.LIST")
        if address1 or address2:
            addr_list = ET.SubElement(mailing_list, "ADDRESS.LIST")
            addr_list.set("TYPE", "String")
            if address1:
                ET.SubElement(addr_list, "ADDRESS").text = address1
            if address2:
                ET.SubElement(addr_list, "ADDRESS").text = address2
        ET.SubElement(mailing_list, "APPLICABLEFROM").text = applicable_from
        if pincode:
            ET.SubElement(mailing_list, "PINCODE").text = pincode
        ET.SubElement(mailing_list, "MAILINGNAME").text = mailing_name
        if pan:
            ET.SubElement(mailing_list, "INCOMETAXNUMBER").text = pan
        if state_name:
            ET.SubElement(mailing_list, "STATE").text = state_name
        ET.SubElement(mailing_list, "COUNTRY").text = country_name

    xml_content = ET.tostring(envelope, encoding="unicode")
    result = _post_xml_to_tally(tally_url, xml_content, timeout=timeout)

    if result.get("success"):
        try:
            created_count = int(float(result.get("created", "0") or "0"))
        except (ValueError, TypeError):
            created_count = 0
        try:
            altered_count = int(float(result.get("altered", "0") or "0"))
        except (ValueError, TypeError):
            altered_count = 0

        if created_count > 0 or altered_count > 0:
            return result

        verify = _fetch_tally_ledger_details_by_name(
            tally_url,
            name,
            timeout=min(timeout, 20),
            company_name=company_name,
        )
        if verify.get("success"):
            return {
                "success": True,
                "created": "0",
                "altered": "1",
                "errors": result.get("errors", "0"),
                "note": "verified-by-fetch",
                "details": verify.get("details") or {},
            }

        return {
            "success": False,
            "error": (
                "Tally did not confirm ledger creation. "
                f"Created={result.get('created', '0')}, "
                f"Altered={result.get('altered', '0')}, "
                f"Errors={result.get('errors', '0')}."
                + (
                    f" Verification also failed: {verify.get('error')}"
                    if verify.get("error")
                    else ""
                )
            ),
        }

    if not result.get("success"):
        err_text = str(result.get("error") or "").lower()
        if "already exists" in err_text:
            return {
                "success": True,
                "created": "0",
                "altered": "1",
                "errors": "0",
                "note": "already-exists",
            }
    return result


def _apply_ledger_mapping_to_xml(xml_content, ledger_mapping):
    mapping_exact = {}
    mapping_by_key = {}
    for src, dst in (ledger_mapping or {}).items():
        src_name = _normalize_ledger_name(src)
        dst_name = _normalize_ledger_name(dst)
        if not src_name or not dst_name:
            continue
        mapping_exact[src_name] = dst_name
        mapping_by_key[_ledger_key(src_name)] = dst_name

    if not mapping_exact:
        return xml_content, 0

    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError:
        replaced = 0
        updated = xml_content
        for src, dst in mapping_exact.items():
            token_old = f">{src}<"
            token_new = f">{dst}<"
            count = updated.count(token_old)
            if count:
                updated = updated.replace(token_old, token_new)
                replaced += count

            # Handle escaped apostrophe form in raw-string fallback.
            escaped_src = src.replace("'", "&apos;")
            escaped_dst = dst.replace("'", "&apos;")
            token_old_esc = f">{escaped_src}<"
            token_new_esc = f">{escaped_dst}<"
            count_esc = updated.count(token_old_esc)
            if count_esc:
                updated = updated.replace(token_old_esc, token_new_esc)
                replaced += count_esc
        return updated, replaced

    replaced = 0
    target_tags = {
        "LEDGERNAME",
        "PARTYLEDGERNAME",
        "PARTYNAME",
        "BASICBASEPARTYNAME",
        "PARTYMAILINGNAME",
    }
    for node in root.iter():
        tag = str(node.tag or "").upper()
        if tag not in target_tags:
            continue
        current = str(node.text or "").strip()
        if not current:
            continue
        normalized_current = _normalize_ledger_name(current)
        replacement = mapping_exact.get(normalized_current)
        if replacement is None:
            replacement = mapping_by_key.get(_ledger_key(normalized_current))
        if replacement is not None:
            node.text = replacement
            replaced += 1

    return ET.tostring(root, encoding="unicode"), replaced


def _build_auto_party_mapping_from_sheet(usage_map, party_ledger_map, existing_ledger_keys):
    """Build PARTYLEDGER remap from mapping sheet when mapped target exists in Tally."""
    remap = {}
    if not usage_map or not party_ledger_map or not existing_ledger_keys:
        return remap

    usage_party_by_key = {}
    for ledger_name, meta in (usage_map or {}).items():
        if not isinstance(meta, dict) or not meta.get("is_party"):
            continue
        normalized = _normalize_ledger_name(ledger_name)
        if not normalized:
            continue
        key = _ledger_key(normalized)
        if key and key not in usage_party_by_key:
            usage_party_by_key[key] = normalized

    for raw_party, raw_target in (party_ledger_map or {}).items():
        party_key = _ledger_key(raw_party)
        src_name = usage_party_by_key.get(party_key)
        if not src_name:
            continue

        target_name = _normalize_ledger_name(raw_target)
        if not target_name:
            continue
        target_key = _ledger_key(target_name)
        if not target_key or target_key == party_key:
            continue
        if target_key in existing_ledger_keys:
            remap[src_name] = target_name

    return remap


def _collect_missing_ledgers_from_usage(usage_map, existing_ledger_keys):
    missing = []
    seen = set()
    for ledger_name in (usage_map or {}).keys():
        normalized = _normalize_ledger_name(ledger_name)
        if not normalized:
            continue
        key = _ledger_key(normalized)
        if not key or key in existing_ledger_keys or key in seen:
            continue
        seen.add(key)
        missing.append(normalized)
    return sorted(missing, key=lambda x: _ledger_key(x))


def _post_xml_with_fallbacks(
    tally_url,
    xml_content,
    timeout=30,
    allow_company_fallback=True,
    allow_date_retry=True,
):
    meta = {
        "forced_date_count": 0,
        "date_retry_used": False,
        "fallback_used": False,
    }
    working_xml = xml_content
    result = _post_xml_to_tally(tally_url, working_xml, timeout=timeout)

    err_text = str(result.get("error", ""))
    if allow_date_retry and (not result.get("success")) and ("voucher date is missing" in err_text.lower()):
        retry_xml, retry_count = _force_voucher_dates_to_today_regex(working_xml)
        retry_result = _post_xml_to_tally(tally_url, retry_xml, timeout=timeout)
        working_xml = retry_xml
        meta["forced_date_count"] = max(meta["forced_date_count"], retry_count)
        if retry_result.get("success"):
            meta["date_retry_used"] = True
        result = retry_result

    err_text = str(result.get("error", ""))
    if allow_company_fallback and (not result.get("success")) and ("svcurrentcompany" in err_text.lower()):
        retry_xml, changed = _remove_svcurrentcompany(working_xml)
        if changed:
            retry_result = _post_xml_to_tally(tally_url, retry_xml, timeout=timeout)
            working_xml = retry_xml
            if retry_result.get("success"):
                meta["fallback_used"] = True
            result = retry_result

    return result, working_xml, meta


def _remove_svcurrentcompany(xml_content):
    """Remove SVCURRENTCOMPANY so import targets currently loaded Tally company."""
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError:
        return xml_content, False

    removed = False
    for parent in root.iter():
        children = list(parent)
        for child in children:
            if child.tag == "SVCURRENTCOMPANY":
                parent.remove(child)
                removed = True

    return ET.tostring(root, encoding="unicode"), removed


def _read_xml_text_safely(xml_path):
    with open(xml_path, "rb") as f:
        raw = f.read()
    if not raw:
        return ""
    for enc in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    return text.replace("\x00", "")


def _normalize_manual_date_to_tally(date_text):
    text = str(date_text or "").strip()
    if not text:
        raise ValueError("Custom date is empty.")

    compact = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\d{8}", compact):
        for fmt in ("%Y%m%d", "%d%m%Y"):
            try:
                return datetime.datetime.strptime(compact, fmt).strftime("%Y%m%d")
            except ValueError:
                continue

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            continue

    raise ValueError("Invalid custom date format. Use DD/MM/YYYY, DD-MM-YYYY, or YYYY-MM-DD.")


def _force_voucher_dates_to_value_regex(xml_content, tally_date):
    target = str(tally_date or "").strip()
    if not (target.isdigit() and len(target) == 8):
        return xml_content, 0

    pattern = re.compile(r"(<VOUCHER\b[^>]*>)(.*?)(</VOUCHER>)", re.IGNORECASE | re.DOTALL)

    def replace_voucher(match):
        head, body, tail = match.group(1), match.group(2), match.group(3)
        body = re.sub(r"<DATE\s*/\s*>", f"<DATE>{target}</DATE>", body, flags=re.IGNORECASE)
        body = re.sub(r"<EFFECTIVEDATE\s*/\s*>", f"<EFFECTIVEDATE>{target}</EFFECTIVEDATE>", body, flags=re.IGNORECASE)
        if re.search(r"<DATE\b", body, flags=re.IGNORECASE):
            body = re.sub(r"<DATE\b[^>]*>.*?</DATE>", f"<DATE>{target}</DATE>", body,
                          flags=re.IGNORECASE | re.DOTALL)
        else:
            body = f"<DATE>{target}</DATE>" + body
        if re.search(r"<EFFECTIVEDATE\b", body, flags=re.IGNORECASE):
            body = re.sub(r"<EFFECTIVEDATE\b[^>]*>.*?</EFFECTIVEDATE>",
                          f"<EFFECTIVEDATE>{target}</EFFECTIVEDATE>", body,
                          flags=re.IGNORECASE | re.DOTALL)
        else:
            body = f"<EFFECTIVEDATE>{target}</EFFECTIVEDATE>" + body
        return head + body + tail

    updated_xml, touched = pattern.subn(replace_voucher, xml_content)
    return updated_xml, touched


def _force_voucher_dates_to_value(xml_content, tally_date):
    target = str(tally_date or "").strip()
    if not (target.isdigit() and len(target) == 8):
        return xml_content, 0

    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError:
        return _force_voucher_dates_to_value_regex(xml_content, target)

    updated_count = 0
    voucher_nodes = list(root.iter("VOUCHER"))
    if not voucher_nodes:
        return _force_voucher_dates_to_value_regex(xml_content, target)

    for voucher in voucher_nodes:
        date_node = voucher.find("DATE")
        if date_node is None:
            date_node = ET.SubElement(voucher, "DATE")
        if (date_node.text or "").strip() != target:
            date_node.text = target
            updated_count += 1
        effective_node = voucher.find("EFFECTIVEDATE")
        if effective_node is None:
            effective_node = ET.SubElement(voucher, "EFFECTIVEDATE")
        effective_node.text = target

    return ET.tostring(root, encoding="unicode"), updated_count


def _force_voucher_dates_to_today_regex(xml_content):
    today = datetime.date.today().strftime("%Y%m%d")
    return _force_voucher_dates_to_value_regex(xml_content, today)


def _force_voucher_dates_to_today(xml_content):
    """Force voucher DATE/EFFECTIVEDATE to current date for push-time reliability."""
    today = datetime.date.today().strftime("%Y%m%d")
    return _force_voucher_dates_to_value(xml_content, today)


def _apply_push_date_mode(xml_content, date_mode="current", custom_tally_date=""):
    mode = str(date_mode or "current").strip().lower()
    if mode == "excel":
        return xml_content, 0
    if mode == "custom":
        target_date = str(custom_tally_date or "").strip()
        if not target_date:
            raise ValueError("Custom date mode selected but custom date is missing.")
        if not (target_date.isdigit() and len(target_date) == 8):
            target_date = _normalize_manual_date_to_tally(target_date)
        return _force_voucher_dates_to_value(xml_content, target_date)
    return _force_voucher_dates_to_today(xml_content)


class GST2BDownloadWorker:
    def __init__(self, app_instance, excel_path, settings):
        self.app = app_instance
        self.excel_path = excel_path
        self.settings = settings
        self.keep_running = True
        self.driver = None
        self.captcha_response = None
        self.captcha_event = threading.Event()
        self.report_data = []

    def log(self, message):
        self.app.update_download2b_log_safe(message)

    def run(self):
        self.log("INITIALIZING GST 2B DOWNLOADER...")
        try:
            if isinstance(self.excel_path, str):
                df = pd.read_excel(self.excel_path)
                clean_cols = {str(c).lower().strip(): c for c in df.columns}
                user_col = next((clean_cols[c] for c in clean_cols if 'user' in c or 'name' in c), None)
                pass_col = next((clean_cols[c] for c in clean_cols if 'pass' in c or 'pwd' in c), None)

                if not user_col or not pass_col:
                    self.app.process_download2b_finished_safe("Column Error: Need Username/Password columns")
                    return
            else:
                # Manual entry
                df = pd.DataFrame(self.excel_path)
                user_col = "Username"
                pass_col = "Password"

            total = len(df)
            self.log(f"Loaded {total} users")

            base_dir = os.path.join(os.getcwd(), "GST_Downloads")
            if not os.path.exists(base_dir):
                os.makedirs(base_dir)

            for index, row in df.iterrows():
                if not self.keep_running:
                    break

                username = str(row[user_col]).strip()
                password = str(row[pass_col]).strip()

                self.app.update_download2b_progress_safe(index / max(1, total))
                self.log(f"Processing: {username}")

                user_root_base = os.path.join(base_dir, username)
                user_root = user_root_base
                counter = 1
                while os.path.exists(user_root):
                    user_root = f"{user_root_base}_{counter}"
                    counter += 1
                os.makedirs(user_root)

                status, reason = self.process_single_user(username, password, user_root)
                self.report_data.append({
                    "Username": username,
                    "Status": status,
                    "Details": reason,
                    "Timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Saved To": os.path.basename(user_root),
                })
                self.log("-" * 40)
                if not self.keep_running:
                    break

            self.generate_excel_report()
            self.app.update_download2b_progress_safe(1.0)
            if not self.keep_running:
                self.log("TASKS STOPPED BY USER")
                self.app.process_download2b_finished_safe("Process Stopped by User")
            else:
                self.log("ALL TASKS COMPLETED")
                self.keep_running = False
                self.app.process_download2b_finished_safe("Batch Completed & Report Saved")

        except Exception as e:
            self.log(f"Critical Error: {e}")
            self.keep_running = False
            self.app.process_download2b_finished_safe("Error Occurred")

    def generate_excel_report(self):
        try:
            if not self.report_data:
                return
            report_df = pd.DataFrame(self.report_data)
            filename = f"GST_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            report_df.to_excel(filename, index=False)
            self.log(f"Summary Report saved: {filename}")
        except Exception as e:
            self.log(f"Failed to save report: {e}")

    def process_single_user(self, username, password, user_root):
        try:
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--disable-infobars")
            options.add_argument("--disable-extensions")
            options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
            options.add_experimental_option("useAutomationExtension", False)
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

            prefs = {
                "download.prompt_for_download": False,
                "directory_upgrade": True,
                "safebrowsing.enabled": True,
                "plugins.always_open_pdf_externally": True,
                "profile.default_content_setting_values.automatic_downloads": 1,
                "credentials_enable_service": False,
                "profile.password_manager_enabled": False
            }
            options.add_experimental_option("prefs", prefs)

            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            
            # Advanced Stealth JS Injection
            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                    Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
                    Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
                    window.chrome = { runtime: {} };
                """
            })

            wait = WebDriverWait(self.driver, 20)
            login_status, login_msg = self.perform_login(username, password, wait)
            if not login_status:
                return "Login Failed", login_msg

            q_map = {
                "Quarter 1 (Apr - Jun)": ["April", "May", "June"],
                "Quarter 2 (Jul - Sep)": ["July", "August", "September"],
                "Quarter 3 (Oct - Dec)": ["October", "November", "December"],
                "Quarter 4 (Jan - Mar)": ["January", "February", "March"],
            }

            tasks = []
            if self.settings["all_quarters"]:
                for q_name, months in q_map.items():
                    for m in months:
                        tasks.append({"q": q_name, "m": m})
                self.log("Mode: All Quarters")
            else:
                selected_q = self.settings["quarter"]
                selected_m = self.settings["month"]
                if selected_m == "Whole Quarter":
                    if selected_q in q_map:
                        for m in q_map[selected_q]:
                            tasks.append({"q": selected_q, "m": m})
                        self.log(f"Mode: Whole {selected_q[:9]}")
                    else:
                        return "Config Error", "Invalid Quarter Data"
                else:
                    tasks.append({"q": selected_q, "m": selected_m})
                    self.log(f"Mode: Single Month ({selected_m})")

            time.sleep(3)
            success_count = 0
            results = []

            fin_year = self.settings["year"]
            year_folder = os.path.join(user_root, fin_year)
            if not os.path.exists(year_folder):
                os.makedirs(year_folder)

            self.driver.execute_cdp_cmd("Page.setDownloadBehavior", {
                "behavior": "allow",
                "downloadPath": year_folder,
            })

            for task in tasks:
                if not self.keep_running:
                    return "Stopped", "User Cancelled"

                q_text = task["q"]
                m_text = task["m"]
                month_success = False
                fail_reason = ""

                for attempt in range(1, 4):
                    self.log(f"Processing {m_text} (Attempt {attempt})")
                    try:
                        if attempt > 1:
                            try:
                                self.driver.get("https://return.gst.gov.in/returns/auth/dashboard")
                                time.sleep(4)
                                # Check if redirected to login
                                here_links = self.driver.find_elements(By.XPATH, "//a[contains(text(), 'here') and contains(@href, 'login')]")
                                if here_links and here_links[0].is_displayed():
                                    self.log("Session expired via 'here' pattern. Resuming login...")
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", here_links[0])
                                    time.sleep(1)
                                    self.driver.execute_script("arguments[0].click();", here_links[0])
                                    time.sleep(2)
                                    login_status, login_msg = self.perform_login(username, password, wait)
                                    if not login_status:
                                        fail_reason = "Re-login failed"
                                        continue
                                    self.driver.execute_script("window.location.href = 'https://return.gst.gov.in/returns/auth/dashboard';")
                                    time.sleep(4)
                                elif "login" in self.driver.current_url.lower() or "username" in self.driver.page_source.lower():
                                    self.log("Session expired. Attempting to re-login...")
                                    login_status, login_msg = self.perform_login(username, password, wait)
                                    if not login_status:
                                        fail_reason = "Re-login failed"
                                        continue
                                    self.driver.get("https://return.gst.gov.in/returns/auth/dashboard")
                                    time.sleep(4)
                            except Exception:
                                pass

                        year_el = wait.until(EC.element_to_be_clickable((By.NAME, "fin")))
                        # Make sure select isn't obscured
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", year_el)
                        # Remove popups
                        self.driver.execute_script("document.querySelectorAll('.modal-backdrop, .modal').forEach(e => e.remove()); document.body.classList.remove('modal-open');")
                        
                        Select(year_el).select_by_visible_text(fin_year)
                        time.sleep(1)

                        qtr_el = self.driver.find_element(By.NAME, "quarter")
                        Select(qtr_el).select_by_visible_text(q_text)
                        time.sleep(1)

                        mon_el = self.driver.find_element(By.NAME, "mon")
                        Select(mon_el).select_by_visible_text(m_text)
                        time.sleep(1)

                        search_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), 'Search') or contains(@id, 'Search')]")
                        self.driver.execute_script("arguments[0].click();", search_btn)
                        time.sleep(random.uniform(4.0, 6.0))

                        dl_status, dl_msg = self.download_gstr2b(wait, year_folder)
                        if dl_status:
                            month_success = True
                            success_count += 1
                            results.append(f"{m_text}: OK")
                            break
                        fail_reason = dl_msg
                        if "Not Generated" in dl_msg:
                            break
                        elif "System Error" in dl_msg or "Tile Missing" in dl_msg or "Gen Button Missing" in dl_msg:
                            # It happens if logged out, force dashboard load next retry
                            self.log("Possible session timeout. Forcing re-login check on next retry.")
                        self.log(f"Attempt {attempt} failed: {dl_msg}")

                    except Exception as e:
                        fail_reason = f"Error: {str(e)[:30]}"
                        self.log(f"Exception: {str(e)[:50]}")
                        try:
                            # if "here" link exists, click it
                            here_links = self.driver.find_elements(By.XPATH, "//a[contains(text(), 'here') and contains(@href, 'login')]")
                            if here_links and here_links[0].is_displayed():
                                self.log("Found 'login here' link. Clicking it to re-authenticate.")
                                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", here_links[0])
                                time.sleep(1)
                                self.driver.execute_script("arguments[0].click();", here_links[0])
                                time.sleep(4)
                                self.log("Session expired. Attempting to re-login...")
                                login_status, login_msg = self.perform_login(username, password, wait)
                                if not login_status:
                                    fail_reason = "Re-login failed"
                                    continue
                                self.driver.execute_script("window.location.href = 'https://return.gst.gov.in/returns/auth/dashboard';")
                                time.sleep(4)
                            else:
                                self.driver.execute_script("window.location.href = 'https://return.gst.gov.in/returns/auth/dashboard';")
                        except Exception:
                            pass

                if not month_success:
                    results.append(f"{m_text}: FAIL ({fail_reason})")

            overall_status = "Success" if success_count == len(tasks) else "Partial"
            if success_count == 0:
                overall_status = "Failed"
            summary = f"Downloaded {success_count}/{len(tasks)}. Details: " + ", ".join(results)
            return overall_status, summary

        except Exception as e:
            return "Error", f"Browser Crash: {str(e)[:30]}"
        finally:
            if self.driver:
                self.driver.quit()

    def perform_login(self, username, password, wait):
        self.log("Opening GST Portal...")
        self.driver.execute_script("window.location.href = 'https://services.gst.gov.in/services/login';")
        time.sleep(2)

        while True:
            if not self.keep_running:
                return False, "Stopped"
            try:
                user_elem = wait.until(EC.visibility_of_element_located((By.ID, "username")))
                user_elem.clear()
                time.sleep(0.5)
                for char in username:
                    user_elem.send_keys(char)
                    time.sleep(random.uniform(0.05, 0.15))
                
                pass_elem = self.driver.find_element(By.ID, "user_pass")
                pass_elem.clear()
                time.sleep(0.5)
                for char in password:
                    pass_elem.send_keys(char)
                    time.sleep(random.uniform(0.05, 0.15))

                captcha_img = wait.until(EC.visibility_of_element_located((By.ID, "imgCaptcha")))
                captcha_path = str(Path(os.getcwd()) / "temp_captcha.png")
                time.sleep(1)
                captcha_img.screenshot(captcha_path)

                self.log("Waiting for Captcha input...")
                self.captcha_response = None
                self.captcha_event.clear()
                self.app.request_download2b_captcha_safe(captcha_path)
                self.captcha_event.wait()

                if not self.captcha_response:
                    return False, "Captcha Cancelled"
                    
                cap_elem = self.driver.find_element(By.ID, "captcha")
                cap_elem.clear()
                time.sleep(0.3)
                for char in self.captcha_response:
                    cap_elem.send_keys(char)
                    time.sleep(random.uniform(0.05, 0.15))
                
                time.sleep(0.5)
                submit_btn = self.driver.find_element(By.XPATH, "//button[@type='submit']")
                self.driver.execute_script("arguments[0].click();", submit_btn)

                time.sleep(4)
                src = self.driver.page_source
                if "Invalid Username or Password" in src:
                    return False, "Invalid Credentials"
                if "Enter valid Letters" in src or "Invalid Captcha" in src:
                    self.log("Invalid Captcha. Retry...")
                    time.sleep(2)
                    continue

                if "Dashboard" in self.driver.title or "Return Dashboard" in src:
                    self.log("Login Successful")
                    self.app.close_download2b_captcha_safe()
                    time.sleep(3)

                    try:
                        aadhaar_skip = self.driver.find_elements(By.XPATH, "//a[contains(text(),'Remind me later')]")
                        if aadhaar_skip and aadhaar_skip[0].is_displayed():
                            aadhaar_skip[0].click()
                            time.sleep(1.5)
                    except Exception:
                        pass

                    try:
                        generic_skip = self.driver.find_elements(By.XPATH, "//button[contains(text(),'Remind Me Later')]")
                        if generic_skip and generic_skip[0].is_displayed():
                            generic_skip[0].click()
                            time.sleep(1.5)
                    except Exception:
                        pass

                    try:
                        dash_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Return Dashboard')]")))
                        dash_btn.click()
                        return True, "Success"
                    except Exception:
                        try:
                            btn = self.driver.find_element(By.XPATH, "//button[contains(., 'Return Dashboard')]")
                            self.driver.execute_script("arguments[0].click();", btn)
                            return True, "Success (JS Click)"
                        except Exception:
                            return False, "Dashboard Nav Failed"

            except Exception as e:
                self.log(f"Login Exception: {e}")
                return False, f"Login Error: {str(e)[:20]}"

    def download_gstr2b(self, wait, download_path):
        self.log("Searching for GSTR-2B tile...")

        xpath_std = "//div[contains(@class,'col-sm-4')]//p[contains(text(),'GSTR2B')]/ancestor::div[contains(@class,'col-sm-4')]//button[contains(text(),'Download')]"
        xpath_qtr = "//p[contains(text(),'Quarterly View')]/ancestor::div[contains(@class,'col-sm-4')]//button[contains(text(),'Download')]"

        found_btn = None
        try:
            found_btn = self.driver.find_element(By.XPATH, xpath_qtr)
            self.log("Found Quarterly View (GSTR-2BQ) tile")
        except Exception:
            try:
                found_btn = self.driver.find_element(By.XPATH, xpath_std)
                self.log("Found Standard GSTR-2B tile")
            except Exception:
                pass

        if not found_btn:
            return False, "Tile Missing"

        try:
            time.sleep(random.uniform(1.0, 2.5))
            self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", found_btn)
            time.sleep(random.uniform(0.5, 1.5))
            self.driver.execute_script("arguments[0].click();", found_btn)
            time.sleep(random.uniform(5.0, 8.0))

            gen_btn_xpath = "//button[contains(text(), 'GENERATE EXCEL FILE TO DOWNLOAD')]"

            if "no record" in self.driver.page_source or "compute your GSTR 2B" in self.driver.page_source:
                self.driver.back()
                time.sleep(random.uniform(1.0, 2.0))
                return False, "Not Generated"

            try:
                final_btn = wait.until(EC.element_to_be_clickable((By.XPATH, gen_btn_xpath)))
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", final_btn)
                time.sleep(random.uniform(1.0, 2.5))
                self.driver.execute_script("arguments[0].click();", final_btn)
            except Exception:
                self.driver.back()
                time.sleep(random.uniform(1.0, 2.0))
                return False, "Gen Button Missing"

            time.sleep(random.uniform(3.0, 5.0))
            if "no record" in self.driver.page_source:
                self.driver.back()
                time.sleep(random.uniform(1.0, 2.0))
                return False, "System Error"

            file_downloaded = False
            for _ in range(20):
                time.sleep(1)
                files = glob.glob(os.path.join(download_path, "*.*"))
                if files:
                    latest = max(files, key=os.path.getctime)
                    if (datetime.datetime.now().timestamp() - os.path.getctime(latest)) < 60:
                        self.log(f"Saved: {os.path.basename(latest)}")
                        file_downloaded = True
                        break
                try:
                    link = self.driver.find_element(By.XPATH, "//a[contains(text(), 'Click here to download')]")
                    if link.is_displayed():
                        self.driver.execute_script("arguments[0].click();", link)
                except Exception:
                    pass

            self.driver.back()
            if not file_downloaded:
                return False, "Timeout"
            return True, "Success"

        except Exception as e:
            try:
                self.driver.back()
            except Exception:
                pass
            return False, f"Script Error: {str(e)[:20]}"


# ─── GST Portal Search ──────────────────────────────────────────────────────

_GST_SEARCH_URL = "https://services.gst.gov.in/services/searchtp"
_GST_MOBILE_UA = (
    "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Mobile Safari/537.36"
)
_GST_DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
)


def _pan_from_gstin(gstin: str) -> str:
    g = (gstin or "").strip().upper()
    return g[2:12] if len(g) == 15 else ""


def _normalize_state_for_portal(value: str) -> str:
    text = str(value or "").strip()
    if text.casefold() in {"not applicable", "* not applicable", "na", "n/a"}:
        return ""
    return text


def _parse_portal_data_to_extra(portal_data: dict, ledger_name: str) -> dict:
    """Convert GST portal result to the 'extra' dict format used by the ledger details form."""
    details    = portal_data.get("details") or {}
    key_d      = details.get("key_details") or {}

    trade_name  = str(key_d.get("Trade Name") or "").strip()
    legal_name  = str(key_d.get("Legal Name of Business") or "").strip()
    mailing     = trade_name or legal_name or ledger_name

    ppob        = str(key_d.get("Principal Place of Business") or "").strip()
    address1    = ""
    address2    = ""
    state_raw   = ""
    pincode_raw = ""

    if ppob:
        parts = [p.strip() for p in ppob.split(",") if p.strip()]
        if parts and re.fullmatch(r"\d{6}", parts[-1]):
            pincode_raw = parts.pop()
        if parts:
            state_raw = parts[-1]
        if len(parts) > 1:
            mid      = parts[:-1]
            address1 = ", ".join(mid[:3])
            address2 = ", ".join(mid[3:])
        elif parts:
            address1 = parts[0]

    taxpayer_type = str(key_d.get("Taxpayer Type") or "").strip()
    gstin_val     = str(portal_data.get("gstin") or "").strip().upper()

    reg_map = {
        "regular": "Regular", "composition": "Composition",
        "unregistered": "Unregistered/Consumer", "consumer": "Unregistered/Consumer",
        "input service distributor": "Input Service Distributor", "isd": "Input Service Distributor",
        "sez unit": "SEZ", "sez developer": "SEZ",
        "embassy / consulate": "Overseas", "overseas": "Overseas",
    }
    reg_type = reg_map.get(taxpayer_type.lower(), "Regular" if gstin_val else "Unregistered/Consumer")

    normalized_state = _normalize_state_for_portal(state_raw)
    if not normalized_state and gstin_val:
        normalized_state = _state_name_from_gstin(gstin_val)

    return {
        "mailing_name": mailing,
        "gstin":        gstin_val,
        "gst_applicable": "Applicable" if gstin_val else "Not Applicable",
        "reg_type":     reg_type,
        "state":        normalized_state,
        "address1":     address1,
        "address2":     address2,
        "pincode":      pincode_raw,
        "country":      "India",
        "billwise":     "Yes",
    }


class _GSTPortalSearcher:
    """Selenium-based GST portal searcher — loads captcha, submits, returns data dict."""

    def __init__(self):
        self.driver            = None
        self.captcha_png_bytes = b""

    def _ensure_driver(self):
        if self.driver is not None:
            return self.driver
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument(f"--user-agent={_GST_MOBILE_UA}")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--window-size=1200,900")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--lang=en-US")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        svc          = Service(ChromeDriverManager().install())
        self.driver  = webdriver.Chrome(service=svc, options=options)
        self.driver.set_page_load_timeout(60)
        try:
            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"})
        except Exception:
            pass
        return self.driver

    def _trigger_events(self, driver, element):
        try:
            driver.execute_script(
                "const e=arguments[0];"
                "e.dispatchEvent(new Event('input',{bubbles:true}));"
                "e.dispatchEvent(new Event('change',{bubbles:true}));"
                "e.dispatchEvent(new Event('blur',{bubbles:true}));", element)
        except Exception:
            pass

    def _find_ready_captcha(self, driver):
        try:
            el  = driver.find_element(By.ID, "imgCaptcha")
            src = (el.get_attribute("src") or "").lower()
            if el.is_displayed() and "captcha" in src:
                return el
        except Exception:
            pass
        return False

    def _capture_captcha_bytes(self, driver, el) -> bytes:
        try:
            WebDriverWait(driver, 15).until(
                lambda d: d.execute_script(
                    "const i=arguments[0];return i&&i.complete&&i.naturalWidth>0;", el))
        except Exception:
            pass
        try:
            png = el.screenshot_as_png
            if png and len(png) > 200:
                return png
        except Exception:
            pass
        return b""

    def _find_captcha_input(self, driver):
        locators = [
            (By.ID, "captcha"), (By.ID, "captchaCode"), (By.ID, "captchaText"),
            (By.NAME, "captcha"), (By.NAME, "captchaCode"),
            (By.CSS_SELECTOR, "input[ng-model*='captcha' i]"),
            (By.CSS_SELECTOR, "input[id*='captcha' i]"),
        ]
        for by, val in locators:
            try:
                el = driver.find_element(by, val)
                if el.is_displayed():
                    return el
            except Exception:
                pass
        return None

    def load_captcha(self, gstin: str) -> bytes:
        driver = self._ensure_driver()
        driver.get(_GST_SEARCH_URL)
        wait   = WebDriverWait(driver, 35)
        inp    = wait.until(EC.element_to_be_clickable((By.ID, "for_gstin")))
        inp.clear()
        inp.send_keys(gstin)
        self._trigger_events(driver, inp)
        el = wait.until(self._find_ready_captcha)
        if not el:
            raise RuntimeError("Timed out waiting for captcha on GST portal.")
        self.captcha_png_bytes = self._capture_captcha_bytes(driver, el)
        return self.captcha_png_bytes

    def fetch(self, gstin: str, captcha_text: str) -> dict:
        driver = self._ensure_driver()
        wait   = WebDriverWait(driver, 25)
        if "searchtp" not in (driver.current_url or "").lower():
            driver.get(_GST_SEARCH_URL)

        inp = wait.until(EC.element_to_be_clickable((By.ID, "for_gstin")))
        inp.clear()
        inp.send_keys(gstin)
        self._trigger_events(driver, inp)

        cap_inp = None
        for _ in range(40):
            cap_inp = self._find_captcha_input(driver)
            if cap_inp:
                break
            time.sleep(0.3)
        if not cap_inp:
            raise RuntimeError("Captcha input not found on page.")
        cap_inp.clear()
        cap_inp.send_keys(captcha_text)

        btn = wait.until(EC.element_to_be_clickable((By.ID, "lotsearch")))
        btn.click()

        end = time.time() + 35
        while time.time() < end:
            for sel in ["#lottable", "#searchResult", ".panel-body", "table"]:
                try:
                    el = driver.find_element(By.CSS_SELECTOR, sel)
                    if el.text.strip():
                        break
                except Exception:
                    pass
            else:
                time.sleep(0.5)
                continue
            break

        details = driver.execute_script("""
            const getText=e=>e?(e.textContent||'').replace(/\\s+/g,' ').trim():'';
            const result={};const container=document.querySelector('#lottable');
            if(!container)return result;
            const pairs={};
            const cols=container.querySelectorAll('.tbl-format .inner .col-sm-4,.tbl-format .inner .col-sm-12');
            cols.forEach(col=>{
                const label=getText(col.querySelector('strong'));if(!label)return;
                const list=col.querySelector('ul.jurisdictList');
                if(list){const items=Array.from(list.querySelectorAll('li')).map(getText).filter(Boolean);
                    if(items.length){pairs[label]=items;return;}}
                const word=col.querySelector('.wordCls');
                if(word){pairs[label]=getText(word);return;}
                const ps=Array.from(col.querySelectorAll('p')).map(getText).filter(Boolean);
                if(ps.length>1)pairs[label]=ps[ps.length-1];
                else if(ps.length===1)pairs[label]=ps[0];
            });
            result.key_details=pairs;
            return result;
        """) or {}

        return {"gstin": gstin, "fetchedAt": datetime.datetime.now().isoformat(), "details": details}

    def quit(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


class _GSTFetchDialog(ctk.CTkToplevel):
    """
    Popup for fetching ledger details from the GST portal via captcha.
    Calls on_result(extra_dict) when user confirms.  extra_dict uses lowercase
    keys matching the 'extra' format used by the missing-ledger details form.
    """

    def __init__(self, parent, ledger_name: str, gstin_hint: str,
                 ledger_parent: str, on_result=None):
        super().__init__(parent)
        self.title(f"GST Portal — {ledger_name}")
        self.geometry("560x620")
        self.resizable(False, False)
        self.grab_set()
        self.lift()
        self.focus_force()

        self._ledger_name   = ledger_name
        self._ledger_parent = ledger_parent
        self._on_result     = on_result
        self._searcher      = _GSTPortalSearcher()
        self._captcha_photo = None

        self._gstin_var    = tk.StringVar(value=gstin_hint)
        self._captcha_var  = tk.StringVar()
        self._status_var   = tk.StringVar(value="Enter GSTIN, click Load Captcha.")
        self._name_var     = tk.StringVar()
        self._state_var    = tk.StringVar()
        self._addr_var     = tk.StringVar()
        self._pin_var      = tk.StringVar()
        self._reg_type_var = tk.StringVar()

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

    def _build_ui(self):
        pad = {"padx": 14, "pady": 6}
        ctk.CTkLabel(self, text="GSTIN / UIN", anchor="w").pack(fill="x", **pad)
        r1 = ctk.CTkFrame(self, fg_color="transparent")
        r1.pack(fill="x", padx=14, pady=(0, 6))
        ctk.CTkEntry(r1, textvariable=self._gstin_var, width=300).pack(side="left", padx=(0, 8))
        ctk.CTkButton(r1, text="Load Captcha", width=130,
                      command=self._on_load_captcha).pack(side="left")

        ctk.CTkLabel(self, text="Captcha Image", anchor="w").pack(fill="x", **pad)
        self._captcha_lbl = ctk.CTkLabel(self, text="— captcha not loaded —",
                                          width=360, height=80,
                                          fg_color=("gray90", "gray20"), corner_radius=6)
        self._captcha_lbl.pack(padx=14, pady=(0, 6))

        ctk.CTkLabel(self, text="Enter Captcha Text", anchor="w").pack(fill="x", **pad)
        r2 = ctk.CTkFrame(self, fg_color="transparent")
        r2.pack(fill="x", padx=14, pady=(0, 6))
        ctk.CTkEntry(r2, textvariable=self._captcha_var, width=200).pack(side="left", padx=(0, 8))
        ctk.CTkButton(r2, text="Fetch Details", width=130,
                      command=self._on_fetch).pack(side="left")

        ctk.CTkLabel(self, textvariable=self._status_var, anchor="w",
                      text_color=("gray40", "gray70"),
                      font=("Segoe UI", 11)).pack(fill="x", padx=14, pady=(0, 8))

        ctk.CTkFrame(self, height=1, fg_color=("gray80", "gray30")).pack(fill="x", padx=14, pady=(0, 8))
        ctk.CTkLabel(self, text="Fetched Details (editable):",
                      anchor="w", font=("Segoe UI", 11, "bold")).pack(fill="x", padx=14, pady=(0, 4))

        for lbl, var in [
            ("Mailing Name",          self._name_var),
            ("GST Registration Type", self._reg_type_var),
            ("State",                 self._state_var),
            ("Address",               self._addr_var),
            ("Pincode",               self._pin_var),
        ]:
            fr = ctk.CTkFrame(self, fg_color="transparent")
            fr.pack(fill="x", padx=14, pady=2)
            ctk.CTkLabel(fr, text=lbl, width=190, anchor="w").pack(side="left")
            ctk.CTkEntry(fr, textvariable=var, width=300).pack(side="left")

        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=14, pady=(12, 8))
        self._confirm_btn = ctk.CTkButton(
            btn_row, text="Use Portal Data",
            fg_color=("#059669", "#10B981"),
            command=self._on_confirm, state="disabled")
        self._confirm_btn.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkButton(btn_row, text="Cancel",
                       fg_color=("gray60", "gray30"),
                       command=self._on_cancel, width=90).pack(side="left")

    def _on_load_captcha(self):
        gstin = self._gstin_var.get().strip().upper()
        if not gstin:
            self._status_var.set("Enter GSTIN first.")
            return
        self._status_var.set("Loading captcha (opening browser)...")
        self._captcha_lbl.configure(text="Loading…")
        threading.Thread(target=self._bg_load_captcha, args=(gstin,), daemon=True).start()

    def _bg_load_captcha(self, gstin):
        try:
            png = self._searcher.load_captcha(gstin)
            self.after(0, self._show_captcha, png)
            self.after(0, self._status_var.set, "Captcha loaded. Enter text and click Fetch Details.")
        except Exception as exc:
            self.after(0, self._status_var.set, f"Error: {exc}")
            self.after(0, lambda: self._captcha_lbl.configure(text="Failed to load captcha"))

    def _show_captcha(self, png_bytes):
        try:
            import io as _io
            img = Image.open(_io.BytesIO(png_bytes))
            img = img.resize(
                (img.width * 2, img.height * 2),
                Image.LANCZOS if hasattr(Image, "LANCZOS") else Image.Resampling.LANCZOS,
            )
            from PIL import ImageTk as _ITk
            self._captcha_photo = _ITk.PhotoImage(img)
            self._captcha_lbl.configure(image=self._captcha_photo, text="")
        except Exception:
            self._captcha_lbl.configure(text="Captcha loaded (render error)")

    def _on_fetch(self):
        gstin        = self._gstin_var.get().strip().upper()
        captcha_text = self._captcha_var.get().strip()
        if not gstin:
            self._status_var.set("GSTIN is required.")
            return
        if not captcha_text:
            self._status_var.set("Enter captcha text first.")
            return
        self._status_var.set("Fetching from GST portal...")
        threading.Thread(target=self._bg_fetch, args=(gstin, captcha_text), daemon=True).start()

    def _bg_fetch(self, gstin, captcha_text):
        try:
            data    = self._searcher.fetch(gstin, captcha_text)
            extra   = _parse_portal_data_to_extra(data, self._ledger_name)
            self.after(0, self._populate_fields, extra)
            self.after(0, self._status_var.set, "Details fetched. Review and click Use Portal Data.")
        except Exception as exc:
            self.after(0, self._status_var.set, f"Fetch failed: {exc}")

    def _populate_fields(self, extra: dict):
        self._name_var.set(extra.get("mailing_name", "") or self._ledger_name)
        self._reg_type_var.set(extra.get("reg_type", ""))
        self._state_var.set(extra.get("state", ""))
        self._addr_var.set(
            ", ".join(filter(None, [extra.get("address1", ""), extra.get("address2", "")])))
        self._pin_var.set(extra.get("pincode", ""))
        self._confirm_btn.configure(state="normal")

    def _on_confirm(self):
        addr_full  = self._addr_var.get().strip()
        addr_parts = [p.strip() for p in addr_full.split(",") if p.strip()]
        gstin_val  = self._gstin_var.get().strip().upper()
        result_def = {
            "mailing_name":   self._name_var.get().strip() or self._ledger_name,
            "gstin":          gstin_val,
            "gst_applicable": "Applicable" if gstin_val else "Not Applicable",
            "reg_type":       self._reg_type_var.get().strip() or "Regular",
            "state":          _normalize_state_for_portal(self._state_var.get().strip()),
            "address1":       addr_parts[0] if addr_parts else "",
            "address2":       ", ".join(addr_parts[1:]) if len(addr_parts) > 1 else "",
            "pincode":        self._pin_var.get().strip(),
            "country":        "India",
            "billwise":       "Yes",
        }
        self._searcher.quit()
        if callable(self._on_result):
            self._on_result(result_def)
        self.destroy()

    def _on_cancel(self):
        self._searcher.quit()
        self.destroy()


# ═══════════════════════════════════════════════════════════
#  MAIN APPLICATION WINDOW
# ═══════════════════════════════════════════════════════════

class GSTR2BTallyApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.engine = GSTR2BEngine()
        self.source_file = ""
        self.mapping_file = ""
        self.itc_download_btn = None
        self.tally_sheet_file = ""
        self.download2b_excel_file = ""
        self.tally_push_xml_file = ""
        self.download2b_worker = None
        self._download2b_captcha_img = None
        self.tally_push_is_running = False
        # Pending push-after-generate state
        self._pending_push = False
        self._pending_push_url = ""
        self._pending_push_timeout = 30
        self._pending_push_date_mode = "current"
        self._pending_push_custom_date = ""
        self._pending_push_company = ""
        self._stock_items_pushed_count = 0  # tracks stock item vouchers for combined total
        self.tally_push_companies = []
        self.tally_push_company_placeholder = "Auto (Loaded Company)"
        self.roundoff_all_ledgers = []   # cached ledger list for round off picker
        self.create_ledger_is_running = False
        self.create_ledger_companies = []
        self.output_dir = ""
        self.current_mode = "gstr2b"
        self.workflow_demo_url = ""  # Add YouTube demo link later.
        self.tally_push_date_mode = ctk.StringVar(value="current")
        self.tally_push_custom_date_var = ctk.StringVar(value="")
        self.tally_push_date_checks = {
            "current": ctk.BooleanVar(value=True),
            "excel": ctk.BooleanVar(value=False),
            "custom": ctk.BooleanVar(value=False),
        }

        self.title("GSTR-2B + Tally Sheet → Tally Converter  |  Studycafe PVT LTD")
        self.geometry("1180x820")
        self.minsize(1000, 700)
        self.configure(fg_color=COLORS["bg_dark"])
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")
        self._build_ui()

    def set_theme(self, mode: str):
        try:
            ctk.set_appearance_mode(mode)
        except Exception:
            pass

    def _build_ui(self):
        self.company_label = ctk.CTkLabel(
            self,
            text="",
            font=("Segoe UI", 11),
            text_color=COLORS["text_secondary"],
        )

        # ═══ MAIN CONTAINER ═══
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=16, pady=12)
        self.left_col = ctk.CTkScrollableFrame(main, fg_color="transparent", width=360)
        self.left_col.pack(side="left", fill="y", padx=(0, 8))
        self.right_col = ctk.CTkFrame(main, fg_color="transparent")
        self.right_col.pack(side="left", fill="both", expand=True, padx=(8, 0))

        # ─── LEFT: Mode Selector ───
        mode_card = ctk.CTkFrame(self.left_col, fg_color=COLORS["bg_card"], corner_radius=12,
                                  border_width=1, border_color=COLORS["border"])
        mode_card.pack(fill="x", pady=(0, 8))
        mode_head = ctk.CTkFrame(mode_card, fg_color="transparent")
        mode_head.pack(fill="x", padx=16, pady=(14, 8))
        ctk.CTkLabel(mode_head, text="Workflow Steps", font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(side="left")
        ctk.CTkButton(
            mode_head,
            text="▶ View Demo",
            width=132,
            height=28,
            font=("Segoe UI", 10, "bold"),
            fg_color="#DC2626",
            hover_color="#B91C1C",
            text_color="#FFFFFF",
            corner_radius=6,
            command=self._view_workflow_demo,
        ).pack(side="right")
        mode_grid = ctk.CTkFrame(mode_card, fg_color="transparent")
        mode_grid.pack(fill="x", padx=16, pady=(0, 14))
        mode_grid.grid_columnconfigure((0, 1), weight=1)
        self.mode_buttons = {}
        mode_labels = ["Step 1: GSTR-2B → XML", "Step 2: Push To Tally", "Create Ledger"]
        for idx, label in enumerate(mode_labels):
            btn = ctk.CTkButton(
                mode_grid,
                text=label,
                height=34,
                font=("Segoe UI", 11, "bold"),
                corner_radius=8,
                fg_color=COLORS["bg_input"],
                hover_color=COLORS["bg_card_hover"],
                text_color=COLORS["text_primary"],
                command=lambda value=label: self._on_mode_change(value),
            )
            btn.grid(row=idx // 2, column=idx % 2, sticky="ew", padx=3, pady=3)
            self.mode_buttons[label] = btn
        self._refresh_mode_selector_text_colors("Step 1: GSTR-2B → XML")

        # ─── LEFT: GSTR-2B Upload Card ───
        self.gstr2b_card = ctk.CTkFrame(self.left_col, fg_color=COLORS["bg_card"], corner_radius=12,
                                         border_width=1, border_color=COLORS["border"])
        self.gstr2b_card.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(self.gstr2b_card, text="Step 1A: Upload GSTR-2B File", font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(anchor="w", padx=16, pady=(14, 8))
        self.upload_zone = ctk.CTkFrame(self.gstr2b_card, fg_color=COLORS["bg_dark"], corner_radius=10,
                                         height=80, border_width=2, border_color=COLORS["border"])
        self.upload_zone.pack(fill="x", padx=16, pady=(0, 6)); self.upload_zone.pack_propagate(False)
        self.file_icon_label = ctk.CTkLabel(self.upload_zone, text="Click to select .xlsx file",
                                             font=("Segoe UI", 12), text_color=COLORS["text_muted"], cursor="hand2")
        self.file_icon_label.place(relx=0.5, rely=0.5, anchor="center")
        self.file_icon_label.bind("<Button-1>", lambda e: self._browse_file())
        self.upload_zone.bind("<Button-1>", lambda e: self._browse_file())
        ctk.CTkButton(self.gstr2b_card, text="Browse File", font=("Segoe UI", 12, "bold"),
                      fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"], text_color="#FFFFFF",
                      corner_radius=8, height=38, command=self._browse_file).pack(fill="x", padx=16, pady=(4, 6))

        # ─── ITC Template Section (inside gstr2b_card) ───
        map_frame = ctk.CTkFrame(self.gstr2b_card, fg_color=COLORS["bg_input"], corner_radius=8)
        map_frame.pack(fill="x", padx=16, pady=(4, 14))
        ctk.CTkLabel(map_frame, text="Step 1B: ITC Template", font=("Segoe UI", 11, "bold"),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=10, pady=(8, 4))
        self.mapping_label = ctk.CTkLabel(map_frame, text="No template uploaded — load 2B file first to download template",
                                           font=("Segoe UI", 10), text_color=COLORS["text_muted"])
        self.mapping_label.pack(anchor="w", padx=10)
        map_btn_frame = ctk.CTkFrame(map_frame, fg_color="transparent")
        map_btn_frame.pack(fill="x", padx=10, pady=(6, 8))

        self.itc_download_btn = ctk.CTkButton(
            map_btn_frame, text="📥  Download Template", font=("Segoe UI", 11), height=32,
            fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"], text_color="#FFFFFF",
            corner_radius=6, state="disabled", command=self._generate_itc_template)
        self.itc_download_btn.pack(fill="x", pady=(0, 4))

        ctk.CTkButton(
            map_btn_frame, text="📤  Upload Template", font=("Segoe UI", 11), height=32,
            fg_color=COLORS["bg_card"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], corner_radius=6,
            command=self._upload_itc_template).pack(fill="x")

        # ─── LEFT: Tally Sheet Upload Card (hidden by default) ───
        self.tally_card = ctk.CTkFrame(self.left_col, fg_color=COLORS["bg_card"], corner_radius=12,
                                        border_width=1, border_color=COLORS["border"])
        ctk.CTkLabel(self.tally_card, text="Upload Tally Sheet", font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(anchor="w", padx=16, pady=(14, 8))
        self.tally_zone = ctk.CTkFrame(self.tally_card, fg_color=COLORS["bg_dark"], corner_radius=10,
                                        height=80, border_width=2, border_color=COLORS["border"])
        self.tally_zone.pack(fill="x", padx=16, pady=(0, 6)); self.tally_zone.pack_propagate(False)
        self.tally_label = ctk.CTkLabel(self.tally_zone, text="Click to select tally sheet .xlsx",
                                         font=("Segoe UI", 12), text_color=COLORS["text_muted"], cursor="hand2")
        self.tally_label.place(relx=0.5, rely=0.5, anchor="center")
        self.tally_label.bind("<Button-1>", lambda e: self._browse_tally_sheet())
        self.tally_zone.bind("<Button-1>", lambda e: self._browse_tally_sheet())
        ctk.CTkButton(self.tally_card, text="Browse Tally Sheet", font=("Segoe UI", 12, "bold"),
                      fg_color=COLORS["xml_accent"], hover_color=COLORS["xml_accent_h"], text_color="#FFFFFF",
                      corner_radius=8, height=38, command=self._browse_tally_sheet).pack(fill="x", padx=16, pady=(4, 14))

        # ─── LEFT: Download 2B Card (hidden by default) ───
        self.download2b_card = ctk.CTkFrame(self.left_col, fg_color=COLORS["bg_card"], corner_radius=12,
                                            border_width=1, border_color=COLORS["border"])
        ctk.CTkLabel(self.download2b_card, text="GST Portal 2B Downloader", font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(anchor="w", padx=16, pady=(14, 4))
                     
        # Tabs for credentials input method
        self.d2b_input_mode = ctk.StringVar(value="Excel File")
        self.d2b_tabview = ctk.CTkSegmentedButton(self.download2b_card, values=["Excel File", "Manual Entry"],
                                                  variable=self.d2b_input_mode,
                                                  command=self._on_d2b_input_mode_change,
                                                  selected_color=COLORS["accent"],
                                                  selected_hover_color=COLORS["accent_hover"])
        self.d2b_tabview.pack(fill="x", padx=16, pady=(0, 10))

        # --- Excel Input Frame ---
        self.d2b_excel_frame = ctk.CTkFrame(self.download2b_card, fg_color="transparent")
        self.d2b_excel_frame.pack(fill="x")
        ctk.CTkLabel(self.d2b_excel_frame, text="Credentials Excel", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        d2b_file_row = ctk.CTkFrame(self.d2b_excel_frame, fg_color="transparent")
        d2b_file_row.pack(fill="x", padx=16, pady=(4, 8))
        self.download2b_file_entry = ctk.CTkEntry(
            d2b_file_row,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Select Excel with Username/Password columns",
            font=("Segoe UI", 11),
            corner_radius=8,
        )
        self.download2b_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkButton(
            d2b_file_row,
            text="...",
            width=42,
            height=34,
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._browse_download2b_excel,
        ).pack(side="right")

        # --- Manual Entry Frame ---
        self.d2b_manual_frame = ctk.CTkFrame(self.download2b_card, fg_color="transparent")
        
        self.d2b_manual_entries = []
        self.d2b_manual_list = ctk.CTkScrollableFrame(self.d2b_manual_frame, fg_color=COLORS["bg_input"], border_color=COLORS["border"], border_width=1, corner_radius=8, height=120)
        self.d2b_manual_list.pack(fill="x", padx=16, pady=(4, 8))
        
        add_btn_frame = ctk.CTkFrame(self.d2b_manual_frame, fg_color="transparent")
        add_btn_frame.pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkButton(add_btn_frame, text="+ Add User", width=100, height=28, font=("Segoe UI", 11, "bold"),
                      fg_color=COLORS["bg_dark"], text_color=COLORS["text_primary"], hover_color=COLORS["bg_card_hover"],
                      command=self._add_d2b_manual_entry).pack(side="left")
        
        self._add_d2b_manual_entry() # Add one by default

        now = datetime.datetime.now()
        start_year = now.year - 1 if now.month < 4 else now.year
        fin_years = [f"{y}-{str(y + 1)[-2:]}" for y in range(start_year - 2, start_year + 2)]
        fin_years.sort(reverse=True)

        self.download2b_year_row_ref = ctk.CTkFrame(self.download2b_card, fg_color="transparent")
        self.download2b_year_row_ref.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(self.download2b_year_row_ref, text="Financial Year", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.download2b_year_cb = ctk.CTkComboBox(self.download2b_year_row_ref, values=fin_years, width=170,
                                                   fg_color=COLORS["bg_input"],
                                                   border_color=COLORS["border"],
                                                   button_color=COLORS["accent"],
                                                   button_hover_color=COLORS["accent_hover"])
        self.download2b_year_cb.set(fin_years[0])
        self.download2b_year_cb.pack(side="right")

        self.download2b_all_quarters_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            self.download2b_card,
            text="Download All Quarters (Apr-Mar)",
            variable=self.download2b_all_quarters_var,
            command=self._toggle_download2b_inputs,
            font=("Segoe UI", 11, "bold"),
            text_color=COLORS["text_secondary"],
        ).pack(anchor="w", padx=16, pady=(2, 6))

        d2b_q_row = ctk.CTkFrame(self.download2b_card, fg_color="transparent")
        d2b_q_row.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(d2b_q_row, text="Quarter", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.download2b_quarter_cb = ctk.CTkComboBox(
            d2b_q_row,
            values=["Quarter 1 (Apr - Jun)", "Quarter 2 (Jul - Sep)", "Quarter 3 (Oct - Dec)", "Quarter 4 (Jan - Mar)"],
            command=self._update_download2b_months,
            width=190,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
        )
        self.download2b_quarter_cb.set("Quarter 1 (Apr - Jun)")
        self.download2b_quarter_cb.pack(side="right")

        d2b_m_row = ctk.CTkFrame(self.download2b_card, fg_color="transparent")
        d2b_m_row.pack(fill="x", padx=16, pady=(0, 10))
        ctk.CTkLabel(d2b_m_row, text="Month", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.download2b_month_cb = ctk.CTkComboBox(
            d2b_m_row,
            values=["Whole Quarter", "April", "May", "June"],
            width=190,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
        )
        self.download2b_month_cb.set("Whole Quarter")
        self.download2b_month_cb.pack(side="right")

        d2b_btn_row = ctk.CTkFrame(self.download2b_card, fg_color="transparent")
        d2b_btn_row.pack(fill="x", padx=16, pady=(0, 14))
        self.download2b_start_btn = ctk.CTkButton(
            d2b_btn_row,
            text="Start Download",
            font=("Segoe UI", 12, "bold"),
            fg_color=COLORS["success"],
            hover_color="#047857",
            text_color="#FFFFFF",
            corner_radius=8,
            height=38,
            command=self._start_download2b_process,
        )
        self.download2b_start_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.download2b_stop_btn = ctk.CTkButton(
            d2b_btn_row,
            text="Stop",
            font=("Segoe UI", 11, "bold"),
            fg_color=COLORS["warning"],
            hover_color="#B45309",
            text_color="#FFFFFF",
            corner_radius=8,
            height=38,
            width=92,
            state="disabled",
            command=self._stop_download2b_process,
        )
        self.download2b_stop_btn.pack(side="left", padx=(4, 0))

        # ─── LEFT: Push To Tally Card (hidden by default) ───
        self.tally_push_card = ctk.CTkFrame(
            self.left_col,
            fg_color=COLORS["bg_card"],
            corner_radius=12,
            border_width=1,
            border_color=COLORS["border"],
        )
        ctk.CTkLabel(
            self.tally_push_card,
            text="Step 2: Push XML To Tally",
            font=("Segoe UI", 13, "bold"),
            text_color=COLORS["text_primary"],
        ).pack(anchor="w", padx=16, pady=(14, 8))

        ctk.CTkLabel(
            self.tally_push_card,
            text="Tally XML File",
            font=("Segoe UI", 11),
            text_color=COLORS["text_secondary"],
        ).pack(anchor="w", padx=16)
        push_file_row = ctk.CTkFrame(self.tally_push_card, fg_color="transparent")
        push_file_row.pack(fill="x", padx=16, pady=(4, 10))
        self.tally_push_file_entry = ctk.CTkEntry(
            push_file_row,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Select generated tally-ready XML file",
            font=("Segoe UI", 11),
            corner_radius=8,
        )
        self.tally_push_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkButton(
            push_file_row,
            text="...",
            width=42,
            height=34,
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._browse_tally_push_xml,
        ).pack(side="right")

        conn_row = ctk.CTkFrame(self.tally_push_card, fg_color="transparent")
        conn_row.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(conn_row, text="Host", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.tally_push_host_entry = ctk.CTkEntry(
            conn_row,
            width=104,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 11),
            corner_radius=8,
        )
        self.tally_push_host_entry.insert(0, "localhost")
        self.tally_push_host_entry.pack(side="left", padx=(6, 10))

        ctk.CTkLabel(conn_row, text="Port", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.tally_push_port_entry = ctk.CTkEntry(
            conn_row,
            width=74,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 11),
            corner_radius=8,
        )
        self.tally_push_port_entry.insert(0, "9000")
        self.tally_push_port_entry.pack(side="left", padx=(6, 0))

        timeout_row = ctk.CTkFrame(self.tally_push_card, fg_color="transparent")
        timeout_row.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(timeout_row, text="Timeout(s)", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.tally_push_timeout_entry = ctk.CTkEntry(
            timeout_row,
            width=90,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 11),
            corner_radius=8,
        )
        self.tally_push_timeout_entry.insert(0, "30")
        self.tally_push_timeout_entry.pack(side="left", padx=(6, 0))

        date_mode_frame = ctk.CTkFrame(self.tally_push_card, fg_color=COLORS["bg_input"], corner_radius=8)
        date_mode_frame.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(
            date_mode_frame,
            text="Voucher Date Mode",
            font=("Segoe UI", 10, "bold"),
            text_color=COLORS["text_secondary"],
        ).pack(anchor="w", padx=10, pady=(8, 4))

        date_checks_row = ctk.CTkFrame(date_mode_frame, fg_color="transparent")
        date_checks_row.pack(fill="x", padx=10, pady=(0, 4))

        self.tally_push_date_current_cb = ctk.CTkCheckBox(
            date_checks_row,
            text="Current Date",
            variable=self.tally_push_date_checks["current"],
            command=lambda: self._set_tally_push_date_mode("current"),
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
        )
        self.tally_push_date_current_cb.pack(side="left", padx=(0, 8))

        self.tally_push_date_excel_cb = ctk.CTkCheckBox(
            date_checks_row,
            text="Excel Date",
            variable=self.tally_push_date_checks["excel"],
            command=lambda: self._set_tally_push_date_mode("excel"),
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
        )
        self.tally_push_date_excel_cb.pack(side="left", padx=(0, 8))

        self.tally_push_date_custom_cb = ctk.CTkCheckBox(
            date_checks_row,
            text="Custom Date",
            variable=self.tally_push_date_checks["custom"],
            command=lambda: self._set_tally_push_date_mode("custom"),
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
        )
        self.tally_push_date_custom_cb.pack(side="left")

        custom_date_row = ctk.CTkFrame(date_mode_frame, fg_color="transparent")
        custom_date_row.pack(fill="x", padx=10, pady=(0, 8))
        self.tally_push_custom_date_entry = ctk.CTkEntry(
            custom_date_row,
            textvariable=self.tally_push_custom_date_var,
            height=30,
            fg_color=COLORS["bg_card"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Custom Date (DD/MM/YYYY)",
            font=("Segoe UI", 10),
            corner_radius=6,
        )
        self.tally_push_custom_date_entry.pack(fill="x")
        self._set_tally_push_date_mode("current")

        company_row = ctk.CTkFrame(self.tally_push_card, fg_color="transparent")
        company_row.pack(fill="x", padx=16, pady=(0, 6))
        company_row.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(company_row, text="Target Company", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).grid(row=0, column=0, sticky="w")
        self.tally_push_company_cb = ctk.CTkComboBox(
            company_row,
            values=[self.tally_push_company_placeholder],
            width=200,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
        )
        self.tally_push_company_cb.set(self.tally_push_company_placeholder)
        self.tally_push_company_cb.grid(row=0, column=1, sticky="ew", padx=(8, 6))

        self.tally_push_company_refresh_btn = ctk.CTkButton(
            company_row,
            text="Refresh",
            width=86,
            height=34,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._fetch_tally_companies_thread,
        )
        self.tally_push_company_refresh_btn.grid(row=0, column=2, sticky="e")

        self.tally_push_company_status = ctk.CTkLabel(
            self.tally_push_card,
            text="Companies: Not fetched",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.tally_push_company_status.pack(anchor="w", padx=16, pady=(0, 4))

        self.tally_push_conn_status = ctk.CTkLabel(
            self.tally_push_card,
            text="Connection: Not checked",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.tally_push_conn_status.pack(anchor="w", padx=16, pady=(0, 8))

        push_btn_row = ctk.CTkFrame(self.tally_push_card, fg_color="transparent")
        push_btn_row.pack(fill="x", padx=16, pady=(0, 14))
        self.tally_push_test_btn = ctk.CTkButton(
            push_btn_row,
            text="Test Connection",
            height=38,
            font=("Segoe UI", 11, "bold"),
            fg_color=COLORS["warning"],
            hover_color="#B45309",
            text_color="#FFFFFF",
            corner_radius=8,
            command=self._check_tally_connection_thread,
        )
        self.tally_push_test_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))

        self.tally_push_post_btn = ctk.CTkButton(
            push_btn_row,
            text="Step 2: Push XML",
            height=38,
            font=("Segoe UI", 12, "bold"),
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            text_color="#FFFFFF",
            corner_radius=8,
            command=self._post_tally_xml_thread,
        )
        self.tally_push_post_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # ─── LEFT: Create Ledger Card (hidden by default) ───
        self.create_ledger_card = ctk.CTkFrame(
            self.left_col,
            fg_color=COLORS["bg_card"],
            corner_radius=12,
            border_width=1,
            border_color=COLORS["border"],
        )
        ctk.CTkLabel(
            self.create_ledger_card,
            text="Create Ledger In Tally",
            font=("Segoe UI", 12, "bold"),
            text_color=COLORS["text_primary"],
        ).pack(anchor="w", padx=16, pady=(14, 8))

        create_conn_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        create_conn_row.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(create_conn_row, text="Host", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.create_ledger_host_entry = ctk.CTkEntry(
            create_conn_row,
            width=96,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_host_entry.insert(0, "localhost")
        self.create_ledger_host_entry.pack(side="left", padx=(6, 8))

        ctk.CTkLabel(create_conn_row, text="Port", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.create_ledger_port_entry = ctk.CTkEntry(
            create_conn_row,
            width=66,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_port_entry.insert(0, "9000")
        self.create_ledger_port_entry.pack(side="left", padx=(6, 8))

        ctk.CTkLabel(create_conn_row, text="Timeout", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.create_ledger_timeout_entry = ctk.CTkEntry(
            create_conn_row,
            width=70,
            height=32,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_timeout_entry.insert(0, "30")
        self.create_ledger_timeout_entry.pack(side="left", padx=(6, 0))

        create_company_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        create_company_row.pack(fill="x", padx=16, pady=(0, 6))
        create_company_row.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(create_company_row, text="Target Company", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).grid(row=0, column=0, sticky="w")
        self.create_ledger_company_cb = ctk.CTkComboBox(
            create_company_row,
            values=[self.tally_push_company_placeholder],
            width=180,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.create_ledger_company_cb.set(self.tally_push_company_placeholder)
        self.create_ledger_company_cb.grid(row=0, column=1, sticky="ew", padx=(8, 6))
        self.create_ledger_company_refresh_btn = ctk.CTkButton(
            create_company_row,
            text="Refresh",
            width=82,
            height=34,
            font=("Segoe UI", 9, "bold"),
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._fetch_create_ledger_companies_thread,
        )
        self.create_ledger_company_refresh_btn.grid(row=0, column=2, sticky="e")

        self.create_ledger_company_status = ctk.CTkLabel(
            self.create_ledger_card,
            text="Companies: Not fetched",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.create_ledger_company_status.pack(anchor="w", padx=16, pady=(0, 2))
        self.create_ledger_conn_status = ctk.CTkLabel(
            self.create_ledger_card,
            text="Connection: Not checked",
            font=("Segoe UI", 10),
            text_color=COLORS["text_muted"],
        )
        self.create_ledger_conn_status.pack(anchor="w", padx=16, pady=(0, 8))

        ctk.CTkLabel(
            self.create_ledger_card,
            text="Ledger Name",
            font=("Segoe UI", 10),
            text_color=COLORS["text_secondary"],
        ).pack(anchor="w", padx=16)
        self.create_ledger_name_entry = ctk.CTkEntry(
            self.create_ledger_card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Enter ledger name",
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_name_entry.pack(fill="x", padx=16, pady=(4, 8))

        parent_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        parent_row.pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkLabel(parent_row, text="Parent", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.create_ledger_parent_cb = ctk.CTkComboBox(
            parent_row,
            values=[
                "Sundry Creditors",
                "Purchase Accounts",
                "Duties & Taxes",
                "Indirect Expenses",
                "Current Liabilities",
                "Sundry Debtors",
            ],
            width=220,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.create_ledger_parent_cb.set("Sundry Creditors")
        self.create_ledger_parent_cb.pack(side="right")

        self.create_ledger_is_party_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            self.create_ledger_card,
            text="Party Ledger (Billwise On)",
            variable=self.create_ledger_is_party_var,
            font=("Segoe UI", 10, "bold"),
            text_color=COLORS["text_secondary"],
        ).pack(anchor="w", padx=16, pady=(0, 8))

        ctk.CTkLabel(self.create_ledger_card, text="Mailing Name", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        self.create_ledger_mailing_entry = ctk.CTkEntry(
            self.create_ledger_card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Optional mailing name",
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_mailing_entry.pack(fill="x", padx=16, pady=(4, 6))

        self.create_ledger_address1_entry = ctk.CTkEntry(
            self.create_ledger_card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Address line 1",
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_address1_entry.pack(fill="x", padx=16, pady=(0, 6))

        self.create_ledger_address2_entry = ctk.CTkEntry(
            self.create_ledger_card,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Address line 2",
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_address2_entry.pack(fill="x", padx=16, pady=(0, 6))

        geo_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        geo_row.pack(fill="x", padx=16, pady=(0, 6))
        self.create_ledger_state_entry = ctk.CTkComboBox(
            geo_row,
            values=[""] + LEDGER_STATE_OPTIONS,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_state_entry.set("")
        self.create_ledger_state_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.create_ledger_country_entry = ctk.CTkComboBox(
            geo_row,
            values=LEDGER_COUNTRY_OPTIONS,
            width=120,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            text_color=COLORS["text_primary"],
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_country_entry.set("India")
        self.create_ledger_country_entry.pack(side="left", padx=(4, 0))

        gst_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        gst_row.pack(fill="x", padx=16, pady=(0, 6))
        self.create_ledger_gstin_entry = ctk.CTkEntry(
            gst_row,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="GSTIN",
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_gstin_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.create_ledger_pincode_entry = ctk.CTkEntry(
            gst_row,
            width=120,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Pincode",
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_pincode_entry.pack(side="left", padx=(4, 0))

        pan_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        pan_row.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(pan_row, text="PAN / IT No.", font=("Segoe UI", 10),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.create_ledger_pan_entry = ctk.CTkEntry(
            pan_row,
            height=34,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Auto-filled from GSTIN",
            font=("Segoe UI", 10),
            corner_radius=8,
        )
        self.create_ledger_pan_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        def _on_create_ledger_gstin_change(_event=None):
            gstin_val = self.create_ledger_gstin_entry.get().strip().upper()
            pan_auto = _pan_from_gstin(gstin_val)
            if pan_auto and not self.create_ledger_pan_entry.get().strip():
                self.create_ledger_pan_entry.delete(0, "end")
                self.create_ledger_pan_entry.insert(0, pan_auto)
            if gstin_val and not self.create_ledger_state_entry.get().strip():
                inferred = _state_name_from_gstin(gstin_val)
                if inferred:
                    self.create_ledger_state_entry.set(inferred)

        self.create_ledger_gstin_entry.bind("<KeyRelease>", _on_create_ledger_gstin_change)
        self.create_ledger_gstin_entry.bind("<FocusOut>", _on_create_ledger_gstin_change)

        type_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        type_row.pack(fill="x", padx=16, pady=(0, 10))
        self.create_ledger_gst_app_cb = ctk.CTkComboBox(
            type_row,
            values=LEDGER_GST_APPLICABLE_OPTIONS,
            width=150,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.create_ledger_gst_app_cb.set("Applicable")
        self.create_ledger_gst_app_cb.pack(side="left", padx=(0, 6))

        self.create_ledger_reg_type_cb = ctk.CTkComboBox(
            type_row,
            values=["Regular", "Composition", "Unregistered", "Consumer", "Unknown"],
            width=170,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.create_ledger_reg_type_cb.set("Regular")
        self.create_ledger_reg_type_cb.pack(side="left", padx=(0, 6))
        self.create_ledger_billwise_cb = ctk.CTkComboBox(
            type_row,
            values=["Yes", "No"],
            width=110,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border"],
            button_color=COLORS["accent"],
            button_hover_color=COLORS["accent_hover"],
            font=("Segoe UI", 10),
        )
        self.create_ledger_billwise_cb.set("Yes")
        self.create_ledger_billwise_cb.pack(side="left")

        create_btn_row = ctk.CTkFrame(self.create_ledger_card, fg_color="transparent")
        create_btn_row.pack(fill="x", padx=16, pady=(0, 14))
        self.create_ledger_test_btn = ctk.CTkButton(
            create_btn_row,
            text="Test Connection",
            height=38,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._check_create_ledger_connection_thread,
        )
        self.create_ledger_test_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.create_ledger_create_btn = ctk.CTkButton(
            create_btn_row,
            text="Create Ledger",
            height=38,
            font=("Segoe UI", 10, "bold"),
            fg_color=COLORS["success"],
            hover_color="#047857",
            text_color="#FFFFFF",
            corner_radius=8,
            command=self._create_ledger_from_tab_thread,
        )
        self.create_ledger_create_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # ─── LEFT: Settings Card ───
        self.settings_card = ctk.CTkFrame(self.left_col, fg_color=COLORS["bg_card"], corner_radius=12,
                                      border_width=1, border_color=COLORS["border"])
        self.settings_card.pack(fill="x", pady=8)
        ctk.CTkLabel(self.settings_card, text="Settings", font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(anchor="w", padx=16, pady=(14, 10))
        ctk.CTkLabel(self.settings_card, text="Company Name (for Tally XML)", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        self.company_entry = ctk.CTkEntry(self.settings_card, height=36, fg_color=COLORS["bg_input"],
                                           border_color=COLORS["border"], text_color=COLORS["text_primary"],
                                           placeholder_text="Auto-detected from file",
                                           font=("Segoe UI", 11), corner_radius=8)
        self.company_entry.pack(fill="x", padx=16, pady=(4, 10))
        ctk.CTkLabel(self.settings_card, text="Company GSTIN (prevents Uncertain in GSTR-3B)", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        self.company_gstin_entry = ctk.CTkEntry(self.settings_card, height=36, fg_color=COLORS["bg_input"],
                                                border_color=COLORS["border"], text_color=COLORS["text_primary"],
                                                placeholder_text="e.g. 07ABDCS9065J1ZV  (auto-detected from GSTR-2B)",
                                                font=("Segoe UI", 11), corner_radius=8)
        self.company_gstin_entry.pack(fill="x", padx=16, pady=(4, 10))
        ctk.CTkLabel(self.settings_card, text="Default Purchase Ledger", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        self.purchase_ledger_entry = ctk.CTkEntry(self.settings_card, height=36, fg_color=COLORS["bg_input"],
                                                    border_color=COLORS["border"], text_color=COLORS["text_primary"],
                                                    font=("Segoe UI", 11), corner_radius=8)
        self.purchase_ledger_entry.pack(fill="x", padx=16, pady=(4, 10))
        self.purchase_ledger_entry.insert(0, "Purchase Account")
        ctk.CTkLabel(self.settings_card, text="Narration Template", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        self.narration_entry = ctk.CTkEntry(self.settings_card, height=36, fg_color=COLORS["bg_input"],
                                             border_color=COLORS["border"], text_color=COLORS["text_primary"],
                                             font=("Segoe UI", 11), corner_radius=8)
        self.narration_entry.pack(fill="x", padx=16, pady=(4, 10))
        self.narration_entry.insert(0, "Being purchase from {party} vide Inv {inv} dt {date}")

        # ── Auto Round Off ──
        ro_sep = ctk.CTkFrame(self.settings_card, fg_color=COLORS["border"], height=1)
        ro_sep.pack(fill="x", padx=16, pady=(4, 8))

        # Checkbox row
        ro_header = ctk.CTkFrame(self.settings_card, fg_color="transparent")
        ro_header.pack(fill="x", padx=16, pady=(0, 6))
        self.roundoff_enabled_var = ctk.BooleanVar(value=False)
        ro_cb = ctk.CTkCheckBox(ro_header, text="Enable Auto Round Off",
                                variable=self.roundoff_enabled_var,
                                font=("Segoe UI", 12), text_color=COLORS["text_primary"])
        ro_cb.pack(side="left")
        ctk.CTkLabel(ro_header, text="Fractional amounts auto-post to this ledger",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left", padx=10)

        # Entry + Fetch button row
        ro_entry_row = ctk.CTkFrame(self.settings_card, fg_color="transparent")
        ro_entry_row.pack(fill="x", padx=16, pady=(0, 10))
        self.roundoff_ledger_entry = ctk.CTkEntry(
            ro_entry_row, height=36, fg_color=COLORS["bg_input"],
            border_color=COLORS["border"], text_color=COLORS["text_primary"],
            placeholder_text="Search or type Round Off ledger name",
            font=("Segoe UI", 11), corner_radius=8, state="disabled")
        self.roundoff_ledger_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.roundoff_fetch_btn = ctk.CTkButton(
            ro_entry_row, text="Fetch Ledgers", width=130, height=36,
            fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"], font=("Segoe UI", 11),
            corner_radius=8, state="disabled")
        self.roundoff_fetch_btn.pack(side="left")

        # Toggle entry+button when checkbox is toggled
        def _toggle_ro():
            if self.roundoff_enabled_var.get():
                self.roundoff_ledger_entry.configure(state="normal")
                self.roundoff_fetch_btn.configure(state="normal")
            else:
                self.roundoff_ledger_entry.configure(state="disabled")
                self.roundoff_fetch_btn.configure(state="disabled")
                self.roundoff_ledger_entry.delete(0, "end")
        ro_cb.configure(command=_toggle_ro)

        # Fetch ledger button logic
        def _do_fetch_roundoff_main():
            try:
                tally_url = self._get_tally_push_url()
            except (ValueError, AttributeError):
                messagebox.showerror("Settings", "Invalid Tally URL. Check host/port in the Push panel.")
                return
            company = self._get_effective_push_company()
            cached = getattr(self, "tally_push_companies", [])
            if not company and len(cached) > 1:
                messagebox.showwarning(
                    "Select Company",
                    "Multiple companies are open in Tally.\n"
                    "Please select a target company from the company dropdown before fetching ledgers.")
                return
            self.roundoff_fetch_btn.configure(state="disabled", text="Fetching...")
            def _worker():
                result = _fetch_tally_ledgers(tally_url, timeout=15, company_name=company)
                def _done():
                    self.roundoff_fetch_btn.configure(state="normal", text="Fetch Ledgers")
                    if result.get("success"):
                        self.roundoff_all_ledgers = result.get("ledgers") or []
                        self._show_roundoff_ledger_picker()
                    else:
                        messagebox.showwarning("Fetch Failed",
                                               "Could not fetch ledgers from Tally.\nIs Tally running?")
                self.after(0, _done)
            threading.Thread(target=_worker, daemon=True).start()
        self.roundoff_fetch_btn.configure(command=_do_fetch_roundoff_main)

        # Click on entry → open picker if ledgers already loaded
        def _ro_entry_click(event):
            if self.roundoff_enabled_var.get() and self.roundoff_all_ledgers:
                self._show_roundoff_ledger_picker()
        self.roundoff_ledger_entry.bind("<Button-1>", _ro_entry_click)

        ctk.CTkLabel(self.settings_card, text="Output Folder", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=16)
        out_frame = ctk.CTkFrame(self.settings_card, fg_color="transparent")
        out_frame.pack(fill="x", padx=16, pady=(4, 14))
        self.output_entry = ctk.CTkEntry(out_frame, height=36, fg_color=COLORS["bg_input"],
                                          border_color=COLORS["border"], text_color=COLORS["text_primary"],
                                          placeholder_text="Same as source file",
                                          font=("Segoe UI", 11), corner_radius=8)
        self.output_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkButton(out_frame, text="...", width=40, height=36, fg_color=COLORS["bg_input"],
                      hover_color=COLORS["bg_card_hover"], text_color=COLORS["text_secondary"],
                      corner_radius=8, command=self._browse_output).pack(side="right")

        # ─── LEFT: Action Buttons ───
        self.action_card = ctk.CTkFrame(self.left_col, fg_color=COLORS["bg_card"], corner_radius=12,
                        border_width=1, border_color=COLORS["border"])
        self.action_card.pack(fill="x", pady=8)
        ctk.CTkLabel(self.action_card, text="Step 1C: Generate Output", font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["text_primary"]).pack(anchor="w", padx=16, pady=(14, 10))
        self.generate_btn = ctk.CTkButton(self.action_card, text="Step 1C: Generate Tally Sheet + XML",
                                           font=("Segoe UI", 13, "bold"), height=44,
                                           fg_color=COLORS["success"], hover_color="#047857",
                                           text_color="#FFFFFF", corner_radius=10,
                                           command=self._generate_output)
        self.generate_btn.pack(fill="x", padx=16, pady=(0, 6))
        # Keep these created (they are referenced elsewhere) but not packed
        self.excel_only_btn = ctk.CTkButton(self.action_card, text="Generate Tally Sheet Only",
                                             font=("Segoe UI", 12), height=38, fg_color=COLORS["bg_input"],
                                             hover_color=COLORS["bg_card_hover"], text_color=COLORS["text_primary"],
                                             corner_radius=8, command=lambda: self._generate_output(xml=False))
        self.xml_only_btn = ctk.CTkButton(self.action_card, text="Generate XML Only",
                                           font=("Segoe UI", 12), height=38, fg_color=COLORS["bg_input"],
                                           hover_color=COLORS["bg_card_hover"], text_color=COLORS["text_primary"],
                                           corner_radius=8, command=lambda: self._generate_output(excel=False))

        # ─── Inline Push to Tally Panel ───────────────────────────────────────
        ctk.CTkFrame(self.action_card, fg_color=COLORS["border"], height=1).pack(fill="x", padx=16, pady=(4, 0))
        ctk.CTkLabel(self.action_card, text="Push to Tally",
                     font=("Segoe UI", 11, "bold"), text_color=COLORS["accent"]).pack(anchor="w", padx=16, pady=(8, 4))

        # Host / Port / Timeout row
        _ipr = ctk.CTkFrame(self.action_card, fg_color="transparent")
        _ipr.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(_ipr, text="Host", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.inline_push_host_entry = ctk.CTkEntry(_ipr, width=104, height=32,
            fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], font=("Segoe UI", 11), corner_radius=8)
        self.inline_push_host_entry.insert(0, "localhost")
        self.inline_push_host_entry.pack(side="left", padx=(6, 10))
        ctk.CTkLabel(_ipr, text="Port", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.inline_push_port_entry = ctk.CTkEntry(_ipr, width=68, height=32,
            fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], font=("Segoe UI", 11), corner_radius=8)
        self.inline_push_port_entry.insert(0, "9000")
        self.inline_push_port_entry.pack(side="left", padx=(6, 10))
        ctk.CTkLabel(_ipr, text="Timeout", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).pack(side="left")
        self.inline_push_timeout_entry = ctk.CTkEntry(_ipr, width=54, height=32,
            fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], font=("Segoe UI", 11), corner_radius=8)
        self.inline_push_timeout_entry.insert(0, "30")
        self.inline_push_timeout_entry.pack(side="left", padx=(6, 0))

        # Date mode row — shares same BooleanVars as Push To Tally tab
        _dmf = ctk.CTkFrame(self.action_card, fg_color=COLORS["bg_input"], corner_radius=8)
        _dmf.pack(fill="x", padx=16, pady=(0, 6))
        ctk.CTkLabel(_dmf, text="Voucher Date Mode", font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["text_secondary"]).pack(anchor="w", padx=10, pady=(8, 4))
        _dcr = ctk.CTkFrame(_dmf, fg_color="transparent")
        _dcr.pack(fill="x", padx=10, pady=(0, 4))
        ctk.CTkCheckBox(_dcr, text="Current Date", variable=self.tally_push_date_checks["current"],
            command=lambda: self._set_tally_push_date_mode("current"),
            font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left", padx=(0, 8))
        ctk.CTkCheckBox(_dcr, text="Excel Date", variable=self.tally_push_date_checks["excel"],
            command=lambda: self._set_tally_push_date_mode("excel"),
            font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left", padx=(0, 8))
        ctk.CTkCheckBox(_dcr, text="Custom Date", variable=self.tally_push_date_checks["custom"],
            command=lambda: self._set_tally_push_date_mode("custom"),
            font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).pack(side="left")
        _cdr = ctk.CTkFrame(_dmf, fg_color="transparent")
        _cdr.pack(fill="x", padx=10, pady=(0, 8))
        ctk.CTkEntry(_cdr, textvariable=self.tally_push_custom_date_var, height=30,
            fg_color=COLORS["bg_card"], border_color=COLORS["border"],
            text_color=COLORS["text_primary"], placeholder_text="Custom Date (DD/MM/YYYY)",
            font=("Segoe UI", 10), corner_radius=6).pack(fill="x")

        # Company selector row
        _cr2 = ctk.CTkFrame(self.action_card, fg_color="transparent")
        _cr2.pack(fill="x", padx=16, pady=(0, 4))
        _cr2.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(_cr2, text="Target Company", font=("Segoe UI", 11),
                     text_color=COLORS["text_secondary"]).grid(row=0, column=0, sticky="w")
        self.inline_push_company_cb = ctk.CTkComboBox(
            _cr2, values=[self.tally_push_company_placeholder], width=200, height=34,
            fg_color=COLORS["bg_input"], border_color=COLORS["border"],
            button_color=COLORS["accent"], button_hover_color=COLORS["accent_hover"])
        self.inline_push_company_cb.set(self.tally_push_company_placeholder)
        self.inline_push_company_cb.grid(row=0, column=1, sticky="ew", padx=(8, 6))
        self.inline_push_company_refresh_btn = ctk.CTkButton(
            _cr2, text="Fetch", width=60, height=34,
            font=("Segoe UI", 10, "bold"), fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"], text_color=COLORS["text_secondary"],
            corner_radius=8, command=self._inline_push_refresh_companies_thread)
        self.inline_push_company_refresh_btn.grid(row=0, column=2, sticky="e")

        self.inline_push_company_status = ctk.CTkLabel(
            self.action_card, text="Companies: Not fetched",
            font=("Segoe UI", 10), text_color=COLORS["text_muted"])
        self.inline_push_company_status.pack(anchor="w", padx=16, pady=(0, 2))
        self.inline_push_conn_status = ctk.CTkLabel(
            self.action_card, text="Connection: Not checked",
            font=("Segoe UI", 10), text_color=COLORS["text_muted"])
        self.inline_push_conn_status.pack(anchor="w", padx=16, pady=(0, 6))

        # Action buttons
        _ibr = ctk.CTkFrame(self.action_card, fg_color="transparent")
        _ibr.pack(fill="x", padx=16, pady=(0, 14))
        self.inline_push_test_btn = ctk.CTkButton(
            _ibr, text="Test Connection", height=40,
            font=("Segoe UI", 11, "bold"), fg_color=COLORS["warning"], hover_color="#B45309",
            text_color="#FFFFFF", corner_radius=8, command=self._inline_push_test_connection)
        self.inline_push_test_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.inline_push_post_btn = ctk.CTkButton(
            _ibr, text="Generate & Push to Tally", height=40,
            font=("Segoe UI", 12, "bold"), fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
            text_color="#FFFFFF", corner_radius=8, command=self._generate_and_push_inline)
        self.inline_push_post_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # Progress
        self.progress_frame = ctk.CTkFrame(self.left_col, fg_color=COLORS["bg_card"], corner_radius=12,
                                            border_width=1, border_color=COLORS["border"])
        self.progress_frame.pack(fill="x", pady=8)
        self.progress_label = ctk.CTkLabel(self.progress_frame, text="Ready", font=("Segoe UI", 11),
                                            text_color=COLORS["text_muted"])
        self.progress_label.pack(anchor="w", padx=16, pady=(10, 4))
        self.progress_bar = ctk.CTkProgressBar(self.progress_frame, height=8, fg_color=COLORS["bg_dark"],
                                                 progress_color=COLORS["accent"], corner_radius=4)
        self.progress_bar.pack(fill="x", padx=16, pady=(0, 12)); self.progress_bar.set(0)

        # ═══ RIGHT COLUMN ═══
        self.stats_frame = ctk.CTkFrame(self.right_col, fg_color="transparent", height=100)
        self.stats_frame.pack(fill="x", pady=(0, 8)); self.stats_frame.pack_propagate(False)
        self.stat_total = StatsCard(self.stats_frame, "N", "Total Invoices", "—", accent=COLORS["accent"])
        self.stat_total.pack(side="left", fill="both", expand=True, padx=(0, 4))
        self.stat_taxable = StatsCard(self.stats_frame, "$", "Taxable Value", "—", accent=COLORS["tally_gold"])
        self.stat_taxable.pack(side="left", fill="both", expand=True, padx=4)
        self.stat_igst = StatsCard(self.stats_frame, "I", "IGST Total", "—", accent="#3B82F6")
        self.stat_igst.pack(side="left", fill="both", expand=True, padx=4)
        self.stat_gst = StatsCard(self.stats_frame, "G", "CGST + SGST", "—", accent=COLORS["warning"])
        self.stat_gst.pack(side="left", fill="both", expand=True, padx=(4, 0))
        self.preview_table = DataPreviewTable(self.right_col)
        self.preview_table.pack(fill="both", expand=True, pady=(0, 8))
        self.log_panel = LogPanel(self.right_col, height=180)
        self.log_panel.pack(fill="x")

        # Download 2B captcha panel (shown only when captcha is required)
        self.download2b_captcha_frame = ctk.CTkFrame(
            self.right_col,
            fg_color=COLORS["warning_bg"],
            border_width=1,
            border_color=COLORS["warning"],
            corner_radius=10,
        )
        ctk.CTkLabel(
            self.download2b_captcha_frame,
            text="CAPTCHA ACTION REQUIRED",
            font=("Segoe UI", 13, "bold"),
            text_color=COLORS["warning"],
        ).pack(anchor="w", padx=12, pady=(10, 6))
        self.download2b_captcha_img_label = ctk.CTkLabel(self.download2b_captcha_frame, text="")
        self.download2b_captcha_img_label.pack(padx=12, pady=(0, 8))
        row = ctk.CTkFrame(self.download2b_captcha_frame, fg_color="transparent")
        row.pack(fill="x", padx=12, pady=(0, 12))
        self.download2b_captcha_entry = ctk.CTkEntry(
            row,
            height=36,
            fg_color=COLORS["bg_card"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
            placeholder_text="Type captcha",
            font=("Consolas", 16),
        )
        self.download2b_captcha_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self.download2b_captcha_entry.bind("<Return>", self._submit_download2b_captcha)
        self.download2b_captcha_submit_btn = ctk.CTkButton(
            row,
            text="Submit",
            width=92,
            height=36,
            font=("Segoe UI", 11, "bold"),
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            text_color="#FFFFFF",
            command=self._submit_download2b_captcha,
        )
        self.download2b_captcha_submit_btn.pack(side="right")

        # ═══ BOTTOM BAR ═══
        bottom = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], height=36, corner_radius=0)
        bottom.pack(fill="x", side="bottom"); bottom.pack_propagate(False)
        ctk.CTkLabel(bottom, text="Studycafe PVT LTD  |  GSTR-2B + Tally Sheet → Tally v4.1",
                     font=("Segoe UI", 10), text_color=COLORS["text_muted"]).pack(side="left", padx=20)
        self.status_label = ctk.CTkLabel(bottom, text="Ready", font=("Segoe UI", 10),
                                          text_color=COLORS["success"])
        self.status_label.pack(side="right", padx=20)

        self.log_panel.log("Application started. Select source type and upload files.", "info")
        self.log_panel.log("GSTR-2B mode: Auto-detects columns + optional party→ledger mapping.", "info")
        self.log_panel.log("Push To Tally mode: Post XML to TallyPrime at localhost:9000.", "info")
        self.log_panel.log("Create Ledger mode: Create a ledger in Tally using the same connection flow.", "info")

    # ─── MODE SWITCHING ───

    def _view_workflow_demo(self):
        demo_url = (self.workflow_demo_url or "").strip()
        if demo_url:
            try:
                webbrowser.open_new_tab(demo_url)
                self.log_panel.log("Opened workflow demo in browser.", "info")
            except Exception as exc:
                messagebox.showerror("View Demo", f"Could not open demo link.\n\n{exc}")
            return

        messagebox.showinfo(
            "View Demo",
            "Demo link is not set yet.\n\n"
            "Set self.workflow_demo_url in code to your YouTube link later.",
        )

    def _refresh_mode_selector_text_colors(self, selected):
        for label, btn in self.mode_buttons.items():
            if label == selected:
                btn.configure(
                    fg_color=COLORS["accent"],
                    hover_color=COLORS["accent_hover"],
                    text_color="#FFFFFF",
                )
            else:
                btn.configure(
                    fg_color=COLORS["bg_input"],
                    hover_color=COLORS["bg_card_hover"],
                    text_color=COLORS["text_primary"],
                )

    def _on_mode_change(self, selected):
        self._refresh_mode_selector_text_colors(selected)
        if selected == "Step 1: GSTR-2B → XML":
            self.current_mode = "gstr2b"
            self.download2b_card.pack_forget()
            self.tally_push_card.pack_forget()
            self.create_ledger_card.pack_forget()
            self.download2b_captcha_frame.pack_forget()
            self.tally_card.pack_forget()
            if not self.settings_card.winfo_manager():
                self.settings_card.pack(fill="x", pady=8)
            if not self.action_card.winfo_manager():
                self.action_card.pack(fill="x", pady=8)
            if not self.progress_frame.winfo_manager():
                self.progress_frame.pack(fill="x", pady=8)
            if not self.stats_frame.winfo_manager():
                self.stats_frame.pack(fill="x", pady=(0, 8))
            if not self.preview_table.winfo_manager():
                self.preview_table.pack(fill="both", expand=True, pady=(0, 8), in_=self.right_col, before=self.log_panel)
            self.gstr2b_card.pack(fill="x", pady=(0, 8), in_=self.left_col, before=self.settings_card)
            self.generate_btn.configure(text="Step 1C: Generate Tally Sheet + XML")
            self.log_panel.log("Switched to Step 1: GSTR-2B → XML.", "info")
        elif selected == "Create Ledger":
            self.current_mode = "create_ledger"
            self.gstr2b_card.pack_forget()
            self.tally_card.pack_forget()
            self.download2b_card.pack_forget()
            self.tally_push_card.pack_forget()
            self.download2b_captcha_frame.pack_forget()
            self.settings_card.pack_forget()
            self.action_card.pack_forget()
            if not self.progress_frame.winfo_manager():
                self.progress_frame.pack(fill="x", pady=8)
            self.stats_frame.pack_forget()
            self.preview_table.pack_forget()
            self.create_ledger_card.pack(fill="x", pady=(0, 8), in_=self.left_col)
            self.progress_bar.set(0)
            self.progress_label.configure(text="Ready to create ledger in Tally", text_color=COLORS["text_muted"])
            self.status_label.configure(text="Ready", text_color=COLORS["success"])
            self.log_panel.log("Switched to Create Ledger mode.", "info")
            self.after(120, lambda: self._fetch_create_ledger_companies_thread(silent=True))
        elif selected == "Step 2: Push To Tally":
            self.current_mode = "push_tally"
            self.gstr2b_card.pack_forget()
            self.tally_card.pack_forget()
            self.download2b_card.pack_forget()
            self.create_ledger_card.pack_forget()
            self.download2b_captcha_frame.pack_forget()
            self.settings_card.pack_forget()
            self.action_card.pack_forget()
            if not self.progress_frame.winfo_manager():
                self.progress_frame.pack(fill="x", pady=8)
            self.stats_frame.pack_forget()
            self.preview_table.pack_forget()
            self.tally_push_card.pack(fill="x", pady=(0, 8), in_=self.left_col)
            self.progress_bar.set(0)
            self.progress_label.configure(text="Ready to test/push XML to Tally", text_color=COLORS["text_muted"])
            self.status_label.configure(text="Ready", text_color=COLORS["success"])
            self.log_panel.log("Switched to Step 2: Push To Tally.", "info")
            self.log_panel.log("Tip: Keep Tally open with HTTP server enabled (default port 9000).", "info")
            self.after(150, lambda: self._fetch_tally_companies_thread(silent=True))
        else:
            self.log_panel.log(f"Unknown mode selection: {selected}", "warning")

    def _browse_download2b_excel(self):
        filepath = filedialog.askopenfilename(
            title="Select Credentials Excel",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if filepath:
            self.download2b_excel_file = filepath
            self.download2b_file_entry.delete(0, "end")
            self.download2b_file_entry.insert(0, filepath)
            self.log_panel.log(f"Download 2B credentials selected: {Path(filepath).name}", "info")

    def _on_d2b_input_mode_change(self, value):
        if value == "Excel File":
            self.d2b_manual_frame.pack_forget()
            self.d2b_excel_frame.pack(fill="x", before=self.download2b_year_row_ref)
        else:
            self.d2b_excel_frame.pack_forget()
            self.d2b_manual_frame.pack(fill="x", before=self.download2b_year_row_ref)

    def _add_d2b_manual_entry(self):
        row = ctk.CTkFrame(self.d2b_manual_list, fg_color="transparent")
        row.pack(fill="x", pady=2)
        user_ent = ctk.CTkEntry(row, placeholder_text="Username", height=28, fg_color=COLORS["bg_card"],
                                border_color=COLORS["border"], text_color=COLORS["text_primary"], width=100)
        user_ent.pack(side="left", fill="x", expand=True, padx=(0, 2))
        pass_ent = ctk.CTkEntry(row, placeholder_text="Password", show="*", height=28, fg_color=COLORS["bg_card"],
                                border_color=COLORS["border"], text_color=COLORS["text_primary"], width=100)
        pass_ent.pack(side="left", fill="x", expand=True, padx=2)
        
        def remove_row():
            if len(self.d2b_manual_entries) > 1:
                row.destroy()
                self.d2b_manual_entries.remove((user_ent, pass_ent))
                
        rem_btn = ctk.CTkButton(row, text="✕", width=28, height=28, fg_color=COLORS["error_bg"], 
                                text_color=COLORS["error"], hover_color="#FECACA", command=remove_row)
        rem_btn.pack(side="left", padx=(2, 0))
        self.d2b_manual_entries.append((user_ent, pass_ent))

    def _toggle_download2b_inputs(self):
        state = "disabled" if self.download2b_all_quarters_var.get() else "normal"
        self.download2b_quarter_cb.configure(state=state)
        self.download2b_month_cb.configure(state=state)

    def _update_download2b_months(self, choice):
        if "Quarter 1" in choice:
            vals = ["Whole Quarter", "April", "May", "June"]
        elif "Quarter 2" in choice:
            vals = ["Whole Quarter", "July", "August", "September"]
        elif "Quarter 3" in choice:
            vals = ["Whole Quarter", "October", "November", "December"]
        elif "Quarter 4" in choice:
            vals = ["Whole Quarter", "January", "February", "March"]
        else:
            vals = ["Whole Quarter"]
        self.download2b_month_cb.configure(values=vals)
        self.download2b_month_cb.set(vals[0])

    def _start_download2b_process(self):
        input_mode = self.d2b_input_mode.get()
        
        credentials_data = None
        if input_mode == "Excel File":
            excel_path = self.download2b_file_entry.get().strip() or getattr(self, "download2b_excel_file", "")
            if not excel_path:
                messagebox.showerror("Missing File", "Please select credentials Excel file first.")
                return
            if not os.path.exists(excel_path):
                messagebox.showerror("Missing File", "Selected credentials Excel file does not exist.")
                return
            credentials_data = excel_path
        else:
            # Manual Entry Mode
            creds = []
            for u_ent, p_ent in self.d2b_manual_entries:
                u, p = u_ent.get().strip(), p_ent.get().strip()
                if u and p:
                    creds.append({"Username": u, "Password": p})
            
            if not creds:
                messagebox.showerror("No Credentials", "Please add at least one username and password.")
                return
            credentials_data = creds

        if getattr(self, "download2b_worker", None) and self.download2b_worker.keep_running:
            messagebox.showinfo("Already Running", "Download process is already running.")
            return

        settings = {
            "year": self.download2b_year_cb.get(),
            "month": self.download2b_month_cb.get(),
            "quarter": self.download2b_quarter_cb.get(),
            "all_quarters": self.download2b_all_quarters_var.get(),
        }

        self.progress_bar.set(0)
        self.progress_label.configure(text="Starting Download 2B...", text_color=COLORS["warning"])
        self.status_label.configure(text="Running", text_color=COLORS["warning"])
        self.download2b_start_btn.configure(state="disabled", text="RUNNING...")
        self.download2b_stop_btn.configure(state="normal")
        self.download2b_worker = GST2BDownloadWorker(self, credentials_data, settings)
        threading.Thread(target=self.download2b_worker.run, daemon=True).start()

    def _stop_download2b_process(self):
        if self.download2b_worker:
            self.download2b_worker.keep_running = False
            # Set the captcha event so it doesn't block indefinitely waiting for user input
            if hasattr(self.download2b_worker, 'captcha_event'):
                self.download2b_worker.captcha_event.set()
            self.log_panel.log("Stop requested for Download 2B process.", "warning")
            self.download2b_stop_btn.configure(state="disabled")

    def _submit_download2b_captcha(self, event=None):
        _ = event
        txt = self.download2b_captcha_entry.get().strip()
        if not txt:
            return
        if not self.download2b_worker:
            return
        self.download2b_captcha_submit_btn.configure(state="disabled", text="VERIFYING...")
        self.download2b_worker.captcha_response = txt
        self.download2b_worker.captcha_event.set()

    def update_download2b_log_safe(self, msg):
        self.after(0, lambda m=msg: self.log_panel.log(m, "info"))

    def update_download2b_progress_safe(self, val):
        self.after(0, lambda p=val: self.progress_bar.set(p))

    def process_download2b_finished_safe(self, msg):
        def done():
            self.status_label.configure(text="Ready", text_color=COLORS["success"])
            self.progress_label.configure(text=msg, text_color=COLORS["success"])
            self.download2b_start_btn.configure(state="normal", text="Start Download")
            self.download2b_stop_btn.configure(state="disabled")
            self.download2b_captcha_frame.pack_forget()
            messagebox.showinfo("Download 2B", msg)
        self.after(0, done)

    def request_download2b_captcha_safe(self, img_path):
        def show():
            try:
                pil_img = Image.open(img_path)
                self._download2b_captcha_img = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(200, 70))
                self.download2b_captcha_img_label.configure(image=self._download2b_captcha_img)
                if not self.download2b_captcha_frame.winfo_manager():
                    self.download2b_captcha_frame.pack(fill="x", pady=(0, 8), in_=self.right_col, before=self.log_panel)
                self.download2b_captcha_submit_btn.configure(state="normal", text="Submit")
                self.download2b_captcha_entry.delete(0, "end")
                self.download2b_captcha_entry.focus_set()
            except Exception as e:
                self.log_panel.log(f"Captcha display error: {e}", "error")
        self.after(0, show)

    def close_download2b_captcha_safe(self):
        self.after(0, lambda: self.download2b_captcha_frame.pack_forget())

    # ─── CREATE LEDGER ───

    def _set_create_ledger_widget_value(self, widget, value):
        text_value = _normalize_ledger_name(value or "")
        try:
            widget.set(text_value)
            return
        except Exception:
            pass
        try:
            widget.delete(0, "end")
            widget.insert(0, text_value)
        except Exception:
            pass

    def _get_create_ledger_url(self):
        host = (self.create_ledger_host_entry.get() or "localhost").strip()
        port = (self.create_ledger_port_entry.get() or "9000").strip()
        if host.startswith("http://"):
            host = host[7:]
        elif host.startswith("https://"):
            host = host[8:]
        host = host.strip("/") or "localhost"
        if "/" in host:
            host = host.split("/", 1)[0]
        if not port.isdigit():
            raise ValueError("Port must be numeric.")
        return f"http://{host}:{port}"

    def _get_create_ledger_timeout(self):
        timeout_txt = (self.create_ledger_timeout_entry.get() or "30").strip()
        try:
            timeout_val = float(timeout_txt)
        except ValueError as exc:
            raise ValueError("Timeout must be a number.") from exc
        if timeout_val <= 0:
            raise ValueError("Timeout must be greater than 0.")
        return timeout_val

    def _set_create_ledger_running_ui(self, running, label_text=None, label_color=None):
        self.create_ledger_is_running = running
        if running:
            if hasattr(self, "create_ledger_fetch_btn"):
                self.create_ledger_fetch_btn.configure(state="disabled")
            self.create_ledger_test_btn.configure(state="disabled")
            self.create_ledger_create_btn.configure(state="disabled", text="PROCESSING...")
            self.create_ledger_company_refresh_btn.configure(state="disabled")
            self.create_ledger_company_cb.configure(state="disabled")
        else:
            if hasattr(self, "create_ledger_fetch_btn"):
                self.create_ledger_fetch_btn.configure(state="normal")
            self.create_ledger_test_btn.configure(state="normal")
            self.create_ledger_create_btn.configure(state="normal", text="Create Ledger")
            self.create_ledger_company_refresh_btn.configure(state="normal")
            self.create_ledger_company_cb.configure(state="normal")
        if label_text is not None:
            self.progress_label.configure(text=label_text, text_color=label_color or COLORS["text_muted"])

    def _set_create_ledger_company_dropdown(self, companies, keep_selection=True):
        current_value = ""
        if keep_selection:
            current_value = _normalize_company_name(self.create_ledger_company_cb.get() or "")

        cleaned = []
        seen = set()
        for name in companies or []:
            txt = _normalize_company_name(name)
            if not _is_valid_company_name(txt):
                continue
            key = _company_key(txt)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(txt)

        cleaned = sorted(cleaned, key=lambda x: _company_key(x))
        values = [self.tally_push_company_placeholder] + cleaned
        self.create_ledger_company_cb.configure(values=values)

        if current_value and _company_key(current_value) in {_company_key(x) for x in cleaned}:
            self.create_ledger_company_cb.set(current_value)
        else:
            self.create_ledger_company_cb.set(self.tally_push_company_placeholder)

        self.create_ledger_companies = cleaned
        self.create_ledger_company_status.configure(
            text=f"Companies: {len(cleaned)} available",
            text_color=COLORS["text_muted"],
        )

    def _get_selected_create_ledger_company(self):
        selected = _normalize_company_name(self.create_ledger_company_cb.get() or "")
        if not selected:
            return ""
        if _company_key(selected) == _company_key(self.tally_push_company_placeholder):
            return ""
        return selected

    def _fetch_create_ledger_companies_thread(self, silent=False):
        if self.create_ledger_is_running:
            return

        try:
            tally_url = self._get_create_ledger_url()
            timeout = self._get_create_ledger_timeout()
        except ValueError as e:
            if not silent:
                messagebox.showerror("Invalid Settings", str(e))
            return

        self.create_ledger_company_refresh_btn.configure(state="disabled", text="Fetching...")
        if not silent:
            self.create_ledger_company_status.configure(text="Companies: Fetching...", text_color=COLORS["warning"])

        threading.Thread(
            target=self._fetch_create_ledger_companies_worker,
            args=(tally_url, timeout, silent),
            daemon=True,
        ).start()

    def _fetch_create_ledger_companies_worker(self, tally_url, timeout, silent):
        result = _fetch_tally_companies(tally_url, timeout=min(timeout, 20))

        def done():
            self.create_ledger_company_refresh_btn.configure(state="normal", text="Refresh")
            if result.get("success"):
                companies = result.get("companies", [])
                self._set_create_ledger_company_dropdown(companies, keep_selection=True)
                self.log_panel.log(f"Create Ledger: fetched Tally companies ({len(companies)})", "info")
            else:
                err = str(result.get("error") or "Unknown error")
                self.create_ledger_company_status.configure(text="Companies: Fetch failed", text_color=COLORS["warning"])
                self.log_panel.log(f"Create Ledger: company fetch failed - {err}", "warning")
                if not silent:
                    messagebox.showwarning("Company Fetch Failed", f"Could not fetch companies from Tally.\n\n{err}")

        self.after(0, done)

    def _check_create_ledger_connection_thread(self):
        if self.create_ledger_is_running:
            return
        try:
            tally_url = self._get_create_ledger_url()
            timeout = self._get_create_ledger_timeout()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return

        self._set_create_ledger_running_ui(True, "Checking Tally connection...", COLORS["warning"])
        self.status_label.configure(text="Checking", text_color=COLORS["warning"])
        self.create_ledger_conn_status.configure(text="Connection: Checking...", text_color=COLORS["warning"])
        threading.Thread(
            target=self._check_create_ledger_connection_worker,
            args=(tally_url, timeout),
            daemon=True,
        ).start()

    def _check_create_ledger_connection_worker(self, tally_url, timeout):
        check_result = _check_tally_connection(tally_url, timeout=min(timeout, 10))
        ok = check_result.get("connected", False)
        error_msg = ""
        if not ok:
            if check_result.get("error") == "ConnectionError":
                error_msg = "Cannot connect to Tally. Check if Tally is running and HTTP server is enabled."
            elif check_result.get("error") == "Timeout":
                error_msg = "Connection check timed out."
            elif "status_code" in check_result:
                error_msg = f"HTTP {check_result.get('status_code')}"
            else:
                error_msg = str(check_result.get("error") or "Unknown error")

        def done():
            self._set_create_ledger_running_ui(False, "Ready to create ledger in Tally", COLORS["text_muted"])
            self.status_label.configure(text="Ready", text_color=COLORS["success"])
            if ok:
                self.create_ledger_conn_status.configure(text=f"Connection: Connected ({tally_url})", text_color=COLORS["success"])
                self.log_panel.log(f"Create Ledger: Tally connection successful at {tally_url}", "success")
                self._fetch_create_ledger_companies_thread(silent=True)
            else:
                self.create_ledger_conn_status.configure(text="Connection: Offline", text_color=COLORS["error"])
                self.log_panel.log(f"Create Ledger: connection failed - {error_msg}", "error")

        self.after(0, done)

    def _fetch_create_ledger_from_gstin_thread(self):
        if self.create_ledger_is_running:
            return

        gstin = _normalize_ledger_name(self.create_ledger_gstin_entry.get() or "").upper()
        if len(gstin) < 15:
            messagebox.showwarning("Invalid GSTIN", "Please enter a valid 15-digit GSTIN first.")
            return

        try:
            tally_url = self._get_create_ledger_url()
            timeout = self._get_create_ledger_timeout()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return

        selected_company = self._get_selected_create_ledger_company()
        if not _normalize_ledger_name(self.create_ledger_country_entry.get() or ""):
            self._set_create_ledger_widget_value(self.create_ledger_country_entry, "India")

        if hasattr(self, "create_ledger_fetch_btn"):
            self.create_ledger_fetch_btn.configure(state="disabled", text="Fetching...")
        self.log_panel.log(f"Create Ledger: trying GST lookup for {gstin}", "process")

        threading.Thread(
            target=self._fetch_create_ledger_from_gstin_worker,
            args=(tally_url, timeout, gstin, selected_company),
            daemon=True,
        ).start()

    def _fetch_create_ledger_from_gstin_worker(self, tally_url, timeout, gstin, selected_company):
        fetch_result = _fetch_tally_party_details_by_gstin(
            tally_url,
            gstin,
            timeout=min(timeout, 20),
            company_name=selected_company,
        )

        def done():
            if hasattr(self, "create_ledger_fetch_btn"):
                self.create_ledger_fetch_btn.configure(state="normal", text="Fetch GSTIN")

            if fetch_result.get("success"):
                fetched = fetch_result.get("details", {})
                if fetched.get("mailing_name"):
                    self._set_create_ledger_widget_value(self.create_ledger_mailing_entry, fetched.get("mailing_name"))
                if fetched.get("address1"):
                    self._set_create_ledger_widget_value(self.create_ledger_address1_entry, fetched.get("address1"))
                if fetched.get("address2"):
                    self._set_create_ledger_widget_value(self.create_ledger_address2_entry, fetched.get("address2"))
                if fetched.get("state"):
                    self._set_create_ledger_widget_value(self.create_ledger_state_entry, fetched.get("state"))
                if fetched.get("country"):
                    self._set_create_ledger_widget_value(self.create_ledger_country_entry, fetched.get("country"))
                if fetched.get("pincode"):
                    self._set_create_ledger_widget_value(self.create_ledger_pincode_entry, fetched.get("pincode"))
                if fetched.get("gstin"):
                    self._set_create_ledger_widget_value(self.create_ledger_gstin_entry, fetched.get("gstin"))
                if fetched.get("gst_applicable"):
                    self._set_create_ledger_widget_value(self.create_ledger_gst_app_cb, fetched.get("gst_applicable"))
                if fetched.get("reg_type"):
                    self._set_create_ledger_widget_value(self.create_ledger_reg_type_cb, fetched.get("reg_type"))
                if fetched.get("billwise"):
                    self._set_create_ledger_widget_value(self.create_ledger_billwise_cb, fetched.get("billwise"))
                self.log_panel.log("Create Ledger: GST details fetched from Tally data.", "success")
                return

            inferred_state = _state_name_from_gstin(gstin)
            if inferred_state and not _normalize_ledger_name(self.create_ledger_state_entry.get() or ""):
                self._set_create_ledger_widget_value(self.create_ledger_state_entry, inferred_state)
            if not _normalize_ledger_name(self.create_ledger_country_entry.get() or ""):
                self._set_create_ledger_widget_value(self.create_ledger_country_entry, "India")
            self._set_create_ledger_widget_value(self.create_ledger_gst_app_cb, "Applicable")
            self._set_create_ledger_widget_value(self.create_ledger_reg_type_cb, "Regular")
            self.log_panel.log(
                f"Create Ledger: GST lookup fallback used ({fetch_result.get('error', 'Unknown')}).",
                "warning",
            )

        self.after(0, done)

    def _create_ledger_from_tab_thread(self):
        if self.create_ledger_is_running:
            return

        ledger_name = _normalize_ledger_name(self.create_ledger_name_entry.get() or "")
        if not ledger_name:
            messagebox.showwarning("Missing Ledger Name", "Please enter ledger name.")
            return

        try:
            tally_url = self._get_create_ledger_url()
            timeout = self._get_create_ledger_timeout()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return

        parent_name = _normalize_ledger_name(self.create_ledger_parent_cb.get() or "") or "Sundry Creditors"
        selected_company = self._get_selected_create_ledger_company()
        gstin = _normalize_ledger_name(self.create_ledger_gstin_entry.get() or "").upper()
        state_name = _normalize_ledger_name(self.create_ledger_state_entry.get() or "")
        if not state_name and gstin:
            state_name = _state_name_from_gstin(gstin)

        extra_info = {
            "mailing_name": _normalize_ledger_name(self.create_ledger_mailing_entry.get() or "") or ledger_name,
            "address1": _normalize_ledger_name(self.create_ledger_address1_entry.get() or ""),
            "address2": _normalize_ledger_name(self.create_ledger_address2_entry.get() or ""),
            "state": state_name,
            "country": _normalize_ledger_name(self.create_ledger_country_entry.get() or "") or "India",
            "pincode": _normalize_ledger_name(self.create_ledger_pincode_entry.get() or ""),
            "gstin": gstin,
            "pan": _normalize_ledger_name(self.create_ledger_pan_entry.get() or "").upper() or _pan_from_gstin(gstin),
            "gst_applicable": _normalize_ledger_name(self.create_ledger_gst_app_cb.get() or "") or ("Applicable" if gstin else "Not Applicable"),
            "reg_type": _normalize_ledger_name(self.create_ledger_reg_type_cb.get() or "") or ("Regular" if gstin else "Unknown"),
            "billwise": _normalize_ledger_name(self.create_ledger_billwise_cb.get() or "") or ("Yes" if self.create_ledger_is_party_var.get() else "No"),
        }
        is_party = bool(self.create_ledger_is_party_var.get())

        self._set_create_ledger_running_ui(True, "Creating ledger in Tally...", COLORS["warning"])
        self.status_label.configure(text="Creating", text_color=COLORS["warning"])
        self.progress_bar.set(0.3)
        self.log_panel.log(f"Create Ledger: creating '{ledger_name}' under '{parent_name}'", "process")
        if selected_company:
            self.log_panel.log(f"Create Ledger: target company '{selected_company}'", "info")

        threading.Thread(
            target=self._create_ledger_from_tab_worker,
            args=(tally_url, timeout, ledger_name, parent_name, is_party, extra_info, selected_company),
            daemon=True,
        ).start()

    def _create_ledger_from_tab_worker(self, tally_url, timeout, ledger_name, parent_name, is_party, extra_info, selected_company):
        create_result = _create_tally_ledger(
            tally_url,
            ledger_name,
            parent_name,
            timeout=min(timeout, 30),
            is_party=is_party,
            extra_info=extra_info,
            company_name=selected_company,
        )
        verify_result = {"success": False, "details": {}}
        if create_result.get("success"):
            verify_result = _fetch_tally_ledger_details_by_name(
                tally_url,
                ledger_name,
                timeout=min(timeout, 20),
                company_name=selected_company,
            )

        def done():
            self._set_create_ledger_running_ui(False, "Ready to create ledger in Tally", COLORS["text_muted"])

            if create_result.get("success"):
                self.progress_bar.set(1.0)
                self.status_label.configure(text="Created", text_color=COLORS["success"])
                self.create_ledger_conn_status.configure(text=f"Connection: Connected ({tally_url})", text_color=COLORS["success"])

                if verify_result.get("success"):
                    fetched = verify_result.get("details") or {}
                    if fetched.get("mailing_name"):
                        self._set_create_ledger_widget_value(self.create_ledger_mailing_entry, fetched.get("mailing_name"))
                    if fetched.get("gst_applicable"):
                        self._set_create_ledger_widget_value(self.create_ledger_gst_app_cb, fetched.get("gst_applicable"))
                    if fetched.get("state"):
                        self._set_create_ledger_widget_value(self.create_ledger_state_entry, fetched.get("state"))
                    if fetched.get("country"):
                        self._set_create_ledger_widget_value(self.create_ledger_country_entry, fetched.get("country"))
                    if fetched.get("pincode"):
                        self._set_create_ledger_widget_value(self.create_ledger_pincode_entry, fetched.get("pincode"))
                    self.log_panel.log(f"Create Ledger: '{ledger_name}' created and verified.", "success")
                else:
                    self.log_panel.log(f"Create Ledger: '{ledger_name}' created. Verification fetch not available.", "warning")

                messagebox.showinfo(
                    "Create Ledger",
                    "Ledger created in Tally successfully.\n\n"
                    f"Name: {ledger_name}\n"
                    f"Company: {selected_company or 'Loaded company in Tally'}",
                )
                return

            self.progress_bar.set(0.0)
            self.status_label.configure(text="Create Failed", text_color=COLORS["error"])
            err = str(create_result.get("error") or "Unknown error")
            self.log_panel.log(f"Create Ledger failed: {err}", "error")
            messagebox.showerror("Create Ledger Failed", err)

        self.after(0, done)

    # ─── PUSH TO TALLY ───

    def _browse_tally_push_xml(self):
        filepath = filedialog.askopenfilename(
            title="Select Tally XML File",
            filetypes=[("XML Files", "*.xml"), ("All Files", "*.*")],
        )
        if filepath:
            self.tally_push_xml_file = filepath
            self.tally_push_file_entry.delete(0, "end")
            self.tally_push_file_entry.insert(0, filepath)
            self.log_panel.log(f"Push XML selected: {Path(filepath).name}", "info")

    def _get_tally_push_url(self):
        host = (self.tally_push_host_entry.get() or "localhost").strip()
        port = (self.tally_push_port_entry.get() or "9000").strip()
        if host.startswith("http://"):
            host = host[7:]
        elif host.startswith("https://"):
            host = host[8:]
        host = host.strip("/") or "localhost"
        if "/" in host:
            host = host.split("/", 1)[0]
        if not port.isdigit():
            raise ValueError("Port must be numeric.")
        return f"http://{host}:{port}"

    def _get_tally_push_timeout(self):
        timeout_txt = (self.tally_push_timeout_entry.get() or "30").strip()
        try:
            timeout_val = float(timeout_txt)
        except ValueError as exc:
            raise ValueError("Timeout must be a number.") from exc
        if timeout_val <= 0:
            raise ValueError("Timeout must be greater than 0.")
        return timeout_val

    def _set_tally_push_date_mode(self, selected_mode):
        mode = str(selected_mode or "current").strip().lower()
        if mode not in {"current", "excel", "custom"}:
            mode = "current"

        self.tally_push_date_mode.set(mode)
        for key, var in self.tally_push_date_checks.items():
            var.set(key == mode)

        if hasattr(self, "tally_push_custom_date_entry"):
            custom_state = "normal" if (mode == "custom" and not self.tally_push_is_running) else "disabled"
            self.tally_push_custom_date_entry.configure(state=custom_state)

    def _format_tally_date_for_display(self, tally_date):
        text = str(tally_date or "").strip()
        if text.isdigit() and len(text) == 8:
            return f"{text[6:8]}/{text[4:6]}/{text[:4]}"
        return text

    def _get_tally_push_date_selection(self):
        mode = str(self.tally_push_date_mode.get() or "current").strip().lower()
        if mode not in {"current", "excel", "custom"}:
            mode = "current"
            self._set_tally_push_date_mode(mode)

        custom_tally_date = ""
        if mode == "custom":
            raw_custom = (self.tally_push_custom_date_var.get() or "").strip()
            if not raw_custom:
                raise ValueError("Enter Custom Date (DD/MM/YYYY) or choose Current Date / Excel Date.")
            custom_tally_date = _normalize_manual_date_to_tally(raw_custom)

        return mode, custom_tally_date

    def _set_tally_push_running_ui(self, running, label_text=None, label_color=None):
        self.tally_push_is_running = running
        if running:
            self.tally_push_test_btn.configure(state="disabled")
            self.tally_push_post_btn.configure(state="disabled", text="PROCESSING...")
            if hasattr(self, "tally_push_company_refresh_btn"):
                self.tally_push_company_refresh_btn.configure(state="disabled")
            if hasattr(self, "tally_push_company_cb"):
                self.tally_push_company_cb.configure(state="disabled")
            if hasattr(self, "tally_push_date_current_cb"):
                self.tally_push_date_current_cb.configure(state="disabled")
            if hasattr(self, "tally_push_date_excel_cb"):
                self.tally_push_date_excel_cb.configure(state="disabled")
            if hasattr(self, "tally_push_date_custom_cb"):
                self.tally_push_date_custom_cb.configure(state="disabled")
            if hasattr(self, "tally_push_custom_date_entry"):
                self.tally_push_custom_date_entry.configure(state="disabled")
            # Inline push panel
            if hasattr(self, "inline_push_post_btn"):
                self.inline_push_post_btn.configure(state="disabled", text="PROCESSING...")
            if hasattr(self, "inline_push_test_btn"):
                self.inline_push_test_btn.configure(state="disabled")
            if hasattr(self, "inline_push_company_refresh_btn"):
                self.inline_push_company_refresh_btn.configure(state="disabled")
        else:
            self.tally_push_test_btn.configure(state="normal")
            self.tally_push_post_btn.configure(state="normal", text="Step 2: Push XML")
            if hasattr(self, "tally_push_company_refresh_btn"):
                self.tally_push_company_refresh_btn.configure(state="normal")
            if hasattr(self, "tally_push_company_cb"):
                self.tally_push_company_cb.configure(state="normal")
            if hasattr(self, "tally_push_date_current_cb"):
                self.tally_push_date_current_cb.configure(state="normal")
            if hasattr(self, "tally_push_date_excel_cb"):
                self.tally_push_date_excel_cb.configure(state="normal")
            if hasattr(self, "tally_push_date_custom_cb"):
                self.tally_push_date_custom_cb.configure(state="normal")
            self._set_tally_push_date_mode(self.tally_push_date_mode.get())
            # Inline push panel
            if hasattr(self, "inline_push_post_btn"):
                self.inline_push_post_btn.configure(state="normal", text="Generate & Push to Tally")
            if hasattr(self, "inline_push_test_btn"):
                self.inline_push_test_btn.configure(state="normal")
            if hasattr(self, "inline_push_company_refresh_btn"):
                self.inline_push_company_refresh_btn.configure(state="normal")
        if label_text is not None:
            self.progress_label.configure(text=label_text, text_color=label_color or COLORS["text_muted"])

    def _set_tally_push_company_dropdown(self, companies, keep_selection=True):
        current_value = ""
        if keep_selection and hasattr(self, "tally_push_company_cb"):
            current_value = _normalize_company_name(self.tally_push_company_cb.get() or "")

        cleaned = []
        seen = set()
        for name in companies or []:
            txt = _normalize_company_name(name)
            if not _is_valid_company_name(txt):
                continue
            key = _company_key(txt)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(txt)

        cleaned = sorted(cleaned, key=lambda x: _company_key(x))
        values = [self.tally_push_company_placeholder] + cleaned
        self.tally_push_company_cb.configure(values=values)

        if current_value and _company_key(current_value) in {_company_key(x) for x in cleaned}:
            self.tally_push_company_cb.set(current_value)
        else:
            self.tally_push_company_cb.set(self.tally_push_company_placeholder)

        self.tally_push_companies = cleaned
        self.tally_push_company_status.configure(
            text=f"Companies: {len(cleaned)} available",
            text_color=COLORS["text_muted"],
        )
        # Also sync inline push panel dropdown
        if hasattr(self, "inline_push_company_cb"):
            inline_cur = ""
            if keep_selection:
                inline_cur = _normalize_company_name(self.inline_push_company_cb.get() or "")
            self.inline_push_company_cb.configure(values=values)
            if inline_cur and _company_key(inline_cur) in {_company_key(x) for x in cleaned}:
                self.inline_push_company_cb.set(inline_cur)
            else:
                self.inline_push_company_cb.set(self.tally_push_company_placeholder)
        if hasattr(self, "inline_push_company_status"):
            self.inline_push_company_status.configure(
                text=f"Companies: {len(cleaned)} available", text_color=COLORS["text_muted"])

    def _get_selected_tally_push_company(self):
        if not hasattr(self, "tally_push_company_cb"):
            return ""
        selected = _normalize_company_name(self.tally_push_company_cb.get() or "")
        if not selected:
            return ""
        if _company_key(selected) == _company_key(self.tally_push_company_placeholder):
            return ""
        return selected

    def _get_effective_push_company(self):
        """Return the company to use for Tally requests.
        If the user selected a specific company, use it.
        If 'Auto' is selected and we have exactly one cached company, use that.
        Returns "" when multiple companies are available and none is chosen (Tally uses active)."""
        company = self._get_selected_tally_push_company()
        if company:
            return company
        cached = getattr(self, "tally_push_companies", [])
        if len(cached) == 1:
            return cached[0]
        return ""

    def _fetch_tally_companies_thread(self, silent=False):
        if self.tally_push_is_running:
            return

        try:
            tally_url = self._get_tally_push_url()
            timeout = self._get_tally_push_timeout()
        except ValueError as e:
            if not silent:
                messagebox.showerror("Invalid Settings", str(e))
            return

        self.tally_push_company_refresh_btn.configure(state="disabled", text="Fetching...")
        if not silent:
            self.tally_push_company_status.configure(text="Companies: Fetching...", text_color=COLORS["warning"])

        threading.Thread(
            target=self._fetch_tally_companies_worker,
            args=(tally_url, timeout, silent),
            daemon=True,
        ).start()

    def _fetch_tally_companies_worker(self, tally_url, timeout, silent):
        result = _fetch_tally_companies(tally_url, timeout=min(timeout, 20))

        def done():
            self.tally_push_company_refresh_btn.configure(state="normal", text="Fetch")
            if result.get("success"):
                companies = result.get("companies", [])
                self._set_tally_push_company_dropdown(companies, keep_selection=True)
                self.log_panel.log(f"Fetched Tally companies: {len(companies)} found", "info")
            else:
                err = str(result.get("error") or "Unknown error")
                self.tally_push_company_status.configure(text="Companies: Fetch failed", text_color=COLORS["warning"])
                self.log_panel.log(f"Could not fetch Tally companies: {err}", "warning")
                if not silent:
                    messagebox.showwarning("Company Fetch Failed", f"Could not fetch companies from Tally.\n\n{err}")

        self.after(0, done)

    def _check_tally_connection_thread(self):
        if self.tally_push_is_running:
            return
        try:
            tally_url = self._get_tally_push_url()
            timeout = self._get_tally_push_timeout()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return

        self._set_tally_push_running_ui(True, "Checking Tally connection...", COLORS["warning"])
        self.status_label.configure(text="Checking", text_color=COLORS["warning"])
        self.tally_push_conn_status.configure(text="Connection: Checking...", text_color=COLORS["warning"])
        threading.Thread(
            target=self._check_tally_connection_worker,
            args=(tally_url, timeout),
            daemon=True,
        ).start()

    def _check_tally_connection_worker(self, tally_url, timeout):
        check_result = _check_tally_connection(tally_url, timeout=min(timeout, 10))
        ok = check_result.get("connected", False)
        error_msg = ""
        if not ok:
            if check_result.get("error") == "ConnectionError":
                error_msg = "Cannot connect to Tally. Check if Tally is running and HTTP server is enabled."
            elif check_result.get("error") == "Timeout":
                error_msg = "Connection check timed out."
            elif "status_code" in check_result:
                error_msg = f"HTTP {check_result.get('status_code')}"
            else:
                error_msg = str(check_result.get("error") or "Unknown error")

        def done():
            self._set_tally_push_running_ui(False, "Ready to test/push XML to Tally", COLORS["text_muted"])
            self.status_label.configure(text="Ready", text_color=COLORS["success"])
            if ok:
                self.tally_push_conn_status.configure(text=f"Connection: Connected ({tally_url})", text_color=COLORS["success"])
                self.log_panel.log(f"Tally connection successful at {tally_url}", "success")
                self._fetch_tally_companies_thread(silent=True)
            else:
                self.tally_push_conn_status.configure(text="Connection: Offline", text_color=COLORS["error"])
                self.log_panel.log(f"Tally connection failed: {error_msg}", "error")

        self.after(0, done)

    def _post_tally_xml_thread(self):
        if self.tally_push_is_running:
            return

        xml_path = self.tally_push_file_entry.get().strip() or self.tally_push_xml_file
        if not xml_path:
            messagebox.showwarning("No File", "Please select XML file first.")
            return
        if not os.path.isfile(xml_path):
            messagebox.showerror("File Missing", "Selected XML file does not exist.")
            return
        try:
            tally_url = self._get_tally_push_url()
            timeout = self._get_tally_push_timeout()
            date_mode, custom_tally_date = self._get_tally_push_date_selection()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return

        selected_company = self._get_selected_tally_push_company()

        self._set_tally_push_running_ui(True, "Posting XML to Tally...", COLORS["warning"])
        self.status_label.configure(text="Posting", text_color=COLORS["warning"])
        self.progress_bar.set(0.25)
        self.log_panel.log(f"Posting XML to {tally_url}", "process")
        self.log_panel.log(f"File: {Path(xml_path).name}", "info")
        if selected_company:
            self.log_panel.log(f"Target company selected: {selected_company}", "info")
        else:
            self.log_panel.log("Target company: currently loaded company in Tally", "info")
        if date_mode == "excel":
            self.log_panel.log("Voucher date mode: Excel Date (kept from XML).", "info")
        elif date_mode == "custom":
            custom_label = self._format_tally_date_for_display(custom_tally_date)
            self.log_panel.log(f"Voucher date mode: Custom Date ({custom_label}).", "info")
        else:
            self.log_panel.log("Voucher date mode: Current Date (today).", "info")
        threading.Thread(
            target=self._post_tally_xml_worker,
            args=(xml_path, tally_url, timeout, selected_company, date_mode, custom_tally_date),
            daemon=True,
        ).start()

    def _post_tally_xml_worker(
        self,
        xml_path,
        tally_url,
        timeout,
        selected_company="",
        date_mode="current",
        custom_tally_date="",
    ):
        result = {"success": False, "error": "Unknown error"}
        forced_date_count = 0
        posted_xml_content = ""
        missing_ledgers = []
        tally_ledgers = []
        usage_map = {}
        ledger_fetch_error = ""
        try:
            xml_content = _read_xml_text_safely(xml_path)
            xml_content, forced_date_count = _apply_push_date_mode(
                xml_content,
                date_mode=date_mode,
                custom_tally_date=custom_tally_date,
            )
            if selected_company:
                xml_content, _ = _set_svcurrentcompany(xml_content, selected_company)
            usage_map = _extract_ledger_usage_from_xml(xml_content)

            fetch_result = _fetch_tally_ledgers(tally_url, timeout=min(timeout, 20), company_name=selected_company)
            if fetch_result.get("success"):
                tally_ledgers = fetch_result.get("ledgers", [])
                existing_keys = {_ledger_key(x) for x in tally_ledgers}
                precheck_missing = _collect_missing_ledgers_from_usage(usage_map, existing_keys)
                if precheck_missing:
                    posted_xml_content = xml_content
                    missing_ledgers = precheck_missing
                    result = {"success": False, "error": "Missing ledgers found before posting."}
                    raise RuntimeError("PRECHECK_MISSING_LEDGERS")
            else:
                ledger_fetch_error = str(fetch_result.get("error") or "")
                result = {
                    "success": False,
                    "error": (
                        "Could not fetch ledger list from Tally for pre-check. "
                        "Posting was not attempted to avoid duplicate entries. "
                        f"Details: {ledger_fetch_error or 'Unknown error'}"
                    ),
                }
                raise RuntimeError("PRECHECK_FETCH_FAILED")

            voucher_count_in_xml = xml_content.count("<TALLYMESSAGE")
            self.after(0, lambda n=voucher_count_in_xml: self.log_panel.log(
                f"Pushing {n} voucher(s) to Tally...", "process"
            ))

            result, posted_xml_content, retry_meta = _post_xml_with_fallbacks(
                tally_url,
                xml_content,
                timeout=timeout,
                allow_company_fallback=not bool(selected_company),
                allow_date_retry=(date_mode == "current"),
            )
            forced_date_count = max(forced_date_count, retry_meta.get("forced_date_count", 0))
            if retry_meta.get("date_retry_used"):
                result["date_retry_used"] = True
            if retry_meta.get("fallback_used"):
                result["fallback_used"] = True

            if not result.get("success"):
                parsed_missing = _extract_missing_ledger_names(result.get("error", ""))
                if parsed_missing:
                    if not tally_ledgers:
                        fetch_result = _fetch_tally_ledgers(tally_url, timeout=min(timeout, 20), company_name=selected_company)
                        if fetch_result.get("success"):
                            tally_ledgers = fetch_result.get("ledgers", [])
                        else:
                            ledger_fetch_error = str(fetch_result.get("error") or "")

                    if tally_ledgers:
                        existing_keys = {_ledger_key(x) for x in tally_ledgers}
                        xml_ledgers = _extract_ledger_names_from_xml(posted_xml_content)
                        for name in xml_ledgers:
                            n_name = _normalize_ledger_name(name)
                            if _ledger_key(n_name) not in existing_keys and _ledger_key(n_name) not in {_ledger_key(x) for x in missing_ledgers}:
                                missing_ledgers.append(n_name)

                    for missing in parsed_missing:
                        n_missing = _normalize_ledger_name(missing)
                        if _ledger_key(n_missing) not in {_ledger_key(x) for x in missing_ledgers}:
                            missing_ledgers.append(n_missing)

                    missing_ledgers = sorted(missing_ledgers, key=lambda x: _ledger_key(x))
                    if not usage_map:
                        usage_map = _extract_ledger_usage_from_xml(posted_xml_content)
        except Exception as e:
            if str(e) not in {"PRECHECK_MISSING_LEDGERS", "PRECHECK_FETCH_FAILED"}:
                result = {"success": False, "error": str(e)}

        def done():
            self._set_tally_push_running_ui(False)

            if result.get("success"):
                self.progress_bar.set(1.0)
                self._show_tally_push_success(
                    result,
                    forced_date_count=forced_date_count,
                    target_company=selected_company,
                    date_mode=date_mode,
                    custom_tally_date=custom_tally_date,
                )
                return

            if missing_ledgers:
                self.progress_bar.set(0.0)
                self.status_label.configure(text="Ledger Mapping Needed", text_color=COLORS["warning"])
                self.progress_label.configure(text="Resolve missing ledgers to continue", text_color=COLORS["warning"])
                self.log_panel.log("Tally rejected XML due to missing ledger(s). Opening resolution dialog...", "warning")
                if ledger_fetch_error:
                    self.log_panel.log(f"Could not fetch full ledger list from Tally: {ledger_fetch_error}", "warning")
                self._show_missing_ledger_resolution_dialog(
                    xml_path=xml_path,
                    tally_url=tally_url,
                    timeout=timeout,
                    base_xml_content=posted_xml_content,
                    missing_ledgers=missing_ledgers,
                    tally_ledgers=tally_ledgers,
                    usage_map=usage_map,
                    forced_date_count=forced_date_count,
                    selected_company=selected_company,
                    date_mode=date_mode,
                    custom_tally_date=custom_tally_date,
                )
                return

            self.progress_bar.set(0.0)
            self.status_label.configure(text="Push Failed", text_color=COLORS["error"])
            self.progress_label.configure(text="Tally push failed", text_color=COLORS["error"])
            self.log_panel.log(f"Tally push failed: {result.get('error', 'Unknown error')}", "error")
            messagebox.showerror("Push To Tally Failed", result.get("error", "Unknown error"))

        self.after(0, done)

    def _show_tally_push_success(
        self,
        result,
        forced_date_count=0,
        created_ledgers=None,
        target_company="",
        date_mode="current",
        custom_tally_date="",
    ):
        created = result.get("created", "0")
        altered = result.get("altered", "0")
        deleted = result.get("deleted", "0")
        cancelled = result.get("cancelled", "0")
        errors = result.get("errors", "0")

        self.status_label.configure(text="Posted", text_color=COLORS["success"])
        self.progress_label.configure(text="XML posted successfully", text_color=COLORS["success"])
        self.log_panel.log(
            f"Tally import success - Created: {created}, Altered: {altered}, Deleted: {deleted}, Cancelled: {cancelled}, Errors: {errors}",
            "success",
        )
        if target_company:
            self.log_panel.log(f"Posted to company: {target_company}", "info")
        if date_mode == "excel":
            self.log_panel.log("Voucher date mode used: Excel Date (kept from XML).", "info")
        elif date_mode == "custom":
            custom_label = self._format_tally_date_for_display(custom_tally_date)
            self.log_panel.log(f"Voucher date mode used: Custom Date ({custom_label}).", "info")
        else:
            self.log_panel.log("Voucher date mode used: Current Date (today).", "info")
        if forced_date_count > 0:
            if date_mode == "custom":
                custom_label = self._format_tally_date_for_display(custom_tally_date)
                self.log_panel.log(
                    f"Voucher date set to custom date {custom_label} for {forced_date_count} voucher(s).",
                    "info",
                )
            else:
                self.log_panel.log(f"Voucher date auto-set to today for {forced_date_count} voucher(s) during push.", "info")
        if result.get("date_retry_used"):
            self.log_panel.log("Date error auto-fixed using strict voucher-date retry.", "warning")
        if result.get("fallback_used"):
            self.log_panel.log("Company mismatch auto-fixed: SVCURRENTCOMPANY removed and posted to loaded company.", "warning")
        if created_ledgers:
            self.log_panel.log(f"Ledgers created in Tally: {', '.join(created_ledgers)}", "success")
        if str(errors) not in {"0", "0.0"}:
            self.log_panel.log(f"Tally reported {errors} error(s). Please verify in Tally.", "warning")

        if date_mode == "excel":
            date_mode_text = "Excel Date"
        elif date_mode == "custom":
            date_mode_text = f"Custom Date ({self._format_tally_date_for_display(custom_tally_date)})"
        else:
            date_mode_text = "Current Date"

        # Include stock item vouchers (pushed earlier via popup) in total count
        stock_pushed = getattr(self, "_stock_items_pushed_count", 0)
        if stock_pushed > 0:
            self._stock_items_pushed_count = 0  # reset after reading
        total_vouchers = int(created) + stock_pushed

        info_lines = [
            "Entries posted to Tally successfully.",
            "",
            f"Target Company: {target_company or 'Loaded company in Tally'}",
            f"Voucher Date Mode: {date_mode_text}",
            "",
        ]
        if stock_pushed > 0:
            info_lines += [
                f"Purchase Item Vouchers (stock): {stock_pushed}",
                f"Purchase Vouchers (accounting): {created}",
                f"Total Vouchers Created: {total_vouchers}",
            ]
        else:
            info_lines += [
                f"Created: {created}",
            ]
        info_lines += [
            f"Altered: {altered}",
            f"Deleted: {deleted}",
            f"Cancelled: {cancelled}",
            f"Errors: {errors}",
        ]
        if created_ledgers:
            info_lines.append("")
            info_lines.append(f"Ledgers created: {', '.join(created_ledgers)}")
        messagebox.showinfo("Push To Tally", "\n".join(info_lines))

    def _show_missing_ledger_resolution_dialog(
        self,
        xml_path,
        tally_url,
        timeout,
        base_xml_content,
        missing_ledgers,
        tally_ledgers,
        usage_map,
        forced_date_count,
        selected_company="",
        date_mode="current",
        custom_tally_date="",
    ):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Resolve Missing Ledgers")
        dialog.geometry("980x620")
        dialog.minsize(840, 460)
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color=COLORS["bg_dark"])

        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 980) // 2
        y = self.winfo_y() + (self.winfo_height() - 620) // 2
        dialog.geometry(f"+{x}+{y}")

        top = ctk.CTkFrame(dialog, fg_color=COLORS["warning_bg"], corner_radius=10)
        top.pack(fill="x", padx=16, pady=(16, 8))
        ctk.CTkLabel(
            top,
            text="Missing Ledger Resolution",
            font=("Segoe UI", 16, "bold"),
            text_color=COLORS["warning"],
        ).pack(anchor="w", padx=14, pady=(10, 2))
        ctk.CTkLabel(
            top,
            text=(
                f"Tally could not find {len(missing_ledgers)} ledger(s). "
                "Map to an existing ledger or create a new ledger, then retry push."
                + (f"\nTarget company: {selected_company}" if selected_company else "")
            ),
            font=("Segoe UI", 11),
            text_color=COLORS["text_secondary"],
            wraplength=920,
            justify="left",
        ).pack(anchor="w", padx=14, pady=(0, 10))

        info_row = ctk.CTkFrame(dialog, fg_color="transparent")
        info_row.pack(fill="x", padx=16, pady=(0, 6))

        def _sanitize_dropdown_values(values, max_items=50000):
            cleaned = []
            seen = set()
            for raw in values or []:
                txt = _normalize_ledger_name(raw)
                if not txt:
                    continue
                key = _ledger_key(txt)
                if key in seen:
                    continue
                seen.add(key)
                cleaned.append(txt)
                if len(cleaned) >= max_items:
                    break
            return cleaned

        current_tally_ledgers = _sanitize_dropdown_values(tally_ledgers)
        tally_count_label = ctk.CTkLabel(
            info_row,
            text=f"Existing ledgers in Tally: {len(current_tally_ledgers)}",
            font=("Segoe UI", 11, "bold"),
            text_color=COLORS["text_primary"],
        )
        tally_count_label.pack(side="left")

        header = ctk.CTkFrame(dialog, fg_color=COLORS["table_header"], corner_radius=6, height=34)
        header.pack(fill="x", padx=16, pady=(0, 2))
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="Missing Ledger", width=210, font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["tally_gold"], anchor="w").pack(side="left", padx=(10, 6))
        ctk.CTkLabel(header, text="Map To Existing", width=250, font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["tally_gold"], anchor="w").pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Create New", width=230, font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["tally_gold"], anchor="w").pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Parent", font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["tally_gold"], anchor="w").pack(side="left", padx=6)

        table = ctk.CTkScrollableFrame(
            dialog,
            fg_color=COLORS["bg_card"],
            corner_radius=8,
            border_width=1,
            border_color=COLORS["border"],
        )
        table.pack(fill="both", expand=True, padx=16, pady=(0, 10))

        parent_options = [
            "Purchase Accounts",
            "Sundry Creditors",
            "Duties & Taxes",
            "Indirect Expenses",
            "Current Liabilities",
            "Sundry Debtors",
        ]

        row_controls = []
        search_placeholder = "Search Ledger"

        def _combo_values():
            return current_tally_ledgers if current_tally_ledgers else [""]

        def _default_extra_for_missing(missing_name):
            base = dict((usage_map.get(missing_name) or {}).get("extra", {}))
            gstin = _normalize_ledger_name(base.get("gstin") or "").upper()
            gst_applicable = _normalize_ledger_name(base.get("gst_applicable") or "")
            if not gst_applicable:
                gst_applicable = "Applicable" if gstin else "Not Applicable"
            reg_type = _normalize_ledger_name(base.get("reg_type") or "")
            if not reg_type:
                reg_type = "Regular" if gstin else "Unknown"
            billwise = _normalize_ledger_name(base.get("billwise") or "")
            if not billwise:
                billwise = "Yes" if bool((usage_map.get(missing_name) or {}).get("is_party")) else "No"
            return {
                "mailing_name": _normalize_ledger_name(base.get("mailing_name") or missing_name),
                "address1": _normalize_ledger_name(base.get("address1") or ""),
                "address2": _normalize_ledger_name(base.get("address2") or ""),
                "state": _normalize_ledger_name(base.get("state") or ""),
                "country": _normalize_ledger_name(base.get("country") or "India") or "India",
                "pincode": _normalize_ledger_name(base.get("pincode") or ""),
                "gstin": gstin,
                "gst_applicable": gst_applicable,
                "reg_type": reg_type,
                "billwise": billwise,
            }

        def _open_create_details_dialog(row_control):
            detail = dict(row_control.get("extra") or {})
            popup = ctk.CTkToplevel(dialog)
            popup.title("Create Ledger Details")
            popup.geometry("640x460")
            popup.transient(dialog)
            popup.grab_set()
            popup.configure(fg_color=COLORS["bg_dark"])

            ctk.CTkLabel(
                popup,
                text=f"Create details for: {row_control['missing']}",
                font=("Segoe UI", 13, "bold"),
                text_color=COLORS["text_primary"],
            ).pack(anchor="w", padx=16, pady=(14, 6))

            ctk.CTkLabel(
                popup,
                text="Fill details like your PartyMaster macro (GST, mailing, state, pincode).",
                font=("Segoe UI", 10),
                text_color=COLORS["text_secondary"],
            ).pack(anchor="w", padx=16, pady=(0, 10))

            form = ctk.CTkFrame(popup, fg_color=COLORS["bg_card"], corner_radius=10)
            form.pack(fill="both", expand=True, padx=16, pady=(0, 10))

            entries = {}
            popup_state = {"closing": False}

            def set_widget_value(widget, value):
                text_value = _normalize_ledger_name(value or "")
                try:
                    widget.set(text_value)
                    return
                except Exception:
                    pass
                try:
                    widget.delete(0, "end")
                    widget.insert(0, text_value)
                except Exception:
                    pass

            def add_field(label, key, row_idx, width=260, values=None):
                ctk.CTkLabel(form, text=label, font=("Segoe UI", 10), text_color=COLORS["text_secondary"]).grid(
                    row=row_idx,
                    column=0,
                    sticky="w",
                    padx=(14, 8),
                    pady=6,
                )
                if values:
                    widget = ctk.CTkComboBox(
                        form,
                        values=values,
                        width=width,
                        fg_color=COLORS["bg_input"],
                        border_color=COLORS["border"],
                        button_color=COLORS["accent"],
                        button_hover_color=COLORS["accent_hover"],
                    )
                    set_widget_value(widget, _normalize_ledger_name(detail.get(key) or values[0]))
                else:
                    widget = ctk.CTkEntry(
                        form,
                        width=width,
                        fg_color=COLORS["bg_input"],
                        border_color=COLORS["border"],
                        text_color=COLORS["text_primary"],
                        font=("Segoe UI", 10),
                    )
                    set_widget_value(widget, _normalize_ledger_name(detail.get(key) or ""))
                widget.grid(row=row_idx, column=1, sticky="w", padx=(0, 14), pady=6)
                entries[key] = widget

            add_field("Mailing Name", "mailing_name", 0)
            add_field("Address 1", "address1", 1)
            add_field("Address 2", "address2", 2)
            add_field("State", "state", 3, values=[""] + LEDGER_STATE_OPTIONS)
            add_field("Country", "country", 4, values=LEDGER_COUNTRY_OPTIONS)
            add_field("Pincode", "pincode", 5)
            add_field("GSTIN", "gstin", 6)
            add_field("PAN / IT No.", "pan", 7)
            add_field("GST Applicable", "gst_applicable", 8, values=LEDGER_GST_APPLICABLE_OPTIONS)
            add_field("Registration Type", "reg_type", 9, values=["Regular", "Composition", "Unregistered", "Consumer", "Unknown"])
            add_field("Billwise", "billwise", 10, values=["Yes", "No"])

            gst_lookup_guard = {"busy": False, "last": ""}

            action_row = ctk.CTkFrame(popup, fg_color="transparent")
            action_row.pack(fill="x", padx=16, pady=(0, 14))

            def fetch_from_gstin(show_messages=True):
                gstin_val = _normalize_ledger_name(entries["gstin"].get() or "").upper()
                if len(gstin_val) < 15:
                    if show_messages:
                        messagebox.showwarning("Invalid GSTIN", "Please enter a valid 15-digit GSTIN first.")
                    return

                if not _normalize_ledger_name(entries["country"].get() or ""):
                    set_widget_value(entries["country"], "India")

                self.log_panel.log(f"Trying GST detail fetch for {gstin_val} from Tally...", "process")
                fetch_result = _fetch_tally_party_details_by_gstin(
                    tally_url,
                    gstin_val,
                    timeout=min(timeout, 20),
                    company_name=selected_company,
                )
                if fetch_result.get("success"):
                    fetched = fetch_result.get("details", {})
                    for key in ["mailing_name", "address1", "address2", "state", "country", "pincode", "gstin", "pan", "gst_applicable", "reg_type", "billwise"]:
                        if key in entries and fetched.get(key):
                            set_widget_value(entries[key], fetched.get(key))
                    # Auto-derive PAN from GSTIN if not returned by Tally
                    if "pan" in entries and not _normalize_ledger_name(entries["pan"].get() or ""):
                        gstin_fetched = _normalize_ledger_name(fetched.get("gstin") or "").upper()
                        pan_auto = _pan_from_gstin(gstin_fetched)
                        if pan_auto:
                            set_widget_value(entries["pan"], pan_auto)
                    self.log_panel.log("GST fetch succeeded from existing Tally ledger data.", "success")
                    if show_messages:
                        messagebox.showinfo("GST Fetch", "Details fetched from existing Tally ledger with same GSTIN.")
                    return

                inferred_state = _state_name_from_gstin(gstin_val)

                if inferred_state:
                    set_widget_value(entries["state"], inferred_state)
                if not _normalize_ledger_name(entries["country"].get() or ""):
                    set_widget_value(entries["country"], "India")
                set_widget_value(entries["gst_applicable"], "Applicable")
                set_widget_value(entries["reg_type"], "Regular")
                self.log_panel.log(
                    f"GST fetch from Tally not available ({fetch_result.get('error', 'Unknown')}). Filled available defaults.",
                    "warning",
                )
                if show_messages:
                    messagebox.showinfo(
                        "GST Fetch",
                        "No matching GSTIN details were returned by Tally API.\n"
                        "State/registration defaults have been filled; please review remaining fields.",
                    )

            def _on_gstin_change(_event=None):
                gstin_now = _normalize_ledger_name(entries["gstin"].get() or "").upper()
                if not _normalize_ledger_name(entries["country"].get() or ""):
                    set_widget_value(entries["country"], "India")
                if len(gstin_now) >= 2 and not _normalize_ledger_name(entries["state"].get() or ""):
                    inferred = _state_name_from_gstin(gstin_now)
                    if inferred:
                        set_widget_value(entries["state"], inferred)
                if len(gstin_now) == 15 and not _normalize_ledger_name(entries["pan"].get() or ""):
                    pan_auto = _pan_from_gstin(gstin_now)
                    if pan_auto:
                        set_widget_value(entries["pan"], pan_auto)
                if len(gstin_now) == 15 and gstin_now != gst_lookup_guard["last"] and not gst_lookup_guard["busy"]:
                    gst_lookup_guard["busy"] = True
                    gst_lookup_guard["last"] = gstin_now

                    def _auto_fetch_run():
                        try:
                            latest = _normalize_ledger_name(entries["gstin"].get() or "").upper()
                            if latest == gstin_now:
                                fetch_from_gstin(show_messages=False)
                        finally:
                            gst_lookup_guard["busy"] = False

                    popup.after(120, _auto_fetch_run)

            entries["gstin"].bind("<KeyRelease>", _on_gstin_change)
            entries["gstin"].bind("<FocusOut>", _on_gstin_change)

            def _collect_updated_details():
                updated = {}
                for key, widget in entries.items():
                    try:
                        val = widget.get()
                    except Exception:
                        val = ""
                    updated[key] = _normalize_ledger_name(val or "")
                updated["country"] = updated.get("country") or "India"
                if not updated.get("state") and updated.get("gstin"):
                    updated["state"] = _state_name_from_gstin(updated.get("gstin"))
                if not updated.get("gst_applicable"):
                    updated["gst_applicable"] = "Applicable" if updated.get("gstin") else "Not Applicable"
                updated["reg_type"] = updated.get("reg_type") or ("Regular" if updated.get("gstin") else "Unknown")
                updated["billwise"] = updated.get("billwise") or ("Yes" if bool(row_control.get("is_party")) else "No")
                return updated

            def _apply_row_detail(updated):
                row_control["extra"] = dict(updated)
                row_control["create_var"].set(True)
                sync_create_ui = row_control.get("sync_create_ui")
                if callable(sync_create_ui):
                    sync_create_ui()

            def _save_details_to_row(close_popup=True):
                updated = _collect_updated_details()
                _apply_row_detail(updated)
                if close_popup and popup.winfo_exists():
                    popup_state["closing"] = True
                    popup.destroy()
                return updated

            def _close_with_save():
                if popup_state["closing"]:
                    return
                _save_details_to_row(close_popup=True)

            popup.protocol("WM_DELETE_WINDOW", _close_with_save)

            def save_details():
                _save_details_to_row(close_popup=True)

            def create_in_tally_now():
                updated = _save_details_to_row(close_popup=False)
                ledger_name = _normalize_ledger_name(row_control["new_name_entry"].get() or row_control["missing"])
                parent_name = _normalize_ledger_name(row_control["parent_combo"].get() or "Sundry Creditors")

                create_result = _create_tally_ledger(
                    tally_url,
                    ledger_name,
                    parent_name,
                    timeout=min(timeout, 30),
                    is_party=bool(row_control.get("is_party", False)),
                    company_name=selected_company,
                    extra_info=updated,
                )
                if not create_result.get("success"):
                    err = create_result.get("error") or "Ledger creation failed"
                    self.log_panel.log(f"Manual create failed for '{ledger_name}': {err}", "error")
                    messagebox.showerror("Create Ledger", f"Could not create ledger in Tally.\n\n{err}")
                    return

                row_control["create_var"].set(True)
                verify_result = _fetch_tally_ledger_details_by_name(
                    tally_url,
                    ledger_name,
                    timeout=min(timeout, 20),
                    company_name=selected_company,
                )
                if verify_result.get("success"):
                    fetched = verify_result.get("details") or {}
                    fetched_gstin = _normalize_ledger_name(fetched.get("gstin") or "-")
                    fetched_state = _normalize_ledger_name(fetched.get("state") or "-")
                    self.log_panel.log(
                        f"Manual create verified for '{ledger_name}' (GST: {fetched_gstin}, State: {fetched_state}).",
                        "success",
                    )
                    messagebox.showinfo(
                        "Create Ledger",
                        "Ledger created and verified from Tally.\n\n"
                        f"Name: {fetched.get('name') or ledger_name}\n"
                        f"GSTIN: {fetched_gstin}\n"
                        f"State: {fetched_state}",
                    )
                    return

                warn = verify_result.get("error") or "Verification fetch failed"
                self.log_panel.log(
                    f"Ledger created for '{ledger_name}', but verification fetch failed: {warn}",
                    "warning",
                )
                messagebox.showwarning(
                    "Create Ledger",
                    "Ledger was created in Tally, but verification fetch did not return details.\n"
                    "You can still continue with Apply & Retry Push.",
                )

            ctk.CTkButton(
                action_row,
                text="Fetch From GSTIN",
                height=34,
                font=("Segoe UI", 11, "bold"),
                fg_color=COLORS["warning"],
                hover_color="#B45309",
                text_color="#FFFFFF",
                command=fetch_from_gstin,
            ).pack(side="left", padx=(0, 6))

            ctk.CTkButton(
                action_row,
                text="Create In Tally",
                height=34,
                font=("Segoe UI", 11, "bold"),
                fg_color=COLORS["accent"],
                hover_color=COLORS["accent_hover"],
                text_color="#FFFFFF",
                command=create_in_tally_now,
            ).pack(side="left", padx=(0, 6))

            ctk.CTkButton(
                action_row,
                text="Save Details",
                height=34,
                font=("Segoe UI", 11, "bold"),
                fg_color=COLORS["success"],
                hover_color="#047857",
                text_color="#FFFFFF",
                command=save_details,
            ).pack(side="left", fill="x", expand=True, padx=(0, 6))

            ctk.CTkButton(
                action_row,
                text="Close & Save",
                height=34,
                font=("Segoe UI", 11),
                fg_color=COLORS["bg_input"],
                hover_color=COLORS["bg_card_hover"],
                text_color=COLORS["text_secondary"],
                command=_close_with_save,
            ).pack(side="right")

        def _show_create_choice_dialog(rc):
            """Choice popup: Fetch from GST Portal  OR  Create Manually."""
            led_name    = str(rc["missing"] or "")
            gstin_hint  = str((rc.get("extra") or {}).get("gstin") or "").upper()
            parent_name = _normalize_ledger_name(rc["parent_combo"].get() or "Sundry Debtors")

            cpop = ctk.CTkToplevel(dialog)
            cpop.title("Create Ledger — Choose Method")
            cpop.geometry("480x250")
            cpop.resizable(False, False)
            cpop.transient(dialog)
            cpop.grab_set()
            cpop.lift()
            cpop.focus_force()

            hdr = ctk.CTkFrame(cpop, fg_color=COLORS["warning_bg"], corner_radius=0)
            hdr.pack(fill="x")
            ctk.CTkLabel(hdr, text="How would you like to create this ledger?",
                          font=("Segoe UI", 13, "bold"),
                          text_color=COLORS["warning"]).pack(padx=16, pady=10)

            ctk.CTkLabel(cpop, text=f'Ledger:  "{led_name}"',
                          font=("Segoe UI", 12),
                          text_color=COLORS["text_primary"]).pack(pady=(12, 2), padx=16, anchor="w")
            if gstin_hint:
                ctk.CTkLabel(cpop, text=f"GSTIN: {gstin_hint}",
                              font=("Segoe UI", 10, "italic"),
                              text_color=COLORS["text_muted"]).pack(padx=16, anchor="w", pady=(0, 6))

            bf = ctk.CTkFrame(cpop, fg_color="transparent")
            bf.pack(fill="x", padx=16, pady=(4, 16))

            def _mark_create():
                rc["create_var"].set(True)
                fn = rc.get("sync_create_ui")
                if callable(fn):
                    fn()

            def _choose_portal():
                cpop.destroy()
                _mark_create()
                def _on_portal_data(extra_result):
                    if extra_result:
                        rc["extra"] = extra_result
                _GSTFetchDialog(dialog, led_name, gstin_hint, parent_name,
                                on_result=_on_portal_data)

            def _choose_manual():
                cpop.destroy()
                _mark_create()
                _open_create_details_dialog(rc)

            ctk.CTkButton(
                bf,
                text="Fetch from GST Portal",
                fg_color=("#2563EB", "#3B82F6"),
                hover_color=("#1D4ED8", "#2563EB"),
                height=44,
                font=("Segoe UI", 12),
                command=_choose_portal,
            ).pack(fill="x", pady=(0, 8))

            ctk.CTkButton(
                bf,
                text="Create Manually",
                fg_color=("gray70", "gray30"),
                hover_color=("gray60", "gray40"),
                height=44,
                font=("Segoe UI", 12),
                command=_choose_manual,
            ).pack(fill="x")

        for idx, missing in enumerate(missing_ledgers):
            row_bg = COLORS["table_row_odd"] if idx % 2 == 0 else COLORS["table_row_even"]
            row = ctk.CTkFrame(table, fg_color=row_bg, corner_radius=4, height=112)
            row.pack(fill="x", pady=1)
            row.pack_propagate(False)

            ctk.CTkLabel(row, text=missing[:36], width=210, font=("Segoe UI", 10),
                         text_color=COLORS["text_primary"], anchor="w").pack(side="left", padx=(10, 6))

            map_col = ctk.CTkFrame(row, fg_color="transparent", width=250)
            map_col.pack(side="left", padx=6)
            map_col.pack_propagate(False)

            map_search_entry = ctk.CTkEntry(
                map_col,
                height=26,
                fg_color=COLORS["bg_input"],
                border_color=COLORS["border"],
                text_color=COLORS["text_primary"],
                placeholder_text=search_placeholder,
                font=("Segoe UI", 10),
            )
            map_search_entry.pack(fill="x", pady=(0, 2))

            initial_values = (_sanitize_dropdown_values(_combo_values()) or [""])[:200]

            def _make_on_combo_select(entry_widget):
                def _on_select(_value):
                    entry_widget.delete(0, "end")
                    entry_widget.insert(0, str(_value or "").strip())
                return _on_select

            on_combo_select = _make_on_combo_select(map_search_entry)
            try:
                map_combo = ctk.CTkComboBox(
                    map_col,
                    values=initial_values,
                    height=28,
                    fg_color=COLORS["bg_input"],
                    border_color=COLORS["border"],
                    button_color=COLORS["accent"],
                    button_hover_color=COLORS["accent_hover"],
                    font=("Segoe UI", 10),
                    command=on_combo_select,
                )
            except tk.TclError:
                # Fallback keeps dialog usable even if Tally returned unsupported text.
                map_combo = ctk.CTkComboBox(
                    map_col,
                    values=[""],
                    height=28,
                    fg_color=COLORS["bg_input"],
                    border_color=COLORS["border"],
                    button_color=COLORS["accent"],
                    button_hover_color=COLORS["accent_hover"],
                    font=("Segoe UI", 10),
                    command=on_combo_select,
                )
                self.log_panel.log("Some ledger names contain unsupported characters for dropdown; using safe mode.", "warning")
            map_combo.pack(fill="x")
            map_combo.set(search_placeholder)

            suggestion_box = tk.Listbox(
                map_col,
                height=3,
                exportselection=False,
                activestyle="none",
                relief="solid",
                borderwidth=1,
                bg=_theme_color("bg_input"),
                fg=_theme_color("text_primary"),
                selectbackground=_theme_color("accent"),
                selectforeground="#FFFFFF",
            )
            suggestion_box.pack(fill="x", pady=(2, 0))

            def _make_suggestion_click_handler(listbox_widget, entry_widget, combo_widget):
                def _handle(_event=None):
                    sel = listbox_widget.curselection()
                    if not sel:
                        return
                    value = str(listbox_widget.get(sel[0]) or "").strip()
                    if not value:
                        return
                    entry_widget.delete(0, "end")
                    entry_widget.insert(0, value)
                    combo_widget.set(value)
                    listbox_widget.delete(0, "end")
                return _handle

            suggestion_click_handler = _make_suggestion_click_handler(suggestion_box, map_search_entry, map_combo)
            suggestion_box.bind("<<ListboxSelect>>", suggestion_click_handler)
            suggestion_box.bind("<Double-Button-1>", suggestion_click_handler)
            suggestion_box.bind("<Return>", suggestion_click_handler)

            create_frame = ctk.CTkFrame(row, fg_color="transparent", width=230)
            create_frame.pack(side="left", padx=6)
            create_frame.pack_propagate(False)
            create_var = ctk.BooleanVar(value=False)
            is_party_row = bool((usage_map.get(missing) or {}).get("is_party"))
            detail_extra = _default_extra_for_missing(missing)

            create_action_btn = ctk.CTkButton(
                create_frame,
                text="Create",
                width=68,
                height=28,
                font=("Segoe UI", 10, "bold"),
                fg_color=COLORS["success"],
                hover_color="#047857",
                text_color="#FFFFFF",
            )
            create_action_btn.pack(side="left", padx=(0, 6))

            new_name_entry = ctk.CTkEntry(
                create_frame,
                width=154,
                height=28,
                fg_color=COLORS["bg_input"],
                border_color=COLORS["border"],
                text_color=COLORS["text_primary"],
                placeholder_text="New Ledger Name",
                font=("Segoe UI", 10),
            )
            new_name_entry.pack(side="left")
            new_name_entry.insert(0, missing)

            guessed_parent = _guess_parent_for_ledger(missing, usage_map.get(missing, {}))
            parent_combo = ctk.CTkComboBox(
                row,
                values=parent_options,
                width=210,
                fg_color=COLORS["bg_input"],
                border_color=COLORS["border"],
                button_color=COLORS["accent"],
                button_hover_color=COLORS["accent_hover"],
            )
            parent_combo.pack(side="left", padx=6)
            parent_combo.set(guessed_parent)

            def _make_sync_create_ui(var=create_var, name_entry=new_name_entry, parent_cb=parent_combo, action_btn=create_action_btn):
                def _sync():
                    state = "normal" if var.get() else "disabled"
                    name_entry.configure(state=state)
                    parent_cb.configure(state=state)
                    action_btn.configure(text=("Edit" if var.get() else "Create"))
                return _sync

            sync_create_ui_fn = _make_sync_create_ui()
            sync_create_ui_fn()

            def _make_combo_filter(combo_widget, entry_widget, listbox_widget):
                def _apply_combo_filter(query_text=None):
                    q = str(query_text if query_text is not None else entry_widget.get() or "").strip().lower()
                    if not q:
                        filtered = list(current_tally_ledgers)
                    else:
                        filtered = [name for name in current_tally_ledgers if q in name.lower()]
                    filtered = _sanitize_dropdown_values(filtered)
                    combo_widget.configure(values=(filtered[:200] if filtered else [""]))

                    listbox_widget.delete(0, "end")
                    if q:
                        for item in filtered[:12]:
                            listbox_widget.insert("end", item)
                return _apply_combo_filter

            apply_combo_filter = _make_combo_filter(map_combo, map_search_entry, suggestion_box)
            map_search_entry.bind(
                "<KeyRelease>",
                lambda e, entry=map_search_entry, fn=apply_combo_filter: fn(entry.get()),
            )
            apply_combo_filter()

            row_control = {
                "missing": missing,
                "map_combo": map_combo,
                "map_search_entry": map_search_entry,
                "suggestion_box": suggestion_box,
                "apply_combo_filter": apply_combo_filter,
                "create_var": create_var,
                "create_action_btn": create_action_btn,
                "new_name_entry": new_name_entry,
                "parent_combo": parent_combo,
                "sync_create_ui": sync_create_ui_fn,
                "extra": detail_extra,
                "is_party": is_party_row,
            }

            def on_create_or_edit(rc=row_control):
                if rc["create_var"].get():
                    # Already marked for creation — go straight to edit details
                    _open_create_details_dialog(rc)
                else:
                    # First click — show choice: GST Portal or Manual
                    _show_create_choice_dialog(rc)

            create_action_btn.configure(command=on_create_or_edit)

            row_controls.append(row_control)

        btn_row = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_row.pack(fill="x", padx=16, pady=(0, 14))

        refresh_btn = ctk.CTkButton(
            btn_row,
            text="Refresh Ledgers From Tally",
            width=210,
            height=38,
            font=("Segoe UI", 11, "bold"),
            fg_color=COLORS["warning"],
            hover_color="#B45309",
            text_color="#FFFFFF",
        )
        refresh_btn.pack(side="left", padx=(0, 6))

        def do_refresh_ledgers():
            refresh_btn.configure(state="disabled", text="Refreshing...")

            def worker():
                fetch = _fetch_tally_ledgers(tally_url, timeout=min(timeout, 20), company_name=selected_company)

                def finish():
                    refresh_btn.configure(state="normal", text="Refresh Ledgers From Tally")
                    if fetch.get("success"):
                        current_tally_ledgers.clear()
                        current_tally_ledgers.extend(_sanitize_dropdown_values(fetch.get("ledgers", [])))
                        for r in row_controls:
                            r["map_search_entry"].delete(0, "end")
                            r["suggestion_box"].delete(0, "end")
                            r["apply_combo_filter"]("")
                            r["map_combo"].set(search_placeholder)
                        tally_count_label.configure(text=f"Existing ledgers in Tally: {len(current_tally_ledgers)}")
                        self.log_panel.log(f"Refreshed Tally ledger list: {len(current_tally_ledgers)} found", "info")
                    else:
                        err = str(fetch.get("error") or "Unknown error")
                        self.log_panel.log(f"Could not refresh ledger list: {err}", "warning")
                        messagebox.showwarning("Refresh Failed", f"Could not fetch ledgers from Tally.\n\n{err}")

                self.after(0, finish)

            threading.Thread(target=worker, daemon=True).start()

        refresh_btn.configure(command=do_refresh_ledgers)
        # Auto-fetch a fresh full list once the dialog is visible.
        self.after(120, do_refresh_ledgers)

        retry_state = {"started": False}
        retry_btn_ref = {"btn": None}

        def do_save_all_details():
            saved_rows = 0
            for row in row_controls:
                if not row.get("create_var") or not row["create_var"].get():
                    continue
                row["extra"] = dict(row.get("extra") or {})
                saved_rows += 1

            self.log_panel.log(f"Saved create details for {saved_rows} row(s).", "info")
            messagebox.showinfo("Details Saved", f"Saved details for {saved_rows} row(s).")

        def do_retry_push():
            if retry_state["started"] or self.tally_push_is_running:
                return

            ledger_mapping = {}
            create_specs = []
            unresolved = []
            existing_keys = {_ledger_key(x) for x in current_tally_ledgers}

            for row in row_controls:
                missing_name = _normalize_ledger_name(row["missing"])
                selected_name = _normalize_ledger_name(row["map_combo"].get() or "")
                if _ledger_key(selected_name) == _ledger_key(search_placeholder):
                    selected_name = ""
                if not selected_name:
                    selected_name = _normalize_ledger_name(row["map_search_entry"].get() or "")
                create_enabled = bool(row["create_var"].get())
                new_name = _normalize_ledger_name(row["new_name_entry"].get() or "")
                parent_name = _normalize_ledger_name(row["parent_combo"].get() or "") or _guess_parent_for_ledger(
                    missing_name,
                    usage_map.get(missing_name, {}),
                )

                if create_enabled:
                    if not new_name:
                        unresolved.append(missing_name)
                        continue
                    create_specs.append({
                        "name": new_name,
                        "parent": parent_name,
                        "is_party": bool(row.get("is_party")),
                        "extra": dict(row.get("extra") or {}),
                    })
                    if _ledger_key(new_name) != _ledger_key(missing_name):
                        ledger_mapping[missing_name] = new_name
                    continue

                if selected_name:
                    if _ledger_key(selected_name) == _ledger_key(missing_name) and _ledger_key(selected_name) not in existing_keys:
                        unresolved.append(missing_name)
                    elif _ledger_key(selected_name) != _ledger_key(missing_name):
                        ledger_mapping[missing_name] = selected_name
                    continue

                unresolved.append(missing_name)

            if unresolved:
                unresolved_text = ", ".join(unresolved[:5])
                if len(unresolved) > 5:
                    unresolved_text += f" and {len(unresolved) - 5} more"
                messagebox.showwarning(
                    "Incomplete Resolution",
                    "Please resolve all missing ledgers before retrying.\n\n"
                    f"Unresolved: {unresolved_text}",
                )
                return

            dedup_create = []
            seen_new = set()
            for spec in create_specs:
                key = _ledger_key(spec["name"])
                if key in seen_new:
                    continue
                seen_new.add(key)
                dedup_create.append(spec)

            retry_state["started"] = True
            if retry_btn_ref["btn"] is not None:
                retry_btn_ref["btn"].configure(state="disabled", text="Retrying...")

            dialog.destroy()
            self._retry_post_with_ledger_resolution(
                xml_path=xml_path,
                tally_url=tally_url,
                timeout=timeout,
                base_xml_content=base_xml_content,
                ledger_mapping=ledger_mapping,
                create_specs=dedup_create,
                forced_date_count=forced_date_count,
                selected_company=selected_company,
                date_mode=date_mode,
                custom_tally_date=custom_tally_date,
            )

        ctk.CTkButton(
            btn_row,
            text="Save Details",
            width=130,
            height=38,
            font=("Segoe UI", 11, "bold"),
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=do_save_all_details,
        ).pack(side="left", padx=6)

        retry_btn = ctk.CTkButton(
            btn_row,
            text="Apply & Retry Push",
            height=38,
            font=("Segoe UI", 12, "bold"),
            fg_color=COLORS["success"],
            hover_color="#047857",
            text_color="#FFFFFF",
            corner_radius=8,
            command=do_retry_push,
        )
        retry_btn.pack(side="left", fill="x", expand=True, padx=6)
        retry_btn_ref["btn"] = retry_btn

        ctk.CTkButton(
            btn_row,
            text="Cancel",
            width=110,
            height=38,
            font=("Segoe UI", 11, "bold"),
            fg_color=COLORS["bg_input"],
            hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=dialog.destroy,
        ).pack(side="right", padx=(6, 0))

    def _retry_post_with_ledger_resolution(
        self,
        xml_path,
        tally_url,
        timeout,
        base_xml_content,
        ledger_mapping,
        create_specs,
        forced_date_count,
        selected_company="",
        date_mode="current",
        custom_tally_date="",
    ):
        if self.tally_push_is_running:
            return

        self._set_tally_push_running_ui(True, "Applying ledger resolution and retrying push...", COLORS["warning"])
        self.status_label.configure(text="Retrying", text_color=COLORS["warning"])
        self.progress_bar.set(0.35)
        if ledger_mapping:
            self.log_panel.log(f"Applying ledger remap for {len(ledger_mapping)} item(s) before retry.", "process")
        if create_specs:
            self.log_panel.log(f"Creating {len(create_specs)} ledger(s) in Tally before retry.", "process")

        threading.Thread(
            target=self._retry_post_with_ledger_resolution_worker,
            args=(
                xml_path,
                tally_url,
                timeout,
                base_xml_content,
                ledger_mapping,
                create_specs,
                forced_date_count,
                selected_company,
                date_mode,
                custom_tally_date,
            ),
            daemon=True,
        ).start()

    def _retry_post_with_ledger_resolution_worker(
        self,
        xml_path,
        tally_url,
        timeout,
        base_xml_content,
        ledger_mapping,
        create_specs,
        forced_date_count,
        selected_company="",
        date_mode="current",
        custom_tally_date="",
    ):
        result = {"success": False, "error": "Unknown error"}
        working_xml = base_xml_content
        created_ledgers = []
        create_failures = []
        replaced_count = 0
        missing_ledgers = []
        tally_ledgers = []
        usage_map = _extract_ledger_usage_from_xml(base_xml_content or "")
        ledger_fetch_error = ""

        try:
            if not working_xml:
                working_xml = _read_xml_text_safely(xml_path)
                working_xml, touched_count = _apply_push_date_mode(
                    working_xml,
                    date_mode=date_mode,
                    custom_tally_date=custom_tally_date,
                )
                forced_date_count = max(forced_date_count, touched_count)

            if selected_company:
                working_xml, _ = _set_svcurrentcompany(working_xml, selected_company)

            if ledger_mapping:
                working_xml, replaced_count = _apply_ledger_mapping_to_xml(working_xml, ledger_mapping)

            for spec in create_specs:
                create_result = _create_tally_ledger(
                    tally_url,
                    spec["name"],
                    spec["parent"],
                    timeout=min(timeout, 30),
                    is_party=bool(spec.get("is_party")),
                    extra_info=spec.get("extra", {}),
                    company_name=selected_company,
                )
                if create_result.get("success"):
                    created_ledgers.append(spec["name"])
                else:
                    create_failures.append((spec, create_result))

            fetch_result = _fetch_tally_ledgers(tally_url, timeout=min(timeout, 20), company_name=selected_company)
            if fetch_result.get("success"):
                tally_ledgers = fetch_result.get("ledgers", [])
                existing_keys = {_ledger_key(x) for x in tally_ledgers}
                existing_keys.update({_ledger_key(x) for x in created_ledgers})

                usage_map = _extract_ledger_usage_from_xml(working_xml)
                precheck_missing = _collect_missing_ledgers_from_usage(usage_map, existing_keys)
                if precheck_missing:
                    missing_ledgers = precheck_missing
                    result = {"success": False, "error": "Missing ledgers found before retry-posting."}
                    raise RuntimeError("PRECHECK_MISSING_LEDGERS")
            else:
                ledger_fetch_error = str(fetch_result.get("error") or "")
                result = {
                    "success": False,
                    "error": (
                        "Could not fetch ledger list from Tally for retry pre-check. "
                        "Posting was not attempted to avoid duplicate entries. "
                        f"Details: {ledger_fetch_error or 'Unknown error'}"
                    ),
                }
                raise RuntimeError("PRECHECK_FETCH_FAILED")

            voucher_count_in_xml = working_xml.count("<TALLYMESSAGE")
            self.after(0, lambda n=voucher_count_in_xml: self.log_panel.log(
                f"Retrying push: {n} voucher(s) in XML...", "process"
            ))

            result, working_xml, retry_meta = _post_xml_with_fallbacks(
                tally_url,
                working_xml,
                timeout=timeout,
                allow_company_fallback=not bool(selected_company),
                allow_date_retry=(date_mode == "current"),
            )
            forced_date_count = max(forced_date_count, retry_meta.get("forced_date_count", 0))
            if retry_meta.get("date_retry_used"):
                result["date_retry_used"] = True
            if retry_meta.get("fallback_used"):
                result["fallback_used"] = True

            if not result.get("success"):
                parsed_missing = _extract_missing_ledger_names(result.get("error", ""))
                if parsed_missing:
                    if not tally_ledgers:
                        fetch_result = _fetch_tally_ledgers(tally_url, timeout=min(timeout, 20), company_name=selected_company)
                        if fetch_result.get("success"):
                            tally_ledgers = fetch_result.get("ledgers", [])
                        else:
                            ledger_fetch_error = str(fetch_result.get("error") or "")

                    if tally_ledgers:
                        existing_keys = {_ledger_key(x) for x in tally_ledgers}
                        xml_ledgers = _extract_ledger_names_from_xml(working_xml)
                        for name in xml_ledgers:
                            n_name = _normalize_ledger_name(name)
                            if _ledger_key(n_name) not in existing_keys and _ledger_key(n_name) not in {_ledger_key(x) for x in missing_ledgers}:
                                missing_ledgers.append(n_name)

                    for missing in parsed_missing:
                        n_missing = _normalize_ledger_name(missing)
                        if _ledger_key(n_missing) not in {_ledger_key(x) for x in missing_ledgers}:
                            missing_ledgers.append(n_missing)

                    missing_ledgers = sorted(missing_ledgers, key=lambda x: _ledger_key(x))
                    if not usage_map:
                        usage_map = _extract_ledger_usage_from_xml(working_xml)
        except Exception as exc:
            if str(exc) not in {"PRECHECK_MISSING_LEDGERS", "PRECHECK_FETCH_FAILED"}:
                result = {"success": False, "error": str(exc)}

        def done():
            self._set_tally_push_running_ui(False)

            if replaced_count > 0:
                self.log_panel.log(f"Applied {replaced_count} XML ledger-name replacement(s).", "info")
            for spec, create_result in create_failures:
                self.log_panel.log(
                    f"Could not create ledger '{spec['name']}' under '{spec['parent']}': {create_result.get('error', 'Unknown error')}",
                    "warning",
                )

            if result.get("success"):
                self.progress_bar.set(1.0)
                self._show_tally_push_success(
                    result,
                    forced_date_count=forced_date_count,
                    created_ledgers=sorted(set(created_ledgers), key=lambda x: x.upper()),
                    target_company=selected_company,
                    date_mode=date_mode,
                    custom_tally_date=custom_tally_date,
                )
                return

            if missing_ledgers:
                self.progress_bar.set(0.0)
                self.status_label.configure(text="Ledger Mapping Needed", text_color=COLORS["warning"])
                self.progress_label.configure(text="Resolve missing ledgers to continue", text_color=COLORS["warning"])
                self.log_panel.log("Retry still failed due to missing ledger(s).", "warning")
                if ledger_fetch_error:
                    self.log_panel.log(f"Could not fetch full ledger list from Tally: {ledger_fetch_error}", "warning")
                self._show_missing_ledger_resolution_dialog(
                    xml_path=xml_path,
                    tally_url=tally_url,
                    timeout=timeout,
                    base_xml_content=working_xml,
                    missing_ledgers=missing_ledgers,
                    tally_ledgers=tally_ledgers,
                    usage_map=usage_map,
                    forced_date_count=forced_date_count,
                    selected_company=selected_company,
                    date_mode=date_mode,
                    custom_tally_date=custom_tally_date,
                )
                return

            self.progress_bar.set(0.0)
            self.status_label.configure(text="Push Failed", text_color=COLORS["error"])
            self.progress_label.configure(text="Tally push failed", text_color=COLORS["error"])
            self.log_panel.log(f"Tally push failed: {result.get('error', 'Unknown error')}", "error")
            messagebox.showerror("Push To Tally Failed", result.get("error", "Unknown error"))

        self.after(0, done)

    # ─── FILE BROWSING ───

    def _browse_file(self):
        filepath = filedialog.askopenfilename(title="Select GSTR-2B Excel File",
                                               filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")])
        if filepath:
            self.source_file = filepath
            filename = Path(filepath).name
            self.file_icon_label.configure(text=f"  {filename}", text_color=COLORS["success"])
            self.upload_zone.configure(border_color=COLORS["success"])
            self.log_panel.log(f"File selected: {filename}", "info")
            if not self.output_entry.get():
                self.output_dir = str(Path(filepath).parent)
            self.preview_table.load_excel(filepath)
            self._parse_gstr2b()

    def _browse_mapping(self):
        filepath = filedialog.askopenfilename(title="Select Mapping Excel",
                                               filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")])
        if filepath:
            self.mapping_file = filepath
            count = self.engine.load_mapping(filepath)
            if count > 0:
                self.mapping_label.configure(
                    text=f"Loaded {count} unique party mappings from {Path(filepath).name}",
                    text_color=COLORS["success"])
                self.log_panel.log(f"Mapping loaded: {count} party→ledger entries from {Path(filepath).name}", "map")
                # Show sample mappings
                shown = 0
                for party, ledger in self.engine.party_ledger_map.items():
                    if shown >= 3:
                        remaining = len(self.engine.party_ledger_map) - 3
                        self.log_panel.log(f"  ... and {remaining} more", "map")
                        break
                    self.log_panel.log(f"  {party[:40]} → {ledger}", "map")
                    shown += 1
            else:
                self.mapping_label.configure(text="Failed to load mapping — check file format", text_color=COLORS["error"])
                self.log_panel.log("Mapping load failed — need PartyLedger + PurchaseLedger columns", "error")

    def _clear_mapping(self):
        self.engine.party_ledger_map = {}
        self.engine.party_tds_ledger_map = {}
        self.engine.party_tds_rate_map = {}
        self.mapping_file = ""
        self.mapping_label.configure(text="No mapping loaded — all → Purchase Account",
                                      text_color=COLORS["text_muted"])
        self.log_panel.log("Mapping cleared — using default purchase ledger.", "info")

    def _download_mapping_template(self):
        initial_dir = self.output_dir or os.getcwd()
        if not os.path.isdir(initial_dir):
            initial_dir = os.getcwd()

        save_path = filedialog.asksaveasfilename(
            title="Save Mapping Template",
            defaultextension=".xlsx",
            initialfile="mapping_template.xlsx",
            initialdir=initial_dir,
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mapping Template"

            headers = ["Party Ledger", "Mapping", "TDS Ledger", "TDS Rate"]
            ws.append(headers)

            # Keep template rows generic; do not include any client data.
            sample_rows = [
                ["SAMPLE PARTY A", "Office Maintenance", "", ""],
                ["SAMPLE PARTY B", "Professional Fees", "TDS Payable On Professional 194J", 10],
                ["SAMPLE PARTY C", "Contract Expense", "TDS Payable On Contract 194C", 2],
            ]
            for row in sample_rows:
                ws.append(row)

            widths = [38, 32, 38, 12]
            for idx, (header, width) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row=1, column=idx)
                cell.value = header
                cell.font = Font(name="Calibri", size=11, bold=True)
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

            for r in range(2, 2 + len(sample_rows)):
                rate_cell = ws.cell(row=r, column=4)
                if rate_cell.value not in ("", None):
                    rate_cell.number_format = "0.##"

            wb.save(save_path)
            wb.close()

            self.log_panel.log(f"Mapping template saved: {Path(save_path).name}", "success")
            messagebox.showinfo(
                "Template Downloaded",
                "Mapping template saved successfully.\n\n"
                "The sample rows are placeholders (no client details).",
            )
        except Exception as exc:
            self.log_panel.log(f"Could not save mapping template: {exc}", "error")
            messagebox.showerror("Template Error", f"Could not save template.\n\n{exc}")

    def _generate_itc_template(self):
        if not self.engine.records:
            messagebox.showwarning("No Data", "Load a GSTR-2B file first.")
            return

        initial_dir = self.output_dir or os.getcwd()
        if not os.path.isdir(initial_dir):
            initial_dir = os.getcwd()
        save_path = filedialog.asksaveasfilename(
            title="Save ITC Template",
            defaultextension=".xlsx",
            initialfile="ITC_Template.xlsx",
            initialdir=initial_dir,
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            from openpyxl.utils import get_column_letter
            from copy import copy as _copy

            # ── ITC B2B tab helpers ───────────────────────────────────────
            # Yellow (auto-filled): A=GSTIN, B=Trade/Legal name, C=Invoice Number,
            #   D=Invoice Date, E=Invoice Value, F=Taxable Value, G=CGST, H=SGST,
            #   I=IGST, J=CESS, K=ITC Availability, L=Supply Attract Reverse Charge
            # Green (user fills): M=ITC to be claimed or not (Yes/No),
            #   N=Whether Contains stock Item (Yes/No),
            #   O=Mapping Ledger, P=TDS Ledger, Q=TDS Rate
            yellow_headers = [
                "GSTIN", "Trade/Legal name", "Invoice Number", "Invoice Date", "Invoice Value",
                "Taxable Value", "CGST", "SGST", "IGST", "CESS",
                "ITC Availability", "Supply Attract Reverse Charge",
            ]
            green_headers = [
                "ITC to be claimed or not", "Whether Contains stock Item",
                "Mapping Ledger", "TDS Ledger", "TDS Rate",
            ]
            all_headers  = yellow_headers + green_headers
            n_yellow     = len(yellow_headers)   # 12
            dv_col_start = n_yellow + 1           # col M (1-based)
            dv_col_end   = n_yellow + 2           # col N (1-based)

            yellow_fill  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            green_fill   = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            header_font  = Font(name="Calibri", size=11, bold=True)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            data_font    = Font(name="Calibri", size=10)
            num_align    = Alignment(horizontal="right")
            col_widths   = [20, 32, 22, 14, 20, 14, 10, 10, 10, 10, 24, 30, 26, 28, 30, 30, 12]

            # Open source GSTR-2B workbook for raw sheet access
            src_wb = openpyxl.load_workbook(self.source_file, data_only=True)
            all_src_upper = {s.upper(): s for s in src_wb.sheetnames}

            # ── Build ITC B2B tab from RAW source B2B sheet (no record skipping) ───
            def _build_itc_b2b_sheet(dest_ws):
                src_name = all_src_upper.get("B2B")
                if not src_name:
                    return 0
                src_ws = src_wb[src_name]

                # Detect header rows and data start using same logic as engine
                mapper = B2BColumnMapper()
                mapper.detect_columns(src_ws)
                data_start = mapper.data_start_row
                has_rate_col = mapper.has("rate")
                # Extend yellow headers with Rate column only if source GSTR-2B has one
                _yellow_hdrs = yellow_headers + (["Rate (%)"] if has_rate_col else [])
                _all_hdrs    = _yellow_hdrs + green_headers
                _n_yellow    = len(_yellow_hdrs)          # 12 without Rate, 13 with Rate
                _dv_col_s    = _n_yellow + 1              # first green col (1-based)
                _dv_col_e    = _n_yellow + 2              # second green col (1-based)
                _col_widths  = col_widths[:12] + ([10] if has_rate_col else []) + col_widths[12:]

                # Write ITC template headers (row 1)
                for col_idx, header in enumerate(_all_hdrs, start=1):
                    cell = dest_ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.fill = yellow_fill if col_idx <= _n_yellow else green_fill
                dest_ws.row_dimensions[1].height = 32

                # Write data rows from raw B2B sheet
                out_row = 2
                _itc_last_gstin      = ""
                _itc_last_trade_name = ""
                _itc_trade_col = mapper.get("trade_name", 1)
                for row_idx in range(data_start, src_ws.max_row + 1):
                    row = [cell.value for cell in src_ws[row_idx]]
                    gstin_col = mapper.get("gstin", 0)
                    # Forward-fill GSTIN/trade name for GSTR-2B merged-cell rows
                    cur_gstin = str(row[gstin_col] or "").strip() if gstin_col < len(row) else ""
                    if cur_gstin and len(cur_gstin) >= 15:
                        _itc_last_gstin = cur_gstin
                        _itc_last_trade_name = str(row[_itc_trade_col] or "").strip() if _itc_trade_col < len(row) else ""
                    elif not cur_gstin and _itc_last_gstin:
                        row = list(row)
                        while len(row) <= max(gstin_col, _itc_trade_col):
                            row.append(None)
                        row[gstin_col] = _itc_last_gstin
                        if not row[_itc_trade_col]:
                            row[_itc_trade_col] = _itc_last_trade_name
                    else:
                        continue
                    try:
                        rec = self.engine._parse_b2b_row(row, row_idx, mapper)
                    except Exception:
                        continue
                    if not rec:
                        continue
                    row_values = [
                        rec.get("gstin", ""),
                        rec.get("trade_name", ""),
                        rec.get("invoice_no", ""),
                        rec.get("invoice_date", ""),
                        rec.get("invoice_value", 0),   # E = Invoice Value
                        rec.get("taxable_value", 0),
                        rec.get("cgst", 0),
                        rec.get("sgst", 0),
                        rec.get("igst", 0),
                        rec.get("cess", 0),
                        rec.get("itc_avail", ""),
                        rec.get("reverse_charge", "No"),
                    ]
                    if has_rate_col:
                        _rate_col_idx = mapper.get("rate")
                        _raw_rate = row[_rate_col_idx] if _rate_col_idx is not None and _rate_col_idx < len(row) else ""
                        row_values.append(_raw_rate)
                    for col_idx, value in enumerate(row_values, start=1):
                        cell = dest_ws.cell(row=out_row, column=col_idx, value=value)
                        cell.font = data_font
                        if isinstance(value, (int, float)):
                            cell.alignment = num_align
                    out_row += 1

                # Yes/No validation on M and N columns
                last_row = max(out_row - 1, 2)
                dv_start = get_column_letter(_dv_col_s)
                dv_end   = get_column_letter(_dv_col_e)
                dv = OpenpyxlDataValidation(
                    type="list", formula1='"Yes,No"', allow_blank=True,
                    showDropDown=False, showErrorMessage=True,
                    errorTitle="Invalid Value",
                    error="Only 'Yes' or 'No' is allowed. Please select from the dropdown.",
                    errorStyle="stop",
                )
                dv.sqref = f"{dv_start}2:{dv_end}{last_row}"
                dest_ws.add_data_validation(dv)

                for col_idx, width in enumerate(_col_widths, start=1):
                    dest_ws.column_dimensions[
                        dest_ws.cell(row=1, column=col_idx).column_letter
                    ].width = width
                dest_ws.freeze_panes = "A2"
                return last_row - 1  # data row count

            # ── Full formatted copy for B2BA / B2B-CDNR sheets ──────────────
            def _copy_sheet_formatted(src_ws, dest_ws):
                """Copy values, styles, column widths, row heights and merged cells."""
                # Copy column widths
                for col_letter, dim in src_ws.column_dimensions.items():
                    dest_ws.column_dimensions[col_letter].width = dim.width
                    dest_ws.column_dimensions[col_letter].hidden = dim.hidden

                # Copy row heights
                for row_num, dim in src_ws.row_dimensions.items():
                    dest_ws.row_dimensions[row_num].height = dim.height
                    dest_ws.row_dimensions[row_num].hidden = dim.hidden

                # Copy merged cells
                for merge in src_ws.merged_cells.ranges:
                    dest_ws.merge_cells(str(merge))

                # Copy cell values and styles
                for row in src_ws.iter_rows():
                    for src_cell in row:
                        dest_cell = dest_ws.cell(row=src_cell.row, column=src_cell.column,
                                                  value=src_cell.value)
                        if src_cell.has_style:
                            dest_cell.font      = _copy(src_cell.font)
                            dest_cell.fill      = _copy(src_cell.fill)
                            dest_cell.border    = _copy(src_cell.border)
                            dest_cell.alignment = _copy(src_cell.alignment)
                            dest_cell.number_format = src_cell.number_format

                dest_ws.freeze_panes = src_ws.freeze_panes
                dest_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines

            # ── Assemble workbook ─────────────────────────────────────────
            wb = openpyxl.Workbook()

            # Tab 1: ITC B2B (processed + green columns)
            ws_b2b = wb.active
            ws_b2b.title = "ITC B2B"
            b2b_count = _build_itc_b2b_sheet(ws_b2b)

            # Tab 2: B2BA — exact formatted copy
            ws_b2ba = wb.create_sheet("B2BA")
            if "B2BA" in all_src_upper:
                _copy_sheet_formatted(src_wb[all_src_upper["B2BA"]], ws_b2ba)

            # Tab 3: B2B CDNR — exact formatted copy
            ws_cdnr = wb.create_sheet("B2B CDNR")
            if "B2B-CDNR" in all_src_upper:
                _copy_sheet_formatted(src_wb[all_src_upper["B2B-CDNR"]], ws_cdnr)

            src_wb.close()
            wb.save(save_path)

            self.mapping_label.configure(
                text=(
                    f"Template saved: {Path(save_path).name} "
                    f"(B2B={b2b_count} rows) — fill green columns and upload back"
                ),
                text_color=COLORS["success"],
            )
            self.log_panel.log(
                f"ITC Template generated: {Path(save_path).name} — "
                f"ITC B2B={b2b_count} rows, B2BA+CDNR exact copy",
                "success",
            )
            messagebox.showinfo(
                "Template Downloaded",
                f"ITC Template saved!\n{save_path}\n\n"
                f"ITC B2B: {b2b_count} rows (ALL B2B entries, no skipping)\n"
                "B2BA: exact copy from GSTR-2B\n"
                "B2B CDNR: exact copy from GSTR-2B\n\n"
                "Fill the green columns (M–Q) in ITC B2B and upload back.",
            )
        except Exception as exc:
            self.log_panel.log(f"ITC template generation failed: {exc}", "error")
            messagebox.showerror("Error", f"Failed to generate template:\n{exc}")

    def _upload_itc_template(self):
        filepath = filedialog.askopenfilename(
            title="Select ITC Template",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if not filepath:
            return

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)

            # ── Work from "ITC B2B" sheet (fall back to first sheet) ──────────
            if "ITC B2B" in wb.sheetnames:
                ws = wb["ITC B2B"]
            else:
                ws = wb.active

            # ITC B2B template column layout (0-indexed):
            # 0=GSTIN, 1=Trade/Legal name, 2=Invoice Number, 3=Invoice Date,
            # 4=Invoice Value, 5=Taxable Value, 6=CGST, 7=SGST, 8=IGST, 9=CESS,
            # 10=ITC Availability, 11=Supply Attract Reverse Charge
            # 12=ITC to be claimed (Yes/No), 13=Whether Contains stock Item (Yes/No)
            # 14=Mapping Ledger, 15=TDS Ledger, 16=TDS Rate

            def _cell(row, idx, default=""):
                if idx < len(row) and row[idx] is not None:
                    return row[idx]
                return default

            def _str(row, idx, default=""):
                return str(_cell(row, idx, default) or "").strip()

            def _float(row, idx, default=0.0):
                v = _cell(row, idx, None)
                if v is None or v == "":
                    return default
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return default

            def _norm_date(raw):
                if isinstance(raw, (datetime.datetime, datetime.date)):
                    return raw.strftime("%d/%m/%Y")
                eng = GSTR2BEngine()
                return eng._normalize_date_str(raw) if raw else ""

            # Reset engine — fresh start
            new_engine = GSTR2BEngine()
            # Inherit company identity so voucher XML gets full GST registration block
            new_engine.company_gstin             = self.engine.company_gstin
            new_engine.company_registration_name  = self.engine.company_registration_name
            new_engine.company_registration_state = self.engine.company_registration_state
            new_engine.company_name   = self.engine.company_name
            new_engine.trade_name     = self.engine.trade_name
            new_engine.financial_year = self.engine.financial_year
            new_engine.tax_period     = self.engine.tax_period

            records       = []
            party_ledger  = {}
            party_tds_ldr = {}
            party_tds_rt  = {}
            itc_map       = {}

            # Detect if this template was generated with a Rate column (between col 12 and green cols).
            # Older templates have green cols starting at index 12; newer ones with Rate start at 13.
            _hdr = [str(c or "").strip().lower() for c in (
                next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
            )]
            _has_rate_col = (
                len(_hdr) > 12
                and any(h in ("rate (%)", "rate(%)", "rate") and "tds" not in h for h in _hdr[11:14])
            )
            _g = 1 if _has_rate_col else 0  # offset applied to all green column indices

            # ── Parse ITC B2B sheet ───────────────────────────────────────────
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue
                gstin      = _str(row, 0)
                trade_name = _str(row, 1)
                # Skip only if BOTH GSTIN and trade name are absent (truly empty row).
                # Rows without a valid GSTIN but with a trade name are manually-added
                # purchases (not in GSTR-2B) and should be included.
                if (not gstin or len(gstin) < 15) and not trade_name:
                    continue
                invoice_no    = _str(row, 2)
                invoice_date  = _norm_date(_cell(row, 3, ""))
                invoice_value = _float(row, 4)           # col E = Invoice Value
                taxable       = _float(row, 5)
                cgst          = _float(row, 6)
                sgst          = _float(row, 7)
                igst          = _float(row, 8)
                cess          = _float(row, 9)
                itc_avail_src = _str(row, 10, "Yes")
                rev_charge    = _str(row, 11, "No")

                # Green columns — indices shift right by 1 when Rate column is present
                itc_claimed    = _str(row, 12 + _g)
                has_stock      = _str(row, 13 + _g)
                mapping_ledger = _str(row, 14 + _g)
                tds_ledger     = _str(row, 15 + _g)
                tds_rate_raw   = _cell(row, 16 + _g, None)
                try:
                    tds_rate = float(tds_rate_raw) if tds_rate_raw not in (None, "") else 0.0
                except (ValueError, TypeError):
                    tds_rate = 0.0

                itc_upper = itc_claimed.upper()
                if itc_upper in ("YES", "Y"):
                    eff_itc = "Yes"
                elif itc_upper in ("NO", "N"):
                    eff_itc = "Ineligible"
                else:
                    eff_itc = itc_avail_src or "Yes"

                if _has_rate_col:
                    # Read rate directly from the Rate (%) column in the template
                    rate_raw = _cell(row, 12, None)
                    try:
                        rate = float(rate_raw) if rate_raw not in (None, "") else 0.0
                    except (ValueError, TypeError):
                        rate = 0.0
                else:
                    # Compute rate from tax amounts when no Rate column in template
                    total_tax   = abs(igst) + abs(cgst) + abs(sgst)
                    abs_taxable = abs(taxable)
                    rate = round((total_tax / abs_taxable) * 100, 0) if abs_taxable > 0 and total_tax > 0 else 0.0

                rec = {
                    "gstin": gstin, "trade_name": trade_name, "party_name": trade_name,
                    "invoice_no": invoice_no, "invoice_date": invoice_date,
                    "invoice_type": "Regular",   # B2B tab = regular invoices
                    "invoice_value": invoice_value,
                    "taxable_value": taxable, "cgst": cgst, "sgst": sgst, "igst": igst, "cess": cess,
                    "rate": rate, "itc_avail": eff_itc, "reverse_charge": rev_charge,
                    "has_stock_item": has_stock.upper() in ("YES", "Y"),
                    "tds_ledger": tds_ledger, "tds_rate": tds_rate, "tds_amount": "",
                    "purchase_ledger": mapping_ledger,
                    "sheet_type": "B2B", "is_amendment": False, "orig_invoice_no": "",
                    "filing_period": "", "place_of_supply": "", "party_state": "",
                    "party_mailing_name": trade_name, "party_address1": "", "party_address2": "",
                    "party_pincode": "", "voucher_date": "", "voucher_no": "",
                    "supplier_invoice_no": invoice_no, "supplier_invoice_date": invoice_date,
                    "narration": "", "row_idx": row_idx,
                }
                records.append(rec)

                key = trade_name.upper()
                if trade_name and mapping_ledger:
                    party_ledger[key] = mapping_ledger
                if trade_name and tds_ledger:
                    party_tds_ldr[key] = tds_ledger
                if trade_name and tds_rate:
                    party_tds_rt[key] = tds_rate
                if invoice_no:
                    itc_map[invoice_no.upper()] = {"itc_claimed": itc_claimed, "has_stock": has_stock}

            b2b_count_raw = len(records)  # B2B rows before filtering

            # ── Parse B2BA sheet (amendment records) ─────────────────────────
            # Collect amended invoice numbers first, then:
            # - If a B2B invoice_no matches a B2BA orig_invoice_no → skip the B2B record
            # - If no match → keep both B2B record AND B2BA record
            b2ba_recs = []          # parsed B2BA records (to append later)
            b2ba_orig_nos = set()   # normalised original invoice numbers from B2BA
            b2ba_ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.upper() == "B2BA":
                    b2ba_ws = wb[sheet_name]
                    break

            if b2ba_ws:
                b2ba_mapper = B2BColumnMapper()
                b2ba_mapper.detect_columns(b2ba_ws)
                if b2ba_mapper.has("gstin"):
                    for row_idx in range(b2ba_mapper.data_start_row, b2ba_ws.max_row + 1):
                        row = [cell.value for cell in b2ba_ws[row_idx]]
                        gstin_col = b2ba_mapper.get("gstin", 0)
                        if gstin_col >= len(row) or not row[gstin_col]:
                            continue
                        try:
                            rec = new_engine._parse_b2b_row(row, row_idx, b2ba_mapper)
                        except Exception:
                            continue
                        if not rec:
                            continue
                        rec["is_amendment"] = True
                        rec["sheet_type"]   = "B2BA"
                        key = rec.get("trade_name", "").upper()
                        rec["purchase_ledger"] = party_ledger.get(key, "")
                        rec["tds_ledger"]      = party_tds_ldr.get(key, "")
                        rec["tds_rate"]        = party_tds_rt.get(key, 0.0)
                        rec["tds_amount"]      = ""
                        rec.setdefault("party_name", rec.get("trade_name", ""))
                        rec.setdefault("party_mailing_name", rec.get("trade_name", ""))
                        rec.setdefault("party_address1", "")
                        rec.setdefault("party_address2", "")
                        rec.setdefault("party_pincode", "")
                        rec.setdefault("voucher_date", "")
                        rec.setdefault("voucher_no", "")
                        supplier_inv = rec.get("supplier_invoice_no") or rec.get("invoice_no", "")
                        rec.setdefault("supplier_invoice_no", supplier_inv)
                        rec.setdefault("supplier_invoice_date", rec.get("invoice_date", ""))
                        rec.setdefault("narration", "")
                        b2ba_recs.append(rec)
                        # Track which B2B invoice numbers are superseded
                        orig_no = rec.get("orig_invoice_no", "").strip().upper()
                        inv_no  = rec.get("invoice_no", "").strip().upper()
                        if orig_no:
                            b2ba_orig_nos.add(orig_no)
                        elif inv_no:
                            b2ba_orig_nos.add(inv_no)

            # Filter B2B records: skip those superseded by a B2BA amendment
            skipped_b2b = 0
            if b2ba_orig_nos:
                filtered_b2b = []
                for r in records:
                    if r.get("sheet_type") == "B2B" and r.get("invoice_no", "").strip().upper() in b2ba_orig_nos:
                        skipped_b2b += 1
                    else:
                        filtered_b2b.append(r)
                records = filtered_b2b

            # Append B2BA records (revised entries)
            records.extend(b2ba_recs)
            b2ba_count = len(b2ba_recs)
            b2b_count  = b2b_count_raw - skipped_b2b
            if skipped_b2b:
                self.log_panel.log(
                    f"B2BA: {b2ba_count} amendment(s) found — "
                    f"{skipped_b2b} B2B invoice(s) superseded and skipped.",
                    "info",
                )

            # ── Parse B2B CDNR sheet (credit / debit notes) ───────────────────
            cdnr_count = 0
            cdnr_ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.upper() in ("B2B CDNR", "B2B-CDNR", "CDNR"):
                    cdnr_ws = wb[sheet_name]
                    break

            if cdnr_ws:
                cdnr_mapper = B2BColumnMapper()
                cdnr_mapper.detect_columns(cdnr_ws)
                if cdnr_mapper.has("gstin"):
                    for row_idx in range(cdnr_mapper.data_start_row, cdnr_ws.max_row + 1):
                        row = [cell.value for cell in cdnr_ws[row_idx]]
                        gstin_col = cdnr_mapper.get("gstin", 0)
                        if gstin_col >= len(row) or not row[gstin_col]:
                            continue
                        try:
                            rec = new_engine._parse_b2b_row(row, row_idx, cdnr_mapper)
                        except Exception:
                            continue
                        if not rec:
                            continue
                        rec["is_amendment"] = False
                        rec["sheet_type"]   = "CDNR"
                        key = rec.get("trade_name", "").upper()
                        rec["purchase_ledger"] = party_ledger.get(key, "")
                        rec["tds_ledger"]      = party_tds_ldr.get(key, "")
                        rec["tds_rate"]        = party_tds_rt.get(key, 0.0)
                        rec["tds_amount"]      = ""
                        rec.setdefault("party_name", rec.get("trade_name", ""))
                        rec.setdefault("party_mailing_name", rec.get("trade_name", ""))
                        rec.setdefault("party_address1", "")
                        rec.setdefault("party_address2", "")
                        rec.setdefault("party_pincode", "")
                        rec.setdefault("voucher_date", "")
                        rec.setdefault("voucher_no", "")
                        supplier_inv = rec.get("supplier_invoice_no") or rec.get("invoice_no", "")
                        rec.setdefault("supplier_invoice_no", supplier_inv)
                        rec.setdefault("supplier_invoice_date", rec.get("invoice_date", ""))
                        rec.setdefault("narration", "")
                        records.append(rec)
                        cdnr_count += 1

            wb.close()


            if not records:
                messagebox.showwarning(
                    "No Data",
                    "No valid records found in the ITC B2B sheet.\n"
                    "Make sure the sheet has GSTIN values in column A."
                )
                return

            # ── Populate engine with parsed records + maps ────────────────────
            new_engine.records            = records
            new_engine.party_ledger_map   = party_ledger
            new_engine.party_tds_ledger_map = party_tds_ldr
            new_engine.party_tds_rate_map = party_tds_rt
            new_engine.itc_map            = itc_map
            new_engine._compute_stats()
            self.engine = new_engine

            # ── Update UI ─────────────────────────────────────────────────────
            s = self.engine.stats
            self.stat_total.update_value(str(s["total_records"]))
            self.stat_taxable.update_value(f"Rs.{s['total_taxable']:,.0f}")
            self.stat_igst.update_value(f"Rs.{s['total_igst']:,.0f}")
            self.stat_gst.update_value(f"Rs.{s['total_cgst'] + s['total_sgst']:,.0f}")
            self.preview_table.load_excel(filepath)

            self.progress_bar.set(1.0)
            self.progress_label.configure(text="ITC Template loaded!", text_color=COLORS["success"])
            self.status_label.configure(text="Ready", text_color=COLORS["success"])

            mapped_count  = len(party_ledger)
            itc_yes_count = sum(1 for v in itc_map.values() if v.get("itc_claimed", "").upper() in ("YES", "Y"))
            itc_no_count  = sum(1 for v in itc_map.values() if v.get("itc_claimed", "").upper() in ("NO", "N"))
            stock_count   = sum(1 for v in itc_map.values() if v.get("has_stock", "").upper() in ("YES", "Y"))

            self.mapping_label.configure(
                text=(
                    f"ITC Template: B2B={b2b_count} | B2BA={b2ba_count} | CDNR={cdnr_count} | "
                    f"Total={len(records)} records | {mapped_count} ledger mappings"
                ),
                text_color=COLORS["success"],
            )
            self.log_panel.log(
                f"ITC Template loaded from {Path(filepath).name}: "
                f"B2B={b2b_count}, B2BA={b2ba_count}, CDNR={cdnr_count} "
                f"(Total={len(records)}), {mapped_count} mappings, "
                f"ITC=Yes:{itc_yes_count}, No:{itc_no_count}, stock={stock_count}",
                "map",
            )
            if mapped_count > 0:
                shown = 0
                for party, ledger in party_ledger.items():
                    if shown >= 3:
                        self.log_panel.log(f"  ... and {len(party_ledger) - 3} more", "map")
                        break
                    self.log_panel.log(f"  {party[:40]} \u2192 {ledger}", "map")
                    shown += 1
            self.log_panel.log(
                "Ready to generate \u2014 click 'Generate & Push' or any export button.", "success"
            )

        except Exception as exc:
            import traceback
            self.mapping_label.configure(text="Failed to load ITC template", text_color=COLORS["error"])
            self.log_panel.log(f"ITC Template upload failed: {exc}", "error")
            self.log_panel.log(traceback.format_exc(), "error")
            messagebox.showerror("Upload Error", f"Failed to load template:\n{exc}")


    def _browse_tally_sheet(self):
        filepath = filedialog.askopenfilename(title="Select Tally Sheet",
                                               filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")])
        if filepath:
            self.tally_sheet_file = filepath
            filename = Path(filepath).name
            self.tally_label.configure(text=f"  {filename}", text_color=COLORS["success"])
            self.tally_zone.configure(border_color=COLORS["success"])
            self.log_panel.log(f"Tally sheet selected: {filename}", "info")
            if not self.output_entry.get():
                self.output_dir = str(Path(filepath).parent)
            self.preview_table.load_excel(filepath)
            self._parse_tally_sheet()

    def _browse_output(self):
        dirpath = filedialog.askdirectory(title="Select Output Folder")
        if dirpath:
            self.output_dir = dirpath
            self.output_entry.delete(0, "end"); self.output_entry.insert(0, dirpath)

    # ─── PARSING ───

    def _parse_gstr2b(self):
        self.log_panel.log("Parsing GSTR-2B file...", "process")
        self.progress_bar.set(0)
        self.progress_label.configure(text="Parsing...", text_color=COLORS["warning"])
        self.status_label.configure(text="Processing", text_color=COLORS["warning"])
        # Preserve mapping across engine reset
        saved_map = dict(self.engine.party_ledger_map)
        saved_tds_ledger_map = dict(self.engine.party_tds_ledger_map)
        saved_tds_rate_map = dict(self.engine.party_tds_rate_map)
        self.engine = GSTR2BEngine()
        self.engine.party_ledger_map = saved_map
        self.engine.party_tds_ledger_map = saved_tds_ledger_map
        self.engine.party_tds_rate_map = saved_tds_rate_map

        def run():
            def cb(pct, msg):
                self.after(0, lambda: self.progress_bar.set(pct))
                self.after(0, lambda: self.progress_label.configure(text=msg))
            success = self.engine.parse_gstr2b(self.source_file, cb)
            self.after(0, lambda: self._on_parse_complete(success))
        threading.Thread(target=run, daemon=True).start()

    def _parse_tally_sheet(self):
        self.log_panel.log("Reading tally sheet...", "process")
        self.progress_bar.set(0)
        self.progress_label.configure(text="Reading...", text_color=COLORS["warning"])
        self.status_label.configure(text="Processing", text_color=COLORS["warning"])
        self.engine = GSTR2BEngine()

        def run():
            def cb(pct, msg):
                self.after(0, lambda: self.progress_bar.set(pct))
                self.after(0, lambda: self.progress_label.configure(text=msg))
            success = self.engine.parse_tally_sheet(self.tally_sheet_file, cb)
            self.after(0, lambda: self._on_parse_complete(success))
        threading.Thread(target=run, daemon=True).start()

    def _reset_stats(self):
        self.stat_total.update_value("—"); self.stat_taxable.update_value("—")
        self.stat_igst.update_value("—"); self.stat_gst.update_value("—")

    def _on_parse_complete(self, success):
        if success:
            s = self.engine.stats
            self.stat_total.update_value(str(s["total_records"]))
            self.stat_taxable.update_value(f"Rs.{s['total_taxable']:,.0f}")
            self.stat_igst.update_value(f"Rs.{s['total_igst']:,.0f}")
            self.stat_gst.update_value(f"Rs.{s['total_cgst'] + s['total_sgst']:,.0f}")
            _src = self.source_file if self.current_mode == "gstr2b" else self.tally_sheet_file
            if _src:
                self.preview_table.load_excel(_src)
            info_parts = []
            dn = self.engine.trade_name or self.engine.company_name
            if dn: info_parts.append(dn)
            if self.engine.company_gstin: info_parts.append(f"GSTIN: {self.engine.company_gstin}")
            if self.engine.tax_period and self.engine.financial_year:
                info_parts.append(f"{self.engine.tax_period} {self.engine.financial_year}")
            self.company_label.configure(text="  |  ".join(info_parts))
            if dn and not self.company_entry.get(): self.company_entry.insert(0, dn)
            if self.engine.company_gstin and not self.company_gstin_entry.get():
                self.company_gstin_entry.delete(0, "end")
                self.company_gstin_entry.insert(0, self.engine.company_gstin.strip().upper())
            self.progress_bar.set(1.0)
            self.progress_label.configure(text="Parsing complete!", text_color=COLORS["success"])
            self.status_label.configure(text="Ready", text_color=COLORS["success"])
            if self.current_mode == "gstr2b" and self.itc_download_btn is not None:
                self.itc_download_btn.configure(state="normal")
            if self.current_mode == "gstr2b":
                mapper = getattr(self.engine, "mapper", None)
                if mapper:
                    self.log_panel.log(f"Column auto-detection: found {len(mapper.column_map)} fields", "detect")
                    self.log_panel.log(f"  Headers at rows {mapper.header_row_1}-{mapper.header_row_2}, data starts row {mapper.data_start_row}", "detect")
                    if not mapper.has("rate"):
                        self.log_panel.log("  Rate(%) not found — auto-calculating from tax amounts", "warning")
                    else:
                        self.log_panel.log(f"  Rate(%) found at index {mapper.get('rate')}", "detect")
                if self.engine.party_ledger_map:
                    mapped = sum(1 for r in self.engine.records
                                 if r["trade_name"].upper().strip() in self.engine.party_ledger_map)
                    self.log_panel.log(f"Party mapping: {mapped}/{s['total_records']} records matched", "map")
            else:
                # Tally sheet mode
                unique_ledgers = set(r.get("purchase_ledger", "?") for r in self.engine.records)
                self.log_panel.log(f"Tally sheet loaded: {s['total_records']} vouchers", "success")
                self.log_panel.log(f"Unique purchase ledgers: {', '.join(sorted(unique_ledgers))}", "info")
            self.log_panel.log(f"Successfully parsed {s['total_records']} records", "success")
            self.log_panel.log(f"IGST: {s['igst_count']}  |  CGST/SGST: {s['cgst_sgst_count']}", "info")
            self.log_panel.log(f"Total taxable: Rs.{s['total_taxable']:,.2f}", "info")
            for w in self.engine.warnings: self.log_panel.log(w, "warning")
            for e in self.engine.errors[:5]: self.log_panel.log(e, "warning")
        else:
            self.progress_label.configure(text="Parse failed!", text_color=COLORS["error"])
            self.status_label.configure(text="Error", text_color=COLORS["error"])
            for err in self.engine.errors: self.log_panel.log(err, "error")

    # ─── UNMAPPED PARTY CHECK ───

    def _find_unmapped_parties(self, records=None):
        """Return list of unique party names not found in mapping sheet.
        A record is considered already mapped if:
          - Its trade_name is in party_ledger_map, OR
          - It has a non-empty 'purchase_ledger' embedded in the record itself
            (i.e. the user filled the Mapping Ledger column in the ITC template).
        """
        source_records = records if records is not None else self.engine.records
        unmapped = set()
        for rec in source_records:
            party = rec.get("trade_name", "").strip()
            if not party:
                continue
            # Already mapped directly in record?
            if rec.get("purchase_ledger", "").strip():
                continue
            # Mapped via party_ledger_map?
            if self.engine.party_ledger_map and party.upper() in self.engine.party_ledger_map:
                continue
            unmapped.add(party)
        return sorted(unmapped)

    def _resolve_output_dir(self):
        output_dir = (self.output_entry.get() or "").strip()
        if output_dir:
            return output_dir
        src = self.source_file if self.current_mode == "gstr2b" else self.tally_sheet_file
        if src:
            return str(Path(src).parent)
        return str(Path.cwd())

    def _save_tax_validation_report(self, output_dir, issues):
        os.makedirs(output_dir, exist_ok=True)
        report_path = _get_unique_path(output_dir, "tax-mismatch-report", ".xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tax Mismatch"

        headers = [
            "Row", "Invoice No", "Party Name", "Taxable Value",
            "IGST", "CGST", "SGST", "Cess", "Computed Tax %", "Issue",
        ]
        ws.append(headers)
        for issue in issues:
            ws.append([
                issue.get("row_idx", ""),
                issue.get("invoice_no", ""),
                issue.get("party_name", ""),
                issue.get("taxable_value", 0.0),
                issue.get("igst", 0.0),
                issue.get("cgst", 0.0),
                issue.get("sgst", 0.0),
                issue.get("cess", 0.0),
                issue.get("computed_rate", 0.0),
                issue.get("issue", ""),
            ])

        for cell in ws[1]:
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 34
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 14
        ws.column_dimensions["J"].width = 56

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=9):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(report_path)
        wb.close()
        return report_path

    def _show_unmapped_dialog(self, unmapped_parties, on_proceed, on_retry):
        """Show a dialog with editable mapping fields for each unmapped party."""
        default_ledger = self.purchase_ledger_entry.get() or "Purchase Account"

        dialog = ctk.CTkToplevel(self)
        dialog.title("Map Unmapped Party Names")
        dialog.geometry("780x580")
        dialog.resizable(True, True)
        dialog.minsize(600, 400)
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color=COLORS["bg_dark"])

        # Center on parent
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 780) // 2
        y = self.winfo_y() + (self.winfo_height() - 580) // 2
        dialog.geometry(f"+{x}+{y}")

        # Cached ledger list for search picker
        _dialog_ledgers = []

        # Header
        hdr = ctk.CTkFrame(dialog, fg_color=COLORS["error_bg"], corner_radius=10, height=60)
        hdr.pack(fill="x", padx=16, pady=(16, 4)); hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="⚠️  Unmapped Party Names", font=("Segoe UI", 15, "bold"),
                     text_color=COLORS["error"]).pack(side="left", padx=16)
        ctk.CTkLabel(hdr, text=f"{len(unmapped_parties)} parties not in mapping sheet",
                     font=("Segoe UI", 11), text_color=COLORS["text_secondary"]).pack(side="right", padx=16)

        # Info + Fetch Ledgers bar
        info_bar = ctk.CTkFrame(dialog, fg_color="transparent")
        info_bar.pack(fill="x", padx=16, pady=(4, 4))
        ctk.CTkLabel(info_bar, text="Type a ledger name or click 🔍 to search from Tally.",
                     font=("Segoe UI", 11), text_color=COLORS["text_secondary"]).pack(side="left")
        fetch_status_lbl = ctk.CTkLabel(info_bar, text="", font=("Segoe UI", 10),
                                        text_color=COLORS["text_muted"])
        fetch_status_lbl.pack(side="left", padx=8)
        fetch_ledger_btn = ctk.CTkButton(info_bar, text="Fetch Ledgers from Tally", width=200, height=30,
                                         font=("Segoe UI", 10, "bold"),
                                         fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
                                         text_color="#FFFFFF", corner_radius=6)
        fetch_ledger_btn.pack(side="right")

        # Column headers
        col_hdr = ctk.CTkFrame(dialog, fg_color=COLORS["table_header"], corner_radius=6, height=32)
        col_hdr.pack(fill="x", padx=16, pady=(0, 2)); col_hdr.pack_propagate(False)
        ctk.CTkLabel(col_hdr, text="#", width=35, font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["tally_gold"], anchor="center").pack(side="left", padx=(8, 0))
        ctk.CTkLabel(col_hdr, text="Party Name", width=280, font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["tally_gold"], anchor="w").pack(side="left", padx=8)
        ctk.CTkLabel(col_hdr, text="Mapping Ledger", font=("Segoe UI", 10, "bold"),
                     text_color=COLORS["tally_gold"], anchor="w").pack(side="left", padx=8, fill="x", expand=True)

        # Scrollable list with editable entries + search buttons
        list_frame = ctk.CTkScrollableFrame(dialog, fg_color=COLORS["bg_card"], corner_radius=8,
                                              border_width=1, border_color=COLORS["border"])
        list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 6))

        entry_map = {}  # party_name -> CTkEntry

        def _open_picker_for_entry(entry_widget):
            if not _dialog_ledgers:
                messagebox.showinfo("Fetch First", "Click 'Fetch Ledgers from Tally' first to load the ledger list.",
                                    parent=dialog)
                return
            _pick_popup = tk.Toplevel(dialog)
            _pick_popup.title("Select Ledger")
            _pick_popup.geometry("360x340")
            _pick_popup.resizable(False, False)
            _pick_popup.transient(dialog)
            _pick_popup.grab_set()
            bg = COLORS["bg_card"][0] if isinstance(COLORS["bg_card"], tuple) else COLORS["bg_card"]
            fg = COLORS["text_primary"][0] if isinstance(COLORS["text_primary"], tuple) else COLORS["text_primary"]
            inp_bg = COLORS["bg_input"][0] if isinstance(COLORS["bg_input"], tuple) else COLORS["bg_input"]
            _pick_popup.configure(bg=bg)

            sf = tk.Frame(_pick_popup, bg=bg)
            sf.pack(fill="x", padx=8, pady=8)
            tk.Label(sf, text="Search:", bg=bg, fg=fg, font=("Segoe UI", 10)).pack(side="left")
            sv = tk.StringVar()
            se = tk.Entry(sf, textvariable=sv, font=("Segoe UI", 10), bg=inp_bg, fg=fg,
                          insertbackground=fg, relief="flat", bd=2)
            se.pack(side="left", padx=(6, 0), fill="x", expand=True)

            lf = tk.Frame(_pick_popup, bg=bg)
            lf.pack(fill="both", expand=True, padx=8, pady=(0, 4))
            sb = tk.Scrollbar(lf)
            sb.pack(side="right", fill="y")
            lb = tk.Listbox(lf, yscrollcommand=sb.set, font=("Segoe UI", 10), selectmode="single",
                            bg=inp_bg, fg=fg, relief="flat", bd=0,
                            selectbackground=COLORS["accent"][0] if isinstance(COLORS["accent"], tuple) else COLORS["accent"],
                            selectforeground="#FFFFFF")
            lb.pack(fill="both", expand=True)
            sb.config(command=lb.yview)

            def _filt(*_):
                q = sv.get().strip().lower()
                lb.delete(0, "end")
                for name in _dialog_ledgers:
                    if not q or q in name.lower():
                        lb.insert("end", name)
            sv.trace_add("write", _filt)
            _filt()

            def _sel(*_):
                sel = lb.curselection()
                if sel:
                    chosen = lb.get(sel[0])
                    entry_widget.delete(0, "end")
                    entry_widget.insert(0, chosen)
                    _pick_popup.destroy()

            br = tk.Frame(_pick_popup, bg=bg)
            br.pack(fill="x", padx=8, pady=(0, 8))
            acc = COLORS["accent"][0] if isinstance(COLORS["accent"], tuple) else COLORS["accent"]
            acc_h = COLORS["accent_hover"][0] if isinstance(COLORS["accent_hover"], tuple) else COLORS["accent_hover"]
            tk.Button(br, text="Select", command=_sel, relief="flat",
                      font=("Segoe UI", 10, "bold"), bg=acc, fg="#FFFFFF",
                      activebackground=acc_h, padx=12, pady=4).pack(side="left", padx=(0, 6))
            tk.Button(br, text="Cancel", command=_pick_popup.destroy, relief="flat",
                      font=("Segoe UI", 10), bg=inp_bg, fg=fg, padx=12, pady=4).pack(side="left")
            lb.bind("<Double-Button-1>", _sel)
            lb.bind("<Return>", _sel)
            se.focus_set()

        for idx, party in enumerate(unmapped_parties):
            bg = COLORS["table_row_odd"] if idx % 2 == 0 else COLORS["table_row_even"]
            row = ctk.CTkFrame(list_frame, fg_color=bg, corner_radius=4, height=36)
            row.pack(fill="x", pady=1); row.pack_propagate(False)

            ctk.CTkLabel(row, text=f"{idx+1}.", width=35, font=("Consolas", 10),
                         text_color=COLORS["text_muted"], anchor="center").pack(side="left", padx=(8, 0))

            ctk.CTkLabel(row, text=party[:38], width=280, font=("Segoe UI", 10),
                         text_color=COLORS["text_primary"], anchor="w").pack(side="left", padx=8)

            pick_btn = ctk.CTkButton(row, text="🔍", width=32, height=26,
                                     fg_color=COLORS["bg_input"], hover_color=COLORS["bg_card_hover"],
                                     text_color=COLORS["accent"], font=("Segoe UI", 11),
                                     corner_radius=4)
            pick_btn.pack(side="right", padx=(0, 6))

            entry = ctk.CTkEntry(row, height=28, fg_color=COLORS["bg_input"],
                                  border_color=COLORS["border"], text_color=COLORS["text_primary"],
                                  font=("Segoe UI", 10), corner_radius=6,
                                  placeholder_text=default_ledger)
            entry.pack(side="left", fill="x", expand=True, padx=(4, 4))
            entry.insert(0, default_ledger)
            entry_map[party] = entry
            pick_btn.configure(command=lambda e=entry: _open_picker_for_entry(e))

        # Fetch ledger button logic
        def _do_fetch_dialog_ledgers():
            try:
                tally_url = self._get_tally_push_url()
            except (ValueError, AttributeError):
                messagebox.showerror("Settings", "Invalid Tally URL. Check host/port in the Push panel.", parent=dialog)
                return
            company = self._get_effective_push_company()
            cached = getattr(self, "tally_push_companies", [])
            if not company and len(cached) > 1:
                messagebox.showwarning(
                    "Select Company",
                    "Multiple companies are open in Tally.\n"
                    "Please select a target company from the company dropdown before fetching ledgers.",
                    parent=dialog)
                return
            fetch_ledger_btn.configure(state="disabled", text="Fetching...")
            fetch_status_lbl.configure(text="Fetching ledgers...", text_color=COLORS["warning"])

            def _worker():
                result = _fetch_tally_ledgers(tally_url, timeout=15, company_name=company)
                def _done():
                    fetch_ledger_btn.configure(state="normal", text="Fetch Ledgers from Tally")
                    if result.get("success"):
                        _dialog_ledgers.clear()
                        _dialog_ledgers.extend(result.get("ledgers") or [])
                        fetch_status_lbl.configure(
                            text=f"✓ {len(_dialog_ledgers)} ledgers loaded — click 🔍 to search",
                            text_color=COLORS["success"])
                    else:
                        fetch_status_lbl.configure(text="✗ Fetch failed — is Tally running?",
                                                   text_color=COLORS["error"])
                self.after(0, _done)
            threading.Thread(target=_worker, daemon=True).start()

        fetch_ledger_btn.configure(command=_do_fetch_dialog_ledgers)

        # Buttons
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent", height=50)
        btn_frame.pack(fill="x", padx=16, pady=(0, 14)); btn_frame.pack_propagate(False)

        def do_apply():
            """Apply typed mappings to engine and proceed with generation."""
            applied = 0
            for party, entry in entry_map.items():
                ledger = entry.get().strip()
                if not ledger:
                    ledger = default_ledger
                self.engine.party_ledger_map[party.upper()] = ledger
                applied += 1
            self.log_panel.log(f"Applied {applied} party mappings from dialog.", "map")
            count = len(self.engine.party_ledger_map)
            self.mapping_label.configure(
                text=f"{count} total party mappings active (incl. {applied} from dialog)",
                text_color=COLORS["success"])
            dialog.destroy()
            on_proceed()

        def do_retry():
            dialog.destroy()
            on_retry()

        def do_proceed_default():
            """Proceed with default Purchase Account for all unmapped (don't save mappings)."""
            dialog.destroy()
            on_proceed()

        ctk.CTkButton(btn_frame, text="✓ Apply & Generate", font=("Segoe UI", 12, "bold"), height=40,
                      fg_color=COLORS["success"], hover_color="#047857",
                      text_color="#FFFFFF", corner_radius=8,
                      command=do_apply).pack(side="left", fill="x", expand=True, padx=(0, 4))

        ctk.CTkButton(btn_frame, text="Upload Mapping Sheet", font=("Segoe UI", 11), height=40,
                      fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
                      text_color="#FFFFFF", corner_radius=8,
                      command=do_retry).pack(side="left", fill="x", expand=True, padx=4)

        ctk.CTkButton(btn_frame, text="Proceed to Purchase", font=("Segoe UI", 11), height=40,
                      fg_color=COLORS["warning"], hover_color="#B45309",
                      text_color="#FFFFFF", corner_radius=8,
                      command=do_proceed_default).pack(side="right", fill="x", expand=True, padx=(4, 0))

    def _show_roundoff_ledger_picker(self):
        """Searchable popup to pick the round off ledger from the cached list."""
        all_ledgers = self.roundoff_all_ledgers
        if not all_ledgers:
            return
        popup = tk.Toplevel(self)
        popup.title("Select Round Off Ledger")
        popup.geometry("380x360")
        popup.resizable(False, False)
        popup.transient(self)
        popup.grab_set()

        bg = COLORS["bg_card"]
        fg = COLORS["text_primary"]
        inp_bg = COLORS["bg_input"]
        acc = COLORS["accent"]
        acc_h = COLORS["accent_hover"]
        if isinstance(bg, tuple):
            mode = ctk.get_appearance_mode().lower()
            idx = 1 if mode == "dark" else 0
            bg, fg, inp_bg, acc, acc_h = bg[idx], fg[idx], inp_bg[idx], acc[idx], acc_h[idx]

        popup.configure(bg=bg)

        sf = tk.Frame(popup, bg=bg)
        sf.pack(fill="x", padx=10, pady=10)
        tk.Label(sf, text="Search:", bg=bg, fg=fg, font=("Segoe UI", 10)).pack(side="left")
        sv = tk.StringVar()
        se = tk.Entry(sf, textvariable=sv, font=("Segoe UI", 10), bg=inp_bg, fg=fg,
                      insertbackground=fg, relief="flat", bd=2)
        se.pack(side="left", padx=(8, 0), fill="x", expand=True)

        lf = tk.Frame(popup, bg=bg)
        lf.pack(fill="both", expand=True, padx=10, pady=(0, 6))
        sb = tk.Scrollbar(lf)
        sb.pack(side="right", fill="y")
        lb = tk.Listbox(lf, yscrollcommand=sb.set, font=("Segoe UI", 10), selectmode="single",
                        bg=inp_bg, fg=fg, relief="flat", bd=0,
                        selectbackground=acc, selectforeground="#FFFFFF")
        lb.pack(fill="both", expand=True)
        sb.config(command=lb.yview)

        def _filt(*_):
            q = sv.get().strip().lower()
            lb.delete(0, "end")
            for name in all_ledgers:
                if not q or q in name.lower():
                    lb.insert("end", name)
        sv.trace_add("write", _filt)
        _filt()

        def _select(*_):
            sel = lb.curselection()
            if sel:
                chosen = lb.get(sel[0])
                self.roundoff_ledger_entry.configure(state="normal")
                self.roundoff_ledger_entry.delete(0, "end")
                self.roundoff_ledger_entry.insert(0, chosen)
                popup.destroy()

        br = tk.Frame(popup, bg=bg)
        br.pack(fill="x", padx=10, pady=(0, 10))
        tk.Button(br, text="Select", command=_select, relief="flat",
                  font=("Segoe UI", 10, "bold"), bg=acc, fg="#FFFFFF",
                  activebackground=acc_h, padx=14, pady=5).pack(side="left", padx=(0, 8))
        tk.Button(br, text="Cancel", command=popup.destroy, relief="flat",
                  font=("Segoe UI", 10), bg=inp_bg, fg=fg,
                  padx=14, pady=5).pack(side="left")

        lb.bind("<Double-Button-1>", _select)
        lb.bind("<Return>", _select)
        se.focus_set()

    def _retry_mapping_and_generate(self, excel, xml, records_to_generate=None):
        """Called when user clicks 'Add & Try Again' — opens file chooser for new mapping."""
        filepath = filedialog.askopenfilename(title="Select Updated Mapping Excel",
                                               filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")])
        if filepath:
            self.mapping_file = filepath
            count = self.engine.load_mapping(filepath)
            if count > 0:
                self.mapping_label.configure(
                    text=f"Loaded {count} unique party mappings from {Path(filepath).name}",
                    text_color=COLORS["success"])
                self.log_panel.log(f"New mapping loaded: {count} entries from {Path(filepath).name}", "map")
            else:
                self.mapping_label.configure(text="Failed to load mapping — check file format", text_color=COLORS["error"])
                self.log_panel.log("Mapping load failed — need PartyLedger + PurchaseLedger columns", "error")
                return
            active_records = records_to_generate if records_to_generate is not None else self.engine.records
            unmapped = self._find_unmapped_parties(active_records)
            if unmapped:
                self._show_unmapped_dialog(
                    unmapped,
                    on_proceed=lambda: self._do_generate(excel, xml, records_to_generate=active_records),
                    on_retry=lambda: self._retry_mapping_and_generate(excel, xml, records_to_generate=active_records),
                )
                return
            self._do_generate(excel, xml, records_to_generate=active_records)
        else:
            self.log_panel.log("Mapping upload cancelled — generation aborted.", "warning")

    # ─── STOCK ITEM PURCHASE POPUP ───

    def _show_stock_item_popup(self, stock_recs, xml_flag=True, regular_recs=None, excel_flag=True):
        """Popup to collect stock item details and push Purchase Item vouchers to Tally.
        Supports multiple items per invoice, per-item rate, fetched ledger dropdown."""
        if not _SPE_AVAILABLE or not _spe_gen_purchase_item_xml:
            messagebox.showerror(
                "Feature Unavailable",
                "sale_purchase_entry.py could not be loaded.\nCannot generate Purchase Item XML."
            )
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Purchase Item Entry — {len(stock_recs)} Stock Invoice(s)")
        dialog.geometry("1280x720")
        dialog.resizable(True, True)
        dialog.minsize(900, 500)
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color=COLORS["bg_dark"])

        dialog.update_idletasks()
        x = self.winfo_x() + max(0, (self.winfo_width() - 1280) // 2)
        y = self.winfo_y() + max(0, (self.winfo_height() - 720) // 2)
        dialog.geometry(f"+{x}+{y}")

        # ── Header ─────────────────────────────────────────────
        hdr = ctk.CTkFrame(dialog, fg_color=COLORS["accent"], corner_radius=10)
        hdr.pack(fill="x", padx=16, pady=(16, 4))
        ctk.CTkLabel(hdr, text=f"Purchase Item Entry — {len(stock_recs)} Invoice(s) with Stock Items",
                     font=("Segoe UI", 15, "bold"), text_color="#FFFFFF").pack(anchor="w", padx=14, pady=(10, 2))
        ctk.CTkLabel(hdr,
                     text="1. Click 'Fetch Items & Ledgers'  2. Select item, qty, rate, per, godown, ledger  "
                          "3. Use '+ Add Item' to add multiple items per invoice  4. Push to Tally",
                     font=("Segoe UI", 11), text_color="#E2E8F0",
                     wraplength=1200, justify="left").pack(anchor="w", padx=14, pady=(0, 10))

        # ── Fetch bar ──────────────────────────────────────────
        fetch_bar = ctk.CTkFrame(dialog, fg_color="transparent")
        fetch_bar.pack(fill="x", padx=16, pady=(4, 0))
        fetch_status_var = ctk.StringVar(value="Click 'Fetch' to load items & ledgers from Tally")
        fetch_btn = ctk.CTkButton(fetch_bar, text="🔄  Fetch Items & Ledgers",
                                  font=("Segoe UI", 12), height=34, width=230,
                                  fg_color=COLORS["accent"], hover_color=COLORS["accent_hover"],
                                  text_color="#FFFFFF", corner_radius=8)
        fetch_btn.pack(side="left")
        ctk.CTkLabel(fetch_bar, textvariable=fetch_status_var,
                     font=("Segoe UI", 11), text_color=COLORS["text_secondary"]).pack(side="left", padx=12)

        # ── Shared data refs (updated on fetch) ────────────────
        stock_items_ref  = [[]]   # stock_items_ref[0] = list of item names
        ledger_items_ref = [[]]   # ledger_items_ref[0] = list of ledger names

        # ── Scrollable invoice blocks ──────────────────────────
        scroll_frame = ctk.CTkScrollableFrame(dialog, fg_color=COLORS["bg_dark"])
        scroll_frame.pack(fill="both", expand=True, padx=16, pady=4)

        # invoice_rows[inv_idx] = list of row_data dicts
        invoice_rows = [[] for _ in stock_recs]

        COL_W = [190, 62, 84, 55, 110, 28]   # Item, Qty, Rate, Per, Godown, Del

        def _make_search_listbox(parent, ref_list, width=190):
            """Create a search entry + suggestion listbox pair. Returns (entry, listbox)."""
            entry = ctk.CTkEntry(
                parent, placeholder_text="Search...",
                height=26, width=width, font=("Segoe UI", 11),
                fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                text_color=COLORS["text_primary"],
            )
            entry.pack(fill="x", pady=(0, 1))

            lb = tk.Listbox(
                parent, height=0, exportselection=False, activestyle="none",
                relief="solid", borderwidth=1,
                bg=_theme_color("bg_input"), fg=_theme_color("text_primary"),
                selectbackground=_theme_color("accent"), selectforeground="#FFFFFF",
                font=("Segoe UI", 10),
            )
            lb.pack(fill="x")

            def _filter(e=None):
                q = entry.get().strip().lower()
                candidates = [n for n in ref_list[0] if q in n.lower()] if q else []
                lb.delete(0, "end")
                if q and candidates:
                    for it in candidates[:8]:
                        lb.insert("end", it)
                    lb.configure(height=min(len(candidates), 4))
                else:
                    lb.configure(height=0)

            def _select(e=None):
                sel = lb.curselection()
                if not sel:
                    return
                val = str(lb.get(sel[0]) or "").strip()
                if not val:
                    return
                entry.delete(0, "end")
                entry.insert(0, val)
                lb.configure(height=0)
                lb.delete(0, "end")

            entry.bind("<KeyRelease>", _filter)
            lb.bind("<<ListboxSelect>>", _select)
            lb.bind("<Double-Button-1>", _select)
            lb.bind("<Return>", _select)
            return entry, lb

        def _add_item_row(inv_idx, container, rec):
            rows_for_inv = invoice_rows[inv_idx]
            ri = len(rows_for_inv)
            bg = COLORS["bg_card"] if ri % 2 == 0 else COLORS["bg_dark"]
            rf = ctk.CTkFrame(container, fg_color=bg, corner_radius=3)
            rf.pack(fill="x", pady=1)
            rd = {"frame": rf}

            top = ctk.CTkFrame(rf, fg_color="transparent")
            top.pack(fill="x", padx=2, pady=(2, 0))

            # ── Item Name (search entry + listbox) ────────────
            item_col = ctk.CTkFrame(top, fg_color="transparent", width=COL_W[0])
            item_col.pack(side="left", padx=(2, 2))
            item_col.pack_propagate(False)
            item_entry, item_lb = _make_search_listbox(item_col, stock_items_ref, width=COL_W[0])
            rd["item_var"] = item_entry   # .get() returns text
            rd["item_lb"]  = item_lb

            # ── Qty ──────────────────────────────────────────
            qty_var = ctk.StringVar(value="1")
            ctk.CTkEntry(top, textvariable=qty_var, width=COL_W[1], height=26,
                         font=("Segoe UI", 11), justify="center").pack(side="left", padx=(2, 2))
            rd["qty_var"] = qty_var

            # ── Rate ₹ ───────────────────────────────────────
            taxable = float(rec.get("taxable_value") or 0.0)
            rate_var = ctk.StringVar(value=f"{taxable:.2f}")
            ctk.CTkEntry(top, textvariable=rate_var, width=COL_W[2], height=26,
                         font=("Segoe UI", 11), justify="right").pack(side="left", padx=(2, 2))
            rd["rate_var"] = rate_var

            # ── Per / UOM ─────────────────────────────────────
            per_var = ctk.StringVar(value="Nos")
            ctk.CTkEntry(top, textvariable=per_var, width=COL_W[3], height=26,
                         font=("Segoe UI", 11), justify="center").pack(side="left", padx=(2, 2))
            rd["per_var"] = per_var

            # ── Godown ───────────────────────────────────────
            godown_var = ctk.StringVar(value="Main Location")
            ctk.CTkEntry(top, textvariable=godown_var, width=COL_W[4], height=26,
                         font=("Segoe UI", 11)).pack(side="left", padx=(2, 2))
            rd["godown_var"] = godown_var

            # ── Purchase Ledger (search entry + listbox) ──────
            led_col = ctk.CTkFrame(top, fg_color="transparent")
            led_col.pack(side="left", padx=(2, 2), fill="x", expand=True)
            default_led = (rec.get("purchase_ledger") or "").strip() or self.purchase_ledger_entry.get() or "Purchase Account"
            led_entry, led_lb = _make_search_listbox(led_col, ledger_items_ref, width=185)
            led_entry.delete(0, "end")
            led_entry.insert(0, default_led)
            rd["ledger_var"] = led_entry  # .get() returns text
            rd["led_lb"]     = led_lb

            # ── Delete row button ─────────────────────────────
            def _del(r=rd, inv=inv_idx):
                if len(invoice_rows[inv]) > 1:
                    invoice_rows[inv].remove(r)
                    r["frame"].destroy()
            ctk.CTkButton(top, text="✕", width=COL_W[5], height=26,
                          font=("Segoe UI", 10), fg_color=("#DC2626", "#991B1B"),
                          hover_color="#7F1D1D", text_color="#FFFFFF",
                          corner_radius=4, command=_del).pack(side="left", padx=(2, 4))

            rows_for_inv.append(rd)

        for inv_idx, rec in enumerate(stock_recs):
            party    = str(rec.get("party_name") or rec.get("trade_name") or "").strip()
            inv_no   = str(rec.get("invoice_no") or "").strip()
            inv_date = str(rec.get("invoice_date") or "").strip()
            gstin    = str(rec.get("gstin") or "").strip()
            taxable  = float(rec.get("taxable_value") or 0.0)
            cgst_v   = abs(float(rec.get("cgst") or 0.0))
            sgst_v   = abs(float(rec.get("sgst") or 0.0))
            igst_v   = abs(float(rec.get("igst") or 0.0))
            abs_tax  = abs(taxable)
            gst_pct  = round(((cgst_v + sgst_v + igst_v) / abs_tax) * 100, 0) if abs_tax > 0 else 0.0

            # Invoice card
            inv_card = ctk.CTkFrame(scroll_frame, fg_color=COLORS["bg_card"], corner_radius=8)
            inv_card.pack(fill="x", pady=(4, 0))

            # Invoice header strip
            inv_hdr_strip = ctk.CTkFrame(inv_card, fg_color=COLORS["accent"], corner_radius=6)
            inv_hdr_strip.pack(fill="x", padx=6, pady=(6, 2))
            hdr_left = ctk.CTkFrame(inv_hdr_strip, fg_color="transparent")
            hdr_left.pack(side="left", fill="x", expand=True, padx=8, pady=4)
            ctk.CTkLabel(hdr_left,
                         text=f"{party[:36]}  |  Invoice: {inv_no[:24]}  |  Taxable: ₹{taxable:,.2f}  |  GST: {gst_pct:.0f}%",
                         font=("Segoe UI", 11, "bold"), text_color="#FFFFFF",
                         anchor="w").pack(anchor="w")
            ctk.CTkLabel(hdr_left,
                         text=f"GSTIN: {gstin or '—'}  |  Date: {inv_date or '—'}",
                         font=("Segoe UI", 10), text_color="#CBD5E1",
                         anchor="w").pack(anchor="w")

            # ── Sub-column headers ─────────────────────────────
            sub_hdr = ctk.CTkFrame(inv_card, fg_color="transparent")
            sub_hdr.pack(fill="x", padx=6)
            for lbl, w in zip(
                ["Item Name", "Qty", "Rate ₹", "Per", "Godown", "Purchase Ledger", ""],
                [COL_W[0], COL_W[1], COL_W[2], COL_W[3], COL_W[4], 185, COL_W[5]],
            ):
                ctk.CTkLabel(sub_hdr, text=lbl, font=("Segoe UI", 10, "bold"),
                             text_color=COLORS["text_secondary"], width=w, anchor="w",
                             ).pack(side="left", padx=(4, 2), pady=2)

            # Item rows container
            item_container = ctk.CTkFrame(inv_card, fg_color="transparent")
            item_container.pack(fill="x", padx=6, pady=(0, 2))

            # First default item row
            _add_item_row(inv_idx, item_container, rec)

            # "+ Add Item" button
            def _make_add(ii, cont, r):
                return lambda: _add_item_row(ii, cont, r)
            ctk.CTkButton(inv_card, text="+ Add Item", font=("Segoe UI", 11), height=26, width=110,
                          fg_color=COLORS["bg_card_hover"], hover_color=COLORS["bg_dark"],
                          text_color=COLORS["accent"], corner_radius=4,
                          command=_make_add(inv_idx, item_container, rec),
                          ).pack(anchor="w", padx=12, pady=(0, 6))

        # ── Fetch worker (items + ledgers together) ────────────
        def do_fetch():
            fetch_btn.configure(state="disabled", text="Fetching...")
            fetch_status_var.set("Contacting Tally...")

            def _worker():
                try:
                    t_url = self._get_tally_push_url()
                    t_co  = self._get_selected_tally_push_company()
                except ValueError as exc:
                    self.after(0, lambda e=str(exc): _done(None, None, e))
                    return
                items_res   = _fetch_stock_items_from_tally(t_url, timeout=20, company_name=t_co)
                ledgers_res = _fetch_tally_ledgers(t_url, timeout=20, company_name=t_co)
                self.after(0, lambda: _done(items_res, ledgers_res, None))

            def _done(ir, lr, err):
                fetch_btn.configure(state="normal", text="🔄  Fetch Items & Ledgers")
                if err:
                    fetch_status_var.set(f"✗ {err[:80]}")
                    return
                ni = nl = 0
                if ir and ir.get("success"):
                    stock_items_ref[0] = ir.get("items", [])
                    ni = len(stock_items_ref[0])
                if lr and lr.get("success"):
                    ledger_items_ref[0] = lr.get("ledgers", [])
                    nl = len(ledger_items_ref[0])
                fetch_status_var.set(f"✓ {ni} items, {nl} ledgers loaded — start typing to search")

            threading.Thread(target=_worker, daemon=True).start()

        fetch_btn.configure(command=do_fetch)

        # ── Bottom bar ─────────────────────────────────────────
        btn_bar = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_bar.pack(fill="x", padx=16, pady=(4, 16))
        push_status_var = ctk.StringVar(value="")
        ctk.CTkLabel(btn_bar, textvariable=push_status_var,
                     font=("Segoe UI", 11), text_color=COLORS["text_secondary"]).pack(side="left", padx=4)

        def do_push():
            rows = []
            for inv_idx, rec in enumerate(stock_recs):
                taxable_total = float(rec.get("taxable_value") or 0.0)
                cgst_v   = abs(float(rec.get("cgst") or 0.0))
                sgst_v   = abs(float(rec.get("sgst") or 0.0))
                igst_v   = abs(float(rec.get("igst") or 0.0))
                abs_tax  = abs(taxable_total)
                rate_pct = round(((cgst_v + sgst_v + igst_v) / abs_tax) * 100, 0) if abs_tax > 0 else 0.0
                if igst_v > 0:
                    cgst_rate, sgst_rate, igst_rate = 0.0, 0.0, rate_pct
                else:
                    half = round(rate_pct / 2, 1)
                    cgst_rate, sgst_rate, igst_rate = half, half, 0.0

                # Party ledger should be the supplier/party name (not the purchase ledger mapping)
                default_party = (
                    str(rec.get("party_name") or rec.get("trade_name") or "").strip()
                    or "Suspense A/c"
                )

                # Validate and collect all items for this invoice first
                inv_items = []
                for row_j, rd in enumerate(invoice_rows[inv_idx]):
                    item_name = rd["item_var"].get().strip()
                    if not item_name or item_name.startswith("—"):
                        messagebox.showwarning("Item Missing",
                            f"Invoice '{rec.get('invoice_no','')}' row {row_j+1}: select or type a stock item name.",
                            parent=dialog)
                        return
                    try:
                        qty = float(rd["qty_var"].get().strip() or "0")
                        if qty <= 0:
                            raise ValueError
                    except ValueError:
                        messagebox.showwarning("Invalid Qty",
                            f"Invoice '{rec.get('invoice_no','')}' row {row_j+1}: qty must be > 0.",
                            parent=dialog)
                        return
                    try:
                        rate = float(rd["rate_var"].get().strip() or "0")
                        if rate < 0:
                            raise ValueError
                    except ValueError:
                        messagebox.showwarning("Invalid Rate",
                            f"Invoice '{rec.get('invoice_no','')}' row {row_j+1}: rate must be ≥ 0.",
                            parent=dialog)
                        return
                    purch_ledger = rd["ledger_var"].get().strip() or self.purchase_ledger_entry.get() or "Purchase Account"
                    inv_items.append({
                        "ItemName":      item_name,
                        "Quantity":      qty,
                        "Rate":          rate,
                        "Per":           rd["per_var"].get().strip() or "Nos",
                        "GodownName":    rd["godown_var"].get().strip() or "Main Location",
                        "PurchaseLedger": purch_ledger,
                    })

                # One row per invoice — items list keeps all items in ONE voucher
                rows.append({
                    "Date":                rec.get("invoice_date") or rec.get("supplier_invoice_date") or "",
                    "VoucherNumber":       rec.get("voucher_no") or "",
                    "PartyLedger":         default_party,
                    "GSTIN":               rec.get("gstin") or "",
                    "TaxableValue":        taxable_total,
                    "CGSTRate":            cgst_rate,
                    "CGSTLedger":          "CGST" if cgst_rate > 0 else "",
                    "SGSTRate":            sgst_rate,
                    "SGSTLedger":          "SGST" if sgst_rate > 0 else "",
                    "IGSTRate":            igst_rate,
                    "IGSTLedger":          "IGST" if igst_rate > 0 else "",
                    "Narration":           rec.get("narration") or "",
                    "SupplierInvoiceNo":   rec.get("invoice_no") or "",
                    "SupplierInvoiceDate": rec.get("invoice_date") or "",
                    "items":               inv_items,
                })

            if not rows:
                messagebox.showwarning("Nothing to Push", "No item rows found.", parent=dialog)
                return

            sample_parties = [
                str(r.get("PartyLedger") or "").strip()
                for r in rows[:3]
                if str(r.get("PartyLedger") or "").strip()
            ]
            if sample_parties:
                self.log_panel.log(
                    f"Purchase Item party ledgers (sample): {', '.join(sample_parties)}",
                    "info",
                )

            try:
                date_mode, custom_tally_date = self._get_tally_push_date_selection()
            except ValueError as exc:
                messagebox.showerror("Invalid Date Selection", str(exc), parent=dialog)
                return

            if date_mode == "custom":
                custom_label = self._format_tally_date_for_display(custom_tally_date)
                self.log_panel.log(f"Purchase Item date mode: Custom ({custom_label})", "info")
            elif date_mode == "excel":
                self.log_panel.log("Purchase Item date mode: Excel Date", "info")
            else:
                self.log_panel.log("Purchase Item date mode: Current Date", "info")
                today_label = datetime.date.today().strftime("%d/%m/%Y")
                self.log_panel.log(f"Purchase Item effective date: {today_label}", "info")

            sample_dates = [
                str(r.get("Date") or "").strip()
                for r in rows[:3]
                if str(r.get("Date") or "").strip()
            ]
            if sample_dates:
                self.log_panel.log(
                    f"Purchase Item source dates (sample): {', '.join(sample_dates)}",
                    "info",
                )

            push_status_var.set(f"Generating XML for {len(rows)} item voucher(s)...")
            push_btn.configure(state="disabled", text="Pushing...")
            cancel_btn.configure(state="disabled")

            def _push_worker():
                try:
                    # Apply UI GSTIN to engine so CMPGSTIN block is always populated
                    if hasattr(self, "company_gstin_entry"):
                        _ui_gstin = self.company_gstin_entry.get().strip().upper()
                        if _ui_gstin:
                            self.engine.company_gstin = _ui_gstin

                    t_url = self._get_tally_push_url()
                    t_co  = self._get_selected_tally_push_company() or self.engine.company_name or ""

                    # Fetch company GST registrations for the CMPGSTIN/GSTREGISTRATION block
                    gst_regs = []
                    if _spe_fetch_gst_regs:
                        try:
                            res = _spe_fetch_gst_regs(t_url, t_co)
                            if res.get("success"):
                                gst_regs = res.get("registrations", [])
                        except Exception:
                            pass
                    # Persist fetched GSTIN into engine so _do_generate (for regular
                    # vouchers called after this push) also includes GSTREGISTRATION.
                    if gst_regs and not self.engine.company_gstin:
                        _fetched_co_gstin = str(gst_regs[0].get("gstin") or "").strip().upper()
                        if _fetched_co_gstin:
                            self.engine.company_gstin = _fetched_co_gstin
                            def _update_gstin_ui(g=_fetched_co_gstin):
                                if hasattr(self, "company_gstin_entry") and not self.company_gstin_entry.get():
                                    self.company_gstin_entry.delete(0, "end")
                                    self.company_gstin_entry.insert(0, g)
                            self.after(0, _update_gstin_ui)
                    # Fallback: build registration from engine.company_gstin so XML is never empty
                    if not gst_regs and self.engine.company_gstin:
                        co_gstin = self.engine.company_gstin.strip().upper()
                        co_state = _state_name_from_gstin(co_gstin)
                        co_name  = (f"{co_state} Registration" if co_state
                                    else (self.engine.company_name or co_gstin))
                        gst_regs = [{"name": co_name, "gstin": co_gstin, "state": co_state}]

                    xml_content = _spe_gen_purchase_item_xml(
                        rows=rows,
                        company=t_co,
                        date_mode=date_mode,
                        custom_tally_date=custom_tally_date,
                        company_gst_registrations=gst_regs,
                        company_gstin=self.engine.company_gstin or "",
                        voucher_type="Purchase",
                    )
                    result = _post_xml_to_tally(t_url, xml_content, timeout=60)
                except Exception as exc:
                    result = {"success": False, "error": str(exc)}
                self.after(0, lambda: _push_done(result, len(rows)))

            def _push_done(result, pushed_count):
                push_btn.configure(state="normal", text="🚀  Push to Tally")
                cancel_btn.configure(state="normal")
                created = int(result.get("created") or 0)
                errors  = int(result.get("errors") or 0)
                # Success: Tally responded OK with no errors
                is_success = result.get("success") and errors == 0
                if is_success:
                    n = created if created > 0 else pushed_count
                    push_status_var.set(f"✓ {n} voucher(s) pushed!")
                    self.log_panel.log(f"Purchase Item: {n} voucher(s) pushed to Tally.", "success")
                    # Track stock item count so final success dialog shows combined total
                    self._stock_items_pushed_count = n
                    dialog.destroy()
                    # Now generate + push regular accounting entries (B2B Purchase vouchers)
                    # stock_recs are included in extra_excel_records so they appear in
                    # tally-sheet.xlsx even though their XML was already pushed separately.
                    if regular_recs:
                        total_expected = n + len(regular_recs)
                        self.log_panel.log(
                            f"Generating {len(regular_recs)} regular voucher(s) "
                            f"(total including stock items: {total_expected})...", "process"
                        )
                        self.after(300, lambda: self._do_generate(
                            excel_flag, xml_flag,
                            records_to_generate=regular_recs,
                            extra_excel_records=stock_recs,
                        ))
                    elif stock_recs:
                        # Only stock records: write Excel if requested, otherwise finish cleanly.
                        if excel_flag:
                            self.after(300, lambda: self._do_generate(
                                excel_flag, False,
                                records_to_generate=[],
                                extra_excel_records=stock_recs,
                            ))
                        else:
                            self.log_panel.log(
                                "Purchase Item push complete; no output files selected.",
                                "success",
                            )
                            messagebox.showinfo(
                                "Success",
                                f"{n} Purchase Item voucher(s) pushed to Tally successfully!",
                            )
                    else:
                        messagebox.showinfo("Success",
                            f"{n} Purchase Item voucher(s) pushed to Tally successfully!")
                else:
                    err = str(result.get("error") or f"Created={created}, Errors={errors}" or "Unknown")
                    push_status_var.set(f"✗ Push failed (created={created}, errors={errors})")
                    self.log_panel.log(f"Stock item push: {err}", "error")
                    messagebox.showerror("Push Failed",
                        f"Tally response (created={created}, errors={errors}):\n\n{err}",
                        parent=dialog)

            threading.Thread(target=_push_worker, daemon=True).start()

        push_btn = ctk.CTkButton(
            btn_bar, text="🚀  Push to Tally",
            font=("Segoe UI", 13, "bold"), height=42, width=180,
            fg_color=COLORS["success"], hover_color="#047857",
            text_color="#FFFFFF", corner_radius=8, command=do_push)
        push_btn.pack(side="right", padx=(4, 0))

        cancel_btn = ctk.CTkButton(
            btn_bar, text="Cancel",
            font=("Segoe UI", 12), height=42, width=100,
            fg_color=COLORS["bg_card"], hover_color=COLORS["bg_card_hover"],
            text_color=COLORS["text_primary"], corner_radius=8, command=dialog.destroy)
        cancel_btn.pack(side="right", padx=(0, 4))

    # ─── RCM LEDGER MAPPING DIALOG ───

    def _show_rcm_ledger_mapping_dialog(self, rcm_count, on_proceed):
        """Show a popup for mapping RCM ledgers before generation."""
        dialog = ctk.CTkToplevel(self)
        dialog.title("RCM Ledger Mapping")
        dialog.geometry("660x560")
        dialog.minsize(560, 440)
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(fg_color=COLORS["bg_dark"])

        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 660) // 2
        y = self.winfo_y() + (self.winfo_height() - 560) // 2
        dialog.geometry(f"+{x}+{y}")

        # Header
        hdr = ctk.CTkFrame(dialog, fg_color=COLORS["warning_bg"], corner_radius=10)
        hdr.pack(fill="x", padx=16, pady=(16, 8))
        ctk.CTkLabel(hdr, text="RCM (Reverse Charge) Ledger Mapping",
                     font=("Segoe UI", 15, "bold"),
                     text_color=COLORS["warning"]).pack(anchor="w", padx=14, pady=(10, 2))
        ctk.CTkLabel(
            hdr,
            text=(f"Found {rcm_count} invoice(s) with Reverse Charge = Yes.\n"
                  "Map the ledgers below. A Journal voucher will be created for each RCM invoice."),
            font=("Segoe UI", 11),
            text_color=COLORS["text_secondary"],
            wraplength=620,
            justify="left",
        ).pack(anchor="w", padx=14, pady=(0, 10))

        # Scrollable form
        form_outer = ctk.CTkScrollableFrame(
            dialog, fg_color=COLORS["bg_card"], corner_radius=8,
            border_width=1, border_color=COLORS["border"],
        )
        form_outer.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        # Pre-fill from existing rcm_ledger_map
        existing = self.engine.rcm_ledger_map or {}

        fields = [
            ("expense",       "Expense / Purchase Ledger (Dr)",      "Dr",  "e.g. Professional Fees, Purchase Account"),
            ("cgst_inward",   "CGST Input RCM Ledger (Dr)",          "Dr",  "e.g. CGST Input RCM"),
            ("sgst_inward",   "SGST Input RCM Ledger (Dr)",          "Dr",  "e.g. SGST Input RCM"),
            ("igst_inward",   "IGST Input RCM Ledger (Dr)  [IGST]",  "Dr",  "e.g. IGST Input RCM  (only for interstate)"),
            ("cgst_outward",  "CGST Output RCM Ledger (Cr)",         "Cr",  "e.g. CGST Output RCM"),
            ("sgst_outward",  "SGST Output RCM Ledger (Cr)",         "Cr",  "e.g. SGST Output RCM"),
            ("igst_outward",  "IGST Output RCM Ledger (Cr) [IGST]",  "Cr",  "e.g. IGST Output RCM  (only for interstate)"),
        ]

        entries = {}
        for key, label, side, placeholder in fields:
            row = ctk.CTkFrame(form_outer, fg_color="transparent")
            row.pack(fill="x", pady=5, padx=8)

            side_badge = ctk.CTkLabel(
                row, text=side, width=36, font=("Segoe UI", 10, "bold"),
                fg_color=COLORS["success"] if side == "Dr" else COLORS["error"],
                text_color="#FFFFFF", corner_radius=6,
            )
            side_badge.pack(side="left", padx=(0, 8))

            ctk.CTkLabel(row, text=label, font=("Segoe UI", 11),
                         text_color=COLORS["text_primary"], width=230, anchor="w").pack(side="left", padx=(0, 8))

            ent = ctk.CTkEntry(row, placeholder_text=placeholder,
                               fg_color=COLORS["bg_input"], border_color=COLORS["border"],
                               font=("Segoe UI", 11))
            ent.pack(side="left", fill="x", expand=True)
            if existing.get(key):
                ent.insert(0, existing[key])
            entries[key] = ent

        # Buttons
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 14))

        def do_apply():
            mapped = {}
            for key, ent in entries.items():
                val = ent.get().strip()
                if val:
                    mapped[key] = val
            # Expense ledger is mandatory
            if not mapped.get("expense"):
                messagebox.showwarning(
                    "Missing Ledger",
                    "Please enter the Expense / Purchase Ledger name (Dr side).",
                    parent=dialog,
                )
                return
            self.engine.rcm_ledger_map = mapped
            self.log_panel.log(
                f"RCM ledger mapping saved: {', '.join(f'{k}={v}' for k, v in mapped.items())}",
                "map",
            )
            dialog.destroy()
            on_proceed()

        def do_skip():
            """Proceed without RCM mapping — RCM invoices will use defaults."""
            self.engine.rcm_ledger_map = {"_skipped": True}
            self.log_panel.log("RCM mapping skipped — using default ledger names.", "warning")
            dialog.destroy()
            on_proceed()

        ctk.CTkButton(
            btn_frame, text="Apply & Generate", font=("Segoe UI", 12, "bold"), height=40,
            fg_color=COLORS["success"], hover_color="#047857",
            text_color="#FFFFFF", corner_radius=8,
            command=do_apply,
        ).pack(side="left", fill="x", expand=True, padx=(0, 6))

        ctk.CTkButton(
            btn_frame, text="Skip (use defaults)", font=("Segoe UI", 11), height=40,
            fg_color=COLORS["warning"], hover_color="#B45309",
            text_color="#FFFFFF", corner_radius=8,
            command=do_skip,
        ).pack(side="right", fill="x", expand=True, padx=(6, 0))

    # ─── GENERATION ───

    def _generate_output(self, excel=True, xml=True):
        if not self.engine.records:
            messagebox.showwarning("No Data", "Please upload and parse a file first.")
            return

        # Apply company GSTIN from UI into engine before any generation/routing
        if hasattr(self, "company_gstin_entry"):
            ui_gstin = self.company_gstin_entry.get().strip().upper()
            if ui_gstin:
                self.engine.company_gstin = ui_gstin

        records_to_generate = list(self.engine.records)

        # ─── Apply ITC template mappings onto records before generation ───
        # When the user has uploaded an ITC template, itc_map contains per-invoice
        # Yes/No flags for ITC claimed and stock items. Apply them back to records
        # so that the XML builder routes each invoice correctly (purchase vs journal).
        if self.engine.itc_map:
            party_has_stock = set()
            # Pass 1: Identify all parties that have at least one stock item mapped in the B2B template
            for rec in records_to_generate:
                inv_key = rec.get("invoice_no", "").strip().upper()
                orig_key = rec.get("orig_invoice_no", "").strip().upper()
                
                itc_info = self.engine.itc_map.get(inv_key)
                if not itc_info and orig_key:
                    itc_info = self.engine.itc_map.get(orig_key)
                    
                if itc_info and itc_info.get("has_stock", "").strip().upper() in ("YES", "Y"):
                    party_has_stock.add(rec.get("trade_name", "").strip().upper())
                    
            # Pass 2: Apply the ITC eligibility and Party-wide stock item flag
            for rec in records_to_generate:
                inv_key = rec.get("invoice_no", "").strip().upper()
                orig_key = rec.get("orig_invoice_no", "").strip().upper()
                trade_key = rec.get("trade_name", "").strip().upper()
                
                itc_info = self.engine.itc_map.get(inv_key)
                if not itc_info and orig_key:
                    itc_info = self.engine.itc_map.get(orig_key)
                
                if itc_info:
                    itc_val = itc_info.get("itc_claimed", "").strip().upper()
                    # "YES" → eligible (Purchase voucher), "NO" → ineligible (Journal)
                    if itc_val in ("YES", "Y"):
                        rec["itc_avail"] = "Yes"
                    elif itc_val in ("NO", "N"):
                        rec["itc_avail"] = "Ineligible"
                
                # If this party has stock items, mark ALL their invoices (B2BA, IMPG, etc.) as stock items.
                # The stock popup will filter out CDNR/notes automatically.
                if trade_key in party_has_stock:
                    rec["has_stock_item"] = True

        # Apply per-party TDS/ledger from itc_map (overrides engine defaults
        # if the template sets a ledger directly per party row)
        # (party_ledger_map / tds maps are already set by _upload_itc_template)

        # In tally sheet mode, only XML generation
        if self.current_mode == "tally":
            excel = False
            xml = True

        # ─── GSTR-2B: mandatory mapping check ───
        if self.current_mode == "gstr2b":
            valid_records, invalid_issues = self.engine.validate_tax_configuration(records_to_generate)
            if invalid_issues:
                output_dir = self._resolve_output_dir()
                report_path = self._save_tax_validation_report(output_dir, invalid_issues)
                self.log_panel.log(
                    f"Tax validation mismatch in {len(invalid_issues)} records. Report: {report_path}",
                    "warning",
                )

                proceed = messagebox.askyesno(
                    "Tax Mismatch Found",
                    f"Found {len(invalid_issues)} record(s) with invalid tax slab/structure.\n\n"
                    "Allowed slabs: 0, 5, 12, 18, 28, 40\n"
                    "Rules: IGST cannot coexist with CGST/SGST.\n\n"
                    f"Mismatch report saved at:\n{report_path}\n\n"
                    f"Yes: Continue with {len(valid_records)} valid record(s) only.\n"
                    f"No: Continue with all {len(self.engine.records)} record(s) (old behavior).",
                )
                if proceed:
                    if valid_records:
                        records_to_generate = valid_records
                        self.log_panel.log(
                            f"Proceeding with {len(records_to_generate)} valid record(s) only.",
                            "warning",
                        )
                    else:
                        records_to_generate = list(self.engine.records)
                        self.log_panel.log(
                            "No valid records found, so proceeding with all records (old behavior).",
                            "warning",
                        )
                else:
                    records_to_generate = list(self.engine.records)
                    self.log_panel.log("User chose No: proceeding with all records (old behavior).", "warning")

            if not self.engine.party_ledger_map:
                # Check whether all records have purchase_ledger embedded
                # (happens when records came from the uploaded ITC template)
                records_with_embedded_ledger = sum(
                    1 for r in records_to_generate if r.get("purchase_ledger", "").strip()
                )
                if records_with_embedded_ledger < len(records_to_generate):
                    # Some records still need a mapping sheet
                    resp = messagebox.askyesno(
                        "Mapping Sheet Required",
                        "No mapping sheet has been loaded!\n\n"
                        "In GSTR-2B mode, a party \u2192 ledger mapping sheet is required \n"
                        "to assign correct purchase ledgers.\n\n"
                        "Click 'Yes' to load a mapping sheet now, or\n"
                        "'No' to proceed with all entries mapped to default Purchase Account.")
                    if resp:
                        self._browse_mapping()
                        if not self.engine.party_ledger_map:
                            return  # Still no mapping → abort
                    # If user clicked No, continue with defaults — but still check unmapped
                # else: all records have embedded ledgers — proceed silently

            # Check for unmapped parties
            unmapped = self._find_unmapped_parties(records_to_generate)
            if unmapped:
                self.log_panel.log(f"Found {len(unmapped)} unmapped party names.", "warning")
                # Capture excel/xml flags for retry callback
                e_flag, x_flag = excel, xml
                self._show_unmapped_dialog(
                    unmapped,
                    on_proceed=lambda: self._do_generate(e_flag, x_flag, records_to_generate=records_to_generate),
                    on_retry=lambda: self._retry_mapping_and_generate(e_flag, x_flag, records_to_generate=records_to_generate)
                )
                return  # Wait for dialog choice

        # Mark RCM as handled so _build_voucher_xml uses default ledger names
        rcm_records = [r for r in records_to_generate if str(r.get("reverse_charge") or "").strip().upper() == "YES"]
        if rcm_records and not self.engine.rcm_ledger_map:
            self.engine.rcm_ledger_map = {"_skipped": True}
            self.log_panel.log(f"Found {len(rcm_records)} RCM invoice(s). Generating RCM Journal entries with default ledger names.", "info")

        # Separate stock-item records — show popup; popup will push ALL entries together.
        # CDNR (credit/debit notes) and records with credit-note invoice type are excluded
        # from the stock item popup even when has_stock_item=True — they are generated as
        # regular Debit Note / Credit Note vouchers, not Purchase Item invoices.
        def _is_note_or_cdnr(r):
            st  = str(r.get("sheet_type") or "").upper()
            it  = str(r.get("invoice_type") or "").lower()
            return st == "CDNR" or "credit note" in it or "debit note" in it

        if self.current_mode == "gstr2b" and _SPE_AVAILABLE:
            stock_recs   = [r for r in records_to_generate if r.get("has_stock_item") and not _is_note_or_cdnr(r)]
            regular_recs = [r for r in records_to_generate if not r.get("has_stock_item") or _is_note_or_cdnr(r)]
            if stock_recs:
                # Pass regular_recs so popup can generate+push them AFTER item data is collected
                self._show_stock_item_popup(stock_recs, xml_flag=xml,
                                            regular_recs=regular_recs, excel_flag=excel)
                return

        self._do_generate(excel, xml, records_to_generate=records_to_generate)

    def _do_generate(self, excel=True, xml=True, records_to_generate=None, extra_excel_records=None):
        """Actually perform the generation (called after mapping validation passes)."""
        source_records = records_to_generate if records_to_generate is not None else self.engine.records
        # Excel gets regular + stock-item records; XML stays accounting-only
        excel_source_records = list(source_records) + list(extra_excel_records or [])
        output_dir = self.output_entry.get() or self.output_dir
        if not output_dir:
            src = self.source_file if self.current_mode == "gstr2b" else self.tally_sheet_file
            if src:
                output_dir = str(Path(src).parent)
            else:
                # Fallback: use the user's Downloads folder
                output_dir = str(Path.home() / "Downloads")
                self.log_panel.log(f"No output folder set — saving to {output_dir}", "info")

        company_name = self.company_entry.get() or self.engine.trade_name or self.engine.company_name or "My Company"
        purchase_ledger = self.purchase_ledger_entry.get() or "Purchase Account"
        narration = self.narration_entry.get() or "Being purchase from {party} vide Inv {inv} dt {date}"
        round_off_ledger = (
            self.roundoff_ledger_entry.get().strip()
            if (hasattr(self, "roundoff_enabled_var") and self.roundoff_enabled_var.get()
                and hasattr(self, "roundoff_ledger_entry"))
            else ""
        )

        # Apply company GSTIN from UI field so GSTREGISTRATION/CMPGSTIN block is
        # always present in the generated XML — preventing "Uncertain" in GSTR-3B.
        ui_gstin = self.company_gstin_entry.get().strip().upper() if hasattr(self, "company_gstin_entry") else ""
        if ui_gstin:
            self.engine.company_gstin = ui_gstin
        if xml and not self.engine.company_gstin:
            self.log_panel.log(
                "Warning: Company GSTIN not set — vouchers may appear Uncertain in GSTR-3B. "
                "Enter your GSTIN in the Settings panel.",
                "warning",
            )

        excel_path = _get_unique_path(output_dir, "tally-sheet", ".xlsx")
        xml_path = _get_unique_path(output_dir, "tally-ready", ".xml")

        self.generate_btn.configure(state="disabled")
        self.excel_only_btn.configure(state="disabled")
        self.xml_only_btn.configure(state="disabled")
        if hasattr(self, "inline_push_post_btn"):
            self.inline_push_post_btn.configure(state="disabled")
        if hasattr(self, "inline_push_test_btn"):
            self.inline_push_test_btn.configure(state="disabled")
        self.progress_bar.set(0)
        self.progress_label.configure(text="Generating...", text_color=COLORS["warning"])
        self.status_label.configure(text="Generating", text_color=COLORS["warning"])
        self.log_panel.log("Starting output generation...", "process")

        def run():
            results = {"excel": False, "xml": False}
            def progress_cb(pct, msg):
                self.after(0, lambda: self.progress_bar.set(pct * 0.5 if excel and xml else pct))
                self.after(0, lambda: self.progress_label.configure(text=msg))
            if excel:
                en = os.path.basename(excel_path)
                self.after(0, lambda n=en: self.log_panel.log(f"Generating {n}...", "process"))
                results["excel"] = self.engine.generate_tally_sheet(
                    excel_path,
                    purchase_ledger,
                    narration,
                    progress_cb,
                    records=excel_source_records,
                )
            if xml:
                def xml_progress(pct, msg):
                    offset = 0.5 if excel else 0
                    self.after(0, lambda: self.progress_bar.set(offset + pct * 0.5))
                    self.after(0, lambda: self.progress_label.configure(text=msg))
                xn = os.path.basename(xml_path)
                self.after(0, lambda n=xn: self.log_panel.log(f"Generating {n}...", "process"))
                results["xml"] = self.engine.generate_tally_xml(
                    xml_path,
                    company_name,
                    purchase_ledger,
                    narration,
                    xml_progress,
                    records=source_records,
                    round_off_ledger=round_off_ledger,
                )
            self.after(0, lambda: self._on_generate_complete(results, excel, xml, excel_path, xml_path, len(excel_source_records)))
        threading.Thread(target=run, daemon=True).start()

    def _on_generate_complete(self, results, do_excel, do_xml, excel_path, xml_path, generated_count):
        self.generate_btn.configure(state="normal")
        self.excel_only_btn.configure(state="normal")
        self.xml_only_btn.configure(state="normal")
        if hasattr(self, "inline_push_post_btn"):
            self.inline_push_post_btn.configure(state="normal")
        if hasattr(self, "inline_push_test_btn"):
            self.inline_push_test_btn.configure(state="normal")
        self.progress_bar.set(1.0)
        success_msgs = []
        if do_excel and results["excel"]:
            success_msgs.append("tally-sheet.xlsx"); self.log_panel.log(f"Excel saved: {excel_path}", "success")
        if do_xml and results["xml"]:
            success_msgs.append("tally-ready.xml"); self.log_panel.log(f"XML saved: {xml_path}", "success")
        if success_msgs:
            self.progress_label.configure(text="Generation complete!", text_color=COLORS["success"])
            self.status_label.configure(text="Complete", text_color=COLORS["success"])
            self.log_panel.log(f"Files generated in: {os.path.dirname(xml_path)}", "success")
            consolidation_msg = getattr(self.engine, "_last_consolidation_msg", "")
            if consolidation_msg:
                self.log_panel.log(consolidation_msg, "info")
            merge_log = getattr(self.engine, "_consolidation_merge_log", [])
            for entry in merge_log[:10]:
                self.log_panel.log(entry, "info")
            if len(merge_log) > 10:
                self.log_panel.log(f"  ... and {len(merge_log) - 10} more merged rows", "info")
            self.engine._consolidation_merge_log = []
            # Auto-push if requested via the new button
            if self._pending_push and do_xml and results["xml"] and os.path.isfile(xml_path):
                self._pending_push = False
                self.log_panel.log("Auto-push to Tally triggered...", "process")
                self.after(400, lambda p=xml_path: self._auto_push_generated_xml(p))
                return  # Skip the messagebox — push flow will show its own result
            msg = ["Files generated successfully!\n"]
            if do_excel and results["excel"]: msg.append(f"Excel: {excel_path}")
            if do_xml and results["xml"]: msg.append(f"XML: {xml_path}")
            msg.append(f"\nTotal records generated: {generated_count}")
            messagebox.showinfo("Success!", "\n".join(msg))
        else:
            self.progress_label.configure(text="Generation failed!", text_color=COLORS["error"])
            self.status_label.configure(text="Error", text_color=COLORS["error"])
            for err in self.engine.errors: self.log_panel.log(err, "error")
            messagebox.showerror("Error", "Failed to generate output files.\nCheck activity log.")


    # ─── INLINE PUSH TO TALLY METHODS ────────────────────────────────────────

    def _inline_push_get_url(self):
        host = (self.inline_push_host_entry.get() or "localhost").strip()
        port_str = (self.inline_push_port_entry.get() or "9000").strip()
        try:
            port = int(port_str)
        except ValueError:
            raise ValueError(f"Invalid port number: '{port_str}'")
        return f"http://{host}:{port}"

    def _inline_push_get_timeout(self):
        t = (self.inline_push_timeout_entry.get() or "30").strip()
        try:
            return int(t)
        except ValueError:
            raise ValueError(f"Invalid timeout value: '{t}'")

    def _inline_push_get_company(self):
        if not hasattr(self, "inline_push_company_cb"):
            return ""
        selected = _normalize_company_name(self.inline_push_company_cb.get() or "")
        if not selected or _company_key(selected) == _company_key(self.tally_push_company_placeholder):
            return ""
        return selected

    def _inline_push_set_company_dropdown(self, companies, keep_selection=True):
        current = ""
        if keep_selection and hasattr(self, "inline_push_company_cb"):
            current = _normalize_company_name(self.inline_push_company_cb.get() or "")
        cleaned = []
        seen = set()
        for name in companies or []:
            txt = _normalize_company_name(name)
            if not _is_valid_company_name(txt):
                continue
            key = _company_key(txt)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(txt)
        cleaned = sorted(cleaned, key=lambda x: _company_key(x))
        values = [self.tally_push_company_placeholder] + cleaned
        self.inline_push_company_cb.configure(values=values)
        if current and _company_key(current) in {_company_key(x) for x in cleaned}:
            self.inline_push_company_cb.set(current)
        else:
            self.inline_push_company_cb.set(self.tally_push_company_placeholder)
        self.inline_push_company_status.configure(
            text=f"Companies: {len(cleaned)} available", text_color=COLORS["text_muted"])

    def _inline_push_refresh_companies_thread(self):
        if self.tally_push_is_running:
            return
        try:
            tally_url = self._inline_push_get_url()
            timeout = self._inline_push_get_timeout()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return
        self.inline_push_company_refresh_btn.configure(state="disabled", text="Fetching...")
        self.inline_push_company_status.configure(text="Companies: Fetching...", text_color=COLORS["warning"])

        def _worker():
            result = _fetch_tally_companies(tally_url, timeout=min(timeout, 20))
            def done():
                self.inline_push_company_refresh_btn.configure(state="normal", text="Fetch")
                if result.get("success"):
                    companies = result.get("companies", [])
                    self._inline_push_set_company_dropdown(companies, keep_selection=True)
                    self.log_panel.log(f"Fetched {len(companies)} Tally companies.", "info")
                else:
                    err = str(result.get("error") or "Unknown error")
                    self.inline_push_company_status.configure(text="Companies: Fetch failed", text_color=COLORS["warning"])
                    self.log_panel.log(f"Could not fetch companies: {err}", "warning")
            self.after(0, done)
        threading.Thread(target=_worker, daemon=True).start()

    def _inline_push_test_connection(self):
        if self.tally_push_is_running:
            return
        try:
            tally_url = self._inline_push_get_url()
            timeout = self._inline_push_get_timeout()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return
        self.inline_push_test_btn.configure(state="disabled", text="Checking...")
        self.inline_push_conn_status.configure(text="Connection: Checking...", text_color=COLORS["warning"])

        def _worker():
            check_result = _check_tally_connection(tally_url, timeout=min(timeout, 10))
            ok = check_result.get("connected", False)
            def done():
                self.inline_push_test_btn.configure(state="normal", text="Test Connection")
                if ok:
                    self.inline_push_conn_status.configure(
                        text=f"Connection: Connected ({tally_url})", text_color=COLORS["success"])
                    self.log_panel.log(f"Tally connected at {tally_url}", "success")
                    self._inline_push_refresh_companies_thread()
                else:
                    err = str(check_result.get("error") or "Offline")
                    self.inline_push_conn_status.configure(text="Connection: Offline", text_color=COLORS["error"])
                    self.log_panel.log(f"Tally offline ({err})", "error")
            self.after(0, done)
        threading.Thread(target=_worker, daemon=True).start()

    def _generate_and_push_inline(self):
        """Generate XML then immediately push it to Tally using inline panel settings."""
        try:
            tally_url = self._inline_push_get_url()
            timeout = self._inline_push_get_timeout()
            date_mode, custom_tally_date = self._get_tally_push_date_selection()
        except ValueError as e:
            messagebox.showerror("Invalid Settings", str(e))
            return
        self._pending_push = True
        self._pending_push_url = tally_url
        self._pending_push_timeout = timeout
        self._pending_push_date_mode = date_mode
        self._pending_push_custom_date = custom_tally_date
        self._pending_push_company = self._inline_push_get_company()
        self.log_panel.log(f"Generate & Push: will post to {tally_url} after XML is ready.", "process")

        # ALWAYS fetch company GSTIN from Tally before generating so GSTREGISTRATION/CMPGSTIN
        # is present in every voucher (prevents "Uncertain" in GSTR-3B regardless of
        # whether the UI GSTIN field is filled or the GSTR-2B file auto-detected it).
        if _spe_fetch_gst_regs:
            t_co_for_fetch = self._pending_push_company or self.engine.company_name or ""
            self.log_panel.log("Fetching company GST registration from Tally...", "info")

            def _fetch_gstin_then_generate():
                fetched_gstin = ""
                fetched_reg_name = ""
                fetched_reg_state = ""
                try:
                    res = _spe_fetch_gst_regs(tally_url, t_co_for_fetch, timeout=10.0)
                    if res.get("success"):
                        regs = res.get("registrations", [])
                        if regs:
                            fetched_gstin = str(regs[0].get("gstin") or "").strip().upper()
                            fetched_reg_name = str(regs[0].get("name") or "").strip()
                            fetched_reg_state = str(regs[0].get("state") or "").strip()
                except Exception:
                    pass

                def on_done():
                    if fetched_gstin:
                        # Always override with authoritative values from Tally
                        self.engine.company_gstin = fetched_gstin
                        self.engine.company_registration_name = fetched_reg_name
                        self.engine.company_registration_state = fetched_reg_state
                        if hasattr(self, "company_gstin_entry"):
                            self.company_gstin_entry.delete(0, "end")
                            self.company_gstin_entry.insert(0, fetched_gstin)
                        reg_label = fetched_reg_name or fetched_gstin
                        self.log_panel.log(
                            f"Company GSTIN: {fetched_gstin}  |  Registration: {reg_label}",
                            "success",
                        )
                    else:
                        # Fall back to whatever is in the UI or engine
                        ui_gstin_now = self.company_gstin_entry.get().strip().upper() if hasattr(self, "company_gstin_entry") else ""
                        if ui_gstin_now:
                            self.engine.company_gstin = ui_gstin_now
                        if not self.engine.company_gstin:
                            self.log_panel.log(
                                "Could not fetch company GSTIN from Tally — "
                                "enter it manually in Settings to prevent Uncertain in GSTR-3B",
                                "warning",
                            )
                    self._generate_output(excel=False)

                self.after(0, on_done)

            threading.Thread(target=_fetch_gstin_then_generate, daemon=True).start()
        else:
            # spe module not available — use UI/engine GSTIN as-is
            self._generate_output(excel=False)

    def _auto_push_generated_xml(self, xml_path: str):
        """Auto-push the freshly generated XML to Tally."""
        if not os.path.isfile(xml_path):
            messagebox.showerror("Auto-Push Failed", f"Generated XML not found:\n{xml_path}")
            return
        tally_url = self._pending_push_url
        timeout = self._pending_push_timeout
        date_mode = self._pending_push_date_mode
        custom_tally_date = self._pending_push_custom_date
        selected_company = getattr(self, "_pending_push_company", "") or ""

        self._set_tally_push_running_ui(True, "Auto-pushing XML to Tally...", COLORS["warning"])
        self.status_label.configure(text="Posting", text_color=COLORS["warning"])
        self.progress_bar.set(0.25)
        self.log_panel.log(f"Auto-pushing: {Path(xml_path).name} to {tally_url}", "process")
        if selected_company:
            self.log_panel.log(f"Target company: {selected_company}", "info")
        else:
            self.log_panel.log("Target company: currently loaded company in Tally", "info")

        threading.Thread(
            target=self._post_tally_xml_worker,
            args=(xml_path, tally_url, timeout, selected_company, date_mode, custom_tally_date),
            daemon=True,
        ).start()


if __name__ == "__main__":
    app = GSTR2BTallyApp()
    app.mainloop()