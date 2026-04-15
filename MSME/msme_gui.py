# =============================================================================
#  MSME Payment Tracker & Email Automation — GUI
#  pip install customtkinter pandas openpyxl xlrd numpy pywin32
# =============================================================================

import os
import threading
from datetime import datetime, timedelta

import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
from openpyxl import load_workbook

try:
    import win32com.client as win32
    OUTLOOK_OK = True
except ImportError:
    OUTLOOK_OK = False


# ── Global theme ──────────────────────────────────────────────────────────────
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# Accepted Excel extensions for the file-open dialog
EXCEL_TYPES = [
    ("Excel files", "*.xlsx *.xlsm *.xls *.xlsb"),
    ("All files",   "*.*"),
]


# =============================================================================
#  Reusable section-card widget
# =============================================================================
class _Card(ctk.CTkFrame):
    """A labelled, rounded card with a thin top-accent stripe."""

    def __init__(self, parent, title: str, **kw):
        super().__init__(parent, corner_radius=14, **kw)
        self.grid_columnconfigure(0, weight=1)

        # Accent bar
        ctk.CTkFrame(self, height=3, corner_radius=3,
                    fg_color=("#2563EB", "#2563EB")).grid(
            row=0, column=0, sticky="ew", padx=0, pady=(0, 0))

        ctk.CTkLabel(self, text=f"  {title}",
                    font=ctk.CTkFont("Segoe UI", 12, "bold"),
                    anchor="w").grid(row=1, column=0, sticky="ew",
                                    padx=16, pady=(10, 6))

        # Thin separator
        ctk.CTkFrame(self, height=1,
                    fg_color=("gray80", "gray30")).grid(
            row=2, column=0, sticky="ew", padx=16, pady=(0, 8))

        # Inner content frame — callers use .body
        self.body = ctk.CTkFrame(self, fg_color="transparent")
        self.body.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 14))
        self.body.grid_columnconfigure(0, weight=1)


# =============================================================================
#  Stat badge (shows a number + label)
# =============================================================================
class _StatBadge(ctk.CTkFrame):
    def __init__(self, parent, label: str, color: str, **kw):
        super().__init__(parent, corner_radius=10, fg_color=("gray90", "gray20"), **kw)
        self._val = ctk.CTkLabel(self, text="—",
                                font=ctk.CTkFont("Segoe UI", 22, "bold"),
                                text_color=color)
        self._val.pack(padx=20, pady=(10, 2))
        ctk.CTkLabel(self, text=label,
                    font=ctk.CTkFont("Segoe UI", 10),
                    text_color="gray").pack(padx=20, pady=(0, 10))

    def set(self, value):
        self._val.configure(text=str(value))


# =============================================================================
#  MAIN APP
# =============================================================================
class MSMEApp(ctk.CTk):

    def __init__(self):
        super().__init__()
        self.title("MSME Payment Tracker")
        self.geometry("860x700")
        self.minsize(780, 620)

        self._df      = None   # result DataFrame
        self._busy    = False

        self._build()
        self._log("Ready.  Browse any Excel file (.xlsx / .xlsm / .xls) and click Run.", "muted")

    # =========================================================================
    #  LAYOUT
    # =========================================================================
    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        # rows: header | file-card | stats | actions | console | footer
        self.grid_rowconfigure(5, weight=1)

        self._build_header()
        self._build_file_card()
        self._build_stats()
        self._build_actions()
        self._build_console()
        self._build_footer()

    # ── Header ────────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=28, pady=(20, 8))
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr,
                    text="MSME Payment Tracker",
                    font=ctk.CTkFont("Segoe UI", 22, "bold")
                    ).grid(row=0, column=0, sticky="w")

        ctk.CTkLabel(hdr,
                    text="45-Day rule · Overdue detection · Outlook email automation",
                    font=ctk.CTkFont("Segoe UI", 11),
                    text_color="gray"
                    ).grid(row=1, column=0, sticky="w", pady=(2, 0))

        # Appearance switcher
        self._mode_var = ctk.StringVar(value="System")
        ctk.CTkSegmentedButton(
            hdr, values=["Light", "Dark", "System"],
            variable=self._mode_var,
            command=lambda v: ctk.set_appearance_mode(v),
            width=200, font=ctk.CTkFont(size=11),
        ).grid(row=0, column=1, rowspan=2, sticky="e")

    # ── File card ─────────────────────────────────────────────────────────────
    def _build_file_card(self):
        card = _Card(self, "File Selection")
        card.grid(row=1, column=0, sticky="ew", padx=28, pady=4)

        body = card.body
        body.grid_columnconfigure(1, weight=1)

        # --- Input
        ctk.CTkLabel(body, text="Input File",
                    font=ctk.CTkFont(size=12), width=100, anchor="w"
                    ).grid(row=0, column=0, sticky="w", pady=6)

        self._in_var = ctk.StringVar()
        self._in_entry = ctk.CTkEntry(body, textvariable=self._in_var,
                                    placeholder_text="Browse any Excel file…",
                                    font=ctk.CTkFont(size=11), height=36, corner_radius=8)
        self._in_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=6)

        ctk.CTkButton(body, text="Browse", width=86, height=36,
                    corner_radius=8, command=self._browse_input
                    ).grid(row=0, column=2, pady=6)

        # --- Output
        ctk.CTkLabel(body, text="Output File",
                    font=ctk.CTkFont(size=12), width=100, anchor="w"
                    ).grid(row=1, column=0, sticky="w", pady=(0, 6))

        self._out_var = ctk.StringVar()
        ctk.CTkEntry(body, textvariable=self._out_var,
                    placeholder_text="Save output as…",
                    font=ctk.CTkFont(size=11), height=36, corner_radius=8
                    ).grid(row=1, column=1, sticky="ew", padx=8, pady=(0, 6))

        ctk.CTkButton(body, text="Save As", width=86, height=36,
                    corner_radius=8, command=self._browse_output
                    ).grid(row=1, column=2, pady=(0, 6))

    # ── Stats row ─────────────────────────────────────────────────────────────
    def _build_stats(self):
        row = ctk.CTkFrame(self, fg_color="transparent")
        row.grid(row=2, column=0, sticky="ew", padx=28, pady=4)
        for i in range(3):
            row.grid_columnconfigure(i, weight=1)

        self._stat_total   = _StatBadge(row, "Total Rows",      "#8ab4f8")
        self._stat_msme    = _StatBadge(row, "MSME Vendors",     "#fdd663")
        self._stat_overdue = _StatBadge(row, "Overdue Invoices", "#f28b82")

        self._stat_total  .grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self._stat_msme   .grid(row=0, column=1, sticky="ew", padx=6)
        self._stat_overdue.grid(row=0, column=2, sticky="ew", padx=(6, 0))

    # ── Action bar ────────────────────────────────────────────────────────────
    def _build_actions(self):
        bar = ctk.CTkFrame(self, fg_color="transparent")
        bar.grid(row=3, column=0, sticky="ew", padx=28, pady=6)
        bar.grid_columnconfigure((0, 1, 2), weight=1)

        self._run_btn = ctk.CTkButton(
            bar, text="▶   Run Calculations",
            height=44, corner_radius=10,
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            fg_color="#2563EB", hover_color="#1D4ED8",
            command=self._on_run,
        )
        self._run_btn.grid(row=0, column=0, sticky="ew", padx=(0, 7))

        self._email_btn = ctk.CTkButton(
            bar, text="✉   Send Overdue Emails",
            height=44, corner_radius=10,
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            fg_color="#DC2626" if OUTLOOK_OK else "#555",
            hover_color="#B91C1C" if OUTLOOK_OK else "#555",
            state="disabled",
            command=self._on_email,
        )
        self._email_btn.grid(row=0, column=1, sticky="ew", padx=7)

        ctk.CTkButton(
            bar, text="🗑  Clear Log",
            height=44, corner_radius=10,
            font=ctk.CTkFont("Segoe UI", 12),
            fg_color="transparent", border_width=2,
            text_color=("gray30", "gray70"),
            hover_color=("gray85", "gray25"),
            command=self._clear_log,
        ).grid(row=0, column=2, sticky="ew", padx=(7, 0))

        # Progress bar — hidden until busy
        self._prog = ctk.CTkProgressBar(self, mode="indeterminate",
                                        corner_radius=3, height=4)
        self._prog.grid(row=4, column=0, sticky="ew", padx=28)
        self._prog.grid_remove()

    # ── Console ───────────────────────────────────────────────────────────────
    def _build_console(self):
        card = _Card(self, "Log Console")
        card.grid(row=5, column=0, sticky="nsew", padx=28, pady=(4, 4))
        card.grid_rowconfigure(3, weight=1)

        # status pill inside the card header row
        self._pill = ctk.CTkLabel(card, text="● Idle",
                                font=ctk.CTkFont(size=11), text_color="gray")
        self._pill.grid(row=1, column=0, sticky="e", padx=16, pady=0)

        self._console = ctk.CTkTextbox(
            card, font=ctk.CTkFont("Consolas", 11),
            wrap="word", corner_radius=8, activate_scrollbars=True,
        )
        self._console.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 4))
        self._console.configure(state="disabled")

        # colour tags on the underlying tk.Text
        for name, colour in {
            "info":    "#8ab4f8",
            "success": "#81c995",
            "warning": "#fdd663",
            "error":   "#f28b82",
            "muted":   "#888",
        }.items():
            self._console._textbox.tag_config(name, foreground=colour)

    # ── Footer ────────────────────────────────────────────────────────────────
    def _build_footer(self):
        ctk.CTkLabel(self,
                    text="MSME Act 2006  ·  Sections 15–17  ·  45-Day Payment Rule",
                    font=ctk.CTkFont(size=10), text_color="gray"
                    ).grid(row=5, column=0, pady=(0, 10))

    # =========================================================================
    #  FILE BROWSING
    # =========================================================================
    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Select Input Excel File",
            filetypes=EXCEL_TYPES,
        )
        if not path:
            return
        self._in_var.set(path)
        # Auto-suggest output next to input
        base = os.path.splitext(path)[0]
        self._out_var.set(base + "_MSME_Output.xlsx")
        self._log(f"Input  → {os.path.basename(path)}", "info")

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Output File",
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self._out_var.set(path)
            self._log(f"Output → {os.path.basename(path)}", "info")

    # =========================================================================
    #  LOGGING
    # =========================================================================
    def _log(self, msg: str, level: str = "info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self._console.configure(state="normal")
        self._console._textbox.insert("end", f"[{ts}]  {msg}\n", level)
        self._console._textbox.see("end")
        self._console.configure(state="disabled")

    def _clear_log(self):
        self._console.configure(state="normal")
        self._console.delete("1.0", "end")
        self._console.configure(state="disabled")
        self._log("Log cleared.", "muted")

    # =========================================================================
    #  BUSY STATE
    # =========================================================================
    def _set_busy(self, on: bool):
        self._busy = on
        self._run_btn.configure(state="disabled" if on else "normal")
        if on:
            self._prog.grid()
            self._prog.start()
            self._pill.configure(text="● Running…", text_color="#fdd663")
        else:
            self._prog.stop()
            self._prog.grid_remove()

    # =========================================================================
    #  BUTTON HANDLERS
    # =========================================================================
    def _on_run(self):
        if self._busy:
            return

        inp = self._in_var.get().strip()
        out = self._out_var.get().strip()

        if not inp:
            messagebox.showwarning("No Input File", "Please browse and select an Excel file first.")
            return
        if not os.path.isfile(inp):
            messagebox.showerror("File Not Found", f"Cannot find:\n{inp}")
            return
        if not out:
            messagebox.showwarning("No Output Path", "Please specify where to save the output.")
            return

        self._log("─" * 58, "muted")
        self._email_btn.configure(state="disabled")
        self._df = None
        for stat in (self._stat_total, self._stat_msme, self._stat_overdue):
            stat.set("—")

        self._set_busy(True)
        threading.Thread(target=self._calc_worker,
                        args=(inp, out), daemon=True).start()

    def _on_email(self):
        if self._busy or self._df is None:
            return
        if not messagebox.askyesno(
            "Confirm",
            "Send overdue payment reminder emails to all flagged clients via Outlook?\n\nThis cannot be undone."
        ):
            return
        self._log("─" * 58, "muted")
        self._set_busy(True)
        threading.Thread(target=self._email_worker, daemon=True).start()

    # =========================================================================
    #  WORKER — CALCULATIONS
    # =========================================================================
    def _calc_worker(self, inp: str, out: str):
        try:
            self._log(f"Reading  →  {os.path.basename(inp)}", "info")

            df = pd.read_excel(inp, sheet_name=0)
            total_rows = len(df)
            self._log(f"Rows loaded : {total_rows}", "info")
            self.after(0, lambda: self._stat_total.set(total_rows))

            # Parse date columns F (idx 5) and G (idx 6)
            df.iloc[:, 5] = pd.to_datetime(df.iloc[:, 5], errors="coerce")
            df.iloc[:, 6] = pd.to_datetime(df.iloc[:, 6], errors="coerce")

            # Add output columns H, I, J if missing
            if len(df.columns) < 10:
                df = df.reindex(columns=list(df.columns) +
                                ["MSME Due Date", "Overdue Status", "Days Delay"])

            self._log("Calculating 45-day due dates…", "info")

            msme_count   = 0
            overdue_count = 0

            for idx, row in df.iterrows():
                status = str(row.iloc[3]).strip().lower()

                if status == "yes":
                    msme_count += 1
                    inv_date = row.iloc[5]
                    pay_date = row.iloc[6]

                    if pd.notnull(inv_date):
                        due_date = inv_date + timedelta(days=45)
                        df.iloc[idx, 7] = due_date.date()

                        if pd.notnull(pay_date):
                            if pay_date > due_date:
                                df.iloc[idx, 8] = "Yes"
                                df.iloc[idx, 9] = int((pay_date - due_date).days)
                                overdue_count += 1
                            else:
                                df.iloc[idx, 8] = "No"
                                df.iloc[idx, 9] = np.nan
                        else:
                            df.iloc[idx, 8] = "No Payment Date"
                            df.iloc[idx, 9] = np.nan
                    else:
                        df.iloc[idx, 7] = "Invalid Date"
                        df.iloc[idx, 8] = np.nan
                        df.iloc[idx, 9] = np.nan
                else:
                    df.iloc[idx, 7] = np.nan
                    df.iloc[idx, 8] = np.nan
                    df.iloc[idx, 9] = np.nan

            # 1. Parse into standard datetimes first to handle weird Excel inputs cleanly
            col_inv = pd.to_datetime(df.iloc[:, 5], errors="coerce")
            col_pay = pd.to_datetime(df.iloc[:, 6], errors="coerce")
            col_due = pd.to_datetime(df.iloc[:, 7], errors="coerce")

            # 2. Force conversion to native Python dates. Swap NaT for None to break the datetime64 lock.
            df.iloc[:, 5] = col_inv.apply(lambda x: x.date() if pd.notna(x) else None)
            df.iloc[:, 6] = col_pay.apply(lambda x: x.date() if pd.notna(x) else None)
            df.iloc[:, 7] = col_due.apply(lambda x: x.date() if pd.notna(x) else None)

            # 3. Save then force date-only number format on date columns via openpyxl
            df.to_excel(out, index=False)
            wb = load_workbook(out)
            ws = wb.active
            for col_letter in ("F", "G", "H"):
                for cell in list(ws[col_letter])[1:]:  # skip header row
                    if cell.value is not None:
                        cell.number_format = "YYYY-MM-DD"
            wb.save(out)
            self._df = df

            self.after(0, lambda: self._stat_msme.set(msme_count))
            self.after(0, lambda: self._stat_overdue.set(overdue_count))
            self._log(f"MSME vendors   : {msme_count}", "info")
            self._log(f"Overdue invoices: {overdue_count}",
                    "warning" if overdue_count else "success")
            self._log(f"Saved  →  {os.path.basename(out)}", "success")
            self._log("✔  Calculations complete!", "success")
            self.after(0, lambda: self._pill.configure(text="● Done", text_color="#81c995"))
            self.after(0, lambda: messagebox.showinfo(
                "Done",
                f"Calculations complete!\n\nMSME Vendors   : {msme_count}\nOverdue Invoices: {overdue_count}\n\nOutput saved to:\n{os.path.basename(out)}"
            ))

            if OUTLOOK_OK and overdue_count > 0:
                self.after(0, lambda: self._email_btn.configure(state="normal"))
            elif overdue_count == 0:
                self._log("No overdue rows — email button remains disabled.", "muted")

        except Exception as e:
            self._log(f"✘  {e}", "error")
            self.after(0, lambda: self._pill.configure(text="● Error", text_color="#f28b82"))
        finally:
            self.after(0, lambda: self._set_busy(False))

    # =========================================================================
    #  WORKER — EMAILS
    # =========================================================================
    def _email_worker(self):
        try:
            self._log("Connecting to Outlook…", "info")
            outlook = win32.Dispatch("outlook.application")
            sent = 0

            for _, row in self._df.iterrows():
                if str(row.iloc[8]).strip().lower() != "yes":
                    continue

                to = str(row.iloc[1]).strip()
                if not to or to.lower() == "nan":
                    continue

                name      = str(row.iloc[0])
                inv_no    = str(row.iloc[2])
                amount    = str(row.iloc[4])
                inv_dt    = row.iloc[5].strftime("%d-%m-%Y")
                due_dt    = row.iloc[7].strftime("%d-%m-%Y")
                delay     = str(int(row.iloc[9]))

                mail         = outlook.CreateItem(0)
                mail.To      = to
                mail.Subject = "Urgent: Overdue Payment Reminder under MSME Act"
                mail.Body    = (
                    f"Dear {name},\n\n"
                    "I hope you are doing well.\n\n"
                    "This is a gentle reminder that the following invoice—raised by us as a registered "
                    "Micro/Small Enterprise under the MSME Development Act, 2006—is now overdue:\n\n"
                    f"  Invoice No.         : {inv_no}\n"
                    f"  Invoice Date        : {inv_dt}\n"
                    f"  MSME Due Date       : {due_dt}\n"
                    f"  Outstanding Amount  : ₹ {amount}\n"
                    f"  Days Overdue        : {delay}\n\n"
                    "As per Section 15 of the MSME Act, payment must be made within 45 days. "
                    "Sections 16 & 17 mandate interest at three times the RBI bank rate, compounded "
                    "monthly, until full settlement.\n\n"
                    "Kindly settle the above within 15 days. If already paid, please share the UTR "
                    "number for our records.\n\n"
                    "Thank you for your cooperation.\n\nWarm regards,"
                )
                mail.Send()
                sent += 1
                self._log(f"  ✉  Sent → {to}", "success")

            self._log(f"✔  {sent} email(s) sent successfully.", "success")
            self.after(0, lambda: self._pill.configure(text="● Done", text_color="#81c995"))

        except Exception as e:
            self._log(f"✘  {e}", "error")
            self.after(0, lambda: self._pill.configure(text="● Error", text_color="#f28b82"))
        finally:
            self.after(0, lambda: self._set_busy(False))


# =============================================================================
if __name__ == "__main__":
    app = MSMEApp()
    app.mainloop()