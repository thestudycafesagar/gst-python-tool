import threading
import time
import os
import json
import pandas as pd
import customtkinter as ctk
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- CONFIGURATION ---
# Commented out: theme is controlled globally by GST_Suite.py
# ctk.set_appearance_mode("System")
# ctk.set_default_color_theme("blue")

# --- GST MAPPING UTILITIES ---
GST_STATE_CODES = {
    "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh", 
    "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh", 
    "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", 
    "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal", 
    "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat", 
    "25": "Daman & Diu", "26": "Dadra & Nagar Haveli", "27": "Maharashtra", "29": "Karnataka", 
    "30": "Goa", "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu", "34": "Puducherry", 
    "35": "Andaman & Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh", "38": "Ladakh", 
    "97": "Other Territory", "96": "Other Country"
}

def get_month_name(fp_string):
    """ Converts '042024' -> 'April' """
    try:
        if not fp_string: return ""
        month_num = int(fp_string[:2])
        return datetime(2000, month_num, 1).strftime('%B')
    except:
        return fp_string

def get_financial_year(fp_string):
    """ Converts '042024' -> '2024 - 2025' """
    try:
        if not fp_string: return ""
        month = int(fp_string[:2])
        year = int(fp_string[2:])
        if month >= 4:
            return f"{year} - {year + 1}"
        else:
            return f"{year - 1} - {year}"
    except:
        return fp_string

def map_state(code):
    return GST_STATE_CODES.get(str(code), str(code))

def clean_reverse_charge(val):
    if val == "Y": return "Yes"
    if val == "N": return "No"
    return val

# --- WORKER CLASS ---
class GstConversionWorker:
    def __init__(self, app_instance, json_files, output_dir):
        self.app = app_instance
        self.json_files = json_files
        self.output_dir = output_dir
        self.keep_running = True

    def log(self, message):
        self.app.update_log_safe(message)

    def format_excel_header(self, file_path, gstin_value, return_period):
        """ Applies the header format: GSTIN, Period, Report Name """
        try:
            wb = load_workbook(file_path)
            bold_font = Font(bold=True)
            
            report_map = {
                "GSTR-1-1A_B2B": "GSTR-1-1A-B2B",
                "GSTR-1-1A_B2CS": "GSTR-1-1A-B2CS",
                "GSTR-1-1A_CDN": "GSTR-1-1A-CDN",
                "GSTR-1-1A_HSNSummary": "GSTR-1-1A-HSN Summary",
                "GSTR-1-1A_DocIssue": "GSTR-1-1A-Document Issue",
                "GSTR-1-1A_B2B Amendment": "GSTR-1-1A-B2B Amendment"
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.insert_rows(1, amount=4)
                
                # Row 1: GSTIN
                ws.cell(row=1, column=2, value="Company GSTIN :").font = bold_font
                ws.cell(row=1, column=3, value=gstin_value).font = bold_font
                
                # Row 2: Return Period
                ws.cell(row=2, column=2, value="Return Period :").font = bold_font
                ws.cell(row=2, column=3, value=return_period).font = bold_font
                
                # Row 3: Report Name
                display_report = report_map.get(sheet_name, sheet_name)
                ws.cell(row=3, column=2, value="Report Name :").font = bold_font
                ws.cell(row=3, column=3, value=display_report).font = bold_font

            wb.save(file_path)
        except Exception as e:
            self.log(f"   ⚠️ Formatting Error: {e}")

    def get_unique_filename(self, target_path):
        """ 
        Generates a unique filename like 'File (1).xlsx' if 'File.xlsx' exists.
        Handles the issue of overwriting previous data.
        """
        if not os.path.exists(target_path):
            return target_path
        
        base, ext = os.path.splitext(target_path)
        counter = 1
        while True:
            new_path = f"{base} ({counter}){ext}"
            if not os.path.exists(new_path):
                return new_path
            counter += 1

    # --- PROCESSORS ---
    def process_b2b(self, json_data, month):
        rows = []
        if 'b2b' in json_data:
            for supplier in json_data['b2b']:
                ctin = supplier.get('ctin', '')
                for inv in supplier.get('inv', []):
                    for item in inv.get('itms', []):
                        det = item.get('itm_det', {})
                        rows.append({
                            "Month": month,
                            "GSTIN/UIN": ctin,
                            "Party Name": "", 
                            "Invoice No": inv.get('inum', ''),
                            "Invoice Date": inv.get('idt', ''),
                            "Invoice Value": inv.get('val', 0),
                            "Rate": det.get('rt', 0),
                            "Total Taxable Value": det.get('txval', 0),
                            "IGST Amount": det.get('iamt', 0),
                            "CGST Amount": det.get('camt', 0),
                            "SGST Amount": det.get('samt', 0),
                            "CESS Amount": det.get('csamt', 0),
                            "Place Of Supply": map_state(inv.get('pos', '')),
                            "Reverse Charge": clean_reverse_charge(inv.get('rchrg', 'N')),
                            "Invoice Type": inv.get('inv_typ', 'Regular'),
                            "E-Commerce GSTIN": inv.get('etin', ''),
                            "Return Type": "GSTR-1"
                        })
        df = pd.DataFrame(rows)
        if not df.empty: df.insert(0, "Sr. #", range(1, len(df) + 1))
        return df

    def process_hsn(self, json_data, month):
        rows = []
        if 'hsn' in json_data:
            hsn_data = json_data['hsn']
            # JSON uses 'hsn_b2b' and 'hsn_b2c' — NOT 'data'
            all_items = hsn_data.get('hsn_b2b', []) + hsn_data.get('hsn_b2c', [])
            for item in all_items:
                rows.append({
                    "Month": month,
                    "HSN": item.get('hsn_sc', ''),
                    "Description": item.get('desc', ''),
                    "User Description": item.get('user_desc', ''),
                    "UQC": item.get('uqc', ''),
                    "Total Quantity": item.get('qty', 0),
                    "GST %": item.get('rt', 0),
                    "Total Taxable Value": item.get('txval', 0),
                    "IGST Amount": item.get('iamt', 0),
                    "CGST Amount": item.get('camt', 0),
                    "SGST Amount": item.get('samt', 0),
                    "Cess Amount": item.get('csamt', 0)
                })
        df = pd.DataFrame(rows)
        if not df.empty: df.insert(0, "Sr. #", range(1, len(df) + 1))
        return df

    def process_doc_issue(self, json_data, month):
        rows = []
        doc_map = {"1": "Invoice for outward supply", "5": "Credit Note", "12": "Delivery Challan"}
        if 'doc_issue' in json_data and 'doc_det' in json_data['doc_issue']:
            for category in json_data['doc_issue']['doc_det']:
                nature = doc_map.get(str(category.get('doc_num', '')), "Others")
                for doc in category.get('docs', []):
                    rows.append({
                        "Month": month,
                        "Nature of Document": nature,
                        "Sr. No. From": doc.get('from', ''),
                        "Sr. No. To": doc.get('to', ''),
                        "Total Number": doc.get('totnum', 0),
                        "Cancelled": doc.get('cancel', 0),
                        "Net Issued": doc.get('net_issue', 0)
                    })
        df = pd.DataFrame(rows)
        if not df.empty: df.insert(0, "Sr. #", range(1, len(df) + 1))
        return df

    def process_cdn(self, json_data, month):
        rows = []
        if 'cdnr' in json_data:
            for supplier in json_data['cdnr']:
                ctin = supplier.get('ctin', '')
                for note in supplier.get('nt', []):
                    nt_type = "Credit Note" if note.get('ntty') == "C" else "Debit Note"
                    for item in note.get('itms', []):
                        det = item.get('itm_det', {})
                        rows.append({
                            "Month": month,
                            "GSTIN/UIN of Recipient": ctin,
                            "Party Name": "",
                            "Type of note (Debit/ Credit)": nt_type,
                            "Pre GST Regime Dr./ Cr. Notes": "No",
                            "Debit Note/ credit note/ Refund voucher No.": note.get('nt_num', ''),
                            "Debit Note/ credit note/ Refund voucher Date": note.get('nt_dt', ''),
                            "Original Invoice No": note.get('inum', ''),
                            "Original Invoice Date": note.get('idt', ''),
                            "Note/Refund Voucher Value": note.get('val', 0),
                            "Rate": det.get('rt', 0),
                            "Taxable Value": det.get('txval', 0),
                            "IGST Amount": det.get('iamt', 0),
                            "CGST Amount": det.get('camt', 0),
                            "SGST Amount": det.get('samt', 0),
                            "CESS Amount": det.get('csamt', 0),
                            "Place of supply": map_state(note.get('pos', '')),
                            "Return Type": "GSTR-1"
                        })
        df = pd.DataFrame(rows)
        if not df.empty: df.insert(0, "Sr. #", range(1, len(df) + 1))
        return df
    
    def process_b2cs(self, json_data, month):
        rows = []
        if 'b2cs' in json_data:
            for item in json_data['b2cs']:
                rows.append({
                    "Month": month,
                    "Supply Type": item.get('sply_ty', ''),
                    "Place of Supply": map_state(item.get('pos', '')),
                    "Rate of Tax": item.get('rt', 0),
                    "Total Taxable Value": item.get('txval', 0),
                    "IGST Amount": item.get('iamt', 0),
                    "CGST Amount": item.get('camt', 0),
                    "SGST Amount": item.get('samt', 0),
                    "Cess Amount": item.get('csamt', 0),
                    "GSTIN of E-commerce Operator": item.get('etin', ''),
                    "Return Type": "GSTR-1"
                })
        df = pd.DataFrame(rows)
        if not df.empty: df.insert(0, "Sr. #", range(1, len(df) + 1))
        return df

    def run(self):
        self.log("🚀 STARTING BATCH PROCESS...")
        
        try:
            if not self.json_files:
                self.app.process_finished_safe("No files queued.")
                return
            
            if not os.path.exists(self.output_dir):
                os.makedirs(self.output_dir)

            total = len(self.json_files)
            
            for i, file_path in enumerate(self.json_files):
                if not self.keep_running: break
                
                filename = os.path.basename(file_path)
                self.log(f"🔹 Processing ({i+1}/{total}): {filename}")
                self.app.update_progress_safe(i / total)
                
                try:
                    with open(file_path, 'r') as f:
                        data = json.load(f)
                    
                    gstin = data.get('gstin', 'Unknown_GSTIN')
                    fp = data.get('fp', '')
                    month_name = get_month_name(fp)
                    fin_year = get_financial_year(fp)
                    
                    self.log(f"   🏢 Found GSTIN: {gstin}")

                    dfs = {
                        'GSTR-1-1A_B2B': self.process_b2b(data, month_name),
                        'GSTR-1-1A_B2CS': self.process_b2cs(data, month_name),
                        'GSTR-1-1A_CDN': self.process_cdn(data, month_name),
                        'GSTR-1-1A_HSNSummary': self.process_hsn(data, month_name),
                        'GSTR-1-1A_DocIssue': self.process_doc_issue(data, month_name),
                        'GSTR-1-1A_B2B Amendment': pd.DataFrame()
                    }
                    
                    # Output Name
                    out_name = f"{gstin}_{month_name}.xlsx"
                    target_path = os.path.join(self.output_dir, out_name)
                    
                    # FILE VERSIONING FIX: Get unique name like 'File (1).xlsx'
                    final_path = self.get_unique_filename(target_path)
                    
                    self.log(f"   ⚙️ Writing to: {os.path.basename(final_path)}")

                    with pd.ExcelWriter(final_path, engine='openpyxl') as writer:
                        for sheet, df in dfs.items():
                            df.to_excel(writer, sheet_name=sheet, index=False)

                    self.format_excel_header(final_path, gstin, fin_year)
                    self.log(f"   ✅ Finished.")

                except Exception as e:
                    self.log(f"   ❌ Error with file: {e}")
            
            self.app.update_progress_safe(1.0)
            self.log("-" * 30)
            self.log("✅ BATCH COMPLETE.")
            self.app.process_finished_safe(f"Processed {total} files.\nSaved in: {self.output_dir}")

        except Exception as e:
            self.log(f"❌ Critical Failure: {e}")
            self.app.process_finished_safe(f"Error: {e}")

# --- MODERN GUI CLASS ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("GSTR-1 Converter Pro")
        self.geometry("900x700")
        
        self.worker = None
        self.selected_files = [] # Stores file paths

        # --- LAYOUT SETUP ---
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # 1. SIDEBAR
        self.sidebar = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(4, weight=1)

        self.logo_label = ctk.CTkLabel(self.sidebar, text="GST TOOL", font=ctk.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        self.lbl_desc = ctk.CTkLabel(self.sidebar, text="GSTR-1 JSON\nto Excel", font=ctk.CTkFont(size=14))
        self.lbl_desc.grid(row=1, column=0, padx=20, pady=10)

        self.btn_clear = ctk.CTkButton(self.sidebar, text="Clear Queue", command=self.clear_queue, fg_color="#DC2626", hover_color="#B91C1C")
        self.btn_clear.grid(row=2, column=0, padx=20, pady=20)

        # Theme controls removed — theme is controlled globally by GST_Suite.py
        # self.appearance_mode_label = ctk.CTkLabel(self.sidebar, text="Appearance Mode:", anchor="w")
        # self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        # self.appearance_mode_optionemenu = ctk.CTkOptionMenu(self.sidebar, values=["Dark", "Light", "System"],
        #                                                      command=self.change_appearance_mode_event)
        # self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 20))

        # 2. MAIN AREA
        self.main_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, sticky="nsew")
        self.main_frame.grid_rowconfigure(2, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)

        # Top Bar
        self.top_bar = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.top_bar.grid(row=0, column=0, sticky="ew", padx=20, pady=20)
        
        self.lbl_status = ctk.CTkLabel(self.top_bar, text="0 Files Selected", font=ctk.CTkFont(size=24, weight="bold"))
        self.lbl_status.pack(side="left")
        
        self.btn_add = ctk.CTkButton(self.top_bar, text="+ Add JSON Files", command=self.add_files, 
                                     font=ctk.CTkFont(size=14, weight="bold"), height=40, width=150)
        self.btn_add.pack(side="right")

        # File List Area
        self.file_list_frame = ctk.CTkScrollableFrame(self.main_frame, label_text="Queue")
        self.file_list_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 20))
        
        # Placeholder text in list
        self.lbl_empty = ctk.CTkLabel(self.file_list_frame, text="No files added yet. Click '+ Add JSON Files' to start.", text_color="gray")
        self.lbl_empty.pack(pady=20)

        # Log & Progress Area
        self.log_frame = ctk.CTkFrame(self.main_frame)
        self.log_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 20))
        self.log_frame.grid_rowconfigure(1, weight=1)
        self.log_frame.grid_columnconfigure(0, weight=1)

        self.lbl_log = ctk.CTkLabel(self.log_frame, text="Execution Logs", font=ctk.CTkFont(weight="bold"))
        self.lbl_log.grid(row=0, column=0, sticky="w", padx=10, pady=5)
        
        self.log_box = ctk.CTkTextbox(self.log_frame, font=("Consolas", 12))
        self.log_box.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.log_box.configure(state="disabled")


        self.prog_bar = ctk.CTkProgressBar(self.main_frame, height=15)
        self.prog_bar.grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 10))
        self.prog_bar.set(0)

        self.btn_start = ctk.CTkButton(self.main_frame, text="START CONVERSION", height=40, font=ctk.CTkFont(size=16, weight="bold"), fg_color="#047857", hover_color="#059669", command=self.start_process)
        self.btn_start.grid(row=4, column=0, sticky="ew", padx=20, pady=(0, 10))

        self.btn_open_folder = ctk.CTkButton(self.main_frame, text="📂 OPEN OUTPUT FOLDER", height=40,
                                       font=ctk.CTkFont(size=14, weight="bold"), fg_color="#64748B",
                                       hover_color="#475569", command=self.open_output_folder)
        self.btn_open_folder.grid(row=5, column=0, sticky="ew", padx=20, pady=(0, 20))
        self.btn_open_folder.grid_remove()

    def change_appearance_mode_event(self, new_appearance_mode: str):
        pass  # Theme controlled by GST_Suite.py

    def add_files(self):
        files = filedialog.askopenfilenames(filetypes=[("JSON Files", "*.json")])
        if files:
            new_count = 0
            for f in files:
                # DUPLICATE CHECK: Only add if not already in list
                if f not in self.selected_files:
                    self.selected_files.append(f)
                    new_count += 1
            
            if new_count > 0:
                self.update_file_list_ui()
                self.log_gui(f"Added {new_count} new files.")
            else:
                self.log_gui("No new files added (Duplicates ignored).")

    def update_file_list_ui(self):
        # Update Count Label
        count = len(self.selected_files)
        self.lbl_status.configure(text=f"{count} File{'s' if count!=1 else ''} Selected")
        
        # Enable/Disable Start Button
        if count > 0:
            self.btn_start.configure(state="normal")
            self.lbl_empty.pack_forget()
        else:
            self.btn_start.configure(state="disabled")
            self.lbl_empty.pack(pady=20)

        # Clear current list widgets
        for widget in self.file_list_frame.winfo_children():
            if widget != self.lbl_empty:
                widget.destroy()

        # Re-populate list
        for i, f in enumerate(self.selected_files):
            fname = os.path.basename(f)
            row = ctk.CTkFrame(self.file_list_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=f"{i+1}. {fname}", anchor="w").pack(side="left", padx=10)

    def clear_queue(self):
        self.selected_files = []
        self.update_file_list_ui()
        self.log_gui("Queue cleared.")

    def log_gui(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def update_log_safe(self, msg):
        self.after(0, lambda: self.log_gui(msg))

    def update_progress_safe(self, val):
        self.after(0, lambda: self.prog_bar.set(val))

    def process_finished_safe(self, msg):
        self.after(0, lambda: messagebox.showinfo("Process Complete", msg))
        self.after(0, lambda: self.btn_start.configure(state="normal"))
        self.after(0, lambda: self.prog_bar.set(0))
        self.after(0, lambda: self.btn_open_folder.grid())

    def open_output_folder(self):
        target = os.path.join(os.getcwd(), "GST Downloaded", "GSTR1 Json to Excel")
        if os.path.exists(target):
            os.startfile(target)
        else:
            messagebox.showinfo("Info", "Output folder not found.")

    def start_process(self):
        if not self.selected_files: return
        
        self.btn_start.configure(state="disabled")
        self.btn_open_folder.grid_remove()
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        
        base = os.path.join(os.getcwd(), "GST Downloaded", "GSTR1 Json to Excel")
        self.worker = GstConversionWorker(self, self.selected_files, base)
        threading.Thread(target=self.worker.run, daemon=True).start()

if __name__ == "__main__":
    app = App()
    app.mainloop()