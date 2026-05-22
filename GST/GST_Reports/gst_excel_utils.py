import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# =============================================================================
# Colour palette & style helpers
# =============================================================================

_NAVY   = "2F75B5"  # Excel Blue
_WHITE  = "FFFFFF"
_ALTBG  = "F3F2F1"  # Office Grey tint
_BORDER_CLR = "D2D0CE"


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _side() -> "Side":
    return Side(style="thin", color=_BORDER_CLR)


def _border() -> "Border":
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h: str = "left", wrap: bool = False) -> "Alignment":
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# =============================================================================
# Numeric / string helpers / state mapping
# =============================================================================

_STATE_CODES = {
    "01": "JAMMU AND KASHMIR", "02": "HIMACHAL PRADESH", "03": "PUNJAB",
    "04": "CHANDIGARH", "05": "UTTARAKHAND", "06": "HARYANA", "07": "DELHI",
    "08": "RAJASTHAN", "09": "UTTAR PRADESH", "10": "BIHAR", "11": "SIKKIM",
    "12": "ARUNACHAL PRADESH", "13": "NAGALAND", "14": "MANIPUR", "15": "MIZORAM",
    "16": "TRIPURA", "17": "MEGHALAYA", "18": "ASSAM", "19": "WEST BENGAL",
    "20": "JHARKHAND", "21": "ODISHA", "22": "CHHATTISGARH", "23": "MADHYA PRADESH",
    "24": "GUJARAT", "25": "DAMAN AND DIU", "26": "DADRA AND NAGAR HAVELI",
    "27": "MAHARASHTRA", "29": "KARNATAKA", "30": "GOA", "31": "LAKSHADWEEP",
    "32": "KERALA", "33": "TAMIL NADU", "34": "PUDUCHERRY",
    "35": "ANDAMAN AND NICOBAR ISLANDS", "36": "TELANGANA", "37": "ANDHRA PRADESH",
    "38": "LADAKH", "97": "OTHER TERRITORY", "99": "OTHER COUNTRY"
}

_DOC_NATURE = {
    1: "Invoices for outward supply",
    2: "Invoices for inward supply from unregistered person",
    3: "Revised Invoice",
    4: "Debit Note",
    5: "Credit Note",
    6: "Receipt voucher",
    7: "Payment Voucher",
    8: "Refund voucher",
    9: "Delivery Challan for job work",
    10: "Delivery Challan for supply on approval",
    11: "Delivery Challan in case of liquid gas",
    12: "Delivery Challan in cases other than by way of supply"
}

def _doc_name(num) -> str:
    try: return _DOC_NATURE.get(int(num), str(num))
    except (ValueError, TypeError): return str(num)

def _state_name(code: str) -> str:
    if not code: return ""
    c = str(code).strip()
    if "-" in c: c = c.split("-")[0].strip()
    return _STATE_CODES.get(c.zfill(2), code)

def _n(v, default: float = 0.0) -> float:
    try:
        return float(v) if v not in (None, "", "null") else default
    except (TypeError, ValueError):
        return default


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


# =============================================================================
# Sheet writer
# =============================================================================

def _write_sheet(ws, headers: list, rows: list, num_col_indices: set = None, 
                 title: str = None, subtitle: str = None):
    """
    Write a styled table (header + data) to a worksheet.
    num_col_indices: 1-based column indices that contain numeric values.
    """
    start_row = 1
    if title: start_row += 1
    if subtitle: start_row += 1

    ws.freeze_panes = f"A{start_row + 1}"
    
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=11, name="Calibri")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 5))
    
    if subtitle:
        ws.cell(row=2 if title else 1, column=1, value=subtitle).font = Font(bold=True, size=10, name="Calibri")
        ws.merge_cells(start_row=2 if title else 1, start_column=1, 
                       end_row=2 if title else 1, end_column=max(len(headers), 5))

    hdr_fill = _fill(_NAVY)
    hdr_font = Font(bold=True, color=_WHITE, size=10, name="Calibri")
    alt_fill = _fill(_ALTBG)
    bdr      = _border()
    num_cols = num_col_indices or set()

    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = bdr
        cell.alignment = _align("center")

    # Data rows
    body_font = Font(size=10, name="Calibri")
    for r, row in enumerate(rows, start_row + 1):
        fill = alt_fill if r % 2 == 0 else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font   = body_font
            cell.border = bdr
            if fill:
                cell.fill = fill
            is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
            if is_num or c in num_cols:
                cell.alignment     = _align("right")
                cell.number_format = "#,##0.00"
            else:
                cell.alignment = _align("left")

    # Column widths
    for c, h in enumerate(headers, 1):
        col_letter = get_column_letter(c)
        if c in num_cols:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = max(14, len(str(h)) + 4)


def _empty_sheet(wb, msg: str = "No data found in JSON."):
    ws = wb.create_sheet("No Data")
    ws.cell(row=1, column=1, value=msg).font = Font(italic=True, color="888888")
    ws.column_dimensions["A"].width = 40


# =============================================================================
# Common B2B / CDN flatteners  (reused across GSTR-1, 2A, 2B)
# =============================================================================

_B2B_HEADERS = [
    "Sr", "GSTIN", "Recipient Name", "State", "Invoice Number", "date",
    "Invoice Value", "Taxable Value", "IGST", "CGST", "SGST", "Cess",
    "Cfs", "RC", "Up By"
]
_B2B_NUM_COLS = {7, 8, 9, 10, 11, 12}

_CDN_HEADERS = ["Supplier GSTIN", "Trade Name", "Note Type", "Note No", "Note Date", "Invoice Type",
                "Rate %", "Taxable Value", "IGST", "CGST", "SGST", "Cess"]
_CDN_NUM_COLS = {7, 8, 9, 10, 11, 12}


def _b2b_rows(b2b_list: list, extra_hdrs=None, extra_fn=None):
    """Flatten supplier→invoice→item list into flat rows."""
    hdrs = _B2B_HEADERS + (extra_hdrs or [])
    rows = []
    for party in (b2b_list or []):
        gstin = _s(party.get("ctin") or party.get("stin", ""))
        name  = _s(party.get("trdnm") or party.get("trade_name") or 
                   party.get("name") or party.get("lgl_nm") or 
                   party.get("legal_name") or "")
        for inv in (party.get("inv") or []):
            inv_no  = _s(inv.get("inum") or inv.get("inv_no") or "")
            inv_dt  = _s(inv.get("idt")  or inv.get("dt") or inv.get("inv_dt") or "")
            inv_val = _n(inv.get("val") or inv.get("inv_val"))
            pos     = _state_name(_s(inv.get("pos")  or ""))
            rchrg   = _s(inv.get("rchrg") or "N")
            cfs     = _s(party.get("cfs") or "")
            updby   = _s(inv.get("updby") or "S")

            itms = inv.get("itms") or []
            if not itms:
                # Handle cases like GSTR-2B where values are directly in the invoice
                ext = extra_fn(party, inv) if extra_fn else []
                rows.append([
                    len(rows) + 1,
                    gstin, name, pos, inv_no, inv_dt,
                    inv_val, _n(inv.get("txval")),
                    _n(inv.get("iamt") or inv.get("igst")),
                    _n(inv.get("camt") or inv.get("cgst")),
                    _n(inv.get("samt") or inv.get("sgst")),
                    _n(inv.get("csamt") or inv.get("cess")),
                    cfs, rchrg, updby
                ] + ext)
            else:
                for itm in itms:
                    d = itm.get("itm_det") or itm
                    ext = extra_fn(party, inv) if extra_fn else []
                    rows.append([
                        len(rows) + 1,
                        gstin, name, pos, inv_no, inv_dt,
                        inv_val, _n(d.get("txval")),
                        _n(d.get("iamt") or d.get("igst")),
                        _n(d.get("camt") or d.get("cgst")),
                        _n(d.get("samt") or d.get("sgst")),
                        _n(d.get("csamt") or d.get("cess")),
                        cfs, rchrg, updby
                    ] + ext)
    return hdrs, rows


def _cdn_rows(cdn_list: list):
    """Flatten credit/debit note list into flat rows."""
    rows = []
    for party in (cdn_list or []):
        gstin = _s(party.get("ctin") or party.get("stin") or "")
        name  = _s(party.get("trdnm") or party.get("trade_name") or 
                   party.get("name") or party.get("lgl_nm") or 
                   party.get("legal_name") or "")
        for note in (party.get("nt") or []):
            ntyp = _s(note.get("ntty") or note.get("typ") or "")
            nno  = _s(note.get("ntnum") or note.get("nt_num") or note.get("nt_no") or "")
            ndt  = _s(note.get("ndt") or note.get("dt") or note.get("nt_dt") or "")
            ityp = _s(note.get("suptyp") or note.get("inv_typ") or "")
            itms = note.get("itms") or []
            if not itms:
                # GSTR-2B often has values directly in the note object
                rows.append([
                    gstin, name, ntyp, nno, ndt, ityp,
                    _n(note.get("rt")),
                    _n(note.get("txval")),
                    _n(note.get("iamt") or note.get("igst")),
                    _n(note.get("camt") or note.get("cgst")),
                    _n(note.get("samt") or note.get("sgst")),
                    _n(note.get("csamt") or note.get("cess")),
                ])
            else:
                for itm in itms:
                    d = itm.get("itm_det") or itm
                    rows.append([
                        gstin, name, ntyp, nno, ndt, ityp,
                        _n(d.get("rt")),
                        _n(d.get("txval")),
                        _n(d.get("iamt") or d.get("igst")),
                        _n(d.get("camt") or d.get("cgst")),
                        _n(d.get("samt") or d.get("sgst")),
                        _n(d.get("csamt") or d.get("cess")),
                    ])
    return _CDN_HEADERS, rows
