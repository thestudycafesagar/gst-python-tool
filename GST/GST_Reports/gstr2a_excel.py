import json
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gst_excel_utils import _state_name, _n, _s

# =============================================================================
# Portal-Style Styles (Matching GSTR-2A)
# =============================================================================

_NAVY   = "002060"  # Portal dark navy (exact match)
_WHITE  = "FFFFFF"
_LGREY  = "F2F2F2"
_BORDER = "D2D0CE"

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, size=11, color="000000"): return Font(name="Calibri", size=size, bold=bold, color=color)
def _hdr_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)
def _border(): return _hdr_border()   # kept for legacy callers
def _align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=True if wrap else False)

_INV_TYP_MAP = {
    "R": "Regular",
    "SEZWP": "SEZ supplies with payment of tax",
    "SEZWOP": "SEZ supplies without payment of tax",
    "DE": "Deemed exports"
}

_MONTH_MAP = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December"
}

# =============================================================================
# Header Definitions
# =============================================================================

B2B_HDR5 = ["GSTIN of supplier", "Trade/Legal name of the Supplier", "Invoice details", None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/IFF/GSTR-1A/5 Filing Status", "GSTR-1/IFF/GSTR-1A/5 Filing Date", "GSTR-1/IFF/GSTR-1A/5 Filing Period", "GSTR-3B Filing Status", "Amendment made, if any", "Tax Period in which Amended", "Effective date of cancellation", "Source", "IRN", "IRN date"]
B2B_HDR6 = [None, None, "Invoice number", "Invoice type", "Invoice Date", "Invoice Value (₹)", None, None, None, None, "Integrated Tax  (₹)", "Central Tax (₹)", "State/UT tax (₹)", "Cess  (₹)", None, None, None, None, None, None, None, None, None, None]

B2BA_HDR5 = ["Original details", None, "Revised details", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]
B2BA_HDR6 = ["Invoice number", "Invoice Date", "GSTIN of Supplier", "Trade/Legal name of the supplier", "Invoice details", None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/IFF/GSTR-1A/5 Filing Status", "GSTR-1/IFF/GSTR-1A/5 Filing Date", "GSTR-1/IFF/GSTR-1A/5 Filing Period", "GSTR-3B Filing Status", "Effective date of cancellation", "Amendment made, if any", "Original tax period in which reported "]

CDN_HDR5 = ["GSTIN of Supplier", "Trade/Legal name of the supplier", "Credit note/Debit note details   \n", None, None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/IFF/GSTR-1A/5 Filing Status", "GSTR-1/IFF/GSTR-1A/5 Filing Date", "GSTR-1/IFF/GSTR-1A/5 Filing Period", "GSTR-3B Filing Status", "Amendment made, if any", "Tax Period in which Amended", "Effective date of cancellation", "Source", "IRN", "IRN date"]
CDN_HDR6 = [None, None, "Note type", "Note number", "Note Supply type ", "Note  date", "Note Value (₹)", None, None, None, None, "Integrated Tax (₹)", "Central Tax (₹)", "State Tax (₹)", "Cess Amount (₹)", None, None, None, None, None, None, None, None, None, None]

CDNRA_HDR5 = ["Original details", None, None, "Revised details", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]
CDNRA_HDR6 = ["Note type", "Note Number", "Note date", "GSTIN of Supplier", "Trade/Legal name of the supplier", "Credit note/Debit note details", None, None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/IFF/GSTR-1A/5 Filing Status", "GSTR-1/IFF/GSTR-1A/5 Filing Date", "GSTR-1/IFF/GSTR-1A/5 Filing Period", "GSTR-3B Filing Status", "Amendment made, if any", "Tax Period in which reported earlier", "Effective date of cancellation"]

# ── Yearly Consolidated Headers ─────────────────────────────────────────────
B2B_Y_HDR5 = ["GSTIN of supplier", "Trade/Legal name of the Supplier", "Invoice details", None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/5 Filing Status", "GSTR-1/5 Filing Date", "GSTR-1/5 Filing Period", "GSTR-3B Filing Status", "Amendment made, if any", "Tax Period in which Amended", "Effective date of cancellation", "Period"]
B2B_Y_HDR6 = [None, None, "Invoice number", "Invoice type", "Invoice Date", "Invoice Value (₹)", None, None, None, None, "Integrated Tax  (₹)", "Central Tax (₹)", "State/UT tax (₹)", "Cess  (₹)", None, None, None, None, None, None, None, None]

B2BA_Y_HDR5 = ["Original details", None, "Revised details", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "Period"]
B2BA_Y_HDR6 = ["Invoice number", "Invoice Date", "GSTIN of Supplier", "Trade/Legal name of the supplier", "Invoice details", None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/5 Filing Status", "GSTR-1/5 Filing Date", "GSTR-1/5 Filing Period", "GSTR-3B Filing Status", "Effective date of cancellation", "Amendment made, if any", "Original tax period in which reported ", None]
B2BA_Y_HDR7 = [None, None, None, None, "Invoice type", "Invoice number", "Invoice Date", "Invoice Value (₹)", None, None, None, None, "Integrated Tax  (₹)", "Central Tax (₹)", "State/UT tax (₹)", "Cess  (₹)", None, None, None, None, None, None, None, None]

CDNR_Y_HDR5 = ["GSTIN of Supplier", "Trade/Legal name of the supplier", "Credit note/Debit note details   \n", None, None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/5 Filing Status", "GSTR-1/5 Filing Date", "GSTR-1/5 Filing Period", "GSTR-3B Filing Status", "Amendment made, if any", "Tax Period in which Amended", "Effective date of cancellation", "Period"]
CDNR_Y_HDR6 = [None, None, "Note type", "Note number", "Note Supply type ", "Note  date", "Note Value (₹)", None, None, None, None, "Integrated Tax (₹)", "Central Tax (₹)", "State Tax (₹)", "Cess Amount (₹)", None, None, None, None, None, None, None, None]

CDNRA_Y_HDR5 = ["Original details", None, None, "Revised details", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "Period"]
CDNRA_Y_HDR6 = ["Note type", "Note Number", "Note date", "GSTIN of Supplier", "Trade/Legal name of the supplier", "Credit note/Debit note details", None, None, None, None, "Place of supply", "Supply Attract Reverse Charge", "Rate (%)", "Taxable Value (₹)", "Tax Amount", None, None, None, "GSTR-1/5 Filing Status", "GSTR-1/5 Filing Date", "GSTR-1/5 Filing Period", "GSTR-3B Filing Status", "Amendment made, if any", "Tax Period in which reported earlier", "Effective date of cancellation", None]
CDNRA_Y_HDR7 = [None, None, None, None, None, "Note type", "Note number", "Note Supply type ", "Note  date", "Note Value (₹)", None, None, None, None, "Integrated Tax (₹)", "Central Tax (₹)", "State Tax (₹)", "Cess Amount (₹)", None, None, None, None, None, None, None, None]

# ── TDS ──────────────────────────────────────────────────────────────────────
TDS_HDR5 = ["GSTIN of Deductor", "Deductor's Name", "Tax period of GSTR 7",
             "Taxable Value (₹)", "Amount of Tax Deducted", None, None]
TDS_HDR6 = [None, None, None, None,
             "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)"]
TDS_Y_HDR5 = ["GSTIN of Deductor", "Deductor's Name", "Tax period of GSTR 7",
               "Taxable Value (₹)", "Amount of Tax Deducted", None, None, "Period"]
TDS_Y_HDR6 = [None, None, None, None,
               "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", None]

# ── TDSA ─────────────────────────────────────────────────────────────────────
TDSA_HDR5 = ["GSTIN of Deductor", "Deductor's Name",
              "Tax period of original GSTR 7", "Tax period of amended GSTR 7",
              "Revised Taxable Value (₹)", "Revised Amount of Tax Deducted", None, None]
TDSA_HDR6 = [None, None, None, None, None,
              "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)"]
TDSA_Y_HDR5 = ["GSTIN of Deductor", "Deductor's Name",
                "Tax period of original GSTR 7", "Tax period of amended GSTR 7",
                "Revised Taxable Value (₹)", "Revised Amount of Tax Deducted", None, None, "Period"]
TDSA_Y_HDR6 = [None, None, None, None, None,
                "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", None]

# ── TCS ──────────────────────────────────────────────────────────────────────
TCS_HDR5 = ["GSTIN of E-com. Operator", "E-com. Operator's Name",
             "Tax period of GSTR 8", "Gross Value of Supplies (₹)",
             "Value of Supplies Returned (₹)", "Net Amount Liable for TCS (₹)",
             "Total TCS Amount", None, None]
TCS_HDR6 = [None, None, None, None, None, None,
             "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)"]
TCS_Y_HDR5 = ["GSTIN of E-com. Operator", "E-com. Operator's Name",
               "Tax period of GSTR 8", "Gross Value of Supplies (₹)",
               "Value of Supplies Returned (₹)", "Net Amount Liable for TCS (₹)",
               "Total TCS Amount", None, None, "Period"]
TCS_Y_HDR6 = [None, None, None, None, None, None,
               "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", None]

# ── ISD ──────────────────────────────────────────────────────────────────────
ISD_HDR5 = ["Eligibility of ITC", "GSTIN of ISD", "Trade/Legal name of the ISD",
             "ISD Document type", "ISD Invoice number", "ISD Invoice date",
             "ISD credit note number", "ISD credit note date",
             "Original Invoice Number", "Original invoice date",
             "Input tax distribution by ISD", None, None, None, "ISD GSTR-6 Filing Status"]
ISD_HDR6 = [None, None, None, None, None, None, None, None, None, None,
             "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Cess (₹)", None]
ISD_Y_HDR5 = ["Eligibility of ITC", "GSTIN of ISD", "Trade/Legal name of the ISD",
               "ISD Document type", "ISD Invoice number", "ISD Invoice date",
               "ISD credit note number", "ISD credit note date",
               "Original Invoice Number", "Original invoice date",
               "Input tax distribution by ISD", None, None, None,
               "ISD GSTR-6 Filing Status", "Period"]
ISD_Y_HDR6 = [None, None, None, None, None, None, None, None, None, None,
               "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)", "Cess (₹)", None, None]

# ── ISDA ─────────────────────────────────────────────────────────────────────
ISDA_HDR5  = ["Original Details", None, None, "Revised Details", None, None, None,
               None, None, None, None, None, None, None, None, None, None, None]
ISDA_HDR6  = ["ISD Document type", "Document Number", "Document date",
               "Eligibility of ITC", "GSTIN of ISD", "Trade/Legal name of the ISD",
               "ISD Document type", "ISD Invoice number", "ISD Invoice date",
               "ISD credit note number", "ISD credit note date",
               "Original Invoice Number", "Original invoice date",
               "Input tax distribution by ISD", None, None, None, "ISD GSTR-6 Filing Status"]
ISDA_HDR7  = [None, None, None, None, None, None, None, None, None, None, None,
               None, None, "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)",
               "Cess (₹)", None]
ISDA_Y_HDR5 = ["Original Details", None, None, "Revised Details", None, None, None,
                None, None, None, None, None, None, None, None, None, None, None, "Period"]
ISDA_Y_HDR6 = ["ISD Document type", "Document Number", "Document date",
                "Eligibility of ITC", "GSTIN of ISD", "Trade/Legal name of the ISD",
                "ISD Document type", "ISD Invoice number", "ISD Invoice date",
                "ISD credit note number", "ISD credit note date",
                "Original Invoice Number", "Original invoice date",
                "Input tax distribution by ISD", None, None, None,
                "ISD GSTR-6 Filing Status", None]
ISDA_Y_HDR7 = [None, None, None, None, None, None, None, None, None, None, None,
                None, None, "Integrated Tax (₹)", "Central Tax (₹)", "State/UT Tax (₹)",
                "Cess (₹)", None, None]

# ── IMPG ─────────────────────────────────────────────────────────────────────
IMPG_HDR5 = ["Reference date (ICEGATE)", None, "Bill of Entry Details", None, None,
              "Amount of Tax (₹)", None, "Amended (Yes)"]
IMPG_HDR6 = [None, "Port code", "Number", "Date", "Taxable Value (₹)",
              "Integrated Tax (₹)", "Cess (₹)", None]
IMPG_Y_HDR5 = ["Reference date (ICEGATE)", None, "Bill of Entry Details", None, None,
                "Amount of Tax (₹)", None, "Amended (Yes)", "Period"]
IMPG_Y_HDR6 = [None, "Port code", "Number", "Date", "Taxable Value (₹)",
                "Integrated Tax (₹)", "Cess (₹)", None, None]

# ── IMPG SEZ ─────────────────────────────────────────────────────────────────
IMPGSEZ_HDR5 = ["GSTIN of Supplier", "Trade/Legal name", "Reference date (ICEGATE)", None,
                 "Bill of Entry Details", None, None, "Amount of Tax (₹)", None, "Amended (Yes)"]
IMPGSEZ_HDR6 = [None, None, None, "Port code", "Number", "Date", "Taxable Value (₹)",
                 "Integrated Tax (₹)", "Cess (₹)", None]
IMPGSEZ_Y_HDR5 = ["GSTIN of Supplier", "Trade/Legal name", "Reference date (ICEGATE)", None,
                   "Bill of Entry Details", None, None, "Amount of Tax (₹)", None,
                   "Amended (Yes)", "Period"]
IMPGSEZ_Y_HDR6 = [None, None, None, "Port code", "Number", "Date", "Taxable Value (₹)",
                   "Integrated Tax (₹)", "Cess (₹)", None, None]

# =============================================================================
# Helper Functions
# =============================================================================

def _fmt_date(s):
    if not s: return ""
    return str(s).strip()

def _dt(s):
    if not s: return None
    import datetime
    s = str(s).strip()
    for fmt in ("%d-%m-%Y", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except:
            continue
    return s

def _parse_section(data, key):
    # GSTR-2A JSON often uses "cdn" for CDNR and "cdna" for CDNRA
    key_map = {"cdnr": "cdn", "cdnra": "cdna", "impg sez": "impgsez"}
    search_key = key_map.get(key.lower(), key.lower())
    
    # Try top-level keys first, then look inside 'sections' wrapper
    val = data.get(search_key.lower()) or data.get(search_key)
    if not val:
        secs = data.get("sections", {})
        val = secs.get(search_key) or secs.get(search_key.upper()) or secs.get(search_key.lower())
    
    if not val: return []

    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            # If parsed result is a dict with the key inside (nested), extract it
            if isinstance(parsed, dict):
                inner = parsed.get(search_key.lower()) or parsed.get(search_key.upper()) or parsed.get(search_key)
                if isinstance(inner, list): return inner
                # Sometimes it's a list directly in the dict without matching the search_key
                for v in parsed.values():
                    if isinstance(v, list): return v
            return parsed if isinstance(parsed, list) else []
        except: return []

    if isinstance(val, dict):
        # Handle dict case where data is one level deeper
        for k in [search_key.lower(), search_key.upper(), search_key]:
            if isinstance(val.get(k), list): return val[k]
        for v in val.values():
            if isinstance(v, list): return v
            
    return val if isinstance(val, list) else []

def _write_portal_sheet(ws, title_row1, title_row4, hdr5, hdr6, rows, num_cols=None, hdr7=None):
    max_cols = max(len(hdr5), len(hdr6) if hdr6 else 0, len(hdr7) if hdr7 else 0)
    
    # Title row 1 — dark navy, Calibri 22, white, merged
    ws.cell(1, 1, title_row1).font = Font(name="Calibri", size=22, color=_WHITE)
    ws.cell(1, 1).fill = _fill(_NAVY)
    ws.cell(1, 1).alignment = _align("left", "center")
    ws.row_dimensions[1].height = 15
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=max_cols)

    # Section title row 4 — light grey fill, bold 11pt dark text
    ws.cell(4, 1, "          " + title_row4).font = Font(name="Calibri", size=11, bold=True, color="000000")
    ws.cell(4, 1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(4, 1).alignment = _align("left", "center")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max_cols)
    ws.row_dimensions[4].height = 15

    # Column header rows 5/6/7 — dark navy, bold 9pt, white
    bg  = _fill(_NAVY)
    bdr = _hdr_border()
    header_rows = [5, 6]
    if hdr7: header_rows.append(7)

    for r in header_rows:
        ws.row_dimensions[r].height = 30
        h_data = hdr5 if r == 5 else (hdr6 if r == 6 else hdr7)
        for c in range(1, max_cols + 1):
            val = h_data[c-1] if c <= len(h_data) else None
            cell = ws.cell(r, c, val)
            cell.fill = bg
            cell.border = bdr
            cell.font = Font(name="Calibri", size=9, bold=True, color=_WHITE)
            cell.alignment = _align(h="center", wrap=True)

    # Merging logic (Reference matching)
    if "registered persons" in title_row4:
        # B2B
        ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
        ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)
        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=6)
        for c in [7, 8, 9, 10]: ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
        ws.merge_cells(start_row=5, start_column=11, end_row=5, end_column=14)
        for c in range(15, max_cols + 1): ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
    elif "Debit/Credit notes (Original)" in title_row4:
        # CDNR
        ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
        ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)
        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=7)
        for c in [8, 9, 10, 11]: ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
        ws.merge_cells(start_row=5, start_column=12, end_row=5, end_column=15)
        for c in range(16, max_cols + 1): ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
    elif "TDS Credit" in title_row4 or "TCS Credit" in title_row4:
        # TDS / TCS: single-value cols span rows 5-6; tax group spans multiple cols row 5
        tax_start = 5 if "TDS Credit" in title_row4 else 7
        tax_end   = 7 if "TDS Credit" in title_row4 else 9
        for c in range(1, tax_start):
            ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
        ws.merge_cells(start_row=5, start_column=tax_start, end_row=5, end_column=tax_end)
        if max_cols > tax_end:
            ws.merge_cells(start_row=5, start_column=max_cols, end_row=6, end_column=max_cols)
    elif "Amendments to TDS" in title_row4:
        # TDSA: cols 1-5 span rows 5-6; tax group cols 6-8
        for c in range(1, 6):
            ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
        ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=8)
        if max_cols > 8:
            ws.merge_cells(start_row=5, start_column=max_cols, end_row=6, end_column=max_cols)
    elif "ISD Credits" in title_row4:
        # ISD: cols 1-10 span rows 5-6; tax group cols 11-14; col 15 spans
        for c in list(range(1, 11)) + [15]:
            ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
        ws.merge_cells(start_row=5, start_column=11, end_row=5, end_column=14)
        if max_cols > 15:
            ws.merge_cells(start_row=5, start_column=max_cols, end_row=6, end_column=max_cols)
    elif "Bill of Entry" in title_row4 and "SEZ" not in title_row4:
        # IMPG: col 1 spans rows 5-6 (with col 2); group for BoE cols 3-5; tax 6-7; cols 8+ span
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=5)
        ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=7)
        ws.merge_cells(start_row=5, start_column=8, end_row=6, end_column=8)
        if max_cols > 8:
            ws.merge_cells(start_row=5, start_column=max_cols, end_row=6, end_column=max_cols)
    elif "SEZ" in title_row4 and "Bill of Entry" in title_row4:
        # IMPG SEZ: cols 1-2 span; refdate group 3-4; BoE group 5-7; tax 8-9; col 10+ span
        for c in [1, 2]:
            ws.merge_cells(start_row=5, start_column=c, end_row=6, end_column=c)
        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=4)
        ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=7)
        ws.merge_cells(start_row=5, start_column=8, end_row=5, end_column=9)
        ws.merge_cells(start_row=5, start_column=10, end_row=6, end_column=10)
        if max_cols > 10:
            ws.merge_cells(start_row=5, start_column=max_cols, end_row=6, end_column=max_cols)
    elif "Amendments" in title_row4 and hdr7:
        # B2BA / CDNRA (3-row headers)
        ws.merge_cells(start_row=5, start_column=1, end_row=7, end_column=1)
        ws.merge_cells(start_row=5, start_column=2, end_row=7, end_column=2)
        # Handle "Revised details" span
        if "previously uploaded invoices" in title_row4:
            # B2BA
            ws.merge_cells(start_row=5, start_column=3, end_row=7, end_column=3)
            ws.merge_cells(start_row=5, start_column=4, end_row=7, end_column=4)
            ws.merge_cells(start_row=5, start_column=5, end_row=6, end_column=max_cols-1)
            ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=8) # Invoice details nested
            for c in range(9, max_cols): ws.merge_cells(start_row=7, start_column=c, end_row=7, end_column=c)
        else:
            # CDNRA
            ws.merge_cells(start_row=5, start_column=3, end_row=7, end_column=3)
            ws.merge_cells(start_row=5, start_column=4, end_row=7, end_column=4)
            ws.merge_cells(start_row=5, start_column=5, end_row=7, end_column=5)
            ws.merge_cells(start_row=5, start_column=6, end_row=6, end_column=max_cols-1)

    # Data Rows — portal style: no fill, no border, Calibri 11
    start_row = 8 if hdr7 else 7
    num_cols = num_cols or set()
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.font = _font()
            if c_idx in num_cols or isinstance(val, (int, float)):
                cell.alignment = _align(h="right")
                cell.number_format = "#,##0.00"
            elif hasattr(val, "year"):
                cell.number_format = "DD-MM-YYYY"
                cell.alignment = _align(h="center")
            else:
                cell.alignment = _align(h="left", wrap=True)
    ws.freeze_panes = f"A{start_row}"
    for c in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

# =============================================================================
# Flatteners
# =============================================================================

def _flatten_b2b(b2b_list):
    rows = []
    if not b2b_list: return []
    for party in b2b_list:
        gstin = _s(party.get("ctin"))
        name = _s(party.get("trdnm") or party.get("lgl_nm") or "")
        fldtr1 = _dt(party.get("fldtr1"))
        flprdr1 = _s(party.get("flprdr1"))
        cfs = "Yes" if party.get("cfs") == "Y" else "No"
        cfs3b = "Yes" if party.get("cfs3b") == "Y" else "No"
        dtcancel = _dt(party.get("dtcancel"))
        for inv in party.get("inv", []):
            inum, typ, idt, val, pos, rev = _s(inv.get("inum")), _INV_TYP_MAP.get(_s(inv.get("inv_typ")), _s(inv.get("inv_typ"))), _dt(inv.get("idt")), _n(inv.get("val")), _state_name(_s(inv.get("pos"))), "Yes" if _s(inv.get("rchrg")) == "Y" else "No"
            srctyp, irn, irndt = _s(inv.get("srctyp") or ""), _s(inv.get("irn") or ""), _dt(inv.get("irngendate"))
            itms = inv.get("itms", [])
            if not itms:
                rows.append([gstin, name, inum, typ, idt, val, pos, rev, 0, _n(inv.get("txval")), _n(inv.get("iamt")), _n(inv.get("camt")), _n(inv.get("samt")), _n(inv.get("csamt")), cfs, fldtr1, flprdr1, cfs3b, "", "", dtcancel])
            else:
                for itm in itms:
                    d = itm.get("itm_det", itm)
                    rows.append([gstin, name, inum, typ, idt, val, pos, rev, _n(d.get("rt")), _n(d.get("txval")), _n(d.get("iamt")), _n(d.get("camt")), _n(d.get("samt")), _n(d.get("csamt")), cfs, fldtr1, flprdr1, cfs3b, "", "", dtcancel])
    return rows

def _flatten_b2ba(b2ba_list):
    rows = []
    if not b2ba_list: return []
    for party in b2ba_list:
        gstin = _s(party.get("ctin"))
        name = _s(party.get("trdnm") or party.get("lgl_nm") or "")
        fldtr1, flprdr1 = _dt(party.get("fldtr1")), _s(party.get("flprdr1"))
        cfs, cfs3b = ("Yes" if party.get("cfs") == "Y" else "No"), ("Yes" if party.get("cfs3b") == "Y" else "No")
        dtcancel = _dt(party.get("dtcancel"))
        for inv in party.get("inv", []):
            oinum, oidt = _s(inv.get("oinum")), _dt(inv.get("oidt"))
            inum, typ, idt, val, pos, rev = _s(inv.get("inum")), _INV_TYP_MAP.get(_s(inv.get("inv_typ")), _s(inv.get("inv_typ"))), _dt(inv.get("idt")), _n(inv.get("val")), _state_name(_s(inv.get("pos"))), "Yes" if _s(inv.get("rchrg")) == "Y" else "No"
            itms = inv.get("itms", [])
            for itm in (itms or [{"itm_det": {"rt": 0, "txval": _n(inv.get("txval")), "iamt": _n(inv.get("iamt")), "camt": _n(inv.get("camt")), "samt": _n(inv.get("samt")), "csamt": _n(inv.get("csamt"))}}]):
                d = itm.get("itm_det", itm)
                rows.append([oinum, oidt, gstin, name, typ, inum, idt, val, pos, rev, _n(d.get("rt")), _n(d.get("txval")), _n(d.get("iamt")), _n(d.get("camt")), _n(d.get("samt")), _n(d.get("csamt")), cfs, fldtr1, flprdr1, cfs3b, "", "", dtcancel])
    return rows

def _flatten_cdn(cdn_list):
    rows = []
    if not cdn_list: return []
    for party in cdn_list:
        gstin, name = _s(party.get("ctin")), _s(party.get("trdnm") or party.get("lgl_nm") or "")
        fldtr1, flprdr1, cfs, cfs3b, dtcancel = _dt(party.get("fldtr1")), _s(party.get("flprdr1")), ("Yes" if party.get("cfs") == "Y" else "No"), ("Yes" if party.get("cfs3b") == "Y" else "No"), _dt(party.get("dtcancel"))
        for note in party.get("nt", []):
            ntyp = "Credit Note" if _s(note.get("ntty")) == "C" else "Debit Note"
            nno = _s(note.get("nt_num")) or _s(note.get("ntnum"))
            ndt = _dt(note.get("nt_dt")) or _dt(note.get("ndt"))
            nval, pos, rev = _n(note.get("val")), _state_name(_s(note.get("pos"))), ("Yes" if _s(note.get("rchrg")) == "Y" else "No")
            typ = _INV_TYP_MAP.get(_s(note.get("inv_typ")), _s(note.get("inv_typ")))
            itms = note.get("itms", [])
            for itm in (itms or [{"itm_det": {"rt": 0, "txval": _n(note.get("txval")), "iamt": _n(note.get("iamt")), "camt": _n(note.get("camt")), "samt": _n(note.get("samt")), "csamt": _n(note.get("csamt"))}}]):
                d = itm.get("itm_det", itm)
                rows.append([gstin, name, ntyp, nno, typ, ndt, nval, pos, rev, _n(d.get("rt")), _n(d.get("txval")), _n(d.get("iamt")), _n(d.get("camt")), _n(d.get("samt")), _n(d.get("csamt")), cfs, fldtr1, flprdr1, cfs3b, "", "", dtcancel])
    return rows

def _flatten_cdnra(cdnra_list):
    rows = []
    if not cdnra_list: return []
    for party in cdnra_list:
        gstin, name = _s(party.get("ctin")), _s(party.get("trdnm") or party.get("lgl_nm") or "")
        fldtr1, flprdr1, cfs, cfs3b, dtcancel = _dt(party.get("fldtr1")), _s(party.get("flprdr1")), ("Yes" if party.get("cfs") == "Y" else "No"), ("Yes" if party.get("cfs3b") == "Y" else "No"), _dt(party.get("dtcancel"))
        for note in party.get("nt", []):
            ontyp = ("Credit Note" if _s(note.get("ontty")) == "C" else "Debit Note")
            onno, ondt = _s(note.get("ont_num")) or _s(note.get("ontnum")), _dt(note.get("ont_dt")) or _dt(note.get("ondt"))
            ntyp = ("Credit Note" if _s(note.get("ntty")) == "C" else "Debit Note")
            nno, ndt = _s(note.get("nt_num")) or _s(note.get("ntnum")), _dt(note.get("nt_dt")) or _dt(note.get("ndt"))
            nval, pos, rev = _n(note.get("val")), _state_name(_s(note.get("pos"))), ("Yes" if _s(note.get("rchrg")) == "Y" else "No")
            typ = _INV_TYP_MAP.get(_s(note.get("inv_typ")), _s(note.get("inv_typ")))
            itms = note.get("itms", [])
            for itm in (itms or [{"itm_det": {"rt": 0, "txval": _n(note.get("txval")), "iamt": _n(note.get("iamt")), "camt": _n(note.get("camt")), "samt": _n(note.get("samt")), "csamt": _n(note.get("csamt"))}}]):
                d = itm.get("itm_det", itm)
                rows.append([ontyp, onno, ondt, gstin, name, ntyp, nno, typ, ndt, nval, pos, rev, _n(d.get("rt")), _n(d.get("txval")), _n(d.get("iamt")), _n(d.get("camt")), _n(d.get("samt")), _n(d.get("csamt")), cfs, fldtr1, flprdr1, cfs3b, "", "", dtcancel])
    return rows

def _flatten_tds(tds_list):
    rows = []
    if not tds_list: return []
    for e in tds_list:
        rows.append([
            _s(e.get("ctin")),
            _s(e.get("trdnm") or e.get("lgl_nm") or ""),
            _s(e.get("rtn_prd") or ""),
            _n(e.get("txval")),
            _n(e.get("iamt")), _n(e.get("camt")), _n(e.get("samt")),
        ])
    return rows

def _flatten_tdsa(tdsa_list):
    rows = []
    if not tdsa_list: return []
    for e in tdsa_list:
        rows.append([
            _s(e.get("ctin")),
            _s(e.get("trdnm") or e.get("lgl_nm") or ""),
            _s(e.get("o_rtn_prd") or e.get("ortn_prd") or ""),
            _s(e.get("rtn_prd") or ""),
            _n(e.get("txval")),
            _n(e.get("iamt")), _n(e.get("camt")), _n(e.get("samt")),
        ])
    return rows

def _flatten_tcs(tcs_list):
    rows = []
    if not tcs_list: return []
    for e in tcs_list:
        rows.append([
            _s(e.get("ctin")),
            _s(e.get("trdnm") or e.get("lgl_nm") or ""),
            _s(e.get("rtn_prd") or ""),
            _n(e.get("grsup")), _n(e.get("suppret")), _n(e.get("netamt")),
            _n(e.get("iamt")), _n(e.get("camt")), _n(e.get("samt")),
        ])
    return rows

def _flatten_isd(isd_list):
    rows = []
    if not isd_list: return []
    for e in isd_list:
        rows.append([
            _s(e.get("elgITC") or e.get("elg") or ""),
            _s(e.get("ctin")),
            _s(e.get("trdnm") or e.get("lgl_nm") or ""),
            _s(e.get("doctyp") or ""),
            _s(e.get("num") or ""),
            _dt(e.get("dt")),
            _s(e.get("cnum") or ""),
            _dt(e.get("cdt")),
            _s(e.get("oinum") or ""),
            _dt(e.get("oidt")),
            _n(e.get("iamt")), _n(e.get("camt")), _n(e.get("samt")), _n(e.get("csamt")),
            _s(e.get("fsts") or ""),
        ])
    return rows

def _flatten_isda(isda_list):
    rows = []
    if not isda_list: return []
    for e in isda_list:
        rows.append([
            _s(e.get("o_doctyp") or e.get("odoctyp") or ""),
            _s(e.get("o_num") or e.get("onum") or ""),
            _dt(e.get("o_dt") or e.get("odt")),
            _s(e.get("elgITC") or e.get("elg") or ""),
            _s(e.get("ctin")),
            _s(e.get("trdnm") or e.get("lgl_nm") or ""),
            _s(e.get("doctyp") or ""),
            _s(e.get("num") or ""),
            _dt(e.get("dt")),
            _s(e.get("cnum") or ""),
            _dt(e.get("cdt")),
            _s(e.get("oinum") or ""),
            _dt(e.get("oidt")),
            _n(e.get("iamt")), _n(e.get("camt")), _n(e.get("samt")), _n(e.get("csamt")),
            _s(e.get("fsts") or ""),
        ])
    return rows

def _flatten_impg(impg_list):
    rows = []
    if not impg_list: return []
    for e in impg_list:
        rows.append([
            _dt(e.get("refdt")),
            _s(e.get("portcd") or ""),
            _s(e.get("benum") or ""),
            _dt(e.get("bedt")),
            _n(e.get("txval")),
            _n(e.get("iamt")), _n(e.get("csamt")),
            "Yes" if _s(e.get("isamd")) == "Y" else "",
        ])
    return rows

def _flatten_impgsez(impgsez_list):
    rows = []
    if not impgsez_list: return []
    for e in impgsez_list:
        rows.append([
            _s(e.get("ctin")),
            _s(e.get("trdnm") or e.get("lgl_nm") or ""),
            _dt(e.get("refdt")),
            _s(e.get("portcd") or ""),
            _s(e.get("benum") or ""),
            _dt(e.get("bedt")),
            _n(e.get("txval")),
            _n(e.get("iamt")), _n(e.get("csamt")),
            "Yes" if _s(e.get("isamd")) == "Y" else "",
        ])
    return rows

# =============================================================================
# Summary and Metadata
# =============================================================================

def _write_readme(wb, profile=None, data=None):
    ws = wb.create_sheet("Read me", 0)
    p = profile or {}
    gstin   = p.get("gstin")  or (data or {}).get("gstin") or ""
    lgl_nm  = p.get("lgl_nm") or p.get("bname") or ""
    trd_nm  = p.get("trdnm")  or p.get("trade_name") or lgl_nm
    fy      = p.get("fy")     or ""
    period  = p.get("period") or fy

    bdr  = _hdr_border()
    bold = Font(name="Calibri", size=11, bold=True)
    norm = Font(name="Calibri", size=11)

    # ── Metadata block (rows 2-4) — portal exact layout ──────────────────────
    meta = [
        (2, "Taxpayer's GSTIN", gstin,   "Tax period",         period),
        (3, "Legal name",       lgl_nm,  "Financial year",     fy),
        (4, "Trade name",       trd_nm,  "Date of generation", datetime.now().strftime("%d-%m-%Y")),
    ]
    for row, l1, v1, l2, v2 in meta:
        for c in range(2, 7): ws.cell(row, c).border = bdr
        c1 = ws.cell(row, 2, l1); c1.font = bold; c1.border = bdr
        c2 = ws.cell(row, 3, v1); c2.font = norm; c2.border = bdr
        c3 = ws.cell(row, 4, l2); c3.font = bold; c3.border = bdr
        c4 = ws.cell(row, 5, v2); c4.font = norm; c4.border = bdr
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

    # ── Section heading row 6 ─────────────────────────────────────────────────
    tc = ws.cell(6, 2, "GSTR-2A Data Entry Instructions")
    tc.font = Font(name="Calibri", size=11, bold=True, color="0070C0")

    # ── Column header row 7 ───────────────────────────────────────────────────
    for c, val in enumerate(["Worksheet Name", "GSTR-2A Table Reference", "Field Name", "Help Instruction"], 2):
        cell = ws.cell(7, c, val)
        cell.font  = Font(name="Calibri", size=11, bold=True, color=_WHITE)
        cell.fill  = _fill(_NAVY)
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Help instruction data (mirrors portal Read me exactly) ────────────────
    _NA = "GSTIN of the supplier"
    _TN = "Trade name of the supplier will be displayed. If trade name is not available then legal name of the supplier"
    _RC = "Supply attract reverse charge divided into two types:\nY- Purchases attract reverse charge\nN- Purchases don't attract reverse charge"
    _IT = ("Invoice type can be derived based on the following types:\n"
           "R- Regular (Other than SEZ supplies and Deemed exports supplies)\n"
           "SEZWP- SEZ supplies with payment of tax\n"
           "SEZWOP- SEZ supplies without payment of tax\n"
           "DE- Deemed exports")
    _FS = "If the counter party filed the GSTR 1/5 return then status will be Y else N"
    _EC = "If your purchases are through E-commerce operator then E-commerce operator GSTIN will be displayed"
    _POS = "Place of supply shall be the place where goods are supplied or services are provided (As determined by the supplier)"

    rows = [
        # B2B
        ("B2B","Taxable inward supplies received from registered persons","GSTIN of supplier",        _NA),
        ("",   "",                                                         "Trade/Legal name",         _TN),
        ("",   "",                                                         "Invoice number",           "Invoice number"),
        ("",   "",                                                         "Invoice type",             _IT),
        ("",   "",                                                         "Invoice date",             "Invoice date format shall be DD-MM-YYYY"),
        ("",   "",                                                         "Invoice value",            "Invoice value (in rupees)"),
        ("",   "",                                                         "Place of supply",          _POS),
        ("",   "",                                                         "Supply attract Reverse charge", _RC),
        ("",   "",                                                         "Rate(%)",                  "Applicable Rate of tax"),
        ("",   "",                                                         "Taxable value",            "Taxable value (In rupees)"),
        ("",   "",                                                         "Integrated Tax",           "Integrated Tax amount (in rupees)"),
        ("",   "",                                                         "Central Tax",              "Central Tax amount (in rupees)"),
        ("",   "",                                                         "State/UT tax",             "State/UT tax amount (in rupees)"),
        ("",   "",                                                         "Cess",                     "Cess amount (In rupees)"),
        ("",   "",                                                         "E-commerce GSTIN",         _EC),
        ("",   "",                                                         "GSTR-1/5 Filing Status",   _FS),
        # B2BA
        ("B2BA","Amendments to previously uploaded invoices by supplier","Invoice date (Original details)", "Original invoice date (Date format shall be DD-MM-YYYY)"),
        ("",    "",                                                        "GSTIN of Supplier",        _NA),
        ("",    "",                                                        "Trade/Legal name",         _TN),
        ("",    "",                                                        "Invoice number",           "Revised Invoice number"),
        ("",    "",                                                        "Invoice type",             _IT),
        ("",    "",                                                        "Invoice date",             "Invoice date format shall be DD-MM-YYYY"),
        ("",    "",                                                        "Invoice value",            "Invoice value (in rupees)"),
        ("",    "",                                                        "Place of supply",          _POS),
        ("",    "",                                                        "Supply attract Reverse charge", _RC),
        ("",    "",                                                        "Rate(%)",                  "Applicable Rate of tax"),
        ("",    "",                                                        "Taxable value",            "Taxable value (In rupees)"),
        ("",    "",                                                        "Integrated Tax",           "Integrated Tax amount (in rupees)"),
        ("",    "",                                                        "Central Tax",              "Central Tax amount (in rupees)"),
        ("",    "",                                                        "State/UT tax",             "State/UT tax amount (in rupees)"),
        ("",    "",                                                        "Cess",                     "Cess amount (In rupees)"),
        ("",    "",                                                        "E-commerce GSTIN",         _EC),
        ("",    "",                                                        "GSTR-1/5 Filing Status",   _FS),
        # CDNR
        ("CDNR","Debit/Credit notes (Original)","GSTIN of supplier",      _NA),
        ("",    "",                              "Trade/Legal name",       _TN),
        ("",    "",                              "Document type",          "Document type can be Debit note or credit note"),
        ("",    "",                              "Document number",        "Credit/Debit note number"),
        ("",    "",                              "Document date",          "Credit/Debit note date format shall be DD-MM-YYYY"),
        ("",    "",                              "Reason",                 "Reason for issuing Credit/Debit note"),
        ("",    "",                              "Rate(%)",                "Applicable Rate of tax"),
        ("",    "",                              "Taxable value",          "Taxable value (in rupees)"),
        ("",    "",                              "Integrated Tax",         "Integrated Tax amount (in rupees)"),
        ("",    "",                              "Central Tax",            "Central Tax amount (in rupees)"),
        ("",    "",                              "State/UT tax",           "State/UT tax amount (in rupees)"),
        ("",    "",                              "Cess",                   "Cess amount (in rupees)"),
        ("",    "",                              "E-commerce GSTIN",       _EC),
        ("",    "",                              "Original Debit/Credit note number", "Revised Debit/Credit note number"),
        ("",    "",                              "GSTR-1/5 Filing Status", _FS),
        # CDNRA
        ("CDNRA","Amendments to previously uploaded Credit/Debit notes by supplier","Document date (Original details)", "Original Debit/Credit note date (Note date format shall be DD-MM-YYYY)"),
        ("",     "",                                                                  "GSTIN of Supplier",   _NA),
        ("",     "",                                                                  "Trade/Legal name",    _TN),
        ("",     "",                                                                  "Document type",       "Document type can be Debit note or credit note"),
        ("",     "",                                                                  "Document number",     "Credit/Debit note number"),
        ("",     "",                                                                  "Document Value",      "Debit/Credit note value"),
        ("",     "",                                                                  "Rate(%)",             "Applicable Rate of tax"),
        ("",     "",                                                                  "Taxable value",       "Taxable value (in rupees)"),
        ("",     "",                                                                  "Integrated Tax",      "Integrated Tax amount (in rupees)"),
        ("",     "",                                                                  "Central Tax",         "Central Tax amount (in rupees)"),
        ("",     "",                                                                  "State/UT tax",        "State/UT tax amount (in rupees)"),
        ("",     "",                                                                  "Cess",                "Cess amount (in rupees)"),
        ("",     "",                                                                  "E-commerce GSTIN",    _EC),
        ("",     "",                                                                  "GSTR-1/5 Filing Status", _FS),
        # ISD
        ("ISD","ISD Credit","Eligibility of ITC",          "Eligibility of ITC are two types:\nY- Yes: Taxpayer can claim ITC on such invoice\nN- No: Taxpayer can't claim ITC on such invoice"),
        ("",   "",           "GSTIN of ISD",               "Input Service Distributor GSTIN"),
        ("",   "",           "ISD name",                   "Name of the ISD will be displayed. If trade name is not available then legal name of the ISD"),
        ("",   "",           "ISD-document type",          "ISD document type can be Invoice or Credit note"),
        ("",   "",           "ISD Invoice number",         "ISD invoice number"),
        ("",   "",           "ISD Invoice date",           "ISD Invoice date format shall be DD-MM-YYYY"),
        ("",   "",           "ISD Credit note number",     "ISD-Credit note number"),
        ("",   "",           "ISD Credit note date",       "ISD Credit note date format shall be DD-MM-YYYY"),
        ("",   "",           "Original ISD Invoice number","This is applicable only if ISD document type is 'Credit note'"),
        ("",   "",           "Original ISD Invoice date",  "This is applicable only if ISD document type is 'Credit note'"),
        ("",   "",           "Integrated Tax",             "Integrated Tax amount (in rupees)"),
        ("",   "",           "Central Tax",                "Central Tax amount (in rupees)"),
        ("",   "",           "State/UT tax",               "State/UT tax amount (in rupees)"),
        ("",   "",           "Cess",                       "Cess amount (in rupees)"),
        ("",   "",           "Status",                     "Status will be filed or not filed"),
        # ISDA
        ("ISDA","Amendments to ISD Credits received","ISD document type",  "ISD document type can be Invoice or Credit note"),
        ("",    "",                                   "ISD document number","ISD invoice number"),
        ("",    "",                                   "ISD Invoice date",   "ISD Invoice date format shall be DD-MM-YYYY"),
        ("",    "",                                   "ISD Credit number",  "ISD-Credit note number"),
        ("",    "",                                   "Original ISD Invoice number","This is applicable only if ISD document type is 'Credit note'"),
        ("",    "",                                   "Original ISD Invoice date",  "This is applicable only if ISD document type is 'Credit note'"),
        ("",    "",                                   "Integrated Tax",     "Integrated Tax amount (in rupees)"),
        ("",    "",                                   "Central Tax",        "Central Tax amount (in rupees)"),
        ("",    "",                                   "State/UT tax",       "State/UT tax amount (in rupees)"),
        ("",    "",                                   "Cess",               "Cess amount (in rupees)"),
        ("",    "",                                   "Status",             "Status will be filed or not filed"),
        # TDS
        ("TDS","TDS Credit received","GSTIN of deductor",       "Deductor's GSTIN"),
        ("",   "",                   "Trade/Legal name",         "Trade name of the Deductor will be displayed. If trade name is not available then legal name of the Deductor"),
        ("",   "",                   "Tax period of GSTR-7",     "Tax period in which TDS details uploaded by Deductor"),
        ("",   "",                   "Taxable value",            "Amount on which tax is deducted"),
        ("",   "",                   "IGST",                     "Central TDS amount"),
        ("",   "",                   "CGST",                     "Central TDS amount"),
        ("",   "",                   "State/UT tax",             "State/UT TDS amount"),
        # TDSA
        ("TDSA","Amendments to TDS Credit received","GSTIN of deductor",           "Deductor's GSTIN"),
        ("",    "",                                  "Trade/Legal name",             "Trade name of the Deductor will be displayed. If trade name is not available then legal name of the Deductor"),
        ("",    "",                                  "Tax period of original GSTR-7","Tax period in which TDS details uploaded by Deductor"),
        ("",    "",                                  "Tax period of amended GSTR-7", "Tax period in which TDS details revised by deductor"),
        ("",    "",                                  "IGST",                         "Revised integrated TDS amount"),
        ("",    "",                                  "CGST",                         "Central TDS amount"),
        ("",    "",                                  "State/UT tax",                 "State/UT TDS amount"),
        # TCS
        ("TCS","Details of supplies made through e-commerce operator (TCS)","GSTIN of E-commerce operator",  "E-commerce operator's GSTIN"),
        ("",   "",                                                            "Trade/Legal name",               "Trade name of the E-commerce operator. If trade name is not available then legal name of the E-commerce operator"),
        ("",   "",                                                            "Tax period of GSTR-8",           "Tax period in which TCS is collected previously"),
        ("",   "",                                                            "Gross value of supplies",        "Total value supplies returned include supplies returned by registered and unregistered persons"),
        ("",   "",                                                            "Value of supplies returned",     "Total value supplies returned include supplies returned by registered and unregistered persons"),
        ("",   "",                                                            "Net amount liable for TCS",      "Net amount liable for TCS shall be difference between Gross value of supplies and Value of supplies returned"),
        ("",   "",                                                            "Integrated Tax",                 "Integrated tax amount"),
        ("",   "",                                                            "Central Tax",                    "Central tax amount"),
        ("",   "",                                                            "State/UT tax",                   "State/UT tax amount"),
        ("",   "",                                                            "Type",                           "Rejected by supplier: If TCS details rejected by supplier then those details will be displayed with type as 'Rejected by supplier'\nUploaded by E-commerce: If you accepted TCS details which are not rejected or no action taken by supplier then those details will be displayed with type as 'Uploaded by E-commerce'"),
        # IMPG
        ("IMPG","Import of Goods from Overseas on Bill of Entry","Reference date (ICEGATE)", "Reference date of ICEGATE format shall be DD-MM-YYYY"),
        ("",    "",                                               "Port code",               "Port code should be 6 digit alphanumeric"),
        ("",    "",                                               "Number",                  "Bill of entry no, format shall be 7 digit number"),
        ("",    "",                                               "Date",                    "Bill of entry date format shall be DD-MM-YYYY"),
        ("",    "",                                               "Taxable value",           "Taxable value (in rupees)"),
        ("",    "",                                               "Integrated tax",          "Integrated tax (in rupees)"),
        ("",    "",                                               "Cess",                    "Cess amount (in rupees)"),
        ("",    "",                                               "Amended (Yes)",           "Amended (Yes) details"),
        # IMPG SEZ
        ("IMPG SEZ","Import of Goods from SEZ Units/Developers on Bill of Entry","GSTIN of supplier",       "GSTIN of the supplier"),
        ("",        "",                                                            "Trade/Legal name",        _TN),
        ("",        "",                                                            "Reference date (ICEGATE)","Reference date of ICEGATE format shall be DD-MM-YYYY"),
        ("",        "",                                                            "Port code",               "Port code should be 6 digit alphanumeric"),
        ("",        "",                                                            "Number",                  "Bill of entry no, format shall be 7 digit number"),
        ("",        "",                                                            "Date",                    "Bill of entry date format shall be DD-MM-YYYY"),
        ("",        "",                                                            "Taxable value",           "Taxable value (in rupees)"),
        ("",        "",                                                            "Integrated tax",          "Integrated tax (in rupees)"),
        ("",        "",                                                            "Cess",                    "Cess amount (in rupees)"),
        ("",        "",                                                            "Amended (Yes)",           "Amended (Yes) details"),
    ]

    for r_idx, (ws_name, table_ref, field, help_txt) in enumerate(rows, 8):
        ws.cell(r_idx, 2, ws_name).font  = Font(name="Calibri", size=11, bold=bool(ws_name))
        ws.cell(r_idx, 3, table_ref).font = Font(name="Calibri", size=11)
        ws.cell(r_idx, 4, field).font     = Font(name="Calibri", size=11)
        hc = ws.cell(r_idx, 5, help_txt)
        hc.font      = Font(name="Calibri", size=11)
        hc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r_idx].height = 30 if "\n" not in (help_txt or "") else 60

    # ── Column widths (portal exact) ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 5

def _write_summary(wb, profile, b2b, cdn, b2ba, cdnra):
    ws = wb.create_sheet("Summary", 1)
    p = profile or {}

    # Metadata block — labels col C bold, values merged D:H (portal exact layout)
    meta = [
        (1, "GSTIN:",      p.get("gstin", "")),
        (2, "Trade Name:", p.get("trdnm") or p.get("lgl_nm") or p.get("bname", "")),
        (3, "F.Y. ",       p.get("fy", "")),
        (4, "Period:",     p.get("period") or p.get("fy", "")),
    ]
    for row, label, value in meta:
        lc = ws.cell(row, 3, label)
        lc.font = Font(name="Calibri", size=11, bold=True)
        vc = ws.cell(row, 4, value)
        vc.font = Font(name="Calibri", size=11, bold=True)
        vc.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)

    ws.row_dimensions[5].height = 20.25
    ws.merge_cells("A5:H5")
    cell_title = ws.cell(5, 1, "GSTR-2A Summary")
    cell_title.font = Font(name="Calibri", size=16, bold=True, color="000000")
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[6].height = 28.5
    headers = ["Type of Invoices", "Status", "Document Count", "Taxable Value", "IGST", "CGST", "SGST", "CESS"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(6, c, h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
        cell.fill = _fill(_NAVY)
        cell.border = _hdr_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _get_stats(rows, tx_idx=9):
        cnt = len({(r[0], r[2]) for r in rows if len(r) > 2})  # Unique by GSTIN + Doc Num
        txv  = sum(r[tx_idx]   for r in rows if isinstance(r[tx_idx],   (int, float)))
        igst = sum(r[tx_idx+1] for r in rows if isinstance(r[tx_idx+1], (int, float)))
        cgst = sum(r[tx_idx+2] for r in rows if isinstance(r[tx_idx+2], (int, float)))
        sgst = sum(r[tx_idx+3] for r in rows if isinstance(r[tx_idx+3], (int, float)))
        cess = sum(r[tx_idx+4] for r in rows if isinstance(r[tx_idx+4], (int, float)))
        return [cnt, txv, igst, cgst, sgst, cess]

    def _get_cdn_stats(rows):
        # CDN row: [gstin(0), name(1), ntyp(2), nno(3), typ(4), ndt(5), nval(6),
        #           pos(7), rev(8), rt(9), txval(10), iamt(11), camt(12), samt(13), csamt(14)]
        # Credit Notes reduce ITC → shown as negative; Debit Notes → positive
        unique_docs = set()
        txv = igst = cgst = sgst = cess = 0.0
        for r in rows:
            if len(r) < 15:
                continue
            unique_docs.add((r[0], r[3]))  # GSTIN + note number
            sign = -1 if r[2] == "Credit Note" else 1
            if isinstance(r[10], (int, float)): txv  += sign * r[10]
            if isinstance(r[11], (int, float)): igst += sign * r[11]
            if isinstance(r[12], (int, float)): cgst += sign * r[12]
            if isinstance(r[13], (int, float)): sgst += sign * r[13]
            if isinstance(r[14], (int, float)): cess += sign * r[14]
        return [len(unique_docs), txv, igst, cgst, sgst, cess]

    def _get_cdnra_stats(rows):
        # CDNRA row: [ontyp(0), onno(1), ondt(2), gstin(3), name(4), ntyp(5), nno(6),
        #             typ(7), ndt(8), nval(9), pos(10), rev(11), rt(12),
        #             txval(13), iamt(14), camt(15), samt(16), csamt(17)]
        unique_docs = set()
        txv = igst = cgst = sgst = cess = 0.0
        for r in rows:
            if len(r) < 18:
                continue
            unique_docs.add((r[3], r[6]))  # GSTIN + revised note number
            sign = -1 if r[5] == "Credit Note" else 1
            if isinstance(r[13], (int, float)): txv  += sign * r[13]
            if isinstance(r[14], (int, float)): igst += sign * r[14]
            if isinstance(r[15], (int, float)): cgst += sign * r[15]
            if isinstance(r[16], (int, float)): sgst += sign * r[16]
            if isinstance(r[17], (int, float)): cess += sign * r[17]
        return [len(unique_docs), txv, igst, cgst, sgst, cess]

    s_a = _get_stats(b2b, 9)
    s_b = _get_cdn_stats(cdn)
    s_c = _get_stats(b2ba, 11)
    s_d = _get_cdnra_stats(cdnra)

    data = [
        ["(A). Inward supplies received from a registered person", "Completed"] + s_a,
        ["(B). Details of Credit/Debit Notes", "Completed"] + s_b,
        ["(C). Amendments to B2B Invoices", "Completed"] + s_c,
        ["(D). Amendments to Credit/Debit Notes", "-"] + s_d,
        ["Total", "Completed"] + [s_a[0]+s_b[0], s_a[1]+s_b[1], s_a[2]+s_b[2], s_a[3]+s_b[3], s_a[4]+s_b[4], s_a[5]+s_b[5]]
    ]
    for r_idx, row in enumerate(data, 7):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.font = Font(name="Calibri", size=11, bold=(r_idx == 11))
            if c_idx > 2:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # Portal column widths (exact)
    ws.column_dimensions["A"].width = 62.18
    ws.column_dimensions["B"].width = 17.82
    ws.column_dimensions["C"].width = 15.73
    ws.column_dimensions["D"].width = 21.54
    ws.column_dimensions["E"].width = 19.45
    ws.column_dimensions["F"].width = 17.18
    ws.column_dimensions["G"].width = 17.45
    ws.column_dimensions["H"].width = 14.82

# =============================================================================
# Main Export Functions
# =============================================================================

def gstr2a_to_excel(data: dict, out_path: str, profile: dict = None):
    wb = Workbook(); wb.remove(wb.active); _write_readme(wb, profile, data)
    sheets = [
        ('B2B',      "Goods and Services Tax  - GSTR 2A", "Taxable inward supplies received from registered persons",                          B2B_HDR5,     B2B_HDR6,     _flatten_b2b,     {6,9,10,11,12,13,14},     None),
        ('B2BA',     "Goods and Services Tax  - GSTR 2A", "Amendments to previously uploaded invoices by supplier",                            B2BA_HDR5,    B2BA_HDR6,    _flatten_b2ba,    {8,11,12,13,14,15,16},    None),
        ('CDNR',     "Goods and Services Tax  - GSTR 2A", "Debit/Credit notes (Original)",                                                     CDN_HDR5,     CDN_HDR6,     _flatten_cdn,     {7,10,11,12,13,14},       None),
        ('CDNRA',    "Goods and Services Tax  - GSTR 2A", "Amendments to previously uploaded Credit/Debit notes by supplier",                  CDNRA_HDR5,   CDNRA_HDR6,   _flatten_cdnra,   {10,13,14,15,16,17,18},   None),
        ('ISD',      "Goods and Services Tax - GSTR2A",   "ISD Credits",                                                                       ISD_HDR5,     ISD_HDR6,     _flatten_isd,     {11,12,13,14},            None),
        ('ISDA',     "Goods and Services Tax - GSTR2A",   "Amendments ISD Credits received",                                                   ISDA_HDR5,    ISDA_HDR6,    _flatten_isda,    {14,15,16,17},            ISDA_HDR7),
        ('TDS',      "Goods and Services Tax  - GSTR 2A", "TDS Credit received",                                                               TDS_HDR5,     TDS_HDR6,     _flatten_tds,     {4,5,6,7},                None),
        ('TDSA',     "Goods and Services Tax  - GSTR 2A", "Amendments to TDS Credit received",                                                 TDSA_HDR5,    TDSA_HDR6,    _flatten_tdsa,    {5,6,7,8},                None),
        ('TCS',      "Goods and Services Tax  - GSTR 2A", "TCS Credit received",                                                               TCS_HDR5,     TCS_HDR6,     _flatten_tcs,     {4,5,6,7,8,9},            None),
        ('IMPG',     "Goods and Services Tax  - GSTR 2A", "Import of Goods from Overseas on Bill of Entry",                                    IMPG_HDR5,    IMPG_HDR6,    _flatten_impg,    {5,6,7},                  None),
        ('IMPG SEZ', "Goods and Services Tax  - GSTR 2A", "Import of Goods from SEZ Units/Developers on Bill of Entry",                        IMPGSEZ_HDR5, IMPGSEZ_HDR6, _flatten_impgsez, {7,8,9},                  None),
    ]
    for sname, t1, t4, h5, h6, flat_fn, nums, h7 in sheets:
        key = "impgsez" if sname == "IMPG SEZ" else sname.lower()
        ws = wb.create_sheet(sname)
        sec_data = _parse_section(data, key)
        rows = flat_fn(sec_data)
        if not rows: rows = [[""] * len(h5)]
        _write_portal_sheet(ws, t1, t4, h5, h6, rows, nums, h7)
    wb.save(out_path)

def gstr2a_consolidated_to_excel(data_list: list, out_path: str, profile: dict = None):
    wb = Workbook(); wb.remove(wb.active)
    read_me_profile = dict(profile or {}); read_me_profile["period"] = read_me_profile.get("fy", "")
    _write_readme(wb, read_me_profile, None)

    sheets_meta = [
        ('B2B',      "Goods and Services Tax  - GSTR 2A", "Taxable inward supplies received from registered persons",               B2B_Y_HDR5,     B2B_Y_HDR6,     _flatten_b2b,     {6,9,10,11,12,13,14},     None),
        ('B2BA',     "Goods and Services Tax - GSTR-2A",  "Amendments to previously uploaded invoices by supplier",                B2BA_Y_HDR5,    B2BA_Y_HDR6,    _flatten_b2ba,    {8,11,12,13,14,15,16},    B2BA_Y_HDR7),
        ('CDNR',     "Goods and Services Tax  - GSTR 2A", "Debit/Credit notes (Original)",                                         CDNR_Y_HDR5,    CDNR_Y_HDR6,    _flatten_cdn,     {7,10,11,12,13,14,15},    None),
        ('CDNRA',    "Goods and Services Tax - GSTR2A",   "Amendments to previously uploaded Credit/Debit notes by supplier",      CDNRA_Y_HDR5,   CDNRA_Y_HDR6,   _flatten_cdnra,   {10,13,14,15,16,17,18},   CDNRA_Y_HDR7),
        ('ISD',      "Goods and Services Tax - GSTR2A",   "ISD Credits",                                                           ISD_Y_HDR5,     ISD_Y_HDR6,     _flatten_isd,     {11,12,13,14},            None),
        ('ISDA',     "Goods and Services Tax - GSTR2A",   "Amendments ISD Credits received",                                       ISDA_Y_HDR5,    ISDA_Y_HDR6,    _flatten_isda,    {14,15,16,17},            ISDA_Y_HDR7),
        ('TDS',      "Goods and Services Tax  - GSTR 2A", "TDS Credit received",                                                   TDS_Y_HDR5,     TDS_Y_HDR6,     _flatten_tds,     {4,5,6,7},                None),
        ('TDSA',     "Goods and Services Tax  - GSTR 2A", "Amendments to TDS Credit received",                                     TDSA_Y_HDR5,    TDSA_Y_HDR6,    _flatten_tdsa,    {5,6,7,8},                None),
        ('TCS',      "Goods and Services Tax  - GSTR 2A", "TCS Credit received",                                                   TCS_Y_HDR5,     TCS_Y_HDR6,     _flatten_tcs,     {4,5,6,7,8,9},            None),
        ('IMPG',     "Goods and Services Tax  - GSTR 2A", "Import of Goods from Overseas on Bill of Entry",                        IMPG_Y_HDR5,    IMPG_Y_HDR6,    _flatten_impg,    {5,6,7},                  None),
        ('IMPG SEZ', "Goods and Services Tax  - GSTR 2A", "Import of Goods from SEZ Units/Developers on Bill of Entry",            IMPGSEZ_Y_HDR5, IMPGSEZ_Y_HDR6, _flatten_impgsez, {7,8,9},                  None),
    ]
    
    all_rows = {s[0]: [] for s in sheets_meta}
    for m in sorted(data_list, key=lambda x: str(x.get("rtn_prd") or "")[2:]+str(x.get("rtn_prd") or "")[:2]):
        rp = str(m.get("rtn_prd") or "")
        period_str = f"{_MONTH_MAP.get(rp[:2], rp[:2])[:3]}- {rp[4:]}" if len(rp) == 6 else rp
        for sname, t1, t4, h5, h6, flat_fn, nums, *extra in sheets_meta:
            sec_data = _parse_section(m, sname.lower()); rows = flat_fn(sec_data)
            for r in rows:
                r_sliced = r[:len(h5)-1]
                if len(r_sliced) < len(h5) - 1: r_sliced.extend([""] * (len(h5) - 1 - len(r_sliced)))
                r_sliced.append(period_str)
                all_rows[sname].append(r_sliced)

    # Global Sort by Date then GSTIN
    for sname in ["B2B", "CDNR", "B2BA", "CDNRA"]:
        # Date column is index 4 for B2B/CDNR, index 6 for B2BA, index 8 for CDNRA
        dt_idx = 4 if sname in ["B2B", "CDNR"] else (6 if sname == "B2BA" else 8)
        gst_idx = 0 if sname in ["B2B", "CDNR"] else 2
        all_rows[sname].sort(key=lambda r: (r[dt_idx] if len(r) > dt_idx and hasattr(r[dt_idx], "year") else datetime(1900,1,1), str(r[gst_idx] or "")))

    _write_summary(wb, profile or {}, all_rows["B2B"], all_rows["CDNR"], all_rows["B2BA"], all_rows["CDNRA"])
    for sname, t1, t4, h5, h6, flat_fn, nums, *extra in sheets_meta:
        ws = wb.create_sheet(sname); h7 = extra[0] if extra else None
        _write_portal_sheet(ws, t1, t4, h5, h6, all_rows[sname] or [[""] * len(h5)], nums, h7)

    wb.save(out_path)
