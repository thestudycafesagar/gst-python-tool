import json
from pathlib import Path
from gst_excel_utils import (
    openpyxl, _write_sheet, _empty_sheet, _s, _n,
    _fill, _side, _border, _align, _NAVY, _WHITE, get_column_letter, Font
)
from openpyxl.styles import PatternFill, Alignment, Border, Side

_3B_OUTWARD_MAP = {
    "osup_det":      "3.1(a)  Taxable Outward Supplies",
    "osup_zero":     "3.1(b)  Zero Rated Outward Supplies",
    "osup_nil_exmp": "3.1(c)  Nil / Exempt Outward Supplies",
    "isup_rev":      "3.1(d)  Inward Supplies (Reverse Charge)",
    "osup_nongst":   "3.1(e)  Non-GST Outward Supplies",
}
_3B_ITC_AVL_MAP = {
    "IMPG": "4(A)(1)  Import of Goods",
    "IMPS": "4(A)(2)  Import of Services",
    "ISRC": "4(A)(3)  Inward Supplies (Reverse Charge)",
    "ISD":  "4(A)(4)  ISD Inward Supplies",
    "OTH":  "4(A)(5)  All Other ITC",
}
_3B_ITC_REV_MAP = {
    "RUL_42_43": "4(B)(1)  Rules 42 & 43 of CGST",
    "OTH":       "4(B)(2)  Others",
}
_3B_INELG_MAP = {
    "RUL_42_43": "4(D)(1)  Rules 42 & 43 of CGST",
    "OTH":       "4(D)(2)  Others",
}
_3B_INWARD_MAP = {
    "GST_INWARD":    "5(a)  GST Inward (Reverse Charge)",
    "NONGST_INWARD": "5(b)  Non-GST Inward Supplies",
}


def gstr3b_to_excel(data: dict, out_path: str, profile: dict = None):
    """Convert parsed GSTR-3B JSON dict to a multi-sheet Excel workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Handle unified structure (summary + payment)
    summary = data.get("summary")
    payment = data.get("payment")

    # Fallback for old structure or direct summary data
    if summary is None and payment is None:
        summary = data

    # Unwrap summary if nested in 'data'
    summary_unwrapped = summary
    if isinstance(summary_unwrapped, dict) and isinstance(summary_unwrapped.get("data"), dict):
        summary_unwrapped = summary_unwrapped["data"]

    # Extract company name and GSTIN if available
    summary_raw = data.get("summary") or data
    profile = data.get("profile") or {}
    
    comp_name = (profile.get("bname") or profile.get("lgl_nm") or profile.get("trdnm") or 
                 profile.get("legal_name") or profile.get("trade_name"))
    gstin = profile.get("gstin") or profile.get("ctin")

    if not comp_name:
        comp_name = (summary_raw.get("bname") or summary_raw.get("lgl_nm") or summary_raw.get("trdnm") or 
                     summary_raw.get("legal_name") or summary_raw.get("trade_name"))
    if not gstin:
        gstin = summary_raw.get("gstin") or summary_raw.get("ctin")
    
    # Check inner data
    s_inner = summary_raw
    if isinstance(s_inner, dict) and isinstance(s_inner.get("data"), dict): s_inner = s_inner["data"]
    if not comp_name:
        comp_name = (s_inner.get("bname") or s_inner.get("lgl_nm") or s_inner.get("trdnm") or 
                     s_inner.get("legal_name") or s_inner.get("trade_name"))
    if not gstin:
        gstin = s_inner.get("gstin") or s_inner.get("ctin")

    title = f"{comp_name}" if comp_name else ""
    subtitle = f"GSTIN: {gstin}" if gstin else ""

    # Use summary as the primary source for existing sheets
    if isinstance(summary_unwrapped, dict):
        # ── 3.1 Outward Supplies ─────────────────────────────────────────────
        ws = wb.create_sheet("3.1 Outward Supplies")
        hdrs = ["Description", "Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"]
        rows = []
        sup = summary_unwrapped.get("sup_details") or {}
        for key, label in _3B_OUTWARD_MAP.items():
            d = sup.get(key) or {}
            rows.append([
                label, _n(d.get("txval")), _n(d.get("iamt")), 
                _n(d.get("camt")), _n(d.get("samt")), _n(d.get("csamt"))
            ])
        _write_sheet(ws, hdrs, rows, {2, 3, 4, 5, 6}, title=title, subtitle=subtitle)

        # ── 3.2 Inter-State Supplies ─────────────────────────────────────────
        ws = wb.create_sheet("3.2 Inter-State")
        hdrs = ["Category", "Place of Supply", "Taxable Value (₹)", "IGST (₹)"]
        rows = []
        inter = summary_unwrapped.get("inter_sup") or {}
        for k, lbl in [("unreg_details", "Unregistered"), ("comp_details", "Composition"), ("uin_details", "UIN Holders")]:
            for item in (inter.get(k) or []):
                rows.append([lbl, _s(item.get("pos")), _n(item.get("txval")), _n(item.get("iamt"))])
        _write_sheet(ws, hdrs, rows, {3, 4}, title=title, subtitle=subtitle)

        # ── 4 Eligible ITC ───────────────────────────────────────────────────
        ws = wb.create_sheet("4 Eligible ITC")
        hdrs = ["Description", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"]
        rows = []
        itc = summary_unwrapped.get("itc_elg") or {}

        def _add_section(lbl, data_list, map_dict):
            rows.append([lbl, "", "", "", ""])
            lookup = {item.get("ty"): item for item in (data_list or [])}
            for k, label in map_dict.items():
                d = lookup.get(k) or {}
                rows.append([label, _n(d.get("iamt")), _n(d.get("camt")), _n(d.get("samt")), _n(d.get("csamt"))])

        _add_section("4(A) ITC Available", itc.get("itc_avl"), _3B_ITC_AVL_MAP)
        _add_section("4(B) ITC Reversed",  itc.get("itc_rev"), _3B_ITC_REV_MAP)
        net = itc.get("itc_net") or {}
        rows.append(["4(C) Net ITC (A-B)", _n(net.get("iamt")), _n(net.get("camt")), _n(net.get("samt")), _n(net.get("csamt"))])
        _add_section("4(D) Ineligible ITC", itc.get("itc_inelg"), _3B_INELG_MAP)
        _write_sheet(ws, hdrs, rows, {2, 3, 4, 5}, title=title, subtitle=subtitle)

        # ── 5 Inward Nil/Exempt ──────────────────────────────────────────────
        ws = wb.create_sheet("5 Inward Nil-Exempt")
        hdrs = ["Description", "Inter-State (₹)", "Intra-State (₹)"]
        rows = []
        inw = summary_unwrapped.get("inward_sup") or {}
        for k, label in _3B_INWARD_MAP.items():
            d = inw.get(k) or {}
            rows.append([label, _n(d.get("inter")), _n(d.get("intra"))])
        _write_sheet(ws, hdrs, rows, {2, 3}, title=title, subtitle=subtitle)

        # ── 5.1 Interest & Late Fee ──────────────────────────────────────────
        ws = wb.create_sheet("5.1 Interest & Late Fee")
        hdrs = ["Description", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)"]
        rows = []
        fees = summary.get("intr_ltfee") or {}
        def _add_fees(d, label):
            d = d or {}
            rows.append([label, _n(d.get("iamt")), _n(d.get("camt")), _n(d.get("samt")), _n(d.get("csamt"))])
        _add_fees(fees.get("intr_details"),  "Interest")
        _add_fees(fees.get("ltfee_details"), "Late Fee")
        _write_sheet(ws, hdrs, rows, {2, 3, 4, 5})

    # ── 7 Payment of Tax (Portal Style) ──────────────────────────────────────
    if payment:
        pay_data = payment.get("data") if isinstance(payment.get("data"), dict) else payment
        if isinstance(pay_data, dict):
            ws = wb.create_sheet("Payment of Tax")
            
            # Header Structure (Single row as requested)
            headers = [
                "Tax Type", 
                "Net Tax Payable - Reverse charge (₹)", "Net Tax Payable - Forward charge (₹)",
                "Paid through ITC (Integrated Tax)", "Paid through ITC (Central Tax)", "Paid through ITC (State/UT Tax)", "Paid through ITC (CESS)",
                "Other than rev charge Tax to be paid in Cash(₹)", "Reverse charge Tax to be paid in Cash(₹)",
                "Interest payable (₹)", "Interest to be paid in cash (₹)",
                "Late Fee Payable (₹)", "Late Fee to be paid in cash (₹)"
            ]
            
            rows = []
            nested = pay_data.get("returnsDbCdredList") or {}
            
            # Organize data by tax type
            tax_types = ["igst", "cgst", "sgst", "cess"]
            labels = {
                "igst": "Integrated Tax (₹)",
                "cgst": "Central Tax (₹)",
                "sgst": "State/UT Tax (₹)",
                "cess": "CESS (₹)"
            }
            
            # Parse components
            tax_pay = nested.get("tax_pay") or []
            paid = nested.get("tax_paid") or {}
            pd_cash = paid.get("pd_by_cash") or []
            pd_itc = (paid.get("pd_by_itc") or [{}])[0]
            
            for t in tax_types:
                # Net Tax Payable
                rev_liab = 0
                fwd_liab = 0
                for tp in tax_pay:
                    amt = _n(tp.get(t, {}).get("tx"))
                    if str(tp.get("trancd")) == "30003": rev_liab += amt
                    else: fwd_liab += amt
                
                # Paid through ITC
                itc_igst = _n(pd_itc.get(f"{t}_igst_amt"))
                itc_cgst = _n(pd_itc.get(f"{t}_cgst_amt"))
                itc_sgst = _n(pd_itc.get(f"{t}_sgst_amt"))
                itc_cess = _n(pd_itc.get(f"{t}_cess_amt"))
                
                # Paid through Cash
                fwd_cash = 0
                rev_cash = 0
                intr_pay = 0
                intr_cash = 0
                lfee_pay = 0
                lfee_cash = 0
                
                for pc in pd_cash:
                    tx_amt = _n(pc.get(t, {}).get("tx"))
                    in_amt = _n(pc.get(t, {}).get("intr"))
                    lf_amt = _n(pc.get(t, {}).get("fee"))
                    
                    if str(pc.get("trancd")) == "30003":
                        rev_cash += tx_amt
                    else:
                        fwd_cash += tx_amt
                    
                    intr_cash += in_amt
                    lfee_cash += lf_amt

                for tp in tax_pay:
                    intr_pay += _n(tp.get(t, {}).get("intr"))
                    lfee_pay += _n(tp.get(t, {}).get("fee"))

                rows.append([
                    labels[t],
                    rev_liab, fwd_liab,
                    itc_igst, itc_cgst, itc_sgst, itc_cess,
                    fwd_cash, rev_cash,
                    intr_pay, intr_cash,
                    lfee_pay, lfee_cash
                ])
            
            _write_sheet(ws, headers, rows, set(range(2, 15)))
            
            # Formatting widths
            ws.column_dimensions["A"].width = 25
            for c in range(2, 15):
                ws.column_dimensions[get_column_letter(c)].width = 20

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

_3B_COMPUTAX_ROWS = [
    ("3.1 Details of Outward Supplies and inward supplies liable to reverse charge (other than those covered in 3.1.1)", "header", ""),
    ("(a) Outward taxable supplies (other than zero rated, nil rated and exempted)", "header", ""),
    ("Total Taxable Value", "3.1a", "txval"),
    ("Integrated Tax", "3.1a", "iamt"),
    ("Central Tax", "3.1a", "camt"),
    ("State Tax", "3.1a", "samt"),
    ("Cess", "3.1a", "csamt"),
    ("(b) Outward taxable supplies (zero rated)", "header", ""),
    ("Total Taxable Value", "3.1b", "txval"),
    ("Integrated Tax", "3.1b", "iamt"),
    ("Cess", "3.1b", "csamt"),
    ("(c) Other outward supplies (Nil Rated, exempted)", "header", ""),
    ("Total Taxable Value", "3.1c", "txval"),
    ("(d) Inward supplies (liable to reverse charge)", "header", ""),
    ("Total Taxable Value", "3.1d", "txval"),
    ("Integrated Tax", "3.1d", "iamt"),
    ("Central Tax", "3.1d", "camt"),
    ("State Tax", "3.1d", "samt"),
    ("Cess", "3.1d", "csamt"),
    ("(e) Non-GST outward supplies", "header", ""),
    ("Total Taxable Value", "3.1e", "txval"),
    ("3.1.1 Details of Supplies notified under section 9(5) of the CGST Act, 2017 and corresponding provisions in IGST/UTGST/SGST Acts", "header", ""),
    ("(i) Taxable supplies on which electronic commerce operator pays tax under sub-section (5) of section 9 [to be furnished by the electronic commerce operator]", "header", ""),
    ("Total Taxable Value", "3.1.1i", "txval"),
    ("Integrated Tax", "3.1.1i", "iamt"),
    ("Central Tax", "3.1.1i", "camt"),
    ("State Tax", "3.1.1i", "samt"),
    ("Cess", "3.1.1i", "csamt"),
    ("(ii) Taxable supplies made by the registered person through electronic commerce operator, on which electronic commerce operator is required to pay tax under sub-section (5) of section 9 [to be furnished by the registered person making supplies through electronic commerce operator].", "header", ""),
    ("Total Taxable Value", "3.1.1ii", "txval"),
    ("Integrated Tax", "3.1.1ii", "iamt"),
    ("Central Tax", "3.1.1ii", "camt"),
    ("State Tax", "3.1.1ii", "samt"),
    ("Cess", "3.1.1ii", "csamt"),
    ("4 Eligible ITC", "header", ""),
    ("(A) ITC Available (whether in full or part)", "header", ""),
    ("(1) Import of goods", "header", ""),
    ("Integrated Tax", "4a1", "iamt"),
    ("Cess", "4a1", "csamt"),
    ("(2) Import of services", "header", ""),
    ("Integrated Tax", "4a2", "iamt"),
    ("Cess", "4a2", "csamt"),
    ("(3) Inward supplies liable to reverse charge (other than 1 & 2 above)", "header", ""),
    ("Integrated Tax", "4a3", "iamt"),
    ("Central Tax", "4a3", "camt"),
    ("State Tax", "4a3", "samt"),
    ("Cess", "4a3", "csamt"),
    ("(4) Inward supplies from ISD", "header", ""),
    ("Integrated Tax", "4a4", "iamt"),
    ("Central Tax", "4a4", "camt"),
    ("State Tax", "4a4", "samt"),
    ("Cess", "4a4", "csamt"),
    ("(5) All other ITC", "header", ""),
    ("Integrated Tax", "4a5", "iamt"),
    ("Central Tax", "4a5", "camt"),
    ("State Tax", "4a5", "samt"),
    ("Cess", "4a5", "csamt"),
    ("(B) ITC Reversed", "header", ""),
    ("(1) As per rules 38,42 & 43 of CGST Rules and section 17(5)", "header", ""),
    ("Integrated Tax", "4b1", "iamt"),
    ("Central Tax", "4b1", "camt"),
    ("State Tax", "4b1", "samt"),
    ("Cess", "4b1", "csamt"),
    ("(2) Others", "header", ""),
    ("Integrated Tax", "4b2", "iamt"),
    ("Central Tax", "4b2", "camt"),
    ("State Tax", "4b2", "samt"),
    ("Cess", "4b2", "csamt"),
    ("(C) Net ITC Available", "header", ""),
    ("Integrated Tax", "4c", "iamt"),
    ("Central Tax", "4c", "camt"),
    ("State Tax", "4c", "samt"),
    ("Cess", "4c", "csamt"),
    ("(D) Other Details", "header", ""),
    ("(1) ITC reclaimed which was reversed under Table 4(B)(2) in earlier tax period", "header", ""),
    ("Integrated Tax", "4d1", "iamt"),
    ("Central Tax", "4d1", "camt"),
    ("State Tax", "4d1", "samt"),
    ("Cess", "4d1", "csamt"),
    ("(2) Ineligible ITC under section 16(4) & ITC restricted due to PoS rules", "header", ""),
    ("Integrated Tax", "4d2", "iamt"),
    ("Central Tax", "4d2", "camt"),
    ("State Tax", "4d2", "samt"),
    ("Cess", "4d2", "csamt"),
    ("5. Values of exempt, nil-rated and non-GST inward supplies", "header", ""),
    ("From a supplier under composition scheme, Exempt and Nil rated supply", "header", ""),
    ("Inter-State supplies", "5a", "inter"),
    ("Intra-state supplies", "5a", "intra"),
    ("Non GST supply", "header", ""),
    ("Inter-State supplies", "5b", "inter"),
    ("Intra-state supplies", "5b", "intra"),
    ("5.1 Interest & Late Fee Payable", "header", ""),
    ("Interest", "header", ""),
    ("Integrated Tax", "5.1_intr", "iamt"),
    ("Central Tax", "5.1_intr", "camt"),
    ("State Tax", "5.1_intr", "samt"),
    ("Cess", "5.1_intr", "csamt"),
    ("Late Fee", "header", ""),
    ("Central Tax", "5.1_fee", "camt"),
    ("State Tax", "5.1_fee", "samt"),
    ("6.1 Payment of tax", "header", ""),
    ("Tax Payable Other than Reverse Charge", "header", ""),
    ("Integrated Tax", "6_pay_other", "igst"),
    ("Central Tax", "6_pay_other", "cgst"),
    ("State Tax", "6_pay_other", "sgst"),
    ("Cess", "6_pay_other", "cess"),
    ("Paid through ITC", "header", ""),
    ("IGST", "header", ""),
    ("Integrated Tax", "6_itc_igst", "igst"),
    ("Central Tax", "6_itc_igst", "cgst"),
    ("State Tax", "6_itc_igst", "sgst"),
    ("Cess", "6_itc_igst", "cess"),
    ("CGST", "header", ""),
    ("Integrated Tax", "6_itc_cgst", "igst"),
    ("Central Tax", "6_itc_cgst", "cgst"),
    ("SGST", "header", ""),
    ("Integrated Tax", "6_itc_sgst", "igst"),
    ("State Tax", "6_itc_sgst", "sgst"),
    ("CESS", "header", ""),
    ("Cess", "6_itc_cess", "cess"),
    ("Balance Payable in Cash", "header", ""),
    ("Integrated Tax", "6_bal", "igst"),
    ("Central Tax", "6_bal", "cgst"),
    ("State Tax", "6_bal", "sgst"),
    ("Cess", "6_bal", "cess"),
    ("Tax Payable On Reverse Charge", "header", ""),
    ("Integrated Tax", "6_pay_rcm", "igst"),
    ("Central Tax", "6_pay_rcm", "cgst"),
    ("State Tax", "6_pay_rcm", "sgst"),
    ("Cess", "6_pay_rcm", "cess"),
    ("Total Tax Payable in Cash", "header", ""),
    ("Integrated Tax", "6_tot_bal", "igst"),
    ("Central Tax", "6_tot_bal", "cgst"),
    ("State Tax", "6_tot_bal", "sgst"),
    ("Cess", "6_tot_bal", "cess"),
    ("Paid in Cash", "header", ""),
    ("Tax", "header", ""),
    ("Integrated Tax", "6_cash_tax", "igst"),
    ("Central Tax", "6_cash_tax", "cgst"),
    ("State Tax", "6_cash_tax", "sgst"),
    ("Cess", "6_cash_tax", "cess"),
    ("Interest", "header", ""),
    ("Integrated Tax", "6_cash_intr", "igst"),
    ("Central Tax", "6_cash_intr", "cgst"),
    ("State Tax", "6_cash_intr", "sgst"),
    ("Cess", "6_cash_intr", "cess"),
    ("Late Fee", "header", ""),
    ("Central Tax", "6_cash_fee", "cgst"),
    ("State Tax", "6_cash_fee", "sgst"),
]

def _extract_3b_val(month_data, section, key):
    summary = month_data.get("summary") or month_data
    if isinstance(summary, dict) and isinstance(summary.get("data"), dict): summary = summary["data"]
    payment = month_data.get("payment")
    if not isinstance(summary, dict): summary = {}

    if section == "3.1a": return _n(summary.get("sup_details", {}).get("osup_det", {}).get(key))
    if section == "3.1b": return _n(summary.get("sup_details", {}).get("osup_zero", {}).get(key))
    if section == "3.1c": return _n(summary.get("sup_details", {}).get("osup_nil_exmp", {}).get(key))
    if section == "3.1d": return _n(summary.get("sup_details", {}).get("isup_rev", {}).get(key))
    if section == "3.1e": return _n(summary.get("sup_details", {}).get("osup_nongst", {}).get(key))
    if section == "3.1.1i": return _n(summary.get("eco_details", {}).get("eco_sup", {}).get(key))
    if section == "3.1.1ii": return _n(summary.get("eco_details", {}).get("reg_sup", {}).get(key))

    itc = summary.get("itc_elg", {})
    avl = {x.get("ty"): x for x in itc.get("itc_avl", [])}
    rev = {x.get("ty"): x for x in itc.get("itc_rev", [])}
    inelg = {x.get("ty"): x for x in itc.get("itc_inelg", [])}

    if section == "4a1": return _n(avl.get("IMPG", {}).get(key))
    if section == "4a2": return _n(avl.get("IMPS", {}).get(key))
    if section == "4a3": return _n(avl.get("ISRC", {}).get(key))
    if section == "4a4": return _n(avl.get("ISD", {}).get(key))
    if section == "4a5": return _n(avl.get("OTH", {}).get(key))
    if section == "4b1": return _n(rev.get("RUL_42_43", {}).get(key)) or _n(rev.get("RUL_38_42_43", {}).get(key))
    if section == "4b2": return _n(rev.get("OTH", {}).get(key))
    if section == "4c": return _n(itc.get("itc_net", {}).get(key))
    if section == "4d1": return _n(inelg.get("ITC_RECL", {}).get(key)) or _n(inelg.get("OTH", {}).get(key))
    if section == "4d2": return _n(inelg.get("POS_REST", {}).get(key)) or _n(inelg.get("RUL_42_43", {}).get(key))

    if section == "5a": return _n(summary.get("inward_sup", {}).get("GST_INWARD", {}).get(key))
    if section == "5b": return _n(summary.get("inward_sup", {}).get("NONGST_INWARD", {}).get(key))

    fees = summary.get("intr_ltfee", {})
    if section == "5.1_intr": return _n(fees.get("intr_details", {}).get(key))
    if section == "5.1_fee": return _n(fees.get("ltfee_details", {}).get(key))

    if payment and section.startswith("6_"):
        pay_data = payment.get("data") if isinstance(payment.get("data"), dict) else payment
        if isinstance(pay_data, dict):
            nested = pay_data.get("returnsDbCdredList") or {}
            tax_pay = nested.get("tax_pay") or []
            paid = nested.get("tax_paid") or {}
            pd_cash = paid.get("pd_by_cash") or []
            pd_itc = (paid.get("pd_by_itc") or [{}])[0]

            def get_liab(rcm=False):
                return sum(_n(tp.get(key, {}).get("tx")) for tp in tax_pay if (str(tp.get("trancd")) == "30003") == rcm)

            if section == "6_pay_other": return get_liab(rcm=False)
            if section == "6_pay_rcm": return get_liab(rcm=True)
            if section == "6_itc_igst": return _n(pd_itc.get(f"{key}_igst_amt"))
            if section == "6_itc_cgst": return _n(pd_itc.get(f"{key}_cgst_amt"))
            if section == "6_itc_sgst": return _n(pd_itc.get(f"{key}_sgst_amt"))
            if section == "6_itc_cess": return _n(pd_itc.get(f"{key}_cess_amt"))

            def get_cash(field):
                return sum(_n(pc.get(key, {}).get(field)) for pc in pd_cash)

            if section == "6_cash_tax": return get_cash("tx")
            if section == "6_cash_intr": return get_cash("intr")
            if section == "6_cash_fee": return get_cash("fee")
            
            if section == "6_bal":
                # Forward charge liability minus ITC = cash balance payable
                liab = get_liab(rcm=False)
                paid_itc = sum(_n(pd_itc.get(f"{key}_{t}_amt")) for t in ["igst", "cgst", "sgst", "cess"])
                return max(0, liab - paid_itc)
            if section == "6_tot_bal":
                # Balance Payable in Cash (forward) + Tax Payable On Reverse Charge
                fwd_liab = get_liab(rcm=False)
                paid_itc = sum(_n(pd_itc.get(f"{key}_{t}_amt")) for t in ["igst", "cgst", "sgst", "cess"])
                fwd_bal = max(0, fwd_liab - paid_itc)
                rcm_liab = get_liab(rcm=True)
                return fwd_bal + rcm_liab
    return 0

def _classify_header(label: str) -> str:
    """Return header level: 'major', 'sub', or 'minor'."""
    majors = ("3.1 ", "3.1.1 ", "4 E", "5. V", "5.1 I", "6.1 P")
    subs   = ("(a)", "(b)", "(c)", "(d)", "(e)",
              "(A)", "(B)", "(C)", "(D)", "(i)", "(ii)")
    if any(label.startswith(p) for p in majors):
        return "major"
    if any(label.startswith(p) for p in subs):
        return "sub"
    return "minor"


def _apply_3b_styles(ws, row_types: list, ncols: int = 14):
    """Apply vibrant styles to the 3B sheet after all rows are appended."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border(color="C8D4E8"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def align(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    # ── Colour palette (Excel Formal) ──────────────────────────────────────────
    C = {
        "title_bg":   fill("2F75B5"),   # Excel Blue
        "title_fg":   Font(bold=True, color="FFFFFF", size=13, name="Calibri"),

        "col_bg":     fill("F3F2F1"),   # Office Grey
        "col_fg":     Font(bold=True, color="323130", size=10, name="Calibri"),

        "major_bg":   fill("D9D9D9"),   # Neutral Medium Grey
        "major_fg":   Font(bold=True, color="000000", size=10, name="Calibri"),

        "sub_bg":     fill("E6E6E6"),   # Light Neutral Grey
        "sub_fg":     Font(bold=True, color="323130", size=10, name="Calibri"),

        "minor_bg":   fill("F2F2F2"),   # Very Light Grey
        "minor_fg":   Font(bold=True, italic=True, color="323130", size=9, name="Calibri"),

        "data_odd":   fill("FFFFFF"),
        "data_even":  fill("FAFAFA"),   # Subtle off-white
        "data_fg":    Font(color="323130", size=10, name="Calibri"),
        "data_fg_b":  Font(bold=True, color="323130", size=10, name="Calibri"),

        "total_bg":   fill("F3F2F1"),   # Consistent grey for totals
        "total_fg":   Font(bold=True, color="2F75B5", size=10, name="Calibri"),
    }

    NUM_FMT      = '#,##0.00;-#,##0.00;""'   # hides zero values
    NUM_FMT_TOT  = '#,##0.00;-#,##0.00;""'
    bdr          = thin_border()
    bdr_hdr      = thin_border("8EAABF")
    total_col    = ncols                      # last column = Total

    data_row_counter = 0

    for r_idx, rtype in enumerate(row_types, 1):
        if rtype == "title":
            ws.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=ncols)
            cell = ws.cell(row=r_idx, column=1)
            cell.fill      = C["title_bg"]
            cell.font      = C["title_fg"]
            cell.alignment = align("center")
            ws.row_dimensions[r_idx].height = 24

        elif rtype == "colheader":
            ws.freeze_panes = f"B{r_idx + 1}"
            ws.row_dimensions[r_idx].height = 22
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c)
                cell.fill      = C["col_bg"]
                cell.font      = C["col_fg"]
                cell.border    = bdr_hdr
                cell.alignment = align("center")

        elif rtype == "major":
            ws.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=ncols)
            ws.row_dimensions[r_idx].height = 18
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c)
                cell.fill      = C["major_bg"]
                cell.font      = C["major_fg"]
                cell.border    = bdr_hdr
                cell.alignment = align("left", wrap=True)

        elif rtype == "sub":
            ws.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=ncols)
            ws.row_dimensions[r_idx].height = 16
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c)
                cell.fill      = C["sub_bg"]
                cell.font      = C["sub_fg"]
                cell.border    = bdr_hdr
                cell.alignment = align("left", wrap=True)

        elif rtype == "minor":
            ws.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=ncols)
            ws.row_dimensions[r_idx].height = 15
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c)
                cell.fill      = C["minor_bg"]
                cell.font      = C["minor_fg"]
                cell.border    = bdr
                cell.alignment = align("left")

        elif rtype == "data":
            data_row_counter += 1
            bg = C["data_even"] if data_row_counter % 2 == 0 else C["data_odd"]
            ws.row_dimensions[r_idx].height = 15
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r_idx, column=c)
                cell.border = bdr
                if c == 1:
                    cell.fill      = bg
                    cell.font      = C["data_fg"]
                    cell.alignment = align("left", wrap=True)
                elif c == total_col:
                    cell.fill          = C["total_bg"]
                    cell.font          = C["total_fg"]
                    cell.alignment     = align("right")
                    cell.number_format = NUM_FMT_TOT
                else:
                    cell.fill          = bg
                    cell.font          = C["data_fg"]
                    cell.alignment     = align("right")
                    cell.number_format = NUM_FMT


def gstr3b_consolidated_to_excel(data_list: list, out_path: str, profile: dict = None):
    """
    Combine multiple months of GSTR-3B data into the exact 'Computax 3b' Excel layout.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if not data_list:
        _empty_sheet(wb)
        wb.save(out_path)
        return

    # Sort data by financial year order: Apr (04) to Mar (03)
    month_order = ["04", "05", "06", "07", "08", "09", "10", "11", "12", "01", "02", "03"]
    month_names = ["April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]

    # Map data to month index
    month_data_map = {}
    fy_str = "2025-2026"
    for data in data_list:
        # ret_period is nested at summary.data.ret_period (format: MMYYYY e.g. "042025")
        fp = str(data.get("fp", ""))
        if not fp:
            s = data.get("summary") or {}
            if isinstance(s, dict) and isinstance(s.get("data"), dict):
                fp = str(s["data"].get("ret_period", ""))
        if len(fp) == 6:
            m = fp[:2]
            y = fp[2:]
            if m in month_order:
                month_data_map[m] = data
            if m == "04":
                fy_str = f"{y}-{str(int(y)+1)}"

    ncols = 14  # Particular + 12 months + Total

    # ─── SHEET 1: 3B ──────────────────────────────────────────
    ws1 = wb.create_sheet("GSTR-3B Yearly")
    row_types = []

    # Extract company name and GSTIN if available
    comp_name = ""
    gstin = ""
    for d in data_list:
        summary_raw = d.get("summary") or d
        profile = d.get("profile") or {}
        
        # Try metadata in profile first (most reliable)
        comp_name = (profile.get("bname") or profile.get("lgl_nm") or profile.get("trdnm") or 
                     profile.get("legal_name") or profile.get("trade_name"))
        gstin = profile.get("gstin") or profile.get("ctin")

        # If not in profile, try summary root
        if not comp_name:
            comp_name = (summary_raw.get("bname") or summary_raw.get("lgl_nm") or summary_raw.get("trdnm") or 
                         summary_raw.get("legal_name") or summary_raw.get("trade_name"))
        if not gstin:
            gstin = summary_raw.get("gstin") or summary_raw.get("ctin")
        
        # If still not found, try inner 'data'
        s = summary_raw
        if isinstance(s, dict) and isinstance(s.get("data"), dict): s = s["data"]
        
        if not comp_name:
            comp_name = (s.get("bname") or s.get("lgl_nm") or s.get("trdnm") or 
                         s.get("legal_name") or s.get("trade_name"))
        if not gstin:
            gstin = s.get("gstin") or s.get("ctin")
            
        if comp_name or gstin: break

    if comp_name or gstin:
        name_part = str(comp_name or "").strip().upper()
        gstin_part = str(gstin or "").strip().upper()
        if name_part and gstin_part:
            title_text = f"{name_part} ({gstin_part})"
        else:
            title_text = name_part or gstin_part
        
        if title_text:
            ws1.append([title_text] + [""] * (ncols - 1))
            row_types.append("title")

    # Row 1: FY title
    ws1.append([f"F.Y. : {fy_str}"] + [""] * (ncols - 1))
    row_types.append("title")

    # Row 2: Column headers
    ws1.append(["Particular"] + month_names + ["Total"])
    row_types.append("colheader")

    # Data rows
    for row_def in _3B_COMPUTAX_ROWS:
        label, section, key = row_def
        if section == "header":
            ws1.append([label] + [""] * (ncols - 1))
            row_types.append(_classify_header(label))
        else:
            row_vals = [label]
            total = 0
            for m in month_order:
                if m in month_data_map:
                    val = _extract_3b_val(month_data_map[m], section, key)
                    row_vals.append(val)
                    total += val
                else:
                    row_vals.append("")
            row_vals.append(total)
            ws1.append(row_vals)
            row_types.append("data")

    # Apply styles
    _apply_3b_styles(ws1, row_types, ncols)

    # Column widths
    ws1.column_dimensions["A"].width = 62
    for c in range(2, ncols + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 14


    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
