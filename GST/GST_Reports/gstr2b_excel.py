import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gst_excel_utils import _state_name, _n, _s

_T_INV = ('Invoice type can be derived based on the following types \n'
          ' R- Regular (Other than SEZ supplies and Deemed exports) \n'
          ' SEZWP- SEZ supplies with payment of tax \n'
          ' SEZWOP- SEZ supplies with out payment of tax \n'
          ' DE- Deemed exports \n'
          ' CBW - Intra-State Supplies attracting IGST')
_T_REV = ('Supply attract reverse charge divided into two types: \n'
          ' Y- Purchases attract reverse charge \n'
          " N- Purchases don’t attract reverse charge")
_T_REV2 = ('Supply attract reverse charge divided in to two types \n'
           ' Y- Purchases attract reverse charge \n'
           " N- Purchases don’t attract reverse charge")
_T_SRC = ("Source of the document shall be displayed. It shall be: \n"
          " a. 'e-invoice', if the document is auto-populated from e-invoice. \n"
          " b. Blank, if the document is uploaded by the supplier")
_T_SRC_ECO = ("Source of the document shall be displayed. It shall be: \n"
              " a. 'e-invoice', if the document is auto-populated from e-invoice. \n"
              " b. Blank, if the document is uploaded by the e commerce operator")
_T_IRN = ('It is the unique Invoice reference number of the document auto-populated from e-invoice.'
          ' For the documents uploaded by the supplier, this shall be blank.')
_T_IRN_ECO = ('It is the unique Invoice reference number of the document auto-populated from e-invoice.'
              ' For the documents uploaded by the ecommerce operator, this shall be blank.')
_T_IRNDT = ('This is the date of invoice reference number, auto-populated from e-invoice.'
            ' For the documents uploaded by the supplier, this shall be blank.')
_T_IRNDT_ECO = ('This is the date of invoice reference number, auto-populated from e-invoice.'
                ' For the documents uploaded by the e commerce operator, this shall be blank.')
_T_APPL = 'If the supply is eligible to be taxed at 65% of the existing rate of tax, it shall be 65%, else blank'
_T_ITC_Y = "Is ITC available or not on the document - 'Yes' or 'No'"
_T_ITC_IF = "If ITC is available, 'Yes', else 'No'"
_T_RSN = "Reason, if ITC availability is 'No'"
_T_NOTE_CDN = ('Note Supply type can be derived based on the following types \n'
               ' R- Regular (Other than SEZ supplies and Deemed exports) \n'
               ' SEZWP- SEZ supplies with payment of tax \n'
               ' SEZWOP- SEZ supplies with out payment of tax \n'
               ' DE- Deemed exports \n'
               ' CBW - Intra-State Supplies attracting IGST')
_T_DOC_ECO = ('Document type can be derived based on the following types \n'
              ' R- Regular (Other than SEZ supplies and Deemed exports) \n'
              ' SEZWP- SEZ supplies with payment of tax \n'
              ' SEZWOP- SEZ supplies with out payment of tax \n'
              ' DE- Deemed exports')
_T_ISD_ELG = ("Eligibility of ITC are two types: \n"
              " Y-Yes. Taxpayer can claim ITC on such invoice \n"
              " N- No. Taxpayer can’t claim ITC on such invoice")
_T_ISD_ELG2 = ("Eligibility of ITC are two types \n"
               " Y-Yes. Taxpayer can claim ITC on such invoice \n"
               " N- No. Taxpayer can’t claim ITC on such invoice")
_T_POS = 'Place of supply shall be the place where goods are supplied or services are provided (As  declared by the supplier)'
_T_POS2 = 'Place of supply shall be the place where goods supplied or services provided (As declared by the supplier)'
_T_POS_ECO = 'Place of supply shall be the place where goods are supplied or services are provided (As  declared by the eco)'
_T_POS_ECO2 = 'Place of supply shall be the place where goods supplied or services provided (As declared by the Eco)'
_T_TRD = 'Trade name of the supplier will be displayed. If trade name is not available, then legal name of the supplier'
_T_TRD2 = 'Trade name of the supplier will be displayed. If trade name is not available then legal name of the supplier'
_T_TRD3 = 'Trade name of the supplier will be displayed. If trade name is not available then legal name of the supplier.'

_README_ROWS = [
    # B2B
    ['B2B', 'Taxable inward supplies received from registered person', 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD],
    [None, None, 'Invoice number', 'Invoice number'],
    [None, None, 'Invoice type', _T_INV],
    [None, None, 'Invoice date', 'Invoice date format shall be DD-MM-YYYY'],
    [None, None, 'Invoice value', 'Invoice value (in rupees)'],
    [None, None, 'Place of supply', _T_POS],
    [None, None, 'Supply attract Reverse charge', _T_REV],
    [None, None, 'Taxable value', 'Taxable value'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_Y],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    [None, None, 'Source', _T_SRC],
    [None, None, 'IRN', _T_IRN],
    [None, None, 'IRN date', _T_IRNDT],
    # B2BA
    ['B2BA', 'Amendments to previously uploaded invoices by supplier', 'Invoice number (Original details)', 'Original invoice number'],
    [None, None, 'Invoice date (Original details)', 'Original invoice date (Date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', 'Trade name of the supplier will be displayed. If trade name is not available then legal name of the supplier.'],
    [None, None, 'Invoice number', 'Revised Invoice number'],
    [None, None, 'Invoice type', _T_INV],
    [None, None, 'Invoice date', 'Invoice date format shall be DD-MM-YYYY'],
    [None, None, 'Invoice value', 'Invoice value (in rupees)'],
    [None, None, 'Place of supply', 'Place of supply shall be the place where goods supplied or services provided (As declared by the supplier)'],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    # B2B-CDNR
    ['B2B-CDNR', 'Debit/Credit notes(Original)', 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD2],
    [None, None, 'Note number', 'Debit/Credit note number'],
    [None, None, 'Note type', 'Document type can be Debit note or credit note'],
    [None, None, 'Note Supply Type', _T_NOTE_CDN],
    [None, None, 'Note date', 'Debit/Credit note date format shall be (DD-MM-YYYY)'],
    [None, None, 'Note Value', 'Debit/Credit note value (In rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    [None, None, 'Source', _T_SRC],
    [None, None, 'IRN', _T_IRN],
    [None, None, 'IRN date', _T_IRNDT],
    # B2B-CDNRA
    ['B2B-CDNRA', 'Amendments to previously uploaded Credit/Debit notes by supplier', 'Note type(Original)', 'Note type can be Debit note or credit note'],
    [None, None, 'Note number(Original)', 'Original Debit/Credit note number'],
    [None, None, 'Note date(Original)', 'Original Debit/Credit note date (Note date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD2],
    [None, None, 'Note number', 'Debit/Credit note number'],
    [None, None, 'Note type', 'Note type can be Debit note or credit note'],
    [None, None, 'Note Supply Type', _T_NOTE_CDN],
    [None, None, 'Note date', 'Debit/Credit note date format shall be (DD-MM-YYYY)'],
    [None, None, 'Note Value', 'Debit/Credit note value (In rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    # ECO
    ['ECO', 'Documents reported by ECO on which ECO is liable to pay tax u/s 9(5)', 'GSTIN of ECO', 'GSTIN of E-commerce operator'],
    [None, None, 'Trade/Legal name', 'Trade name of the E-commerce operator will be displayed. If trade name is not available, then legal name of the E-commerce operator'],
    [None, None, 'Document number', 'Document number'],
    [None, None, 'Document type', _T_DOC_ECO],
    [None, None, 'Document date', 'Document date format shall be DD-MM-YYYY'],
    [None, None, 'Document value', 'Document value (in rupees)'],
    [None, None, 'Place of supply', _T_POS_ECO],
    [None, None, 'Taxable value', 'Taxable value'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A period', 'Period for which GSTR-1/IFF has been filed'],
    [None, None, 'GSTR-1/IFF/1A filing date', 'Date on which GSTR-1/IFF has been filed'],
    [None, None, 'ITC availability', _T_ITC_Y],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Source', _T_SRC_ECO],
    [None, None, 'IRN', _T_IRN_ECO],
    [None, None, 'IRN date', _T_IRNDT_ECO],
    # ECOA
    ['ECOA', 'Amendments to documents reported by ECO on which ECO is liable to pay tax u/s 9(5)', 'Document number (Original details)', 'Document invoice number'],
    [None, None, 'Document date (Original details)', 'Document invoice date (Date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of ECO', 'GSTIN of E-commerce operator'],
    [None, None, 'Trade/Legal name', 'Trade name of the E-commerce operator will be displayed. If trade name is not available then legal name of the E-commerce operator.'],
    [None, None, 'Document number', 'Revised Document number'],
    [None, None, 'Document type', _T_DOC_ECO],
    [None, None, 'Document date', 'Document date format shall be DD-MM-YYYY'],
    [None, None, 'Document value', 'Document value (in rupees)'],
    [None, None, 'Place of supply', _T_POS_ECO2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A Period', 'Period for which GSTR-1/IFF has been filed'],
    [None, None, 'GSTR-1/IFF/1A filing date', 'Date on which GSTR-1/IFF has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    # ISD
    ['ISD', 'ISD Credit', 'GSTIN of ISD', 'Input Service Distributor GSTIN'],
    [None, None, 'Trade/Legal name of the ISD', 'Trade name of the ISD will be displayed. If trade name is not available then legal name of the ISD'],
    [None, None, 'ISD Document type', 'ISD document type can be Invoice or Credit note'],
    [None, None, 'ISD Document number', 'ISD invoice / ISD Credit note number'],
    [None, None, 'ISD Document date', 'ISD Document date format will be DD-MM-YYYY'],
    [None, None, 'Original ISD Invoice number', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Original ISD Invoice date', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'ISD GSTR-6 Period', 'Period for which GSTR-6 is to be filed.'],
    [None, None, 'ISD GSTR-6 Filing date', 'Date on which GSTR-6 has been filed.'],
    [None, None, 'Eligibilty of ITC', _T_ISD_ELG],
    # ISDA
    ['ISDA', 'Amendments to ISD Credits received', 'ISD Document type (Original)', 'ISD document type can be Invoice or Credit note'],
    [None, None, 'ISD Document Number (Original)', 'Invoice/Credit note number'],
    [None, None, 'ISD Document date (Original)', 'Invoice/Credit note date'],
    [None, None, 'GSTIN of ISD', 'GSTIN of the Input Service Distributor'],
    [None, None, 'Trade/Legal name of the ISD', 'Trade name of the ISD will be displayed. If trade name is not available then legal name of the ISD'],
    [None, None, 'ISD Document type', 'ISD document type can be Invoice or Credit note'],
    [None, None, 'ISD Document number', 'ISD invoice / ISD Credit note number'],
    [None, None, 'ISD Document date', 'ISD Document date format will be DD-MM-YYYY'],
    [None, None, 'Original ISD Invoice number', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Original ISD Invoice date', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'ISD GSTR-6 Period', 'Period for which GSTR-6 is to be filed.'],
    [None, None, 'ISD GSTR-6 Filing date', 'Date on which GSTR-6 has been filed.'],
    [None, None, 'Eligibilty of ITC', _T_ISD_ELG2],
    # IMPG
    ['IMPG', 'Import of goods from overseas on bill of entry', 'ICEGATE Reference date', 'Relevant date for availing credit on the bill of entry'],
    [None, None, 'Port Code', 'Port code'],
    [None, None, 'Bill of Entry number', 'Bill of Entry number'],
    [None, None, 'Bill of Entry date', 'Bill of Entry date format shall be DD-MM-YYYY'],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'Amended (Yes)', "Has the bill of entry been amended. 'Yes' or 'No'"],
    # IMPGSEZ
    ['IMPGSEZ', 'Import of goods from SEZ units/developers on bill of entry', 'GSTIN of Supplier', 'GSTIN of SEZ supplier'],
    [None, None, 'Trade/Legal name', 'Trade name of the SEZ supplier will be displayed. If trade name is not available then legal name of the SEZ supplier'],
    [None, None, 'ICEGATE Reference date', 'Relevant date for availing credit on the bill of entry'],
    [None, None, 'Port Code', 'Port code'],
    [None, None, 'Bill of Entry number', 'Bill of Entry number'],
    [None, None, 'Bill of Entry date', 'Bill of Entry date format shall be DD-MM-YYYY'],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'Amended (Yes)', "Has the bill of entry been amended. 'Yes' or 'No'"],
    # B2B(ITC Reversal)
    ['B2B(ITC Reversal)', 'ITC Reversed - Others', 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD],
    [None, None, 'Invoice number', 'Invoice number'],
    [None, None, 'Invoice type', _T_INV],
    [None, None, 'Invoice date', 'Invoice date format shall be DD-MM-YYYY'],
    [None, None, 'Invoice value', 'Invoice value (in rupees)'],
    [None, None, 'Place of supply', _T_POS],
    [None, None, 'Supply attract Reverse charge', _T_REV],
    [None, None, 'Taxable value', 'Taxable value'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A Period', 'Period for which GSTR-1/IFF/1A has been filed'],
    [None, None, 'GSTR-1/IFF/1A Filing Date', 'Date on which GSTR-1/IFF/1A has been filed'],
    [None, None, 'ITC Availability', _T_ITC_Y],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    [None, None, 'Source', _T_SRC],
    [None, None, 'IRN', _T_IRN],
    [None, None, 'IRN date', _T_IRNDT],
    # B2BA(ITC Reversal)
    ['B2BA(ITC Reversal)', 'Amendments to previously filed invoices by supplier (ITC reversal)', 'Invoice number (Original details)', 'Original invoice number'],
    [None, None, 'Invoice date (Original details)', 'Original invoice date (Date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD3],
    [None, None, 'Invoice number', 'Revised Invoice number'],
    [None, None, 'Invoice type', _T_INV],
    [None, None, 'Invoice date', 'Invoice date format shall be DD-MM-YYYY'],
    [None, None, 'Invoice value', 'Invoice value (in rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/GSTR-1A Period', 'Period for which GSTR-1/IFF/GSTR-1A has been filed'],
    [None, None, 'GSTR-1/IFF/GSTR-1A Filing Date', 'Date on which GSTR-1/IFF/GSTR-1A has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    # B2B-DNR
    ['B2B-DNR', 'Debit notes(Original)', 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD2],
    [None, None, 'Note number', 'Debit note number'],
    [None, None, 'Note type', 'Document type can be Debit note '],
    [None, None, 'Note Supply Type', _T_NOTE_CDN],
    [None, None, 'Note date', 'Debit note date format shall be (DD-MM-YYYY)'],
    [None, None, 'Note Value', 'Debit note value (In rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A Period', 'Period for which GSTR-1/IFF/1A has been filed'],
    [None, None, 'GSTR-1/IFF/1A Filing Date', 'Date on which GSTR-1/IFF/GSTR-1A has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    [None, None, 'Source', _T_SRC],
    [None, None, 'IRN', _T_IRN],
    [None, None, 'IRN date', _T_IRNDT],
    # B2B-DNRA
    ['B2B-DNRA', 'Amendments to previously uploaded Debit notes by supplier', 'Note type(Original)', 'Note type can be Debit note'],
    [None, None, 'Note number(Original)', 'Original Debit note number'],
    [None, None, 'Note date(Original)', 'Original Debit note date (Note date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD2],
    [None, None, 'Note number', 'Debit note number'],
    [None, None, 'Note type', 'Note type can be Debit note '],
    [None, None, 'Note Supply Type', _T_NOTE_CDN],
    [None, None, 'Note date', 'Debit note date format shall be (DD-MM-YYYY)'],
    [None, None, 'Note Value', 'Debit note value (In rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A Period', 'Period for which GSTR-1/IFF/1A has been filed'],
    [None, None, 'GSTR-1/IFF/1A Filing Date', 'Date on which GSTR-1/IFF/1A has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    # B2B-Rejected
    ['B2B-Rejected', 'ITC Rejected for taxable inward supplies received from registered persons', 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD],
    [None, None, 'Invoice number', 'Invoice number'],
    [None, None, 'Invoice type', _T_INV],
    [None, None, 'Invoice date', 'Invoice date format shall be DD-MM-YYYY'],
    [None, None, 'Invoice value', 'Invoice value (in rupees)'],
    [None, None, 'Place of supply', _T_POS],
    [None, None, 'Supply attract Reverse charge', _T_REV],
    [None, None, 'Taxable value', 'Taxable value'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_Y],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    [None, None, 'Source', _T_SRC],
    [None, None, 'IRN', _T_IRN],
    [None, None, 'IRN date', _T_IRNDT],
    # B2BA-Rejected
    ['B2BA-Rejected', 'ITC Rejected for amendments to previously filed invoices by supplier', 'Invoice number (Original details)', 'Original invoice number'],
    [None, None, 'Invoice date (Original details)', 'Original invoice date (Date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD3],
    [None, None, 'Invoice number', 'Revised Invoice number'],
    [None, None, 'Invoice type', _T_INV],
    [None, None, 'Invoice date', 'Invoice date format shall be DD-MM-YYYY'],
    [None, None, 'Invoice value', 'Invoice value (in rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    # B2B-CDNR-Rejected
    ['B2B-CDNR-Rejected', 'ITC Rejected for Debit/Credit notes (Original)', 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD2],
    [None, None, 'Note number', 'Debit/Credit note number'],
    [None, None, 'Note type', 'Document type can be Debit note or credit note'],
    [None, None, 'Note Supply Type', _T_NOTE_CDN],
    [None, None, 'Note date', 'Debit/Credit note date format shall be (DD-MM-YYYY)'],
    [None, None, 'Note Value', 'Debit/Credit note value (In rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    [None, None, 'Source', _T_SRC],
    [None, None, 'IRN', _T_IRN],
    [None, None, 'IRN date', _T_IRNDT],
    # B2B-CDNRA-Rejected
    ['B2B-CDNRA-Rejected', 'ITC Rejected for amendments to previously filed Credit/Debit notes by supplier', 'Note type(Original)', 'Note type can be Debit note or credit note'],
    [None, None, 'Note number(Original)', 'Original Debit/Credit note number'],
    [None, None, 'Note date(Original)', 'Original Debit/Credit note date (Note date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of Supplier', 'GSTIN of supplier'],
    [None, None, 'Trade/Legal name', _T_TRD2],
    [None, None, 'Note number', 'Debit/Credit note number'],
    [None, None, 'Note type', 'Note type can be Debit note or credit note'],
    [None, None, 'Note Supply Type', _T_NOTE_CDN],
    [None, None, 'Note date', 'Debit/Credit note date format shall be (DD-MM-YYYY)'],
    [None, None, 'Note Value', 'Debit/Credit note value (In rupees)'],
    [None, None, 'Place of supply', _T_POS2],
    [None, None, 'Supply attract Reverse charge', _T_REV2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Period', 'Period for which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'GSTR-1/IFF/1A/GSTR-5 Filing Date', 'Date on which GSTR-1/IFF/1A/GSTR-5 has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Applicable % of Tax Rate', _T_APPL],
    # ECO-Rejected
    ['ECO-Rejected', 'ITC Rejected for documents reported by ECO on which ECO is liable to pay tax us 9(5)', 'GSTIN of ECO', 'GSTIN of E-commerce operator'],
    [None, None, 'Trade/Legal name', 'Trade name of the E-commerce operator will be displayed. If trade name is not available, then legal name of the E-commerce operator'],
    [None, None, 'Document number', 'Document number'],
    [None, None, 'Document type', _T_DOC_ECO],
    [None, None, 'Document date', 'Document date format shall be DD-MM-YYYY'],
    [None, None, 'Document value', 'Document value (in rupees)'],
    [None, None, 'Place of supply', _T_POS_ECO],
    [None, None, 'Taxable value', 'Taxable value'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A period', 'Period for which GSTR-1/IFF has been filed'],
    [None, None, 'GSTR-1/IFF/1A filing date', 'Date on which GSTR-1/IFF has been filed'],
    [None, None, 'ITC availability', _T_ITC_Y],
    [None, None, 'Reason', _T_RSN],
    [None, None, 'Source', _T_SRC_ECO],
    [None, None, 'IRN', _T_IRN_ECO],
    [None, None, 'IRN date', _T_IRNDT_ECO],
    # ECOA-Rejected
    ['ECOA-Rejected', 'ITC Rejected for amendments to documents reported by ECO on which ECO is liable to pay tax u/s 9(5)', 'Document number (Original details)', 'Document invoice number'],
    [None, None, 'Document date (Original details)', 'Document invoice date (Date format shall be DD-MM-YYYY)'],
    [None, None, 'GSTIN of ECO', 'GSTIN of E-commerce operator'],
    [None, None, 'Trade/Legal name', 'Trade name of the E-commerce operator will be displayed. If trade name is not available then legal name of the E-commerce operator.'],
    [None, None, 'Document number', 'Revised Document number'],
    [None, None, 'Document type', _T_DOC_ECO],
    [None, None, 'Document date', 'Document date format shall be DD-MM-YYYY'],
    [None, None, 'Document value', 'Document value (in rupees)'],
    [None, None, 'Place of supply', _T_POS_ECO2],
    [None, None, 'Taxable value', 'Taxable value (In rupees)'],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'GSTR-1/IFF/1A Period', 'Period for which GSTR-1/IFF has been filed'],
    [None, None, 'GSTR-1/IFF/1A filing date', 'Date on which GSTR-1/IFF has been filed'],
    [None, None, 'ITC Availability', _T_ITC_IF],
    [None, None, 'Reason', _T_RSN],
    # ISD-Rejected
    ['ISD-Rejected', 'ITC Rejected for ISD Credits', 'GSTIN of ISD', 'Input Service Distributor GSTIN'],
    [None, None, 'Trade/Legal name of the ISD', 'Trade name of the ISD will be displayed. If trade name is not available then legal name of the ISD'],
    [None, None, 'ISD Document type', 'ISD document type can be Invoice or Credit note'],
    [None, None, 'ISD Document number', 'ISD invoice / ISD Credit note number'],
    [None, None, 'ISD Document date', 'ISD Document date format will be DD-MM-YYYY'],
    [None, None, 'Original ISD Invoice number', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Original ISD Invoice date', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'ISD GSTR-6 Period', 'Period for which GSTR-6 is to be filed.'],
    [None, None, 'ISD GSTR-6 Filing date', 'Date on which GSTR-6 has been filed.'],
    [None, None, 'Eligibilty of ITC', _T_ISD_ELG],
    # ISDA-Rejected
    ['ISDA-Rejected', 'ITC Rejected for amendments of ISD Credits received', 'ISD Document type (Original)', 'ISD document type can be Invoice or Credit note'],
    [None, None, 'ISD Document Number (Original)', 'Invoice/Credit note number'],
    [None, None, 'ISD Document date (Original)', 'Invoice/Credit note date'],
    [None, None, 'GSTIN of ISD', 'GSTIN of the Input Service Distributor'],
    [None, None, 'Trade/Legal name of the ISD', 'Trade name of the ISD will be displayed. If trade name is not available then legal name of the ISD'],
    [None, None, 'ISD Document type', 'ISD document type can be Invoice or Credit note'],
    [None, None, 'ISD Document number', 'ISD invoice / ISD Credit note number'],
    [None, None, 'ISD Document date', 'ISD Document date format will be DD-MM-YYYY'],
    [None, None, 'Original ISD Invoice number', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Original ISD Invoice date', "This is applicable only if ISD document type is 'Credit note' is linked to invoice"],
    [None, None, 'Integrated Tax', 'Integrated Tax amount (In rupees)'],
    [None, None, 'Central Tax', 'Central Tax amount (In rupees)'],
    [None, None, 'State/UT tax', 'State/UT tax amount (In rupees)'],
    [None, None, 'Cess', 'Cess amount (In rupees)'],
    [None, None, 'ISD GSTR-6 Period', 'Period for which GSTR-6 is to be filed.'],
    [None, None, 'ISD GSTR-6 Filing date', 'Date on which GSTR-6 has been filed.'],
    [None, None, 'Eligibilty of ITC', _T_ISD_ELG2],
]

_NAVY = "2F75B5"
_WHITE = "FFFFFF"
_BORDER_CLR = "D2D0CE"

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _side(): return Side(style="thin", color=_BORDER_CLR)
def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _align(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_MONTH_ABBR = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun',
               '07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
_MONTH_FULL = {'01':'January','02':'February','03':'March','04':'April','05':'May',
               '06':'June','07':'July','08':'August','09':'September','10':'October',
               '11':'November','12':'December'}

def _format_period(fp):
    if not fp or len(fp) != 6: return fp or ''
    mm, yy = fp[:2], fp[4:]
    return f"{_MONTH_FULL.get(mm, mm)}'{yy[-2:]}"

def _period_short(fp):
    """042025 -> 'Apr- 25'"""
    if not fp or len(fp) != 6: return fp or ''
    mm, yy = fp[:2], fp[4:]
    return f"{_MONTH_ABBR.get(mm, mm)}- {yy[-2:]}"

def _fmt_slash(d):
    return str(d).replace('-', '/') if d else ''

def _state_title(code):
    name = _state_name(code)
    return name.title() if name else ''

_INV_TYPES = {'R': 'Regular', 'SEWP': 'SEZ supplies with payment',
              'SEWOP': 'SEZ supplies without payment', 'DE': 'Deemed Exports',
              'CBW': 'Customs Bonded Warehouse'}
def _inv_type(code):
    k = str(code).strip().upper() if code else ''
    return _INV_TYPES.get(k, str(code) if code else '')

def _yn(code):
    s = str(code).strip().upper() if code else ''
    return 'Yes' if s == 'Y' else ('No' if s == 'N' else s)

_STD_RATES = {0, 0.1, 0.25, 1, 1.5, 3, 5, 6, 7.5, 12, 18, 28}
def _calc_rate(igst, cgst, sgst, txval):
    if not txval: return ''
    pct = round(((igst or 0) + (cgst or 0) + (sgst or 0)) / txval * 100, 1)
    for r in _STD_RATES:
        if abs(pct - r) < 0.5:
            return str(int(r)) if r == int(r) else str(r)
    return 'Multi Rate'

def _write_data_rows(ws, start_row, rows):
    bdr = _border()
    body_font = Font(size=11, name="Calibri")
    for r_idx, row in enumerate(rows):
        fill = _fill("F3F2F1") if (start_row + r_idx) % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            cell.border = bdr
            cell.font = body_font
            if fill:
                cell.fill = fill
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = "#,##0.00"
                cell.alignment = _align("right")
            else:
                cell.alignment = _align("left")


def _apply_header_merges(ws, hdr_rows, start_row):
    """Merge header cells based on None patterns:
    - Non-None followed by Nones in same row → horizontal span
    - Non-None in row1, None below in row2 → vertical span (2 rows)
    """
    if not hdr_rows:
        return
    row1 = hdr_rows[0]
    row2 = hdr_rows[1] if len(hdr_rows) > 1 else []
    r1, r2 = start_row, start_row + 1

    i = 0
    while i < len(row1):
        if row1[i] is None:
            i += 1
            continue
        # find end of horizontal span
        end = i + 1
        while end < len(row1) and row1[end] is None:
            end += 1
        h_span = end - i - 1
        col_s = i + 1  # 1-based
        col_e = col_s + h_span
        if h_span > 0:
            ws.merge_cells(start_row=r1, start_column=col_s,
                           end_row=r1, end_column=col_e)
        elif row2 and i < len(row2) and row2[i] is None:
            ws.merge_cells(start_row=r1, start_column=col_s,
                           end_row=r2, end_column=col_s)
        i = end


def _setup_data_sheet(ws, section_name, col_headers):
    """
    Rows 1-3 : merged title  (dark navy  203764, white size-22)
    Row  4   : section name  (yellow FFF2CC, bold size-11)
    Row  5   : [amended only] Original / Revised super-header (yellow / lavender)
    Row 5/6+ : navy column headers with span merges
    Returns first data row number.
    """
    _TITLE_BG  = "203764"
    _SECT_BG   = "FFF2CC"
    _ORIG_BG   = "FFF2CC"
    _REV_BG    = "DCC8DC"
    _HDR_BG    = "203764"
    bdr = _border()

    # ── detect amended super-header ─────────────────────────────────────────
    has_super = (bool(col_headers) and bool(col_headers[0]) and
                 col_headers[0][0] == 'Original Details')
    hdr_rows  = col_headers[1:] if has_super else col_headers
    n_cols    = max((len(r) for r in col_headers), default=1)
    last_col  = get_column_letter(n_cols)

    # ── rows 1-3 : title ────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}3")
    c = ws.cell(1, 1, "Goods and Services Tax  - GSTR-2B")
    c.font      = Font(color=_WHITE, size=22, name="Calibri")
    c.fill      = _fill(_TITLE_BG)
    c.alignment = _align("left")
    ws.row_dimensions[1].height = 42
    for r in range(1, 4):
        for col in range(1, n_cols + 1):
            ws.cell(r, col).fill   = _fill(_TITLE_BG)
            ws.cell(r, col).border = bdr

    # ── row 4 : section name ────────────────────────────────────────────────
    ws.merge_cells(f"A4:{last_col}4")
    c = ws.cell(4, 1, section_name)
    c.font      = Font(bold=True, size=11, name="Calibri")
    c.fill      = _fill(_SECT_BG)
    c.alignment = _align("left")
    ws.row_dimensions[4].height = 18
    for col in range(1, n_cols + 1):
        ws.cell(4, col).fill   = _fill(_SECT_BG)
        ws.cell(4, col).border = bdr

    # ── row 5 : amended "Original / Revised Details" super-header ───────────
    hdr_start = 5
    if has_super:
        super_row = col_headers[0]
        # find where Revised Details begins (first non-None after index 0)
        orig_end = 1
        for j in range(1, len(super_row)):
            if super_row[j] is None:
                orig_end = j + 1
            else:
                break  # found 'Revised Details'
        rev_start = orig_end + 1

        # Original Details span
        if orig_end >= 1:
            ws.merge_cells(start_row=5, start_column=1,
                           end_row=5, end_column=orig_end)
        c = ws.cell(5, 1, 'Original Details')
        c.font = Font(bold=True, size=11, name="Calibri")
        c.fill = _fill(_ORIG_BG)
        c.alignment = _align("left")
        for col in range(1, orig_end + 1):
            ws.cell(5, col).fill   = _fill(_ORIG_BG)
            ws.cell(5, col).border = bdr

        # Revised Details span
        if rev_start <= n_cols:
            ws.merge_cells(start_row=5, start_column=rev_start,
                           end_row=5, end_column=n_cols)
            c = ws.cell(5, rev_start, 'Revised Details')
            c.font = Font(bold=True, size=11, name="Calibri")
            c.fill = _fill(_REV_BG)
            c.alignment = _align("left")
            for col in range(rev_start, n_cols + 1):
                ws.cell(5, col).fill   = _fill(_REV_BG)
                ws.cell(5, col).border = bdr
        hdr_start = 6

    # ── regular header rows (navy) ──────────────────────────────────────────
    hdr_font = Font(bold=True, color=_WHITE, size=9, name="Calibri")
    hdr_fill = _fill(_HDR_BG)
    for r_idx, row_hdrs in enumerate(hdr_rows):
        row = hdr_start + r_idx
        for c_idx, val in enumerate(row_hdrs, 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.border    = bdr
            cell.alignment = _align("center", wrap=True)

    _apply_header_merges(ws, hdr_rows, hdr_start)

    data_start = hdr_start + len(hdr_rows)
    if data_start > 6:
        ws.freeze_panes = f"A{data_start}"
    return data_start

def _setup_summary_sheet(ws, title, sub_title, section_name, col_headers):
    """Summary sheets: title=row1, sub=row2, section=row5, headers=row6. Returns first data row."""
    ws.cell(1, 1, title).font = Font(bold=True, size=11)
    if sub_title:
        ws.cell(2, 1, sub_title).font = Font(size=9)
    if section_name:
        ws.cell(5, 1, section_name).font = Font(bold=True, size=10)
    _write_hdr_rows(ws, col_headers, 6)
    return 6 + len(col_headers)


def _write_summary_sheet(wb, name, section_title, itcsumm, mode='avl'):
    ws = wb.create_sheet(name)

    _DARK_NAVY = "002060"
    _PEACH     = "F4B084"
    _DARK_BLUE = "1F4E78"
    _ORANGE    = "C65911"
    bdr = _border()

    # helpers
    def gv(d, key=None):
        src = d.get(key) if (key and d) else d
        if not src: return [0.0, 0.0, 0.0, 0.0]
        return [_n(src.get("igst")), _n(src.get("cgst")), _n(src.get("sgst")), _n(src.get("cess"))]
    def z(): return [0.0, 0.0, 0.0, 0.0]

    def sc(r, c, val=None, font=None, fill=None, align=None, num_fmt=None):
        cell = ws.cell(r, c, val)
        cell.border = bdr
        if font:    cell.font    = font
        if fill:    cell.fill    = fill
        if align:   cell.alignment = align
        if num_fmt: cell.number_format = num_fmt
        return cell

    def fill_row(row, hex_color, start=1, end=10):
        for col in range(start, end + 1):
            ws.cell(row, col).fill = _fill(hex_color)
            ws.cell(row, col).border = bdr

    # ── Row 1: Title ────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    sc(1, 1, "FORM GSTR-2B",
       font=Font(bold=True, color=_WHITE, size=11, name="Calibri"),
       fill=_fill(_DARK_NAVY),
       align=_align("center"))
    fill_row(1, _DARK_NAVY)

    # ── Rows 2-4: Subtitle ──────────────────────────────────────────────────
    if mode == 'rej':
        sub = ("FORM GSTR-2B has been generated on the basis of the information furnished by your suppliers "
               "in their respective FORMS GSTR-1/IFF including E-Commerce supplies, GSTR-1A, 5 and 6. "
               "It also contains information on imports of goods from the ICEGATE system. "
               "This information is for guidance purposes only.")
    else:
        sub = ("FORM GSTR-2B has been generated on the basis of the information furnished by your suppliers "
               "in their respective FORMS GSTR-1,5 and 6. It also contains information on imports of goods "
               "from the ICEGATE system. This information is for guidance purposes only.")
    ws.merge_cells("A2:J4")
    sc(2, 1, sub,
       font=Font(bold=True, size=10, name="Calibri"),
       fill=_fill(_PEACH),
       align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    for r in range(2, 5):
        fill_row(r, _PEACH)
    ws.row_dimensions[2].height = 60

    # ── Row 5: Section title ─────────────────────────────────────────────────
    ws.merge_cells("A5:J5")
    sc(5, 1, section_title,
       font=Font(bold=True, color=_WHITE, size=10, name="Calibri"),
       fill=_fill(_DARK_BLUE),
       align=_align("center"))
    fill_row(5, _DARK_BLUE)

    # ── Row 6: Column headers ────────────────────────────────────────────────
    ws.merge_cells("H6:J6")
    hdr_font = Font(bold=True, color=_WHITE, size=9, name="Calibri")
    hdr_fill = _fill(_DARK_NAVY)
    hdr_vals = ["S.no.", "Heading", "GSTR-3B\ntable",
                "Integrated\nTax  (₹)", "Central\nTax (₹)",
                "State/UT\nTax (₹)", "Cess  (₹)", "Advisory"]
    for ci, val in enumerate(hdr_vals, 1):
        sc(6, ci, val,
           font=hdr_font, fill=hdr_fill,
           align=_align("center", wrap=True))
    # cols 9-10 are part of H6:J6 merge — just apply fill/border
    for ci in (9, 10):
        ws.cell(6, ci).fill   = hdr_fill
        ws.cell(6, ci).border = bdr
    ws.row_dimensions[6].height = 33.75

    # ── Column widths ────────────────────────────────────────────────────────
    for col, w in zip("ABCDEFGHIJ",
                      [6.14, 29, 6.57, 17, 13, 13, 13, 13, 13, 15.43]):
        ws.column_dimensions[col].width = w

    # ── Build content: list of tagged tuples ─────────────────────────────────
    # ('banner', text)
    # ('part',   label, text)
    # ('section', num_str, heading, gstr3b, vals, advisory, [(detail_heading, vals), ...])

    if mode == 'avl':
        avl       = itcsumm.get("itcavl") or {}
        ns        = avl.get("nonrevsup") or {}
        os        = avl.get("othersup")  or {}
        rs        = avl.get("revsup")    or {}
        isd_d     = avl.get("isd")       or {}
        impg_d    = avl.get("impg")      or {}
        impgsez_d = avl.get("impgsez")   or {}
        content = [
            ('banner', 'Credit which may be availed under FORM GSTR-3B'),
            ('part', 'Part A',
             'ITC Available - Credit may be claimed in relevant headings in GSTR-3B'),
            ('section', 'I',
             'All other ITC - Supplies from registered persons other than reverse charge',
             '4(A)(5)', gv(ns),
             'Net input tax credit may be availed under Table 4(A)(5) of FORM GSTR-3B.',
             [('B2B - Invoices (IMS)',               gv(ns, 'b2b')),
              ('B2B - Debit notes (IMS)',             gv(ns, 'cdnr')),
              ('ECO - Documents (IMS)',               gv(ns, 'eco')),
              ('B2B - Invoices (Amendment) (IMS)',    gv(ns, 'b2ba')),
              ('B2B - Debit notes (Amendment) (IMS)', gv(ns, 'cdnra')),
              ('ECO - Documents (Amendment) (IMS)',   gv(ns, 'ecoa'))]),
            ('section', 'II',
             'Inward Supplies from ISD',
             '4(A)(4)', gv(isd_d),
             'Net input tax credit may be availed under Table 4(A)(4) of FORM GSTR-3B.',
             [('ISD - Invoices',              gv(isd_d, 'isd')),
              ('ISD - Invoices (Amendment)',  gv(isd_d, 'isda'))]),
            ('section', 'III',
             'Inward Supplies liable for reverse charge',
             '3.1(d)\n4(A)(3)', gv(rs),
             ('These supplies shall be declared in Table 3.1(d) of FORM GSTR-3B for payment of tax. '
              'Net input tax credit may be availed under Table 4(A)(3) of FORM GSTR-3B on payment of tax.'),
             [('B2B - Invoices',               gv(rs, 'b2b')),
              ('B2B - Debit notes',             z()),
              ('B2B - Invoices (Amendment)',    gv(rs, 'b2ba')),
              ('B2B - Debit notes (Amendment)', z())]),
            ('section', 'IV',
             'Import of Goods',
             '4(A)(1)', gv(impg_d),
             'Net input tax credit may be availed under Table 4(A)(1) of FORM GSTR-3B.',
             [('IMPG - Import of goods from overseas',  gv(impg_d, 'impg')),
              ('IMPG (Amendment)',                       gv(impg_d, 'impga')),
              ('IMPGSEZ - Import of goods from SEZ',    gv(impgsez_d, 'impgsez')),
              ('IMPGSEZ (Amendment)',                    gv(impgsez_d, 'impgseza'))]),
            ('part', 'Part B',
             'ITC Available - Credit notes should be net off against relevant ITC available headings in GSTR-3B'),
            ('section', 'I',
             'Others',
             '4(A)', gv(os),
             'Credit Notes shall be net-off against relevant ITC available headings in GSTR-3B.',
             [('B2B - Credit notes (IMS)',                          '4(A)(5)', gv(os, 'cdnr')),
              ('B2B - Credit notes (Amendment) (IMS)',               '4(A)(5)', gv(os, 'cdnra')),
              ('B2B - Credit notes (Reverse charge)',                '3.1(d)\n4(A)(3)', z()),
              ('B2B - Credit notes (Reverse charge)(Amendment)',     '3.1(d)\n4(A)(3)', z()),
              ('ISD - Credit notes',                                 '4(A)(4)', z()),
              ('ISD - Credit notes (Amendment)',                     '4(A)(4)', z())]),
        ]

    elif mode == 'nonavl':
        nonavl = itcsumm.get("itcnonavl") or {}
        ns     = nonavl.get("nonrevsup") or {}
        os     = nonavl.get("othersup")  or {}
        rs     = nonavl.get("revsup")    or {}
        isd_d  = nonavl.get("isd")       or {}
        content = [
            ('banner', 'Credit which may not be availed under FORM GSTR-3B'),
            ('part', 'Part A', 'ITC Not Available'),
            ('section', 'I',
             'All other ITC - Supplies from registered persons other than reverse charge',
             '4(D)(2)', gv(ns),
             'Such credit shall not be taken and has to be reported under Table 4(D)(2) of FORM GSTR-3B.',
             [('B2B - Invoices',               gv(ns, 'b2b')),
              ('ECO - Documents',               z()),
              ('B2B - Debit notes',             z()),
              ('B2B - Invoices (Amendment)',    gv(ns, 'b2ba')),
              ('B2B - Debit notes (Amendment)', z()),
              ('ECO - Documents (Amendment)',   z())]),
            ('section', 'II',
             'Inward Supplies from ISD',
             '4(D)(2)', gv(isd_d),
             'Such credit shall not be taken and has to be reported under Table 4(D)(2) of FORM GSTR-3B.',
             [('ISD - Invoices',             gv(isd_d, 'isd')),
              ('ISD - Invoices (Amendment)', gv(isd_d, 'isda'))]),
            ('section', 'III',
             'Inward Supplies liable for reverse charge',
             '3.1(d)\n4(D)(2)', gv(rs),
             'These supplies shall be declared in Table 3.1(d) of FORM GSTR-3B for payment of tax.',
             [('B2B - Invoices',               gv(rs, 'b2b')),
              ('B2B - Debit notes',             z()),
              ('B2B - Invoices (Amendment)',    gv(rs, 'b2ba')),
              ('B2B - Debit notes (Amendment)', z())]),
            ('part', 'Part B',
             'ITC Not Available - Credit notes should be net off against relevant ITC available headings in GSTR-3B'),
            ('section', 'I',
             'Others',
             '4(A)', gv(os),
             'Credit Notes should be net-off against relevant ITC available headings in GSTR-3B.',
             [('B2B - Credit notes',                            '4(A)(5)', gv(os, 'cdnr')),
              ('B2B - Credit notes (Amendment)',                 '4(A)(5)', gv(os, 'cdnra')),
              ('B2B - Credit notes (Reverse charge)',            '4(A)(3)', z()),
              ('B2B - Credit notes (Reverse charge)(Amendment)', '4(A)(3)', z()),
              ('ISD - Credit notes',                             '4(A)(4)', z()),
              ('ISD - Credit notes (Amendment)',                 '4(A)(4)', z())]),
        ]

    elif mode == 'rej':
        rej   = itcsumm.get("itcRejected") or {}
        ns    = rej.get("nonrevsup") or {}
        os    = rej.get("othersup")  or {}
        isd_d = rej.get("isd")       or {}
        content = [
            ('banner', 'Credit which is rejected on IMS Dashboard'),
            ('part', 'Part A', 'ITC Rejected - Others'),
            ('section', 'I',
             'All other ITC - Supplies from registered persons other than reverse charge (IMS)',
             'NA', gv(ns),
             'Input tax credit cannot be availed in FORM GSTR-3B.',
             [('B2B - Invoices (IMS)',               gv(ns, 'b2b')),
              ('B2B - Debit notes (IMS)',             z()),
              ('ECO - Documents (IMS)',               z()),
              ('B2B - Invoices (Amendment) (IMS)',    gv(ns, 'b2ba')),
              ('B2B - Debit notes (Amendment) (IMS)', z()),
              ('ECO - Documents (Amendment) (IMS)',   z())]),
            ('section', 'II',
             'Inward Supplies from ISD',
             'NA', gv(isd_d),
             'Input tax credit cannot be availed in FORM GSTR-3B.',
             [('ISD - Invoices',             gv(isd_d, 'isd')),
              ('ISD - Invoices (Amendment)', gv(isd_d, 'isda'))]),
            ('part', 'Part B',
             'Rejected Records - Credit notes rejected on IMS Dashboard'),
            ('section', 'I',
             'Others',
             'NA', gv(os),
             'These Credit Notes are not eligible to net-off against relevant ITC available headings in GSTR-3B.',
             [('B2B - Credit notes (IMS)',             'NA', gv(os, 'cdnr')),
              ('B2B - Credit notes (Amendment) (IMS)', 'NA', gv(os, 'cdnra')),
              ('ISD - Credit notes',                   'NA', z()),
              ('ISD - Credit notes (Amendment)',        'NA', z())]),
        ]

    else:  # mode == 'rev'
        rev = itcsumm.get("itcrev") or itcsumm.get("itcRev") or {}
        ns  = rev.get("nonrevsup") or {}
        content = [
            ('banner', 'Credit which may not be availed under FORM GSTR-3B'),
            ('part', 'Part A', 'ITC Reversed - Others'),
            ('section', 'I',
             'ITC Reversal on account of Rule 37A',
             '4(B)(2)', gv(ns),
             'Such credit shall be reversed and has to be reported under Table 4(B)(2) of FORM GSTR-3B.',
             [('B2B - Invoices',               gv(ns, 'b2b')),
              ('B2B - Debit notes',             z()),
              ('B2B - Invoices (Amendment)',    gv(ns, 'b2ba')),
              ('B2B - Debit notes (Amendment)', z())]),
        ]

    # ── Render rows starting at row 7 ────────────────────────────────────────
    cur = 7
    body_font  = Font(size=10, name="Calibri")
    bold_font  = Font(bold=True, size=10, name="Calibri")
    num_font   = Font(size=10, name="Calibri")
    rot_align  = Alignment(horizontal="center", vertical="center",
                           text_rotation=90, wrap_text=True)

    for item in content:
        if item[0] == 'banner':
            ws.merge_cells(f"A{cur}:J{cur}")
            sc(cur, 1, item[1],
               font=Font(bold=True, color=_WHITE, size=10, name="Calibri"),
               fill=_fill(_ORANGE),
               align=_align("left"))
            fill_row(cur, _ORANGE)
            cur += 1

        elif item[0] == 'part':
            _, label, text = item
            sc(cur, 1, label, font=bold_font, fill=_fill(_PEACH), align=_align("center"))
            ws.merge_cells(f"B{cur}:J{cur}")
            sc(cur, 2, text, font=bold_font, fill=_fill(_PEACH),
               align=_align("left", wrap=True))
            fill_row(cur, _PEACH, start=3)
            cur += 1

        elif item[0] == 'section':
            _, num_str, heading, gstr3b, vals, advisory, details = item

            # ── section header row ──────────────────────────────────────────
            sec_r = cur
            ws.merge_cells(f"H{sec_r}:J{sec_r}")
            sc(sec_r, 1, num_str, font=body_font, align=_align("center"))
            sc(sec_r, 2, heading, font=body_font, align=_align("left", wrap=True))
            sc(sec_r, 3, gstr3b,  font=body_font, align=_align("center", wrap=True))
            for ci, v in enumerate(vals, 4):
                sc(sec_r, ci, v, font=num_font,
                   align=_align("right"), num_fmt="#,##0.00")
            sc(sec_r, 8, advisory, font=body_font,
               align=_align("left", wrap=True))
            ws.cell(sec_r, 9).border  = bdr
            ws.cell(sec_r, 10).border = bdr
            ws.row_dimensions[sec_r].height = 24.75
            cur += 1

            # ── detail rows ─────────────────────────────────────────────────
            if details:
                det_start = cur
                has_det_gstr3b = any(len(d) >= 3 for d in details)
                for det in details:
                    if len(det) == 3:
                        d_hdr, d_gstr3b, d_vals = det
                    else:
                        d_hdr, d_vals = det
                        d_gstr3b = None
                    sc(cur, 1, None, align=rot_align)
                    sc(cur, 2, d_hdr, font=body_font, align=_align("left", wrap=True))
                    sc(cur, 3, d_gstr3b, font=body_font, align=_align("center", wrap=True))
                    for ci, v in enumerate(d_vals, 4):
                        sc(cur, ci, v, font=num_font,
                           align=_align("right"), num_fmt="#,##0.00")
                    for col in range(8, 11):
                        ws.cell(cur, col).border = bdr
                    cur += 1
                det_end = cur - 1

                # merge A, H:J across all detail rows; merge C only if no per-row gstr3b
                if det_end >= det_start:
                    if det_end > det_start:
                        ws.merge_cells(f"A{det_start}:A{det_end}")
                        if not has_det_gstr3b:
                            ws.merge_cells(f"C{det_start}:C{det_end}")
                    ws.merge_cells(f"H{det_start}:J{det_end}")

                # set "Details" rotated in the top-left merged cell
                c_a = ws.cell(det_start, 1)
                c_a.value     = "Details"
                c_a.alignment = rot_align
                c_a.font      = Font(size=9, name="Calibri")


def _write_data_sheet(wb, name, section_title, data_list, rtnprd='', is_b2b=False,
                      is_cdn=False, is_isd=False, is_impg=False, is_impgsez=False,
                      is_ecomm=False, is_amended=False, is_rejected_ims=False):
    ws = wb.create_sheet(name)
    # ---- Build headers based on sheet type ----
    if is_b2b:
        if is_rejected_ims:
            if is_amended:
                h = [
                    ['Original Details', None, 'Revised Details'] + [None] * 15,
                    ['Invoice number', 'Invoice Date', 'GSTIN of supplier', 'Trade/Legal name',
                     'Invoice Details', None, None, None,
                     'Place of supply', 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                     'GSTR-1/IFF/GSTR-5 Period', 'GSTR-1/IFF/GSTR-5 Filing Date', 'Applicable % of Tax Rate', 'Period'],
                    [None, None, None, None, 'Invoice number', 'Invoice type', 'Invoice Date',
                     'Invoice Value(₹)', None, None, 'Integrated Tax(₹)', 'Central Tax(₹)',
                     'State/UT Tax(₹)', 'Cess(₹)', None, None, None, None],
                ]
            else:
                h = [
                    ['GSTIN of supplier', 'Trade/Legal name', 'Invoice Details', None, None, None,
                     'Place of supply', 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                     'GSTR-1/IFF/GSTR-5 Period', 'GSTR-1/IFF/GSTR-5 Filing Date',
                     'Applicable % of Tax Rate', 'Source', 'IRN', 'IRN Date', 'Period'],
                    [None, None, 'Invoice number', 'Invoice type', 'Invoice Date', 'Invoice Value(₹)',
                     None, None, 'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                     None, None, None, None, None, None, None],
                ]
        elif is_amended:
            h = [
                ['Original Details', None, 'Revised Details'] + [None] * 19,
                ['Invoice number', 'Invoice Date', 'GSTIN of supplier', 'Trade/Legal name',
                 'Invoice Details', None, None, None,
                 'Place of supply', 'Supply Attract Reverse Charge', 'Rate(%)',
                 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                 'GSTR-1/5 Period', 'GSTR-1/5 Filing Date', 'ITC Availability', 'Reason',
                 'Applicable % of Tax Rate', 'Period'],
                [None, None, None, None, 'Invoice number', 'Invoice type', 'Invoice Date',
                 'Invoice Value(₹)', None, None, None, None,
                 'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                 None, None, None, None, None, None],
            ]
        else:
            h = [
                ['GSTIN of supplier', 'Trade/Legal name', 'Invoice Details', None, None, None,
                 'Place of supply', 'Supply Attract Reverse Charge', 'Rate(%)',
                 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                 'GSTR-1/5 Period', 'GSTR-1/5 Filing Date', 'ITC Availability', 'Reason',
                 'Applicable % of Tax Rate', 'Source', 'IRN', 'IRN Date', 'Period'],
                [None, None, 'Invoice number', 'Invoice type', 'Invoice Date', 'Invoice Value(₹)',
                 None, None, None, None,
                 'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                 None, None, None, None, None, None, None, None, None],
            ]

    elif is_cdn:
        if is_rejected_ims:
            if is_amended:
                h = [
                    ['Original Details', None, None, 'Revised Details'] + [None] * 15,
                    ['Note type', 'Note number', 'Note date', 'GSTIN of supplier', 'Trade/Legal name',
                     'Credit note/Debit note details', None, None, None, None,
                     'Place of supply', 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                     'GSTR-1/IFF/GSTR-5 Period', 'GSTR-1/IFF/GSTR-5 Filing Date', 'Applicable % of Tax Rate', 'Period'],
                    [None] * 6 + ['Note number', 'Note type', 'Note Supply type', 'Note date',
                                  None, None, 'Integrated Tax(₹)', 'Central Tax(₹)',
                                  'State/UT Tax(₹)', 'Cess(₹)', None, None, None, None],
                ]
            else:
                h = [
                    ['GSTIN of supplier', 'Trade/Legal name', 'Credit note/Debit note details',
                     None, None, None, None,
                     'Place of supply', 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                     'GSTR-1/IFF/GSTR-5 Period', 'GSTR-1/IFF/GSTR-5 Filing Date',
                     'Applicable % of Tax Rate', 'Source', 'IRN', 'IRN Date', 'Period'],
                    [None, None, 'Note number', 'Note type', 'Note Supply type', 'Note date',
                     'Note Value (₹)', None, None,
                     'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                     None, None, None, None, None, None, None],
                ]
        elif is_amended:
            h = [
                ['Original Details', None, None, 'Revised Details'] + [None] * 20,
                ['Note type', 'Note number', 'Note date', 'GSTIN of supplier', 'Trade/Legal name',
                 'Credit note/Debit note details', None, None, None, None,
                 'Place of supply', 'Supply Attract Reverse Charge', 'Rate(%)',
                 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                 'GSTR-1/5 Period', 'GSTR-1/5 Filing Date', 'ITC Availability', 'Reason',
                 'Applicable % of Tax Rate', 'Period'],
                [None] * 6 + ['Note number', 'Note type', 'Note Supply type', 'Note date',
                              'Note Value (₹)', None, None, None,
                              'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                              None, None, None, None, None, None],
            ]
        else:
            h = [
                ['GSTIN of supplier', 'Trade/Legal name', 'Credit note/Debit note details',
                 None, None, None, None,
                 'Place of supply', 'Supply Attract Reverse Charge', 'Rate(%)',
                 'Taxable Value (₹)', 'Tax Amount', None, None, None,
                 'GSTR-1/5 Period', 'GSTR-1/5 Filing Date', 'ITC Availability', 'Reason',
                 'Applicable % of Tax Rate', 'Source', 'IRN', 'IRN Date', 'Period'],
                [None, None, 'Note number', 'Note type', 'Note Supply type', 'Note date',
                 'Note Value (₹)', None, None, None, None,
                 'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                 None, None, None, None, None, None, None, None, None],
            ]

    elif is_isd:
        if is_rejected_ims:
            if is_amended:
                h = [
                    ['Original Details', None, None, 'Revised Details'] + [None] * 13,
                    ['ISD Document type', 'Document Number', 'Document date',
                     'GSTIN of ISD', 'Trade/Legal name', 'ISD Document type',
                     'ISD Document number', 'ISD Document date', 'Original Invoice Number',
                     'Original invoice date', 'Input tax distribution by ISD', None, None, None,
                     'ISD GSTR-6 Period', 'ISD GSTR-6 Filing Date', 'Period'],
                    [None] * 11 + ['Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                                   None, None, None],
                ]
            else:
                h = [
                    ['GSTIN of ISD', 'Trade/Legal name', 'ISD Document type',
                     'ISD Document number', 'ISD Document date', 'Original Invoice Number',
                     'Original invoice date', 'Input tax distribution by ISD', None, None, None,
                     'ISD GSTR-6 Period', 'ISD GSTR-6 Filing Date', 'Period'],
                    [None] * 8 + ['Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                                  None, None, None],
                ]
        elif is_amended:
            h = [
                ['Original Details', None, None, 'Revised Details'] + [None] * 13,
                ['ISD Document type', 'Document Number', 'Document date',
                 'GSTIN of ISD', 'Trade/Legal name', 'ISD Document type',
                 'ISD Document number', 'ISD Document date', 'Original Invoice Number',
                 'Original invoice date', 'Input tax distribution by ISD', None, None, None,
                 'ISD GSTR-6 Period', 'ISD GSTR-6 Filing Date', 'Eligibility of ITC', 'Period'],
                [None] * 11 + ['Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                               None, None, None, None],
            ]
        else:
            h = [
                ['GSTIN of ISD', 'Trade/Legal name', 'ISD Document type',
                 'ISD Document number', 'ISD Document date', 'Original Invoice Number',
                 'Original invoice date', 'Input tax distribution by ISD', None, None, None,
                 'ISD GSTR-6 Period', 'ISD GSTR-6 Filing Date', 'Eligibility of ITC', 'Period'],
                [None] * 8 + ['Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                              None, None, None, None],
            ]

    elif is_ecomm:
        if is_rejected_ims:
            if is_amended:
                h = [
                    ['Original Details', None, 'Revised Details'] + [None] * 13,
                    ['Document number', 'Document date', 'GSTIN of ECO', 'Trade/Legal name',
                     'Document details', None, None, None,
                     'Place of supply', 'Taxable value (₹)', 'Tax amount', None, None, None,
                     'GSTR-1/1A/IFF period', 'GSTR-1/1A/IFF filing date', 'Period'],
                    [None] * 5 + ['Document number', 'Document type', 'Document date',
                                  None, None,
                                  'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                                  None, None, None],
                ]
            else:
                h = [
                    ['GSTIN of ECO', 'Trade/Legal name', 'Document details', None, None, None,
                     'Place of supply', 'Taxable value (₹)', 'Tax amount', None, None, None,
                     'GSTR-1/1A/IFF period', 'GSTR-1/1A/IFF filing date', 'Source', 'IRN', 'IRN Date', 'Period'],
                    [None, None, 'Document number', 'Document type', 'Document date',
                     'Document value(₹)', None, None,
                     'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                     None, None, None, None, None, None],
                ]
        elif is_amended:
            h = [
                ['Original Details', None, 'Revised Details'] + [None] * 19,
                ['Document number', 'Document date', 'GSTIN of ECO', 'Trade/Legal name',
                 'Document details', None, None, None,
                 'Place of supply', 'Rate(%)', 'Taxable value (₹)', 'Tax amount', None, None, None,
                 'GSTR-1/IFF/GSTR-1A period', 'GSTR-1/IFF/GSTR-1A filing date',
                 'ITC availability', 'Reason', 'Source', 'IRN', 'IRN Date', 'Period'],
                [None] * 5 + ['Document number', 'Document type', 'Document date',
                              'Document value(₹)', None, None,
                              'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                              None, None, None, None, None, None, None, None],
            ]
        else:
            h = [
                ['GSTIN of ECO', 'Trade/Legal name', 'Document details', None, None, None,
                 'Place of supply', 'Rate(%)', 'Taxable value (₹)', 'Tax amount', None, None, None,
                 'GSTR-1/IFF/GSTR-1A period', 'GSTR-1/IFF/GSTR-1A filing date',
                 'ITC availability', 'Reason', 'Source', 'IRN', 'IRN Date', 'Period'],
                [None, None, 'Document number', 'Document type', 'Document date',
                 'Document value(₹)', None, None, None,
                 'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'Cess(₹)',
                 None, None, None, None, None, None, None, None],
            ]

    elif is_impg:
        h = [
            ['Icegate Reference Date', 'Port Code', 'Bill of Entry Details', None, None,
             'Amount of tax (₹)', None, 'Amended (Yes)', 'Period'],
            [None, None, 'Number', 'Date', 'Taxable Value',
             'Integrated Tax(₹)', 'Cess(₹)', None, None],
        ]

    elif is_impgsez:
        h = [
            ['GSTIN of supplier', 'Trade/Legal name', 'Icegate Reference Date', 'Port Code',
             'Bill of Entry Details', None, None, 'Amount of tax (₹)', None, 'Amended (Yes)', 'Period'],
            [None, None, None, None, 'Number', 'Date', 'Taxable Value',
             'Integrated Tax(₹)', 'Cess(₹)', None, None],
        ]
    else:
        h = [['Data']]

    next_row = _setup_data_sheet(ws, section_title, h)

    # ---- Build data rows ----
    rows = []
    # rtn_period = _period_short(rtnprd)  # Removed fixed period

    for party in data_list:
        ctin = _s(party.get("ctin") or party.get("stin") or "")
        trdnm = _s(party.get("trdnm") or party.get("lgl_nm") or "")
        
        # Use injected period if available (for consolidation), else fallback to arg
        curr_rtn_prd = party.get("rtn_prd") or rtnprd
        supprd = _format_period(party.get("supprd") or curr_rtn_prd)
        supfildt = _s(party.get("supfildt") or "")
        rtn_period_row = _period_short(curr_rtn_prd)

        if is_b2b:
            for inv in (party.get("inv") or []):
                igst = _n(inv.get("igst") or inv.get("iamt"))
                cgst = _n(inv.get("cgst") or inv.get("camt"))
                sgst = _n(inv.get("sgst") or inv.get("samt"))
                cess = _n(inv.get("cess") or inv.get("csamt"))
                txval = _n(inv.get("txval"))
                irn = _s(inv.get("irn") or "")
                irndt = _s(inv.get("irngendate") or "")
                srctyp = _s(inv.get("srctyp") or "")

                if is_rejected_ims:
                    if is_amended:
                        row = [_s(inv.get("oinum")), _s(inv.get("oidt")),
                               ctin, trdnm,
                               _s(inv.get("inum")), _inv_type(inv.get("typ")),
                               _s(inv.get("dt")), _n(inv.get("val")),
                               _state_title(inv.get("pos")),
                               txval, igst, cgst, sgst, cess,
                               supprd, supfildt, '100%', rtn_period_row]
                    else:
                        row = [ctin, trdnm,
                               _s(inv.get("inum")), _inv_type(inv.get("typ")),
                               _s(inv.get("dt")), _n(inv.get("val")),
                               _state_title(inv.get("pos")),
                               txval, igst, cgst, sgst, cess,
                               supprd, supfildt, '100%', srctyp or None, irn or None, irndt or None, rtn_period_row]
                elif is_amended:
                    row = [_s(inv.get("oinum")), _s(inv.get("oidt")),
                           ctin, trdnm,
                           _s(inv.get("inum")), _inv_type(inv.get("typ")),
                           _fmt_slash(inv.get("dt")), _n(inv.get("val")),
                           _state_title(inv.get("pos")), _yn(inv.get("rev")),
                           _calc_rate(igst, cgst, sgst, txval),
                           txval, igst, cgst, sgst, cess,
                           supprd, _fmt_slash(supfildt),
                           _yn(inv.get("itcavl")), _s(inv.get("rsn") or ""),
                           None, rtn_period_row]
                else:
                    row = [ctin, trdnm,
                           _s(inv.get("inum")), _inv_type(inv.get("typ")),
                           _fmt_slash(inv.get("dt")), _n(inv.get("val")),
                           _state_title(inv.get("pos")), _yn(inv.get("rev")),
                           _calc_rate(igst, cgst, sgst, txval),
                           txval, igst, cgst, sgst, cess,
                           supprd, _fmt_slash(supfildt),
                           _yn(inv.get("itcavl")), _s(inv.get("rsn") or ""),
                           None, srctyp or None, irn or None, irndt or None, rtn_period_row]
                rows.append(row)

        elif is_cdn:
            for nt in (party.get("nt") or []):
                igst = _n(nt.get("igst") or nt.get("iamt"))
                cgst = _n(nt.get("cgst") or nt.get("camt"))
                sgst = _n(nt.get("sgst") or nt.get("samt"))
                cess = _n(nt.get("cess") or nt.get("csamt"))
                txval = _n(nt.get("txval"))
                irn = _s(nt.get("irn") or "")
                irndt = _s(nt.get("irngendate") or "")
                srctyp = _s(nt.get("srctyp") or "")

                if is_rejected_ims:
                    if is_amended:
                        row = [_s(nt.get("ontyp")), _s(nt.get("onum")), _s(nt.get("odt")),
                               ctin, trdnm,
                               _s(nt.get("ntnum")), _s(nt.get("ntty")), _s(nt.get("suptyp")),
                               _s(nt.get("ndt")), _n(nt.get("val")),
                               _state_title(nt.get("pos")),
                               txval, igst, cgst, sgst, cess,
                               supprd, supfildt, '100%', rtn_period_row]
                    else:
                        row = [ctin, trdnm,
                               _s(nt.get("ntnum")), _s(nt.get("ntty")), _s(nt.get("suptyp")),
                               _s(nt.get("ndt")), _n(nt.get("val")),
                               _state_title(nt.get("pos")),
                               txval, igst, cgst, sgst, cess,
                               supprd, supfildt, '100%',
                               srctyp or None, irn or None, irndt or None, rtn_period_row]
                elif is_amended:
                    row = [_s(nt.get("ontyp")), _s(nt.get("onum")), _s(nt.get("odt")),
                           ctin, trdnm,
                           _s(nt.get("ntnum")), _s(nt.get("ntty")), _s(nt.get("suptyp")),
                           _s(nt.get("ndt")), _n(nt.get("val")),
                           _state_title(nt.get("pos")), _yn(nt.get("rev")),
                           _calc_rate(igst, cgst, sgst, txval),
                           txval, igst, cgst, sgst, cess,
                           supprd, _fmt_slash(supfildt),
                           _yn(nt.get("itcavl")), _s(nt.get("rsn") or ""), None, rtn_period_row]
                else:
                    row = [ctin, trdnm,
                           _s(nt.get("ntnum")), _s(nt.get("ntty")), _s(nt.get("suptyp")),
                           _s(nt.get("ndt")), _n(nt.get("val")),
                           _state_title(nt.get("pos")), _yn(nt.get("rev")),
                           _calc_rate(igst, cgst, sgst, txval),
                           txval, igst, cgst, sgst, cess,
                           supprd, _fmt_slash(supfildt),
                           _yn(nt.get("itcavl")), _s(nt.get("rsn") or ""),
                           None, srctyp or None, irn or None, irndt or None, rtn_period_row]
                rows.append(row)

        elif is_isd:
            for doc in (party.get("docdet") or party.get("inv") or [party]):
                igst = _n(doc.get("igst") or doc.get("iamt"))
                cgst = _n(doc.get("cgst") or doc.get("camt"))
                sgst = _n(doc.get("sgst") or doc.get("samt"))
                cess = _n(doc.get("cess") or doc.get("csamt"))
                if is_rejected_ims:
                    if is_amended:
                        row = [_s(doc.get("odoc_typ")), _s(doc.get("odoc_num")), _s(doc.get("odoc_dt")),
                               ctin, trdnm,
                               _s(doc.get("doc_typ")), _s(doc.get("doc_num")), _s(doc.get("doc_dt")),
                               _s(doc.get("oinum")), _s(doc.get("oidt")),
                               igst, cgst, sgst, cess,
                               supprd, supfildt, rtn_period_row]
                    else:
                        row = [ctin, trdnm,
                               _s(doc.get("doc_typ")), _s(doc.get("doc_num")), _s(doc.get("doc_dt")),
                               _s(doc.get("oinum")), _s(doc.get("oidt")),
                               igst, cgst, sgst, cess,
                               supprd, supfildt, rtn_period_row]
                elif is_amended:
                    row = [_s(doc.get("odoc_typ")), _s(doc.get("odoc_num")), _s(doc.get("odoc_dt")),
                           ctin, trdnm,
                           _s(doc.get("doc_typ")), _s(doc.get("doc_num")), _s(doc.get("doc_dt")),
                           _s(doc.get("oinum")), _s(doc.get("oidt")),
                           igst, cgst, sgst, cess,
                           supprd, supfildt, _s(doc.get("itcavl") or ""), rtn_period_row]
                else:
                    row = [ctin, trdnm,
                           _s(doc.get("doc_typ")), _s(doc.get("doc_num")), _s(doc.get("doc_dt")),
                           _s(doc.get("oinum")), _s(doc.get("oidt")),
                           igst, cgst, sgst, cess,
                           supprd, supfildt, _s(doc.get("itcavl") or ""), rtn_period_row]
                rows.append(row)

        elif is_ecomm:
            for doc in (party.get("doc") or party.get("inv") or []):
                igst = _n(doc.get("igst") or doc.get("iamt"))
                cgst = _n(doc.get("cgst") or doc.get("camt"))
                sgst = _n(doc.get("sgst") or doc.get("samt"))
                cess = _n(doc.get("cess") or doc.get("csamt"))
                txval = _n(doc.get("txval"))
                irn = _s(doc.get("irn") or "")
                irndt = _s(doc.get("irngendate") or "")
                srctyp = _s(doc.get("srctyp") or "")
                if is_rejected_ims:
                    if is_amended:
                        row = [_s(doc.get("odoc_num")), _s(doc.get("odoc_dt")),
                               ctin, trdnm,
                               _s(doc.get("doc_num")), _s(doc.get("doc_typ")),
                               _s(doc.get("doc_dt")), _n(doc.get("val")),
                               _state_title(doc.get("pos")),
                               txval, igst, cgst, sgst, cess,
                               supprd, supfildt, rtn_period_row]
                    else:
                        row = [ctin, trdnm,
                               _s(doc.get("doc_num")), _s(doc.get("doc_typ")),
                               _s(doc.get("doc_dt")), _n(doc.get("val")),
                               _state_title(doc.get("pos")),
                               txval, igst, cgst, sgst, cess,
                               supprd, supfildt,
                               srctyp or None, irn or None, irndt or None, rtn_period_row]
                elif is_amended:
                    row = [_s(doc.get("odoc_num")), _s(doc.get("odoc_dt")),
                           ctin, trdnm,
                           _s(doc.get("doc_num")), _s(doc.get("doc_typ")),
                           _s(doc.get("doc_dt")), _n(doc.get("val")),
                           _state_title(doc.get("pos")),
                           _calc_rate(igst, cgst, sgst, txval),
                           txval, igst, cgst, sgst, cess,
                           supprd, _fmt_slash(supfildt),
                           _yn(doc.get("itcavl")), _s(doc.get("rsn") or ""),
                           srctyp or None, irn or None, irndt or None, rtn_period_row]
                else:
                    row = [ctin, trdnm,
                           _s(doc.get("doc_num")), _s(doc.get("doc_typ")),
                           _s(doc.get("doc_dt")), _n(doc.get("val")),
                           _state_title(doc.get("pos")),
                           _calc_rate(igst, cgst, sgst, txval),
                           txval, igst, cgst, sgst, cess,
                           supprd, _fmt_slash(supfildt),
                           _yn(doc.get("itcavl")), _s(doc.get("rsn") or ""),
                           srctyp or None, irn or None, irndt or None, rtn_period_row]
                rows.append(row)

        elif is_impg or is_impgsez:
            item = party
            igst = _n(item.get("igst") or item.get("iamt"))
            cess = _n(item.get("cess") or item.get("csamt"))
            row = []
            if is_impgsez:
                row += [ctin, trdnm]
            row += [_s(item.get("icedt")), _s(item.get("portcode")),
                    _s(item.get("benum")), _s(item.get("bedt")),
                    _n(item.get("txval")), igst, cess,
                    _s(item.get("isamd") or ""), rtn_period_row]
            rows.append(row)

    _write_data_rows(ws, next_row, rows)

    # ---- Column widths ─────────────────────────────────────────────────────
    # GSTIN / ISD GSTIN columns get 20, trade name gets 20,
    # dates 13, amounts 14, long-text cols (reason, advisory) 35, others 13
    _W = {
        'A': 20, 'B': 13, 'C': 16, 'D': 14, 'E': 13, 'F': 20, 'G': 20,
        'H': 13, 'I': 13, 'J': 13, 'K': 13, 'L': 13, 'M': 13, 'N': 13,
        'O': 20, 'P': 13, 'Q': 13, 'R': 35, 'S': 20, 'T': 13, 'U': 13,
        'V': 13, 'W': 13, 'X': 13,
    }
    n_hdr_cols = max((len(r) for r in h), default=1)
    for ci in range(1, n_hdr_cols + 1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = _W.get(col_letter, 13)


def _compute_itcsumm(docdata: dict, docRejdata: dict, docRevdata: dict) -> dict:
    """
    Build an itcsumm-compatible dict by aggregating tax amounts from raw invoice data.
    Used when the JSON doesn't carry a pre-computed itcsumm (monthly API downloads).
    """
    def _empty_bucket():
        return {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}

    def _add(bucket, inv):
        bucket["igst"] += _n(inv.get("igst"))
        bucket["cgst"] += _n(inv.get("cgst"))
        bucket["sgst"] += _n(inv.get("sgst"))
        bucket["cess"] += _n(inv.get("cess"))

    # ── itcavl ───────────────────────────────────────────────────────────────
    avl_ns_b2b   = _empty_bucket()  # nonrevsup b2b  (rev=N, itcavl=Y)
    avl_ns_b2ba  = _empty_bucket()  # nonrevsup b2ba (rev=N, itcavl=Y)
    avl_ns_cdn   = _empty_bucket()  # nonrevsup cdn  (rev=N, itcavl=Y)
    avl_rs_b2b   = _empty_bucket()  # revsup b2b     (rev=Y, itcavl=Y)
    avl_rs_b2ba  = _empty_bucket()  # revsup b2ba    (rev=Y, itcavl=Y)
    avl_isd_isd  = _empty_bucket()  # isd
    avl_isd_isda = _empty_bucket()  # isda
    avl_impg     = _empty_bucket()  # impg
    avl_impga    = _empty_bucket()  # impga
    avl_impgsez  = _empty_bucket()  # impgsez
    avl_impgseza = _empty_bucket()  # impgseza

    # ── itcnonavl ────────────────────────────────────────────────────────────
    nav_ns_b2b   = _empty_bucket()
    nav_ns_b2ba  = _empty_bucket()
    nav_ns_cdn   = _empty_bucket()
    nav_rs_b2b   = _empty_bucket()
    nav_rs_b2ba  = _empty_bucket()
    nav_isd_isd  = _empty_bucket()
    nav_isd_isda = _empty_bucket()

    # ── itcRejected ──────────────────────────────────────────────────────────
    rej_ns_b2b   = _empty_bucket()
    rej_ns_b2ba  = _empty_bucket()
    rej_ns_cdn   = _empty_bucket()
    rej_isd_isd  = _empty_bucket()
    rej_isd_isda = _empty_bucket()

    # ── itcrev (Rule 37A) ────────────────────────────────────────────────────
    rev_ns_b2b   = _empty_bucket()
    rev_ns_b2ba  = _empty_bucket()

    # ── Process docdata.b2b ──────────────────────────────────────────────────
    for party in (docdata.get("b2b") or []):
        for inv in (party.get("inv") or []):
            is_rev  = _s(inv.get("rev")).upper() == "Y"
            is_avl  = _s(inv.get("itcavl")).upper() == "Y"
            if is_avl:
                _add(avl_rs_b2b if is_rev else avl_ns_b2b, inv)
            else:
                _add(nav_rs_b2b if is_rev else nav_ns_b2b, inv)

    # ── Process docdata.b2ba ─────────────────────────────────────────────────
    for party in (docdata.get("b2ba") or []):
        for inv_grp in (party.get("inv") or []):
            for inv in (inv_grp.get("itms") or [inv_grp]):
                is_rev = _s(inv.get("rev")).upper() == "Y"
                is_avl = _s(inv.get("itcavl")).upper() == "Y"
                if is_avl:
                    _add(avl_rs_b2ba if is_rev else avl_ns_b2ba, inv)
                else:
                    _add(nav_rs_b2ba if is_rev else nav_ns_b2ba, inv)

    # ── Process docdata.cdn (credit/debit notes) ─────────────────────────────
    for party in (docdata.get("cdn") or docdata.get("cdnr") or []):
        for nt in (party.get("nt") or []):
            is_avl = _s(nt.get("itcavl")).upper() == "Y"
            if is_avl:
                _add(avl_ns_cdn, nt)
            else:
                _add(nav_ns_cdn, nt)

    # ── Process docdata.isd ──────────────────────────────────────────────────
    for isd in (docdata.get("isd") or []):
        for doc in (isd.get("docdet") or isd.get("inv") or [isd]):
            is_avl = _s(doc.get("itcElg") or doc.get("elg") or "Y").upper() != "N"
            if is_avl:
                _add(avl_isd_isd, doc)
            else:
                _add(nav_isd_isd, doc)

    # ── Process docdata.isda ─────────────────────────────────────────────────
    for isd in (docdata.get("isda") or []):
        for doc in (isd.get("docdet") or isd.get("inv") or [isd]):
            is_avl = _s(doc.get("itcElg") or doc.get("elg") or "Y").upper() != "N"
            if is_avl:
                _add(avl_isd_isda, doc)
            else:
                _add(nav_isd_isda, doc)

    # ── Process docdata.impg ─────────────────────────────────────────────────
    for rec in (docdata.get("impg") or []):
        _add(avl_impg, rec)

    for rec in (docdata.get("impga") or []):
        _add(avl_impga, rec)

    # ── Process docdata.impgsez ──────────────────────────────────────────────
    for rec in (docdata.get("impgsez") or []):
        _add(avl_impgsez, rec)

    for rec in (docdata.get("impgseza") or []):
        _add(avl_impgseza, rec)

    # ── Process docRejdata.b2b ───────────────────────────────────────────────
    for party in (docRejdata.get("b2b") or []):
        for inv in (party.get("inv") or []):
            _add(rej_ns_b2b, inv)

    for party in (docRejdata.get("b2ba") or []):
        for inv_grp in (party.get("inv") or []):
            for inv in (inv_grp.get("itms") or [inv_grp]):
                _add(rej_ns_b2ba, inv)

    for party in (docRejdata.get("cdn") or docRejdata.get("cdnr") or []):
        for nt in (party.get("nt") or []):
            _add(rej_ns_cdn, nt)

    for isd in (docRejdata.get("isd") or []):
        for doc in (isd.get("docdet") or isd.get("inv") or [isd]):
            _add(rej_isd_isd, doc)

    for isd in (docRejdata.get("isda") or []):
        for doc in (isd.get("docdet") or isd.get("inv") or [isd]):
            _add(rej_isd_isda, doc)

    # ── Process docRevdata.b2b (Rule 37A reversals) ──────────────────────────
    for party in (docRevdata.get("b2b") or []):
        for inv in (party.get("inv") or []):
            _add(rev_ns_b2b, inv)

    for party in (docRevdata.get("b2ba") or []):
        for inv_grp in (party.get("inv") or []):
            for inv in (inv_grp.get("itms") or [inv_grp]):
                _add(rev_ns_b2ba, inv)

    # ── Aggregate totals ──────────────────────────────────────────────────────
    def _sum_buckets(*buckets):
        total = _empty_bucket()
        for b in buckets:
            for k in total:
                total[k] += b[k]
        return total

    avl_ns_total = _sum_buckets(avl_ns_b2b, avl_ns_b2ba)
    avl_rs_total = _sum_buckets(avl_rs_b2b, avl_rs_b2ba)
    avl_isd_total = _sum_buckets(avl_isd_isd, avl_isd_isda)
    avl_impg_total = _sum_buckets(avl_impg, avl_impga)
    avl_impgsez_total = _sum_buckets(avl_impgsez, avl_impgseza)

    nav_ns_total = _sum_buckets(nav_ns_b2b, nav_ns_b2ba, nav_ns_cdn)
    nav_rs_total = _sum_buckets(nav_rs_b2b, nav_rs_b2ba)
    nav_isd_total = _sum_buckets(nav_isd_isd, nav_isd_isda)

    rej_ns_total = _sum_buckets(rej_ns_b2b, rej_ns_b2ba, rej_ns_cdn)
    rej_isd_total = _sum_buckets(rej_isd_isd, rej_isd_isda)

    rev_ns_total = _sum_buckets(rev_ns_b2b, rev_ns_b2ba)

    return {
        "itcavl": {
            "nonrevsup": {**avl_ns_total, "b2b": avl_ns_b2b, "b2ba": avl_ns_b2ba, "cdn": avl_ns_cdn},
            "revsup":    {**avl_rs_total, "b2b": avl_rs_b2b, "b2ba": avl_rs_b2ba},
            "isd":       {**avl_isd_total, "isd": avl_isd_isd, "isda": avl_isd_isda},
            "impg":      {**avl_impg_total, "impg": avl_impg, "impga": avl_impga},
            "impgsez":   {**avl_impgsez_total, "impgsez": avl_impgsez, "impgseza": avl_impgseza},
        },
        "itcnonavl": {
            "nonrevsup": {**nav_ns_total, "b2b": nav_ns_b2b, "b2ba": nav_ns_b2ba, "cdn": nav_ns_cdn},
            "revsup":    {**nav_rs_total, "b2b": nav_rs_b2b, "b2ba": nav_rs_b2ba},
            "isd":       {**nav_isd_total, "isd": nav_isd_isd, "isda": nav_isd_isda},
        },
        "itcRejected": {
            "nonrevsup": {**rej_ns_total, "b2b": rej_ns_b2b, "b2ba": rej_ns_b2ba, "cdn": rej_ns_cdn},
            "isd":       {**rej_isd_total, "isd": rej_isd_isd, "isda": rej_isd_isda},
        },
        "itcrev": {
            "nonrevsup": {**rev_ns_total, "b2b": rev_ns_b2b, "b2ba": rev_ns_b2ba},
        },
    }


def gstr2b_to_excel(data: dict, out_path: str, profile: dict = None):
    wb = Workbook()
    wb.remove(wb.active)

    inner = data.get("data", data)
    docdata = inner.get("docdata") or {}
    itcsumm = inner.get("itcsumm") or {}
    docRejdata = inner.get("docRejdata") or {}
    docRevdata = inner.get("docRevdata") or inner.get("docrevdata") or {}

    if not itcsumm:
        itcsumm = _compute_itcsumm(docdata, docRejdata, docRevdata)
    rtnprd = inner.get("rtnprd") or ""
    gstin_main = inner.get("gstin") or ""

    lgl_nm = ""
    trd_nm = ""
    if profile:
        lgl_nm = profile.get("lgl_nm") or profile.get("bname") or ""
        trd_nm = profile.get("trdnm") or profile.get("bname") or ""
    if not lgl_nm:
        lgl_nm = inner.get("lgl_nm") or inner.get("trdnm") or ""
    if not trd_nm:
        trd_nm = inner.get("trdnm") or inner.get("lgl_nm") or ""

    fy_str = ""
    if rtnprd and len(rtnprd) == 6:
        mm = int(rtnprd[:2])
        yyyy = int(rtnprd[2:])
        fy_str = f"{yyyy} - {yyyy + 1}" if mm >= 4 else f"{yyyy - 1} - {yyyy}"

    # Read me sheet
    ws_rm = wb.create_sheet("Read me")
    _TITLE_BG = "203764"
    _YELLOW_BG = "FFF2CC"
    _PEACH_BG = "FCE4D6"
    _GREY_BG = "E4E4E4"

    ws_rm.merge_cells("A1:F3")
    for r in range(1, 4):
        for c in range(1, 7):
            cell = ws_rm.cell(r, c)
            cell.fill = _fill(_TITLE_BG)
            if r == 1 and c == 1:
                cell.value = "Goods and Services Tax  - GSTR-2B"
                cell.font = Font(bold=True, color=_WHITE, size=16)
                cell.alignment = _align("center", "center")

    rows_info = [
        ("Financial Year", fy_str, _YELLOW_BG),
        ("Tax Period", "Consolidated (Yearly)" if data.get("is_consolidated") else (_format_period(rtnprd).split("'")[0] if rtnprd else ""), _YELLOW_BG),
        ("GSTIN", gstin_main, _PEACH_BG),
        ("Legal Name", lgl_nm, _PEACH_BG),
        ("Trade Name (if any)", trd_nm, _PEACH_BG),
        ("Date of generation", datetime.now().strftime("%d/%m/%Y"), _PEACH_BG),
    ]
    for i, (label, val, bg) in enumerate(rows_info, 4):
        ws_rm.merge_cells(f"A{i}:B{i}")
        ws_rm.merge_cells(f"C{i}:F{i}")
        ws_rm.cell(i, 1, label).font = Font(bold=True, size=10)
        ws_rm.cell(i, 3, val)
        for col in range(1, 7):
            c = ws_rm.cell(i, col)
            c.fill = _fill(bg)
            c.border = _border()

    ws_rm.column_dimensions["A"].width = 11
    ws_rm.column_dimensions["B"].width = 25
    ws_rm.column_dimensions["C"].width = 26
    ws_rm.column_dimensions["D"].width = 29
    ws_rm.column_dimensions["E"].width = 22

    ws_rm.merge_cells("A11:F11")
    instr_hdr = ws_rm.cell(11, 1, "GSTR-2B Data Entry Instructions")
    instr_hdr.font = Font(bold=True)
    instr_hdr.fill = _fill(_GREY_BG)
    instr_hdr.alignment = _align("center")
    for c_idx, h in enumerate(["Worksheet Name", "GSTR-2B Table Reference", "Field Name", "Instructions"], 1):
        cell = ws_rm.cell(12, c_idx, h)
        cell.font = Font(bold=True)
        cell.fill = _fill(_GREY_BG)
        cell.border = _border()

    bdr = _border()
    body_font = Font(size=10)
    for r_idx, row in enumerate(_README_ROWS, 13):
        for c_idx, val in enumerate(row, 1):
            cell = ws_rm.cell(r_idx, c_idx, val)
            cell.font = body_font
            cell.border = bdr
            cell.alignment = _align("left", wrap=True)
    ws_rm.column_dimensions["A"].width = 20
    ws_rm.column_dimensions["B"].width = 38
    ws_rm.column_dimensions["C"].width = 30
    ws_rm.column_dimensions["D"].width = 60

    # Summary sheets
    _write_summary_sheet(wb, "ITC Available", "FORM SUMMARY - ITC Available", itcsumm, mode='avl')
    _write_summary_sheet(wb, "ITC not available", "FORM SUMMARY - ITC Not Available", itcsumm, mode='nonavl')
    _write_summary_sheet(wb, "ITC Reversal", "FORM SUMMARY - ITC Reversal", itcsumm, mode='rev')
    _write_summary_sheet(wb, "ITC Rejected", "FORM SUMMARY - ITC Rejected", itcsumm, mode='rej')

    # Data sheets
    kw = dict(rtnprd=rtnprd)
    _write_data_sheet(wb, "B2B", "Taxable inward supplies received from registered persons",
                      docdata.get("b2b") or [], **kw, is_b2b=True)
    _write_data_sheet(wb, "B2BA", "Amendments to previously filed invoices by supplier",
                      docdata.get("b2ba") or [], **kw, is_b2b=True, is_amended=True)
    _write_data_sheet(wb, "B2B-CDNR", "Debit/Credit notes (Original)",
                      docdata.get("cdn") or docdata.get("cdnr") or [], **kw, is_cdn=True)
    _write_data_sheet(wb, "B2B-CDNRA", "Amendments to previously filed Credit/Debit notes by supplier",
                      docdata.get("cdna") or docdata.get("cdnra") or [], **kw, is_cdn=True, is_amended=True)
    _write_data_sheet(wb, "ISD", "ISD Credits",
                      docdata.get("isd") or [], **kw, is_isd=True)
    _write_data_sheet(wb, "ISDA", "Amendments ISD Credits received",
                      docdata.get("isda") or [], **kw, is_isd=True, is_amended=True)
    _write_data_sheet(wb, "IMPG", "Import of goods from overseas on bill of entry",
                      docdata.get("impg") or [], **kw, is_impg=True)
    _write_data_sheet(wb, "IMPGA", "Import of goods from overseas on bill of entry (Amendment)",
                      docdata.get("impga") or [], **kw, is_impg=True)
    _write_data_sheet(wb, "IMPGSEZ", "Import of goods from SEZ units/developers on bill of entry",
                      docdata.get("impgsez") or [], **kw, is_impgsez=True)
    _write_data_sheet(wb, "IMPGSEZA", "Import of goods from SEZ units/developers on bill of entry (Amendment)",
                      docdata.get("impgseza") or [], **kw, is_impgsez=True)
    _write_data_sheet(wb, "Ecomm", "Documents reported by ECO on which ECO is liable to pay tax u/s 9(5)",
                      docdata.get("ecomm") or docdata.get("eco") or [], **kw, is_ecomm=True)
    _write_data_sheet(wb, "EcommA",
                      "Amendments to documents reported by ECO on which ECO is liable to pay tax u/s 9(5)",
                      docdata.get("ecomma") or docdata.get("ecoa") or [], **kw, is_ecomm=True, is_amended=True)
    # ITC Reversal data sheets (Rule 37A)
    _write_data_sheet(wb, "B2B (ITC Reversal)", "ITC Reversed - Others",
                      docRevdata.get("b2b") or [], **kw, is_b2b=True)
    _write_data_sheet(wb, "B2BA (ITC Reversal)",
                      "Amendments to previously filed invoices by supplier (ITC reversal)",
                      docRevdata.get("b2ba") or [], **kw, is_b2b=True, is_amended=True)
    _write_data_sheet(wb, "Debit notes (Original)", "Debit notes (Original)",
                      docRevdata.get("cdn") or docRevdata.get("dnr") or [], **kw, is_cdn=True)
    _write_data_sheet(wb, "B2B-DNRA",
                      "Amendments to previously filed Debit notes by supplier",
                      docRevdata.get("cdna") or docRevdata.get("dnra") or [],
                      **kw, is_cdn=True, is_amended=True)
    _write_data_sheet(wb, "B2B(Rejected)",
                      "ITC Rejected for taxable inward supplies received from registered persons",
                      docRejdata.get("b2b") or [], **kw, is_b2b=True, is_rejected_ims=True)
    _write_data_sheet(wb, "B2BA(Rejected)",
                      "ITC Rejected for amendments to previously filed invoices by supplier",
                      docRejdata.get("b2ba") or [], **kw, is_b2b=True, is_amended=True, is_rejected_ims=True)
    _write_data_sheet(wb, "B2B-CDNR(Rejected)", "ITC Rejected for Debit/Credit notes (Original)",
                      docRejdata.get("cdn") or docRejdata.get("cdnr") or [],
                      **kw, is_cdn=True, is_rejected_ims=True)
    _write_data_sheet(wb, "B2B-CDNRA(Rejected)",
                      "ITC Rejected for amendments to previously filed Credit/Debit notes by supplier",
                      docRejdata.get("cdna") or docRejdata.get("cdnra") or [],
                      **kw, is_cdn=True, is_amended=True, is_rejected_ims=True)
    _write_data_sheet(wb, "ECO(Rejected)",
                      "ITC Rejected for documents reported by ECO on which ECO is liable to pay tax us 9(5)",
                      docRejdata.get("ecomm") or docRejdata.get("eco") or [],
                      **kw, is_ecomm=True, is_rejected_ims=True)
    _write_data_sheet(wb, "ECOA(Rejected)",
                      "  ITC Rejected for amendments to documents reported by ECO on which ECO is liable to pay tax u/s 9(5)",
                      docRejdata.get("ecomma") or docRejdata.get("ecoa") or [],
                      **kw, is_ecomm=True, is_amended=True, is_rejected_ims=True)
    _write_data_sheet(wb, "ISD(Rejected)", "ITC Rejected for ISD Credits",
                      docRejdata.get("isd") or [], **kw, is_isd=True, is_rejected_ims=True)
    _write_data_sheet(wb, "ISDA(Rejected)", "ITC Rejected for amendments of ISD Credits received",
                      docRejdata.get("isda") or [], **kw, is_isd=True, is_amended=True, is_rejected_ims=True)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _deep_sum_dicts(target, source):
    """Recursively sums numeric values in nested dictionaries."""
    for k, v in source.items():
        if isinstance(v, dict):
            target[k] = _deep_sum_dicts(target.get(k, {}), v)
        elif isinstance(v, (int, float)):
            target[k] = target.get(k, 0.0) + v
        else:
            # For non-numeric, just keep the last value or first value
            if k not in target:
                target[k] = v
    return target

def gstr2b_consolidated_to_excel(data_list: list, out_path: str, profile: dict = None):
    """
    Consolidates multiple GSTR-2B JSON objects into a single yearly Excel workbook.
    Aggregates ITC summaries and concatenates document lists with period tracking.
    """
    if not data_list:
        wb = Workbook()
        wb.save(out_path)
        return

    # Master container for aggregated data
    master = {
        "itcsumm": {},
        "docdata": {},
        "docRejdata": {},
        "docRevdata": {},
        "gstin": data_list[0].get("gstin") or (data_list[0].get("data", {}).get("gstin") if isinstance(data_list[0].get("data"), dict) else ""),
        "lgl_nm": data_list[0].get("lgl_nm") or (data_list[0].get("data", {}).get("lgl_nm") if isinstance(data_list[0].get("data"), dict) else ""),
        "trdnm": data_list[0].get("trdnm") or (data_list[0].get("data", {}).get("trdnm") if isinstance(data_list[0].get("data"), dict) else ""),
        "is_consolidated": True
    }

    # Find the most common financial year from the list
    periods = []
    for data in data_list:
        inner = data.get("data", data)
        p = inner.get("rtnprd")
        if p: periods.append(p)
    
    if periods:
        # Just use the first one for the "Read Me" calculation
        master["rtnprd"] = periods[0]

    for data in data_list:
        inner = data.get("data", data)
        prd = inner.get("rtnprd") or ""
        
        # 1. Sum ITC Summary
        _deep_sum_dicts(master["itcsumm"], inner.get("itcsumm") or {})
        
        # 2. Merge Document Data (docdata)
        docdata_src = inner.get("docdata") or {}
        for section, parties in docdata_src.items():
            if section not in master["docdata"]:
                master["docdata"][section] = []
            
            # Parties is a list of objects (GSTIN level)
            for party in parties:
                # We copy the party and inject the return period so the 
                # row-level logic in _write_data_sheet can pick it up.
                p_copy = party.copy()
                p_copy["rtn_prd"] = prd
                master["docdata"][section].append(p_copy)

        # 3. Merge Rejected Document Data (docRejdata)
        docRej_src = inner.get("docRejdata") or {}
        for section, parties in docRej_src.items():
            if section not in master["docRejdata"]:
                master["docRejdata"][section] = []
            for party in parties:
                p_copy = party.copy()
                p_copy["rtn_prd"] = prd
                master["docRejdata"][section].append(p_copy)

        # 4. Merge Reversal Document Data (docRevdata)
        docRev_src = inner.get("docRevdata") or inner.get("docrevdata") or {}
        for section, parties in docRev_src.items():
            if section not in master["docRevdata"]:
                master["docRevdata"][section] = []
            for party in parties:
                p_copy = party.copy()
                p_copy["rtn_prd"] = prd
                master["docRevdata"][section].append(p_copy)

    # Finally, call the standard to_excel function with our master object
    gstr2b_to_excel(master, out_path, profile=profile)
