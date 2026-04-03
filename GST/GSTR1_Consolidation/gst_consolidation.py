import os
import glob
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────────────────────────────────────────────
# TARGET SHEET NAMES (same as VBA)
# ─────────────────────────────────────────────────────────
TARGET_SHEETS = [
    "b2b,sez,de", "b2ba", "b2cl", "b2cla", "b2cs", "b2csa",
    "cdnr", "cdnra", "cdnur", "cdnura", "exp", "expa",
    "at", "ata", "atadj", "atadja", "exemp", "hsn", "docs"
]

# ─────────────────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────────────────
BG_DARK    = "#0D0D1A"
BG_CARD    = "#141428"
BG_PANEL   = "#1A1A35"
ACCENT1    = "#FF6B35"   # orange
ACCENT2    = "#00D4FF"   # cyan
ACCENT3    = "#A855F7"   # purple
SUCCESS    = "#22C55E"   # green
WARNING    = "#F59E0B"   # amber
ERROR_CLR  = "#EF4444"   # red
TEXT_MAIN  = "#F0F4FF"
TEXT_DIM   = "#8892AA"
BTN_HOVER  = "#FF8C5A"

# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def find_sheet_ci(wb: Workbook, name: str):
    """Case-insensitive sheet lookup."""
    for sh in wb.sheetnames:
        if sh.strip().lower() == name.strip().lower():
            return wb[sh]
    return None


def get_or_create_sheet(wb: Workbook, name: str):
    sh = find_sheet_ci(wb, name)
    if sh is None:
        safe = name[:31]
        sh = wb.create_sheet(title=safe)
    return sh


def last_row(ws) -> int:
    """Return 0 if sheet is empty, else the last used row index."""
    for row in reversed(range(1, ws.max_row + 1)):
        for cell in ws[row]:
            if cell.value is not None:
                return row
    return 0


def sheet_has_data(ws) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                return True
    return False


def read_sheet_values(ws):
    """Return all rows as list-of-lists."""
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    return data


def strip_trailing_empty(rows):
    """Remove fully-None rows at end."""
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    return rows

# ─────────────────────────────────────────────────────────
# CONSOLIDATION ENGINE
# ─────────────────────────────────────────────────────────
def consolidate(folder_path: str, output_path: str, log_fn, progress_fn, done_fn):
    """Run in a background thread."""
    try:
        xls_files = glob.glob(os.path.join(folder_path, "*.xls*"))
        xls_files = [f for f in xls_files if not os.path.basename(f).startswith("~$")]
        # Exclude output file itself if it's in the same folder
        xls_files = [f for f in xls_files if os.path.abspath(f) != os.path.abspath(output_path)]

        if not xls_files:
            log_fn("⚠  No Excel files found in the selected folder.", WARNING)
            done_fn(False)
            return

        log_fn(f"📂  Found {len(xls_files)} file(s) to process.", ACCENT2)

        # Load or create destination workbook
        if os.path.exists(output_path):
            wb_dest = load_workbook(output_path)
            log_fn(f"📋  Loaded existing output file: {os.path.basename(output_path)}", ACCENT2)
        else:
            wb_dest = Workbook()
            # Remove default sheet
            if "Sheet" in wb_dest.sheetnames:
                del wb_dest["Sheet"]
            log_fn(f"🆕  Created new output file: {os.path.basename(output_path)}", ACCENT2)

        # Ensure all target sheets exist
        for t in TARGET_SHEETS:
            get_or_create_sheet(wb_dest, t)

        total = len(xls_files)
        sheets_copied = 0
        files_processed = 0

        for idx, fpath in enumerate(xls_files):
            fname = os.path.basename(fpath)
            log_fn(f"\n📄  Opening: {fname}", TEXT_MAIN)
            progress_fn(int((idx / total) * 90))

            try:
                wb_src = load_workbook(fpath, read_only=True, data_only=True)
            except Exception as e:
                log_fn(f"   ❌  Could not open {fname}: {e}", ERROR_CLR)
                continue

            for t in TARGET_SHEETS:
                ws_src = find_sheet_ci(wb_src, t)
                if ws_src is None:
                    continue

                rows = read_sheet_values(ws_src)
                rows = strip_trailing_empty(rows)

                if not rows:
                    continue

                ws_dest = get_or_create_sheet(wb_dest, t)
                dest_lr = last_row(ws_dest)

                if dest_lr == 0:
                    # Destination empty → copy all rows (headers + data)
                    copy_rows = rows
                    start_row = 1
                else:
                    # Skip header (first row) of source
                    copy_rows = rows[1:] if len(rows) > 1 else []
                    start_row = dest_lr + 1

                if not copy_rows:
                    continue

                for ri, row_data in enumerate(copy_rows):
                    for ci, val in enumerate(row_data):
                        ws_dest.cell(row=start_row + ri, column=ci + 1, value=val)

                log_fn(
                    f"   ✅  Sheet '{t}' → {len(copy_rows)} row(s) appended.", SUCCESS
                )
                sheets_copied += 1

            wb_src.close()
            files_processed += 1

        # Remove unused empty sheets (sheets that were created but got no data)
        for t in TARGET_SHEETS:
            ws = find_sheet_ci(wb_dest, t)
            if ws and not sheet_has_data(ws):
                del wb_dest[ws.title]

        wb_dest.save(output_path)
        progress_fn(100)
        log_fn(
            f"\n🎉  Done! Processed {files_processed} file(s), "
            f"consolidated {sheets_copied} sheet(s).",
            SUCCESS,
        )
        log_fn(f"💾  Saved to: {output_path}", ACCENT2)
        done_fn(True)

    except Exception as e:
        log_fn(f"\n💥  Unexpected error: {e}", ERROR_CLR)
        done_fn(False)


# ─────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────
class ChallExtractorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GST Sheet Consolidation Tool")
        self.geometry("1200x800")
        self.minsize(960, 680)
        self.configure(bg=BG_DARK)
        self.resizable(True, True)

        self.folder_var  = tk.StringVar(value="")
        self.output_var  = tk.StringVar(value="")
        self._running    = False

        self._build_ui()
        self._center()

    # ── layout ────────────────────────────────────────────
    def _build_ui(self):
        # ── Header bar ────────────────────────────────────
        hdr = tk.Frame(self, bg=BG_CARD, height=90)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        logo_lbl = tk.Label(
            hdr, text="⚡ GST SHEET CONSOLIDATION TOOL",
            bg=BG_CARD, fg=ACCENT1,
            font=("Consolas", 26, "bold"),
            padx=20
        )
        logo_lbl.pack(side="left", pady=15)

        sub_lbl = tk.Label(
            hdr, text="GST Sheet Consolidation Tool",
            bg=BG_CARD, fg=TEXT_DIM,
            font=("Consolas", 13),
        )
        sub_lbl.pack(side="left", pady=20)

        ts_lbl = tk.Label(
            hdr, text="v2.0", bg=BG_CARD, fg=ACCENT3,
            font=("Consolas", 14, "bold"), padx=20
        )
        ts_lbl.pack(side="right", pady=20)

        # ── Accent strip ──────────────────────────────────
        strip = tk.Frame(self, bg=ACCENT1, height=3)
        strip.pack(fill="x")

        # ── Main body ─────────────────────────────────────
        body = tk.Frame(self, bg=BG_DARK)
        body.pack(fill="both", expand=True, padx=24, pady=18)

        # Left panel (controls)
        left = tk.Frame(body, bg=BG_PANEL, bd=0, relief="flat")
        left.pack(side="left", fill="y", padx=(0, 16), ipadx=16, ipady=14)

        # ── Folder picker ─────────────────────────────────
        self._section_label(left, "📁  SOURCE FOLDER")
        self._path_row(
            left, self.folder_var,
            btn_text="Browse…", cmd=self._pick_folder
        )

        # ── Output file picker ────────────────────────────
        self._section_label(left, "💾  OUTPUT EXCEL FILE")
        self._path_row(
            left, self.output_var,
            btn_text="Save As…", cmd=self._pick_output
        )

        # ── Sheet list ────────────────────────────────────
        self._section_label(left, "🗂  SHEETS TO CONSOLIDATE")
        sheet_frame = tk.Frame(left, bg=BG_PANEL)
        sheet_frame.pack(fill="x", padx=8, pady=(0, 8))

        sb = tk.Scrollbar(sheet_frame, orient="vertical", bg=BG_PANEL)
        sheet_lb = tk.Listbox(
            sheet_frame, yscrollcommand=sb.set,
            bg=BG_DARK, fg=ACCENT2,
            font=("Consolas", 12),
            selectbackground=ACCENT3,
            activestyle="none",
            height=12, width=26,
            relief="flat", bd=0,
            highlightthickness=1,
            highlightcolor=ACCENT3,
            highlightbackground=BG_PANEL
        )
        for t in TARGET_SHEETS:
            sheet_lb.insert("end", f"  {t}")
        sb.config(command=sheet_lb.yview)
        sheet_lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # ── Run button ────────────────────────────────────
        self.run_btn = self._make_button(
            left, "▶  START CONSOLIDATION",
            bg=ACCENT1, fg="white",
            cmd=self._start
        )
        self.run_btn.pack(fill="x", padx=8, pady=(10, 4))

        self.clear_btn = self._make_button(
            left, "🧹  CLEAR LOG",
            bg=BG_DARK, fg=TEXT_DIM,
            cmd=self._clear_log,
            font_size=11
        )
        self.clear_btn.pack(fill="x", padx=8, pady=(0, 6))

        # Right panel (log + progress)
        right = tk.Frame(body, bg=BG_CARD, bd=0)
        right.pack(side="left", fill="both", expand=True)

        log_hdr = tk.Frame(right, bg=BG_CARD)
        log_hdr.pack(fill="x", padx=10, pady=(10, 4))
        tk.Label(
            log_hdr, text="📋  ACTIVITY LOG",
            bg=BG_CARD, fg=TEXT_DIM,
            font=("Consolas", 13, "bold")
        ).pack(side="left")

        self.ts_label = tk.Label(
            log_hdr, text="", bg=BG_CARD, fg=TEXT_DIM, font=("Consolas", 11)
        )
        self.ts_label.pack(side="right")

        # Log text widget
        log_frame = tk.Frame(right, bg=BG_CARD)
        log_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        log_scroll = tk.Scrollbar(log_frame, bg=BG_DARK)
        self.log_box = tk.Text(
            log_frame,
            bg=BG_DARK, fg=TEXT_MAIN,
            font=("Consolas", 12),
            relief="flat", bd=0,
            insertbackground=ACCENT2,
            wrap="word",
            state="disabled",
            yscrollcommand=log_scroll.set,
            highlightthickness=0,
            padx=6, pady=6
        )
        log_scroll.config(command=self.log_box.yview)
        self.log_box.pack(side="left", fill="both", expand=True)
        log_scroll.pack(side="right", fill="y")

        # Colour tags
        self.log_box.tag_config("cyan",    foreground=ACCENT2)
        self.log_box.tag_config("orange",  foreground=ACCENT1)
        self.log_box.tag_config("green",   foreground=SUCCESS)
        self.log_box.tag_config("red",     foreground=ERROR_CLR)
        self.log_box.tag_config("amber",   foreground=WARNING)
        self.log_box.tag_config("dim",     foreground=TEXT_DIM)
        self.log_box.tag_config("white",   foreground=TEXT_MAIN)

        # Progress bar
        prog_frame = tk.Frame(right, bg=BG_CARD)
        prog_frame.pack(fill="x", padx=10, pady=(0, 10))

        style = ttk.Style(self)
        style.theme_use("default")
        style.configure(
            "Neon.Horizontal.TProgressbar",
            troughcolor=BG_DARK,
            background=ACCENT1,
            thickness=12
        )
        self.progress = ttk.Progressbar(
            prog_frame, orient="horizontal",
            length=100, mode="determinate",
            style="Neon.Horizontal.TProgressbar"
        )
        self.progress.pack(fill="x")

        self.status_lbl = tk.Label(
            right, text="Ready — Select folder and output file, then click Start.",
            bg=BG_CARD, fg=TEXT_DIM, font=("Consolas", 11), anchor="w", padx=10
        )
        self.status_lbl.pack(fill="x", pady=(0, 6))

        # Initial welcome log
        self._log("Welcome to GST Sheet Consolidation Tool  ⚡", ACCENT1)
        self._log("Select a source folder and an output file to begin.\n", TEXT_DIM)

    # ── Helper widgets ────────────────────────────────────
    def _section_label(self, parent, text):
        tk.Label(
            parent, text=text,
            bg=BG_PANEL, fg=ACCENT2,
            font=("Consolas", 12, "bold"),
            anchor="w", padx=10
        ).pack(fill="x", pady=(16, 4))

    def _path_row(self, parent, var, btn_text, cmd):
        frame = tk.Frame(parent, bg=BG_PANEL)
        frame.pack(fill="x", padx=8, pady=(0, 6))

        entry = tk.Entry(
            frame, textvariable=var,
            bg=BG_DARK, fg=ACCENT2,
            font=("Consolas", 11),
            relief="flat", bd=0,
            insertbackground=ACCENT2,
            highlightthickness=1,
            highlightcolor=ACCENT3,
            highlightbackground=BG_DARK
        )
        entry.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 4))

        btn = tk.Button(
            frame, text=btn_text,
            bg=ACCENT3, fg="white",
            font=("Consolas", 11, "bold"),
            relief="flat", bd=0,
            cursor="hand2",
            padx=12, pady=6,
            command=cmd,
            activebackground=BTN_HOVER,
            activeforeground="white"
        )
        btn.pack(side="right")

    def _make_button(self, parent, text, bg, fg, cmd, font_size=13):
        btn = tk.Button(
            parent, text=text,
            bg=bg, fg=fg,
            font=("Consolas", font_size, "bold"),
            relief="flat", bd=0,
            cursor="hand2",
            pady=13,
            command=cmd,
            activebackground=BTN_HOVER,
            activeforeground="white"
        )
        btn.bind("<Enter>", lambda e: btn.config(bg=BTN_HOVER if bg == ACCENT1 else BG_PANEL))
        btn.bind("<Leave>", lambda e: btn.config(bg=bg))
        return btn

    # ── Actions ───────────────────────────────────────────
    def _pick_folder(self):
        path = filedialog.askdirectory(title="Select Source Folder")
        if path:
            self.folder_var.set(path)
            self._log(f"📁  Folder set: {path}", ACCENT2)
            # Auto-suggest output path
            if not self.output_var.get():
                default_out = os.path.join(
                    path,
                    f"GST_Consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                self.output_var.set(default_out)

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Output Excel As",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")]
        )
        if path:
            self.output_var.set(path)
            self._log(f"💾  Output set: {path}", ACCENT2)

    def _start(self):
        if self._running:
            return
        folder  = self.folder_var.get().strip()
        output  = self.output_var.get().strip()

        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Error", "Please select a valid source folder.")
            return
        if not output:
            messagebox.showerror("Error", "Please specify an output Excel file path.")
            return

        self._running = True
        self.run_btn.config(state="disabled", text="⏳  Processing…", bg=WARNING)
        self.progress["value"] = 0
        self.ts_label.config(text=datetime.now().strftime("%H:%M:%S"))
        self._log("\n" + "─" * 52, TEXT_DIM)
        self._log(f"🚀  Starting consolidation at {datetime.now().strftime('%H:%M:%S')}", ACCENT1)

        thread = threading.Thread(
            target=consolidate,
            args=(folder, output, self._log_thread, self._progress_thread, self._done),
            daemon=True
        )
        thread.start()

    def _clear_log(self):
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")
        self.progress["value"] = 0
        self.status_lbl.config(text="Log cleared. Ready.")

    # ── Thread-safe callbacks ─────────────────────────────
    def _log_thread(self, msg, colour):
        self.after(0, lambda m=msg, c=colour: self._log(m, c))

    def _progress_thread(self, val):
        self.after(0, lambda v=val: self._set_progress(v))

    def _done(self, success: bool):
        self.after(0, lambda: self._on_done(success))

    def _log(self, msg, colour=TEXT_MAIN):
        tag = {
            ACCENT2:   "cyan",
            ACCENT1:   "orange",
            SUCCESS:   "green",
            ERROR_CLR: "red",
            WARNING:   "amber",
            TEXT_DIM:  "dim",
        }.get(colour, "white")

        self.log_box.config(state="normal")
        self.log_box.insert("end", msg + "\n", tag)
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    def _set_progress(self, val):
        self.progress["value"] = val
        self.status_lbl.config(text=f"Progress: {val}%")

    def _on_done(self, success: bool):
        self._running = False
        self.run_btn.config(
            state="normal",
            text="▶  START CONSOLIDATION",
            bg=ACCENT1
        )
        if success:
            self.status_lbl.config(
                text=f"✅  Consolidation complete — {datetime.now().strftime('%H:%M:%S')}",
            )
            messagebox.showinfo(
                "Success",
                f"Consolidation complete!\n\nOutput saved to:\n{self.output_var.get()}"
            )
        else:
            self.status_lbl.config(text="❌  Consolidation failed — check the log.")

    # ── Utilities ─────────────────────────────────────────
    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw   = self.winfo_screenwidth()
        sh   = self.winfo_screenheight()
        x    = (sw - w) // 2
        y    = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")


# ─────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = ChallExtractorApp()
    app.mainloop()