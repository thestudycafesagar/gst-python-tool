import customtkinter as ctk
from tkinter import filedialog, messagebox
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import os
import sys
import threading
import time

# --- Configuration ---
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class GSTR3BConverterPro(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Window Setup
        self.title("GSTR-3B Pro Converter")
        self.geometry("700x650")
        self.resizable(True, True)

        # Variables
        self.selected_files = []
        self.merge_mode = ctk.BooleanVar(value=False) 
        
        # UI Layout
        self.create_widgets()

    def create_widgets(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # CONTENT AREA (SCROLLABLE)
        self.scroll_container = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll_container.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.scroll_container.grid_columnconfigure(0, weight=1)

        # Header
        self.header_frame = ctk.CTkFrame(self.scroll_container, fg_color="transparent")
        self.header_frame.pack(pady=20)
        
        self.title_label = ctk.CTkLabel(
            self.header_frame, 
            text="GSTR-3B Batch Processor", 
            font=("Segoe UI", 24)
        )
        self.title_label.pack()
        
        self.subtitle = ctk.CTkLabel(
            self.header_frame,
            text="Saves to: /GST Downloaded/ folder",
            text_color="#10B981"
        )
        self.subtitle.pack()

        # File Selection Area
        self.file_frame = ctk.CTkFrame(self.scroll_container)
        self.file_frame.pack(pady=10, padx=40, fill="x")

        self.select_btn = ctk.CTkButton(
            self.file_frame, 
            text="Select PDF Files", 
            command=self.select_files,
            width=200,
            height=40,
            font=("Segoe UI", 14)
        )
        self.select_btn.pack(pady=15)

        self.file_count_label = ctk.CTkLabel(
            self.file_frame,
            text="No files selected",
            text_color="gray"
        )
        self.file_count_label.pack(pady=(0, 10))

        self.btn_demo = ctk.CTkButton(
            self.file_frame, text="▶ View Demo", command=self.open_demo_link,
            fg_color="#DC2626", hover_color="#B91C1C", height=28,
            font=("Segoe UI", 12, "bold"), width=140
        )
        self.btn_demo.pack(pady=(0, 15))

        # --- Options Area ---
        self.options_frame = ctk.CTkFrame(self.scroll_container, fg_color="transparent")
        self.options_frame.pack(pady=5)

        self.merge_check = ctk.CTkCheckBox(
            self.options_frame, 
            text="Bunch Converter (Merge all into one Excel)", 
            variable=self.merge_mode,
            font=("Segoe UI", 12),
            checkbox_height=24,
            checkbox_width=24,
            border_width=2
        )
        self.merge_check.pack()
        
        self.hint_label = ctk.CTkLabel(
            self.options_frame,
            text="(Ideal for combining 12 months or 4 quarters into one sheet)",
            font=("Segoe UI", 10),
            text_color="gray"
        )
        self.hint_label.pack()
        self.convert_btn = ctk.CTkButton(
            self.scroll_container,
            text="Convert to Excel Now",
            command=self.process_files,
            state="disabled",
            height=50,
            width=250,
            font=("Segoe UI", 16, "bold"),
            fg_color="#059669",
            hover_color="#047857"
        )
        self.convert_btn.pack(pady=20)

        self.open_folder_btn = ctk.CTkButton(
            self.scroll_container,
            text="📂 Open Output Folder",
            command=self.open_output_folder,
            height=40,
            width=200,
            font=("Segoe UI", 14),
            fg_color="#64748B",
            hover_color="#475569"
        )
        self.open_folder_btn.pack(pady=(0, 10))
        self.open_folder_btn.pack_forget() 

        # Log/Status Area
        self.textbox = ctk.CTkTextbox(self.scroll_container, height=200, width=600)
        self.textbox.pack(pady=10)
        self.textbox.insert("0.0", "Status log will appear here...\n")
        self.textbox.configure(state="disabled")

    def log(self, message):
        self.after(0, self._log_to_gui, message)

    def _log_to_gui(self, message):
        self.textbox.configure(state="normal")
        self.textbox.insert("end", f"[{time.strftime('%H:%M:%S')}] {message}\n")
        self.textbox.see("end")
        self.textbox.configure(state="disabled")

    def open_demo_link(self):
        import webbrowser
        webbrowser.open_new_tab("https://youtu.be/zEggEXMjL-w")

    def open_output_folder(self):
        target = os.path.join(os.getcwd(), "GST Downloaded", "GSTR 3B to Excel")
        if os.path.exists(target):
            os.startfile(target)
        else:
            messagebox.showinfo("Info", "Output folder not found.")

    def select_files(self):
        filetypes = (("PDF files", "*.pdf"), ("All files", "*.*"))
        filenames = filedialog.askopenfilenames(title="Select GSTR-3B PDFs", filetypes=filetypes)
        
        if filenames:
            self.selected_files = filenames
            count = len(filenames)
            self.file_count_label.configure(text=f"{count} file(s) selected", text_color="white")
            self.convert_btn.configure(state="normal")
            self.log(f"Selected {count} files.")

    def extract_data(self, pdf_path):
        """Extracts ALL specific tables and returns them in a structured dictionary."""
        data = {
            "meta": {"GSTIN": "Unknown", "Year": "Unknown", "Month": "Unknown"},
            "supplies": [],        # 3.1
            "supplies_9_5": [],    # 3.1.1
            "inter_state": [],     # 3.2
            "itc": [],             # 4
            "nil": [],             # 5
            "interest": [],        # 5.1
            "payment": []          # 6.1
        }
        
        current_6_1_block = "(A) Other than reverse charge" # Default State Tracker

        try:
            with pdfplumber.open(pdf_path) as pdf:
                full_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text: full_text += text + "\n"
                    
                    tables = page.extract_tables()
                    for table in tables:
                        if not table: continue
                        clean_table = [[str(c).replace('\n', ' ').strip() if c else "0.00" for c in row] for row in table]
                        table_text_lower = " ".join([" ".join(row) for row in clean_table]).lower()

                        # 1. Identify Table 3.1 (Supplies)
                        if "outward taxable supplies" in table_text_lower and "zero rated" in table_text_lower:
                            for row in clean_table:
                                r0 = row[0].lower()
                                if "(a)" in r0 or "(b)" in r0 or "(c)" in r0 or "(d)" in r0 or "(e)" in r0:
                                    data["supplies"].append(row)

                        # 2. Identify Table 3.1.1 (Supplies under 9(5))
                        elif "electronic commerce operator pays tax u/s 9(5)" in table_text_lower:
                            for row in clean_table:
                                if "electronic commerce" in row[0].lower():
                                    data["supplies_9_5"].append(row)

                        # 3. Identify Table 3.2 (Inter-state Supplies)
                        elif "unregistered persons" in table_text_lower and "composition" in table_text_lower:
                            for row in clean_table:
                                r0 = row[0].lower()
                                if "unregistered persons" in r0 or "composition taxable" in r0 or "uin holders" in r0:
                                    data["inter_state"].append(row)

                        # 4. Identify Table 4 (ITC)
                        elif "itc available" in table_text_lower or "itc reversed" in table_text_lower:
                            for row in clean_table:
                                r0 = row[0].lower()
                                if "import" in r0 or "inward" in r0 or "all other itc" in r0 or "as per rules" in r0 or "others" in r0 or "net itc" in r0 or "itc reclaimed" in r0 or "ineligible itc" in r0:
                                    data["itc"].append(row)

                        # 5. Identify Table 5 (Nil Rated)
                        elif "from a supplier under composition" in table_text_lower and "non gst" in table_text_lower:
                            for row in clean_table:
                                r0 = row[0].lower()
                                if "composition" in r0 or "non gst" in r0:
                                    data["nil"].append(row)

                        # 6. Identify Table 6.1 (MUST Process BEFORE 5.1 to prevent Keyword Conflict)
                        elif ("tax payable" in table_text_lower and "paid in cash" in table_text_lower) or \
                             ("other than reverse charge" in table_text_lower and "integrated tax" in table_text_lower) or \
                             ("adjustment" in table_text_lower and "negative liability" in table_text_lower):
                            for row in clean_table:
                                row_str = " ".join(row).lower()
                                
                                # Skip header/title rows completely
                                if "tax payable" in row_str or "paid in cash" in row_str or "adjustment" in row_str:
                                    continue
                                
                                # Identify Blocks (Updates Current State)
                                if "other than reverse charge" in row_str:
                                    current_6_1_block = "(A) Other than reverse charge"
                                elif "reverse charge" in row_str and "other than" not in row_str:
                                    current_6_1_block = "(B) Reverse charge"
                                
                                # Extract Tax Head regardless of which column it got pushed into
                                tax_head = None
                                if "integrated tax" in row_str: tax_head = "Integrated Tax"
                                elif "central tax" in row_str: tax_head = "Central Tax"
                                elif "state/ut tax" in row_str: tax_head = "State/UT Tax"
                                elif "cess" in row_str: tax_head = "Cess"

                                if tax_head:
                                    # Store block, head, and raw row values
                                    data["payment"].append([current_6_1_block, tax_head] + row)

                        # 7. Identify Table 5.1 (Interest & Late fee)
                        elif "system computed interest" in table_text_lower or ("interest paid" in table_text_lower and "late fee" in table_text_lower):
                            for row in clean_table:
                                r0 = row[0].lower()
                                if "interest" in r0 or "late fee" in r0:
                                    data["interest"].append(row)

                # Metadata Extraction
                gstin_match = re.search(r"GSTIN.*?(\d{2}[A-Z]{5}\d{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1})", full_text)
                year_match = re.search(r"Year\s*(\d{4}-\d{2})", full_text)
                period_match = re.search(r"Period\s*([A-Za-z]{3}-[A-Za-z]{3}|[A-Za-z]+)", full_text)
                
                if gstin_match: data["meta"]["GSTIN"] = gstin_match.group(1)
                if year_match: data["meta"]["Year"] = year_match.group(1)
                if period_match: data["meta"]["Month"] = period_match.group(1)
                
        except Exception as e:
            self.log(f"Error parsing PDF: {e}")
            
        return data

    def prepare_rows(self, data):
        """Flattens the data structure for all tables."""
        meta = data["meta"]
        
        def clean_float(val):
            if not val or str(val).strip() in ("0.00", "0", ""): return 0.00
            cleaned = str(val).replace(",", "").replace(" ", "")
            try:
                return float(cleaned)
            except ValueError:
                match = re.search(r"[-+]?\d*\.\d+|\d+", cleaned)
                return float(match.group()) if match else 0.00

        prepared = {
            "supplies": [], "supplies_9_5": [], "inter_state": [], 
            "itc": [], "nil": [], "interest": [], "payment": []
        }

        # 3.1 Supplies & 3.1.1 Supplies 9(5)
        for target_key in ["supplies", "supplies_9_5"]:
            for row in data[target_key]:
                nature = row[0]
                val = clean_float(row[1]) if len(row) > 1 else 0
                igst = clean_float(row[2]) if len(row) > 2 else 0
                cgst = clean_float(row[3]) if len(row) > 3 else 0
                sgst = clean_float(row[4]) if len(row) > 4 else 0
                cess = clean_float(row[5]) if len(row) > 5 else 0
                prepared[target_key].append([meta["GSTIN"], meta["Year"], meta["Month"], nature, val, igst, cgst, sgst, cess])

        # 3.2 Inter-state
        for row in data["inter_state"]:
            cols = [clean_float(x) for x in row[1:] if str(x).strip() != ""]
            val = cols[0] if len(cols) > 0 else 0.0
            igst = cols[1] if len(cols) > 1 else 0.0
            prepared["inter_state"].append([meta["GSTIN"], meta["Year"], meta["Month"], row[0], val, igst])

        # 4 ITC & 5.1 Interest
        for target_key in ["itc", "interest"]:
            for row in data[target_key]:
                details = row[0]
                igst = clean_float(row[1]) if len(row) > 1 else 0
                cgst = clean_float(row[2]) if len(row) > 2 else 0
                sgst = clean_float(row[3]) if len(row) > 3 else 0
                cess = clean_float(row[4]) if len(row) > 4 else 0
                prepared[target_key].append([meta["GSTIN"], meta["Year"], meta["Month"], details, igst, cgst, sgst, cess])

        # 5 Nil
        for row in data["nil"]:
            nature = row[0]
            inter = clean_float(row[1]) if len(row) > 1 else 0
            intra = clean_float(row[2]) if len(row) > 2 else 0
            prepared["nil"].append([meta["GSTIN"], meta["Year"], meta["Month"], nature, inter, intra])

        # 6.1 Payment - SMART ROW EXTRACTION
        for row in data["payment"]:
            block = row[0]
            tax_head = row[1]
            raw_cells = row[2:]
            
            vals = []
            for cell in raw_cells:
                cell_str = str(cell).strip().lower()
                # Dynamically skip cells containing alphabet characters, isolating purely numeric columns
                if re.search(r'[a-z]{3,}', cell_str):
                    continue
                vals.append(clean_float(cell))
            
            # Pad or trim to ensure exactly 10 data points per row
            vals += [0.0] * 10
            prepared["payment"].append([meta["GSTIN"], meta["Year"], meta["Month"], block, tax_head] + vals[:10])

        return prepared

    def write_excel(self, combined_data, output_path):
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def setup_sheet(sheet_name, headers, rows):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Write Headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            
            # Write Rows
            for r_idx, row_data in enumerate(rows, 2):
                for c_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=r_idx, column=c_idx, value=val)
                    c.border = border

        # Define Sheet Structure
        sheets_config = [
            ("GSTR3B_3.1_Supplies", ["GSTIN", "Year", "Month", "Nature of Supplies", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"], combined_data["supplies"]),
            ("GSTR3B_3.1.1_Supplies_9_5", ["GSTIN", "Year", "Month", "Nature of Supplies", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"], combined_data["supplies_9_5"]),
            ("GSTR3B_3.2_InterState", ["GSTIN", "Year", "Month", "Nature of Supplies", "Total Taxable Value", "Integrated Tax"], combined_data["inter_state"]),
            ("GSTR3B_4_ITC", ["GSTIN", "Year", "Month", "Details", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"], combined_data["itc"]),
            ("GSTR3B_5_NilRated", ["GSTIN", "Year", "Month", "Nature of Supplies", "Inter-State Supplies", "Intra-State Supplies"], combined_data["nil"]),
            ("GSTR3B_5.1_Interest", ["GSTIN", "Year", "Month", "Details", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"], combined_data["interest"]),
            ("GSTR3B_6.1_Payment", ["GSTIN", "Year", "Month", "Block", "Tax Head", "Tax Payable", "Adjustment", "Net Tax Payable", "Paid via IGST", "Paid via CGST", "Paid via SGST", "Paid via Cess", "Tax Paid in Cash", "Interest Paid in Cash", "Late Fee Paid in Cash"], combined_data["payment"])
        ]

        # Generate Sheets
        for sheet_name, headers, data_rows in sheets_config:
            if data_rows: # Only create sheet if there's data to process
                setup_sheet(sheet_name, headers, data_rows)

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1: 
            wb.remove(wb["Sheet"])
            
        try:
            wb.save(output_path)
        except PermissionError:
            raise PermissionError(f"The file '{os.path.basename(output_path)}' is currently open in Excel.\n\nPlease close it and try again.")

    def process_files(self):
        if not self.selected_files:
            return

        self.convert_btn.configure(state="disabled")
        self.open_folder_btn.pack_forget()
        
        # Run the heavy processing in a background thread to prevent UI freezing
        threading.Thread(target=self._run_conversion, daemon=True).start()

    def _run_conversion(self):
        try:
            # --- Output Folder Setup ---
            output_folder = os.path.join(os.getcwd(), "GST Downloaded", "GSTR 3B to Excel")
            if not os.path.exists(output_folder):
                os.makedirs(output_folder, exist_ok=True)
            
            is_merge = self.merge_mode.get()
            
            if is_merge:
                # === BUNCH CONVERTER MODE ===
                self.log(f"Starting Bunch Merge for {len(self.selected_files)} files...")
                
                master_data = {
                    "supplies": [], "supplies_9_5": [], "inter_state": [], 
                    "itc": [], "nil": [], "interest": [], "payment": []
                }
                first_meta = None

                for idx, file_path in enumerate(self.selected_files):
                    filename = os.path.basename(file_path)
                    self.log(f"Reading ({idx+1}/{len(self.selected_files)}): {filename}...")
                    
                    # Extract & Prepare
                    raw_data = self.extract_data(file_path)
                    if not first_meta: first_meta = raw_data["meta"]
                    flat_rows = self.prepare_rows(raw_data)
                    
                    # Merge all keys natively
                    for key in master_data.keys():
                        master_data[key].extend(flat_rows[key])

                # Save Consolidated File
                out_name = "GSTR3B_Consolidated.xlsx"
                if first_meta and first_meta["Year"] != "Unknown":
                    out_name = f"GSTR3B_Consolidated_{first_meta['Year']}.xlsx"
                
                out_path = os.path.join(output_folder, out_name)
                self.write_excel(master_data, out_path)
                
                self.log("----------------")
                self.log(f"MERGE SUCCESSFUL!")
                self.log(f"Saved: {out_name}")
                self.after(0, lambda: messagebox.showinfo("Success", f"Merged {len(self.selected_files)} files into:\n{out_name}"))
                self.after(0, lambda: self.open_folder_btn.pack(pady=(0, 10), before=self.textbox))

            else:
                # === INDIVIDUAL MODE ===
                self.log(f"Starting Individual Conversion...")
                
                for idx, file_path in enumerate(self.selected_files):
                    filename = os.path.basename(file_path)
                    self.log(f"Processing ({idx+1}/{len(self.selected_files)}): {filename}")
                    
                    # Extract & Prepare
                    raw_data = self.extract_data(file_path)
                    flat_rows = self.prepare_rows(raw_data)
                    
                    # Naming
                    safe_month = raw_data["meta"]["Month"].replace(" ", "_")
                    safe_year = raw_data["meta"]["Year"].replace("-", "_")
                    if safe_month == "Unknown":
                        out_name = filename.replace(".pdf", ".xlsx")
                    else:
                        out_name = f"GSTR3B_{safe_month}_{safe_year}.xlsx"
                    
                    out_path = os.path.join(output_folder, out_name)
                    
                    # Write
                    self.write_excel(flat_rows, out_path)
                    self.log(f" -> Saved: {out_name}")

                self.log("----------------")
                self.log("Batch Complete.")
                self.after(0, lambda: self.open_folder_btn.pack(pady=(0, 10), before=self.textbox))
                self.after(0, lambda: messagebox.showinfo("Done", f"Processed {len(self.selected_files)} files."))

        except PermissionError as pe:
            self.log(f"❌ Save Error: File is open.")
            self.after(0, lambda: messagebox.showerror("File Open Error", str(pe)))
            
        except Exception as e:
            self.log(f"❌ Critical Error: {str(e)}")
            self.after(0, lambda: messagebox.showerror("Error", f"An unexpected error occurred:\n{e}"))
        
        finally:
            self.after(0, self._reset_ui)

    def _reset_ui(self):
        self.convert_btn.configure(state="normal")
        self.selected_files = []
        self.file_count_label.configure(text="No files selected", text_color="gray")


if __name__ == "__main__":
    app = GSTR3BConverterPro()
    app.mainloop()