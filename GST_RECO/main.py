import os
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
FY_PATTERNS = [
    r"FY2023[-/]?24", r"2023[-/]24", r"FY2324", r"FY202324", r"23[-/]24", r"2324[-/]",
    r"FY2024[-/]?25", r"2024[-/]25", r"FY2425", r"FY202425", r"24[-/]25", r"2425[-/]",
    r"FY2025[-/]?26", r"2025[-/]26", r"FY2526", r"FY202526", r"25[-/]26", r"2526[-/]",
]

def clean_invoice(inv_str: str) -> str:
    """Strip FY patterns and non-digit chars; return numeric-only tail (no leading zeros)."""
    s = str(inv_str).strip()
    for pat in FY_PATTERNS:
        s = re.sub(pat, "", s, flags=re.IGNORECASE)
    digits = re.sub(r"\D", "", s)
    return digits.lstrip("0") or digits

def make_gstinv(gstin: str, inv_str: str) -> str:
    return str(gstin).strip() + clean_invoice(inv_str)

def safe_float(val) -> float:
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


# ─────────────────────────────────────────────
#  PARSE GSTR-2B  (reads the "B2B" sheet)
# ─────────────────────────────────────────────
def parse_gstr2b(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # ── Strict file identity check ──────────────────────────────────────
    sheets = wb.sheetnames
    if "B2B" not in sheets:
        wb.close()
        if len(sheets) == 1 and sheets[0] in ("Sheet1", "Sheet"):
            raise ValueError(
                "The file you selected for GSTR-2B looks like a Tally export "
                "(it has no 'B2B' sheet).\n\nPlease make sure you put the GSTR-2B "
                "file in the top slot and the Tally file in the second slot."
            )
        raise ValueError(
            f"GSTR-2B file has no 'B2B' sheet.\nSheets found: {sheets}\n"
            "Please select the correct GSTR-2B file downloaded from the GST portal."
        )

    ws = wb["B2B"]
    rows = list(ws.iter_rows(values_only=True))

    # Confirm row 4 looks like expected GSTR-2B header
    header_check = str(rows[4][0] if rows[4] else "").strip()
    if "GSTIN" not in header_check.upper():
        wb.close()
        raise ValueError(
            "The 'B2B' sheet does not have the expected GSTR-2B header format.\n"
            "Please check you have selected the correct GSTR-2B portal download file."
        )

    # Data starts at row index 6 (rows[6])
    records = []
    for row in rows[6:]:
        gstin = row[0]
        if not gstin or str(gstin).strip() == "":
            continue
        inv_no   = str(row[2] or "").strip()
        inv_date = row[4]
        inv_val  = safe_float(row[5])
        taxable  = safe_float(row[8])
        igst     = safe_float(row[9])
        cgst     = safe_float(row[10])
        sgst     = safe_float(row[11])
        new_inv  = clean_invoice(inv_no)
        gstinv   = make_gstinv(gstin, inv_no)
        records.append({
            "gstin": str(gstin).strip(),
            "name":  str(row[1] or "").strip(),
            "original_inv": inv_no,
            "new_inv":  new_inv,
            "gstinv":   gstinv,
            "inv_date": inv_date,
            "taxable":  taxable,
            "igst":     igst,
            "cgst":     cgst,
            "sgst":     sgst,
            "total":    inv_val if inv_val else taxable + igst + cgst + sgst,
            "source":   "GSTR2B",
        })
    wb.close()
    return records


# ─────────────────────────────────────────────
#  PARSE TALLY / OUR DATA
# ─────────────────────────────────────────────
def parse_tally(filepath: str) -> list:
    """
    Reads the Tally GSTR-2A Reconciliation export (B2B Invoices voucher register).
    Expected header row contains: 'Party GSTIN/UIN', 'Doc No.', 'Taxable'
    Data columns (0-based from header row):
      0=Date  1=Particulars  2=Party GSTIN/UIN  5=Vch No.  6=Doc No.
      11=Taxable Amount  12=IGST  13=CGST  14=SGST  16=Tax Amount  17=Invoice Amount
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets = wb.sheetnames

    # ── Strict file identity check ──────────────────────────────────────
    if "B2B" in sheets and len(sheets) > 5:
        wb.close()
        raise ValueError(
            "The file you selected for Tally / Our Data looks like a GSTR-2B portal "
            "download (it contains a 'B2B' sheet).\n\nPlease make sure the GSTR-2B file "
            "is in the top slot and the Tally export is in the second slot."
        )

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # ── Find the header row — must contain 'Party GSTIN/UIN' exactly ────
    header_row_idx = None
    for i, row in enumerate(all_rows):
        row_text = " ".join(str(c).upper() for c in row if c)
        if "PARTY GSTIN" in row_text and "DOC NO" in row_text.replace(".", "").replace(" ", ""):
            header_row_idx = i
            break

    if header_row_idx is None:
        # Fallback: any row with GSTIN/UIN text
        for i, row in enumerate(all_rows):
            if any(c and "GSTIN" in str(c).upper() for c in row):
                header_row_idx = i
                break

    if header_row_idx is None:
        wb.close()
        raise ValueError(
            "Could not find the Tally header row (expected 'Party GSTIN/UIN' and 'Doc No.').\n"
            "Please check you have selected the correct Tally B2B voucher register export."
        )

    # Validate column positions from the header row
    header = all_rows[header_row_idx]
    col_gstin   = next((i for i, c in enumerate(header) if c and "GSTIN" in str(c).upper()), 2)
    col_doc_no  = next((i for i, c in enumerate(header) if c and "DOC" in str(c).upper() and "NO" in str(c).upper()), 6)
    col_taxable = next((i for i, c in enumerate(header) if c and "TAXABLE" in str(c).upper()), 11)
    # Tax columns are always right after Taxable in Tally export
    col_igst  = col_taxable + 1
    col_cgst  = col_taxable + 2
    col_sgst  = col_taxable + 3
    col_inv   = col_taxable + 6   # Invoice Amount = Taxable + 6 in standard Tally layout

    # Skip header row + sub-header row (+2)
    records = []
    for row in all_rows[header_row_idx + 2:]:
        if len(row) <= col_gstin:
            continue
        gstin = row[col_gstin]
        if not gstin or str(gstin).strip() in ("", "nan", "None"):
            continue
        gstin_str = str(gstin).strip()
        # Must look like a valid GSTIN (2 digits + letters)
        if not re.match(r"^\d{2}[A-Z]", gstin_str):
            continue

        inv_no  = str(row[col_doc_no] if len(row) > col_doc_no else "").strip()
        if not inv_no or inv_no in ("None", "nan"):
            # fallback to Vch No. (col 5)
            inv_no = str(row[5] if len(row) > 5 else "").strip()

        inv_date = row[0]
        taxable  = safe_float(row[col_taxable]  if len(row) > col_taxable  else None)
        igst     = safe_float(row[col_igst]     if len(row) > col_igst     else None)
        cgst     = safe_float(row[col_cgst]     if len(row) > col_cgst     else None)
        sgst     = safe_float(row[col_sgst]     if len(row) > col_sgst     else None)
        inv_val  = safe_float(row[col_inv]      if len(row) > col_inv      else None)
        new_inv  = clean_invoice(inv_no)
        gstinv   = make_gstinv(gstin_str, inv_no)

        records.append({
            "gstin": gstin_str,
            "name":  str(row[1] or "").strip(),
            "original_inv": inv_no,
            "new_inv":  new_inv,
            "gstinv":   gstinv,
            "inv_date": inv_date,
            "taxable":  taxable,
            "igst":     igst,
            "cgst":     cgst,
            "sgst":     sgst,
            "total":    inv_val if inv_val else taxable + igst + cgst + sgst,
            "source":   "OUR DATA",
        })
    wb.close()

    if not records:
        raise ValueError(
            "No valid invoice records found in the Tally file.\n"
            "Please check the file is a Tally B2B Invoices voucher register export."
        )
    return records


# ─────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────
FILL_MATCH   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_NOMINAL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_DIFF    = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
FILL_NO_OUR  = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
FILL_NO_GSTR = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT    = Font(name="Arial", size=9)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center")
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "Matched":            FILL_MATCH,
    "Nominal Difference": FILL_NOMINAL,
    "Difference":         FILL_DIFF,
    "Not in Our Data":    FILL_NO_OUR,
    "Not in GSTR2B":      FILL_NO_GSTR,
}

HEADERS = [
    "GSTIN of Supplier", "Trade/Legal Name",
    "Original Invoice", "New Invoice No.", "GSTinv",
    "Invoice Date",
    "Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Total (₹)",
    "Source", "Reconciliation Status",
]
COL_WIDTHS = [22, 32, 20, 16, 30, 14, 16, 14, 14, 14, 14, 12, 22]


def _write_header(ws, col_widths):
    ws.row_dimensions[1].height = 30
    for ci, (hdr, w) in enumerate(zip(HEADERS, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def _write_row(ws, ri, rec, status):
    fill = STATUS_FILL.get(status)
    values = [
        rec["gstin"], rec["name"],
        rec["original_inv"], rec["new_inv"], rec["gstinv"],
        rec["inv_date"],
        rec["taxable"], rec["igst"], rec["cgst"], rec["sgst"], rec["total"],
        rec["source"], status,
    ]
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = DATA_FONT
        c.border = BORDER
        c.alignment = LEFT
        if fill:
            c.fill = fill
    fmt_num  = "#,##0.00"
    fmt_date = "DD-MMM-YYYY"
    for ci in range(7, 12):
        ws.cell(row=ri, column=ci).number_format = fmt_num
    if rec["inv_date"]:
        ws.cell(row=ri, column=6).number_format = fmt_date


# ─────────────────────────────────────────────
#  MAIN RECONCILIATION
# ─────────────────────────────────────────────
def reconcile_and_write(gstr2b_file: str, tally_file: str, output_file: str,
                        progress_cb=None):
    def log(msg):
        if progress_cb:
            progress_cb(msg)
        else:
            print(msg)

    log("Parsing GSTR-2B…")
    gstr_recs = parse_gstr2b(gstr2b_file)
    log(f"  → {len(gstr_recs)} B2B invoices in GSTR-2B")

    log("Parsing Tally / Our Data…")
    tally_recs = parse_tally(tally_file)
    log(f"  → {len(tally_recs)} invoices in Tally file")

    all_recs = gstr_recs + tally_recs

    # Aggregate total per (gstinv, source)
    agg: dict = {}
    for r in all_recs:
        key = r["gstinv"]
        src = r["source"]
        agg.setdefault(key, {})
        agg[key][src] = agg[key].get(src, 0.0) + r["total"]

    def get_status(gstinv: str) -> str:
        d = agg.get(gstinv, {})
        has_g = "GSTR2B"   in d
        has_o = "OUR DATA" in d
        if has_g and has_o:
            diff = abs(d["GSTR2B"] - d["OUR DATA"])
            if diff == 0:       return "Matched"
            elif diff <= 20:    return "Nominal Difference"
            else:               return "Difference"
        elif has_g:             return "Not in Our Data"
        elif has_o:             return "Not in GSTR2B"
        return ""

    log("Building output workbook…")
    wb = openpyxl.Workbook()

    # ── Sheet: All Data ───────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Data"
    _write_header(ws_all, COL_WIDTHS)

    sorted_recs = (
        sorted(gstr_recs,  key=lambda r: r["gstinv"]) +
        sorted(tally_recs, key=lambda r: r["gstinv"])
    )
    for ri, rec in enumerate(sorted_recs, 2):
        _write_row(ws_all, ri, rec, get_status(rec["gstinv"]))

    # ── Sheet: Exceptions ────────────────────────────────────────
    ws_exc = wb.create_sheet("Exceptions")
    _write_header(ws_exc, COL_WIDTHS)
    exc_ri = 2
    for rec in sorted_recs:
        s = get_status(rec["gstinv"])
        if s in ("Difference", "Not in Our Data", "Not in GSTR2B"):
            _write_row(ws_exc, exc_ri, rec, s)
            exc_ri += 1

    # ── Sheet: Summary ───────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 14

    # Count unique gstinv keys per status
    seen: dict = {}
    for r in all_recs:
        k = r["gstinv"]
        if k not in seen:
            seen[k] = get_status(k)

    statuses = ["Matched", "Nominal Difference", "Difference",
                "Not in Our Data", "Not in GSTR2B"]
    counts = {s: sum(1 for v in seen.values() if v == s) for s in statuses}

    rows_data = [
        ("GST Reconciliation Summary", None),
        (None, None),
        ("Status", "Invoice Count"),
    ] + [(s, counts[s]) for s in statuses] + [
        (None, None),
        ("GSTR-2B Records",   len(gstr_recs)),
        ("Our Data Records",  len(tally_recs)),
        ("Total Combined",    len(all_recs)),
    ]

    for ri, (lbl, val) in enumerate(rows_data, 1):
        cl = ws_sum.cell(row=ri, column=1, value=lbl)
        cv = ws_sum.cell(row=ri, column=2, value=val)
        bold = ri in (1, 3, len(rows_data) - 2, len(rows_data) - 1, len(rows_data))
        cl.font = Font(name="Arial", size=11 if ri == 1 else 10, bold=bold or ri == 1)
        cv.font = Font(name="Arial", size=10, bold=bold)
        # Color status rows
        fill = STATUS_FILL.get(str(lbl))
        if fill:
            cl.fill = fill
            cv.fill = fill

    # Legend
    leg_start = len(rows_data) + 3
    ws_sum.cell(row=leg_start, column=1, value="Colour Legend").font = Font(
        name="Arial", size=10, bold=True)
    for i, (lbl, fill) in enumerate([
        ("Matched",                    FILL_MATCH),
        ("Nominal Difference (≤ ₹20)", FILL_NOMINAL),
        ("Difference",                 FILL_DIFF),
        ("Not in Our Data",            FILL_NO_OUR),
        ("Not in GSTR2B",              FILL_NO_GSTR),
    ], leg_start + 1):
        c = ws_sum.cell(row=i, column=1, value=lbl)
        c.fill = fill
        c.font = Font(name="Arial", size=9)

    log(f"Saving → {output_file}")
    wb.save(output_file)
    log("Reconciliation complete ✓")


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────
def launch_gui():
    root = tk.Tk()
    root.title("GST Reconciliation Tool")
    root.geometry("700x500")
    root.resizable(False, False)

    notebook = ttk.Notebook(root)
    notebook.pack(pady=10, expand=True, fill="both")

    tab = ttk.Frame(notebook)
    notebook.add(tab, text="GST Reconciliation (2 Files)")

    frame = tk.Frame(tab, padx=20, pady=15)
    frame.pack(fill="both", expand=True)

    lbl_kw = {"font": ("Arial", 10, "bold"), "anchor": "w"}

    # Row helpers
    def file_row(row_idx, label, entry_var):
        tk.Label(frame, text=label, **lbl_kw).grid(
            row=row_idx, column=0, sticky="w", pady=5)
        entry = tk.Entry(frame, width=46)
        entry.grid(row=row_idx, column=1, padx=8)
        return entry

    gstr_entry  = file_row(0, "GSTR-2B File (.xlsx):",         None)
    tally_entry = file_row(1, "Tally / Our Data File (.xlsx):", None)

    def make_browse(entry, title):
        ft = [("Excel", "*.xlsx *.xlsm"), ("All", "*.*")]
        def cb():
            f = filedialog.askopenfilename(title=title, filetypes=ft)
            if f:
                entry.delete(0, tk.END)
                entry.insert(0, f)
        return cb

    tk.Button(frame, text="Browse", command=make_browse(gstr_entry,  "Select GSTR-2B File"),          width=9).grid(row=0, column=2)
    tk.Button(frame, text="Browse", command=make_browse(tally_entry, "Select Tally / Our Data File"), width=9).grid(row=1, column=2)

    tk.Label(frame, text="Log:", **lbl_kw).grid(row=2, column=0, sticky="nw", pady=(14, 0))
    log_box = tk.Text(frame, height=9, width=60, state=tk.DISABLED,
                      font=("Courier", 9), bg="#f5f5f5")
    log_box.grid(row=2, column=1, columnspan=2, pady=(14, 0))

    def append_log(msg):
        log_box.config(state=tk.NORMAL)
        log_box.insert(tk.END, msg + "\n")
        log_box.see(tk.END)
        log_box.config(state=tk.DISABLED)
        root.update_idletasks()

    run_btn = tk.Button(frame, text="▶  Run Reconciliation",
                        font=("Arial", 12, "bold"),
                        bg="#4CAF50", fg="white", width=24, pady=7)
    run_btn.grid(row=3, column=0, columnspan=3, pady=18)

    def run_reco():
        g = gstr_entry.get().strip()
        t = tally_entry.get().strip()

        if not g or not os.path.exists(g):
            messagebox.showerror("Error", "Please select a valid GSTR-2B file.")
            return
        if not t or not os.path.exists(t):
            messagebox.showerror("Error", "Please select a valid Tally / Our Data file.")
            return

        o = os.path.join(os.path.dirname(g), "GST_Reconciled_Output.xlsx")

        def process():
            run_btn.config(state=tk.DISABLED, text="Processing…")
            log_box.config(state=tk.NORMAL)
            log_box.delete("1.0", tk.END)
            log_box.config(state=tk.DISABLED)
            try:
                reconcile_and_write(g, t, o, progress_cb=append_log)
                messagebox.showinfo("Done", f"Reconciliation complete!\n\nSaved to:\n{o}")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred:\n{e}")
                append_log(f"ERROR: {e}")
            finally:
                run_btn.config(state=tk.NORMAL, text="▶  Run Reconciliation")

        threading.Thread(target=process, daemon=True).start()

    run_btn.config(command=run_reco)
    root.mainloop()


if __name__ == "__main__":
    launch_gui()
