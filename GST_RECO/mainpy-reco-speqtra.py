import os
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import customtkinter as ctk

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  HELPERS & METADATA EXTRACTOR
# ─────────────────────────────────────────────
FY_PATTERNS = [
    r"FY2023[-/]?24", r"(?<![A-Za-z\d])2023[-/]24", r"FY2324", r"FY202324", r"(?<![A-Za-z\d])23[-/]24", r"(?<![A-Za-z\d])2324[-/]",
    r"FY2024[-/]?25", r"(?<![A-Za-z\d])2024[-/]25", r"FY2425", r"FY202425", r"(?<![A-Za-z\d])24[-/]25", r"(?<![A-Za-z\d])2425[-/]",
    r"FY2025[-/]?26", r"(?<![A-Za-z\d])2025[-/]26", r"FY2526", r"FY202526", r"(?<![A-Za-z\d])25[-/]26", r"(?<![A-Za-z\d])2526[-/]",
]

def clean_invoice(inv_str: str) -> str:
    """Strip FY patterns and non-digit chars; return numeric-only tail (no leading zeros)."""
    s = str(inv_str).strip()
    for pat in FY_PATTERNS:
        s = re.sub(pat, "", s, flags=re.IGNORECASE)
    digits = re.sub(r"\D", "", s)
    return digits.lstrip("0") or digits

def make_gstinv(gstin: str, inv_str: str) -> str:
    return str(gstin).strip() + clean_invoice(inv_str)

def safe_float(val) -> float:
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def digit_tail(s: str) -> str:
    """Return only the digit characters from a string (already-cleaned invoice key)."""
    return re.sub(r"\D", "", str(s).strip())

def get_gstr_info(filepath: str, log_cb) -> dict:
    """Bulletproof metadata extractor that skips empty columns to find data."""
    info = {"gstin": "", "trade_name": "", "fy": "", "period": ""}
    try:
        # Load the workbook reading actual values
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        target_sheet = None
        for sheet in wb.sheetnames:
            if "readme" in sheet.lower().replace(" ", "").replace("_", ""):
                target_sheet = sheet
                break
        
        if target_sheet:
            ws = wb[target_sheet]
            # Scan a wide area: first 20 rows, first 10 columns
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True):
                for i, cell in enumerate(row):
                    if not cell:
                        continue
                        
                    label = str(cell).strip().lower()
                    
                    # Instead of just taking the immediate next column, 
                    # scan to the right and grab the FIRST non-blank value
                    val = ""
                    for j in range(i + 1, len(row)):
                        if row[j] and str(row[j]).strip():
                            val = str(row[j]).strip()
                            break
                            
                    # If we found a value, map it to the right dictionary key
                    if val:
                        if "gstin" in label and not info["gstin"]:
                            info["gstin"] = val
                        elif ("trade name" in label or "legal name" in label) and not info["trade_name"]:
                            info["trade_name"] = val
                        elif ("financial year" in label or "f.y" in label) and not info["fy"]:
                            info["fy"] = val
                        elif ("period" in label) and not info["period"]:
                            info["period"] = val
                            
        wb.close()
    except Exception as e:
        log_cb(f"Warning: Could not extract Read me metadata: {e}")
        
    return info
    """Bulletproof metadata extractor that scans a grid to find headers."""
    info = {"gstin": "", "trade_name": "", "fy": "", "period": ""}
    try:
        # Removed read_only=True to prevent blank reads on portal-generated files
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        target_sheet = None
        for sheet in wb.sheetnames:
            if "readme" in sheet.lower().replace(" ", "").replace("_", ""):
                target_sheet = sheet
                break
        
        if target_sheet:
            ws = wb[target_sheet]
            # Scan a wide area: first 20 rows, first 10 columns
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True):
                for i, cell in enumerate(row):
                    if not cell:
                        continue
                        
                    label = str(cell).strip().lower()
                    
                    # If we find a label, grab the value in the immediate next column
                    if i + 1 < len(row):
                        val = str(row[i+1] or "").strip()
                        
                        if label == "gstin" and not info["gstin"]:
                            info["gstin"] = val
                        elif label in ("trade name", "legal name") and not info["trade_name"]:
                            info["trade_name"] = val
                        elif label in ("financial year", "f.y.") and not info["fy"]:
                            info["fy"] = val
                        elif label in ("return period", "period") and not info["period"]:
                            info["period"] = val
        wb.close()
    except Exception as e:
        log_cb(f"Warning: Could not extract Read me metadata: {e}")
        
    return info

# ─────────────────────────────────────────────
#  PARSE GSTR-2B
# ─────────────────────────────────────────────
def parse_gstr2b(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets = wb.sheetnames
    if "B2B" not in sheets:
        wb.close()
        raise ValueError("GSTR-2B file has no 'B2B' sheet.")

    ws = wb["B2B"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[6:]:
        gstin = row[0]
        if not gstin or str(gstin).strip() == "":
            continue
        inv_no   = str(row[2] or "").strip()
        inv_date = row[4]
        inv_val  = safe_float(row[5])
        taxable  = safe_float(row[8])
        igst     = safe_float(row[9])
        cgst     = safe_float(row[10])
        sgst     = safe_float(row[11])
        cess     = safe_float(row[12] if len(row) > 12 else None)
        new_inv  = clean_invoice(inv_no)
        gstinv   = make_gstinv(gstin, inv_no)
        records.append({
            "gstin": str(gstin).strip(),
            "name":  str(row[1] or "").strip(),
            "original_inv": inv_no,
            "new_inv":  new_inv,
            "gstinv":   gstinv,
            "inv_date": inv_date,
            "taxable":  taxable,
            "igst":     igst,
            "cgst":     cgst,
            "sgst":     sgst,
            "cess":     cess,
            "total":    inv_val if inv_val else taxable + igst + cgst + sgst + cess,
            "source":   "GSTR2B",
        })
    wb.close()
    return records

# ─────────────────────────────────────────────
#  PARSE TALLY / OUR DATA
# ─────────────────────────────────────────────
def parse_tally(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for i, row in enumerate(all_rows):
        row_text = " ".join(str(c).upper() for c in row if c)
        if "PARTY GSTIN" in row_text and "DOC NO" in row_text.replace(".", "").replace(" ", ""):
            header_row_idx = i
            break
            
    if header_row_idx is None:
        for i, row in enumerate(all_rows):
            if any(c and "GSTIN" in str(c).upper() for c in row):
                header_row_idx = i
                break

    if header_row_idx is None:
        wb.close()
        raise ValueError("Could not find the Tally header row.")

    header = all_rows[header_row_idx]
    col_gstin   = next((i for i, c in enumerate(header) if c and "GSTIN" in str(c).upper()), 2)
    col_doc_no  = next((i for i, c in enumerate(header) if c and "DOC" in str(c).upper() and "NO" in str(c).upper()), 6)
    col_taxable = next((i for i, c in enumerate(header) if c and "TAXABLE" in str(c).upper()), 11)
    
    col_igst  = col_taxable + 1
    col_cgst  = col_taxable + 2
    col_sgst  = col_taxable + 3
    col_cess  = col_taxable + 4
    col_inv   = col_taxable + 6

    records = []
    for row in all_rows[header_row_idx + 2:]:
        if len(row) <= col_gstin:
            continue
        gstin = row[col_gstin]
        if not gstin or str(gstin).strip() in ("", "nan", "None"):
            continue
        gstin_str = str(gstin).strip()
        if not re.match(r"^\d{2}[A-Z]", gstin_str):
            continue

        inv_no  = str(row[col_doc_no] if len(row) > col_doc_no else "").strip()
        if not inv_no or inv_no in ("None", "nan"):
            inv_no = str(row[5] if len(row) > 5 else "").strip()

        inv_date = row[0]
        taxable  = safe_float(row[col_taxable]  if len(row) > col_taxable  else None)
        igst     = safe_float(row[col_igst]     if len(row) > col_igst     else None)
        cgst     = safe_float(row[col_cgst]     if len(row) > col_cgst     else None)
        sgst     = safe_float(row[col_sgst]     if len(row) > col_sgst     else None)
        cess     = safe_float(row[col_cess]     if len(row) > col_cess     else None)
        inv_val  = safe_float(row[col_inv]      if len(row) > col_inv      else None)
        new_inv  = clean_invoice(inv_no)
        gstinv   = make_gstinv(gstin_str, inv_no)

        records.append({
            "gstin": gstin_str,
            "name":  str(row[1] or "").strip(),
            "original_inv": inv_no,
            "new_inv":  new_inv,
            "gstinv":   gstinv,
            "inv_date": inv_date,
            "taxable":  taxable,
            "igst":     igst,
            "cgst":     cgst,
            "sgst":     sgst,
            "cess":     cess,
            "total":    inv_val if inv_val else taxable + igst + cgst + sgst + cess,
            "source":   "OUR DATA",
        })
    wb.close()
    return records

# ─────────────────────────────────────────────
#  STYLES & EXCEL WRITER
# ─────────────────────────────────────────────
FILL_MATCH    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_NOMINAL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_DIFF     = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
FILL_NO_OUR   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
FILL_NO_GSTR  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
FILL_AI_MATCH = PatternFill(start_color="D8B4FE", end_color="D8B4FE", fill_type="solid")

STATUS_FILL = {
    "Matched":                 FILL_MATCH,
    "Nominal Difference":      FILL_NOMINAL,
    "Difference":              FILL_DIFF,
    "Not in Our Data":         FILL_NO_OUR,
    "Not in GSTR2B":           FILL_NO_GSTR,
    "Matched with AI":         FILL_AI_MATCH,
    "Nominal Difference (AI)": FILL_NOMINAL,
    "Difference (AI)":         FILL_DIFF,
}

# ─────────────────────────────────────────────
#  MAIN RECONCILIATION
# ─────────────────────────────────────────────
def reconcile_and_write(gstr2b_file: str, tally_file: str, output_file: str, progress_cb=None,
                        matched_threshold: float = 2.0, nominal_threshold: float = 20.0):
    def log(msg):
        if progress_cb: progress_cb(msg)
        else: print(msg)

    # 0. Extract Dynamic Metadata
    log("Extracting metadata from 'Read me' sheet...")
    meta = get_gstr_info(gstr2b_file, log)

    log("Parsing GSTR-2B…")
    gstr_recs = parse_gstr2b(gstr2b_file)
    log(f"  → {len(gstr_recs)} invoices in GSTR-2B")

    log("Parsing Tally / Our Data…")
    tally_recs = parse_tally(tally_file)
    log(f"  → {len(tally_recs)} invoices in Tally file")

    all_recs = gstr_recs + tally_recs

    # 1. Calculate Status for each GSTINV
    FIELDS = ("taxable", "igst", "cgst", "sgst", "cess")
    agg: dict = {}
    for r in all_recs:
        key = r["gstinv"]
        src = r["source"]
        agg.setdefault(key, {})
        if src not in agg[key]:
            agg[key][src] = {f: 0.0 for f in FIELDS}
        for f in FIELDS:
            agg[key][src][f] += r.get(f, 0.0)

    def get_status(gstinv: str) -> str:
        d = agg.get(gstinv, {})
        has_g = "GSTR2B"   in d
        has_o = "OUR DATA" in d
        if has_g and has_o:
            g, o = d["GSTR2B"], d["OUR DATA"]
            max_diff = max(abs(g[f] - o[f]) for f in FIELDS)
            if max_diff <= matched_threshold:   return "Matched"
            elif max_diff <= nominal_threshold: return "Nominal Difference"
            else:                               return "Difference"
        elif has_g:             return "Not in Our Data"
        elif has_o:             return "Not in GSTR2B"
        return ""

    # 1b. AI Second-Pass: suffix digit matching for unmatched records
    log("Running AI suffix match for unmatched invoices...")

    # Build lookup: gstinv -> record (first seen), for unmatched only
    seen_gstinv: dict = {}
    for r in all_recs:
        k = r["gstinv"]
        if k not in seen_gstinv:
            seen_gstinv[k] = r

    gstr_unmatched: dict = {}   # gstin -> [(gstinv, digit_string)]
    tally_unmatched: dict = {}  # gstin -> [(gstinv, digit_string)]

    for k, r in seen_gstinv.items():
        st = get_status(k)
        if st == "Not in Our Data":
            gstr_unmatched.setdefault(r["gstin"], []).append((k, r["new_inv"]))
        elif st == "Not in GSTR2B":
            tally_unmatched.setdefault(r["gstin"], []).append((k, r["new_inv"]))

    ai_remap: dict = {}        # tally gstinv -> gstr gstinv
    ai_matched_keys: set = set()  # gstr gstinv keys that got AI-matched

    for gstin, g_list in gstr_unmatched.items():
        t_list = tally_unmatched.get(gstin, [])
        if not t_list:
            continue
        used_t: set = set()
        for g_key, g_digits in g_list:
            if not g_digits:
                continue
            for t_key, t_digits in t_list:
                if t_key in used_t or not t_digits:
                    continue
                if min(len(g_digits), len(t_digits)) >= 4 and (g_digits.endswith(t_digits) or t_digits.endswith(g_digits)):
                    ai_remap[t_key] = g_key
                    ai_matched_keys.add(g_key)
                    used_t.add(t_key)
                    break

    # Merge OUR DATA amounts from tally key into gstr key in agg
    for t_key, g_key in ai_remap.items():
        if t_key in agg and "OUR DATA" in agg[t_key]:
            agg[g_key]["OUR DATA"] = agg[t_key].pop("OUR DATA")

    log(f"  → {len(ai_matched_keys)} invoice(s) AI-matched via suffix digit logic")

    # 2. Group Data into Side-By-Side Format
    log("Grouping records for side-by-side view...")
    grouped_data = {}
    
    for r in all_recs:
        k = r["gstinv"]
        # Re-key AI-matched tally records to their GSTR-2B partner key
        if k in ai_remap:
            k = ai_remap[k]
        if k not in grouped_data:
            grouped_data[k] = {"books": {}, "gstr": {}}

        src_key = "gstr" if r["source"] == "GSTR2B" else "books"
        target = grouped_data[k][src_key]

        if not target:
            target.update(r)
        else:
            target["taxable"] += r["taxable"]
            target["igst"]    += r["igst"]
            target["cgst"]    += r["cgst"]
            target["sgst"]    += r["sgst"]
            target["cess"]    += r.get("cess", 0.0)
            target["total"]   += r["total"]

    # 3. Write Output Workbook
    log("Building output workbook…")
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "All Invoice Reco Status"
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Top Headers - DYNAMICALLY POPULATED
    ws_main["C1"], ws_main["D1"] = "GSTIN:", meta["gstin"]
    ws_main["C2"], ws_main["D2"] = "Trade Name:", meta["trade_name"]
    ws_main["C3"], ws_main["D3"] = "F.Y.:", meta["fy"]
    ws_main["C4"], ws_main["D4"] = "Period:", meta["period"]

    for row in range(1, 5):
        ws_main[f"C{row}"].font = bold_font
        ws_main[f"D{row}"].font = bold_font

    ws_main["A5"] = "GSTR-2B Reconciliation :: All Invoices Reconciliation Report"
    ws_main["A5"].font = Font(bold=True, size=14)
    
    ws_main["D6"] = "As Per Books of Accounts [A]"
    ws_main.merge_cells("D6:J6")
    ws_main["K6"] = "As Per GSTR-2B"
    ws_main.merge_cells("K6:Q6")
    ws_main["R6"] = "Difference"
    ws_main.merge_cells("R6:V6")
    
    for col in ["D6", "K6", "R6"]:
        ws_main[col].font = bold_font
        ws_main[col].alignment = center_align
        ws_main[col].fill = header_fill
        ws_main[col].border = thin_border

    headers = [
        "S.No.", "GSTIN", "Supplier Name",
        "Invoice No.", "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", "Cess",
        "Invoice No.", "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", "Cess",
        "Taxable Value", "IGST", "CGST", "SGST", "Cess",
        "Status"
    ]
    ws_main.append(headers) 
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws_main.cell(row=7, column=col_idx)
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
        ws_main.column_dimensions[get_column_letter(col_idx)].width = 14

    ws_main.column_dimensions["C"].width = 25 
    ws_main.freeze_panes = "A8"

    log("Writing grouped data rows...")
    for index, (gstinv, data) in enumerate(grouped_data.items(), start=1):
        b = data["books"]
        g = data["gstr"]
        
        if gstinv in ai_matched_keys:
            d = agg.get(gstinv, {})
            if "GSTR2B" in d and "OUR DATA" in d:
                g_amt, o_amt = d["GSTR2B"], d["OUR DATA"]
                max_diff = max(abs(g_amt[f] - o_amt[f]) for f in FIELDS)
                if max_diff <= matched_threshold:   status = "Matched with AI"
                elif max_diff <= nominal_threshold: status = "Nominal Difference (AI)"
                else:                               status = "Difference (AI)"
            else:
                status = "Matched with AI"
        else:
            status = get_status(gstinv)
        gstin_val = b.get("gstin") if b else g.get("gstin", "")
        name_val = b.get("name") if b else g.get("name", "")
        
        b_tax  = b.get("taxable", 0)
        b_igst = b.get("igst", 0)
        b_cgst = b.get("cgst", 0)
        b_sgst = b.get("sgst", 0)
        b_cess = b.get("cess", 0)

        g_tax  = g.get("taxable", 0)
        g_igst = g.get("igst", 0)
        g_cgst = g.get("cgst", 0)
        g_sgst = g.get("sgst", 0)
        g_cess = g.get("cess", 0)

        row_data = [
            index, gstin_val, name_val,

            b.get("original_inv", "") if b else "",
            b.get("inv_date", "") if b else "",
            b_tax, b_igst, b_cgst, b_sgst, b_cess,

            g.get("original_inv", "") if g else "",
            g.get("inv_date", "") if g else "",
            g_tax, g_igst, g_cgst, g_sgst, g_cess,

            round(b_tax  - g_tax,  2),
            round(b_igst - g_igst, 2),
            round(b_cgst - g_cgst, 2),
            round(b_sgst - g_sgst, 2),
            round(b_cess - g_cess, 2),

            status
        ]
        
        ws_main.append(row_data)
        
        current_row = ws_main.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_main.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if isinstance(cell.value, (int, float)) and col_idx not in (1, 4, 11):
                cell.number_format = "#,##0.00"
            if col_idx in (5, 12) and cell.value:
                cell.number_format = "DD-MM-YY"

        status_cell = ws_main.cell(row=current_row, column=23)
        if status in STATUS_FILL:
            status_cell.fill = STATUS_FILL[status]


    # 4. Sheet: Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 14

    seen: dict = {}
    for r in all_recs:
        k = r["gstinv"]
        rk = ai_remap.get(k, k)   # use remapped key if AI-matched
        if rk not in seen:
            seen[rk] = get_status(rk)

    statuses = ["Matched", "Nominal Difference", "Difference", "Not in Our Data", "Not in GSTR2B"]
    counts = {s: sum(1 for v in seen.values() if v == s) for s in statuses}
    ai_count = len(ai_matched_keys)

    rows_data = [
        ("GST Reconciliation Summary", None),
        (None, None),
        ("Status", "Invoice Count"),
    ] + [(s, counts[s]) for s in statuses] + [
        ("Matched with AI", ai_count),
        (None, None),
        ("GSTR-2B Records", len(gstr_recs)),
        ("Our Data Records", len(tally_recs)),
        ("Total Combined", len(all_recs)),
    ]

    for ri, (lbl, val) in enumerate(rows_data, 1):
        cl = ws_sum.cell(row=ri, column=1, value=lbl)
        cv = ws_sum.cell(row=ri, column=2, value=val)
        bold = ri in (1, 3, len(rows_data) - 2, len(rows_data) - 1, len(rows_data))
        cl.font = Font(name="Arial", size=11 if ri == 1 else 10, bold=bold or ri == 1)
        cv.font = Font(name="Arial", size=10, bold=bold)
        fill = STATUS_FILL.get(str(lbl))
        if fill:
            cl.fill, cv.fill = fill, fill
        elif str(lbl) == "Matched with AI":
            cl.fill, cv.fill = FILL_AI_MATCH, FILL_AI_MATCH

    log(f"Saving → {output_file}")
    wb.save(output_file)
    log("Reconciliation complete ✓")

# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────
def launch_gui(embedded=False):
    # ── Appearance ────────────────────────────────────────────────────────
    if not embedded:
        ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")

    # ── Root window ───────────────────────────────────────────────────────
    root = ctk.CTk()
    root.title("GST Reconciliation Tool")
    if not embedded:
        root.state("zoomed")
    root.resizable(True, True)
    root.columnconfigure(0, weight=1)

    content_row = 0
    if not embedded:
        root.rowconfigure(1, weight=1)
        content_row = 1

        # ── Header bar ────────────────────────────────────────────────────
        header_bar = ctk.CTkFrame(root, height=58, corner_radius=0,
                                   fg_color=("#1F2937", "#0F172A"))
        header_bar.grid(row=0, column=0, sticky="ew")
        header_bar.grid_propagate(False)
        header_bar.columnconfigure(1, weight=1)

        ctk.CTkLabel(header_bar,
                     text="  GST Reconciliation Tool",
                     font=("Arial", 18, "bold"),
                     text_color="#FFFFFF",
                     ).grid(row=0, column=0, sticky="w", padx=18)

        def _toggle_theme():
            new = "dark" if ctk.get_appearance_mode() == "Light" else "light"
            ctk.set_appearance_mode(new)
            _theme_btn.configure(text="☀  Light" if new == "dark" else "🌙  Dark")

        _theme_btn = ctk.CTkButton(
            header_bar, text="🌙  Dark", width=110, height=32,
            corner_radius=8, fg_color="transparent", border_width=1,
            border_color="#6B7280", text_color="#D1D5DB",
            hover_color="#374151", command=_toggle_theme,
        )
        _theme_btn.grid(row=0, column=2, padx=18, pady=13, sticky="e")

    else:
        root.rowconfigure(0, weight=1)

    # ── Main content (non-scrollable, expands to fill window) ─────────────
    main_frame = ctk.CTkFrame(root, corner_radius=0, fg_color=("gray92", "gray10"))
    main_frame.grid(row=content_row, column=0, sticky="nsew")
    main_frame.columnconfigure(0, weight=1)
    main_frame.rowconfigure(4, weight=1)   # log card row expands

    _CP = {"padx": 20, "pady": (10, 0)}   # common card padding

    # ═══════════════════════════════════════════════════
    # CARD 1 — Input / Output Files
    # ═══════════════════════════════════════════════════
    files_card = ctk.CTkFrame(main_frame, corner_radius=12)
    files_card.grid(row=0, column=0, sticky="ew", **_CP)
    files_card.columnconfigure(1, weight=1)

    ctk.CTkLabel(files_card, text="Input Files",
                 font=("Arial", 13, "bold"),
                 ).grid(row=0, column=0, columnspan=5, sticky="w", padx=16, pady=(12, 2))
    ctk.CTkFrame(files_card, height=1, fg_color=("gray75", "gray35"),
                 ).grid(row=1, column=0, columnspan=5, sticky="ew", padx=16, pady=(0, 8))

    def _file_row(parent, r, label_text):
        ctk.CTkLabel(parent, text=label_text, font=("Arial", 12), anchor="w",
                     ).grid(row=r, column=0, sticky="w", padx=(16, 8), pady=5)
        e = ctk.CTkEntry(parent, font=("Arial", 11), height=34)
        e.grid(row=r, column=1, sticky="ew", padx=(0, 8), pady=5)
        return e

    gstr_entry  = _file_row(files_card, 2, "GSTR-2B File (.xlsx):")
    tally_entry = _file_row(files_card, 3, "Tally / Our Data File (.xlsx):")

    def _browse_open(entry, title):
        def cb():
            f = filedialog.askopenfilename(
                title=title, filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])
            if f:
                entry.delete(0, tk.END)
                entry.insert(0, f)
        return cb

    # Sample file embedded as base64 so it works even if deleted from disk
    _SAMPLE_B64 = (
        'UEsDBBQABgAIAAAAIQBi7p1oXgEAAJAEAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAAC'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACs'
        'lMtOwzAQRfdI/EPkLUrcskAINe2CxxIqUT7AxJPGqmNbnmlp/56J+xBCoRVqN7ESz9x7MvHNaLJu'
        'bbaCiMa7UgyLgcjAVV4bNy/Fx+wlvxcZknJaWe+gFBtAMRlfX41mmwCYcbfDUjRE4UFKrBpoFRY+'
        'gOOd2sdWEd/GuQyqWqg5yNvB4E5W3hE4yqnTEOPRE9RqaSl7XvPjLUkEiyJ73BZ2XqVQIVhTKWJS'
        'uXL6l0u+cyi4M9VgYwLeMIaQvQ7dzt8Gu743Hk00GrKpivSqWsaQayu/fFx8er8ojov0UPq6NhVo'
        'Xy1bnkCBIYLS2ABQa4u0Fq0ybs99xD8Vo0zL8MIg3fsl4RMcxN8bZLqej5BkThgibSzgpceeRE85'
        'NyqCfqfIybg4wE/tYxx8bqbRB+QERfj/FPYR6brzwEIQycAhJH2H7eDI6Tt77NDlW4Pu8ZbpfzL+'
        'BgAA//8DAFBLAwQUAAYACAAAACEAtVUwI/QAAABMAgAACwAIAl9yZWxzLy5yZWxzIKIEAiigAAIA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKyS'
        'TU/DMAyG70j8h8j31d2QEEJLd0FIuyFUfoBJ3A+1jaMkG92/JxwQVBqDA0d/vX78ytvdPI3qyCH2'
        '4jSsixIUOyO2d62Gl/pxdQcqJnKWRnGs4cQRdtX11faZR0p5KHa9jyqruKihS8nfI0bT8USxEM8u'
        'VxoJE6UchhY9mYFaxk1Z3mL4rgHVQlPtrYawtzeg6pPPm3/XlqbpDT+IOUzs0pkVyHNiZ9mufMhs'
        'IfX5GlVTaDlpsGKecjoieV9kbMDzRJu/E/18LU6cyFIiNBL4Ms9HxyWg9X9atDTxy515xDcJw6vI'
        '8MmCix+o3gEAAP//AwBQSwMEFAAGAAgAAAAhAHWxC3Z1AwAAwQgAAA8AAAB4bC93b3JrYm9vay54'
        'bWysVW1vozgQ/n7S/QfEd4rNSwKodBUCaCu1qyrNtndSpcoFU6wC5oxpUlX7328MIW03p1Oue1Fi'
        'x57h8TMzj4fTL9u60p6p6BhvQh2fIF2jTcZz1jyG+vd1ani61knS5KTiDQ31F9rpX85+/+10w8XT'
        'A+dPGgA0XaiXUraBaXZZSWvSnfCWNmApuKiJhKV4NLtWUJJ3JaWyrkwLoZlZE9boI0IgjsHgRcEy'
        'GvOsr2kjRxBBKyKBfleytpvQ6uwYuJqIp741Ml63APHAKiZfBlBdq7Pg/LHhgjxUEPYWu9pWwHcG'
        'P4xgsKaTwHRwVM0ywTteyBOANkfSB/FjZGL8IQXbwxwch+SYgj4zVcM9KzH7JKvZHmv2BobRL6Nh'
        'kNaglQCS90k0d8/N0s9OC1bRm1G6Gmnbb6RWlap0rSKdTHImaR7qc1jyDf2wIfo26lkFVsv3LEs3'
        'z/ZyvhJaTgvSV3INQp7gwRFZNkLKE4SxqCQVDZF0yRsJOtzF9auaG7CXJQeFayv6V88EhYsF+oJY'
        'YSRZQB66KyJLrRdVqC+Du+8dhH/39eou5pum4nC77t7Jkhzegf8gTJKpaE0Id6Q0/v85dGAmgkl8'
        'V1Jo8P88voACXJNnKAcUPd/d1nPIN7bvm0wE+P7V82xnhiJsxC6ODMdBluGjJDV8nCwthJGFneUP'
        'CEbMgoyTXpa7SivoUHegrAemS7KdLBgFPcvfaLyi3cdQ80/DZPuhAlY97YbRTfemCbXUtresyfkm'
        '1A2slPzycbkZjLcslyUE6TsWuIx7Xyl7LIExRko/qnEoZqH+mnqxZ7sL10h9LzWcJPWMKLJtI0os'
        'FyXWwpm7zsDIfEdp6J5AbZi1ZlD8teqoGNq0mock65oI1BniPMdDEafHMlJloHA1DY4+5NlXHnQr'
        'Lzo5zCAuBvSwgxZz5DsGSmzXcDzfMjzHtoylE1uJO0/iJHJVfVT3D/6PHjhoPJheK4plSYRcC5I9'
        'wctoRYuIdCCoMSDg+55s5HoRsoGik2LIJfYR5HLmGG6c2u4cx8vETd/IqvCLT3YgzxyepkT2cDvV'
        'xRzWgRrT3e5+sxg3dnX6cPeCVazyvnv63xyvIfqKHumc3hzpuPx2ub480vciWd/fpsc6Ly6jeHG8'
        '/2K1Wvy5Tv6YjjD/MaHmUHA1DjI1J5mc/Q0AAP//AwBQSwMEFAAGAAgAAAAhAIE+lJfzAAAAugIA'
        'ABoACAF4bC9fcmVscy93b3JrYm9vay54bWwucmVscyCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAKxSTUvEMBC9C/6HMHebdhUR2XQvIuxV6w8IybQp2yYhM3703xsqul1Y1ksvA2+G'
        'ee/Nx3b3NQ7iAxP1wSuoihIEehNs7zsFb83zzQMIYu2tHoJHBRMS7Orrq+0LDppzE7k+ksgsnhQ4'
        '5vgoJRmHo6YiRPS50oY0as4wdTJqc9Adyk1Z3su05ID6hFPsrYK0t7cgmilm5f+5Q9v2Bp+CeR/R'
        '8xkJSTwNeQDR6NQhK/jBRfYI8rz8Zk15zmvBo/oM5RyrSx6qNT18hnQgh8hHH38pknPlopm7Ve/h'
        'dEL7yim/2/Isy/TvZuTJx9XfAAAA//8DAFBLAwQUAAYACAAAACEAl9tT0pMIAAAQKgAAGAAAAHhs'
        'L3dvcmtzaGVldHMvc2hlZXQxLnhtbJyaW3OjPBKG77dq/4OL+9jm5FPF+SoxPp8Pu3tNHJJQY5ss'
        'kGSmvtr/vi2EMLwiGTFTM2Py5FWrpW61QOb2r5/nU+3DCyM/uPQ1vd7Uat7lGDz5l5e+9q/D6Kaj'
        '1aLYvTy5p+Di9bVfXqT9dffPf9x+BuGP6NXz4hpZuER97TWO33qNRnR89c5uVA/evAv95jkIz25M'
        'P4Yvjegt9NynpNH51DCazVbj7PoXjVvohSo2gudn/+g5wfH97F1ibiT0Tm5M/kev/lskrJ2PKubO'
        'bvjj/e3mGJzfyMSjf/LjX4lRrXY+9qYvlyB0H0807p+65R5rP0P6a9A/U3STcKmns38Mgyh4jutk'
        'ucF9loffbXQb7jGzJI9fyYxuNULvw2cBvJoy/swl3c5sGVdj5h8aa2XG2HSFvXf/qa/9bd6bI10f'
        'OjfN+07nxho69zfdkd29MY3hYNAZNm1naP1Pu7t98inCbFS10Hvua/d676DbWuPuNkmgf/veZ5S7'
        'rsXu4947ecfYo050rRYHbwvvOR54pxM1bms1lrCPQfCDtZySpkl9REkL1od7jP0Pj6sfSB39N+mV'
        'LqnHRtZl/lp0P0pSfBPWnrxn9/0U74LPiee/vMbkh1W3aegsd3pPvxwvOlLSUtd1MxnJMTiREfq/'
        'dvbZ6qOkc38mn5/+U/xKV2a9o7esZssgM49eFI98ZlWrHd+jODj/J1UxHzMrFLjECn2mVqx23bJN'
        'vYINildigz6FJ3a9W8kNKzVBn8IEeaQ4Bhpt0j99po3b1SeilRqhz9RIq+o8UCIkfrD0See6XW+Z'
        'RqfKZFIFTYzQZzaYanPZTS3Qp3Cj8lB0quw8u+giSy/liOhZctLFNTurjUMXuckuUiOdull1QnWR'
        'nlQPrlNaMcV1kR7sQiyU6kmmixRhF1l4q/oiUkS/5sifzIvIE/2aKO2Ky9YQacIusoWrmvMNXtGS'
        'cum4sXt3GwafNdrYKGmiN5fdJug9ZjkpjvbXxZHqGWt2z9r1NZPmlupeRBX7485q3TY+qA4fU81D'
        'pmFVkLUaSMThRE8KJZMMEYwQjBFMEEwRzBDMESwQLBGsEKwRbBBsEewQ7BEccqBBIcriROuyEKfy'
        'rUtEh6kpOrSArtFpQ3QyTRYdiTic5KKDYIRgjGCCYIpghmCOYIFgiWCFYI1gg2CLYIdgj+CQA4Xo'
        'UN2rEB2mxuh0IDqZJouORBxOctFBMEIwRjBBMEUwQzBHsECwRLBCsEawQbBFsEOwR3DIgUJ06E6n'
        'QnSYmqJDBf+6dmyITqbJoiMRh5NcdBCMEIwRTBBMEcwQzBEsECwRrBCsEWwQbBHsEOwRHHKgEB26'
        'VfijHYi1ozjRjprFqQlhyiRZmCTicJILE4IRgjGCCYIpghmCOYIFgiWCFYI1gg2CLYIdgj2CQw4U'
        'wkR7SYVFxNRQ4nQITibJgiMRh5NccBCMEIwRTBBMEcwQzBEsECwRrBCsEWwQbBHsEOwRHHKgEBz2'
        'XJy/i/v+7oCp+xrtZdnCMSA2XEH19FoCzaJkwCXX0DgIhghGCMYIJgimCGYI5ggWCJYIVgjWCDYI'
        'tgh2CPYIDjlQiBU7t1OPFVMXYwWBeOCKQqwgVFyRCxWCIYIRgjGCCYIpghmCOYIFgiWCFYI1gg2C'
        'LYIdgj2CQw4UQkV7S4VQMXVfo20uWzR428AVVGEzBTwyDbiClnmmgNt2hysouTIF3DoOuYLGlCm6'
        'xYQYyZ7qsHWOuSS/uepQwCclEqgj0xIJpO+sRGIV3Z2XSGBmFyUSmNpliQTmdlUigcldl0hgdjey'
        'xIDZ3ZZIYHZ3XMIOh65VGqZ3X6aB+T2Uaa4TXEh31lmFfE/kdG5Aj6xXFzHlhaiw24BooCJySruD'
        'IA9VLI1UHB8LUb6wGuD4pEwEmqmKoZmKaF4qghxelIogi5cq3a1UROtSEa4IFUtbFcd3QpSvsgaM'
        'bl8mMmEJHkpF10VYXBjsVEZ9z2Yrgh2o8a8S2KnWQ4ryyWTCah6IZtlZmCORoUToGxveWdZqLJGJ'
        'RKYSmUlkLpGFRJYSWUlkLZGNRLYS2UlkL5FDnhQDxg5qKgQsPddp8XNMu23hcQw7F4f7MAu2qYGC'
        'xkk1dLSdnJjKPQ1LrJhQ00fCSv4uwAR3xkKU38NNqYCl42LfudEJruzPVPiTr/ImlN1ZqQgq01w4'
        'lL8zMWGiF8KSkThk2rh1LoXAzBJ+BW30Vh137d8qNrLZbbFRyzTrMOydUFh89nQbYrAXgvwtnwnV'
        '8VAmyu00xcRmZ1wVEpsfidHB/hfhfWBf2bDEzodXzuxUlH8ehKE6wtCXXQ3LupJTOz3Ey2ethTen'
        'qSX2bcr1ARVvT3/r0LTUISm1SybIlFI79fr71E4t8dTWm027CQNbCo/yyV1o1W1asITXog03Kws2'
        'stFtsY3eaXbBk51Q8NTW9Y6Nkr2QfJ/cqft50ZfJzY4IKyQ3P1E0WlkleNBTlC+MlrTPlojwzs4R'
        'lq7Gh2XGsd1IiPKZgKKximgiezBV8WCmIpqriBZC1OW12DQoY+stu5v+gUWwBLnRNbt13RDqLj5s'
        'FeW6aXY79TasvrWKaJOKzGaWCNtiO7vbNlvkefG5cifa6Xx8XaPZbtVtfMgSqu+eYA6/ERVrODti'
        'rZDm/ETW/O5B5IF94c4OXAv3xdJjloLIEZa+626o0t1IxdJYRTRREU1VfJqpiOYqooWKT0sV0UpF'
        'tFYRbVREWxXRToi+KNvJt/l7lXlir4Z9l5l8ZfD3uPiLCWcvfEne94pqx+CdvZXFbpYzmr501uoN'
        'ktKMXO/RSwb0jQJyozcwyrjZGySbLeqt3sAq09u9QfJY17i6eXf75r54Szd88S9R7UTvtrEXyWgL'
        'Cvm7Zsk1vfWWUJqNxyCmV8TET6/05qVHj4j06plWew6CWPzAXm/L3uW8+z8AAAD//wMAUEsDBBQA'
        'BgAIAAAAIQD2YLRBuAcAABEiAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxazY8btxW/B8j/QMxd'
        '1szoe2E50Kc39u564ZVd5EhJlIZeznBAUrsrFAEK59RLgQJp0UuB3nooigZogAa55I8xYCNN/4g8'
        'ckaa4YqKvf5AkmJ3LzPU7z3+5r3HxzePc/eTq5ihCyIk5UnXC+74HiLJjM9psux6TybjSttDUuFk'
        'jhlPSNdbE+l9cu/jj+7iAxWRmCCQT+QB7nqRUulBtSpnMIzlHZ6SBH5bcBFjBbdiWZ0LfAl6Y1YN'
        'fb9ZjTFNPJTgGNQ+WizojKCJVund2ygfMbhNlNQDMybOtGpiSRjs/DzQCLmWAybQBWZdD+aZ88sJ'
        'uVIeYlgq+KHr+ebPq967W8UHuRBTe2RLcmPzl8vlAvPz0MwpltPtpP4obNeDrX4DYGoXN2rr/60+'
        'A8CzGTxpxqWsM2g0/XaYY0ug7NKhu9MKaja+pL+2wznoNPth3dJvQJn++u4zjjujYcPCG1CGb+zg'
        'e37Y79QsvAFl+OYOvj7qtcKRhTegiNHkfBfdbLXbzRy9hSw4O3TCO82m3xrm8AIF0bCNLj3Fgidq'
        'X6zF+BkXYwBoIMOKJkitU7LAM4jiXqq4REMqU4bXHkpxwiUM+2EQQOjV/XD7byyODwguSWtewETu'
        'DGk+SM4ETVXXewBavRLk5TffvHj+9Yvn/3nxxRcvnv8LHdFlpDJVltwhTpZluR/+/sf//fV36L//'
        '/tsPX/7JjZdl/Kt//v7Vt9/9lHpYaoUpXv75q1dff/XyL3/4/h9fOrT3BJ6W4RMaE4lOyCV6zGN4'
        'QGMKmz+ZiptJTCJMLQkcgW6H6pGKLODJGjMXrk9sEz4VkGVcwPurZxbXs0isFHXM/DCKLeAx56zP'
        'hdMAD/VcJQtPVsnSPblYlXGPMb5wzT3AieXg0SqF9EpdKgcRsWieMpwovCQJUUj/xs8JcTzdZ5Ra'
        'dj2mM8ElXyj0GUV9TJ0mmdCpFUiF0CGNwS9rF0FwtWWb46eoz5nrqYfkwkbCssDMQX5CmGXG+3il'
        'cOxSOcExKxv8CKvIRfJsLWZl3Egq8PSSMI5GcyKlS+aRgOctOf0hhsTmdPsxW8c2Uih67tJ5hDkv'
        'I4f8fBDhOHVypklUxn4qzyFEMTrlygU/5vYK0ffgB5zsdfdTSix3vz4RPIEEV6ZUBIj+ZSUcvrxP'
        'uL0e12yBiSvL9ERsZdeeoM7o6K+WVmgfEcLwJZ4Tgp586mDQ56ll84L0gwiyyiFxBdYDbMeqvk+I'
        'hDJJ1zW7KfKISitkz8iS7+FzvL6WeNY4ibHYp/kEvG6F7lTAYnRQeMRm52XgCYXyD+LFaZRHEnSU'
        'gnu0T+tphK29S99Ld7yuheW/N1ljsC6f3XRdggy5sQwk9je2zQQza4IiYCaYoiNXugURy/2FiN5X'
        'jdjKKbewF23hBiiMrHonpsnrip8TLAS//Hlqnw9W9bgVv0u9sy+vHF6rcvbhfoW1zRCvklMC28lu'
        '4rotbW5LG+//vrTZt5ZvC5rbgua2oHG9gn2QgqaoYaC8KVo9pvET7+37LChjZ2rNyJE0rR8JrzXz'
        'MQyanpRpTG77gGkEl/p5YAILtxTYyCDB1W+ois4inEJ/KDBdzKXMVS8lSrmEtpEZNv1Uck23aT6t'
        '4mM+z9qdpr/kZyaUWBXjfgMaT9k4tKpUhm628kHNb0PdsF2aVuuGgJa9CYnSZDaJmoNEazP4GhK6'
        'c/Z+WHQcLNpa/cZVO6YAaluvwHs3grf1rteoZ4ygIwc1+lz7KXP1xrvaOe/V0/uMycoRAK3FXU93'
        'NNe9j6efLgu1N/C0RcI4JQsrm4TxlSnwZARvw3l0lvvuPxVwN/V1p3CpRU+bYrMaChqt9ofwtU4i'
        '13IDS8qZgiXoEtZ4CIvOQzOcdr0F9I3hMk4heKR+98JsCYcvMyWyFf82qSUVUg2xjDKLm6yT+Sem'
        'igjEaNz19PNvw4ElJolk5DqwdH+p5EK94H5p5MDrtpfJYkFmquz30oi2dHYLKT5LFs5fjfjbg7Uk'
        'X4G7z6L5JZqylXiMIcQarUB7d04lHB8EmavnFM7DtpmsiL9rO1Oe/a1DriIfY5ZGON9Sytk8g5sN'
        'ZUvH3G1tULrLnxkMumvC6VLvsO+87b5+r9aWK/bHTrFpWmlFb5vubPrhdvkSq2IXtVhluft6zu1s'
        'kh0EqnObePe9v0StmMyiphnv5mGdtPNRm9p7rAhKu09zj922m4TTEm+79YPc9ajVO8SmsDSBbw7O'
        'y2fbfPoMkscQThFXLDvtZgncmdIyPRXGt1M+X+eXTGaJJvO5LkqzVP6YLBCdX3W90FU55ofHeTXA'
        'EkCbmhdW2FbQWe3Zgnqzy0WzBbsVzsrYa/WqLbyV2ByzboVNa9FFW11tTtR1rW5m1g7LntqkYWMp'
        'uNq1IrTJBYbSOTvMzXIv5JkrlVfacIVWgna93/qNXn0QNgYVv90YVeq1ul9pN3q1Sq/RqAWjRuAP'
        '++HnQE9FcdDIvnwYw2kQW+ffP5jxnW8g4s2B150Zj6vcfONQNd4330AE4f5vIMCRQCscBfWwFw4q'
        'g2HQrNTDYbPSbtV6lUHYHIY92LSb497nHrow4KA/HI7HjbDSHACu7vcalV6/Nqg026N+OA5G9aEP'
        '4Hz7uYK3GJ1zc1vApeF170cAAAD//wMAUEsDBBQABgAIAAAAIQCEeGYmBQUAAL8mAAANAAAAeGwv'
        'c3R5bGVzLnhtbNRaX2/iOBB/P+m+Q+SHe6P5U8IWjrAqbSOttFed1J50ryZxwFrH5hzTDXu6735j'
        'h0BogSYsUOgDTRx75ueZ8Xg84/7nPGXWC5EZFTxA7pWDLMIjEVM+DtBfz2HrBlmZwjzGTHASoDnJ'
        '0OfBr7/0MzVn5GlCiLKABM8CNFFq2rPtLJqQFGdXYko4fEmETLGCVzm2s6kkOM70oJTZnuN07BRT'
        'jgoKvTSqQyTF8tts2opEOsWKjiijam5oISuNel/GXEg8YgA1d9s4snK3Iz0rlyUT0/qGT0ojKTKR'
        'qCuga4skoRF5C7drd20crSgB5f0oub7teGtzz+WelNq2JC9Uqw8N+nyWhqnKrEjMuAqQt2yyii9f'
        'YtBxp42sQit3IgY5/fbPTKjfi1/nynGQXRJaG+XvGKWH2Avmg34i+ApDB8SlFdH7xsV3HupPgAGA'
        '6V6DfvbDesEMWlxNIxJMSEuBAQEu08JxSooet1MlMusRSym+674JTimbF9883WAMb9E5pWAGBlXB'
        '5hUzM8edzCTFbCMXu0pwpNmWMzAgDkC0JNh9VyLNQR6ZphyPAhSGDvyFobGjlfreA2sEm4FeKWNr'
        'FqwbBn1Y7IpIHsKLtXh+nk/BTDj4pULTpt87vccSz13PrwywDcNBfyRkDH6wXDtdMNGiadBnJFGg'
        'aUnHE/1fiSn8joRS4CsG/ZjiseCY6TVQjqgxEtwqeNAAqQl4wNL0KY9JTmCVwiIt6GkmTXnUIg2z'
        'aDCJWiS1nNbEdAZAakEwmjWKPRvRLS3soHbSVI0NpXceoJsv0VOJZYu3WLgNcH4RYexJ+4W/k6Un'
        '0ptlnlT2YojR9CakN3P9CG5v8Vh4neJFe48qtYJ2hWzb2YuulSdLBttQuVtRLUdbeDplcx0O6I2+'
        'eLtldMxTUjQN+rh81fGpopEOE8BpGb+YJ69k0u6uhHJdl/3jLB0RGZoIdQXjOKB00LVZU69kckpQ'
        '2yTlQdC2UHOhmjqghmYDraHOiZD0B2heK9R4XtRUw9uEeTrceqt7DdvSu7e23s77NvoxM2i4kD5e'
        'zPuZx8eY9YGEC+BPtPYiMFYCB+MDAQdruUzgsP9cpLc7Lu4DWQWkDS7TKnTi4jR7YIN1WAn9agcU'
        'NeOZ/XZl16SFimD0NHHXz0cP54yzouBtMD+9Nsx1BdcOww4UVb+BcwaxYh2rPEfcdU4xx8VdwygW'
        'CeTda/4sraJTw1mdJ3CTst8tcSjUnHUUs82dHRd3DXuusae+icn3dLnHO/ke99RwcUeb+ufGNW+2'
        'LaY63ZmsCfCKN7ss4BVvtg346Q6Th81AHRd3w3WoDWRTzvF0Z7L9hPsxuA8k3NOdGw+bHINtsHn8'
        'cKiUPSRuLyc7flZ5l22SO24qfPNaMQUnKDFV6lhrVaxlPcrS1zoC9KgrL6xid6MZZYryDRUsoBnn'
        'q5qYucCh9D0fUy1bcgEjjkmCZ0w9Lz8GaPX8B4npLAXZLHr9SV+EMiQCtHr+qgv9rsnik1x9zaCo'
        'Df+tmaQB+vdh+Kl7/xB6rRtneNNqXxO/1fWH9y2/fTe8vw+7jufc/Ve5bfQTd43M5SgodLntXsbg'
        'RpJcTHYB/mnVFqDKSwHf1CAAdhV71+s4t77rtMJrx221O/imddO59luh73r3nfbwwQ/9CnZ/zztJ'
        'ju26xe0mDd7vKZoSRnmpq1JD1VZQErzumIRdasJe3Twb/A8AAP//AwBQSwMEFAAGAAgAAAAhAC51'
        'i6VZAgAA7gUAABQAAAB4bC9zaGFyZWRTdHJpbmdzLnhtbHxU32/aMBB+n7T/4ZSHqZMWkrDSrQzo'
        'KHQTU2EIAip788wB1hw7sx0E++t3aZm02XR5ifzdfffL37lzcygk7NFYoVU3yhppBKi4Xgu17UaL'
        '/FP8PgLrmFozqRV2oyPa6Kb38kXHWgfEVbYb7Zwr20li+Q4LZhu6REWWjTYFc3Q028SWBtna7hBd'
        'IZNmml4lBRMqAq4r5ShvmkVQKfGzwsETcnkd9TpW9Dqu93mez+LmLcyQa8WFFMxRsRDDUleU0pBh'
        'K6xD00lcr5PUpCdiFvdLEzdb4DS8TU8H34minwKYx7ht3+GUxYLeBLZbKmuk9lpwtD5vyBz62JQZ'
        'J3glmQnca9MRqJrRJFmMJj7zbrNB7sQe4VzcJd9BfiyDfDU+0Y2gNM2fgX3PU3Nng08l40hTgXlV'
        'lvLoU/sECs6+SwSSgm/N2aE2BfloAD42OIPNCUsCR7TBWCnPMz0FcD7woRkyS0qj8mEUWJN/tFb7'
        'zNBVRoVB6vVCGOyY2QYN94ta7z5nkZ/pub9nQp4b2kKxJxPthgvv4Q8Nvip5BKHgVusfwaC+aKqd'
        'Sb+Spg+MaX2brbh5lby7Dlvd1tqG/64Fyp3wiRPtI6twoVpBbadC0uwykJd2YS9ba5yj/i8GuiiZ'
        'EfXdknwfxwGvWFF+gKk2RIQlkxXa137Uw/GXD93FY7oVaNNLphzj7iMeKJDEhgiE8LD6BlMj9vQu'
        'wL0ohMN1oOBRvnoDw3pEkGVpmgUjXgxX/TGlq//9cTy8j9Pr+OH0wcVYcKOTOZp9/ST91UBC73Xv'
        'NwAAAP//AwBQSwMEFAAGAAgAAAAhAGyctWRUAQAAZwIAABEACAFkb2NQcm9wcy9jb3JlLnhtbCCi'
        'BAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAISSX0vDMBTF3wW/Q8l7m/7Zioa2Ayfz'
        'xYFoRfEtJHdrsUlDktnt25t2W+1Q8DH3nPxyziXZYi8a7wu0qVuZoygIkQeStbyW2xy9liv/BnnG'
        'Uslp00rI0QEMWhTXVxlThLUannSrQNsajOdI0hCmclRZqwjGhlUgqAmcQzpx02pBrTvqLVaUfdIt'
        '4DgMUyzAUk4txT3QVyMRnZCcjUi1080A4AxDAwKkNTgKIvzjtaCF+fPCoEycorYH5Tqd4k7ZnB3F'
        '0b039Wjsui7okiGGyx/h9/Xjy1DVr2W/KwaoyDgjTAO1rS6euPewU5YmGZ5M+w021Ni1W/amBn53'
        'KJ7bqrbesqK7isoM/zY46lDiiAbuuVjkWOKsvCXL+3KFijiMUz9M/Dgso4jMIjJPP/r3L+73MY8D'
        'cUrxL3Hmx3EZpiS5Jcl8QjwDiiH35dcovgEAAP//AwBQSwMEFAAGAAgAAAAhAGFJCRCJAQAAEQMA'
        'ABAACAFkb2NQcm9wcy9hcHAueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
        'nJJBb9swDIXvA/ofDN0bOd1QDIGsYkhX9LBhAZK2Z02mY6GyJIiskezXj7bR1Nl66o3ke3j6REnd'
        'HDpf9JDRxVCJ5aIUBQQbaxf2lXjY3V1+FQWSCbXxMUAljoDiRl98UpscE2RygAVHBKxES5RWUqJt'
        'oTO4YDmw0sTcGeI272VsGmfhNtqXDgLJq7K8lnAgCDXUl+kUKKbEVU8fDa2jHfjwcXdMDKzVt5S8'
        's4b4lvqnszlibKj4frDglZyLium2YF+yo6MulZy3amuNhzUH68Z4BCXfBuoezLC0jXEZtepp1YOl'
        'mAt0f3htV6L4bRAGnEr0JjsTiLEG29SMtU9IWT/F/IwtAKGSbJiGYzn3zmv3RS9HAxfnxiFgAmHh'
        'HHHnyAP+ajYm0zvEyznxyDDxTjjbgW86c843XplP+id7HbtkwpGFU/XDhWd8SLt4awhe13k+VNvW'
        'ZKj5BU7rPg3UPW8y+yFk3Zqwh/rV878wPP7j9MP18npRfi75XWczJd/+sv4LAAD//wMAUEsBAi0A'
        'FAAGAAgAAAAhAGLunWheAQAAkAQAABMAAAAAAAAAAAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54'
        'bWxQSwECLQAUAAYACAAAACEAtVUwI/QAAABMAgAACwAAAAAAAAAAAAAAAACXAwAAX3JlbHMvLnJl'
        'bHNQSwECLQAUAAYACAAAACEAdbELdnUDAADBCAAADwAAAAAAAAAAAAAAAAC8BgAAeGwvd29ya2Jv'
        'b2sueG1sUEsBAi0AFAAGAAgAAAAhAIE+lJfzAAAAugIAABoAAAAAAAAAAAAAAAAAXgoAAHhsL19y'
        'ZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAJfbU9KTCAAAECoAABgAAAAAAAAA'
        'AAAAAAAAkQwAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQItABQABgAIAAAAIQD2YLRBuAcA'
        'ABEiAAATAAAAAAAAAAAAAAAAAFoVAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAi0AFAAGAAgAAAAh'
        'AIR4ZiYFBQAAvyYAAA0AAAAAAAAAAAAAAAAAQx0AAHhsL3N0eWxlcy54bWxQSwECLQAUAAYACAAA'
        'ACEALnWLpVkCAADuBQAAFAAAAAAAAAAAAAAAAABzIgAAeGwvc2hhcmVkU3RyaW5ncy54bWxQSwEC'
        'LQAUAAYACAAAACEAbJy1ZFQBAABnAgAAEQAAAAAAAAAAAAAAAAD+JAAAZG9jUHJvcHMvY29yZS54'
        'bWxQSwECLQAUAAYACAAAACEAYUkJEIkBAAARAwAAEAAAAAAAAAAAAAAAAACJJwAAZG9jUHJvcHMv'
        'YXBwLnhtbFBLBQYAAAAACgAKAIACAABIKgAAAAA='
    )

    def _download_tally_sample():
        import base64

        save_path = filedialog.asksaveasfilename(
            title="Save Tally/Our Data Sample",
            initialfile="input2 - sample.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
        )
        if not save_path:
            return

        try:
            data = base64.b64decode("".join(_SAMPLE_B64.split()))
            with open(save_path, "wb") as f:
                f.write(data)
            messagebox.showinfo("Success", f"Sample saved to:\n{save_path}")
        except Exception as exc:
            messagebox.showerror("Save Failed", f"Could not save sample:\n{exc}")

    ctk.CTkButton(files_card, text="Browse", width=90, height=34,
                  command=_browse_open(gstr_entry, "Select GSTR-2B File"),
                  ).grid(row=2, column=2, padx=(0, 16), pady=5)
    ctk.CTkButton(files_card, text="Browse", width=90, height=34,
                  command=_browse_open(tally_entry, "Select Tally / Our Data File"),
                  ).grid(row=3, column=2, padx=(0, 16), pady=(5, 14))
    ctk.CTkButton(files_card, text="Download Sample", width=150, height=34,
                  command=_download_tally_sample,
                  ).grid(row=3, column=3, padx=(0, 16), pady=(5, 14))

    def _open_demo_link():
        import webbrowser
        webbrowser.open_new_tab("https://www.youtube.com/watch?v=invalid")

    ctk.CTkButton(files_card, text="▶ View Demo", width=120, height=34,
                  fg_color="#DC2626", hover_color="#B91C1C", text_color="white",
                  command=_open_demo_link,
                  ).grid(row=2, column=3, padx=(0, 16), pady=5)

    # ═══════════════════════════════════════════════════
    # CARD 2 — Reconciliation Thresholds
    # ═══════════════════════════════════════════════════
    thresh_card = ctk.CTkFrame(main_frame, corner_radius=12)
    thresh_card.grid(row=1, column=0, sticky="ew", **_CP)

    ctk.CTkLabel(thresh_card, text="Reconciliation Thresholds (₹)",
                 font=("Arial", 13, "bold"),
                 ).grid(row=0, column=0, columnspan=6, sticky="w", padx=16, pady=(12, 2))
    ctk.CTkFrame(thresh_card, height=1, fg_color=("gray75", "gray35"),
                 ).grid(row=1, column=0, columnspan=6, sticky="ew", padx=16, pady=(0, 8))

    ctk.CTkLabel(thresh_card, text="Matched up to (₹):", font=("Arial", 12),
                 ).grid(row=2, column=0, sticky="w", padx=(16, 8), pady=(4, 14))
    matched_var = tk.StringVar(value="2")
    ctk.CTkEntry(thresh_card, textvariable=matched_var, width=110, height=34,
                 font=("Arial", 12),
                 ).grid(row=2, column=1, sticky="w", padx=(0, 30), pady=(4, 14))

    ctk.CTkLabel(thresh_card, text="Nominal Difference up to (₹):", font=("Arial", 12),
                 ).grid(row=2, column=2, sticky="w", padx=(0, 8), pady=(4, 14))
    nominal_var = tk.StringVar(value="20")
    ctk.CTkEntry(thresh_card, textvariable=nominal_var, width=110, height=34,
                 font=("Arial", 12),
                 ).grid(row=2, column=3, sticky="w", padx=(0, 16), pady=(4, 14))

    run_btn = ctk.CTkButton(
        thresh_card,
        text="▶   Run Reconciliation",
        font=("Arial", 12, "bold"),
        height=36, width=200, corner_radius=10,
        fg_color="#059669", hover_color="#047857",
        text_color="#FFFFFF"
    )
    run_btn.grid(row=2, column=4, sticky="e", padx=(8, 16), pady=(4, 14))

    def _open_out_folder():
        d = os.path.join(os.getcwd(), "GST Downloaded", "GST Reco")
        if os.path.exists(d):
            os.startfile(d)
        else:
            messagebox.showinfo("Info", "Output folder not found.")

    open_folder_btn = ctk.CTkButton(
        thresh_card,
        text="📂 Open Output Folder",
        font=("Arial", 12, "bold"),
        height=36, width=180, corner_radius=10,
        fg_color="#3B82F6", hover_color="#2563EB",
        text_color="#FFFFFF",
        command=_open_out_folder
    )
    open_folder_btn.grid(row=2, column=5, sticky="e", padx=(0, 16), pady=(4, 14))

    progress_bar = ctk.CTkProgressBar(thresh_card, mode="indeterminate",
                                      height=8, corner_radius=4)
    progress_bar.grid(row=3, column=0, columnspan=6, sticky="ew", padx=16, pady=(0, 10))
    progress_bar.grid_remove()   # hidden until processing starts

    # ═══════════════════════════════════════════════════
    # CARD 3 — Summary Dashboard
    # ═══════════════════════════════════════════════════
    _TILE_COLORS = {
        "Matched":            "#C6EFCE",
        "Nominal Difference": "#FFEB9C",
        "Difference":         "#FFC0CB",
        "Not in Our Data":    "#ADD8E6",
        "Not in GSTR2B":      "#FFFF99",
        "Matched with AI":    "#D8B4FE",
    }

    sum_card = ctk.CTkFrame(main_frame, corner_radius=12)
    sum_card.grid(row=2, column=0, sticky="ew", **_CP)
    sum_card.columnconfigure((0, 1, 2), weight=1)

    ctk.CTkLabel(sum_card, text="Reconciliation Summary",
                 font=("Arial", 13, "bold"),
                 ).grid(row=0, column=0, columnspan=3, sticky="w", padx=16, pady=(12, 2))
    ctk.CTkFrame(sum_card, height=1, fg_color=("gray75", "gray35"),
                 ).grid(row=1, column=0, columnspan=3, sticky="ew", padx=16, pady=(0, 8))

    sum_labels: dict = {}
    for _si, _st in enumerate(_TILE_COLORS):
        _c = _si % 3
        _r = _si // 3 + 2
        _bg = _TILE_COLORS[_st]
        _px_l = 16 if _c == 0 else 6
        _px_r = 16 if _c == 2 else 6
        _tile = ctk.CTkFrame(sum_card, corner_radius=10, fg_color=_bg,
                             border_width=1, border_color=("#C0C0C0", "#505050"))
        _tile.grid(row=_r, column=_c, sticky="ew",
                   padx=(_px_l, _px_r), pady=4)
        _tile.columnconfigure(0, weight=1)
        ctk.CTkLabel(_tile, text=_st, font=("Arial", 10, "bold"),
                     text_color="#1F2937", fg_color=_bg, anchor="center",
                     ).grid(row=0, column=0, sticky="ew", padx=6, pady=(8, 1))
        _clbl = ctk.CTkLabel(_tile, text="—", font=("Arial", 22, "bold"),
                             text_color="#1F2937", fg_color=_bg, anchor="center")
        _clbl.grid(row=1, column=0, sticky="ew", padx=6, pady=(1, 8))
        sum_labels[_st] = _clbl
    ctk.CTkLabel(sum_card, text="", height=4,
                 ).grid(row=100, column=0, columnspan=3)

    # ═══════════════════════════════════════════════════
    # CARD 4 — Console Log  (expands to fill remaining height)
    # ═══════════════════════════════════════════════════
    log_card = ctk.CTkFrame(main_frame, corner_radius=12)
    log_card.grid(row=4, column=0, sticky="nsew", padx=20, pady=(10, 0))
    log_card.columnconfigure(0, weight=1)
    log_card.rowconfigure(2, weight=1)

    ctk.CTkLabel(log_card, text="Console Log",
                 font=("Arial", 13, "bold"),
                 ).grid(row=0, column=0, sticky="w", padx=16, pady=(12, 2))
    ctk.CTkFrame(log_card, height=1, fg_color=("gray75", "gray35"),
                 ).grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 6))

    log_box = ctk.CTkTextbox(log_card, font=("Courier", 11),
                              state="disabled", wrap="word",
                              corner_radius=6)
    log_box.grid(row=2, column=0, sticky="nsew", padx=16, pady=(0, 14))

    def append_log(msg: str):
        log_box.configure(state="normal")
        log_box.insert(tk.END, msg + "\n")
        log_box.see(tk.END)
        log_box.configure(state="disabled")
        root.update_idletasks()


    def run_reco():
        g = gstr_entry.get().strip()
        t = tally_entry.get().strip()

        if not g or not os.path.exists(g):
            messagebox.showerror("Error", "Please select a valid GSTR-2B file.")
            return
        if not t or not os.path.exists(t):
            messagebox.showerror("Error", "Please select a valid Tally / Our Data file.")
            return

        base_dir = os.path.join(os.getcwd(), "GST Downloaded", "GST Reco")
        if not os.path.exists(base_dir): os.makedirs(base_dir, exist_ok=True)
        
        filename = os.path.basename(g)
        name_only, ext = os.path.splitext(filename)
        o = os.path.join(base_dir, name_only + "_reco_report" + (ext if ext else ".xlsx"))
        try:
            matched_thr = float(matched_var.get())
            nominal_thr = float(nominal_var.get())
        except ValueError:
            messagebox.showerror("Error", "Threshold values must be valid numbers.")
            return
        if matched_thr >= nominal_thr:
            messagebox.showerror("Error",
                                 "'Matched' threshold must be less than 'Nominal Difference' threshold.")
            return

        for _lbl in sum_labels.values():
            _lbl.configure(text="—")

        def process():
            run_btn.configure(state="disabled", text="  Processing…")
            progress_bar.grid()
            progress_bar.start()
            log_box.configure(state="normal")
            log_box.delete("0.0", tk.END)
            log_box.configure(state="disabled")
            try:
                reconcile_and_write(g, t, o, progress_cb=append_log,
                                    matched_threshold=matched_thr,
                                    nominal_threshold=nominal_thr)
                try:
                    wb_out = openpyxl.load_workbook(o, data_only=True)
                    if "Summary" in wb_out.sheetnames:
                        ws_s = wb_out["Summary"]
                        for _row in ws_s.iter_rows(values_only=True):
                            if _row[0] and _row[1] is not None:
                                _stn = str(_row[0]).strip()
                                if _stn in sum_labels:
                                    sum_labels[_stn].configure(text=str(_row[1]))
                    wb_out.close()
                except Exception:
                    pass
                ans = messagebox.askyesno("Done", f"Reconciliation complete!\n\nSaved to:\n{o}\n\nWould you like to open the output folder?")
                if ans:
                    os.startfile(os.path.dirname(o))
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred:\n{e}")
                append_log(f"ERROR: {e}")
            finally:
                progress_bar.stop()
                progress_bar.grid_remove()
                run_btn.configure(state="normal", text="▶   Run Reconciliation")

        threading.Thread(target=process, daemon=True).start()

    run_btn.configure(command=run_reco)

    # ── Theme handle for GST Suite integration ────────────────────────────
    class _Handle:
        def set_theme(self, suite_mode: str):
            """Called by GST Suite when the user toggles dark/light."""
            m = "dark" if suite_mode == "Dark" else "light"
            ctk.set_appearance_mode(m)

    _handle = _Handle()
    root.mainloop()
    return _handle


class App:
    """Entry-point class used by GST Suite to embed this tool."""
    def __new__(cls):
        return launch_gui(embedded=True)


if __name__ == "__main__":
    launch_gui()