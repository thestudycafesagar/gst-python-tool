import os
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  HELPERS & METADATA EXTRACTOR
# ─────────────────────────────────────────────
FY_PATTERNS = [
    r"FY2023[-/]?24", r"(?<![A-Za-z\d])2023[-/]24", r"FY2324", r"FY202324", r"(?<![A-Za-z\d])23[-/]24", r"(?<![A-Za-z\d])2324[-/]",
    r"FY2024[-/]?25", r"(?<![A-Za-z\d])2024[-/]25", r"FY2425", r"FY202425", r"(?<![A-Za-z\d])24[-/]25", r"(?<![A-Za-z\d])2425[-/]",
    r"FY2025[-/]?26", r"(?<![A-Za-z\d])2025[-/]26", r"FY2526", r"FY202526", r"(?<![A-Za-z\d])25[-/]26", r"(?<![A-Za-z\d])2526[-/]",
]

def clean_invoice(inv_str: str) -> str:
    """Strip FY patterns and non-digit chars; return numeric-only tail (no leading zeros)."""
    s = str(inv_str).strip()
    for pat in FY_PATTERNS:
        s = re.sub(pat, "", s, flags=re.IGNORECASE)
    digits = re.sub(r"\D", "", s)
    return digits.lstrip("0") or digits

def make_gstinv(gstin: str, inv_str: str) -> str:
    return str(gstin).strip() + clean_invoice(inv_str)

def safe_float(val) -> float:
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def digit_tail(s: str) -> str:
    """Return only the digit characters from a string (already-cleaned invoice key)."""
    return re.sub(r"\D", "", str(s).strip())

def get_gstr_info(filepath: str, log_cb) -> dict:
    """Bulletproof metadata extractor that skips empty columns to find data."""
    info = {"gstin": "", "trade_name": "", "fy": "", "period": ""}
    try:
        # Load the workbook reading actual values
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        target_sheet = None
        for sheet in wb.sheetnames:
            if "readme" in sheet.lower().replace(" ", "").replace("_", ""):
                target_sheet = sheet
                break
        
        if target_sheet:
            ws = wb[target_sheet]
            # Scan a wide area: first 20 rows, first 10 columns
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True):
                for i, cell in enumerate(row):
                    if not cell:
                        continue
                        
                    label = str(cell).strip().lower()
                    
                    # Instead of just taking the immediate next column, 
                    # scan to the right and grab the FIRST non-blank value
                    val = ""
                    for j in range(i + 1, len(row)):
                        if row[j] and str(row[j]).strip():
                            val = str(row[j]).strip()
                            break
                            
                    # If we found a value, map it to the right dictionary key
                    if val:
                        if "gstin" in label and not info["gstin"]:
                            info["gstin"] = val
                        elif ("trade name" in label or "legal name" in label) and not info["trade_name"]:
                            info["trade_name"] = val
                        elif ("financial year" in label or "f.y" in label) and not info["fy"]:
                            info["fy"] = val
                        elif ("period" in label) and not info["period"]:
                            info["period"] = val
                            
        wb.close()
    except Exception as e:
        log_cb(f"Warning: Could not extract Read me metadata: {e}")
        
    return info
    """Bulletproof metadata extractor that scans a grid to find headers."""
    info = {"gstin": "", "trade_name": "", "fy": "", "period": ""}
    try:
        # Removed read_only=True to prevent blank reads on portal-generated files
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        target_sheet = None
        for sheet in wb.sheetnames:
            if "readme" in sheet.lower().replace(" ", "").replace("_", ""):
                target_sheet = sheet
                break
        
        if target_sheet:
            ws = wb[target_sheet]
            # Scan a wide area: first 20 rows, first 10 columns
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True):
                for i, cell in enumerate(row):
                    if not cell:
                        continue
                        
                    label = str(cell).strip().lower()
                    
                    # If we find a label, grab the value in the immediate next column
                    if i + 1 < len(row):
                        val = str(row[i+1] or "").strip()
                        
                        if label == "gstin" and not info["gstin"]:
                            info["gstin"] = val
                        elif label in ("trade name", "legal name") and not info["trade_name"]:
                            info["trade_name"] = val
                        elif label in ("financial year", "f.y.") and not info["fy"]:
                            info["fy"] = val
                        elif label in ("return period", "period") and not info["period"]:
                            info["period"] = val
        wb.close()
    except Exception as e:
        log_cb(f"Warning: Could not extract Read me metadata: {e}")
        
    return info

# ─────────────────────────────────────────────
#  PARSE GSTR-2B
# ─────────────────────────────────────────────
def parse_gstr2b(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets = wb.sheetnames
    if "B2B" not in sheets:
        wb.close()
        raise ValueError("GSTR-2B file has no 'B2B' sheet.")

    ws = wb["B2B"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[6:]:
        gstin = row[0]
        if not gstin or str(gstin).strip() == "":
            continue
        inv_no   = str(row[2] or "").strip()
        inv_date = row[4]
        inv_val  = safe_float(row[5])
        taxable  = safe_float(row[8])
        igst     = safe_float(row[9])
        cgst     = safe_float(row[10])
        sgst     = safe_float(row[11])
        cess     = safe_float(row[12] if len(row) > 12 else None)
        new_inv  = clean_invoice(inv_no)
        gstinv   = make_gstinv(gstin, inv_no)
        records.append({
            "gstin": str(gstin).strip(),
            "name":  str(row[1] or "").strip(),
            "original_inv": inv_no,
            "new_inv":  new_inv,
            "gstinv":   gstinv,
            "inv_date": inv_date,
            "taxable":  taxable,
            "igst":     igst,
            "cgst":     cgst,
            "sgst":     sgst,
            "cess":     cess,
            "total":    inv_val if inv_val else taxable + igst + cgst + sgst + cess,
            "source":   "GSTR2B",
        })
    wb.close()
    return records

# ─────────────────────────────────────────────
#  PARSE TALLY / OUR DATA
# ─────────────────────────────────────────────
def parse_tally(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for i, row in enumerate(all_rows):
        row_text = " ".join(str(c).upper() for c in row if c)
        if "PARTY GSTIN" in row_text and "DOC NO" in row_text.replace(".", "").replace(" ", ""):
            header_row_idx = i
            break
            
    if header_row_idx is None:
        for i, row in enumerate(all_rows):
            if any(c and "GSTIN" in str(c).upper() for c in row):
                header_row_idx = i
                break

    if header_row_idx is None:
        wb.close()
        raise ValueError("Could not find the Tally header row.")

    header = all_rows[header_row_idx]
    col_gstin   = next((i for i, c in enumerate(header) if c and "GSTIN" in str(c).upper()), 2)
    col_doc_no  = next((i for i, c in enumerate(header) if c and "DOC" in str(c).upper() and "NO" in str(c).upper()), 6)
    col_taxable = next((i for i, c in enumerate(header) if c and "TAXABLE" in str(c).upper()), 11)
    
    col_igst  = col_taxable + 1
    col_cgst  = col_taxable + 2
    col_sgst  = col_taxable + 3
    col_cess  = col_taxable + 4
    col_inv   = col_taxable + 6

    records = []
    for row in all_rows[header_row_idx + 2:]:
        if len(row) <= col_gstin:
            continue
        gstin = row[col_gstin]
        if not gstin or str(gstin).strip() in ("", "nan", "None"):
            continue
        gstin_str = str(gstin).strip()
        if not re.match(r"^\d{2}[A-Z]", gstin_str):
            continue

        inv_no  = str(row[col_doc_no] if len(row) > col_doc_no else "").strip()
        if not inv_no or inv_no in ("None", "nan"):
            inv_no = str(row[5] if len(row) > 5 else "").strip()

        inv_date = row[0]
        taxable  = safe_float(row[col_taxable]  if len(row) > col_taxable  else None)
        igst     = safe_float(row[col_igst]     if len(row) > col_igst     else None)
        cgst     = safe_float(row[col_cgst]     if len(row) > col_cgst     else None)
        sgst     = safe_float(row[col_sgst]     if len(row) > col_sgst     else None)
        cess     = safe_float(row[col_cess]     if len(row) > col_cess     else None)
        inv_val  = safe_float(row[col_inv]      if len(row) > col_inv      else None)
        new_inv  = clean_invoice(inv_no)
        gstinv   = make_gstinv(gstin_str, inv_no)

        records.append({
            "gstin": gstin_str,
            "name":  str(row[1] or "").strip(),
            "original_inv": inv_no,
            "new_inv":  new_inv,
            "gstinv":   gstinv,
            "inv_date": inv_date,
            "taxable":  taxable,
            "igst":     igst,
            "cgst":     cgst,
            "sgst":     sgst,
            "cess":     cess,
            "total":    inv_val if inv_val else taxable + igst + cgst + sgst + cess,
            "source":   "OUR DATA",
        })
    wb.close()
    return records

# ─────────────────────────────────────────────
#  STYLES & EXCEL WRITER
# ─────────────────────────────────────────────
FILL_MATCH    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_NOMINAL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_DIFF     = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
FILL_NO_OUR   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
FILL_NO_GSTR  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
FILL_AI_MATCH = PatternFill(start_color="D8B4FE", end_color="D8B4FE", fill_type="solid")

STATUS_FILL = {
    "Matched":                 FILL_MATCH,
    "Nominal Difference":      FILL_NOMINAL,
    "Difference":              FILL_DIFF,
    "Not in Our Data":         FILL_NO_OUR,
    "Not in GSTR2B":           FILL_NO_GSTR,
    "Matched with AI":         FILL_AI_MATCH,
    "Nominal Difference (AI)": FILL_NOMINAL,
    "Difference (AI)":         FILL_DIFF,
}

# ─────────────────────────────────────────────
#  MAIN RECONCILIATION
# ─────────────────────────────────────────────
def reconcile_and_write(gstr2b_file: str, tally_file: str, output_file: str, progress_cb=None,
                        matched_threshold: float = 2.0, nominal_threshold: float = 20.0):
    def log(msg):
        if progress_cb: progress_cb(msg)
        else: print(msg)

    # 0. Extract Dynamic Metadata
    log("Extracting metadata from 'Read me' sheet...")
    meta = get_gstr_info(gstr2b_file, log)

    log("Parsing GSTR-2B…")
    gstr_recs = parse_gstr2b(gstr2b_file)
    log(f"  → {len(gstr_recs)} invoices in GSTR-2B")

    log("Parsing Tally / Our Data…")
    tally_recs = parse_tally(tally_file)
    log(f"  → {len(tally_recs)} invoices in Tally file")

    all_recs = gstr_recs + tally_recs

    # 1. Calculate Status for each GSTINV
    FIELDS = ("taxable", "igst", "cgst", "sgst", "cess")
    agg: dict = {}
    for r in all_recs:
        key = r["gstinv"]
        src = r["source"]
        agg.setdefault(key, {})
        if src not in agg[key]:
            agg[key][src] = {f: 0.0 for f in FIELDS}
        for f in FIELDS:
            agg[key][src][f] += r.get(f, 0.0)

    def get_status(gstinv: str) -> str:
        d = agg.get(gstinv, {})
        has_g = "GSTR2B"   in d
        has_o = "OUR DATA" in d
        if has_g and has_o:
            g, o = d["GSTR2B"], d["OUR DATA"]
            max_diff = max(abs(g[f] - o[f]) for f in FIELDS)
            if max_diff <= matched_threshold:   return "Matched"
            elif max_diff <= nominal_threshold: return "Nominal Difference"
            else:                               return "Difference"
        elif has_g:             return "Not in Our Data"
        elif has_o:             return "Not in GSTR2B"
        return ""

    # 1b. AI Second-Pass: suffix digit matching for unmatched records
    log("Running AI suffix match for unmatched invoices...")

    # Build lookup: gstinv -> record (first seen), for unmatched only
    seen_gstinv: dict = {}
    for r in all_recs:
        k = r["gstinv"]
        if k not in seen_gstinv:
            seen_gstinv[k] = r

    gstr_unmatched: dict = {}   # gstin -> [(gstinv, digit_string)]
    tally_unmatched: dict = {}  # gstin -> [(gstinv, digit_string)]

    for k, r in seen_gstinv.items():
        st = get_status(k)
        if st == "Not in Our Data":
            gstr_unmatched.setdefault(r["gstin"], []).append((k, r["new_inv"]))
        elif st == "Not in GSTR2B":
            tally_unmatched.setdefault(r["gstin"], []).append((k, r["new_inv"]))

    ai_remap: dict = {}        # tally gstinv -> gstr gstinv
    ai_matched_keys: set = set()  # gstr gstinv keys that got AI-matched

    for gstin, g_list in gstr_unmatched.items():
        t_list = tally_unmatched.get(gstin, [])
        if not t_list:
            continue
        used_t: set = set()
        for g_key, g_digits in g_list:
            if not g_digits:
                continue
            for t_key, t_digits in t_list:
                if t_key in used_t or not t_digits:
                    continue
                if min(len(g_digits), len(t_digits)) >= 4 and (g_digits.endswith(t_digits) or t_digits.endswith(g_digits)):
                    ai_remap[t_key] = g_key
                    ai_matched_keys.add(g_key)
                    used_t.add(t_key)
                    break

    # Merge OUR DATA amounts from tally key into gstr key in agg
    for t_key, g_key in ai_remap.items():
        if t_key in agg and "OUR DATA" in agg[t_key]:
            agg[g_key]["OUR DATA"] = agg[t_key].pop("OUR DATA")

    log(f"  → {len(ai_matched_keys)} invoice(s) AI-matched via suffix digit logic")

    # 2. Group Data into Side-By-Side Format
    log("Grouping records for side-by-side view...")
    grouped_data = {}
    
    for r in all_recs:
        k = r["gstinv"]
        # Re-key AI-matched tally records to their GSTR-2B partner key
        if k in ai_remap:
            k = ai_remap[k]
        if k not in grouped_data:
            grouped_data[k] = {"books": {}, "gstr": {}}

        src_key = "gstr" if r["source"] == "GSTR2B" else "books"
        target = grouped_data[k][src_key]

        if not target:
            target.update(r)
        else:
            target["taxable"] += r["taxable"]
            target["igst"]    += r["igst"]
            target["cgst"]    += r["cgst"]
            target["sgst"]    += r["sgst"]
            target["cess"]    += r.get("cess", 0.0)
            target["total"]   += r["total"]

    # 3. Write Output Workbook
    log("Building output workbook…")
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "All Invoice Reco Status"
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Top Headers - DYNAMICALLY POPULATED
    ws_main["C1"], ws_main["D1"] = "GSTIN:", meta["gstin"]
    ws_main["C2"], ws_main["D2"] = "Trade Name:", meta["trade_name"]
    ws_main["C3"], ws_main["D3"] = "F.Y.:", meta["fy"]
    ws_main["C4"], ws_main["D4"] = "Period:", meta["period"]

    for row in range(1, 5):
        ws_main[f"C{row}"].font = bold_font
        ws_main[f"D{row}"].font = bold_font

    ws_main["A5"] = "GSTR-2B Reconciliation :: All Invoices Reconciliation Report"
    ws_main["A5"].font = Font(bold=True, size=14)
    
    ws_main["D6"] = "As Per Books of Accounts [A]"
    ws_main.merge_cells("D6:J6")
    ws_main["K6"] = "As Per GSTR-2B"
    ws_main.merge_cells("K6:Q6")
    ws_main["R6"] = "Difference"
    ws_main.merge_cells("R6:V6")
    
    for col in ["D6", "K6", "R6"]:
        ws_main[col].font = bold_font
        ws_main[col].alignment = center_align
        ws_main[col].fill = header_fill
        ws_main[col].border = thin_border

    headers = [
        "S.No.", "GSTIN", "Supplier Name",
        "Invoice No.", "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", "Cess",
        "Invoice No.", "Invoice Date", "Taxable Value", "IGST", "CGST", "SGST", "Cess",
        "Taxable Value", "IGST", "CGST", "SGST", "Cess",
        "Status"
    ]
    ws_main.append(headers) 
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws_main.cell(row=7, column=col_idx)
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
        ws_main.column_dimensions[get_column_letter(col_idx)].width = 14

    ws_main.column_dimensions["C"].width = 25 
    ws_main.freeze_panes = "A8"

    log("Writing grouped data rows...")
    for index, (gstinv, data) in enumerate(grouped_data.items(), start=1):
        b = data["books"]
        g = data["gstr"]
        
        if gstinv in ai_matched_keys:
            d = agg.get(gstinv, {})
            if "GSTR2B" in d and "OUR DATA" in d:
                g_amt, o_amt = d["GSTR2B"], d["OUR DATA"]
                max_diff = max(abs(g_amt[f] - o_amt[f]) for f in FIELDS)
                if max_diff <= matched_threshold:   status = "Matched with AI"
                elif max_diff <= nominal_threshold: status = "Nominal Difference (AI)"
                else:                               status = "Difference (AI)"
            else:
                status = "Matched with AI"
        else:
            status = get_status(gstinv)
        gstin_val = b.get("gstin") if b else g.get("gstin", "")
        name_val = b.get("name") if b else g.get("name", "")
        
        b_tax  = b.get("taxable", 0)
        b_igst = b.get("igst", 0)
        b_cgst = b.get("cgst", 0)
        b_sgst = b.get("sgst", 0)
        b_cess = b.get("cess", 0)

        g_tax  = g.get("taxable", 0)
        g_igst = g.get("igst", 0)
        g_cgst = g.get("cgst", 0)
        g_sgst = g.get("sgst", 0)
        g_cess = g.get("cess", 0)

        row_data = [
            index, gstin_val, name_val,

            b.get("original_inv", "") if b else "",
            b.get("inv_date", "") if b else "",
            b_tax, b_igst, b_cgst, b_sgst, b_cess,

            g.get("original_inv", "") if g else "",
            g.get("inv_date", "") if g else "",
            g_tax, g_igst, g_cgst, g_sgst, g_cess,

            round(b_tax  - g_tax,  2),
            round(b_igst - g_igst, 2),
            round(b_cgst - g_cgst, 2),
            round(b_sgst - g_sgst, 2),
            round(b_cess - g_cess, 2),

            status
        ]
        
        ws_main.append(row_data)
        
        current_row = ws_main.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_main.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if isinstance(cell.value, (int, float)) and col_idx not in (1, 4, 11):
                cell.number_format = "#,##0.00"
            if col_idx in (5, 12) and cell.value:
                cell.number_format = "DD-MM-YY"

        status_cell = ws_main.cell(row=current_row, column=23)
        if status in STATUS_FILL:
            status_cell.fill = STATUS_FILL[status]


    # 4. Sheet: Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 14

    seen: dict = {}
    for r in all_recs:
        k = r["gstinv"]
        rk = ai_remap.get(k, k)   # use remapped key if AI-matched
        if rk not in seen:
            seen[rk] = get_status(rk)

    statuses = ["Matched", "Nominal Difference", "Difference", "Not in Our Data", "Not in GSTR2B"]
    counts = {s: sum(1 for v in seen.values() if v == s) for s in statuses}
    ai_count = len(ai_matched_keys)

    rows_data = [
        ("GST Reconciliation Summary", None),
        (None, None),
        ("Status", "Invoice Count"),
    ] + [(s, counts[s]) for s in statuses] + [
        ("Matched with AI", ai_count),
        (None, None),
        ("GSTR-2B Records", len(gstr_recs)),
        ("Our Data Records", len(tally_recs)),
        ("Total Combined", len(all_recs)),
    ]

    for ri, (lbl, val) in enumerate(rows_data, 1):
        cl = ws_sum.cell(row=ri, column=1, value=lbl)
        cv = ws_sum.cell(row=ri, column=2, value=val)
        bold = ri in (1, 3, len(rows_data) - 2, len(rows_data) - 1, len(rows_data))
        cl.font = Font(name="Arial", size=11 if ri == 1 else 10, bold=bold or ri == 1)
        cv.font = Font(name="Arial", size=10, bold=bold)
        fill = STATUS_FILL.get(str(lbl))
        if fill:
            cl.fill, cv.fill = fill, fill
        elif str(lbl) == "Matched with AI":
            cl.fill, cv.fill = FILL_AI_MATCH, FILL_AI_MATCH

    log(f"Saving → {output_file}")
    wb.save(output_file)
    log("Reconciliation complete ✓")

# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────
def launch_gui():
    # ── Colour palettes ──────────────────────────────────────────────────
    THEMES = {
        "light": {
            "root_bg":    "#F0F4F8",
            "frame_bg":   "#FFFFFF",
            "lf_bg":      "#FFFFFF",
            "lf_fg":      "#374151",
            "label_bg":   "#FFFFFF",
            "label_fg":   "#1F2937",
            "entry_bg":   "#FFFFFF",
            "entry_fg":   "#1F2937",
            "text_bg":    "#F5F5F5",
            "text_fg":    "#374151",
            "btn_bg":     "#E5E7EB",
            "btn_fg":     "#1F2937",
            "btn_act":    "#D1D5DB",
            "header_bg":  "#1F2937",
            "header_fg":  "#FFFFFF",
            "tab_bg":     "#D1D5DB",
            "tab_fg":     "#1F2937",
            "sel_tab_bg": "#FFFFFF",
            "toggle_txt": "🌙  Dark",
        },
        "dark": {
            "root_bg":    "#111827",
            "frame_bg":   "#1F2937",
            "lf_bg":      "#1F2937",
            "lf_fg":      "#9CA3AF",
            "label_bg":   "#1F2937",
            "label_fg":   "#F9FAFB",
            "entry_bg":   "#374151",
            "entry_fg":   "#F9FAFB",
            "text_bg":    "#111827",
            "text_fg":    "#D1FAE5",
            "btn_bg":     "#374151",
            "btn_fg":     "#F9FAFB",
            "btn_act":    "#4B5563",
            "header_bg":  "#0F172A",
            "header_fg":  "#F9FAFB",
            "tab_bg":     "#374151",
            "tab_fg":     "#D1D5DB",
            "sel_tab_bg": "#1F2937",
            "toggle_txt": "☀  Light",
        },
    }

    mode     = ["light"]   # mutable so nested functions can write to it
    skip_ids = set()       # widget ids that must keep their own colors

    def apply_theme(m):
        T = THEMES[m]
        root.config(bg=T["root_bg"])
        header_bar.config(bg=T["header_bg"])
        header_lbl.config(bg=T["header_bg"], fg=T["header_fg"])
        style.configure("TNotebook",     background=T["root_bg"],    borderwidth=0)
        style.configure("TNotebook.Tab", background=T["tab_bg"],     foreground=T["tab_fg"],
                        padding=[10, 4])
        style.map("TNotebook.Tab",
                  background=[("selected", T["sel_tab_bg"])],
                  foreground=[("selected", T["label_fg"])])
        style.configure("TFrame", background=T["frame_bg"])

        def walk(widget):
            if id(widget) in skip_ids:
                return
            cls = widget.winfo_class()
            try:
                if cls == "Frame":
                    widget.config(bg=T["frame_bg"])
                elif cls == "Labelframe":
                    widget.config(bg=T["lf_bg"], fg=T["lf_fg"])
                elif cls == "Label":
                    widget.config(bg=T["label_bg"], fg=T["label_fg"])
                elif cls == "Entry":
                    widget.config(bg=T["entry_bg"], fg=T["entry_fg"],
                                  insertbackground=T["entry_fg"])
                elif cls == "Text":
                    widget.config(bg=T["text_bg"], fg=T["text_fg"],
                                  insertbackground=T["text_fg"])
                elif cls == "Button":
                    widget.config(bg=T["btn_bg"], fg=T["btn_fg"],
                                  activebackground=T["btn_act"],
                                  activeforeground=T["btn_fg"])
            except tk.TclError:
                pass
            for child in widget.winfo_children():
                walk(child)

        for child in notebook.winfo_children():
            walk(child)

    # ── Root window ───────────────────────────────────────────────────────
    root = tk.Tk()
    root.title("GST Reconciliation Tool")
    root.state("zoomed")
    root.resizable(True, True)

    style = ttk.Style(root)
    style.theme_use("clam")

    # ── Header bar ────────────────────────────────────────────────────────
    header_bar = tk.Frame(root, height=54)
    header_bar.pack(fill="x", side="top")
    header_bar.pack_propagate(False)

    header_lbl = tk.Label(header_bar, text="GST Reconciliation Tool",
                          font=("Arial", 16, "bold"))
    header_lbl.pack(side="left", padx=18)

    # ── Notebook ──────────────────────────────────────────────────────────
    notebook = ttk.Notebook(root)
    notebook.pack(pady=6, expand=True, fill="both")

    # ═══════════════════════════════════════════════════
    # TAB 1 : GST Reconciliation (2 Files)
    # ═══════════════════════════════════════════════════
    tab = ttk.Frame(notebook)
    notebook.add(tab, text="GST Reconciliation (2 Files)")

    frame = tk.Frame(tab, padx=30, pady=22)
    frame.pack(fill="both", expand=True)
    frame.columnconfigure(1, weight=1)
    frame.rowconfigure(3, weight=1)

    lbl_kw = {"font": ("Arial", 13, "bold"), "anchor": "w"}

    def file_row(row_idx, label, entry_var):
        tk.Label(frame, text=label, **lbl_kw).grid(row=row_idx, column=0, sticky="w", pady=8)
        entry = tk.Entry(frame, font=("Arial", 12))
        entry.grid(row=row_idx, column=1, padx=10, sticky="ew")
        return entry

    gstr_entry  = file_row(0, "GSTR-2B File (.xlsx):",         None)
    tally_entry = file_row(1, "Tally / Our Data File (.xlsx):", None)

    thresh_frame = tk.LabelFrame(frame, text="Reconciliation Thresholds (₹)",
                                 font=("Arial", 11, "bold"), padx=14, pady=10)
    thresh_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(14, 0))

    tk.Label(thresh_frame, text="Matched up to (₹):", font=("Arial", 12)).grid(
        row=0, column=0, sticky="w", padx=(0, 10))
    matched_var = tk.StringVar(value="2")
    tk.Entry(thresh_frame, textvariable=matched_var, width=12, font=("Arial", 12)).grid(row=0, column=1, sticky="w")

    tk.Label(thresh_frame, text="  Nominal Difference up to (₹):", font=("Arial", 12)).grid(
        row=0, column=2, sticky="w", padx=(24, 10))
    nominal_var = tk.StringVar(value="20")
    tk.Entry(thresh_frame, textvariable=nominal_var, width=12, font=("Arial", 12)).grid(row=0, column=3, sticky="w")

    def make_browse(entry, title):
        ft = [("Excel", "*.xlsx *.xlsm"), ("All", "*.*")]
        def cb():
            f = filedialog.askopenfilename(title=title, filetypes=ft)
            if f:
                entry.delete(0, tk.END)
                entry.insert(0, f)
        return cb

    tk.Button(frame, text="Browse", font=("Arial", 11),
              command=make_browse(gstr_entry,  "Select GSTR-2B File"),          width=10).grid(row=0, column=2, padx=(6, 0))
    tk.Button(frame, text="Browse", font=("Arial", 11),
              command=make_browse(tally_entry, "Select Tally / Our Data File"), width=10).grid(row=1, column=2, padx=(6, 0))

    tk.Label(frame, text="Log:", **lbl_kw).grid(row=3, column=0, sticky="nw", pady=(18, 0))
    log_box = tk.Text(frame, state=tk.DISABLED, font=("Courier", 12))
    log_box.grid(row=3, column=1, columnspan=2, pady=(18, 0), sticky="nsew")

    def append_log(msg):
        log_box.config(state=tk.NORMAL)
        log_box.insert(tk.END, msg + "\n")
        log_box.see(tk.END)
        log_box.config(state=tk.DISABLED)
        root.update_idletasks()

    run_btn = tk.Button(frame, text="▶  Run Reconciliation", font=("Arial", 15, "bold"),
                        bg="#2E7D32", fg="white", activebackground="#1B5E20",
                        activeforeground="white", width=28, pady=12)
    run_btn.grid(row=4, column=0, columnspan=3, pady=22)
    skip_ids.add(id(run_btn))

    def run_reco():
        g = gstr_entry.get().strip()
        t = tally_entry.get().strip()
        if not g or not os.path.exists(g):
            messagebox.showerror("Error", "Please select a valid GSTR-2B file.")
            return
        if not t or not os.path.exists(t):
            messagebox.showerror("Error", "Please select a valid Tally / Our Data file.")
            return
        try:
            matched_thr = float(matched_var.get())
            nominal_thr = float(nominal_var.get())
        except ValueError:
            messagebox.showerror("Error", "Threshold values must be valid numbers.")
            return
        if matched_thr >= nominal_thr:
            messagebox.showerror("Error", "'Matched' threshold must be less than 'Nominal Difference' threshold.")
            return
        o = os.path.join(os.path.dirname(g), "GST_Reconciled_Output.xlsx")
        def process():
            run_btn.config(state=tk.DISABLED, text="Processing…")
            log_box.config(state=tk.NORMAL)
            log_box.delete("1.0", tk.END)
            log_box.config(state=tk.DISABLED)
            try:
                reconcile_and_write(g, t, o, progress_cb=append_log,
                                    matched_threshold=matched_thr, nominal_threshold=nominal_thr)
                messagebox.showinfo("Done", f"Reconciliation complete!\n\nSaved to:\n{o}")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred:\n{e}")
                append_log(f"ERROR: {e}")
            finally:
                run_btn.config(state=tk.NORMAL, text="▶  Run Reconciliation")
        threading.Thread(target=process, daemon=True).start()

    run_btn.config(command=run_reco)

    # ═══════════════════════════════════════════════════
    # TAB 2 : Reconciliation (custom output + in-app summary)
    # ═══════════════════════════════════════════════════
    tab2 = ttk.Frame(notebook)
    notebook.add(tab2, text="Reconciliation")
    notebook.hide(tab2)

    frame2 = tk.Frame(tab2, padx=30, pady=22)
    frame2.pack(fill="both", expand=True)
    frame2.columnconfigure(1, weight=1)
    frame2.rowconfigure(4, weight=1)

    lbl_kw2 = {"font": ("Arial", 13, "bold"), "anchor": "w"}

    def file_row2(row_idx, label):
        tk.Label(frame2, text=label, **lbl_kw2).grid(row=row_idx, column=0, sticky="w", pady=8)
        entry = tk.Entry(frame2, font=("Arial", 12))
        entry.grid(row=row_idx, column=1, padx=10, sticky="ew")
        return entry

    gstr_entry2  = file_row2(0, "GSTR-2B File (.xlsx):")
    tally_entry2 = file_row2(1, "Tally / Our Data File (.xlsx):")
    out_entry2   = file_row2(2, "Output File (.xlsx):")

    thresh2 = tk.LabelFrame(frame2, text="Thresholds (₹)", font=("Arial", 11, "bold"), padx=14, pady=10)
    thresh2.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(14, 0))
    tk.Label(thresh2, text="Matched up to (₹):", font=("Arial", 12)).grid(row=0, column=0, sticky="w")
    matched_var2 = tk.StringVar(value="2")
    tk.Entry(thresh2, textvariable=matched_var2, width=12, font=("Arial", 12)).grid(row=0, column=1, sticky="w")
    tk.Label(thresh2, text="  Nominal Difference up to (₹):", font=("Arial", 12)).grid(
        row=0, column=2, sticky="w", padx=(24, 10))
    nominal_var2 = tk.StringVar(value="20")
    tk.Entry(thresh2, textvariable=nominal_var2, width=12, font=("Arial", 12)).grid(row=0, column=3, sticky="w")

    def make_browse2(entry, title):
        ft = [("Excel", "*.xlsx *.xlsm"), ("All", "*.*")]
        def cb():
            f = filedialog.askopenfilename(title=title, filetypes=ft)
            if f:
                entry.delete(0, tk.END)
                entry.insert(0, f)
        return cb

    def browse_output2():
        f = filedialog.asksaveasfilename(
            title="Save Reconciliation Output As",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
            initialfile="GST_Reconciled_Output.xlsx",
        )
        if f:
            out_entry2.delete(0, tk.END)
            out_entry2.insert(0, f)

    tk.Button(frame2, text="Browse", font=("Arial", 11),
              command=make_browse2(gstr_entry2,  "Select GSTR-2B File"),          width=10).grid(row=0, column=2, padx=(6, 0))
    tk.Button(frame2, text="Browse", font=("Arial", 11),
              command=make_browse2(tally_entry2, "Select Tally / Our Data File"), width=10).grid(row=1, column=2, padx=(6, 0))
    tk.Button(frame2, text="Browse", font=("Arial", 11),
              command=browse_output2,                                              width=10).grid(row=2, column=2, padx=(6, 0))

    tk.Label(frame2, text="Log:", **lbl_kw2).grid(row=4, column=0, sticky="nw", pady=(14, 0))
    log_box2 = tk.Text(frame2, state=tk.DISABLED, font=("Courier", 12))
    log_box2.grid(row=4, column=1, columnspan=2, pady=(14, 0), sticky="nsew")

    def append_log2(msg):
        log_box2.config(state=tk.NORMAL)
        log_box2.insert(tk.END, msg + "\n")
        log_box2.see(tk.END)
        log_box2.config(state=tk.DISABLED)
        root.update_idletasks()

    # Summary tiles — these keep their status colors in both modes
    sum_frame = tk.LabelFrame(frame2, text="Reconciliation Summary",
                              font=("Arial", 11, "bold"), padx=12, pady=10)
    sum_frame.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(14, 0))

    STATUS_COLORS = {
        "Matched":            "#C6EFCE",
        "Nominal Difference": "#FFEB9C",
        "Difference":         "#FFC0CB",
        "Not in Our Data":    "#ADD8E6",
        "Not in GSTR2B":      "#FFFF99",
        "Matched with AI":    "#D8B4FE",
    }
    sum_labels = {}
    status_list = ["Matched", "Nominal Difference", "Difference",
                   "Not in Our Data", "Not in GSTR2B", "Matched with AI"]
    for si, st in enumerate(status_list):
        col   = si % 3
        row_i = si // 3
        bg    = STATUS_COLORS[st]
        tf = tk.Frame(sum_frame, bg=bg, bd=1, relief="solid")
        tf.grid(row=row_i, column=col, padx=8, pady=6, sticky="ew")
        nl = tk.Label(tf, text=st,  bg=bg, font=("Arial", 11, "bold"), width=22)
        nl.pack(pady=(6, 2))
        cl = tk.Label(tf, text="—", bg=bg, font=("Arial", 14, "bold"))
        cl.pack(pady=(0, 6))
        sum_labels[st] = cl
        # Protect all three from the theme walker
        skip_ids.add(id(tf))
        skip_ids.add(id(nl))
        skip_ids.add(id(cl))
    for ci in range(3):
        sum_frame.columnconfigure(ci, weight=1)

    run_btn2 = tk.Button(frame2, text="▶  Run Reconciliation", font=("Arial", 15, "bold"),
                         bg="#1565C0", fg="white", activebackground="#0D47A1",
                         activeforeground="white", width=28, pady=12)
    run_btn2.grid(row=6, column=0, columnspan=3, pady=22)
    skip_ids.add(id(run_btn2))

    def run_reco2():
        g = gstr_entry2.get().strip()
        t = tally_entry2.get().strip()
        o = out_entry2.get().strip()
        if not g or not os.path.exists(g):
            messagebox.showerror("Error", "Please select a valid GSTR-2B file.")
            return
        if not t or not os.path.exists(t):
            messagebox.showerror("Error", "Please select a valid Tally / Our Data file.")
            return
        if not o:
            o = os.path.join(os.path.dirname(g), "GST_Reconciled_Output.xlsx")
            out_entry2.insert(0, o)
        try:
            matched_thr = float(matched_var2.get())
            nominal_thr = float(nominal_var2.get())
        except ValueError:
            messagebox.showerror("Error", "Threshold values must be valid numbers.")
            return
        if matched_thr >= nominal_thr:
            messagebox.showerror("Error", "'Matched' threshold must be less than 'Nominal Difference' threshold.")
            return
        for lbl in sum_labels.values():
            lbl.config(text="—")
        def process():
            run_btn2.config(state=tk.DISABLED, text="Processing…")
            log_box2.config(state=tk.NORMAL)
            log_box2.delete("1.0", tk.END)
            log_box2.config(state=tk.DISABLED)
            try:
                gstr_recs2  = parse_gstr2b(g)
                tally_recs2 = parse_tally(t)
                append_log2(f"GSTR-2B: {len(gstr_recs2)} invoices | Books: {len(tally_recs2)} invoices")
                reconcile_and_write(g, t, o, progress_cb=append_log2,
                                    matched_threshold=matched_thr, nominal_threshold=nominal_thr)
                try:
                    wb_out = openpyxl.load_workbook(o, data_only=True)
                    if "Summary" in wb_out.sheetnames:
                        ws_s = wb_out["Summary"]
                        for row in ws_s.iter_rows(values_only=True):
                            if row[0] and row[1] is not None:
                                st_name = str(row[0]).strip()
                                if st_name in sum_labels:
                                    sum_labels[st_name].config(text=str(row[1]))
                    wb_out.close()
                except Exception:
                    pass
                messagebox.showinfo("Done", f"Reconciliation complete!\n\nSaved to:\n{o}")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred:\n{e}")
                append_log2(f"ERROR: {e}")
            finally:
                run_btn2.config(state=tk.NORMAL, text="▶  Run Reconciliation")
        threading.Thread(target=process, daemon=True).start()

    run_btn2.config(command=run_reco2)

    # ── Apply initial light theme & start ─────────────────────────────────
    apply_theme("light")

    # ── Theme handle for GST Suite integration ────────────────────────────
    class _Handle:
        def set_theme(self, suite_mode: str):
            """Called by GST Suite when the user toggles dark/light."""
            m = "dark" if suite_mode == "Dark" else "light"
            mode[0] = m
            try:
                apply_theme(m)
            except Exception:
                pass

    _handle = _Handle()
    root.mainloop()   # no-op when embedded via _EmbeddedTkFrame
    return _handle


class App:
    """Entry-point class used by GST Suite to embed this tool."""
    def __new__(cls):
        return launch_gui()


if __name__ == "__main__":
    launch_gui()