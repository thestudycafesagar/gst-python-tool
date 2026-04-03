"""
compile_excel_data.py  —  Excel Folder Compiler (GUI)
------------------------------------------------------
Merges GSTR-2B monthly files into one Compiled_Data.xlsx.
Preserves the exact original layout: header rows, merged cells,
column widths and formatting — only data rows are appended per month.

Non-GST Excel files are also supported (generic flat merge).
"""

import copy
import re
import threading
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook


# ---------------------------------------------------------------------------
# GST 2B constants
# ---------------------------------------------------------------------------

# sheet_name → first data row (1-indexed, Excel row number)
# Rows before this are the header block that is copied once and preserved.
GST_SECTIONS = {
    "B2B":       7,
    "B2B-CDNR":  7,
    "ECO":       7,
    "ISD":       7,
    "IMPG":      7,
    "IMPGSEZ":   7,
    "B2BA":      8,   # amended sheets have one extra header row
    "B2B-CDNRA": 8,
    "ECOA":      8,
    "ISDA":      8,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def is_gst_2b_file(file: Path) -> bool:
    """Return True if the first cell contains the GST 2B title."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        wb.close()
        return first is not None and "Goods and Services Tax" in str(first)
    except Exception:
        return False


def _sort_key_by_date(file: Path) -> tuple:
    """Sort GSTR-2B files chronologically using MMYYYY_ filename prefix."""
    m = re.match(r"^(\d{2})(\d{4})_", file.name)
    if m:
        return (int(m.group(2)), int(m.group(1)))   # (year, month)
    return (9999, 99)


def _copy_cell_style(src, dst):
    """Copy font, fill, border, alignment and number format from src to dst cell."""
    if src.has_style:
        dst.font      = copy.copy(src.font)
        dst.fill      = copy.copy(src.fill)
        dst.border    = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format


# ---------------------------------------------------------------------------
# Core GST compile  (openpyxl direct copy — exact layout preserved)
# ---------------------------------------------------------------------------

def compile_gst(folder: Path, log, progress_cb):
    """
    Open every GSTR-2B file in folder (sorted by month), copy each section's
    header block once, then append data rows from every month.

    Returns (output_path, total_rows, skipped_files).
    """
    excel_files = sorted(
        (f for f in folder.glob("*.xls*") if f.name != "Compiled_Data.xlsx"),
        key=_sort_key_by_date,
    )
    if not excel_files:
        log("No Excel files found.", "warn")
        return None, 0, []

    log(f"Found {len(excel_files)} file(s).\n", "info")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)          # drop the default blank sheet

    # Track state per section: ws, header copied?, next write row
    section_ws:           dict[str, object] = {}
    section_header_done:  dict[str, bool]   = {}
    section_next_row:     dict[str, int]    = {}

    total_rows   = 0
    skipped      = []
    n            = len(excel_files)

    for file_idx, file in enumerate(excel_files, 1):
        log(f"Processing ({file_idx}/{n}): {file.name}", "info")

        try:
            wb_src = load_workbook(file, data_only=True)
        except Exception as exc:
            log(f"  [SKIP] Cannot open '{file.name}': {exc}", "warn")
            skipped.append(file.name)
            progress_cb(file_idx / n * 90)
            continue

        for sheet_name, data_start_row in GST_SECTIONS.items():
            if sheet_name not in wb_src.sheetnames:
                continue

            ws_src = wb_src[sheet_name]

            # Check if this sheet has any data rows before doing anything
            has_data = any(
                not all(c.value is None for c in row)
                for row in ws_src.iter_rows(
                    min_row=data_start_row, max_row=ws_src.max_row)
            )
            if not has_data:
                log(f"  [SKIP] '{file.name}' > '{sheet_name}' is blank.", "warn")
                continue

            # ── First encounter: create output sheet + copy column widths ──
            if not section_header_done.get(sheet_name):
                ws_out = wb_out.create_sheet(title=sheet_name)
                section_ws[sheet_name]          = ws_out
                section_header_done[sheet_name] = True
                section_next_row[sheet_name]    = 1

                # Copy column widths once (from first file)
                for col_letter, col_dim in ws_src.column_dimensions.items():
                    ws_out.column_dimensions[col_letter].width = col_dim.width

            ws_out   = section_ws[sheet_name]
            next_row = section_next_row[sheet_name]

            # ── Copy header block before every file's data ────────────────
            hdr_rows = data_start_row - 1   # number of header rows to repeat
            for r in range(1, data_start_row):
                dest_r = next_row + (r - 1)
                for cell in ws_src[r]:
                    dst = ws_out.cell(row=dest_r, column=cell.column,
                                      value=cell.value)
                    _copy_cell_style(cell, dst)
                if r in ws_src.row_dimensions:
                    ws_out.row_dimensions[dest_r].height = \
                        ws_src.row_dimensions[r].height

            # Reproduce merged cells for this header block at current position
            offset = next_row - 1   # shift merged-cell rows by this amount
            for rng in ws_src.merged_cells.ranges:
                if rng.min_row < data_start_row:
                    new_min = rng.min_row + offset
                    new_max = rng.max_row + offset
                    try:
                        ws_out.merge_cells(
                            start_row=new_min, start_column=rng.min_col,
                            end_row=new_max,   end_column=rng.max_col,
                        )
                    except Exception:
                        pass

            next_row += hdr_rows
            rows_written = 0

            # ── Append data rows ──────────────────────────────────────────
            for src_row in ws_src.iter_rows(
                    min_row=data_start_row, max_row=ws_src.max_row):
                if all(c.value is None for c in src_row):
                    continue
                for cell in src_row:
                    ws_out.cell(row=next_row, column=cell.column,
                                value=cell.value)
                next_row     += 1
                rows_written += 1

            section_next_row[sheet_name] = next_row
            total_rows += rows_written
            log(f"  [OK]   '{file.name}' > '{sheet_name}' "
                f"({rows_written} rows)", "ok")

        wb_src.close()
        progress_cb(file_idx / n * 90)

    if not wb_out.sheetnames:
        return None, 0, skipped

    output_path = folder / "Compiled_Data.xlsx"
    wb_out.save(output_path)
    return output_path, total_rows, skipped


# ---------------------------------------------------------------------------
# Generic (non-GST) compile — pandas flat merge
# ---------------------------------------------------------------------------

def compile_generic(folder: Path, log, progress_cb):
    """Merge all sheets from all non-GST Excel files into one flat table."""
    excel_files = sorted(
        (f for f in folder.glob("*.xls*") if f.name != "Compiled_Data.xlsx"),
        key=_sort_key_by_date,
    )
    all_frames = []
    skipped    = []
    n          = len(excel_files)

    for i, file in enumerate(excel_files, 1):
        try:
            xl = pd.ExcelFile(file, engine="openpyxl")
        except Exception as exc:
            log(f"  [SKIP] Cannot open '{file.name}': {exc}", "warn")
            skipped.append(file.name)
            progress_cb(i / n * 90)
            continue

        for sheet_name in xl.sheet_names:
            try:
                df = xl.parse(sheet_name, header=0, dtype=str)
            except Exception as exc:
                log(f"  [SKIP] '{file.name}' > '{sheet_name}': {exc}", "warn")
                continue

            df.dropna(how="all", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)
            if df.empty:
                log(f"  [SKIP] '{file.name}' > '{sheet_name}' is blank.", "warn")
                continue

            df.insert(0, "__source_sheet__", sheet_name)
            df.insert(0, "__source_file__",  file.name)
            all_frames.append(df)
            log(f"  [OK]   '{file.name}' > '{sheet_name}' ({len(df)} rows)", "ok")

        progress_cb(i / n * 90)

    return all_frames, skipped


# ---------------------------------------------------------------------------
# Top-level compile dispatcher
# ---------------------------------------------------------------------------

def compile_folder(folder: Path, log, progress_cb):
    """
    Auto-detect whether the folder contains GSTR-2B files or generic Excel files
    and dispatch to the appropriate compile function.
    """
    excel_files = [
        f for f in folder.glob("*.xls*")
        if f.name != "Compiled_Data.xlsx"
    ]
    if not excel_files:
        log("No Excel files found in the selected folder.", "warn")
        return None, 0, []

    # Peek at the first file to decide mode
    first = sorted(excel_files, key=_sort_key_by_date)[0]
    if is_gst_2b_file(first):
        log("GSTR-2B files detected — using exact-layout merge.\n", "info")
        return compile_gst(folder, log, progress_cb)
    else:
        log("Generic Excel files detected — using flat merge.\n", "info")
        frames, skipped = compile_generic(folder, log, progress_cb)
        if not frames:
            return None, 0, skipped
        combined = pd.concat(frames, ignore_index=True, sort=False)
        output_path = folder / "Compiled_Data.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            combined.to_excel(writer, sheet_name="Compile Data Here", index=False)
        progress_cb(100)
        return output_path, len(combined), skipped


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Folder Compiler  (GST 2B Ready)")
        self.geometry("700x480")
        self.resizable(True, True)
        self.configure(bg="#f0f0f0")
        self._build_ui()

    def _build_ui(self):
        # ── Top bar ──────────────────────────────────────────────────────────
        top = tk.Frame(self, bg="#f0f0f0", padx=12, pady=10)
        top.pack(fill="x")
        tk.Label(top, text="Excel Folder Compiler  (GST 2B Ready)",
                 font=("Segoe UI", 14, "bold"), bg="#f0f0f0").pack(side="left")

        # ── Folder selection ─────────────────────────────────────────────────
        folder_frame = tk.Frame(self, bg="#f0f0f0", padx=12, pady=4)
        folder_frame.pack(fill="x")
        tk.Label(folder_frame, text="Source Folder:", bg="#f0f0f0",
                 font=("Segoe UI", 10)).pack(side="left")
        self.folder_var = tk.StringVar(value="No folder selected")
        tk.Label(folder_frame, textvariable=self.folder_var, bg="#ffffff",
                 relief="sunken", anchor="w", font=("Segoe UI", 9),
                 width=55, padx=4).pack(side="left", padx=6)
        self.browse_btn = tk.Button(
            folder_frame, text="Browse...", command=self._browse,
            font=("Segoe UI", 9), bg="#0078d4", fg="white", relief="flat",
            padx=10, pady=4, cursor="hand2")
        self.browse_btn.pack(side="left")

        # ── Run button ───────────────────────────────────────────────────────
        btn_frame = tk.Frame(self, bg="#f0f0f0", padx=12, pady=8)
        btn_frame.pack(fill="x")
        self.run_btn = tk.Button(
            btn_frame, text="Compile Excel Files", command=self._run,
            font=("Segoe UI", 10, "bold"), bg="#107c10", fg="white",
            relief="flat", padx=16, pady=6, cursor="hand2", state="disabled")
        self.run_btn.pack(side="left")
        self.status_lbl = tk.Label(btn_frame, text="", bg="#f0f0f0",
                                   font=("Segoe UI", 9), fg="#555")
        self.status_lbl.pack(side="left", padx=12)

        # ── Progress bar ─────────────────────────────────────────────────────
        prog_frame = tk.Frame(self, bg="#f0f0f0", padx=12, pady=2)
        prog_frame.pack(fill="x")
        self.progress = ttk.Progressbar(prog_frame, mode="determinate",
                                        length=660, maximum=100)
        self.progress.pack(fill="x")

        # ── Log panel ────────────────────────────────────────────────────────
        log_frame = tk.Frame(self, bg="#f0f0f0", padx=12, pady=6)
        log_frame.pack(fill="both", expand=True)
        tk.Label(log_frame, text="Log", bg="#f0f0f0",
                 font=("Segoe UI", 9, "bold")).pack(anchor="w")
        self.log_box = tk.Text(log_frame, height=14, state="disabled",
                               font=("Consolas", 9), bg="#1e1e1e", fg="#d4d4d4",
                               relief="flat", wrap="none")
        self.log_box.pack(fill="both", expand=True)
        scroll = ttk.Scrollbar(log_frame, command=self.log_box.yview)
        self.log_box["yscrollcommand"] = scroll.set
        scroll.pack(side="right", fill="y")
        self.log_box.tag_config("ok",   foreground="#4ec9b0")
        self.log_box.tag_config("warn", foreground="#ce9178")
        self.log_box.tag_config("info", foreground="#9cdcfe")
        self.log_box.tag_config("done", foreground="#dcdcaa")

        # ── Open output button ────────────────────────────────────────────────
        self.open_btn = tk.Button(
            self, text="Open Output File", command=self._open_output,
            font=("Segoe UI", 9), bg="#5c5c5c", fg="white", relief="flat",
            padx=12, pady=4, state="disabled")
        self.open_btn.pack(pady=(0, 10))
        self._output_path = None

    def _browse(self):
        folder = filedialog.askdirectory(title="Select Folder Containing Excel Files")
        if folder:
            self.folder_var.set(folder)
            self.run_btn.config(state="normal")
            self._log("Folder selected: " + folder, "info")

    def _log(self, msg: str, tag: str = "info"):
        def _append():
            self.log_box.config(state="normal")
            self.log_box.insert("end", msg + "\n", tag)
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _append)

    def _set_progress(self, val: float):
        self.after(0, lambda: self.progress.config(value=val))

    def _set_status(self, msg: str):
        self.after(0, lambda: self.status_lbl.config(text=msg))

    def _run(self):
        folder = Path(self.folder_var.get())
        if not folder.is_dir():
            messagebox.showerror("Error", "Please select a valid folder first.")
            return
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")
        self.progress["value"] = 0
        self.run_btn.config(state="disabled")
        self.open_btn.config(state="disabled")
        self._output_path = None
        threading.Thread(target=self._worker, args=(folder,), daemon=True).start()

    def _worker(self, folder: Path):
        self._set_status("Running...")
        result = compile_folder(folder, self._log, self._set_progress)
        out, total_rows, skipped = result if result[0] else (None, 0, [])

        if out is None:
            self._set_status("Failed — no data compiled.")
            self.after(0, lambda: self.run_btn.config(state="normal"))
            return

        self._set_progress(100)
        self._output_path = out
        summary = (
            f"\nDone!  Rows compiled: {total_rows:,} | "
            f"Files skipped: {len(skipped)}\n"
            f"Saved to: {out}"
        )
        self._log(summary, "done")
        self._set_status(f"Done — {total_rows:,} rows compiled.")
        self.after(0, lambda: self.run_btn.config(state="normal"))
        self.after(0, lambda: self.open_btn.config(state="normal"))

    def _open_output(self):
        if self._output_path and self._output_path.exists():
            import os
            os.startfile(self._output_path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = App()
    app.mainloop()
